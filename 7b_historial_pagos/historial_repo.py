"""
7b_historial_pagos/historial_repo.py — Único writer del ledger de pagos crudos

API PÚBLICA:
    registrar_pago(mz, lt, nombre, canal, monto, fecha, mes_ciclo, referencia,
                    *, origen_archivo, estado="identificado")              → dict
    registrar_correccion(mz_orig, lt_orig, mz_dest, lt_dest, canal, monto,
                          motivo, audit_ref, *, mes_ciclo, source,
                          ciclo_correccion=1)                               → dict (2 filas: -monto/+monto)
    identificar_blanco(mz, lt, nombre, monto, canal, referencia_original,
                        *, mes_ciclo, audit_ref, source)                    → dict
    leer_eventos(mes_ciclo)                                                 → DataFrame

INVARIANTES:
    - Único escritor del store (shared/reporte_acumulado_procesado/YYYY-MM_historial.xlsx).
    - Append-only: un evento nunca se modifica ni se borra.
    - FUENTE=pago se idempotiza por ORIGEN_ARCHIVO (re-importar el mismo libro no duplica).
    - FUENTE=correccion se idempotiza por (AUDIT_REF, MZ, LT, CANAL).
    - Conservación: registrar_correccion siempre escribe 2 filas (−monto en origen,
      +monto en destino) — la plata que sale de un predio entra al otro, nunca se
      crea ni se destruye.
    - Sin concepto: 1 fila = 1 pago crudo por canal. El reparto por concepto
      (convenio/multa/consumo) es responsabilidad de shared/seguimiento_repo.

Contrato visual: docs/formato_evento_pago.html
Diseño (Fase 1): README.md
"""

import os
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

STORE_DIR  = Path(__file__).parent.parent / "shared" / "reporte_acumulado_procesado"
SHEET_NAME = "Eventos"

CANALES_VALIDOS = {"efectivo", "yape"}
ESTADOS_VALIDOS = {"identificado", "blanco"}
FUENTES_VALIDAS = {"pago", "correccion"}

# ── Paleta (contrato: docs/formato_evento_pago.html) ───────────────────────────
_SEC_ID    = ("EBF5FB", "1A5276", "F4FAFF", "1A5276")
_SEC_PAGO  = ("E6F1FB", "0C447C", "F0F8FF", "0C447C")
_SEC_REF   = ("E9F7EF", "1E5C3A", "F4FBF7", "1E5C3A")
_SEC_AUDIT = ("F3E8FF", "5B21B6", "FAF5FF", "5B21B6")

_COLS = [
    ("MZ",               _SEC_ID,    6,  "center"),
    ("LT",               _SEC_ID,    7,  "center"),
    ("NOMBRE",           _SEC_ID,    28, "left"),
    ("CANAL",            _SEC_PAGO,  10, "center"),
    ("MONTO",            _SEC_PAGO,  12, "right"),
    ("FECHA",            _SEC_PAGO,  18, "center"),
    ("MES_CICLO",        _SEC_PAGO,  12, "center"),
    ("REFERENCIA",       _SEC_REF,   26, "left"),
    ("ESTADO",           _SEC_AUDIT, 14, "center"),
    ("FUENTE",           _SEC_AUDIT, 12, "center"),
    ("MOTIVO",           _SEC_AUDIT, 40, "left"),
    ("AUDIT_REF",        _SEC_AUDIT, 26, "left"),
    ("CICLO_CORRECCION", _SEC_AUDIT, 16, "center"),
    ("ORIGEN_ARCHIVO",   _SEC_AUDIT, 34, "left"),
    ("TIMESTAMP",        _SEC_AUDIT, 18, "center"),
]

_SECCIONES = [
    ("¿Quién es?",  "MZ",     "NOMBRE"),
    ("¿Qué pagó?",  "CANAL",  "MES_CICLO"),
    ("Referencia",  "REFERENCIA", "REFERENCIA"),
    ("Auditoría",   "ESTADO", "TIMESTAMP"),
]

_COLS_TEXTO = ("MZ", "LT", "NOMBRE", "CANAL", "FECHA", "MES_CICLO", "REFERENCIA",
               "ESTADO", "FUENTE", "MOTIVO", "AUDIT_REF", "ORIGEN_ARCHIVO", "TIMESTAMP")
_COLS_NUM   = ("MONTO", "CICLO_CORRECCION")

# Nombres de columna, en orden — público: lo usa consultar.py para construir un
# DataFrame vacío con el schema correcto cuando un predio no tiene eventos.
COLUMNAS = [c[0] for c in _COLS]

# ── Rutas ─────────────────────────────────────────────────────────────────────

def _ruta_store(mes_ciclo: str) -> Path:
    return STORE_DIR / f"{mes_ciclo}_historial.xlsx"

# ── Helpers internos (mismo patrón que shared/seguimiento_repo.py) ─────────────

def _save_atomic(wb, path: Path) -> None:
    """Guarda de forma atómica: temp en el mismo dir + os.replace(). Reintenta ante
    PermissionError transitorio (antivirus/indexador con el archivo agarrado)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.stem}.tmp{path.suffix}")
    wb.save(tmp)
    ultimo_error = None
    for intento in range(5):
        try:
            os.replace(tmp, path)
            return
        except PermissionError as e:
            ultimo_error = e
            time.sleep(0.3 * (intento + 1))
    raise ultimo_error


def _norm(v) -> str:
    if v is None:
        return ""
    return str(v).strip().upper().replace(" ", "")


def _argb(hex6: str) -> str:
    return "FF" + hex6.lstrip("#")


def _fill(hex6):
    return PatternFill("solid", fgColor=_argb(hex6))


def _hdr(cell, bg, fg, texto):
    cell.value = texto
    cell.fill  = _fill(bg)
    cell.font  = Font(color=_argb(fg), bold=True, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _dat(cell, valor, bg, fg, align="center"):
    cell.value = valor
    cell.fill  = _fill(bg)
    cell.font  = Font(color=_argb(fg), size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if isinstance(valor, (int, float)):
        cell.number_format = "#,##0.00"


def _write_headers(ws) -> None:
    col_idx = {c[0]: i + 1 for i, c in enumerate(_COLS)}
    for label, start, end in _SECCIONES:
        c1, c2 = col_idx[start], col_idx[end]
        sec = next(s for n, s, _, _ in _COLS if n == start)
        if c1 != c2:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        _hdr(ws.cell(row=1, column=c1), sec[0], sec[1], label)
    ws.row_dimensions[1].height = 18

    for i, (nombre, sec, ancho, _align) in enumerate(_COLS, start=1):
        _hdr(ws.cell(row=2, column=i), sec[0], sec[1], nombre)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"


def leer_eventos(mes_ciclo: str) -> pd.DataFrame:
    """Lee todos los eventos del mes como DataFrame. Vacío si el archivo no existe."""
    path = _ruta_store(mes_ciclo)
    if not path.exists():
        return pd.DataFrame(columns=[c[0] for c in _COLS])
    dtype_map = {c: str for c in _COLS_TEXTO}
    df = pd.read_excel(path, sheet_name=SHEET_NAME, header=1, dtype=dtype_map).fillna("")
    for col in _COLS_NUM:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def _ya_registrado_pago(df: pd.DataFrame, origen_archivo: str) -> bool:
    if df.empty:
        return False
    return bool((df["ORIGEN_ARCHIVO"].astype(str).str.strip() == origen_archivo).any())


def _ya_registrado_correccion(df: pd.DataFrame, audit_ref: str, mz: str, lt: str, canal: str) -> bool:
    if df.empty:
        return False
    mask = (
        (df["AUDIT_REF"].astype(str).str.strip() == audit_ref) &
        (df["MZ"].astype(str).map(_norm) == _norm(mz)) &
        (df["LT"].astype(str).map(_norm) == _norm(lt)) &
        (df["CANAL"].astype(str).str.strip() == canal)
    )
    return bool(mask.any())


def _append_filas(mes_ciclo: str, filas: list[dict]) -> None:
    path = _ruta_store(mes_ciclo)
    if path.exists():
        wb = load_workbook(path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        next_row = max(ws.max_row + 1, 3)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        _write_headers(ws)
        next_row = 3

    for fila in filas:
        for ci, (nombre, sec, _ancho, align) in enumerate(_COLS, start=1):
            _dat(ws.cell(row=next_row, column=ci), fila.get(nombre, ""), sec[2], sec[3], align=align)
        next_row += 1

    _save_atomic(wb, path)

# ── API pública: escritura ───────────────────────────────────────────────────

def registrar_pago(mz, lt, nombre, canal, monto, fecha, mes_ciclo, referencia,
                    *, origen_archivo: str, estado: str = "identificado") -> dict:
    """Observación cruda de un pago — nunca se edita después de escrita.

    fecha: string ya formateado ("DD/MM/YYYY" para efectivo, "DD/MM/YYYY HH:MM:SS"
    para yape — la precisión de segundos es lo que, junto con `referencia`, identifica
    el pago yape de forma única).
    """
    canal = str(canal).strip().lower()
    if canal not in CANALES_VALIDOS:
        raise ValueError(f"canal inválido: {canal!r} — válidos: {sorted(CANALES_VALIDOS)}")
    if estado not in ESTADOS_VALIDOS:
        raise ValueError(f"estado inválido: {estado!r} — válidos: {sorted(ESTADOS_VALIDOS)}")
    if not origen_archivo:
        raise ValueError("origen_archivo no puede ser vacío — es la clave de idempotencia")
    if estado == "identificado" and not (_norm(mz) and _norm(lt)):
        raise ValueError("estado=identificado requiere MZ y LT")
    if not referencia:
        raise ValueError("referencia no puede ser vacía")

    df = leer_eventos(mes_ciclo)
    if _ya_registrado_pago(df, origen_archivo):
        return {"skipped": True}

    fila = {
        "MZ": _norm(mz), "LT": _norm(lt), "NOMBRE": nombre or "",
        "CANAL": canal, "MONTO": float(monto), "FECHA": fecha, "MES_CICLO": mes_ciclo,
        "REFERENCIA": referencia, "ESTADO": estado, "FUENTE": "pago",
        "MOTIVO": "", "AUDIT_REF": "", "CICLO_CORRECCION": "",
        "ORIGEN_ARCHIVO": origen_archivo,
        "TIMESTAMP": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    _append_filas(mes_ciclo, [fila])
    return {"skipped": False}


def registrar_correccion(mz_orig, lt_orig, mz_dest, lt_dest, canal, monto, motivo,
                          audit_ref, *, mes_ciclo, source: str, ciclo_correccion: int = 1) -> dict:
    """Reatribución de pago entre dos predios — 2 filas (conservación): −monto en
    origen, +monto en destino. Nunca edita el evento de pago original."""
    canal = str(canal).strip().lower()
    if canal not in CANALES_VALIDOS:
        raise ValueError(f"canal inválido: {canal!r}")
    if not audit_ref:
        raise ValueError("audit_ref no puede ser vacío")
    if not motivo:
        raise ValueError("motivo no puede ser vacío")
    monto = float(monto)

    df = leer_eventos(mes_ciclo)
    ya_orig = _ya_registrado_correccion(df, audit_ref, mz_orig, lt_orig, canal)
    ya_dest = _ya_registrado_correccion(df, audit_ref, mz_dest, lt_dest, canal)
    if ya_orig and ya_dest:
        return {"skipped": True}

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    base = {
        "CANAL": canal, "FECHA": datetime.now().strftime("%d/%m/%Y"), "MES_CICLO": mes_ciclo,
        "REFERENCIA": "", "ESTADO": "identificado", "FUENTE": "correccion",
        "MOTIVO": motivo, "AUDIT_REF": audit_ref, "CICLO_CORRECCION": ciclo_correccion,
        "ORIGEN_ARCHIVO": source, "TIMESTAMP": ts,
    }
    filas = []
    if not ya_orig:
        filas.append({**base, "MZ": _norm(mz_orig), "LT": _norm(lt_orig), "NOMBRE": "", "MONTO": -monto})
    if not ya_dest:
        filas.append({**base, "MZ": _norm(mz_dest), "LT": _norm(lt_dest), "NOMBRE": "", "MONTO": monto})
    _append_filas(mes_ciclo, filas)
    return {"skipped": False}


def identificar_blanco(mz, lt, nombre, monto, canal, referencia_original,
                        *, mes_ciclo, audit_ref: str, source: str) -> dict:
    """Enlaza un pago yape en blanco a un predio — 1 fila FUENTE=correccion nueva.
    El evento blanco original NO se edita (queda como observación física)."""
    if not audit_ref:
        raise ValueError("audit_ref no puede ser vacío")
    if not (_norm(mz) and _norm(lt)):
        raise ValueError("identificar_blanco requiere MZ y LT de destino")

    df = leer_eventos(mes_ciclo)
    if _ya_registrado_correccion(df, audit_ref, mz, lt, canal):
        return {"skipped": True}

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fila = {
        "MZ": _norm(mz), "LT": _norm(lt), "NOMBRE": nombre or "",
        "CANAL": str(canal).strip().lower(), "MONTO": float(monto), "FECHA": ts,
        "MES_CICLO": mes_ciclo, "REFERENCIA": referencia_original,
        "ESTADO": "identificado", "FUENTE": "correccion",
        "MOTIVO": f"identificación de pago en blanco ({referencia_original})",
        "AUDIT_REF": audit_ref, "CICLO_CORRECCION": 1,
        "ORIGEN_ARCHIVO": source, "TIMESTAMP": ts,
    }
    _append_filas(mes_ciclo, [fila])
    return {"skipped": False}
