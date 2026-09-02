"""Genera la vista futura del ledger desde el snapshot validado del ciclo activo."""
import hashlib
import json
import sys
from pathlib import Path

import ciclo
import seguimiento_repo as repo
import utils_estado_ciclo as estado
from openpyxl import load_workbook


SHARED = Path(__file__).parent
ROOT = SHARED.parent
ESTADO_PATH = SHARED / "reporte_acumulado_procesado" / "estado_ciclo.json"
VALIDACION_PATH = ROOT / "5b_validacion" / "outputs" / "validacion_diferencias.xlsx"


def _alertas_validacion() -> str:
    if not VALIDACION_PATH.exists():
        return "5b no selló el snapshot; falta su reporte de validación"
    ws = load_workbook(VALIDACION_PATH, data_only=True, read_only=True)["resumen"]
    alertas = []
    for fila in ws.iter_rows(values_only=True):
        if "ALERTA" in fila:
            alertas.append(f"{fila[0]}: diferencia {float(fila[3]):+.2f}")
    return "; ".join(alertas) or "5b no selló el snapshot"


def main() -> Path:
    mes = ciclo.activo(path=SHARED / "ciclo_activo.json")
    ruta_snapshot = ROOT / "5_cobranza" / "outputs" / f"snapshot_ledger_{mes}.json"
    if not ruta_snapshot.exists():
        raise FileNotFoundError(f"Falta {ruta_snapshot}; correr 5_cobranza --force")

    documento = json.loads(ruta_snapshot.read_text(encoding="utf-8"))
    snapshot_hash = documento.pop("snapshot_hash", "")
    normalizado = json.dumps(documento, ensure_ascii=True, sort_keys=True,
                             separators=(",", ":")).encode("utf-8")
    if hashlib.sha256(normalizado).hexdigest() != snapshot_hash:
        raise RuntimeError("El hash del snapshot no coincide con su contenido")
    estado_ciclo = json.loads(ESTADO_PATH.read_text(encoding="utf-8"))
    arrastre = estado_ciclo.get(mes, {}).get("arrastre", {})
    if arrastre.get("snapshot_hash") != snapshot_hash:
        raise RuntimeError("El snapshot no coincide con el registrado en estado_ciclo.json")
    validado = estado.ciclo_validado(mes, ruta=ESTADO_PATH, snapshot_hash=snapshot_hash)
    permitir_no_validado = "--permitir-no-validado" in sys.argv
    if not validado and not permitir_no_validado:
        raise RuntimeError(f"El snapshot {snapshot_hash[:12]} de {mes} no está validado por 5b")

    salida = repo.generar_vista_provisional(
        mes, snapshot_hash, documento,
        estado_validacion="VALIDADO" if validado else "NO VALIDADO",
        alerta="" if validado else _alertas_validacion(),
    )
    print(f"Vista provisional -> {salida}")
    return salida


if __name__ == "__main__":
    main()
