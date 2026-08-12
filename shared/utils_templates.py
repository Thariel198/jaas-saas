"""
shared/utils_templates.py — Primitivos para templates Excel de input manual

Contrato visual: 4_pagos/efectivo/docs/formato_registro.html
Usado por: 4_pagos/efectivo/crear_templates.py (setup inicial) y 7_cierre
(reset al cerrar el período — ver docs/metodologia_desarrollo.md, "cross-módulo
nunca es import directo").
"""
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Secciones y columnas de mesa_N.xlsx (contrato formato_registro.html) ────
_GRUPOS = [
    ("¿Quién cobró?",     2, "EFF6FF", "1D4ED8"),
    ("¿Dónde vive?",      2, "E1F5EE", "085041"),
    ("¿Cuánto y cuándo?", 4, "FEF9E7", "7D6608"),
    ("¿Alguna nota?",     1, "F4ECF7", "5B21B6"),
    ("¿Qué tipo?",        1, "FFF7ED", "9A3412"),
    ("¿Qué pasó?",        1, "FCE4EC", "880E4F"),
]

_COLUMNAS = [
    ("COBRADOR",        "EFF6FF", "1D4ED8", 20),
    ("FECHA_REGISTRO",  "EFF6FF", "1D4ED8", 16),
    ("MZ",              "E1F5EE", "085041",  8),
    ("LT",              "E1F5EE", "085041",  8),
    ("MONTO",           "FEF9E7", "7D6608", 12),
    ("MONTO_EFECTIVO",  "FEF9E7", "7D6608", 16),
    ("MONTO_YAPE",      "FEF9E7", "7D6608", 14),
    ("FECHA",           "FEF9E7", "7D6608", 14),
    ("COMENTARIO",      "F4ECF7", "5B21B6", 30),
    ("CONCEPTO",        "FFF7ED", "9A3412", 14),
    ("CATEGORIA",       "FCE4EC", "880E4F", 14),
]

_EJEMPLO = ["María García", "10/06/2026", "A", "8C", "38.00", "20.00", "18.00", "03/06/2026", "", "", ""]

CATEGORIAS_VALIDAS = ["reclamo", "compromiso", "exoneracion", "otros"]

_HOJAS = ["registro_1", "registro_2", "registro_3"]


def _borde() -> Border:
    s = Side(style="thin", color="FFFFFF")
    return Border(left=s, right=s, top=s, bottom=s)


def _celda_grupo(cell, texto, bg, txt):
    cell.value     = texto
    cell.font      = Font(name="Arial", bold=True, size=9, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _borde()


def _celda_col(cell, texto, bg, txt):
    cell.value     = texto
    cell.font      = Font(name="Arial", bold=True, size=10, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _borde()


def _celda_ejemplo(cell, valor):
    cell.value     = valor
    cell.font      = Font(name="Arial", size=10, color="9CA3AF", italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _construir_hoja(ws) -> None:
    col = 1
    for texto, n_cols, bg, txt in _GRUPOS:
        if n_cols > 1:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + n_cols - 1)
        _celda_grupo(ws.cell(row=1, column=col), texto, bg, txt)
        col += n_cols
    ws.row_dimensions[1].height = 20

    for idx, (nombre, bg, txt, ancho) in enumerate(_COLUMNAS, start=1):
        _celda_col(ws.cell(row=2, column=idx), nombre, bg, txt)
        ws.column_dimensions[get_column_letter(idx)].width = ancho
    ws.row_dimensions[2].height = 22

    for idx, valor in enumerate(_EJEMPLO, start=1):
        _celda_ejemplo(ws.cell(row=3, column=idx), valor)
    ws.row_dimensions[3].height = 18

    agregar_dropdown_categoria(ws)
    aplicar_wrap_comentario(ws)

    ws.freeze_panes = "A3"


def agregar_dropdown_categoria(ws) -> None:
    """Dropdown de CATEGORIA (col 11) filas 4-500. También lo usa la migración
    v3 para agregar la validación a mesas con datos existentes."""
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(CATEGORIAS_VALIDAS) + '"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="CATEGORIA inválida",
        error="Valores: reclamo, compromiso, exoneracion, otros — o dejar vacío (pago normal).",
    )
    ws.add_data_validation(dv)
    col = get_column_letter(len(_COLUMNAS))
    dv.add(f"{col}4:{col}500")


def aplicar_wrap_comentario(ws) -> None:
    """COMENTARIO (col 9) filas 4-500: wrap_text + altura fija 18 = el texto
    largo queda en UNA línea recortada, sin desbordar sobre CONCEPTO/CATEGORIA;
    el contenido completo se ve al hacer click en la celda (barra de fórmulas)."""
    col_comentario = next(i for i, (n, *_r) in enumerate(_COLUMNAS, start=1) if n == "COMENTARIO")
    for fila in range(4, 501):
        c = ws.cell(row=fila, column=col_comentario)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        # altura manual (customHeight) desactiva el auto-fit → la fila no crece
        ws.row_dimensions[fila].height = 18


def filas_con_datos(ruta: Path) -> int:
    """Cuántas filas de cobro reales tiene un mesa_N.xlsx (fila 4+ de cada hoja).

    La fila 3 es el ejemplo guía del template y no cuenta. Sirve para saber si
    un archivo que se va a sobrescribir tiene trabajo del cobrador adentro."""
    if not ruta.exists():
        return 0
    try:
        wb = load_workbook(ruta, read_only=True, data_only=True)
    except Exception:
        return 0
    n = 0
    for hoja in wb.sheetnames:
        for fila in list(wb[hoja].values)[3:]:
            if fila and any(c not in (None, "") for c in fila[:6]):
                n += 1
    wb.close()
    return n


def respaldar_si_tiene_datos(ruta: Path) -> Path | None:
    """Copia el archivo a backup/ antes de pisarlo, si tiene filas de cobro.

    Por qué existe: el 26/07/2026 se corrió crear_templates.py para preparar el
    ciclo de agosto y sobrescribió las 7 mesas de julio —374 filas escritas a
    mano, con el split yape/efectivo y los comentarios del cobrador— sin
    backup. Se recuperaron de una copia suelta del repo por pura suerte. El
    respaldo va acá, en el primitivo, para que lo tengan TODOS los que
    resetean mesas (4_pagos/efectivo y 7_cierre), no solo el que se acuerde."""
    n = filas_con_datos(ruta)
    if n == 0:
        return None
    destino = ruta.parent.parent / "backup" / f"mesas_pre_reset_{datetime.now():%Y%m%d_%H%M%S}"
    destino.mkdir(parents=True, exist_ok=True)
    copia = destino / ruta.name
    shutil.copy2(ruta, copia)
    return copia


def crear_mesa_vacio(ruta: Path) -> None:
    """Escribe un mesa_N.xlsx vacío (3 hojas + fila-ejemplo) en `ruta`.
    Usado tanto para el setup inicial (4_pagos/efectivo) como para el
    reset de cierre de período (7_cierre).

    Si el archivo ya tenía filas de cobro, se respalda antes de pisarlo."""
    respaldar_si_tiene_datos(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    for i, hoja in enumerate(_HOJAS):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = hoja
        _construir_hoja(ws)
    wb.save(ruta)
