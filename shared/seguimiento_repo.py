"""
shared/seguimiento_repo.py — Único writer de seguimiento_pueblo.xlsx

API PÚBLICA:
    registrar_cargo(mz, lt, concepto, mes, monto, *, source, audit_ref, clase=, motivo=)  → dict
    registrar_pago(mz, lt, concepto, mes, monto, *, source, audit_ref, clase=, motivo=)   → dict
    registrar_ajuste(mz, lt, concepto, mes, monto, *, source, audit_ref, motivo, clase=)  → dict
    get_saldo(mz, lt, concepto, mes)          → float
    get_saldos_bulk(concepto, mes)            → dict {(mz,lt): saldo} (todos los predios, 1 sola lectura)
    pago_registrado(mz, lt, concepto, mes)    → float (Σ PAGO ya anotado — para reconciliación por delta)
    estado_cuenta(mz, lt, concepto)           → DataFrame (pivot ancho: DEUDA·PAGO·DECLARADO·AJUSTE·SALDO por mes, 1 predio)
    generar_vista(ruta=None)                  → Path (vista_seguimiento_pueblo.xlsx, 3 hojas + Ajustes + historial, TODOS los predios)
    exportar_vista_pdf(...)                   → Path (vista_seguimiento_pueblo.pdf, imprimible para la mesa)

INVARIANTES:
    - Único escritor de seguimiento_pueblo.xlsx (single writer).
    - Append-only real: un evento nunca se modifica ni se borra.
    - Idempotencia por (source, audit_ref, mz, lt, concepto) — mismo evento no duplica.
    - SALDO = Σcargos − Σpagos ± Σajustes, siempre derivado, nunca celda mutable a mano.
    - Génesis = el primer evento CARGO de un predio. No existe ledger de génesis aparte.
    - TIPO_EVENTO dice CÓMO se mueve el saldo; CLASE dice POR QUÉ pasó. El reporte de
      plata que entró = Σ PAGO con CLASE ∈ CLASES_SUMAN_CAJA — nunca Σ PAGO a secas.

Contrato visual: docs/formato_seguimiento_pueblo.html
Decisión de diseño: docs/decisiones/seguimiento_pueblo.md
"""

import logging
import hashlib
import json
import os
import shutil
import tempfile
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ── Rutas ─────────────────────────────────────────────────────────────────────
_SHARED = Path(__file__).parent

SEGUIMIENTO_PATH = _SHARED / "seguimiento_pueblo.xlsx"
ANULACIONES_PATH = _SHARED / "anulaciones_ledger.json"
SHEET_NAME       = "Eventos"

CONCEPTOS_VALIDOS = {
    "AGUA", "MANTENIMIENTO", "CORTE_RECONEXION",
    "MULTA", "ACUERDOS", "CONVENIO", "OTROS",
}
CONCEPTOS_ORDEN = (
    "AGUA", "MANTENIMIENTO", "CORTE_RECONEXION",
    "CONVENIO", "ACUERDOS", "MULTA", "OTROS",
)
TIPOS_EVENTO      = ("CARGO", "PAGO", "AJUSTE")
TOL_SALDO         = 0.005  # redondeo de céntimos: por debajo de esto el saldo es 0

# CLASE — qué HECHO representa el evento. TIPO_EVENTO dice cómo se mueve el
# saldo (CARGO sube, PAGO baja, AJUSTE corrige); CLASE dice por qué pasó.
# Sin esto, un AJUSTE de −75 no distingue "la directiva exoneró" de "pagó hace
# dos años" de "revertimos un bug" — las tres colapsan en la misma fila y el
# reporte de caja se vuelve indefendible.
CLASES_VALIDAS = {
    "COBRANZA",           # PAGO normal: plata del ciclo, vino por la cascada
    "ABONO_REZAGADO",     # plata real de un ciclo viejo que la caja recién ve
    "DECLARACION",        # "ya pagó" declarado, sin monto/fecha/canal que ubicar
    "DECLARACION_SECRETARIA",  # la secretaria confirma que el vecino ya pagó, y
                          # TODAVÍA NO SE INVESTIGÓ si esa plata es un exceso que ya
                          # estaba en caja o un pago nuevo que hay que sembrar en
                          # abonos_rezagados.xlsx. Marca de trabajo pendiente: no
                          # suma a caja mientras no se resuelva (ver LEER_ANTES.md)
    "EXONERACION",        # la directiva decidió no cobrar — nunca hubo plata
    "CORRECCION_SISTEMA", # defecto del código (pago fantasma, fórmula mal)
    "REASIGNACION",       # la deuda se movió entre conceptos, sin plata nueva
    "GENESIS",            # nació el cargo (siembra o génesis tardía)
    "SIN_CLASIFICAR",     # default de compatibilidad — visible a propósito
}

# Las únicas dos clases que representan plata que SUMA al total de caja.
# DECLARACION no está acá a propósito: esa plata ya entró y ya se contó (como
# exceso sin atribuir), lo que cambia es el dueño, no el total. Ver
# docs/diario/2026-08-03_solucion_precursor_mas_ledger.html
CLASES_SUMAN_CAJA = {"COBRANZA", "ABONO_REZAGADO"}

_CLASE_DEFAULT = {"CARGO": "GENESIS", "PAGO": "COBRANZA", "AJUSTE": "SIN_CLASIFICAR"}

# Predios de convenio de instalación (conexión nueva) — su deuda ya está
# completa en arrastre_consolidado (junio la cargó full desde DATA_boletas),
# NUNCA entran a seguimiento_pueblo para CONVENIO. Fuente única acá: tanto
# sembrar_seguimiento_pueblo.py (no los siembra) como 5_cobranza (no les
# reconcilia pagos de CONVENIO) importan esta lista — evita que ambos lados
# se desincronicen. Ver docs/decisiones/seguimiento_pueblo.md.
PREDIOS_INSTALACION_EXCLUIDOS = {("B", "20"), ("C", "43"), ("C", "35"), ("F1", "11"), ("G", "21"), ("W", "2")}

# ── Paleta (contrato: formato_seguimiento_pueblo.html) ────────────────────────
_SEC_PREDIO = ("EBF5FB", "1A5276", "F4FAFF", "1A5276")
_SEC_EVENTO = ("FEF3C7", "92400E", "FFFBEB", "92400E")
_SEC_MONTO  = ("F3E8FF", "5B21B6", "FAF5FF", "5B21B6")
_SEC_QUIEN  = ("F3F4F6", "374151", "F9FAFB", "374151")

_COLS = [
    ("MZ",          _SEC_PREDIO, 6,  "center"),
    ("LT",          _SEC_PREDIO, 7,  "center"),
    ("CONCEPTO",    _SEC_EVENTO, 12, "center"),
    ("MES",         _SEC_EVENTO, 10, "center"),
    ("TIPO_EVENTO", _SEC_EVENTO, 12, "center"),
    ("CARGO",       _SEC_MONTO,  12, "right"),
    ("PAGO",        _SEC_MONTO,  12, "right"),
    ("AJUSTE",      _SEC_MONTO,  12, "right"),
    ("SALDO",       _SEC_MONTO,  12, "right"),
    ("SOURCE",      _SEC_QUIEN,  24, "center"),
    ("AUDIT_REF",   _SEC_QUIEN,  28, "left"),
    ("TIMESTAMP",   _SEC_QUIEN,  20, "center"),
    # CLASE/MOTIVO van al FINAL a propósito: _append_evento escribe por índice
    # de _COLS, así que insertarlas en el medio desalinearía las 1.563 filas
    # ya escritas (su TIMESTAMP quedaría bajo el encabezado de otra columna).
    ("CLASE",       _SEC_QUIEN,  20, "center"),
    ("MOTIVO",      _SEC_QUIEN,  46, "left"),
]

_SECCIONES = [
    ("Predio",         "MZ",     "LT"),
    ("Evento",         "CONCEPTO", "TIPO_EVENTO"),
    ("Movimiento",     "CARGO",  "SALDO"),
    ("Quién / por qué", "SOURCE", "MOTIVO"),
]

# ── Helpers internos ─────────────────────────────────────────────────────────

def _save_atomic(wb, path: Path) -> None:
    """Guarda el workbook de forma atómica: escribe a un archivo temporal en el
    mismo directorio y lo reemplaza con os.replace() (atómico en Windows y POSIX
    — o el reemplazo pasa completo, o no pasa nada). Si el proceso se corta a
    mitad de camino, el archivo real nunca queda a medio escribir (un .xlsx
    corrupto). El temporal usa el mismo directorio para que el replace sea
    atómico incluso si /tmp está en otro filesystem.

    Reintenta el replace ante PermissionError — en Windows, un antivirus o
    indexador puede tener el archivo agarrado un instante justo después de
    escribirlo; es transitorio, no un fallo real (visto en producción:
    WinError 5 en la corrida real de la siembra, 2026-07-02)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.stem}.tmp{path.suffix}")
    wb.save(tmp)
    ultimo_error = None
    for intento in range(5):
        try:
            os.replace(tmp, path)
            return
        except PermissionError as e:
            ultimo_error = e
            time.sleep(0.3 * (intento + 1))
    raise ultimo_error


def _norm(v) -> str:
    if v is None:
        return ""
    return str(v).strip().upper().replace(" ", "")


# str(nan)='NAN' es truthy — sin esta lista, una fila de TOTALES (MZ/LT vacíos)
# pasa como predio fantasma (visto 2026-07-04: cargo CONVENIO NAN-NAN de S/1,334).
_CLAVES_INVALIDAS = {"", "NAN", "NONE", "NAT"}


def _clave_valida(mz_n: str, lt_n: str) -> bool:
    return mz_n not in _CLAVES_INVALIDAS and lt_n not in _CLAVES_INVALIDAS


def _argb(hex6: str) -> str:
    return "FF" + hex6.lstrip("#")


def _fill(hex6):
    return PatternFill("solid", fgColor=_argb(hex6))


def _hdr(cell, bg, fg, texto):
    cell.value = texto
    cell.fill  = _fill(bg)
    cell.font  = Font(color=_argb(fg), bold=True, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _dat(cell, valor, bg, fg, align="center"):
    cell.value = valor
    cell.fill  = _fill(bg)
    cell.font  = Font(color=_argb(fg), size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if isinstance(valor, (int, float)):
        cell.number_format = "#,##0.00"


def _validar_concepto(concepto: str) -> str:
    c = str(concepto).strip().upper()
    if c not in CONCEPTOS_VALIDOS:
        raise ValueError(f"concepto inválido: {concepto!r} — válidos: {sorted(CONCEPTOS_VALIDOS)}")
    return c


def _write_headers(ws) -> None:
    col_idx = {c[0]: i + 1 for i, c in enumerate(_COLS)}
    for label, start, end in _SECCIONES:
        c1, c2 = col_idx[start], col_idx[end]
        sec = next(s for n, s, _, _ in _COLS if n == start)
        if c1 != c2:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        _hdr(ws.cell(row=1, column=c1), sec[0], sec[1], label)
    ws.row_dimensions[1].height = 18

    for i, (nombre, sec, ancho, _align) in enumerate(_COLS, start=1):
        _hdr(ws.cell(row=2, column=i), sec[0], sec[1], nombre)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"


_COLS_TEXTO = ("MZ", "LT", "CONCEPTO", "MES", "TIPO_EVENTO", "SOURCE", "AUDIT_REF", "TIMESTAMP",
               "CLASE", "MOTIVO")
_COLS_NUM   = ("CARGO", "PAGO", "AJUSTE", "SALDO")


def _audit_refs_anulados() -> set[str]:
    """Referencias anuladas lógicamente; el ledger físico permanece append-only."""
    if not ANULACIONES_PATH.exists():
        return set()
    try:
        data = json.loads(ANULACIONES_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise RuntimeError(f"No se pudo leer {ANULACIONES_PATH}: {exc}") from exc

    refs = set()
    for anulacion in data.get("anulaciones", []):
        if anulacion.get("estado", "ACTIVA") == "REVOCADA":
            continue
        for evento in anulacion.get("eventos", []):
            ref = str(evento.get("audit_ref", "")).strip()
            if ref:
                refs.add(ref)
    return refs


def _leer_eventos(incluir_anulados: bool = False) -> pd.DataFrame:
    """Lee todos los eventos como DataFrame. Vacío (sin filas) si el archivo no existe.
    MZ/LT y demás columnas de texto se leen forzadas a str — sin esto, Excel infiere LT
    numérico ("6" → int64) y rompe la normalización downstream."""
    if not SEGUIMIENTO_PATH.exists():
        return pd.DataFrame(columns=[c[0] for c in _COLS])
    dtype_map = {c: str for c in _COLS_TEXTO}
    df = pd.read_excel(SEGUIMIENTO_PATH, sheet_name=SHEET_NAME, header=1, dtype=dtype_map)
    # CLASE/MOTIVO se agregaron cuando el archivo ya tenía historia. Un archivo
    # sin migrar no las trae — se materializan vacías para que ningún lector
    # tenga que preguntarse si la columna existe.
    for col in _COLS_TEXTO:
        if col not in df.columns:
            df[col] = ""
    for col in _COLS_NUM:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    if not incluir_anulados:
        refs = _audit_refs_anulados()
        if refs:
            df = df[~df["AUDIT_REF"].astype(str).str.strip().isin(refs)].copy()
    return df


def _ya_registrado(source: str, audit_ref: str, mz: str, lt: str, concepto: str) -> bool:
    """Idempotencia: mismo (source, audit_ref, mz, lt, concepto) no duplica."""
    df = _leer_eventos(incluir_anulados=True)
    if df.empty:
        return False
    mask = (
        (df["SOURCE"].astype(str).str.strip() == source) &
        (df["AUDIT_REF"].astype(str).str.strip() == audit_ref) &
        (df["MZ"].astype(str).map(_norm) == _norm(mz)) &
        (df["LT"].astype(str).map(_norm) == _norm(lt)) &
        (df["CONCEPTO"].astype(str).str.strip().str.upper() == concepto)
    )
    return bool(mask.any())


def _saldo_previo(mz: str, lt: str, concepto: str, mes: str) -> float:
    """Saldo acumulado ANTES de aplicar un evento nuevo — última fila del predio/concepto
    con MES <= mes dado, ordenado por TIMESTAMP. 0.0 si no hay eventos previos."""
    df = _leer_eventos()
    if df.empty:
        return 0.0
    mz_n, lt_n = _norm(mz), _norm(lt)
    sub = df[
        (df["MZ"].astype(str).map(_norm) == mz_n) &
        (df["LT"].astype(str).map(_norm) == lt_n) &
        (df["CONCEPTO"].astype(str).str.strip().str.upper() == concepto) &
        (df["MES"].astype(str) <= str(mes))
    ]
    if sub.empty:
        return 0.0
    sub = sub.sort_values(["MES", "TIMESTAMP"])
    return float(sub.iloc[-1]["SALDO"])


def _append_evento(mz, lt, concepto, mes, tipo_evento, cargo, pago, ajuste, saldo,
                   source, audit_ref, clase, motivo) -> None:
    if SEGUIMIENTO_PATH.exists():
        wb = load_workbook(SEGUIMIENTO_PATH)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        next_row = max(ws.max_row + 1, 3)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        _write_headers(ws)
        next_row = 3

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fila = {
        "MZ": _norm(mz), "LT": _norm(lt), "CONCEPTO": concepto, "MES": str(mes),
        "TIPO_EVENTO": tipo_evento,
        "CARGO": cargo, "PAGO": pago, "AJUSTE": ajuste, "SALDO": saldo,
        "SOURCE": source, "AUDIT_REF": audit_ref, "TIMESTAMP": ts,
        "CLASE": clase, "MOTIVO": motivo or "",
    }
    for ci, (nombre, sec, _ancho, align) in enumerate(_COLS, start=1):
        _dat(ws.cell(row=next_row, column=ci), fila[nombre], sec[2], sec[3], align=align)

    _save_atomic(wb, SEGUIMIENTO_PATH)

# ── API pública: escritura ───────────────────────────────────────────────────

def _registrar(mz, lt, concepto, mes, monto, tipo_evento, *, source, audit_ref,
               motivo=None, clase=None) -> dict:
    if not source:
        raise ValueError("source no puede ser vacío")
    if not audit_ref:
        raise ValueError("audit_ref no puede ser vacío")
    concepto = _validar_concepto(concepto)
    mz_n, lt_n = _norm(mz), _norm(lt)
    if not _clave_valida(mz_n, lt_n):
        raise ValueError(f"MZ/LT inválidos ({mz!r}, {lt!r})")
    if tipo_evento == "AJUSTE" and not motivo:
        raise ValueError("registrar_ajuste: motivo no puede ser vacío")
    clase = str(clase or _CLASE_DEFAULT[tipo_evento]).strip().upper()
    if clase not in CLASES_VALIDAS:
        raise ValueError(f"clase inválida: {clase!r} — válidas: {sorted(CLASES_VALIDAS)}")

    if _ya_registrado(source, audit_ref, mz_n, lt_n, concepto):
        log.info(f"seguimiento_repo: skip (idempotente) — {source}/{audit_ref} {mz_n}-{lt_n} {concepto}")
        return {"saldo_resultante": None, "skipped": True}

    monto = float(monto)
    saldo_previo = _saldo_previo(mz_n, lt_n, concepto, mes)

    cargo = pago = ajuste = None
    if tipo_evento == "CARGO":
        cargo = monto
        saldo_nuevo = saldo_previo + monto
    elif tipo_evento == "PAGO":
        pago = monto
        saldo_nuevo = saldo_previo - monto
    else:  # AJUSTE
        ajuste = monto
        saldo_nuevo = saldo_previo + monto

    # Un PAGO que deja el saldo por debajo de 0 es plata de más, y siempre hay que
    # mirarlo: puede ser exceso real (devolución o saldo a favor) o un error de
    # captura. Los dos casos que se dan seguido (06/08/2026):
    #   · la secretaria declara un monto mayor a la deuda — si lo que quería decir
    #     era "no debe nada", el pago tenía que ser exactamente la deuda;
    #   · 5_cobranza acredita plata del ciclo sobre un concepto que una declaración
    #     ya había saldado — la misma deuda pagada dos veces.
    # No se topa el monto a propósito: recortarlo en silencio borraría el exceso.
    # Se avisa y se devuelve en `exceso` para que el llamador pueda actuar.
    exceso = 0.0
    if tipo_evento == "PAGO" and saldo_nuevo < -TOL_SALDO:
        exceso = round(-saldo_nuevo, 2)
        log.warning(
            f"seguimiento_repo: {mz_n}-{lt_n} {concepto} {mes} · PAGO {monto:.2f} sobre una "
            f"deuda de {saldo_previo:.2f} → EXCESO {exceso:.2f}, el saldo queda en "
            f"{saldo_nuevo:.2f} (source={source}). ¿El exceso es real (devolución o saldo a "
            f"favor) o la intención era dejar la deuda en 0? En ese caso el pago debía ser "
            f"{max(saldo_previo, 0.0):.2f}.")

    _append_evento(mz_n, lt_n, concepto, mes, tipo_evento, cargo, pago, ajuste, saldo_nuevo,
                   source, audit_ref, clase, motivo)
    log.info(f"seguimiento_repo: {tipo_evento} {mz_n}-{lt_n} {concepto} {mes} "
             f"monto={monto} saldo={saldo_nuevo} clase={clase}")
    return {"saldo_resultante": saldo_nuevo, "skipped": False, "exceso": exceso}


def registrar_cargo(mz, lt, concepto, mes, monto, *, source: str, audit_ref: str,
                    clase: str | None = None, motivo: str | None = None) -> dict:
    """Deuda nueva — siembra inicial o cargo del mes (faena/reunión/asamblea/convenio).
    clase por defecto GENESIS."""
    return _registrar(mz, lt, concepto, mes, monto, "CARGO", source=source, audit_ref=audit_ref,
                      clase=clase, motivo=motivo)


def registrar_pago(mz, lt, concepto, mes, monto, *, source: str, audit_ref: str,
                   clase: str | None = None, motivo: str | None = None) -> dict:
    """5_cobranza registra la porción de un pago que va a este concepto.
    clase por defecto COBRANZA (plata del ciclo). Usar ABONO_REZAGADO cuando la
    plata es de un ciclo viejo que la caja recién ve."""
    return _registrar(mz, lt, concepto, mes, monto, "PAGO", source=source, audit_ref=audit_ref,
                      clase=clase, motivo=motivo)


def registrar_ajuste(mz, lt, concepto, mes, monto, *, source: str, audit_ref: str, motivo: str,
                     clase: str | None = None) -> dict:
    """Corrección manual — monto puede ser negativo. Nunca edita un evento pasado.
    `motivo` es obligatorio y AHORA SE PERSISTE (antes se exigía y se descartaba).
    `clase` dice qué hecho fue (EXONERACION · DECLARACION · CORRECCION_SISTEMA ·
    REASIGNACION); sin ella queda SIN_CLASIFICAR, visible a propósito."""
    return _registrar(mz, lt, concepto, mes, monto, "AJUSTE", source=source, audit_ref=audit_ref,
                      motivo=motivo, clase=clase)


def reconciliar_objetivos_batch(mes: str, snapshot_hash: str,
                                objetivos: list[dict], cargos: list[dict] | None = None) -> dict:
    """Compromete en un solo write el objetivo mensual producido por 5_cobranza.

    `objetivos` expresa SET_DEBE por (source,mz,lt,concepto). Para la transición
    también incluye claves que ya existan en el ledger y ya no estén en el
    snapshot, con objetivo cero. El archivo se guarda atómicamente y cada evento
    usa una referencia determinista ligada al hash del snapshot.
    """
    if not snapshot_hash:
        raise ValueError("snapshot_hash no puede ser vacío")
    cargos = cargos or []
    df = _leer_eventos()
    df_todos = _leer_eventos(incluir_anulados=True)
    cols = [c[0] for c in _COLS]
    if df.empty:
        df = pd.DataFrame(columns=cols)

    targets: dict[tuple[str, str, str, str], tuple[float, str]] = {}
    for obj in objetivos:
        source = str(obj["source"]).strip()
        mz, lt = _norm(obj["mz"]), _norm(obj["lt"])
        concepto = _validar_concepto(obj["concepto"])
        monto = round(float(obj["monto_objetivo"]), 2)
        if not source or not _clave_valida(mz, lt) or monto < -TOL_SALDO:
            raise ValueError(f"objetivo inválido: {obj}")
        targets[(source, mz, lt, concepto)] = (max(monto, 0.0), str(obj.get("clase") or "COBRANZA"))

    # Una fuente que escribió provisionalmente y desapareció del resultado final
    # debe converger a cero durante el primer cierre con este contrato.
    if not df.empty:
        prev = df[
            (df["MES"].astype(str) == str(mes)) &
            (df["SOURCE"].astype(str).isin({"5_cobranza", "abonos_rezagados"}))
        ]
        for _, row in prev.iterrows():
            key = (str(row["SOURCE"]).strip(), _norm(row["MZ"]), _norm(row["LT"]),
                   _validar_concepto(row["CONCEPTO"]))
            targets.setdefault(key, (0.0, "ABONO_REZAGADO" if key[0] == "abonos_rezagados" else "COBRANZA"))

    existentes = {
        (str(r["SOURCE"]).strip(), str(r["AUDIT_REF"]).strip(), _norm(r["MZ"]),
         _norm(r["LT"]), str(r["CONCEPTO"]).strip().upper())
        for _, r in df_todos.iterrows()
    }
    saldos: dict[tuple[str, str, str], float] = {}
    if not df.empty:
        orden = df.sort_values(["MZ", "LT", "CONCEPTO", "MES", "TIMESTAMP"])
        for _, row in orden.iterrows():
            if str(row["MES"]) <= str(mes):
                saldos[(_norm(row["MZ"]), _norm(row["LT"]), str(row["CONCEPTO"]).strip().upper())] = float(row["SALDO"])

    nuevas = []
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def agregar(mz, lt, concepto, tipo, monto, source, audit_ref, clase, motivo="",
                mes_evento=None):
        mz_n, lt_n = _norm(mz), _norm(lt)
        concepto_n = _validar_concepto(concepto)
        identidad = (source, audit_ref, mz_n, lt_n, concepto_n)
        if identidad in existentes:
            return
        key_saldo = (mz_n, lt_n, concepto_n)
        previo = saldos.get(key_saldo, 0.0)
        cargo = pago = ajuste = None
        if tipo == "CARGO":
            cargo, nuevo = monto, previo + monto
        elif tipo == "PAGO":
            pago, nuevo = monto, previo - monto
        else:
            ajuste, nuevo = monto, previo + monto
        saldos[key_saldo] = round(nuevo, 2)
        nuevas.append({
            "MZ": mz_n, "LT": lt_n, "CONCEPTO": concepto_n, "MES": str(mes_evento or mes),
            "TIPO_EVENTO": tipo, "CARGO": cargo, "PAGO": pago, "AJUSTE": ajuste,
            "SALDO": round(nuevo, 2), "SOURCE": source, "AUDIT_REF": audit_ref,
            "TIMESTAMP": ts, "CLASE": clase, "MOTIVO": motivo,
        })
        existentes.add(identidad)

    for cargo in sorted(cargos, key=lambda x: (str(x["mz"]), str(x["lt"]), str(x["concepto"]))):
        agregar(cargo["mz"], cargo["lt"], cargo["concepto"], "CARGO",
                round(float(cargo["monto"]), 2), str(cargo["source"]),
                str(cargo["audit_ref"]), str(cargo.get("clase") or "GENESIS"),
                str(cargo.get("motivo") or ""), cargo.get("mes"))

    for (source, mz, lt, concepto), (objetivo, clase) in sorted(targets.items()):
        sub = df[
            (df["MES"].astype(str) == str(mes)) &
            (df["SOURCE"].astype(str).str.strip() == source) &
            (df["MZ"].astype(str).map(_norm) == mz) &
            (df["LT"].astype(str).map(_norm) == lt) &
            (df["CONCEPTO"].astype(str).str.strip().str.upper() == concepto)
        ]
        ya = round(float(sub["PAGO"].fillna(0).sum() - sub["AJUSTE"].fillna(0).sum()), 2)
        delta = round(objetivo - ya, 2)
        historia = "|".join(sorted(
            f"{str(r['TIPO_EVENTO'])}:{str(r['AUDIT_REF'])}:{r.get('PAGO')}:{r.get('AJUSTE')}"
            for _, r in sub.iterrows()
        ))
        predecesor = hashlib.sha256(historia.encode("utf-8")).hexdigest()[:12]
        ref_base = (f"cierre|{mes}|{snapshot_hash[:16]}|{source}|{mz}|{lt}|{concepto}"
                    f"|{predecesor}|{ya:.2f}->{objetivo:.2f}")
        if delta > TOL_SALDO:
            agregar(mz, lt, concepto, "PAGO", delta, source, ref_base + "|PAGO", clase)
        elif delta < -TOL_SALDO:
            agregar(mz, lt, concepto, "AJUSTE", -delta, source, ref_base + "|AJUSTE",
                    "CORRECCION_SISTEMA", "cierre: el objetivo final es menor que lo acreditado provisionalmente")

    if not nuevas:
        return {"eventos": 0, "skipped": True, "backup": None}

    backup_dir = SEGUIMIENTO_PATH.parent / "backups_ledger" / "commits"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup = None
    if SEGUIMIENTO_PATH.exists():
        backup = backup_dir / f"seguimiento_pueblo_pre_commit_{mes}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        shutil.copy2(SEGUIMIENTO_PATH, backup)
        wb = load_workbook(SEGUIMIENTO_PATH)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        _write_headers(ws)

    for fila in nuevas:
        next_row = max(ws.max_row + 1, 3)
        for ci, (nombre, sec, _ancho, align) in enumerate(_COLS, start=1):
            _dat(ws.cell(row=next_row, column=ci), fila[nombre], sec[2], sec[3], align=align)
    _save_atomic(wb, SEGUIMIENTO_PATH)
    verificacion = _leer_eventos()
    refs_guardadas = set(verificacion["AUDIT_REF"].astype(str).str.strip())
    faltantes = {fila["AUDIT_REF"] for fila in nuevas} - refs_guardadas
    if faltantes:
        raise RuntimeError(f"Commit ledger incompleto; faltan referencias: {sorted(faltantes)}")
    return {"eventos": len(nuevas), "skipped": False, "backup": backup}

# ── API pública: lectura ─────────────────────────────────────────────────────

def pago_registrado(mz, lt, concepto, mes, source: str | None = None) -> float:
    """Σ PAGO ya registrado para (mz, lt, concepto, mes) — no incluye AJUSTE.
    Es el "SET_TIENE" de la reconciliación por delta que hace 5_cobranza en
    cada corrida (ver docs/decisiones/seguimiento_pueblo.md).

    `source` acota el conteo a lo que escribió ESE writer, igual que hace
    `ajuste_reconciliado`. 5_cobranza debe pasarlo: su reconciliación compara
    contra lo que él mismo dejó, no contra todo lo que haya en el mes. Sin el
    filtro contaba como propios los PAGO manuales — las declaraciones de la
    secretaria del 28/07/2026 — veía "de más", emitía un AJUSTE negativo a
    ciegas y alguien tenía que estabilizarlo a mano: 44 filas de ruido en julio
    2026 (ver LEER_ANTES.md § "los pagos que declaró la secretaria")."""
    concepto = _validar_concepto(concepto)
    df = _leer_eventos()
    if df.empty:
        return 0.0
    mz_n, lt_n = _norm(mz), _norm(lt)
    sub = df[
        (df["MZ"].astype(str).map(_norm) == mz_n) &
        (df["LT"].astype(str).map(_norm) == lt_n) &
        (df["CONCEPTO"].astype(str).str.strip().str.upper() == concepto) &
        (df["MES"].astype(str) == str(mes)) &
        (df["TIPO_EVENTO"] == "PAGO")
    ]
    if source is not None:
        sub = sub[sub["SOURCE"].astype(str).str.strip() == str(source).strip()]
    if sub.empty:
        return 0.0
    return float(sub["PAGO"].fillna(0).sum())


def ajuste_reconciliado(mz, lt, concepto, mes, source) -> float:
    """Σ AJUSTE ya registrado por `source` para (mz, lt, concepto, mes).
    Complementa a pago_registrado() en la reconciliación por delta: el "SET_TIENE"
    real que dejó una corrida es Σ PAGO + Σ AJUSTE(source propio). Filtra por
    `source` para NO contar ajustes de otro origen (ej. correccion_genesis_formula
    corrige la deuda, no el pago). Sin este término el branch AJUSTE no es
    idempotente y se re-dispara en cada corrida (ver docs/aprendizaje contador tuerto)."""
    concepto = _validar_concepto(concepto)
    df = _leer_eventos()
    if df.empty:
        return 0.0
    mz_n, lt_n = _norm(mz), _norm(lt)
    sub = df[
        (df["MZ"].astype(str).map(_norm) == mz_n) &
        (df["LT"].astype(str).map(_norm) == lt_n) &
        (df["CONCEPTO"].astype(str).str.strip().str.upper() == concepto) &
        (df["MES"].astype(str) == str(mes)) &
        (df["TIPO_EVENTO"] == "AJUSTE") &
        (df["SOURCE"].astype(str) == str(source))
    ]
    if sub.empty:
        return 0.0
    return float(sub["AJUSTE"].fillna(0).sum())


def get_saldo(mz, lt, concepto, mes) -> float:
    """Saldo de (mz, lt, concepto) al cierre de `mes` (inclusive). 0.0 si no hay eventos."""
    concepto = _validar_concepto(concepto)
    df = _leer_eventos()
    if df.empty:
        return 0.0
    mz_n, lt_n = _norm(mz), _norm(lt)
    sub = df[
        (df["MZ"].astype(str).map(_norm) == mz_n) &
        (df["LT"].astype(str).map(_norm) == lt_n) &
        (df["CONCEPTO"].astype(str).str.strip().str.upper() == concepto) &
        (df["MES"].astype(str) <= str(mes))
    ]
    if sub.empty:
        return 0.0
    sub = sub.sort_values(["MES", "TIMESTAMP"])
    return float(sub.iloc[-1]["SALDO"])


def get_saldos_bulk(concepto, mes) -> dict:
    """{(mz, lt): saldo} al cierre de `mes`, para TODOS los predios de una vez.

    2_planilla necesita el saldo de ~575 predios por concepto — llamar
    get_saldo() por predio sería 575 lecturas completas del registro (mismo
    patrón O(n) por llamada que hizo lenta la siembra). Esta función lee el
    registro una sola vez y calcula todos los saldos de un pase."""
    concepto = _validar_concepto(concepto)
    df = _leer_eventos()
    if df.empty:
        return {}
    sub = df[
        (df["CONCEPTO"].astype(str).str.strip().str.upper() == concepto) &
        (df["MES"].astype(str) <= str(mes))
    ]
    if sub.empty:
        return {}
    sub = sub.sort_values(["MZ", "LT", "MES", "TIMESTAMP"])
    ultimo = sub.groupby(["MZ", "LT"], as_index=False).last()
    return {(_norm(r["MZ"]), _norm(r["LT"])): float(r["SALDO"]) for _, r in ultimo.iterrows()}


def estado_cuenta(mz, lt, concepto) -> pd.DataFrame:
    """Pivot a la vista ancha — una fila (mz, lt), columnas DEUDA/PAGO/SALDO por mes.
    Regenerable siempre desde el registro largo — nunca es la fuente."""
    concepto = _validar_concepto(concepto)
    df = _leer_eventos()
    mz_n, lt_n = _norm(mz), _norm(lt)
    sub = df[
        (df["MZ"].astype(str).map(_norm) == mz_n) &
        (df["LT"].astype(str).map(_norm) == lt_n) &
        (df["CONCEPTO"].astype(str).str.strip().str.upper() == concepto)
    ].sort_values(["MES", "TIMESTAMP"])

    meses = sorted(sub["MES"].astype(str).unique())
    filas = []
    for mes in meses:
        del_mes = sub[sub["MES"].astype(str) == mes]
        deuda = float(del_mes["CARGO"].fillna(0).sum())
        pago, declarado = _partir_pago(del_mes)
        ajuste = float(del_mes["AJUSTE"].fillna(0).sum())
        saldo = float(del_mes.iloc[-1]["SALDO"])
        filas.append({"MES": mes, "DEUDA": deuda, "PAGO": pago,
                      "DECLARADO": declarado, "AJUSTE": ajuste, "SALDO": saldo})
    return pd.DataFrame(filas, columns=["MES", "DEUDA", "PAGO", "DECLARADO", "AJUSTE", "SALDO"])


def _partir_pago(del_mes: pd.DataFrame) -> tuple[float, float]:
    """(plata que entró, plata solo declarada) de un bloque de eventos.

    En el ledger los dos son TIPO_EVENTO=PAGO — mueven el saldo igual. Lo que los
    distingue es la CLASE, y eso es un atributo de la fila, no una columna. Al
    pivotear a formato ancho la celda agrega varios eventos y no hay dónde llevar
    ese atributo, así que la dimensión se convierte en dos columnas.

    No es cosmético: a agosto/2026 son S/1,257 en 28 filas DECLARACION_SECRETARIA
    (la secretaria dice que ya pagó, sin verificar de dónde salió la plata), y hay
    5 celdas que mezclan las dos cosas en el mismo mes. El corte sale de
    CLASES_SUMAN_CAJA, el mismo que usa el resto del código."""
    pagos = del_mes[del_mes["TIPO_EVENTO"].astype(str).str.strip() == "PAGO"]
    if pagos.empty:
        return 0.0, 0.0
    es_caja = pagos["CLASE"].fillna("").astype(str).str.strip().str.upper().isin(CLASES_SUMAN_CAJA)
    return (float(pagos[es_caja]["PAGO"].fillna(0).sum()),
            float(pagos[~es_caja]["PAGO"].fillna(0).sum()))


def deudores(concepto, minimo: float = 0.0) -> pd.DataFrame:
    """Lotes con saldo vigente (última fila por mz,lt,concepto) >= minimo."""
    concepto = _validar_concepto(concepto)
    df = _leer_eventos()
    if df.empty:
        return pd.DataFrame(columns=["MZ", "LT", "SALDO"])
    sub = df[df["CONCEPTO"].astype(str).str.strip().str.upper() == concepto].copy()
    if sub.empty:
        return pd.DataFrame(columns=["MZ", "LT", "SALDO"])
    sub = sub.sort_values(["MZ", "LT", "MES", "TIMESTAMP"])
    ultimo = sub.groupby(["MZ", "LT"], as_index=False).last()[["MZ", "LT", "SALDO"]]
    return ultimo[ultimo["SALDO"] >= minimo].reset_index(drop=True)

# ── Vista ancha regenerable ────────────────────────────────────────────────────
# Contrato: docs/formato_vista_seguimiento_pueblo.html

VISTA_PATH = _SHARED / "vista_seguimiento_pueblo.xlsx"
VISTA_PROVISIONAL_PATH = _SHARED / "vista_seguimiento_provicional.xlsx"

_SEC_PREDIO_V = ("EBF5FB", "1A5276", "F4FAFF", "1A5276")
# (header_bg, header_fg, deuda_fg, pago_fg, saldo_bg, saldo_fg, ajuste_fg, declarado_fg)
# — cicla cada 3 meses
_PALETAS_MES = [
    ("FEF3C7", "92400E", "92400E", "065F46", "FDE68A", "78350F", "5B21B6", "B45309"),
    ("DBEAFE", "1E3A8A", "1E3A8A", "065F46", "BFDBFE", "1E3A8A", "5B21B6", "B45309"),
    ("EDE9FE", "5B21B6", "5B21B6", "065F46", "DDD6FE", "4C1D95", "92400E", "B45309"),
]
# AJUSTE sin MOTIVO: el ledger no sabe por qué se movió ese saldo. Rojo a propósito.
_AJUSTE_SIN_MOTIVO_FG = "B91C1C"
_DATA_BOLETAS_PATH_VISTA = _SHARED.parent / "3_boletas" / "inputs" / "DATA_boletas.xlsx"
_PADRON_PATH_VISTA = _SHARED.parent / "0_padron" / "02_matching" / "outputs" / "padron_reconciliado.xlsx"


def _lookup_nombres() -> dict:
    """(mz, lt) → NOMBRES. Base: padrón reconciliado (cubre TODOS los predios,
    también sin servicio); encima DATA_boletas (verdad post-reclamo, solo los
    facturados del mes). {} si no existe ninguno."""
    out = {}
    for path, sheet, col in ((_PADRON_PATH_VISTA, 0, "Nombres"),
                             (_DATA_BOLETAS_PATH_VISTA, "Data", "NOMBRES")):
        if not path.exists():
            continue
        try:
            df = pd.read_excel(path, sheet_name=sheet, dtype=str)
        except Exception:
            continue
        for _, r in df.iterrows():
            mz, lt, nombre = r.get("MZ"), r.get("LT"), str(r.get(col, "")).strip()
            if mz and lt and nombre and nombre.upper() != "NAN":
                out[(_norm(mz), _norm(lt))] = nombre
    return out


_COLS_POR_MES = 5  # DEUDA · PAGO · DECLARADO · AJUSTE · SALDO


def _escribir_hoja_ajustes(ws, df: pd.DataFrame, nombres: dict) -> None:
    """Una fila por AJUSTE, con su MOTIVO completo. La grilla de las hojas por
    concepto muestra el monto; el porqué vive acá — hay celdas (predio·concepto·mes)
    con hasta 3 ajustes, así que no entra en el grid sin comprimir algo."""
    cols = [("MZ", 6), ("LT", 7), ("NOMBRE", 26), ("CONCEPTO", 12), ("MES", 10),
            ("AJUSTE", 11), ("CLASE", 22), ("SOURCE", 22), ("MOTIVO", 90)]

    _hdr(ws.cell(row=1, column=1), *_SEC_PREDIO_V[:2], "Predio")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    _hdr(ws.cell(row=1, column=4), _SEC_EVENTO[0], _SEC_EVENTO[1], "Ajuste")
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)
    _hdr(ws.cell(row=1, column=7), _SEC_QUIEN[0], _SEC_QUIEN[1], "Por qué")
    ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=9)
    ws.row_dimensions[1].height = 18

    for i, (nombre, ancho) in enumerate(cols, start=1):
        sec = _SEC_PREDIO_V[:2] if i <= 3 else (_SEC_EVENTO[:2] if i <= 6 else _SEC_QUIEN[:2])
        _hdr(ws.cell(row=2, column=i), sec[0], sec[1], nombre)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "D3"

    aj = df[df["TIPO_EVENTO"].astype(str).str.strip() == "AJUSTE"].copy()
    if aj.empty:
        return
    aj = aj.sort_values(["MZ", "LT", "CONCEPTO", "MES", "TIMESTAMP"])

    for r_off, (_, ev) in enumerate(aj.iterrows()):
        row = r_off + 3
        mz, lt = _norm(ev["MZ"]), _norm(ev["LT"])
        motivo = str(ev["MOTIVO"] if pd.notna(ev["MOTIVO"]) else "").strip()
        mudo = motivo == ""
        valores = [mz, lt, nombres.get((mz, lt), ""), str(ev["CONCEPTO"]).strip(),
                   str(ev["MES"]).strip(), float(ev["AJUSTE"] or 0),
                   str(ev["CLASE"] if pd.notna(ev["CLASE"]) else "").strip(),
                   str(ev["SOURCE"] if pd.notna(ev["SOURCE"]) else "").strip(),
                   "(SIN MOTIVO — falta explicar por qué se movió este saldo)" if mudo else motivo]
        for i, val in enumerate(valores, start=1):
            cell = ws.cell(row=row, column=i, value=val)
            cell.fill = _fill("F4FAFF" if i <= 3 else ("FFFBEB" if i <= 6 else "F9FAFB"))
            rojo = mudo and i in (7, 9)
            cell.font = Font(color=_argb(_AJUSTE_SIN_MOTIVO_FG if rojo else "374151"),
                             bold=(i == 6 or rojo), size=10)
            cell.alignment = Alignment(horizontal="right" if i == 6 else
                                       ("center" if i in (1, 2, 4, 5) else "left"),
                                       vertical="center")
            if i == 6:
                cell.number_format = "+#,##0.00;-#,##0.00;0.00"


def _escribir_hoja_vista(ws, concepto: str, df_concepto: pd.DataFrame, nombres: dict) -> None:
    predios = sorted({(r["MZ"], r["LT"]) for _, r in df_concepto.iterrows()
                      if _clave_valida(_norm(r["MZ"]), _norm(r["LT"]))})
    meses = sorted(df_concepto["MES"].astype(str).unique())

    col_predio = [("MZ", 6, "center"), ("LT", 7, "center"), ("NOMBRE", 26, "left")]
    cols = list(col_predio)
    for mes in meses:
        cols += [(f"{mes}|DEUDA", 11, "right"), (f"{mes}|PAGO", 11, "right"),
                 (f"{mes}|DECLARADO", 12, "right"), (f"{mes}|AJUSTE", 11, "right"),
                 (f"{mes}|SALDO", 11, "right")]

    # ── Fila 1: secciones (Predio + 1 por mes) ──
    _hdr(ws.cell(row=1, column=1), *_SEC_PREDIO_V[:2], "Predio")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    for i, mes in enumerate(meses):
        pal = _PALETAS_MES[i % 3]
        c1 = 4 + i * _COLS_POR_MES
        ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c1 + _COLS_POR_MES - 1)
        _hdr(ws.cell(row=1, column=c1), pal[0], pal[1], mes)
    ws.row_dimensions[1].height = 18

    # ── Fila 2: columnas ──
    for i, (nombre, ancho, _align) in enumerate(cols, start=1):
        etiqueta = nombre.split("|")[-1]
        if i <= 3:
            _hdr(ws.cell(row=2, column=i), *_SEC_PREDIO_V[:2], etiqueta)
        else:
            mes_idx = (i - 4) // _COLS_POR_MES
            pal = _PALETAS_MES[mes_idx % 3]
            _hdr(ws.cell(row=2, column=i), pal[0], pal[1], etiqueta)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "D3"

    if not predios:
        return

    # ── Filas de datos ──
    for r_off, (mz, lt) in enumerate(predios):
        row = r_off + 3
        _dat(ws.cell(row=row, column=1), mz, *_SEC_PREDIO_V[2:], align="center")
        _dat(ws.cell(row=row, column=2), lt, *_SEC_PREDIO_V[2:], align="center")
        _dat(ws.cell(row=row, column=3), nombres.get((mz, lt), ""), *_SEC_PREDIO_V[2:], align="left")

        sub_predio = df_concepto[(df_concepto["MZ"] == mz) & (df_concepto["LT"] == lt)]
        for i, mes in enumerate(meses):
            pal = _PALETAS_MES[i % 3]
            base = 4 + i * _COLS_POR_MES
            c_deuda, c_pago, c_decl = base, base + 1, base + 2
            c_ajuste, c_saldo = base + 3, base + 4
            del_mes = sub_predio[sub_predio["MES"].astype(str) == mes]

            if del_mes.empty:
                for c in (c_deuda, c_pago, c_decl, c_ajuste, c_saldo):
                    cell = ws.cell(row=row, column=c, value="—")
                    cell.fill = _fill("F3F4F6")
                    cell.font = Font(color=_argb("9CA3AF"), italic=True, size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            # DEUDA es solo lo que se le CARGÓ. El AJUSTE va aparte: fundirlo acá
            # (como se hacía hasta el 06/08/2026) hacía que una corrección se
            # leyera como deuda nueva — P-6 parecía tener un cargo de 58 en julio
            # que en realidad era el ajuste que tapaba un desfase de mes.
            deuda = float(del_mes["CARGO"].fillna(0).sum())
            pago, declarado = _partir_pago(del_mes)
            ajustes = del_mes[del_mes["TIPO_EVENTO"].astype(str).str.strip() == "AJUSTE"]
            ajuste = float(ajustes["AJUSTE"].fillna(0).sum())
            sin_motivo = (not ajustes.empty and
                          ajustes["MOTIVO"].fillna("").astype(str).str.strip().eq("").all())
            saldo = float(del_mes.sort_values("TIMESTAMP").iloc[-1]["SALDO"])

            _dat(ws.cell(row=row, column=c_deuda), deuda, "FFFBEB", pal[2], align="right")
            _dat(ws.cell(row=row, column=c_pago), pago, "FFFBEB", pal[3], align="right")
            cell_d = ws.cell(row=row, column=c_decl)
            if declarado <= TOL_SALDO:
                cell_d.value = "·"
                cell_d.fill = _fill("FFFBEB")
                cell_d.font = Font(color=_argb("D1D5DB"), size=10)
                cell_d.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell_d.value = declarado
                cell_d.fill = _fill("FFFBEB")
                cell_d.font = Font(color=_argb(pal[7]), bold=True, size=10)
                cell_d.alignment = Alignment(horizontal="right", vertical="center")
                cell_d.number_format = "#,##0.00"
            cell_a = ws.cell(row=row, column=c_ajuste)
            if ajustes.empty:
                cell_a.value = "·"
                cell_a.fill = _fill("FFFBEB")
                cell_a.font = Font(color=_argb("D1D5DB"), size=10)
                cell_a.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell_a.value = ajuste
                cell_a.fill = _fill("FFFBEB")
                cell_a.font = Font(color=_argb(_AJUSTE_SIN_MOTIVO_FG if sin_motivo else pal[6]),
                                   bold=True, size=10)
                cell_a.alignment = Alignment(horizontal="right", vertical="center")
                cell_a.number_format = "+#,##0.00;-#,##0.00;0.00"
            cell_s = ws.cell(row=row, column=c_saldo)
            cell_s.value = saldo
            cell_s.fill = _fill(pal[4])
            cell_s.font = Font(color=_argb(pal[5]), bold=True, size=10)
            cell_s.alignment = Alignment(horizontal="right", vertical="center")
            cell_s.number_format = "#,##0.00"


_MEDIDOR_SALDO_PATH_VISTA = _SHARED / "genesis_inputs" / "medidor_saldo.xlsx"

# (mes_label, columna_en_medidor_saldo)
_CUOTAS_HISTORIAL = [
    ("DIC-25", "Diciembre-2025"), ("ENE-26", "Enero-2025"), ("FEB-26", "Febrero-2026"),
    ("MAR-26", "Marzo-2026"), ("ABR-26", "Abril-2026"), ("MAY-26", "Mayo-2026"),
]


_MESES_ES = {"01": "ENE", "02": "FEB", "03": "MAR", "04": "ABR", "05": "MAY", "06": "JUN",
             "07": "JUL", "08": "AGO", "09": "SEP", "10": "OCT", "11": "NOV", "12": "DIC"}


def _mes_label(mes: str) -> str:
    a, m = str(mes).split("-")
    return f"{_MESES_ES.get(m, m)}-{a[2:]}"


def _escribir_hoja_historial_convenio(ws, nombres: dict, df_conv: pd.DataFrame) -> None:
    """Historia COMPLETA del convenio de medidor por predio: cuotas dic-2025 a
    may-2026 desde el snapshot de génesis (el ledger colapsó ese período en un
    solo CARGO — el snapshot es el único lugar con ese detalle) + pagos de
    junio-2026 en adelante desde el ledger (una columna por mes, crece sola)
    + SALDO ACTUAL. MULTA/ACUERDOS no tienen detalle pre-ledger (entraron como
    monto total)."""
    df = pd.read_excel(_MEDIDOR_SALDO_PATH_VISTA, sheet_name="Cobro medidores")

    # pagos por mes y saldo vigente desde el ledger
    meses_ledger = sorted(df_conv["MES"].astype(str).unique())
    pagos_mes: dict = {}
    saldo_actual: dict = {}
    for (mz, lt), sub in df_conv.groupby(
            [df_conv["MZ"].astype(str).map(_norm), df_conv["LT"].astype(str).map(_norm)]):
        for mes, del_mes in sub.groupby(sub["MES"].astype(str)):
            pagos_mes[(mz, lt, mes)] = float(del_mes["PAGO"].fillna(0).sum())
        saldo_actual[(mz, lt)] = float(sub.sort_values(["MES", "TIMESTAMP"]).iloc[-1]["SALDO"])

    def _lt_txt(v):
        try:
            f = float(v)
            if f == int(f):
                return str(int(f))
        except (ValueError, TypeError):
            pass
        return _norm(v)

    pal = _PALETAS_MES[0]
    pal2 = _PALETAS_MES[1]
    n_cuotas = len(_CUOTAS_HISTORIAL)
    cols = [("MZ", 6), ("LT", 7), ("NOMBRE", 26), ("DEUDA", 11)]
    cols += [(lbl, 11) for lbl, _ in _CUOTAS_HISTORIAL]
    cols += [(f"PAGO {_mes_label(m)}", 12) for m in meses_ledger]
    cols += [("SALDO ACTUAL", 14)]
    c_ledger_ini = 5 + n_cuotas          # primera columna de pagos del ledger

    _hdr(ws.cell(row=1, column=1), *_SEC_PREDIO_V[:2], "Predio")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=c_ledger_ini - 1)
    _hdr(ws.cell(row=1, column=4), pal[0], pal[1],
         "Cuotas antes del ledger (snapshot génesis)")
    ws.merge_cells(start_row=1, start_column=c_ledger_ini, end_row=1, end_column=len(cols))
    _hdr(ws.cell(row=1, column=c_ledger_ini), pal2[0], pal2[1],
         "Pagos en el ledger (jun-2026 →) y saldo vigente")
    ws.row_dimensions[1].height = 18
    for i, (nombre, ancho) in enumerate(cols, start=1):
        p = _SEC_PREDIO_V[:2] if i <= 3 else ((pal[0], pal[1]) if i < c_ledger_ini else (pal2[0], pal2[1]))
        _hdr(ws.cell(row=2, column=i), *p, nombre)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "D3"

    filas = []
    for _, r in df.iterrows():
        mz, lt = _norm(r.get("MZ")), _lt_txt(r.get("LT"))
        deuda = pd.to_numeric(r.get("Deuda"), errors="coerce")
        if not _clave_valida(mz, lt) or pd.isna(deuda) or deuda <= 0:
            continue
        cuotas = [float(pd.to_numeric(r.get(col), errors="coerce") or 0) for _, col in _CUOTAS_HISTORIAL]
        saldo_gen = float(pd.to_numeric(r.get("Saldo"), errors="coerce") or 0)
        pagos = [pagos_mes.get((mz, lt, m), 0.0) for m in meses_ledger]
        saldo = saldo_actual.get((mz, lt), saldo_gen)
        filas.append((mz, lt, float(deuda), cuotas, pagos, saldo))

    for row, (mz, lt, deuda, cuotas, pagos, saldo) in enumerate(sorted(filas, key=lambda f: (f[0], f[1])), start=3):
        _dat(ws.cell(row=row, column=1), mz, *_SEC_PREDIO_V[2:], align="center")
        _dat(ws.cell(row=row, column=2), lt, *_SEC_PREDIO_V[2:], align="center")
        _dat(ws.cell(row=row, column=3), nombres.get((mz, lt), ""), *_SEC_PREDIO_V[2:], align="left")
        _dat(ws.cell(row=row, column=4), deuda, "FFFBEB", pal[2], align="right")
        for i, cuota in enumerate(cuotas):
            _dat(ws.cell(row=row, column=5 + i), cuota, "FFFBEB", pal[3], align="right")
        for i, pago in enumerate(pagos):
            _dat(ws.cell(row=row, column=c_ledger_ini + i), pago, "EFF6FF", pal2[3], align="right")
        cell_s = ws.cell(row=row, column=len(cols))
        cell_s.value = saldo
        cell_s.fill = _fill(pal2[4])
        cell_s.font = Font(color=_argb(pal2[5]), bold=True, size=10)
        cell_s.alignment = Alignment(horizontal="right", vertical="center")
        cell_s.number_format = "#,##0.00"


def generar_vista(ruta: Path | None = None) -> Path:
    """Regenera vista_seguimiento_pueblo.xlsx completa desde el registro
    (una hoja por concepto + CONVENIO_HISTORIAL desde el snapshot de génesis).
    Nunca es la fuente — se puede borrar y recrear en cualquier momento."""
    ruta = ruta or VISTA_PATH
    df = _leer_eventos()
    nombres = _lookup_nombres()

    wb = Workbook()
    wb.remove(wb.active)
    for concepto in CONCEPTOS_ORDEN:
        ws = wb.create_sheet(concepto)
        df_c = df[df["CONCEPTO"].astype(str).str.strip().str.upper() == concepto]
        _escribir_hoja_vista(ws, concepto, df_c, nombres)

    _escribir_hoja_ajustes(wb.create_sheet("Ajustes"), df, nombres)

    if _MEDIDOR_SALDO_PATH_VISTA.exists():
        df_conv = df[df["CONCEPTO"].astype(str).str.strip().str.upper() == "CONVENIO"]
        _escribir_hoja_historial_convenio(wb.create_sheet("CONVENIO_HISTORIAL"), nombres, df_conv)

    _save_atomic(wb, ruta)
    log.info(f"vista_seguimiento_pueblo regenerada -> {ruta}")
    return ruta


def generar_vista_provisional(mes: str, snapshot_hash: str, snapshot: dict,
                              ruta: Path | None = None, estado_validacion: str = "VALIDADO",
                              alerta: str = "") -> Path:
    """Renderiza el resultado futuro de un snapshot sin mutar el ledger real."""
    ruta = ruta or VISTA_PROVISIONAL_PATH
    ledger_real = SEGUIMIENTO_PATH
    with tempfile.TemporaryDirectory(prefix="vista_provisional_") as tmp:
        ledger_temporal = Path(tmp) / "estado_cuenta_provisional.xlsx"
        if ledger_real.exists():
            shutil.copy2(ledger_real, ledger_temporal)
        globals()["SEGUIMIENTO_PATH"] = ledger_temporal
        try:
            reconciliar_objetivos_batch(
                mes, snapshot_hash, snapshot.get("objetivos", []), snapshot.get("cargos", []))
            generar_vista(ruta)
        finally:
            globals()["SEGUIMIENTO_PATH"] = ledger_real

    wb = load_workbook(ruta)
    if "PROVISIONAL" in wb.sheetnames:
        del wb["PROVISIONAL"]
    ws = wb.create_sheet("PROVISIONAL", 0)
    ws["A1"] = "VISTA PROVISIONAL - NO ES EL LEDGER OFICIAL"
    ws["A2"] = "MES"
    ws["B2"] = mes
    ws["A3"] = "SNAPSHOT"
    ws["B3"] = snapshot_hash
    ws["A4"] = "ESTADO"
    ws["B4"] = estado_validacion
    ws["A5"] = "ALERTA"
    ws["B5"] = alerta or "Sin alertas de validación"
    ws["A6"] = "ALCANCE"
    ws["B6"] = "Simula exactamente el batch que 7_cierre comprometerá si las fuentes no cambian."
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90
    _save_atomic(wb, ruta)
    return ruta


# ── PDF imprimible de la vista ────────────────────────────────────────────────

VISTA_PDF_PATH = _SHARED / "vista_seguimiento_pueblo.pdf"

_PDF_PAGE_W, _PDF_PAGE_H = 842, 595  # A4 horizontal (pts)
_PDF_M = 30
_PDF_ROW_H = 13.5
_PDF_HDR_H = 26


def exportar_vista_pdf(ruta_xlsx: Path | None = None, ruta_pdf: Path | None = None) -> Path:
    """PDF imprimible de la vista (A4 horizontal, cabeceras repetidas por página,
    saldo en rojo si debe / verde si está en 0). Para la mesa de cobro.
    Se dibuja directo con PyMuPDF — Excel COM exige una impresora instalada
    y esta máquina no tiene ninguna. Regenerable siempre, nunca es fuente."""
    import fitz  # lazy: que importar el repo no exija PyMuPDF

    ruta_xlsx = ruta_xlsx or VISTA_PATH
    ruta_pdf = ruta_pdf or VISTA_PDF_PATH

    azul, azul_bg = (26/255, 82/255, 118/255), (235/255, 245/255, 251/255)
    ambar_bg, ambar_fg = (254/255, 243/255, 199/255), (146/255, 64/255, 14/255)
    zebra, gris, negro = (243/255, 244/255, 246/255), (0.42, 0.45, 0.5), (0.12, 0.16, 0.22)
    rojo, verde = (0.7, 0.1, 0.1), (0.02, 0.37, 0.27)

    def _fmt(v):
        if pd.isna(v) or v == "—":
            return ""
        if isinstance(v, (int, float)):
            return f"{v:,.2f}"
        return str(v)

    hoy = datetime.now().strftime("%d/%m/%Y")
    xl = pd.ExcelFile(ruta_xlsx)
    doc = fitz.open()

    for hoja in xl.sheet_names:
        # "Ajustes" no va al PDF: es texto largo (MOTIVO) y este PDF es la hoja de
        # saldos para la mesa de cobro, no el expediente de correcciones.
        if hoja == "Ajustes":
            continue
        grupos = pd.read_excel(xl, hoja, header=None, nrows=2).iloc[0].ffill().fillna("")
        df = pd.read_excel(xl, hoja, header=1)
        cols = df.columns.tolist()

        fijas = {"MZ": 32, "LT": 40, "NOMBRE": 200}
        w_num = min(85.0, (_PDF_PAGE_W - 2*_PDF_M - 272) / max(len(cols) - 3, 1))
        ws = [fijas.get(c, w_num) for c in cols]

        page, y = None, 0.0

        def _nueva_pagina():
            nonlocal page, y
            page = doc.new_page(width=_PDF_PAGE_W, height=_PDF_PAGE_H)
            page.insert_text((_PDF_M, _PDF_M + 4), f"{hoja} — seguimiento del pueblo",
                             fontsize=11, fontname="helv", color=azul)
            page.insert_text((_PDF_PAGE_W - _PDF_M - 200, _PDF_M + 4),
                             f"generado {hoy} desde seguimiento_pueblo",
                             fontsize=7, fontname="helv", color=gris)
            y = _PDF_M + 14
            # cabecera: primero todos los fondos, después los textos
            # (si no, el fondo de la columna siguiente tapa el texto del grupo)
            x = _PDF_M
            for i in range(len(cols)):
                page.draw_rect(fitz.Rect(x, y, x + ws[i], y + _PDF_HDR_H),
                               fill=azul_bg if i < 3 else ambar_bg, color=(1, 1, 1), width=0.5)
                x += ws[i]
            x = _PDF_M
            for i, c in enumerate(cols):
                fg = azul if i < 3 else ambar_fg
                gr = str(grupos.iloc[i]) if i >= 3 else ""
                gr_prev = str(grupos.iloc[i-1]) if i > 3 else ""
                if gr and gr != "nan" and gr != gr_prev:
                    page.insert_text((x + 3, y + 10), gr[:110], fontsize=6.5,
                                     fontname="helv", color=fg)
                page.insert_text((x + 3, y + _PDF_HDR_H - 5), str(c), fontsize=7,
                                 fontname="hebo", color=fg)
                x += ws[i]
            y += _PDF_HDR_H

        _nueva_pagina()
        for n, (_, fila) in enumerate(df.iterrows()):
            if y + _PDF_ROW_H > _PDF_PAGE_H - _PDF_M:
                _nueva_pagina()
            if n % 2 == 1:
                page.draw_rect(fitz.Rect(_PDF_M, y, _PDF_M + sum(ws), y + _PDF_ROW_H),
                               fill=zebra, color=None)
            x = _PDF_M
            for i, c in enumerate(cols):
                v = _fmt(fila[c])
                es_saldo = str(c).upper().startswith("SALDO")
                font = "hebo" if es_saldo else "helv"
                color = negro
                if es_saldo and isinstance(fila[c], (int, float)) and not pd.isna(fila[c]):
                    color = rojo if fila[c] > 0 else verde
                if i >= 3:
                    tw = fitz.get_text_length(v, fontname=font, fontsize=7.5)
                    page.insert_text((x + ws[i] - 4 - tw, y + _PDF_ROW_H - 4), v,
                                     fontsize=7.5, fontname=font, color=color)
                else:
                    page.insert_text((x + 3, y + _PDF_ROW_H - 4), v[:44],
                                     fontsize=7.5, fontname=font, color=color)
                x += ws[i]
            y += _PDF_ROW_H

    for i, page in enumerate(doc, start=1):
        page.insert_text((_PDF_PAGE_W - _PDF_M - 60, _PDF_PAGE_H - 12),
                         f"Pag. {i} de {len(doc)}", fontsize=7, fontname="helv", color=gris)

    doc.save(str(ruta_pdf))
    doc.close()
    log.info(f"vista_seguimiento_pueblo PDF -> {ruta_pdf} ({len(xl.sheet_names)} hojas)")
    return ruta_pdf
