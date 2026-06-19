"""
shared/data_boletas_repo.py — Único writer de DATA_boletas.xlsx con audit centralizado

API PÚBLICA:
    read_padron()                    → DataFrame
    get_predio(mz, lt)               → dict | None
    get_predio_lookup()              → dict[(mz_norm, lt_norm), dict]  (bulk helper)
    apply_correction(...)            → dict {valor_antes, valor_despues, ts, skipped}

INVARIANTES:
    - Único escritor de DATA_boletas.xlsx (single writer).
    - Audit obligatorio en cada write → shared/data_boletas_audit.xlsx
    - Backup automático antes de escribir → 3_boletas/backup/DATA_boletas/
    - Idempotencia por (source, audit_ref, mz, lt, campo) en audit log.
    - Validación: rechaza si predio o campo no existen.

Contratos visuales: docs/diagrama_repo_pattern.html
                    docs/formato_data_boletas_audit.html
"""

import logging
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ── Rutas ─────────────────────────────────────────────────────────────────────
_SHARED = Path(__file__).parent
_BASE   = _SHARED.parent

DATA_BOLETAS_PATH = _BASE / "3_boletas" / "inputs" / "DATA_boletas.xlsx"
AUDIT_PATH        = _SHARED / "data_boletas_audit.xlsx"
BACKUP_DIR        = _BASE / "3_boletas" / "backup" / "DATA_boletas"

DATA_BOLETAS_SHEET = "Data"
AUDIT_SHEET        = "Audit"

# ── Paleta del audit log (contrato: formato_data_boletas_audit.html) ─────────
# (header_bg, header_fg, data_bg, data_fg)
_SEC_CUANDO = ("F3F4F6", "374151", "F9FAFB", "374151")
_SEC_PREDIO = ("EBF5FB", "1A5276", "F4FAFF", "1A5276")
_SEC_CAMBIO = ("FEF0E0", "7C2D12", "FFFAF5", "7C2D12")
_SEC_QUIEN  = ("F3E8FF", "5B21B6", "FAF5FF", "5B21B6")
_SEC_MOTIVO = ("FEF9E7", "7D6608", "FFFBEB", "7D6608")

_AUDIT_COLS = [
    ("TIMESTAMP",     _SEC_CUANDO, 20, "center", None),
    ("MZ",            _SEC_PREDIO, 6,  "center", None),
    ("LT",            _SEC_PREDIO, 7,  "center", None),
    ("CAMPO",         _SEC_CAMBIO, 26, "center", None),
    ("VALOR_ANTES",   _SEC_CAMBIO, 14, "right",  '#,##0.00'),
    ("VALOR_DESPUES", _SEC_CAMBIO, 14, "right",  '#,##0.00'),
    ("SOURCE",        _SEC_QUIEN,  20, "center", None),
    ("AUDIT_REF",     _SEC_QUIEN,  26, "left",   None),
    ("MOTIVO",        _SEC_MOTIVO, 50, "left",   None),
]

_AUDIT_SECCIONES = [
    ("Cuándo",        "TIMESTAMP",   "TIMESTAMP"),
    ("Predio",        "MZ",          "LT"),
    ("Qué cambió",    "CAMPO",       "VALOR_DESPUES"),
    ("Quién lo hizo", "SOURCE",      "AUDIT_REF"),
    ("Por qué",       "MOTIVO",      "MOTIVO"),
]

# ── Helpers internos ─────────────────────────────────────────────────────────

def _norm(v) -> str:
    """Normaliza MZ/LT para matching: upper + strip + sin espacios."""
    if v is None:
        return ""
    return str(v).strip().upper().replace(" ", "")


def _clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("nan", "None", "NaT") else s


def _try_float(v):
    if v is None:
        return None
    s = str(v).strip().replace(",", "")
    if s in ("", "nan", "None", "NaT"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _argb(hex6: str) -> str:
    return "FF" + hex6.lstrip("#")


def _fill(hex6):
    return PatternFill("solid", fgColor=_argb(hex6))


def _hdr(cell, bg, fg, texto):
    cell.value = texto
    cell.fill  = _fill(bg)
    cell.font  = Font(color=_argb(fg), bold=True, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _dat(cell, valor, bg, fg, align="center", fmt=None):
    cell.value = valor
    cell.fill  = _fill(bg)
    cell.font  = Font(color=_argb(fg), size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt and isinstance(valor, (int, float)):
        cell.number_format = fmt

# ── API pública: lectura ─────────────────────────────────────────────────────

def read_padron() -> pd.DataFrame:
    """Lee DATA_boletas como DataFrame (dtype=str). Raises FileNotFoundError."""
    if not DATA_BOLETAS_PATH.exists():
        raise FileNotFoundError(f"DATA_boletas no encontrado: {DATA_BOLETAS_PATH}")
    return pd.read_excel(DATA_BOLETAS_PATH, sheet_name=DATA_BOLETAS_SHEET, dtype=str).fillna("")


def get_predio(mz, lt) -> dict | None:
    """Retorna fila del predio (MZ, LT) o None. Matching case-insensitive sin espacios."""
    mz_n, lt_n = _norm(mz), _norm(lt)
    if not mz_n or not lt_n:
        return None
    df = read_padron()
    for _, row in df.iterrows():
        if _norm(row.get("MZ", "")) == mz_n and _norm(row.get("LT", "")) == lt_n:
            return row.to_dict()
    return None


def get_predio_lookup() -> dict:
    """
    Helper para operaciones bulk: retorna {(mz_norm, lt_norm): row_dict}.
    Útil cuando hay que verificar/leer muchos predios de una sola pasada.
    """
    df = read_padron()
    lookup = {}
    for _, row in df.iterrows():
        mz_n = _norm(row.get("MZ", ""))
        lt_n = _norm(row.get("LT", ""))
        if mz_n and lt_n:
            lookup[(mz_n, lt_n)] = row.to_dict()
    return lookup

# ── API pública: escritura ───────────────────────────────────────────────────

def apply_correction(mz, lt, campo, valor, *, source: str, audit_ref: str, motivo: str = "") -> dict:
    """
    Mutación atómica de DATA_boletas:
        1. Valida inputs (predio + campo existen; source + audit_ref no vacíos)
        2. Idempotencia: si (source, audit_ref, mz, lt, campo) ya en audit → skip
        3. Backup completo de DATA_boletas con timestamp
        4. Lee VALOR_ANTES en celda destino
        5. Escribe nuevo valor en (MZ, LT)[CAMPO]
        6. Append fila al audit log

    Retorna {valor_antes, valor_despues, ts, skipped}.
    Raises ValueError si inputs inválidos.
    """
    if not source:
        raise ValueError("apply_correction: source no puede ser vacío")
    if not audit_ref:
        raise ValueError("apply_correction: audit_ref no puede ser vacío")

    mz_n, lt_n = _norm(mz), _norm(lt)
    if not mz_n or not lt_n:
        raise ValueError(f"apply_correction: MZ/LT inválidos ({mz!r}, {lt!r})")

    campo = str(campo).strip()
    if not campo:
        raise ValueError("apply_correction: campo no puede ser vacío")

    # Idempotencia
    if _audit_already_applied(source, audit_ref, mz_n, lt_n, campo):
        log.info(f"apply_correction: skip (idempotente) — {source}/{audit_ref} mz={mz_n} lt={lt_n} campo={campo}")
        return {"valor_antes": None, "valor_despues": None, "ts": None, "skipped": True}

    if not DATA_BOLETAS_PATH.exists():
        raise FileNotFoundError(f"DATA_boletas no encontrado: {DATA_BOLETAS_PATH}")

    wb = load_workbook(DATA_BOLETAS_PATH)
    ws = wb[DATA_BOLETAS_SHEET] if DATA_BOLETAS_SHEET in wb.sheetnames else wb.active

    # Mapa nombre_columna → índice (1-based) leyendo row 1
    col_idx = {}
    for c, cell in enumerate(ws[1], start=1):
        nombre = str(cell.value).strip() if cell.value is not None else ""
        if nombre:
            col_idx[nombre] = c

    if campo not in col_idx:
        raise ValueError(
            f"apply_correction: campo {campo!r} no existe en DATA_boletas. "
            f"Columnas válidas: {list(col_idx.keys())}"
        )

    mz_col = col_idx.get("MZ")
    lt_col = col_idx.get("LT")
    if mz_col is None or lt_col is None:
        raise ValueError("apply_correction: DATA_boletas no tiene columnas MZ y LT")

    # Buscar fila destino
    target_row = None
    for r in range(2, ws.max_row + 1):
        if _norm(ws.cell(row=r, column=mz_col).value) == mz_n and \
           _norm(ws.cell(row=r, column=lt_col).value) == lt_n:
            target_row = r
            break

    if target_row is None:
        raise ValueError(f"apply_correction: predio (MZ={mz!r}, LT={lt!r}) no existe en DATA_boletas")

    # Backup ANTES del write
    backup_path = _backup_data_boletas()
    log.info(f"apply_correction: backup -> {backup_path.name}")

    # Read-then-write
    campo_cell = ws.cell(row=target_row, column=col_idx[campo])
    valor_antes_raw = campo_cell.value

    valor_num = _try_float(valor)
    nuevo_valor = valor_num if valor_num is not None else valor

    campo_cell.value = nuevo_valor
    wb.save(DATA_BOLETAS_PATH)

    # Audit append
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _audit_append({
        "TIMESTAMP":     ts,
        "MZ":            mz_n,
        "LT":            lt_n,
        "CAMPO":         campo,
        "VALOR_ANTES":   _try_float(valor_antes_raw) if _try_float(valor_antes_raw) is not None else _clean(valor_antes_raw),
        "VALOR_DESPUES": nuevo_valor,
        "SOURCE":        source,
        "AUDIT_REF":     audit_ref,
        "MOTIVO":        motivo,
    })

    return {
        "valor_antes":   valor_antes_raw,
        "valor_despues": nuevo_valor,
        "ts":            ts,
        "skipped":       False,
    }

# ── Internals: backup ────────────────────────────────────────────────────────

def _backup_data_boletas() -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    dest = BACKUP_DIR / f"DATA_boletas_{ts}.xlsx"
    shutil.copy2(DATA_BOLETAS_PATH, dest)
    return dest

# ── Internals: audit log ─────────────────────────────────────────────────────

def _audit_already_applied(source: str, audit_ref: str, mz: str, lt: str, campo: str) -> bool:
    """True si (source, audit_ref, mz, lt, campo) ya está en el audit log."""
    if not AUDIT_PATH.exists():
        return False
    try:
        df = pd.read_excel(AUDIT_PATH, sheet_name=AUDIT_SHEET, header=1, dtype=str).fillna("")
    except Exception:
        return False
    if df.empty:
        return False
    mask = (
        (df["SOURCE"].str.strip()    == source) &
        (df["AUDIT_REF"].str.strip() == audit_ref) &
        (df["MZ"].str.strip()        == mz) &
        (df["LT"].str.strip()        == lt) &
        (df["CAMPO"].str.strip()     == campo)
    )
    return bool(mask.any())


def _audit_append(fila: dict) -> None:
    """Append una fila al audit log; crea el archivo con headers si no existe."""
    if AUDIT_PATH.exists():
        wb = load_workbook(AUDIT_PATH)
        ws = wb[AUDIT_SHEET] if AUDIT_SHEET in wb.sheetnames else wb.active
        next_row = max(ws.max_row + 1, 3)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = AUDIT_SHEET
        _write_audit_headers(ws)
        next_row = 3

    for ci, (nombre, sec, _, align, fmt) in enumerate(_AUDIT_COLS, start=1):
        val = fila.get(nombre, "")
        if nombre in ("VALOR_ANTES", "VALOR_DESPUES"):
            num = _try_float(val)
            val = num if num is not None else _clean(val)
        _dat(ws.cell(row=next_row, column=ci), val, sec[2], sec[3], align=align, fmt=fmt)

    wb.save(AUDIT_PATH)


def _write_audit_headers(ws) -> None:
    """Row 1 = secciones, Row 2 = columnas, freeze A3."""
    col_idx = {c[0]: i + 1 for i, c in enumerate(_AUDIT_COLS)}

    for label, start, end in _AUDIT_SECCIONES:
        c1 = col_idx[start]
        c2 = col_idx[end]
        sec = _AUDIT_COLS[c1 - 1][1]
        if c1 != c2:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        _hdr(ws.cell(row=1, column=c1), sec[0], sec[1], label)
    ws.row_dimensions[1].height = 18

    for i, (nombre, sec, ancho, _, _f) in enumerate(_AUDIT_COLS, start=1):
        _hdr(ws.cell(row=2, column=i), sec[0], sec[1], nombre)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"
