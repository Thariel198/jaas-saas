"""Create an operational map for shared/abonos_rezagados.xlsx."""
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


ROOT = Path(__file__).resolve().parent.parent
PATH = ROOT / "shared" / "abonos_rezagados.xlsx"


def key(row):
    def clean(value):
        text = str(value or "").strip()
        return "" if text.upper() in {"NAN", "NONE", "NAT"} else text

    return (
        clean(row.get("MZ")).upper().replace(" ", ""),
        clean(row.get("LT")).upper().replace(" ", ""),
        round(float(row.get("MONTO") or 0), 2),
        clean(row.get("MES_CICLO"))[:7],
        clean(row.get("MES_ANO_APLICA"))[:7],
    )


def classify(row):
    estado = str(row.get("ESTADO") or "").strip().upper()
    motivo = str(row.get("MOTIVO_ESTADO") or "").strip()
    if estado == "CONFIRMADO":
        return estado, "DRY_RUN_DIRIGIDO", "PENDIENTE_APROBACION", motivo
    if estado == "CONFIRMADO_SIN_APLICACION":
        return estado, "NO_APLICAR", "SIN_EVENTO_MONETARIO", motivo
    if estado == "DESCARTADO":
        return estado, "NO_APLICAR", "CONSERVAR_HISTORIA", motivo
    raise ValueError(f"ESTADO sin contrato: {estado!r}")


def main():
    wb = load_workbook(PATH)
    raw = wb[wb.sheetnames[0]]
    raw.title = "Abonos_Raw"
    headers = {str(raw.cell(2, c).value).strip(): c for c in range(1, raw.max_column + 1)}
    source_rows = []
    for r in range(3, raw.max_row + 1):
        row = {name: raw.cell(r, col).value for name, col in headers.items()}
        k = key(row)
        situation, action, ledger, reason = classify(row)
        source_rows.append({
            "ID_ABONO": row.get("ID_ABONO") or f"{k[0]}-{k[1]}|{k[2]:.2f}|{k[3]}|{k[4]}",
            "MZ": k[0], "LT": k[1], "MONTO": k[2], "BALDE_ORIGINAL": row.get("BALDE", ""),
            "CANAL": row.get("CANAL", ""), "MES_CICLO": k[3], "MES_ANO_APLICA": k[4],
            "RETENIDO_POR": row.get("RETENIDO_POR", ""), "EVIDENCIA": row.get("EVIDENCIA", ""),
            "MOTIVO_ORIGINAL": row.get("MOTIVO", ""), "SITUACION": situation,
            "PROXIMA_ACCION": action, "ESTADO_LEDGER": ledger, "MOTIVO_ESTADO": reason,
            "CONCEPTO_DESTINO": row.get("CONCEPTO_DESTINO", ""),
        })

    for name in ("Mapa_Abonos", "Categorias"):
        if name in wb.sheetnames:
            del wb[name]
    mapa = wb.create_sheet("Mapa_Abonos")
    cols = list(source_rows[0])
    mapa.append(cols)
    for row in source_rows:
        mapa.append([row[c] for c in cols])
    categorias = wb.create_sheet("Categorias")
    categorias.append(["SITUACION", "SIGNIFICADO", "PROXIMA_ACCION", "ESTADO_LEDGER", "REGLA"])
    categorias.append(["CONFIRMADO", "Fila aprobada", "DRY_RUN_DIRIGIDO", "PENDIENTE_APROBACION", "Requiere monto y destino validos."])
    categorias.append(["DESCARTADO", "Historia conservada", "NO_APLICAR", "CONSERVAR_HISTORIA", "Nunca entra a calculos o reportes operativos."])
    categorias.append(["CONFIRMADO_SIN_APLICACION", "Constancia auditiva", "NO_APLICAR", "SIN_EVENTO_MONETARIO", "No genera movimiento de dinero."])

    for ws in (mapa, categorias):
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="174A63")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in ws.columns:
            ws.column_dimensions[column[0].column_letter].width = min(max(max(len(str(c.value or "")) for c in column) + 2, 12), 42)
    mapa.auto_filter.ref = mapa.dimensions
    wb.save(PATH)
    print(f"Mapa creado: {len(source_rows)} filas")


if __name__ == "__main__":
    main()
