import argparse
import json
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parent
JSON_PATH = ROOT / "anulaciones_ledger.json"
XLSX_PATH = ROOT / "anulaciones_ledger.xlsx"

COLUMNAS = [
    "CAUSA_ID", "ESTADO_ANULACION", "MES_ANULACION", "ANULADO_EN",
    "AUTORIZACION", "MOTIVO_ANULACION", "FILA_EXCEL", "AUDIT_REF",
    "MZ", "LT", "CONCEPTO", "MES", "TIPO_EVENTO", "MONTO", "SOURCE",
    "CLASE",
]
ANCHOS = [48, 18, 14, 22, 45, 95, 12, 70, 8, 8, 15, 12, 15, 14, 24, 24]


def _filas(data: dict) -> list[list]:
    filas = []
    referencias = set()
    for anulacion in data.get("anulaciones", []):
        for evento in anulacion.get("eventos", []):
            referencia = str(evento.get("audit_ref", "")).strip()
            if not referencia:
                raise ValueError(f"Evento sin audit_ref en {anulacion.get('id')}")
            if referencia in referencias:
                raise ValueError(f"AUDIT_REF duplicado: {referencia}")
            referencias.add(referencia)
            filas.append([
                anulacion.get("id"),
                anulacion.get("estado", "ACTIVA"),
                anulacion.get("mes"),
                anulacion.get("anulado_en"),
                anulacion.get("autorizacion"),
                anulacion.get("motivo"),
                evento.get("fila_excel"),
                referencia,
                evento.get("mz"),
                evento.get("lt"),
                evento.get("concepto"),
                evento.get("mes"),
                evento.get("tipo_evento"),
                evento.get("monto"),
                evento.get("source"),
                evento.get("clase"),
            ])
    return filas


def exportar(json_path: Path = JSON_PATH, xlsx_path: Path = XLSX_PATH) -> Path:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    if data.get("schema") != 1:
        raise ValueError(f"Schema no soportado: {data.get('schema')}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Eventos"
    ws.append(COLUMNAS)
    for fila in _filas(data):
        ws.append(fila)

    encabezado = PatternFill("solid", fgColor="1F4E78")
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = encabezado
        celda.alignment = Alignment(horizontal="center", vertical="center")

    for fila in ws.iter_rows(min_row=2):
        for celda in fila:
            celda.alignment = Alignment(vertical="top")
        fila[4].alignment = Alignment(vertical="top", wrap_text=True)
        fila[5].alignment = Alignment(vertical="top", wrap_text=True)
        fila[13].number_format = "#,##0.00;-#,##0.00;0.00"

    for columna, ancho in zip(ws.iter_cols(min_row=1, max_row=1), ANCHOS):
        ws.column_dimensions[columna[0].column_letter].width = ancho
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{ws.max_row}"

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    temporal = xlsx_path.with_suffix(".tmp.xlsx")
    wb.save(temporal)
    os.replace(temporal, xlsx_path)
    return xlsx_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Exporta anulaciones_ledger.json al Excel de auditoria."
    )
    parser.add_argument("--json", type=Path, default=JSON_PATH)
    parser.add_argument("--salida", type=Path, default=XLSX_PATH)
    args = parser.parse_args()
    salida = exportar(args.json, args.salida)
    print(f"OK: {salida}")


if __name__ == "__main__":
    main()
