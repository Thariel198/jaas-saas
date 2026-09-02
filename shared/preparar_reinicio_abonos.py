"""Prepara la fase 1 del reinicio de abonos; nunca escribe en el ledger."""
import argparse
import hashlib
import json
import shutil
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

import abonos_rezagados_repo as abonos_repo
import seguimiento_repo as seguimiento


ROOT = Path(__file__).resolve().parent.parent
SOURCE = ROOT / "shared" / "abonos_rezagados.xlsx"
LEDGER = ROOT / "shared" / "seguimiento_pueblo.xlsx"
ANULACIONES = ROOT / "shared" / "anulaciones_ledger.json"
REPORT = ROOT / "shared" / "abono_rezagado" / "reporte_reinicio_abonos_2026-08-30.xlsx"
BACKUPS = ROOT / "shared" / "backups_abono_rezagado"
FECHA_ESTADO = "2026-08-30"
MOTIVO_DESCARTE = "reinicio controlado; registro historico pendiente de revalidacion"
MOTIVO_CORRECCION = "correccion historica de abono de julio omitido en seguimiento_pueblo"

APLICACIONES = (
    ("D", "16", "ACUERDOS", 50.0),
    ("D", "16", "CONVENIO", 25.0),
    ("D1", "6", "MULTA", 13.0),
    ("L", "4", "MULTA", 3.0),
    ("L", "4", "ACUERDOS", 25.0),
    ("F", "9", "CONVENIO", 25.0),
)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _norm(value) -> str:
    text = str(value or "").strip().upper().replace(" ", "")
    return text[:-2] if text.endswith(".0") else text


def _id_historico(row: dict, ordinal: int) -> str:
    canon = "|".join(str(row.get(c, "") or "").strip() for c in (
        "MZ", "LT", "MONTO", "MES_CICLO", "MES_ANO_APLICA",
    ))
    return f"hist-{hashlib.sha256(f'{ordinal}|{canon}'.encode()).hexdigest()[:12]}"


def _copiar_estilo(origen, destino) -> None:
    if origen.has_style:
        destino._style = copy(origen._style)
    destino.number_format = origen.number_format
    destino.alignment = copy(origen.alignment)


def _escribir_tabla(ws, filas: list[dict]) -> None:
    if not filas:
        return
    columnas = list(filas[0])
    ws.append(columnas)
    for fila in filas:
        ws.append([fila.get(col) for col in columnas])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="174A63")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in ws.columns:
        ancho = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[column[0].column_letter].width = min(max(ancho, 12), 48)


def _clasificacion_previa(wb) -> dict[tuple[str, str, float, str, str], dict]:
    if "Mapa_Abonos" not in wb.sheetnames:
        return {}
    ws = wb["Mapa_Abonos"]
    headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
    salida = {}
    for r in range(2, ws.max_row + 1):
        row = {name: ws.cell(r, col).value for name, col in headers.items()}
        try:
            key = (_norm(row.get("MZ")), _norm(row.get("LT")), round(float(row.get("MONTO") or 0), 2),
                   str(row.get("MES_CICLO") or "")[:7], str(row.get("MES_ANO_APLICA") or "")[:7])
        except (TypeError, ValueError):
            continue
        salida[key] = row
    return salida


def _cruce_ledger(filas: list[dict], clasificacion: dict) -> list[dict]:
    fisicos = pd.read_excel(LEDGER, sheet_name="Eventos", header=1)
    activos = seguimiento._leer_eventos()
    anulaciones = json.loads(ANULACIONES.read_text(encoding="utf-8"))
    refs_anuladas = set()
    for causa in anulaciones.get("anulaciones", []):
        if str(causa.get("estado", "ACTIVA")).upper() != "ACTIVA":
            continue
        refs_anuladas.update(str(ref).strip() for ref in causa.get("audit_refs", []))
        refs_anuladas.update(
            str(evento.get("audit_ref", "")).strip()
            for evento in causa.get("eventos", [])
            if evento.get("audit_ref")
        )
    salida = []
    for ordinal, row in enumerate(filas, start=1):
        mz, lt = _norm(row["MZ"]), _norm(row["LT"])
        monto = round(float(row["MONTO"]), 2)
        key = (mz, lt, monto, str(row.get("MES_CICLO") or "")[:7],
               str(row.get("MES_ANO_APLICA") or "")[:7])
        previo = clasificacion.get(key, {})
        f = fisicos[(fisicos["MZ"].map(_norm) == mz) & (fisicos["LT"].map(_norm) == lt)]
        a = activos[(activos["MZ"].map(_norm) == mz) & (activos["LT"].map(_norm) == lt)]
        refs_fisicas = [str(v).strip() for v in f.get("AUDIT_REF", pd.Series(dtype=str)).dropna()]
        salida.append({
            "ID_ABONO": _id_historico(row, ordinal), "MZ": mz, "LT": lt, "MONTO_ORIGINAL": monto,
            "SITUACION_ANTERIOR": previo.get("SITUACION", "SIN_MAPA"),
            "ESTADO_NUEVO": "DESCARTADO", "MOTIVO_ESTADO": MOTIVO_DESCARTE,
            "EVENTOS_FISICOS_PREDIO": len(f), "EVENTOS_ACTIVOS_PREDIO": len(a),
            "REFS_ANULADAS_PREDIO": sum(ref in refs_anuladas for ref in refs_fisicas),
            "NOTA": "Clasificacion de fuente; no modifica ni interpreta eventos del ledger.",
        })
    return salida


def _dry_run() -> list[dict]:
    salida = []
    for mz, lt, concepto, monto in APLICACIONES:
        saldo = round(float(seguimiento.get_saldo(mz, lt, concepto, "2026-08")), 2)
        salida.append({
            "MZ": mz, "LT": lt, "CONCEPTO": concepto, "MONTO_PROPUESTO": monto,
            "SALDO_ACTIVO_ANTES": saldo, "SALDO_PROPUESTO_DESPUES": round(saldo - monto, 2),
            "RESULTADO": "OK" if saldo + 0.005 >= monto else "BLOQUEADO_EXCESO",
        })
    return salida


def preparar() -> tuple[Path, Path]:
    ledger_hash = _sha256(LEDGER)
    wb = load_workbook(SOURCE)
    raw = wb["Abonos_Raw"]
    headers = {str(raw.cell(2, c).value or "").strip(): c for c in range(1, raw.max_column + 1)}
    filas = [
        {name: raw.cell(r, col).value for name, col in headers.items()}
        for r in range(3, raw.max_row + 1)
        if raw.cell(r, headers["MZ"]).value is not None
    ]
    if len(filas) != 43 or round(sum(float(row["MONTO"]) for row in filas), 2) != 2281.0:
        raise RuntimeError("La fuente ya no coincide con el punto de partida: 43 filas y S/2,281")

    previo = _clasificacion_previa(wb)
    cruce = _cruce_ledger(filas, previo)
    backup = BACKUPS / f"reinicio_fuente_{datetime.now():%Y%m%d_%H%M%S}"
    backup.mkdir(parents=True)
    shutil.copy2(SOURCE, backup / SOURCE.name)

    control = ("ID_ABONO", "ESTADO", "MODO_APLICACION", "CONCEPTO_DESTINO",
               "MOTIVO_ESTADO", "FECHA_ESTADO", "AUTORIZADO_POR")
    inicio_control = raw.max_column + 1
    raw.cell(1, inicio_control, "Control operativo")
    raw.merge_cells(start_row=1, start_column=inicio_control, end_row=1,
                    end_column=inicio_control + len(control) - 1)
    for offset, nombre in enumerate(control):
        col = inicio_control + offset
        raw.cell(2, col, nombre)
        _copiar_estilo(raw.cell(2, inicio_control - 1), raw.cell(2, col))

    for ordinal, r in enumerate(range(3, 46), start=1):
        valores = (_id_historico(filas[ordinal - 1], ordinal), "DESCARTADO", "NO_APLICA", "",
                   MOTIVO_DESCARTE, FECHA_ESTADO, "usuario")
        for offset, valor in enumerate(valores):
            raw.cell(r, inicio_control + offset, valor)
            _copiar_estilo(raw.cell(r, inicio_control - 1), raw.cell(r, inicio_control + offset))

    originales = {(_norm(row["MZ"]), _norm(row["LT"])): row for row in filas}
    nuevos = []
    for mz, lt, concepto, monto in APLICACIONES:
        base = dict(originales[(mz, lt)])
        base.update({
            "MONTO": monto,
            "MOTIVO": MOTIVO_CORRECCION,
            "EVIDENCIA": (f"planilla_cobrado_julio.xlsx + arrastre_consolidado_2026-07.xlsx; "
                          f"diferencial confirmado para {concepto}"),
            "RESPALDO": "documentado",
            "ID_ABONO": f"corr-2026-07-{mz}-{lt}-{concepto.lower()}",
            "ESTADO": "CONFIRMADO", "MODO_APLICACION": "DIRIGIDO",
            "CONCEPTO_DESTINO": concepto, "MOTIVO_ESTADO": MOTIVO_CORRECCION,
            "FECHA_ESTADO": FECHA_ESTADO, "AUTORIZADO_POR": "usuario",
        })
        nuevos.append(base)

    s5 = dict(originales[("S", "5")])
    s5.update({
        "MONTO": 0.0, "MOTIVO": "constancia: el abono se consumio en MES_ANTERIOR y CORTE",
        "ID_ABONO": "constancia-2026-07-S-5", "ESTADO": "CONFIRMADO_SIN_APLICACION",
        "MODO_APLICACION": "NO_APLICA", "CONCEPTO_DESTINO": "",
        "MOTIVO_ESTADO": "abono original S/71 aplicado a MES_ANTERIOR 46 + CORTE 25; diferencial pueblo S/0",
        "FECHA_ESTADO": FECHA_ESTADO, "AUTORIZADO_POR": "usuario",
    })
    nuevos.append(s5)

    columnas_raw = [str(raw.cell(2, c).value or "").strip() for c in range(1, raw.max_column + 1)]
    fila_estilo = 45
    for nuevo in nuevos:
        r = raw.max_row + 1
        for c, nombre in enumerate(columnas_raw, start=1):
            raw.cell(r, c, nuevo.get(nombre, ""))
            _copiar_estilo(raw.cell(fila_estilo, c), raw.cell(r, c))

    for name in ("Mapa_Abonos", "Categorias"):
        if name in wb.sheetnames:
            del wb[name]
    mapa = wb.create_sheet("Mapa_Abonos")
    _escribir_tabla(mapa, cruce + [{
        "ID_ABONO": row["ID_ABONO"], "MZ": _norm(row["MZ"]), "LT": _norm(row["LT"]),
        "MONTO_ORIGINAL": row["MONTO"], "SITUACION_ANTERIOR": "NUEVO",
        "ESTADO_NUEVO": row["ESTADO"], "MOTIVO_ESTADO": row["MOTIVO_ESTADO"],
        "EVENTOS_FISICOS_PREDIO": "", "EVENTOS_ACTIVOS_PREDIO": "",
        "REFS_ANULADAS_PREDIO": "", "NOTA": row.get("CONCEPTO_DESTINO") or "S/0",
    } for row in nuevos])
    categorias = wb.create_sheet("Categorias")
    _escribir_tabla(categorias, [
        {"ESTADO": "CONFIRMADO", "PROCESABLE": "SI", "REGLA": "Monto positivo y validacion completa."},
        {"ESTADO": "DESCARTADO", "PROCESABLE": "NO", "REGLA": "Conserva la evidencia historica."},
        {"ESTADO": "CONFIRMADO_SIN_APLICACION", "PROCESABLE": "NO", "REGLA": "Constancia sin evento monetario."},
    ])
    wb.save(SOURCE)

    activos = abonos_repo.leer_abonos(SOURCE)
    if len(activos) != 6 or round(pd.to_numeric(activos["MONTO"]).sum(), 2) != 141.0:
        raise RuntimeError("Validacion posterior fallo: se esperaban 6 activos por S/141")
    dry_run = _dry_run()
    if any(row["RESULTADO"] != "OK" for row in dry_run):
        raise RuntimeError("Dry-run bloqueado: una aplicacion supera el saldo activo")

    REPORT.parent.mkdir(parents=True, exist_ok=True)
    resumen = pd.DataFrame([{
        "FILAS_HISTORICAS_DESCARTADAS": 43, "MONTO_HISTORICO_PRESERVADO": 2281.0,
        "APLICACIONES_CONFIRMADAS": 6, "MONTO_CONFIRMADO": 141.0,
        "CONSTANCIAS_SIN_APLICACION": 1, "LEDGER_SHA256_ANTES": ledger_hash,
        "LEDGER_SHA256_DESPUES": _sha256(LEDGER), "LEDGER_CAMBIO": "NO",
    }])
    with pd.ExcelWriter(REPORT, engine="openpyxl") as writer:
        resumen.to_excel(writer, sheet_name="Resumen", index=False)
        pd.DataFrame(cruce).to_excel(writer, sheet_name="Inventario_43", index=False)
        activos.to_excel(writer, sheet_name="Nuevos_Activos", index=False)
        pd.DataFrame(dry_run).to_excel(writer, sheet_name="Dry_Run_Ledger", index=False)
    if _sha256(LEDGER) != ledger_hash:
        raise RuntimeError("El ledger cambio durante la fase 1")
    return backup, REPORT


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--aplicar", action="store_true")
    args = parser.parse_args()
    if not args.aplicar:
        sys.exit("Use --aplicar para modificar solo abonos_rezagados.xlsx; el ledger nunca se escribe.")
    backup_path, report_path = preparar()
    print(f"BACKUP={backup_path}")
    print(f"REPORTE={report_path}")
