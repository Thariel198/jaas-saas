"""
shared/utils_lote.py — Primitivos para correcciones de lote

Fuente: 5_cobranza/inputs/correcciones_lote.xlsx
        (gestionado por 5_cobranza, consumido por cualquier módulo que
         necesite remapear MZ+LT antes de buscar en planilla o DATA_boletas)
"""
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_BASE = Path(__file__).parent.parent
CORR_LOTE_PATH = _BASE / "5_cobranza" / "inputs" / "correcciones_lote.xlsx"

# Paleta verde — corrección editable por el operador (igual a GH_DC_CORR/TD_DC_CORR
# en 5_cobranza/main.py; ambos llamadores deben verse idénticos)
_HEADER_BG, _HEADER_FG = "ECFDF5", "065F46"
_DATA_BG = "ECFDF5"
_COLS = ["MZ_ORIGEN", "LT_ORIGEN", "MZ_DESTINO", "LT_DESTINO", "MOTIVO", "CICLO", "FECHA"]
_ANCHOS = [8, 8, 10, 10, 40, 8, 18]


def _borde() -> Border:
    s = Side(style="thin", color="FFFFFF")
    return Border(left=s, right=s, top=s, bottom=s)


def _norm(v: str) -> str:
    return str(v).strip().upper().replace(" ", "")


def leer_correcciones_lote() -> dict:
    """
    Lee correcciones_lote.xlsx de 5_cobranza.
    Retorna {(mz_orig, lt_orig): (mz_dest, lt_dest)}.
    Retorna {} si el archivo no existe.
    """
    if not CORR_LOTE_PATH.exists():
        return {}
    df = pd.read_excel(CORR_LOTE_PATH, header=0, dtype=str)
    df.columns = [str(c).strip().upper() for c in df.columns]
    corr = {}
    for _, row in df.iterrows():
        mo = _norm(row.get("MZ_ORIGEN", ""))
        lo = _norm(row.get("LT_ORIGEN", ""))
        md = _norm(row.get("MZ_DESTINO", ""))
        ld = _norm(row.get("LT_DESTINO", ""))
        if mo and lo and md and ld:
            corr[(mo, lo)] = (md, ld)
    return corr


def escribir_correcciones_lote(ruta: Path, filas: list[dict]) -> None:
    """
    Escribe correcciones_lote.xlsx con el formato del contrato (verde, ver
    5_cobranza/docs). filas=[] escribe solo el header — usado por 7_cierre
    para resetear el archivo al cerrar el período.
    """
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "correcciones_lote"

    for ci, col in enumerate(_COLS, 1):
        c = ws.cell(1, ci, col)
        c.font = Font(name="Arial", bold=True, size=9, color=_HEADER_FG)
        c.fill = PatternFill("solid", start_color=_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _borde()
        ws.column_dimensions[get_column_letter(ci)].width = _ANCHOS[ci - 1]
    ws.row_dimensions[1].height = 20

    for ri, row in enumerate(filas, 2):
        for ci, col in enumerate(_COLS, 1):
            c = ws.cell(ri, ci, row.get(col, ""))
            c.font = Font(name="Consolas", size=9)
            c.alignment = Alignment(horizontal="center" if ci <= 4 or ci == 6 else "left",
                                     vertical="center")
            c.border = _borde()
            if ci <= 4:
                c.fill = PatternFill("solid", start_color=_DATA_BG)
        ws.row_dimensions[ri].height = 16

    wb.save(ruta)
