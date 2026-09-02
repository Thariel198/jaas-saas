from pathlib import Path
import sys

import pandas as pd
import pytest


ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(ROOT / "shared"))
import seguimiento_repo as repo  # noqa: E402


def _evento(mz, lt, concepto, mes, tipo, monto, orden, *, source="2_planilla", clase="GENESIS"):
    return {
        "MZ": mz,
        "LT": lt,
        "CONCEPTO": concepto,
        "MES": mes,
        "TIPO_EVENTO": tipo,
        "CARGO": monto if tipo == "CARGO" else 0,
        "PAGO": monto if tipo == "PAGO" else 0,
        "AJUSTE": monto if tipo == "AJUSTE" else 0,
        "SALDO": 0,
        "SOURCE": source,
        "AUDIT_REF": f"ref-{orden}",
        "TIMESTAMP": f"{mes}-01 00:00:{orden:02d}",
        "CLASE": clase,
        "MOTIVO": "",
    }


def _fila(proyeccion, hoja, mes="2026-08"):
    return proyeccion[hoja].loc[proyeccion[hoja]["MES"] == mes].iloc[0]


def test_i9_separa_anterior_actual_y_mantenimiento():
    eventos = [
        _evento("I", "9", "AGUA", "2026-08", "CARGO", 8, 1, source="saldo_inicial"),
        _evento("I", "9", "AGUA", "2026-08", "CARGO", 5, 2),
        _evento("I", "9", "MANTENIMIENTO", "2026-08", "CARGO", 3, 3),
        _evento("I", "9", "AGUA", "2026-08", "PAGO", 13, 4,
                source="5_cobranza", clase="COBRANZA"),
        _evento("I", "9", "MANTENIMIENTO", "2026-08", "PAGO", 3, 5,
                source="5_cobranza", clase="COBRANZA"),
    ]

    proyeccion = repo._proyectar_consumo_temporal(pd.DataFrame(eventos))

    assert _fila(proyeccion, "MES_ANTERIOR")[["DEUDA", "PAGO", "SALDO"]].tolist() == [8, 8, 0]
    assert _fila(proyeccion, "MES_ACTUAL")[["DEUDA", "PAGO", "SALDO"]].tolist() == [5, 5, 0]
    assert _fila(proyeccion, "MANTENIMIENTO")[["DEUDA", "PAGO", "SALDO"]].tolist() == [3, 3, 0]


@pytest.mark.parametrize(
    ("pago", "pago_anterior", "saldo_anterior", "pago_actual", "saldo_actual"),
    [(6, 6, 2, 0, 5), (10, 8, 0, 2, 3)],
)
def test_pago_aplica_fifo_dentro_de_agua(
        pago, pago_anterior, saldo_anterior, pago_actual, saldo_actual):
    eventos = [
        _evento("A", "1", "AGUA", "2026-08", "CARGO", 8, 1, source="saldo_inicial"),
        _evento("A", "1", "AGUA", "2026-08", "CARGO", 5, 2),
        _evento("A", "1", "AGUA", "2026-08", "PAGO", pago, 3,
                source="5_cobranza", clase="COBRANZA"),
    ]

    proyeccion = repo._proyectar_consumo_temporal(pd.DataFrame(eventos))

    assert _fila(proyeccion, "MES_ANTERIOR")[["PAGO", "SALDO"]].tolist() == [pago_anterior, saldo_anterior]
    assert _fila(proyeccion, "MES_ACTUAL")[["PAGO", "SALDO"]].tolist() == [pago_actual, saldo_actual]


def test_rollover_pasa_saldo_actual_al_mes_anterior():
    eventos = [
        _evento("A", "1", "AGUA", "2026-08", "CARGO", 10, 1),
        _evento("A", "1", "AGUA", "2026-08", "PAGO", 4, 2,
                source="5_cobranza", clase="COBRANZA"),
        _evento("A", "1", "AGUA", "2026-09", "CARGO", 5, 3),
    ]

    primera = repo._proyectar_consumo_temporal(pd.DataFrame(eventos))
    segunda = repo._proyectar_consumo_temporal(pd.DataFrame(eventos))

    assert _fila(primera, "MES_ANTERIOR", "2026-09")[["DEUDA", "SALDO"]].tolist() == [6, 6]
    assert _fila(primera, "MES_ACTUAL", "2026-09")[["DEUDA", "SALDO"]].tolist() == [5, 5]
    for hoja in primera:
        pd.testing.assert_frame_equal(primera[hoja], segunda[hoja])


def test_pago_y_declarado_se_conservan_separados_con_orden_estable():
    eventos = [
        _evento("A", "1", "AGUA", "2026-08", "CARGO", 8, 1, source="saldo_inicial"),
        _evento("A", "1", "AGUA", "2026-08", "CARGO", 5, 2),
        _evento("A", "1", "AGUA", "2026-08", "PAGO", 3, 3,
                source="secretaria", clase="DECLARACION_SECRETARIA"),
        _evento("A", "1", "AGUA", "2026-08", "PAGO", 7, 4,
                source="5_cobranza", clase="COBRANZA"),
    ]
    eventos[2]["TIMESTAMP"] = eventos[3]["TIMESTAMP"]

    proyeccion = repo._proyectar_consumo_temporal(pd.DataFrame(eventos))

    assert _fila(proyeccion, "MES_ANTERIOR")[["PAGO", "DECLARADO", "SALDO"]].tolist() == [5, 3, 0]
    assert _fila(proyeccion, "MES_ACTUAL")[["PAGO", "DECLARADO", "SALDO"]].tolist() == [2, 0, 3]


def test_ajuste_ambiguo_no_reemplaza_vista(tmp_path, monkeypatch):
    eventos = pd.DataFrame([
        _evento("A", "1", "AGUA", "2026-08", "AJUSTE", 2, 1,
                source="manual", clase="CORRECCION_SISTEMA")
    ])
    salida = tmp_path / "vista.xlsx"
    salida.write_bytes(b"vista anterior")
    monkeypatch.setattr(repo, "_leer_eventos", lambda: eventos)
    monkeypatch.setattr(repo, "_lookup_nombres", lambda: {})

    with pytest.raises(repo.ProyeccionTemporalAmbiguaError, match="A-1.*ref-1"):
        repo.generar_vista(salida)

    assert salida.read_bytes() == b"vista anterior"


def test_pago_mayor_que_deuda_falla_sin_recortar():
    eventos = [
        _evento("A", "1", "AGUA", "2026-08", "CARGO", 5, 1),
        _evento("A", "1", "AGUA", "2026-08", "PAGO", 6, 2,
                source="5_cobranza", clase="COBRANZA"),
    ]

    with pytest.raises(ValueError, match="PAGO excede deuda conocida"):
        repo._proyectar_consumo_temporal(pd.DataFrame(eventos))


def test_excel_y_pdf_exponen_las_mismas_hojas_temporales(tmp_path, monkeypatch):
    eventos = pd.DataFrame([
        _evento("I", "9", "AGUA", "2026-08", "CARGO", 8, 1, source="saldo_inicial"),
        _evento("I", "9", "AGUA", "2026-08", "PAGO", 8, 2,
                source="5_cobranza", clase="COBRANZA"),
    ])
    xlsx = tmp_path / "vista.xlsx"
    pdf = tmp_path / "vista.pdf"
    monkeypatch.setattr(repo, "_leer_eventos", lambda: eventos)
    monkeypatch.setattr(repo, "_lookup_nombres", lambda: {("I", "9"): "Caso de control"})
    monkeypatch.setattr(repo, "_MEDIDOR_SALDO_PATH_VISTA", tmp_path / "no_existe.xlsx")

    repo.generar_vista(xlsx)
    repo.exportar_vista_pdf(xlsx, pdf)

    from openpyxl import load_workbook
    import fitz

    wb = load_workbook(xlsx, data_only=True)
    assert wb.sheetnames[:3] == ["MES_ANTERIOR", "MES_ACTUAL", "MANTENIMIENTO"]
    texto_pdf = "\n".join(page.get_text() for page in fitz.open(pdf))
    assert "MES_ANTERIOR" in texto_pdf and "MES_ACTUAL" in texto_pdf and "MANTENIMIENTO" in texto_pdf
    assert "8.00" in texto_pdf and "Caso de control" in texto_pdf
    wb.close()
    xlsx.unlink()
    assert not xlsx.exists()
