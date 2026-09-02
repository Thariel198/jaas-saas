import json
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "shared"))

from exportar_anulaciones_ledger import COLUMNAS, exportar


def test_exporta_lotes_activos_y_revocados(tmp_path):
    origen = tmp_path / "anulaciones.json"
    salida = tmp_path / "anulaciones.xlsx"
    origen.write_text(json.dumps({
        "schema": 1,
        "anulaciones": [
            {
                "id": "CAUSA-1",
                "estado": "ACTIVA",
                "mes": "2026-06",
                "eventos": [{
                    "audit_ref": "ref-1", "mz": "F", "lt": "12",
                    "tipo_evento": "AJUSTE", "monto": -50,
                }],
            },
            {
                "id": "CAUSA-2",
                "estado": "REVOCADA",
                "mes": "2026-08",
                "eventos": [{
                    "audit_ref": "ref-2", "mz": "X", "lt": "33",
                    "tipo_evento": "PAGO", "monto": 3,
                }],
            },
        ],
    }), encoding="utf-8")

    exportar(origen, salida)

    ws = load_workbook(salida)["Eventos"]
    assert [celda.value for celda in ws[1]] == COLUMNAS
    assert ws.max_row == 3
    assert ws["A2"].value == "CAUSA-1"
    assert ws["B3"].value == "REVOCADA"
    assert ws["H2"].value == "ref-1"
    assert ws["N2"].value == -50
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:P3"
    assert ws.column_dimensions["F"].width == 95


def test_rechaza_audit_ref_duplicado(tmp_path):
    origen = tmp_path / "anulaciones.json"
    salida = tmp_path / "anulaciones.xlsx"
    origen.write_text(json.dumps({
        "schema": 1,
        "anulaciones": [{
            "id": "CAUSA",
            "eventos": [{"audit_ref": "repetida"}, {"audit_ref": "repetida"}],
        }],
    }), encoding="utf-8")

    try:
        exportar(origen, salida)
    except ValueError as exc:
        assert "duplicado" in str(exc)
    else:
        raise AssertionError("Debio rechazar el AUDIT_REF duplicado")
