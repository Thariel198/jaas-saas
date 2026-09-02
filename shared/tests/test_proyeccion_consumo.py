from pathlib import Path
import sys

import pandas as pd
import pytest


ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(ROOT / "shared"))
import seguimiento_repo as repo  # noqa: E402


def _evento(mz, lt, concepto, mes, tipo, monto, orden, *, source="2_planilla", clase="GENESIS"):
    return {
        "MZ": mz,
        "LT": lt,
        "CONCEPTO": concepto,
        "MES": mes,
        "TIPO_EVENTO": tipo,
        "CARGO": monto if tipo == "CARGO" else 0,
        "PAGO": monto if tipo == "PAGO" else 0,
        "AJUSTE": monto if tipo == "AJUSTE" else 0,
        "SALDO": 0,
        "SOURCE": source,
        "AUDIT_REF": f"ref-{orden}",
        "TIMESTAMP": f"{mes}-01 00:00:{orden:02d}",
        "CLASE": clase,
        "MOTIVO": "",
    }


def _fila(proyeccion, hoja, mes="2026-08"):
    return proyeccion[hoja].loc[proyeccion[hoja]["MES"] == mes].iloc[0]


def test_i9_separa_anterior_actual_y_mantenimiento():
    eventos = [
        _evento("I", "9", "AGUA", "2026-08", "CARGO", 8, 1, source="saldo_inicial"),
        _evento("I", "9", "AGUA", "2026-08", "CARGO", 5, 2),
        _evento("I", "9", "MANTENIMIENTO", "2026-08", "CARGO", 3, 3),
        _evento("I", "9", "AGUA", "2026-08", "PAGO", 13, 4,
                source="5_cobranza", clase="COBRANZA"),
        _evento("I", "9", "MANTENIMIENTO", "2026-08", "PAGO", 3, 5,
                source="5_cobranza", clase="COBRANZA"),
    ]

    proyeccion = repo._proyectar_consumo_temporal(pd.DataFrame(eventos))

    assert _fila(proyeccion, "MES_ANTERIOR")[["DEUDA", "PAGO", "SALDO"]].tolist() == [8, 8, 0]
    assert _fila(proyeccion, "MES_ACTUAL")[["DEUDA", "PAGO", "SALDO"]].tolist() == [5, 5, 0]
    assert _fila(proyeccion, "MANTENIMIENTO")[["DEUDA", "PAGO", "SALDO"]].tolist() == [3, 3, 0]


def test_resumen_deudas_suma_conceptos_y_conserva_predio_en_cero():
    consumo = {
        concepto: pd.DataFrame([{
            "MZ": "A", "LT": "1", "MES": "2026-08", "SALDO": saldo,
        }])
        for concepto, saldo in (("MES_ANTERIOR", 8), ("MES_ACTUAL", 5),
                                ("MANTENIMIENTO", 3))
    }
    eventos = pd.DataFrame([
        {"MZ": "A", "LT": "1", "CONCEPTO": concepto, "MES": "2026-08",
         "TIMESTAMP": "2026-08-01", "SALDO": saldo}
        for concepto, saldo in (("CORTE_RECONEXION", 4), ("CONVENIO", 25),
                                ("ACUERDOS", 10), ("MULTA", 20), ("OTROS", 2))
    ])

    resumen = repo._resumir_deudas_vista(
        eventos, consumo, {("A", "1"): "Con deuda", ("B", "2"): "Sin deuda"})

    assert resumen.columns.tolist() == list(repo.COLUMNAS_RESUMEN_DEUDAS)
    assert resumen.loc[0, list(repo.CONCEPTOS_VISTA_ORDEN)].tolist() == [8, 5, 3, 4, 25, 10, 20, 2]
    assert resumen.loc[0, "TOTAL_DEUDA"] == 77
    assert resumen.loc[1, list(repo.CONCEPTOS_VISTA_ORDEN) + ["TOTAL_DEUDA"]].tolist() == [0] * 9


@pytest.mark.parametrize(
    ("pago", "pago_anterior", "saldo_anterior", "pago_actual", "saldo_actual"),
    [(6, 6, 2, 0, 5), (10, 8, 0, 2, 3)],
)
def test_pago_aplica_fifo_dentro_de_agua(
        pago, pago_anterior, saldo_anterior, pago_actual, saldo_actual):
    eventos = [
        _evento("A", "1", "AGUA", "2026-08", "CARGO", 8, 1, source="saldo_inicial"),
        _evento("A", "1", "AGUA", "2026-08", "CARGO", 5, 2),
        _evento("A", "1", "AGUA", "2026-08", "PAGO", pago, 3,
                source="5_cobranza", clase="COBRANZA"),
    ]

    proyeccion = repo._proyectar_consumo_temporal(pd.DataFrame(eventos))

    assert _fila(proyeccion, "MES_ANTERIOR")[["PAGO", "SALDO"]].tolist() == [pago_anterior, saldo_anterior]
    assert _fila(proyeccion, "MES_ACTUAL")[["PAGO", "SALDO"]].tolist() == [pago_actual, saldo_actual]


def test_rollover_pasa_saldo_actual_al_mes_anterior():
    eventos = [
        _evento("A", "1", "AGUA", "2026-08", "CARGO", 10, 1),
        _evento("A", "1", "AGUA", "2026-08", "PAGO", 4, 2,
                source="5_cobranza", clase="COBRANZA"),
        _evento("A", "1", "AGUA", "2026-09", "CARGO", 5, 3),
    ]

    primera = repo._proyectar_consumo_temporal(pd.DataFrame(eventos))
    segunda = repo._proyectar_consumo_temporal(pd.DataFrame(eventos))

    assert _fila(primera, "MES_ANTERIOR", "2026-09")[["DEUDA", "SALDO"]].tolist() == [6, 6]
    assert _fila(primera, "MES_ACTUAL", "2026-09")[["DEUDA", "SALDO"]].tolist() == [5, 5]
    for hoja in primera:
        pd.testing.assert_frame_equal(primera[hoja], segunda[hoja])


def test_pago_abono_y_declarado_se_conservan_separados_con_orden_estable():
    eventos = [
        _evento("A", "1", "AGUA", "2026-08", "CARGO", 8, 1, source="saldo_inicial"),
        _evento("A", "1", "AGUA", "2026-08", "CARGO", 5, 2),
        _evento("A", "1", "AGUA", "2026-08", "PAGO", 3, 3,
                source="secretaria", clase="DECLARACION_SECRETARIA"),
        _evento("A", "1", "AGUA", "2026-08", "PAGO", 7, 4,
                source="5_cobranza", clase="COBRANZA"),
        _evento("A", "1", "AGUA", "2026-08", "PAGO", 2, 5,
                source="abonos_rezagados", clase="ABONO_REZAGADO"),
    ]
    eventos[2]["TIMESTAMP"] = eventos[3]["TIMESTAMP"]

    proyeccion = repo._proyectar_consumo_temporal(pd.DataFrame(eventos))

    campos = ["PAGO", "ABONO_REZAGADO", "DECLARADO", "SALDO"]
    assert _fila(proyeccion, "MES_ANTERIOR")[campos].tolist() == [5, 0, 3, 0]
    assert _fila(proyeccion, "MES_ACTUAL")[campos].tolist() == [2, 2, 0, 1]


def test_ajuste_ambiguo_no_reemplaza_vista(tmp_path, monkeypatch):
    eventos = pd.DataFrame([
        _evento("A", "1", "AGUA", "2026-08", "AJUSTE", 2, 1,
                source="manual", clase="CORRECCION_SISTEMA")
    ])
    salida = tmp_path / "vista.xlsx"
    salida.write_bytes(b"vista anterior")
    monkeypatch.setattr(repo, "_leer_eventos", lambda: eventos)
    monkeypatch.setattr(repo, "_lookup_nombres", lambda: {})

    with pytest.raises(repo.ProyeccionTemporalAmbiguaError, match="A-1.*ref-1"):
        repo.generar_vista(salida)

    assert salida.read_bytes() == b"vista anterior"


def test_reversion_cierre_restaura_el_ultimo_tramo_pagado():
    eventos = [
        _evento("A", "1", "AGUA", "2026-08", "CARGO", 8, 1, source="saldo_inicial"),
        _evento("A", "1", "AGUA", "2026-08", "CARGO", 5, 2),
        _evento("A", "1", "AGUA", "2026-08", "PAGO", 10, 3,
                source="5_cobranza", clase="COBRANZA"),
        _evento("A", "1", "AGUA", "2026-08", "AJUSTE", 3, 4,
                source="5_cobranza", clase="CORRECCION_SISTEMA"),
    ]
    eventos[-1]["AUDIT_REF"] = "cierre|2026-08|hash|5_cobranza|A|1|AGUA|0|10.00->7.00|AJUSTE"
    eventos[-1]["MOTIVO"] = "cierre: el objetivo final es menor que lo acreditado provisionalmente"

    proyeccion = repo._proyectar_consumo_temporal(pd.DataFrame(eventos))

    assert _fila(proyeccion, "MES_ANTERIOR")[["PAGO", "AJUSTE", "SALDO"]].tolist() == [8, 1, 1]
    assert _fila(proyeccion, "MES_ACTUAL")[["PAGO", "AJUSTE", "SALDO"]].tolist() == [2, 2, 5]


def test_pago_mayor_que_deuda_falla_sin_recortar():
    eventos = [
        _evento("A", "1", "AGUA", "2026-08", "CARGO", 5, 1),
        _evento("A", "1", "AGUA", "2026-08", "PAGO", 6, 2,
                source="5_cobranza", clase="COBRANZA"),
    ]

    with pytest.raises(ValueError, match="PAGO excede deuda conocida"):
        repo._proyectar_consumo_temporal(pd.DataFrame(eventos))


def test_excel_y_pdf_exponen_las_mismas_hojas_temporales(tmp_path, monkeypatch):
    eventos = pd.DataFrame([
        _evento("I", "9", "AGUA", "2026-08", "CARGO", 8, 1, source="saldo_inicial"),
        _evento("I", "9", "AGUA", "2026-08", "PAGO", 8, 2,
                source="5_cobranza", clase="COBRANZA"),
    ])
    xlsx = tmp_path / "vista.xlsx"
    pdf = tmp_path / "vista.pdf"
    monkeypatch.setattr(repo, "_leer_eventos", lambda: eventos)
    monkeypatch.setattr(repo, "_lookup_nombres", lambda: {("I", "9"): "Caso de control"})
    monkeypatch.setattr(repo, "_MEDIDOR_SALDO_PATH_VISTA", tmp_path / "no_existe.xlsx")

    repo.generar_vista(xlsx)
    repo.exportar_vista_pdf(xlsx, pdf)

    from openpyxl import load_workbook
    import fitz

    wb = load_workbook(xlsx, data_only=True)
    assert wb.sheetnames[:4] == [
        "RESUMEN_DEUDAS", "MES_ANTERIOR", "MES_ACTUAL", "MANTENIMIENTO"]
    assert [wb["RESUMEN_DEUDAS"].cell(2, c).value for c in range(1, 13)] == list(
        repo.COLUMNAS_RESUMEN_DEUDAS)
    texto_pdf = "\n".join(page.get_text() for page in fitz.open(pdf))
    assert "RESUMEN_DEUDAS" in texto_pdf and "MES_ANTERIOR" in texto_pdf
    assert "MES_ACTUAL" in texto_pdf and "MANTENIMIENTO" in texto_pdf
    assert "8.00" in texto_pdf and "Caso de control" in texto_pdf
    wb.close()
    xlsx.unlink()
    assert not xlsx.exists()
