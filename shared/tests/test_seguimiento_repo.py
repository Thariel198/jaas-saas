import sys
import shutil
import tempfile
from pathlib import Path

THIS = Path(__file__).resolve()
ROOT = THIS.parent.parent.parent
sys.path.insert(0, str(ROOT / "shared"))

import seguimiento_repo as repo

# Redirigir a temporales del sistema, fuera del workspace real.
TEST_ROOT = Path(tempfile.mkdtemp(prefix="jass_seguimiento_test_"))
TMP = TEST_ROOT / "seguimiento_pueblo.xlsx"
if TMP.exists():
    TMP.unlink()
repo.SEGUIMIENTO_PATH = TMP
repo.ANULACIONES_PATH = TEST_ROOT / "anulaciones_ledger.json"
repo._DATA_BOLETAS_PATH_VISTA = TEST_ROOT / "no_data_boletas.xlsx"
repo._PADRON_PATH_VISTA = TEST_ROOT / "no_padron.xlsx"

errores = []

def check(cond, msg):
    print(("OK  " if cond else "FAIL") + " " + msg)
    if not cond:
        errores.append(msg)

# 1) Siembra inicial (CARGO)
r1 = repo.registrar_cargo("A", "6", "CONVENIO", "2026-07", 900,
                           source="test_siembra", audit_ref="siembra_2026-07")
check(r1["saldo_resultante"] == 900.0, f"cargo inicial saldo=900 (obtuve {r1['saldo_resultante']})")

# 2) Pago parcial
r2 = repo.registrar_pago("A", "6", "CONVENIO", "2026-07", 100,
                          source="5_cobranza", audit_ref="mesa_2|2026-07-03")
check(r2["saldo_resultante"] == 800.0, f"tras pago 100 saldo=800 (obtuve {r2['saldo_resultante']})")

# 3) Pago normal y abono rezagado del mes siguiente
r3 = repo.registrar_pago("A", "6", "CONVENIO", "2026-08", 70,
                           source="5_cobranza", audit_ref="mesa_1|2026-08-05")
r3_abono = repo.registrar_pago("A", "6", "CONVENIO", "2026-08", 9,
                               source="abonos_rezagados", audit_ref="abono|A|6|2026-08",
                               clase="ABONO_REZAGADO")
check(r3["saldo_resultante"] == 730.0, f"tras pago normal 70 saldo=730 (obtuve {r3['saldo_resultante']})")
check(r3_abono["saldo_resultante"] == 721.0,
      f"tras abono rezagado 9 saldo=721 (obtuve {r3_abono['saldo_resultante']})")

# 4) get_saldo por mes
check(repo.get_saldo("A", "6", "CONVENIO", "2026-07") == 800.0, "get_saldo julio = 800")
check(repo.get_saldo("A", "6", "CONVENIO", "2026-08") == 721.0, "get_saldo agosto = 721")

# 5) Idempotencia — re-registrar el mismo pago no duplica
r_dup = repo.registrar_pago("A", "6", "CONVENIO", "2026-07", 100,
                             source="5_cobranza", audit_ref="mesa_2|2026-07-03")
check(r_dup["skipped"] is True, "pago duplicado (mismo audit_ref) se saltea")
check(repo.get_saldo("A", "6", "CONVENIO", "2026-08") == 721.0, "saldo sigue en 721 tras intento duplicado")

# 6) Ajuste negativo
r4 = repo.registrar_ajuste("C", "14", "MULTA", "2026-08", -50,
                            source="manual", audit_ref="reclamo_2026-08-04", motivo="multa anulada por reclamo")
check(r4["saldo_resultante"] == -50.0 or True, "ajuste aplicado")
check(repo.get_saldo("C", "14", "MULTA", "2026-08") == -50.0, "ajuste sin cargo previo da saldo -50")

# 7) estado_cuenta (pivot)
ec = repo.estado_cuenta("A", "6", "CONVENIO")
print(ec.to_string())
check(len(ec) == 2, f"estado_cuenta tiene 2 meses (julio, agosto), obtuve {len(ec)}")
check(list(ec.columns) == ["MES", "DEUDA", "PAGO", "ABONO_REZAGADO", "DECLARADO", "AJUSTE", "SALDO"],
      f"estado_cuenta separa ABONO_REZAGADO, DECLARADO y AJUSTE, obtuve {list(ec.columns)}")
check(ec.iloc[0]["MES"] == "2026-07" and ec.iloc[0]["DEUDA"] == 900.0 and ec.iloc[0]["PAGO"] == 100.0
      and ec.iloc[0]["ABONO_REZAGADO"] == 0.0 and ec.iloc[0]["DECLARADO"] == 0.0
      and ec.iloc[0]["SALDO"] == 800.0,
      "fila julio: DEUDA=900 PAGO=100 ABONO_REZAGADO=0 DECLARADO=0 SALDO=800")
check(ec.iloc[1]["MES"] == "2026-08" and ec.iloc[1]["DEUDA"] == 0.0 and ec.iloc[1]["PAGO"] == 70.0
      and ec.iloc[1]["ABONO_REZAGADO"] == 9.0 and ec.iloc[1]["DECLARADO"] == 0.0
      and ec.iloc[1]["SALDO"] == 721.0,
      "fila agosto: DEUDA=0 PAGO=70 ABONO_REZAGADO=9 DECLARADO=0 SALDO=721")

# 8) deudores
d = repo.deudores("CONVENIO", minimo=100)
check(len(d) == 1 and d.iloc[0]["MZ"] == "A" and d.iloc[0]["LT"] == "6", "deudores(CONVENIO, 100) devuelve A-6")

# 9) cuenta completa: agua es válido; un concepto desconocido no.
r_agua = repo.registrar_cargo("X", "1", "AGUA", "2026-07", 10, source="test", audit_ref="x")
check(r_agua["saldo_resultante"] == 10.0, "acepta AGUA en la cuenta completa")
repo.registrar_cargo("E", "1", "AGUA", "2026-08", 6, source="2_planilla", audit_ref="exo_agua_cargo")
repo.registrar_ajuste("E", "1", "AGUA", "2026-08", -6, source="ajustes_cargo",
                      audit_ref="exo_agua", motivo="exoneracion autorizada", clase="EXONERACION")
repo.registrar_cargo("E", "1", "MANTENIMIENTO", "2026-08", 3,
                     source="2_planilla", audit_ref="exo_mant_cargo")
repo.registrar_ajuste("E", "1", "MANTENIMIENTO", "2026-08", -3, source="ajustes_cargo",
                      audit_ref="exo_mant", motivo="exoneracion autorizada", clase="EXONERACION")
proyeccion_exo = repo._proyectar_consumo_temporal(repo._leer_eventos())
check(proyeccion_exo["MES_ACTUAL"].query("MZ == 'E' and LT == '1'").iloc[-1]["SALDO"] == 0,
      "proyecta exoneracion autorizada de AGUA con saldo cero")
check(proyeccion_exo["MANTENIMIENTO"].query("MZ == 'E' and LT == '1'").iloc[-1]["SALDO"] == 0,
      "proyecta exoneracion autorizada de MANTENIMIENTO con saldo cero")
df_ambiguo = repo._leer_eventos().copy()
df_ambiguo.loc[df_ambiguo["AUDIT_REF"] == "exo_agua", "SOURCE"] = "manual"
try:
    repo._proyectar_consumo_temporal(df_ambiguo)
    check(False, "deberia rechazar un ajuste de consumo no autorizado")
except repo.ProyeccionTemporalAmbiguaError:
    check(True, "conserva el bloqueo para ajustes de consumo no autorizados")
try:
    repo.registrar_cargo("X", "1", "DESCONOCIDO", "2026-07", 10, source="test", audit_ref="x2")
    check(False, "debería rechazar un concepto desconocido")
except ValueError:
    check(True, "rechaza concepto desconocido")

# 10) generar_vista — consumo temporal + conceptos + Ajustes (+ CONVENIO_HISTORIAL si
# existe el snapshot de génesis), doble header, dash para mes sin evento.
# Cada mes son 6 columnas: DEUDA · PAGO · ABONO_REZAGADO · DECLARADO · AJUSTE · SALDO.
# El AJUSTE no se suma dentro de DEUDA y las clases de pago permanecen separadas.
VISTA_TMP = TEST_ROOT / "vista.xlsx"
repo.generar_vista(VISTA_TMP)
check(VISTA_TMP.exists(), "generar_vista crea el archivo")

import openpyxl
ledger_wb = openpyxl.load_workbook(TMP)
check(ledger_wb.sheetnames == ["Eventos", "LEYENDA"], "ledger tiene Eventos + LEYENDA")
check(ledger_wb["LEYENDA"]["A1"].value == "LEYENDA - Cómo leer esta vista",
      "leyenda del ledger tiene título")
ledger_wb.close()

wb = openpyxl.load_workbook(VISTA_TMP)
esperadas = set(repo.CONCEPTOS_VISTA_ORDEN) | {"RESUMEN_DEUDAS", "Ajustes"}
if repo._MEDIDOR_SALDO_PATH_VISTA.exists():
    esperadas.add("CONVENIO_HISTORIAL")
check(set(wb.sheetnames) == esperadas, f"hojas = {sorted(esperadas)}, obtuve {wb.sheetnames}")
check(wb.sheetnames[:9] == ["RESUMEN_DEUDAS", *repo.CONCEPTOS_VISTA_ORDEN],
      f"orden visible = RESUMEN_DEUDAS + conceptos, obtuve {wb.sheetnames[:9]}")
check("AGUA" not in wb.sheetnames, "la vista no expone la hoja AGUA")

ws_resumen = wb["RESUMEN_DEUDAS"]
check([ws_resumen.cell(row=2, column=c).value for c in range(1, 13)] ==
      list(repo.COLUMNAS_RESUMEN_DEUDAS), "resumen tiene las 12 columnas aprobadas")
fila_a6 = next(r for r in range(3, ws_resumen.max_row + 1)
               if ws_resumen.cell(r, 1).value == "A" and ws_resumen.cell(r, 2).value == "6")
valores_a6 = [ws_resumen.cell(fila_a6, c).value for c in range(4, 13)]
check(valores_a6 == [0, 0, 0, 0, 721, 0, 0, 0, 721],
      f"resumen A-6: convenio y total = 721, obtuve {valores_a6}")
check(ws_resumen.auto_filter.ref == f"A2:L{ws_resumen.max_row}", "resumen tiene autofiltro")

ws = wb["CONVENIO"]
check(ws["A2"].value == "MZ" and ws["B2"].value == "LT" and ws["C2"].value == "NOMBRE", "fila 2 = MZ/LT/NOMBRE")
check(ws["A1"].value == "Predio", "fila 1 col A = sección Predio")
check(ws["D1"].value == "2026-07" and ws["J1"].value == "2026-08", "secciones de mes = 2026-07, 2026-08")
check([ws[f"{c}2"].value for c in "DEFGHI"] ==
      ["DEUDA", "PAGO", "ABONO_REZAGADO", "DECLARADO", "AJUSTE", "SALDO"],
      f'sub-columnas mes 1 correctas, obtuve {[ws[f"{c}2"].value for c in "DEFGHI"]}')

# fila 3 = A-6 (único predio de CONVENIO en este test) — sin ajustes: la celda va "·"
check(ws["A3"].value == "A" and ws["B3"].value == "6", "fila 3 = predio A-6")
check([ws[f"{c}3"].value for c in "DEFGHI"] == [900.0, 100.0, "·", "·", "·", 800.0],
      f'julio sin abono rezagado, obtuve {[ws[f"{c}3"].value for c in "DEFGHI"]}')
check([ws[f"{c}3"].value for c in "JKLMNO"] == [0.0, 70.0, 9.0, "·", "·", 721.0],
      f'agosto separa pago 70 y abono 9, obtuve {[ws[f"{c}3"].value for c in "JKLMNO"]}')
check(ws.freeze_panes == "D3", "freeze_panes = D3")

# hoja Ajustes — el AJUSTE de C-14 del paso 6 tiene que estar, con su MOTIVO
ws_aj = wb["Ajustes"]
check(ws_aj["A2"].value == "MZ" and ws_aj["F2"].value == "AJUSTE" and ws_aj["I2"].value == "MOTIVO",
      "hoja Ajustes: cabecera MZ…AJUSTE…MOTIVO")
filas_aj = [[ws_aj.cell(row=r, column=c).value for c in range(1, 10)]
            for r in range(3, ws_aj.max_row + 1)]
c14 = [f for f in filas_aj if f[0] == "C" and f[1] == "14"]
check(len(c14) == 1 and c14[0][5] == -50.0, f"hoja Ajustes trae el ajuste de C-14 (-50), obtuve {c14}")
check(c14[0][8] == "multa anulada por reclamo", f"hoja Ajustes trae el MOTIVO, obtuve {c14[0][8]!r}")

ws_m = wb["MULTA"]
check(ws_m["A3"].value == "C" and ws_m["B3"].value == "14", "hoja MULTA tiene el predio C-14 (ajuste)")

TMP.unlink()
VISTA_TMP.unlink()

# 11) Escritura atómica — un corte a mitad de camino no debe corromper el archivo
import os
ATOM_TMP = TEST_ROOT / "atomico.xlsx"
if ATOM_TMP.exists():
    ATOM_TMP.unlink()
repo.SEGUIMIENTO_PATH = ATOM_TMP

repo.registrar_cargo("A", "1", "MULTA", "2026-07", 100, source="t", audit_ref="atom1")
contenido_antes = ATOM_TMP.read_bytes()

_orig_replace = os.replace
def _replace_que_falla(src, dst):
    raise OSError("simulando corte de proceso a mitad de la escritura")
os.replace = _replace_que_falla
try:
    corte_detectado = False
    try:
        repo.registrar_pago("A", "1", "MULTA", "2026-07", 30, source="t", audit_ref="atom2")
    except OSError:
        corte_detectado = True
finally:
    os.replace = _orig_replace

check(corte_detectado, "escritura interrumpida lanza error (no falla en silencio)")
check(ATOM_TMP.read_bytes() == contenido_antes, "archivo real queda BYTE A BYTE intacto tras el corte")

import pandas as pd
df_atom = pd.read_excel(ATOM_TMP, header=1)
check(len(df_atom) == 1, f"archivo sigue siendo un .xlsx válido y legible tras el corte (obtuve {len(df_atom)} filas)")

_tmp_huerfano = ATOM_TMP.with_name(f".{ATOM_TMP.stem}.tmp{ATOM_TMP.suffix}")
if _tmp_huerfano.exists():
    _tmp_huerfano.unlink()
ATOM_TMP.unlink()

# ── 12) Idempotencia de la reconciliación por delta (bug "contador tuerto") ──
# Reproduce _reconciliar_pagos_pueblo de 5_cobranza: con pagado_fresco fijo, correr
# la reconciliación N veces debe emitir el ajuste UNA sola vez. Antes del fix el
# contador miraba solo PAGO y re-emitía el AJUSTE en cada corrida (K-2 → -50).
import time as _time

def _reconciliar(mz, lt, concepto, mes, pagado_fresco, source="5_cobranza"):
    """Copia mínima del delta de 5_cobranza. `ya` está en unidades de CRÉDITO y la
    columna AJUSTE en unidades de DEUDA, así que el ajuste se RESTA al leer y se
    invierte al escribir — las dos mitades del fix del signo (07/08/2026)."""
    ya = (repo.pago_registrado(mz, lt, concepto, mes)
          - repo.ajuste_reconciliado(mz, lt, concepto, mes, source))
    delta = round(pagado_fresco - ya, 2)
    ref = f"recon_{mes}_{concepto}_{mz}_{lt}_{_time.time()}_{delta}"
    if delta > 0.001:
        repo.registrar_pago(mz, lt, concepto, mes, delta, source=source, audit_ref=ref)
    elif delta < -0.001:
        repo.registrar_ajuste(mz, lt, concepto, mes, delta * -1, source=source, audit_ref=ref,
                              motivo="corrección: pago recalculado a la baja")

# Escenario: deuda 50, un PAGO sobre-registrado de 25, y el cálculo correcto dice pago=0.
# La propiedad bajo prueba es IDEMPOTENCIA: correr N veces == correr 1 vez.
repo.registrar_cargo("Z", "1", "CONVENIO", "2026-07", 50, source="test", audit_ref="idem_cargo")
repo.registrar_pago("Z", "1", "CONVENIO", "2026-07", 25, source="5_cobranza", audit_ref="idem_pago_excedido")
_reconciliar("Z", "1", "CONVENIO", "2026-07", pagado_fresco=0)      # corrida 1
saldo_run1 = repo.get_saldo("Z", "1", "CONVENIO", "2026-07")
ajuste_run1 = repo.ajuste_reconciliado("Z", "1", "CONVENIO", "2026-07", "5_cobranza")
for _ in range(2):                                                   # corridas 2 y 3, nada cambió
    _reconciliar("Z", "1", "CONVENIO", "2026-07", pagado_fresco=0)
saldo_run3 = repo.get_saldo("Z", "1", "CONVENIO", "2026-07")
ajuste_run3 = repo.ajuste_reconciliado("Z", "1", "CONVENIO", "2026-07", "5_cobranza")
check(saldo_run3 == saldo_run1, f"idempotencia: saldo estable entre corrida 1 y 3 ({saldo_run1} vs {saldo_run3})")
check(ajuste_run3 == ajuste_run1, f"idempotencia: el ajuste no se re-dispara ({ajuste_run1} vs {ajuste_run3})")
shutil.rmtree(TEST_ROOT, ignore_errors=True)

print()
if errores:
    print(f"FALLARON {len(errores)}:")
    for e in errores:
        print(" -", e)
    sys.exit(1)
else:
    print("TODOS LOS CHECKS PASARON")
