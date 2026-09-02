import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).parents[2]
sys.path.insert(0, str(ROOT / "shared"))
import seguimiento_repo as repo  # noqa: E402


def test_vista_provisional_no_muta_ledger(tmp_path, monkeypatch):
    ledger = tmp_path / "ledger.xlsx"
    vista = tmp_path / "vista_seguimiento_provicional.xlsx"
    monkeypatch.setattr(repo, "SEGUIMIENTO_PATH", ledger)
    monkeypatch.setattr(repo, "_lookup_nombres", lambda: {("Q", "5"): "Caso Q-5"})

    for concepto, monto in (("CONVENIO", 25), ("ACUERDOS", 50), ("MULTA", 19)):
        repo.registrar_cargo("Q", "5", concepto, "2026-07", monto,
                             source="test", audit_ref=f"cargo-{concepto}")
    repo.registrar_pago("Q", "5", "CONVENIO", "2026-08", 25,
                        source="abonos_rezagados", audit_ref="viejo-convenio")
    repo.registrar_pago("Q", "5", "ACUERDOS", "2026-08", 50,
                        source="abonos_rezagados", audit_ref="viejo-acuerdos")
    repo.registrar_pago("Q", "5", "MULTA", "2026-08", 3,
                        source="abonos_rezagados", audit_ref="viejo-multa")
    original = ledger.read_bytes()

    conceptos = ("AGUA", "MANTENIMIENTO", "CORTE_RECONEXION", "CONVENIO", "ACUERDOS", "MULTA")
    objetivos = [{
        "mz": "Q", "lt": "5", "concepto": concepto, "source": "abonos_rezagados",
        "monto_objetivo": {"AGUA": 20, "CONVENIO": 25, "ACUERDOS": 50, "MULTA": 19}.get(concepto, 0),
        "clase": "ABONO_REZAGADO",
    } for concepto in conceptos]
    cargos = [
        {"mz": "Q", "lt": "5", "concepto": "AGUA", "mes": "2026-08", "monto": 20,
         "source": "saldo_inicial", "audit_ref": "apertura-agua", "clase": "GENESIS"},
        {"mz": "Q", "lt": "5", "concepto": "AGUA", "mes": "2026-08", "monto": 13,
         "source": "2_planilla", "audit_ref": "agua-agosto", "clase": "GENESIS"},
        {"mz": "Q", "lt": "5", "concepto": "MANTENIMIENTO", "mes": "2026-08", "monto": 3,
         "source": "2_planilla", "audit_ref": "mant-agosto", "clase": "GENESIS"},
    ]
    snapshot = {"schema": 2, "mes": "2026-08", "objetivos": objetivos, "cargos": cargos}
    repo.generar_vista_provisional(
        "2026-08", "a" * 64, snapshot, vista,
        estado_validacion="NO VALIDADO", alerta="Diferencia conocida",
    )

    assert ledger.read_bytes() == original
    wb = load_workbook(vista, data_only=True)
    assert wb.sheetnames[:6] == [
        "PROVISIONAL", "RESUMEN_DEUDAS", "MES_ANTERIOR", "MES_ACTUAL",
        "MANTENIMIENTO", "CORTE_RECONEXION"]
    assert wb["PROVISIONAL"]["B4"].value == "NO VALIDADO"
    assert wb["PROVISIONAL"]["B5"].value == "Diferencia conocida"
    ws_resumen = wb["RESUMEN_DEUDAS"]
    assert ws_resumen["A3"].value == "Q" and ws_resumen["B3"].value == "5"
    assert [ws_resumen.cell(3, c).value for c in range(4, 13)] == [
        0, 13, 3, 0, 0, 0, 0, 0, 16]
    ws_anterior = wb["MES_ANTERIOR"]
    assert ws_anterior["A3"].value == "Q" and ws_anterior["B3"].value == "5"
    assert ws_anterior["D3"].value == 20 and ws_anterior["F3"].value == 20
    ws_actual = wb["MES_ACTUAL"]
    assert ws_actual["D3"].value == 13 and ws_actual["I3"].value == 13
    ws_mant = wb["MANTENIMIENTO"]
    assert ws_mant["D3"].value == 3 and ws_mant["I3"].value == 3
