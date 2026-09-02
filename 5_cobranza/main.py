"""5_cobranza/main.py — Carga pagos en planilla · genera estado de cobro

Lee planilla (de 2_planilla) + pagos_yape + pagos_efectivo (de 4_pagos).
Genera outputs provisionales y un snapshot determinista para el cierre:
  · planilla_cobrado_YYYY-MM.xlsx  — copia enriquecida con pagos + SALDO + ESTADO
  · trazabilidad_cobranza.xlsx     — un registro por pago cargado (acumulada)
  · resumen_recaudacion.xlsx       — totales del mes
  · arrastre_deuda_YYYY-MM.xlsx    — SALDO>0  → 2_planilla del próximo mes
  · planilla_cobrado_YYYY-MM.xlsx — planilla + arrastres consolidado/devolucion
  · snapshot_ledger_YYYY-MM.json — objetivo del ledger; no lo escribe

SALDO sale como columna explícita en planilla_cobrado — la lista de corte la
genera 6_corte/generar_lista.py leyendo SALDO + MES_ANTERIOR desde acá.

Idempotente: si los pagos no cambiaron respecto a la trazabilidad existente,
sale sin modificar nada.
"""
import json
import hashlib
import logging
import re
import shutil
import sys
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── CONFIG ────────────────────────────────────────────────────────────────────
ROOT          = Path(__file__).parent

# seguimiento_pueblo: writer único es shared/seguimiento_repo.py (patrón repo,
# igual que data_boletas_repo). Se importa por ruta física — no depende de
# SHARED_DIR (que sí se monkey-patchea en tests) porque el módulo vive siempre
# en el shared/ real, sea cual sea la data que apunte ahí en un test.
sys.path.insert(0, str(ROOT.parent / "shared"))
import seguimiento_repo as repo  # noqa: E402
import utils_estado_ciclo as repo_estado  # noqa: E402
import ciclo as ciclo_activo  # noqa: E402
INPUTS_DIR    = ROOT / "inputs"
OUTPUTS_DIR   = ROOT / "outputs"
SHARED_DIR    = ROOT.parent / "shared"

# Single source of truth: la planilla del mes vive en shared/planilla_mes/.
# 2_planilla la publica ahí, 6_corte/aplicar_penalidad le suma CORTE_RECONEXION,
# y 4_pagos también la lee desde ahí. Leemos el master para que las penalidades
# (y cualquier cargo posterior) se reflejen sin copiar la planilla a mano.
PLAN_DIR      = SHARED_DIR / "planilla_mes"
# Single source of truth: los pagos se leen en vivo desde los outputs de 4_pagos,
# no desde copias en inputs/ (que quedaban stale y perdían pagos agregados después
# de copiar — ej. el 2º pago Yape de un predio). 5_cobranza retroescribe
# CICLO_COBRANZA en estos archivos (per contrato, le corresponde llenarla).
YAPE_DIR      = ROOT.parent / "4_pagos" / "yape" / "motor_matching" / "outputs"
EFEC_DIR      = ROOT.parent / "4_pagos" / "efectivo" / "outputs"
BLANCOS_PATH  = SHARED_DIR / "blancos_acumulados.xlsx"

# Overlay de penalidad (Modelo A · writer único). 6_corte y 6b ya NO escriben
# CORTE_RECONEXION en shared/planilla_mes — la penalidad vive SOLO en sus audits.
# 5_cobranza la re-deriva en vivo sumando el neto de PENALIDAD_APLICADA (col 5,
# ya viene con signo +/-) por predio. shared queda como base pura de 2_planilla.
AUDIT_CORTE_PATH = ROOT.parent / "6_corte" / "outputs" / "audit_penalidad.xlsx"
AUDIT_MULTA_PATH = ROOT.parent / "6b_corte_multas" / "outputs" / "audit_penalidad_multas.xlsx"

# Overlay de devoluciones aplicadas — créditos de exceso de ciclos anteriores
# que se deciden aplicar a un concepto de deuda en vez de esperar reclamo.
# Writer único humano (no lo escribe ningún main.py). No toca mesa_N.xlsx ni
# seguimiento_pueblo.xlsx — es su propia fuente, igual patrón que la penalidad.
DEVOLUCIONES_APLICADAS_PATH = SHARED_DIR / "devoluciones_aplicadas.xlsx"

# Overlay de ajustes de cargo — un CARGO nació (multa, corte...) y después se
# determinó que NO correspondía. Distinto de devoluciones_aplicadas: ahí SÍ
# hubo plata de más (exceso) que se redirige; acá NO hay plata de más — el
# cargo mismo se anula. Ancla el hecho original vía REF_AUDIT (ej. la fila
# APLICADO en 6_corte/outputs/audit_penalidad.xlsx) para que el backfill
# cuente la historia completa: nace el CARGO en MES_ANO_ORIGEN, nace el AJUSTE
# que lo anula en MES_ANO_APLICA — dos hechos reales, no uno solo. Precursor
# manual de registrar_ajuste (libro_mayor/estado_cuenta).
AJUSTES_CARGO_PATH = SHARED_DIR / "ajustes_cargo.xlsx"

# Overlay de génesis tardía — un CARGO que nació BIEN (el cobro fue real, en
# mesa) pero su fuente en obligaciones/ se corrigió DESPUÉS de que el ciclo ya
# estaba congelado — no llegó a tiempo a la génesis de 2_planilla. Distinto de
# ajustes_cargo (ese anula un cargo que nació MAL); acá el cargo es legítimo,
# solo tardío. Se aplica una vez, en el ciclo congelado; el siguiente ciclo ya
# lee la fuente arreglada directo, sin necesitar esta fila. Writer único
# humano. Precursor de registrar_cargo (libro_mayor/estado_cuenta).
GENESIS_TARDIA_PATH = SHARED_DIR / "genesis_tardia.xlsx"

# Overlay de reidentificación — un pago de un ciclo anterior quedó anotado en
# el predio equivocado (typo/confusión de nombre en la mesa), no es exceso
# real. Distinto de devoluciones_aplicadas: ancla la TRANSACCIÓN específica
# (origen_archivo+fila), no un concepto de deuda ya reconocida. Precursor
# manual de reasignar_abono (libro_mayor/dominio, decisión I1-I4).
REIDENTIFICACION_PATH = SHARED_DIR / "reidentificacion.xlsx"

# Overlay de correcciones de deuda — lado espejo de reidentificación: cuando el
# pago mal atribuido SÍ pagaba deuda real del lote de origen (a diferencia de
# los casos donde solo generaba exceso), hay que devolverle esa deuda al origen
# además de acreditar al destino real. Referencia la misma transacción que su
# fila hermana en reidentificacion.xlsx (REF_TRANSACCION), pero con su propio
# MES_ANO_APLICA — normalmente un ciclo después, para no chocar con boletas ya
# impresas. Writer único humano.
DEUDA_CORRECCIONES_PATH = SHARED_DIR / "deuda_correcciones.xlsx"

# Overlay de abonos rezagados — abonos de un ciclo anterior que NO llegaron a la
# caja JASS en su momento (el cobrador recibió el yape del vecino pero no lo
# transfirió) y se regularizan en efectivo un mes después. NO es exceso ni pago
# mal atribuido: es plata que el vecino sí pagó a tiempo, que la caja recién ve
# ahora. Salda deuda del ciclo viejo (en el mes en curso = MES_ANTERIOR), por eso
# usa la misma cascada que reidentificación (sin mes_actual/mantenimiento).
# Precursor manual de caja.MovimientoCaja (libro_mayor). Writer único humano.
ABONOS_REZAGADOS_PATH = SHARED_DIR / "abonos_rezagados.xlsx"
ABONOS_MANIFEST_PATH = INPUTS_DIR / "abonos_rezagados_manifest_2026-08.json"

# Overlay de blancos de efectivo — pagos en efectivo que entraron a la caja de
# un ciclo anterior sin MZ/LT (bug B6: efectivo no ruteaba blancos, a
# diferencia de yape) y quedaron descartados en silencio. Distinto de
# abonos_rezagados: la plata SÍ estaba en la caja del ciclo original, lo que
# faltaba era la atribución al predio, no el movimiento. Solo las filas con
# MZ/LT ya identificado aplican; las que aún no tienen dueño se saltean sin
# tocar nada (quedan visibles en el archivo, esperando identificación).
# Precursor manual de identificar_abono (libro_mayor). Writer único humano.
BLANCOS_EFECTIVO_PATH = SHARED_DIR / "blancos_efectivo.xlsx"

# Overlay de aportes al tanque manuales — plata que el pagador destinó al
# tanque comunitario (voluntario, NO es deuda) pero que 5_cobranza no sabía
# distinguir del pago de deuda normal: entraba al mismo total_pagado y la
# cascada P1-P6 la consumía como si fuera agua/corte/multa/acuerdos/convenio.
# consolidar_tanque.py ya lee este archivo para el reporte de aportes; este
# overlay es el que faltaba — restar el monto del total_pagado ANTES de la
# cascada, para que no se cuente dos veces (una como aporte voluntario, otra
# como pago de deuda). Writer único humano, no lo escribe ningún main.py.
# Precursor de caja.MovimientoCaja BALDE=tanque (libro_mayor).
APORTES_TANQUE_MANUALES_PATH = SHARED_DIR / "aportes_tanque_manuales.xlsx"

# Overlay de reasignación de aplicación — un ABONO que la cascada P1-P6
# aplicaría de oficio a un CARGO (CONCEPTO_ORIGEN) porque le toca por
# prioridad, pero el pagador especificó otro concepto de destino al momento
# de pagar. Distinto de ajustes_cargo (ahí el CARGO no correspondía y se
# anula) y de devoluciones_aplicadas (ahí la plata era sobrante sin cargo que
# la interceptara): acá el CARGO de origen sigue vigente y abierto — solo se
# le quita esta porción de pago y se manda al destino que pidió el pagador.
# Writer único humano. Precursor de reasignar_aplicacion (libro_mayor/estado_cuenta).
REASIGNACIONES_APLICACION_PATH = SHARED_DIR / "reasignaciones_aplicacion.xlsx"

# Gate del arrastre_consolidado: solo se emite tras el sello de 5b_validacion.
ESTADO_CICLO_PATH = SHARED_DIR / "reporte_acumulado_procesado" / "estado_ciclo.json"

# Los outputs de 4_pagos llevan el ciclo en el nombre desde 2026-08
# (pagos_yape_tepago_2026-08.xlsx). El nombre viejo sin periodo se acepta SOLO
# para los ciclos anteriores a esa frontera (ciclo.acepta_legacy): para el ciclo
# en curso, si 4_pagos todavía no generó su archivo, esto tiene que fallar —
# aceptar el pelado sería volver a leer los pagos del mes pasado, que es lo que
# sembró 15 pagos fantasma el 06/07/2026.
_YAPE_BASE         = "pagos_yape_tepago"
_EFEC_BASE         = "pagos_efectivo"
_YAPE_DEV_BASE     = "pagos_yape_devolucion"
_YAPE_RETORNO_BASE = "pagos_yape_retorno"
_EFEC_DEV_BASE     = "pagos_efectivo_devolucion"


def _pago_path(carpeta: Path, base: str, mes_ano: str | None = None) -> Path:
    mes = mes_ano or ciclo_activo.activo(default=None, path=SHARED_DIR / "ciclo_activo.json")
    if mes is None:
        return carpeta / f"{base}.xlsx"
    return ciclo_activo.resolver(carpeta, base, mes,
                                 legacy_sin_periodo=ciclo_activo.acepta_legacy(mes))

CORR_LOTE_PATH = INPUTS_DIR / "correcciones_lote.xlsx"

PENALIDAD     = 20.0   # S/ por reconexión
ARRASTRE_MIN  = 8.0    # MES_ANTERIOR ≥ 8 confirma no pago anterior
TOL           = 0.005  # tolerancia de redondeo

# Posiciones en shared/blancos_acumulados.xlsx (1-indexed)
_BL_MZ   = 13
_BL_LOTE = 14
_BL_EST  = 18
_BL_MES_APLICADO = 19

# ── PALETA — coincide con planilla_cobrado_diseno.html ───────────────────────
GH_QUIEN = ("EBF5FB", "1A5276")
GH_LEC   = ("E6F1FB", "0C447C")
GH_COB   = ("E9F7EF", "1E5C3A")
GH_DESC  = ("EDE9FE", "4C1D95")
GH_TOTAL = ("1E8449", "FFFFFF")
GH_PAGO  = ("F3E8FF", "5B21B6")
GH_TRAZ  = ("FEF9E7", "7D6608")

TD_QUIEN = "F4FAFF"
TD_LEC   = "F0F8FF"
TD_COB   = "F4FBF7"
TD_DESC  = "F5F3FF"
TD_TOTAL = "D5F5E3"
TD_PAGO  = "FAF5FF"
TD_TRAZ  = "FEFCE8"

ESTADO_BG  = {"CANCELADO": "E1F5EE", "EXCESO": "EFF6FF",
              "PARCIAL":   "FAEEDA", "PENDIENTE": "FEF2F2"}
ESTADO_TXT = {"CANCELADO": "085041", "EXCESO": "1D4ED8",
              "PARCIAL":   "854F0B", "PENDIENTE": "991B1B"}

# Badges para RETORNO (medio en que se devolvio el pago — puntero a archivo retorno)
RETORNO_BG  = {"yape": "E1F5EE", "efectivo": "EFF6FF", "mixto": "FEF9E7"}
RETORNO_TXT = {"yape": "085041", "efectivo": "1D4ED8", "mixto": "854F0B"}

# Paleta trazabilidad — ¿Quién es? (morado/lila)
GH_TZ_QUIEN  = ("F4ECF7", "5B21B6")
TD_TZ_QUIEN  = "FAF5FF"

# Paleta trazabilidad — ¿Cómo verificarlo? (ámbar) · ¿De qué ciclo? (azul)
GH_TZ_VERIF  = ("FFF3CD", "7C2D12")   # REFERENCIA · FECHA · COMENTARIO
TD_TZ_VERIF  = "FFFBEC"
GH_TZ_CICLO  = ("EFF6FF", "1E40AF")   # CICLO_CORRECCION_ORIGEN · CICLO_COBRANZA · FECHA_CARGA
TD_TZ_CICLO  = "EFF6FF"

# Paleta arrastre_deuda (coincide con arrastre_deuda_diseno.html)
GH_AD_QUIEN  = ("E8F8F5", "0E6655")
GH_AD_MONTO  = ("EAF2FF", "1A5276")
GH_AD_TRAZ   = ("E8F8F5", "0E6655")
TD_AD_QUIEN  = "F0FFF8"
TD_AD_MONTO  = "EBF5FB"
TD_AD_TRAZ   = "F0FFF8"

# arrastre_consolidado — P1..P5 por prioridad (formato_arrastre_consolidado.html)
GH_AC_QUIEN = ("E8F8F5", "0E6655")
GH_AC_TOTAL = ("1E8449", "FFFFFF")
# (header_bg, header_txt, cell_bg, cell_txt) por componente en orden de prioridad
# El orden es POSICIONAL contra sin_cubrir de _descomponer_saldo — ver su docstring
_AC_P = [
    ("DEUDA_AGUA",        "EAF2FF", "1A5276", "F4F9FF", "1A5276"),  # P1
    ("CORTE_RECONEXION",  "FEF3C7", "92400E", "FFFBEB", "92400E"),  # P2
    ("CONVENIO",          "E0F2FE", "0C4A6E", "F0F9FF", "0369A1"),  # P3
    ("ACUERDOS_ASAMBLEA", "EDE9FE", "4C1D95", "F5F3FF", "4C1D95"),  # P4
    ("MULTA",             "FEF2F2", "991B1B", "FFF5F5", "991B1B"),  # P5
]
TD_AC_QUIEN = "F0FFF8"
TD_AC_TOTAL = "D5F5E3"
TD_AC_ZERO  = "F3F4F6"

# Paleta arrastre_devolucion (paleta EXCESO — azul)
GH_AV_QUIEN  = ("EFF6FF", "1D4ED8")
GH_AV_MONTO  = ("EFF6FF", "1D4ED8")
GH_AV_TRAZ   = ("EFF6FF", "1D4ED8")
TD_AV_QUIEN  = "F5F9FF"
TD_AV_MONTO  = "F5F9FF"
TD_AV_TRAZ   = "F5F9FF"
GH_AV_REVIS  = ("ECFDF5", "065F46")   # verde — revisión editable por el operador (¿legítimo o error?)
TD_AV_REVIS  = "F9FAFB"               # gris claro — celda vacía esperando input

# Paleta discrepancias_cobranza (coincide con formato_discrepancias_cobranza.html)
GH_DC_PREDIO = ("FEF2F2", "991B1B")   # rojo — el MZ+LT que no existe
GH_DC_PAGO   = ("EBF5FB", "1A5276")   # azul — monto y fecha
GH_DC_ORIGEN = ("FEF9E7", "7D6608")   # ambar — pista de origen (ORIGEN o MESA+COBRADOR)
GH_DC_TRAZ   = ("F4ECF7", "5B21B6")   # morado — ciclo y motivo
GH_DC_CORR   = ("ECFDF5", "065F46")   # verde — corrección editable por el operador
TD_DC_PREDIO = "FFF5F5"
TD_DC_PAGO   = "F4FAFF"
TD_DC_ORIGEN = "FFFDF5"
TD_DC_TRAZ   = "FAF5FF"
TD_DC_CORR   = "ECFDF5"   # verde claro — celda con corrección ingresada
TD_DC_CORR_V = "F9FAFB"   # gris muy claro — celda vacía esperando input


# ── LOGGING ───────────────────────────────────────────────────────────────────
def _init_logging():
    OUTPUTS_DIR.mkdir(exist_ok=True)
    # force=True: si el root logger ya esta configurado (tests, imports previos),
    # basicConfig sin force es no-op y el FileHandler nunca se agrega (metodologia v1.9).
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(OUTPUTS_DIR / "run.log", encoding="utf-8"),
        ],
        force=True,
    )

log = logging.getLogger(__name__)


# ── ESTILO HELPERS ────────────────────────────────────────────────────────────
def _borde():
    b = Side(style="thin", color="CCCCCC")
    return Border(left=b, right=b, top=b, bottom=b)

def _c(ws, row, col, value=None, bg=None, txt="333333",
       bold=False, align="left", mono=False, size=9, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Consolas" if mono else "Arial",
                       size=size, bold=bold, color=txt)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _borde()
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    if fmt:
        c.number_format = fmt
    return c

def _gh(ws, row, cs, ce, texto, bg, txt):
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    c = ws.cell(row=row, column=cs, value=texto)
    c.font      = Font(name="Arial", size=8, bold=True, color=txt)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", start_color=bg)
    c.border    = _borde()

def _ch(ws, row, col, texto, bg, txt):
    _c(ws, row, col, texto, bg=bg, txt=txt, bold=True, align="center")

def _sep(ws, col, row_end):
    letra = get_column_letter(col)
    ws.column_dimensions[letra].width = 0.8
    b_sep = Border(left=Side(style="thin", color="D1D5DB"),
                   right=Side(style="thin", color="D1D5DB"))
    for r in range(1, row_end + 1):
        c = ws.cell(row=r, column=col)
        c.fill   = PatternFill("solid", start_color="F3F4F6")
        c.border = b_sep

def _w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── UTILIDADES ────────────────────────────────────────────────────────────────
def _norm_mz(val) -> str:
    s = str(val).strip().upper()
    return "" if not s or s in ("NAN", "NONE") else s

def _norm_lt(val) -> str:
    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN"):
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.upper().replace(" ", "")

def _float(val) -> float:
    if val is None:
        return 0.0
    try:
        f = float(str(val).replace(",", ".").strip())
        return 0.0 if f != f else f  # NaN guard
    except (ValueError, TypeError):
        return 0.0

def _txt(val) -> str:
    """Texto limpio: '' para None/nan/nat/none (evita 'nan' literal en celdas)."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("nan", "nat", "none") else s

def _ref_efectivo(mesa, cobrador) -> str:
    """REFERENCIA de pago en efectivo: 'MESA / COBRADOR' (omite el lado vacío)."""
    m, c = _txt(mesa), _txt(cobrador)
    if m and c:
        return f"{m} / {c}"
    return m or c

def _norm_cols(df) -> list[str]:
    cols = []
    for i, c in enumerate(df.columns):
        s = str(c).strip().upper()
        if not s or s.startswith("UNNAMED"):
            cols.append(f"_C{i}")
        else:
            cols.append(s)
    return cols

def _estado(saldo: float, total_pagado: float) -> str:
    if saldo < -TOL:
        return "EXCESO"
    if abs(saldo) <= TOL:
        return "CANCELADO"
    return "PARCIAL" if total_pagado > TOL else "PENDIENTE"

def _fecha_str(val) -> str:
    """Convierte fecha a string DD/MM/YYYY. Soporta str, Timestamp, datetime, NaN."""
    if val is None:
        return ""
    if isinstance(val, str):
        s = val.strip()
        if not s or s.upper() in ("NAN", "NAT", "NONE"):
            return ""
        m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
        if m:
            d, mes, a = m.groups()
            return f"{int(d):02d}/{int(mes):02d}/{a}"
        try:
            return pd.to_datetime(s).strftime("%d/%m/%Y")
        except Exception:
            return s[:10]
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    try:
        return pd.to_datetime(val).strftime("%d/%m/%Y")
    except Exception:
        return str(val)[:10]

def _fecha_hora_str(val) -> str:
    """Como _fecha_str pero preserva la hora si existe: DD/MM/YYYY HH:MM.
    Yape trae '21/06/2026 12:08:29' → '21/06/2026 12:08'. Efectivo sin hora → DD/MM/YYYY."""
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    # Timestamp/datetime ya parseado: formatear directo. NO hacer str()+dayfirst,
    # eso reinterpreta '2026-06-05' (5 jun) como 6 may y rompe la idempotencia.
    if isinstance(val, (pd.Timestamp, datetime)):
        dt = val
    else:
        s = str(val).strip()
        if not s or s.upper() in ("NAN", "NAT", "NONE"):
            return ""
        es_iso = bool(re.match(r"\d{4}-\d{1,2}-\d{1,2}", s))
        dt = pd.to_datetime(s, dayfirst=not es_iso, errors="coerce")
        if pd.isna(dt):
            return _fecha_str(val)
    # sin componente horario → solo fecha
    if dt.hour == 0 and dt.minute == 0 and dt.second == 0:
        return dt.strftime("%d/%m/%Y")
    return dt.strftime("%d/%m/%Y %H:%M")


def _fecha_hora_seg_str(val) -> str:
    """Como _fecha_hora_str pero con segundos: DD/MM/YYYY HH:MM:SS.
    Usado en REFERENCIA de arrastre_devolucion — ahí sí importa el segundo exacto."""
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(val, (pd.Timestamp, datetime)):
        dt = val
    else:
        s = str(val).strip()
        if not s or s.upper() in ("NAN", "NAT", "NONE"):
            return ""
        es_iso = bool(re.match(r"\d{4}-\d{1,2}-\d{1,2}", s))
        dt = pd.to_datetime(s, dayfirst=not es_iso, errors="coerce")
        if pd.isna(dt):
            return _fecha_str(val)
    if dt.hour == 0 and dt.minute == 0 and dt.second == 0:
        return dt.strftime("%d/%m/%Y")
    return dt.strftime("%d/%m/%Y %H:%M:%S")

def _fecha_max(fechas: list[str]) -> str:
    """Devuelve la fecha más reciente en formato DD/MM/YYYY."""
    fechas = [f for f in fechas if f]
    if not fechas:
        return ""
    dts = []
    for f in fechas:
        try:
            dts.append(datetime.strptime(f, "%d/%m/%Y"))
        except ValueError:
            pass
    if not dts:
        return fechas[0]
    return max(dts).strftime("%d/%m/%Y")


# ── VALIDACIÓN ────────────────────────────────────────────────────────────────
def _localizar_planilla() -> Path:
    candidatos = sorted(PLAN_DIR.glob("planilla_*.xlsx"))
    if not candidatos:
        raise FileNotFoundError(
            f"Falta planilla_YYYY-MM.xlsx en {PLAN_DIR}\n"
            f"  → Correr 2_planilla (publica la planilla en shared/planilla_mes/)"
        )

    # El ciclo lo declara 1_lecturas (shared/ciclo_activo.json). Elegir
    # "el último por orden alfabético" es lo que dejaba cobrar el mes
    # equivocado cuando conviven planillas de varios meses en la carpeta:
    # el 05/08/2026 había planilla_2026-06, -07 y -08 y una corrida habría
    # cobrado agosto con los pagos de julio.
    mes = ciclo_activo.activo(default=None, path=SHARED_DIR / "ciclo_activo.json")
    if mes is not None:
        esperada = PLAN_DIR / f"planilla_{mes}.xlsx"
        if not esperada.exists():
            raise FileNotFoundError(
                f"El ciclo activo es {mes} pero falta {esperada.name} en {PLAN_DIR}\n"
                f"  → Correr 2_planilla para el ciclo {mes}"
            )
        return esperada

    if len(candidatos) > 1:
        log.warning(f"Sin ciclo activo declarado (shared/ciclo_activo.json) y "
                    f"múltiples planillas en {PLAN_DIR} → usando {candidatos[-1].name}")
    return candidatos[-1]

def _validar_inputs() -> Path:
    plan = _localizar_planilla()
    requeridos = [
        (_pago_path(YAPE_DIR, _YAPE_BASE), "Correr 4_pagos/yape/motor_matching para este ciclo"),
        (_pago_path(EFEC_DIR, _EFEC_BASE), "Correr 4_pagos/efectivo para este ciclo"),
    ]
    errores = []
    for ruta, sug in requeridos:
        if not ruta.exists():
            errores.append(f"Falta: {ruta}\n  → {sug}")
    if errores:
        for e in errores:
            log.error(e)
        raise FileNotFoundError("Inputs faltantes — ver errores arriba")
    log.info(f"Inputs OK · planilla = {plan.name}")
    return plan


# ── CARGA: PENALIDADES (overlay Modelo A) ────────────────────────────────────
def _cargar_penalidades(mes_ano: str) -> dict[tuple[str, str], float]:
    """Neto de penalidad de corte por predio, sumado de los audits de 6_corte y 6b.

    Fuente única de la penalidad: sus audit_*.xlsx. Se suma PENALIDAD_APLICADA
    (col 5, ya trae signo: +APLICADO / −REVERTIDO) → net delta. Robusto a que la
    base cambie: no depende del absoluto (col 6). Solo el ciclo MES_ANO actual.
    """
    neto: dict[tuple[str, str], float] = {}
    for p in (AUDIT_CORTE_PATH, AUDIT_MULTA_PATH):
        if not p.exists():
            continue
        df = pd.read_excel(p, header=1)
        df.columns = _norm_cols(df)
        if "PENALIDAD_APLICADA" not in df.columns:
            continue
        for _, f in df.iterrows():
            if str(f.get("MES_ANO", "")).strip() != mes_ano:
                continue
            mz = _norm_mz(f.get("MZ"))
            lt = _norm_lt(f.get("LT"))
            if not mz or not lt:
                continue
            key = (mz, lt)
            neto[key] = round(neto.get(key, 0.0) + _float(f.get("PENALIDAD_APLICADA")), 2)
    return {k: v for k, v in neto.items() if abs(v) > TOL}


_CONCEPTO_DEVOLUCION_A_CAMPO = {
    "CONVENIO":          "convenio",
    "MULTA":             "multa",
    "ACUERDOS":          "acuerdos_asamblea",
    "ACUERDOS_ASAMBLEA": "acuerdos_asamblea",
    "CORTE_RECONEXION":  "corte_reconexion",
    "AGUA":              "mes_actual",
    "MANTENIMIENTO":     "mantenimiento",
    "MES_ANTERIOR":      "mes_anterior",
}


def _cargar_devoluciones_aplicadas(mes_ano: str) -> dict[tuple[str, str], list[tuple[str, float]]]:
    """Créditos de exceso de ciclos anteriores que se decidió aplicar a un
    concepto de deuda en vez de esperar reclamo/devolución (ver
    shared/devoluciones_aplicadas.xlsx). Filtra por MES_ANO_APLICA: cada
    crédito baja el saldo solo en el ciclo al que se decidió aplicarlo, no en
    todos los ciclos futuros — evita doble-aplicación en corridas siguientes.
    """
    if not DEVOLUCIONES_APLICADAS_PATH.exists():
        return {}
    df = pd.read_excel(DEVOLUCIONES_APLICADAS_PATH, header=1)
    df.columns = _norm_cols(df)
    por_predio: dict[tuple[str, str], list[tuple[str, float]]] = {}
    for _, f in df.iterrows():
        if str(f.get("MES_ANO_APLICA", "")).strip() != mes_ano:
            continue
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        campo = _CONCEPTO_DEVOLUCION_A_CAMPO.get(str(f.get("CONCEPTO", "")).strip().upper())
        monto = _float(f.get("MONTO"))
        if not mz or not lt or not campo or monto <= TOL:
            continue
        por_predio.setdefault((mz, lt), []).append((campo, monto))
    return por_predio


def _cargar_ajustes_cargo(mes_ano: str) -> dict[tuple[str, str], list[tuple[str, float]]]:
    """CARGOs que nacieron (ver shared/ajustes_cargo.xlsx) y después se
    determinó que no correspondían — se anulan. Filtra por MES_ANO_APLICA:
    el ajuste se hace una sola vez, no en cada ciclo futuro. A diferencia de
    devoluciones_aplicadas, acá no hubo plata de más — el cargo mismo se resta.
    """
    if not AJUSTES_CARGO_PATH.exists():
        return {}
    df = pd.read_excel(AJUSTES_CARGO_PATH, header=1)
    df.columns = _norm_cols(df)
    por_predio: dict[tuple[str, str], list[tuple[str, float]]] = {}
    for _, f in df.iterrows():
        if str(f.get("MES_ANO_APLICA", "")).strip() != mes_ano:
            continue
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        campo = _CONCEPTO_DEVOLUCION_A_CAMPO.get(str(f.get("CONCEPTO", "")).strip().upper())
        monto = _float(f.get("MONTO"))
        if not mz or not lt or not campo or monto <= TOL:
            continue
        por_predio.setdefault((mz, lt), []).append((campo, monto))
    return por_predio


def _cargar_genesis_tardia(mes_ano: str) -> dict[tuple[str, str], list[tuple[str, float]]]:
    """CARGOs legítimos (ver shared/genesis_tardia.xlsx) cuya fuente en
    obligaciones/ se corrigió después de que el ciclo ya estaba congelado —
    se suman a la génesis de este ciclo. Filtra por MES_ANO_APLICA: se aplica
    una sola vez; el siguiente ciclo ya lee la fuente arreglada directo.
    """
    if not GENESIS_TARDIA_PATH.exists():
        return {}
    df = pd.read_excel(GENESIS_TARDIA_PATH, header=1)
    df.columns = _norm_cols(df)
    por_predio: dict[tuple[str, str], list[tuple[str, float]]] = {}
    for _, f in df.iterrows():
        if str(f.get("MES_ANO_APLICA", "")).strip() != mes_ano:
            continue
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        campo = _CONCEPTO_DEVOLUCION_A_CAMPO.get(str(f.get("CONCEPTO", "")).strip().upper())
        monto = _float(f.get("MONTO"))
        if not mz or not lt or not campo or monto <= TOL:
            continue
        por_predio.setdefault((mz, lt), []).append((campo, monto))
    return por_predio


# Un pago de ciclo cerrado cubre primero la deuda anterior y los conceptos que
# quedaron pendientes; si sobra, continúa con el consumo del ciclo vigente.
_CAMPOS_WATERFALL_REIDENTIFICACION = (
    "mes_anterior", "corte_reconexion", "convenio", "acuerdos_asamblea", "multa",
    "mes_actual", "mantenimiento",
)

# Cascada COMPLETA (P1 entero + P2-P5) para el crédito cuyo MES_CICLO es el ciclo
# en curso. Un abono rezagado NO es "plata de un ciclo anterior": es plata real que
# no puede entrar por la mesa porque llegó fuera de la ventana del registro manual
# — y eso pasa también días después del cobro, dentro del mismo ciclo. Con la tupla
# de arriba esa plata no encontraba balde donde caer (mes_actual y mantenimiento no
# están) y el sobrante se descartaba en silencio: 5 predios, S/99, el 14/08/2026.
# Los 3 campos de agua viven todos en P1 de _descomponer_saldo, así que repartir
# entre ellos no cambia ningún total ni cruza al ledger de pueblo.
_CAMPOS_WATERFALL_CICLO_VIGENTE = (
    "mes_anterior", "mes_actual", "mantenimiento",
    "corte_reconexion", "convenio", "acuerdos_asamblea", "multa",
)


def _aplicar_waterfall(u: dict, monto: float, campos: tuple[str, ...]) -> float:
    """Descuenta `monto` de los campos de deuda de `u` en orden. Devuelve lo que
    sobró (0.0 si se aplicó todo). El sobrante es plata real sin deuda contra la
    cual imputarse — quien llama decide qué hacer, pero nunca lo ignora en silencio."""
    restante = monto
    for c in campos:
        if restante <= TOL:
            break
        aplicar = min(u[c], restante)
        if aplicar > TOL:
            u[c] = round(u[c] - aplicar, 2)
            restante = round(restante - aplicar, 2)
    return restante


def _cargar_reidentificaciones(mes_ano: str) -> dict[tuple[str, str], list[tuple[float, str | None]]]:
    """Pagos de un ciclo anterior anotados en el predio equivocado (typo o
    confusión de nombre en la mesa) — se acreditan al predio real. Filtra por
    MES_ANO_APLICA: se aplica una sola vez, no en cada ciclo futuro. Si
    CONCEPTO_DESTINO viene vacío, se reparte en cascada (ver
    _CAMPOS_WATERFALL_REIDENTIFICACION) en vez de a un solo concepto.
    """
    if not REIDENTIFICACION_PATH.exists():
        return {}
    df = pd.read_excel(REIDENTIFICACION_PATH, header=1)
    df.columns = _norm_cols(df)
    por_predio: dict[tuple[str, str], list[tuple[float, str | None]]] = {}
    for _, f in df.iterrows():
        if str(f.get("MES_ANO_APLICA", "")).strip() != mes_ano:
            continue
        mz = _norm_mz(f.get("MZ_CORRECTO"))
        lt = _norm_lt(f.get("LT_CORRECTO"))
        monto = _float(f.get("MONTO"))
        if not mz or not lt or monto <= TOL:
            continue
        concepto = str(f.get("CONCEPTO_DESTINO", "")).strip().upper()
        campo = _CONCEPTO_DEVOLUCION_A_CAMPO.get(concepto) if concepto and concepto not in ("NAN", "NONE") else None
        por_predio.setdefault((mz, lt), []).append((monto, campo))
    return por_predio


def _cargar_correcciones_deuda(mes_ano: str) -> dict[tuple[str, str], list[tuple[str, float]]]:
    """Deuda real que hay que devolverle al lote de origen de una
    reidentificación (ver DEUDA_CORRECCIONES_PATH) — el pago mal atribuido sí
    había pagado deuda real suya, no solo generado exceso. Filtra por
    MES_ANO_APLICA, normalmente distinto (posterior) al de la reidentificación
    hermana, para no afectar boletas de un ciclo ya impreso.
    """
    if not DEUDA_CORRECCIONES_PATH.exists():
        return {}
    df = pd.read_excel(DEUDA_CORRECCIONES_PATH, header=1)
    df.columns = _norm_cols(df)
    por_predio: dict[tuple[str, str], list[tuple[str, float]]] = {}
    for _, f in df.iterrows():
        if str(f.get("MES_ANO_APLICA", "")).strip() != mes_ano:
            continue
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        campo = str(f.get("CAMPO", "")).strip().lower()
        monto = _float(f.get("MONTO"))
        if not mz or not lt or not campo or monto <= TOL:
            continue
        por_predio.setdefault((mz, lt), []).append((campo, monto))
    return por_predio


def _abono_manifest_key(row) -> tuple[str, str, float, str, str]:
    return (
        _norm_mz(row.get("MZ")),
        _norm_lt(row.get("LT")),
        round(_float(row.get("MONTO")), 2),
        str(row.get("MES_CICLO", "")).strip()[:7],
        str(row.get("MES_ANO_APLICA", "")).strip()[:7],
    )


def _validar_abonos_manifest(df: pd.DataFrame, mes_ano: str) -> pd.DataFrame:
    """Fail closed if the active source differs from the approved manifest."""
    if not ABONOS_MANIFEST_PATH.exists():
        raise RuntimeError(f"Falta manifest de abonos: {ABONOS_MANIFEST_PATH}")
    manifest = json.loads(ABONOS_MANIFEST_PATH.read_text(encoding="utf-8"))
    expected = {
        _abono_manifest_key(row): row.get("ESTADO", "CONFIRMADO")
        for row in manifest
        if str(row.get("MES_ANO_APLICA", "")).strip()[:7] == mes_ano
    }
    active = df[df["MES_ANO_APLICA"].astype(str).str.strip().str[:7] == mes_ano].copy()
    keys = [_abono_manifest_key(row) for row in active.to_dict("records")]
    counts = pd.Series(keys).value_counts() if keys else pd.Series(dtype="int64")
    unexpected = [key for key in keys if key not in expected]
    duplicated = [key for key, count in counts.items() if count > 1]
    missing = [key for key, status in expected.items() if status == "CONFIRMADO" and key not in counts]
    blocked = [key for key in keys if expected.get(key) == "BLOQUEADO"]
    if unexpected or duplicated or missing or blocked:
        raise RuntimeError(
            "Manifest de abonos no coincide: "
            f"inesperados={unexpected} duplicados={duplicated} "
            f"faltantes={missing} bloqueados={blocked}"
        )
    confirmed = active[[key in expected and expected[key] == "CONFIRMADO" for key in keys]]
    log.info(f"Manifest abonos OK · {len(confirmed)} fila(s) confirmada(s) para {mes_ano}")
    return confirmed


def _cargar_abonos_rezagados(mes_ano: str) -> dict[tuple[str, str], tuple[float, float]]:
    """Pagos normales que llegaron fuera de la ventana de cobro.

    En esta etapa solo entran filas BALDE=agua; los destinos especiales se
    mantienen fuera hasta resolver su overlay propio. Filtra por
    MES_ANO_APLICA: se aplica una sola vez.

    Devuelve {(mz, lt): (monto_ciclo_cerrado, monto_ciclo_vigente)} — la fila declara
    a qué ciclo pertenece en MES_CICLO y eso decide con qué cascada baja:

        MES_CICLO <  mes_ano   deuda de un ciclo cerrado → _CAMPOS_WATERFALL_REIDENTIFICACION
                               (sin mes_actual/mantenimiento: no cancela el consumo vigente)
        MES_CICLO == mes_ano   pago del ciclo en curso    → _CAMPOS_WATERFALL_CICLO_VIGENTE
                               (P1 completo, igual que un pago de mesa)

    MES_CICLO vacío o ilegible cae en "cerrado" — el comportamiento previo.
    El [:7] tolera que Excel guarde la celda como fecha ("2026-08-01 00:00:00").
    Suma por predio y por ciclo (un predio puede tener >1 abono, de ciclos distintos).
    """
    if not ABONOS_REZAGADOS_PATH.exists():
        return {}
    df = pd.read_excel(ABONOS_REZAGADOS_PATH, header=1)
    df.columns = _norm_cols(df)
    df = _validar_abonos_manifest(df, mes_ano)
    por_predio: dict[tuple[str, str], tuple[float, float]] = {}
    for _, f in df.iterrows():
        if str(f.get("MES_ANO_APLICA", "")).strip() != mes_ano:
            continue
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        monto = _float(f.get("MONTO"))
        if not mz or not lt or monto <= TOL:
            continue
        vigente = str(f.get("MES_CICLO", "")).strip()[:7] == mes_ano
        cerr, vig = por_predio.get((mz, lt), (0.0, 0.0))
        if vigente:
            por_predio[(mz, lt)] = (cerr, round(vig + monto, 2))
        else:
            por_predio[(mz, lt)] = (round(cerr + monto, 2), vig)
    return por_predio


def _cargar_blancos_efectivo(mes_ano: str) -> dict[tuple[str, str], float]:
    """Pagos en efectivo de un ciclo anterior que entraron a la caja como
    BLANCO (bug B6) y ya se identificó a qué predio pertenecen (columnas
    MZ/LT llenas). Filtra por MES_ANO_APLICA: se aplica una sola vez. Igual
    tratamiento que abonos_rezagados: salda deuda del ciclo viejo, cascada sin
    mes_actual/mantenimiento. Filas sin MZ/LT (aún sin identificar) se
    ignoran — no tienen predio al cual aplicarse.
    """
    if not BLANCOS_EFECTIVO_PATH.exists():
        return {}
    df = pd.read_excel(BLANCOS_EFECTIVO_PATH, header=1)
    df.columns = _norm_cols(df)
    por_predio: dict[tuple[str, str], float] = {}
    for _, f in df.iterrows():
        if str(f.get("MES_ANO_APLICA", "")).strip() != mes_ano:
            continue
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        monto = _float(f.get("MONTO"))
        if not mz or not lt or monto <= TOL:
            continue
        por_predio[(mz, lt)] = round(por_predio.get((mz, lt), 0.0) + monto, 2)
    return por_predio


def _cargar_reasignaciones_aplicacion(mes_ano: str) -> dict[tuple[str, str], list[tuple[str, str, float]]]:
    """Abonos que la cascada P1-P6 aplicaría de oficio a un CARGO distinto al
    que el pagador especificó (ver REASIGNACIONES_APLICACION_PATH). No toca
    _descomponer_saldo (el CARGO de origen sigue abierto en la boleta/deuda);
    solo redirige, en la reconciliación hacia seguimiento_pueblo, qué CONCEPTO
    recibe el PAGO. Filtra por MES_ANO — vive en el mismo ciclo del pago
    (a diferencia de devoluciones_aplicadas, que cruza ciclos).
    """
    if not REASIGNACIONES_APLICACION_PATH.exists():
        return {}
    df = pd.read_excel(REASIGNACIONES_APLICACION_PATH, header=1)
    df.columns = _norm_cols(df)
    por_predio: dict[tuple[str, str], list[tuple[str, str, float]]] = {}
    for _, f in df.iterrows():
        if str(f.get("MES_ANO", "")).strip() != mes_ano:
            continue
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        origen = str(f.get("CONCEPTO_ORIGEN", "")).strip().upper()
        destino = str(f.get("CONCEPTO_DESTINO", "")).strip().upper()
        monto = _float(f.get("MONTO"))
        if not mz or not lt or not origen or not destino or monto <= TOL:
            continue
        por_predio.setdefault((mz, lt), []).append((origen, destino, monto))
    return por_predio


def _cargar_aportes_tanque_manuales(mes_ano: str) -> dict[str, float]:
    """Aportes voluntarios al tanque comunitario (ver
    APORTES_TANQUE_MANUALES_PATH) que llegaron mezclados con el pago de deuda
    normal — el pagador los distinguió en su mensaje ("tanque"/"tanke
    adelanto") pero el motor no separa MONTO_PAGO por concepto, así que ese
    dinero entraba entero a total_pagado y la cascada P1-P6 lo consumía como
    si fuera agua/corte/multa/acuerdos/convenio. Se resta ANTES de la cascada
    (ver _calcular) para que no se cuente dos veces: una como aporte
    voluntario (consolidar_tanque.py ya lee este mismo archivo) y otra como
    pago de deuda. Filtra por MES_ANO_APLICA, solo filas BALDE=tanque.
    Clave 'MZ-LT' (igual que _cargar_blancos), no tupla.
    """
    if not APORTES_TANQUE_MANUALES_PATH.exists():
        return {}
    df = pd.read_excel(APORTES_TANQUE_MANUALES_PATH, header=1)
    df.columns = _norm_cols(df)
    por_predio: dict[str, float] = {}
    for _, f in df.iterrows():
        if str(f.get("MES_ANO_APLICA", "")).strip() != mes_ano:
            continue
        if str(f.get("BALDE", "")).strip().lower() != "tanque":
            continue
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        monto = _float(f.get("MONTO"))
        if not mz or not lt or monto <= TOL:
            continue
        key = f"{mz}-{lt}"
        por_predio[key] = round(por_predio.get(key, 0.0) + monto, 2)
    return por_predio


# ── CARGA: PLANILLA ──────────────────────────────────────────────────────────
def _cargar_planilla(plan_path: Path) -> tuple[list[dict], str]:
    df = pd.read_excel(plan_path, header=1)
    df.columns = _norm_cols(df)
    requeridas = {"MZ", "LT", "NOMBRE", "MES_ANO", "MARC_ANT", "MARC_ACT", "M3",
                  "MES_ACTUAL", "MANTENIMIENTO", "MES_ANTERIOR", "CORTE_RECONEXION",
                  "CONVENIO", "MULTA", "ACUERDOS_ASAMBLEA",
                  "BLANCO", "DEVOLUCION", "TOTAL_A_PAGAR"}
    faltantes = requeridas - set(df.columns)
    if faltantes:
        raise ValueError(f"Planilla — columnas faltantes: {sorted(faltantes)}")

    usuarios, mes_ano = [], ""
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        if not mes_ano:
            mes_ano = str(f.get("MES_ANO", "")).strip()
        usuarios.append({
            "mz":       mz,
            "lt":       lt,
            "key":      f"{mz}-{lt}",
            "nombre":   str(f.get("NOMBRE", "")).strip(),
            "mes_ano":  str(f.get("MES_ANO", "")).strip(),
            "marc_ant": _float(f.get("MARC_ANT")),
            "marc_act": _float(f.get("MARC_ACT")),
            "m3":       _float(f.get("M3")),
            "mes_actual":        _float(f.get("MES_ACTUAL")),
            "mantenimiento":     _float(f.get("MANTENIMIENTO")),
            "mes_anterior":      _float(f.get("MES_ANTERIOR")),
            "corte_reconexion":  _float(f.get("CORTE_RECONEXION")),
            "corte_reconexion_base": _float(f.get("CORTE_RECONEXION")),
            "convenio":          _float(f.get("CONVENIO")),
            "multa":             _float(f.get("MULTA")),
            "acuerdos_asamblea": _float(f.get("ACUERDOS_ASAMBLEA")),
            "blanco_inicial":    _float(f.get("BLANCO")),
            "devolucion":        _float(f.get("DEVOLUCION")),
        })
    # Overlay Modelo A: sumar penalidad de corte (audits de 6_corte/6b) sobre la
    # base leída de shared. shared ya no la trae — se re-deriva acá en cada corrida.
    penalidades = _cargar_penalidades(mes_ano)
    if penalidades:
        n = 0
        for u in usuarios:
            d = penalidades.get((u["mz"], u["lt"]), 0.0)
            if abs(d) > TOL:
                u["corte_reconexion"] = round(u["corte_reconexion"] + d, 2)
                n += 1
        log.info(f"Overlay penalidad · {n} predios con delta de audit (CORTE_RECONEXION)")

    devoluciones = _cargar_devoluciones_aplicadas(mes_ano)
    if devoluciones:
        n = 0
        for u in usuarios:
            for campo, monto in devoluciones.get((u["mz"], u["lt"]), []):
                u[campo] = round(u[campo] - monto, 2)
                n += 1
        log.info(f"Overlay devoluciones aplicadas · {n} crédito(s) aplicado(s)")

    ajustes_cargo = _cargar_ajustes_cargo(mes_ano)
    if ajustes_cargo:
        n = 0
        for u in usuarios:
            for campo, monto in ajustes_cargo.get((u["mz"], u["lt"]), []):
                u[campo] = round(u[campo] - monto, 2)
                n += 1
        log.info(f"Overlay ajustes de cargo · {n} ajuste(s) aplicado(s)")

    reidentificaciones = _cargar_reidentificaciones(mes_ano)
    if reidentificaciones:
        n = 0
        for u in usuarios:
            for monto, campo in reidentificaciones.get((u["mz"], u["lt"]), []):
                if campo:
                    u[campo] = round(u[campo] - monto, 2)
                else:
                    restante = monto
                    for c in _CAMPOS_WATERFALL_REIDENTIFICACION:
                        if restante <= TOL:
                            break
                        aplicar = min(u[c], restante)
                        if aplicar > TOL:
                            u[c] = round(u[c] - aplicar, 2)
                            restante = round(restante - aplicar, 2)
                n += 1
        log.info(f"Overlay reidentificación · {n} crédito(s) aplicado(s)")

    correcciones_deuda = _cargar_correcciones_deuda(mes_ano)
    if correcciones_deuda:
        n = 0
        for u in usuarios:
            for campo, monto in correcciones_deuda.get((u["mz"], u["lt"]), []):
                u[campo] = round(u[campo] + monto, 2)
                n += 1
        log.info(f"Overlay corrección de deuda · {n} ajuste(s) aplicado(s)")

    genesis_tardia = _cargar_genesis_tardia(mes_ano)
    if genesis_tardia:
        n = 0
        for u in usuarios:
            for campo, monto in genesis_tardia.get((u["mz"], u["lt"]), []):
                u[campo] = round(u[campo] + monto, 2)
                n += 1
        log.info(f"Overlay génesis tardía · {n} cargo(s) sembrado(s)")

    blancos_efectivo = _cargar_blancos_efectivo(mes_ano)
    if blancos_efectivo:
        n = 0
        for u in usuarios:
            monto = blancos_efectivo.get((u["mz"], u["lt"]), 0.0)
            if monto <= TOL:
                continue
            restante = monto
            for c in _CAMPOS_WATERFALL_REIDENTIFICACION:
                if restante <= TOL:
                    break
                aplicar = min(u[c], restante)
                if aplicar > TOL:
                    u[c] = round(u[c] - aplicar, 2)
                    restante = round(restante - aplicar, 2)
            n += 1
        log.info(f"Overlay blancos efectivo · {n} crédito(s) aplicado(s)")

    log.info(f"Planilla {mes_ano} → {len(usuarios)} usuarios")
    return usuarios, mes_ano


# ── CARGA: PAGOS YAPE ────────────────────────────────────────────────────────
def _cargar_pagos_yape() -> list[dict]:
    df = pd.read_excel(_pago_path(YAPE_DIR, _YAPE_BASE), header=1)
    df.columns = _norm_cols(df)
    ciclo_col = "CICLO_CORRECCION" if "CICLO_CORRECCION" in df.columns else "CICLO"

    pagos, sin_id = [], 0
    for idx, f in df.iterrows():
        if str(f.get("TIPO", "")).strip().upper() != "TE PAGÓ":
            continue
        mz   = _norm_mz(f.get("MZ"))
        lt   = _norm_lt(f.get("LOTE"))
        conc = str(f.get("CONCEPTO", "")).strip().upper()
        # Pago con CONCEPTO (tanque, honorario, gasto...) NO es pago de agua —
        # 5_cobranza solo procesa agua. Lo maneja consolidar_tanque por separado.
        if conc and conc not in ("NAN", "NONE"):
            continue
        if not mz or not lt:
            sin_id += 1
            continue
        monto = _float(f.get("MONTO_ASIGNADO"))
        if monto <= TOL:
            monto = _float(f.get("MONTO_PAGO"))
        if monto <= TOL:
            continue
        concepto = conc.lower() if conc and conc not in ("NAN", "NONE") else ""
        pagos.append({
            "row":    idx + 3,                # fila Excel (filas 1-2 son cabeceras)
            "mz":     mz,
            "lt":     lt,
            "key":    f"{mz}-{lt}",
            "nombre": str(f.get("NOMBRE", "")).strip(),
            "monto":  round(monto, 2),
            "fecha":  _fecha_str(f.get("FECHA")),
            "ciclo_correccion": int(_float(f.get(ciclo_col)) or 1),
            "fuente": "yape",
            "concepto": concepto,
            # ORIGEN: nombre que el banco asigna al remitente (ej: "Wilder Tru*").
            # Sirve para rastrear pagos huérfanos en discrepancias_cobranza.xlsx.
            "origen": str(f.get("ORIGEN", "")).strip(),
            # Trazabilidad para auditoría (trazabilidad_cobranza.xlsx):
            # REFERENCIA = ORIGEN del banco · COMENTARIO = MENSAJE del pagador.
            "referencia": str(f.get("ORIGEN", "")).strip(),
            "comentario": str(f.get("MENSAJE", "")).strip(),
            "fecha_hora": _fecha_hora_str(f.get("FECHA")),
            "fecha_hora_seg": _fecha_hora_seg_str(f.get("FECHA")),
        })
    log.info(f"Pagos Yape → {len(pagos)} filas · {sin_id} sin identificar")
    return pagos


# ── CARGA: PAGOS EFECTIVO ────────────────────────────────────────────────────
def _cargar_pagos_efectivo() -> list[dict]:
    df = pd.read_excel(_pago_path(EFEC_DIR, _EFEC_BASE), header=1)
    df.columns = _norm_cols(df)
    ciclo_col = "CICLO_CORRECCION" if "CICLO_CORRECCION" in df.columns else "CICLO"

    pagos = []
    for idx, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        monto = _float(f.get("MONTO"))
        if monto <= TOL:
            continue
        concepto = str(f.get("CONCEPTO") or "").strip().lower()
        if concepto.upper() in ("NAN", "NONE"):
            concepto = ""
        # Pago con CONCEPTO (tanque, honorario, gasto...) NO es pago de agua —
        # 5_cobranza solo procesa agua. Lo maneja consolidar_tanque por separado.
        if concepto:
            continue
        pagos.append({
            "row":    idx + 3,
            "mz":     mz,
            "lt":     lt,
            "key":    f"{mz}-{lt}",
            "nombre": "",
            "monto":  round(monto, 2),
            "fecha":  _fecha_str(f.get("FECHA")),
            "ciclo_correccion": int(_float(f.get(ciclo_col)) or 1),
            "fuente": "efectivo",
            "concepto": concepto,
            # MESA + COBRADOR: pista física para rastrear pagos huérfanos.
            # MESA dice en cuál mesa_N.xlsx buscar, COBRADOR quién lo registró.
            "mesa":     str(f.get("MESA", "")).strip(),
            "cobrador": str(f.get("COBRADOR", "")).strip(),
            # Trazabilidad para auditoría (trazabilidad_cobranza.xlsx):
            # REFERENCIA = "MESA / COBRADOR" · COMENTARIO = nota del cobrador.
            "referencia": _ref_efectivo(f.get("MESA"), f.get("COBRADOR")),
            "comentario": _txt(f.get("COMENTARIO")),
            "fecha_hora": _fecha_hora_str(f.get("FECHA")),
        })
    log.info(f"Pagos Efectivo → {len(pagos)} filas")
    return pagos


# ── CARGA: BLANCOS ───────────────────────────────────────────────────────────
def _cargar_blancos(mes_ano: str) -> dict:
    """Blancos identificados (MZ/LOTE llenos) pendientes de aplicar. Un blanco
    marcado ESTADO=aplicado se salta SALVO que se haya aplicado en este MISMO
    mes_ano — planilla_cobrado se regenera entera cada corrida, así que
    re-correr el mismo ciclo no puede perder un crédito ya usado ese ciclo
    (solo se descarta si quedó aplicado en un ciclo DISTINTO, para no
    duplicarlo hacia adelante)."""
    if not BLANCOS_PATH.exists():
        log.info("blancos_acumulados.xlsx no encontrado → sin blancos")
        return {}
    df = pd.read_excel(BLANCOS_PATH, header=1)
    df.columns = _norm_cols(df)
    blancos = {}
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LOTE") if "LOTE" in df.columns else f.get("LT"))
        if not mz or not lt:
            continue
        est = str(f.get("ESTADO", "")).strip().lower()
        mes_aplicado = str(f.get("MES_ANO_APLICADO", "")).strip()
        if est == "aplicado" and mes_aplicado != mes_ano:
            continue
        key = f"{mz}-{lt}"
        blancos[key] = blancos.get(key, 0.0) + _float(f.get("MONTO"))
    if blancos:
        log.info(f"Blancos pendientes → {len(blancos)} lotes")
    return blancos


# ── CARGA: RETORNOS (yape + efectivo) ────────────────────────────────────────
# Los retornos reducen MONTO_YAPE (cualquier medio de devolucion).
# El badge RETORNO en planilla_cobrado y trazabilidad apunta al archivo origen.
# Ambos archivos son opcionales — si no existen, no hay retornos este ciclo.

def _cargar_retornos_yape() -> dict:
    """Retorna {key MZ-LT: monto_total}. Lee pagos_yape_retorno.xlsx. Archivo opcional."""
    path = _pago_path(YAPE_DIR, _YAPE_RETORNO_BASE)
    if not path.exists():
        log.info(f"{path.name} no encontrado → sin retornos Yape")
        return {}
    df = pd.read_excel(path, header=1)
    df.columns = _norm_cols(df)
    devs = {}
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LOTE"))
        if not mz or not lt:
            continue
        monto = _float(f.get("MONTO"))
        if monto <= TOL:
            continue
        k = f"{mz}-{lt}"
        devs[k] = devs.get(k, 0.0) + round(monto, 2)
    if devs:
        log.info(f"Retornos Yape → {len(devs)} lotes · S/ {sum(devs.values()):.2f}")
    return devs


def _cargar_devueltos_yape() -> dict:
    """Retorna {key MZ-LT: monto_total}. Lee pagos_yape_devolucion.xlsx. Archivo opcional."""
    path = _pago_path(YAPE_DIR, _YAPE_DEV_BASE)
    if not path.exists():
        log.info(f"{path.name} no encontrado → sin devoluciones Yape")
        return {}
    df = pd.read_excel(path, header=1)
    df.columns = _norm_cols(df)
    devs = {}
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LOTE"))
        if not mz or not lt:
            continue
        monto = _float(f.get("MONTO"))
        if monto <= TOL:
            continue
        k = f"{mz}-{lt}"
        devs[k] = devs.get(k, 0.0) + round(monto, 2)
    if devs:
        log.info(f"Devoluciones Yape → {len(devs)} lotes · S/ {sum(devs.values()):.2f}")
    return devs


def _cargar_retornos_efectivo() -> dict:
    """Retorna {key MZ-LT: monto_total}. Archivo opcional."""
    path = _pago_path(EFEC_DIR, _EFEC_DEV_BASE)
    if not path.exists():
        log.info(f"{path.name} no encontrado → sin retornos Efectivo")
        return {}
    df = pd.read_excel(path, header=1)
    df.columns = _norm_cols(df)
    devs = {}
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LOTE"))
        if not mz or not lt:
            continue
        monto = _float(f.get("MONTO"))
        if monto <= TOL:
            continue
        k = f"{mz}-{lt}"
        devs[k] = devs.get(k, 0.0) + round(monto, 2)
    if devs:
        log.info(f"Retornos Efectivo → {len(devs)} lotes · S/ {sum(devs.values()):.2f}")
    return devs


def _retorno_badge(yape_dev: float, efec_dev: float) -> str | None:
    """Calcula el badge RETORNO para un lote a partir de sus retornos."""
    has_y = yape_dev > TOL
    has_e = efec_dev > TOL
    if has_y and has_e:
        return "mixto"
    if has_y:
        return "yape"
    if has_e:
        return "efectivo"
    return None


def _retornos_por_lote(dev_yape: dict, dev_efec: dict) -> dict:
    """Combina los retornos yape y efectivo en {key: badge}. Solo incluye lotes con retorno."""
    keys = set(dev_yape) | set(dev_efec)
    out = {}
    for k in keys:
        badge = _retorno_badge(dev_yape.get(k, 0.0), dev_efec.get(k, 0.0))
        if badge:
            out[k] = badge
    return out


def _devueltos_por_lote(dev_devuelto: dict) -> dict:
    """Devuelve {key: 'yape'} para lotes con devolución Yape. Solo incluye lotes con devuelto."""
    return {k: "yape" for k, v in dev_devuelto.items() if v > TOL}


def _retornos_planilla_previa(mes_ano: str) -> dict:
    """Lee el estado RETORNO de la planilla_cobrado anterior — para detectar cambios e idempotencia."""
    path = ciclo_activo.resolver(
        OUTPUTS_DIR, "planilla_cobrado", mes_ano,
        legacy_sin_periodo=ciclo_activo.acepta_legacy(mes_ano),
    )
    if not path.exists():
        return {}
    try:
        df = pd.read_excel(path, header=1)
        df.columns = _norm_cols(df)
    except Exception:
        return {}
    if "RETORNO" not in df.columns:
        return {}
    out = {}
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        val = f.get("RETORNO")
        if val is None:
            continue
        badge = str(val).strip().lower()
        if badge in ("yape", "efectivo", "mixto"):
            out[f"{mz}-{lt}"] = badge
    return out


# ── CARGA: TRAZABILIDAD PREVIA (para idempotencia + ciclo) ───────────────────
def _identidad_pago(p: dict) -> tuple:
    # Usa el lote original como identidad para que la idempotencia funcione
    # aunque en runs futuros el key se remapee a otro lote vía correcciones_lote.
    return (p.get("mz_origen", p["mz"]),
            p.get("lt_origen", p["lt"]),
            p["monto"], p["fuente"],
            p["fecha"], p["ciclo_correccion"])

def _cargar_trazabilidad_previa(pagos_actuales: list[dict] | None = None) -> tuple[set, int]:
    """Retorna (set de identidades ya cargadas, max CICLO_COBRANZA usado)."""
    p = OUTPUTS_DIR / "trazabilidad_cobranza.xlsx"
    if not p.exists():
        return set(), 0
    df = pd.read_excel(p, header=1)
    df.columns = _norm_cols(df)
    fechas_fuente: dict[tuple, set[str]] = {}
    for pago in pagos_actuales or []:
        base = (pago.get("mz_origen", pago["mz"]),
                pago.get("lt_origen", pago["lt"]),
                pago["monto"], pago["fuente"], pago["ciclo_correccion"])
        fechas_fuente.setdefault(base, set()).add(pago["fecha"])

    ids, mx = set(), 0
    for _, f in df.iterrows():
        # Si el pago fue corregido, MZ_ORIGEN/LT_ORIGEN guardan la identidad real
        # (lo que el cobrador escribió). Usar eso para que la idempotencia coincida.
        mz = _norm_mz(f.get("MZ_ORIGEN")) or _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT_ORIGEN")) or _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        monto = round(_float(f.get("MONTO")), 2)
        fuente = str(f.get("FUENTE", "")).strip().lower()
        ciclo_origen = int(_float(f.get("CICLO_CORRECCION_ORIGEN")) or 0)
        fecha = _fecha_str(f.get("FECHA")) if "FECHA" in df.columns else ""
        candidatas = fechas_fuente.get((mz, lt, monto, fuente, ciclo_origen), set())
        if len(candidatas) == 1:
            fecha = next(iter(candidatas))
        ident = (mz, lt, monto, fuente, fecha, ciclo_origen)
        ids.add(ident)
        mx = max(mx, int(_float(f.get("CICLO_COBRANZA")) or 0))
    return ids, mx


# ── CORRECCIONES DE LOTE ─────────────────────────────────────────────────────
# El operador llena MZ_CORRECTO + LT_CORRECTO en discrepancias_cobranza.xlsx.
# El módulo las absorbe, las persiste en inputs/correcciones_lote.xlsx y las
# aplica a todos los pagos antes del matching — el registro de origen no se toca.

def _leer_correcciones() -> dict:
    """Lee correcciones persistidas en correcciones_lote.xlsx.
    Retorna {(mz_origen, lt_origen): (mz_destino, lt_destino)}.
    """
    if not CORR_LOTE_PATH.exists():
        return {}
    df = pd.read_excel(CORR_LOTE_PATH, header=0)
    df.columns = _norm_cols(df)
    corr = {}
    for _, f in df.iterrows():
        mo = _norm_mz(f.get("MZ_ORIGEN"))
        lo = _norm_lt(f.get("LT_ORIGEN"))
        md = _norm_mz(f.get("MZ_DESTINO"))
        ld = _norm_lt(f.get("LT_DESTINO"))
        if mo and lo and md and ld:
            corr[(mo, lo)] = (md, ld)
    if corr:
        log.info(f"correcciones_lote.xlsx → {len(corr)} remapeos activos")
    return corr


def _escribir_correcciones_lote(filas: list[dict]):
    """Escribe correcciones_lote.xlsx — delega el dibujo al primitivo compartido
    (shared/utils_lote.py) para que 7_cierre pueda resetearlo con el mismo formato."""
    from utils_lote import escribir_correcciones_lote
    escribir_correcciones_lote(CORR_LOTE_PATH, filas)


def _recuperar_correcciones_trazabilidad(existentes: dict, ciclo: int,
                                         keys_validos: set = frozenset()) -> dict:
    """Auto-sana correcciones_lote desde trazabilidad_cobranza.xlsx.
    Cada corrección aplicada queda grabada en la trazabilidad (MZ_ORIGEN/LT_ORIGEN
    → MZ/LT). Si correcciones_lote se revirtió (git) y perdió alguna, se recupera
    desde ahí — el trabajo manual no depende de un solo archivo mutable.

    Guarda de seguridad (bug C1-9→C1-17, 10/08/2026): si el ORIGEN de la
    corrección a recuperar es HOY un predio real en la planilla, NO se auto-aplica
    — un origen real puede volver a recibir un pago legítimo en un ciclo futuro
    (verificar_lotes.py u otro fix ya pudo haber corregido la fuente desde
    entonces) y una resurrección ciega se lo robaría en silencio. Se avisa y
    se salta; si sigue haciendo falta, un humano la vuelve a escribir a mano.
    """
    traz = OUTPUTS_DIR / "trazabilidad_cobranza.xlsx"
    if not traz.exists():
        return existentes
    try:
        df = pd.read_excel(traz, header=1, dtype=str)
    except Exception as e:
        log.warning(f"No se pudo leer trazabilidad para recuperar correcciones: {e}")
        return existentes
    df.columns = _norm_cols(df)
    if not {"MZ_ORIGEN", "LT_ORIGEN", "MZ", "LT"} <= set(df.columns):
        return existentes

    nuevas = {}
    for _, f in df.iterrows():
        mo = _norm_mz(f.get("MZ_ORIGEN"))
        lo = _norm_lt(f.get("LT_ORIGEN"))
        md = _norm_mz(f.get("MZ"))
        ld = _norm_lt(f.get("LT"))
        if not (mo and lo and md and ld) or (mo, lo) in existentes:
            continue
        if f"{mo}-{lo}" in keys_validos:
            log.warning(f"  correccion NO recuperada (origen {mo}-{lo} es un predio "
                        f"real hoy — revisar a mano si todavia hace falta)")
            continue
        nuevas[(mo, lo)] = (md, ld)
    if not nuevas:
        return existentes

    ahora = datetime.now().strftime("%d/%m/%Y %H:%M")
    filas_prev = []
    if CORR_LOTE_PATH.exists():
        df_prev = pd.read_excel(CORR_LOTE_PATH, header=0)
        df_prev.columns = _norm_cols(df_prev)
        filas_prev = df_prev.to_dict("records")
    filas_nuevas = [
        {"MZ_ORIGEN": mo, "LT_ORIGEN": lo, "MZ_DESTINO": md, "LT_DESTINO": ld,
         "MOTIVO": "Recuperado desde trazabilidad_cobranza.xlsx",
         "CICLO": ciclo, "FECHA": ahora}
        for (mo, lo), (md, ld) in nuevas.items()
    ]
    _escribir_correcciones_lote(filas_prev + filas_nuevas)
    log.info(f"correcciones_lote.xlsx → {len(nuevas)} recuperada(s) desde trazabilidad")
    for (mo, lo), (md, ld) in nuevas.items():
        log.info(f"  recuperada: {mo}-{lo} → {md}-{ld}")
    return {**existentes, **nuevas}


def _absorber_correcciones_discrepancias(existentes: dict, ciclo: int) -> dict:
    """Lee MZ_CORRECTO+LT_CORRECTO llenados en discrepancias_cobranza.xlsx.
    Guarda las nuevas en correcciones_lote.xlsx y retorna el mapa combinado.
    """
    ruta = OUTPUTS_DIR / "discrepancias_cobranza.xlsx"
    if not ruta.exists():
        return existentes

    nuevas = {}
    try:
        wb_disc = load_workbook(ruta, data_only=True)
        for sheet in wb_disc.sheetnames:
            ws = wb_disc[sheet]
            hdrs = {str(ws.cell(2, c).value or "").strip().upper(): c
                    for c in range(1, ws.max_column + 1)}
            col_mzo = hdrs.get("MZ")
            col_lto = hdrs.get("LT")
            col_mzc = hdrs.get("MZ_CORRECTO")
            col_ltc = hdrs.get("LT_CORRECTO")
            if not all([col_mzo, col_lto, col_mzc, col_ltc]):
                continue
            for r in range(3, ws.max_row + 1):
                mo = _norm_mz(ws.cell(r, col_mzo).value)
                lo = _norm_lt(ws.cell(r, col_lto).value)
                mc = _norm_mz(ws.cell(r, col_mzc).value)
                lc = _norm_lt(ws.cell(r, col_ltc).value)
                if mo and lo and mc and lc and (mo, lo) not in existentes:
                    nuevas[(mo, lo)] = (mc, lc)
    except Exception as e:
        log.warning(f"No se pudo leer correcciones de discrepancias_cobranza.xlsx: {e}")
        return existentes

    if not nuevas:
        return existentes

    # Persistir en correcciones_lote.xlsx
    CORR_LOTE_PATH.parent.mkdir(exist_ok=True)
    ahora = datetime.now().strftime("%d/%m/%Y %H:%M")
    filas_prev = []
    if CORR_LOTE_PATH.exists():
        df_prev = pd.read_excel(CORR_LOTE_PATH, header=0)
        df_prev.columns = _norm_cols(df_prev)
        filas_prev = df_prev.to_dict("records")

    filas_nuevas = [
        {"MZ_ORIGEN": mo, "LT_ORIGEN": lo,
         "MZ_DESTINO": md, "LT_DESTINO": ld,
         "MOTIVO": "Corregido desde discrepancias_cobranza.xlsx",
         "CICLO": ciclo, "FECHA": ahora}
        for (mo, lo), (md, ld) in nuevas.items()
    ]

    _escribir_correcciones_lote(filas_prev + filas_nuevas)
    log.info(f"correcciones_lote.xlsx → {len(nuevas)} nueva(s) guardada(s) · "
             f"total {len(filas_prev) + len(filas_nuevas)}")

    return {**existentes, **nuevas}


def _aplicar_correcciones_lote(pagos: list[dict], correcciones: dict) -> list[dict]:
    """Remapea MZ+LT de pagos según correcciones. Preserva mz_origen/lt_origen."""
    if not correcciones:
        return pagos
    for p in pagos:
        key = (p["mz"], p["lt"])
        if key in correcciones:
            mzd, ltd = correcciones[key]
            p["mz_origen"] = p["mz"]
            p["lt_origen"] = p["lt"]
            p["mz"]  = mzd
            p["lt"]  = ltd
            p["key"] = f"{mzd}-{ltd}"
            log.info(f"  corrección aplicada: {key[0]}-{key[1]} → {mzd}-{ltd}")
    return pagos


# ── CÁLCULO ──────────────────────────────────────────────────────────────────
def _calcular(usuarios: list[dict],
              pagos_yape: list[dict],
              pagos_efectivo: list[dict],
              blancos: dict,
              dev_yape: dict,
              dev_efec: dict,
              dev_devuelto: dict,
              ciclo_nuevo: int,
              pagos_nuevos: set,
              aportes_tanque: dict | None = None,
              abonos_rezagados: dict | None = None) -> tuple[list[dict], set]:
    aportes_tanque = aportes_tanque or {}
    abonos_rezagados = abonos_rezagados or {}
    yape_por_key: dict[str, list[dict]] = {}
    efec_por_key: dict[str, list[dict]] = {}
    for p in pagos_yape:
        yape_por_key.setdefault(p["key"], []).append(p)
    for p in pagos_efectivo:
        efec_por_key.setdefault(p["key"], []).append(p)

    keys_validos = {u["key"] for u in usuarios}
    huerfanos = (set(yape_por_key) | set(efec_por_key)) - keys_validos
    for k in sorted(huerfanos):
        log.warning(f"Pago para {k} pero no está en planilla → discrepancias_cobranza.xlsx")

    # Lotes con retornos pero sin estar en planilla — anomalía
    huerfanos_dev = (set(dev_yape) | set(dev_efec)) - keys_validos
    for k in sorted(huerfanos_dev):
        log.warning(f"Retorno para {k} pero no está en planilla — ignorado")

    resultado = []
    blancos_usados = set()
    for u in usuarios:
        k  = u["key"]
        ys = yape_por_key.get(k, [])
        es = efec_por_key.get(k, [])
        ys_agua = [p for p in ys if not p.get("concepto")]
        es_agua = [p for p in es if not p.get("concepto")]
        yape_sum = round(sum(p["monto"] for p in ys_agua), 2)
        efec_sum = round(sum(p["monto"] for p in es_agua), 2)

        # Retornos y devoluciones — ambos reducen MONTO_YAPE.
        # RETORNO badge: desde pagos_yape_retorno.xlsx + pagos_efectivo_devolucion.xlsx
        # DEVUELTO badge: desde pagos_yape_devolucion.xlsx
        yape_dev_lote     = round(dev_yape.get(k, 0.0), 2)
        efec_dev_lote     = round(dev_efec.get(k, 0.0), 2)
        devuelto_lote     = round(dev_devuelto.get(k, 0.0), 2)
        total_dev         = round(yape_dev_lote + efec_dev_lote + devuelto_lote, 2)
        yape_neto         = round(yape_sum - total_dev, 2)
        if yape_neto < -TOL:
            log.warning(f"Lote {k}: retorno+devuelto (S/ {total_dev:.2f}) excede pago Yape "
                        f"(S/ {yape_sum:.2f}) → MONTO_YAPE quedará negativo")
        retorno_badge = _retorno_badge(yape_dev_lote, efec_dev_lote)
        devuelto_badge = "yape" if devuelto_lote > TOL else None

        pagado_normal = round(yape_neto + efec_sum, 2)

        abono_cerrado, abono_vigente = abonos_rezagados.get(
            (u["mz"], u["lt"]), (0.0, 0.0)
        )
        abono_rezagado = round(abono_cerrado + abono_vigente, 2)
        pagado = round(pagado_normal + abono_rezagado, 2)

        # Aporte al tanque mezclado en el mismo pago — se saca de total_pagado
        # ANTES de la cascada P1-P6 (ver _cargar_aportes_tanque_manuales),
        # para que no se cuente como pago de deuda además de aporte voluntario.
        aporte_tanque_aplicar = round(aportes_tanque.get(k, 0.0), 2)
        if aporte_tanque_aplicar > TOL:
            pagado_normal = round(max(pagado_normal - aporte_tanque_aplicar, 0.0), 2)
            pagado = round(pagado_normal + abono_rezagado, 2)

        blanco_aplicar = round(blancos.get(k, 0.0), 2)
        if blanco_aplicar > TOL:
            blancos_usados.add(k)
        # BLANCO en planilla queda negativo (reduce el total)
        blanco_final = round(u["blanco_inicial"] - blanco_aplicar, 2)

        total = round(
            u["mes_actual"] + u["mantenimiento"]
            + u["mes_anterior"] + u["corte_reconexion"]
            + u["convenio"] + u["multa"] + u["acuerdos_asamblea"]
            + blanco_final + u["devolucion"],
            2,
        )
        saldo = round(total - pagado, 2)
        fechas = [p["fecha"] for p in (ys_agua + es_agua)]

        # CICLO_COBRANZA del usuario: ciclo del pago más reciente cargado.
        # Si alguno de sus pagos es "nuevo" este run → ciclo_nuevo.
        # Si solo tiene pagos viejos → ciclo lo dejamos en None (se respeta el previo).
        ciclo_user = ciclo_nuevo if any(
            _identidad_pago(p) in pagos_nuevos for p in (ys + es)
        ) else None
        # Si no tiene pagos → vacío
        if not (ys or es):
            ciclo_user = None

        resultado.append({
            **u,
            "blanco_final":   blanco_final,
            "total_a_pagar":  total,
            "monto_yape":     yape_neto,
            "monto_efectivo": efec_sum,
            "total_pagado":   pagado,
            "total_pagado_normal": pagado_normal,
            "abono_rezagado": abono_rezagado,
            "abono_rezagado_cerrado": abono_cerrado,
            "abono_rezagado_vigente": abono_vigente,
            "saldo":          saldo,
            "estado":         _estado(saldo, pagado),
            "fecha_pago":     _fecha_max(fechas),
            "ciclo_cobranza": ciclo_user,
            "retorno":        retorno_badge,
            "devuelto":       devuelto_badge,
            "pagos_yape":     ys,
            "pagos_efectivo": es,
        })
    cnt = {e: sum(1 for r in resultado if r["estado"] == e)
           for e in ("CANCELADO", "EXCESO", "PARCIAL", "PENDIENTE")}
    log.info(f"Estados → CANCELADO={cnt['CANCELADO']} EXCESO={cnt['EXCESO']} "
             f"PARCIAL={cnt['PARCIAL']} PENDIENTE={cnt['PENDIENTE']}")
    n_retornos  = sum(1 for r in resultado if r["retorno"])
    n_devueltos = sum(1 for r in resultado if r["devuelto"])
    if n_retornos:
        log.info(f"Lotes con retorno → {n_retornos}")
    if n_devueltos:
        log.info(f"Lotes con devuelto → {n_devueltos}")
    return resultado, blancos_usados


# ─────────────────────────────────────────────────────────────────────────────
#  OUTPUT 1 — planilla_cobrado_YYYY-MM.xlsx
# ─────────────────────────────────────────────────────────────────────────────
# Layout (matching planilla_cobrado_diseno.html):
#   1- 4  ¿Quién es?     MZ LT NOMBRE MES_ANO
#   5      sep
#   6- 8  Lectura        MARC_ANT MARC_ACT M3
#   9      sep
#  10-16  Cobro—cargos   MES_ACTUAL MANTENIMIENTO MES_ANTERIOR CORTE_RECONEXION
#                         CONVENIO MULTA ACUERDOS_ASAMBLEA
#  17-18  Descuentos     BLANCO DEVOLUCION
#  19     Total          TOTAL_A_PAGAR
#  20     sep
#  21-28  Pago→5_cob     MONTO_YAPE MONTO_EFECTIVO ABONO_REZAGADO SALDO
#                         RETORNO DEVUELTO ESTADO FECHA_PAGO
#  29     sep
#  30     ¿Cuándo?       CICLO_COBRANZA
#
# SALDO se expone como columna del contrato — la consumen 6_corte y futuras
# tools sin re-implementar la fórmula (total − pagado).
# ─────────────────────────────────────────────────────────────────────────────

_PC_GRUPOS = [
    (1,  4,  "¿Quién es?",          *GH_QUIEN),
    (6,  8,  "Lectura",              *GH_LEC),
    (10, 16, "Cobro — cargos",       *GH_COB),
    (17, 18, "Descuentos",           *GH_DESC),
    (19, 19, "Total",                *GH_TOTAL),
    (21, 28, "Pago → 5_cobranza",    *GH_PAGO),
    (30, 30, "¿Cuándo?",             *GH_TRAZ),
]
_PC_COLS = [
    (1,  "MZ",                *GH_QUIEN,   6),
    (2,  "LT",                *GH_QUIEN,   6),
    (3,  "NOMBRE",            *GH_QUIEN,  26),
    (4,  "MES_ANO",           *GH_QUIEN,  10),
    (6,  "MARC_ANT",          *GH_LEC,     9),
    (7,  "MARC_ACT",          *GH_LEC,     9),
    (8,  "M3",                *GH_LEC,     6),
    (10, "MES_ACTUAL",        *GH_COB,    11),
    (11, "MANTENIMIENTO",     *GH_COB,    13),
    (12, "MES_ANTERIOR",      *GH_COB,    12),
    (13, "CORTE_RECONEXION",  *GH_COB,    16),
    (14, "CONVENIO",          *GH_COB,    10),
    (15, "MULTA",             *GH_COB,     8),
    (16, "ACUERDOS_ASAMBLEA", *GH_COB,    17),
    (17, "BLANCO",            *GH_DESC,    9),
    (18, "DEVOLUCION",        *GH_DESC,   11),
    (19, "TOTAL_A_PAGAR",     *GH_TOTAL,  13),
    (21, "MONTO_YAPE",        *GH_PAGO,   11),
    (22, "MONTO_EFECTIVO",    *GH_PAGO,   14),
    (23, "ABONO_REZAGADO",    *GH_PAGO,   16),
    (24, "SALDO",             *GH_PAGO,   12),
    (25, "RETORNO",           *GH_PAGO,   10),
    (26, "DEVUELTO",          *GH_PAGO,   10),
    (27, "ESTADO",            *GH_PAGO,   11),
    (28, "FECHA_PAGO",        *GH_PAGO,   11),
    (30, "CICLO_COBRANZA",    *GH_TRAZ,   14),
]
_PC_SEP_COLS = [5, 9, 20, 29]


def _exportar_planilla_cobrado(resultado: list[dict], mes_ano: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "planilla_cobrado"
    ws.freeze_panes = "A3"
    last_row = len(resultado) + 2

    for cs, ce, texto, bg, txt in _PC_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for sc in _PC_SEP_COLS:
        _sep(ws, sc, last_row)
    for col, nombre, bg, txt, ancho in _PC_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    MONEY = '"S/ "#,##0.00'

    for ri, r in enumerate(resultado, 3):
        # Quién es?
        _c(ws, ri,  1, r["mz"],      TD_QUIEN, "1A5276", mono=True, align="center")
        _c(ws, ri,  2, r["lt"],      TD_QUIEN, "1A5276", mono=True, align="center")
        _c(ws, ri,  3, r["nombre"],  TD_QUIEN, "333333", align="left")
        _c(ws, ri,  4, r["mes_ano"], TD_QUIEN, "1A5276", mono=True, align="center")

        # Lectura
        _c(ws, ri,  6, r["marc_ant"], TD_LEC, "0C447C", mono=True, align="right")
        _c(ws, ri,  7, r["marc_act"], TD_LEC, "0C447C", mono=True, align="right")
        _c(ws, ri,  8, r["m3"],       TD_LEC, "065F46", mono=True, align="right", bold=True)

        # Cobro — cargos (None si es 0 → muestra vacío)
        def _cob(col, val):
            val_disp = val if val > TOL else None
            _c(ws, ri, col, val_disp, TD_COB, "1E5C3A",
               mono=True, align="right", fmt=MONEY if val_disp else None)

        _cob(10, r["mes_actual"])
        _cob(11, r["mantenimiento"])
        _cob(12, r["mes_anterior"])
        _cob(13, r["corte_reconexion"])
        _cob(14, r["convenio"])
        _cob(15, r["multa"])
        _cob(16, r["acuerdos_asamblea"])

        # Descuentos (negativos)
        def _desc(col, val):
            val_disp = val if abs(val) > TOL else None
            _c(ws, ri, col, val_disp, TD_DESC, "4C1D95",
               mono=True, align="right", bold=True,
               fmt=MONEY if val_disp else None)

        _desc(17, r["blanco_final"])
        _desc(18, r["devolucion"] if r["devolucion"] >= 0 else -r["devolucion"])

        # Total (fórmula Excel: J:R = cols 10-18)
        formula = f"=SUM(J{ri}:R{ri})"
        c_tot = ws.cell(row=ri, column=19, value=formula)
        c_tot.font          = Font(name="Consolas", size=10, bold=True, color="1E5C3A")
        c_tot.fill          = PatternFill("solid", start_color=TD_TOTAL)
        c_tot.alignment     = Alignment(horizontal="right", vertical="center")
        c_tot.border        = _borde()
        c_tot.number_format = MONEY

        # Pago → 5_cobranza
        def _pag(col, val):
            val_disp = val if val > TOL else None
            _c(ws, ri, col, val_disp, TD_PAGO, "5B21B6",
               mono=True, align="right", fmt=MONEY if val_disp else None)

        # MONTO_YAPE puede ser negativo si el retorno excede el pago — no lo ocultes
        _pag_yape_val = r["monto_yape"]
        if abs(_pag_yape_val) > TOL:
            _c(ws, ri, 21, _pag_yape_val, TD_PAGO, "5B21B6",
               mono=True, align="right", bold=(_pag_yape_val < 0),
               fmt=MONEY)
        else:
            _c(ws, ri, 21, None, TD_PAGO, "5B21B6", mono=True, align="right")
        _pag(22, r["monto_efectivo"])
        _pag(23, r["abono_rezagado"])

        # SALDO — total − pagado. Puede ser positivo (debe), 0 (cancelado),
        # negativo (exceso). Lo expone como columna del contrato para que
        # 6_corte y otras tools lean directo sin re-computar.
        saldo_val = r["saldo"]
        if abs(saldo_val) > TOL:
            _c(ws, ri, 24, saldo_val, TD_PAGO, "5B21B6",
               mono=True, align="right", bold=True, fmt=MONEY)
        else:
            _c(ws, ri, 24, 0, TD_PAGO, "5B21B6",
               mono=True, align="right", fmt=MONEY)

        # RETORNO badge (vacio si no hubo retorno)
        if r["retorno"]:
            ret_bg  = RETORNO_BG.get(r["retorno"], "FFFFFF")
            ret_txt = RETORNO_TXT.get(r["retorno"], "333333")
            c_ret = ws.cell(row=ri, column=25, value=r["retorno"])
            c_ret.font      = Font(name="Arial", size=9, bold=True, color=ret_txt)
            c_ret.fill      = PatternFill("solid", start_color=ret_bg)
            c_ret.alignment = Alignment(horizontal="center", vertical="center")
            c_ret.border    = _borde()
        else:
            _c(ws, ri, 25, None, TD_PAGO, "5B21B6", mono=True, align="center")

        # DEVUELTO badge — PAGASTE devolucion aplicado a este lote
        if r.get("devuelto"):
            dev_bg  = RETORNO_BG.get(r["devuelto"], "FFFFFF")
            dev_txt = RETORNO_TXT.get(r["devuelto"], "333333")
            c_dev = ws.cell(row=ri, column=26, value=r["devuelto"])
            c_dev.font      = Font(name="Arial", size=9, bold=True, color=dev_txt)
            c_dev.fill      = PatternFill("solid", start_color=dev_bg)
            c_dev.alignment = Alignment(horizontal="center", vertical="center")
            c_dev.border    = _borde()
        else:
            _c(ws, ri, 26, None, TD_PAGO, "5B21B6", mono=True, align="center")

        # ESTADO badge
        est_bg  = ESTADO_BG.get(r["estado"], "FFFFFF")
        est_txt = ESTADO_TXT.get(r["estado"], "333333")
        c_est = ws.cell(row=ri, column=27, value=r["estado"])
        c_est.font      = Font(name="Arial", size=9, bold=True, color=est_txt)
        c_est.fill      = PatternFill("solid", start_color=est_bg)
        c_est.alignment = Alignment(horizontal="center", vertical="center")
        c_est.border    = _borde()

        _c(ws, ri, 28, r["fecha_pago"] or None, TD_PAGO, "7C3AED",
           mono=True, align="center")

        # CICLO_COBRANZA
        _c(ws, ri, 30, r["ciclo_cobranza"], TD_TRAZ, "7D6608",
           mono=True, align="center", bold=True)

        ws.row_dimensions[ri].height = 17

    return wb, OUTPUTS_DIR / f"planilla_cobrado_{mes_ano}.xlsx"


# ─────────────────────────────────────────────────────────────────────────────
#  OUTPUT 2 — trazabilidad_cobranza.xlsx (acumulada)
# ─────────────────────────────────────────────────────────────────────────────
# Layout (matching trazabilidad_cobranza.html v2.0):
#   1-3   ¿Quién es?          MZ LT NOMBRE
#   4     sep
#   5-7   ¿Qué pagó?          MONTO FUENTE RETORNO
#   8     sep
#   9-11  ¿Cómo verificarlo?  REFERENCIA FECHA COMENTARIO
#   12    sep
#   13-15 ¿De qué ciclo?      CICLO_CORRECCION_ORIGEN CICLO_COBRANZA FECHA_CARGA
#   16    sep
#   17-18 Lote corregido      MZ_ORIGEN LT_ORIGEN
# REFERENCIA: ORIGEN del banco (Yape) ó "MESA / COBRADOR" (efectivo).
# FECHA: con hora para Yape (DD/MM/YYYY HH:MM), solo fecha para efectivo.
# COMENTARIO: MENSAJE del pagador (Yape) ó nota del cobrador (efectivo).

_TZ_GRUPOS = [
    (1,  3,  "¿Quién es?",          *GH_TZ_QUIEN),
    (5,  9,  "¿Qué pagó?",          *GH_COB),
    (11, 13, "¿Cómo verificarlo?",  *GH_TZ_VERIF),
    (15, 17, "¿De qué ciclo?",      *GH_TZ_CICLO),
    (19, 20, "Lote corregido",      *GH_DC_CORR),
]
_TZ_COLS = [
    (1,  "MZ",                       *GH_TZ_QUIEN,   6),
    (2,  "LT",                       *GH_TZ_QUIEN,   6),
    (3,  "NOMBRE",                   *GH_TZ_QUIEN, 26),
    (5,  "MONTO",                    *GH_COB,      11),
    (6,  "FUENTE",                   *GH_COB,      10),
    (7,  "CONCEPTO",                 *GH_COB,      14),
    (8,  "RETORNO",                  *GH_COB,      10),
    (9,  "DEVUELTO",                 *GH_COB,      10),
    (11, "REFERENCIA",               *GH_TZ_VERIF, 24),
    (12, "FECHA",                    *GH_TZ_VERIF, 18),
    (13, "COMENTARIO",               *GH_TZ_VERIF, 30),
    (15, "CICLO_CORRECCION_ORIGEN",  *GH_TZ_CICLO, 22),
    (16, "CICLO_COBRANZA",           *GH_TZ_CICLO, 16),
    (17, "FECHA_CARGA",              *GH_TZ_CICLO, 18),
    (19, "MZ_ORIGEN",                *GH_DC_CORR,   9),
    (20, "LT_ORIGEN",                *GH_DC_CORR,   9),
]
_TZ_SEP_COLS = [4, 10, 14, 18]

FUENTE_BG  = {"yape": "E1F5EE", "efectivo": "EFF6FF"}
FUENTE_TXT = {"yape": "085041", "efectivo": "1D4ED8"}
CONCEPTO_BG  = {"tanque": "FFF7ED", "honorario": "F5F3FF", "gasto": "ECFDF5", "comunitario": "FEF9E7"}
CONCEPTO_TXT = {"tanque": "9A3412", "honorario": "5B21B6", "gasto": "065F46", "comunitario": "7D6608"}


def _exportar_trazabilidad_cobranza(
    resultado: list[dict],
    pagos_yape: list[dict],
    pagos_efectivo: list[dict],
    ciclo_nuevo: int,
    pagos_nuevos: set,
    trazabilidad_path: Path,
    retornos_por_lote: dict,
    devueltos_por_lote: dict,
):
    """Append: lee filas previas, agrega solo las nuevas con CICLO_COBRANZA=ciclo_nuevo.
    RETORNO se recalcula en cada run desde el estado actual de retornos_por_lote —
    todas las filas del mismo lote muestran el mismo badge."""
    # Mapa key → nombre (de planilla)
    nombre_de = {r["key"]: r["nombre"] for r in resultado}

    # Filas previas (preservar)
    previas = []
    if trazabilidad_path.exists():
        df = pd.read_excel(trazabilidad_path, header=1)
        df.columns = _norm_cols(df)
        for _, f in df.iterrows():
            mz = _norm_mz(f.get("MZ"))
            lt = _norm_lt(f.get("LT"))
            if not mz or not lt:
                continue
            previas.append({
                "mz":     mz,
                "lt":     lt,
                "nombre": str(f.get("NOMBRE", "")).strip(),
                "monto":  round(_float(f.get("MONTO")), 2),
                "fuente": str(f.get("FUENTE", "")).strip().lower(),
                "concepto":   _txt(f.get("CONCEPTO")) if "CONCEPTO" in df.columns else "",
                "referencia": _txt(f.get("REFERENCIA")) if "REFERENCIA" in df.columns else "",
                "fecha_hora":  _txt(f.get("FECHA")) if "FECHA" in df.columns else "",
                "comentario":  _txt(f.get("COMENTARIO")) if "COMENTARIO" in df.columns else "",
                "ciclo_correccion_origen": int(_float(f.get("CICLO_CORRECCION_ORIGEN")) or 0),
                "ciclo_cobranza":          int(_float(f.get("CICLO_COBRANZA")) or 0),
                "fecha_carga":             str(f.get("FECHA_CARGA", "")).strip(),
                "mz_origen":               _norm_mz(f.get("MZ_ORIGEN")) if "MZ_ORIGEN" in df.columns else "",
                "lt_origen":               _norm_lt(f.get("LT_ORIGEN")) if "LT_ORIGEN" in df.columns else "",
            })

    # Backfill in-place: las filas de ciclos antiguos no tenían REFERENCIA/COMENTARIO.
    # Se rellenan cruzando contra los pagos fuente por (mz, lt, monto, fuente) sin
    # tocar su ciclo ni fecha_carga. Idempotente: re-correr rellena igual.
    enriq = {}
    for p in (pagos_yape + pagos_efectivo):
        enriq[(p["mz"], p["lt"], round(p["monto"], 2), p["fuente"])] = {
            "referencia": p.get("referencia", ""),
            "comentario": p.get("comentario", ""),
            "fecha_hora": p.get("fecha_hora", "") or p.get("fecha", ""),
        }
    for t in previas:
        src = enriq.get((t["mz"], t["lt"], round(t["monto"], 2), t["fuente"]))
        if not src:
            continue
        if not t.get("referencia"):
            t["referencia"] = src["referencia"]
        if not t.get("comentario"):
            t["comentario"] = src["comentario"]
        if not t.get("fecha_hora") or len(t["fecha_hora"]) <= 10:
            t["fecha_hora"] = src["fecha_hora"] or t.get("fecha_hora", "")

    # Filas nuevas (solo pagos imputados — huérfanos van a discrepancias, no aquí)
    ahora = datetime.now().strftime("%d/%m/%Y %H:%M")
    nuevas = []
    for p in (pagos_yape + pagos_efectivo):
        if _identidad_pago(p) not in pagos_nuevos:
            continue
        if p["key"] not in nombre_de:
            continue  # huérfano — no está en planilla, no se imputa
        nuevas.append({
            "mz":     p["mz"],
            "lt":     p["lt"],
            "nombre": nombre_de.get(p["key"], p["nombre"]),
            "monto":  p["monto"],
            "fuente": p["fuente"],
            "concepto":   p.get("concepto", ""),
            "referencia": p.get("referencia", ""),
            "fecha_hora": p.get("fecha_hora", "") or p.get("fecha", ""),
            "comentario": p.get("comentario", ""),
            "ciclo_correccion_origen": p["ciclo_correccion"],
            "ciclo_cobranza":          ciclo_nuevo,
            "fecha_carga":             ahora,
            "mz_origen":               p.get("mz_origen", ""),
            "lt_origen":               p.get("lt_origen", ""),
        })

    todas = previas + nuevas
    last_row = max(len(todas) + 2, 3)

    wb = Workbook()
    ws = wb.active
    ws.title = "trazabilidad_cobranza"
    ws.freeze_panes = "A3"

    for cs, ce, texto, bg, txt in _TZ_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for sc in _TZ_SEP_COLS:
        _sep(ws, sc, last_row)
    for col, nombre, bg, txt, ancho in _TZ_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    MONEY = '"S/ "#,##0.00'
    for ri, t in enumerate(todas, 3):
        _c(ws, ri, 1, t["mz"],     TD_TZ_QUIEN, "5B21B6", mono=True, align="center")
        _c(ws, ri, 2, t["lt"],     TD_TZ_QUIEN, "5B21B6", mono=True, align="center")
        _c(ws, ri, 3, t["nombre"], TD_TZ_QUIEN, "333333", align="left")
        _c(ws, ri, 5, t["monto"],  TD_COB,      "065F46",
           mono=True, align="right", bold=True, fmt=MONEY)

        # FUENTE como badge
        f_bg  = FUENTE_BG.get(t["fuente"], "F3F4F6")
        f_txt = FUENTE_TXT.get(t["fuente"], "374151")
        c_f = ws.cell(row=ri, column=6, value=t["fuente"])
        c_f.font      = Font(name="Arial", size=9, bold=True, color=f_txt)
        c_f.fill      = PatternFill("solid", start_color=f_bg)
        c_f.alignment = Alignment(horizontal="center", vertical="center")
        c_f.border    = _borde()

        # CONCEPTO badge (tipo de pago: tanque, honorario, gasto, comunitario, o vacío=agua)
        conc_val = t.get("concepto", "") or ""
        if conc_val:
            c_conc = ws.cell(row=ri, column=7, value=conc_val)
            c_conc.font      = Font(name="Arial", size=9, bold=True,
                                    color=CONCEPTO_TXT.get(conc_val, "9A3412"))
            c_conc.fill      = PatternFill("solid",
                                           start_color=CONCEPTO_BG.get(conc_val, "FFF7ED"))
            c_conc.alignment = Alignment(horizontal="center", vertical="center")
            c_conc.border    = _borde()
        else:
            _c(ws, ri, 7, None, TD_COB, "065F46", mono=True, align="center")

        # RETORNO badge (puntero al archivo de retornos — vacio si no hubo)
        retorno_lote = retornos_por_lote.get(f"{t['mz']}-{t['lt']}")
        if retorno_lote:
            r_bg  = RETORNO_BG.get(retorno_lote, "FFFFFF")
            r_txt = RETORNO_TXT.get(retorno_lote, "333333")
            c_r = ws.cell(row=ri, column=8, value=retorno_lote)
            c_r.font      = Font(name="Arial", size=9, bold=True, color=r_txt)
            c_r.fill      = PatternFill("solid", start_color=r_bg)
            c_r.alignment = Alignment(horizontal="center", vertical="center")
            c_r.border    = _borde()
        else:
            _c(ws, ri, 8, None, TD_COB, "065F46", mono=True, align="center")

        # DEVUELTO badge — pagos_yape_devolucion aplicado a este lote
        devuelto_lote = devueltos_por_lote.get(f"{t['mz']}-{t['lt']}")
        if devuelto_lote:
            dv_bg  = RETORNO_BG.get(devuelto_lote, "FFFFFF")
            dv_txt = RETORNO_TXT.get(devuelto_lote, "333333")
            c_dv = ws.cell(row=ri, column=9, value=devuelto_lote)
            c_dv.font      = Font(name="Arial", size=9, bold=True, color=dv_txt)
            c_dv.fill      = PatternFill("solid", start_color=dv_bg)
            c_dv.alignment = Alignment(horizontal="center", vertical="center")
            c_dv.border    = _borde()
        else:
            _c(ws, ri, 9, None, TD_COB, "065F46", mono=True, align="center")

        # ¿Cómo verificarlo? — REFERENCIA · FECHA (con hora si Yape) · COMENTARIO
        _c(ws, ri, 11, t.get("referencia") or None, TD_TZ_VERIF, "92400E",
           mono=True, align="left")
        _c(ws, ri, 12, t.get("fecha_hora") or None, TD_TZ_VERIF, "7C2D12",
           mono=True, align="center")
        _c(ws, ri, 13, t.get("comentario") or None, TD_TZ_VERIF, "78350F",
           align="left")
        # ¿De qué ciclo? — corrección origen · cobranza · timestamp de carga
        _c(ws, ri, 15, t["ciclo_correccion_origen"], TD_TZ_CICLO, "1E40AF",
           mono=True, align="center")
        _c(ws, ri, 16, t["ciclo_cobranza"],          TD_TZ_CICLO, "1E40AF",
           mono=True, align="center", bold=True)
        _c(ws, ri, 17, t["fecha_carga"],             TD_TZ_CICLO, "1E40AF",
           mono=True, align="center")
        # Lote corregido — solo para pagos con remapeo de correcciones_lote
        mz_orig = t.get("mz_origen") or ""
        lt_orig = t.get("lt_origen") or ""
        bg_orig = TD_DC_CORR if mz_orig else TD_DC_CORR_V
        _c(ws, ri, 19, mz_orig or None, bg_orig, GH_DC_CORR[1], mono=True, align="center")
        _c(ws, ri, 20, lt_orig or None, bg_orig, GH_DC_CORR[1], mono=True, align="center")
        ws.row_dimensions[ri].height = 17

    wb.save(trazabilidad_path)
    log.info(f"trazabilidad_cobranza.xlsx → {len(todas)} filas "
             f"({len(nuevas)} nuevas en ciclo {ciclo_nuevo})")


# ─────────────────────────────────────────────────────────────────────────────
#  OUTPUT 3 — resumen_recaudacion.xlsx
# ─────────────────────────────────────────────────────────────────────────────
def _exportar_resumen(resultado: list[dict], n_corte: int,
                      mes_ano: str, ciclo_nuevo: int):
    tot = {
        "deuda":    round(sum(r["total_a_pagar"]  for r in resultado), 2),
        "yape":     round(sum(r["monto_yape"]     for r in resultado), 2),
        "efectivo": round(sum(r["monto_efectivo"] for r in resultado), 2),
        "pagado":   round(sum(r["total_pagado"]   for r in resultado), 2),
        "saldo":    round(sum(r["saldo"] for r in resultado if r["saldo"] > 0), 2),
        "exceso":   round(sum(-r["saldo"] for r in resultado if r["saldo"] < 0), 2),
    }
    cnt = {e: sum(1 for r in resultado if r["estado"] == e)
           for e in ("CANCELADO", "EXCESO", "PARCIAL", "PENDIENTE")}

    filas = [
        ("── RECAUDACIÓN ──────────────────",         None,           None),
        ("Total a pagar (planilla)",                  tot["deuda"],    "S/"),
        ("Recaudado Yape",                            tot["yape"],     "S/"),
        ("Recaudado Efectivo",                        tot["efectivo"], "S/"),
        ("Recaudado total",                           tot["pagado"],   "S/"),
        ("Saldo pendiente",                           tot["saldo"],    "S/"),
        ("Exceso pagado (a devolver)",                tot["exceso"],   "S/"),
        ("── ESTADOS ──────────────────────",         None,           None),
        ("CANCELADO",                                 cnt["CANCELADO"], "usuarios"),
        ("EXCESO",                                    cnt["EXCESO"],    "usuarios"),
        ("PARCIAL",                                   cnt["PARCIAL"],   "usuarios"),
        ("PENDIENTE",                                 cnt["PENDIENTE"], "usuarios"),
        ("── CORTE ────────────────────────",         None,           None),
        ("Elegibles para corte (SALDO>0 & MES_ANT>=8)", n_corte,        "usuarios"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "resumen"
    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1,
                value=f"Resumen cobranza · {mes_ano} · ciclo {ciclo_nuevo}")
    t.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    t.fill      = PatternFill("solid", start_color="4C1D95")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10

    for ri, (concepto, valor, unidad) in enumerate(filas, 2):
        es_sep = valor is None
        bg = "EDE9FE" if es_sep else "FAF5FF"
        _c(ws, ri, 1, concepto, bg=bg, bold=es_sep, align="left",
           txt="4C1D95" if es_sep else "333333")
        if isinstance(valor, float):
            _c(ws, ri, 2, valor, bg=bg, align="right", mono=True,
               fmt='"S/ "#,##0.00' if unidad == "S/" else "#,##0")
        else:
            _c(ws, ri, 2, valor, bg=bg, align="right", mono=not es_sep)
        _c(ws, ri, 3, unidad, bg=bg, align="left", txt="888888")
        ws.row_dimensions[ri].height = 17

    wb.save(OUTPUTS_DIR / "resumen_recaudacion.xlsx")
    log.info("resumen_recaudacion.xlsx generado")


# ─────────────────────────────────────────────────────────────────────────────
#  OUTPUT 4 — arrastre_deuda_YYYY-MM.xlsx
# ─────────────────────────────────────────────────────────────────────────────
_AD_GRUPOS = [
    (1, 3, "¿Quién es?",      *GH_AD_QUIEN),
    (5, 5, "¿Cuánto debe?",   *GH_AD_MONTO),
    (7, 7, "¿De qué mes?",    *GH_AD_TRAZ),
]
_AD_COLS = [
    (1, "MZ",             *GH_AD_QUIEN,   6),
    (2, "LT",             *GH_AD_QUIEN,   6),
    (3, "NOMBRE",         *GH_AD_QUIEN, 26),
    (5, "monto",          *GH_AD_MONTO, 12),
    (7, "MES_ANO_ORIGEN", *GH_AD_TRAZ,  16),
]
_AD_SEP_COLS = [4, 6]


def _exportar_arrastre_deuda(resultado: list[dict], mes_ano: str):
    pendientes = [r for r in resultado if r["saldo"] > TOL]
    last_row = max(len(pendientes) + 2, 3)

    wb = Workbook()
    ws = wb.active
    ws.title = f"arrastre_deuda_{mes_ano}"[:31]
    ws.freeze_panes = "A3"

    for cs, ce, texto, bg, txt in _AD_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for sc in _AD_SEP_COLS:
        _sep(ws, sc, last_row)
    for col, nombre, bg, txt, ancho in _AD_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    MONEY = '"S/ "#,##0.00'
    for ri, r in enumerate(pendientes, 3):
        _c(ws, ri, 1, r["mz"],         TD_AD_QUIEN, "065F46", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],         TD_AD_QUIEN, "065F46", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"],     TD_AD_QUIEN, "333333", align="left")
        _c(ws, ri, 5, r["saldo"],      TD_AD_MONTO, "1A5276",
           mono=True, align="right", bold=True, fmt=MONEY)
        _c(ws, ri, 7, mes_ano,         TD_AD_TRAZ,  "0E6655", mono=True, align="center")
        ws.row_dimensions[ri].height = 17

    nombre = f"arrastre_deuda_{mes_ano}.xlsx"
    wb.save(OUTPUTS_DIR / nombre)
    log.info(f"{nombre} → {len(pendientes)} usuarios con SALDO>0")


# ─────────────────────────────────────────────────────────────────────────────
#  HOJA 2 — arrastre_devolucion dentro de planilla_cobrado_YYYY-MM.xlsx
#  Paralelo a arrastre_deuda · misma estructura · paleta azul EXCESO.
#  monto = |saldo| (positivo · lo que la JASS le debe al usuario).
# ─────────────────────────────────────────────────────────────────────────────
_AV_GRUPOS = [
    (1, 3, "¿Quién es?",      *GH_AV_QUIEN),
    (5, 5, "¿Cuánto sobra?",  *GH_AV_MONTO),
    (7, 7, "¿De qué mes?",    *GH_AV_TRAZ),
    (9, 10, "¿Cómo ubicarlo?", *GH_AV_TRAZ),
    (12, 13, "¿Revisado?",    *GH_AV_REVIS),
]
_AV_COLS = [
    (1, "MZ",             *GH_AV_QUIEN,   6),
    (2, "LT",             *GH_AV_QUIEN,   6),
    (3, "NOMBRE",         *GH_AV_QUIEN, 26),
    (5, "monto",          *GH_AV_MONTO, 12),
    (7, "MES_ANO_ORIGEN", *GH_AV_TRAZ,  16),
    (9, "REFERENCIA",     *GH_AV_TRAZ,  32),
    (10, "COMENTARIO",    *GH_AV_TRAZ,  28),
    (12, "REVISION",      *GH_AV_REVIS, 32),
    (13, "ESTADO",        *GH_AV_REVIS, 16),
]
_AV_SEP_COLS = [4, 6, 8, 11]
_AV_ESTADO_OPCIONES = ["resuelto", "pendiente"]


def _backup_arrastre_devolucion(ruta: Path) -> Path | None:
    """Copia el archivo actual a outputs/backups/ antes de sobreescribirlo —
    capa 1 de preservación de trabajo manual (columna REVISION, Regla 9)."""
    if not ruta.exists():
        return None
    bk_dir = OUTPUTS_DIR / "backups"
    bk_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    bk = bk_dir / f"{ruta.stem}_{ts}{ruta.suffix}"
    shutil.copy2(ruta, bk)
    return bk


def _leer_revision_previa(ruta: Path, hoja: str | None = None) -> dict | None:
    """Lee REVISION y ESTADO ya tipeados en arrastre_devolucion, keyed por
    (mz, lt) → {"revision": ..., "estado": ...}, para no borrar el trabajo
    manual del operador al regenerar el archivo (Regla 9 — 3 capas)."""
    previo = {}
    if not ruta.exists():
        return None if hoja else previo
    try:
        wb = load_workbook(ruta, data_only=True)
    except Exception as e:
        log.warning(f"No se pudo leer REVISION/ESTADO previo: {e}")
        return None if hoja else previo
    if hoja and hoja not in wb.sheetnames:
        return None
    ws = wb[hoja] if hoja else wb.active
    hdrs = {str(ws.cell(2, c).value or "").strip().upper(): c
            for c in range(1, ws.max_column + 1)}
    cmz, clt = hdrs.get("MZ"), hdrs.get("LT")
    crev, cest = hdrs.get("REVISION"), hdrs.get("ESTADO")
    if not all([cmz, clt]) or not (crev or cest):
        return previo
    for r in range(3, ws.max_row + 1):
        mz = _norm_mz(ws.cell(r, cmz).value)
        lt = _norm_lt(ws.cell(r, clt).value)
        if not (mz and lt):
            continue
        rev = ws.cell(r, crev).value if crev else None
        est = ws.cell(r, cest).value if cest else None
        if rev not in (None, "") or est not in (None, ""):
            previo[(mz, lt)] = {"revision": rev, "estado": est}
    return previo


def _exportar_arrastre_devolucion(wb: Workbook, resultado: list[dict], mes_ano: str,
                                  previo: dict | None = None,
                                  disc_yape: list[dict] = None, disc_efec: list[dict] = None):
    excesos = [r for r in resultado if r["saldo"] < -TOL]
    # No identificados: plata real cobrada (mesa/yape) cuyo MZ+LT no existe en
    # planilla — no tienen usuario ni SALDO, así que nunca entrarían a `excesos`.
    # Sin esto quedaban visibles solo en discrepancias_cobranza.xlsx, un archivo
    # que nadie revisa buscando plata pendiente de devolver/reidentificar.
    no_identificados = list(disc_efec or []) + list(disc_yape or [])
    last_row = max(len(excesos) + len(no_identificados) + 2, 3)

    previo = previo or {}
    ws = wb.create_sheet("arrastre_devolucion")
    ws.freeze_panes = "A3"

    for cs, ce, texto, bg, txt in _AV_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for sc in _AV_SEP_COLS:
        _sep(ws, sc, last_row)
    for col, nombre, bg, txt, ancho in _AV_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    MONEY = '"S/ "#,##0.00'
    for ri, r in enumerate(excesos, 3):
        monto = round(abs(r["saldo"]), 2)
        _c(ws, ri, 1, r["mz"],     TD_AV_QUIEN, "1D4ED8", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_AV_QUIEN, "1D4ED8", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_AV_QUIEN, "333333", align="left")
        _c(ws, ri, 5, monto,       TD_AV_MONTO, "1D4ED8",
           mono=True, align="right", bold=True, fmt=MONEY)
        _c(ws, ri, 7, mes_ano,     TD_AV_TRAZ,  "1D4ED8", mono=True, align="center")

        # REFERENCIA/COMENTARIO — solo pagos agua (sin CONCEPTO), son los que generan el exceso.
        refs, coments = [], []
        for p in r["pagos_efectivo"]:
            if p.get("concepto"):
                continue
            refs.append(f"{p.get('mesa','')} / {p.get('cobrador','')} / {p.get('fecha','')}")
            coments.append(p.get("comentario", ""))
        for p in r["pagos_yape"]:
            if p.get("concepto"):
                continue
            refs.append(f"{p.get('referencia','')} / S/{p['monto']:g} / {p.get('fecha_hora_seg','')}")
            coments.append(p.get("comentario", ""))
        _c(ws, ri, 9,  " · ".join(refs),    TD_AV_TRAZ, "1D4ED8", mono=True, align="left")
        _c(ws, ri, 10, " · ".join(coments), TD_AV_TRAZ, "555555", align="left")

        prev = previo.get((r["mz"], r["lt"]), {})
        rev, est = prev.get("revision"), prev.get("estado")
        bg_rev = TD_DC_CORR if rev else TD_AV_REVIS
        bg_est = TD_DC_CORR if est else TD_AV_REVIS
        _c(ws, ri, 12, rev, bg_rev, GH_AV_REVIS[1], align="left")
        _c(ws, ri, 13, est, bg_est, GH_AV_REVIS[1], align="left")
        ws.row_dimensions[ri].height = 17

    # No identificados — mismo color rojo-huérfano que discrepancias_cobranza,
    # para que no se confundan con un exceso de un usuario real.
    fila_base = 3 + len(excesos)
    for i, p in enumerate(sorted(no_identificados,
                                  key=lambda x: (x["mz"], x["lt"], x["fecha"])),
                           fila_base):
        _c(ws, i, 1, p["mz"],                 TD_DC_PREDIO, "991B1B", mono=True, align="center", bold=True)
        _c(ws, i, 2, p["lt"],                 TD_DC_PREDIO, "991B1B", mono=True, align="center", bold=True)
        _c(ws, i, 3, "(no identificado)",     TD_DC_PREDIO, "991B1B", align="left")
        _c(ws, i, 5, round(p["monto"], 2),    TD_AV_MONTO,  "991B1B",
           mono=True, align="right", bold=True, fmt=MONEY)
        _c(ws, i, 7, mes_ano,                 TD_AV_TRAZ,   "991B1B", mono=True, align="center")
        ref = (f"{p.get('mesa','')} / {p.get('cobrador','')} / {p.get('fecha','')}"
               if p.get("fuente") == "efectivo"
               else f"{p.get('referencia','')} / S/{p['monto']:g} / {p.get('fecha_hora','')}")
        _c(ws, i, 9,  ref,                                TD_AV_TRAZ, "991B1B", mono=True, align="left")
        _c(ws, i, 10, "predio no encontrado en planilla",  TD_AV_TRAZ, "991B1B", align="left")

        prev = previo.get((p["mz"], p["lt"]), {})
        rev, est = prev.get("revision"), prev.get("estado")
        bg_rev = TD_DC_CORR if rev else TD_AV_REVIS
        bg_est = TD_DC_CORR if est else TD_AV_REVIS
        _c(ws, i, 12, rev, bg_rev, GH_AV_REVIS[1], align="left")
        _c(ws, i, 13, est, bg_est, GH_AV_REVIS[1], align="left")
        ws.row_dimensions[i].height = 17

    if last_row >= 3:
        dv = DataValidation(type="list", formula1=f'"{",".join(_AV_ESTADO_OPCIONES)}"',
                             allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"M3:M{last_row}")

    log.info(f"arrastre_devolucion → {len(excesos)} usuarios con SALDO<0 · "
              f"{len(no_identificados)} no identificado(s) (esperan reclamo/reidentificación)")


# ─────────────────────────────────────────────────────────────────────────────
#  HOJA 1 — arrastre_consolidado dentro de planilla_cobrado_YYYY-MM.xlsx
#  Consolida en un archivo lo que eran 3 arrastres separados (deuda + corte +
#  multa). Descompone el SALDO pendiente por componente en orden de prioridad
#  P1 DEUDA_AGUA → P2 CORTE → P3 CONVENIO → P4 ACUERDOS → P5 MULTA: cada columna
#  muestra lo que quedó SIN cubrir tras aplicar el pago en ese orden.
#  Solo filas TOTAL_ARRASTRE>0. Gate: estado_ciclo.json[mes].arrastre.validado.
#  2_planilla del próximo mes lo lee como fuente única de multa/acuerdos/convenio.
# ─────────────────────────────────────────────────────────────────────────────

def _ciclo_validado(mes_ano: str) -> bool:
    return repo_estado.ciclo_validado(mes_ano, ruta=ESTADO_CICLO_PATH)


def _marcar_generado(mes_ano: str, snapshot_hash: str):
    """Registra la proyección vigente e invalida sellos de una versión anterior."""
    repo_estado.marcar_generado(mes_ano, ruta=ESTADO_CICLO_PATH,
                                snapshot_hash=snapshot_hash)
    log.info(f"estado_ciclo.json → proyección generada · {mes_ano} · {snapshot_hash[:12]}")


_CORTE_LEDGER = "2026-08"
_CONCEPTO_COMPONENTE = {
    "AGUA_ANT": "AGUA", "MANT_ANT": "MANTENIMIENTO",
    "AGUA_ACT": "AGUA", "MANT_ACT": "MANTENIMIENTO",
    "CORTE_ANT": "CORTE_RECONEXION", "CORTE_ACT": "CORTE_RECONEXION",
    "CONVENIO": "CONVENIO", "ACUERDOS": "ACUERDOS", "MULTA": "MULTA",
}
_ORDEN_ABONO_CERRADO = (
    "AGUA_ANT", "MANT_ANT", "CORTE_ANT", "CORTE_ACT",
    "CONVENIO", "ACUERDOS", "MULTA", "AGUA_ACT", "MANT_ACT",
)
_ORDEN_CICLO = (
    "AGUA_ANT", "MANT_ANT", "AGUA_ACT", "MANT_ACT",
    "CORTE_ANT", "CORTE_ACT", "CONVENIO", "ACUERDOS", "MULTA",
)


def _mes_anterior(mes: str) -> str:
    year, month = map(int, mes.split("-"))
    month -= 1
    if month == 0:
        year, month = year - 1, 12
    return f"{year:04d}-{month:02d}"


def _componentes_cuenta(r: dict, mes_ano: str) -> dict[str, float]:
    """Deuda disponible por concepto y antigüedad para aplicar pagos por FIFO."""
    if mes_ano == _CORTE_LEDGER:
        agua_ant = max(round(r.get("mes_anterior", 0.0), 2), 0.0)
        mant_ant = 0.0
        corte_ant = min(max(round(r.get("corte_reconexion_base", r.get("corte_reconexion", 0.0)), 2), 0.0),
                        max(round(r.get("corte_reconexion", 0.0), 2), 0.0))
    else:
        anterior = _mes_anterior(mes_ano)
        agua_ant = max(round(repo.get_saldo(r["mz"], r["lt"], "AGUA", anterior), 2), 0.0)
        mant_ant = max(round(repo.get_saldo(r["mz"], r["lt"], "MANTENIMIENTO", anterior), 2), 0.0)
        corte_ant = max(round(repo.get_saldo(r["mz"], r["lt"], "CORTE_RECONEXION", anterior), 2), 0.0)

    partes_agua = {
        "AGUA_ACT": max(round(r.get("mes_actual", 0.0), 2), 0.0),
        "MANT_ACT": max(round(r.get("mantenimiento", 0.0), 2), 0.0),
        "AGUA_ANT": agua_ant,
        "MANT_ANT": mant_ant,
    }
    descuento = max(0.0, -round(r.get("blanco_final", 0.0), 2)) + max(
        0.0, -round(r.get("devolucion", 0.0), 2))
    for nombre in ("AGUA_ACT", "MANT_ACT", "AGUA_ANT", "MANT_ANT"):
        usado = min(partes_agua[nombre], descuento)
        partes_agua[nombre] = round(partes_agua[nombre] - usado, 2)
        descuento = round(descuento - usado, 2)

    corte_total = max(round(r.get("corte_reconexion", 0.0), 2), 0.0)
    return {
        **partes_agua,
        "CORTE_ANT": corte_ant, "CORTE_ACT": max(round(corte_total - corte_ant, 2), 0.0),
        "CONVENIO": max(round(r.get("convenio", 0.0), 2), 0.0),
        "ACUERDOS": max(round(r.get("acuerdos_asamblea", 0.0), 2), 0.0),
        "MULTA": max(round(r.get("multa", 0.0), 2), 0.0),
    }


def _aplicar_componentes(saldos: dict[str, float], monto: float,
                         orden: tuple[str, ...]) -> dict[str, float]:
    aplicado = {c: 0.0 for c in repo.CONCEPTOS_VALIDOS}
    restante = max(round(monto, 2), 0.0)
    for componente in orden:
        usar = min(saldos[componente], restante)
        saldos[componente] = round(saldos[componente] - usar, 2)
        concepto = _CONCEPTO_COMPONENTE[componente]
        aplicado[concepto] = round(aplicado[concepto] + usar, 2)
        restante = round(restante - usar, 2)
        if restante <= TOL:
            break
    return aplicado


def _aplicaciones_por_fuente(r: dict, mes_ano: str) -> tuple[dict[str, dict[str, float]], dict[str, float]]:
    saldos = _componentes_cuenta(r, mes_ano)
    cerrado = r.get("abono_rezagado_cerrado", 0.0)
    vigente = r.get("abono_rezagado_vigente",
                    r.get("abono_rezagado", 0.0) if "abono_rezagado_cerrado" not in r else 0.0)
    normal = r.get("total_pagado_normal", r.get("total_pagado", 0.0))
    pago_cerrado = _aplicar_componentes(saldos, cerrado, _ORDEN_ABONO_CERRADO)
    pago_normal = _aplicar_componentes(saldos, normal, _ORDEN_CICLO)
    pago_vigente = _aplicar_componentes(saldos, vigente, _ORDEN_CICLO)
    pago_abono = {c: round(pago_cerrado[c] + pago_vigente[c], 2)
                  for c in repo.CONCEPTOS_VALIDOS}
    return {"5_cobranza": pago_normal, "abonos_rezagados": pago_abono}, saldos


def _descomponer_pago(r: dict, monto: float) -> dict[str, float]:
    saldos = _componentes_cuenta(r, r.get("mes_ano", _CORTE_LEDGER))
    return _aplicar_componentes(saldos, monto, _ORDEN_CICLO)


def _descomponer_saldo(r: dict) -> tuple[list[float], list[float], float]:
    """Aplica total_pagado en prioridad P1→P5. Devuelve (comps[5], sin_cubrir[5], total).

    ⚠ El ORDEN de esta lista es un contrato con 3 lugares acoplados POR POSICIÓN:
    `_AC_P` (columnas de arrastre_consolidado), `_CONCEPTOS_PUEBLO` (índice →
    concepto del ledger) y `_CAMPOS_WATERFALL_REIDENTIFICACION`. Si se reordena
    acá, los tres se reordenan igual o los montos caen en la columna equivocada.
    """
    mes_ano = r.get("mes_ano", _CORTE_LEDGER)
    inicial = _componentes_cuenta(r, mes_ano)
    _por_fuente, pendientes = _aplicaciones_por_fuente(r, mes_ano)
    grupos = (
        ("AGUA_ANT", "MANT_ANT", "AGUA_ACT", "MANT_ACT"),
        ("CORTE_ANT", "CORTE_ACT"), ("CONVENIO",), ("ACUERDOS",), ("MULTA",),
    )
    comps = [round(sum(inicial[c] for c in grupo), 2) for grupo in grupos]
    sin_cubrir = [round(sum(pendientes[c] for c in grupo), 2) for grupo in grupos]
    return comps, sin_cubrir, round(sum(sin_cubrir), 2)


# ─────────────────────────────────────────────────────────────────────────────
#  Reconciliación de pagos hacia seguimiento_pueblo (MULTA/ACUERDOS/CONVENIO)
#  5_cobranza recalcula el mes ENTERO en cada corrida — no procesa pago por
#  pago. seguimiento_pueblo es append-only (nunca "sobreescribe" un total).
#  Por eso cada corrida reconcilia por DELTA: SET_DEBE (recién calculado) −
#  SET_TIENE (Σ PAGO ya en el registro) = lo que falta anotar.
#  No gateada por _ciclo_validado — los pagos son hechos, no esperan el
#  sello de fin de mes (a diferencia de arrastre_consolidado).
#  Ver docs/decisiones/seguimiento_pueblo.md
# ─────────────────────────────────────────────────────────────────────────────

# índice POSICIONAL en comps/sin_cubrir de _descomponer_saldo → concepto del ledger
_CONCEPTOS_PUEBLO = ((2, "CONVENIO"), (3, "ACUERDOS"), (4, "MULTA"))

# genesis_tardia usa nombres de CONCEPTO del lado planilla (ACUERDOS_ASAMBLEA)
# y el estado de cuenta usa ACUERDOS.
_GENESIS_TARDIA_CONCEPTO_A_PUEBLO = {
    "MULTA": "MULTA",
    "ACUERDOS": "ACUERDOS",
    "ACUERDOS_ASAMBLEA": "ACUERDOS",
    "CONVENIO": "CONVENIO",
}


def _cargos_genesis_tardia_snapshot(mes_ano: str) -> list[dict]:
    """CARGOs tardíos incluidos en la propuesta; no toca el ledger."""
    if not GENESIS_TARDIA_PATH.exists():
        return []
    df = pd.read_excel(GENESIS_TARDIA_PATH, header=1)
    df.columns = _norm_cols(df)
    cargos = []
    for _, f in df.iterrows():
        if str(f.get("MES_ANO_APLICA", "")).strip() != mes_ano:
            continue
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        concepto = _GENESIS_TARDIA_CONCEPTO_A_PUEBLO.get(str(f.get("CONCEPTO", "")).strip().upper())
        monto = _float(f.get("MONTO"))
        mes_origen = str(f.get("MES_ANO_ORIGEN", "")).strip() or mes_ano
        if not mz or not lt or not concepto or monto <= TOL:
            continue
        ref = f"genesis_tardia_{mes_ano}_{concepto}_{mz}_{lt}"
        cargos.append({
            "mz": mz, "lt": lt, "concepto": concepto, "mes": mes_origen,
            "monto": round(monto, 2), "source": "genesis_tardia",
            "audit_ref": ref, "clase": "GENESIS",
        })
    return cargos


def _objetivos_ledger(resultado: list[dict], mes_ano: str) -> list[dict]:
    """Construye SET_DEBE completo sin consultar ni modificar el ledger."""
    objetivos = []
    reasignaciones = _cargar_reasignaciones_aplicacion(mes_ano)
    for r in resultado:
        pagos_por_fuente, _pendientes = _aplicaciones_por_fuente(r, mes_ano)
        # Overlay reasignaciones_aplicacion: el pagador pidió que su abono NO
        # cubra CONCEPTO_ORIGEN (aunque le toque por prioridad de cascada) sino
        # CONCEPTO_DESTINO. El CARGO de origen sigue abierto (ver
        # _descomponer_saldo, no tocado acá) — solo se redirige a qué concepto
        # se anota el PAGO en seguimiento_pueblo.
        for origen, destino, monto in reasignaciones.get((r["mz"], r["lt"]), []):
            pagado_por_concepto = pagos_por_fuente["5_cobranza"]
            if origen not in pagado_por_concepto or destino not in pagado_por_concepto:
                continue
            delta = min(monto, pagado_por_concepto[origen])
            if delta <= TOL:
                continue
            pagado_por_concepto[origen] = round(pagado_por_concepto[origen] - delta, 2)
            pagado_por_concepto[destino] = round(pagado_por_concepto[destino] + delta, 2)
        for source, pagado_por_concepto in pagos_por_fuente.items():
            if source == "abonos_rezagados" and r.get("abono_rezagado", 0.0) <= TOL:
                continue
            clase_pago = "ABONO_REZAGADO" if source == "abonos_rezagados" else None
            for concepto in sorted(repo.CONCEPTOS_VALIDOS - {"OTROS"}):
                if concepto == "CONVENIO" and (r["mz"], r["lt"]) in repo.PREDIOS_INSTALACION_EXCLUIDOS:
                    continue
                objetivos.append({
                    "mz": r["mz"], "lt": r["lt"], "concepto": concepto,
                    "source": source, "monto_objetivo": round(pagado_por_concepto[concepto], 2),
                    "clase": clase_pago or "COBRANZA",
                })
    return sorted(objetivos, key=lambda x: (x["source"], x["mz"], x["lt"], x["concepto"]))


def _cargos_cuenta_snapshot(resultado: list[dict], mes_ano: str) -> list[dict]:
    """Cargos de apertura y del ciclo. Agosto es la frontera sin backfill."""
    cargos = []
    for r in resultado:
        componentes = _componentes_cuenta(r, mes_ano)
        candidatos = [
            ("AGUA", componentes["AGUA_ACT"], "2_planilla", f"cargo|{mes_ano}|AGUA|{r['mz']}|{r['lt']}",
             "consumo del ciclo calculado desde lecturas"),
            ("MANTENIMIENTO", componentes["MANT_ACT"], "2_planilla",
             f"cargo|{mes_ano}|MANTENIMIENTO|{r['mz']}|{r['lt']}", "mantenimiento del ciclo"),
            ("CORTE_RECONEXION", componentes["CORTE_ACT"], "6_corte",
             f"cargo|{mes_ano}|CORTE_RECONEXION|{r['mz']}|{r['lt']}", "penalidad emitida por 6_corte"),
        ]
        if mes_ano == _CORTE_LEDGER:
            candidatos.extend([
                ("AGUA", componentes["AGUA_ANT"], "saldo_inicial",
                 f"apertura|{mes_ano}|AGUA|{r['mz']}|{r['lt']}", "saldo de agua al cierre de julio"),
                ("CORTE_RECONEXION", componentes["CORTE_ANT"], "saldo_inicial",
                 f"apertura|{mes_ano}|CORTE_RECONEXION|{r['mz']}|{r['lt']}",
                 "saldo de corte/reconexión al cierre de julio"),
            ])
        for concepto, monto, source, audit_ref, motivo in candidatos:
            if monto > TOL:
                cargos.append({
                    "mz": r["mz"], "lt": r["lt"], "concepto": concepto, "mes": mes_ano,
                    "monto": round(monto, 2), "source": source, "audit_ref": audit_ref,
                    "clase": "GENESIS", "motivo": motivo,
                })
    return cargos


def _exportar_snapshot_ledger(resultado: list[dict], mes_ano: str) -> str:
    payload = {
        "schema": 2,
        "mes": mes_ano,
        "objetivos": _objetivos_ledger(resultado, mes_ano),
        "cargos": sorted(_cargos_genesis_tardia_snapshot(mes_ano)
                         + _cargos_cuenta_snapshot(resultado, mes_ano),
                         key=lambda x: (x["mz"], x["lt"], x["concepto"], x["audit_ref"])),
    }
    normalizado = json.dumps(payload, ensure_ascii=True, sort_keys=True,
                             separators=(",", ":")).encode("utf-8")
    snapshot_hash = hashlib.sha256(normalizado).hexdigest()
    documento = {**payload, "snapshot_hash": snapshot_hash}
    ruta = OUTPUTS_DIR / f"snapshot_ledger_{mes_ano}.json"
    tmp = ruta.with_suffix(".json.tmp")
    tmp.write_text(json.dumps(documento, ensure_ascii=True, sort_keys=True, indent=2),
                   encoding="utf-8")
    tmp.replace(ruta)
    log.info(f"{ruta.name} → {len(payload['objetivos'])} objetivos · hash {snapshot_hash[:12]}")
    return snapshot_hash


def _exportar_arrastre_consolidado(wb: Workbook, resultado: list[dict], mes_ano: str):
    filas = []
    for r in resultado:
        _comps, sin_cubrir, total_arr = _descomponer_saldo(r)
        if total_arr > TOL:
            filas.append((r["mz"], r["lt"], r["nombre"], sin_cubrir, total_arr))

    last_row = max(len(filas) + 2, 3)
    ws = wb.create_sheet("arrastre_consolidado")
    ws.freeze_panes = "A3"

    # Group headers (fila 1)
    _gh(ws, 1, 1, 3, "¿Quién es?", *GH_AC_QUIEN)
    for i, (_, hbg, htx, _, _) in enumerate(_AC_P):
        _gh(ws, 1, 5 + i, 5 + i, f"P{i + 1}", hbg, htx)
    _gh(ws, 1, 11, 11, "Cierre", *GH_AC_TOTAL)

    # Column headers (fila 2)
    _ch(ws, 2, 1, "MZ",     *GH_AC_QUIEN)
    _ch(ws, 2, 2, "LT",     *GH_AC_QUIEN)
    _ch(ws, 2, 3, "NOMBRE", *GH_AC_QUIEN)
    for i, (nombre, hbg, htx, _, _) in enumerate(_AC_P):
        _ch(ws, 2, 5 + i, nombre, hbg, htx)
    _ch(ws, 2, 11, "TOTAL_ARRASTRE", *GH_AC_TOTAL)

    for sc in (4, 10):
        _sep(ws, sc, last_row)
    # anchos por posición P1..P5 = columnas 5..9 (P3 CONVENIO, P5 MULTA)
    for col, ancho in ((1, 6), (2, 6), (3, 26),
                       (5, 12), (6, 16), (7, 10), (8, 17), (9, 8), (11, 15)):
        _w(ws, col, ancho)
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    MONEY = '"S/ "#,##0.00'
    for ri, (mz, lt, nombre, sin_cubrir, total_arr) in enumerate(filas, 3):
        _c(ws, ri, 1, mz,     TD_AC_QUIEN, "065F46", mono=True, align="center")
        _c(ws, ri, 2, lt,     TD_AC_QUIEN, "065F46", mono=True, align="center")
        _c(ws, ri, 3, nombre, TD_AC_QUIEN, "333333", align="left")
        for i, val in enumerate(sin_cubrir):
            _, _, _, cbg, ctx = _AC_P[i]
            if val > TOL:
                _c(ws, ri, 5 + i, val, cbg, ctx,
                   mono=True, align="right", bold=True, fmt=MONEY)
            else:
                _c(ws, ri, 5 + i, "—", TD_AC_ZERO, "9CA3AF", align="center")
        _c(ws, ri, 11, total_arr, TD_AC_TOTAL, "1E5C3A",
           mono=True, align="right", bold=True, fmt=MONEY)
        ws.row_dimensions[ri].height = 17

    log.info(f"arrastre_consolidado → {len(filas)} usuarios con TOTAL_ARRASTRE>0")


def _guardar_planilla_cobrado(wb: Workbook, ruta: Path, mes_ano: str) -> None:
    _backup_arrastre_devolucion(ruta)
    temporal = ruta.with_name(f".{ruta.stem}.tmp.xlsx")
    wb.save(temporal)
    temporal.replace(ruta)

    for nombre in (f"arrastre_consolidado_{mes_ano}.xlsx",
                   f"arrastre_devolucion_{mes_ano}.xlsx"):
        legado = OUTPUTS_DIR / nombre
        if legado.exists():
            if nombre.startswith("arrastre_devolucion_"):
                _backup_arrastre_devolucion(legado)
            legado.unlink()
    log.info(f"{ruta.name} → hojas: {', '.join(wb.sheetnames)}")


# ─────────────────────────────────────────────────────────────────────────────
#  OUTPUT 7 — discrepancias_cobranza.xlsx
#  Pagos cuyo MZ+LT no existe en planilla — no pudieron imputarse a un usuario.
#  Layout (matching formato_discrepancias_cobranza.html):
#
#  Hoja discrepancias_pago_yape (cols 1..9):
#    1-2  predio    MZ LT                              (rojo — el que no existe)
#    3    sep
#    4-5  pago      MONTO FECHA                        (azul)
#    6    sep
#    7    origen    ORIGEN                             (ambar — pista de origen)
#    8    sep
#    9-10 traz      CICLO_CORRECCION MOTIVO            (morado)
#
#  Hoja discrepancias_pago_efectivo (cols 1..11):
#    1-2  predio    MZ LT
#    3    sep
#    4-5  pago      MONTO FECHA
#    6    sep
#    7-8  origen    MESA COBRADOR                      (pista fisica)
#    9    sep
#    10-11 traz     CICLO_CORRECCION MOTIVO
#
#  Si no hay discrepancias en ninguna fuente → borrar el archivo si existe
#  (su presencia es la senal de que hay trabajo pendiente, como en 4_pagos/efectivo).
# ─────────────────────────────────────────────────────────────────────────────

_DC_MOTIVO = "predio no encontrado en planilla"

_DC_YAPE_GRUPOS = [
    (1,  2,  "¿Dónde vive?",      *GH_DC_PREDIO),
    (4,  5,  "¿Cuánto y cuándo?", *GH_DC_PAGO),
    (7,  7,  "¿Quién pagó?",      *GH_DC_ORIGEN),
    (9,  10, "Trazabilidad",      *GH_DC_TRAZ),
    (12, 13, "¿Corrección?",      *GH_DC_CORR),
]
_DC_YAPE_COLS = [
    (1,  "MZ",               *GH_DC_PREDIO,  8),
    (2,  "LT",               *GH_DC_PREDIO,  8),
    (4,  "MONTO",            *GH_DC_PAGO,   12),
    (5,  "FECHA",            *GH_DC_PAGO,   12),
    (7,  "ORIGEN",           *GH_DC_ORIGEN, 22),
    (9,  "CICLO_CORRECCION", *GH_DC_TRAZ,   16),
    (10, "MOTIVO",           *GH_DC_TRAZ,   34),
    (12, "MZ_CORRECTO",      *GH_DC_CORR,    9),
    (13, "LT_CORRECTO",      *GH_DC_CORR,    9),
]
_DC_YAPE_SEP_COLS = [3, 6, 8, 11]

_DC_EFEC_GRUPOS = [
    (1,  2,  "¿Dónde vive?",      *GH_DC_PREDIO),
    (4,  5,  "¿Cuánto y cuándo?", *GH_DC_PAGO),
    (7,  8,  "¿De qué mesa?",     *GH_DC_ORIGEN),
    (10, 11, "Trazabilidad",      *GH_DC_TRAZ),
    (13, 14, "¿Corrección?",      *GH_DC_CORR),
]
_DC_EFEC_COLS = [
    (1,  "MZ",               *GH_DC_PREDIO,  8),
    (2,  "LT",               *GH_DC_PREDIO,  8),
    (4,  "MONTO",            *GH_DC_PAGO,   12),
    (5,  "FECHA",            *GH_DC_PAGO,   12),
    (7,  "MESA",             *GH_DC_ORIGEN, 12),
    (8,  "COBRADOR",         *GH_DC_ORIGEN, 22),
    (10, "CICLO_CORRECCION", *GH_DC_TRAZ,   16),
    (11, "MOTIVO",           *GH_DC_TRAZ,   34),
    (13, "MZ_CORRECTO",      *GH_DC_CORR,    9),
    (14, "LT_CORRECTO",      *GH_DC_CORR,    9),
]
_DC_EFEC_SEP_COLS = [3, 6, 9, 12]


def _leer_correcciones_tipeadas(ruta) -> dict:
    """Lee MZ_CORRECTO/LT_CORRECTO ya tipeados en discrepancias_cobranza.xlsx,
    keyed por (sheet, mz, lt), para re-mostrarlos al regenerar el archivo y no
    borrar el trabajo manual de un huérfano que persiste entre corridas."""
    tipeadas = {}
    if not ruta.exists():
        return tipeadas
    try:
        wb = load_workbook(ruta, data_only=True)
    except Exception as e:
        log.warning(f"No se pudo leer correcciones tipeadas: {e}")
        return tipeadas
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        hdrs = {str(ws.cell(2, c).value or "").strip().upper(): c
                for c in range(1, ws.max_column + 1)}
        cmzo, clto = hdrs.get("MZ"), hdrs.get("LT")
        cmzc, cltc = hdrs.get("MZ_CORRECTO"), hdrs.get("LT_CORRECTO")
        if not all([cmzo, clto, cmzc, cltc]):
            continue
        for r in range(3, ws.max_row + 1):
            mo = _norm_mz(ws.cell(r, cmzo).value)
            lo = _norm_lt(ws.cell(r, clto).value)
            mc = ws.cell(r, cmzc).value
            lc = ws.cell(r, cltc).value
            if mo and lo and (mc not in (None, "") or lc not in (None, "")):
                tipeadas[(sheet, mo, lo)] = (mc, lc)
    return tipeadas


def _exportar_discrepancias_cobranza(disc_yape: list[dict], disc_efec: list[dict]):
    """
    Genera 5_cobranza/outputs/discrepancias_cobranza.xlsx con dos hojas:
      - discrepancias_pago_yape:     pagos Yape huerfanos (ORIGEN como pista)
      - discrepancias_pago_efectivo: cobros en mesa huerfanos (MESA+COBRADOR como pista)
    Si no hay ninguna discrepancia → borra el archivo si existe.
    La presencia del archivo es la senal de que hay trabajo pendiente.
    """
    ruta = OUTPUTS_DIR / "discrepancias_cobranza.xlsx"

    if not disc_yape and not disc_efec:
        if ruta.exists():
            ruta.unlink()
            log.info("discrepancias_cobranza.xlsx eliminado — todo imputado")
        return

    # Preservar correcciones ya tipeadas por el operador (no borrarlas al regenerar)
    tipeadas = _leer_correcciones_tipeadas(ruta)

    MONEY = '"S/ "#,##0.00'
    wb = Workbook()

    # ── Hoja 1: discrepancias_pago_yape ──────────────────────────────────────
    ws = wb.active
    ws.title = "discrepancias_pago_yape"
    ws.freeze_panes = "A3"

    last_row = max(len(disc_yape) + 2, 3)
    for cs, ce, texto, bg, txt in _DC_YAPE_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for sc in _DC_YAPE_SEP_COLS:
        _sep(ws, sc, last_row)
    for col, nombre, bg, txt, ancho in _DC_YAPE_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    # Orden estable: MZ, LT, FECHA, ORIGEN — para que re-runs produzcan el mismo archivo
    for ri, p in enumerate(sorted(disc_yape,
                                   key=lambda x: (x["mz"], x["lt"], x["fecha"], x["origen"])),
                            3):
        _c(ws, ri, 1,  p["mz"],               TD_DC_PREDIO, "991B1B", mono=True, align="center", bold=True)
        _c(ws, ri, 2,  p["lt"],               TD_DC_PREDIO, "991B1B", mono=True, align="center", bold=True)
        _c(ws, ri, 4,  p["monto"],            TD_DC_PAGO,   "1A5276", mono=True, align="right",  fmt=MONEY)
        _c(ws, ri, 5,  p["fecha"],            TD_DC_PAGO,   "1A5276", mono=True, align="center")
        _c(ws, ri, 7,  p.get("origen", ""),   TD_DC_ORIGEN, "7D6608", align="left")
        _c(ws, ri, 9,  p["ciclo_correccion"], TD_DC_TRAZ,   "5B21B6", mono=True, align="center")
        _c(ws, ri, 10, _DC_MOTIVO,            TD_DC_TRAZ,   "5B21B6", align="left")
        mzc, ltc = tipeadas.get(("discrepancias_pago_yape", _norm_mz(p["mz"]), _norm_lt(p["lt"])), (None, None))
        _c(ws, ri, 12, mzc, TD_DC_CORR_V, GH_DC_CORR[1], mono=True, align="center")
        _c(ws, ri, 13, ltc, TD_DC_CORR_V, GH_DC_CORR[1], mono=True, align="center")
        ws.row_dimensions[ri].height = 17

    # ── Hoja 2: discrepancias_pago_efectivo ──────────────────────────────────
    ws = wb.create_sheet("discrepancias_pago_efectivo")
    ws.freeze_panes = "A3"

    last_row = max(len(disc_efec) + 2, 3)
    for cs, ce, texto, bg, txt in _DC_EFEC_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for sc in _DC_EFEC_SEP_COLS:
        _sep(ws, sc, last_row)
    for col, nombre, bg, txt, ancho in _DC_EFEC_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    for ri, p in enumerate(sorted(disc_efec,
                                   key=lambda x: (x["mz"], x["lt"], x["fecha"], x["mesa"])),
                            3):
        _c(ws, ri, 1,  p["mz"],               TD_DC_PREDIO, "991B1B", mono=True, align="center", bold=True)
        _c(ws, ri, 2,  p["lt"],               TD_DC_PREDIO, "991B1B", mono=True, align="center", bold=True)
        _c(ws, ri, 4,  p["monto"],            TD_DC_PAGO,   "1A5276", mono=True, align="right",  fmt=MONEY)
        _c(ws, ri, 5,  p["fecha"],            TD_DC_PAGO,   "1A5276", mono=True, align="center")
        _c(ws, ri, 7,  p.get("mesa", ""),     TD_DC_ORIGEN, "7D6608", align="center")
        _c(ws, ri, 8,  p.get("cobrador", ""), TD_DC_ORIGEN, "7D6608", align="left")
        _c(ws, ri, 10, p["ciclo_correccion"], TD_DC_TRAZ,   "5B21B6", mono=True, align="center")
        _c(ws, ri, 11, _DC_MOTIVO,            TD_DC_TRAZ,   "5B21B6", align="left")
        mzc, ltc = tipeadas.get(("discrepancias_pago_efectivo", _norm_mz(p["mz"]), _norm_lt(p["lt"])), (None, None))
        _c(ws, ri, 13, mzc, TD_DC_CORR_V, GH_DC_CORR[1], mono=True, align="center")
        _c(ws, ri, 14, ltc, TD_DC_CORR_V, GH_DC_CORR[1], mono=True, align="center")
        ws.row_dimensions[ri].height = 17

    wb.save(ruta)
    log.info(f"discrepancias_cobranza.xlsx → "
             f"{len(disc_yape)} yape · {len(disc_efec)} efectivo")


# ─────────────────────────────────────────────────────────────────────────────
#  RETROESCRITURA — CICLO_COBRANZA en pagos_yape y pagos_efectivo
# ─────────────────────────────────────────────────────────────────────────────
def _retroescribir_ciclo(path: Path, ciclo_col_nombre: str,
                         filas_nuevas: list[int], ciclo_nuevo: int):
    """Agrega o actualiza columna CICLO_COBRANZA en el archivo de pagos.
    filas_nuevas: filas (1-indexed) que recibieron ciclo_nuevo en este run.
    """
    wb = load_workbook(path)
    ws = wb.active

    # Detectar columna CICLO_COBRANZA en fila 2 (cabecera de columnas).
    cob_col = None
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        if str(ws.cell(row=2, column=col).value or "").strip().upper() == "CICLO_COBRANZA":
            cob_col = col
            break

    if cob_col is None:
        # Crear nueva columna al final
        cob_col = max_col + 1
        # Header grupo (fila 1) — usar el mismo grupo que la columna ciclo_correccion
        ref_col = None
        for col in range(1, max_col + 1):
            if str(ws.cell(row=2, column=col).value or "").strip().upper() == ciclo_col_nombre:
                ref_col = col
                break
        if ref_col:
            ref_header_row1 = ws.cell(row=1, column=ref_col).value
            c1 = ws.cell(row=1, column=cob_col, value=ref_header_row1 or "¿Cuándo?")
            c1.font      = Font(name="Arial", size=8, bold=True, color="7D6608")
            c1.fill      = PatternFill("solid", start_color="FEF9E7")
            c1.alignment = Alignment(horizontal="center", vertical="center")
        c2 = ws.cell(row=2, column=cob_col, value="CICLO_COBRANZA")
        c2.font      = Font(name="Arial", size=9, bold=True, color="7D6608")
        c2.fill      = PatternFill("solid", start_color="FEF9E7")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(cob_col)].width = 14

    filas_nuevas_set = set(filas_nuevas)
    for r in filas_nuevas_set:
        if r > ws.max_row:
            continue
        cell = ws.cell(row=r, column=cob_col)
        # Solo escribir si está vacío (para preservar ciclos previos)
        if cell.value in (None, ""):
            cell.value     = ciclo_nuevo
            cell.font      = Font(name="Consolas", size=9, color="7D6608")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill      = PatternFill("solid", start_color="FEFCE8")

    wb.save(path)
    log.info(f"{path.name} → CICLO_COBRANZA={ciclo_nuevo} en {len(filas_nuevas_set)} filas")


# ─────────────────────────────────────────────────────────────────────────────
#  ACTUALIZAR BLANCOS APLICADOS
# ─────────────────────────────────────────────────────────────────────────────
def _actualizar_blancos(blancos_aplicados: set, mes_ano: str):
    if not blancos_aplicados or not BLANCOS_PATH.exists():
        return
    wb = load_workbook(BLANCOS_PATH)
    ws = wb.active
    if str(ws.cell(2, _BL_MES_APLICADO).value or "").strip().upper() != "MES_ANO_APLICADO":
        ws.cell(2, _BL_MES_APLICADO).value = "MES_ANO_APLICADO"
    for r in range(3, ws.max_row + 1):
        mz = _norm_mz(ws.cell(r, _BL_MZ).value)
        lt = _norm_lt(ws.cell(r, _BL_LOTE).value)
        if f"{mz}-{lt}" in blancos_aplicados:
            ws.cell(r, _BL_EST).value = "aplicado"
            ws.cell(r, _BL_MES_APLICADO).value = mes_ano
    wb.save(BLANCOS_PATH)
    log.info(f"blancos_acumulados.xlsx → {len(blancos_aplicados)} marcados aplicado/{mes_ano}")


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "═" * 60)
    print("  5_cobranza — Carga de pagos · estado de cobro")
    print("═" * 60)
    _init_logging()

    print("\n[1/6] Validando inputs...")
    plan_path = _validar_inputs()

    print("\n[2/6] Cargando datos...")
    usuarios, mes_ano = _cargar_planilla(plan_path)
    pagos_yape        = _cargar_pagos_yape()
    pagos_efectivo    = _cargar_pagos_efectivo()
    blancos           = _cargar_blancos(mes_ano)
    aportes_tanque    = _cargar_aportes_tanque_manuales(mes_ano)
    abonos_rezagados  = _cargar_abonos_rezagados(mes_ano)
    dev_yape          = _cargar_retornos_yape()
    dev_efec          = _cargar_retornos_efectivo()
    dev_devuelto      = _cargar_devueltos_yape()
    traz_path         = OUTPUTS_DIR / "trazabilidad_cobranza.xlsx"
    ids_previas, max_ciclo = _cargar_trazabilidad_previa(pagos_yape + pagos_efectivo)

    print("\n[2b/6] Aplicando correcciones de lote...")
    correcciones   = _leer_correcciones()
    _keys_validos_correcciones = {u["key"] for u in usuarios}
    correcciones   = _recuperar_correcciones_trazabilidad(correcciones, max_ciclo,
                                                           _keys_validos_correcciones)
    correcciones   = _absorber_correcciones_discrepancias(correcciones, max_ciclo)
    pagos_yape     = _aplicar_correcciones_lote(pagos_yape,     correcciones)
    pagos_efectivo = _aplicar_correcciones_lote(pagos_efectivo, correcciones)

    print("\n[3/6] Detectando ciclo de cobranza...")
    # Huérfanos: pagos cuyo MZ+LT no existe en planilla. Nunca se escriben en
    # trazabilidad (ver _exportar_trazabilidad_cobranza) → se excluyen de la
    # identidad de idempotencia para que pagos_nuevos quede vacío sin cambios.
    _keys_validos = {u["key"] for u in usuarios}
    _huerfanos    = ({p["key"] for p in pagos_yape} |
                    {p["key"] for p in pagos_efectivo}) - _keys_validos

    ids_actuales = {_identidad_pago(p) for p in (pagos_yape + pagos_efectivo)
                    if p["key"] not in _huerfanos}
    pagos_nuevos = ids_actuales - ids_previas

    # Discrepancias: se exportan siempre — el archivo actúa como señal de trabajo
    # pendiente independientemente de si el ciclo es idempotente o no.
    disc_yape = [p for p in pagos_yape     if p["key"] in _huerfanos]
    disc_efec = [p for p in pagos_efectivo if p["key"] in _huerfanos]
    _exportar_discrepancias_cobranza(disc_yape, disc_efec)

    # Idempotencia: tambien comparar retornos contra el estado previo.
    # Si pagos no cambian pero retornos si → re-generar sin avanzar ciclo.
    retornos_actuales  = _retornos_por_lote(dev_yape, dev_efec)
    devueltos_actuales = _devueltos_por_lote(dev_devuelto)
    retornos_previos  = _retornos_planilla_previa(mes_ano)
    retornos_cambiados = retornos_actuales != retornos_previos

    force = "--force" in sys.argv
    if not pagos_nuevos and not retornos_cambiados and not force:
        log.info(f"Sin cambios (pagos ni retornos) · ciclo {max_ciclo} → idempotente")
        print(f"\n  Idempotencia: no hay pagos ni retornos nuevos")
        print(f"  Último ciclo cargado: {max_ciclo}")
        if disc_yape or disc_efec:
            print(f"  · discrepancias_cobranza.xlsx     "
                  f"({len(disc_yape)} yape · {len(disc_efec)} efectivo)")
            print(f"    → Pagos cuyo MZ+LT no existe en planilla — corregir el archivo de origen")
        print("\n" + "═" * 60 + "\n")
        return

    if pagos_nuevos:
        ciclo_nuevo = max_ciclo + 1
        log.info(f"Ciclo nuevo = {ciclo_nuevo} · pagos nuevos = {len(pagos_nuevos)}")
    else:
        ciclo_nuevo = max_ciclo if max_ciclo > 0 else 1
        log.info(f"Sin pagos nuevos · retornos cambiaron → re-generando en ciclo {ciclo_nuevo}")

    print("\n[4/6] Calculando cobranza...")
    resultado, blancos_usados = _calcular(
        usuarios, pagos_yape, pagos_efectivo, blancos,
        dev_yape, dev_efec, dev_devuelto,
        ciclo_nuevo, pagos_nuevos,
        aportes_tanque=aportes_tanque,
        abonos_rezagados=abonos_rezagados,
    )
    print("\n[5/6] Exportando outputs...")
    planilla_path = OUTPUTS_DIR / f"planilla_cobrado_{mes_ano}.xlsx"
    previo = _leer_revision_previa(planilla_path, "arrastre_devolucion")
    if previo is None:
        previo = _leer_revision_previa(
            OUTPUTS_DIR / f"arrastre_devolucion_{mes_ano}.xlsx"
        ) or {}
    wb_planilla, planilla_path = _exportar_planilla_cobrado(resultado, mes_ano)
    # Contar elegibles para corte — la lista_corte la genera 6_corte leyendo
    # SALDO + MES_ANTERIOR desde planilla_cobrado. Acá solo se reporta el conteo.
    n_corte = sum(1 for r in resultado
                  if r["saldo"] > TOL and r["mes_anterior"] >= ARRASTRE_MIN - TOL)
    _exportar_trazabilidad_cobranza(
        resultado, pagos_yape, pagos_efectivo,
        ciclo_nuevo, pagos_nuevos, traz_path,
        retornos_actuales, devueltos_actuales,
    )
    _exportar_resumen(resultado, n_corte, mes_ano, ciclo_nuevo)
    _exportar_arrastre_deuda(resultado, mes_ano)
    _exportar_arrastre_consolidado(wb_planilla, resultado, mes_ano)
    _exportar_arrastre_devolucion(
        wb_planilla, resultado, mes_ano, previo, disc_yape, disc_efec
    )
    _guardar_planilla_cobrado(wb_planilla, planilla_path, mes_ano)
    snapshot_hash = _exportar_snapshot_ledger(resultado, mes_ano)
    _marcar_generado(mes_ano, snapshot_hash)  # 5b sella exactamente esta versión

    print("\n[6/6] Retroescritura y blancos...")
    filas_yape_nuevas = [p["row"] for p in pagos_yape
                         if _identidad_pago(p) in pagos_nuevos]
    filas_efec_nuevas = [p["row"] for p in pagos_efectivo
                         if _identidad_pago(p) in pagos_nuevos]
    if filas_yape_nuevas:
        _retroescribir_ciclo(_pago_path(YAPE_DIR, _YAPE_BASE), "CICLO",
                             filas_yape_nuevas, ciclo_nuevo)
    if filas_efec_nuevas:
        _retroescribir_ciclo(_pago_path(EFEC_DIR, _EFEC_BASE), "CICLO_CORRECCION",
                             filas_efec_nuevas, ciclo_nuevo)
    _actualizar_blancos(blancos_usados, mes_ano)

    n_pend = sum(1 for r in resultado if r["estado"] in ("PARCIAL", "PENDIENTE"))
    print("\n" + "═" * 60)
    print(f"  Cobranza completada · ciclo {ciclo_nuevo} · {mes_ano}")
    print(f"  Outputs → 5_cobranza/outputs/")
    print(f"  · planilla_cobrado_{mes_ano}.xlsx  "
          f"({len(resultado)} usuarios · 3 hojas)")
    print(f"  · trazabilidad_cobranza.xlsx")
    print(f"  · resumen_recaudacion.xlsx")
    n_exceso = sum(1 for r in resultado if r["saldo"] < -TOL)
    print(f"  · arrastre_deuda_{mes_ano}.xlsx      ({sum(1 for r in resultado if r['saldo'] > TOL)} pendientes)")
    print(f"    hojas: planilla_cobrado · arrastre_consolidado · arrastre_devolucion "
          f"({n_exceso} excesos)")
    if disc_yape or disc_efec:
        print(f"  · discrepancias_cobranza.xlsx     "
              f"({len(disc_yape)} yape · {len(disc_efec)} efectivo)")
        print(f"    → Pagos cuyo MZ+LT no existe en planilla — corregir el archivo de origen")
    if n_corte:
        print(f"\n  → {n_corte} usuarios elegibles para corte "
              f"(SALDO>0 & MES_ANT>=8) — correr 6_corte/generar_lista.py")
    print("═" * 60 + "\n")


if __name__ == "__main__":
    main()
