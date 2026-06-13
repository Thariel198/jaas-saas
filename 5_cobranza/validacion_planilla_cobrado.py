"""5_cobranza/validacion_planilla_cobrado.py — coherencia de outputs

Lee los outputs de 5_cobranza y corre 4 bloques de validacion:

  Bloque 1 — Re-calculo SALDO en planilla_cobrado
  Bloque 2 — Trazabilidad ↔ planilla_cobrado (sumas por fuente · huerfanos)
  Bloque 3 — Arrastres ↔ planilla_cobrado (deuda y devolucion, espejo)
  Bloque 4 — Reporte: consola, run_validacion.log, validacion_errores.xlsx

Uso:
    python validacion_planilla_cobrado.py

Exit code 0 = todo OK · 1 = se encontraron errores.
"""
import logging
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side

ROOT          = Path(__file__).resolve().parent
OUTPUTS_DIR   = ROOT / "outputs"
PLANILLA_PATH = OUTPUTS_DIR / "planilla_cobrado.xlsx"
TRAZ_PATH     = OUTPUTS_DIR / "trazabilidad_cobranza.xlsx"

TOL = 0.005
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
#  UTILIDADES (paridad con main.py)
# ─────────────────────────────────────────────────────────────────────────────
def _norm_mz(val) -> str:
    s = str(val).strip().upper()
    return "" if not s or s in ("NAN", "NONE") else s

def _norm_lt(val) -> str:
    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN"):
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.upper()

def _float(val) -> float:
    if val is None:
        return 0.0
    try:
        f = float(str(val).replace(",", ".").strip())
        return 0.0 if f != f else f
    except (ValueError, TypeError):
        return 0.0

def _norm_cols(df) -> list[str]:
    return [str(c).strip().upper() for c in df.columns]

def _estado_esperado(saldo: float, pagado: float) -> str:
    if saldo < -TOL:
        return "EXCESO"
    if abs(saldo) <= TOL:
        return "CANCELADO"
    return "PARCIAL" if pagado > TOL else "PENDIENTE"

def _saldo_de(r: dict) -> float:
    """Re-calcula SALDO desde los 9 conceptos · no depende de TOTAL_A_PAGAR (formula)."""
    cargos = (r["mes_actual"] + r["mantenimiento"] + r["mes_anterior"]
              + r["corte_reconexion"] + r["convenio"] + r["multa"]
              + r["acuerdos"])
    descuentos = r["blanco"] + r["devolucion"]   # vienen negativos
    total = round(cargos + descuentos, 2)
    return round(total - r["monto_yape"] - r["monto_efectivo"], 2)


# ─────────────────────────────────────────────────────────────────────────────
#  LECTORES
# ─────────────────────────────────────────────────────────────────────────────
def _leer_planilla() -> list[dict]:
    """planilla_cobrado.xlsx con openpyxl + data_only=True.

    TOTAL_A_PAGAR es formula `=SUM(J:R)`. Si el archivo fue escrito por
    openpyxl sin abrirse en Excel, no hay cache → llega como None.
    El re-calculo desde conceptos es la fuente de verdad; TOTAL_A_PAGAR
    solo se compara cuando hay cache.
    """
    if not PLANILLA_PATH.exists():
        raise FileNotFoundError(f"Falta: {PLANILLA_PATH}")

    wb = load_workbook(PLANILLA_PATH, data_only=True)
    ws = wb.active

    # Fila 1 = grupos · Fila 2 = nombres reales de columnas
    headers = [str(c.value).strip().upper() if c.value else "" for c in ws[2]]
    col_idx = {h: i for i, h in enumerate(headers) if h}

    requeridas = {"MZ", "LT", "TOTAL_A_PAGAR", "MONTO_YAPE",
                  "MONTO_EFECTIVO", "ESTADO"}
    faltantes = requeridas - set(col_idx)
    if faltantes:
        raise ValueError(f"Columnas faltantes en planilla_cobrado: {faltantes}")

    def _g(row, col):
        i = col_idx.get(col)
        return row[i] if i is not None else None

    filas = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        mz = _norm_mz(_g(row, "MZ"))
        lt = _norm_lt(_g(row, "LT"))
        if not mz or not lt:
            continue
        filas.append({
            "mz":               mz,
            "lt":               lt,
            "nombre":           str(_g(row, "NOMBRE") or "").strip(),
            "mes_ano":          str(_g(row, "MES_ANO") or "").strip(),
            "mes_actual":       _float(_g(row, "MES_ACTUAL")),
            "mantenimiento":    _float(_g(row, "MANTENIMIENTO")),
            "mes_anterior":     _float(_g(row, "MES_ANTERIOR")),
            "corte_reconexion": _float(_g(row, "CORTE_RECONEXION")),
            "convenio":         _float(_g(row, "CONVENIO")),
            "multa":            _float(_g(row, "MULTA")),
            "acuerdos":         _float(_g(row, "ACUERDOS_ASAMBLEA")),
            "blanco":           _float(_g(row, "BLANCO")),
            "devolucion":       _float(_g(row, "DEVOLUCION")),
            "total_a_pagar_raw": _g(row, "TOTAL_A_PAGAR"),   # None si formula sin cache
            "monto_yape":       _float(_g(row, "MONTO_YAPE")),
            "monto_efectivo":   _float(_g(row, "MONTO_EFECTIVO")),
            "estado":           str(_g(row, "ESTADO") or "").strip().upper(),
        })
    return filas


def _leer_trazabilidad() -> list[dict]:
    if not TRAZ_PATH.exists():
        log.warning(f"No existe {TRAZ_PATH.name} → Bloque 2 sin datos")
        return []
    df = pd.read_excel(TRAZ_PATH, header=1)
    df.columns = _norm_cols(df)
    filas = []
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        filas.append({
            "mz":     mz,
            "lt":     lt,
            "nombre": str(f.get("NOMBRE", "")).strip(),
            "monto":  round(_float(f.get("MONTO")), 2),
            "fuente": str(f.get("FUENTE", "")).strip().lower(),
        })
    return filas


def _leer_arrastre(path: Path) -> list[dict]:
    if not path.exists():
        return []
    df = pd.read_excel(path, header=1)
    df.columns = _norm_cols(df)
    filas = []
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        filas.append({
            "mz":             mz,
            "lt":             lt,
            "nombre":         str(f.get("NOMBRE", "")).strip(),
            "monto":          round(_float(f.get("MONTO")), 2),
            "mes_ano_origen": str(f.get("MES_ANO_ORIGEN", "")).strip(),
        })
    return filas


def _buscar_arrastre(prefijo: str, mes_ano: str) -> Path:
    """outputs/{prefijo}_{mes_ano}.xlsx · fallback al mas reciente que matchee."""
    exacto = OUTPUTS_DIR / f"{prefijo}_{mes_ano}.xlsx"
    if exacto.exists():
        return exacto
    candidatos = sorted(OUTPUTS_DIR.glob(f"{prefijo}_*.xlsx"), reverse=True)
    return candidatos[0] if candidatos else exacto   # path inexistente → falla limpio


# ─────────────────────────────────────────────────────────────────────────────
#  BLOQUE 1 — Re-calculo SALDO + ESTADO
# ─────────────────────────────────────────────────────────────────────────────
def _validar_saldo(planilla: list[dict]) -> list[dict]:
    errores = []
    for r in planilla:
        cargos = (r["mes_actual"] + r["mantenimiento"] + r["mes_anterior"]
                  + r["corte_reconexion"] + r["convenio"] + r["multa"]
                  + r["acuerdos"])
        descuentos = r["blanco"] + r["devolucion"]
        total_esperado = round(cargos + descuentos, 2)
        pagado = round(r["monto_yape"] + r["monto_efectivo"], 2)
        saldo_esperado = round(total_esperado - pagado, 2)
        estado_esperado = _estado_esperado(saldo_esperado, pagado)

        # TOTAL_A_PAGAR es formula. Solo verifico si hay cache.
        total_raw = r["total_a_pagar_raw"]
        total_leido = None
        dif_total = 0.0
        if total_raw is not None and not isinstance(total_raw, str):
            total_leido = round(_float(total_raw), 2)
            dif_total = round(abs(total_leido - total_esperado), 2)

        estado_mismatch = (r["estado"] != estado_esperado)

        if dif_total > TOL or estado_mismatch:
            errores.append({
                "MZ":              r["mz"],
                "LT":              r["lt"],
                "NOMBRE":          r["nombre"],
                "TOTAL_LEIDO":     total_leido,
                "TOTAL_ESPERADO":  total_esperado,
                "DIF_TOTAL":       dif_total,
                "PAGADO":          pagado,
                "SALDO_ESPERADO":  saldo_esperado,
                "ESTADO_LEIDO":    r["estado"],
                "ESTADO_ESPERADO": estado_esperado,
            })
    return errores


# ─────────────────────────────────────────────────────────────────────────────
#  BLOQUE 2 — Trazabilidad ↔ planilla
# ─────────────────────────────────────────────────────────────────────────────
def _validar_trazabilidad(planilla: list[dict], traza: list[dict]) -> list[dict]:
    """sum(MONTO trazabilidad por fuente) == MONTO_YAPE / MONTO_EFECTIVO en planilla.
    Huerfanos: (MZ,LT) en trazabilidad sin usuario en planilla.
    Faltantes / mismatch: planilla con monto que no cuadra con trazabilidad.
    """
    plan_by_key = {(r["mz"], r["lt"]): r for r in planilla}

    traz_sum: dict[tuple, float] = {}
    nombres_traz: dict[tuple, str] = {}
    for t in traza:
        key3 = (t["mz"], t["lt"], t["fuente"])
        traz_sum[key3] = traz_sum.get(key3, 0.0) + t["monto"]
        nombres_traz[(t["mz"], t["lt"])] = t["nombre"]

    errores = []
    keys_planilla = set(plan_by_key.keys())
    keys_traz = {(mz, lt) for (mz, lt, _) in traz_sum.keys()}

    # 1. Huerfanos: en trazabilidad pero no en planilla
    for key in sorted(keys_traz - keys_planilla):
        mz, lt = key
        suma_yape = round(traz_sum.get((mz, lt, "yape"), 0.0), 2)
        suma_efec = round(traz_sum.get((mz, lt, "efectivo"), 0.0), 2)
        errores.append({
            "MZ":               mz,
            "LT":               lt,
            "NOMBRE":           nombres_traz.get(key, ""),
            "TIPO":             "HUERFANO",
            "MONTO_PLAN_YAPE":  None,
            "MONTO_TRAZ_YAPE":  suma_yape,
            "DIF_YAPE":         None,
            "MONTO_PLAN_EFECT": None,
            "MONTO_TRAZ_EFECT": suma_efec,
            "DIF_EFECT":        None,
        })

    # 2. Faltantes / Mismatch desde la planilla
    for key, r in plan_by_key.items():
        suma_yape = round(traz_sum.get((r["mz"], r["lt"], "yape"), 0.0), 2)
        suma_efec = round(traz_sum.get((r["mz"], r["lt"], "efectivo"), 0.0), 2)
        plan_yape = round(r["monto_yape"], 2)
        plan_efec = round(r["monto_efectivo"], 2)

        dif_yape = round(plan_yape - suma_yape, 2)
        dif_efec = round(plan_efec - suma_efec, 2)

        if abs(dif_yape) <= TOL and abs(dif_efec) <= TOL:
            continue

        total_plan = plan_yape + plan_efec
        total_traz = suma_yape + suma_efec
        if total_plan > TOL and total_traz <= TOL:
            tipo = "FALTANTE_EN_TRAZ"
        elif total_plan <= TOL and total_traz > TOL:
            tipo = "FALTANTE_EN_PLAN"
        else:
            tipo = "MISMATCH"

        errores.append({
            "MZ":               r["mz"],
            "LT":               r["lt"],
            "NOMBRE":           r["nombre"],
            "TIPO":             tipo,
            "MONTO_PLAN_YAPE":  plan_yape,
            "MONTO_TRAZ_YAPE":  suma_yape,
            "DIF_YAPE":         dif_yape,
            "MONTO_PLAN_EFECT": plan_efec,
            "MONTO_TRAZ_EFECT": suma_efec,
            "DIF_EFECT":        dif_efec,
        })

    return errores


# ─────────────────────────────────────────────────────────────────────────────
#  BLOQUE 3 — Arrastres ↔ planilla (simetrico)
# ─────────────────────────────────────────────────────────────────────────────
def _validar_arrastre(planilla: list[dict],
                      arrastre: list[dict],
                      *,
                      tipo_arrastre: str,
                      filtro_saldo,         # (saldo) -> bool
                      monto_esperado_de,    # (saldo) -> float
                      mes_ano: str) -> list[dict]:
    """Validacion generica · sirve para arrastre_deuda y arrastre_devolucion.

    Para arrastre_deuda:
        filtro_saldo      = lambda s: s > TOL
        monto_esperado_de = lambda s: s
    Para arrastre_devolucion:
        filtro_saldo      = lambda s: s < -TOL
        monto_esperado_de = lambda s: abs(s)
    """
    plan_by_key = {(r["mz"], r["lt"]): r for r in planilla}
    arr_by_key  = {(a["mz"], a["lt"]): a for a in arrastre}

    errores = []
    tag = tipo_arrastre.upper()

    # 1. Recorrer arrastre: huerfanos + mismatch + mes_ano incorrecto
    for key, a in arr_by_key.items():
        if key not in plan_by_key:
            errores.append({
                "MZ":              a["mz"],
                "LT":              a["lt"],
                "NOMBRE":          a["nombre"],
                "TIPO":            "HUERFANO_SIN_PLANILLA",
                "MONTO_ARRASTRE":  a["monto"],
                "MONTO_ESPERADO":  None,
                "SALDO_PLANILLA":  None,
                "DIF":             None,
                "MES_ANO_ORIGEN":  a["mes_ano_origen"],
            })
            continue

        r = plan_by_key[key]
        saldo = _saldo_de(r)

        if not filtro_saldo(saldo):
            errores.append({
                "MZ":              a["mz"],
                "LT":              a["lt"],
                "NOMBRE":          a["nombre"],
                "TIPO":            f"HUERFANO_SALDO_NO_{tag}",
                "MONTO_ARRASTRE":  a["monto"],
                "MONTO_ESPERADO":  None,
                "SALDO_PLANILLA":  saldo,
                "DIF":             None,
                "MES_ANO_ORIGEN":  a["mes_ano_origen"],
            })
            continue

        esperado = round(monto_esperado_de(saldo), 2)
        dif = round(a["monto"] - esperado, 2)
        if abs(dif) > TOL:
            errores.append({
                "MZ":              a["mz"],
                "LT":              a["lt"],
                "NOMBRE":          a["nombre"],
                "TIPO":            "MONTO_MISMATCH",
                "MONTO_ARRASTRE":  a["monto"],
                "MONTO_ESPERADO":  esperado,
                "SALDO_PLANILLA":  saldo,
                "DIF":             dif,
                "MES_ANO_ORIGEN":  a["mes_ano_origen"],
            })

        if mes_ano and a["mes_ano_origen"] != mes_ano:
            errores.append({
                "MZ":              a["mz"],
                "LT":              a["lt"],
                "NOMBRE":          a["nombre"],
                "TIPO":            "MES_ANO_INCORRECTO",
                "MONTO_ARRASTRE":  a["monto"],
                "MONTO_ESPERADO":  None,
                "SALDO_PLANILLA":  saldo,
                "DIF":             None,
                "MES_ANO_ORIGEN":  a["mes_ano_origen"],
            })

    # 2. Recorrer planilla: usuarios con saldo en rango pero ausentes del arrastre
    for key, r in plan_by_key.items():
        if key in arr_by_key:
            continue
        saldo = _saldo_de(r)
        if filtro_saldo(saldo):
            esperado = round(monto_esperado_de(saldo), 2)
            errores.append({
                "MZ":              r["mz"],
                "LT":              r["lt"],
                "NOMBRE":          r["nombre"],
                "TIPO":            "FALTANTE_EN_ARRASTRE",
                "MONTO_ARRASTRE":  None,
                "MONTO_ESPERADO":  esperado,
                "SALDO_PLANILLA":  saldo,
                "DIF":             None,
                "MES_ANO_ORIGEN":  "",
            })

    return errores


# ─────────────────────────────────────────────────────────────────────────────
#  BLOQUE 4 — Reporte
# ─────────────────────────────────────────────────────────────────────────────
def _escribir_errores_xlsx(secciones: dict[str, list[dict]]):
    """Una hoja por bloque con errores · solo hojas con filas."""
    wb = Workbook()
    wb.remove(wb.active)

    H_BG  = "FECACA"
    H_TXT = "991B1B"
    D_BG  = "FFF7F7"

    fill_h = PatternFill("solid", start_color=H_BG)
    fill_d = PatternFill("solid", start_color=D_BG)
    font_h = Font(name="Arial", size=10, bold=True, color=H_TXT)
    font_d = Font(name="Arial", size=10, color="333333")
    border = Border(
        top=Side(border_style="thin", color="EEEEEE"),
        bottom=Side(border_style="thin", color="EEEEEE"),
    )

    for nombre, filas in secciones.items():
        if not filas:
            continue
        ws = wb.create_sheet(title=nombre[:31])
        cols = list(filas[0].keys())

        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.fill = fill_h
            c.font = font_h
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
            ws.column_dimensions[c.column_letter].width = max(14, len(col) + 2)

        for ri, f in enumerate(filas, 2):
            for ci, col in enumerate(cols, 1):
                c = ws.cell(row=ri, column=ci, value=f[col])
                c.fill = fill_d
                c.font = font_d
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = border

        ws.freeze_panes = "A2"

    if not wb.sheetnames:
        return
    out = OUTPUTS_DIR / "validacion_errores.xlsx"
    wb.save(out)
    log.info(f"validacion_errores.xlsx → {sum(len(v) for v in secciones.values())} "
             f"errores en {len([k for k, v in secciones.items() if v])} hojas")


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main() -> int:
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(OUTPUTS_DIR / "run_validacion.log", encoding="utf-8"),
        ],
        force=True,
    )

    print("\n" + "═" * 60)
    print("  validacion_planilla_cobrado — coherencia de outputs")
    print("═" * 60 + "\n")

    # ── Cargar outputs ──
    print("[1/4] Cargando outputs...")
    planilla = _leer_planilla()
    traza    = _leer_trazabilidad()
    mes_ano  = planilla[0]["mes_ano"] if planilla else ""

    arr_deuda_path = _buscar_arrastre("arrastre_deuda", mes_ano)
    arr_devol_path = _buscar_arrastre("arrastre_devolucion", mes_ano)
    arr_deuda = _leer_arrastre(arr_deuda_path)
    arr_devol = _leer_arrastre(arr_devol_path)

    log.info(f"planilla_cobrado.xlsx → {len(planilla)} usuarios · mes {mes_ano}")
    log.info(f"trazabilidad_cobranza.xlsx → {len(traza)} pagos")
    log.info(f"{arr_deuda_path.name} → {len(arr_deuda)} filas")
    log.info(f"{arr_devol_path.name} → {len(arr_devol)} filas")

    # ── Bloque 1 ──
    print("\n[2/4] Bloque 1 — re-calculo SALDO + ESTADO...")
    e_saldo = _validar_saldo(planilla)
    log.info(f"  Bloque 1 → {len(e_saldo)} errores")

    # ── Bloque 2 ──
    print("[3/4] Bloque 2 — trazabilidad ↔ planilla...")
    e_traz = _validar_trazabilidad(planilla, traza)
    log.info(f"  Bloque 2 → {len(e_traz)} errores")

    # ── Bloque 3 ──
    print("[4/4] Bloque 3 — arrastres ↔ planilla (deuda + devolucion)...")
    e_ad = _validar_arrastre(
        planilla, arr_deuda,
        tipo_arrastre="deuda",
        filtro_saldo=lambda s: s > TOL,
        monto_esperado_de=lambda s: s,
        mes_ano=mes_ano,
    )
    log.info(f"  arrastre_deuda → {len(e_ad)} errores")

    e_av = _validar_arrastre(
        planilla, arr_devol,
        tipo_arrastre="devolucion",
        filtro_saldo=lambda s: s < -TOL,
        monto_esperado_de=lambda s: abs(s),
        mes_ano=mes_ano,
    )
    log.info(f"  arrastre_devolucion → {len(e_av)} errores")

    # ── Reporte ──
    secciones = {
        "errores_saldo":               e_saldo,
        "errores_trazabilidad":        e_traz,
        "errores_arrastre_deuda":      e_ad,
        "errores_arrastre_devolucion": e_av,
    }
    total = sum(len(v) for v in secciones.values())

    print("\n" + "═" * 60)
    if total == 0:
        print("  OK — todos los chequeos pasaron")
        print(f"  · planilla_cobrado: {len(planilla)} usuarios")
        print(f"  · trazabilidad: {len(traza)} pagos")
        print(f"  · arrastre_deuda: {len(arr_deuda)} usuarios")
        print(f"  · arrastre_devolucion: {len(arr_devol)} usuarios")
        print("═" * 60 + "\n")
        return 0

    _escribir_errores_xlsx(secciones)
    print(f"  {total} errores encontrados:")
    print(f"   · Bloque 1 (re-calculo saldo+estado) → {len(e_saldo)}")
    print(f"   · Bloque 2 (trazabilidad)            → {len(e_traz)}")
    print(f"   · Bloque 3 (arrastre_deuda)          → {len(e_ad)}")
    print(f"   · Bloque 3 (arrastre_devolucion)     → {len(e_av)}")
    print(f"\n  → outputs/validacion_errores.xlsx")
    print("═" * 60 + "\n")
    return 1


if __name__ == "__main__":
    sys.exit(main())
