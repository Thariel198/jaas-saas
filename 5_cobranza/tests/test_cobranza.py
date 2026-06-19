"""tests/test_cobranza.py — test de integración con fixtures sintéticos

Construye una planilla mínima + pagos diseñados para forzar cada regla del módulo:

  CANCELADO       — paga exacto                                  (A-1)
  EXCESO          — paga de más → arrastre_devolucion            (A-2)
  PARCIAL nuevo   — paga algo · sin arrastre → NO elegible corte (A-3)
  PENDIENTE nuevo — no paga    · sin arrastre → NO elegible corte (A-4)
  CORTE PARCIAL   — paga algo · MES_ANTERIOR>=8 → elegible corte (B-1)
  CORTE PENDIENTE — no paga    · MES_ANTERIOR>=8 → elegible corte (B-2)
  BLANCO aplicado — blanco descuenta, paga el resto              (C-1)
  PAGO HUERFANO   — MZ-LT no en planilla → warning, ignorado     (X-99)
  Idempotencia    — 2da corrida no agrega filas a trazabilidad

Nota: lista_corte.xlsx ya no es output de 5_cobranza — la genera 6_corte
leyendo SALDO + MES_ANTERIOR de planilla_cobrado. Este test verifica que
planilla_cobrado expone los datos correctos para que 6_corte decida.

Patrón sigue metodología 3.6: fixtures en _tmp_integracion/, monkey-patch
de paths antes de importar main, verificación por conteo de instancias.

Uso:
    python tests/test_cobranza.py
"""
import shutil
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook

THIS = Path(__file__).resolve()
sys.path.insert(0, str(THIS.parent.parent))   # 5_cobranza/ → para importar main

TEST_ROOT = THIS.parent / "_tmp_integracion"
MES_TEST  = "2099-12"


# ─────────────────────────────────────────────────────────────────────────────
#  FIXTURE DATA
# ─────────────────────────────────────────────────────────────────────────────
USUARIOS = [
    # mz, lt, nombre,                  mes_actual, mantenimiento, mes_anterior
    ("A", "1", "CANCELADO_YAPE",       20.00, 3.00,  0.00),
    ("A", "2", "EXCESO_YAPE",          15.00, 3.00,  0.00),
    ("A", "3", "PARCIAL_NEW",          20.00, 3.00,  0.00),
    ("A", "4", "PENDIENTE_NEW",        15.00, 3.00,  0.00),
    ("B", "1", "CORTE_PARCIAL",        15.00, 3.00, 12.00),
    ("B", "2", "CORTE_PENDIENTE",      15.00, 3.00, 12.00),
    ("C", "1", "CANCEL_CON_BLANCO",    15.00, 3.00,  0.00),
    ("D", "1", "YA_ESTABA_OK",          0.00, 0.00,  0.00),
]

PAGOS_YAPE = [
    # mz, lt, monto
    ("A", "1", 23.00),    # debe 23, paga 23 → CANCELADO
    ("A", "2", 30.00),    # debe 18, paga 30 → EXCESO (saldo -12)
]

PAGOS_EFECTIVO = [
    ("A", "3", 10.00),    # debe 23, paga 10 → PARCIAL saldo 13
    ("B", "1",  5.00),    # debe 30 (15+3+12), paga 5 → PARCIAL saldo 25 → elegible corte
    ("C", "1", 15.00),    # con blanco -3 → debe 15, paga 15 → CANCELADO
    ("X", "99", 99.00),   # HUERFANO: MZ-LT no existe en planilla
]

BLANCOS = [
    ("C", "1", 3.00),     # blanco S/3 que se aplica a C-1
]

ESPERADO = {
    "n_planilla":            8,
    "estados": {
        "CANCELADO": 3,    # A-1, C-1, D-1
        "EXCESO":    1,    # A-2
        "PARCIAL":   2,    # A-3, B-1
        "PENDIENTE": 2,    # A-4, B-2
    },
    "n_elegibles_corte":     2,   # B-1, B-2 (SALDO>0 AND MES_ANTERIOR>=8)
    "n_arrastre_deuda":      4,   # A-3, A-4, B-1, B-2
    "n_arrastre_devolucion": 1,   # A-2
    "n_blancos_aplicados":   1,   # C-1
    "n_trazabilidad":        6,   # 2 yape + 4 efectivo (incluye huérfano X-99)
}


# ─────────────────────────────────────────────────────────────────────────────
#  SETUP + FIXTURE BUILDERS
# ─────────────────────────────────────────────────────────────────────────────
def _setup_paths():
    """Limpia y recrea TEST_ROOT · monkey-patchea las constantes de main."""
    if TEST_ROOT.exists():
        shutil.rmtree(TEST_ROOT)
    (TEST_ROOT / "inputs" / "planilla").mkdir(parents=True)
    (TEST_ROOT / "inputs" / "pagos_yape").mkdir(parents=True)
    (TEST_ROOT / "inputs" / "pagos_efectivo").mkdir(parents=True)
    (TEST_ROOT / "outputs").mkdir(parents=True)
    (TEST_ROOT / "shared").mkdir(parents=True)

    import main as mod
    mod.INPUTS_DIR   = TEST_ROOT / "inputs"
    mod.OUTPUTS_DIR  = TEST_ROOT / "outputs"
    mod.SHARED_DIR   = TEST_ROOT / "shared"
    mod.PLAN_DIR     = TEST_ROOT / "inputs" / "planilla"
    mod.YAPE_DIR     = TEST_ROOT / "inputs" / "pagos_yape"
    mod.EFEC_DIR     = TEST_ROOT / "inputs" / "pagos_efectivo"
    mod.BLANCOS_PATH = TEST_ROOT / "shared" / "blancos_acumulados.xlsx"
    return mod


def _write_planilla():
    """planilla_{MES_TEST}.xlsx con doble header · fila 1 grupos, fila 2 nombres."""
    cols = ["MZ", "LT", "NOMBRE", "MES_ANO", "MARC_ANT", "MARC_ACT", "M3",
            "MES_ACTUAL", "MANTENIMIENTO", "MES_ANTERIOR", "CORTE_RECONEXION",
            "CONVENIO", "MULTA", "ACUERDOS_ASAMBLEA",
            "BLANCO", "DEVOLUCION", "TOTAL_A_PAGAR",
            "MONTO_YAPE", "MONTO_EFECTIVO", "ESTADO", "FECHA_PAGO"]
    wb = Workbook()
    ws = wb.active
    ws.title = "planilla"
    ws.append([""] * len(cols))   # fila 1 = grupos (vacíos en test)
    ws.append(cols)               # fila 2 = nombres reales

    for mz, lt, nom, mes_act, mant, mes_ant in USUARIOS:
        cargos = mes_act + mant + mes_ant
        ws.append([
            mz, lt, nom, MES_TEST,
            0, 0, 0,                   # MARC_ANT, MARC_ACT, M3
            mes_act, mant, mes_ant,
            0, 0, 0, 0,                # CORTE_RECONEXION, CONVENIO, MULTA, ACUERDOS
            0, 0,                      # BLANCO, DEVOLUCION (llegan en 0 de 2_planilla)
            cargos,                    # TOTAL_A_PAGAR (main.py lo recalcula)
            None, None, None, None,    # MONTO_YAPE, MONTO_EFECTIVO, ESTADO, FECHA_PAGO
        ])
    wb.save(TEST_ROOT / "inputs" / "planilla" / f"planilla_{MES_TEST}.xlsx")


def _write_yape():
    cols = ["TIPO", "MZ", "LOTE", "NOMBRE", "CONCEPTO",
            "MONTO_PAGO", "MONTO_ASIGNADO", "FECHA", "CICLO_CORRECCION"]
    wb = Workbook()
    ws = wb.active
    ws.append([""] * len(cols))
    ws.append(cols)
    for mz, lt, monto in PAGOS_YAPE:
        ws.append(["TE PAGÓ", mz, lt, "", "", monto, monto, "01/12/2099", 1])
    wb.save(TEST_ROOT / "inputs" / "pagos_yape" / "pagos_yape_tepago.xlsx")


def _write_efectivo():
    cols = ["MZ", "LT", "MONTO", "FECHA", "CICLO_CORRECCION"]
    wb = Workbook()
    ws = wb.active
    ws.append([""] * len(cols))
    ws.append(cols)
    for mz, lt, monto in PAGOS_EFECTIVO:
        ws.append([mz, lt, monto, "01/12/2099", 1])
    wb.save(TEST_ROOT / "inputs" / "pagos_efectivo" / "pagos_efectivo.xlsx")


def _write_blancos():
    """Esquema fijo de shared/blancos_acumulados.xlsx · 19 columnas con MZ en col 13,
    LOTE en col 14, ESTADO en col 18, MES_APLICADO en col 19 (main._BL_*).
    """
    # row 1 = grupos · row 2 = nombres reales
    grupos = [None] * 19
    nombres = [None] * 19
    nombres[0]  = "MES"          # col 1
    nombres[2]  = "USER_ID"      # col 3
    nombres[3]  = "NOMBRE"       # col 4
    nombres[5]  = "TIPO"         # col 6
    nombres[6]  = "ORIGEN"       # col 7
    nombres[7]  = "DESTINO"      # col 8
    nombres[8]  = "MONTO"        # col 9
    nombres[9]  = "MENSAJE"      # col 10
    nombres[10] = "FECHA"        # col 11
    nombres[12] = "MZ"           # col 13  ← _BL_MZ
    nombres[13] = "LOTE"         # col 14  ← _BL_LOTE
    nombres[14] = "CONCEPTO"     # col 15
    nombres[15] = "MOTIVO"       # col 16
    nombres[17] = "ESTADO"       # col 18  ← _BL_EST
    nombres[18] = "MES_APLICADO" # col 19  ← _BL_MES

    wb = Workbook()
    ws = wb.active
    ws.append(grupos)
    ws.append(nombres)
    for mz, lt, monto in BLANCOS:
        fila = [None] * 19
        fila[0]  = MES_TEST
        fila[8]  = monto
        fila[12] = mz
        fila[13] = lt
        fila[17] = "pendiente"
        ws.append(fila)
    wb.save(TEST_ROOT / "shared" / "blancos_acumulados.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
#  LECTORES DE OUTPUT
# ─────────────────────────────────────────────────────────────────────────────
def _leer_dh(path: Path) -> tuple[list[str], list[tuple]]:
    """Lee xlsx con doble header (fila 1=grupos, fila 2=cols, fila 3+=datos)."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip().upper() if c.value else "" for c in ws[2]]
    filas = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if any(c is not None and str(c).strip() for c in row):
            filas.append(row)
    return headers, filas


def _col(headers, row, nombre):
    if nombre in headers:
        return row[headers.index(nombre)]
    return None


# ─────────────────────────────────────────────────────────────────────────────
#  VERIFICACIONES
# ─────────────────────────────────────────────────────────────────────────────
def _verificar_primera(mod) -> tuple[bool, list[str]]:
    """Chequeos sobre los outputs después de la 1ra corrida."""
    msgs = []
    ok = True

    def chk(condicion: bool, etiqueta: str, esperado, obtenido):
        nonlocal ok
        marca = "OK " if condicion else "FAIL"
        msgs.append(f"  [{marca}] {etiqueta:<40} esperado={esperado}  obtenido={obtenido}")
        if not condicion:
            ok = False

    # ── planilla_cobrado ──
    plan_path = mod.OUTPUTS_DIR / "planilla_cobrado.xlsx"
    headers, filas = _leer_dh(plan_path)
    chk(len(filas) == ESPERADO["n_planilla"],
        "planilla_cobrado · filas", ESPERADO["n_planilla"], len(filas))

    estados = Counter(str(_col(headers, r, "ESTADO") or "").strip().upper() for r in filas)
    for k, v in ESPERADO["estados"].items():
        chk(estados.get(k, 0) == v, f"estado {k}", v, estados.get(k, 0))

    # C-1: BLANCO debe ser -3 (negativo) y SALDO=0
    fila_c1 = [r for r in filas
               if _col(headers, r, "MZ") == "C" and str(_col(headers, r, "LT")) == "1"][0]
    blanco_c1 = _col(headers, fila_c1, "BLANCO") or 0
    chk(abs(blanco_c1 - (-3)) < 0.01, "C-1 BLANCO aplicado (-3)", -3, blanco_c1)

    def _saldo_calc(r):
        cargos = sum(_col(headers, r, c) or 0 for c in
                     ("MES_ACTUAL", "MANTENIMIENTO", "MES_ANTERIOR",
                      "CORTE_RECONEXION", "CONVENIO", "MULTA", "ACUERDOS_ASAMBLEA"))
        descuentos = (_col(headers, r, "BLANCO") or 0) + (_col(headers, r, "DEVOLUCION") or 0)
        total = round(cargos + descuentos, 2)
        pagado = (_col(headers, r, "MONTO_YAPE") or 0) + (_col(headers, r, "MONTO_EFECTIVO") or 0)
        return round(total - pagado, 2)

    chk(abs(_saldo_calc(fila_c1) - 0) < 0.01, "C-1 SALDO=0 tras blanco", 0, _saldo_calc(fila_c1))

    # ── elegibles para corte ──
    # planilla_cobrado debe exponer SALDO y MES_ANTERIOR para que 6_corte
    # decida quién va a la lista. Acá contamos los que cumplen el filtro.
    elegibles = sum(
        1 for r in filas
        if (_col(headers, r, "SALDO") or 0) > 0.005
        and (_col(headers, r, "MES_ANTERIOR") or 0) >= 8 - 0.005
    )
    chk(elegibles == ESPERADO["n_elegibles_corte"],
        "elegibles corte (SALDO>0 & MES_ANT>=8)",
        ESPERADO["n_elegibles_corte"], elegibles)

    # ── arrastre_deuda ──
    ad_path = mod.OUTPUTS_DIR / f"arrastre_deuda_{MES_TEST}.xlsx"
    _, filas_ad = _leer_dh(ad_path)
    chk(len(filas_ad) == ESPERADO["n_arrastre_deuda"],
        "arrastre_deuda · filas", ESPERADO["n_arrastre_deuda"], len(filas_ad))

    # ── arrastre_devolucion ──
    av_path = mod.OUTPUTS_DIR / f"arrastre_devolucion_{MES_TEST}.xlsx"
    h_av, filas_av = _leer_dh(av_path)
    chk(len(filas_av) == ESPERADO["n_arrastre_devolucion"],
        "arrastre_devolucion · filas", ESPERADO["n_arrastre_devolucion"], len(filas_av))

    if filas_av:
        # El monto debe ser |saldo| = 12, positivo
        monto_av = _col(h_av, filas_av[0], "MONTO") or 0
        chk(monto_av > 0 and abs(monto_av - 12) < 0.01,
            "arrastre_devolucion · monto=|saldo|", 12.0, monto_av)

    # ── trazabilidad ──
    traz_path = mod.OUTPUTS_DIR / "trazabilidad_cobranza.xlsx"
    _, filas_traz = _leer_dh(traz_path)
    chk(len(filas_traz) == ESPERADO["n_trazabilidad"],
        "trazabilidad · filas (incluye huérfano)", ESPERADO["n_trazabilidad"], len(filas_traz))

    # ── blancos_acumulados: ESTADO=aplicado ──
    h_bl, filas_bl = _leer_dh(mod.BLANCOS_PATH)
    aplicados = sum(1 for r in filas_bl
                    if str(_col(h_bl, r, "ESTADO") or "").strip().lower() == "aplicado")
    chk(aplicados == ESPERADO["n_blancos_aplicados"],
        "blancos · ESTADO=aplicado", ESPERADO["n_blancos_aplicados"], aplicados)

    return ok, msgs


def _snapshot(mod) -> dict:
    """Conteos rápidos de los outputs principales · para verificar idempotencia."""
    def _n(p):
        if not p.exists():
            return 0
        _, filas = _leer_dh(p)
        return len(filas)
    return {
        "planilla":            _n(mod.OUTPUTS_DIR / "planilla_cobrado.xlsx"),
        "arrastre_deuda":      _n(mod.OUTPUTS_DIR / f"arrastre_deuda_{MES_TEST}.xlsx"),
        "arrastre_devolucion": _n(mod.OUTPUTS_DIR / f"arrastre_devolucion_{MES_TEST}.xlsx"),
        "trazabilidad":        _n(mod.OUTPUTS_DIR / "trazabilidad_cobranza.xlsx"),
    }


def _verificar_idempotencia(antes: dict, despues: dict) -> tuple[bool, list[str]]:
    msgs = []
    ok = True
    for k in antes:
        cond = antes[k] == despues[k]
        marca = "OK " if cond else "FAIL"
        msgs.append(f"  [{marca}] idempotencia · {k:<28} antes={antes[k]}  después={despues[k]}")
        if not cond:
            ok = False
    return ok, msgs


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "=" * 60)
    print("  test_cobranza — fixtures sintéticos")
    print("=" * 60 + "\n")

    print("[Setup] Limpiando _tmp_integracion y monkey-patcheando paths...")
    mod = _setup_paths()

    print("[Fixture] Escribiendo planilla, yape, efectivo, blancos...")
    _write_planilla()
    _write_yape()
    _write_efectivo()
    _write_blancos()

    print("\n[Run 1] main.main() — primera corrida")
    print("-" * 60)
    mod.main()

    print("\n[Check 1] Verificando outputs")
    print("-" * 60)
    ok1, msgs1 = _verificar_primera(mod)
    for m in msgs1:
        print(m)

    print("\n[Run 2] main.main() — segunda corrida (debe ser idempotente)")
    print("-" * 60)
    antes = _snapshot(mod)
    mod.main()
    despues = _snapshot(mod)

    print("\n[Check 2] Verificando idempotencia")
    print("-" * 60)
    ok2, msgs2 = _verificar_idempotencia(antes, despues)
    for m in msgs2:
        print(m)

    print("\n" + "=" * 60)
    if ok1 and ok2:
        print("  TODOS LOS CHEQUEOS PASARON")
        print("=" * 60 + "\n")
        return 0
    print("  HUBO ERRORES — revisar marcas [FAIL] arriba")
    print("=" * 60 + "\n")
    return 1


if __name__ == "__main__":
    sys.exit(main())
