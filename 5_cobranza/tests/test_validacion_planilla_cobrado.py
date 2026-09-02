import importlib.util
from pathlib import Path

from openpyxl import Workbook


MODULE_PATH = Path(__file__).resolve().parents[1] / "validacion_planilla_cobrado.py"


def _cargar_modulo():
    spec = importlib.util.spec_from_file_location("validacion_planilla_cobrado", MODULE_PATH)
    modulo = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(modulo)
    return modulo


def _crear_workbook(ruta: Path):
    wb = Workbook()
    wb.active.title = "otra_hoja"

    planilla = wb.create_sheet("planilla_cobrado")
    planilla.append([])
    planilla.append([
        "MZ", "LT", "NOMBRE", "MES_ANO", "TOTAL_A_PAGAR",
        "MONTO_YAPE", "MONTO_EFECTIVO", "ABONO_REZAGADO", "SALDO", "ESTADO",
    ])
    planilla.append(["A", "1", "Usuario", "2026-08", 10, 2, 3, 1, 4, "PARCIAL"])

    consolidado = wb.create_sheet("arrastre_consolidado")
    consolidado.append([])
    consolidado.append(["MZ", "LT", "NOMBRE", "TOTAL_ARRASTRE"])
    consolidado.append(["A", "1", "Usuario", 4])

    devolucion = wb.create_sheet("arrastre_devolucion")
    devolucion.append([])
    devolucion.append(["MZ", "LT", "NOMBRE", "MONTO", "MES_ANO_ORIGEN"])
    devolucion.append(["B", "2", "Otro", 7, "2026-08"])
    wb.save(ruta)


def test_lee_hojas_del_workbook_canonico_sin_usar_hoja_activa(tmp_path, monkeypatch):
    modulo = _cargar_modulo()
    workbook = tmp_path / "planilla_cobrado_2026-08.xlsx"
    _crear_workbook(workbook)
    monkeypatch.setattr(modulo, "PLANILLA_PATH", workbook)

    planilla = modulo._leer_planilla()
    consolidado = modulo._leer_arrastre(
        workbook, "arrastre_consolidado", "TOTAL_ARRASTRE")
    devolucion = modulo._leer_arrastre(
        workbook, "arrastre_devolucion", "MONTO")

    assert [(fila["mz"], fila["lt"]) for fila in planilla] == [("A", "1")]
    assert consolidado[0]["monto"] == 4
    assert devolucion[0]["monto"] == 7
    assert modulo.PLANILLA_PATH == workbook


def test_ruta_canonica_es_el_periodo_exacto():
    modulo = _cargar_modulo()

    assert modulo.PLANILLA_PATH == (
        modulo.OUTPUTS_DIR / f"planilla_cobrado_{modulo.MES_CICLO}.xlsx")


def test_ignora_trazabilidad_historica(tmp_path, monkeypatch):
    modulo = _cargar_modulo()
    ruta = tmp_path / "trazabilidad_cobranza.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([])
    ws.append([
        "MZ", "LT", "NOMBRE", "MONTO", "FUENTE", "CONCEPTO", "FECHA",
        "CICLO_CORRECCION_ORIGEN", "MZ_ORIGEN", "LT_ORIGEN",
    ])
    ws.append(["A", "1", "Actual", 10, "yape", "", "01/08/2026", 1, "", ""])
    ws.append(["B", "2", "Historico", 20, "yape", "", "01/07/2026", 1, "", ""])
    wb.save(ruta)
    monkeypatch.setattr(modulo, "TRAZ_PATH", ruta)

    identidad = {("A", "1", 10.0, "yape", "01/08/2026", 1)}
    filas = modulo._leer_trazabilidad(identidad)

    assert [(f["mz"], f["lt"], f["monto"]) for f in filas] == [("A", "1", 10.0)]


def test_aporte_tanque_no_crea_saldo_ni_arrastre_falso():
    modulo = _cargar_modulo()
    base = {
        "mz": "I", "lt": "13", "nombre": "Usuario", "mes_ano": "2026-08",
        "mes_actual": 18.0, "mantenimiento": 3.0, "mes_anterior": 0.0,
        "corte_reconexion": 0.0, "convenio": 0.0, "multa": 0.0, "acuerdos": 0.0,
        "blanco": 0.0, "devolucion": 0.0, "total_a_pagar_raw": None,
        "monto_yape": 221.0, "monto_efectivo": 0.0, "abono_rezagado": 0.0,
        "saldo": 0.0, "aporte_tanque": 200.0, "estado": "CANCELADO",
    }

    assert modulo._validar_saldo([base]) == []
    assert modulo._validar_arrastre(
        [base], [], tipo_arrastre="devolucion",
        filtro_saldo=lambda saldo: saldo < -modulo.TOL,
        monto_esperado_de=abs, mes_ano="2026-08",
    ) == []
