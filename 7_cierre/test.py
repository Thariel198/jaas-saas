"""
Test de integración — fixtures en tests/_tmp_integracion/, no toca inputs/outputs reales.
Cada corrida borra y recrea la carpeta temporal → idempotente.

Escenarios cubiertos:
  - GATE bloquea si el ciclo no está validado
  - COSECHAR copia canónicos + fuentes manuales + fuentes auto a archivo/{mes}/
  - FREEZE sella estado_ciclo[mes].estado = CERRADO
  - LIMPIAR resetea SOLO las fuentes manuales (mesa_1, correcciones_lote) a template
    vacío — las fuentes auto (yape) y los canónicos NO se tocan en outputs/
"""
import builtins
import hashlib
import json
import shutil
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).parent))
import config

_TMP = Path(__file__).parent / "tests" / "_tmp_integracion"


def _setup_paths() -> None:
    if _TMP.exists():
        shutil.rmtree(_TMP)
    for sub in ["archivo", "outputs", "cobranza/outputs", "cobranza/inputs",
                "efectivo/inputs", "yape_outputs"]:
        (_TMP / sub).mkdir(parents=True)

    config.ARCHIVO_DIR       = _TMP / "archivo"
    config.OUTPUTS_DIR       = _TMP / "outputs"
    config.COBRANZA_DIR      = _TMP / "cobranza"
    config.EFECTIVO_DIR      = _TMP / "efectivo"
    config.YAPE_DIR          = _TMP / "yape_outputs"
    config.ESTADO_CICLO_PATH = _TMP / "estado_ciclo.json"
    mod.repo.SEGUIMIENTO_PATH = _TMP / "seguimiento_pueblo.xlsx"


import consolidar_cierre as mod  # noqa: E402 — después de definir _setup_paths

MES = "2026-06"


def _wb_simple(ruta: Path, filas: list) -> None:
    wb = Workbook()
    ws = wb.active
    for row in filas:
        ws.append(row)
    wb.save(ruta)


def _build_fixture() -> None:
    out = config.COBRANZA_DIR / "outputs"
    planilla_path = out / f"planilla_cobrado_{MES}.xlsx"
    _wb_simple(planilla_path, [["MZ", "LT", "SALDO"], ["A", "7", 0]])
    wb = load_workbook(planilla_path)
    wb.active.title = "planilla_cobrado"
    wb.create_sheet("arrastre_consolidado").append(["MZ", "LT", "TOTAL"])
    wb.create_sheet("arrastre_devolucion").append(["MZ", "LT", "EXCESO"])
    wb.save(planilla_path)
    _wb_simple(out / "trazabilidad_cobranza_pre_dedup_20260629_141501.xlsx", [["basura"]])
    snapshot = {"schema": 1, "mes": MES, "objetivos": [], "cargos": []}
    normalizado = json.dumps(snapshot, ensure_ascii=True, sort_keys=True,
                             separators=(",", ":")).encode("utf-8")
    snapshot_hash = hashlib.sha256(normalizado).hexdigest()
    (out / f"snapshot_ledger_{MES}.json").write_text(
        json.dumps({**snapshot, "snapshot_hash": snapshot_hash}), encoding="utf-8")

    inputs_efec = config.EFECTIVO_DIR / "inputs"
    _wb_simple(inputs_efec / "mesa_1.xlsx", [["COBRADOR"], ["dato real de junio — debe desaparecer"]])
    for n in range(2, 8):
        _wb_simple(inputs_efec / f"mesa_{n}.xlsx", [["COBRADOR"]])

    inputs_cob = config.COBRANZA_DIR / "inputs"
    _wb_simple(inputs_cob / "correcciones_lote.xlsx",
               [["MZ_ORIGEN", "LT_ORIGEN", "MZ_DESTINO", "LT_DESTINO", "MOTIVO", "CICLO", "FECHA"],
                ["C", "88", "C", "8B", "corrección real", "1", "01/06/2026"]])

    yape = config.YAPE_DIR
    _wb_simple(yape / "pagos_yape_tepago.xlsx", [["MZ", "LT", "MONTO"], ["B", "3", 20]])
    _wb_simple(yape / "pagos_yape_devolucion.xlsx", [["MZ", "LT", "MONTO"]])
    _wb_simple(yape / "pagos_yape_retorno.xlsx", [["MZ", "LT", "MONTO"]])

    config.ESTADO_CICLO_PATH.parent.mkdir(parents=True, exist_ok=True)
    config.ESTADO_CICLO_PATH.write_text(
        json.dumps({MES: {"estado": "ABIERTO", "arrastre": {
            "generado": True, "validado": True, "snapshot_hash": snapshot_hash}}}),
        encoding="utf-8")


def main() -> int:
    ok = True
    print("=" * 60)
    print("  7_cierre — test de integración")
    print("=" * 60)

    print("\n[1] Fixture...")
    _setup_paths()
    mod.paso0_preparar = lambda _mes: None
    mod.repo.generar_vista = lambda: None
    mod.repo.exportar_vista_pdf = lambda: None
    _build_fixture()

    print("\n[2] GATE — ciclo no validado debe abortar...")
    config.ESTADO_CICLO_PATH.write_text(
        '{"2099-01": {"estado": "ABIERTO", "arrastre": {"generado": true, "validado": false}}}',
        encoding="utf-8")
    try:
        mod.paso1_gate("2099-01")
        print("  [FAIL] debía lanzar CicloNoValidadoError")
        ok = False
    except mod.CicloNoValidadoError:
        print("  [OK] gate abortó correctamente")

    print(f"\n[3a] Corriendo main({MES}) SIN --confirmar (dry-run)...")
    _build_fixture()  # re-set estado_ciclo con MES validado
    mod.main(MES, confirmar=False)

    dry_run_intacto = (not mod.repo_estado.ciclo_cerrado(MES, ruta=config.ESTADO_CICLO_PATH)
                       and (config.EFECTIVO_DIR / "inputs" / "mesa_1.xlsx").exists())
    from openpyxl import load_workbook as _lw
    mesa1_sin_resetear = _lw(config.EFECTIVO_DIR / "inputs" / "mesa_1.xlsx").sheetnames != \
        ["registro_1", "registro_2", "registro_3"]
    print(f"  [{'OK' if dry_run_intacto else 'FAIL'}] dry-run NO selló estado_ciclo")
    print(f"  [{'OK' if mesa1_sin_resetear else 'FAIL'}] dry-run NO reseteó mesa_1.xlsx")
    ok = ok and dry_run_intacto and mesa1_sin_resetear

    print(f"\n[3b] Corriendo main({MES}, confirmar=True) — simulando 'SI' tipeado...")
    builtins.input = lambda _: "SI"
    try:
        mod.main(MES, confirmar=True)
    finally:
        builtins.input = input

    print("\n[4] Verificando COSECHAR...")
    destino = config.archivo_mes_dir(MES)
    esperados = [f"planilla_cobrado_{MES}.xlsx", f"snapshot_ledger_{MES}.json",
                  "mesa_1.xlsx", "correcciones_lote.xlsx",
                 "pagos_yape_tepago.xlsx", "pagos_yape_devolucion.xlsx", "pagos_yape_retorno.xlsx"]
    for nombre in esperados:
        existe = (destino / nombre).exists()
        print(f"  [{'OK' if existe else 'FAIL'}] archivo/{MES}/{nombre}")
        ok = ok and existe

    print("\n[5] Verificando FREEZE...")
    cerrado = mod.repo_estado.ciclo_cerrado(MES, ruta=config.ESTADO_CICLO_PATH)
    print(f"  [{'OK' if cerrado else 'FAIL'}] estado_ciclo[{MES}].estado == CERRADO")
    ok = ok and cerrado

    print("\n[6] Verificando LIMPIAR (reset SOLO fuentes manuales)...")
    from openpyxl import load_workbook
    wb = load_workbook(config.EFECTIVO_DIR / "inputs" / "mesa_1.xlsx")
    hojas_ok = wb.sheetnames == ["registro_1", "registro_2", "registro_3"]
    print(f"  [{'OK' if hojas_ok else 'FAIL'}] mesa_1.xlsx reseteado a template (3 hojas registro_*)")
    ok = ok and hojas_ok

    wb2 = load_workbook(config.COBRANZA_DIR / "inputs" / "correcciones_lote.xlsx")
    ws2 = wb2.active
    sin_datos = ws2.max_row == 1  # solo header
    print(f"  [{'OK' if sin_datos else 'FAIL'}] correcciones_lote.xlsx reseteado (solo header)")
    ok = ok and sin_datos

    basura_borrada = not (config.COBRANZA_DIR / "outputs" /
                          "trazabilidad_cobranza_pre_dedup_20260629_141501.xlsx").exists()
    print(f"  [{'OK' if basura_borrada else 'FAIL'}] basura pre_dedup borrada")
    ok = ok and basura_borrada

    tepago_intacto = (config.YAPE_DIR / "pagos_yape_tepago.xlsx").exists()
    print(f"  [{'OK' if tepago_intacto else 'FAIL'}] pagos_yape_tepago.xlsx NO se borró (solo se cosecha, auto-reset lo hace motor_matching)")
    ok = ok and tepago_intacto

    print("\n" + "=" * 60)
    print("PASÓ — 7_cierre transiciona correctamente." if ok else "HUBO ERRORES — revisar [FAIL] arriba.")
    print("=" * 60)
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
