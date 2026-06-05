import logging
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
ROOT               = Path(__file__).parent
INPUTS_DIR         = ROOT / "inputs"
OUTPUTS_DIR        = ROOT / "outputs"
SHARED_DIR         = ROOT.parent / "shared"

PLANILLA_BASE_DIR  = INPUTS_DIR / "planilla_base"
PAGOS_YAPE_DIR     = INPUTS_DIR / "pagos_yape"
PAGOS_EFECTIVO_DIR = INPUTS_DIR / "pagos_efectivo"
BLANCOS_PATH       = SHARED_DIR / "blancos_acumulados.xlsx"

PLANILLA_BASE_FILE  = "planilla_base.xlsx"
PAGOS_YAPE_FILE     = "pagos_yape_tepago.xlsx"
PAGOS_EFECTIVO_FILE = "pagos_efectivo.xlsx"

COSTO_M3   = 1.0   # S/ por m3 — ajustar según tarifa vigente
PENALIDAD  = 20.0  # S/ cargo por reconexión
MES_ACTUAL = datetime.now().strftime("%Y-%m")

# Orden de prioridad para asignar pagos parciales (mayor prioridad primero)
PRIORIDAD_CONCEPTOS = [
    "total_mes", "arrastre", "mant", "corte_reconexion",
    "convenio", "reunion_faena", "techado",
]

# Columnas fijas en blancos_acumulados.xlsx (estructura definida en shared/)
_BLANCOS_COL_MZ   = 13
_BLANCOS_COL_LOTE = 14
_BLANCOS_COL_EST  = 18
_BLANCOS_COL_MES  = 19

# ── PALETA ────────────────────────────────────────────────────────────────────
GH_QUIEN  = ("F4ECF7", "5B21B6")
GH_CONSU  = ("E6F1FB", "0C447C")
GH_DEUDA  = ("FEF9E7", "7D6608")
GH_PAGO   = ("FEF3E8", "7C3003")
GH_QUEDA  = ("E1F5EE", "085041")
GH_PORQUE = ("FEF9E7", "7D6608")
GH_PAGAR  = ("FEF2F2", "991B1B")
GH_PLAN   = ("E0F2FE", "0369A1")

TD_QUIEN  = "FAF5FF"
TD_CONSU  = "F0F8FF"
TD_DEUDA  = "FEFCE8"
TD_PAGO   = "FEF6EE"
TD_QUEDA  = "F0FFF8"
TD_PORQUE = "FEFCE8"
TD_PAGAR  = "FEF2F2"
TD_PLAN   = "F0F9FF"

ESTADO_BG  = {"CANCELADO": "E1F5EE", "EXCESO": "EFF6FF",
               "PARCIAL":   "FAEEDA", "PENDIENTE": "FEF2F2"}
ESTADO_TXT = {"CANCELADO": "085041", "EXCESO": "1D4ED8",
               "PARCIAL":  "854F0B", "PENDIENTE": "991B1B"}

# ── LOGGING ───────────────────────────────────────────────────────────────────
def _init_logging():
    OUTPUTS_DIR.mkdir(exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(OUTPUTS_DIR / "run.log", encoding="utf-8"),
        ],
    )

log = logging.getLogger(__name__)

# ── ESTILO ────────────────────────────────────────────────────────────────────
def _borde():
    b = Side(style="thin", color="CCCCCC")
    return Border(left=b, right=b, top=b, bottom=b)

def _c(ws, row, col, value=None, bg=None, txt="333333",
       bold=False, align="left", mono=False, size=9):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Consolas" if mono else "Arial",
                       size=size, bold=bold, color=txt)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _borde()
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    return c

def _gh(ws, row, cs, ce, texto, bg, txt):
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    c = ws.cell(row=row, column=cs, value=texto)
    c.font      = Font(name="Arial", size=8, bold=True, color=txt)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", start_color=bg)
    c.border    = _borde()

def _ch(ws, row, col, texto, bg, txt):
    _c(ws, row, col, texto, bg=bg, txt=txt, bold=True, align="center")

def _sep(ws, col, row_end):
    letra = get_column_letter(col)
    ws.column_dimensions[letra].width = 0.8
    b_sep = Border(left=Side(style="thin", color="D1D5DB"),
                   right=Side(style="thin", color="D1D5DB"))
    for r in range(1, row_end + 1):
        c = ws.cell(row=r, column=col)
        c.fill   = PatternFill("solid", start_color="F3F4F6")
        c.border = b_sep

def _w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ── UTILIDADES ────────────────────────────────────────────────────────────────
def _norm_lt(val) -> str:
    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN"):
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.upper()

def _key(mz, lt) -> str:
    mz = str(mz).strip().upper()
    lt = _norm_lt(lt)
    return f"{mz}-{lt}" if mz and lt else ""

def _float(val) -> float:
    try:
        return float(str(val).replace(",", ".").strip())
    except (ValueError, TypeError):
        return 0.0

def _estado(saldo: float, total_pagado: float) -> str:
    if saldo < -0.005:
        return "EXCESO"
    if abs(saldo) <= 0.005:
        return "CANCELADO"
    return "PARCIAL" if total_pagado > 0.005 else "PENDIENTE"

def _priorizar_pago(u: dict) -> dict:
    """Distribuye el saldo sin pagar desde el concepto de menor prioridad hacia arriba.
    Retorna los montos a llevar a la planilla del mes siguiente por concepto."""
    saldo_neto = round(max(0.0, u["saldo"]), 2)
    pendiente = {}
    for concepto in reversed(PRIORIDAD_CONCEPTOS):
        cargo = round(max(0.0, u.get(concepto, 0.0)), 2)
        asignado = round(min(saldo_neto, cargo), 2)
        pendiente[concepto] = asignado
        saldo_neto = round(saldo_neto - asignado, 2)
    # mant impago va al arrastre igual que consumo: se acumula como deuda de servicios
    arrastre_nvo    = round(pendiente["total_mes"] + pendiente["arrastre"] + pendiente["mant"], 2)
    devolucion_nvo  = round(abs(u["saldo"]), 2) if u["estado"] == "EXCESO" else 0.0
    return {
        "arrastre_nvo":          arrastre_nvo,
        "corte_reconexion_nvo":  pendiente["corte_reconexion"],
        "convenio_nvo":          pendiente["convenio"],
        "reunion_faena_nvo":     pendiente["reunion_faena"],
        "techado_nvo":           pendiente["techado"],
        "devolucion_nvo":        devolucion_nvo,
    }

def _col_names(df) -> list:
    return [str(c).strip().upper() if str(c) not in ("NAN", "NONE") else f"_C{i}"
            for i, c in enumerate(df.columns)]

# ── VALIDACIÓN ────────────────────────────────────────────────────────────────
def _validar_inputs():
    errores = []
    checks = [
        (PLANILLA_BASE_DIR  / PLANILLA_BASE_FILE,  "Usar crear_template.py y llenar planilla_base.xlsx"),
        (PAGOS_YAPE_DIR     / PAGOS_YAPE_FILE,     "Copiar desde 03_pagos/yape/motor_matching/outputs/"),
        (PAGOS_EFECTIVO_DIR / PAGOS_EFECTIVO_FILE, "Copiar desde 03_pagos/efectivo/outputs/"),
    ]
    for ruta, sugerencia in checks:
        if not ruta.exists():
            errores.append(f"Falta: {ruta}\n  → {sugerencia}")
    if errores:
        for e in errores:
            log.error(e)
        raise FileNotFoundError("Inputs faltantes — ver errores arriba")

    df = pd.read_excel(PLANILLA_BASE_DIR / PLANILLA_BASE_FILE, dtype=str)
    requeridas = {"MZ", "LT", "NOMBRE", "MARC_ANT", "MARC_ACT",
                  "ARRASTRE", "CONVENIO", "MANT", "CORTE_RECONEXION",
                  "REUNION_FAENA", "TECHADO", "DEVOLUCION", "AJUSTE"}
    faltantes = requeridas - {c.strip().upper() for c in df.columns}
    if faltantes:
        raise ValueError(f"planilla_base.xlsx — columnas faltantes: {faltantes}")
    log.info("Inputs validados correctamente")

# ── CARGA PLANILLA ────────────────────────────────────────────────────────────
def _cargar_planilla_base() -> list[dict]:
    df = pd.read_excel(PLANILLA_BASE_DIR / PLANILLA_BASE_FILE, dtype=str)
    df.columns = [c.strip().upper() for c in df.columns]
    usuarios = []
    for _, f in df.iterrows():
        mz = str(f.get("MZ", "")).strip().upper()
        lt = _norm_lt(f.get("LT", ""))
        if not mz or not lt:
            continue
        marc_ant = _float(f.get("MARC_ANT", 0))
        marc_act = _float(f.get("MARC_ACT", 0))
        m3       = round(marc_act - marc_ant, 3)
        usuarios.append({
            "mz":            mz,
            "lt":            lt,
            "key":           f"{mz}-{lt}",
            "nombre":        str(f.get("NOMBRE", "")).strip(),
            "marc_ant":      marc_ant,
            "marc_act":      marc_act,
            "m3":            m3,
            "total_mes":     round(m3 * COSTO_M3, 2),
            "arrastre":      _float(f.get("ARRASTRE", 0)),
            "convenio":      _float(f.get("CONVENIO", 0)),
            "mant":              _float(f.get("MANT", 0)),
            "corte_reconexion":  _float(f.get("CORTE_RECONEXION", 0)),
            "reunion_faena":     _float(f.get("REUNION_FAENA", 0)),
            "techado":       _float(f.get("TECHADO", 0)),
            "devolucion":    _float(f.get("DEVOLUCION", 0)),
            "ajuste":        _float(f.get("AJUSTE", 0)),
        })
    log.info(f"Planilla base: {len(usuarios)} usuarios cargados")
    return usuarios

# ── CARGA BLANCOS ─────────────────────────────────────────────────────────────
def _cargar_blancos() -> dict:
    if not BLANCOS_PATH.exists():
        log.info("blancos_acumulados.xlsx no encontrado — sin blancos este mes")
        return {}
    df = pd.read_excel(BLANCOS_PATH, header=1, dtype=str)
    df.columns = _col_names(df)
    blancos = {}
    for _, fila in df.iterrows():
        mz     = str(fila.get("MZ",    "")).strip().upper()
        lote   = _norm_lt(fila.get("LOTE", ""))
        estado = str(fila.get("ESTADO", "")).strip().lower()
        if not mz or mz in ("NAN", "") or not lote:
            continue
        if estado == "aplicado":
            continue
        key = f"{mz}-{lote}"
        blancos[key] = blancos.get(key, 0.0) + _float(fila.get("MONTO", 0))
    if blancos:
        log.info(f"Blancos pendientes: {len(blancos)} lote(s) → {blancos}")
    return blancos

def _actualizar_blancos(blancos_aplicados: set):
    if not blancos_aplicados or not BLANCOS_PATH.exists():
        return
    wb = load_workbook(BLANCOS_PATH)
    ws = wb.active
    # Datos desde fila 3 (filas 1-2 son cabeceras dobles)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        mz   = str(row[_BLANCOS_COL_MZ   - 1].value or "").strip().upper()
        lote = _norm_lt(row[_BLANCOS_COL_LOTE - 1].value)
        if f"{mz}-{lote}" in blancos_aplicados:
            row[_BLANCOS_COL_EST - 1].value = "aplicado"
            row[_BLANCOS_COL_MES - 1].value = MES_ACTUAL
    wb.save(BLANCOS_PATH)
    log.info(f"blancos_acumulados: {len(blancos_aplicados)} registro(s) → aplicado / {MES_ACTUAL}")

# ── CARGA PAGOS YAPE ──────────────────────────────────────────────────────────
def _cargar_pagos_yape() -> dict:
    df = pd.read_excel(PAGOS_YAPE_DIR / PAGOS_YAPE_FILE, header=1, dtype=str)
    df.columns = _col_names(df)
    totales, sin_id = {}, 0
    for _, fila in df.iterrows():
        if str(fila.get("TIPO", "")).strip().upper() != "TE PAGÓ":
            continue
        mz   = str(fila.get("MZ",   "")).strip().upper()
        lote = _norm_lt(fila.get("LOTE", ""))
        conc = str(fila.get("CONCEPTO", "")).strip().upper()
        if not mz or mz == "NAN" or not lote:
            sin_id += 1
            continue
        if conc and conc not in ("NAN", "NONE", ""):
            continue  # gasto comunitario, no pago de lote
        key = f"{mz}-{lote}"
        totales[key] = totales.get(key, 0.0) + _float(fila.get("MONTO_ASIGNA", 0))
    log.info(f"Pagos Yape: {len(totales)} lotes · {sin_id} sin identificar")
    return totales

# ── CARGA PAGOS EFECTIVO ──────────────────────────────────────────────────────
def _cargar_pagos_efectivo() -> dict:
    df = pd.read_excel(PAGOS_EFECTIVO_DIR / PAGOS_EFECTIVO_FILE, header=1, dtype=str)
    df.columns = _col_names(df)
    totales = {}
    for _, fila in df.iterrows():
        mz = str(fila.get("MZ", "")).strip().upper()
        lt = _norm_lt(fila.get("LT", ""))
        if not mz or not lt:
            continue
        key = f"{mz}-{lt}"
        totales[key] = totales.get(key, 0.0) + _float(fila.get("MONTO", 0))
    log.info(f"Pagos Efectivo: {len(totales)} lotes")
    return totales

# ── CALCULAR ──────────────────────────────────────────────────────────────────
def _calcular(usuarios: list[dict],
              pagos_yape: dict,
              pagos_efectivo: dict,
              blancos: dict) -> tuple[list[dict], set]:
    resultado, blancos_usados = [], set()
    for u in usuarios:
        key    = u["key"]
        blanco = round(blancos.get(key, 0.0), 2)
        if blanco:
            blancos_usados.add(key)
        total = round(
            u["total_mes"] + u["arrastre"] + u["convenio"] + u["mant"]
            + u["corte_reconexion"]
            + u["reunion_faena"] + u["techado"]
            - u["devolucion"] - blanco
            + u["ajuste"],
            2,
        )
        yape     = round(pagos_yape.get(key, 0.0), 2)
        efectivo = round(pagos_efectivo.get(key, 0.0), 2)
        pagado   = round(yape + efectivo, 2)
        saldo    = round(total - pagado, 2)
        base = {
            **u,
            "blanco":       blanco,
            "total_deuda":  total,
            "yape":         yape,
            "efectivo":     efectivo,
            "total_pagado": pagado,
            "saldo":        saldo,
            "estado":       _estado(saldo, pagado),
        }
        resultado.append({**base, **_priorizar_pago(base)})
    cnt = {e: sum(1 for r in resultado if r["estado"] == e)
           for e in ("CANCELADO", "EXCESO", "PARCIAL", "PENDIENTE")}
    log.info(f"Resultado: CANCELADO={cnt['CANCELADO']} · EXCESO={cnt['EXCESO']} "
             f"· PARCIAL={cnt['PARCIAL']} · PENDIENTE={cnt['PENDIENTE']}")
    return resultado, blancos_usados

# ── EXPORT: COBRANZA FINAL ────────────────────────────────────────────────────
# Col layout:
#  1- 3  ¿Quién es?        MZ LT NOMBRE
#  4      sep
#  5- 9  ¿Cuánto consumió? MARC_ANT MARC_ACT m3 COSTO_m3 TOTAL_MES
# 10      sep
# 11-20  ¿Cuánto debía?    ARRASTRE CONVENIO MANT CORTE_RECONEXION
#                           REUNION_FAENA TECHADO DEVOLUCION BLANCOS AJUSTE TOTAL
# 21      sep
# 22-24  ¿Cómo pagó?       YAPE EFECTIVO TOTAL_PAGADO
# 25      sep
# 26-27  ¿Qué queda?       SALDO ESTADO

_CF_GRUPOS = [
    (1,  3,  "¿Quién es?",        *GH_QUIEN),
    (5,  9,  "¿Cuánto consumió?", *GH_CONSU),
    (11, 20, "¿Cuánto debía?",    *GH_DEUDA),
    (22, 24, "¿Cómo pagó?",       *GH_PAGO),
    (26, 27, "¿Qué queda?",       *GH_QUEDA),
]
_CF_COLS = [
    # (col, nombre, bg, txt, ancho)
    (1,  "MZ",               *GH_QUIEN,  6),
    (2,  "LT",               *GH_QUIEN,  7),
    (3,  "NOMBRE",           *GH_QUIEN, 26),
    (5,  "MARC_ANT",         *GH_CONSU, 10),
    (6,  "MARC_ACT",         *GH_CONSU, 10),
    (7,  "m3",               *GH_CONSU,  7),
    (8,  "COSTO_m3",         *GH_CONSU,  9),
    (9,  "TOTAL_MES",        *GH_CONSU, 11),
    (11, "ARRASTRE",         *GH_DEUDA, 10),
    (12, "CONVENIO",         *GH_DEUDA, 10),
    (13, "MANT",             *GH_DEUDA,  8),
    (14, "CORTE_RECONEXION", *GH_DEUDA, 16),
    (15, "REUNION_FAENA",    *GH_DEUDA, 14),
    (16, "TECHADO",          *GH_DEUDA, 10),
    (17, "DEVOLUCION",       *GH_DEUDA, 12),
    (18, "BLANCOS",          *GH_DEUDA, 10),
    (19, "AJUSTE",           *GH_DEUDA,  9),
    (20, "TOTAL",            *GH_DEUDA, 11),
    (22, "YAPE",             *GH_PAGO,  10),
    (23, "EFECTIVO",         *GH_PAGO,  10),
    (24, "TOTAL_PAGADO",     *GH_PAGO,  13),
    (26, "SALDO",            *GH_QUEDA, 10),
    (27, "ESTADO",           *GH_QUEDA, 12),
]
_CF_SEP_COLS = [4, 10, 21, 25]

def _exportar_cobranza_final(resultado: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title       = "cobranza_final"
    ws.freeze_panes = "A3"
    last_row = len(resultado) + 2

    for cs, ce, texto, bg, txt in _CF_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for sc in _CF_SEP_COLS:
        _sep(ws, sc, last_row)
    for col, nombre, bg, txt, ancho in _CF_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 20

    for ri, r in enumerate(resultado, 3):
        _c(ws, ri, 1, r["mz"],     TD_QUIEN, "5B21B6", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_QUIEN, "5B21B6", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_QUIEN, "333333", align="left")

        _c(ws, ri, 5, r["marc_ant"],  TD_CONSU, "185FA5", mono=True, align="right")
        _c(ws, ri, 6, r["marc_act"],  TD_CONSU, "185FA5", mono=True, align="right")
        _c(ws, ri, 7, r["m3"],        TD_CONSU, "065F46", mono=True, align="right")
        _c(ws, ri, 8, COSTO_M3,       TD_CONSU, "185FA5", mono=True, align="right")
        _c(ws, ri, 9, r["total_mes"], TD_CONSU, "065F46", mono=True, align="right")

        def _d(col, val, reduccion=False):
            txt = "065F46" if (reduccion and val > 0.005) else "7D6608"
            _c(ws, ri, col, val or None, TD_DEUDA, txt, mono=True, align="right")

        _d(11, r["arrastre"])
        _d(12, r["convenio"])
        _d(13, r["mant"])
        _d(14, r["corte_reconexion"])
        _d(15, r["reunion_faena"])
        _d(16, r["techado"])
        _d(17, r["devolucion"],  reduccion=True)
        _d(18, r["blanco"],      reduccion=True)
        _c(ws, ri, 19, r["ajuste"] or None,  TD_DEUDA, "7D6608", mono=True, align="right")
        _c(ws, ri, 20, r["total_deuda"],      TD_DEUDA, "7D6608", bold=True, mono=True, align="right")

        _c(ws, ri, 22, r["yape"]        or None, TD_PAGO, "7C3003", mono=True, align="right")
        _c(ws, ri, 23, r["efectivo"]    or None, TD_PAGO, "7C3003", mono=True, align="right")
        _c(ws, ri, 24, r["total_pagado"]or None, TD_PAGO, "7C3003", bold=True, mono=True, align="right")

        saldo_txt = ("A32D2D" if r["saldo"] > 0.005
                     else "1D4ED8" if r["saldo"] < -0.005
                     else "065F46")
        _c(ws, ri, 26, r["saldo"], TD_QUEDA, saldo_txt, bold=True, mono=True, align="right")

        est_bg  = ESTADO_BG.get(r["estado"],  "FFFFFF")
        est_txt = ESTADO_TXT.get(r["estado"], "333333")
        c_est = ws.cell(row=ri, column=27, value=r["estado"])
        c_est.font      = Font(name="Arial", size=9, bold=True, color=est_txt)
        c_est.fill      = PatternFill("solid", start_color=est_bg)
        c_est.alignment = Alignment(horizontal="center", vertical="center")
        c_est.border    = _borde()
        ws.row_dimensions[ri].height = 17

    wb.save(OUTPUTS_DIR / "cobranza_final.xlsx")
    log.info(f"cobranza_final.xlsx → {len(resultado)} filas")

# ── EXPORT: LISTA CORTE ───────────────────────────────────────────────────────
# Col layout:
#  1-3  ¿Quién es?          MZ LT NOMBRE
#  4     sep
#  5-6  ¿Por qué va corte?  DEUDA_ARRASTRE SALDO
#  7     sep
#  8-9  ¿Qué debe pagar?    PENALIDAD TOTAL_A_PAGAR

_LC_GRUPOS = [
    (1, 3, "¿Quién es?",           *GH_QUIEN),
    (5, 6, "¿Por qué va a corte?", *GH_PORQUE),
    (8, 9, "¿Qué debe pagar?",     *GH_PAGAR),
]
_LC_COLS = [
    (1, "MZ",             *GH_QUIEN,   6),
    (2, "LT",             *GH_QUIEN,   7),
    (3, "NOMBRE",         *GH_QUIEN,  26),
    (5, "DEUDA_ARRASTRE", *GH_PORQUE, 14),
    (6, "SALDO",          *GH_PORQUE, 11),
    (8, "PENALIDAD",      *GH_PAGAR,  11),
    (9, "TOTAL_A_PAGAR",  *GH_PAGAR,  14),
]
_LC_SEP_COLS = [4, 7]

def _exportar_lista_corte(resultado: list[dict]) -> int:
    # Condición: saldo pendiente Y arrastre >= 8 (mínimo mensual = S/8 mant, confirma que no pagó nada el mes anterior)
    corte    = [r for r in resultado if r["saldo"] > 0.005 and r["arrastre"] >= 8.0 - 0.005]
    last_row = len(corte) + 2

    wb = Workbook()
    ws = wb.active
    ws.title       = "lista_corte"
    ws.freeze_panes = "A3"

    for cs, ce, texto, bg, txt in _LC_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for sc in _LC_SEP_COLS:
        _sep(ws, sc, last_row)
    for col, nombre, bg, txt, ancho in _LC_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 20

    for ri, r in enumerate(corte, 3):
        tap = round(r["saldo"] + PENALIDAD, 2)
        _c(ws, ri, 1, r["mz"],       TD_QUIEN,  "5B21B6", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],       TD_QUIEN,  "5B21B6", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"],   TD_QUIEN,  "333333", align="left")
        _c(ws, ri, 5, r["arrastre"], TD_PORQUE, "92400E", mono=True, align="right")
        _c(ws, ri, 6, r["saldo"],    TD_PORQUE, "92400E", mono=True, align="right")
        _c(ws, ri, 8, PENALIDAD,     TD_PAGAR,  "991B1B", mono=True, align="right")
        _c(ws, ri, 9, tap,           TD_PAGAR,  "7F1D1D", bold=True, mono=True, align="right")
        ws.row_dimensions[ri].height = 17

    wb.save(OUTPUTS_DIR / "lista_corte.xlsx")
    log.info(f"lista_corte.xlsx → {len(corte)} usuarios en corte")
    return len(corte)

# ── EXPORT: RESUMEN ───────────────────────────────────────────────────────────
def _exportar_resumen(resultado: list[dict], n_corte: int):
    tot = {
        "deuda":    sum(r["total_deuda"]  for r in resultado),
        "yape":     sum(r["yape"]         for r in resultado),
        "efectivo": sum(r["efectivo"]     for r in resultado),
        "pagado":   sum(r["total_pagado"] for r in resultado),
        "saldo":    sum(r["saldo"] for r in resultado if r["saldo"] > 0),
    }
    cnt = {e: sum(1 for r in resultado if r["estado"] == e)
           for e in ("CANCELADO", "EXCESO", "PARCIAL", "PENDIENTE")}

    filas = [
        ("── RECAUDACIÓN ──────────────────", None,             None),
        ("Total deuda planilla",              tot["deuda"],     "S/"),
        ("Total recaudado Yape",              tot["yape"],      "S/"),
        ("Total recaudado Efectivo",          tot["efectivo"],  "S/"),
        ("Total recaudado (combinado)",       tot["pagado"],    "S/"),
        ("Saldo pendiente total",             tot["saldo"],     "S/"),
        ("── ESTADOS ──────────────────────", None,             None),
        ("Usuarios CANCELADO",                cnt["CANCELADO"], "usuarios"),
        ("Usuarios EXCESO",                   cnt["EXCESO"],    "usuarios"),
        ("Usuarios PARCIAL",                  cnt["PARCIAL"],   "usuarios"),
        ("Usuarios PENDIENTE",                cnt["PENDIENTE"], "usuarios"),
        ("── CORTE ────────────────────────", None,             None),
        ("Usuarios lista_corte (penalidad S/20)", n_corte,      "usuarios"),
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "resumen"
    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value=f"Resumen cobranza  {MES_ACTUAL}")
    t.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    t.fill      = PatternFill("solid", start_color="4A235A")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10

    for ri, (concepto, valor, unidad) in enumerate(filas, 2):
        es_sep = valor is None
        bg = "E8DAEF" if es_sep else "FAF5FF"
        _c(ws, ri, 1, concepto, bg=bg, bold=es_sep, align="left",
           txt="5B21B6" if es_sep else "333333")
        _c(ws, ri, 2, valor,    bg=bg, align="right", mono=not es_sep)
        _c(ws, ri, 3, unidad,   bg=bg, align="left",  txt="888888")
        ws.row_dimensions[ri].height = 17

    wb.save(OUTPUTS_DIR / "resumen_recaudacion.xlsx")
    log.info("resumen_recaudacion.xlsx generado")

# ── EXPORT: PLANILLA SIGUIENTE ────────────────────────────────────────────────
# Solo usuarios PARCIAL / PENDIENTE / EXCESO.
# Col layout:
#  1- 3  ¿Quién es?                   MZ LT NOMBRE
#  4      sep
#  5- 6  ¿Qué quedó este mes?         ESTADO SALDO
#  7      sep
#  8-13  Llevar a planilla siguiente  ARRASTRE CORTE_RECONEXION CONVENIO
#                                     REUNION_FAENA TECHADO DEVOLUCION

_PS_GRUPOS = [
    (1,  3,  "¿Quién es?",                *GH_QUIEN),
    (5,  6,  "¿Qué quedó este mes?",       *GH_QUEDA),
    (8,  13, "Llevar a planilla siguiente", *GH_PLAN),
]
_PS_COLS = [
    (1,  "MZ",               *GH_QUIEN,  6),
    (2,  "LT",               *GH_QUIEN,  7),
    (3,  "NOMBRE",           *GH_QUIEN, 26),
    (5,  "ESTADO",           *GH_QUEDA, 12),
    (6,  "SALDO",            *GH_QUEDA, 10),
    (8,  "ARRASTRE",         *GH_PLAN,  10),
    (9,  "CORTE_RECONEXION", *GH_PLAN,  16),
    (10, "CONVENIO",         *GH_PLAN,  10),
    (11, "REUNION_FAENA",    *GH_PLAN,  13),
    (12, "TECHADO",          *GH_PLAN,  10),
    (13, "DEVOLUCION",       *GH_PLAN,  12),
]
_PS_SEP_COLS = [4, 7]

def _exportar_planilla_siguiente(resultado: list[dict]):
    pendientes = [r for r in resultado if r["estado"] in ("PARCIAL", "PENDIENTE", "EXCESO")]
    if not pendientes:
        log.info("planilla_siguiente_prep: todos cancelados — archivo no generado")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "planilla_siguiente"
    ws.freeze_panes = "A3"
    last_row = len(pendientes) + 2

    for cs, ce, texto, bg, txt in _PS_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for sc in _PS_SEP_COLS:
        _sep(ws, sc, last_row)
    for col, nombre, bg, txt, ancho in _PS_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 20

    for ri, r in enumerate(pendientes, 3):
        est_bg  = ESTADO_BG.get(r["estado"],  "FFFFFF")
        est_txt = ESTADO_TXT.get(r["estado"], "333333")

        _c(ws, ri, 1, r["mz"],     TD_QUIEN, "5B21B6", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_QUIEN, "5B21B6", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_QUIEN, "333333", align="left")

        c_est = ws.cell(row=ri, column=5, value=r["estado"])
        c_est.font      = Font(name="Arial", size=9, bold=True, color=est_txt)
        c_est.fill      = PatternFill("solid", start_color=est_bg)
        c_est.alignment = Alignment(horizontal="center", vertical="center")
        c_est.border    = _borde()

        saldo_txt = ("A32D2D" if r["saldo"] > 0.005
                     else "1D4ED8" if r["saldo"] < -0.005
                     else "065F46")
        _c(ws, ri, 6, r["saldo"], TD_QUEDA, saldo_txt, bold=True, mono=True, align="right")

        def _pn(col, val, ri=ri):
            _c(ws, ri, col, val or None, TD_PLAN, "0369A1", mono=True, align="right")

        _pn(8,  r["arrastre_nvo"])
        _pn(9,  r["corte_reconexion_nvo"])
        _pn(10, r["convenio_nvo"])
        _pn(11, r["reunion_faena_nvo"])
        _pn(12, r["techado_nvo"])
        _pn(13, r["devolucion_nvo"])

        ws.row_dimensions[ri].height = 17

    wb.save(OUTPUTS_DIR / "planilla_siguiente_prep.xlsx")
    log.info(f"planilla_siguiente_prep.xlsx → {len(pendientes)} fila(s) con saldo pendiente")

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "═" * 55)
    print("  04_cobranza — Procesamiento de cobranza")
    print("═" * 55)
    _init_logging()

    print("\n[1/5] Validando inputs...")
    _validar_inputs()

    print("\n[2/5] Cargando datos...")
    usuarios       = _cargar_planilla_base()
    pagos_yape     = _cargar_pagos_yape()
    pagos_efectivo = _cargar_pagos_efectivo()
    blancos        = _cargar_blancos()

    print("\n[3/5] Calculando cobranza...")
    resultado, blancos_usados = _calcular(usuarios, pagos_yape, pagos_efectivo, blancos)

    print("\n[4/5] Actualizando blancos...")
    _actualizar_blancos(blancos_usados)

    print("\n[5/5] Exportando outputs...")
    _exportar_cobranza_final(resultado)
    n_corte = _exportar_lista_corte(resultado)
    _exportar_resumen(resultado, n_corte)
    _exportar_planilla_siguiente(resultado)

    n_pend = sum(1 for r in resultado if r["estado"] in ("PARCIAL", "PENDIENTE", "EXCESO"))
    print("\n" + "═" * 55)
    print("  Cobranza completada → revisar outputs/")
    print("  Pasar lista_corte.xlsx → 05_corte")
    if n_pend:
        print(f"  {n_pend} con saldo → revisar planilla_siguiente_prep.xlsx")
    print("═" * 55 + "\n")


if __name__ == "__main__":
    main()
