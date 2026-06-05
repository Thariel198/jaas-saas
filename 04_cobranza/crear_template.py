# Crea planilla_base.xlsx vacía con las columnas correctas.
# Correr una sola vez: python crear_template.py
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DESTINO = Path(__file__).parent / "inputs" / "planilla_base" / "planilla_base.xlsx"

COLS = [
    # (nombre, desc, ancho)
    ("MZ",            "Manzana — ej: A, B, G1",                        8),
    ("LT",            "Lote — ej: 1, 7, 11A",                          8),
    ("NOMBRE",        "Nombre completo del usuario",                   30),
    ("MARC_ANT",      "Marcación anterior (lectura mes pasado)",        14),
    ("MARC_ACT",      "Marcación actual (lectura este mes)",            14),
    ("ARRASTRE",      "Deuda del mes anterior (0 si al día)",           12),
    ("CONVENIO",      "Cuota de convenio de pago activo (0 si no)",     12),
    ("MANT",              "Mantenimiento mensual fijo (ej: 3)",              8),
    ("CORTE_RECONEXION", "Penalidad corte y reconexión del mes pasado (0 si no)", 16),
    ("REUNION_FAENA",    "Multa por reunión/faena no asistida (0 si no)",  15),
    ("TECHADO",       "Cuota techado si aplica (0 si no)",              12),
    ("DEVOLUCION",    "Exceso anterior reconocido — reduce deuda",      13),
    ("AJUSTE",        "Ajuste manual + o - (0 si no aplica)",           10),
]

EJEMPLO = ["A", "1", "JUAN PEREZ GARCIA", 100, 112, 0, 0, 3, 0, 0, 0, 0, 0]

# Paleta coherente con el módulo (púrpura JASS)
C_HEAD_BG  = "4A235A"
C_DESC_BG  = "F4ECF7"
C_EJMP_BG  = "FAFAFA"


def main():
    DESTINO.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title       = "planilla_base"
    ws.freeze_panes = "A3"

    b = Side(style="thin", color="CCCCCC")
    borde = Border(left=b, right=b, top=b, bottom=b)

    for ci, (nombre, _, ancho) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=nombre)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", start_color=C_HEAD_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = borde
        ws.column_dimensions[c.column_letter].width = ancho

    for ci, (_, desc, _) in enumerate(COLS, 1):
        c = ws.cell(row=2, column=ci, value=desc)
        c.font      = Font(name="Arial", italic=True, color="666666", size=9)
        c.fill      = PatternFill("solid", start_color=C_DESC_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = borde

    for ci, val in enumerate(EJEMPLO, 1):
        c = ws.cell(row=3, column=ci, value=val)
        c.font      = Font(name="Arial", color="999999", size=9, italic=True)
        c.fill      = PatternFill("solid", start_color=C_EJMP_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = borde

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    nota = ws.cell(row=4, column=1,
                   value="← Llenar desde aquí. Eliminar filas 2 y 3 antes de correr main.py")
    nota.font = Font(name="Arial", color="AAAAAA", italic=True, size=8)

    wb.save(DESTINO)
    print(f"Template creado: {DESTINO}")
    print("Llenar desde fila 4. Eliminar filas 2 y 3 antes de correr main.py.")
    print("m3 y TOTAL_MES los calcula main.py a partir de MARC_ANT y MARC_ACT.")


if __name__ == "__main__":
    main()
