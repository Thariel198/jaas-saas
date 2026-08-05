import glob
import logging
import shutil
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config

# seguimiento_pueblo: writer único es shared/seguimiento_repo.py (patrón repo).
# Se importa por ruta física, no depende de config (que sí se monkey-patchea
# en tests) porque el módulo vive siempre en el shared/ real.
sys.path.insert(0, str(config.BASE_DIR.parent / "shared"))
import seguimiento_repo as repo  # noqa: E402
import utils_estado_ciclo as repo_estado  # noqa: E402
import ciclo as ciclo_activo  # noqa: E402

log = logging.getLogger(__name__)


# ── Normalización de clave (MZ, LT) ──────────────────────────────────────

def _norm_mz(v) -> str:
    return str(v).strip().upper()


def _norm_lt(v) -> str:
    s = str(v).strip()
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
    except ValueError:
        pass
    return s.upper()


def _add_key(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_mz"] = df["MZ"].map(_norm_mz)
    df["_lt"] = df["LT"].map(_norm_lt)
    return df


# ── Carga de inputs ───────────────────────────────────────────────────────

def _load_lecturas(mes: str) -> pd.DataFrame:
    """lecturas_planilla_YYYY-MM.xlsx (1_lecturas) trae la fila de grupos en la 1
    y los nombres de columna en la 2 → header=1 (mismo formato que arrastre_consolidado,
    ver _load_consolidado)."""
    path = config.lecturas_path(mes)
    df = pd.read_excel(path, header=1, dtype=str)
    missing = [c for c in config.COLS_LECTURAS if c not in df.columns]
    if missing:
        raise ValueError(f"lecturas_planilla: columnas faltantes: {missing}")
    return _add_key(df)


def _load_optional(path: Path, required_cols: list, label: str) -> pd.DataFrame | None:
    if not path.exists():
        log.warning(f"{label}: archivo no encontrado -> valores = 0")
        return None
    df = pd.read_excel(path, dtype=str)
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        log.warning(f"{label}: columnas faltantes {missing} -> valores = 0")
        return None
    return _add_key(df)


# ── Fuente única de arrastres — arrastre_consolidado del mes anterior ──────

def _mes_anterior(mes: str) -> str:
    y, m = mes.split("-")
    y, m = int(y), int(m)
    m -= 1
    if m == 0:
        m, y = 12, y - 1
    return f"{y:04d}-{m:02d}"


def _ciclo_validado(mes: str) -> bool:
    return repo_estado.ciclo_validado(mes, ruta=config.ESTADO_CICLO_PATH)


def _load_consolidado(mes: str) -> pd.DataFrame | None:
    """Lee arrastre_consolidado del mes anterior (writer único = 5_cobranza/outputs).

    Devuelve None si no existe (mes de génesis / pipeline sin cerrar → arrastres=0).
    Aborta si existe pero el ciclo anterior no está validado (dato no confiable).
    El archivo trae la fila de grupos en la 1 y los nombres de columna en la 2
    → header=1. Los guiones (—) de componentes en 0 se coercionan a 0 en el join.
    """
    mes_ant = _mes_anterior(mes)
    path = config.consolidado_path(mes_ant)
    if not path.exists():
        log.warning(f"arrastre_consolidado {mes_ant}: no encontrado -> arrastres = 0 "
                    f"(mes de génesis o ciclo anterior sin cerrar)")
        return None
    if not _ciclo_validado(mes_ant):
        raise ValueError(
            f"arrastre_consolidado_{mes_ant}.xlsx existe pero el ciclo {mes_ant} NO está "
            f"validado en estado_ciclo.json — correr 5b_validacion antes de generar la planilla."
        )
    df = pd.read_excel(path, header=1, dtype=str)
    missing = [c for c in config.COLS_CONSOLIDADO if c not in df.columns]
    if missing:
        raise ValueError(f"arrastre_consolidado {mes_ant}: columnas faltantes {missing}")
    log.info(f"arrastre_consolidado {mes_ant} -> {len(df)} filas de arrastre")
    return _add_key(df)


# ── Join de arrastres ─────────────────────────────────────────────────────

def _join_optional(base: pd.DataFrame, src: pd.DataFrame | None,
                   src_col: str, dest_col: str, warn: bool = True) -> pd.DataFrame:
    if src is None:
        base = base.copy()
        base[dest_col] = 0.0
        return base

    # Advertir sobre filas en src sin match en base
    if warn:
        check = src.merge(base[["_mz", "_lt"]], on=["_mz", "_lt"], how="left", indicator=True)
        sin_match = (check["_merge"] == "left_only").sum()
        if sin_match:
            log.warning(f"arrastre: {sin_match} fila(s) sin match en lecturas -> ignoradas")

    merged = base.merge(
        src[["_mz", "_lt", src_col]].rename(columns={src_col: dest_col}),
        on=["_mz", "_lt"],
        how="left",
    )
    merged[dest_col] = pd.to_numeric(merged[dest_col], errors="coerce").fillna(0)
    return merged


# ── Build del dataframe de planilla ──────────────────────────────────────

def _join_saldo_pueblo(df: pd.DataFrame, concepto: str, mes_ant: str, dest_col: str) -> pd.DataFrame:
    """MULTA/ACUERDOS_ASAMBLEA/CONVENIO vienen de seguimiento_pueblo, no del
    consolidado — writer único es seguimiento_repo (ver shared/README.md).
    get_saldos_bulk lee el registro UNA vez para todos los predios."""
    saldos = repo.get_saldos_bulk(concepto, mes_ant)
    df = df.copy()
    df[dest_col] = [saldos.get((mz, lt), 0.0) for mz, lt in zip(df["_mz"], df["_lt"])]
    return df


def _nombres_padron() -> dict:
    """Fallback de nombre para predios que solo tienen deuda de pueblo (MULTA/
    ACUERDOS/CONVENIO) y por eso no aparecen en arrastre_consolidado (que solo
    trae agua+corte) — sin esto, get_saldos_bulk() los agrega sin nombre."""
    path = config.BASE_DIR.parent / "0_padron" / "02_matching" / "outputs" / "padron_reconciliado.xlsx"
    if not path.exists():
        return {}
    df = pd.read_excel(path, sheet_name="cobranza")
    return {(_norm_mz(r["MZ"]), _norm_lt(r["LT"])): str(r["Nombres"]).strip()
            for _, r in df.iterrows() if pd.notna(r.get("Nombres"))}


def _extra_keys_deuda_pueblo(df_cons: pd.DataFrame | None, mes_ant: str) -> dict:
    """Predios sin lectura este ciclo (SIN_MEDIDOR/sin_servicio) pero con deuda
    de agua arrastrada o deuda del pueblo (multa/acuerdos/convenio) pendiente.
    Esa deuda no depende del medidor y no debe perderse solo porque el predio
    no tiene fila en lecturas_planilla. Retorna {(mz, lt): nombre}."""
    extra: dict = {}
    nombres: dict = {}
    if df_cons is not None:
        nombres = {(r["_mz"], r["_lt"]): str(r.get("NOMBRE") or "")
                  for _, r in df_cons.iterrows()}
        for _, r in df_cons.iterrows():
            deuda = _to_num(r.get("DEUDA_AGUA")) + _to_num(r.get("CORTE_RECONEXION"))
            if deuda > 0:
                extra[(r["_mz"], r["_lt"])] = nombres.get((r["_mz"], r["_lt"]), "")
    nombres_padron = _nombres_padron()
    for concepto in ("MULTA", "ACUERDOS", "CONVENIO"):
        for (mz, lt), monto in repo.get_saldos_bulk(concepto, mes_ant).items():
            if monto:
                key = (_norm_mz(mz), _norm_lt(lt))
                extra[key] = extra.get(key) or nombres.get(key, "") or nombres_padron.get(key, "")
    return extra


def _to_num(v) -> float:
    n = pd.to_numeric(v, errors="coerce")
    return 0.0 if pd.isna(n) else float(n)


def build_planilla(mes: str) -> pd.DataFrame:
    df = _load_lecturas(mes)
    log.info(f"Lecturas cargadas: {len(df)} usuarios · mes {mes}")

    for col in ["MARC_ANT", "MARC_ACT", "M3"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Agua + corte: el consolidado del mes anterior (5_cobranza) los trae ya
    # descompuestos por prioridad. DEUDA_AGUA es la deuda de agua no cubierta
    # → alimenta MES_ANTERIOR.
    df_cons = _load_consolidado(mes)

    # Predios sin lectura (SIN_MEDIDOR) con deuda de agua/pueblo pendiente:
    # se agregan con consumo y mantenimiento en 0 — no tienen servicio activo,
    # pero su deuda del pueblo (multa/acuerdos/convenio) no depende del medidor.
    mes_ant = _mes_anterior(mes)
    keys_existentes = set(zip(df["_mz"], df["_lt"]))
    extra = {k: v for k, v in _extra_keys_deuda_pueblo(df_cons, mes_ant).items()
             if k not in keys_existentes}
    sin_lectura_keys = set(extra.keys())
    if extra:
        filas_extra = [{"MZ": mz, "LT": lt, "NOMBRE": nom, "MES_ANO": mes,
                        "MARC_ANT": 0.0, "MARC_ACT": 0.0, "M3": 0.0,
                        "_mz": mz, "_lt": lt}
                       for (mz, lt), nom in sorted(extra.items())]
        df = pd.concat([df, pd.DataFrame(filas_extra)], ignore_index=True)
        log.info(f"{len(extra)} predio(s) sin lectura (SIN_MEDIDOR) agregados por "
                 f"deuda de agua/pueblo pendiente: {sorted(extra.keys())}")

    df = _join_optional(df, df_cons, "DEUDA_AGUA",        "MES_ANTERIOR")
    df = _join_optional(df, df_cons, "CORTE_RECONEXION",  "CORTE_RECONEXION",  warn=False)

    # Pueblo (MULTA/ACUERDOS/CONVENIO): fuente única es seguimiento_pueblo,
    # saldo al cierre del mes anterior — el consolidado ya no los carga.
    df = _join_saldo_pueblo(df, "MULTA",    mes_ant, "MULTA")
    df = _join_saldo_pueblo(df, "ACUERDOS", mes_ant, "ACUERDOS_ASAMBLEA")
    df = _join_saldo_pueblo(df, "CONVENIO", mes_ant, "CONVENIO")

    df["MES_ACTUAL"]    = df["M3"].apply(
        lambda m: max(float(m) * config.TARIFA_M3, config.TARIFA_MIN)
    )
    df["MANTENIMIENTO"] = float(config.MANT_FIJO)
    # Sin servicio de agua → sin consumo ni mantenimiento, solo deuda del pueblo.
    if sin_lectura_keys:
        mask_sin_lectura = df.apply(lambda r: (r["_mz"], r["_lt"]) in sin_lectura_keys, axis=1)
        df.loc[mask_sin_lectura, ["MES_ACTUAL", "MANTENIMIENTO"]] = 0.0
    df["BLANCO"]        = 0.0
    df["DEVOLUCION"]    = 0.0

    # TOTAL_A_PAGAR y columnas de pago son fórmula/vacío — se escriben en Excel
    df["TOTAL_A_PAGAR"]  = None
    df["MONTO_YAPE"]     = None
    df["MONTO_EFECTIVO"] = None
    df["ESTADO"]         = None
    df["FECHA_PAGO"]     = None

    return df


# ── Escritura Excel ───────────────────────────────────────────────────────

_SECTION_LABELS = [
    ("¿Quién es?",      "MZ",            "MES_ANO"),
    ("Lectura",         "MARC_ANT",      "M3"),
    ("Cobro — cargos",  "MES_ACTUAL",    "ACUERDOS_ASAMBLEA"),
    ("Descuentos",      "BLANCO",        "DEVOLUCION"),
    ("Total",           "TOTAL_A_PAGAR", "TOTAL_A_PAGAR"),
    ("Pago → 4_pagos", "MONTO_YAPE",    "FECHA_PAGO"),
]

# Columnas con valores numéricos → alineación derecha en datos
_RIGHT_COLS = {
    "MARC_ANT", "MARC_ACT", "M3",
    "MES_ACTUAL", "MANTENIMIENTO", "MES_ANTERIOR", "CORTE_RECONEXION",
    "CONVENIO", "MULTA", "ACUERDOS_ASAMBLEA",
    "BLANCO", "DEVOLUCION", "TOTAL_A_PAGAR",
    "MONTO_YAPE", "MONTO_EFECTIVO",
}


def _argb(hex6: str) -> str:
    return "FF" + hex6.lstrip("#")


def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=_argb(hex6))


def _cell_align(col: str) -> str:
    sec = config.COL_SECTION[col]
    if sec is config.SEC_PAGO:
        return "center"
    if col in _RIGHT_COLS:
        return "right"
    if col == "NOMBRE":
        return "left"
    return "center"


def write_excel(df: pd.DataFrame, mes: str) -> None:
    cols = config.OUTPUT_COLS
    col_idx = {col: i + 1 for i, col in enumerate(cols)}  # 1-indexed

    # Columnas sumandas: MES_ACTUAL … DEVOLUCION (OUTPUT_COLS índices 7–15)
    summand_cols = cols[7:16]

    wb = Workbook()
    ws = wb.active
    ws.title = config.OUTPUT_SHEET

    # ── Fila 1: secciones ──────────────────────────────────────────────
    for label, start, end in _SECTION_LABELS:
        c1 = col_idx[start]
        c2 = col_idx[end]
        sec = config.COL_SECTION[start]
        if c1 != c2:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        cell = ws.cell(row=1, column=c1, value=label)
        cell.fill = _fill(sec["header_bg"])
        cell.font = Font(color=_argb(sec["header_fg"]), bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = config.ROW_HEIGHT_SECTIONS

    # ── Fila 2: nombres de columna ─────────────────────────────────────
    for i, col in enumerate(cols, start=1):
        sec = config.COL_SECTION[col]
        cell = ws.cell(row=2, column=i, value=col)
        cell.fill = _fill(sec["header_bg"])
        cell.font = Font(color=_argb(sec["header_fg"]), bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = config.ROW_HEIGHT_COLS

    # ── Filas de datos (desde fila 3) ─────────────────────────────────
    for r_offset, (_, row) in enumerate(df.iterrows()):
        xl_row = r_offset + 3

        for col in cols:
            c = col_idx[col]
            sec = config.COL_SECTION[col]
            is_total = col == "TOTAL_A_PAGAR"
            is_pago  = sec is config.SEC_PAGO

            if is_total:
                val = "=" + "+".join(
                    f"{get_column_letter(col_idx[sc])}{xl_row}"
                    for sc in summand_cols
                )
            else:
                val = row.get(col)
                try:
                    if pd.isna(val):
                        val = None
                except TypeError:
                    pass

            cell = ws.cell(row=xl_row, column=c, value=val)
            cell.fill = _fill(sec["data_bg"])
            cell.number_format = config.COL_FORMAT[col]
            cell.alignment = Alignment(horizontal=_cell_align(col), vertical="center")
            cell.font = Font(
                color=_argb(sec["data_fg"]),
                bold=is_total,
                italic=is_pago,
                size=10,
            )

    # ── Anchos de columna ──────────────────────────────────────────────
    for i, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = config.COL_WIDTH[col]

    # ── Congelar paneles ───────────────────────────────────────────────
    ws.freeze_panes = config.FREEZE_PANES

    # ── Guardar ────────────────────────────────────────────────────────
    out = config.output_path(mes)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    log.info(f"Planilla guardada: {out}")


# ── Punto de entrada ──────────────────────────────────────────────────────

def main() -> None:
    config.OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(config.OUTPUTS_DIR / "run.log", encoding="utf-8"),
        ],
        force=True,
    )

    pattern = str(config.INPUTS_DIR / "lecturas" / "lecturas_planilla_*.xlsx")
    matches = sorted(glob.glob(pattern))
    if not matches:
        log.error("No se encontró ningún archivo lecturas_planilla_YYYY-MM.xlsx en inputs/lecturas/")
        sys.exit(1)

    # El mes lo DECLARA 1_lecturas (columna MES_ANO de la plantilla del operario)
    # en shared/ciclo_activo.json. Deducirlo del nombre del último archivo
    # alfabético era la heurística que dejaba trabajar con el mes equivocado
    # cuando conviven varios ciclos en la carpeta.
    mes = ciclo_activo.activo(default=None,
                              path=config.BASE_DIR.parent / "shared" / "ciclo_activo.json")
    disponibles = [Path(m).stem.replace("lecturas_planilla_", "") for m in matches]
    if mes is None:
        mes = disponibles[-1]
        log.warning(f"Sin ciclo activo declarado (shared/ciclo_activo.json) — "
                    f"usando el último archivo de lecturas: {mes}. Correr 1_lecturas para declararlo.")
    elif mes not in disponibles:
        log.error(f"El ciclo activo es {mes} pero no hay lecturas_planilla_{mes}.xlsx "
                  f"en inputs/lecturas/ (hay: {', '.join(disponibles)})")
        sys.exit(1)
    log.info(f"Mes del ciclo: {mes}")

    df = build_planilla(mes)
    write_excel(df, mes)
    publicar_a_shared(mes)


def publicar_a_shared(mes: str) -> None:
    """
    Copia planilla_{mes}.xlsx a shared/planilla_mes/ — donde 4_pagos espera leerla.
    Sobreescribe el archivo del mismo mes; no toca archivos de otros meses.
    """
    src = config.output_path(mes)
    dest = config.shared_planilla_path(mes)
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dest)
    log = logging.getLogger(__name__)
    log.info(f"Publicada a shared: {dest}")


if __name__ == "__main__":
    main()
