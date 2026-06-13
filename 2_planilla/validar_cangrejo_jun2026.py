# ╔══════════════════════════════════════════════════════════════╗
# ║  SCRIPT DESECHABLE - SOLO CICLO JUNIO 2026                   ║
# ║                                                              ║
# ║  Valida que planilla_2026-06.xlsx (generada por cangrejo)    ║
# ║  cuadre con DATA_boletas.xlsx (verdad de oro manual).        ║
# ║                                                              ║
# ║  Compara TOTAL_A_PAGAR (planilla) vs "Importe a pagar"       ║
# ║  (DATA) usando clave MZ-LT. NO modifica nada — solo reporta. ║
# ║                                                              ║
# ║  Al cerrar el ciclo de junio -> mover a backup/              ║
# ╚══════════════════════════════════════════════════════════════╝

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).parent

PLAN_PATH   = ROOT / "outputs" / "planilla_2026-06.xlsx"
DATA_PATH   = ROOT.parent / "3_boletas" / "inputs" / "DATA_boletas.xlsx"
REPORT_PATH = ROOT / "outputs" / "reporte_validacion_cangrejo.xlsx"

TOL = 0.01  # tolerancia en soles para diferencias de float


def clave(mz, lt) -> str:
    return f"{str(mz).strip().upper()}-{str(lt).strip().upper()}"


def main():
    print("\n" + "=" * 55)
    print("  validar_cangrejo_jun2026 — planilla vs DATA_boletas")
    print("=" * 55)

    if not PLAN_PATH.exists():
        raise FileNotFoundError(f"No se encontró: {PLAN_PATH}")
    if not DATA_PATH.exists():
        raise FileNotFoundError(f"No se encontró: {DATA_PATH}")

    # planilla: write_excel pone título en fila 1, headers en fila 2
    df_plan = pd.read_excel(PLAN_PATH, header=1)
    df_data = pd.read_excel(DATA_PATH, sheet_name="Data")

    # Validación mínima de columnas
    for col in ("MZ", "LT", "NOMBRE", "TOTAL_A_PAGAR"):
        if col not in df_plan.columns:
            raise ValueError(f"Falta columna '{col}' en planilla")
    for col in ("MZ", "LT", "NOMBRES", "Importe a pagar"):
        if col not in df_data.columns:
            raise ValueError(f"Falta columna '{col}' en DATA_boletas")

    # Indexar por clave
    plan = {}
    for _, r in df_plan.iterrows():
        k = clave(r["MZ"], r["LT"])
        plan[k] = {
            "MZ":     str(r["MZ"]).strip(),
            "LT":     str(r["LT"]).strip(),
            "NOMBRE": r["NOMBRE"],
            "total":  pd.to_numeric(r["TOTAL_A_PAGAR"], errors="coerce"),
        }

    data = {}
    for _, r in df_data.iterrows():
        k = clave(r["MZ"], r["LT"])
        data[k] = {
            "MZ":     str(r["MZ"]).strip(),
            "LT":     str(r["LT"]).strip(),
            "NOMBRE": r["NOMBRES"],
            "total":  pd.to_numeric(r["Importe a pagar"], errors="coerce"),
        }

    # Clasificar
    discrepancias = []  # filas para el Excel
    n_ok = 0
    n_diff = 0
    n_falta_plan = 0
    n_falta_data = 0
    suma_diff = 0.0

    todas_las_claves = sorted(set(plan) | set(data))
    for k in todas_las_claves:
        en_plan = k in plan
        en_data = k in data

        if en_plan and en_data:
            tp = plan[k]["total"]
            td = data[k]["total"]
            if pd.isna(tp) or pd.isna(td) or abs(tp - td) > TOL:
                diff = (tp if not pd.isna(tp) else 0) - (td if not pd.isna(td) else 0)
                suma_diff += abs(diff)
                n_diff += 1
                discrepancias.append({
                    "tipo":                   "monto_diff",
                    "clave":                  k,
                    "MZ":                     plan[k]["MZ"],
                    "LT":                     plan[k]["LT"],
                    "NOMBRE":                 plan[k]["NOMBRE"],
                    "TOTAL_A_PAGAR_planilla": tp,
                    "IMPORTE_A_PAGAR_data":   td,
                    "diferencia":             diff,
                })
            else:
                n_ok += 1
        elif en_data and not en_plan:
            n_falta_plan += 1
            discrepancias.append({
                "tipo":                   "falta_en_planilla",
                "clave":                  k,
                "MZ":                     data[k]["MZ"],
                "LT":                     data[k]["LT"],
                "NOMBRE":                 data[k]["NOMBRE"],
                "TOTAL_A_PAGAR_planilla": None,
                "IMPORTE_A_PAGAR_data":   data[k]["total"],
                "diferencia":             None,
            })
        else:  # en_plan and not en_data
            n_falta_data += 1
            discrepancias.append({
                "tipo":                   "falta_en_data",
                "clave":                  k,
                "MZ":                     plan[k]["MZ"],
                "LT":                     plan[k]["LT"],
                "NOMBRE":                 plan[k]["NOMBRE"],
                "TOTAL_A_PAGAR_planilla": plan[k]["total"],
                "IMPORTE_A_PAGAR_data":   None,
                "diferencia":             None,
            })

    # ── Reporte consola ────────────────────────────────────────
    print(f"\n  Total filas planilla : {len(plan)}")
    print(f"  Total filas DATA     : {len(data)}")
    print(f"    ok               : {n_ok}")
    print(f"    monto_diff       : {n_diff}  (suma |diferencia|: S/. {suma_diff:,.2f})")
    print(f"    falta_en_planilla: {n_falta_plan}")
    print(f"    falta_en_data    : {n_falta_data}")

    # ── Reporte Excel ──────────────────────────────────────────
    if not discrepancias:
        print("\n  OK Sin discrepancias -- no se genera reporte Excel")
        print("=" * 55 + "\n")
        return

    # Orden: tipo asc, luego |diferencia| desc
    orden_tipo = {"monto_diff": 0, "falta_en_planilla": 1, "falta_en_data": 2}
    discrepancias.sort(key=lambda d: (
        orden_tipo[d["tipo"]],
        -abs(d["diferencia"]) if d["diferencia"] is not None else 0,
    ))

    wb = Workbook()
    ws = wb.active
    ws.title = "discrepancias"

    headers = [
        "tipo", "clave", "MZ", "LT", "NOMBRE",
        "TOTAL_A_PAGAR_planilla", "IMPORTE_A_PAGAR_data", "diferencia",
    ]
    header_fill = PatternFill("solid", fgColor="1E5C3A")
    header_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    fmt_monto = '"S/"\\ #,##0.00'
    for r, d in enumerate(discrepancias, start=2):
        ws.cell(row=r, column=1, value=d["tipo"])
        ws.cell(row=r, column=2, value=d["clave"])
        ws.cell(row=r, column=3, value=d["MZ"])
        ws.cell(row=r, column=4, value=d["LT"])
        ws.cell(row=r, column=5, value=d["NOMBRE"])
        for col_idx, key in ((6, "TOTAL_A_PAGAR_planilla"),
                             (7, "IMPORTE_A_PAGAR_data"),
                             (8, "diferencia")):
            v = d[key]
            cell = ws.cell(row=r, column=col_idx, value=v)
            if v is not None:
                cell.number_format = fmt_monto

    widths = [18, 10, 6, 6, 32, 22, 22, 14]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w
    ws.freeze_panes = "A2"

    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(REPORT_PATH)
    print(f"\n  Reporte: {REPORT_PATH}")
    print("=" * 55 + "\n")


if __name__ == "__main__":
    main()
