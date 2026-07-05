"""
validar_arrastres.py — reconciliación post-build de 2_planilla.

Corre DESPUÉS de main.py. Vuelve a leer las 3 fuentes en vivo (arrastre_consolidado
de 5_cobranza + seguimiento_repo) y compara, predio por predio, contra lo que quedó
escrito en planilla_{mes}.xlsx. No corrige nada — solo reporta discrepancias.
Contrato de formato: docs/formato_validacion_de_arrastres.html
"""
import glob
import logging
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import config
from main import _norm_mz, _norm_lt, _mes_anterior

sys.path.insert(0, str(config.BASE_DIR.parent / "shared"))
import seguimiento_repo as repo  # noqa: E402

log = logging.getLogger(__name__)

TOLERANCIA = 0.005

# (columna en planilla, etiqueta de fuente, tipo de fuente, columna/concepto en fuente, tipo de dato)
CONCEPTOS = [
    ("NOMBRE",            "lecturas_planilla.NOMBRE",              "lecturas",    "NOMBRE",           "txt"),
    ("MARC_ANT",          "lecturas_planilla.MARC_ANT",            "lecturas",    "MARC_ANT",         "num"),
    ("MARC_ACT",          "lecturas_planilla.MARC_ACT",            "lecturas",    "MARC_ACT",         "num"),
    ("M3",                "lecturas_planilla.M3",                  "lecturas",    "M3",               "num"),
    ("MES_ANTERIOR",      "arrastre_consolidado.DEUDA_AGUA",       "consolidado", "DEUDA_AGUA",       "num"),
    ("CORTE_RECONEXION",  "arrastre_consolidado.CORTE_RECONEXION", "consolidado", "CORTE_RECONEXION", "num"),
    ("MULTA",             "seguimiento_repo",                      "repo",        "MULTA",            "num"),
    ("CONVENIO",          "seguimiento_repo",                      "repo",        "CONVENIO",         "num"),
    ("ACUERDOS_ASAMBLEA", "seguimiento_repo",                      "repo",        "ACUERDOS",         "num"),
]

# Columnas cuya discrepancia se muestra sin formato moneda (son lecturas, no soles)
_CONCEPTOS_ENTEROS = {"MARC_ANT", "MARC_ACT", "M3"}

# Origen verdadero de las lecturas — el output de 1_lecturas, no la copia
# manual en 2_planilla/inputs/ (así se detecta una copia mal hecha).
def _lecturas_origen_path(mes: str) -> Path:
    return config.BASE_DIR.parent / "1_lecturas" / "outputs" / f"lecturas_planilla_{mes}.xlsx"


def _detectar_mes() -> str:
    pattern = str(config.OUTPUTS_DIR / "planilla_*.xlsx")
    matches = sorted(glob.glob(pattern))
    if not matches:
        log.error("No se encontró ninguna planilla_YYYY-MM.xlsx en outputs/ — correr main.py primero")
        sys.exit(1)
    return Path(matches[-1]).stem.replace("planilla_", "")


def _cargar_planilla(mes: str) -> pd.DataFrame:
    df = pd.read_excel(config.output_path(mes), header=1, dtype=str)
    df["_mz"] = df["MZ"].map(_norm_mz)
    df["_lt"] = df["LT"].map(_norm_lt)
    for col, _, _, _, tipo in CONCEPTOS:
        if tipo == "num":
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    return df


def _cargar_lecturas_origen(mes: str) -> pd.DataFrame | None:
    path = _lecturas_origen_path(mes)
    if not path.exists():
        log.warning(f"lecturas_planilla {mes} (1_lecturas/outputs): no encontrado — "
                    f"NOMBRE/MARC_ANT/MARC_ACT/M3 y cobertura se saltan")
        return None
    df = pd.read_excel(path, header=1, dtype=str)
    df["_mz"] = df["MZ"].map(_norm_mz)
    df["_lt"] = df["LT"].map(_norm_lt)
    return df


def _cargar_consolidado(mes_ant: str) -> pd.DataFrame | None:
    path = config.consolidado_path(mes_ant)
    if not path.exists():
        log.warning(f"arrastre_consolidado {mes_ant}: no encontrado — MES_ANTERIOR/CORTE_RECONEXION "
                    f"se saltan (mes de génesis o ciclo anterior sin cerrar)")
        return None
    df = pd.read_excel(path, header=1, dtype=str)
    df["_mz"] = df["MZ"].map(_norm_mz)
    df["_lt"] = df["LT"].map(_norm_lt)
    for c in ("DEUDA_AGUA", "CORTE_RECONEXION"):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    return df


def _fuente_dict(src_kind: str, src_key: str, df_cons: pd.DataFrame | None,
                  df_lect: pd.DataFrame | None, mes_ant: str) -> dict | None:
    if src_kind == "lecturas":
        if df_lect is None:
            return None
        return {(r["_mz"], r["_lt"]): r[src_key] for _, r in df_lect.iterrows()}
    if src_kind == "consolidado":
        if df_cons is None:
            return None
        return {(r["_mz"], r["_lt"]): r[src_key] for _, r in df_cons.iterrows()}
    # seguimiento_repo
    return repo.get_saldos_bulk(src_key, mes_ant)


def comparar(mes: str) -> tuple[list[dict], list[dict]]:
    """Devuelve (resumen, discrepancias) — ver formato_validacion_de_arrastres.html."""
    mes_ant = _mes_anterior(mes)
    df_planilla = _cargar_planilla(mes)
    df_cons = _cargar_consolidado(mes_ant)
    df_lect = _cargar_lecturas_origen(mes)

    resumen = []
    discrepancias = []

    for dest_col, fuente_label, src_kind, src_key, tipo in CONCEPTOS:
        fuente = _fuente_dict(src_kind, src_key, df_cons, df_lect, mes_ant)
        if fuente is None:
            resumen.append({
                "CONCEPTO": dest_col, "FUENTE": fuente_label,
                "TOTAL_PREDIOS": 0, "COINCIDEN": 0, "DISCREPANCIAS": 0,
                "NOTA": "fuente no disponible",
            })
            continue

        total = 0
        coinciden = 0
        for _, row in df_planilla.iterrows():
            key = (row["_mz"], row["_lt"])
            # lecturas: solo claves presentes en la fuente — la cobertura
            # (abajo) reporta las que faltan, sin duplicar ruido acá.
            if src_kind == "lecturas" and key not in fuente:
                continue
            total += 1
            if tipo == "txt":
                v_fuente = str(fuente.get(key, "") or "").strip()
                v_planilla = str(row[dest_col] or "").strip()
                if v_fuente == v_planilla:
                    coinciden += 1
                else:
                    discrepancias.append({
                        "MZ": row["MZ"], "LT": row["LT"], "CONCEPTO": dest_col,
                        "VALOR_FUENTE": v_fuente, "VALOR_PLANILLA": v_planilla,
                        "DIFERENCIA": "TEXTO_DISTINTO",
                    })
                continue
            valor_fuente = pd.to_numeric(fuente.get(key, 0.0), errors="coerce")
            valor_fuente = 0.0 if pd.isna(valor_fuente) else float(valor_fuente)
            valor_planilla = float(row[dest_col])
            if abs(valor_fuente - valor_planilla) <= TOLERANCIA:
                coinciden += 1
            else:
                discrepancias.append({
                    "MZ": row["MZ"], "LT": row["LT"], "CONCEPTO": dest_col,
                    "VALOR_FUENTE": round(valor_fuente, 2),
                    "VALOR_PLANILLA": round(valor_planilla, 2),
                    "DIFERENCIA": round(valor_planilla - valor_fuente, 2),
                })

        resumen.append({
            "CONCEPTO": dest_col, "FUENTE": fuente_label,
            "TOTAL_PREDIOS": total, "COINCIDEN": coinciden,
            "DISCREPANCIAS": total - coinciden, "NOTA": "",
        })

    # ── Cobertura de predios: lecturas origen ↔ planilla ────────────────
    if df_lect is not None:
        keys_lect = set(zip(df_lect["_mz"], df_lect["_lt"]))
        keys_pla = set(zip(df_planilla["_mz"], df_planilla["_lt"]))
        for k in sorted(keys_lect - keys_pla):
            discrepancias.append({"MZ": k[0], "LT": k[1], "CONCEPTO": "(MZ, LT)",
                                  "VALOR_FUENTE": "presente", "VALOR_PLANILLA": "ausente",
                                  "DIFERENCIA": "FALTA_EN_PLANILLA"})
        for k in sorted(keys_pla - keys_lect):
            discrepancias.append({"MZ": k[0], "LT": k[1], "CONCEPTO": "(MZ, LT)",
                                  "VALOR_FUENTE": "ausente", "VALOR_PLANILLA": "presente",
                                  "DIFERENCIA": "SOBRA_EN_PLANILLA"})
        n_cob = len(keys_lect ^ keys_pla)
        resumen.append({
            "CONCEPTO": "(MZ, LT)", "FUENTE": "lecturas_planilla (cobertura)",
            "TOTAL_PREDIOS": len(keys_lect | keys_pla),
            "COINCIDEN": len(keys_lect & keys_pla),
            "DISCREPANCIAS": n_cob, "NOTA": "",
        })

    return resumen, discrepancias


# ── Escritura Excel ──────────────────────────────────────────────────────

_HEAD_BG = "1E5C3A"
_OK_BG   = "E9F7EF"
_BAD_BG  = "FEE2E2"
_BAD_FG  = "991B1B"


def _hoja_resumen(wb: Workbook, resumen: list[dict]) -> None:
    ws = wb.active
    ws.title = "Resumen"
    cols = ["CONCEPTO", "FUENTE", "TOTAL_PREDIOS", "COINCIDEN", "DISCREPANCIAS", "NOTA"]
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.fill = PatternFill("solid", fgColor="FF" + _HEAD_BG)
        cell.font = Font(color="FFFFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
    for r, fila in enumerate(resumen, start=2):
        for i, c in enumerate(cols, start=1):
            val = fila.get(c, "")
            cell = ws.cell(row=r, column=i, value=val)
            if c == "DISCREPANCIAS" and isinstance(val, int) and val > 0:
                cell.fill = PatternFill("solid", fgColor="FF" + _BAD_BG)
                cell.font = Font(color="FF" + _BAD_FG, bold=True)
            elif c == "COINCIDEN":
                cell.fill = PatternFill("solid", fgColor="FF" + _OK_BG)
    widths = [20, 34, 14, 12, 14, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"


def _hoja_discrepancias(wb: Workbook, discrepancias: list[dict]) -> None:
    ws = wb.create_sheet("Discrepancias")
    cols = ["MZ", "LT", "CONCEPTO", "VALOR_FUENTE", "VALOR_PLANILLA", "DIFERENCIA"]
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.fill = PatternFill("solid", fgColor="FF" + _HEAD_BG)
        cell.font = Font(color="FFFFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
    for r, fila in enumerate(discrepancias, start=2):
        for i, c in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=i, value=fila[c])
            if c == "DIFERENCIA":
                cell.fill = PatternFill("solid", fgColor="FF" + _BAD_BG)
                cell.font = Font(color="FF" + _BAD_FG, bold=True)
            if c in ("VALOR_FUENTE", "VALOR_PLANILLA", "DIFERENCIA") and isinstance(fila[c], (int, float)):
                cell.number_format = "0" if fila["CONCEPTO"] in _CONCEPTOS_ENTEROS else '"S/" #,##0.00'
    widths = [6, 7, 20, 16, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"


def escribir_reporte(mes: str, resumen: list[dict], discrepancias: list[dict]) -> Path:
    wb = Workbook()
    _hoja_resumen(wb, resumen)
    _hoja_discrepancias(wb, discrepancias)
    out = config.OUTPUTS_DIR / f"validacion_de_arrastres_{mes}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)s  %(message)s")
    mes = _detectar_mes()
    log.info(f"Validando arrastres de planilla_{mes}.xlsx")
    resumen, discrepancias = comparar(mes)
    out = escribir_reporte(mes, resumen, discrepancias)

    total_disc = sum(r["DISCREPANCIAS"] for r in resumen)
    log.info(f"Reporte: {out}")
    if total_disc:
        log.warning(f"{total_disc} discrepancia(s) encontrada(s) — revisar hoja Discrepancias")
    else:
        log.info("Sin discrepancias — todos los arrastres coinciden con su fuente")


if __name__ == "__main__":
    main()
