# ╔══════════════════════════════════════════════════════════════╗
# ║  SCRIPT DESECHABLE - SOLO CICLO JUNIO 2026                  ║
# ║                                                              ║
# ║  Tecnica del cangrejo: convierte DATA_boletas.xlsx           ║
# ║  (generado manualmente) en planilla_2026-06.xlsx usando el   ║
# ║  mismo write_excel() de 2_planilla/main.py para que el       ║
# ║  archivo tenga formato identico al pipeline normal.          ║
# ║                                                              ║
# ║  Usar UNA SOLA VEZ para arrancar el ciclo digital de junio.  ║
# ║  Desde julio, 2_planilla/main.py genera la planilla          ║
# ║  directamente desde 1_lecturas. Este script ya no aplica.    ║
# ║                                                              ║
# ║  Al cerrar el ciclo de junio -> mover a backup/             ║
# ╚══════════════════════════════════════════════════════════════╝

import shutil
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))

import config                                  # noqa: E402
from main import write_excel                   # noqa: E402

DATA_PATH    = ROOT.parent / "3_boletas" / "inputs" / "DATA_boletas.xlsx"
SHARED_DEST  = ROOT.parent / "shared" / "planilla_mes" / "planilla_2026-06.xlsx"
MES_ANO      = "2026-06"

# Columnas de DATA_boletas -> nombres canonicos de planilla
MAPEO = {
    "MZ":                      "MZ",
    "LT":                      "LT",
    "NOMBRES":                 "NOMBRE",
    "Marcación anterior":      "MARC_ANT",
    "Marcacion altual":        "MARC_ACT",   # typo original en enriquecimiento/main.py
    "M3":                      "M3",
    "Total mes actual":        "MES_ACTUAL",
    "Mantenimiento":           "MANTENIMIENTO",
    "MES ANTERIOR":            "MES_ANTERIOR",
    "Corte y reconexion":      "CORTE_RECONEXION",
    "Convenio":                "CONVENIO",
    "Multa (faena + reunión)": "MULTA",
    "Cuota directa":           "ACUERDOS_ASAMBLEA",
    "Importe a pagar":         "TOTAL_A_PAGAR",   # valor literal — no formula
}

# Columnas numericas que deben ir como float (no string)
COLS_NUM = [
    "MARC_ANT", "MARC_ACT", "M3",
    "MES_ACTUAL", "MANTENIMIENTO", "MES_ANTERIOR", "CORTE_RECONEXION",
    "CONVENIO", "MULTA", "ACUERDOS_ASAMBLEA",
    "TOTAL_A_PAGAR",
]


def main():
    print("\n" + "=" * 55)
    print("  cangrejo_jun2026 - DATA_boletas -> planilla")
    print("=" * 55)

    if not DATA_PATH.exists():
        raise FileNotFoundError(f"No se encontro: {DATA_PATH}")

    df_raw = pd.read_excel(DATA_PATH, sheet_name="Data")
    print(f"\n  Leidos {len(df_raw)} registros de DATA_boletas.xlsx")

    faltantes = [c for c in MAPEO if c not in df_raw.columns]
    if faltantes:
        raise ValueError(f"Columnas faltantes en DATA_boletas: {faltantes}")

    # Seleccionar y renombrar
    df = df_raw[list(MAPEO.keys())].rename(columns=MAPEO).copy()

    # MES_ANO fijo del ciclo
    df["MES_ANO"] = MES_ANO

    # Tipar columnas numericas (DATA_boletas las trae como str/object)
    for col in COLS_NUM:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Descuentos en 0 - 4_pagos los actualizara con valores negativos
    df["BLANCO"]     = 0.0
    df["DEVOLUCION"] = 0.0

    # Columnas vacias - 4_pagos las llena
    df["MONTO_YAPE"]     = None
    df["MONTO_EFECTIVO"] = None
    df["ESTADO"]         = None
    df["FECHA_PAGO"]     = None

    # Reordenar segun contrato de planilla
    df = df[config.OUTPUT_COLS]

    # Escribir con formato oficial (colores, anchos, freeze panes)
    # write_excel sobreescribe TOTAL_A_PAGAR con formula =H+I+J+...
    # → la patcheamos despues con el valor literal de DATA_boletas
    write_excel(df, MES_ANO)
    out_oficial = config.output_path(MES_ANO)
    print(f"  Planilla oficial: {out_oficial}")

    # Patch: TOTAL_A_PAGAR como valor literal (no formula)
    # Razon: openpyxl con data_only=True devuelve None para formulas sin cache,
    # y el motor lo lee asi -> deuda=0 -> matching incorrecto.
    print(f"  Patcheando TOTAL_A_PAGAR con valor literal...")
    wb = load_workbook(out_oficial)
    ws = wb.active
    col_total = config.OUTPUT_COLS.index("TOTAL_A_PAGAR") + 1  # 1-indexed
    for r_offset, total_val in enumerate(df["TOTAL_A_PAGAR"].tolist(), start=3):
        ws.cell(row=r_offset, column=col_total).value = float(total_val)
    wb.save(out_oficial)
    print(f"  OK {len(df)} celdas TOTAL_A_PAGAR escritas como valor")

    # Copiar a shared/ donde 4_pagos lee la planilla del mes
    SHARED_DEST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(out_oficial, SHARED_DEST)
    print(f"  Copiada a:        {SHARED_DEST}")

    print(f"\n  {len(df)} usuarios listos para 4_pagos")
    print("  Siguiente paso: python 4_pagos/main.py")
    print("=" * 55 + "\n")


if __name__ == "__main__":
    main()
