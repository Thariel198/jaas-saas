"""6_corte/aplicar_penalidad.py v2 — Reconciliación bidireccional (Día 0)

Compara dos sets en cada corrida y modifica planilla_mes para que coincida:

  SET_DEBE  = (MZ, LT) con EJECUTAR_CORTE=SI en lista_corte.xlsx
  SET_TIENE = (MZ, LT) con cargo activo en audit (APLICADO − REVERTIDO) este ciclo

  DEBE − TIENE → NUEVOS    → +S/20 en planilla · escribe APLICADO en audit
  TIENE − DEBE → SOBRANTES → −S/20 en planilla · escribe REVERTIDO en audit
  DEBE ∩ TIENE → CORRECTOS → skip (idempotente)

Único writer del sistema sobre la columna CORTE_RECONEXION en shared/planilla_mes.
Antes de tocar nada hace backup de la planilla. El audit es el log de auditoría.

Retrocompatibilidad: filas de audit v1 sin columna ACCION se tratan como APLICADO.
La primera corrida v2 migra el archivo automáticamente (extiende sección Trazabilidad).

Re-correr siempre produce el estado correcto: aplica los que faltan, revierte
los sobrantes, skipea los correctos.

Uso:
    python aplicar_penalidad.py
"""
import logging
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
import config

SOURCE = "6_corte/aplicar_penalidad"

# ── VALORES de la columna ACCION (audit v2) ─────────────────────────────────
ACCION_APLICADO  = "APLICADO"
ACCION_REVERTIDO = "REVERTIDO"

# ── PALETA AUDIT (azul/verde/morado — registro de auditoría) ────────────────
GH_AUD_ID    = ("EBF5FB", "1A5276")
GH_AUD_APL   = ("E9F7EF", "1E5C3A")
GH_AUD_TRAZ  = ("F3E8FF", "5B21B6")
TD_AUD_ID    = "F4FAFF"
TD_AUD_APL   = "F4FBF7"
TD_AUD_TRAZ  = "FAF5FF"

# Colores condicionales para la celda ACCION
TD_AUD_APL_VAL = "D5F5E3"; TX_AUD_APL_VAL = "145A32"   # APLICADO
TD_AUD_REV_VAL = "FADBD8"; TX_AUD_REV_VAL = "7B241C"   # REVERTIDO

_AUDIT_GRUPOS = [
    (1, 4, "Identidad",         *GH_AUD_ID),
    (5, 6, "Aplicado",          *GH_AUD_APL),
    (7, 9, "Trazabilidad",      *GH_AUD_TRAZ),
]
_AUDIT_COLS = [
    (1, "MZ",                 *GH_AUD_ID,    6),
    (2, "LT",                 *GH_AUD_ID,    7),
    (3, "NOMBRE",             *GH_AUD_ID,   28),
    (4, "MES_ANO",            *GH_AUD_ID,   10),
    (5, "PENALIDAD_APLICADA", *GH_AUD_APL,  18),
    (6, "CORTE_RECON_DESPUES",*GH_AUD_APL,  20),
    (7, "FECHA_APLICACION",   *GH_AUD_TRAZ, 18),
    (8, "SOURCE",              *GH_AUD_TRAZ, 28),
    (9, "ACCION",              *GH_AUD_TRAZ, 12),
]


# ── HELPERS DE ESTILO ────────────────────────────────────────────────────────
def _borde():
    b = Side(style="thin", color="CCCCCC")
    return Border(left=b, right=b, top=b, bottom=b)

def _c(ws, row, col, value=None, bg=None, txt="333333",
       bold=False, align="left", mono=False, size=9, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Consolas" if mono else "Arial",
                       size=size, bold=bold, color=txt)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _borde()
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    if fmt:
        c.number_format = fmt
    return c

def _gh(ws, row, cs, ce, texto, bg, txt):
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    c = ws.cell(row=row, column=cs, value=texto)
    c.font      = Font(name="Arial", size=8, bold=True, color=txt)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", start_color=bg)
    c.border    = _borde()

def _ch(ws, row, col, texto, bg, txt):
    _c(ws, row, col, texto, bg=bg, txt=txt, bold=True, align="center")

def _w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── PARSING ──────────────────────────────────────────────────────────────────
def _norm_mz(val) -> str:
    if val is None:
        return ""
    s = str(val).strip().upper()
    return "" if not s or s in ("NAN", "NONE") else s

def _norm_lt(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN"):
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.upper().replace(" ", "")

def _float(val) -> float:
    if val is None:
        return 0.0
    try:
        f = float(str(val).replace(",", ".").strip())
        return 0.0 if f != f else f
    except (ValueError, TypeError):
        return 0.0


# ── VALIDACIÓN DE INPUTS ─────────────────────────────────────────────────────
def _localizar_planilla_mes(log: logging.Logger) -> Path:
    """Busca planilla_*.xlsx en shared/planilla_mes/. Debe haber exactamente una."""
    if not config.PLANILLA_MES_DIR.exists():
        raise FileNotFoundError(
            f"Falta directorio: {config.PLANILLA_MES_DIR}"
        )
    candidatos = sorted(config.PLANILLA_MES_DIR.glob("planilla_*.xlsx"))
    if not candidatos:
        raise FileNotFoundError(
            f"No hay planilla_*.xlsx en {config.PLANILLA_MES_DIR}\n"
            f"  -> 2_planilla debe haber generado la planilla del mes"
        )
    if len(candidatos) > 1:
        log.warning(f"Multiples planillas en planilla_mes/ -> usando {candidatos[-1].name}")
    return candidatos[-1]

def _leer_lista_corte(log: logging.Logger) -> list[dict]:
    if not config.LISTA_CORTE_PATH.exists():
        raise FileNotFoundError(
            f"Falta: {config.LISTA_CORTE_PATH}\n"
            f"  -> Correr generar_lista.py primero"
        )
    df = pd.read_excel(config.LISTA_CORTE_PATH, header=1)
    df.columns = [str(c).strip().upper() for c in df.columns]
    requeridas = {"MZ", "LT", "NOMBRE"}
    faltantes = requeridas - set(df.columns)
    if faltantes:
        raise ValueError(f"lista_corte.xlsx · faltan columnas {sorted(faltantes)}")

    total = len(df)
    if "EJECUTAR_CORTE" in df.columns:
        bloqueados = (df["EJECUTAR_CORTE"].astype(str).str.strip().str.upper() != "SI").sum()
        df = df[df["EJECUTAR_CORTE"].astype(str).str.strip().str.upper() == "SI"]
        if bloqueados:
            log.info(f"lista_corte.xlsx · {total} total · {bloqueados} omitidos (EJECUTAR_CORTE=NO)")
    else:
        log.warning("lista_corte.xlsx · columna EJECUTAR_CORTE no encontrada · procesando todos")

    filas = []
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        filas.append({
            "mz": mz,
            "lt": lt,
            "nombre": str(f.get("NOMBRE", "")).strip(),
        })
    return filas


# ── AUDIT LOG — net APLICADO − REVERTIDO ─────────────────────────────────────
def _net_aplicados(mes_ano: str, log: logging.Logger) -> set[tuple[str, str]]:
    """Set de (MZ, LT) con cargo activo en MES_ANO = #APLICADO − #REVERTIDO > 0.

    Filas sin columna ACCION (audit v1) se cuentan como APLICADO (retrocompat).
    """
    p = config.AUDIT_PENALIDAD_PATH
    if not p.exists():
        return set()
    try:
        df = pd.read_excel(p, header=1, dtype=str).fillna("")
    except Exception:
        return set()
    if df.empty:
        return set()
    df.columns = [str(c).strip().upper() for c in df.columns]
    tiene_accion = "ACCION" in df.columns
    if not tiene_accion:
        log.info("audit_penalidad sin columna ACCION (v1) — filas tratadas como APLICADO")

    net: dict[tuple[str, str], int] = {}
    for _, f in df.iterrows():
        if str(f.get("MES_ANO", "")).strip() != mes_ano:
            continue
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        accion = str(f.get("ACCION", "")).strip().upper() if tiene_accion else ""
        if not accion:
            accion = ACCION_APLICADO
        key = (mz, lt)
        if accion == ACCION_APLICADO:
            net[key] = net.get(key, 0) + 1
        elif accion == ACCION_REVERTIDO:
            net[key] = net.get(key, 0) - 1

    return {k for k, v in net.items() if v > 0}


# ── RECONCILIACIÓN — SET_DEBE vs SET_TIENE ───────────────────────────────────
def _reconciliar(
    lista: list[dict],
    mes_ano: str,
    log: logging.Logger,
) -> tuple[list[dict], list[tuple[str, str]]]:
    """Devuelve (nuevos_a_aplicar, sobrantes_a_revertir) comparando DEBE vs TIENE.

    nuevos    = items de `lista` cuya (mz, lt) no está en el audit (DEBE − TIENE)
    sobrantes = (mz, lt) en el audit que ya no están en `lista` (TIENE − DEBE)
    """
    set_debe  = {(item["mz"], item["lt"]) for item in lista}
    set_tiene = _net_aplicados(mes_ano, log)

    nuevos_keys    = set_debe - set_tiene
    sobrantes_keys = set_tiene - set_debe
    correctos      = set_debe & set_tiene

    item_de = {(item["mz"], item["lt"]): item for item in lista}
    nuevos  = [item_de[k] for k in sorted(nuevos_keys) if k in item_de]

    log.info(f"Reconciliacion · DEBE={len(set_debe)} · TIENE={len(set_tiene)} · "
             f"NUEVOS={len(nuevos)} · SOBRANTES={len(sobrantes_keys)} · "
             f"CORRECTOS={len(correctos)}")
    return nuevos, sorted(sobrantes_keys)


def _migrar_audit_v1_a_v2(ws, log: logging.Logger) -> bool:
    """Si el audit no tiene columna ACCION, la agrega y marca existentes como APLICADO.

    Retorna True si migró, False si ya estaba en v2.
    """
    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    headers_upper = [str(h).strip().upper() if h else "" for h in headers]
    if "ACCION" in headers_upper:
        return False

    log.info("audit_penalidad: migrando v1->v2 (agregando columna ACCION)")
    col_accion = 9

    # 1) Extender el merge "Trazabilidad" de 7-8 a 7-9
    for merged_range in list(ws.merged_cells.ranges):
        if (merged_range.min_row == 1 and merged_range.max_row == 1
                and merged_range.min_col == 7 and merged_range.max_col == 8):
            ws.unmerge_cells(str(merged_range))
            break
    _gh(ws, 1, 7, 9, "Trazabilidad", GH_AUD_TRAZ[0], GH_AUD_TRAZ[1])

    # 2) Header de columna ACCION en fila 2
    _ch(ws, 2, col_accion, "ACCION", GH_AUD_TRAZ[0], GH_AUD_TRAZ[1])
    _w(ws, col_accion, 12)

    # 3) Marcar cada fila existente con ACCION=APLICADO
    migradas = 0
    for r in range(3, ws.max_row + 1):
        if ws.cell(row=r, column=1).value:   # fila con datos
            _c(ws, r, col_accion, ACCION_APLICADO,
               TD_AUD_APL_VAL, TX_AUD_APL_VAL, bold=True, align="center")
            migradas += 1
    log.info(f"audit_penalidad: {migradas} filas existentes marcadas como APLICADO")
    return True


def _audit_append(filas_nuevas: list[dict], log: logging.Logger) -> None:
    """Append a audit_penalidad.xlsx — crea con headers o migra v1→v2 si hace falta."""
    p = config.AUDIT_PENALIDAD_PATH
    config.OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

    if p.exists():
        wb = load_workbook(p)
        ws = wb.active
        _migrar_audit_v1_a_v2(ws, log)
        next_row = max(ws.max_row + 1, 3)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "audit_penalidad"
        ws.freeze_panes = "A3"
        for cs, ce, texto, bg, txt in _AUDIT_GRUPOS:
            _gh(ws, 1, cs, ce, texto, bg, txt)
        for col, nombre, bg, txt, ancho in _AUDIT_COLS:
            _ch(ws, 2, col, nombre, bg, txt)
            _w(ws, col, ancho)
        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 22
        next_row = 3

    MONEY = '"S/ "#,##0.00'
    for ri_offset, fila in enumerate(filas_nuevas):
        r = next_row + ri_offset
        _c(ws, r, 1, fila["mz"],     TD_AUD_ID, "1A5276", mono=True, align="center")
        _c(ws, r, 2, fila["lt"],     TD_AUD_ID, "1A5276", mono=True, align="center")
        _c(ws, r, 3, fila["nombre"], TD_AUD_ID, "333333", align="left")
        _c(ws, r, 4, fila["mes_ano"],TD_AUD_ID, "1A5276", mono=True, align="center")
        _c(ws, r, 5, fila["penalidad"], TD_AUD_APL, "1E5C3A",
           mono=True, align="right", bold=True, fmt=MONEY)
        _c(ws, r, 6, fila["corte_recon_despues"], TD_AUD_APL, "1E5C3A",
           mono=True, align="right", fmt=MONEY)
        _c(ws, r, 7, fila["fecha"], TD_AUD_TRAZ, "5B21B6", mono=True, align="center")
        _c(ws, r, 8, fila["source"], TD_AUD_TRAZ, "5B21B6", mono=True, align="left")
        # ACCION — color condicional según el valor
        if fila["accion"] == ACCION_APLICADO:
            _c(ws, r, 9, ACCION_APLICADO,
               TD_AUD_APL_VAL, TX_AUD_APL_VAL, bold=True, align="center")
        else:
            _c(ws, r, 9, ACCION_REVERTIDO,
               TD_AUD_REV_VAL, TX_AUD_REV_VAL, bold=True, align="center")
        ws.row_dimensions[r].height = 17

    wb.save(p)


# ── BACKUP ───────────────────────────────────────────────────────────────────
def _backup_planilla(plan_path: Path) -> Path:
    backup_dir = config.BACKUP_DIR / "planilla_mes"
    backup_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    dest = backup_dir / f"{plan_path.stem}_{ts}.xlsx"
    shutil.copy2(plan_path, dest)
    return dest


# ── ESCRITURA EN PLANILLA — bidireccional (aplicar + revertir) ───────────────
def _modificar_planilla(
    plan_path: Path,
    nuevos: list[dict],
    sobrantes: list[tuple[str, str]],
    log: logging.Logger,
) -> tuple[list[dict], list[dict], list[str]]:
    """Aplica +PENALIDAD a `nuevos` y −PENALIDAD a `sobrantes` en una sola apertura.

    Devuelve (filas_aplicadas, filas_revertidas, no_encontrados).
    El audit solo recibe lo que efectivamente se modificó en la planilla.
    """
    wb = load_workbook(plan_path)
    ws = wb.active

    # Headers en fila 2 (fila 1 = grupos). Buscamos columna por nombre — robusto
    # frente a reordenamientos del layout de 2_planilla.
    header_row = 2
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=col).value
        if v:
            headers[str(v).strip().upper()] = col

    for req in ("MZ", "LT", "CORTE_RECONEXION", "MES_ANO"):
        if req not in headers:
            raise ValueError(f"planilla_mes · falta columna {req} en fila {header_row}")

    col_mz   = headers["MZ"]
    col_lt   = headers["LT"]
    col_cr   = headers["CORTE_RECONEXION"]
    col_mes  = headers["MES_ANO"]
    col_nom  = headers.get("NOMBRE", 0)

    # Index (mz, lt) → row para localizar rápido sin re-escanear N veces
    fila_de: dict[tuple[str, str], int] = {}
    mes_ano_planilla = ""
    for r in range(3, ws.max_row + 1):
        mz = _norm_mz(ws.cell(row=r, column=col_mz).value)
        lt = _norm_lt(ws.cell(row=r, column=col_lt).value)
        if not mz or not lt:
            continue
        fila_de[(mz, lt)] = r
        if not mes_ano_planilla:
            mes_ano_planilla = str(ws.cell(row=r, column=col_mes).value or "").strip()

    if not mes_ano_planilla:
        raise ValueError("planilla_mes · no se pudo detectar MES_ANO en la fila 3")

    aplicados: list[dict]  = []
    revertidos: list[dict] = []
    no_encontrados: list[str] = []
    ahora = datetime.now().strftime("%d/%m/%Y %H:%M")

    def _nombre_de(r: int, fallback: str = "") -> str:
        if col_nom:
            n = str(ws.cell(row=r, column=col_nom).value or "").strip()
            if n:
                return n
        return fallback

    # ── NUEVOS — sumar +PENALIDAD ──
    for p in nuevos:
        key = (p["mz"], p["lt"])
        r = fila_de.get(key)
        if r is None:
            no_encontrados.append(f"{p['mz']}-{p['lt']} (NUEVO)")
            continue
        actual = _float(ws.cell(row=r, column=col_cr).value)
        nuevo  = round(actual + config.PENALIDAD, 2)
        ws.cell(row=r, column=col_cr).value = nuevo
        aplicados.append({
            "mz":     p["mz"],
            "lt":     p["lt"],
            "nombre": _nombre_de(r, p["nombre"]),
            "mes_ano": mes_ano_planilla,
            "penalidad": config.PENALIDAD,
            "corte_recon_despues": nuevo,
            "fecha":  ahora,
            "source": SOURCE,
            "accion": ACCION_APLICADO,
        })
        log.info(f"  +S/{config.PENALIDAD:.0f} -> {p['mz']}-{p['lt']} "
                 f"(CORTE_RECON: {actual:.2f} -> {nuevo:.2f})")

    # ── SOBRANTES — restar PENALIDAD ──
    for mz, lt in sobrantes:
        key = (mz, lt)
        r = fila_de.get(key)
        if r is None:
            no_encontrados.append(f"{mz}-{lt} (SOBRANTE)")
            continue
        actual = _float(ws.cell(row=r, column=col_cr).value)
        nuevo  = round(actual - config.PENALIDAD, 2)
        ws.cell(row=r, column=col_cr).value = nuevo
        revertidos.append({
            "mz":     mz,
            "lt":     lt,
            "nombre": _nombre_de(r),
            "mes_ano": mes_ano_planilla,
            "penalidad": -config.PENALIDAD,   # negativo en el audit
            "corte_recon_despues": nuevo,
            "fecha":  ahora,
            "source": SOURCE,
            "accion": ACCION_REVERTIDO,
        })
        log.info(f"  -S/{config.PENALIDAD:.0f} -> {mz}-{lt} "
                 f"(CORTE_RECON: {actual:.2f} -> {nuevo:.2f})")

    wb.save(plan_path)
    return aplicados, revertidos, no_encontrados


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main() -> None:
    config.OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(config.OUTPUTS_DIR / "run.log", encoding="utf-8"),
        ],
        force=True,
    )
    log = logging.getLogger(__name__)
    log.info("aplicar_penalidad.py · iniciando")

    print("=" * 60)
    print("  6_corte/aplicar_penalidad.py")
    print("=" * 60)

    print("\n[1/5] Validando inputs...")
    lista = _leer_lista_corte(log)
    log.info(f"lista_corte.xlsx · {len(lista)} usuarios elegibles (EJECUTAR_CORTE=SI)")
    plan_path = _localizar_planilla_mes(log)
    log.info(f"planilla_mes · {plan_path.name}")

    # Detectar MES_ANO de la planilla (necesario para reconciliar contra el audit)
    df_plan_head = pd.read_excel(plan_path, header=1, nrows=1)
    df_plan_head.columns = [str(c).strip().upper() for c in df_plan_head.columns]
    if "MES_ANO" not in df_plan_head.columns:
        raise ValueError("planilla_mes · falta columna MES_ANO")
    mes_ano_actual = str(df_plan_head.iloc[0]["MES_ANO"]).strip() if len(df_plan_head) else ""
    if not mes_ano_actual:
        raise ValueError("planilla_mes · MES_ANO vacío en la primera fila de datos")
    log.info(f"Ciclo detectado · {mes_ano_actual}")

    print(f"\n[2/5] Reconciliando SET_DEBE (lista_corte) vs SET_TIENE (audit) · ciclo {mes_ano_actual}...")
    nuevos, sobrantes = _reconciliar(lista, mes_ano_actual, log)

    if not nuevos and not sobrantes:
        print(f"\n  Nada que reconciliar · audit y lista_corte ya están alineados")
        print(f"  (idempotencia desde {config.AUDIT_PENALIDAD_PATH.name})")
        print("\n" + "=" * 60)
        print("  aplicar_penalidad.py completado · 0 escrituras")
        print("=" * 60 + "\n")
        return

    print(f"\n[3/5] Backup de planilla del mes ({plan_path.name})...")
    backup_path = _backup_planilla(plan_path)
    log.info(f"Backup -> {backup_path}")

    print(f"\n[4/5] Modificando CORTE_RECONEXION en planilla...")
    if nuevos:
        print(f"  NUEVOS:    +S/{config.PENALIDAD:.0f} a {len(nuevos)} usuarios")
    if sobrantes:
        print(f"  SOBRANTES: -S/{config.PENALIDAD:.0f} a {len(sobrantes)} usuarios")
    aplicados, revertidos, no_encontrados = _modificar_planilla(
        plan_path, nuevos, sobrantes, log)

    if no_encontrados:
        log.warning(f"No encontrados en planilla ({len(no_encontrados)}): "
                    f"{', '.join(no_encontrados[:10])}"
                    f"{'...' if len(no_encontrados) > 10 else ''}")

    print(f"\n[5/5] Append a audit log (APLICADO + REVERTIDO)...")
    filas_audit = aplicados + revertidos
    if filas_audit:
        _audit_append(filas_audit, log)
        log.info(f"audit_penalidad.xlsx · +{len(aplicados)} APLICADO · "
                 f"+{len(revertidos)} REVERTIDO")

    print("\n" + "=" * 60)
    print(f"  aplicar_penalidad.py completado")
    print(f"  -> {len(aplicados)} APLICADOS (+S/{config.PENALIDAD:.0f})")
    print(f"  -> {len(revertidos)} REVERTIDOS (-S/{config.PENALIDAD:.0f})")
    if no_encontrados:
        print(f"  -> {len(no_encontrados)} no encontrados en planilla (revisar)")
    print(f"  -> Backup: {backup_path}")
    print(f"  -> Audit:  {config.AUDIT_PENALIDAD_PATH}")
    if aplicados:
        print(f"\n  Esperar ventana de gracia (2 días) y luego: python seguimiento.py")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
