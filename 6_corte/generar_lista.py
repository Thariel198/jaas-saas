"""6_corte/generar_lista.py — Genera lista_corte.xlsx (Día 0)

Lee planilla_cobrado.xlsx de 5_cobranza, filtra usuarios con
SALDO > 0 AND MES_ANTERIOR >= 8.

Cruza por (MZ, LT) con:
  - 4b_reclamos/outputs/resolucion_reclamos_YYYY-MM.xlsx (decisiones)
  - 4b_reclamos/outputs/reclamos_YYYY-MM.xlsx           (lista cruda)

Los usuarios con reclamo EN_REVISION quedan en la lista pero con
EJECUTAR_CORTE = NO. Los FUNDADO/RECHAZADO no bloquean (ya resueltos).

Genera UN solo archivo `lista_corte.xlsx` (auditoría completa).
La columna EJECUTAR_CORTE permite a `aplicar_penalidad.py` filtrar
quiénes entran al período de gracia.

Idempotente: mismo input → mismo output, byte por byte aproximadamente
(salvo metadata de openpyxl).

Uso:
    python generar_lista.py
"""
import logging
import importlib.util
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
import config

# ── PALETA — matching docs/formato_lista_corte.html v2.1 ────────────────────
GH_ID    = ("EBF5FB", "1A5276")   # azul — ¿Quién es el deudor?
GH_DEUDA = ("E9F7EF", "1E5C3A")   # verde claro — ¿Cuánto debe?
GH_PEN   = ("1E8449", "FFFFFF")   # verde oscuro — Penalidad aplicada
GH_REC   = ("5B21B6", "FFFFFF")   # morado — ¿Ejecutar corte?
GH_PAGO  = ("D97706", "FFFFFF")   # ámbar — ¿Pagó efectivo este mes?

TD_ID    = "F4FAFF"
TD_DEUDA = "F4FBF7"
TD_PEN   = "D5F5E3"
TD_REC   = "F3E8FF"
TD_PAGO  = "FFFBEB"

# EJECUTAR_CORTE — color por valor
TD_SI = "D5F5E3"; TX_SI = "145A32"
TD_NO = "FADBD8"; TX_NO = "7B241C"

# ── LAYOUT ───────────────────────────────────────────────────────────────────
_LC_GRUPOS = [
    (1,  3,  "¿Quién es el deudor?",     *GH_ID),
    (4,  5,  "¿Cuánto debe?",             *GH_DEUDA),
    (6,  7,  "Penalidad aplicada",        *GH_PEN),
    (8,  10, "¿Ejecutar corte?",          *GH_REC),
    (11, 12, "¿Pagó efectivo este mes?",  *GH_PAGO),
]
_LC_COLS = [
    (1,  "MZ",                 *GH_ID,     6),
    (2,  "LT",                 *GH_ID,     7),
    (3,  "NOMBRE",             *GH_ID,    28),
    (4,  "SALDO",              *GH_DEUDA, 14),
    (5,  "MES_ANTERIOR",       *GH_DEUDA, 14),
    (6,  "PENALIDAD",          *GH_PEN,   14),
    (7,  "TOTAL_A_PAGAR",      *GH_PEN,   16),
    (8,  "ESTADO_RECLAMO",     *GH_REC,   16),
    (9,  "EJECUTAR_CORTE",     *GH_REC,   14),
    (10, "MOTIVO_NO_EJECUTAR", *GH_REC,   24),
    (11, "MESA",               *GH_PAGO,  12),
    (12, "COBRADOR",           *GH_PAGO,  20),
]


# ── HELPERS DE ESTILO ────────────────────────────────────────────────────────
def _borde():
    b = Side(style="thin", color="CCCCCC")
    return Border(left=b, right=b, top=b, bottom=b)

def _c(ws, row, col, value=None, bg=None, txt="333333",
       bold=False, align="left", mono=False, size=9, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Consolas" if mono else "Arial",
                       size=size, bold=bold, color=txt)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _borde()
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    if fmt:
        c.number_format = fmt
    return c

def _gh(ws, row, cs, ce, texto, bg, txt):
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    c = ws.cell(row=row, column=cs, value=texto)
    c.font      = Font(name="Arial", size=8, bold=True, color=txt)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", start_color=bg)
    c.border    = _borde()

def _ch(ws, row, col, texto, bg, txt):
    _c(ws, row, col, texto, bg=bg, txt=txt, bold=True, align="center")

def _w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── HELPERS DE PARSING ───────────────────────────────────────────────────────
def _norm_mz(val) -> str:
    if val is None:
        return ""
    s = str(val).strip().upper()
    return "" if not s or s in ("NAN", "NONE") else s

def _norm_lt(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN"):
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.upper().replace(" ", "")

def _float(val) -> float:
    if val is None:
        return 0.0
    try:
        f = float(str(val).replace(",", ".").strip())
        return 0.0 if f != f else f   # NaN guard
    except (ValueError, TypeError):
        return 0.0


# ── DETECCIÓN DE MES_ANO ─────────────────────────────────────────────────────
def _detectar_mes_ano(df: pd.DataFrame) -> str:
    """MES_ANO único del ciclo, leído de planilla_cobrado.xlsx."""
    if "MES_ANO" not in df.columns:
        raise ValueError("planilla_cobrado.xlsx · falta columna MES_ANO")
    valores = (df["MES_ANO"].dropna().astype(str).str.strip()
                                  .replace({"": None, "nan": None, "NaN": None})
                                  .dropna().unique().tolist())
    if not valores:
        raise ValueError("planilla_cobrado.xlsx · columna MES_ANO vacía")
    if len(valores) > 1:
        raise ValueError(f"planilla_cobrado.xlsx · MES_ANO inconsistente: {valores}")
    return valores[0]


# ── CANDADO — Día 0 ya penalizado ────────────────────────────────────────────
def _ciclo_ya_penalizado(mes_ano: str, log: logging.Logger) -> bool:
    """True si aplicar_penalidad.py ya cargó penalidad para este ciclo.

    lista_corte.xlsx es el snapshot Día 0 que seguimiento.py necesita intacto
    para comparar contra planilla_cobrado del Día 2 (resta algebraica del pago
    en ventana). Regenerarlo después de que Día 0 ya ocurrió destruye ese ancla
    sin aviso ni backup — pasó el 2026-07-26 (snapshot del 22/07 perdido,
    reconstruido a mano desde una copia vieja + audit + reclamos)."""
    p = config.AUDIT_PENALIDAD_PATH
    if not p.exists():
        return False
    df = pd.read_excel(p, header=1, dtype=str).fillna("")
    df.columns = [str(c).strip().upper() for c in df.columns]
    if "MES_ANO" not in df.columns:
        return False
    return bool((df["MES_ANO"].astype(str).str.strip() == mes_ano).any())


# ── VALIDACIÓN DE INPUTS ─────────────────────────────────────────────────────
def _validar_input(log: logging.Logger) -> tuple[pd.DataFrame, str]:
    if not config.PLANILLA_COBRADO_PATH.exists():
        raise FileNotFoundError(
            f"Falta: {config.PLANILLA_COBRADO_PATH}\n"
            f"  -> Correr 5_cobranza/main.py primero"
        )
    df = pd.read_excel(config.PLANILLA_COBRADO_PATH, header=1)
    df.columns = [str(c).strip().upper() for c in df.columns]

    requeridas = {"MZ", "LT", "NOMBRE", "SALDO", "MES_ANTERIOR", "MES_ANO"}
    faltantes = requeridas - set(df.columns)
    if faltantes:
        raise ValueError(
            f"planilla_cobrado.xlsx · columnas faltantes: {sorted(faltantes)}\n"
            f"  -> SALDO/MES_ANO deben venir expuestos por 5_cobranza. "
            f"Re-correr 5_cobranza/main.py para regenerar."
        )
    mes_ano = _detectar_mes_ano(df)
    log.info(f"planilla_cobrado.xlsx leida · {len(df)} filas · ciclo {mes_ano}")
    return df, mes_ano


# ── CARGA DE RECLAMOS (cruza por MZ+LT) ──────────────────────────────────────
def _cargar_reclamos_map(mes_ano: str, log: logging.Logger) -> dict:
    """
    Devuelve {(mz, lt) -> 'EN_REVISION'} para usuarios bloqueados.

    Reglas:
    - resolucion_reclamos (con decisión) manda sobre reclamos raw
    - FUNDADO o RECHAZADO en resolucion → NO bloquea (no entra al mapa)
    - EN_REVISION en cualquiera de los dos → bloquea
    """
    mapa: dict = {}

    # 1) resolucion_reclamos — fuente principal
    res_path = config.resolucion_reclamos_path(mes_ano)
    if res_path.exists():
        df_res = pd.read_excel(res_path, sheet_name="Correcciones",
                               header=1, dtype=str).fillna("")
        df_res.columns = [str(c).strip().upper() for c in df_res.columns]
        n_blocked = 0
        for _, f in df_res.iterrows():
            mz = _norm_mz(f.get("MZ"))
            lt = _norm_lt(f.get("LT"))
            if not mz or not lt:
                continue
            estado = str(f.get("ESTADO", "")).strip().upper()
            if estado in ("FUNDADO", "RECHAZADO"):
                continue                              # ya resuelto, no bloquea
            if estado in ("EN_REVISION", "PENDIENTE", ""):
                mapa[(mz, lt)] = "EN_REVISION"
                n_blocked += 1
        log.info(f"resolucion_reclamos_{mes_ano}.xlsx · {len(df_res)} filas · "
                 f"{n_blocked} bloquean corte")
    else:
        log.info(f"resolucion_reclamos_{mes_ano}.xlsx no existe — solo reclamos raw")

    # 2) reclamos raw — complementa los que aún no llegaron a resolucion
    rec_path = config.reclamos_path(mes_ano)
    if rec_path.exists():
        df_rec = pd.read_excel(rec_path, sheet_name="Reclamos",
                               header=1, dtype=str).fillna("")
        df_rec.columns = [str(c).strip().upper() for c in df_rec.columns]
        n_extra = 0
        for _, f in df_rec.iterrows():
            mz = _norm_mz(f.get("MZ"))
            lt = _norm_lt(f.get("LT"))
            if not mz or not lt:
                continue
            if (mz, lt) in mapa:
                continue                              # ya manejado por resolucion
            estado = str(f.get("ESTADO", "")).strip().upper()
            if estado in ("EN_REVISION", "PENDIENTE", ""):
                mapa[(mz, lt)] = "EN_REVISION"
                n_extra += 1
        log.info(f"reclamos_{mes_ano}.xlsx · {len(df_rec)} filas · "
                 f"+{n_extra} bloqueos adicionales")
    else:
        log.info(f"reclamos_{mes_ano}.xlsx no existe — solo resolucion")

    log.info(f"Total usuarios bloqueados por reclamo · {len(mapa)}")
    return mapa


# ── CARGA DE CORTADOS ACTIVOS ────────────────────────────────────────────────
def _cargar_cortados_activos(log: logging.Logger) -> set[tuple[str, str]]:
    """Retorna set (MZ, LT) excluidos: ESTADO=CORTADO o EXONERADO en registro_cortes.xlsx."""
    p = config.REGISTRO_CORTES_PATH
    if not p.exists():
        log.info("registro_cortes.xlsx no existe — primer ciclo, 0 excluidos")
        return set()
    df = pd.read_excel(p, header=1, dtype=str).fillna("")
    df.columns = [str(c).strip().upper() for c in df.columns]
    if "ESTADO" not in df.columns:
        log.warning("registro_cortes.xlsx · falta columna ESTADO — ignorando archivo")
        return set()
    excluidos: set[tuple[str, str]] = set()
    n_cortados = n_exonerados = 0
    for _, f in df.iterrows():
        estado = str(f.get("ESTADO", "")).strip().upper()
        if estado not in ("CORTADO", "EXONERADO"):
            continue
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if mz and lt:
            excluidos.add((mz, lt))
            if estado == "CORTADO":
                n_cortados += 1
            else:
                n_exonerados += 1
    log.info(f"registro_cortes.xlsx · {len(excluidos)} excluidos "
             f"(CORTADO={n_cortados} · EXONERADO={n_exonerados})")
    return excluidos


def _cargar_sin_servicio_activos(log: logging.Logger) -> set[tuple[str, str]]:
    """Retorna set (MZ, LT) de predios sin servicio de agua (1_lecturas/sin_servicio).
    Nunca son elegibles para corte — no tienen agua que cortar (ej. SIN_MEDIDOR)."""
    p = config.LISTA_SIN_SERVICIO_PATH
    if not p.exists():
        log.info("lista_sin_servicio.xlsx no existe — 0 excluidos por sin-servicio")
        return set()
    df = pd.read_excel(p, header=1, dtype=str).fillna("")
    df.columns = [str(c).strip().upper() for c in df.columns]
    excluidos: set[tuple[str, str]] = set()
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if mz and lt:
            excluidos.add((mz, lt))
    log.info(f"lista_sin_servicio.xlsx · {len(excluidos)} excluidos (sin servicio de agua)")
    return excluidos


# ── CARGA DE PAGOS EFECTIVO (MESA + COBRADOR) ────────────────────────────────
def _cargar_pagos_map(log: logging.Logger) -> dict:
    """Retorna {(mz, lt): (mesa, cobrador)} desde pagos_efectivo.xlsx."""
    p = config.PAGOS_EFECTIVO_PATH
    if not p.exists():
        log.info("pagos_efectivo.xlsx no existe — MESA/COBRADOR quedarán vacíos")
        return {}
    df = pd.read_excel(p, header=1)
    df.columns = [str(c).strip().upper() for c in df.columns]
    mapa: dict = {}
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        mesa     = str(f.get("MESA",     "")).strip()
        cobrador = str(f.get("COBRADOR", "")).strip()
        if (mz, lt) not in mapa:
            mapa[(mz, lt)] = (mesa, cobrador)
        else:
            em, ec = mapa[(mz, lt)]
            if mesa and mesa not in em:
                mesa = em + "/" + mesa
            else:
                mesa = em
            if cobrador and cobrador not in ec:
                cobrador = ec + "/" + cobrador
            else:
                cobrador = ec
            mapa[(mz, lt)] = (mesa, cobrador)
    log.info(f"pagos_efectivo.xlsx · {len(df)} filas · {len(mapa)} usuarios únicos")
    return mapa


def _cargar_abonos_map(mes_ano: str, log: logging.Logger) -> dict[tuple[str, str], float]:
    """Reutiliza el loader manifestado de 5_cobranza y suma por predio."""
    main_path = config.ROOT.parent / "5_cobranza" / "main.py"
    spec = importlib.util.spec_from_file_location("cobranza_main_corte", main_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"No se pudo cargar: {main_path}")
    cobranza = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(cobranza)
    por_ciclo = cobranza._cargar_abonos_rezagados(mes_ano)
    mapa = {
        predio: round(cerrado + vigente, 2)
        for predio, (cerrado, vigente) in por_ciclo.items()
        if cerrado + vigente > config.TOL
    }
    log.info(f"abonos_rezagados manifestados · {len(mapa)} usuarios")
    return mapa


# ── FILTRO ───────────────────────────────────────────────────────────────────
def _filtrar_corte(df: pd.DataFrame, mapa_reclamos: dict,
                   cortados_activos: set, pagos_map: dict,
                   abonos_map: dict[tuple[str, str], float],
                   log: logging.Logger) -> list[dict]:
    corte: list[dict] = []
    n_saldos_no_cero = 0
    n_bloqueados_reclamo = 0
    n_bloqueados_pago    = 0
    n_excluidos_cortados = 0
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        if (mz, lt) in cortados_activos:
            n_excluidos_cortados += 1
            continue
        saldo      = round(_float(f.get("SALDO")), 2)
        mes_ant    = round(_float(f.get("MES_ANTERIOR")), 2)
        monto_yape = round(_float(f.get("MONTO_YAPE")), 2)
        abono       = abonos_map.get((mz, lt), 0.0)

        if saldo > config.TOL:
            n_saldos_no_cero += 1
        if saldo <= config.TOL:
            continue
        if mes_ant < config.MES_ANTERIOR_MIN - config.TOL:
            continue

        mesa, cobrador = pagos_map.get((mz, lt), ("", ""))

        # Reclamo activo → bloquea corte (prioridad sobre pago parcial)
        bloqueado = mapa_reclamos.get((mz, lt))
        if bloqueado == "EN_REVISION":
            estado_reclamo = "EN_REVISION"
            ejecutar       = "NO"
            motivo         = "Reclamo en revision"
            n_bloqueados_reclamo += 1
        elif mesa or monto_yape > config.TOL or abono > config.TOL:
            # Pagó algo este mes (efectivo/mesa, yape o abono) → se salva del corte,
            # la deuda se acumula. La regla vale para cualquier canal: es injusto
            # cortar a quien pagó parcial por yape y no a quien pagó parcial en mesa.
            estado_reclamo = "SIN_RECLAMO"
            ejecutar       = "NO"
            if abono > config.TOL:
                motivo = "Abono rezagado"
            else:
                motivo = "Pago parcial en mesa" if mesa else "Pago parcial por yape"
            n_bloqueados_pago += 1
        else:
            estado_reclamo = "SIN_RECLAMO"
            ejecutar       = "SI"
            motivo         = ""

        corte.append({
            "mz":             mz,
            "lt":             lt,
            "nombre":         str(f.get("NOMBRE", "")).strip(),
            "saldo":          saldo,
            "mes_anterior":   mes_ant,
            "total_a_pagar":  round(saldo + config.PENALIDAD, 2),
            "estado_reclamo": estado_reclamo,
            "ejecutar_corte": ejecutar,
            "motivo":         motivo,
            "mesa":           mesa,
            "cobrador":       cobrador,
        })

    if n_saldos_no_cero == 0 and len(df) > 0:
        log.warning("Ninguna fila con SALDO>0 — verificar que 5_cobranza "
                    "escribio la columna SALDO correctamente")

    n_no = n_bloqueados_reclamo + n_bloqueados_pago
    log.info(f"Excluidos (servicio ya cortado o sin servicio de agua): {n_excluidos_cortados}")
    log.info(f"Elegibles · {len(corte)} usuarios "
             f"(SALDO>{config.TOL:.3f} AND MES_ANTERIOR>={config.MES_ANTERIOR_MIN})")
    log.info(f"  · EJECUTAR_CORTE = SI · {len(corte) - n_no}")
    log.info(f"  · EJECUTAR_CORTE = NO · {n_no} "
             f"(reclamo={n_bloqueados_reclamo} · pago_parcial={n_bloqueados_pago})")
    return corte


# ── EXPORT ───────────────────────────────────────────────────────────────────
def _exportar(corte: list[dict]) -> None:
    config.OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "lista_corte"
    ws.freeze_panes = "A3"

    for cs, ce, texto, bg, txt in _LC_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for col, nombre, bg, txt, ancho in _LC_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    MONEY = '"S/ "#,##0.00'
    for ri, r in enumerate(corte, 3):
        _c(ws, ri, 1, r["mz"],     TD_ID,    "1A5276", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_ID,    "1A5276", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_ID,    "333333", align="left")
        _c(ws, ri, 4, r["saldo"],            TD_DEUDA, "1E5C3A",
           mono=True, align="right", fmt=MONEY)
        _c(ws, ri, 5, int(r["mes_anterior"]), TD_DEUDA, "1E5C3A",
           mono=True, align="right")
        _c(ws, ri, 6, config.PENALIDAD,      TD_PEN, "145A32",
           mono=True, align="right", bold=True, fmt=MONEY)
        _c(ws, ri, 7, r["total_a_pagar"],    TD_PEN, "145A32",
           mono=True, align="right", bold=True, size=10, fmt=MONEY)

        # Sección ¿Ejecutar corte?
        _c(ws, ri, 8, r["estado_reclamo"], TD_REC, "4A235A", align="center")
        if r["ejecutar_corte"] == "SI":
            _c(ws, ri, 9, "SI", TD_SI, TX_SI, bold=True, align="center")
        else:
            _c(ws, ri, 9, "NO", TD_NO, TX_NO, bold=True, align="center")
        _c(ws, ri, 10, r["motivo"],   TD_REC,  "4A235A", align="left")
        # Sección ¿Pagó efectivo este mes?
        _c(ws, ri, 11, r["mesa"],     TD_PAGO, "92400E", align="center")
        _c(ws, ri, 12, r["cobrador"], TD_PAGO, "92400E", align="left")

        ws.row_dimensions[ri].height = 17

    wb.save(config.LISTA_CORTE_PATH)


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main() -> None:
    # Logging con force=True para que el FileHandler funcione aunque otro
    # logger esté activo (test runner, imports previos).
    config.OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(config.OUTPUTS_DIR / "run.log", encoding="utf-8"),
        ],
        force=True,
    )
    log = logging.getLogger(__name__)
    log.info("generar_lista.py · iniciando")

    print("=" * 60)
    print("  6_corte/generar_lista.py")
    print("=" * 60)

    print("\n[1/5] Validando inputs...")
    df, mes_ano = _validar_input(log)

    if _ciclo_ya_penalizado(mes_ano, log) and "--force" not in sys.argv:
        log.error(f"Ciclo {mes_ano} ya tiene penalidad aplicada (audit_penalidad.xlsx)")
        raise SystemExit(
            f"\nBLOQUEADO: el ciclo {mes_ano} ya pasó por aplicar_penalidad.py.\n"
            f"lista_corte.xlsx es el snapshot Dia 0 — seguimiento.py lo necesita "
            f"intacto para comparar contra planilla_cobrado del Dia 2.\n"
            f"Regenerarlo ahora destruye ese snapshot sin poder recuperarlo despues.\n"
            f"Si estas seguro (ej. seguimiento.py ya corrio y este es un ciclo nuevo "
            f"que aun no pasa por aplicar_penalidad), corre con --force.\n"
        )

    print("\n[2/5] Cargando cortados activos (registro_cortes.xlsx)...")
    cortados_activos = _cargar_cortados_activos(log)
    sin_servicio_activos = _cargar_sin_servicio_activos(log)
    excluidos_previos = cortados_activos | sin_servicio_activos

    print(f"\n[3/5] Cargando reclamos del ciclo {mes_ano}...")
    mapa_reclamos = _cargar_reclamos_map(mes_ano, log)

    print("\n[4/5] Cargando pagos efectivo y abonos rezagados...")
    pagos_map = _cargar_pagos_map(log)
    abonos_map = _cargar_abonos_map(mes_ano, log)

    print("\n[5/5] Filtrando elegibles y exportando lista_corte.xlsx...")
    corte = _filtrar_corte(
        df, mapa_reclamos, excluidos_previos, pagos_map, abonos_map, log
    )
    _exportar(corte)
    log.info(f"{config.LISTA_CORTE_PATH.name} -> {len(corte)} usuarios")

    n_si           = sum(1 for r in corte if r["ejecutar_corte"] == "SI")
    n_no_reclamo   = sum(1 for r in corte if r["motivo"] == "Reclamo en revision")
    n_no            = sum(1 for r in corte if r["ejecutar_corte"] == "NO")
    n_no_pago       = n_no - n_no_reclamo

    print("\n" + "=" * 60)
    print(f"  generar_lista.py completado")
    print(f"  -> {config.LISTA_CORTE_PATH}")
    print(f"  -> {len(corte)} en lista · EJECUTAR=SI: {n_si} · NO: {n_no}")
    print(f"     NO por reclamo: {n_no_reclamo} · NO por pago parcial: {n_no_pago}")
    if cortados_activos:
        print(f"  -> {len(cortados_activos)} excluidos (CORTADO o EXONERADO en registro_cortes)")
    if sin_servicio_activos:
        print(f"  -> {len(sin_servicio_activos)} excluidos (sin servicio de agua, lista_sin_servicio)")
    if n_si > 0:
        print(f"\n  Siguiente paso: python aplicar_penalidad.py")
        print(f"    (solo procesa filas con EJECUTAR_CORTE = SI)")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
