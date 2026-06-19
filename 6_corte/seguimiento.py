"""6_corte/seguimiento.py — Clasifica resultado post-ventana (Día 2)

Lee lista_corte.xlsx (snapshot Día 0) + planilla_cobrado.xlsx ciclo 2 (Día 2).
Por cada usuario en lista_corte, calcula cuánto pagó durante la ventana de
gracia y clasifica:

  pagado >= S/20  → SALVADO → pagaron_penalidad.xlsx
  pagado <  S/20  → CORTADO → corte_fisico.xlsx + arrastre_corte_YYYY-MM.xlsx

Cómo se calcula PAGADO_CORTE_RECONEXION:
  La planilla no rastrea pagos por componente — solo MONTO_PAGADO total.
  Pero como aplicar_penalidad sumó +20 al cargo CORTE_RECONEXION en el Día 0,
  y SALDO = TOTAL_CARGOS - TOTAL_PAGADO, entonces:

      SALDO_ciclo2 = lista.SALDO + 20 - pagado_en_ventana

  Despejando:
      pagado_en_ventana = lista.SALDO - SALDO_ciclo2 + 20

  Esa es la cantidad que el usuario pagó durante los 2 días, que conceptualmente
  se asigna al cargo CORTE_RECONEXION (que era 0 antes y 20 después).

Idempotente: re-correr regenera los 3 outputs sin estado entre runs.

Uso:
    python seguimiento.py
"""
import logging
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
import config

# ── PALETAS ──────────────────────────────────────────────────────────────────
# pagaron_penalidad — verdes/azules (salvados)
GH_P_ID    = ("EBF5FB", "1A5276")
GH_P_DEUDA = ("E9F7EF", "1E5C3A")
GH_P_PAGO  = ("1E8449", "FFFFFF")
GH_P_RES   = ("F3F4F6", "374151")
TD_P_ID    = "F4FAFF"
TD_P_DEUDA = "F4FBF7"
TD_P_PAGO  = "D5F5E3"
TD_P_RES   = "F9FAFB"

# corte_fisico — ámbar (urgencia operario)
GH_C_ID    = ("EBF5FB", "1A5276")
GH_C_DEUDA = ("E9F7EF", "1E5C3A")
GH_C_CORTE = ("92400E", "FFFFFF")
GH_C_RES   = ("F3F4F6", "374151")
TD_C_ID    = "F4FAFF"
TD_C_DEUDA = "F4FBF7"
TD_C_CORTE = "FEF3C7"
TD_C_KEY   = "D97706"   # ARRASTRE_CORTE - ámbar oscuro, columna clave
TD_C_RES   = "F9FAFB"

# arrastre_corte — morado (pendiente, lo consume 2_planilla del mes siguiente)
GH_A_ID    = ("EBF5FB", "1A5276")
GH_A_ORIG  = ("F3F4F6", "374151")
GH_A_ARR   = ("5B21B6", "FFFFFF")
TD_A_ID    = "F4FAFF"
TD_A_ORIG  = "F9FAFB"
TD_A_ARR   = "EDE9FE"

# registro_cortes — estado persistente
GH_R_ID  = ("EBF5FB", "1A5276"); TD_R_ID  = "F4FAFF"
GH_R_PER = ("FEF3C7", "78350F"); TD_R_PER = "FFFBEB"
GH_R_EST = ("1E8449", "FFFFFF")
TD_R_COR = "FADBD8"; TX_R_COR = "7B241C"
GH_R_TRZ = ("F3E8FF", "5B21B6"); TD_R_TRZ = "FAF5FF"

_R_GRUPOS = [
    (1, 3, "¿Quién es el usuario?", *GH_R_ID),
    (4, 5, "Período del corte",      *GH_R_PER),
    (6, 6, "Estado actual",          *GH_R_EST),
    (7, 9, "Trazabilidad",           *GH_R_TRZ),
]
_R_COLS = [
    (1, "MZ",               *GH_R_ID,   6),
    (2, "LT",               *GH_R_ID,   7),
    (3, "NOMBRE",           *GH_R_ID,  28),
    (4, "MES_INICIO_CORTE", *GH_R_PER, 18),
    (5, "MES_REACTIVACION", *GH_R_PER, 18),
    (6, "ESTADO",           *GH_R_EST, 14),
    (7, "OBSERVACIONES",    *GH_R_TRZ, 30),
    (8, "FECHA_REGISTRO",   *GH_R_TRZ, 18),
    (9, "SOURCE",           *GH_R_TRZ, 20),
]

# Fila 3 — ejemplo guía en gris itálico (ESTADO=EJEMPLO para que el código la ignore)
_R_EJEMPLO = [
    "B",                                        # MZ
    "5",                                        # LT
    "Rosa Mamani",                              # NOMBRE
    "2026-04",                                  # MES_INICIO_CORTE
    "",                                         # MES_REACTIVACION (vacío si sigue cortado)
    "EJEMPLO",                                  # ESTADO (cambiar a CORTADO al cargar real)
    "Borrar esta fila antes de cargar reales",  # OBSERVACIONES
    "2026-04-18",                               # FECHA_REGISTRO
    "ejemplo",                                  # SOURCE
]


# ── HELPERS DE ESTILO ────────────────────────────────────────────────────────
def _borde():
    b = Side(style="thin", color="CCCCCC")
    return Border(left=b, right=b, top=b, bottom=b)

def _c(ws, row, col, value=None, bg=None, txt="333333",
       bold=False, align="left", mono=False, size=9, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Consolas" if mono else "Arial",
                       size=size, bold=bold, color=txt)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _borde()
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    if fmt:
        c.number_format = fmt
    return c

def _gh(ws, row, cs, ce, texto, bg, txt):
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    c = ws.cell(row=row, column=cs, value=texto)
    c.font      = Font(name="Arial", size=8, bold=True, color=txt)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", start_color=bg)
    c.border    = _borde()

def _ch(ws, row, col, texto, bg, txt):
    _c(ws, row, col, texto, bg=bg, txt=txt, bold=True, align="center")

def _w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def _e(ws, row, col, value):
    """Celda de ejemplo guía — gris itálico, sin fondo (igual que crear_templates.py)."""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=10, color="9CA3AF", italic=True)
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c


# ── PARSING ──────────────────────────────────────────────────────────────────
def _norm_mz(val) -> str:
    if val is None:
        return ""
    s = str(val).strip().upper()
    return "" if not s or s in ("NAN", "NONE") else s

def _norm_lt(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN"):
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.upper().replace(" ", "")

def _float(val) -> float:
    if val is None:
        return 0.0
    try:
        f = float(str(val).replace(",", ".").strip())
        return 0.0 if f != f else f
    except (ValueError, TypeError):
        return 0.0


# ── INPUTS ───────────────────────────────────────────────────────────────────
def _leer_lista_corte() -> dict[tuple[str, str], dict]:
    """Snapshot del Día 0 — clave (MZ, LT) → {nombre, saldo}."""
    p = config.LISTA_CORTE_PATH
    if not p.exists():
        raise FileNotFoundError(
            f"Falta: {p}\n  -> Correr generar_lista.py primero"
        )
    df = pd.read_excel(p, header=1)
    df.columns = [str(c).strip().upper() for c in df.columns]
    requeridas = {"MZ", "LT", "NOMBRE", "SALDO"}
    faltantes = requeridas - set(df.columns)
    if faltantes:
        raise ValueError(f"lista_corte.xlsx · faltan columnas {sorted(faltantes)}")

    snapshot = {}
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        snapshot[(mz, lt)] = {
            "nombre": str(f.get("NOMBRE", "")).strip(),
            "saldo":  round(_float(f.get("SALDO")), 2),
        }
    return snapshot


def _leer_planilla_cobrado() -> tuple[dict[tuple[str, str], dict], str]:
    """Estado ciclo 2 — clave (MZ, LT) → {saldo_actual}, también devuelve MES_ANO."""
    p = config.PLANILLA_COBRADO_PATH
    if not p.exists():
        raise FileNotFoundError(
            f"Falta: {p}\n  -> Re-correr 5_cobranza/main.py (ciclo 2) primero"
        )
    df = pd.read_excel(p, header=1)
    df.columns = [str(c).strip().upper() for c in df.columns]
    requeridas = {"MZ", "LT", "SALDO", "MES_ANO"}
    faltantes = requeridas - set(df.columns)
    if faltantes:
        raise ValueError(
            f"planilla_cobrado.xlsx · faltan columnas {sorted(faltantes)}"
        )

    estado = {}
    mes_ano = ""
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        if not mes_ano:
            mes_ano = str(f.get("MES_ANO", "")).strip()
        estado[(mz, lt)] = {
            "saldo_actual": round(_float(f.get("SALDO")), 2),
        }
    return estado, mes_ano


# ── CLASIFICACIÓN ────────────────────────────────────────────────────────────
def _clasificar(
    snapshot: dict[tuple[str, str], dict],
    estado: dict[tuple[str, str], dict],
    fecha_seg: str,
    log: logging.Logger,
) -> tuple[list[dict], list[dict]]:
    """Devuelve (salvados, cortados)."""
    salvados: list[dict] = []
    cortados: list[dict] = []
    no_encontrados: list[str] = []

    for (mz, lt), snap in snapshot.items():
        e = estado.get((mz, lt))
        if e is None:
            no_encontrados.append(f"{mz}-{lt}")
            continue

        # pagado durante la ventana = lista.SALDO - SALDO_ciclo2 + PENALIDAD
        # (clamped a 0 — el usuario nunca "paga negativo")
        pagado = max(0.0, round(
            snap["saldo"] - e["saldo_actual"] + config.PENALIDAD, 2
        ))

        fila = {
            "mz":     mz,
            "lt":     lt,
            "nombre": snap["nombre"],
            "saldo":  snap["saldo"],
            "pagado_corte_reconexion": pagado,
            "fecha_seguimiento": fecha_seg,
        }

        if pagado >= config.PENALIDAD - config.TOL:
            fila["estado"] = "SALVADO"
            salvados.append(fila)
        else:
            # ARRASTRE_CORTE = lo que falta del cargo escalado de S/40.
            arrastre = max(0.0, round(config.PENALIDAD_FINAL - pagado, 2))
            fila["estado"]            = "CORTADO"
            fila["penalidad_final"]   = config.PENALIDAD_FINAL
            fila["arrastre_corte"]    = arrastre
            cortados.append(fila)

    if no_encontrados:
        log.warning(
            f"{len(no_encontrados)} usuarios en lista_corte no aparecen en "
            f"planilla_cobrado ciclo 2: {', '.join(no_encontrados[:10])}"
            f"{'...' if len(no_encontrados) > 10 else ''}"
        )

    log.info(f"Clasificacion · SALVADOS={len(salvados)} · CORTADOS={len(cortados)}")
    return salvados, cortados


# ── EXPORT: pagaron_penalidad.xlsx ───────────────────────────────────────────
_P_GRUPOS = [
    (1, 3, "¿Quién es?",       *GH_P_ID),
    (4, 4, "¿Cuánto debía?",   *GH_P_DEUDA),
    (5, 5, "¿Qué pagó?",       *GH_P_PAGO),
    (6, 7, "Resultado",        *GH_P_RES),
]
_P_COLS = [
    (1, "MZ",                       *GH_P_ID,     6),
    (2, "LT",                       *GH_P_ID,     7),
    (3, "NOMBRE",                   *GH_P_ID,    28),
    (4, "SALDO",                    *GH_P_DEUDA, 14),
    (5, "PAGADO_CORTE_RECONEXION",  *GH_P_PAGO,  24),
    (6, "ESTADO",                   *GH_P_RES,   12),
    (7, "FECHA_SEGUIMIENTO",        *GH_P_RES,   18),
]

def _exportar_pagaron(salvados: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "salvados"
    ws.freeze_panes = "A3"

    for cs, ce, texto, bg, txt in _P_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for col, nombre, bg, txt, ancho in _P_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    MONEY = '"S/ "#,##0.00'
    for ri, r in enumerate(salvados, 3):
        _c(ws, ri, 1, r["mz"],     TD_P_ID,    "1A5276", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_P_ID,    "1A5276", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_P_ID,    "333333", align="left")
        _c(ws, ri, 4, r["saldo"],  TD_P_DEUDA, "1E5C3A",
           mono=True, align="right", fmt=MONEY)
        _c(ws, ri, 5, r["pagado_corte_reconexion"], TD_P_PAGO, "145A32",
           mono=True, align="right", bold=True, fmt=MONEY)
        _c(ws, ri, 6, r["estado"], TD_P_RES, "1E5C3A",
           mono=True, align="center", bold=True)
        _c(ws, ri, 7, r["fecha_seguimiento"], TD_P_RES, "374151",
           mono=True, align="center")
        ws.row_dimensions[ri].height = 17

    wb.save(config.PAGARON_PATH)


# ── EXPORT: corte_fisico.xlsx ────────────────────────────────────────────────
_C_GRUPOS = [
    (1, 3, "¿A dónde va el operario?",         *GH_C_ID),
    (4, 4, "Deuda de consumo",                  *GH_C_DEUDA),
    (5, 7, "Penalidad de corte (escalada a S/40)", *GH_C_CORTE),
    (8, 9, "Cierre",                            *GH_C_RES),
]
_C_COLS = [
    (1, "MZ",                       *GH_C_ID,     6),
    (2, "LT",                       *GH_C_ID,     7),
    (3, "NOMBRE",                   *GH_C_ID,    28),
    (4, "SALDO",                    *GH_C_DEUDA, 14),
    (5, "PAGADO_CORTE_RECONEXION",  *GH_C_CORTE, 24),
    (6, "PENALIDAD_FINAL",          *GH_C_CORTE, 16),
    (7, "ARRASTRE_CORTE",           *GH_C_CORTE, 16),
    (8, "ESTADO",                   *GH_C_RES,   12),
    (9, "FECHA_SEGUIMIENTO",        *GH_C_RES,   18),
]

def _exportar_corte_fisico(cortados: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "cortes"
    ws.freeze_panes = "A3"

    for cs, ce, texto, bg, txt in _C_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for col, nombre, bg, txt, ancho in _C_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    MONEY = '"S/ "#,##0.00'
    for ri, r in enumerate(cortados, 3):
        _c(ws, ri, 1, r["mz"],     TD_C_ID,    "1A5276", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_C_ID,    "1A5276", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_C_ID,    "333333", align="left")
        _c(ws, ri, 4, r["saldo"],  TD_C_DEUDA, "1E5C3A",
           mono=True, align="right", fmt=MONEY)
        _c(ws, ri, 5, r["pagado_corte_reconexion"], TD_C_CORTE, "78350F",
           mono=True, align="right", fmt=MONEY)
        _c(ws, ri, 6, r["penalidad_final"], TD_C_CORTE, "78350F",
           mono=True, align="right", fmt=MONEY)
        # ARRASTRE_CORTE — columna clave, ámbar oscuro
        _c(ws, ri, 7, r["arrastre_corte"], TD_C_KEY, "FFFFFF",
           mono=True, align="right", bold=True, size=10, fmt=MONEY)
        _c(ws, ri, 8, r["estado"], TD_C_RES, "78350F",
           mono=True, align="center", bold=True)
        _c(ws, ri, 9, r["fecha_seguimiento"], TD_C_RES, "374151",
           mono=True, align="center")
        ws.row_dimensions[ri].height = 17

    wb.save(config.CORTE_FISICO_PATH)


# ── EXPORT: arrastre_corte_YYYY-MM.xlsx ──────────────────────────────────────
_A_GRUPOS = [
    (1, 3, "¿Quién es?",              *GH_A_ID),
    (4, 4, "¿De qué mes?",             *GH_A_ORIG),
    (5, 5, "Deuda de reconexión",      *GH_A_ARR),
]
_A_COLS = [
    (1, "MZ",             *GH_A_ID,    6),
    (2, "LT",             *GH_A_ID,    7),
    (3, "NOMBRE",         *GH_A_ID,   28),
    (4, "MES_ORIGEN",     *GH_A_ORIG, 14),
    (5, "ARRASTRE_CORTE", *GH_A_ARR,  18),
]

def _exportar_arrastre(cortados: list[dict], mes_ano: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "arrastre"
    ws.freeze_panes = "A3"

    for cs, ce, texto, bg, txt in _A_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for col, nombre, bg, txt, ancho in _A_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    MONEY = '"S/ "#,##0.00'
    for ri, r in enumerate(cortados, 3):
        _c(ws, ri, 1, r["mz"],     TD_A_ID,    "1A5276", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_A_ID,    "1A5276", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_A_ID,    "333333", align="left")
        _c(ws, ri, 4, mes_ano,     TD_A_ORIG,  "374151", mono=True, align="center")
        _c(ws, ri, 5, r["arrastre_corte"], TD_A_ARR, "5B21B6",
           mono=True, align="right", bold=True, size=10, fmt=MONEY)
        ws.row_dimensions[ri].height = 17

    path = config.arrastre_corte_path(mes_ano)
    wb.save(path)


# ── REGISTRO CORTES (append idempotente) ────────────────────────────────────
def _leer_cortados_existentes(log: logging.Logger) -> set[tuple[str, str, str]]:
    """Retorna set (MZ, LT, MES_INICIO_CORTE) ya registrados — para idempotencia.
    Ignora la fila de ejemplo guía (ESTADO=EJEMPLO)."""
    p = config.REGISTRO_CORTES_PATH
    if not p.exists():
        return set()
    df = pd.read_excel(p, header=1, dtype=str).fillna("")
    df.columns = [str(c).strip().upper() for c in df.columns]
    existentes: set[tuple[str, str, str]] = set()
    for _, f in df.iterrows():
        estado = str(f.get("ESTADO", "")).strip().upper()
        if estado not in ("CORTADO", "REACTIVADO"):
            continue   # salta fila de ejemplo u otros estados no válidos
        mz  = _norm_mz(f.get("MZ"))
        lt  = _norm_lt(f.get("LT"))
        mes = str(f.get("MES_INICIO_CORTE", "")).strip()
        if mz and lt and mes:
            existentes.add((mz, lt, mes))
    log.info(f"registro_cortes.xlsx · {len(existentes)} registros existentes")
    return existentes


def _appendar_registro_cortes(
    cortados: list[dict], mes_ano: str, log: logging.Logger
) -> int:
    """Agrega filas ESTADO=CORTADO al archivo persistente. Retorna cantidad de filas nuevas."""
    p = config.REGISTRO_CORTES_PATH
    config.INPUTS_DIR.mkdir(parents=True, exist_ok=True)

    existentes = _leer_cortados_existentes(log)
    fecha_registro = datetime.now().strftime("%Y-%m-%d")

    if p.exists():
        wb = load_workbook(p)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Cortes"
        ws.freeze_panes = "A3"
        for cs, ce, texto, bg, txt in _R_GRUPOS:
            _gh(ws, 1, cs, ce, texto, bg, txt)
        for col, nombre, bg, txt, ancho in _R_COLS:
            _ch(ws, 2, col, nombre, bg, txt)
            _w(ws, col, ancho)
        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 22
        # Fila 3 — ejemplo guía en gris itálico (supervisor la borra al cargar reales)
        for idx, valor in enumerate(_R_EJEMPLO, start=1):
            _e(ws, 3, idx, valor)
        ws.row_dimensions[3].height = 18

    n_nuevos = 0
    next_row = ws.max_row + 1
    for r in cortados:
        clave = (r["mz"], r["lt"], mes_ano)
        if clave in existentes:
            continue
        ri = next_row + n_nuevos
        _c(ws, ri, 1, r["mz"],     TD_R_ID,  "1A5276", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_R_ID,  "1A5276", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_R_ID,  "333333", align="left")
        _c(ws, ri, 4, mes_ano,     TD_R_PER, "78350F", mono=True, align="center")
        _c(ws, ri, 5, "",          TD_R_PER, "78350F", align="center")
        _c(ws, ri, 6, "CORTADO",   TD_R_COR, TX_R_COR, bold=True, align="center")
        _c(ws, ri, 7, f"Corte físico confirmado {mes_ano}", TD_R_TRZ, "4A235A", align="left")
        _c(ws, ri, 8, fecha_registro, TD_R_TRZ, "4A235A", mono=True, align="center")
        _c(ws, ri, 9, "seguimiento.py",        TD_R_TRZ, "4A235A", align="left")
        ws.row_dimensions[ri].height = 17
        n_nuevos += 1

    wb.save(p)
    return n_nuevos


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main() -> None:
    config.OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(config.OUTPUTS_DIR / "run.log", encoding="utf-8"),
        ],
        force=True,
    )
    log = logging.getLogger(__name__)
    log.info("seguimiento.py · iniciando")

    print("=" * 60)
    print("  6_corte/seguimiento.py")
    print("=" * 60)

    print("\n[1/5] Validando inputs...")
    snapshot = _leer_lista_corte()
    log.info(f"lista_corte.xlsx · {len(snapshot)} usuarios (snapshot Dia 0)")
    estado, mes_ano = _leer_planilla_cobrado()
    log.info(f"planilla_cobrado.xlsx ciclo 2 · {len(estado)} usuarios · MES_ANO={mes_ano}")
    if not mes_ano:
        raise ValueError("No se pudo detectar MES_ANO en planilla_cobrado")

    print("\n[2/5] Clasificando segun pago durante la ventana...")
    fecha_seg = datetime.now().strftime("%d/%m/%Y")
    salvados, cortados = _clasificar(snapshot, estado, fecha_seg, log)

    print(f"\n[3/5] Escribiendo outputs...")
    _exportar_pagaron(salvados)
    log.info(f"{config.PAGARON_PATH.name} -> {len(salvados)} salvados")

    _exportar_corte_fisico(cortados)
    log.info(f"{config.CORTE_FISICO_PATH.name} -> {len(cortados)} cortes fisicos")

    arrastre_path = config.arrastre_corte_path(mes_ano)
    _exportar_arrastre(cortados, mes_ano)
    log.info(f"{arrastre_path.name} -> {len(cortados)} arrastres")

    print(f"\n[4/5] Registrando {len(cortados)} cortes fisicos en registro_cortes.xlsx...")
    n_nuevos = _appendar_registro_cortes(cortados, mes_ano, log)
    log.info(f"registro_cortes.xlsx · {n_nuevos} filas nuevas agregadas")

    print("\n[5/5] Resumen del ciclo")
    total = len(snapshot)
    print(f"  · Total en lista_corte: {total}")
    print(f"  · Salvados (pagaron >= S/{config.PENALIDAD:.0f}): {len(salvados)}")
    print(f"  · Cortados (pagaron <  S/{config.PENALIDAD:.0f}): {len(cortados)}")
    if total > 0:
        pct_salvados = 100.0 * len(salvados) / total
        print(f"  · Tasa de salvacion: {pct_salvados:.1f}%")
    print(f"  · Nuevos en registro_cortes.xlsx: {n_nuevos}")

    print("\n" + "=" * 60)
    print(f"  seguimiento.py completado")
    print(f"  -> {config.PAGARON_PATH}")
    print(f"  -> {config.CORTE_FISICO_PATH}  (entregar al operario)")
    print(f"  -> {arrastre_path}  (pasar a 2_planilla del mes siguiente)")
    print(f"  -> {config.REGISTRO_CORTES_PATH}  ({n_nuevos} nuevos CORTADO)")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
