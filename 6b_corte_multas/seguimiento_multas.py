"""6b_corte_multas/seguimiento_multas.py — Clasifica resultado post-ventana (Día 2)

Espejo de 6_corte/seguimiento.py. Lee lista_multas.xlsx (snapshot Día 0, solo
EJECUTAR_CORTE=SI) + planilla_cobrado.xlsx ciclo 2 (Día 2). Por cada usuario
elegible clasifica:

  SALVADO  ⟺ saldó TODA su deuda menos convenio  ⟺  SALDO_ciclo2 ≤ CONVENIO
  CORTADO  ⟺ aún le queda deuda no-convenio

  Por qué SALDO_ciclo2 ≤ CONVENIO: convenio es lo más bajo en la cascada de
  prioridad (agua → corte → multa → acuerdos → convenio). Si todo lo demás
  está pagado, el único saldo que puede quedar es convenio. Convenio es deuda
  grande de pago opcional — no impide salvarse del corte.

Condición MÁS ESTRICTA que 6_corte: allá basta pagar S/20 de penalidad; aquí
hay que saldar agua + corte (incl. la penalidad de S/40 en CORTE_RECONEXION) +
multa + acuerdos. Convenio queda fuera.

Genera:
  - pagaron_penalidad_multas.xlsx  (salvados + trazabilidad del pago)
  - corte_fisico_multas.xlsx       (cortados → operario)
  - arrastre_multa_YYYY-MM.xlsx    (penalidad/reconexión pendiente → 2_planilla)
  - append a shared/registro_cortes.xlsx (nuevos CORTADO por multa)

Idempotente: re-correr regenera los outputs sin estado entre runs.

Uso:
    python seguimiento_multas.py
"""
import logging
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
import config

# ── PALETAS ──────────────────────────────────────────────────────────────────
# pagaron_penalidad_multas
GH_P_ID    = ("EBF5FB", "1A5276")
GH_P_DEBE  = ("FDEBD0", "7C2D12")   # naranja — ¿Cuánto tenía que saldar?
GH_P_SALDA = ("D1FAE5", "065F46")   # verde — ¿Lo saldó?
GH_P_TRAZ  = ("FEF9E7", "7D6608")   # ámbar — ¿Cómo pagó?
TD_P_ID    = "F4FAFF"
TD_P_DEBE  = "FFF7ED"
TD_P_TOTAL = "FDEBD0"               # TOTAL_A_SALDAR — naranja oscuro (key)
TD_P_SALDA = "D1FAE5"
TD_P_RES   = "F9FAFB"
TD_P_TRAZ  = "FFFDF5"
TX_TIPO_YAPE = "065F46"
TX_TIPO_EF   = "1E40AF"
TX_TIPO_NO   = "6B7280"

# corte_fisico_multas — rojo (urgencia operario)
GH_C_ID    = ("EBF5FB", "1A5276")
GH_C_DEUDA = ("FDEBD0", "7C2D12")
GH_C_CORTE = ("B91C1C", "FFFFFF")
GH_C_RES   = ("F3F4F6", "374151")
TD_C_ID    = "F4FAFF"
TD_C_DEUDA = "FFF7ED"
TD_C_CORTE = "FEE2E2"
TD_C_KEY   = "DC2626"   # ARRASTRE_MULTA — rojo, columna clave
TD_C_RES   = "F9FAFB"

# arrastre_multa — morado
GH_A_ID    = ("EBF5FB", "1A5276")
GH_A_ORIG  = ("F3F4F6", "374151")
GH_A_ARR   = ("5B21B6", "FFFFFF")
TD_A_ID    = "F4FAFF"
TD_A_ORIG  = "F9FAFB"
TD_A_ARR   = "EDE9FE"

# registro_cortes — estado persistente (mismo formato que 6_corte)
GH_R_ID  = ("EBF5FB", "1A5276"); TD_R_ID  = "F4FAFF"
GH_R_PER = ("FEF3C7", "78350F"); TD_R_PER = "FFFBEB"
GH_R_EST = ("1E8449", "FFFFFF")
TD_R_COR = "FADBD8"; TX_R_COR = "7B241C"
GH_R_TRZ = ("F3E8FF", "5B21B6"); TD_R_TRZ = "FAF5FF"

_R_GRUPOS = [
    (1, 3, "¿Quién es el usuario?", *GH_R_ID),
    (4, 5, "Período del corte",      *GH_R_PER),
    (6, 6, "Estado actual",          *GH_R_EST),
    (7, 9, "Trazabilidad",           *GH_R_TRZ),
]
_R_COLS = [
    (1, "MZ",               *GH_R_ID,   6),
    (2, "LT",               *GH_R_ID,   7),
    (3, "NOMBRE",           *GH_R_ID,  28),
    (4, "MES_INICIO_CORTE", *GH_R_PER, 18),
    (5, "MES_REACTIVACION", *GH_R_PER, 18),
    (6, "ESTADO",           *GH_R_EST, 14),
    (7, "OBSERVACIONES",    *GH_R_TRZ, 30),
    (8, "FECHA_REGISTRO",   *GH_R_TRZ, 18),
    (9, "SOURCE",           *GH_R_TRZ, 20),
]
_R_EJEMPLO = [
    "B", "5", "Rosa Mamani", "2026-04", "", "EJEMPLO",
    "Borrar esta fila antes de cargar reales", "2026-04-18", "ejemplo",
]


# ── HELPERS DE ESTILO ────────────────────────────────────────────────────────
def _borde():
    b = Side(style="thin", color="CCCCCC")
    return Border(left=b, right=b, top=b, bottom=b)

def _c(ws, row, col, value=None, bg=None, txt="333333",
       bold=False, align="left", mono=False, size=9, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Consolas" if mono else "Arial",
                       size=size, bold=bold, color=txt)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _borde()
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    if fmt:
        c.number_format = fmt
    return c

def _gh(ws, row, cs, ce, texto, bg, txt):
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    c = ws.cell(row=row, column=cs, value=texto)
    c.font      = Font(name="Arial", size=8, bold=True, color=txt)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", start_color=bg)
    c.border    = _borde()

def _ch(ws, row, col, texto, bg, txt):
    _c(ws, row, col, texto, bg=bg, txt=txt, bold=True, align="center")

def _w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def _e(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=10, color="9CA3AF", italic=True)
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c


# ── PARSING ──────────────────────────────────────────────────────────────────
def _norm_mz(val) -> str:
    if val is None:
        return ""
    s = str(val).strip().upper()
    return "" if not s or s in ("NAN", "NONE") else s

def _norm_lt(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN"):
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.upper().replace(" ", "")

def _float(val) -> float:
    if val is None:
        return 0.0
    try:
        f = float(str(val).replace(",", ".").strip())
        return 0.0 if f != f else f
    except (ValueError, TypeError):
        return 0.0


# ── INPUTS ───────────────────────────────────────────────────────────────────
def _leer_lista_multas() -> tuple[dict[tuple[str, str], dict], int]:
    """Snapshot del Día 0 — solo EJECUTAR_CORTE=SI · retorna (snapshot, omitidos)."""
    p = config.LISTA_MULTAS_PATH
    if not p.exists():
        raise FileNotFoundError(f"Falta: {p}\n  -> Correr generar_lista_multas.py primero")
    df = pd.read_excel(p, header=1)
    df.columns = [str(c).strip().upper() for c in df.columns]
    requeridas = {"MZ", "LT", "NOMBRE", "DEUDA_MULTA", "TOTAL_A_PAGAR", "EJECUTAR_CORTE"}
    faltantes = requeridas - set(df.columns)
    if faltantes:
        raise ValueError(f"lista_multas.xlsx · faltan columnas {sorted(faltantes)}")

    snapshot = {}
    omitidos = 0
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        if str(f.get("EJECUTAR_CORTE", "")).strip().upper() != "SI":
            omitidos += 1
            continue
        snapshot[(mz, lt)] = {
            "nombre":         str(f.get("NOMBRE", "")).strip(),
            "deuda_multa":    round(_float(f.get("DEUDA_MULTA")), 2),
            "total_a_saldar": round(_float(f.get("TOTAL_A_PAGAR")), 2),
        }
    return snapshot, omitidos


def _leer_planilla_cobrado() -> tuple[dict[tuple[str, str], dict], str]:
    """Estado ciclo 2 — clave (MZ, LT) → {saldo, convenio, corte_reconexion}."""
    p = config.PLANILLA_COBRADO_PATH
    if not p.exists():
        raise FileNotFoundError(
            f"Falta: {p}\n  -> Re-correr 5_cobranza/main.py (ciclo 2) primero"
        )
    df = pd.read_excel(p, header=1)
    df.columns = [str(c).strip().upper() for c in df.columns]
    requeridas = {"MZ", "LT", "SALDO", "CONVENIO", "CORTE_RECONEXION", "MES_ANO"}
    faltantes = requeridas - set(df.columns)
    if faltantes:
        raise ValueError(f"planilla_cobrado.xlsx · faltan columnas {sorted(faltantes)}")

    estado = {}
    mes_ano = ""
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        if not mes_ano:
            mes_ano = str(f.get("MES_ANO", "")).strip()
        estado[(mz, lt)] = {
            "saldo_actual":     round(_float(f.get("SALDO")), 2),
            "convenio":         round(_float(f.get("CONVENIO")), 2),
            "corte_reconexion": round(_float(f.get("CORTE_RECONEXION")), 2),
        }
    return estado, mes_ano


# ── GUARD: planilla_cobrado refleja las penalidades aplicadas ────────────────
def _verificar_planilla_sincronizada(
    estado: dict[tuple[str, str], dict],
    mes_ano: str,
    log: logging.Logger,
) -> None:
    """Aborta si planilla_cobrado no refleja las penalidades ya aplicadas.

    aplicar_penalidad_multas registra +PENALIDAD en audit_penalidad_multas.
    5_cobranza debe re-correrse para overlayar ese cargo en planilla_cobrado. Si no
    se hizo, planilla_cobrado queda desincronizada y la clasificación daría falsos
    SALVADOS.

    TODO (B7 Fase 2, deferido): bajo Modelo A este guard compara el ABSOLUTO
    (CORTE_RECON_DESPUES, col 6) que se pudre si la base cambia. El reemplazo
    correcto compara el NET DELTA (col 5) igual que 5_cobranza._cargar_penalidades,
    pero necesita separar base de delta (planilla_cobrado no expone la base).
    Diseñar cuando se active 6b (no se ejerce en junio: sin audit → sale temprano).
    """
    p = config.AUDIT_PENALIDAD_PATH
    if not p.exists():
        return
    df = pd.read_excel(p, header=1, dtype=str).fillna("")
    df.columns = [str(c).strip().upper() for c in df.columns]
    if "MES_ANO" not in df.columns or "CORTE_RECON_DESPUES" not in df.columns:
        log.warning("audit_penalidad_multas sin columnas esperadas — guard de sync omitido")
        return

    tiene_accion = "ACCION" in df.columns
    neto: dict[tuple[str, str], int] = {}
    esperado_cr: dict[tuple[str, str], float] = {}
    for _, f in df.iterrows():
        if str(f.get("MES_ANO", "")).strip() != mes_ano:
            continue
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        accion = (str(f.get("ACCION", "")).strip().upper()
                  if tiene_accion else "APLICADO") or "APLICADO"
        key = (mz, lt)
        if accion == "APLICADO":
            neto[key] = neto.get(key, 0) + 1
            esperado_cr[key] = round(_float(f.get("CORTE_RECON_DESPUES")), 2)
        elif accion == "REVERTIDO":
            neto[key] = neto.get(key, 0) - 1

    aplicados = [k for k, v in neto.items() if v > 0]
    if not aplicados:
        return

    desync = []
    for key in aplicados:
        esperado = esperado_cr.get(key, 0.0)
        if esperado <= config.TOL:
            continue
        actual = round(estado.get(key, {}).get("corte_reconexion", 0.0), 2)
        if abs(esperado - actual) > config.TOL:
            desync.append((key, esperado, actual))

    if not desync:
        log.info(f"planilla_cobrado sincronizada · {len(aplicados)} penalidad(es) reflejada(s)")
        return

    log.error(
        f"DESYNC planilla_cobrado · {len(desync)} de {len(aplicados)} penalidad(es) "
        f"NO reflejada(s) (CORTE_RECONEXION)"
    )
    print()
    print("=" * 60)
    print("  BLOQUEADO — planilla_cobrado desactualizada")
    print()
    print(f"  aplicar_penalidad_multas.py aplicó penalidades del ciclo {mes_ano} en")
    print(f"  shared/planilla_mes, pero {len(desync)} no llegaron a planilla_cobrado:")
    print(f"  su SALDO no incluye el +S/{config.PENALIDAD:.0f}.")
    print()
    print("  Predio · CORTE_RECONEXION esperado · en planilla_cobrado:")
    for (mz, lt), esp, act in desync[:5]:
        print(f"    {mz}-{lt}:  esperado S/{esp:.0f}  ·  actual S/{act:.0f}")
    if len(desync) > 5:
        print(f"    ... y {len(desync) - 5} más")
    print()
    print("  -> Re-corre 5_cobranza/main.py (ciclo 2) para propagar el cargo,")
    print("     luego vuelve a correr seguimiento_multas.py")
    print("=" * 60)
    print()
    sys.exit(1)


# ── TRAZABILIDAD DEL PAGO (Yape / Efectivo) ──────────────────────────────────
def _ancla_periodo(log: logging.Logger):
    """Inicio del periodo de cobro = fecha máxima del reporte procesado más
    reciente en shared/reporte_acumulado_procesado/ (criterio de motor_matching)."""
    archivos = sorted(
        p for p in config.SHARED_PROCESADO_DIR.glob("*_procesado.xlsx")
        if not p.name.startswith("~$")
    )
    if not archivos:
        log.warning("Sin reportes procesados — no se filtra por periodo de cobro")
        return None

    archivo = archivos[-1]
    try:
        wb = load_workbook(archivo, read_only=True, data_only=True)
        ws = wb["TE_PAGÓ"] if "TE_PAGÓ" in wb.sheetnames else wb.active
        datos = list(ws.values)
        wb.close()
    except Exception as e:
        log.warning(f"No se pudo leer ancla de {archivo.name}: {e}")
        return None

    if len(datos) < 2:
        return None

    headers = [str(h).strip().lower() if h else "" for h in datos[0]]
    col_fecha = next((i for i, h in enumerate(headers) if "fecha" in h), None)
    if col_fecha is None and len(datos) > 2:
        headers   = [str(h).strip().lower() if h else "" for h in datos[1]]
        col_fecha = next((i for i, h in enumerate(headers) if "fecha" in h), None)
        filas = datos[2:]
    else:
        filas = datos[1:]
    if col_fecha is None:
        log.warning(f"{archivo.name} · sin columna fecha — no se filtra por periodo")
        return None

    fechas = []
    for fila in filas:
        if not fila or col_fecha >= len(fila):
            continue
        dt = _parse_fecha(fila[col_fecha])
        if dt is not None:
            fechas.append(dt)
    if not fechas:
        return None

    ancla = max(fechas)
    log.info(f"Ancla periodo de cobro · {archivo.name} · {ancla:%d/%m/%Y %H:%M}")
    return ancla


def _parse_fecha(val):
    if val is None:
        return None
    if isinstance(val, pd.Timestamp):
        return None if pd.isna(val) else val
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return None
    dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
    return None if pd.isna(dt) else dt


def _cargar_trazabilidad_pagos(log: logging.Logger, ancla=None) -> dict[tuple[str, str], dict]:
    """Cruza (MZ, LT) contra pagos_yape_tepago.xlsx y pagos_efectivo.xlsx,
    filtrando al periodo de cobro actual (FECHA >= ancla). Devuelve el pago más
    reciente por predio."""
    candidatos: dict[tuple[str, str], list[dict]] = defaultdict(list)

    def _en_periodo(dt) -> bool:
        if ancla is None:
            return True
        return dt is not None and dt >= ancla

    py = config.PAGOS_YAPE_TEPAGO_PATH
    if py.exists():
        df = pd.read_excel(py, header=1)
        df.columns = [str(c).strip().upper() for c in df.columns]
        n = n_fuera = 0
        for _, f in df.iterrows():
            mz = _norm_mz(f.get("MZ"))
            lt = _norm_lt(f.get("LOTE"))   # Yape usa LOTE
            if not mz or not lt:
                continue
            dt = _parse_fecha(f.get("FECHA"))
            if not _en_periodo(dt):
                n_fuera += 1
                continue
            origen = str(f.get("ORIGEN", "")).strip()
            if origen.lower() in ("nan", "none"):
                origen = ""
            mensaje = str(f.get("MENSAJE", "")).strip()
            if mensaje.lower() in ("nan", "none"):
                mensaje = ""
            candidatos[(mz, lt)].append({
                "tipo":       "YAPE",
                "referencia": origen or "(sin origen)",
                "monto":      round(_float(f.get("MONTO_PAGO")), 2),
                "fecha_dt":   dt,
                "fecha_str":  dt.strftime("%d/%m/%Y %H:%M") if dt is not None else "",
                "comentario": mensaje or (f"Yape de {origen}" if origen else "Pago Yape"),
            })
            n += 1
        log.info(f"pagos_yape_tepago.xlsx · {n} pagos en periodo · {n_fuera} fuera")
    else:
        log.info("pagos_yape_tepago.xlsx no existe — pagos Yape no trazables")

    pe = config.PAGOS_EFECTIVO_PATH
    if pe.exists():
        df = pd.read_excel(pe, header=1)
        df.columns = [str(c).strip().upper() for c in df.columns]
        n = n_fuera = 0
        for _, f in df.iterrows():
            mz = _norm_mz(f.get("MZ"))
            lt = _norm_lt(f.get("LT"))
            if not mz or not lt:
                continue
            dt = _parse_fecha(f.get("FECHA"))
            if not _en_periodo(dt):
                n_fuera += 1
                continue
            mesa     = str(f.get("MESA", "")).strip()
            cobrador = str(f.get("COBRADOR", "")).strip()
            com      = str(f.get("COMENTARIO", "")).strip()
            if com.lower() in ("nan", "none"):
                com = ""
            ref = " / ".join(x for x in (mesa, cobrador) if x) or "(sin mesa)"
            candidatos[(mz, lt)].append({
                "tipo":       "EFECTIVO",
                "referencia": ref,
                "monto":      round(_float(f.get("MONTO")), 2),
                "fecha_dt":   dt,
                "fecha_str":  dt.strftime("%d/%m/%Y") if dt is not None else "",
                "comentario": com or (f"Pago en {mesa}" if mesa else "Pago en efectivo"),
            })
            n += 1
        log.info(f"pagos_efectivo.xlsx · {n} pagos en periodo · {n_fuera} fuera")
    else:
        log.info("pagos_efectivo.xlsx no existe — pagos efectivo no trazables")

    traz: dict[tuple[str, str], dict] = {}
    for key, lst in candidatos.items():
        lst.sort(key=lambda d: (d["fecha_dt"] is not None,
                                d["fecha_dt"] or pd.Timestamp.min))
        traz[key] = lst[-1]
    log.info(f"trazabilidad de pagos · {len(traz)} predios con pago localizado")
    return traz


def _enriquecer_salvados(salvados: list[dict], traz: dict[tuple[str, str], dict],
                         log: logging.Logger) -> None:
    """Agrega las 5 columnas de trazabilidad a cada SALVADO (in-place)."""
    n_no_id = 0
    for r in salvados:
        t = traz.get((r["mz"], r["lt"]))
        if t:
            r["tipo_pago"]  = t["tipo"]
            r["referencia"] = t["referencia"]
            r["monto_pago"] = t["monto"]
            r["fecha_pago"] = t["fecha_str"]
            r["comentario"] = t["comentario"]
        else:
            r["tipo_pago"]  = "NO_IDENTIFICADO"
            r["referencia"] = ""
            r["monto_pago"] = None
            r["fecha_pago"] = ""
            r["comentario"] = ""
            n_no_id += 1
    if n_no_id:
        log.warning(
            f"{n_no_id} salvados sin pago localizado en Yape/efectivo "
            f"-> TIPO_PAGO=NO_IDENTIFICADO (revisar manualmente)"
        )


# ── CLASIFICACIÓN ────────────────────────────────────────────────────────────
def _clasificar(
    snapshot: dict[tuple[str, str], dict],
    estado: dict[tuple[str, str], dict],
    fecha_seg: str,
    log: logging.Logger,
) -> tuple[list[dict], list[dict]]:
    """Devuelve (salvados, cortados).

    SALVADO ⟺ saldó todo menos convenio ⟺ deuda_no_convenio_ciclo2 ≤ 0
            ⟺ SALDO_ciclo2 ≤ CONVENIO_ciclo2.
    """
    salvados: list[dict] = []
    cortados: list[dict] = []
    no_encontrados: list[str] = []

    for (mz, lt), snap in snapshot.items():
        e = estado.get((mz, lt))
        if e is None:
            no_encontrados.append(f"{mz}-{lt}")
            continue

        saldo2    = e["saldo_actual"]
        convenio2 = e["convenio"]
        # Deuda no-convenio que queda sin pagar (convenio absorbe la base del saldo)
        deuda_nc_2 = round(max(0.0, max(0.0, saldo2) - max(0.0, convenio2)), 2)

        total_a_saldar = snap["total_a_saldar"]
        # Pagado hacia la deuda no-convenio = lo que tenía que saldar menos lo que queda
        total_pagado = round(min(total_a_saldar, max(0.0, total_a_saldar - deuda_nc_2)), 2)

        fila = {
            "mz":             mz,
            "lt":             lt,
            "nombre":         snap["nombre"],
            "deuda_multa":    snap["deuda_multa"],
            "penalidad":      config.PENALIDAD,
            "total_a_saldar": total_a_saldar,
            "total_pagado":   total_pagado,
            "fecha_seguimiento": fecha_seg,
        }

        if deuda_nc_2 <= config.TOL:
            fila["estado"] = "SALVADO"
            salvados.append(fila)
        else:
            # ARRASTRE_MULTA = penalidad/reconexión pendiente (espejo de 6_corte:
            # max(0, PENALIDAD_FINAL − pagado)). Es el cargo de reconexión que
            # arrastra a 2_planilla del mes siguiente.
            arrastre = max(0.0, round(config.PENALIDAD_FINAL - total_pagado, 2))
            fila["estado"]          = "CORTADO"
            fila["penalidad_final"] = config.PENALIDAD_FINAL
            fila["arrastre_multa"]  = arrastre
            cortados.append(fila)

    if no_encontrados:
        log.warning(
            f"{len(no_encontrados)} usuarios en lista_multas no aparecen en "
            f"planilla_cobrado ciclo 2: {', '.join(no_encontrados[:10])}"
            f"{'...' if len(no_encontrados) > 10 else ''}"
        )

    log.info(f"Clasificacion · SALVADOS={len(salvados)} · CORTADOS={len(cortados)}")
    return salvados, cortados


# ── EXPORT: pagaron_penalidad_multas.xlsx ────────────────────────────────────
_P_GRUPOS = [
    (1,  3,  "¿Quién es?",                 *GH_P_ID),
    (4,  6,  "¿Cuánto tenía que saldar?",   *GH_P_DEBE),
    (7,  9,  "¿Lo saldó?",                  *GH_P_SALDA),
    (10, 14, "¿Cómo pagó?",                 *GH_P_TRAZ),
]
_P_COLS = [
    (1,  "MZ",                *GH_P_ID,    6),
    (2,  "LT",                *GH_P_ID,    7),
    (3,  "NOMBRE",            *GH_P_ID,   28),
    (4,  "DEUDA_MULTA",       *GH_P_DEBE, 14),
    (5,  "PENALIDAD_MULTA",   *GH_P_DEBE, 16),
    (6,  "TOTAL_A_SALDAR",    *GH_P_DEBE, 16),
    (7,  "TOTAL_PAGADO",      *GH_P_SALDA, 14),
    (8,  "ESTADO",            *GH_P_SALDA, 12),
    (9,  "FECHA_SEGUIMIENTO", *GH_P_SALDA, 18),
    (10, "TIPO_PAGO",         *GH_P_TRAZ, 16),
    (11, "REFERENCIA",        *GH_P_TRAZ, 26),
    (12, "MONTO_PAGO",        *GH_P_TRAZ, 14),
    (13, "FECHA_PAGO",        *GH_P_TRAZ, 22),
    (14, "COMENTARIO",        *GH_P_TRAZ, 38),
]
_TX_TIPO = {
    "YAPE":            TX_TIPO_YAPE,
    "EFECTIVO":        TX_TIPO_EF,
    "NO_IDENTIFICADO": TX_TIPO_NO,
}

def _exportar_pagaron(salvados: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "salvados"
    ws.freeze_panes = "A3"

    for cs, ce, texto, bg, txt in _P_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for col, nombre, bg, txt, ancho in _P_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    MONEY = '"S/ "#,##0.00'
    for ri, r in enumerate(salvados, 3):
        _c(ws, ri, 1, r["mz"],     TD_P_ID, "1A5276", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_P_ID, "1A5276", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_P_ID, "333333", align="left")
        _c(ws, ri, 4, r["deuda_multa"], TD_P_DEBE, "92400E",
           mono=True, align="right", fmt=MONEY)
        _c(ws, ri, 5, r["penalidad"], TD_P_DEBE, "92400E",
           mono=True, align="right", fmt=MONEY)
        # TOTAL_A_SALDAR — columna clave, naranja oscuro
        _c(ws, ri, 6, r["total_a_saldar"], TD_P_TOTAL, "7C2D12",
           mono=True, align="right", bold=True, fmt=MONEY)
        _c(ws, ri, 7, r["total_pagado"], TD_P_SALDA, "064E3B",
           mono=True, align="right", bold=True, fmt=MONEY)
        _c(ws, ri, 8, r["estado"], TD_P_RES, "065F46",
           mono=True, align="center", bold=True)
        _c(ws, ri, 9, r["fecha_seguimiento"], TD_P_RES, "374151",
           mono=True, align="center")

        # ¿Cómo pagó? — trazabilidad
        tipo = r.get("tipo_pago", "NO_IDENTIFICADO")
        _c(ws, ri, 10, tipo, TD_P_TRAZ, _TX_TIPO.get(tipo, TX_TIPO_NO),
           mono=True, align="center", bold=True)
        _c(ws, ri, 11, r.get("referencia", ""), TD_P_TRAZ, "374151", align="left")
        monto = r.get("monto_pago")
        if monto is None:
            _c(ws, ri, 12, "", TD_P_TRAZ, "7D6608", align="right")
        else:
            _c(ws, ri, 12, monto, TD_P_TRAZ, "7D6608",
               mono=True, align="right", bold=True, fmt=MONEY)
        _c(ws, ri, 13, r.get("fecha_pago", ""), TD_P_TRAZ, "374151",
           mono=True, align="center")
        _c(ws, ri, 14, r.get("comentario", ""), TD_P_TRAZ, "6B7280", align="left")

        ws.row_dimensions[ri].height = 17

    wb.save(config.PAGARON_PATH)


# ── EXPORT: corte_fisico_multas.xlsx ─────────────────────────────────────────
_C_GRUPOS = [
    (1, 3, "¿A dónde va el operario?",         *GH_C_ID),
    (4, 4, "Deuda de multas",                   *GH_C_DEUDA),
    (5, 7, "Penalidad (escalada a S/40)",       *GH_C_CORTE),
    (8, 9, "Cierre",                            *GH_C_RES),
]
_C_COLS = [
    (1, "MZ",                     *GH_C_ID,     6),
    (2, "LT",                     *GH_C_ID,     7),
    (3, "NOMBRE",                 *GH_C_ID,    28),
    (4, "DEUDA_MULTA",            *GH_C_DEUDA, 14),
    (5, "PAGADO_PENALIDAD_MULTA", *GH_C_CORTE, 24),
    (6, "PENALIDAD_FINAL",        *GH_C_CORTE, 16),
    (7, "ARRASTRE_MULTA",         *GH_C_CORTE, 16),
    (8, "ESTADO",                 *GH_C_RES,   12),
    (9, "FECHA_SEGUIMIENTO",      *GH_C_RES,   18),
]

def _exportar_corte_fisico(cortados: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "cortes"
    ws.freeze_panes = "A3"

    for cs, ce, texto, bg, txt in _C_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for col, nombre, bg, txt, ancho in _C_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    MONEY = '"S/ "#,##0.00'
    for ri, r in enumerate(cortados, 3):
        _c(ws, ri, 1, r["mz"],     TD_C_ID, "1A5276", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_C_ID, "1A5276", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_C_ID, "333333", align="left")
        _c(ws, ri, 4, r["deuda_multa"], TD_C_DEUDA, "92400E",
           mono=True, align="right", fmt=MONEY)
        _c(ws, ri, 5, r["total_pagado"], TD_C_CORTE, "7F1D1D",
           mono=True, align="right", fmt=MONEY)
        _c(ws, ri, 6, r["penalidad_final"], TD_C_CORTE, "7F1D1D",
           mono=True, align="right", fmt=MONEY)
        # ARRASTRE_MULTA — columna clave, rojo
        _c(ws, ri, 7, r["arrastre_multa"], TD_C_KEY, "FFFFFF",
           mono=True, align="right", bold=True, size=10, fmt=MONEY)
        _c(ws, ri, 8, r["estado"], TD_C_RES, "7F1D1D",
           mono=True, align="center", bold=True)
        _c(ws, ri, 9, r["fecha_seguimiento"], TD_C_RES, "374151",
           mono=True, align="center")
        ws.row_dimensions[ri].height = 17

    wb.save(config.CORTE_FISICO_PATH)


# ── EXPORT: arrastre_multa_YYYY-MM.xlsx ──────────────────────────────────────
_A_GRUPOS = [
    (1, 3, "¿Quién es?",                   *GH_A_ID),
    (4, 4, "¿De qué mes?",                  *GH_A_ORIG),
    (5, 5, "Penalidad de multa pendiente",  *GH_A_ARR),
]
_A_COLS = [
    (1, "MZ",             *GH_A_ID,    6),
    (2, "LT",             *GH_A_ID,    7),
    (3, "NOMBRE",         *GH_A_ID,   28),
    (4, "MES_ORIGEN",     *GH_A_ORIG, 14),
    (5, "ARRASTRE_MULTA", *GH_A_ARR,  18),
]

def _exportar_arrastre(cortados: list[dict], mes_ano: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "arrastre"
    ws.freeze_panes = "A3"

    for cs, ce, texto, bg, txt in _A_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for col, nombre, bg, txt, ancho in _A_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    MONEY = '"S/ "#,##0.00'
    for ri, r in enumerate(cortados, 3):
        _c(ws, ri, 1, r["mz"],     TD_A_ID,   "1A5276", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_A_ID,   "1A5276", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_A_ID,   "333333", align="left")
        _c(ws, ri, 4, mes_ano,     TD_A_ORIG, "374151", mono=True, align="center")
        _c(ws, ri, 5, r["arrastre_multa"], TD_A_ARR, "5B21B6",
           mono=True, align="right", bold=True, size=10, fmt=MONEY)
        ws.row_dimensions[ri].height = 17

    path = config.arrastre_multa_path(mes_ano)
    wb.save(path)


# ── REGISTRO CORTES (append idempotente · shared) ────────────────────────────
def _leer_cortados_existentes(log: logging.Logger) -> set[tuple[str, str, str]]:
    """Set (MZ, LT, MES_INICIO_CORTE) ya registrados — para idempotencia."""
    p = config.REGISTRO_CORTES_PATH
    if not p.exists():
        return set()
    df = pd.read_excel(p, header=1, dtype=str).fillna("")
    df.columns = [str(c).strip().upper() for c in df.columns]
    existentes: set[tuple[str, str, str]] = set()
    for _, f in df.iterrows():
        estado = str(f.get("ESTADO", "")).strip().upper()
        if estado not in ("CORTADO", "REACTIVADO"):
            continue
        mz  = _norm_mz(f.get("MZ"))
        lt  = _norm_lt(f.get("LT"))
        mes = str(f.get("MES_INICIO_CORTE", "")).strip()
        if mz and lt and mes:
            existentes.add((mz, lt, mes))
    log.info(f"registro_cortes.xlsx (shared) · {len(existentes)} registros existentes")
    return existentes


def _appendar_registro_cortes(cortados: list[dict], mes_ano: str,
                              log: logging.Logger) -> int:
    """Agrega filas ESTADO=CORTADO (por multa) al archivo persistente compartido."""
    p = config.REGISTRO_CORTES_PATH
    p.parent.mkdir(parents=True, exist_ok=True)

    existentes = _leer_cortados_existentes(log)
    fecha_registro = datetime.now().strftime("%Y-%m-%d")

    if p.exists():
        wb = load_workbook(p)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Cortes"
        ws.freeze_panes = "A3"
        for cs, ce, texto, bg, txt in _R_GRUPOS:
            _gh(ws, 1, cs, ce, texto, bg, txt)
        for col, nombre, bg, txt, ancho in _R_COLS:
            _ch(ws, 2, col, nombre, bg, txt)
            _w(ws, col, ancho)
        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 22
        for idx, valor in enumerate(_R_EJEMPLO, start=1):
            _e(ws, 3, idx, valor)
        ws.row_dimensions[3].height = 18

    n_nuevos = 0
    next_row = ws.max_row + 1
    for r in cortados:
        clave = (r["mz"], r["lt"], mes_ano)
        if clave in existentes:
            continue
        ri = next_row + n_nuevos
        _c(ws, ri, 1, r["mz"],     TD_R_ID,  "1A5276", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_R_ID,  "1A5276", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_R_ID,  "333333", align="left")
        _c(ws, ri, 4, mes_ano,     TD_R_PER, "78350F", mono=True, align="center")
        _c(ws, ri, 5, "",          TD_R_PER, "78350F", align="center")
        _c(ws, ri, 6, "CORTADO",   TD_R_COR, TX_R_COR, bold=True, align="center")
        _c(ws, ri, 7, f"Corte por multa confirmado {mes_ano}", TD_R_TRZ, "4A235A", align="left")
        _c(ws, ri, 8, fecha_registro, TD_R_TRZ, "4A235A", mono=True, align="center")
        _c(ws, ri, 9, "seguimiento_multas.py", TD_R_TRZ, "4A235A", align="left")
        ws.row_dimensions[ri].height = 17
        n_nuevos += 1

    wb.save(p)
    return n_nuevos


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main() -> None:
    config.OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(config.OUTPUTS_DIR / "run.log", encoding="utf-8"),
        ],
        force=True,
    )
    log = logging.getLogger(__name__)
    log.info("seguimiento_multas.py · iniciando")

    print("=" * 60)
    print("  6b_corte_multas/seguimiento_multas.py")
    print("=" * 60)

    print("\n[1/5] Validando inputs...")
    snapshot, omitidos = _leer_lista_multas()
    log.info(
        f"lista_multas.xlsx · {len(snapshot)} elegibles (EJECUTAR_CORTE=SI) · "
        f"{omitidos} omitidos (bloqueados por reclamo)"
    )
    estado, mes_ano = _leer_planilla_cobrado()
    log.info(f"planilla_cobrado.xlsx ciclo 2 · {len(estado)} usuarios · MES_ANO={mes_ano}")
    if not mes_ano:
        raise ValueError("No se pudo detectar MES_ANO en planilla_cobrado")

    _verificar_planilla_sincronizada(estado, mes_ano, log)

    print("\n[2/5] Clasificando (SALVADO ⟺ saldó todo menos convenio)...")
    fecha_seg = datetime.now().strftime("%d/%m/%Y")
    salvados, cortados = _clasificar(snapshot, estado, fecha_seg, log)

    ancla = _ancla_periodo(log)
    traz = _cargar_trazabilidad_pagos(log, ancla=ancla)
    _enriquecer_salvados(salvados, traz, log)

    print(f"\n[3/5] Escribiendo outputs...")
    _exportar_pagaron(salvados)
    log.info(f"{config.PAGARON_PATH.name} -> {len(salvados)} salvados")

    _exportar_corte_fisico(cortados)
    log.info(f"{config.CORTE_FISICO_PATH.name} -> {len(cortados)} cortes fisicos")

    arrastre_path = config.arrastre_multa_path(mes_ano)
    _exportar_arrastre(cortados, mes_ano)
    log.info(f"{arrastre_path.name} -> {len(cortados)} arrastres")

    print(f"\n[4/5] Registrando {len(cortados)} cortes en shared/registro_cortes.xlsx...")
    n_nuevos = _appendar_registro_cortes(cortados, mes_ano, log)
    log.info(f"registro_cortes.xlsx · {n_nuevos} filas nuevas agregadas")

    print("\n[5/5] Resumen del ciclo")
    total = len(snapshot)
    print(f"  · Total en lista_multas (EJECUTAR=SI): {total}")
    print(f"  · Salvados (saldaron todo menos convenio): {len(salvados)}")
    print(f"  · Cortados (deuda no-convenio pendiente):  {len(cortados)}")
    if total > 0:
        print(f"  · Tasa de salvacion: {100.0 * len(salvados) / total:.1f}%")
    print(f"  · Nuevos en registro_cortes.xlsx: {n_nuevos}")

    print("\n" + "=" * 60)
    print(f"  seguimiento_multas.py completado")
    print(f"  -> {config.PAGARON_PATH}")
    print(f"  -> {config.CORTE_FISICO_PATH}  (entregar al operario)")
    print(f"  -> {arrastre_path}  (pasar a 2_planilla del mes siguiente)")
    print(f"  -> {config.REGISTRO_CORTES_PATH}  ({n_nuevos} nuevos CORTADO)")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
