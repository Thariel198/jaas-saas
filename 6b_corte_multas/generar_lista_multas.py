"""6b_corte_multas/generar_lista_multas.py — Genera lista_multas.xlsx (Día 0)

Espejo de 6_corte/generar_lista.py, pero penaliza la deuda de MULTA +
ACUERDOS_ASAMBLEA en lugar del saldo de consumo de agua.

Lee planilla_cobrado.xlsx de 5_cobranza y calcula DEUDA_MULTA por el MODELO DE
EXCEDENTE DE AGUA: el pago del mes cubre PRIMERO el cargo de agua (consumo +
mantenimiento + arrastre de agua del mes anterior). Lo que sobra (excedente) se
abona a la deuda de multa. Lo que quede sin cubrir de la multa dispara penalidad.

  CARGO_AGUA     = MES_ACTUAL + MANTENIMIENTO + MES_ANTERIOR
  EXCEDENTE_AGUA = max(0, PAGADO_MES − CARGO_AGUA)
  DEUDA_MULTA    = max(0, (MULTA + ACUERDOS_ASAMBLEA) − EXCEDENTE_AGUA)

CONVENIO se excluye del cálculo: es deuda grande de pago opcional por cuotas,
nunca dispara corte. La separación de qué saldo es de multa vs convenio es un
problema aparte que resuelve el arrastre consolidado (5_cobranza), no esta lista.

Elegibilidad: PAGÓ algo este mes (MONTO_YAPE o MONTO_EFECTIVO > 0) Y DEUDA_MULTA > 0.

  Política "hecha la ley, hecha la trampa": el corte por multa es SOLO para
  quien pagó PARCIAL — pagó algo pero no saldó su multa. El pago parcial es la
  evidencia de intención: puede pagar y eligió evadir la multa. Quien no pagó
  NADA este mes no entra aquí; su evasión total la captura el corte por AGUA
  (ley de 2 meses consecutivos sin pagar), reforzado por el cobro presencial.
  Saldar la multa (DEUDA_MULTA=0) es "completo" — convenio no cuenta.

Exclusión cruzada (única): lista_corte.xlsx de 6_corte · EJECUTAR_CORTE=SI → ya van
a corte por AGUA este ciclo. El S/40 de CORTE_RECONEXION es el MISMO corte físico,
no se cobra dos veces.

NO bloquean aquí (a diferencia de 6_corte): reclamos de 4b_reclamos y los estados
EXONERADO/CORTADO de registro_cortes. Pertenecen al mundo del corte por AGUA; la
multa es una sanción de asamblea independiente del consumo.

Cruza por (MZ, LT) con compromisos.xlsx: compromiso VIGENTE → EJECUTAR_CORTE=NO.
Es la ÚNICA exoneración del corte por multa (firmar compromiso de pago de la multa).

PHASE GATE: si aplicar_penalidad_multas.py ya corrió para este ciclo
(audit_penalidad_multas.xlsx con filas APLICADO para el MES_ANO actual),
este script aborta. Para corregir: editar lista_multas.xlsx directamente.

Uso:
    python generar_lista_multas.py
"""
import logging
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
import config

# ── PALETA ───────────────────────────────────────────────────────────────────
GH_ID    = ("EBF5FB", "1A5276")   # azul — ¿Quién es?
GH_MULTA = ("FED7AA", "7C2D12")   # naranja claro — ¿Qué debe en multas?
GH_AGUA  = ("D6D3D1", "292524")   # gris cálido — ¿Pagó algo que descuenta?
GH_REAL  = ("B45309", "FFFFFF")   # naranja oscuro — ¿Cuánto queda real?
GH_PEN   = ("B91C1C", "FFFFFF")   # rojo — Penalidad
GH_REC   = ("5B21B6", "FFFFFF")   # morado — ¿Se ejecuta el corte?

TD_ID    = "F4FAFF"
TD_MULTA = "FFF7ED"
TD_AGUA  = "F5F5F4"
TD_REAL  = "FDEBD0"
TD_PEN   = "FEE2E2"
TD_REC   = "F3E8FF"

# EJECUTAR_CORTE — color por valor
TD_SI = "D5F5E3"; TX_SI = "145A32"
TD_NO = "FADBD8"; TX_NO = "7B241C"

# ── LAYOUT ───────────────────────────────────────────────────────────────────
_LM_GRUPOS = [
    (1,  3,  "¿Quién es?",                 *GH_ID),
    (4,  5,  "¿Qué debe en multas?",        *GH_MULTA),
    (6,  8,  "¿Pagó algo que descuenta?",   *GH_AGUA),
    (9,  9,  "¿Cuánto queda real?",         *GH_REAL),
    (10, 11, "Penalidad",                   *GH_PEN),
    (12, 14, "¿Se ejecuta el corte?",       *GH_REC),
]
_LM_COLS = [
    (1,  "MZ",                 *GH_ID,    6),
    (2,  "LT",                 *GH_ID,    7),
    (3,  "NOMBRE",             *GH_ID,   28),
    (4,  "MULTA",              *GH_MULTA, 14),
    (5,  "ACUERDOS_ASAMBLEA",  *GH_MULTA, 20),
    (6,  "CARGO_AGUA",         *GH_AGUA,  14),
    (7,  "PAGADO_MES",         *GH_AGUA,  14),
    (8,  "EXCEDENTE_AGUA",     *GH_AGUA,  16),
    (9,  "DEUDA_MULTA",        *GH_REAL,  14),
    (10, "PENALIDAD_MULTA",    *GH_PEN,   16),
    (11, "TOTAL_A_PAGAR",      *GH_PEN,   16),
    (12, "ESTADO_COMPROMISO",  *GH_REC,   18),
    (13, "EJECUTAR_CORTE",     *GH_REC,   14),
    (14, "MOTIVO_NO_EJECUTAR", *GH_REC,   24),
]


# ── HELPERS DE ESTILO ────────────────────────────────────────────────────────
def _borde():
    b = Side(style="thin", color="CCCCCC")
    return Border(left=b, right=b, top=b, bottom=b)

def _c(ws, row, col, value=None, bg=None, txt="333333",
       bold=False, align="left", mono=False, size=9, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Consolas" if mono else "Arial",
                       size=size, bold=bold, color=txt)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _borde()
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    if fmt:
        c.number_format = fmt
    return c

def _gh(ws, row, cs, ce, texto, bg, txt):
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    c = ws.cell(row=row, column=cs, value=texto)
    c.font      = Font(name="Arial", size=8, bold=True, color=txt)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", start_color=bg)
    c.border    = _borde()

def _ch(ws, row, col, texto, bg, txt):
    _c(ws, row, col, texto, bg=bg, txt=txt, bold=True, align="center")

def _w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── HELPERS DE PARSING ───────────────────────────────────────────────────────
def _norm_mz(val) -> str:
    if val is None:
        return ""
    s = str(val).strip().upper()
    return "" if not s or s in ("NAN", "NONE") else s

def _norm_lt(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN"):
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.upper().replace(" ", "")

def _float(val) -> float:
    if val is None:
        return 0.0
    try:
        f = float(str(val).replace(",", ".").strip())
        return 0.0 if f != f else f   # NaN guard
    except (ValueError, TypeError):
        return 0.0


# ── MODELO DE EXCEDENTE DE AGUA ──────────────────────────────────────────────
def deuda_multa_excedente(pagado_mes: float, cargo_agua: float,
                          multa: float, acuerdos: float) -> tuple[float, float]:
    """El pago del mes cubre primero el cargo de agua; el excedente abona la multa.

      EXCEDENTE_AGUA = max(0, PAGADO_MES − CARGO_AGUA)
      DEUDA_MULTA    = max(0, (MULTA + ACUERDOS_ASAMBLEA) − EXCEDENTE_AGUA)

    Devuelve (EXCEDENTE_AGUA, DEUDA_MULTA).
    """
    excedente   = round(max(0.0, pagado_mes - cargo_agua), 2)
    deuda_bruta = max(0.0, multa) + max(0.0, acuerdos)
    deuda_multa = round(max(0.0, deuda_bruta - excedente), 2)
    return excedente, deuda_multa


# ── DETECCIÓN DE MES_ANO ─────────────────────────────────────────────────────
def _detectar_mes_ano(df: pd.DataFrame) -> str:
    if "MES_ANO" not in df.columns:
        raise ValueError("planilla_cobrado.xlsx · falta columna MES_ANO")
    valores = (df["MES_ANO"].dropna().astype(str).str.strip()
                                  .replace({"": None, "nan": None, "NaN": None})
                                  .dropna().unique().tolist())
    if not valores:
        raise ValueError("planilla_cobrado.xlsx · columna MES_ANO vacía")
    if len(valores) > 1:
        raise ValueError(f"planilla_cobrado.xlsx · MES_ANO inconsistente: {valores}")
    return valores[0]


# ── PHASE GATE ───────────────────────────────────────────────────────────────
def _verificar_phase_gate(mes_ano: str, log: logging.Logger) -> None:
    """Aborta si aplicar_penalidad_multas.py ya corrió para este ciclo."""
    p = config.AUDIT_PENALIDAD_PATH
    if not p.exists():
        return
    try:
        df = pd.read_excel(p, header=1, dtype=str).fillna("")
    except Exception:
        return
    df.columns = [str(c).strip().upper() for c in df.columns]
    if "MES_ANO" not in df.columns:
        return
    tiene_accion = "ACCION" in df.columns
    aplicados = sum(
        1 for _, f in df.iterrows()
        if str(f.get("MES_ANO", "")).strip() == mes_ano
        and (str(f.get("ACCION", "")).strip().upper() if tiene_accion else "APLICADO")
        in ("APLICADO", "")
    )
    if aplicados == 0:
        return
    log.error(
        f"PHASE GATE: ciclo {mes_ano} comprometido — "
        f"{aplicados} penalidad(es) APLICADA(s) en audit_penalidad_multas.xlsx"
    )
    print()
    print("=" * 60)
    print(f"  BLOQUEADO — ciclo {mes_ano} ya comprometido")
    print(f"  audit_penalidad_multas.xlsx registra {aplicados} penalidad(es) APLICADA(s)")
    print()
    print("  aplicar_penalidad_multas.py ya corrió → lista_multas no puede")
    print("  regenerarse para evitar revertir cargos ya cobrados.")
    print()
    print("  Para corregir la lista: editar lista_multas.xlsx directamente.")
    print("=" * 60)
    print()
    sys.exit(1)


# ── VALIDACIÓN DE INPUTS ─────────────────────────────────────────────────────
def _validar_input(log: logging.Logger) -> tuple[pd.DataFrame, str]:
    if not config.PLANILLA_COBRADO_PATH.exists():
        raise FileNotFoundError(
            f"Falta: {config.PLANILLA_COBRADO_PATH}\n"
            f"  -> Correr 5_cobranza/main.py primero"
        )
    df = pd.read_excel(config.PLANILLA_COBRADO_PATH, header=1)
    df.columns = [str(c).strip().upper() for c in df.columns]

    requeridas = {"MZ", "LT", "NOMBRE", "MES_ANO",
                  "MES_ACTUAL", "MANTENIMIENTO", "MES_ANTERIOR",
                  "MULTA", "ACUERDOS_ASAMBLEA",
                  "MONTO_YAPE", "MONTO_EFECTIVO"}
    faltantes = requeridas - set(df.columns)
    if faltantes:
        raise ValueError(
            f"planilla_cobrado.xlsx · columnas faltantes: {sorted(faltantes)}\n"
            f"  -> MES_ACTUAL/MANTENIMIENTO/MES_ANTERIOR/MULTA/ACUERDOS_ASAMBLEA/"
            f"MONTO_YAPE/MONTO_EFECTIVO deben venir expuestos por 5_cobranza. "
            f"Re-correr 5_cobranza/main.py."
        )
    mes_ano = _detectar_mes_ano(df)
    log.info(f"planilla_cobrado.xlsx leida · {len(df)} filas · ciclo {mes_ano}")
    return df, mes_ano


# ── CARGA DE COMPROMISOS VIGENTES (cruza por MZ+LT) ──────────────────────────
def _cargar_compromisos(log: logging.Logger) -> dict[tuple[str, str], str]:
    """Devuelve {(mz, lt) -> FECHA_LIMITE} de compromisos VIGENTES.

    Un compromiso de pago VIGENTE exonera del corte por multa (EJECUTAR_CORTE=NO).
    Es la ÚNICA exoneración aquí: reclamos y EXONERADO/CORTADO de registro_cortes
    son del mundo agua y no aplican a la multa.
    """
    p = config.COMPROMISOS_PATH
    if not p.exists():
        log.info("compromisos.xlsx no existe — 0 exonerados (correr generar_compromisos.py)")
        return {}
    df = pd.read_excel(p, header=1, dtype=str).fillna("")
    df.columns = [str(c).strip().upper() for c in df.columns]
    if "ESTADO" not in df.columns:
        log.warning("compromisos.xlsx · falta columna ESTADO — ignorando archivo")
        return {}
    mapa: dict[tuple[str, str], str] = {}
    for _, f in df.iterrows():
        if str(f.get("ESTADO", "")).strip().upper() != "VIGENTE":
            continue
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if mz and lt:
            mapa[(mz, lt)] = str(f.get("FECHA_LIMITE", "")).strip()
    log.info(f"compromisos.xlsx · {len(mapa)} compromisos VIGENTES (exoneran corte por multa)")
    return mapa


# ── CARGA DE CORTE DE AGUA (6_corte/lista_corte) ─────────────────────────────
def _cargar_corte_agua(log: logging.Logger) -> set[tuple[str, str]]:
    """Set (MZ, LT) que ya van a corte por AGUA este ciclo (EJECUTAR_CORTE=SI en
    lista_corte de 6_corte). Se excluyen: un solo corte físico, un solo cargo."""
    p = config.LISTA_CORTE_AGUA_PATH
    if not p.exists():
        log.info("6_corte/lista_corte.xlsx no existe — sin exclusión por corte de agua")
        return set()
    df = pd.read_excel(p, header=1, dtype=str).fillna("")
    df.columns = [str(c).strip().upper() for c in df.columns]
    if "EJECUTAR_CORTE" not in df.columns:
        log.warning("6_corte/lista_corte.xlsx · sin columna EJECUTAR_CORTE — no se excluye")
        return set()
    s: set[tuple[str, str]] = set()
    for _, f in df.iterrows():
        if str(f.get("EJECUTAR_CORTE", "")).strip().upper() != "SI":
            continue
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if mz and lt:
            s.add((mz, lt))
    log.info(f"6_corte/lista_corte.xlsx · {len(s)} predios ya en corte de agua (excluidos)")
    return s


# ── FILTRO ───────────────────────────────────────────────────────────────────
def _filtrar_multas(df: pd.DataFrame, compromisos: dict,
                    corte_agua: set, log: logging.Logger) -> list[dict]:
    lista: list[dict] = []
    n_compromisos = 0
    n_excl_agua = 0
    n_excl_no_pago = 0   # tenía deuda de multa pero NO pagó nada → difiere a corte de agua
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        if (mz, lt) in corte_agua:
            n_excl_agua += 1
            continue

        acuerdos = round(_float(f.get("ACUERDOS_ASAMBLEA")), 2)
        multa    = round(_float(f.get("MULTA")), 2)

        # CARGO_AGUA = lo que el pago debe cubrir antes de tocar la multa:
        # consumo del mes + mantenimiento + arrastre de agua del mes anterior.
        cargo_agua = round(
            _float(f.get("MES_ACTUAL"))
            + _float(f.get("MANTENIMIENTO"))
            + _float(f.get("MES_ANTERIOR")), 2)

        pago_yape = round(_float(f.get("MONTO_YAPE")), 2)
        pago_efec = round(_float(f.get("MONTO_EFECTIVO")), 2)
        pagado_mes = round(pago_yape + pago_efec, 2)

        excedente, deuda_multa = deuda_multa_excedente(
            pagado_mes, cargo_agua, multa, acuerdos)

        if deuda_multa <= config.TOL:
            continue   # el excedente cubrió toda la multa → "completo" → no elegible

        # Pago parcial = pagó algo este mes pero NO saldó su multa. Es la
        # evidencia de intención: puede pagar y eligió evadir la multa. Quien no
        # pagó NADA no entra — su evasión total la captura el corte por AGUA.
        if pagado_mes <= config.TOL:
            n_excl_no_pago += 1
            continue   # no pagó nada → difiere a corte de agua (ley de 2 meses)

        total_a_pagar = round(deuda_multa + config.PENALIDAD, 2)

        fecha_limite = compromisos.get((mz, lt))
        if fecha_limite is not None:
            estado_compromiso = "VIGENTE"
            ejecutar          = "NO"
            motivo            = f"Compromiso de pago {fecha_limite}".strip()
            n_compromisos += 1
        else:
            estado_compromiso = "SIN_COMPROMISO"
            ejecutar          = "SI"
            motivo            = ""

        lista.append({
            "mz":                mz,
            "lt":                lt,
            "nombre":            str(f.get("NOMBRE", "")).strip(),
            "multa":             multa,
            "acuerdos":          acuerdos,
            "cargo_agua":        cargo_agua,
            "pagado_mes":        pagado_mes,
            "excedente":         excedente,
            "deuda_multa":       deuda_multa,
            "total_a_pagar":     total_a_pagar,
            "estado_compromiso": estado_compromiso,
            "ejecutar_corte":    ejecutar,
            "motivo":            motivo,
        })

    n_si = sum(1 for r in lista if r["ejecutar_corte"] == "SI")
    log.info(f"Excluidos · corte_agua={n_excl_agua} · "
             f"sin_pago(difieren a corte agua)={n_excl_no_pago}")
    log.info(f"Elegibles · {len(lista)} usuarios (PAGÓ PARCIAL Y DEUDA_MULTA > {config.TOL:.3f})")
    log.info(f"  · EJECUTAR_CORTE = SI · {n_si}")
    log.info(f"  · EJECUTAR_CORTE = NO · {n_compromisos} (compromiso de pago VIGENTE)")
    return lista


# ── EXPORT ───────────────────────────────────────────────────────────────────
def _exportar(lista: list[dict]) -> None:
    config.OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "lista_multas"
    ws.freeze_panes = "A3"

    for cs, ce, texto, bg, txt in _LM_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for col, nombre, bg, txt, ancho in _LM_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 22

    MONEY = '"S/ "#,##0.00'
    for ri, r in enumerate(lista, 3):
        _c(ws, ri, 1, r["mz"],     TD_ID, "1A5276", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_ID, "1A5276", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_ID, "333333", align="left")
        _c(ws, ri, 4, r["multa"],    TD_MULTA, "92400E", mono=True, align="right", fmt=MONEY)
        _c(ws, ri, 5, r["acuerdos"], TD_MULTA, "92400E", mono=True, align="right", fmt=MONEY)
        # ¿Pagó algo que descuenta? — CARGO_AGUA, PAGADO_MES, EXCEDENTE_AGUA
        _c(ws, ri, 6, r["cargo_agua"], TD_AGUA, "44403C", mono=True, align="right", fmt=MONEY)
        _c(ws, ri, 7, r["pagado_mes"], TD_AGUA, "44403C", mono=True, align="right", fmt=MONEY)
        _c(ws, ri, 8, r["excedente"],  TD_AGUA, "44403C", mono=True, align="right", fmt=MONEY)
        # DEUDA_MULTA — multa+acuerdos menos el excedente de agua, columna clave
        _c(ws, ri, 9, r["deuda_multa"], TD_REAL, "7C2D12",
           mono=True, align="right", bold=True, fmt=MONEY)
        _c(ws, ri, 10, config.PENALIDAD, TD_PEN, "7F1D1D",
           mono=True, align="right", bold=True, fmt=MONEY)
        _c(ws, ri, 11, r["total_a_pagar"], TD_PEN, "7F1D1D",
           mono=True, align="right", bold=True, size=10, fmt=MONEY)

        # ¿Se ejecuta el corte?
        _c(ws, ri, 12, r["estado_compromiso"], TD_REC, "4A235A", align="center")
        if r["ejecutar_corte"] == "SI":
            _c(ws, ri, 13, "SI", TD_SI, TX_SI, bold=True, align="center")
        else:
            _c(ws, ri, 13, "NO", TD_NO, TX_NO, bold=True, align="center")
        _c(ws, ri, 14, r["motivo"], TD_REC, "4A235A", align="left")

        ws.row_dimensions[ri].height = 17

    wb.save(config.LISTA_MULTAS_PATH)


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main() -> None:
    config.OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(config.OUTPUTS_DIR / "run.log", encoding="utf-8"),
        ],
        force=True,
    )
    log = logging.getLogger(__name__)
    log.info("generar_lista_multas.py · iniciando")

    print("=" * 60)
    print("  6b_corte_multas/generar_lista_multas.py")
    print("=" * 60)

    print("\n[1/5] Validando inputs...")
    df, mes_ano = _validar_input(log)
    _verificar_phase_gate(mes_ano, log)

    print("\n[2/5] Cargando corte de agua de 6_corte (única exclusión cruzada)...")
    corte_agua = _cargar_corte_agua(log)

    print("\n[3/5] Cargando compromisos de pago VIGENTES (exoneran corte por multa)...")
    compromisos = _cargar_compromisos(log)

    print("\n[4/5] Calculando DEUDA_MULTA por excedente de agua y filtrando elegibles...")
    lista = _filtrar_multas(df, compromisos, corte_agua, log)

    print("\n[5/5] Exportando lista_multas.xlsx...")
    _exportar(lista)
    log.info(f"{config.LISTA_MULTAS_PATH.name} -> {len(lista)} usuarios")

    n_si  = sum(1 for r in lista if r["ejecutar_corte"] == "SI")
    n_no  = len(lista) - n_si

    print("\n" + "=" * 60)
    print(f"  generar_lista_multas.py completado")
    print(f"  -> {config.LISTA_MULTAS_PATH}")
    print(f"  -> {len(lista)} en lista · EJECUTAR=SI: {n_si} · NO: {n_no}")
    if corte_agua:
        print(f"  -> {len(corte_agua)} excluidos por corte de agua (mismo S/40 físico)")
    if compromisos:
        print(f"  -> {len(compromisos)} con compromiso VIGENTE (EJECUTAR=NO, no se cortan)")
    if n_si > 0:
        print(f"\n  Siguiente paso: python aplicar_penalidad_multas.py")
        print(f"    (solo procesa filas con EJECUTAR_CORTE = SI)")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
