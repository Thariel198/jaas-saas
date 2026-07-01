"""6b_corte_multas/generar_compromisos.py — Fuente de verdad de compromisos de pago

Promueve el tag suelto "compromiso" de la columna COMENTARIO de pagos_efectivo.xlsx
a un archivo de primera clase: compromisos.xlsx. Quien firmó compromiso de pago de
su multa se exonera del corte por multa (generar_lista_multas lo marca EJECUTAR=NO).

Solo aplica al corte por MULTA. No exonera del corte por agua (6_corte) ni depende
de reclamos — la multa es una sanción de asamblea independiente del consumo.

APPEND-IDEMPOTENTE (clave MZ, LT): re-correr no pisa registros existentes. Los
nuevos entran como VIGENTE; los que ya están conservan su ESTADO (incl. el que el
seguimiento haya movido a CUMPLIDO/INCUMPLIDO). Mismo patrón que registro_cortes.

Fuente: pagos_efectivo.xlsx · COMENTARIO contiene "compromi" (cubre "compromiso"
y "compromizo"). NOMBRE y MONTO_COMPROMETIDO se cruzan con la planilla por (MZ, LT).

Uso:
    python generar_compromisos.py
"""
import logging
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
import config

SOURCE = "generar_compromisos.py"

# ── PALETA (espejo de docs/formato_compromisos.html) ─────────────────────────
GH_ID   = ("EBF5FB", "1A5276")   # azul — ¿Quién firmó?
GH_COMP = ("FED7AA", "7C2D12")   # naranja — ¿Qué se comprometió?
GH_EST  = ("065F46", "FFFFFF")   # verde — Estado
GH_TRAZ = ("FEF9E7", "7D6608")   # ámbar — ¿Dónde y cuándo se registró?

TD_ID   = "F4FAFF"
TD_COMP = "FFF7ED"
TD_KEY  = "B45309"   # MONTO_COMPROMETIDO — naranja oscuro, texto blanco
TD_EST  = "D1FAE5"
TD_TRAZ = "FFFDF5"

_C_GRUPOS = [
    (1,  3,  "¿Quién firmó?",                 *GH_ID),
    (4,  8,  "¿Qué se comprometió?",          *GH_COMP),
    (9,  9,  "Estado",                        *GH_EST),
    (10, 13, "¿Dónde y cuándo se registró?",  *GH_TRAZ),
]
_C_COLS = [
    (1,  "MZ",                 *GH_ID,    6),
    (2,  "LT",                 *GH_ID,    7),
    (3,  "NOMBRE",             *GH_ID,   28),
    (4,  "FECHA_FIRMA",        *GH_COMP, 14),
    (5,  "MONTO_PAGO_FIRMA",   *GH_COMP, 16),
    (6,  "MONTO_COMPROMETIDO", *GH_COMP, 18),
    (7,  "FECHA_LIMITE",       *GH_COMP, 14),
    (8,  "TIPO",               *GH_COMP, 22),
    (9,  "ESTADO",             *GH_EST,  12),
    (10, "MESA",               *GH_TRAZ, 10),
    (11, "COBRADOR",           *GH_TRAZ, 20),
    (12, "FECHA_REGISTRO",     *GH_TRAZ, 16),
    (13, "SOURCE",             *GH_TRAZ, 24),
]


# ── HELPERS DE ESTILO ────────────────────────────────────────────────────────
def _borde():
    b = Side(style="thin", color="CCCCCC")
    return Border(left=b, right=b, top=b, bottom=b)

def _c(ws, row, col, value=None, bg=None, txt="333333",
       bold=False, align="left", mono=False, size=9, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Consolas" if mono else "Arial",
                       size=size, bold=bold, color=txt)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _borde()
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    if fmt:
        c.number_format = fmt
    return c

def _gh(ws, row, cs, ce, texto, bg, txt):
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    c = ws.cell(row=row, column=cs, value=texto)
    c.font      = Font(name="Arial", size=8, bold=True, color=txt)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", start_color=bg)
    c.border    = _borde()

def _ch(ws, row, col, texto, bg, txt):
    _c(ws, row, col, texto, bg=bg, txt=txt, bold=True, align="center")

def _w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── PARSING ──────────────────────────────────────────────────────────────────
def _norm_mz(val) -> str:
    if val is None:
        return ""
    s = str(val).strip().upper()
    return "" if not s or s in ("NAN", "NONE") else s

def _norm_lt(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN"):
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.upper().replace(" ", "")

def _float(val) -> float:
    if val is None:
        return 0.0
    try:
        f = float(str(val).replace(",", ".").strip())
        return 0.0 if f != f else f
    except (ValueError, TypeError):
        return 0.0

def _fecha_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, pd.Timestamp):
        return "" if pd.isna(val) else val.strftime("%d/%m/%Y")
    s = str(val).strip()
    if not s or s.lower() in ("nat", "nan", "none"):
        return ""
    dt = pd.to_datetime(s, errors="coerce")
    return s if pd.isna(dt) else dt.strftime("%d/%m/%Y")


# ── CARGA DE DEUDA/NOMBRE DESDE PLANILLA (cruce por MZ, LT) ───────────────────
def _cargar_planilla_map(log: logging.Logger) -> dict[tuple[str, str], dict]:
    """{(mz, lt) -> {nombre, deuda_multa}} desde planilla_cobrado."""
    p = config.PLANILLA_COBRADO_PATH
    if not p.exists():
        log.warning(f"{p.name} no existe — NOMBRE y MONTO_COMPROMETIDO quedarán vacíos")
        return {}
    df = pd.read_excel(p, header=1)
    df.columns = [str(c).strip().upper() for c in df.columns]
    mapa: dict[tuple[str, str], dict] = {}
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        deuda = round(_float(f.get("MULTA")) + _float(f.get("ACUERDOS_ASAMBLEA")), 2)
        mapa[(mz, lt)] = {"nombre": str(f.get("NOMBRE", "")).strip(), "deuda_multa": deuda}
    log.info(f"{p.name} · {len(mapa)} predios para cruce de NOMBRE/deuda")
    return mapa


# ── EXTRACCIÓN DE COMPROMISOS DESDE pagos_efectivo ───────────────────────────
def _extraer_compromisos(plan_map: dict, log: logging.Logger) -> dict[tuple[str, str], dict]:
    """Filtra pagos_efectivo por tag 'compromi' y arma un registro por predio
    (el compromiso más reciente si hay varios pagos)."""
    p = config.PAGOS_EFECTIVO_OUT_PATH
    if not p.exists():
        raise FileNotFoundError(
            f"Falta: {p}\n  -> Correr 4_pagos/efectivo primero (genera pagos_efectivo.xlsx)"
        )
    df = pd.read_excel(p, header=1)
    df.columns = [str(c).strip().upper() for c in df.columns]
    if "COMENTARIO" not in df.columns:
        raise ValueError("pagos_efectivo.xlsx · falta columna COMENTARIO")

    mask = df["COMENTARIO"].astype(str).str.contains(config.TAG_COMPROMISO, case=False, na=False)
    df = df[mask]
    log.info(f"pagos_efectivo.xlsx · {len(df)} pagos con tag '{config.TAG_COMPROMISO}'")

    encontrados: dict[tuple[str, str], dict] = {}
    sin_planilla = 0
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if not mz or not lt:
            continue
        com = str(f.get("COMENTARIO", "")).strip()
        tipo = ("COMPROMISO_EXONERACION"
                if config.TAG_EXONERACION in com.lower() else "COMPROMISO")
        fecha_dt = pd.to_datetime(f.get("FECHA"), errors="coerce")
        info = plan_map.get((mz, lt), {})
        if (mz, lt) not in plan_map:
            sin_planilla += 1
        reg = {
            "mz":                 mz,
            "lt":                 lt,
            "nombre":             info.get("nombre", ""),
            "fecha_firma":        _fecha_str(f.get("FECHA")),
            "monto_pago_firma":   round(_float(f.get("MONTO")), 2),
            "monto_comprometido": info.get("deuda_multa", 0.0),
            "tipo":               tipo,
            "mesa":               str(f.get("MESA", "")).strip(),
            "cobrador":           str(f.get("COBRADOR", "")).strip(),
            "_fecha_dt":          fecha_dt,
        }
        prev = encontrados.get((mz, lt))
        # conserva el compromiso más reciente por fecha de firma
        if prev is None or (pd.notna(fecha_dt) and (
                pd.isna(prev["_fecha_dt"]) or fecha_dt >= prev["_fecha_dt"])):
            encontrados[(mz, lt)] = reg

    if sin_planilla:
        log.warning(f"{sin_planilla} compromisos sin match en planilla (NOMBRE/deuda vacíos)")
    log.info(f"Compromisos únicos extraídos · {len(encontrados)} predios")
    return encontrados


# ── ESTADO EXISTENTE (idempotencia) ──────────────────────────────────────────
def _leer_existentes(log: logging.Logger) -> set[tuple[str, str]]:
    """Set (MZ, LT) ya presentes en compromisos.xlsx — no se vuelven a escribir."""
    p = config.COMPROMISOS_PATH
    if not p.exists():
        return set()
    df = pd.read_excel(p, header=1, dtype=str).fillna("")
    df.columns = [str(c).strip().upper() for c in df.columns]
    existentes: set[tuple[str, str]] = set()
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ"))
        lt = _norm_lt(f.get("LT"))
        if mz and lt:
            existentes.add((mz, lt))
    log.info(f"compromisos.xlsx · {len(existentes)} registros existentes (se preservan)")
    return existentes


# ── APPEND ───────────────────────────────────────────────────────────────────
def _appendar(nuevos: list[dict], log: logging.Logger) -> int:
    p = config.COMPROMISOS_PATH
    p.parent.mkdir(parents=True, exist_ok=True)
    fecha_reg = datetime.now().strftime("%d/%m/%Y")

    if p.exists():
        wb = load_workbook(p)
        ws = wb.active
        next_row = max(ws.max_row + 1, 3)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Compromisos"
        ws.freeze_panes = "A3"
        for cs, ce, texto, bg, txt in _C_GRUPOS:
            _gh(ws, 1, cs, ce, texto, bg, txt)
        for col, nombre, bg, txt, ancho in _C_COLS:
            _ch(ws, 2, col, nombre, bg, txt)
            _w(ws, col, ancho)
        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 22
        next_row = 3

    MONEY = '"S/ "#,##0.00'
    fecha_lim = pd.to_datetime(config.COMPROMISO_FECHA_LIMITE, errors="coerce")
    fecha_lim_str = (fecha_lim.strftime("%d/%m/%Y")
                     if pd.notna(fecha_lim) else config.COMPROMISO_FECHA_LIMITE)

    for i, r in enumerate(nuevos):
        ri = next_row + i
        _c(ws, ri, 1, r["mz"],     TD_ID, "1A5276", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_ID, "1A5276", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_ID, "333333", align="left")
        _c(ws, ri, 4, r["fecha_firma"], TD_COMP, "92400E", mono=True, align="center")
        _c(ws, ri, 5, r["monto_pago_firma"], TD_COMP, "92400E",
           mono=True, align="right", fmt=MONEY)
        # MONTO_COMPROMETIDO — base del seguimiento, naranja oscuro
        _c(ws, ri, 6, r["monto_comprometido"], TD_KEY, "FFFFFF",
           mono=True, align="right", bold=True, fmt=MONEY)
        _c(ws, ri, 7, fecha_lim_str, TD_COMP, "92400E", mono=True, align="center")
        _c(ws, ri, 8, r["tipo"], TD_COMP, "7C2D12", align="center")
        _c(ws, ri, 9, "VIGENTE", TD_EST, "064E3B", bold=True, align="center")
        _c(ws, ri, 10, r["mesa"],     TD_TRAZ, "374151", align="center")
        _c(ws, ri, 11, r["cobrador"], TD_TRAZ, "374151", align="left")
        _c(ws, ri, 12, fecha_reg,     TD_TRAZ, "374151", mono=True, align="center")
        _c(ws, ri, 13, SOURCE,        TD_TRAZ, "9CA3AF", align="left")
        ws.row_dimensions[ri].height = 17

    wb.save(p)
    return len(nuevos)


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
        force=True,
    )
    log = logging.getLogger(__name__)

    print("=" * 60)
    print("  6b_corte_multas/generar_compromisos.py")
    print("=" * 60)

    print("\n[1/4] Cruzando NOMBRE/deuda con la planilla...")
    plan_map = _cargar_planilla_map(log)

    print("\n[2/4] Extrayendo compromisos de pagos_efectivo (tag 'compromi')...")
    extraidos = _extraer_compromisos(plan_map, log)

    print("\n[3/4] Filtrando contra compromisos ya registrados (idempotencia)...")
    existentes = _leer_existentes(log)
    nuevos = [r for (k), r in extraidos.items() if k not in existentes]
    nuevos.sort(key=lambda r: (r["mz"], r["lt"]))
    log.info(f"Nuevos a registrar · {len(nuevos)} (de {len(extraidos)} extraídos)")

    print("\n[4/4] Append a compromisos.xlsx...")
    n = _appendar(nuevos, log)

    n_exo = sum(1 for r in nuevos if r["tipo"] == "COMPROMISO_EXONERACION")
    print("\n" + "=" * 60)
    print(f"  generar_compromisos.py completado")
    print(f"  -> {config.COMPROMISOS_PATH}")
    print(f"  -> {len(existentes)} preservados · {n} nuevos agregados (VIGENTE)")
    if n_exo:
        print(f"  -> {n_exo} marcados COMPROMISO_EXONERACION")
    print(f"  -> Total en archivo: {len(existentes) + n}")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
