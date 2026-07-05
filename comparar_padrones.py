import openpyxl
import unicodedata

def norm(s):
    if s is None: return ''
    s = str(s).strip().upper()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return ' '.join(s.split())

# registro_operario_acumulado
wb1 = openpyxl.load_workbook(
    r'C:\Users\wilde\PycharmProjects\jass_system\1_lecturas\inputs\registro_operario_acumulado.xlsx',
    read_only=True, data_only=True)
ws1 = wb1['Acumulada']
mapa_operario = {}
for r in ws1.iter_rows(min_row=2, values_only=True):
    mz, lt, nombre = r[0], r[1], r[2]
    if mz is not None or lt is not None:
        key = (norm(str(mz)), norm(str(lt)))
        mapa_operario[key] = norm(nombre)
wb1.close()

# padron_reconciliado
wb2 = openpyxl.load_workbook(
    r'C:\Users\wilde\PycharmProjects\jass_system\0_padron\02_matching\outputs\padron_reconciliado.xlsx',
    read_only=True, data_only=True)
ws2 = wb2['cobranza']
mapa_padron = {}
for r in ws2.iter_rows(min_row=2, values_only=True):
    nombre, mz, lt = r[1], r[2], r[3]
    if mz is not None or lt is not None:
        key = (norm(str(mz)), norm(str(lt)))
        mapa_padron[key] = norm(nombre)
wb2.close()

keys_op  = set(mapa_operario.keys())
keys_pad = set(mapa_padron.keys())
solo_operario = keys_op - keys_pad
solo_padron   = keys_pad - keys_op
comunes       = keys_op & keys_pad

dif_nombre = []
for k in sorted(comunes):
    n1 = mapa_operario[k]
    n2 = mapa_padron[k]
    if n1 != n2:
        dif_nombre.append((k[0], k[1], n1, n2))

print(f"Registros en operario : {len(mapa_operario)}")
print(f"Registros en padron   : {len(mapa_padron)}")
print(f"Claves comunes        : {len(comunes)}")
print()

if solo_operario:
    print(f"=== SOLO en operario ({len(solo_operario)}) ===")
    for k in sorted(solo_operario):
        print(f"  MZ={k[0]}  LT={k[1]}  nombre={mapa_operario[k]}")
else:
    print("Sin claves exclusivas en operario.")

print()
if solo_padron:
    print(f"=== SOLO en padron ({len(solo_padron)}) ===")
    for k in sorted(solo_padron):
        print(f"  MZ={k[0]}  LT={k[1]}  nombre={mapa_padron[k]}")
else:
    print("Sin claves exclusivas en padron.")

print()
if dif_nombre:
    print(f"=== NOMBRE diferente ({len(dif_nombre)}) ===")
    print(f"  {'MZ':<6} {'LT':<6} {'OPERARIO':<40} PADRON")
    for mz, lt, n1, n2 in dif_nombre:
        print(f"  {mz:<6} {lt:<6} {n1:<40} {n2}")
else:
    print("Todos los nombres coinciden en claves comunes.")
