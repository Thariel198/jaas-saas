import logging
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
ROOT               = Path(__file__).parent
COBRANZA_DIR       = ROOT.parent / "04_cobranza"
OUTPUTS_DIR        = ROOT / "outputs"

LISTA_CORTE_PATH     = COBRANZA_DIR / "outputs" / "lista_corte.xlsx"
COBRANZA_FINAL_PATH  = COBRANZA_DIR / "outputs" / "cobranza_final.xlsx"
PAGOS_YAPE_PATH      = COBRANZA_DIR / "inputs" / "pagos_yape" / "pagos_yape_tepago.xlsx"
PAGOS_EFEC_PATH      = COBRANZA_DIR / "inputs" / "pagos_efectivo" / "pagos_efectivo.xlsx"
CORTE_ACUMULADO_PATH = OUTPUTS_DIR / "corte_acumulado.xlsx"

PENALIDAD  = 20.0
RECONEXION = 40.0
TOLERANCIA = 0.005
MES_ACTUAL = datetime.now().strftime("%Y-%m")

# ── PALETA ────────────────────────────────────────────────────────────────────
GH_QUIEN  = ("F4ECF7", "5B21B6")
GH_DEUDA  = ("FEF9E7", "7D6608")
GH_PAGO   = ("FEF3E8", "7C3003")
GH_QUEDA  = ("E1F5EE", "085041")
GH_CORTE  = ("FEF2F2", "991B1B")
GH_CUANDO = ("F3F4F6", "374151")

TD_QUIEN  = "FAF5FF"
TD_DEUDA  = "FEFCE8"
TD_PAGO   = "FEF6EE"
TD_QUEDA  = "F0FFF8"
TD_CORTE  = "FEF2F2"
TD_CUANDO = "F9FAFB"

ESTADO_BG_PP  = {"CANCELADO": "E1F5EE", "EXCESO": "EFF6FF", "PARCIAL": "FAEEDA"}
ESTADO_TXT_PP = {"CANCELADO": "085041", "EXCESO": "1D4ED8", "PARCIAL": "854F0B"}
ESTADO_BG_CA  = {"CORTADO": "FEF2F2", "RECONECTADO": "E1F5EE"}
ESTADO_TXT_CA = {"CORTADO": "991B1B", "RECONECTADO": "085041"}

# ── LOGGING ───────────────────────────────────────────────────────────────────
def _init_logging():
    OUTPUTS_DIR.mkdir(exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(OUTPUTS_DIR / "run.log", encoding="utf-8"),
        ],
    )

log = logging.getLogger(__name__)

# ── ESTILO ────────────────────────────────────────────────────────────────────
def _borde():
    b = Side(style="thin", color="CCCCCC")
    return Border(left=b, right=b, top=b, bottom=b)

def _c(ws, row, col, value=None, bg=None, txt="333333",
       bold=False, align="left", mono=False, size=9):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Consolas" if mono else "Arial",
                       size=size, bold=bold, color=txt)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _borde()
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    return c

def _gh(ws, row, cs, ce, texto, bg, txt):
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    c = ws.cell(row=row, column=cs, value=texto)
    c.font      = Font(name="Arial", size=8, bold=True, color=txt)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", start_color=bg)
    c.border    = _borde()

def _ch(ws, row, col, texto, bg, txt):
    _c(ws, row, col, texto, bg=bg, txt=txt, bold=True, align="center")

def _sep(ws, col, row_end):
    letra = get_column_letter(col)
    ws.column_dimensions[letra].width = 0.8
    b_sep = Border(left=Side(style="thin", color="D1D5DB"),
                   right=Side(style="thin", color="D1D5DB"))
    for r in range(1, row_end + 1):
        c = ws.cell(row=r, column=col)
        c.fill   = PatternFill("solid", start_color="F3F4F6")
        c.border = b_sep

def _w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ── UTILIDADES ────────────────────────────────────────────────────────────────
def _norm_lt(val) -> str:
    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN"):
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.upper()

def _key(mz, lt) -> str:
    mz = str(mz).strip().upper()
    lt = _norm_lt(lt)
    return f"{mz}-{lt}" if mz and lt else ""

def _float(val) -> float:
    try:
        return float(str(val).replace(",", ".").strip())
    except (ValueError, TypeError):
        return 0.0

def _col_names(df) -> list:
    return [str(c).strip().upper() if str(c) not in ("NAN", "NONE") else f"_C{i}"
            for i, c in enumerate(df.columns)]

def _estado_penalidad(saldo_final: float) -> str:
    if saldo_final < -TOLERANCIA:
        return "EXCESO"
    if abs(saldo_final) <= TOLERANCIA:
        return "CANCELADO"
    return "PARCIAL"

# ── VALIDACIÓN ────────────────────────────────────────────────────────────────
def _validar_inputs():
    errores = []
    for ruta, desc in [
        (LISTA_CORTE_PATH,    "correr 04_cobranza primero"),
        (COBRANZA_FINAL_PATH, "correr 04_cobranza primero"),
        (PAGOS_YAPE_PATH,     "re-correr 03_pagos con el reporte extendido del banco"),
        (PAGOS_EFEC_PATH,     "actualizar pagos_efectivo con cobros del período de penalidad"),
    ]:
        if not ruta.exists():
            errores.append(f"Falta: {ruta}\n  → {desc}")
    if errores:
        for e in errores:
            log.error(e)
        raise FileNotFoundError("Inputs faltantes — ver errores arriba")
    log.info("Inputs validados correctamente")

# ── CARGA ─────────────────────────────────────────────────────────────────────
def _cargar_lista_corte() -> list[dict]:
    df = pd.read_excel(LISTA_CORTE_PATH, header=1, dtype=str)
    df.columns = _col_names(df)
    resultado = []
    for _, f in df.iterrows():
        mz = str(f.get("MZ", "")).strip().upper()
        lt = _norm_lt(f.get("LT", ""))
        if not mz or not lt:
            continue
        saldo     = _float(f.get("SALDO", 0))
        penalidad = _float(f.get("PENALIDAD", PENALIDAD))
        resultado.append({
            "mz":            mz,
            "lt":            lt,
            "key":           f"{mz}-{lt}",
            "nombre":        str(f.get("NOMBRE", "")).strip(),
            "saldo":         saldo,
            "penalidad":     penalidad,
            "total_a_pagar": _float(f.get("TOTAL_A_PAGAR", round(saldo + penalidad, 2))),
        })
    log.info(f"Lista corte: {len(resultado)} usuarios")
    return resultado

def _cargar_cobranza_final() -> dict:
    df = pd.read_excel(COBRANZA_FINAL_PATH, header=1, dtype=str)
    df.columns = _col_names(df)
    result = {}
    for _, f in df.iterrows():
        mz = str(f.get("MZ", "")).strip().upper()
        lt = _norm_lt(f.get("LT", ""))
        if not mz or not lt:
            continue
        result[f"{mz}-{lt}"] = {
            "yape":             _float(f.get("YAPE", 0)),
            "efectivo":         _float(f.get("EFECTIVO", 0)),
            "corte_reconexion": _float(f.get("CORTE_RECONEXION", 0)),
            "saldo":            _float(f.get("SALDO", 0)),
            "estado":           str(f.get("ESTADO", "")).strip().upper(),
        }
    log.info(f"Cobranza final: {len(result)} usuarios cargados")
    return result

def _cargar_pagos_yape() -> dict:
    df = pd.read_excel(PAGOS_YAPE_PATH, header=1, dtype=str)
    df.columns = _col_names(df)
    totales = {}
    for _, f in df.iterrows():
        if str(f.get("TIPO", "")).strip().upper() != "TE PAGÓ":
            continue
        mz   = str(f.get("MZ",   "")).strip().upper()
        lote = _norm_lt(f.get("LOTE", ""))
        conc = str(f.get("CONCEPTO", "")).strip().upper()
        if not mz or mz == "NAN" or not lote:
            continue
        if conc and conc not in ("NAN", "NONE", ""):
            continue
        key = f"{mz}-{lote}"
        totales[key] = totales.get(key, 0.0) + _float(f.get("MONTO_ASIGNA", 0))
    log.info(f"Pagos Yape actualizados: {len(totales)} lotes")
    return totales

def _cargar_pagos_efectivo() -> dict:
    df = pd.read_excel(PAGOS_EFEC_PATH, header=1, dtype=str)
    df.columns = _col_names(df)
    totales = {}
    for _, f in df.iterrows():
        mz = str(f.get("MZ", "")).strip().upper()
        lt = _norm_lt(f.get("LT", ""))
        if not mz or not lt:
            continue
        key = f"{mz}-{lt}"
        totales[key] = totales.get(key, 0.0) + _float(f.get("MONTO", 0))
    log.info(f"Pagos Efectivo actualizados: {len(totales)} lotes")
    return totales

# ── CLASIFICAR ────────────────────────────────────────────────────────────────
def _clasificar(lista_corte: list[dict],
                cobranza_orig: dict,
                pagos_yape: dict,
                pagos_efectivo: dict) -> tuple[list[dict], list[dict]]:
    pagaron, no_pagaron = [], []
    for u in lista_corte:
        key  = u["key"]
        orig = cobranza_orig.get(key, {})
        orig_pagado = round(
            _float(orig.get("yape", 0)) + _float(orig.get("efectivo", 0)), 2
        )
        nuevo_total = round(
            pagos_yape.get(key, 0.0) + pagos_efectivo.get(key, 0.0), 2
        )
        pago_nuevo = round(max(0.0, nuevo_total - orig_pagado), 2)

        if pago_nuevo >= u["penalidad"] - TOLERANCIA:
            saldo_final = round(u["saldo"] - max(0.0, pago_nuevo - u["penalidad"]), 2)
            pagaron.append({
                **u,
                "pago_nuevo":  pago_nuevo,
                "saldo_final": saldo_final,
                "estado":      _estado_penalidad(saldo_final),
            })
        else:
            no_pagaron.append({
                **u,
                "pago_nuevo":       pago_nuevo,
                "reconexion":       RECONEXION,
                "total_reconexion": round(u["saldo"] + RECONEXION, 2),
            })

    log.info(f"Pagaron penalidad: {len(pagaron)} · Corte físico: {len(no_pagaron)}")
    return pagaron, no_pagaron

# ── EXPORT: PAGARON PENALIDAD ────────────────────────────────────────────────
# Col layout:
#  1- 3  ¿Quién es?               MZ LT NOMBRE
#  4      sep
#  5- 7  ¿Cuánto debía al corte?  SALDO PENALIDAD TOTAL_A_PAGAR
#  8      sep
#  9      ¿Cuánto pagó en ventana? PAGO_NUEVO
# 10      sep
# 11-12  ¿Qué queda?              SALDO_FINAL ESTADO

_PP_GRUPOS = [
    (1,  3,  "¿Quién es?",                *GH_QUIEN),
    (5,  7,  "¿Cuánto debía al corte?",   *GH_DEUDA),
    (9,  9,  "¿Cuánto pagó en ventana?",  *GH_PAGO),
    (11, 12, "¿Qué queda?",               *GH_QUEDA),
]
_PP_COLS = [
    (1,  "MZ",            *GH_QUIEN,  6),
    (2,  "LT",            *GH_QUIEN,  7),
    (3,  "NOMBRE",        *GH_QUIEN, 26),
    (5,  "SALDO",         *GH_DEUDA, 10),
    (6,  "PENALIDAD",     *GH_DEUDA, 10),
    (7,  "TOTAL_A_PAGAR", *GH_DEUDA, 13),
    (9,  "PAGO_NUEVO",    *GH_PAGO,  11),
    (11, "SALDO_FINAL",   *GH_QUEDA, 11),
    (12, "ESTADO",        *GH_QUEDA, 12),
]
_PP_SEP = [4, 8, 10]

def _exportar_pagaron(pagaron: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title       = "pagaron_penalidad"
    ws.freeze_panes = "A3"
    last_row = len(pagaron) + 2

    for cs, ce, texto, bg, txt in _PP_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for sc in _PP_SEP:
        _sep(ws, sc, last_row)
    for col, nombre, bg, txt, ancho in _PP_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 20

    for ri, r in enumerate(pagaron, 3):
        _c(ws, ri, 1, r["mz"],     TD_QUIEN, "5B21B6", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_QUIEN, "5B21B6", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_QUIEN, "333333")
        _c(ws, ri, 5, r["saldo"],         TD_DEUDA, "7D6608", mono=True, align="right")
        _c(ws, ri, 6, r["penalidad"],     TD_DEUDA, "7D6608", mono=True, align="right")
        _c(ws, ri, 7, r["total_a_pagar"], TD_DEUDA, "7D6608", bold=True, mono=True, align="right")
        _c(ws, ri, 9, r["pago_nuevo"],    TD_PAGO,  "7C3003", bold=True, mono=True, align="right")

        sf = r["saldo_final"]
        sf_txt = ("065F46" if abs(sf) <= TOLERANCIA
                  else "1D4ED8" if sf < -TOLERANCIA
                  else "7D6608")
        _c(ws, ri, 11, sf or None, TD_QUEDA, sf_txt, mono=True, align="right")

        est  = r["estado"]
        c_e  = ws.cell(row=ri, column=12, value=est)
        c_e.font      = Font(name="Arial", size=9, bold=True,
                             color=ESTADO_TXT_PP.get(est, "333333"))
        c_e.fill      = PatternFill("solid",
                                    start_color=ESTADO_BG_PP.get(est, "FFFFFF"))
        c_e.alignment = Alignment(horizontal="center", vertical="center")
        c_e.border    = _borde()
        ws.row_dimensions[ri].height = 17

    wb.save(OUTPUTS_DIR / "pagaron_penalidad.xlsx")
    log.info(f"pagaron_penalidad.xlsx → {len(pagaron)} filas")

# ── EXPORT: CORTE FÍSICO ──────────────────────────────────────────────────────
# Col layout:
#  1-3  ¿Quién es?           MZ LT NOMBRE
#  4     sep
#  5-6  ¿Por qué se corta?   SALDO PAGO_EN_VENTANA
#  7     sep
#  8-9  ¿Para reconectarse?  RECONEXION TOTAL_RECONEXION

_CF_GRUPOS = [
    (1, 3, "¿Quién es?",          *GH_QUIEN),
    (5, 6, "¿Por qué se corta?",  *GH_DEUDA),
    (8, 9, "¿Para reconectarse?", *GH_CORTE),
]
_CF_COLS = [
    (1, "MZ",               *GH_QUIEN,  6),
    (2, "LT",               *GH_QUIEN,  7),
    (3, "NOMBRE",           *GH_QUIEN, 26),
    (5, "SALDO",            *GH_DEUDA, 10),
    (6, "PAGO_EN_VENTANA",  *GH_DEUDA, 14),
    (8, "RECONEXION",       *GH_CORTE, 11),
    (9, "TOTAL_RECONEXION", *GH_CORTE, 16),
]
_CF_SEP_COLS = [4, 7]

def _exportar_corte_fisico(no_pagaron: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title       = "corte_fisico"
    ws.freeze_panes = "A3"
    last_row = len(no_pagaron) + 2

    for cs, ce, texto, bg, txt in _CF_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for sc in _CF_SEP_COLS:
        _sep(ws, sc, last_row)
    for col, nombre, bg, txt, ancho in _CF_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 20

    for ri, r in enumerate(no_pagaron, 3):
        _c(ws, ri, 1, r["mz"],     TD_QUIEN, "5B21B6", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_QUIEN, "5B21B6", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_QUIEN, "333333")
        _c(ws, ri, 5, r["saldo"],                    TD_DEUDA, "92400E", mono=True, align="right")
        _c(ws, ri, 6, r["pago_nuevo"] or None,        TD_DEUDA, "92400E", mono=True, align="right")
        _c(ws, ri, 8, r["reconexion"],                TD_CORTE, "991B1B", mono=True, align="right")
        _c(ws, ri, 9, r["total_reconexion"],           TD_CORTE, "7F1D1D", bold=True, mono=True, align="right")
        ws.row_dimensions[ri].height = 17

    wb.save(OUTPUTS_DIR / "corte_fisico.xlsx")
    log.info(f"corte_fisico.xlsx → {len(no_pagaron)} filas")

# ── EXPORT/UPDATE: CORTE ACUMULADO ───────────────────────────────────────────
# Col layout:
#  1- 3  ¿Quién?              MZ LT NOMBRE
#  4      sep
#  5- 6  ¿Cuándo cortado?     MES_CORTE SALDO_AL_CORTE
#  7      sep
#  8- 9  ¿Para reconectarse?  RECONEXION TOTAL_RECONEXION
# 10      sep
# 11-12  ¿Estado actual?      ESTADO MES_RECONEXION

_CA_GRUPOS = [
    (1,  3,  "¿Quién?",              *GH_QUIEN),
    (5,  6,  "¿Cuándo fue cortado?", *GH_CUANDO),
    (8,  9,  "¿Para reconectarse?",  *GH_CORTE),
    (11, 12, "¿Estado actual?",      *GH_QUEDA),
]
_CA_COLS = [
    (1,  "MZ",               *GH_QUIEN,  6),
    (2,  "LT",               *GH_QUIEN,  7),
    (3,  "NOMBRE",           *GH_QUIEN, 26),
    (5,  "MES_CORTE",        *GH_CUANDO, 10),
    (6,  "SALDO_AL_CORTE",   *GH_CUANDO, 13),
    (8,  "RECONEXION",       *GH_CORTE,  11),
    (9,  "TOTAL_RECONEXION", *GH_CORTE,  16),
    (11, "ESTADO",           *GH_QUEDA,  12),
    (12, "MES_RECONEXION",   *GH_QUEDA,  14),
]
_CA_SEP_COLS = [4, 7, 10]

def _actualizar_corte_acumulado(no_pagaron: list[dict], cobranza_orig: dict):
    registros = []

    if CORTE_ACUMULADO_PATH.exists():
        df = pd.read_excel(CORTE_ACUMULADO_PATH, header=1, dtype=str)
        df.columns = _col_names(df)
        for _, f in df.iterrows():
            mz = str(f.get("MZ", "")).strip().upper()
            lt = _norm_lt(f.get("LT", ""))
            if not mz or not lt:
                continue
            key    = f"{mz}-{lt}"
            estado = str(f.get("ESTADO", "")).strip().upper()
            mes_r  = str(f.get("MES_RECONEXION", "")).strip()
            mes_r  = "" if mes_r in ("NAN", "NONE", "—", "") else mes_r

            # Detectar reconexión: cobró CORTE_RECONEXION este mes y quedó al día
            if estado == "CORTADO":
                cob = cobranza_orig.get(key, {})
                if (_float(cob.get("corte_reconexion", 0)) > TOLERANCIA
                        and cob.get("estado", "") in ("CANCELADO", "EXCESO")):
                    estado = "RECONECTADO"
                    mes_r  = MES_ACTUAL

            registros.append({
                "mz":               mz,
                "lt":               lt,
                "nombre":           str(f.get("NOMBRE", "")).strip(),
                "mes_corte":        str(f.get("MES_CORTE", "")).strip(),
                "saldo_al_corte":   _float(f.get("SALDO_AL_CORTE", 0)),
                "reconexion":       _float(f.get("RECONEXION", RECONEXION)),
                "total_reconexion": _float(f.get("TOTAL_RECONEXION", 0)),
                "estado":           estado,
                "mes_reconexion":   mes_r,
            })
        log.info(f"corte_acumulado existente: {len(registros)} registros cargados")

    claves_cortadas = {r["mz"] + "-" + r["lt"] for r in registros if r["estado"] == "CORTADO"}
    nuevos = 0
    for u in no_pagaron:
        if u["key"] in claves_cortadas:
            continue
        registros.append({
            "mz":               u["mz"],
            "lt":               u["lt"],
            "nombre":           u["nombre"],
            "mes_corte":        MES_ACTUAL,
            "saldo_al_corte":   u["saldo"],
            "reconexion":       u["reconexion"],
            "total_reconexion": u["total_reconexion"],
            "estado":           "CORTADO",
            "mes_reconexion":   "",
        })
        nuevos += 1

    registros.sort(key=lambda r: (r["mes_corte"], r["mz"], r["lt"]))

    wb = Workbook()
    ws = wb.active
    ws.title       = "corte_acumulado"
    ws.freeze_panes = "A3"
    last_row = len(registros) + 2

    for cs, ce, texto, bg, txt in _CA_GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)
    for sc in _CA_SEP_COLS:
        _sep(ws, sc, last_row)
    for col, nombre, bg, txt, ancho in _CA_COLS:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 20

    for ri, r in enumerate(registros, 3):
        _c(ws, ri, 1, r["mz"],     TD_QUIEN, "5B21B6", mono=True, align="center")
        _c(ws, ri, 2, r["lt"],     TD_QUIEN, "5B21B6", mono=True, align="center")
        _c(ws, ri, 3, r["nombre"], TD_QUIEN, "333333")
        _c(ws, ri, 5, r["mes_corte"],      TD_CUANDO, "374151", mono=True, align="center")
        _c(ws, ri, 6, r["saldo_al_corte"], TD_CORTE,  "991B1B", mono=True, align="right")
        _c(ws, ri, 8, r["reconexion"],       TD_CORTE, "991B1B", mono=True, align="right")
        _c(ws, ri, 9, r["total_reconexion"], TD_CORTE, "7F1D1D", bold=True, mono=True, align="right")

        est  = r["estado"]
        c_e  = ws.cell(row=ri, column=11, value=est)
        c_e.font      = Font(name="Arial", size=9, bold=True,
                             color=ESTADO_TXT_CA.get(est, "333333"))
        c_e.fill      = PatternFill("solid",
                                    start_color=ESTADO_CA_BG.get(est, "FFFFFF"))
        c_e.alignment = Alignment(horizontal="center", vertical="center")
        c_e.border    = _borde()

        mes_r = r["mes_reconexion"]
        _c(ws, ri, 12, mes_r or None,
           TD_QUEDA, "085041" if mes_r else "AAAAAA", mono=True, align="center")
        ws.row_dimensions[ri].height = 17

    wb.save(CORTE_ACUMULADO_PATH)
    reconectados = sum(1 for r in registros if r["estado"] == "RECONECTADO")
    log.info(f"corte_acumulado.xlsx → {len(registros)} total · "
             f"{nuevos} nuevos · {reconectados} reconectados")

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "═" * 55)
    print("  05_corte — Clasificación penalidad y corte físico")
    print("═" * 55)
    _init_logging()

    print("\n[1/5] Validando inputs...")
    _validar_inputs()

    print("\n[2/5] Cargando datos...")
    lista_corte   = _cargar_lista_corte()
    cobranza_orig = _cargar_cobranza_final()
    pagos_yape    = _cargar_pagos_yape()
    pagos_efec    = _cargar_pagos_efectivo()

    print("\n[3/5] Clasificando...")
    pagaron, no_pagaron = _clasificar(lista_corte, cobranza_orig, pagos_yape, pagos_efec)

    print("\n[4/5] Exportando resultados...")
    _exportar_pagaron(pagaron)
    _exportar_corte_fisico(no_pagaron)

    print("\n[5/5] Actualizando corte_acumulado...")
    _actualizar_corte_acumulado(no_pagaron, cobranza_orig)

    print("\n" + "═" * 55)
    if no_pagaron:
        print(f"  {len(pagaron)} pagaron la penalidad")
        print(f"  {len(no_pagaron)} van a corte físico → corte_fisico.xlsx")
        print(f"  Pasar corte_fisico.xlsx al operario")
    else:
        print(f"  {len(pagaron)} pagaron la penalidad")
        print("  Sin corte físico este mes")
    print("═" * 55 + "\n")


if __name__ == "__main__":
    main()
