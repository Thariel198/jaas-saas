import logging
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── CONFIG ────────────────────────────────────────────────────────────────────
ROOT                = Path(__file__).parent
COBRANZA_DIR        = ROOT.parent / "04_cobranza"
INPUTS_DIR          = ROOT / "inputs"
OUTPUTS_DIR         = ROOT / "outputs"
BACKUP_DIR          = ROOT / "backup"

OVERRIDES_PATH      = INPUTS_DIR / "overrides.xlsx"
PAGOS_YAPE_PATH     = COBRANZA_DIR / "inputs" / "pagos_yape" / "pagos_yape_tepago.xlsx"
PAGOS_EFEC_PATH     = COBRANZA_DIR / "inputs" / "pagos_efectivo" / "pagos_efectivo.xlsx"
COBRANZA_FINAL_PATH = COBRANZA_DIR / "outputs" / "cobranza_final.xlsx"

MES_ACTUAL        = datetime.now().strftime("%Y-%m")
TRAZABILIDAD_PATH = OUTPUTS_DIR / f"trazabilidad_overrides_{MES_ACTUAL}.xlsx"

# Posiciones columnas cobranza_final (1-indexed, fila 1=grupos, fila 2=cabeceras, fila 3+=datos)
_CF_ROW_DATA  = 3
_CF_COL_MZ    = 1
_CF_COL_LT    = 2
_CF_COL_TOTAL = 20   # TOTAL deuda
_CF_COL_YAPE  = 22
_CF_COL_EFEC  = 23
_CF_COL_TOPAG = 24
_CF_COL_SALDO = 26
_CF_COL_ESTADO= 27

ESTADO_BG  = {"CANCELADO": "E1F5EE", "EXCESO": "EFF6FF",
               "PARCIAL":   "FAEEDA", "PENDIENTE": "FEF2F2"}
ESTADO_TXT = {"CANCELADO": "085041", "EXCESO": "1D4ED8",
               "PARCIAL":  "854F0B", "PENDIENTE": "991B1B"}

# ── LOGGING ───────────────────────────────────────────────────────────────────
def _init_logging():
    OUTPUTS_DIR.mkdir(exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(OUTPUTS_DIR / "run.log", encoding="utf-8"),
        ],
    )

log = logging.getLogger(__name__)

# ── UTILIDADES ────────────────────────────────────────────────────────────────
def _float(val) -> float:
    try:
        return float(str(val).replace(",", ".").strip())
    except (ValueError, TypeError):
        return 0.0

def _norm_lt(val) -> str:
    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN"):
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.upper()

def _estado(saldo: float, total_pagado: float) -> str:
    if saldo < -0.005:
        return "EXCESO"
    if abs(saldo) <= 0.005:
        return "CANCELADO"
    return "PARCIAL" if total_pagado > 0.005 else "PENDIENTE"

def _norm_fecha(val) -> str:
    if val is None:
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d %H:%M")
    s = str(val).strip().split(".")[0]
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M")
        except ValueError:
            continue
    return s

def _calcular_ajuste(cur_pago: float, cur_otro: float,
                     total_deuda: float, monto: float,
                     signo: int) -> tuple[float, float, float, str]:
    """Recalcula pago, total_pagado, saldo y estado tras mover `monto` (signo: -1 resta / +1 suma)."""
    nuevo_pago   = round(max(0.0, cur_pago + signo * monto), 2)
    nuevo_pagado = round(nuevo_pago + cur_otro, 2)
    nuevo_saldo  = round(total_deuda - nuevo_pagado, 2)
    return nuevo_pago, nuevo_pagado, nuevo_saldo, _estado(nuevo_saldo, nuevo_pagado)

def _borde():
    b = Side(style="thin", color="CCCCCC")
    return Border(left=b, right=b, top=b, bottom=b)

def _find_col(ws, header_row: int, name: str) -> int | None:
    for cell in ws[header_row]:
        if str(cell.value or "").strip().upper() == name.strip().upper():
            return cell.column
    return None

# ── VALIDACIÓN ────────────────────────────────────────────────────────────────
def _validar_inputs():
    errores = []
    for ruta, desc in [
        (OVERRIDES_PATH,      "crear overrides.xlsx en inputs/ y llenarlo"),
        (PAGOS_YAPE_PATH,     "asegurar que 03_pagos ya corrió"),
        (PAGOS_EFEC_PATH,     "asegurar que pagos_efectivo.xlsx existe"),
        (COBRANZA_FINAL_PATH, "asegurar que 04_cobranza ya corrió"),
    ]:
        if not ruta.exists():
            errores.append(f"Falta: {ruta}\n  → {desc}")
    if errores:
        for e in errores:
            log.error(e)
        raise FileNotFoundError("Inputs faltantes — ver errores arriba")
    log.info("Inputs validados correctamente")

# ── CARGA OVERRIDES ───────────────────────────────────────────────────────────
def _cargar_overrides() -> list[dict]:
    df = pd.read_excel(OVERRIDES_PATH, dtype=str)
    df.columns = [str(c).strip().upper() for c in df.columns]
    pendientes = []
    for idx, f in df.iterrows():
        if str(f.get("ESTADO", "")).strip().lower() != "pendiente":
            continue
        medio  = str(f.get("MEDIO", "")).strip().upper()
        mz_ant = str(f.get("MZ_ANTERIOR", "")).strip().upper()
        lt_ant = _norm_lt(f.get("LOTE_ANTERIOR", ""))
        mz_nvo = str(f.get("MZ_NUEVO", "")).strip().upper()
        lt_nvo = _norm_lt(f.get("LOTE_NUEVO", ""))

        if not mz_ant or not lt_ant or not mz_nvo or not lt_nvo:
            log.warning(f"Override fila {idx + 2}: MZ/LOTE incompleto — saltando")
            continue
        if medio not in ("YAPE", "EFECTIVO"):
            log.warning(f"Override fila {idx + 2}: MEDIO '{medio}' no reconocido (usar YAPE o EFECTIVO)")
            continue

        pendientes.append({
            "row_xlsx": idx + 2,   # fila real en Excel (pandas idx es 0-based, +1 header, +1 offset)
            "fecha":    f.get("FECHA"),
            "medio":    medio,
            "mz_ant":   mz_ant,
            "lt_ant":   lt_ant,
            "mz_nvo":   mz_nvo,
            "lt_nvo":   lt_nvo,
            "motivo":   str(f.get("MOTIVO", "")).strip(),
        })

    log.info(f"Overrides pendientes: {len(pendientes)}")
    return pendientes

# ── BACKUP ────────────────────────────────────────────────────────────────────
def _backup(ruta: Path):
    BACKUP_DIR.mkdir(exist_ok=True)
    ts   = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    dest = BACKUP_DIR / f"{ruta.stem}_{ts}{ruta.suffix}"
    shutil.copy2(ruta, dest)
    log.info(f"Backup: {dest.name}")

# ── APLICAR YAPE ──────────────────────────────────────────────────────────────
def _aplicar_yape(override: dict) -> float:
    wb = load_workbook(PAGOS_YAPE_PATH)
    ws = wb.active

    col_tipo  = _find_col(ws, 2, "TIPO")
    col_fecha = _find_col(ws, 2, "FECHA")
    col_mz    = _find_col(ws, 2, "MZ")
    col_lote  = _find_col(ws, 2, "LOTE")
    col_monto = _find_col(ws, 2, "MONTO_ASIGNA")

    if not all([col_tipo, col_fecha, col_mz, col_lote, col_monto]):
        log.error("pagos_yape_tepago: columnas TIPO/FECHA/MZ/LOTE/MONTO_ASIGNA no encontradas")
        return 0.0

    target_fecha = _norm_fecha(override["fecha"])
    monto_total, filas = 0.0, 0

    for r in range(3, ws.max_row + 1):
        tipo  = str(ws.cell(r, col_tipo).value  or "").strip().upper()
        fecha = _norm_fecha(ws.cell(r, col_fecha).value)
        mz    = str(ws.cell(r, col_mz).value   or "").strip().upper()
        lote  = _norm_lt(ws.cell(r, col_lote).value)

        if (tipo == "TE PAGÓ"
                and fecha == target_fecha
                and mz    == override["mz_ant"]
                and lote  == override["lt_ant"]):
            monto_total += _float(ws.cell(r, col_monto).value)
            ws.cell(r, col_mz).value   = override["mz_nvo"]
            ws.cell(r, col_lote).value = override["lt_nvo"]
            filas += 1

    if filas == 0:
        log.warning(f"YAPE: sin coincidencia para {override['mz_ant']}-{override['lt_ant']} "
                    f"FECHA={target_fecha}")
        return 0.0

    wb.save(PAGOS_YAPE_PATH)
    log.info(f"YAPE: {filas} fila(s) "
             f"{override['mz_ant']}-{override['lt_ant']} → {override['mz_nvo']}-{override['lt_nvo']} "
             f"S/{monto_total:.2f}")
    return round(monto_total, 2)

# ── APLICAR EFECTIVO ──────────────────────────────────────────────────────────
def _aplicar_efectivo(override: dict) -> float:
    wb = load_workbook(PAGOS_EFEC_PATH)
    ws = wb.active

    col_mz    = _find_col(ws, 2, "MZ")
    col_lt    = _find_col(ws, 2, "LT")
    col_monto = _find_col(ws, 2, "MONTO")

    if not all([col_mz, col_lt, col_monto]):
        log.error("pagos_efectivo: columnas MZ/LT/MONTO no encontradas")
        return 0.0

    monto_total, filas = 0.0, 0

    for r in range(3, ws.max_row + 1):
        mz   = str(ws.cell(r, col_mz).value or "").strip().upper()
        lt   = _norm_lt(ws.cell(r, col_lt).value)

        if mz == override["mz_ant"] and lt == override["lt_ant"]:
            monto_total += _float(ws.cell(r, col_monto).value)
            ws.cell(r, col_mz).value = override["mz_nvo"]
            ws.cell(r, col_lt).value = override["lt_nvo"]
            filas += 1

    if filas == 0:
        log.warning(f"EFECTIVO: sin coincidencia para {override['mz_ant']}-{override['lt_ant']}")
        return 0.0

    wb.save(PAGOS_EFEC_PATH)
    log.info(f"EFECTIVO: {filas} fila(s) "
             f"{override['mz_ant']}-{override['lt_ant']} → {override['mz_nvo']}-{override['lt_nvo']} "
             f"S/{monto_total:.2f}")
    return round(monto_total, 2)

# ── ACTUALIZAR COBRANZA FINAL ─────────────────────────────────────────────────
def _actualizar_cobranza(override: dict, monto: float):
    if monto <= 0:
        log.warning("Monto = 0 — cobranza_final no modificada")
        return

    col_pago = _CF_COL_YAPE if override["medio"] == "YAPE" else _CF_COL_EFEC
    col_otro = _CF_COL_EFEC if override["medio"] == "YAPE" else _CF_COL_YAPE

    wb = load_workbook(COBRANZA_FINAL_PATH)
    ws = wb.active

    for mz, lt, signo in [(override["mz_ant"], override["lt_ant"], -1),
                           (override["mz_nvo"], override["lt_nvo"], +1)]:
        row = None
        for r in range(_CF_ROW_DATA, ws.max_row + 1):
            if (str(ws.cell(r, _CF_COL_MZ).value or "").strip().upper() == mz
                    and _norm_lt(ws.cell(r, _CF_COL_LT).value) == lt):
                row = r
                break

        if row is None:
            log.warning(f"cobranza_final: no se encontró {mz}-{lt}")
            continue

        total_deuda = _float(ws.cell(row, _CF_COL_TOTAL).value)
        cur_pago    = _float(ws.cell(row, col_pago).value)
        cur_otro    = _float(ws.cell(row, col_otro).value)

        nuevo_pago, nuevo_topag, nuevo_saldo, nuevo_estado = _calcular_ajuste(
            cur_pago, cur_otro, total_deuda, monto, signo
        )

        TD_PAGO  = "FEF6EE"
        TD_QUEDA = "F0FFF8"

        def _cell(col, val, bg, txt, bold=False, mono=False, align="right"):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name="Consolas" if mono else "Arial",
                               size=9, bold=bold, color=txt)
            c.fill      = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border    = _borde()

        _cell(col_pago,     nuevo_pago   or None, TD_PAGO,  "7C3003", mono=True)
        _cell(_CF_COL_TOPAG, nuevo_topag or None, TD_PAGO,  "7C3003", bold=True, mono=True)

        saldo_txt = ("A32D2D" if nuevo_saldo > 0.005
                     else "1D4ED8" if nuevo_saldo < -0.005
                     else "065F46")
        _cell(_CF_COL_SALDO, nuevo_saldo, TD_QUEDA, saldo_txt, bold=True, mono=True)

        c_est = ws.cell(row=row, column=_CF_COL_ESTADO, value=nuevo_estado)
        c_est.font      = Font(name="Arial", size=9, bold=True,
                               color=ESTADO_TXT.get(nuevo_estado, "333333"))
        c_est.fill      = PatternFill("solid",
                                      start_color=ESTADO_BG.get(nuevo_estado, "FFFFFF"))
        c_est.alignment = Alignment(horizontal="center", vertical="center")
        c_est.border    = _borde()

        log.info(f"cobranza_final {mz}-{lt}: "
                 f"{'YAPE' if override['medio']=='YAPE' else 'EFEC'} "
                 f"{cur_pago:.2f} → {nuevo_pago:.2f} · "
                 f"SALDO {nuevo_saldo:.2f} · {nuevo_estado}")

    wb.save(COBRANZA_FINAL_PATH)

# ── TRAZABILIDAD ──────────────────────────────────────────────────────────────
_TRAZ_HEADERS = ["APLICADO_EN", "MEDIO", "FECHA_PAGO",
                 "MZ_ANTERIOR", "LOTE_ANTERIOR", "MZ_NUEVO", "LOTE_NUEVO",
                 "MONTO_MOVIDO", "MOTIVO"]
_TRAZ_WIDTHS  = [18, 10, 18, 12, 14, 10, 10, 13, 40]
_TRAZ_BG      = "F8FAFC"
_TRAZ_HEAD_BG = "1E3A5F"

def _registrar_trazabilidad(override: dict, monto: float):
    if TRAZABILIDAD_PATH.exists():
        wb = load_workbook(TRAZABILIDAD_PATH)
        ws = wb["Overrides"] if "Overrides" in wb.sheetnames else wb.create_sheet("Overrides")
        if ws.max_row < 2 or ws.cell(1, 1).value is None:
            _escribir_cabecera_traz(ws)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Overrides"
        _escribir_cabecera_traz(ws)

    fecha_pago = _norm_fecha(override["fecha"]) if override["medio"] == "YAPE" else "—"
    fila = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        override["medio"],
        fecha_pago,
        override["mz_ant"],
        override["lt_ant"],
        override["mz_nvo"],
        override["lt_nvo"],
        monto,
        override["motivo"],
    ]

    b = Side(style="thin", color="D1D5DB")
    borde = Border(left=b, right=b, top=b, bottom=b)
    ri = ws.max_row + 1
    for ci, val in enumerate(fila, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font      = Font(name="Consolas" if ci in (1, 3, 8) else "Arial", size=9)
        c.fill      = PatternFill("solid", start_color=_TRAZ_BG)
        c.alignment = Alignment(horizontal="right" if ci == 8 else "center" if ci in (2, 4, 5, 6, 7) else "left",
                                vertical="center")
        c.border    = borde
        ws.row_dimensions[ri].height = 16

    OUTPUTS_DIR.mkdir(exist_ok=True)
    wb.save(TRAZABILIDAD_PATH)

def _escribir_cabecera_traz(ws):
    b = Side(style="thin", color="D1D5DB")
    borde = Border(left=b, right=b, top=b, bottom=b)
    for ci, (hdr, ancho) in enumerate(zip(_TRAZ_HEADERS, _TRAZ_WIDTHS), 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font      = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", start_color=_TRAZ_HEAD_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = borde
        ws.column_dimensions[c.column_letter].width = ancho
    ws.row_dimensions[1].height = 20

# ── MARCAR APLICADO ───────────────────────────────────────────────────────────
def _marcar_aplicado(row_xlsx: int):
    wb = load_workbook(OVERRIDES_PATH)
    ws = wb.active
    estado_col = _find_col(ws, 1, "ESTADO")
    if estado_col:
        c = ws.cell(row=row_xlsx, column=estado_col, value="aplicado")
        c.font      = Font(name="Arial", size=10, color="085041")
        c.fill      = PatternFill("solid", start_color="E1F5EE")
        c.alignment = Alignment(horizontal="center", vertical="center")
    wb.save(OVERRIDES_PATH)

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "═" * 55)
    print("  05b_override — Corrección de pagos mal asignados")
    print("═" * 55)
    _init_logging()

    print("\n[1/4] Validando inputs...")
    _validar_inputs()

    print("\n[2/4] Cargando overrides pendientes...")
    pendientes = _cargar_overrides()
    if not pendientes:
        print("\n  Sin overrides pendientes — nada que hacer.")
        print("═" * 55 + "\n")
        return

    print(f"\n[3/4] Aplicando {len(pendientes)} override(s)...")

    # Backup único por ejecución (antes del primer cambio)
    _backup(PAGOS_YAPE_PATH)
    _backup(PAGOS_EFEC_PATH)
    _backup(COBRANZA_FINAL_PATH)

    aplicados, errores = 0, 0
    for ov in pendientes:
        print(f"\n  → {ov['medio']} {ov['mz_ant']}-{ov['lt_ant']} → {ov['mz_nvo']}-{ov['lt_nvo']}")
        try:
            if ov["medio"] == "YAPE":
                monto = _aplicar_yape(ov)
            else:
                monto = _aplicar_efectivo(ov)

            if monto > 0:
                _actualizar_cobranza(ov, monto)
                _registrar_trazabilidad(ov, monto)
                _marcar_aplicado(ov["row_xlsx"])
                print(f"     S/{monto:.2f} movidos — aplicado")
                aplicados += 1
            else:
                print(f"     SIN COINCIDENCIA — override saltado (revisar FECHA/MZ/LOTE)")
                errores += 1
        except Exception as e:
            log.error(f"Error en override fila {ov['row_xlsx']}: {e}")
            print(f"     ERROR: {e}")
            errores += 1

    print(f"\n[4/4] Guardando trazabilidad...")

    print("\n" + "═" * 55)
    print(f"  {aplicados} override(s) aplicados correctamente")
    if errores:
        print(f"  {errores} override(s) con error — revisar run.log")
    print(f"  Trazabilidad: {TRAZABILIDAD_PATH.name}")
    print(f"  Backup en: backup/")
    print("═" * 55 + "\n")


if __name__ == "__main__":
    main()
