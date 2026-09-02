from pathlib import Path
import sys

from openpyxl import Workbook


MODULO = Path(__file__).parents[1]
sys.path.insert(0, str(MODULO))
import recibos_medidor_pagado as recibos  # noqa: E402


def test_carga_convenio_historial_sin_depender_de_hoja_agua(tmp_path, monkeypatch):
    vista = tmp_path / "vista.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("MES_ANTERIOR")
    ws = wb.create_sheet("CONVENIO_HISTORIAL")
    ws.append(["Predio", None, None, "Historia", None, "Saldo"])
    ws.append(["MZ", "LT", "NOMBRE", "DEUDA", "PAGO AGO-26", "SALDO ACTUAL"])
    ws.append(["A", "1", "Persona", 25, 25, 0])
    wb.save(vista)
    monkeypatch.setattr(recibos, "VISTA_PATH", vista)

    filas = recibos._cargar_pagados()

    assert len(filas) == 1
    assert filas[0]["NRO"] == "MP-001"
    assert filas[0]["PAGOS"] == [("AGO-26", 25.0)]
