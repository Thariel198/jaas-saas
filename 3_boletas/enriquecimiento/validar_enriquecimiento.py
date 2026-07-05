"""
validar_enriquecimiento.py — reconciliación planilla → DATA_boletas.

Corre DESPUÉS de enriquecimiento/main.py y ANTES de 3_boletas/main.py.
Relee planilla_{mes}.xlsx y DATA_boletas.xlsx y compara predio por predio,
columna por columna. No corrige nada — solo reporta discrepancias.
Contrato de formato: docs/formato_validacion_enriquecimiento.html
"""
import logging
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from main import (
    PLANILLA_PATH, OUTPUT_PATH as DATA_BOLETAS_PATH,
    _cargar_config, _fecha_pago_str, _float, _norm_lt,
)

log = logging.getLogger(__name__)

TOLERANCIA = 0.005
OUTPUTS_DIR = Path(__file__).parent / "outputs"

# (columna en DATA_boletas, columna en planilla, tipo: "num" | "txt")
COLS_1A1 = [
    ("NOMBRES",                 "NOMBRE",            "txt"),
    ("Marcación anterior",      "MARC_ANT",          "num"),
    ("Marcacion altual",        "MARC_ACT",          "num"),
    ("M3",                      "M3",                "num"),
    ("Total mes actual",        "MES_ACTUAL",        "num"),
    ("MES ANTERIOR",            "MES_ANTERIOR",      "num"),
    ("Corte y reconexion",      "CORTE_RECONEXION",  "num"),
    ("Convenio",                "CONVENIO",          "num"),
    ("Mantenimiento",           "MANTENIMIENTO",     "num"),
    ("Multa (faena + reunión)", "MULTA",             "num"),
    ("Cuota directa",           "ACUERDOS_ASAMBLEA", "num"),
]

# Componentes de la fórmula Total (columnas de planilla)
COMPONENTES_TOTAL = [
    "MES_ACTUAL", "MES_ANTERIOR", "CORTE_RECONEXION",
    "CONVENIO", "MANTENIMIENTO", "MULTA", "ACUERDOS_ASAMBLEA",
]

# (columna en DATA_boletas, clave en config — None si el valor esperado se arma aparte)
COLS_CONFIG = [
    ("PERIODO",              "PERIODO"),
    ("FECHA DE VENCIMIENTO", "FECHA_VENCIMIENTO"),
    ("FECHA DE EMISIÓN",     "FECHA_EMISION"),
    ("LECTURA ANTERIOR",     "LECTURA_ANT_FECHA"),
    ("LECTURA ACTUAL",       "LECTURA_ACT_FECHA"),
    ("fecha pago",           None),  # valor esperado = _fecha_pago_str(config)
]


def _key(mz, lt) -> tuple:
    return (str(mz).strip().upper(), _norm_lt(lt))


def _cargar_planilla() -> pd.DataFrame:
    df = pd.read_excel(PLANILLA_PATH, header=1, dtype=str)
    df.columns = [str(c).strip().upper() for c in df.columns]
    df["_key"] = [_key(mz, lt) for mz, lt in zip(df["MZ"], df["LT"])]
    return df[df["_key"].map(lambda k: bool(k[0]) and bool(k[1]))]


def _cargar_boletas() -> pd.DataFrame:
    df = pd.read_excel(DATA_BOLETAS_PATH, sheet_name="Data", dtype=str)
    df["_key"] = [_key(mz, lt) for mz, lt in zip(df["MZ"], df["LT"])]
    return df


def comparar() -> tuple[list[dict], list[dict], list[dict]]:
    """Devuelve (resumen, discrepancias, consistencia_config)."""
    df_pla = _cargar_planilla()
    df_bol = _cargar_boletas()
    config = _cargar_config()

    pla_por_key = {r["_key"]: r for _, r in df_pla.iterrows()}
    bol_por_key = {}
    duplicados = []
    for _, r in df_bol.iterrows():
        if r["_key"] in bol_por_key:
            duplicados.append(r["_key"])
        bol_por_key[r["_key"]] = r

    resumen = []
    discrepancias = []

    # ── Chequeo 1: columnas 1:1 ──────────────────────────────────────────
    comunes = [k for k in pla_por_key if k in bol_por_key]
    for col_bol, col_pla, tipo in COLS_1A1:
        malas = 0
        for k in comunes:
            v_pla, v_bol = pla_por_key[k].get(col_pla), bol_por_key[k].get(col_bol)
            if tipo == "num":
                ok = abs(_float(v_pla) - _float(v_bol)) <= TOLERANCIA
            else:
                ok = str(v_pla or "").strip() == str(v_bol or "").strip()
            if not ok:
                malas += 1
                discrepancias.append({
                    "MZ": k[0], "LT": k[1], "COLUMNA": col_bol,
                    "VALOR_PLANILLA": v_pla, "VALOR_BOLETAS": v_bol,
                    "DIFERENCIA_TIPO": round(_float(v_bol) - _float(v_pla), 2) if tipo == "num" else "TEXTO_DISTINTO",
                })
        resumen.append({"CHEQUEO": "1:1", "COLUMNA": col_bol,
                        "TOTAL_FILAS": len(comunes), "COINCIDEN": len(comunes) - malas,
                        "DISCREPANCIAS": malas})

    # ── Chequeo 2: fórmulas recalculadas ────────────────────────────────
    for col_bol in ("Total", "Importe a pagar"):
        malas = 0
        for k in comunes:
            esperado = max(0.0, sum(_float(pla_por_key[k].get(c)) for c in COMPONENTES_TOTAL))
            v_bol = _float(bol_por_key[k].get(col_bol))
            if abs(esperado - v_bol) > TOLERANCIA:
                malas += 1
                discrepancias.append({
                    "MZ": k[0], "LT": k[1], "COLUMNA": col_bol,
                    "VALOR_PLANILLA": round(esperado, 2), "VALOR_BOLETAS": round(v_bol, 2),
                    "DIFERENCIA_TIPO": round(v_bol - esperado, 2),
                })
        resumen.append({"CHEQUEO": "FORMULA", "COLUMNA": col_bol,
                        "TOTAL_FILAS": len(comunes), "COINCIDEN": len(comunes) - malas,
                        "DISCREPANCIAS": malas})

    # ── Chequeo 3: cobertura de predios ─────────────────────────────────
    faltan = [k for k in pla_por_key if k not in bol_por_key]
    sobran = [k for k in bol_por_key if k not in pla_por_key]
    for k in faltan:
        discrepancias.append({"MZ": k[0], "LT": k[1], "COLUMNA": "(MZ, LT)",
                              "VALOR_PLANILLA": "presente", "VALOR_BOLETAS": "ausente",
                              "DIFERENCIA_TIPO": "FALTA_EN_BOLETAS"})
    for k in sobran:
        discrepancias.append({"MZ": k[0], "LT": k[1], "COLUMNA": "(MZ, LT)",
                              "VALOR_PLANILLA": "ausente", "VALOR_BOLETAS": "presente",
                              "DIFERENCIA_TIPO": "SOBRA_EN_BOLETAS"})
    for k in duplicados:
        discrepancias.append({"MZ": k[0], "LT": k[1], "COLUMNA": "(MZ, LT)",
                              "VALOR_PLANILLA": "1 fila", "VALOR_BOLETAS": "2+ filas",
                              "DIFERENCIA_TIPO": "DUPLICADO"})
    n_cob = len(faltan) + len(sobran) + len(duplicados)
    resumen.append({"CHEQUEO": "COBERTURA", "COLUMNA": "(MZ, LT)",
                    "TOTAL_FILAS": len(pla_por_key), "COINCIDEN": len(comunes),
                    "DISCREPANCIAS": n_cob})

    # ── Chequeo 4: consistencia de config ───────────────────────────────
    consistencia = []
    for col_bol, key_cfg in COLS_CONFIG:
        esperado = _fecha_pago_str(config) if key_cfg is None else config[key_cfg]
        valores = df_bol[col_bol].fillna("").map(lambda s: str(s).strip())
        no_coinciden = int((valores != str(esperado).strip()).sum())
        consistencia.append({
            "COLUMNA": col_bol, "VALOR_CONFIG": esperado,
            "VALORES_DISTINTOS_EN_BOLETAS": int(valores.nunique()),
            "FILAS_QUE_NO_COINCIDEN": no_coinciden,
            "OK": "SÍ" if no_coinciden == 0 else "NO",
        })
        resumen.append({"CHEQUEO": "CONFIG", "COLUMNA": col_bol,
                        "TOTAL_FILAS": len(df_bol), "COINCIDEN": len(df_bol) - no_coinciden,
                        "DISCREPANCIAS": no_coinciden})

    # NUMERO DE RECIBO correlativo desde NUMERO_RECIBO_INICIO
    inicio = config["NUMERO_RECIBO_INICIO"]
    recibos = df_bol["NUMERO DE RECIBO"].map(lambda v: int(_float(v)))
    esperados = list(range(inicio, inicio + len(df_bol)))
    malos_recibo = int((recibos != pd.Series(esperados, index=recibos.index)).sum())
    consistencia.append({
        "COLUMNA": "NUMERO DE RECIBO", "VALOR_CONFIG": f"inicio {inicio}",
        "VALORES_DISTINTOS_EN_BOLETAS": "correlativo" if malos_recibo == 0 else "con saltos",
        "FILAS_QUE_NO_COINCIDEN": malos_recibo,
        "OK": "SÍ" if malos_recibo == 0 else "NO",
    })
    resumen.append({"CHEQUEO": "CONFIG", "COLUMNA": "NUMERO DE RECIBO",
                    "TOTAL_FILAS": len(df_bol), "COINCIDEN": len(df_bol) - malos_recibo,
                    "DISCREPANCIAS": malos_recibo})

    return resumen, discrepancias, consistencia


# ── Escritura Excel ──────────────────────────────────────────────────────

_HEAD_BG = "1E5C3A"
_OK_BG   = "E9F7EF"
_BAD_BG  = "FEE2E2"
_BAD_FG  = "991B1B"


def _escribir_hoja(ws, cols: list[str], filas: list[dict], widths: list[int],
                   col_mal: str, es_mala) -> None:
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.fill = PatternFill("solid", fgColor="FF" + _HEAD_BG)
        cell.font = Font(color="FFFFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
    for r, fila in enumerate(filas, start=2):
        for i, c in enumerate(cols, start=1):
            val = fila.get(c, "")
            cell = ws.cell(row=r, column=i, value=val)
            if c == col_mal and es_mala(fila):
                cell.fill = PatternFill("solid", fgColor="FF" + _BAD_BG)
                cell.font = Font(color="FF" + _BAD_FG, bold=True)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"


def escribir_reporte(mes: str, resumen, discrepancias, consistencia) -> Path:
    wb = Workbook()

    ws = wb.active
    ws.title = "Resumen"
    _escribir_hoja(ws, ["CHEQUEO", "COLUMNA", "TOTAL_FILAS", "COINCIDEN", "DISCREPANCIAS"],
                   resumen, [12, 26, 12, 12, 15],
                   "DISCREPANCIAS", lambda f: f["DISCREPANCIAS"] > 0)

    ws = wb.create_sheet("Discrepancias")
    _escribir_hoja(ws, ["MZ", "LT", "COLUMNA", "VALOR_PLANILLA", "VALOR_BOLETAS", "DIFERENCIA_TIPO"],
                   discrepancias, [6, 7, 26, 18, 18, 18],
                   "DIFERENCIA_TIPO", lambda f: True)

    ws = wb.create_sheet("Consistencia_Config")
    _escribir_hoja(ws, ["COLUMNA", "VALOR_CONFIG", "VALORES_DISTINTOS_EN_BOLETAS",
                        "FILAS_QUE_NO_COINCIDEN", "OK"],
                   consistencia, [22, 44, 28, 22, 6],
                   "OK", lambda f: f["OK"] != "SÍ")

    out = OUTPUTS_DIR / f"validacion_enriquecimiento_{mes}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)s  %(message)s")
    mes = PLANILLA_PATH.stem.replace("planilla_", "")
    log.info(f"Validando enriquecimiento: planilla_{mes}.xlsx vs DATA_boletas.xlsx")
    resumen, discrepancias, consistencia = comparar()
    out = escribir_reporte(mes, resumen, discrepancias, consistencia)

    total_disc = sum(r["DISCREPANCIAS"] for r in resumen)
    log.info(f"Reporte: {out}")
    if total_disc:
        log.warning(f"{total_disc} discrepancia(s) — revisar hojas Discrepancias / Consistencia_Config")
    else:
        log.info("Sin discrepancias — DATA_boletas refleja fielmente la planilla y la config")


if __name__ == "__main__":
    main()
