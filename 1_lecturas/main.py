"""1_lecturas/main.py — valida lecturas mensuales y orquesta el ciclo de corrección.

Lee:
    inputs/registro_operario_mes.xlsx       — operario llena MARC_ACT, M3, obs_operario
    inputs/registro_operario_acumulado.xlsx — historial con doble header (MARCACION, M3)
    outputs/correcciones_YYYY-MM.xlsx       — solo en ciclo 2+ (cola de maquillaje)

Detecta 11 anomalías (7 bloqueantes + 4 informativas) y genera:
    outputs/correcciones_YYYY-MM.xlsx     — bloqueantes pendientes (si quedan)
    outputs/trazabilidad_YYYY-MM.xlsx     — registro permanente acumulativo
    outputs/lecturas_planilla_YYYY-MM.xlsx — input limpio para 2_planilla (solo al cerrar)
    inputs/registro_operario_acumulado.xlsx — actualizado con par del ciclo (solo al cerrar)

La detección de ciclo se basa en la existencia de correcciones_YYYY-MM.xlsx:
    no existe → Ciclo 1 (primer pass del mes)
    existe    → Ciclo 2+ (lee maquillaje y regenera o cierra)
"""
from __future__ import annotations

import logging
import statistics
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
import sys as _sys
_sys.path.insert(0, str(Path(__file__).parent.parent))  # para shared/
import config
from pdf_orden import generar_pdf
from shared.utils_sort_mz_lt import clave_orden
import formato_excel as fe

log = logging.getLogger(__name__)

# ── Tipos de anomalía ─────────────────────────────────────────────────────────
BLOQUEANTES = {
    "MEDIDOR_INVERTIDO", "POSIBLE_CAMBIO_MEDIDOR",
    "DIFERENCIA_M3", "EXCESIVO", "SIN_LECTURA",
    "DUPLICADO", "USUARIO_FANTASMA", "MARC_ACT_NO_NUMERICO",
}
INFORMATIVAS = {
    "SIN_HISTORIAL", "CONSUMO_CERO", "SALTO_HISTORICO",
    "MEDIDOR_CAMBIADO", "FUGA_REPORTADA",
}

# Mensajes legibles para columna motivo_detectado del reporte
def _msg_anomalia(tipo: str, **kw) -> str:
    if tipo == "MEDIDOR_INVERTIDO":
        return (f"MARC_ACT ({kw['act']}) < MARC_ANT ({kw['ant']}) · diferencia {kw['act']-kw['ant']:+.1f} m³ "
                f"(≤ umbral {config.UMBRAL_INVERSION}) — posible medidor instalado al revés")
    if tipo == "POSIBLE_CAMBIO_MEDIDOR":
        return (f"MARC_ACT ({kw['act']}) << MARC_ANT ({kw['ant']}) · diferencia {kw['act']-kw['ant']:+.1f} m³ "
                f"(> umbral {config.UMBRAL_INVERSION}) — probable medidor cambiado sin reportar")
    if tipo == "DIFERENCIA_M3":
        return f"calc={kw['calc']} (act−ant) pero operario anotó M3={kw['m3']} · |Δ|>0.05"
    if tipo == "EXCESIVO":
        return f"consumo calculado {kw['calc']:.1f} m³ > umbral {config.M3_EXCESIVO}"
    if tipo == "SIN_LECTURA":
        return "MARC_ACT vacío · sin código obs_operario que justifique"
    if tipo == "DUPLICADO":
        return f"la fila {kw['mz']}-{kw['lt']} aparece {kw['n']} veces en el template"
    if tipo == "USUARIO_FANTASMA":
        return "presente en acumulado pero falta en registro_operario_mes"
    if tipo == "MARC_ACT_NO_NUMERICO":
        return f"MARC_ACT={kw['raw']!r} contiene caracteres no numéricos"
    if tipo == "SIN_HISTORIAL":
        return "usuario nuevo · sin ciclos anteriores en acumulado"
    if tipo == "CONSUMO_CERO":
        return "act = ant y M3 = 0 · sin mínimo aplicado"
    if tipo == "SALTO_HISTORICO":
        return f"M3={kw['m3']:.1f} > {config.SALTO_FACTOR}× promedio últimos {config.SALTO_MESES} ciclos ({kw['prom']:.1f})"
    if tipo == "MEDIDOR_CAMBIADO":
        return f"MARC_ACT ({kw['act']}) < MARC_ANT ({kw['ant']}) · obs=M legitima (medidor cambiado)"
    if tipo == "FUGA_REPORTADA":
        return "obs=F · operario reportó fuga visible en el predio"
    return tipo


# ── LOGGING ───────────────────────────────────────────────────────────────────
def _init_logging() -> None:
    config.LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(config.LOG_PATH, encoding="utf-8"),
        ],
        force=True,
    )


# ── HELPERS DE PARSING ────────────────────────────────────────────────────────
def _str_clean(val) -> str:
    """Normaliza celdas vacías de Excel (NaN, None, 'nan') a ''."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("", "nan", "none") else s


def _norm_lt(val) -> str:
    s = _str_clean(val)
    if not s:
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.upper()


def _try_float(s: str) -> float | None:
    """Intenta convertir a float. Devuelve None si no es numérico válido."""
    if not s:
        return None
    try:
        return float(str(s).replace(",", "."))
    except (ValueError, TypeError):
        return None


# ── CARGA: REGISTRO MES ───────────────────────────────────────────────────────
def _cargar_registro_mes() -> list[dict]:
    """Lee el template del operario, salta la leyenda y devuelve las filas válidas.

    El archivo tiene leyenda M/F/P en las primeras filas y luego una fila con los
    headers reales (MZ, LT, NOMBRE, MES_ANO, MARC_ANT, MARC_ACT, M3, obs_operario).
    Localiza el header buscando la fila que contiene "MZ" en columna 1.
    """
    wb = load_workbook(config.REGISTRO_MES_PATH, data_only=True)
    ws = wb.active

    # Buscar fila de header
    header_row = None
    for ri in range(1, min(ws.max_row, 15) + 1):
        if str(ws.cell(ri, 1).value or "").strip().upper() == "MZ":
            header_row = ri
            break
    if header_row is None:
        raise ValueError(
            f"No se encontró el header (fila con 'MZ' en columna 1) en {config.REGISTRO_MES_PATH}"
        )

    headers = [str(ws.cell(header_row, c).value or "").strip().upper()
               for c in range(1, ws.max_column + 1)]

    requeridas = {"MZ", "LT", "NOMBRE", "MES_ANO", "MARC_ANT", "MARC_ACT", "M3"}
    faltantes = requeridas - set(headers)
    if faltantes:
        raise ValueError(f"Columnas faltantes en {config.REGISTRO_MES_PATH.name}: {faltantes}")

    idx = {h: i for i, h in enumerate(headers)}
    # obs_operario es opcional
    has_obs = "OBS_OPERARIO" in idx

    filas = []
    for ri in range(header_row + 1, ws.max_row + 1):
        row_vals = [ws.cell(ri, c).value for c in range(1, ws.max_column + 1)]
        mz = str(row_vals[idx["MZ"]] or "").strip().upper()
        lt = _norm_lt(row_vals[idx["LT"]])
        if not mz or not lt:
            continue
        obs = _str_clean(row_vals[idx["OBS_OPERARIO"]]).upper() if has_obs else ""
        if obs and obs not in config.OBS_CODIGOS:
            raise ValueError(
                f"obs_operario tiene valor inválido: {obs!r} en fila {ri} "
                f"({mz}-{lt}) · solo se aceptan {sorted(config.OBS_CODIGOS)} o vacío"
            )
        filas.append({
            "mz":           mz,
            "lt":           lt,
            "nombre":       _str_clean(row_vals[idx["NOMBRE"]]),
            "mes_ano":      _str_clean(row_vals[idx["MES_ANO"]]),
            "marc_ant":     _str_clean(row_vals[idx["MARC_ANT"]]),
            "marc_act":     _str_clean(row_vals[idx["MARC_ACT"]]),
            "m3":           _str_clean(row_vals[idx["M3"]]),
            "obs_operario": obs,
        })
    log.info(f"Lecturas cargadas: {len(filas)} filas")
    return filas


# ── CARGA: REGISTRO ACUMULADO (doble header) ──────────────────────────────────
def _cargar_acumulado() -> tuple[dict, list[str]]:
    """Devuelve (historial, meses_orden):
        historial: {(mz, lt): {"nombre": str, "ciclos": {mes: {"marc": float, "m3": float|None}}}}
        meses_orden: lista ordenada de meses presentes (YYYY-MM).

    Si el archivo no existe (primer ciclo del sistema), devuelve ({}, []).
    """
    if not config.REGISTRO_ACUMULADO_PATH.exists():
        return {}, []

    wb = load_workbook(config.REGISTRO_ACUMULADO_PATH, data_only=True)
    ws = wb.active

    # Detectar schema v2 (col 4 = SIN_SERVICIO) vs legacy
    cab_col4 = str(ws.cell(1, 4).value or "").strip().upper()
    tiene_sin_servicio = cab_col4 == "SIN_SERVICIO"
    col_inicio_meses = 5 if tiene_sin_servicio else 4

    n_cols = ws.max_column
    fila1 = [ws.cell(1, c).value for c in range(1, n_cols + 1)]
    fila2 = [ws.cell(2, c).value for c in range(1, n_cols + 1)]

    # Mapear meses al par de columnas (MARCACION, M3) — a partir de col_inicio_meses
    mes_cols: dict[str, tuple[int, int]] = {}
    mes_actual = None
    marc_col = None
    for ci in range(col_inicio_meses, n_cols + 1):
        v1 = fila1[ci - 1]
        v2 = fila2[ci - 1]
        if isinstance(v1, str) and len(v1) == 7 and v1[4] == "-":
            mes_actual = v1
            marc_col = None
        if mes_actual and v2 == "MARCACION":
            marc_col = ci
        elif mes_actual and v2 == "M3" and marc_col is not None:
            mes_cols[mes_actual] = (marc_col, ci)
            mes_actual = None
            marc_col = None

    meses_orden = sorted(mes_cols.keys())

    historial: dict = {}
    for ri in range(3, ws.max_row + 1):
        mz = str(ws.cell(ri, 1).value or "").strip().upper()
        lt = _norm_lt(ws.cell(ri, 2).value)
        if not mz or not lt:
            continue
        nombre = _str_clean(ws.cell(ri, 3).value)
        sin_serv = ""
        if tiene_sin_servicio:
            sin_serv_raw = str(ws.cell(ri, 4).value or "").strip().lower()
            sin_serv = "Si" if sin_serv_raw.startswith("si") else ""
        ciclos = {}
        for mes, (col_marc, col_m3) in mes_cols.items():
            marc = _try_float(_str_clean(ws.cell(ri, col_marc).value))
            m3 = _try_float(_str_clean(ws.cell(ri, col_m3).value))
            if marc is not None:
                ciclos[mes] = {"marc": marc, "m3": m3}
        historial[(mz, lt)] = {"nombre": nombre, "sin_servicio": sin_serv, "ciclos": ciclos}

    log.info(f"Acumulado: {len(historial)} usuarios · {len(meses_orden)} ciclos previos · "
             f"schema {'v2 (SIN_SERVICIO)' if tiene_sin_servicio else 'v1 (legacy)'}")
    return historial, meses_orden


# ── CICLO ─────────────────────────────────────────────────────────────────────
def _detectar_ciclo(mes_ano: str) -> int:
    """Devuelve el número de ciclo actual.

    Regla: si correcciones_YYYY-MM.xlsx no existe → Ciclo 1.
    Si existe → leer el max(ciclo) del archivo y sumar 1.
    """
    path = config.correcciones_path(mes_ano)
    if not path.exists():
        return 1
    try:
        existentes = fe.leer_filas_existentes(path, fe.GRUPOS_CORRECCIONES)
        ciclos = []
        for r in existentes:
            try:
                ciclos.append(int(r.get("ciclo")))
            except (TypeError, ValueError):
                pass
        return (max(ciclos) + 1) if ciclos else 2
    except Exception as e:
        log.warning(f"No se pudo leer ciclo de correcciones existente ({e}) · asumo Ciclo 2")
        return 2


# ── APLICAR CORRECCIONES (Ciclo 2+) ───────────────────────────────────────────
def _aplicar_correcciones(filas: list[dict], mes_ano: str) -> tuple[list[dict], list[dict], dict]:
    """Si existe correcciones_YYYY-MM.xlsx, lee los maquillajes resueltos y los aplica
    sobre 'filas' (sustituyendo MARC_ACT y M3). Devuelve la lista de filas modificadas,
    la lista de cerradas para trazabilidad, y un mapa (mz,lt) → resuelto_por para que
    _detectar_anomalias sepa qué casos bypassear (ej: acepta_original).
    """
    path = config.correcciones_path(mes_ano)
    if not path.exists():
        return filas, [], {}

    # Leer con el helper que entiende el formato de grupos (headers en fila 2, datos desde fila 3)
    existentes = fe.leer_filas_existentes(path, fe.GRUPOS_CORRECCIONES)

    # Map (mz, lt) -> fila completa con campos relevantes
    correcciones_map = {}
    for r in existentes:
        mz = str(r.get("MZ") or "").strip().upper()
        lt = _norm_lt(r.get("LT"))
        if not mz or not lt:
            continue
        correcciones_map[(mz, lt)] = {
            "resuelto":   _str_clean(r.get("resuelto_por")),
            "marc_corr":  _try_float(_str_clean(r.get("MARC_ACT_corregido"))),
            "m3_corr":    _try_float(_str_clean(r.get("M3_corregido"))),
            "motivo":     _str_clean(r.get("motivo_correccion")),
            "tipo":       _str_clean(r.get("tipo_anomalia")),
            "ciclo":      _str_clean(r.get("ciclo")),
            "fecha":      _str_clean(r.get("fecha_correccion")),
            "mes_ano":    _str_clean(r.get("MES_ANO")),
            "marc_ant":   _str_clean(r.get("MARC_ANT")),
            "marc_orig":  _str_clean(r.get("MARC_ACT_original")),
            "m3_orig":    _str_clean(r.get("M3_original")),
            "obs_orig":   _str_clean(r.get("obs_operario_original")),
            "nombre":     _str_clean(r.get("NOMBRE")),
        }

    # Aplicar maquillaje a filas del mes
    cerradas = []
    filas_out = []
    eliminar_keys = set()
    resoluciones = {}  # (mz, lt) -> resuelto_por (para que _detectar_anomalias sepa qué bypassear)
    for f in filas:
        key = (f["mz"], f["lt"])
        c = correcciones_map.get(key)
        if c and c["resuelto"]:
            resoluciones[key] = c["resuelto"]
            if c["resuelto"] in ("borra_duplicado", "marca_baja"):
                eliminar_keys.add(key)
            else:
                if c["marc_corr"] is not None:
                    f["marc_act"] = str(c["marc_corr"])
                if c["m3_corr"] is not None:
                    f["m3"] = str(c["m3_corr"])
            cerradas.append({**c, "mz": key[0], "lt": key[1]})
        filas_out.append(f)

    # Eliminar las marcadas para baja/duplicado (mantener solo una vez en caso de duplicados)
    if eliminar_keys:
        vistos = set()
        filtradas = []
        for f in filas_out:
            k = (f["mz"], f["lt"])
            if k in eliminar_keys:
                if k in vistos:
                    continue
                vistos.add(k)
                # Si es marca_baja, no aparece. Si es borra_duplicado, dejamos UNA.
                resuelto = correcciones_map.get(k, {}).get("resuelto", "")
                if resuelto == "marca_baja":
                    continue
                # borra_duplicado: dejamos esta única instancia
                filtradas.append(f)
            else:
                filtradas.append(f)
        filas_out = filtradas

    log.info(f"Correcciones aplicadas: {len(cerradas)} filas cerradas")
    return filas_out, cerradas, resoluciones


# ── DETECCIÓN DE ANOMALÍAS ────────────────────────────────────────────────────
def _detectar_anomalias(
    filas: list[dict],
    historial: dict,
    meses_orden: list[str],
    resoluciones: dict | None = None,
) -> tuple[list[dict], list[dict], list[dict]]:
    """Devuelve (confirmados, bloqueantes, informativas).

    Aplica las 11 reglas según el orden documentado.
    `resoluciones` mapea (mz, lt) → resuelto_por para filas que ya pasaron por
    correcciones; los valores 'acepta_original' bypasan la detección y van directo
    a confirmados (con origen='corregido') aunque sigan disparando reglas.
    """
    resoluciones = resoluciones or {}
    confirmados, bloqueantes, informativas = [], [], []

    # Regla DUPLICADO: detectar (MZ, LT) que aparecen 2+ veces
    counts = Counter((f["mz"], f["lt"]) for f in filas)
    duplicados_keys = {k for k, n in counts.items() if n > 1}

    # Regla USUARIO_FANTASMA: usuarios en acumulado que no aparecen en mes
    keys_en_mes = {(f["mz"], f["lt"]) for f in filas}
    for key, datos in historial.items():
        if key not in keys_en_mes:
            informativas_o_bloqueantes = bloqueantes  # USUARIO_FANTASMA es bloqueante
            informativas_o_bloqueantes.append({
                "mz": key[0], "lt": key[1],
                "nombre": datos.get("nombre", ""),
                "mes_ano": filas[0]["mes_ano"] if filas else "",
                "marc_ant": "", "marc_act": "", "m3": "", "obs_operario": "",
                "tipo": "USUARIO_FANTASMA",
                "motivo": _msg_anomalia("USUARIO_FANTASMA"),
                "calc_m3": None,
            })

    # Procesar cada fila del mes
    for f in filas:
        key = (f["mz"], f["lt"])

        # Bypass: si ya se resolvió por correcciones con acepta_original/maquillaje,
        # tratar como confirmado sin re-evaluar las reglas.
        resuelto = resoluciones.get(key)
        if resuelto in ("acepta_original", "maquillaje", "campo", "corrige_dato"):
            ciclos_h = historial.get(key, {}).get("ciclos", {})
            meses_prev = [m for m in ciclos_h if m < f["mes_ano"]] if f["mes_ano"] else list(ciclos_h)
            marc_ant_hist = ciclos_h[max(meses_prev)]["marc"] if meses_prev else None
            confirmados.append({
                **f, "marc_ant_hist": marc_ant_hist,
                "marc_act_val": _try_float(f["marc_act"]),
                "m3_val": _try_float(f["m3"]),
                "calc_m3": None, "origen": "corregido",
            })
            continue

        # DUPLICADO
        if key in duplicados_keys:
            bloqueantes.append({
                **f,
                "tipo": "DUPLICADO",
                "motivo": _msg_anomalia("DUPLICADO", mz=f["mz"], lt=f["lt"], n=counts[key]),
                "calc_m3": None,
            })
            continue

        # Resolver MARC_ANT del historial (último ciclo previo al actual)
        ciclos = historial.get(key, {}).get("ciclos", {})
        meses_prev = [m for m in ciclos if m < f["mes_ano"]] if f["mes_ano"] else list(ciclos)
        marc_ant_hist = None
        if meses_prev:
            ultimo = max(meses_prev)
            marc_ant_hist = ciclos[ultimo]["marc"]

        # MARC_ACT_NO_NUMERICO: hay string pero no parsea
        if f["marc_act"] and _try_float(f["marc_act"]) is None:
            bloqueantes.append({
                **f, "marc_ant_hist": marc_ant_hist,
                "tipo": "MARC_ACT_NO_NUMERICO",
                "motivo": _msg_anomalia("MARC_ACT_NO_NUMERICO", raw=f["marc_act"]),
                "calc_m3": None,
            })
            continue

        marc_act_val = _try_float(f["marc_act"])
        m3_op        = _try_float(f["m3"])
        obs          = f["obs_operario"]

        # SIN_LECTURA (con excepción obs=P → informativo)
        if marc_act_val is None:
            if obs == "P":
                informativas.append({
                    **f, "marc_ant_hist": marc_ant_hist,
                    "tipo": "SIN_LECTURA", "categoria": "informativa",
                    "motivo": "MARC_ACT vacío · obs=P legitima (predio cerrado o inaccesible)",
                    "resuelto_por": "obs_operario_legitima",
                    "calc_m3": None,
                })
            else:
                bloqueantes.append({
                    **f, "marc_ant_hist": marc_ant_hist,
                    "tipo": "SIN_LECTURA",
                    "motivo": _msg_anomalia("SIN_LECTURA"),
                    "calc_m3": None,
                })
            continue

        # SIN_HISTORIAL (informativa) — usuario nuevo, aceptar
        if marc_ant_hist is None:
            informativas.append({
                **f, "marc_ant_hist": None,
                "tipo": "SIN_HISTORIAL", "categoria": "informativa",
                "motivo": _msg_anomalia("SIN_HISTORIAL"),
                "resuelto_por": "informativa_aceptada",
                "calc_m3": None,
            })
            confirmados.append({**f, "marc_ant_hist": None, "marc_act_val": marc_act_val,
                               "m3_val": m3_op, "calc_m3": None, "origen": "informativa_legitimada"})
            continue

        calc = round(marc_act_val - marc_ant_hist, 3)

        # FUGA_REPORTADA (informativa, independiente del consumo)
        if obs == "F":
            informativas.append({
                **f, "marc_ant_hist": marc_ant_hist,
                "tipo": "FUGA_REPORTADA", "categoria": "informativa",
                "motivo": _msg_anomalia("FUGA_REPORTADA"),
                "resuelto_por": "obs_operario_legitima",
                "calc_m3": calc,
            })
            # F no impide procesar las demás reglas — se evalúa además

        # MARC_ACT < MARC_ANT: distinguir 3 casos
        # 1. obs=M legitima → MEDIDOR_CAMBIADO informativo
        # 2. diferencia ≤ UMBRAL_INVERSION → MEDIDOR_INVERTIDO bloqueante
        # 3. diferencia > UMBRAL_INVERSION → POSIBLE_CAMBIO_MEDIDOR bloqueante
        if marc_act_val < marc_ant_hist:
            if obs == "M":
                informativas.append({
                    **f, "marc_ant_hist": marc_ant_hist,
                    "tipo": "MEDIDOR_CAMBIADO", "categoria": "informativa",
                    "motivo": _msg_anomalia("MEDIDOR_CAMBIADO", act=marc_act_val, ant=marc_ant_hist),
                    "resuelto_por": "obs_operario_legitima",
                    "calc_m3": None,
                })
                confirmados.append({**f, "marc_ant_hist": marc_ant_hist, "marc_act_val": marc_act_val,
                                   "m3_val": m3_op, "calc_m3": None, "origen": "informativa_legitimada"})
                continue

            diff_abs = marc_ant_hist - marc_act_val
            tipo_inv = "MEDIDOR_INVERTIDO" if diff_abs <= config.UMBRAL_INVERSION else "POSIBLE_CAMBIO_MEDIDOR"
            bloqueantes.append({
                **f, "marc_ant_hist": marc_ant_hist,
                "tipo": tipo_inv,
                "motivo": _msg_anomalia(tipo_inv, act=marc_act_val, ant=marc_ant_hist),
                "calc_m3": calc,
            })
            continue

        # EXCESIVO (calc > M3_EXCESIVO)
        if calc > config.M3_EXCESIVO:
            bloqueantes.append({
                **f, "marc_ant_hist": marc_ant_hist,
                "tipo": "EXCESIVO",
                "motivo": _msg_anomalia("EXCESIVO", calc=calc),
                "calc_m3": calc,
            })
            continue

        # DIFERENCIA_M3 (con excepción mínimo aplicado)
        minimo_aplicado = (calc < config.M3_MINIMO) and (m3_op is not None and abs(m3_op - config.M3_MINIMO) < 0.001)
        if not minimo_aplicado and m3_op is not None and abs(calc - m3_op) > 0.05:
            bloqueantes.append({
                **f, "marc_ant_hist": marc_ant_hist,
                "tipo": "DIFERENCIA_M3",
                "motivo": _msg_anomalia("DIFERENCIA_M3", calc=calc, m3=m3_op),
                "calc_m3": calc,
            })
            continue

        # CONSUMO_CERO (informativa)
        if calc == 0 and m3_op is not None and m3_op == 0:
            informativas.append({
                **f, "marc_ant_hist": marc_ant_hist,
                "tipo": "CONSUMO_CERO", "categoria": "informativa",
                "motivo": _msg_anomalia("CONSUMO_CERO"),
                "resuelto_por": "informativa_aceptada",
                "calc_m3": 0,
            })

        # SALTO_HISTORICO (informativa) — requiere SALTO_MESES ciclos previos con M3
        m3_finals = [ciclos[m]["m3"] for m in meses_prev if ciclos[m].get("m3") is not None]
        if len(m3_finals) >= config.SALTO_MESES and m3_op is not None:
            ultimos = sorted(meses_prev)[-config.SALTO_MESES:]
            valores = [ciclos[m]["m3"] for m in ultimos if ciclos[m].get("m3") is not None]
            if len(valores) >= config.SALTO_MESES:
                prom = statistics.mean(valores)
                if prom > 0 and m3_op > config.SALTO_FACTOR * prom:
                    informativas.append({
                        **f, "marc_ant_hist": marc_ant_hist,
                        "tipo": "SALTO_HISTORICO", "categoria": "informativa",
                        "motivo": _msg_anomalia("SALTO_HISTORICO", m3=m3_op, prom=prom),
                        "resuelto_por": "informativa_aceptada",
                        "calc_m3": calc,
                    })

        # Confirmado (con o sin anotación de mínimo)
        origen = "minimo_aplicado" if minimo_aplicado else "directo"
        confirmados.append({**f, "marc_ant_hist": marc_ant_hist, "marc_act_val": marc_act_val,
                           "m3_val": m3_op, "calc_m3": calc, "origen": origen})

    return confirmados, bloqueantes, informativas


# Borde común reutilizado por _actualizar_acumulado (los exportadores usan formato_excel)
def _borde():
    b = Side(style="thin", color="CCCCCC")
    return Border(left=b, right=b, top=b, bottom=b)


# ── EXPORTAR: correcciones_YYYY-MM.xlsx ───────────────────────────────────────
def _exportar_correcciones(bloqueantes: list[dict], ciclo: int, mes_ano: str,
                          cerradas_previas: list[dict]) -> None:
    """Genera correcciones_YYYY-MM.xlsx con los bloqueantes que aún quedan pendientes.

    Aplica formato con grupos narrativos según contrato_correcciones.html.
    Si bloqueantes está vacío, elimina el archivo (señal de mes cerrado).
    """
    path = config.correcciones_path(mes_ano)

    if not bloqueantes:
        if path.exists():
            path.unlink()
            log.info(f"correcciones_{mes_ano}.xlsx eliminado · no quedan bloqueantes")
        return

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Correcciones"

    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    filas = []
    for b in bloqueantes:
        filas.append({
            "MZ": b["mz"], "LT": b["lt"], "NOMBRE": b.get("nombre", ""),
            "MES_ANO": b.get("mes_ano", ""),
            "MARC_ANT": b.get("marc_ant_hist") if b.get("marc_ant_hist") is not None else b.get("marc_ant", ""),
            "MARC_ACT_original": b.get("marc_act", ""),
            "M3_original": b.get("m3", ""),
            "obs_operario_original": b.get("obs_operario", ""),
            "tipo_anomalia": b["tipo"],
            "motivo_detectado": b["motivo"],
            "MARC_ACT_corregido": "", "M3_corregido": "",
            "motivo_correccion": "", "resuelto_por": "",
            "estado": "pendiente", "ciclo": ciclo, "fecha_correccion": ts,
        })

    fe.escribir_con_grupos(ws, fe.GRUPOS_CORRECCIONES, filas)
    wb.save(path)
    log.info(f"correcciones_{mes_ano}.xlsx generado · {len(bloqueantes)} bloqueantes pendientes")


# ── EXPORTAR: trazabilidad_YYYY-MM.xlsx ───────────────────────────────────────
def _exportar_trazabilidad(informativas: list[dict], cerradas: list[dict],
                          ciclo: int, mes_ano: str) -> None:
    """Agrega filas al archivo de trazabilidad. Nunca borra. Evita duplicados por (MZ, LT, tipo).

    Estrategia: leer existentes en memoria, agregar nuevas, reescribir todo el archivo
    con formato de grupos. Para 100s de filas es rápido y deja el archivo siempre
    consistente con el contrato.
    """
    path = config.trazabilidad_path(mes_ano)
    path.parent.mkdir(parents=True, exist_ok=True)

    # Cargar existentes (si hay)
    filas_total: list[dict] = []
    existentes = set()  # (mz, lt, tipo)
    if path.exists():
        for fila in fe.leer_filas_existentes(path, fe.GRUPOS_TRAZABILIDAD):
            mz = str(fila.get("MZ") or "").strip().upper()
            lt = _norm_lt(fila.get("LT"))
            tipo = str(fila.get("tipo_anomalia") or "").strip()
            existentes.add((mz, lt, tipo))
            filas_total.append(fila)

    ts = datetime.now().strftime("%d/%m/%Y %H:%M")
    nuevas = 0

    # Informativas
    for inf in informativas:
        key = (inf["mz"], inf["lt"], inf["tipo"])
        if key in existentes:
            continue
        filas_total.append({
            "MZ": inf["mz"], "LT": inf["lt"], "NOMBRE": inf.get("nombre", ""),
            "MES_ANO": inf.get("mes_ano", ""),
            "MARC_ANT": inf.get("marc_ant_hist") if inf.get("marc_ant_hist") is not None else inf.get("marc_ant", ""),
            "MARC_ACT_original": inf.get("marc_act", ""),
            "M3_original": inf.get("m3", ""),
            "obs_operario_original": inf.get("obs_operario", ""),
            "categoria": inf.get("categoria", "informativa"),
            "tipo_anomalia": inf["tipo"],
            "motivo_detectado": inf["motivo"],
            "MARC_ACT_final": inf.get("marc_act", ""),
            "M3_final": inf.get("m3", ""),
            "motivo_correccion": "",
            "resuelto_por": inf.get("resuelto_por", "informativa_aceptada"),
            "ciclo": ciclo,
            "fecha_correccion": ts,
        })
        existentes.add(key)
        nuevas += 1

    # Bloqueantes cerradas
    for c in cerradas:
        key = (c["mz"], c["lt"], c.get("tipo", ""))
        if key in existentes:
            continue
        filas_total.append({
            "MZ": c["mz"], "LT": c["lt"], "NOMBRE": c.get("nombre", ""),
            "MES_ANO": c.get("mes_ano", ""),
            "MARC_ANT": c.get("marc_ant", ""),
            "MARC_ACT_original": c.get("marc_orig", ""),
            "M3_original": c.get("m3_orig", ""),
            "obs_operario_original": c.get("obs_orig", ""),
            "categoria": "bloqueante",
            "tipo_anomalia": c.get("tipo", ""),
            "motivo_detectado": "",
            "MARC_ACT_final": c.get("marc_corr") if c.get("marc_corr") is not None else "",
            "M3_final": c.get("m3_corr") if c.get("m3_corr") is not None else "",
            "motivo_correccion": c.get("motivo", ""),
            "resuelto_por": c.get("resuelto", ""),
            "ciclo": c.get("ciclo", ciclo),
            "fecha_correccion": c.get("fecha", ts),
        })
        existentes.add(key)
        nuevas += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Trazabilidad"
    fe.escribir_con_grupos(ws, fe.GRUPOS_TRAZABILIDAD, filas_total)
    wb.save(path)
    log.info(f"trazabilidad_{mes_ano}.xlsx: +{nuevas} filas (total: {len(filas_total)})")


# ── EXPORTAR: lecturas_planilla_YYYY-MM.xlsx ──────────────────────────────────
def _exportar_lecturas_planilla(confirmados: list[dict], ciclo: int, mes_ano: str) -> None:
    """Genera el archivo final con lecturas facturables. Solo se llama si bloqueantes = 0.

    Aplica formato con grupos narrativos según contrato_lecturas_planilla.html.
    """
    path = config.lecturas_planilla_path(mes_ano)
    path.parent.mkdir(parents=True, exist_ok=True)

    filas = []
    for c in confirmados:
        filas.append({
            "MZ": c["mz"], "LT": c["lt"], "NOMBRE": c.get("nombre", ""),
            "MES_ANO": c.get("mes_ano", ""),
            "MARC_ANT": c.get("marc_ant_hist") if c.get("marc_ant_hist") is not None else c.get("marc_ant", ""),
            "MARC_ACT": c.get("marc_act", ""),
            "M3": c.get("m3_val") if c.get("m3_val") is not None else c.get("m3", ""),
            "origen": c.get("origen", "directo"),
            "ciclo": ciclo,
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "LecturasPlanilla"
    fe.escribir_con_grupos(ws, fe.GRUPOS_LECTURAS_PLANILLA, filas)
    wb.save(path)
    log.info(f"lecturas_planilla_{mes_ano}.xlsx generado · {len(confirmados)} usuarios")


# ── ACTUALIZAR ACUMULADO (doble header) ───────────────────────────────────────
def _actualizar_acumulado(confirmados: list[dict], historial: dict, meses_orden: list[str],
                         mes_ano: str) -> None:
    """Agrega el par de columnas (MARCACION, M3) del ciclo al acumulado.

    Si el mes ya existe (re-corrida), sobrescribe el par.
    """
    # Construir mapa nuevo por usuario
    nuevos = {(c["mz"], c["lt"]): (c.get("marc_act_val"), c.get("m3_val"), c.get("nombre", ""))
              for c in confirmados}

    # Construir todos los usuarios (historial + nuevos)
    todas_keys = set(historial.keys()) | set(nuevos.keys())

    # Lista de meses incluyendo el actual
    meses_final = sorted(set(meses_orden) | {mes_ano})

    wb = Workbook()
    ws = wb.active
    ws.title = "Acumulada"
    borde = _borde()

    # Cabecera fila 1: MZ, LT, NOMBRE, SIN_SERVICIO + YYYY-MM por cada mes (merge horizontal)
    # Cabecera fila 2: MARCACION + M3 por cada mes
    headers_fijos = list(config.COLS_FIJAS)  # MZ, LT, NOMBRE, SIN_SERVICIO
    for ci, h in enumerate(headers_fijos, 1):
        c1 = ws.cell(row=1, column=ci, value=h)
        c2 = ws.cell(row=2, column=ci, value="")
        for c in (c1, c2):
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
            c.fill = PatternFill("solid", start_color="1E3A5F")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borde
        ws.merge_cells(start_row=1, start_column=ci, end_row=2, end_column=ci)

    # Anchos fijos
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 14  # SIN_SERVICIO

    # Pares de columnas por mes
    col_mes = {}  # mes -> (col_marc, col_m3)
    next_col = len(headers_fijos) + 1
    ultimo_mes = meses_final[-1]
    for mes in meses_final:
        color_mes = "0369A1" if mes == ultimo_mes else "5DADE2"
        color_sub = "AED6F1"
        # Fila 1: nombre del mes (merge)
        c1 = ws.cell(row=1, column=next_col, value=mes)
        c1.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c1.fill = PatternFill("solid", start_color=color_mes)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = borde
        ws.merge_cells(start_row=1, start_column=next_col, end_row=1, end_column=next_col + 1)
        # Fila 2: sub-headers
        for offset, sub in enumerate(["MARCACION", "M3"]):
            c2 = ws.cell(row=2, column=next_col + offset, value=sub)
            c2.font = Font(name="Arial", bold=True, color="1A5276", size=9)
            c2.fill = PatternFill("solid", start_color=color_sub)
            c2.alignment = Alignment(horizontal="center", vertical="center")
            c2.border = borde
        ws.column_dimensions[get_column_letter(next_col)].width = 13
        ws.column_dimensions[get_column_letter(next_col + 1)].width = 8
        col_mes[mes] = (next_col, next_col + 1)
        next_col += 2

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18

    # Datos
    nombre_por_key = {}
    for key, datos in historial.items():
        nombre_por_key[key] = datos.get("nombre", "")
    for key, (marc, m3, nom) in nuevos.items():
        if nom:
            nombre_por_key[key] = nom

    # Preservar SIN_SERVICIO del historial (lo gobierna aplicar_sincronizacion.py)
    sin_servicio_por_key = {
        k: v.get("sin_servicio", "") for k, v in historial.items()
    }

    # Orden canónico del pipeline (mismo que padron_reconciliado.xlsx)
    keys_ordenadas = sorted(todas_keys, key=lambda k: clave_orden(k[0], k[1]))
    for ri, key in enumerate(keys_ordenadas, 3):
        mz, lt = key
        nom = nombre_por_key.get(key, "")
        sin_serv = sin_servicio_por_key.get(key, "")
        # Cols fijas: MZ, LT, NOMBRE, SIN_SERVICIO
        for ci, val in enumerate([mz, lt, nom, sin_serv], 1):
            c = ws.cell(row=ri, column=ci, value=val if val else None)
            es_sin_serv_marcado = (ci == 4 and val == "Si")
            c.font = Font(name="Arial", size=9,
                          bold=es_sin_serv_marcado,
                          color="7B241C" if es_sin_serv_marcado else "333333")
            c.alignment = Alignment(horizontal="left" if ci == 3 else "center", vertical="center")
            c.border = borde
            if es_sin_serv_marcado:
                c.fill = PatternFill("solid", start_color="FADBD8")
        # Por mes
        ciclos_hist = historial.get(key, {}).get("ciclos", {})
        for mes in meses_final:
            col_marc, col_m3 = col_mes[mes]
            if mes == mes_ano:
                if key in nuevos:
                    marc, m3, _ = nuevos[key]
                    c1 = ws.cell(row=ri, column=col_marc, value=marc)
                    c2 = ws.cell(row=ri, column=col_m3, value=m3)
                else:
                    c1 = ws.cell(row=ri, column=col_marc, value=None)
                    c2 = ws.cell(row=ri, column=col_m3, value=None)
            else:
                cdata = ciclos_hist.get(mes, {})
                c1 = ws.cell(row=ri, column=col_marc, value=cdata.get("marc"))
                c2 = ws.cell(row=ri, column=col_m3, value=cdata.get("m3"))
            for c in (c1, c2):
                c.font = Font(name="Arial", size=9)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = borde
        ws.row_dimensions[ri].height = 15

    ws.freeze_panes = "E3"  # 4 cols fijas (MZ, LT, NOMBRE, SIN_SERVICIO)
    wb.save(config.REGISTRO_ACUMULADO_PATH)
    log.info(f"registro_operario_acumulado.xlsx actualizado · "
             f"{len(todas_keys)} usuarios · {len(meses_final)} ciclos")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "═" * 65)
    print("  1_lecturas — procesamiento de lecturas del ciclo")
    print("═" * 65)
    _init_logging()

    # 1. Validar inputs
    print("\n[1/6] Validando inputs...")
    if not config.REGISTRO_MES_PATH.exists():
        raise FileNotFoundError(
            f"Falta: {config.REGISTRO_MES_PATH}\n"
            f"  → correr crear_template.py primero"
        )

    # 2. Cargar
    print("[2/6] Cargando registro_operario_mes y acumulado...")
    filas = _cargar_registro_mes()
    historial, meses_orden = _cargar_acumulado()

    mes_ano = filas[0]["mes_ano"] if filas else datetime.now().strftime("%Y-%m")
    ciclo = _detectar_ciclo(mes_ano)
    print(f"       Mes: {mes_ano} · Ciclo: {ciclo}")

    # 3. Aplicar correcciones del ciclo previo (si las hay)
    print("[3/6] Aplicando correcciones del ciclo previo (si las hay)...")
    filas, cerradas, resoluciones = _aplicar_correcciones(filas, mes_ano)

    # 4. Detectar anomalías
    print("[4/6] Detectando anomalías (11 reglas)...")
    confirmados, bloqueantes, informativas = _detectar_anomalias(
        filas, historial, meses_orden, resoluciones
    )

    by_tipo_bloq = Counter(b["tipo"] for b in bloqueantes)
    by_tipo_inf  = Counter(i["tipo"] for i in informativas)

    # 5. Exportar
    print("[5/6] Exportando outputs...")
    _exportar_correcciones(bloqueantes, ciclo, mes_ano, cerradas)
    _exportar_trazabilidad(informativas, cerradas, ciclo, mes_ano)
    if bloqueantes:
        generar_pdf(bloqueantes, mes_ano)

    cerrado = (len(bloqueantes) == 0)
    if cerrado:
        _exportar_lecturas_planilla(confirmados, ciclo, mes_ano)
        _actualizar_acumulado(confirmados, historial, meses_orden, mes_ano)

    # 6. Reporte
    print("[6/6] Reporte final\n")
    print("═" * 65)
    print(f"  Mes: {mes_ano} · Ciclo: {ciclo}")
    print(f"  Confirmados: {len(confirmados)}")
    if bloqueantes:
        print(f"  Bloqueantes: {len(bloqueantes)}")
        for tipo, n in by_tipo_bloq.most_common():
            print(f"      {tipo:<24} {n}")
    if informativas:
        print(f"  Informativas: {len(informativas)}")
        for tipo, n in by_tipo_inf.most_common():
            print(f"      {tipo:<24} {n}")
    print()
    if cerrado:
        print(f"  CICLO CERRADO ✓")
        print(f"  → outputs/lecturas_planilla_{mes_ano}.xlsx")
        print(f"  → registro_operario_acumulado.xlsx actualizado (col {mes_ano})")
        print(f"  Siguiente: python ../2_planilla/main.py")
    else:
        print(f"  ⚠ Corregir bloqueantes en outputs/correcciones_{mes_ano}.xlsx")
        print(f"  Volver a correr python main.py después de aplicar correcciones")
    print("═" * 65 + "\n")


if __name__ == "__main__":
    main()
