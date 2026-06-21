"""Genera el template del operario (registro_operario_mes.xlsx).

Lee registro_operario_acumulado.xlsx con doble header y pre-carga MZ, LT, NOMBRE
y MARC_ANT (última MARCACION del último ciclo). Agrega columna obs_operario para
que el operario pueda legitimar situaciones de campo con códigos M/F/P.

Pone una leyenda de códigos M/F/P en las primeras filas (encima de la tabla) para
que el operario la tenga visible.

Correr una sola vez al inicio del ciclo: python crear_template.py [YYYY-MM]
Si no se pasa el mes, usa la fecha del sistema.
"""
from datetime import date
from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
import config

# ── Estilo ────────────────────────────────────────────────────────────────────
C_HEAD_BG  = "1E3A5F"
C_LOCK_BG  = "EFF6FF"  # columnas pre-cargadas
C_INPUT_BG = "FFFDE7"  # columnas que el operario llena
C_OBS_BG   = "F3E8FF"  # obs_operario (opcional)
C_LEYENDA_BG = "FEF3C7"

COLS = [
    # (key, header, ancho, tipo_celda)
    ("MZ",           "MZ",            8,  "lock"),
    ("LT",           "LT",            6,  "lock"),
    ("NOMBRE",       "NOMBRE",        28, "lock"),
    ("MES_ANO",      "MES_ANO",       11, "lock"),
    ("MARC_ANT",     "MARC_ANT",      11, "lock"),
    ("MARC_ACT",     "MARC_ACT",      11, "input"),
    ("M3",           "M3",            8,  "input"),
    ("obs_operario", "obs_operario",  14, "obs"),
]


def _side():
    b = Side(style="thin", color="CCCCCC")
    return Border(left=b, right=b, top=b, bottom=b)


def _cargar_usuarios_base() -> list[dict]:
    """Lee MZ, LT, NOMBRE y la MARCACION del último ciclo del acumulado.

    Saltea filas con SIN_SERVICIO=Si (usuarios marcados como baja por el sync).

    Schema esperado del acumulado:
        Fila 1: MZ | LT | NOMBRE | SIN_SERVICIO | YYYY-MM (merge)        | ...
        Fila 2:                                  | MARCACION | M3        | ...

    Schema legacy (sin SIN_SERVICIO en col 4): se detecta y procesa también.
    """
    if not config.REGISTRO_ACUMULADO_PATH.exists():
        return []

    wb = load_workbook(config.REGISTRO_ACUMULADO_PATH, data_only=True)
    ws = wb.active

    # Detectar schema: col 4 = SIN_SERVICIO (nuevo) o YYYY-MM (legacy)
    cab_col4 = str(ws.cell(1, 4).value or "").strip().upper()
    tiene_sin_servicio = cab_col4 == "SIN_SERVICIO"
    col_inicio_meses = 5 if tiene_sin_servicio else 4

    n_cols = ws.max_column
    fila1 = [ws.cell(1, c).value for c in range(1, n_cols + 1)]
    fila2 = [ws.cell(2, c).value for c in range(1, n_cols + 1)]

    # Mapear meses al índice de su sub-columna MARCACION (a partir de col_inicio_meses)
    meses_marc = {}
    mes_actual = None
    for ci in range(col_inicio_meses, n_cols + 1):
        v1 = fila1[ci - 1]
        v2 = fila2[ci - 1]
        if v1 and isinstance(v1, str) and len(v1) == 7 and v1[4] == "-":
            mes_actual = v1
        if mes_actual and v2 == "MARCACION":
            meses_marc[mes_actual] = ci

    if not meses_marc:
        return []

    ultimo_mes = sorted(meses_marc.keys())[-1]
    col_marc = meses_marc[ultimo_mes]

    usuarios = []
    saltados_sin_servicio = 0
    for ri in range(3, ws.max_row + 1):
        mz   = str(ws.cell(ri, 1).value or "").strip().upper()
        lt   = str(ws.cell(ri, 2).value or "").strip()
        if not mz or not lt or mz in ("NONE",) or lt.lower() in ("none", "nan"):
            continue
        if tiene_sin_servicio:
            sin_serv = str(ws.cell(ri, 4).value or "").strip().lower()
            if sin_serv.startswith("si"):
                saltados_sin_servicio += 1
                continue
        nom  = str(ws.cell(ri, 3).value or "").strip()
        marc = ws.cell(ri, col_marc).value
        marc_str = str(marc).strip() if marc is not None else ""
        if marc_str.lower() in ("nan", "none"):
            marc_str = ""
        usuarios.append({"mz": mz, "lt": lt, "nombre": nom, "marc_ant": marc_str})

    if saltados_sin_servicio:
        print(f"  ({saltados_sin_servicio} usuarios con SIN_SERVICIO=Si omitidos del template)")
    return usuarios


def _escribir_leyenda(ws, fila_inicio: int = 1) -> int:
    """Escribe la leyenda M/F/P en las filas superiores y devuelve la fila siguiente libre."""
    title = ws.cell(row=fila_inicio, column=1, value="LEYENDA — códigos válidos en obs_operario")
    title.font = Font(name="Arial", bold=True, color="78350F", size=11)
    title.fill = PatternFill("solid", start_color=C_LEYENDA_BG)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=fila_inicio, start_column=1, end_row=fila_inicio, end_column=8)
    ws.row_dimensions[fila_inicio].height = 20

    leyenda = [
        ("M", "Medidor cambiado",         "MARC_ACT < MARC_ANT → no es retroceso, se acepta"),
        ("F", "Fuga visible reportada",   "Alerta informativa al supervisor"),
        ("P", "Predio cerrado / inaccesible", "MARC_ACT vacío → no es SIN_LECTURA, se acepta"),
        ("",  "(vacío)",                  "Lectura normal sin observación"),
    ]
    for offset, (codigo, signif, cuando) in enumerate(leyenda, 1):
        r = fila_inicio + offset
        ccod = ws.cell(row=r, column=1, value=codigo)
        ccod.font = Font(name="Arial", bold=True, size=11, color="78350F")
        ccod.fill = PatternFill("solid", start_color=C_LEYENDA_BG)
        ccod.alignment = Alignment(horizontal="center", vertical="center")

        csig = ws.cell(row=r, column=2, value=signif)
        csig.font = Font(name="Arial", bold=True, size=10, color="78350F")
        csig.fill = PatternFill("solid", start_color=C_LEYENDA_BG)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)

        ccuando = ws.cell(row=r, column=4, value=cuando)
        ccuando.font = Font(name="Arial", size=9, color="78350F", italic=True)
        ccuando.fill = PatternFill("solid", start_color=C_LEYENDA_BG)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)

        ws.row_dimensions[r].height = 16

    # Fila vacía como separador
    return fila_inicio + len(leyenda) + 2


def main():
    config.INPUTS_DIR.mkdir(parents=True, exist_ok=True)

    mes_ano = sys.argv[1] if len(sys.argv) > 1 else date.today().strftime("%Y-%m")
    usuarios = _cargar_usuarios_base()

    wb = Workbook()
    ws = wb.active
    ws.title = "Lecturas"

    borde = _side()

    # Leyenda M/F/P encima de la tabla — regla R2 del contrato registro_operario_mes
    fila_tabla = _escribir_leyenda(ws, fila_inicio=1)

    # Encabezados de la tabla
    for ci, (_, header, ancho, _) in enumerate(COLS, 1):
        c = ws.cell(row=fila_tabla, column=ci, value=header)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", start_color=C_HEAD_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = borde
        ws.column_dimensions[get_column_letter(ci)].width = ancho
    ws.row_dimensions[fila_tabla].height = 22

    if usuarios:
        for offset, u in enumerate(usuarios):
            ri = fila_tabla + 1 + offset
            vals = {
                "MZ":           u["mz"],
                "LT":           u["lt"],
                "NOMBRE":       u["nombre"],
                "MES_ANO":      mes_ano,
                "MARC_ANT":     u["marc_ant"],
                "MARC_ACT":     "",
                "M3":           "",
                "obs_operario": "",
            }
            for ci, (key, _, _, tipo) in enumerate(COLS, 1):
                value = vals[key]
                # Intentar numérico en MARC_ANT
                if key == "MARC_ANT" and value:
                    try:
                        value = float(value)
                    except (ValueError, TypeError):
                        pass
                c = ws.cell(row=ri, column=ci, value=value if value != "" else None)
                c.font      = Font(name="Arial", size=9)
                c.alignment = Alignment(horizontal="left" if key == "NOMBRE" else "center",
                                        vertical="center")
                c.border    = borde
                bg = {"lock": C_LOCK_BG, "input": C_INPUT_BG, "obs": C_OBS_BG}[tipo]
                c.fill = PatternFill("solid", start_color=bg)
            ws.row_dimensions[ri].height = 16
    else:
        nota = ws.cell(row=fila_tabla + 1, column=1,
                       value="Sin acumulado previo (primer ciclo) — llenar manualmente MZ/LT/NOMBRE/MES_ANO/MARC_ACT/M3")
        nota.font = Font(name="Arial", color="78350F", italic=True, size=9)
        ws.merge_cells(start_row=fila_tabla + 1, start_column=1,
                       end_row=fila_tabla + 1, end_column=8)

    ws.freeze_panes = ws.cell(row=fila_tabla + 1, column=1)

    wb.save(config.REGISTRO_MES_PATH)

    n = len(usuarios)
    print(f"Template creado: {config.REGISTRO_MES_PATH}")
    if n:
        print(f"   {n} usuarios pre-cargados desde el acumulado.")
    else:
        print("   Sin historial previo — llenar manualmente.")
    print("   Completar MARC_ACT, M3 y (opcional) obs_operario antes de correr main.py.")


if __name__ == "__main__":
    main()
