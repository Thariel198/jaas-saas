"""1_lecturas/proponer_sincronizacion.py — Paso 0a: diff padrón vs operario.

Compara dos sets en cada corrida y propone deltas:

  PADRÓN   = (MZ, LT, NOMBRE) en 0_padron/02_matching/outputs/padron_reconciliado.xlsx
  OPERARIO = (MZ, LT, NOMBRE, SIN_SERVICIO) en inputs/registro_operario_acumulado.xlsx

Clasifica cada delta en uno de 4 tipos:
  🟢 AGREGADO     — (MZ, LT) en padrón, NO en operario
  🟡 SIN_SERVICIO — (MZ, LT) en operario, NO en padrón (no se borra: se marca)
  🔵 RENAME       — mismo (MZ, LT), NOMBRE difiere normalizado
  🔴 MOVIMIENTO   — mismo NOMBRE, (MZ, LT) cambió (cruza AGREGADO+SIN_SERVICIO por nombre)

Solo MOVIMIENTO requiere autorización manual. Los demás se auto-marcan
REVISADO=Si (auto) · AUTORIZAR=Si (auto).

Si reporte_sincronizacion.xlsx existe del ciclo anterior, preserva REVISADO,
AUTORIZAR y APLICADO de filas pendientes (sticky entre ciclos).

Este script NO muta el acumulado. Solo escribe el reporte.
aplicar_sincronizacion.py es el único writer del acumulado.

Uso:
    python proponer_sincronizacion.py
"""
import logging
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
import config

SOURCE = "1_lecturas/proponer_sincronizacion"

# ── COLUMNAS del reporte ────────────────────────────────────────────────────
COLS_REPORTE = [
    "TIPO",
    "MZ_PADRON", "LT_PADRON", "NOMBRE_PADRON",
    "MZ_OPERARIO", "LT_OPERARIO", "NOMBRE_OPERARIO",
    "REVISADO", "AUTORIZAR", "FECHA_REVISION",
    "APLICADO", "FECHA_APLICADO",
]

# Paleta del reporte (igual que contrato_reporte_sincronizacion.html)
PAL = {
    "AGREGADO":     ("D5F5E3", "145A32"),
    "SIN_SERVICIO": ("FEF3C7", "7D5A00"),
    "RENAME":       ("D6EAF8", "1A5276"),
    "MOVIMIENTO":   ("FADBD8", "7B241C"),
}


# ── NORMALIZACIÓN ───────────────────────────────────────────────────────────
def _norm(s) -> str:
    """strip · UPPER · NFD sin marcas · espacios colapsados."""
    if s is None:
        return ""
    s = str(s).strip().upper()
    if not s or s in ("NAN", "NONE"):
        return ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


def _norm_lt(val) -> str:
    """LT acepta 8A o números: '7.0' -> '7', '8A' -> '8A'."""
    if val is None:
        return ""
    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN"):
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.upper().replace(" ", "")


# ── CARGADORES ──────────────────────────────────────────────────────────────
def _cargar_padron(log: logging.Logger) -> dict[tuple[str, str], str]:
    """{(mz_norm, lt_norm): nombre_norm} desde padron_reconciliado.xlsx hoja 'cobranza'."""
    if not config.PADRON_RECONCILIADO_PATH.exists():
        raise FileNotFoundError(
            f"No se encontró {config.PADRON_RECONCILIADO_PATH}\n"
            f"  -> Correr primero el módulo 0_padron"
        )

    wb = load_workbook(config.PADRON_RECONCILIADO_PATH, read_only=True, data_only=True)
    ws = wb["cobranza"] if "cobranza" in wb.sheetnames else wb.active

    # Header en fila 1: ..., Nombres, MZ, LT (índices 1, 2, 3 según comparar_padrones.py)
    mapa: dict[tuple[str, str], str] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if len(r) < 4:
            continue
        nombre, mz, lt = r[1], r[2], r[3]
        mz_n = _norm(mz)
        lt_n = _norm_lt(lt)
        if not mz_n or not lt_n:
            continue
        mapa[(mz_n, lt_n)] = _norm(nombre)
    wb.close()
    log.info(f"padron_reconciliado.xlsx · {len(mapa)} usuarios cargados")
    return mapa


def _cargar_acumulado(log: logging.Logger) -> dict[tuple[str, str], dict]:
    """{(mz_norm, lt_norm): {nombre, sin_servicio, mz_orig, lt_orig, nombre_orig}}.

    Tolerante a schema legacy sin SIN_SERVICIO (col 4 = primer YYYY-MM).
    """
    if not config.REGISTRO_ACUMULADO_PATH.exists():
        log.warning("registro_operario_acumulado.xlsx no existe — primer sync del sistema")
        return {}

    wb = load_workbook(config.REGISTRO_ACUMULADO_PATH, data_only=True)
    ws = wb.active

    # Detectar si col 4 es SIN_SERVICIO (nuevo) o YYYY-MM (legacy)
    cab_col4 = str(ws.cell(1, 4).value or "").strip().upper()
    tiene_sin_servicio = cab_col4 == "SIN_SERVICIO"
    if not tiene_sin_servicio:
        log.info("acumulado en schema legacy (sin columna SIN_SERVICIO) — se agregará en el siguiente apply")

    mapa: dict[tuple[str, str], dict] = {}
    for ri in range(3, ws.max_row + 1):
        mz_orig = ws.cell(ri, 1).value
        lt_orig = ws.cell(ri, 2).value
        nombre_orig = ws.cell(ri, 3).value
        sin_serv = ws.cell(ri, 4).value if tiene_sin_servicio else None

        mz_n = _norm(mz_orig)
        lt_n = _norm_lt(lt_orig)
        if not mz_n or not lt_n:
            continue
        mapa[(mz_n, lt_n)] = {
            "nombre": _norm(nombre_orig),
            "sin_servicio": str(sin_serv or "").strip().lower() == "si",
            "mz_orig": str(mz_orig or "").strip(),
            "lt_orig": str(lt_orig or "").strip(),
            "nombre_orig": str(nombre_orig or "").strip(),
        }
    wb.close()
    activos = sum(1 for v in mapa.values() if not v["sin_servicio"])
    inactivos = len(mapa) - activos
    log.info(f"registro_operario_acumulado.xlsx · {len(mapa)} usuarios · "
             f"{activos} activos · {inactivos} SIN_SERVICIO")
    return mapa


def _cargar_reporte_previo(log: logging.Logger) -> dict[tuple, dict]:
    """{(tipo, mz_p, lt_p, mz_o, lt_o): {REVISADO, AUTORIZAR, FECHA_REVISION,
                                         APLICADO, FECHA_APLICADO}} para preservar sticky.

    Si no existe, dict vacío.
    """
    if not config.REPORTE_SYNC_PATH.exists():
        return {}
    try:
        wb = load_workbook(config.REPORTE_SYNC_PATH, data_only=True)
    except Exception as e:
        log.warning(f"No se pudo leer reporte previo: {e} — partiendo de cero")
        return {}
    ws = wb.active
    # header en fila 2 (fila 1 = grupos)
    headers = {str(ws.cell(2, c).value or "").strip().upper(): c
               for c in range(1, ws.max_column + 1)}
    requeridas = {"TIPO", "MZ_PADRON", "LT_PADRON", "MZ_OPERARIO", "LT_OPERARIO",
                  "REVISADO", "AUTORIZAR", "APLICADO"}
    faltantes = requeridas - set(headers.keys())
    if faltantes:
        log.warning(f"Reporte previo sin columnas {faltantes} — partiendo de cero")
        return {}

    sticky: dict[tuple, dict] = {}
    for r in range(3, ws.max_row + 1):
        tipo = str(ws.cell(r, headers["TIPO"]).value or "").strip().upper()
        if not tipo:
            continue
        # Limpiar emojis del TIPO si quedaron
        for emoji in ("🟢", "🟡", "🔵", "🔴"):
            tipo = tipo.replace(emoji, "").strip()
        key = (
            tipo,
            _norm(ws.cell(r, headers["MZ_PADRON"]).value),
            _norm_lt(ws.cell(r, headers["LT_PADRON"]).value),
            _norm(ws.cell(r, headers["MZ_OPERARIO"]).value),
            _norm_lt(ws.cell(r, headers["LT_OPERARIO"]).value),
        )
        sticky[key] = {
            "REVISADO":       str(ws.cell(r, headers["REVISADO"]).value or "").strip(),
            "AUTORIZAR":      str(ws.cell(r, headers["AUTORIZAR"]).value or "").strip(),
            "FECHA_REVISION": str(ws.cell(r, headers.get("FECHA_REVISION", 0)).value or "").strip() if "FECHA_REVISION" in headers else "",
            "APLICADO":       str(ws.cell(r, headers["APLICADO"]).value or "").strip(),
            "FECHA_APLICADO": str(ws.cell(r, headers.get("FECHA_APLICADO", 0)).value or "").strip() if "FECHA_APLICADO" in headers else "",
        }
    wb.close()
    log.info(f"reporte_sincronizacion previo · {len(sticky)} filas (sticky)")
    return sticky


# ── DETECCIÓN DE DELTAS ─────────────────────────────────────────────────────
def _detectar_deltas(
    padron: dict[tuple[str, str], str],
    operario: dict[tuple[str, str], dict],
    log: logging.Logger,
) -> list[dict]:
    """Clasifica diferencias en AGREGADO, SIN_SERVICIO, RENAME, MOVIMIENTO."""
    keys_padron = set(padron.keys())
    keys_operario_activos = {k for k, v in operario.items() if not v["sin_servicio"]}
    keys_operario_inactivos = {k for k, v in operario.items() if v["sin_servicio"]}

    solo_padron   = keys_padron - keys_operario_activos - keys_operario_inactivos
    solo_operario = keys_operario_activos - keys_padron
    comunes       = keys_padron & keys_operario_activos

    # MOVIMIENTO: usuarios que aparecen en padrón y operario con (MZ, LT) distintos.
    # Lo detectamos cruzando NOMBRE_PADRON ↔ NOMBRE_OPERARIO entre solo_padron y solo_operario.
    nombres_solo_padron = {padron[k]: k for k in solo_padron if padron[k]}
    nombres_solo_operario = {operario[k]["nombre"]: k for k in solo_operario if operario[k]["nombre"]}
    nombres_movidos = set(nombres_solo_padron.keys()) & set(nombres_solo_operario.keys())

    deltas: list[dict] = []

    # 🔴 MOVIMIENTO — cancela el AGREGADO + SIN_SERVICIO correspondientes
    keys_padron_movidos = {nombres_solo_padron[n] for n in nombres_movidos}
    keys_operario_movidos = {nombres_solo_operario[n] for n in nombres_movidos}
    for nombre in nombres_movidos:
        k_pad = nombres_solo_padron[nombre]
        k_ope = nombres_solo_operario[nombre]
        deltas.append({
            "TIPO": "MOVIMIENTO",
            "MZ_PADRON":   k_pad[0], "LT_PADRON":   k_pad[1], "NOMBRE_PADRON":   padron[k_pad],
            "MZ_OPERARIO": k_ope[0], "LT_OPERARIO": k_ope[1], "NOMBRE_OPERARIO": operario[k_ope]["nombre"],
            "auto_autorizado": False,
        })

    # 🟢 AGREGADO — en padrón, no en operario, no es movimiento
    for k in sorted(solo_padron - keys_padron_movidos):
        deltas.append({
            "TIPO": "AGREGADO",
            "MZ_PADRON": k[0], "LT_PADRON": k[1], "NOMBRE_PADRON": padron[k],
            "MZ_OPERARIO": "", "LT_OPERARIO": "", "NOMBRE_OPERARIO": "",
            "auto_autorizado": True,
        })

    # 🟡 SIN_SERVICIO — en operario, no en padrón, no es movimiento, no estaba ya marcado
    for k in sorted(solo_operario - keys_operario_movidos):
        deltas.append({
            "TIPO": "SIN_SERVICIO",
            "MZ_PADRON": "", "LT_PADRON": "", "NOMBRE_PADRON": "",
            "MZ_OPERARIO": k[0], "LT_OPERARIO": k[1], "NOMBRE_OPERARIO": operario[k]["nombre"],
            "auto_autorizado": True,
        })

    # 🔵 RENAME — mismo (MZ, LT), nombres distintos
    for k in sorted(comunes):
        n_pad = padron[k]
        n_ope = operario[k]["nombre"]
        if n_pad and n_ope and n_pad != n_ope:
            deltas.append({
                "TIPO": "RENAME",
                "MZ_PADRON": k[0], "LT_PADRON": k[1], "NOMBRE_PADRON": n_pad,
                "MZ_OPERARIO": k[0], "LT_OPERARIO": k[1], "NOMBRE_OPERARIO": n_ope,
                "auto_autorizado": True,
            })

    log.info(
        f"Deltas detectados · AGREGADO={sum(1 for d in deltas if d['TIPO']=='AGREGADO')} "
        f"· SIN_SERVICIO={sum(1 for d in deltas if d['TIPO']=='SIN_SERVICIO')} "
        f"· RENAME={sum(1 for d in deltas if d['TIPO']=='RENAME')} "
        f"· MOVIMIENTO={sum(1 for d in deltas if d['TIPO']=='MOVIMIENTO')}"
    )
    return deltas


def _fusionar_con_sticky(
    deltas: list[dict],
    sticky: dict[tuple, dict],
    log: logging.Logger,
) -> list[dict]:
    """Aplica REVISADO/AUTORIZAR/APLICADO del reporte previo a los deltas regenerados."""
    preservados = 0
    for d in deltas:
        key = (
            d["TIPO"],
            _norm(d["MZ_PADRON"]),   _norm_lt(d["LT_PADRON"]),
            _norm(d["MZ_OPERARIO"]), _norm_lt(d["LT_OPERARIO"]),
        )
        previo = sticky.get(key)
        if previo and (previo["REVISADO"] or previo["AUTORIZAR"] or previo["APLICADO"]):
            d["REVISADO"]       = previo["REVISADO"]
            d["AUTORIZAR"]      = previo["AUTORIZAR"]
            d["FECHA_REVISION"] = previo["FECHA_REVISION"]
            d["APLICADO"]       = previo["APLICADO"]
            d["FECHA_APLICADO"] = previo["FECHA_APLICADO"]
            preservados += 1
        else:
            # Defaults según auto_autorizado
            if d["auto_autorizado"]:
                d["REVISADO"]       = "Si (auto)"
                d["AUTORIZAR"]      = "Si (auto)"
                d["FECHA_REVISION"] = ""
            else:
                d["REVISADO"]       = ""
                d["AUTORIZAR"]      = ""
                d["FECHA_REVISION"] = ""
            d["APLICADO"]       = ""
            d["FECHA_APLICADO"] = ""
    if preservados:
        log.info(f"Sticky preservado · {preservados} filas del reporte anterior")
    return deltas


# ── ESCRITURA DEL REPORTE ───────────────────────────────────────────────────
def _borde():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, r, c, val, bg=None, txt="333333", bold=False, align="center", mono=False, size=9):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(name="Consolas" if mono else "Arial", size=size, bold=bold, color=txt)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    cell.border = _borde()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    return cell


def _escribir_reporte(deltas: list[dict], log: logging.Logger) -> None:
    config.SYNC_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "reporte"

    # Grupos en fila 1 (matches contrato_reporte_sincronizacion.html)
    GRUPOS = [
        (1, 1, "Tipo",          "F4ECF7", "5B21B6"),
        (2, 4, "Padrón",        "EBF5FB", "1A5276"),
        (5, 7, "Operario",      "FEF3E8", "7C3003"),
        (8, 10, "Decisión",     "D5F5E3", "145A32"),
        (11, 12, "Estado",      "FEF9C3", "7D6608"),
    ]
    for cs, ce, texto, bg, txt in GRUPOS:
        ws.merge_cells(start_row=1, start_column=cs, end_row=1, end_column=ce)
        c = ws.cell(row=1, column=cs, value=texto)
        c.font = Font(name="Arial", size=9, bold=True, color=txt)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", start_color=bg)
        c.border = _borde()

    # Headers en fila 2
    HEADERS = [
        ("TIPO",            "F4ECF7", "5B21B6", 16),
        ("MZ_PADRON",       "EBF5FB", "1A5276",  8),
        ("LT_PADRON",       "EBF5FB", "1A5276",  8),
        ("NOMBRE_PADRON",   "EBF5FB", "1A5276", 32),
        ("MZ_OPERARIO",     "FEF3E8", "7C3003",  8),
        ("LT_OPERARIO",     "FEF3E8", "7C3003",  8),
        ("NOMBRE_OPERARIO", "FEF3E8", "7C3003", 32),
        ("REVISADO",        "D5F5E3", "145A32", 12),
        ("AUTORIZAR",       "D5F5E3", "145A32", 12),
        ("FECHA_REVISION",  "D5F5E3", "145A32", 14),
        ("APLICADO",        "FEF9C3", "7D6608", 10),
        ("FECHA_APLICADO",  "FEF9C3", "7D6608", 14),
    ]
    for ci, (nombre, bg, txt, ancho) in enumerate(HEADERS, 1):
        _cell(ws, 2, ci, nombre, bg=bg, txt=txt, bold=True, align="center")
        ws.column_dimensions[get_column_letter(ci)].width = ancho

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    # Orden de filas: MOVIMIENTO primero (lo que necesita atención), luego el resto
    orden_tipo = {"MOVIMIENTO": 0, "AGREGADO": 1, "SIN_SERVICIO": 2, "RENAME": 3}
    deltas_sorted = sorted(
        deltas,
        key=lambda d: (orden_tipo.get(d["TIPO"], 99),
                       d.get("MZ_PADRON", "") or d.get("MZ_OPERARIO", ""),
                       d.get("LT_PADRON", "") or d.get("LT_OPERARIO", "")),
    )

    EMOJI = {"AGREGADO": "🟢", "SIN_SERVICIO": "🟡", "RENAME": "🔵", "MOVIMIENTO": "🔴"}
    for ri, d in enumerate(deltas_sorted, 3):
        tipo = d["TIPO"]
        bg_tipo, txt_tipo = PAL[tipo]
        _cell(ws, ri, 1, f"{EMOJI[tipo]} {tipo}", bg=bg_tipo, txt=txt_tipo, bold=True)

        # Padrón (col 2-4)
        _cell(ws, ri, 2, d["MZ_PADRON"] or "—",     bg="F4FAFF", txt="1A5276", mono=True)
        _cell(ws, ri, 3, d["LT_PADRON"] or "—",     bg="F4FAFF", txt="1A5276", mono=True)
        _cell(ws, ri, 4, d["NOMBRE_PADRON"] or "—", bg="F4FAFF", txt="1A5276", align="left")

        # Operario (col 5-7)
        _cell(ws, ri, 5, d["MZ_OPERARIO"] or "—",     bg="FEF9EC", txt="7C3003", mono=True)
        _cell(ws, ri, 6, d["LT_OPERARIO"] or "—",     bg="FEF9EC", txt="7C3003", mono=True)
        _cell(ws, ri, 7, d["NOMBRE_OPERARIO"] or "—", bg="FEF9EC", txt="7C3003", align="left")

        # Decisión (col 8-10) — colores condicionales para vacíos
        bg_dec = "F0FFF8"; txt_dec = "065F46"
        if tipo == "MOVIMIENTO" and not d["AUTORIZAR"]:
            bg_dec = "FADBD8"; txt_dec = "7B241C"
        _cell(ws, ri, 8,  d["REVISADO"]  or "❌ vacío", bg=bg_dec, txt=txt_dec, bold=True)
        _cell(ws, ri, 9,  d["AUTORIZAR"] or "❌ vacío", bg=bg_dec, txt=txt_dec, bold=True)
        _cell(ws, ri, 10, d["FECHA_REVISION"] or "—",   bg="F0FFF8", txt="065F46", mono=True)

        # Estado (col 11-12)
        bg_est = "FEF9EC"; txt_est = "7D5A00"
        if d["APLICADO"].lower().startswith("si"):
            bg_est = "D5F5E3"; txt_est = "145A32"
        _cell(ws, ri, 11, d["APLICADO"]       or "No", bg=bg_est, txt=txt_est, bold=True)
        _cell(ws, ri, 12, d["FECHA_APLICADO"] or "—",  bg="FEF9EC", txt="7D5A00", mono=True)

        ws.row_dimensions[ri].height = 17

    wb.save(config.REPORTE_SYNC_PATH)
    log.info(f"reporte_sincronizacion.xlsx · {len(deltas)} filas escritas")


# ── MAIN ────────────────────────────────────────────────────────────────────
def main() -> None:
    # Windows cmd usa cp1252 por defecto y revienta con emojis. Forzar UTF-8.
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    config.OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    config.SYNC_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(config.LOG_PATH, encoding="utf-8"),
        ],
        force=True,
    )
    log = logging.getLogger(__name__)
    log.info("proponer_sincronizacion.py · iniciando")

    print("=" * 60)
    print("  1_lecturas/proponer_sincronizacion.py")
    print("=" * 60)

    print("\n[1/4] Cargando padrón reconciliado...")
    padron = _cargar_padron(log)

    print("\n[2/4] Cargando registro_operario_acumulado...")
    operario = _cargar_acumulado(log)

    print("\n[3/4] Cargando reporte previo (sticky)...")
    sticky = _cargar_reporte_previo(log)

    print("\n[4/4] Detectando deltas y escribiendo reporte...")
    deltas = _detectar_deltas(padron, operario, log)
    deltas = _fusionar_con_sticky(deltas, sticky, log)
    _escribir_reporte(deltas, log)

    # Resumen por tipo
    conteo = {t: 0 for t in ("AGREGADO", "SIN_SERVICIO", "RENAME", "MOVIMIENTO")}
    pendientes_autorizar = 0
    for d in deltas:
        conteo[d["TIPO"]] = conteo.get(d["TIPO"], 0) + 1
        if d["TIPO"] == "MOVIMIENTO" and not d["AUTORIZAR"]:
            pendientes_autorizar += 1

    print("\n" + "=" * 60)
    print(f"  Resumen ({sum(conteo.values())} deltas)")
    print(f"    🟢 AGREGADO     {conteo['AGREGADO']:>4}   (auto-autorizado)")
    print(f"    🟡 SIN_SERVICIO {conteo['SIN_SERVICIO']:>4}   (auto-autorizado)")
    print(f"    🔵 RENAME       {conteo['RENAME']:>4}   (auto-autorizado)")
    print(f"    🔴 MOVIMIENTO   {conteo['MOVIMIENTO']:>4}   ({pendientes_autorizar} pendientes de autorizar)")
    print(f"\n  -> {config.REPORTE_SYNC_PATH}")
    if pendientes_autorizar:
        print(f"\n  ⚠  Abrir el reporte, marcar AUTORIZAR=Si/No en MOVIMIENTOs y guardar.")
        print(f"     Luego correr: python aplicar_sincronizacion.py")
    else:
        print(f"\n  Sin MOVIMIENTOs pendientes — ejecutar:")
        print(f"     python aplicar_sincronizacion.py")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
