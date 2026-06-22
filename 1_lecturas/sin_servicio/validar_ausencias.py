"""sin_servicio/validar_ausencias.py — core del submódulo.

Importable desde 1_lecturas/main.py. Recibe la lista de casos que main.py ya identificó
(usuarios sin MARC_ACT o fantasmas en el historial), los enriquece con datos del
acumulado, los cruza con lista_sin_servicio.xlsx y escribe el reporte para el supervisor.

Escribe:
    outputs/validacion_ausencias_YYYY-MM.xlsx

Retorna a main.py:
    (justificadas, revisar)
        justificadas — DECISIÓN ya llena (pre-llenada o supervisor previo) → no bloquea
        revisar      — DECISIÓN vacía → siguen siendo bloqueantes del ciclo
"""
from __future__ import annotations

import logging
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

# Asegurar que 1_lecturas/ y la raíz del proyecto están en sys.path para que
# `from sin_servicio.config import ...` y `from shared.*` resuelvan al correr
# tanto desde main.py como en standalone (tests, CLI).
sys.path.insert(0, str(Path(__file__).parent.parent))
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from sin_servicio.config import (
    COLS_LISTA,
    COLS_VALIDACION,
    DECISIONES_VALIDAS,
    LISTA_PATH,
    MESES_NUNCA_TUVO,
    OUTPUTS_DIR,
    PREFIJO_AUTO,
    PREFIJO_DEFAULT_SIN_AGUA,
    SHEET_LISTA,
    TIPOS_VALIDOS,
    UMBRAL_AUTO_PROMOCION,
    sheet_ausencias,
    validacion_ausencias_path,
)
from formato_excel import escribir_con_grupos

try:
    from shared.utils_sort_mz_lt import clave_orden
except ImportError:
    def clave_orden(mz, lt):
        return (str(mz), str(lt))


# ── Grupos narrativos (espejan formato_validacion_ausencias.html) ─────────────
GRUPOS_VALIDACION_AUSENCIAS = [
    {"nombre": "¿Quién es?", "tipo": "quien",
     "cols": ["MZ", "LT", "NOMBRE"]},
    {"nombre": "Evidencia — sistema llena", "tipo": "operario",
     "cols": ["ESCENARIO", "ÚLTIMO_MES_CON_LECTURA", "MARCACION_ÚLTIMA",
              "MESES_SIN_LECTURA", "ESTADO_LISTA"]},
    {"nombre": "Decisión — supervisor llena", "tipo": "fix",
     "cols": ["DECISIÓN", "NOTAS"]},
    {"nombre": "Trazabilidad", "tipo": "cuando",
     "cols": ["CICLO", "FECHA_DETECCIÓN"]},
]

log = logging.getLogger(__name__)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _fmt_marc(v) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _detectar_header_row(rows: list, primera_col_esperada: str) -> int:
    """Devuelve el índice (0-based) de la fila HEADER.

    v1 plano (legacy): fila 0 es header → data desde fila 1.
    v2 con grupos:     fila 0 es group label, fila 1 es header → data desde fila 2.

    Devuelve -1 si no encuentra el header.
    """
    if not rows:
        return -1
    fila0 = [_cell_str(c) for c in rows[0]]
    if fila0 and fila0[0] == primera_col_esperada:
        return 0
    if len(rows) >= 2:
        fila1 = [_cell_str(c) for c in rows[1]]
        if fila1 and fila1[0] == primera_col_esperada:
            return 1
    return -1


# ── Cargar lista_sin_servicio ─────────────────────────────────────────────────

def cargar_lista() -> dict[tuple[str, str], dict]:
    """Devuelve {(mz, lt): {nombre, tipo, fecha_inicio, ultima_revision, meses, notas}}.

    Falla ruidoso si:
        - el header no coincide con COLS_LISTA
        - alguna fila tiene TIPO fuera de TIPOS_VALIDOS

    Devuelve {} si el archivo no existe (avisa por log).
    """
    if not LISTA_PATH.exists():
        log.warning(
            f"{LISTA_PATH.name} no existe — lista vacía. "
            f"¿Se ejecutó migrar_lista.py?"
        )
        return {}

    wb = load_workbook(LISTA_PATH, read_only=True, data_only=True)
    if SHEET_LISTA not in wb.sheetnames:
        raise ValueError(
            f"{LISTA_PATH.name} no contiene la hoja '{SHEET_LISTA}'. "
            f"Hojas: {wb.sheetnames}"
        )
    ws = wb[SHEET_LISTA]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header_row = _detectar_header_row(rows, "MZ")
    if header_row < 0:
        raise ValueError(
            f"{LISTA_PATH.name}: no se encontró fila de header (esperado 'MZ' en col 0)."
        )
    header = [_cell_str(c) for c in rows[header_row]]
    if header != list(COLS_LISTA):
        raise ValueError(
            f"Header de {LISTA_PATH.name} no coincide con COLS_LISTA.\n"
            f"  esperado:   {list(COLS_LISTA)}\n"
            f"  encontrado: {header}"
        )

    out: dict[tuple[str, str], dict] = {}
    data_start = header_row + 1
    for fila_idx, r in enumerate(rows[data_start:], start=data_start + 1):
        if all(c is None for c in r):
            continue
        mz = _cell_str(r[0])
        lt = _cell_str(r[1])
        if not (mz and lt):
            continue
        tipo = _cell_str(r[3])
        if tipo not in TIPOS_VALIDOS:
            raise ValueError(
                f"{LISTA_PATH.name} fila {fila_idx}: "
                f"TIPO={tipo!r} no está en {TIPOS_VALIDOS}"
            )
        meses_raw = r[6]
        try:
            meses = int(meses_raw) if meses_raw is not None else MESES_NUNCA_TUVO
        except (TypeError, ValueError):
            meses = MESES_NUNCA_TUVO
        out[(mz, lt)] = {
            "nombre":          _cell_str(r[2]),
            "tipo":            tipo,
            "fecha_inicio":    _cell_str(r[4]),
            "ultima_revision": _cell_str(r[5]),
            "meses":           meses,
            "notas":           _cell_str(r[7]),
        }
    log.info(f"lista_sin_servicio: {len(out)} usuarios cargados")
    return out


# ── Helper público para main.py ───────────────────────────────────────────────

def justifica_ausencia(key: tuple[str, str], lista: dict) -> bool:
    """True si (mz, lt) está catalogado como SIN_MEDIDOR — caso cerrado, no bloquea.

    Los demás TIPOs (SIN_AGUA, EN_INVESTIGACIÓN) NO justifican: el supervisor debe
    decidir cada ciclo a través del reporte validacion_ausencias_YYYY-MM.xlsx.
    """
    entry = lista.get(key)
    return entry is not None and entry["tipo"] == "SIN_MEDIDOR"


# ── Cálculos sobre el historial ───────────────────────────────────────────────

def _meses_sin_lectura(
    key: tuple[str, str],
    historial: dict,
    meses_orden: list[str],
    mes_ano: str,
) -> int:
    """Cuenta ciclos consecutivos sin MARC_ACT desde el mes actual hacia atrás.

    Convención: incluye el mes actual (que es por definición sin lectura, ya que
    estamos validando una ausencia). Si el usuario nunca tuvo lectura en el
    acumulado devuelve MESES_NUNCA_TUVO (99).
    """
    count = 1  # mes actual cuenta
    entry = historial.get(key)
    if not entry:
        return MESES_NUNCA_TUVO
    ciclos = entry.get("ciclos", {})
    # Iterar meses cerrados de más reciente al más antiguo, excluyendo mes_ano.
    for mes in reversed(meses_orden):
        if mes == mes_ano:
            continue
        ciclo_data = ciclos.get(mes)
        if ciclo_data and ciclo_data.get("marc") is not None:
            return count
        count += 1
    return MESES_NUNCA_TUVO


def _ultimo_mes_con_lectura(
    key: tuple[str, str],
    historial: dict,
    meses_orden: list[str],
    mes_ano: str,
) -> tuple[str, str]:
    """Devuelve (mes_yyyy_mm, marc_str) del último ciclo CERRADO con lectura.

    ('', '') si el usuario no está en historial o nunca tuvo lectura.
    """
    entry = historial.get(key)
    if not entry:
        return "", ""
    ciclos = entry.get("ciclos", {})
    for mes in reversed(meses_orden):
        if mes == mes_ano:
            continue
        ciclo_data = ciclos.get(mes)
        if ciclo_data and ciclo_data.get("marc") is not None:
            return mes, _fmt_marc(ciclo_data["marc"])
    return "", ""


# ── Clasificación y pre-llenado ───────────────────────────────────────────────

def _clasificar(
    entry_lista: dict | None,
    meses_sin_lectura: int,
    fecha_hoy: str,
) -> dict:
    """Decide ESTADO_LISTA + DECISIÓN pre-llenada + prefijo de NOTAS.

    Reglas (espejan formato_validacion_ausencias.html y el README):
        NO_ESTÁ          → vacío (bloqueante)
        SIN_AGUA         → pre-llena 'sin_servicio' con prefijo default
        EN_INVEST < 3    → vacío (bloqueante)
        EN_INVEST ≥ 3    → pre-llena 'sin_servicio' con prefijo auto (auto-promo)
        SIN_MEDIDOR      → no debería llegar acá (main.py lo filtra antes)
    """
    if entry_lista is None:
        return {"estado_lista": "NO_ESTÁ", "decision_default": "", "notas_default": ""}
    tipo = entry_lista["tipo"]
    if tipo == "SIN_MEDIDOR":
        # Defensa en profundidad — main.py debería haberlo filtrado como informativa.
        return {"estado_lista": "NO_ESTÁ", "decision_default": "", "notas_default": ""}
    if tipo == "SIN_AGUA":
        return {
            "estado_lista":     "SIN_AGUA",
            "decision_default": "sin_servicio",
            "notas_default":    f"{PREFIJO_DEFAULT_SIN_AGUA} · {fecha_hoy}",
        }
    if tipo == "EN_INVESTIGACIÓN":
        if meses_sin_lectura >= UMBRAL_AUTO_PROMOCION:
            return {
                "estado_lista":     "EN_INVESTIGACIÓN",
                "decision_default": "sin_servicio",
                "notas_default":    f"{PREFIJO_AUTO} · {fecha_hoy}",
            }
        return {
            "estado_lista":     "EN_INVESTIGACIÓN",
            "decision_default": "",
            "notas_default":    "",
        }
    raise ValueError(f"TIPO inesperado en lista: {tipo!r}")


# ── Preservación entre re-ejecuciones ─────────────────────────────────────────

def _leer_previo(mes_ano: str) -> dict[tuple[str, str], dict]:
    """Lee el xlsx previo si existe.

    Devuelve {(mz, lt): {decision, notas}} solo de filas con DECISIÓN llena —
    son las que se preservan entre re-ejecuciones.
    """
    path = validacion_ausencias_path(mes_ano)
    if not path.exists():
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = sheet_ausencias(mes_ano)
    if sheet not in wb.sheetnames:
        return {}
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {}

    header_row = _detectar_header_row(rows, "MZ")
    if header_row < 0:
        log.warning(f"Header no encontrado en {path.name} — no se preservan decisiones")
        return {}
    header = [_cell_str(c) for c in rows[header_row]]
    try:
        i_mz  = header.index("MZ")
        i_lt  = header.index("LT")
        i_dec = header.index("DECISIÓN")
        i_not = header.index("NOTAS")
    except ValueError:
        log.warning(f"Header inesperado en {path.name} — no se preservan decisiones")
        return {}

    out: dict[tuple[str, str], dict] = {}
    for r in rows[header_row + 1:]:
        mz  = _cell_str(r[i_mz])
        lt  = _cell_str(r[i_lt])
        dec = _cell_str(r[i_dec])
        if not (mz and lt and dec):
            continue
        if dec not in DECISIONES_VALIDAS:
            log.warning(
                f"DECISIÓN inválida en {path.name} para {mz}-{lt}: {dec!r} — se ignora"
            )
            continue
        out[(mz, lt)] = {"decision": dec, "notas": _cell_str(r[i_not])}
    return out


# ── Escritura del xlsx ────────────────────────────────────────────────────────

def _escribir_xlsx(filas: list[dict], mes_ano: str) -> Path:
    """Escribe validacion_ausencias_YYYY-MM.xlsx con grupos narrativos. Sobreescribe si existe."""
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    path = validacion_ausencias_path(mes_ano)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_ausencias(mes_ano)
    escribir_con_grupos(ws, GRUPOS_VALIDACION_AUSENCIAS, filas)
    wb.save(path)
    return path


# ── Función pública ──────────────────────────────────────────────────────────

def validar_ausencias(
    mes_ano: str,
    ciclo: int,
    casos: list[dict],
    lista: dict,
    historial: dict,
    meses_orden: list[str],
) -> tuple[list[dict], list[dict]]:
    """Procesa los casos pendientes y escribe el reporte.

    Args:
        mes_ano: 'YYYY-MM' del ciclo en curso.
        ciclo: número de ciclo dentro del mes (1, 2, ...).
        casos: lista de {mz, lt, nombre, escenario} — los SIN_LECTURA y USUARIO_FANTASMA
               que NO están justificados como SIN_MEDIDOR (main.py los filtró antes).
        lista: resultado de cargar_lista().
        historial: {(mz,lt): {nombre, ciclos: {mes: {marc, m3}}}} de main._cargar_acumulado.
        meses_orden: lista de meses presentes en el acumulado, ordenada ascendente.

    Returns:
        (justificadas, revisar) — cada uno es lista de {mz, lt, nombre, escenario, decision}.
            justificadas: filas con DECISIÓN llena (pre-llenada o supervisor previo).
            revisar:      filas con DECISIÓN vacía (siguen bloqueantes).

    Side effect:
        Escribe outputs/validacion_ausencias_YYYY-MM.xlsx.
    """
    previo = _leer_previo(mes_ano)
    fecha_hoy = datetime.now().strftime("%Y-%m-%d")
    fecha_deteccion = datetime.now().strftime("%Y-%m-%d %H:%M")

    filas: list[dict] = []
    justificadas: list[dict] = []
    revisar: list[dict] = []

    # Orden determinístico para que re-runs no muevan filas.
    casos_ord = sorted(
        casos,
        key=lambda c: clave_orden(c["mz"], c["lt"]),
    )

    for caso in casos_ord:
        mz, lt = caso["mz"], caso["lt"]
        key = (mz, lt)
        nombre = caso.get("nombre") or (
            lista.get(key, {}).get("nombre")
            or historial.get(key, {}).get("nombre", "")
        )
        escenario = caso["escenario"]
        entry_lista = lista.get(key)

        meses_sin = _meses_sin_lectura(key, historial, meses_orden, mes_ano)
        ultimo_mes, marc_ultima = _ultimo_mes_con_lectura(
            key, historial, meses_orden, mes_ano
        )

        clas = _clasificar(entry_lista, meses_sin, fecha_hoy)
        estado_lista = clas["estado_lista"]
        decision_default = clas["decision_default"]
        notas_default = clas["notas_default"]

        # Preservar trabajo del supervisor entre re-ejecuciones.
        if key in previo:
            decision_final = previo[key]["decision"]
            notas_final = previo[key]["notas"]
        else:
            decision_final = decision_default
            notas_final = notas_default

        fila = {
            "MZ": mz,
            "LT": lt,
            "NOMBRE": nombre,
            "ESCENARIO": escenario,
            "ÚLTIMO_MES_CON_LECTURA": ultimo_mes,
            "MARCACION_ÚLTIMA": marc_ultima,
            "MESES_SIN_LECTURA": meses_sin,
            "ESTADO_LISTA": estado_lista,
            "DECISIÓN": decision_final,
            "NOTAS": notas_final,
            "CICLO": ciclo,
            "FECHA_DETECCIÓN": fecha_deteccion,
        }
        filas.append(fila)

        resumen = {
            "mz": mz, "lt": lt, "nombre": nombre,
            "escenario": escenario, "decision": decision_final,
            "estado_lista": estado_lista,
        }
        if decision_final:
            justificadas.append(resumen)
        else:
            revisar.append(resumen)

    path = _escribir_xlsx(filas, mes_ano)
    log.info(
        f"validar_ausencias: {len(filas)} filas → {path.name} · "
        f"{len(justificadas)} con DECISIÓN llena · {len(revisar)} pendientes"
    )
    return justificadas, revisar
