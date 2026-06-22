"""Tests para sin_servicio/validar_ausencias.py.

Cubre:
    cargar_lista        — lectura básica, archivo ausente, TIPO inválido
    justifica_ausencia  — SIN_MEDIDOR justifica, otros no
    _meses_sin_lectura  — nunca tuvo, lectura mes pasado, lectura más atrás
    _clasificar         — 4 ramas (NO_ESTÁ, SIN_AGUA, EN_INVEST <3, ≥3)
    validar_ausencias   — end-to-end: genera xlsx, preserva DECISIÓN entre re-runs
"""
import pytest
from openpyxl import Workbook, load_workbook

from sin_servicio import config, validar_ausencias as va
from sin_servicio.validar_ausencias import (
    _clasificar,
    _meses_sin_lectura,
    _ultimo_mes_con_lectura,
    cargar_lista,
    justifica_ausencia,
    validar_ausencias,
)


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture
def tmp_paths(tmp_path, monkeypatch):
    """Aisla LISTA_PATH y OUTPUTS_DIR en tmp_path."""
    inputs = tmp_path / "inputs"
    outputs = tmp_path / "outputs"
    inputs.mkdir()
    outputs.mkdir()
    lista_path = inputs / "lista_sin_servicio.xlsx"

    monkeypatch.setattr(va, "LISTA_PATH", lista_path)
    monkeypatch.setattr(va, "OUTPUTS_DIR", outputs)
    monkeypatch.setattr(
        va, "validacion_ausencias_path",
        lambda mes: outputs / f"validacion_ausencias_{mes}.xlsx",
    )
    return {"inputs": inputs, "outputs": outputs, "lista": lista_path}


def _escribir_lista(path, filas):
    """Helper para escribir una lista_sin_servicio.xlsx con filas dadas."""
    wb = Workbook()
    ws = wb.active
    ws.title = config.SHEET_LISTA
    ws.append(list(config.COLS_LISTA))
    for f in filas:
        ws.append(f)
    wb.save(path)


@pytest.fixture
def lista_sample(tmp_paths):
    _escribir_lista(tmp_paths["lista"], [
        ["A", "1", "JUAN",  "SIN_MEDIDOR",      "2026-01-01", "2026-01-01", 99, ""],
        ["B", "2", "MARIA", "SIN_AGUA",         "2026-03-01", "2026-03-01",  2, ""],
        ["C", "3", "PEDRO", "EN_INVESTIGACIÓN", "2026-04-01", "2026-04-01",  1, ""],
    ])
    return tmp_paths


@pytest.fixture
def historial_sample():
    return {
        ("A", "1"): {"nombre": "JUAN",  "ciclos": {}},
        ("B", "2"): {"nombre": "MARIA", "ciclos": {
            "2026-05": {"marc": 100.0, "m3": 5.0},
        }},
        ("C", "3"): {"nombre": "PEDRO", "ciclos": {
            "2026-04": {"marc": 50.0, "m3": 3.0},
        }},
        ("D", "4"): {"nombre": "LUCIA", "ciclos": {
            "2026-05": {"marc": 200.0, "m3": 8.0},
            "2026-06": {"marc": 210.0, "m3": 10.0},
        }},
    }


# ── cargar_lista ──────────────────────────────────────────────────────────────

class TestCargarLista:
    def test_lectura_basica(self, lista_sample):
        lista = cargar_lista()
        assert len(lista) == 3
        assert lista[("A", "1")]["tipo"] == "SIN_MEDIDOR"
        assert lista[("B", "2")]["tipo"] == "SIN_AGUA"
        assert lista[("C", "3")]["tipo"] == "EN_INVESTIGACIÓN"
        assert lista[("A", "1")]["meses"] == 99

    def test_archivo_ausente_devuelve_vacio(self, tmp_paths):
        assert cargar_lista() == {}

    def test_tipo_invalido_falla_ruidoso(self, tmp_paths):
        _escribir_lista(tmp_paths["lista"], [
            ["A", "1", "X", "INVENTADO", "2026-01-01", "2026-01-01", 0, ""],
        ])
        with pytest.raises(ValueError, match="no está en"):
            cargar_lista()

    def test_header_incorrecto_falla(self, tmp_paths):
        wb = Workbook()
        ws = wb.active
        ws.title = config.SHEET_LISTA
        ws.append(["MZ", "LT", "FALTA_COLUMNA"])  # header malo
        wb.save(tmp_paths["lista"])
        with pytest.raises(ValueError, match="no coincide con COLS_LISTA"):
            cargar_lista()


# ── justifica_ausencia ───────────────────────────────────────────────────────

class TestJustificaAusencia:
    def test_sin_medidor_justifica(self):
        lista = {("A", "1"): {"tipo": "SIN_MEDIDOR", "nombre": "X"}}
        assert justifica_ausencia(("A", "1"), lista) is True

    def test_sin_agua_no_justifica(self):
        lista = {("B", "2"): {"tipo": "SIN_AGUA", "nombre": "Y"}}
        assert justifica_ausencia(("B", "2"), lista) is False

    def test_en_investigacion_no_justifica(self):
        lista = {("C", "3"): {"tipo": "EN_INVESTIGACIÓN", "nombre": "Z"}}
        assert justifica_ausencia(("C", "3"), lista) is False

    def test_no_en_lista_no_justifica(self):
        assert justifica_ausencia(("X", "99"), {}) is False


# ── _meses_sin_lectura ───────────────────────────────────────────────────────

class TestMesesSinLectura:
    def test_nunca_tuvo_lectura(self, historial_sample):
        result = _meses_sin_lectura(
            ("A", "1"), historial_sample, ["2026-05", "2026-06"], "2026-07"
        )
        assert result == config.MESES_NUNCA_TUVO

    def test_no_aparece_en_historial(self, historial_sample):
        result = _meses_sin_lectura(
            ("Z", "99"), historial_sample, ["2026-05", "2026-06"], "2026-07"
        )
        assert result == config.MESES_NUNCA_TUVO

    def test_lectura_mes_anterior_cuenta_1(self, historial_sample):
        # ('D', '4') tuvo lectura en 2026-06 (último ciclo cerrado).
        # Mes actual sin lectura cuenta 1; al iterar encuentra marc en 2026-06 → break.
        result = _meses_sin_lectura(
            ("D", "4"), historial_sample, ["2026-05", "2026-06"], "2026-07"
        )
        assert result == 1

    def test_lectura_mas_atras_acumula(self, historial_sample):
        # ('C', '3') tuvo lectura en 2026-04. meses_orden cubre 2026-04, -05, -06.
        # mes actual (1) + 2026-06 sin marc (2) + 2026-05 sin marc (3) + 2026-04 con marc → 3.
        result = _meses_sin_lectura(
            ("C", "3"), historial_sample,
            ["2026-04", "2026-05", "2026-06"], "2026-07",
        )
        assert result == 3


# ── _ultimo_mes_con_lectura ──────────────────────────────────────────────────

class TestUltimoMesConLectura:
    def test_nunca_tuvo(self, historial_sample):
        mes, marc = _ultimo_mes_con_lectura(
            ("A", "1"), historial_sample, ["2026-05", "2026-06"], "2026-07"
        )
        assert mes == "" and marc == ""

    def test_devuelve_ultimo_cerrado(self, historial_sample):
        mes, marc = _ultimo_mes_con_lectura(
            ("D", "4"), historial_sample, ["2026-05", "2026-06"], "2026-07"
        )
        assert mes == "2026-06"
        assert marc == "210"


# ── _clasificar ──────────────────────────────────────────────────────────────

class TestClasificar:
    def test_no_esta_no_pre_llena(self):
        c = _clasificar(None, 5, "2026-06-21")
        assert c["estado_lista"] == "NO_ESTÁ"
        assert c["decision_default"] == ""
        assert c["notas_default"] == ""

    def test_sin_agua_pre_llena_sin_servicio(self):
        c = _clasificar({"tipo": "SIN_AGUA"}, 1, "2026-06-21")
        assert c["estado_lista"] == "SIN_AGUA"
        assert c["decision_default"] == "sin_servicio"
        assert c["notas_default"].startswith(config.PREFIJO_DEFAULT_SIN_AGUA)

    def test_en_invest_bajo_umbral_no_pre_llena(self):
        c = _clasificar({"tipo": "EN_INVESTIGACIÓN"}, 2, "2026-06-21")
        assert c["estado_lista"] == "EN_INVESTIGACIÓN"
        assert c["decision_default"] == ""

    def test_en_invest_alcanza_umbral_auto_promociona(self):
        c = _clasificar(
            {"tipo": "EN_INVESTIGACIÓN"},
            config.UMBRAL_AUTO_PROMOCION,
            "2026-06-21",
        )
        assert c["estado_lista"] == "EN_INVESTIGACIÓN"
        assert c["decision_default"] == "sin_servicio"
        assert c["notas_default"].startswith(config.PREFIJO_AUTO)

    def test_sin_medidor_defensa_devuelve_vacio(self):
        # No debería llegar acá (main.py filtra), pero por defensa devuelve no-op.
        c = _clasificar({"tipo": "SIN_MEDIDOR"}, 1, "2026-06-21")
        assert c["decision_default"] == ""


# ── validar_ausencias end-to-end ─────────────────────────────────────────────

class TestValidarAusenciasIntegration:
    def test_genera_xlsx_y_clasifica_justificadas_vs_revisar(
        self, lista_sample, historial_sample, tmp_paths
    ):
        lista = cargar_lista()
        casos = [
            {"mz": "X", "lt": "1", "nombre": "FUERA",  "escenario": "SIN_LECTURA"},
            {"mz": "B", "lt": "2", "nombre": "MARIA",  "escenario": "SIN_LECTURA"},
            {"mz": "C", "lt": "3", "nombre": "PEDRO",  "escenario": "SIN_LECTURA"},
        ]
        justif, revisar = validar_ausencias(
            "2026-07", 1, casos, lista, historial_sample,
            ["2026-04", "2026-05", "2026-06"],
        )
        # FUERA = NO_ESTÁ → revisar
        # MARIA = SIN_AGUA → pre-llena → justif
        # PEDRO = EN_INVEST con 3 meses sin lectura (≥ umbral) → auto-promo → justif
        assert len(revisar) == 1
        assert revisar[0]["mz"] == "X"
        decisiones = {(j["mz"], j["lt"]): j["decision"] for j in justif}
        assert decisiones == {("B", "2"): "sin_servicio", ("C", "3"): "sin_servicio"}
        path = tmp_paths["outputs"] / "validacion_ausencias_2026-07.xlsx"
        assert path.exists()

    def test_preserva_decision_supervisor_en_re_run(
        self, lista_sample, historial_sample, tmp_paths
    ):
        lista = cargar_lista()
        casos = [{"mz": "X", "lt": "1", "nombre": "FUERA", "escenario": "SIN_LECTURA"}]
        # Run 1 — sin decisión, queda en revisar.
        validar_ausencias(
            "2026-07", 1, casos, lista, historial_sample, ["2026-06"],
        )
        # Supervisor llena DECISIÓN.
        path = tmp_paths["outputs"] / "validacion_ausencias_2026-07.xlsx"
        wb = load_workbook(path)
        ws = wb.active
        header = [c.value for c in ws[1]]
        i_dec = header.index("DECISIÓN") + 1
        i_not = header.index("NOTAS") + 1
        ws.cell(row=2, column=i_dec).value = "investigar"
        ws.cell(row=2, column=i_not).value = "nota del supervisor"
        wb.save(path)
        # Run 2 — debe preservar DECISIÓN.
        justif, revisar = validar_ausencias(
            "2026-07", 1, casos, lista, historial_sample, ["2026-06"],
        )
        assert len(revisar) == 0
        assert len(justif) == 1
        assert justif[0]["decision"] == "investigar"
