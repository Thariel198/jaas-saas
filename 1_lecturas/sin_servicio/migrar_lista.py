"""sin_servicio/migrar_lista.py — one-shot: migra el archivo plano legacy al schema nuevo.

Lee:
    1_lecturas/inputs/Lista de usuarios sin servicio.xlsx   (NOMBRES · MZ · LT)

Escribe:
    sin_servicio/inputs/lista_sin_servicio.xlsx              (schema COLS_LISTA)

Archiva tras éxito:
    Mueve el legacy a 1_lecturas/inputs/backups/Lista_de_usuarios_sin_servicio_LEGACY_pre_migracion_YYYYMMDD_HHMMSS.xlsx
    para que no quede ruido en inputs/ activos.

Todas las filas migran con:
    TIPO              = SIN_MEDIDOR
    FECHA_INICIO      = ÚLTIMA_REVISIÓN = hoy
    MESES_SIN_LECTURA = 99 (MESES_NUNCA_TUVO)
    NOTAS             = ""

Casos según estado de los archivos:
    legacy + (no new)        → migra + archiva legacy
    legacy + new + --force   → backup del new + migra + archiva legacy
    legacy + new (sin force) → abort (avisa de la inconsistencia)
    (no legacy) + new        → "ya migrada", exit 0 (idempotente)
    (no legacy) + (no new)   → error: no hay fuente
    --archive-only           → solo archiva el legacy si está presente

Uso:
    python sin_servicio/migrar_lista.py                  # migración normal
    python sin_servicio/migrar_lista.py --force          # re-migrar pisando new
    python sin_servicio/migrar_lista.py --archive-only   # solo archivar legacy
"""
from __future__ import annotations

import argparse
import logging
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

# El submódulo se ejecuta como script — agrego 1_lecturas/ a sys.path para
# poder hacer `from sin_servicio.config import ...` (import explícito que evita
# colisión con el config.py del padre).
sys.path.insert(0, str(Path(__file__).parent.parent))
from sin_servicio.config import (
    BACKUPS_DIR,
    COLS_LISTA,
    INPUTS_DIR,
    LISTA_LEGACY_PATH,
    LISTA_PATH,
    MESES_NUNCA_TUVO,
    SHEET_LISTA,
)

log = logging.getLogger(__name__)


def _cell_str(v) -> str:
    """Excel a veces devuelve floats (6.0) donde el operario tipeó '6'. Normalizar."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _leer_lista_plana(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"{path.name} está vacío")

    header = [_cell_str(c) for c in rows[0]]
    esperado = {"NOMBRES", "MZ", "LT"}
    if not esperado.issubset(set(header)):
        raise ValueError(
            f"Header inesperado en {path.name}: {header} — se esperaba {esperado}"
        )
    i_nombre = header.index("NOMBRES")
    i_mz     = header.index("MZ")
    i_lt     = header.index("LT")

    filas = []
    for fila_idx, r in enumerate(rows[1:], start=2):
        if all(c is None for c in r):
            continue
        nombre = _cell_str(r[i_nombre])
        mz     = _cell_str(r[i_mz])
        lt     = _cell_str(r[i_lt])
        if not (nombre or mz or lt):
            continue
        if not (mz and lt):
            log.warning(f"  fila {fila_idx} descartada (MZ/LT vacío): nombre={nombre!r}")
            continue
        filas.append({"NOMBRE": nombre, "MZ": mz, "LT": lt})
    return filas


def _dedupe_por_mz_lt(filas: list[dict]) -> list[dict]:
    """Si hay (MZ, LT) repetido en el origen, conservar la primera ocurrencia y warn."""
    visto: dict[tuple, str] = {}
    out: list[dict] = []
    for f in filas:
        k = (f["MZ"], f["LT"])
        if k in visto:
            log.warning(
                f"  duplicado descartado {k[0]}-{k[1]}: {f['NOMBRE']!r} "
                f"(ya estaba {visto[k]!r})"
            )
            continue
        visto[k] = f["NOMBRE"]
        out.append(f)
    return out


def _backup_existente(dst: Path) -> Path:
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = BACKUPS_DIR / f"{dst.stem}_{ts}{dst.suffix}"
    shutil.copy2(dst, backup)
    return backup


def _archivar_legacy(src: Path) -> Path:
    """Mueve el legacy a backups/ del módulo padre con sufijo de fecha."""
    legacy_backups = src.parent / "backups"
    legacy_backups.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    # nombre seguro (sin espacios) para el archivo
    stem_seguro = src.stem.replace(" ", "_")
    dst = legacy_backups / f"{stem_seguro}_LEGACY_pre_migracion_{ts}{src.suffix}"
    shutil.move(str(src), str(dst))
    return dst


def _escribir_lista_nueva(filas: list[dict], dst: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_LISTA
    ws.append(list(COLS_LISTA))

    hoy = datetime.now().strftime("%Y-%m-%d")
    for f in filas:
        ws.append([
            f["MZ"], f["LT"], f["NOMBRE"],
            "SIN_MEDIDOR", hoy, hoy,
            MESES_NUNCA_TUVO,
            "",
        ])

    INPUTS_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(dst)


def main(argv: list[str] | None = None) -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    p = argparse.ArgumentParser(description="Migración one-shot lista_sin_servicio")
    p.add_argument(
        "--force", action="store_true",
        help="Sobreescribir destino existente (crea backup automático)",
    )
    p.add_argument(
        "--archive-only", action="store_true",
        help="No migra; solo archiva el legacy si está presente (útil si la "
             "migración ya se hizo manualmente y el legacy sigue en inputs/).",
    )
    args = p.parse_args(argv)

    src = LISTA_LEGACY_PATH
    dst = LISTA_PATH

    # ── --archive-only: solo mover el legacy si está ─────────────────────
    if args.archive_only:
        if not src.exists():
            log.info(f"Legacy no presente en {src.parent}/ — nada que archivar.")
            return 0
        if not dst.exists():
            log.error(
                f"Se solicitó --archive-only pero el destino {dst.name} no existe.\n"
                f"  Negarse a archivar el legacy sin un destino confirmado."
            )
            return 1
        archivado = _archivar_legacy(src)
        log.info(f"Legacy archivado → {archivado.relative_to(archivado.parent.parent.parent)}")
        return 0

    # ── Casos según presencia de archivos ────────────────────────────────
    if not src.exists() and not dst.exists():
        log.error(
            f"No hay fuente para migrar:\n"
            f"  Legacy no existe: {src}\n"
            f"  Destino no existe: {dst}\n"
            f"  Restaurar legacy desde backups/ si necesitas re-migrar."
        )
        return 1

    if not src.exists() and dst.exists():
        log.info(f"Ya migrada: {dst.name} existe y el legacy fue archivado previamente.")
        return 0

    if src.exists() and dst.exists() and not args.force:
        log.error(
            f"Estado inconsistente:\n"
            f"  Destino existe: {dst.name}\n"
            f"  Legacy aún en inputs/: {src.name}\n"
            f"  Opciones:\n"
            f"    --archive-only  → solo archiva el legacy (recomendado si la lista ya está actualizada)\n"
            f"    --force         → re-migra desde el legacy pisando el destino actual"
        )
        return 1

    # ── Migración (caso A o B con --force) ──────────────────────────────
    log.info(f"Leyendo {src.name}")
    filas = _leer_lista_plana(src)
    log.info(f"  {len(filas)} filas válidas leídas")

    filas = _dedupe_por_mz_lt(filas)
    log.info(f"  {len(filas)} filas tras deduplicar por (MZ, LT)")

    if dst.exists():
        backup = _backup_existente(dst)
        log.info(f"Backup del destino: backups/{backup.name}")

    _escribir_lista_nueva(filas, dst)
    hoy = datetime.now().strftime("%Y-%m-%d")
    log.info(f"Escrito {dst.name}: {len(filas)} filas")
    log.info(f"  TIPO=SIN_MEDIDOR · FECHA_INICIO=ÚLTIMA_REVISIÓN={hoy} · MESES_SIN_LECTURA={MESES_NUNCA_TUVO}")

    archivado = _archivar_legacy(src)
    log.info(f"Legacy archivado → inputs/backups/{archivado.name}")
    log.info("OK — migración completa")
    return 0


if __name__ == "__main__":
    sys.exit(main())
