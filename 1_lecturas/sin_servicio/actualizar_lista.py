"""sin_servicio/actualizar_lista.py — writer único de lista_sin_servicio.xlsx.

Lee:
    outputs/validacion_ausencias_YYYY-MM.xlsx   (con DECISIÓN llena en todas las filas)
    inputs/lista_sin_servicio.xlsx              (estado actual de la lista)

Escribe:
    inputs/lista_sin_servicio.xlsx              (actualizada, con backup automático)
    outputs/audit_lista.xlsx                    (append-only, registro de cada cambio)

Aplica una acción por cada fila del reporte:
    DECISIÓN=sin_servicio + entry vacío    → INSERT  TIPO=SIN_MEDIDOR
    DECISIÓN=sin_servicio + entry SIN_AGUA → UPDATE  SIN_AGUA → SIN_MEDIDOR
    DECISIÓN=sin_servicio + entry EN_INVEST→ UPDATE  EN_INVEST → SIN_MEDIDOR
    DECISIÓN=sin_servicio + entry SIN_MED. → SKIP_DUPLICATE (idempotente)
    DECISIÓN=investigar + entry vacío      → INSERT  TIPO=EN_INVESTIGACIÓN
    DECISIÓN=investigar + entry presente   → SKIP_DUPLICATE (idempotente)
    DECISIÓN=error_captura                 → NO_OP (solo audit)
    DECISIÓN=ignorar                       → NO_OP (solo audit)
    DECISIÓN=(vacío)                       → ABORT (refuse, exit 1)

Uso:
    python sin_servicio/actualizar_lista.py --mes 2026-07
    python sin_servicio/actualizar_lista.py --mes 2026-07 --dry-run

Idempotente: re-correr no duplica filas. Errores deciden por exit code:
    0 = OK (incluye SKIP_DUPLICATE)
    1 = DECISIÓN faltante o archivo no encontrado
    2 = error de schema
"""
from __future__ import annotations

import argparse
import logging
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent))
from sin_servicio.config import (
    AUDIT_LISTA_PATH,
    BACKUPS_DIR,
    COLS_AUDIT,
    COLS_LISTA,
    COLS_VALIDACION,
    DECISIONES_VALIDAS,
    LISTA_PATH,
    OUTPUTS_DIR,
    PREFIJO_AUTO,
    PREFIJO_DEFAULT_SIN_AGUA,
    SHEET_AUDIT,
    SHEET_LISTA,
    TIPOS_VALIDOS,
    sheet_ausencias,
    validacion_ausencias_path,
)
from sin_servicio.validar_ausencias import _detectar_header_row, cargar_lista
from formato_excel import escribir_con_grupos

log = logging.getLogger(__name__)


# ── Grupos narrativos (espejan formato_lista_sin_servicio.html / formato_audit_lista.html) ──
GRUPOS_LISTA = [
    {"nombre": "¿Quién es?", "tipo": "quien",
     "cols": ["MZ", "LT", "NOMBRE"]},
    {"nombre": "Estado en el catálogo — sistema gestiona", "tipo": "operario",
     "cols": ["TIPO", "FECHA_INICIO", "ÚLTIMA_REVISIÓN"]},
    {"nombre": "Historial", "tipo": "cuando",
     "cols": ["MESES_SIN_LECTURA"]},
    {"nombre": "Notas", "tipo": "fix",
     "cols": ["NOTAS"]},
]

GRUPOS_AUDIT = [
    {"nombre": "¿Quién es?", "tipo": "quien",
     "cols": ["MZ", "LT", "NOMBRE"]},
    {"nombre": "¿Qué se hizo?", "tipo": "fix",
     "cols": ["ACCION", "TIPO_NUEVO"]},
    {"nombre": "Origen de la decisión", "tipo": "operario",
     "cols": ["DECISION_ORIGEN", "ORIGEN_PROPUESTA", "NOTAS"]},
    {"nombre": "¿Cuándo?", "tipo": "cuando",
     "cols": ["FECHA", "CICLO_ORIGEN"]},
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _detectar_origen(notas: str) -> str:
    """Infiera ORIGEN_PROPUESTA del prefijo de NOTAS."""
    if notas.startswith(PREFIJO_AUTO):
        return "auto"
    if notas.startswith(PREFIJO_DEFAULT_SIN_AGUA):
        return "default_sin_agua"
    return "supervisor"


def _listar_meses_disponibles() -> list[str]:
    """Para mostrar opciones en el error si --mes falta o no existe."""
    if not OUTPUTS_DIR.exists():
        return []
    meses = []
    for p in OUTPUTS_DIR.glob("validacion_ausencias_*.xlsx"):
        stem = p.stem.replace("validacion_ausencias_", "")
        meses.append(stem)
    return sorted(meses)


# ── Lectura del reporte ──────────────────────────────────────────────────────

def _leer_validacion(mes_ano: str) -> list[dict]:
    """Lee validacion_ausencias_YYYY-MM.xlsx y devuelve lista de dicts."""
    path = validacion_ausencias_path(mes_ano)
    if not path.exists():
        raise FileNotFoundError(f"No existe {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = sheet_ausencias(mes_ano)
    if sheet not in wb.sheetnames:
        raise ValueError(f"{path.name} no contiene la hoja '{sheet}'. Hojas: {wb.sheetnames}")
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    header_row = _detectar_header_row(rows, "MZ")
    if header_row < 0:
        raise ValueError(f"{path.name}: no se encontró fila de header (esperado 'MZ' en col 0).")
    header = [_cell_str(c) for c in rows[header_row]]
    if header != list(COLS_VALIDACION):
        raise ValueError(
            f"Header de {path.name} no coincide con COLS_VALIDACION.\n"
            f"  esperado:   {list(COLS_VALIDACION)}\n"
            f"  encontrado: {header}"
        )

    filas = []
    for r in rows[header_row + 1:]:
        if all(c is None for c in r):
            continue
        d = {col: _cell_str(r[i]) for i, col in enumerate(COLS_VALIDACION)}
        if not (d["MZ"] and d["LT"]):
            continue
        filas.append(d)
    return filas


def _validar_decisiones(filas: list[dict]) -> list[dict]:
    """Devuelve filas con DECISIÓN vacía o inválida (para abortar si hay)."""
    invalidas = []
    for f in filas:
        dec = f["DECISIÓN"]
        if not dec:
            invalidas.append({**f, "_motivo": "vacía"})
            continue
        if dec not in DECISIONES_VALIDAS:
            invalidas.append({**f, "_motivo": f"valor inválido {dec!r}"})
    return invalidas


# ── Planificación de cambios ─────────────────────────────────────────────────

def _planificar(filas: list[dict], lista: dict) -> list[dict]:
    """Para cada fila del reporte decide la acción a aplicar.

    Devuelve lista de planes con keys:
        accion         — INSERT | UPDATE | SKIP_DUPLICATE | NO_OP_ERROR_CAPTURA | NO_OP_IGNORAR
        tipo_nuevo     — TIPO resultante (solo INSERT/UPDATE)
        origen         — auto | default_sin_agua | supervisor
        mz, lt, nombre — identificación
        decision, notas, ciclo — datos originales del reporte
    """
    plan = []
    for f in filas:
        mz, lt = f["MZ"], f["LT"]
        key = (mz, lt)
        decision = f["DECISIÓN"]
        notas = f["NOTAS"]
        origen = _detectar_origen(notas)
        entry = lista.get(key)

        base = {
            "mz": mz, "lt": lt, "nombre": f["NOMBRE"],
            "decision": decision, "notas": notas,
            "origen": origen, "ciclo": f["CICLO"],
        }

        if decision == "sin_servicio":
            if entry is None:
                plan.append({**base, "accion": "INSERT", "tipo_nuevo": "SIN_MEDIDOR"})
            elif entry["tipo"] == "SIN_MEDIDOR":
                plan.append({**base, "accion": "SKIP_DUPLICATE", "tipo_nuevo": "SIN_MEDIDOR"})
            else:
                # SIN_AGUA o EN_INVESTIGACIÓN → promociona a SIN_MEDIDOR
                plan.append({**base, "accion": "UPDATE", "tipo_nuevo": "SIN_MEDIDOR"})
        elif decision == "investigar":
            if entry is None:
                plan.append({**base, "accion": "INSERT", "tipo_nuevo": "EN_INVESTIGACIÓN"})
            else:
                # Ya está en lista — no se re-clasifica para abajo desde acá.
                plan.append({**base, "accion": "SKIP_DUPLICATE", "tipo_nuevo": entry["tipo"]})
        elif decision == "error_captura":
            plan.append({**base, "accion": "NO_OP_ERROR_CAPTURA", "tipo_nuevo": ""})
        elif decision == "ignorar":
            plan.append({**base, "accion": "NO_OP_IGNORAR", "tipo_nuevo": ""})
        else:
            # _validar_decisiones ya filtró, no debería llegar acá.
            raise ValueError(f"Decisión inesperada {decision!r} para {mz}-{lt}")
    return plan


# ── Aplicación a la lista ────────────────────────────────────────────────────

def _aplicar_a_lista(plan: list[dict], lista: dict, hoy: str) -> dict:
    """Devuelve una nueva lista con los INSERT/UPDATE aplicados (no muta `lista`)."""
    nueva = {k: dict(v) for k, v in lista.items()}
    for p in plan:
        key = (p["mz"], p["lt"])
        if p["accion"] == "INSERT":
            nueva[key] = {
                "nombre":          p["nombre"],
                "tipo":            p["tipo_nuevo"],
                "fecha_inicio":    hoy,
                "ultima_revision": hoy,
                "meses":           99 if p["tipo_nuevo"] == "SIN_MEDIDOR" else 1,
                "notas":           p["notas"],
            }
        elif p["accion"] == "UPDATE":
            entry = nueva[key]
            entry["tipo"] = p["tipo_nuevo"]
            entry["ultima_revision"] = hoy
            # NOTAS: concatena la nota nueva sin sobreescribir el historial
            if p["notas"] and p["notas"] != entry.get("notas", ""):
                if entry.get("notas"):
                    entry["notas"] = f"{entry['notas']} | {p['notas']}"
                else:
                    entry["notas"] = p["notas"]
    return nueva


# ── Persistencia ──────────────────────────────────────────────────────────────

def _backup_lista() -> Path | None:
    if not LISTA_PATH.exists():
        return None
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    bk = BACKUPS_DIR / f"{LISTA_PATH.stem}_{ts}{LISTA_PATH.suffix}"
    shutil.copy2(LISTA_PATH, bk)
    return bk


def _escribir_lista(lista: dict) -> None:
    LISTA_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_LISTA
    # Orden determinístico por (MZ, LT) — usa fallback simple si no hay clave_orden.
    try:
        from shared.utils_sort_mz_lt import clave_orden
    except ImportError:
        def clave_orden(mz, lt): return (str(mz), str(lt))
    filas = []
    for key in sorted(lista.keys(), key=lambda k: clave_orden(k[0], k[1])):
        entry = lista[key]
        filas.append({
            "MZ": key[0], "LT": key[1], "NOMBRE": entry["nombre"],
            "TIPO": entry["tipo"],
            "FECHA_INICIO": entry["fecha_inicio"],
            "ÚLTIMA_REVISIÓN": entry["ultima_revision"],
            "MESES_SIN_LECTURA": entry["meses"],
            "NOTAS": entry["notas"],
        })
    escribir_con_grupos(ws, GRUPOS_LISTA, filas)
    wb.save(LISTA_PATH)


def _es_lista_v1() -> bool:
    """True si lista_sin_servicio.xlsx existe pero está en formato plano (header en fila 1).

    Se usa para forzar el re-write a v2 aunque no haya cambios en lista — así la
    migración se aplica de manera self-healing al primer toque después de actualizar el código.
    """
    if not LISTA_PATH.exists():
        return False
    wb = load_workbook(LISTA_PATH, read_only=True, data_only=True)
    if SHEET_LISTA not in wb.sheetnames:
        return False
    ws = wb[SHEET_LISTA]
    rows = list(ws.iter_rows(values_only=True, max_row=2))
    if not rows:
        return False
    fila0 = [_cell_str(c) for c in rows[0]]
    # v1: header (MZ) directamente en fila 0
    return fila0 and fila0[0] == "MZ"


def _audit_necesita_migracion() -> bool:
    """True si audit existe pero su header no coincide con el COLS_AUDIT canónico.

    Cubre tres casos de drift:
    - v1 plano (header en fila 0)
    - v2 con orden viejo (header existe pero las columnas están en otro orden)
    - cualquier cambio futuro de COLS_AUDIT
    """
    if not AUDIT_LISTA_PATH.exists():
        return False
    wb = load_workbook(AUDIT_LISTA_PATH, read_only=True, data_only=True)
    if SHEET_AUDIT not in wb.sheetnames:
        return False
    ws = wb[SHEET_AUDIT]
    rows = list(ws.iter_rows(values_only=True, max_row=2))
    if not rows:
        return False
    # Buscar header (puede estar en fila 0 = v1 o fila 1 = v2). Reconocible por contener "MZ" o "FECHA".
    for r in rows:
        header = [_cell_str(c) for c in r if _cell_str(c)]
        if header and (header[0] == "MZ" or header[0] == "FECHA"):
            return tuple(header) != tuple(COLS_AUDIT)
    return False


def _migrar_audit() -> None:
    """Re-escribe audit_lista.xlsx con GRUPOS_AUDIT actual sin agregar filas."""
    existentes = _leer_audit_existente()
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_AUDIT
    escribir_con_grupos(ws, GRUPOS_AUDIT, existentes)
    wb.save(AUDIT_LISTA_PATH)
    log.info(f"audit_lista.xlsx migrado al orden canónico ({len(existentes)} filas preservadas)")


def _leer_audit_existente() -> list[dict]:
    """Lee audit existente (tolera v1 plano y v2 con grupos). Devuelve list[dict]."""
    if not AUDIT_LISTA_PATH.exists():
        return []
    wb = load_workbook(AUDIT_LISTA_PATH, read_only=True, data_only=True)
    if SHEET_AUDIT not in wb.sheetnames:
        return []
    ws = wb[SHEET_AUDIT]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row = _detectar_header_row(rows, "FECHA")
    if header_row < 0:
        log.warning(f"{AUDIT_LISTA_PATH.name}: header no detectado — se ignoran filas previas")
        return []
    header = [_cell_str(c) for c in rows[header_row]]
    # Tolerancia: el audit v1 tenía COLS_AUDIT en orden viejo; mapeamos por nombre.
    out = []
    for r in rows[header_row + 1:]:
        if all(c is None for c in r):
            continue
        d = {}
        for i, col in enumerate(header):
            if i < len(r):
                d[col] = _cell_str(r[i])
        # Solo conservar filas válidas con FECHA
        if d.get("FECHA"):
            out.append(d)
    return out


def _append_audit(plan: list[dict], mes_ano: str) -> None:
    """Append-only semántico: lee filas existentes, agrega plan, reescribe con grupos.

    Migra automáticamente v1 plano → v2 con grupos. La idempotencia y el carácter
    append-only se preservan: nunca borramos filas previas.
    """
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    existentes = _leer_audit_existente()
    nuevas = [
        {
            "FECHA":            fecha,
            "CICLO_ORIGEN":     f"{mes_ano} · ciclo {p['ciclo']}",
            "MZ":               p["mz"],
            "LT":               p["lt"],
            "NOMBRE":           p["nombre"],
            "ACCION":           p["accion"],
            "TIPO_NUEVO":       p["tipo_nuevo"],
            "DECISION_ORIGEN":  p["decision"],
            "ORIGEN_PROPUESTA": p["origen"],
            "NOTAS":            p["notas"],
        }
        for p in plan
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_AUDIT
    escribir_con_grupos(ws, GRUPOS_AUDIT, existentes + nuevas)
    wb.save(AUDIT_LISTA_PATH)


# ── Entrypoint ────────────────────────────────────────────────────────────────

def _resumen(plan: list[dict]) -> dict[str, int]:
    out: dict[str, int] = {}
    for p in plan:
        out[p["accion"]] = out.get(p["accion"], 0) + 1
    return out


def main(argv: list[str] | None = None) -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    parser = argparse.ArgumentParser(
        description="Aplica decisiones de validacion_ausencias_YYYY-MM.xlsx a la lista."
    )
    parser.add_argument(
        "--mes", required=True,
        help="Mes/año del reporte a aplicar (formato YYYY-MM). "
             "Ej: --mes 2026-07",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Muestra el plan sin escribir nada (ni lista ni audit).",
    )
    args = parser.parse_args(argv)

    # Auto-migración de audit_lista.xlsx si su orden de columnas drifteó del canónico.
    # Self-healing: ocurre en cualquier corrida, sin paso manual.
    if _audit_necesita_migracion():
        _migrar_audit()

    path_reporte = validacion_ausencias_path(args.mes)
    if not path_reporte.exists():
        disponibles = _listar_meses_disponibles()
        msg = f"No existe {path_reporte.name}"
        if disponibles:
            msg += f"\n  Meses disponibles: {', '.join(disponibles)}"
        else:
            msg += f"\n  No hay reportes en {OUTPUTS_DIR}"
        log.error(msg)
        return 1

    log.info(f"Leyendo {path_reporte.name}")
    try:
        filas = _leer_validacion(args.mes)
    except ValueError as e:
        log.error(f"Schema inválido: {e}")
        return 2

    if not filas:
        log.info("El reporte no tiene filas — nada que aplicar.")
        return 0

    log.info(f"  {len(filas)} filas leídas")

    invalidas = _validar_decisiones(filas)
    if invalidas:
        log.error(
            f"{len(invalidas)} filas con DECISIÓN no llena o inválida — abortando.\n"
            f"  Llenar DECISIÓN en {path_reporte.name} y volver a correr."
        )
        for f in invalidas[:10]:
            log.error(f"    {f['MZ']}-{f['LT']} {f['NOMBRE']!r}: {f['_motivo']}")
        if len(invalidas) > 10:
            log.error(f"    ... y {len(invalidas) - 10} más")
        return 1

    log.info("Cargando lista actual")
    lista = cargar_lista()
    log.info(f"  {len(lista)} usuarios en lista")

    plan = _planificar(filas, lista)
    resumen = _resumen(plan)
    log.info("Plan de cambios:")
    for accion, n in sorted(resumen.items()):
        log.info(f"  {accion:<22} {n}")

    if args.dry_run:
        log.info("DRY-RUN — no se escribió nada. Quitar --dry-run para aplicar.")
        return 0

    # Solo hacer backup + escribir lista si hay cambios reales sobre la lista,
    # O si está en formato v1 (self-healing migration al primer toque).
    cambios_lista = resumen.get("INSERT", 0) + resumen.get("UPDATE", 0)
    migrar_v1 = _es_lista_v1()
    if cambios_lista > 0:
        hoy = datetime.now().strftime("%Y-%m-%d")
        nueva = _aplicar_a_lista(plan, lista, hoy)
        bk = _backup_lista()
        if bk:
            log.info(f"Backup: backups/{bk.name}")
        _escribir_lista(nueva)
        log.info(f"lista_sin_servicio.xlsx: {len(lista)} → {len(nueva)} usuarios")
    elif migrar_v1:
        bk = _backup_lista()
        if bk:
            log.info(f"Backup pre-migración v1→v2: backups/{bk.name}")
        _escribir_lista(lista)
        log.info(f"lista_sin_servicio.xlsx migrada a formato v2 ({len(lista)} usuarios)")
    else:
        log.info("Sin cambios sobre la lista (solo NO_OP / SKIP_DUPLICATE)")

    _append_audit(plan, args.mes)
    log.info(f"audit_lista.xlsx: +{len(plan)} entradas")
    log.info("OK")
    return 0


if __name__ == "__main__":
    sys.exit(main())
