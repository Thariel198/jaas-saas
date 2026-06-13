"""Helpers de formato Excel — aplica los grupos narrativos de los contratos HTML.

Cada output del módulo se diseñó como tabla con grupos por color:
    ¿Quién es?            — lila    (identificación del usuario)
    ¿Qué entregó operario?— azul    (datos crudos del campo, inmutables)
    ¿Cuál es la anomalía? — rojo    (problema detectado)
    ¿Cómo se resuelve?    — verde   (maquillaje o valor final)
    ¿Estado / Cuándo?     — naranja (metadata del workflow)

Este módulo replica esos grupos en los Excel reales:
- Fila 1: nombre del grupo con merge horizontal sobre sus columnas
- Fila 2: nombres de las columnas individuales
- Filas 3+: datos con fondo coloreado según su grupo
"""
from __future__ import annotations

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta por grupo ──────────────────────────────────────────────────────────
# Espeja los HTMLs de docs/contrato_*.html
PALETA = {
    "quien":    {"head_bg": "5B21B6", "head_fg": "FFFFFF", "sub_bg": "F4ECF7", "data_bg": "FAF5FF", "fg": "5B21B6"},
    "operario": {"head_bg": "0C447C", "head_fg": "FFFFFF", "sub_bg": "E6F1FB", "data_bg": "F0F8FF", "fg": "185FA5"},
    "anomalia": {"head_bg": "A32D2D", "head_fg": "FFFFFF", "sub_bg": "FCEBEB", "data_bg": "FFF5F5", "fg": "A32D2D"},
    "fix":      {"head_bg": "065F46", "head_fg": "FFFFFF", "sub_bg": "E1F5EE", "data_bg": "F0FFF8", "fg": "065F46"},
    "cuando":   {"head_bg": "7C3003", "head_fg": "FFFFFF", "sub_bg": "FEF3E8", "data_bg": "FEF8F0", "fg": "7C3003"},
}


def _borde() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


# ── Definición de grupos por output (espeja los contratos HTML) ───────────────
GRUPOS_CORRECCIONES = [
    {"nombre": "¿Quién es?", "tipo": "quien",
     "cols": ["MZ", "LT", "NOMBRE"]},
    {"nombre": "¿Qué entregó el operario? — no editar", "tipo": "operario",
     "cols": ["MES_ANO", "MARC_ANT", "MARC_ACT_original", "M3_original", "obs_operario_original"]},
    {"nombre": "¿Cuál es la anomalía?", "tipo": "anomalia",
     "cols": ["tipo_anomalia", "motivo_detectado"]},
    {"nombre": "¿Cómo se corrige? — supervisor llena", "tipo": "fix",
     "cols": ["MARC_ACT_corregido", "M3_corregido", "motivo_correccion", "resuelto_por"]},
    {"nombre": "¿Estado y cuándo?", "tipo": "cuando",
     "cols": ["estado", "ciclo", "fecha_correccion"]},
]

GRUPOS_TRAZABILIDAD = [
    {"nombre": "¿Quién es?", "tipo": "quien",
     "cols": ["MZ", "LT", "NOMBRE"]},
    {"nombre": "¿Qué entregó el operario? — inmutable", "tipo": "operario",
     "cols": ["MES_ANO", "MARC_ANT", "MARC_ACT_original", "M3_original", "obs_operario_original"]},
    {"nombre": "¿Cuál fue la anomalía?", "tipo": "anomalia",
     "cols": ["categoria", "tipo_anomalia", "motivo_detectado"]},
    {"nombre": "¿Cómo se resolvió?", "tipo": "fix",
     "cols": ["MARC_ACT_final", "M3_final", "motivo_correccion", "resuelto_por"]},
    {"nombre": "¿Cuándo?", "tipo": "cuando",
     "cols": ["ciclo", "fecha_correccion"]},
]

GRUPOS_LECTURAS_PLANILLA = [
    {"nombre": "¿Quién es?", "tipo": "quien",
     "cols": ["MZ", "LT", "NOMBRE", "MES_ANO"]},
    {"nombre": "¿Qué se factura?", "tipo": "fix",
     "cols": ["MARC_ANT", "MARC_ACT", "M3"]},
    {"nombre": "¿Cómo se llegó?", "tipo": "cuando",
     "cols": ["origen", "ciclo"]},
]


# ── Función principal de escritura ────────────────────────────────────────────
def escribir_con_grupos(ws, grupos: list[dict], filas: list[dict]) -> list[str]:
    """Escribe la hoja con cabecera de 2 filas (grupo + columna) y datos coloreados.

    grupos:
        [{"nombre": "...", "tipo": "quien|operario|anomalia|fix|cuando",
          "cols": ["MZ", "LT", ...]}, ...]
    filas:
        [{"MZ": "A", "LT": "1", ...}, ...]

    Devuelve la lista plana de nombres de columna (para introspección externa).
    """
    borde = _borde()

    # Fila 1: nombres de grupo
    col = 1
    for g in grupos:
        n = len(g["cols"])
        pal = PALETA[g["tipo"]]
        c = ws.cell(row=1, column=col, value=g["nombre"])
        c.font = Font(name="Arial", bold=True, color=pal["head_fg"], size=10)
        c.fill = PatternFill("solid", start_color=pal["head_bg"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borde
        if n > 1:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + n - 1)
            # aplicar borde a las celdas merge restantes
            for j in range(col + 1, col + n):
                ws.cell(row=1, column=j).border = borde
        col += n
    ws.row_dimensions[1].height = 22

    # Fila 2: nombres de columnas individuales
    todas_cols = []
    col = 1
    for g in grupos:
        pal = PALETA[g["tipo"]]
        for nombre_col in g["cols"]:
            c = ws.cell(row=2, column=col, value=nombre_col)
            c.font = Font(name="Arial", bold=True, color=pal["head_bg"], size=9)
            c.fill = PatternFill("solid", start_color=pal["sub_bg"])
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borde
            ws.column_dimensions[get_column_letter(col)].width = max(13, len(nombre_col) + 2)
            todas_cols.append(nombre_col)
            col += 1
    ws.row_dimensions[2].height = 20

    # Filas 3+: datos con color de fondo según grupo
    for ri_offset, fila in enumerate(filas, 3):
        col = 1
        for g in grupos:
            pal = PALETA[g["tipo"]]
            for nombre_col in g["cols"]:
                val = fila.get(nombre_col)
                if val == "" or val is None:
                    val = None
                c = ws.cell(row=ri_offset, column=col, value=val)
                c.font = Font(name="Arial", size=9, color=pal["fg"])
                c.alignment = Alignment(
                    horizontal="left" if nombre_col == "NOMBRE" else "center",
                    vertical="center",
                )
                c.border = borde
                c.fill = PatternFill("solid", start_color=pal["data_bg"])
                col += 1
        ws.row_dimensions[ri_offset].height = 15

    # Freeze panes debajo de los 2 headers
    ws.freeze_panes = "A3"
    return todas_cols


def leer_filas_existentes(path, grupos: list[dict]) -> list[dict]:
    """Lee filas de un archivo ya escrito con escribir_con_grupos.

    Asume:
        Fila 1: nombres de grupo
        Fila 2: nombres de columna
        Fila 3+: datos

    Devuelve lista de dicts {nombre_col: value}.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # Reconstruir nombres de columna desde fila 2
    todas_cols = []
    for g in grupos:
        todas_cols.extend(g["cols"])

    n_cols = len(todas_cols)
    filas = []
    for ri in range(3, ws.max_row + 1):
        row_vals = [ws.cell(ri, c).value for c in range(1, n_cols + 1)]
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue
        filas.append(dict(zip(todas_cols, row_vals)))
    return filas
