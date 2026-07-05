"""Test de integración — una resolución ya aplicada nunca se vuelve a pedir.

Bug real (2026-07-03): cuando una anomalía se resuelve (resuelto_por + M3_corregido
en correcciones_YYYY-MM.xlsx), main.py la mueve a trazabilidad_YYYY-MM.xlsx y la
saca de correcciones — pero el dato fuente (registro_operario_mes.xlsx) nunca se
corrige. Si main.py se vuelve a correr más adelante (recuperación de datos, re-abrir
un ciclo ya cerrado, etc.) sin volver a pegar la corrección en correcciones, la
misma anomalía se detecta de cero — como si nunca se hubiera resuelto.

Este test fuerza exactamente esa secuencia:
  Ciclo 1 → bloqueante DIFERENCIA_M3
  Supervisor resuelve (M3_corregido=10, resuelto_por=corrige_dato)
  Ciclo 2 → se resuelve, cierra el mes (bloqueantes=0)
  Ciclo 3 (re-correr sin tocar nada) → NO debe reaparecer el bloqueante

Correr:
    python tests/test_persistencia_resoluciones.py
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

THIS = Path(__file__).resolve()
sys.path.insert(0, str(THIS.parent.parent))
import config  # noqa: E402
import formato_excel as fe  # noqa: E402

TEST_ROOT = THIS.parent / "_tmp_persistencia"
MES_ACTUAL = "2026-06"
MES_PREVIO = "2026-05"

MZ, LT, NOMBRE = "Z", "9", "TEST PERSISTENCIA"
MARC_ANT = 100
MARC_ACT = 110       # calc real = 10
M3_MAL_ANOTADO = 15  # el operario anotó mal — dispara DIFERENCIA_M3
M3_CORREGIDO = 10    # lo que el supervisor confirma que es correcto


def _setup_paths() -> Path:
    if TEST_ROOT.exists():
        shutil.rmtree(TEST_ROOT)
    inputs = TEST_ROOT / "inputs"
    outputs = TEST_ROOT / "outputs"
    inputs.mkdir(parents=True)
    outputs.mkdir(parents=True)

    config.INPUTS_DIR = inputs
    config.OUTPUTS_DIR = outputs
    config.BACKUPS_DIR = inputs / "backups"
    config.REGISTRO_MES_PATH = inputs / "registro_operario_mes.xlsx"
    config.REGISTRO_ACUMULADO_PATH = inputs / "registro_operario_acumulado.xlsx"
    config.LOG_PATH = outputs / "run.log"
    config.correcciones_path = lambda m: outputs / f"correcciones_{m}.xlsx"
    config.trazabilidad_path = lambda m: outputs / f"trazabilidad_{m}.xlsx"
    config.lecturas_planilla_path = lambda m: outputs / f"lecturas_planilla_{m}.xlsx"
    config.orden_verificacion_path = lambda m: outputs / f"orden_verificacion_{m}.pdf"
    config.CORRECCIONES_HISTORICAS_PATH = outputs / "correcciones_historicas.xlsx"
    return outputs


def _build_acumulado() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Acumulada"
    borde = Border(*[Side(style="thin", color="CCCCCC")] * 4)
    for ci, h in enumerate(["MZ", "LT", "NOMBRE"], 1):
        c = ws.cell(1, ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1E3A5F")
        ws.merge_cells(start_row=1, start_column=ci, end_row=2, end_column=ci)
    c1 = ws.cell(1, 4, value=MES_PREVIO)
    c1.font = Font(bold=True, color="FFFFFF")
    c1.fill = PatternFill("solid", start_color="0369A1")
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
    for offset, sub in enumerate(["MARCACION", "M3"]):
        cs = ws.cell(2, 4 + offset, value=sub)
        cs.font = Font(bold=True, color="1A5276")
        cs.fill = PatternFill("solid", start_color="AED6F1")
    ws.cell(3, 1, value=MZ)
    ws.cell(3, 2, value=LT)
    ws.cell(3, 3, value=NOMBRE)
    ws.cell(3, 4, value=MARC_ANT)
    wb.save(config.REGISTRO_ACUMULADO_PATH)


def _build_mes() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lecturas"
    ws.cell(1, 1, value="LEYENDA")
    headers = ["MZ", "LT", "NOMBRE", "MES_ANO", "MARC_ANT", "MARC_ACT", "M3", "obs_operario"]
    for ci, h in enumerate(headers, 1):
        ws.cell(7, ci, value=h)
    ws.cell(8, 1, value=MZ)
    ws.cell(8, 2, value=LT)
    ws.cell(8, 3, value=NOMBRE)
    ws.cell(8, 4, value=MES_ACTUAL)
    ws.cell(8, 5, value=MARC_ANT)
    ws.cell(8, 6, value=MARC_ACT)
    ws.cell(8, 7, value=M3_MAL_ANOTADO)
    ws.cell(8, 8, value="")
    wb.save(config.REGISTRO_MES_PATH)


def _resolver_en_correcciones() -> None:
    """Simula al supervisor llenando el bloque verde de correcciones_2026-06.xlsx."""
    path = config.correcciones_path(MES_ACTUAL)
    wb = load_workbook(path)
    ws = wb.active
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers)}
    for ri in range(3, ws.max_row + 1):
        if ws.cell(ri, idx["MZ"]).value == MZ:
            ws.cell(ri, idx["M3_corregido"], value=M3_CORREGIDO)
            ws.cell(ri, idx["motivo_correccion"], value="operario anotó mal el M3")
            ws.cell(ri, idx["resuelto_por"], value="corrige_dato")
    wb.save(path)


def _correr_main() -> None:
    import main as mod_main
    try:
        mod_main.main()
    except SystemExit:
        pass


def _tiene_bloqueante_diferencia_m3(outputs: Path) -> bool:
    path = config.correcciones_path(MES_ACTUAL)
    if not path.exists():
        return False
    filas = fe.leer_filas_existentes(path, fe.GRUPOS_CORRECCIONES)
    return any(f.get("MZ") == MZ and f.get("tipo_anomalia") == "DIFERENCIA_M3" for f in filas)


def main() -> int:
    print("\n" + "═" * 70)
    print("  Test — una resolución ya aplicada nunca se vuelve a pedir")
    print("═" * 70)

    outputs = _setup_paths()
    _build_acumulado()
    _build_mes()
    ok = True

    print("\n[1] Ciclo 1 — detecta DIFERENCIA_M3")
    _correr_main()
    if _tiene_bloqueante_diferencia_m3(outputs):
        print("  ✓ DIFERENCIA_M3 detectado como se esperaba")
    else:
        print("  ✗ no se detectó la anomalía esperada")
        ok = False

    print("\n[2] Supervisor resuelve · Ciclo 2 — cierra el mes")
    _resolver_en_correcciones()
    _correr_main()
    if not config.correcciones_path(MES_ACTUAL).exists():
        print("  ✓ correcciones_2026-06.xlsx desapareció — ciclo cerrado")
    else:
        print("  ✗ correcciones_2026-06.xlsx todavía existe — no cerró")
        ok = False

    print("\n[3] Ciclo 3 — re-correr SIN volver a tocar nada (el caso del bug real)")
    _correr_main()
    if _tiene_bloqueante_diferencia_m3(outputs):
        print("  ✗ DIFERENCIA_M3 reapareció — la resolución se perdió (regresión del bug)")
        ok = False
    else:
        print("  ✓ NO reapareció — main.py recordó la resolución vía trazabilidad")

    print("\n" + "═" * 70)
    print("  ✓ TODOS LOS CASOS PASARON" if ok else "  ✗ HUBO FALLAS — revisar arriba")
    print("═" * 70 + "\n")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
