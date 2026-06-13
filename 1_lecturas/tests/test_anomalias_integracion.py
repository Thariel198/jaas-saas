"""Test de integración — fuerza cada anomalía con datos sintéticos.

Diseñado para complementar `tests/test_lecturas.py` (unitario).
Aquí construimos un módulo 1_lecturas completo en una carpeta temporal aislada
(`tests/_tmp_integracion/`), corremos `main.main()` end-to-end, y verificamos que:

  - Las 8 anomalías bloqueantes se detectan (1 ó 2 instancias cada una)
  - Las 6 anomalías informativas se detectan
  - lecturas_planilla NO se genera (porque quedan bloqueantes)
  - orden_verificacion PDF SÍ se genera con los casos de campo

Correr:
    python tests/test_anomalias_integracion.py

Pass = todos los chequeos verdes.
"""
from __future__ import annotations

import shutil
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Importar el módulo padre
THIS = Path(__file__).resolve()
sys.path.insert(0, str(THIS.parent.parent))
import config  # noqa: E402
import formato_excel as fe  # noqa: E402

# ── Setup de paths temporales ─────────────────────────────────────────────────
TEST_ROOT = THIS.parent / "_tmp_integracion"

MES_ACTUAL = "2026-06"
MESES_PREVIOS = ["2026-02", "2026-03", "2026-04", "2026-05"]


def _setup_paths():
    if TEST_ROOT.exists():
        shutil.rmtree(TEST_ROOT)
    inputs = TEST_ROOT / "inputs"
    outputs = TEST_ROOT / "outputs"
    inputs.mkdir(parents=True)
    outputs.mkdir(parents=True)

    config.INPUTS_DIR = inputs
    config.OUTPUTS_DIR = outputs
    config.REGISTRO_MES_PATH = inputs / "registro_operario_mes.xlsx"
    config.REGISTRO_ACUMULADO_PATH = inputs / "registro_operario_acumulado.xlsx"
    config.LOG_PATH = outputs / "run.log"
    config.correcciones_path = lambda m: outputs / f"correcciones_{m}.xlsx"
    config.trazabilidad_path = lambda m: outputs / f"trazabilidad_{m}.xlsx"
    config.lecturas_planilla_path = lambda m: outputs / f"lecturas_planilla_{m}.xlsx"
    config.orden_verificacion_path = lambda m: outputs / f"orden_verificacion_{m}.pdf"


# ── Construcción de fixtures ──────────────────────────────────────────────────
def _build_acumulado(usuarios: dict) -> None:
    """Construye registro_operario_acumulado.xlsx con doble header.

    usuarios: {(mz, lt): {"nombre": str, "ciclos": {mes: (marc, m3)}}}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Acumulada"
    borde = Border(*[Side(style="thin", color="CCCCCC")] * 4)

    # Headers fijos con merge vertical
    for ci, h in enumerate(["MZ", "LT", "NOMBRE"], 1):
        c = ws.cell(1, ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1E3A5F")
        c.alignment = Alignment(horizontal="center")
        c.border = borde
        ws.merge_cells(start_row=1, start_column=ci, end_row=2, end_column=ci)

    # Pares de columnas por mes
    mes_cols = {}
    col = 4
    ultimo = MESES_PREVIOS[-1]
    for mes in MESES_PREVIOS:
        color = "0369A1" if mes == ultimo else "5DADE2"
        c = ws.cell(1, col, value=mes)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=color)
        c.alignment = Alignment(horizontal="center")
        c.border = borde
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        for offset, sub in enumerate(["MARCACION", "M3"]):
            cs = ws.cell(2, col + offset, value=sub)
            cs.font = Font(bold=True, color="1A5276")
            cs.fill = PatternFill("solid", start_color="AED6F1")
            cs.alignment = Alignment(horizontal="center")
            cs.border = borde
        mes_cols[mes] = (col, col + 1)
        col += 2

    # Datos
    for ri, (key, datos) in enumerate(sorted(usuarios.items()), 3):
        mz, lt = key
        ws.cell(ri, 1, value=mz)
        ws.cell(ri, 2, value=lt)
        ws.cell(ri, 3, value=datos["nombre"])
        for mes, (marc, m3) in datos.get("ciclos", {}).items():
            if mes in mes_cols:
                cm, cm3 = mes_cols[mes]
                if marc is not None:
                    ws.cell(ri, cm, value=marc)
                if m3 is not None:
                    ws.cell(ri, cm3, value=m3)

    wb.save(config.REGISTRO_ACUMULADO_PATH)


def _build_mes(filas: list[tuple]) -> None:
    """Construye registro_operario_mes.xlsx con leyenda + header + datos.

    filas: lista de (mz, lt, nombre, mes_ano, marc_ant, marc_act, m3, obs_operario)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Lecturas"

    # Leyenda (filas 1-5)
    ws.cell(1, 1, value="LEYENDA — códigos válidos en obs_operario")
    leyenda = [
        ("M", "Medidor cambiado"),
        ("F", "Fuga visible reportada"),
        ("P", "Predio cerrado o inaccesible"),
        ("", "(vacío) lectura normal"),
    ]
    for offset, (codigo, signif) in enumerate(leyenda, 2):
        ws.cell(offset, 1, value=codigo)
        ws.cell(offset, 2, value=signif)

    # Fila 6 vacía (separador), fila 7 = header
    headers = ["MZ", "LT", "NOMBRE", "MES_ANO", "MARC_ANT", "MARC_ACT", "M3", "obs_operario"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(7, ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1E3A5F")

    # Datos desde fila 8
    for ri, fila in enumerate(filas, 8):
        for ci, val in enumerate(fila, 1):
            ws.cell(ri, ci, value=val)

    wb.save(config.REGISTRO_MES_PATH)


# ── Casos sintéticos ──────────────────────────────────────────────────────────
ACUMULADO = {
    ("A", "1"):  {"nombre": "JUAN PEREZ (directo + mínimo)",     "ciclos": {"2026-05": (100, 8)}},
    ("A", "2"):  {"nombre": "ROSA M. (MEDIDOR_INVERTIDO)",        "ciclos": {"2026-05": (80, 8)}},
    ("A", "3"):  {"nombre": "LUIS G. (POSIBLE_CAMBIO_MEDIDOR)",   "ciclos": {"2026-05": (1600, 10)}},
    ("A", "4"):  {"nombre": "ANA T. (DIFERENCIA_M3)",             "ciclos": {"2026-05": (200, 8)}},
    ("A", "5"):  {"nombre": "ESTEBAN V. (EXCESIVO)",              "ciclos": {"2026-05": (300, 10)}},
    ("A", "6"):  {"nombre": "PEDRO S. (SIN_LECTURA bloq)",        "ciclos": {"2026-05": (90, 8)}},
    ("A", "7"):  {"nombre": "ELENA M. (DUPLICADO)",               "ciclos": {"2026-05": (120, 8)}},
    ("A", "8"):  {"nombre": "CARLOS R. (USUARIO_FANTASMA)",       "ciclos": {"2026-05": (150, 6)}},
    ("A", "9"):  {"nombre": "JULIA R. (MARC_ACT_NO_NUMERICO)",    "ciclos": {"2026-05": (250, 8)}},
    ("A", "10"): {"nombre": "MARIA A. (CONSUMO_CERO)",            "ciclos": {"2026-05": (100, 5)}},
    ("A", "11"): {"nombre": "FELIX C. (SALTO_HISTORICO)",         "ciclos": {
        "2026-02": (170, 8), "2026-03": (180, 10),
        "2026-04": (190, 12), "2026-05": (200, 10),
    }},
    ("A", "12"): {"nombre": "ENRIQUE B. (MEDIDOR_CAMBIADO obs M)", "ciclos": {"2026-05": (500, 8)}},
    ("A", "13"): {"nombre": "ROCIO P. (FUGA_REPORTADA obs F)",   "ciclos": {"2026-05": (300, 12)}},
    ("A", "14"): {"nombre": "TOMAS L. (SIN_LECTURA legítimo P)", "ciclos": {"2026-05": (70, 5)}},
}

MES_FILAS = [
    # (mz, lt, nombre, mes_ano, marc_ant, marc_act, m3, obs_operario)
    ("A", "1",  "JUAN PEREZ",      MES_ACTUAL, 100,  104,   5,    ""),   # directo (calc=4<5, M3=5 → mínimo)
    ("A", "2",  "ROSA M.",         MES_ACTUAL, 80,   75,    0,    ""),   # MEDIDOR_INVERTIDO (diff −5)
    ("A", "3",  "LUIS G.",         MES_ACTUAL, 1600, 12,    0,    ""),   # POSIBLE_CAMBIO_MEDIDOR (diff −1588)
    ("A", "4",  "ANA T.",          MES_ACTUAL, 200,  215,   10,   ""),   # DIFERENCIA_M3 (calc=15, M3=10)
    ("A", "5",  "ESTEBAN V.",      MES_ACTUAL, 300,  365,   65,   ""),   # EXCESIVO (calc=65>50)
    ("A", "6",  "PEDRO S.",        MES_ACTUAL, 90,   None,  None, ""),   # SIN_LECTURA bloqueante
    ("A", "7",  "ELENA M.",        MES_ACTUAL, 120,  130,   10,   ""),   # DUPLICADO (1ra fila)
    ("A", "7",  "ELENA M.",        MES_ACTUAL, 120,  130,   10,   ""),   # DUPLICADO (2da fila)
    # A-8 NO está en el mes → USUARIO_FANTASMA
    ("A", "9",  "JULIA R.",        MES_ACTUAL, 250,  "2O5", None, ""),   # MARC_ACT_NO_NUMERICO
    ("A", "10", "MARIA A.",        MES_ACTUAL, 100,  100,   0,    ""),   # CONSUMO_CERO (informativa)
    ("A", "11", "FELIX C.",        MES_ACTUAL, 200,  245,   45,   ""),   # SALTO_HISTORICO (45 > 3×10)
    ("A", "12", "ENRIQUE B.",      MES_ACTUAL, 500,  30,    30,   "M"),  # MEDIDOR_CAMBIADO (obs=M)
    ("A", "13", "ROCIO P.",        MES_ACTUAL, 300,  312,   12,   "F"),  # FUGA_REPORTADA (obs=F)
    ("A", "14", "TOMAS L.",        MES_ACTUAL, 70,   None,  None, "P"),  # SIN_LECTURA informativa (obs=P)
    ("A", "15", "NUEVO USUARIO",   MES_ACTUAL, None, 50,    5,    ""),   # SIN_HISTORIAL informativa
]

# Resultados esperados (después de Ciclo 1)
ESPERADO_BLOQUEANTES = {
    "MEDIDOR_INVERTIDO":      1,
    "POSIBLE_CAMBIO_MEDIDOR": 1,
    "DIFERENCIA_M3":          1,
    "EXCESIVO":               1,
    "SIN_LECTURA":            1,  # solo A-6 (A-14 con P pasa a informativa)
    "DUPLICADO":              2,  # las 2 filas de A-7
    "USUARIO_FANTASMA":       1,  # A-8
    "MARC_ACT_NO_NUMERICO":   1,
}

ESPERADO_INFORMATIVAS = {
    "SIN_HISTORIAL":   1,
    "CONSUMO_CERO":    1,
    "SALTO_HISTORICO": 1,
    "MEDIDOR_CAMBIADO": 1,
    "FUGA_REPORTADA":  1,
    "SIN_LECTURA":     1,  # A-14 con obs=P (informativa)
}


# ── Verificación ──────────────────────────────────────────────────────────────
def _verificar() -> tuple[bool, list[str]]:
    """Compara correcciones y trazabilidad generadas vs lo esperado.

    Devuelve (ok, mensajes).
    """
    msgs = []
    ok = True

    # 1. correcciones.xlsx
    corr_path = config.correcciones_path(MES_ACTUAL)
    if not corr_path.exists():
        return False, ["correcciones.xlsx NO se generó (se esperaban bloqueantes)"]
    filas_corr = fe.leer_filas_existentes(corr_path, fe.GRUPOS_CORRECCIONES)
    tipos_bloq = Counter(str(f.get("tipo_anomalia")) for f in filas_corr)
    for tipo, esperado in ESPERADO_BLOQUEANTES.items():
        obtenido = tipos_bloq.get(tipo, 0)
        marca = "✓" if obtenido == esperado else "✗"
        msgs.append(f"  {marca} bloqueante {tipo:<24} esperado={esperado}  obtenido={obtenido}")
        if obtenido != esperado:
            ok = False
    extras = set(tipos_bloq) - set(ESPERADO_BLOQUEANTES)
    if extras:
        msgs.append(f"  ✗ bloqueantes inesperados en correcciones: {extras}")
        ok = False

    # 2. trazabilidad.xlsx
    traz_path = config.trazabilidad_path(MES_ACTUAL)
    if not traz_path.exists():
        msgs.append("✗ trazabilidad.xlsx NO se generó (se esperaban informativas)")
        return False, msgs
    filas_traz = fe.leer_filas_existentes(traz_path, fe.GRUPOS_TRAZABILIDAD)
    tipos_inf = Counter(str(f.get("tipo_anomalia")) for f in filas_traz
                        if str(f.get("categoria")) == "informativa")
    for tipo, esperado in ESPERADO_INFORMATIVAS.items():
        obtenido = tipos_inf.get(tipo, 0)
        marca = "✓" if obtenido == esperado else "✗"
        msgs.append(f"  {marca} informativa {tipo:<24} esperado={esperado}  obtenido={obtenido}")
        if obtenido != esperado:
            ok = False

    # 3. lecturas_planilla NO debe existir (hay bloqueantes)
    lp_path = config.lecturas_planilla_path(MES_ACTUAL)
    if lp_path.exists():
        msgs.append(f"  ✗ lecturas_planilla.xlsx existe (no debería con bloqueantes pendientes)")
        ok = False
    else:
        msgs.append("  ✓ lecturas_planilla.xlsx no existe (correcto, hay bloqueantes)")

    # 4. PDF de campo
    pdf_path = config.orden_verificacion_path(MES_ACTUAL)
    docx_path = pdf_path.with_suffix(".docx")
    if pdf_path.exists() or docx_path.exists():
        msgs.append("  ✓ orden_verificacion generado")
    else:
        msgs.append("  ✗ orden_verificacion NO se generó")
        ok = False

    return ok, msgs


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "═" * 70)
    print("  Test de integración — anomalías forzadas")
    print("═" * 70)

    print("\n[1] Setup paths temporales en", TEST_ROOT)
    _setup_paths()

    print("[2] Construyendo registro_operario_acumulado (14 usuarios, 4 ciclos)...")
    _build_acumulado(ACUMULADO)

    print(f"[3] Construyendo registro_operario_mes ({len(MES_FILAS)} filas, casos forzados)...")
    _build_mes(MES_FILAS)

    print("[4] Ejecutando main.py contra fixtures...")
    # Importar main aquí (después de monkey-patch de config) para que use los paths nuevos
    import main as mod_main
    try:
        mod_main.main()
    except SystemExit:
        pass  # main() puede llamar sys.exit en algunos paths

    print("\n[5] Verificando resultados...")
    ok, msgs = _verificar()
    for m in msgs:
        print(m)

    print("\n" + "═" * 70)
    if ok:
        print("  ✓ TODOS LOS CASOS PASARON")
    else:
        print("  ✗ HUBO FALLAS — revisar arriba")
    print("═" * 70 + "\n")

    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
