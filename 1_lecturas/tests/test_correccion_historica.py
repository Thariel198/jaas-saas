"""Test de integración — corrección de un error histórico en el acumulado.

Construye un caso mínimo donde el ciclo anterior (2026-05) tiene un MARC_ANT
mal cargado (9998 en vez de 998). Verifica:

  1. Sin corrección histórica: la anomalía (POSIBLE_CAMBIO_MEDIDOR) bloquea el ciclo.
  2. Con una fila en correcciones_historicas.xlsx (ESTADO=activo): la anomalía
     desaparece sola al re-correr — valor_efectivo() la aplica sin tocar el acumulado.
  3. registro_operario_acumulado.xlsx NUNCA se edita — el 9998 sigue ahí.

Correr:
    python tests/test_correccion_historica.py

Pass = todos los chequeos verdes.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

THIS = Path(__file__).resolve()
sys.path.insert(0, str(THIS.parent.parent))
import config  # noqa: E402
import formato_excel as fe  # noqa: E402

TEST_ROOT = THIS.parent / "_tmp_correccion_historica"
MES_ACTUAL = "2026-06"
MES_PREVIO = "2026-05"

MZ, LT, NOMBRE = "Z", "1", "TEST CORRECCION HISTORICA"
MARC_ANT_ERRONEO = 9998   # dígito de más — cargado mal en mayo, nunca corregido en el acumulado
MARC_ANT_REAL    = 998    # lo que el supervisor confirma que debió ser
MARC_ACT_JUNIO   = 1030
M3_JUNIO         = 32     # 1030 - 998 = 32 (plausible, < M3_EXCESIVO)


def _setup_paths() -> Path:
    if TEST_ROOT.exists():
        shutil.rmtree(TEST_ROOT)
    inputs = TEST_ROOT / "inputs"
    outputs = TEST_ROOT / "outputs"
    inputs.mkdir(parents=True)
    outputs.mkdir(parents=True)

    config.INPUTS_DIR = inputs
    config.OUTPUTS_DIR = outputs
    config.BACKUPS_DIR = inputs / "backups"
    config.REGISTRO_MES_PATH = inputs / "registro_operario_mes.xlsx"
    config.REGISTRO_ACUMULADO_PATH = inputs / "registro_operario_acumulado.xlsx"
    config.LOG_PATH = outputs / "run.log"
    config.correcciones_path = lambda m: outputs / f"correcciones_{m}.xlsx"
    config.trazabilidad_path = lambda m: outputs / f"trazabilidad_{m}.xlsx"
    config.lecturas_planilla_path = lambda m: outputs / f"lecturas_planilla_{m}.xlsx"
    config.orden_verificacion_path = lambda m: outputs / f"orden_verificacion_{m}.pdf"
    config.CORRECCIONES_HISTORICAS_PATH = outputs / "correcciones_historicas.xlsx"
    return outputs


def _build_acumulado() -> None:
    """registro_operario_acumulado.xlsx (schema legacy, 1 usuario, 1 ciclo previo)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Acumulada"
    borde = Border(*[Side(style="thin", color="CCCCCC")] * 4)

    for ci, h in enumerate(["MZ", "LT", "NOMBRE"], 1):
        c = ws.cell(1, ci, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1E3A5F")
        c.alignment = Alignment(horizontal="center")
        c.border = borde
        ws.merge_cells(start_row=1, start_column=ci, end_row=2, end_column=ci)

    c1 = ws.cell(1, 4, value=MES_PREVIO)
    c1.font = Font(bold=True, color="FFFFFF")
    c1.fill = PatternFill("solid", start_color="0369A1")
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
    for offset, sub in enumerate(["MARCACION", "M3"]):
        cs = ws.cell(2, 4 + offset, value=sub)
        cs.font = Font(bold=True, color="1A5276")
        cs.fill = PatternFill("solid", start_color="AED6F1")

    ws.cell(3, 1, value=MZ)
    ws.cell(3, 2, value=LT)
    ws.cell(3, 3, value=NOMBRE)
    ws.cell(3, 4, value=MARC_ANT_ERRONEO)  # el error histórico — nunca se edita

    wb.save(config.REGISTRO_ACUMULADO_PATH)


def _build_mes() -> None:
    """registro_operario_mes.xlsx — el operario lee normal en junio, el error viene de mayo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Lecturas"
    ws.cell(1, 1, value="LEYENDA")
    headers = ["MZ", "LT", "NOMBRE", "MES_ANO", "MARC_ANT", "MARC_ACT", "M3", "obs_operario"]
    for ci, h in enumerate(headers, 1):
        ws.cell(7, ci, value=h)
    ws.cell(8, 1, value=MZ)
    ws.cell(8, 2, value=LT)
    ws.cell(8, 3, value=NOMBRE)
    ws.cell(8, 4, value=MES_ACTUAL)
    ws.cell(8, 5, value=MARC_ANT_ERRONEO)
    ws.cell(8, 6, value=MARC_ACT_JUNIO)
    ws.cell(8, 7, value=M3_JUNIO)
    ws.cell(8, 8, value="")
    wb.save(config.REGISTRO_MES_PATH)


def _agregar_correccion_historica() -> None:
    """Escribe la fila que el supervisor agregaría a mano en correcciones_historicas.xlsx."""
    path = config.CORRECCIONES_HISTORICAS_PATH
    wb = Workbook()
    ws = wb.active
    ws.title = "CorreccionesHistoricas"
    fila = {
        "MZ": MZ, "LT": LT, "NOMBRE": NOMBRE,
        "CICLO_CORREGIDO": MES_PREVIO, "CAMPO": "MARCACION",
        "VALOR_ORIGINAL": MARC_ANT_ERRONEO, "VALOR_CORREGIDO": MARC_ANT_REAL,
        "MOTIVO": "test: dígito de más cargado en mayo",
        "DETECTADO_EN_CICLO": MES_ACTUAL, "FECHA": "03/07/2026 09:00",
        "ESTADO": "activo",
    }
    fe.escribir_con_grupos(ws, fe.GRUPOS_CORRECCIONES_HISTORICAS, [fila])
    wb.save(path)


def _correr_main() -> None:
    import main as mod_main
    try:
        mod_main.main()
    except SystemExit:
        pass


def main() -> int:
    print("\n" + "═" * 70)
    print("  Test — corrección de error histórico en el acumulado")
    print("═" * 70)

    outputs = _setup_paths()
    _build_acumulado()
    _build_mes()

    print("\n[1] Ciclo 1 — SIN corrección histórica")
    _correr_main()

    corr_path = config.correcciones_path(MES_ACTUAL)
    ok = True
    if not corr_path.exists():
        print("  ✗ correcciones_2026-06.xlsx no se generó (se esperaba POSIBLE_CAMBIO_MEDIDOR)")
        ok = False
    else:
        filas = fe.leer_filas_existentes(corr_path, fe.GRUPOS_CORRECCIONES)
        tipos = [f.get("tipo_anomalia") for f in filas]
        if tipos == ["POSIBLE_CAMBIO_MEDIDOR"]:
            print("  ✓ POSIBLE_CAMBIO_MEDIDOR detectado (MARC_ANT erróneo hereda del acumulado)")
        else:
            print(f"  ✗ tipos inesperados en correcciones: {tipos}")
            ok = False

    print("\n[2] Supervisor agrega fila en correcciones_historicas.xlsx (ESTADO=activo)")
    _agregar_correccion_historica()

    print("[3] Ciclo 2 — re-correr con la corrección histórica activa")
    _correr_main()

    if corr_path.exists():
        filas = fe.leer_filas_existentes(corr_path, fe.GRUPOS_CORRECCIONES)
        print(f"  ✗ correcciones_2026-06.xlsx sigue existiendo con {len(filas)} fila(s) — no se resolvió")
        ok = False
    else:
        print("  ✓ correcciones_2026-06.xlsx desapareció — la anomalía se resolvió sola")

    lp_path = config.lecturas_planilla_path(MES_ACTUAL)
    if not lp_path.exists():
        print("  ✗ lecturas_planilla_2026-06.xlsx no se generó")
        ok = False
    else:
        filas = fe.leer_filas_existentes(lp_path, fe.GRUPOS_LECTURAS_PLANILLA)
        fila = filas[0] if filas else {}
        marc_ant_usado = fila.get("MARC_ANT")
        if marc_ant_usado == MARC_ANT_REAL:
            print(f"  ✓ lecturas_planilla usa MARC_ANT={marc_ant_usado} (corregido, no el {MARC_ANT_ERRONEO} crudo)")
        else:
            print(f"  ✗ MARC_ANT esperado={MARC_ANT_REAL} obtenido={marc_ant_usado}")
            ok = False

    print("\n[4] El acumulado NUNCA se edita — el valor crudo de mayo sigue siendo el erróneo")
    import main as mod_main
    historial, _ = mod_main._cargar_acumulado()
    valor_crudo = historial.get((MZ, LT), {}).get("ciclos", {}).get(MES_PREVIO, {}).get("marc")
    if valor_crudo == MARC_ANT_ERRONEO:
        print(f"  ✓ registro_operario_acumulado.xlsx sigue con el valor original ({valor_crudo}) — intacto")
    else:
        print(f"  ✗ el acumulado fue editado — ahora dice {valor_crudo} (debería seguir en {MARC_ANT_ERRONEO})")
        ok = False

    print("\n" + "═" * 70)
    print("  ✓ TODOS LOS CASOS PASARON" if ok else "  ✗ HUBO FALLAS — revisar arriba")
    print("═" * 70 + "\n")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
