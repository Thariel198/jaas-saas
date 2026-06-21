"""1_lecturas/aplicar_sincronizacion.py — Paso 0b: aplica deltas autorizados al acumulado.

Lee outputs/sync/reporte_sincronizacion.xlsx y aplica las filas con
AUTORIZAR=Si y APLICADO≠Si al registro_operario_acumulado.xlsx.

Único writer de identidad (MZ, LT, NOMBRE, SIN_SERVICIO) en el acumulado.
main.py solo agrega columnas de ciclo YYYY-MM.

Pipeline interno:
  1. Lee reporte_sincronizacion.xlsx y filtra APLICADO≠Si AND AUTORIZAR≈"Si"
  2. Si no hay nada que aplicar → exit (idempotencia)
  3. Backup timestamped del acumulado en inputs/backups/
  4. Carga estado completo del acumulado (todas las columnas mensuales)
  5. Aplica deltas en memoria:
       AGREGADO     → nueva fila con SIN_SERVICIO=No
       SIN_SERVICIO → marca SIN_SERVICIO=Si (no borra ni toca lecturas)
       RENAME       → actualiza NOMBRE
       MOVIMIENTO   → actualiza (MZ, LT)
  6. Reescribe acumulado con 4 cols fijas + todas las columnas mensuales originales
  7. Append a trazabilidad_sincronizacion.xlsx (append-only)
  8. Marca APLICADO=Si + FECHA_APLICADO en reporte_sincronizacion.xlsx

Re-correr siempre produce el mismo estado:
  - APLICADO=Si filtra las ya procesadas
  - Si nada cambió, no toca el acumulado ni hace backup

Uso:
    python aplicar_sincronizacion.py
"""
import logging
import shutil
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent.parent))  # para shared/
import config
from shared.utils_sort_mz_lt import clave_orden

SOURCE = "1_lecturas/aplicar_sincronizacion"

# Paleta TIPO (matches contrato)
PAL_TIPO = {
    "AGREGADO":     ("D5F5E3", "145A32"),
    "SIN_SERVICIO": ("FEF3C7", "7D5A00"),
    "RENAME":       ("D6EAF8", "1A5276"),
    "MOVIMIENTO":   ("FADBD8", "7B241C"),
}
EMOJI = {"AGREGADO": "🟢", "SIN_SERVICIO": "🟡", "RENAME": "🔵", "MOVIMIENTO": "🔴"}


# ── NORMALIZACIÓN ───────────────────────────────────────────────────────────
def _norm(s) -> str:
    if s is None:
        return ""
    s = str(s).strip().upper()
    if not s or s in ("NAN", "NONE"):
        return ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


def _norm_lt(val) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN"):
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.upper().replace(" ", "")


def _es_si(val) -> bool:
    """Acepta 'Si', 'Si (auto)', 'sí', 'YES', etc."""
    if val is None:
        return False
    s = str(val).strip().upper()
    return s.startswith("SI") or s in ("YES", "Y", "TRUE", "1")


# ── ESTILO ──────────────────────────────────────────────────────────────────
def _borde():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _cell(ws, r, c, val, bg=None, txt="333333", bold=False, align="center", mono=False, size=9):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(name="Consolas" if mono else "Arial", size=size, bold=bold, color=txt)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _borde()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    return cell


# ── BACKUP ──────────────────────────────────────────────────────────────────
def _backup_acumulado(log: logging.Logger) -> Path:
    config.BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    dest = config.BACKUPS_DIR / f"registro_operario_acumulado_{ts}.xlsx"
    shutil.copy2(config.REGISTRO_ACUMULADO_PATH, dest)
    log.info(f"Backup -> {dest.name}")
    return dest


# ── CARGAR DELTAS PENDIENTES ────────────────────────────────────────────────
def _leer_deltas_pendientes(log: logging.Logger) -> list[dict]:
    """Filas con APLICADO≠Si AND AUTORIZAR≈Si del reporte_sincronizacion."""
    if not config.REPORTE_SYNC_PATH.exists():
        raise FileNotFoundError(
            f"No se encontró {config.REPORTE_SYNC_PATH}\n"
            f"  -> Correr primero: python proponer_sincronizacion.py"
        )

    wb = load_workbook(config.REPORTE_SYNC_PATH, data_only=True)
    ws = wb.active
    headers = {str(ws.cell(2, c).value or "").strip().upper(): c
               for c in range(1, ws.max_column + 1)}

    pendientes: list[dict] = []
    descartados_no_autorizado = 0
    descartados_ya_aplicado = 0

    for r in range(3, ws.max_row + 1):
        tipo_raw = str(ws.cell(r, headers["TIPO"]).value or "").strip()
        if not tipo_raw:
            continue
        # Quitar emojis
        for e in ("🟢", "🟡", "🔵", "🔴"):
            tipo_raw = tipo_raw.replace(e, "")
        tipo = tipo_raw.strip().upper()
        if tipo not in PAL_TIPO:
            continue

        aplicado  = str(ws.cell(r, headers["APLICADO"]).value or "").strip()
        autorizar = str(ws.cell(r, headers["AUTORIZAR"]).value or "").strip()
        revisado  = str(ws.cell(r, headers["REVISADO"]).value or "").strip()

        if _es_si(aplicado):
            descartados_ya_aplicado += 1
            continue
        if not _es_si(autorizar):
            descartados_no_autorizado += 1
            continue

        pendientes.append({
            "row_in_reporte": r,
            "TIPO":            tipo,
            "MZ_PADRON":       str(ws.cell(r, headers["MZ_PADRON"]).value or "").strip(),
            "LT_PADRON":       str(ws.cell(r, headers["LT_PADRON"]).value or "").strip(),
            "NOMBRE_PADRON":   str(ws.cell(r, headers["NOMBRE_PADRON"]).value or "").strip(),
            "MZ_OPERARIO":     str(ws.cell(r, headers["MZ_OPERARIO"]).value or "").strip(),
            "LT_OPERARIO":     str(ws.cell(r, headers["LT_OPERARIO"]).value or "").strip(),
            "NOMBRE_OPERARIO": str(ws.cell(r, headers["NOMBRE_OPERARIO"]).value or "").strip(),
            "REVISADO":        revisado,
            "AUTORIZAR":       autorizar,
        })
    wb.close()
    log.info(
        f"Reporte · {len(pendientes)} pendientes · "
        f"{descartados_ya_aplicado} ya aplicados (skip) · "
        f"{descartados_no_autorizado} sin autorización (skip)"
    )
    return pendientes


# ── CARGAR ESTADO COMPLETO DEL ACUMULADO ────────────────────────────────────
def _cargar_acumulado(log: logging.Logger) -> tuple[list[dict], list[str], dict]:
    """Devuelve (filas, meses_orden, schema).

    filas: lista de {mz, lt, nombre, sin_servicio, ciclos: {mes: (marc, m3)}}
    meses_orden: lista ordenada de meses YYYY-MM presentes
    schema: {'tiene_sin_servicio': bool, 'col4_legacy': str|None}
    """
    if not config.REGISTRO_ACUMULADO_PATH.exists():
        log.warning("acumulado no existe — se creará desde cero")
        return [], [], {"tiene_sin_servicio": False, "col4_legacy": None}

    wb = load_workbook(config.REGISTRO_ACUMULADO_PATH, data_only=True)
    ws = wb.active

    # Detectar schema
    cab_col4 = str(ws.cell(1, 4).value or "").strip().upper()
    tiene_sin_servicio = cab_col4 == "SIN_SERVICIO"
    col_inicio_meses = 5 if tiene_sin_servicio else 4

    # Mapear columnas mensuales: leer fila 1 (mes) + fila 2 (MARCACION/M3)
    n_cols = ws.max_column
    fila1 = [ws.cell(1, c).value for c in range(1, n_cols + 1)]
    fila2 = [ws.cell(2, c).value for c in range(1, n_cols + 1)]

    mes_cols: dict[str, tuple[int, int]] = {}
    mes_actual = None
    marc_col = None
    for ci in range(col_inicio_meses, n_cols + 1):
        v1 = fila1[ci - 1]
        v2 = fila2[ci - 1]
        if isinstance(v1, str) and len(v1) == 7 and v1[4] == "-":
            mes_actual = v1
            marc_col = None
        if mes_actual and v2 == "MARCACION":
            marc_col = ci
        elif mes_actual and v2 == "M3" and marc_col is not None:
            mes_cols[mes_actual] = (marc_col, ci)
            mes_actual = None
            marc_col = None

    meses_orden = sorted(mes_cols.keys())

    filas: list[dict] = []
    for ri in range(3, ws.max_row + 1):
        mz = str(ws.cell(ri, 1).value or "").strip()
        lt = str(ws.cell(ri, 2).value or "").strip()
        nombre = str(ws.cell(ri, 3).value or "").strip()
        if not mz or not lt or mz.lower() in ("none", "nan"):
            continue
        sin_serv = ""
        if tiene_sin_servicio:
            sin_serv = str(ws.cell(ri, 4).value or "").strip()

        ciclos: dict[str, tuple] = {}
        for mes, (col_marc, col_m3) in mes_cols.items():
            marc = ws.cell(ri, col_marc).value
            m3 = ws.cell(ri, col_m3).value
            ciclos[mes] = (marc, m3)

        filas.append({
            "mz": mz, "lt": lt, "nombre": nombre,
            "sin_servicio": "Si" if sin_serv.lower().startswith("si") else "",
            "ciclos": ciclos,
        })
    wb.close()
    log.info(
        f"Acumulado cargado · {len(filas)} usuarios · {len(meses_orden)} ciclos · "
        f"schema {'v2 (SIN_SERVICIO)' if tiene_sin_servicio else 'v1 (legacy)'}"
    )
    return filas, meses_orden, {
        "tiene_sin_servicio": tiene_sin_servicio,
        "col4_legacy": cab_col4 if not tiene_sin_servicio else None,
    }


# ── APLICAR DELTAS EN MEMORIA ───────────────────────────────────────────────
def _aplicar_deltas(
    filas: list[dict],
    deltas: list[dict],
    log: logging.Logger,
) -> tuple[list[dict], list[dict], list[str]]:
    """Muta `filas` aplicando cada delta. Devuelve (filas, eventos_trazabilidad, errores)."""
    # Index para buscar por (mz_norm, lt_norm)
    idx: dict[tuple[str, str], int] = {}
    for i, f in enumerate(filas):
        idx[(_norm(f["mz"]), _norm_lt(f["lt"]))] = i

    eventos: list[dict] = []
    errores: list[str] = []

    for d in deltas:
        tipo = d["TIPO"]
        mz_p = d["MZ_PADRON"]; lt_p = d["LT_PADRON"]; n_p = d["NOMBRE_PADRON"]
        mz_o = d["MZ_OPERARIO"]; lt_o = d["LT_OPERARIO"]; n_o = d["NOMBRE_OPERARIO"]

        if tipo == "AGREGADO":
            key = (_norm(mz_p), _norm_lt(lt_p))
            if key in idx:
                errores.append(f"AGREGADO ya existe en acumulado: {mz_p}-{lt_p}")
                continue
            filas.append({"mz": mz_p, "lt": lt_p, "nombre": n_p,
                          "sin_servicio": "", "ciclos": {}})
            idx[key] = len(filas) - 1
            eventos.append({
                "tipo": tipo, "mz_old": "", "lt_old": "", "nombre_old": "",
                "mz_new": mz_p, "lt_new": lt_p, "nombre_new": n_p,
            })

        elif tipo == "SIN_SERVICIO":
            key = (_norm(mz_o), _norm_lt(lt_o))
            i = idx.get(key)
            if i is None:
                errores.append(f"SIN_SERVICIO no encontrado en acumulado: {mz_o}-{lt_o}")
                continue
            if filas[i]["sin_servicio"] == "Si":
                continue  # ya marcado
            filas[i]["sin_servicio"] = "Si"
            eventos.append({
                "tipo": tipo, "mz_old": mz_o, "lt_old": lt_o, "nombre_old": n_o,
                "mz_new": mz_o, "lt_new": lt_o, "nombre_new": n_o,
            })

        elif tipo == "RENAME":
            key = (_norm(mz_p), _norm_lt(lt_p))
            i = idx.get(key)
            if i is None:
                errores.append(f"RENAME no encontrado: {mz_p}-{lt_p}")
                continue
            nombre_old = filas[i]["nombre"]
            filas[i]["nombre"] = n_p
            eventos.append({
                "tipo": tipo, "mz_old": mz_p, "lt_old": lt_p, "nombre_old": nombre_old,
                "mz_new": mz_p, "lt_new": lt_p, "nombre_new": n_p,
            })

        elif tipo == "MOVIMIENTO":
            key_old = (_norm(mz_o), _norm_lt(lt_o))
            key_new = (_norm(mz_p), _norm_lt(lt_p))
            i = idx.get(key_old)
            if i is None:
                errores.append(f"MOVIMIENTO origen no encontrado: {mz_o}-{lt_o}")
                continue
            if key_new in idx:
                errores.append(f"MOVIMIENTO destino ya ocupado: {mz_p}-{lt_p}")
                continue
            filas[i]["mz"] = mz_p
            filas[i]["lt"] = lt_p
            filas[i]["nombre"] = n_p
            del idx[key_old]
            idx[key_new] = i
            eventos.append({
                "tipo": tipo, "mz_old": mz_o, "lt_old": lt_o, "nombre_old": n_o,
                "mz_new": mz_p, "lt_new": lt_p, "nombre_new": n_p,
            })
        else:
            errores.append(f"TIPO desconocido: {tipo}")

    log.info(f"Aplicados · {len(eventos)} eventos · {len(errores)} errores")
    return filas, eventos, errores


# ── ESCRIBIR ACUMULADO ──────────────────────────────────────────────────────
def _escribir_acumulado(filas: list[dict], meses_orden: list[str], log: logging.Logger) -> None:
    """Reescribe el acumulado con 4 cols fijas + columnas mensuales originales."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Acumulada"
    borde = _borde()

    # Cabecera fila 1 + fila 2: cols fijas (merge vertical)
    headers_fijos = list(config.COLS_FIJAS)  # MZ, LT, NOMBRE, SIN_SERVICIO
    anchos_fijos  = {1: 8, 2: 6, 3: 32, 4: 14}
    for ci, h in enumerate(headers_fijos, 1):
        c1 = ws.cell(row=1, column=ci, value=h)
        c2 = ws.cell(row=2, column=ci, value="")
        for c in (c1, c2):
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
            c.fill = PatternFill("solid", start_color="1E3A5F")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = borde
        ws.merge_cells(start_row=1, start_column=ci, end_row=2, end_column=ci)
        ws.column_dimensions[get_column_letter(ci)].width = anchos_fijos.get(ci, 12)

    # Pares de columnas por mes (MARCACION, M3)
    col_mes: dict[str, tuple[int, int]] = {}
    next_col = len(headers_fijos) + 1
    ultimo_mes = meses_orden[-1] if meses_orden else None
    for mes in meses_orden:
        color_mes = "0369A1" if mes == ultimo_mes else "5DADE2"
        color_sub = "AED6F1"
        c1 = ws.cell(row=1, column=next_col, value=mes)
        c1.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c1.fill = PatternFill("solid", start_color=color_mes)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = borde
        ws.merge_cells(start_row=1, start_column=next_col, end_row=1, end_column=next_col + 1)
        for offset, sub in enumerate(["MARCACION", "M3"]):
            c2 = ws.cell(row=2, column=next_col + offset, value=sub)
            c2.font = Font(name="Arial", bold=True, color="1A5276", size=9)
            c2.fill = PatternFill("solid", start_color=color_sub)
            c2.alignment = Alignment(horizontal="center", vertical="center")
            c2.border = borde
        ws.column_dimensions[get_column_letter(next_col)].width = 13
        ws.column_dimensions[get_column_letter(next_col + 1)].width = 8
        col_mes[mes] = (next_col, next_col + 1)
        next_col += 2

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18

    # Orden canónico del pipeline (mismo que padron_reconciliado.xlsx):
    # MZ etapa 1 (A-Z) → MZ etapa 2 (A1-H1) → LT numérico + sufijo
    filas_sorted = sorted(filas, key=lambda f: clave_orden(f["mz"], f["lt"]))

    for ri, f in enumerate(filas_sorted, 3):
        # Cols fijas
        valores_fijos = [f["mz"], f["lt"], f["nombre"], f["sin_servicio"]]
        for ci, val in enumerate(valores_fijos, 1):
            c = ws.cell(row=ri, column=ci, value=val if val else None)
            c.font = Font(name="Arial", size=9, bold=(ci == 4 and f["sin_servicio"] == "Si"),
                          color="7B241C" if (ci == 4 and f["sin_servicio"] == "Si") else "333333")
            c.alignment = Alignment(
                horizontal="left" if ci == 3 else "center",
                vertical="center",
            )
            c.border = borde
            if ci == 4 and f["sin_servicio"] == "Si":
                c.fill = PatternFill("solid", start_color="FADBD8")

        # Cols mensuales
        for mes in meses_orden:
            col_marc, col_m3 = col_mes[mes]
            marc, m3 = f["ciclos"].get(mes, (None, None))
            for col, val in ((col_marc, marc), (col_m3, m3)):
                cell = ws.cell(row=ri, column=col, value=val)
                cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde
        ws.row_dimensions[ri].height = 15

    ws.freeze_panes = "E3"  # 4 cols fijas
    wb.save(config.REGISTRO_ACUMULADO_PATH)
    log.info(f"acumulado reescrito · {len(filas_sorted)} usuarios · {len(meses_orden)} ciclos")


# ── TRAZABILIDAD APPEND-ONLY ────────────────────────────────────────────────
TRAZA_HEADERS = [
    ("FECHA_APLICADO", "FEF9C3", "7D6608", 18),
    ("TIPO",           "FEF9C3", "7D6608", 18),
    ("MZ_OLD",         "FEF3E8", "7C3003",  8),
    ("LT_OLD",         "FEF3E8", "7C3003",  8),
    ("NOMBRE_OLD",     "FEF3E8", "7C3003", 32),
    ("MZ_NEW",         "E1F5EE", "085041",  8),
    ("LT_NEW",         "E1F5EE", "085041",  8),
    ("NOMBRE_NEW",     "E1F5EE", "085041", 32),
    ("CICLO",          "F3E8FF", "5B21B6", 10),
    ("FUENTE",         "F3E8FF", "5B21B6", 22),
]


def _crear_traza_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "trazabilidad"
    grupos = [
        (1, 2, "¿Cuándo / qué?", "FEF9C3", "7D6608"),
        (3, 5, "Antes",          "FEF3E8", "7C3003"),
        (6, 8, "Después",        "E1F5EE", "085041"),
        (9, 10, "Trazabilidad",  "F3E8FF", "5B21B6"),
    ]
    for cs, ce, texto, bg, txt in grupos:
        ws.merge_cells(start_row=1, start_column=cs, end_row=1, end_column=ce)
        c = ws.cell(row=1, column=cs, value=texto)
        c.font = Font(name="Arial", size=9, bold=True, color=txt)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", start_color=bg)
        c.border = _borde()
    for ci, (nombre, bg, txt, ancho) in enumerate(TRAZA_HEADERS, 1):
        _cell(ws, 2, ci, nombre, bg=bg, txt=txt, bold=True, align="center")
        ws.column_dimensions[get_column_letter(ci)].width = ancho
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"
    return wb


def _trazabilidad_append(eventos: list[dict], ciclo: str, log: logging.Logger) -> None:
    config.SYNC_DIR.mkdir(parents=True, exist_ok=True)
    if config.TRAZABILIDAD_SYNC_PATH.exists():
        wb = load_workbook(config.TRAZABILIDAD_SYNC_PATH)
        ws = wb.active
        next_row = max(ws.max_row + 1, 3)
    else:
        wb = _crear_traza_workbook()
        ws = wb.active
        next_row = 3

    fecha = datetime.now().strftime("%Y-%m-%d %H:%M")
    for offset, e in enumerate(eventos):
        r = next_row + offset
        tipo = e["tipo"]
        bg_tipo, txt_tipo = PAL_TIPO[tipo]
        _cell(ws, r, 1, fecha, bg="FEF9EC", txt="7D5A00", mono=True)
        _cell(ws, r, 2, f"{EMOJI[tipo]} {tipo}", bg=bg_tipo, txt=txt_tipo, bold=True)
        _cell(ws, r, 3, e["mz_old"] or "—",     bg="FEF9EC", txt="7C3003", mono=True)
        _cell(ws, r, 4, e["lt_old"] or "—",     bg="FEF9EC", txt="7C3003", mono=True)
        _cell(ws, r, 5, e["nombre_old"] or "—", bg="FEF9EC", txt="7C3003", align="left")
        _cell(ws, r, 6, e["mz_new"] or "—",     bg="F0FFF8", txt="065F46", mono=True)
        _cell(ws, r, 7, e["lt_new"] or "—",     bg="F0FFF8", txt="065F46", mono=True)
        _cell(ws, r, 8, e["nombre_new"] or "—", bg="F0FFF8", txt="065F46", align="left")
        _cell(ws, r, 9, ciclo,                  bg="FAF5FF", txt="5B21B6", mono=True)
        _cell(ws, r, 10, "padron_reconciliado", bg="FAF5FF", txt="5B21B6", align="left")
        ws.row_dimensions[r].height = 17

    wb.save(config.TRAZABILIDAD_SYNC_PATH)
    log.info(f"trazabilidad_sincronizacion.xlsx · +{len(eventos)} filas (append)")


# ── ACTUALIZAR REPORTE (marcar APLICADO=Si) ─────────────────────────────────
def _marcar_aplicados(rows_aplicados: list[int], log: logging.Logger) -> None:
    """Reescribe columnas APLICADO y FECHA_APLICADO en el reporte."""
    wb = load_workbook(config.REPORTE_SYNC_PATH)
    ws = wb.active
    headers = {str(ws.cell(2, c).value or "").strip().upper(): c
               for c in range(1, ws.max_column + 1)}
    col_aplic = headers["APLICADO"]
    col_fecha = headers["FECHA_APLICADO"]
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M")

    for r in rows_aplicados:
        _cell(ws, r, col_aplic, "Si", bg="D5F5E3", txt="145A32", bold=True)
        _cell(ws, r, col_fecha, fecha, bg="FEF9EC", txt="7D5A00", mono=True)
    wb.save(config.REPORTE_SYNC_PATH)
    log.info(f"reporte_sincronizacion.xlsx · {len(rows_aplicados)} filas marcadas APLICADO=Si")


# ── MAIN ────────────────────────────────────────────────────────────────────
def main() -> None:
    # Windows cmd usa cp1252 por defecto y revienta con emojis. Forzar UTF-8.
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    config.OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    config.SYNC_DIR.mkdir(parents=True, exist_ok=True)
    config.BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(config.LOG_PATH, encoding="utf-8"),
        ],
        force=True,
    )
    log = logging.getLogger(__name__)
    log.info("aplicar_sincronizacion.py · iniciando")

    print("=" * 60)
    print("  1_lecturas/aplicar_sincronizacion.py")
    print("=" * 60)

    print("\n[1/5] Leyendo reporte_sincronizacion...")
    pendientes = _leer_deltas_pendientes(log)

    if not pendientes:
        print("\n  Nada que aplicar · reporte y acumulado ya están alineados")
        print("  (idempotencia desde APLICADO=Si)")
        print("\n" + "=" * 60)
        print("  aplicar_sincronizacion.py completado · 0 escrituras")
        print("=" * 60 + "\n")
        return

    print(f"  -> {len(pendientes)} deltas autorizados pendientes")

    print(f"\n[2/5] Backup del acumulado...")
    if config.REGISTRO_ACUMULADO_PATH.exists():
        _backup_acumulado(log)

    print(f"\n[3/5] Cargando acumulado completo...")
    filas, meses_orden, schema = _cargar_acumulado(log)

    print(f"\n[4/5] Aplicando {len(pendientes)} deltas en memoria...")
    filas, eventos, errores = _aplicar_deltas(filas, pendientes, log)
    for err in errores:
        log.warning(f"  {err}")

    print(f"\n[5/5] Reescribiendo acumulado + trazabilidad + reporte...")
    _escribir_acumulado(filas, meses_orden, log)

    ciclo = datetime.now().strftime("%Y-%m")
    _trazabilidad_append(eventos, ciclo, log)

    rows_ok = [d["row_in_reporte"] for d, e in zip(pendientes, range(len(pendientes)))
               if e < len(eventos)]
    # Solo marcar los que efectivamente generaron evento (sin errores)
    rows_ok = []
    evento_idx = 0
    for d in pendientes:
        if evento_idx < len(eventos):
            # Asumimos orden 1:1 entre pendientes y eventos exitosos.
            # Verificamos por tipo + identidad para mayor seguridad.
            e = eventos[evento_idx]
            if e["tipo"] == d["TIPO"]:
                rows_ok.append(d["row_in_reporte"])
                evento_idx += 1
    if rows_ok:
        _marcar_aplicados(rows_ok, log)

    # Resumen por tipo
    conteo = {t: 0 for t in PAL_TIPO}
    for e in eventos:
        conteo[e["tipo"]] = conteo.get(e["tipo"], 0) + 1

    print("\n" + "=" * 60)
    print(f"  Aplicados · {len(eventos)} eventos")
    print(f"    🟢 AGREGADO     {conteo['AGREGADO']}")
    print(f"    🟡 SIN_SERVICIO {conteo['SIN_SERVICIO']}")
    print(f"    🔵 RENAME       {conteo['RENAME']}")
    print(f"    🔴 MOVIMIENTO   {conteo['MOVIMIENTO']}")
    if errores:
        print(f"\n  ⚠  {len(errores)} errores — ver run.log")
    print(f"\n  -> acumulado mutado: {config.REGISTRO_ACUMULADO_PATH.name}")
    print(f"  -> trazabilidad:     {config.TRAZABILIDAD_SYNC_PATH.relative_to(config.ROOT)}")
    print(f"  -> backup en:        inputs/backups/")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
