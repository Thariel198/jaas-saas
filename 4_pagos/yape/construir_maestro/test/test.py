# =========================IMPORTS===========================
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent))
from utils.consolidar import filtrar_origenes_validos

# ========================CONFIGURACION======================
BASE_DIR   = Path(__file__).parent
CASOS_PATH = BASE_DIR / "casos" / "test_casos.xlsx"

# ========================COLORES============================
COLOR_OK   = "D1FAE5"; TXT_OK   = "065F46"
COLOR_FAIL = "FEE2E2"; TXT_FAIL = "991B1B"
COLOR_RES  = "FAF5FF"; TXT_RES  = "5B21B6"

# ========================UTILIDADES=========================
def lado(c="CCCCCC"):
    return Side(style="thin", color=c)

def cel(cell, bg, txt, mono=False, align="left", bold=False):
    b = lado()
    cell.border    = Border(left=b, right=b, top=b, bottom=b)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.font      = Font(name="Courier New" if mono else "Arial",
                          bold=bold, size=10, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)

def normalizar(raw) -> list:
    if not raw:
        return []
    return [o.strip() for o in str(raw).split("|") if o.strip()]

def comparar(obtenidos: list, esperados: list) -> bool:
    return sorted([o.upper() for o in obtenidos]) == \
           sorted([o.upper() for o in esperados])

# ========================CARGA Y EJECUCION==================
def correr_tests() -> tuple[int, int]:
    if not CASOS_PATH.exists():
        raise FileNotFoundError(f"No se encontró {CASOS_PATH}")

    wb   = load_workbook(CASOS_PATH)
    ws   = wb.active
    rows = list(ws.values)

    if len(rows) < 3:
        raise ValueError("test_casos.xlsx vacío o sin datos")

    headers        = [str(h).strip() if h else "" for h in rows[1]]
    col_obtenidos  = headers.index("ORIGENES_OBTENIDOS") + 1
    col_veredicto  = headers.index("VEREDICTO")          + 1

    pasaron = 0
    fallaron = 0

    print(f"\n  {'CASO':<12} {'DESCRIPCION':<35} {'VEREDICTO'}")
    print(f"  {'-'*12} {'-'*35} {'-'*20}")

    for ri, fila in enumerate(rows[2:], start=3):
        if not fila or not any(fila):
            continue

        row      = dict(zip(headers, fila))
        caso_id  = str(row.get("CASO_ID",            "")).strip()
        nombre   = str(row.get("NOMBRE",             "")).strip()
        brutos   = normalizar(row.get("ORIGENES_BRUTOS",    ""))
        esperados = normalizar(row.get("ORIGENES_ESPERADOS", ""))

        if not caso_id or not brutos:
            continue

        # Correr algoritmo
        obtenidos = filtrar_origenes_validos(brutos, nombre)

        ok = comparar(obtenidos, esperados)
        if ok:
            veredicto = "✅"
            pasaron  += 1
            bg_v, txt_v = COLOR_OK, TXT_OK
        else:
            veredicto = "❌ origenes"
            fallaron += 1
            bg_v, txt_v = COLOR_FAIL, TXT_FAIL

        print(f"  {caso_id:<12} {row.get('DESCRIPCION',''):<35} {veredicto}")
        if not ok:
            print(f"    esperado : {sorted(esperados)}")
            print(f"    obtenido : {sorted(obtenidos)}")

        # Escribir en Excel
        c_obt = ws.cell(row=ri, column=col_obtenidos, value="|".join(obtenidos))
        cel(c_obt, COLOR_RES, TXT_RES, mono=True)

        c_ver = ws.cell(row=ri, column=col_veredicto, value=veredicto)
        cel(c_ver, bg_v, txt_v, mono=True, align="center", bold=True)

    wb.save(CASOS_PATH)
    return pasaron, fallaron

# ======================MAIN=================================
def main():
    print("=" * 55)
    print("  TEST — validación de filtrar_origenes_validos")
    print("=" * 55)

    pasaron, fallaron = correr_tests()

    print(f"\n  {'='*55}")
    print(f"  Pasaron : {pasaron}")
    print(f"  Fallaron: {fallaron}")
    print(f"  Total   : {pasaron + fallaron}")
    if fallaron == 0:
        print(f"  ✅ Todos los casos pasaron")
    else:
        print(f"  ❌ {fallaron} caso(s) fallaron · revisa test_casos.xlsx")
    print(f"  {'='*55}")

if __name__ == "__main__":
    main()
