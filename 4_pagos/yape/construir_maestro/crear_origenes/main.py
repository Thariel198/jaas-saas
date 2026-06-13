# =========================IMPORTS===========================
import pandas as pd
from pathlib import Path
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========================CONFIGURACION======================
BASE_DIR         = Path(__file__).parent
SHARED_DIR       = BASE_DIR.parent.parent.parent.parent / "shared"
REPORTE_ACUM_DIR = SHARED_DIR / "reporte_acumulado_procesado"
USUARIOS_ID_PATH = SHARED_DIR / "usuarios_id.xlsx"
OUTPUT_DIR       = BASE_DIR / "outputs"

HOJA_REPORTE = "reporte"
TIPO_PAGO    = "TE PAGÓ"

ALIAS_TIPO   = ["tipo de transacción", "tipo de transaccion", "tipo"]
ALIAS_ORIGEN = ["origen"]
ALIAS_MZ     = ["mz", "manzana"]
ALIAS_LT     = ["lt", "lote"]

# ========================UTILIDADES CARPETAS================
def reset_output_folder(path: Path):
    import shutil
    if path.exists():
        shutil.rmtree(path)
    path.mkdir()

# ========================COLORES============================
COLOR_HEADER  = "4A235A"
COLOR_ID_H    = "F4ECF7"
COLOR_ID_C    = "FAF5FF"
COLOR_ID_TXT  = "5B21B6"
COLOR_UBI_H   = "E1F5EE"
COLOR_UBI_C   = "F0FFF8"
COLOR_UBI_TXT = "085041"
COLOR_ORI_H   = "FEF3E2"
COLOR_ORI_C   = "FFFBF0"
COLOR_ORI_TXT = "92400E"
COLOR_MET_H   = "EFF6FF"
COLOR_MET_C   = "EFF6FF"
COLOR_MET_TXT = "1D4ED8"
COLOR_CTL_H   = "F9FAFB"
COLOR_CTL_C   = "F9FAFB"
COLOR_CTL_TXT = "374151"
COLOR_SEP     = "F3F4F6"

# ========================UTILIDADES=========================
def lado(color="CCCCCC"):
    return Side(style="thin", color=color)

def estilo(cell, bold=False, bg=None, txt=None, align="left", mono=False, size=10):
    b = lado()
    cell.border    = Border(left=b, right=b, top=b, bottom=b)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    font_name      = "Courier New" if mono else "Arial"
    cell.font      = Font(name=font_name, bold=bold, size=size,
                          color=txt if txt else "000000")
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)

def estilo_header(cell, bg, txt, bold=True, size=10):
    b = lado("FFFFFF")
    cell.border    = Border(left=b, right=b, top=b, bottom=b)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font      = Font(name="Arial", bold=bold, size=size, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)

def estilo_sep(cell):
    cell.fill      = PatternFill("solid", start_color=COLOR_SEP)
    cell.border    = Border(
        left=lado("CCCCCC"), right=lado("AAAAAA"),
        top=lado("CCCCCC"), bottom=lado("CCCCCC")
    )

def limpiar_lote(val) -> str:
    s = str(val).strip() if val else ""
    if not s or s.upper() in ("NONE", "NAN", "LT", "LOTE", ""):
        return ""
    try:
        return str(int(float(s)))
    except:
        return s.upper()

def col_map(df: pd.DataFrame, aliases: list) -> str | None:
    cols_lower = {c.lower().strip(): c for c in df.columns if c}
    for a in aliases:
        if a in cols_lower:
            return cols_lower[a]
    return None

# ========================CARGA USUARIOS=====================
def cargar_usuarios_id() -> dict:
    """
    Devuelve dict: {(MZ_upper, LOTE_str): (USER_ID, NOMBRE)}
    Soporta columnas MZ/LOTE y MZ2/LOTE2 para usuarios con 2 lotes.
    """
    if not USUARIOS_ID_PATH.exists():
        raise FileNotFoundError(f"No se encontró {USUARIOS_ID_PATH}")

    df = pd.read_excel(USUARIOS_ID_PATH, dtype=str).fillna("")
    cols = {c.upper().strip(): c for c in df.columns}

    def get(row, key):
        return str(row[cols[key]]).strip() if key in cols else ""

    mapa = {}
    for _, row in df.iterrows():
        uid    = get(row, "USER_ID")
        nombre = get(row, "NOMBRE")
        mz     = get(row, "MZ").upper()
        lt     = limpiar_lote(get(row, "LOTE"))
        if uid and mz and lt:
            mapa[(mz, lt)] = (uid, nombre)

        mz2 = get(row, "MZ2").upper()
        lt2 = limpiar_lote(get(row, "LOTE2"))
        if uid and mz2 and lt2:
            mapa[(mz2, lt2)] = (uid, nombre)

    print(f"  ✔ {len(mapa)} pares MZ-LOTE cargados desde usuarios_id.xlsx")
    return mapa

# ========================CARGA REPORTES=====================
def cargar_reportes() -> tuple[list, dict]:
    """
    Retorna:
        registros      — lista de dicts {origen, mz, lote, archivo}
        no_encontrados — dict {(mz, lote): {origenes, archivos}}
    """
    archivos = sorted(REPORTE_ACUM_DIR.glob("*.xlsx"))
    if not archivos:
        raise FileNotFoundError(f"No hay .xlsx en {REPORTE_ACUM_DIR}")

    usuarios_id    = cargar_usuarios_id()
    registros      = []
    no_encontrados = defaultdict(lambda: {"origenes": set(), "archivos": set()})

    for archivo in archivos:
        nombre = archivo.name
        print(f"\n  Leyendo: {nombre}")

        try:
            hojas = pd.read_excel(archivo, sheet_name=None, dtype=str)
        except Exception as e:
            print(f"  ⚠ No se pudo abrir: {e}")
            continue

        hoja_key = next((k for k in hojas if k.lower() == HOJA_REPORTE), None)
        if not hoja_key:
            print(f"  ⚠ Saltado — sin hoja '{HOJA_REPORTE}'")
            continue

        df = hojas[hoja_key].fillna("")
        if df.empty:
            print(f"  ⚠ Saltado — hoja vacía")
            continue

        col_tipo   = col_map(df, ALIAS_TIPO)
        col_origen = col_map(df, ALIAS_ORIGEN)
        col_mz     = col_map(df, ALIAS_MZ)
        col_lt     = col_map(df, ALIAS_LT)

        if not all([col_tipo, col_origen, col_mz, col_lt]):
            faltantes = [a for a, c in zip(
                ["tipo", "origen", "mz", "lt"],
                [col_tipo, col_origen, col_mz, col_lt]
            ) if not c]
            print(f"  ⚠ Saltado — columnas faltantes: {faltantes}")
            continue

        df = df[df[col_tipo].str.strip().str.upper() == TIPO_PAGO.upper()]
        if df.empty:
            print(f"  ⚠ Saltado — sin filas '{TIPO_PAGO}'")
            continue

        print(f"  ✔ {len(df)} filas TE PAGÓ")

        for _, fila in df.iterrows():
            mz     = str(fila[col_mz]).strip().upper()
            lote   = limpiar_lote(fila[col_lt])
            origen = str(fila[col_origen]).strip()

            if not mz or mz in ("NONE", "NAN", "") or mz.startswith("="):
                continue
            if not lote:
                continue
            if not origen or origen in ("NONE", "NAN", ""):
                continue

            clave = (mz, lote)
            if clave in usuarios_id:
                registros.append({
                    "origen":  origen,
                    "mz":      mz,
                    "lote":    lote,
                    "archivo": nombre,
                })
            else:
                no_encontrados[clave]["origenes"].add(origen)
                no_encontrados[clave]["archivos"].add(nombre)

    print(f"\n  Total registros válidos  : {len(registros)}")
    print(f"  MZ-LOTE no encontrados   : {len(no_encontrados)}")
    return registros, no_encontrados, usuarios_id

# ========================PROCESAMIENTO======================
def construir_origenes(registros: list, usuarios_id: dict) -> list:
    grupos = defaultdict(lambda: {"origenes": [], "archivos": set()})

    for r in registros:
        clave = (r["mz"], r["lote"])
        if r["origen"] not in grupos[clave]["origenes"]:
            grupos[clave]["origenes"].append(r["origen"])
        grupos[clave]["archivos"].add(r["archivo"])

    resultado = []
    for (mz, lote), data in sorted(grupos.items()):
        uid, nombre = usuarios_id.get((mz, lote), ("", ""))
        resultado.append({
            "user_id":            uid,
            "nombre":             nombre,
            "mz":                 mz,
            "lote":               lote,
            "origenes":           data["origenes"],
            "total_apariciones":  len([r for r in registros
                                       if r["mz"] == mz and r["lote"] == lote]),
            "meses_en_registros": len(data["archivos"]),
            "fecha_registro":     datetime.today().strftime("%d/%m/%Y"),
        })

    print(f"  MZ-LOTE en origenes_yape : {len(resultado)}")
    return resultado

# ========================EXPORTAR ORIGENES==================
def exportar_origenes_yape(datos: list):
    path = OUTPUT_DIR / "origenes_yape.xlsx"

    max_origenes = max((len(d["origenes"]) for d in datos), default=1)

    # Definicion de columnas y grupos
    cols_fijas = ["USER_ID", "NOMBRE", "MZ", "LOTE"]
    cols_ori   = [f"ORIGEN_{i+1}" for i in range(max_origenes)]
    cols_met   = ["TOTAL_APARICIONES", "MESES_EN_REGISTROS"]
    cols_ctl   = ["FECHA_REGISTRO"]
    todas      = cols_fijas + cols_ori + cols_met + cols_ctl

    # Grupos para cabecera doble
    grupos = [
        ("¿QUIÉN ES?",                    2,  COLOR_ID_H,  COLOR_ID_TXT),
        (None,                             0,  None,        None),   # sep
        ("¿DÓNDE VIVE?",                  2,  COLOR_UBI_H, COLOR_UBI_TXT),
        (None,                             0,  None,        None),   # sep
        ("¿QUIÉN HA PAGADO DESDE ESTE LOTE?", max_origenes, COLOR_ORI_H, COLOR_ORI_TXT),
        (None,                             0,  None,        None),   # sep
        ("¿CUÁNTO LO HEMOS VISTO?",       2,  COLOR_MET_H, COLOR_MET_TXT),
        (None,                             0,  None,        None),   # sep
        ("¿CUÁNDO SE REGISTRÓ?",          1,  COLOR_CTL_H, COLOR_CTL_TXT),
    ]

    # Mapeo col → grupo para colorear celdas
    col_grupo = {}
    for c in ["USER_ID", "NOMBRE"]:
        col_grupo[c] = ("id",  COLOR_ID_C,  COLOR_ID_TXT,  True)
    for c in ["MZ", "LOTE"]:
        col_grupo[c] = ("ubi", COLOR_UBI_C, COLOR_UBI_TXT, True)
    for c in cols_ori:
        col_grupo[c] = ("ori", COLOR_ORI_C, COLOR_ORI_TXT, True)
    for c in cols_met:
        col_grupo[c] = ("met", COLOR_MET_C, COLOR_MET_TXT, True)
    for c in cols_ctl:
        col_grupo[c] = ("ctl", COLOR_CTL_C, COLOR_CTL_TXT, False)

    wb = Workbook()
    ws = wb.active
    ws.title = "origenes_yape"
    ws.freeze_panes = "A3"

    # Construir lista de columnas con separadores
    cols_con_sep = []
    for col in todas:
        if col == "MZ":
            cols_con_sep.append("__SEP__")
        elif col == cols_ori[0]:
            cols_con_sep.append("__SEP__")
        elif col == "TOTAL_APARICIONES":
            cols_con_sep.append("__SEP__")
        elif col == "FECHA_REGISTRO":
            cols_con_sep.append("__SEP__")
        cols_con_sep.append(col)

    n_cols = len(cols_con_sep)

    # ---- Fila 1: grupos ----
    col_idx = 1
    for nombre_g, span, bg_h, txt_h in grupos:
        if nombre_g is None:
            # separador
            c = ws.cell(row=1, column=col_idx, value="")
            estilo_sep(c)
            ws.column_dimensions[get_column_letter(col_idx)].width = 1
            col_idx += 1
        else:
            if span == 1:
                c = ws.cell(row=1, column=col_idx, value=nombre_g)
                estilo_header(c, bg_h, txt_h, bold=True, size=9)
            else:
                ws.merge_cells(
                    start_row=1, start_column=col_idx,
                    end_row=1,   end_column=col_idx + span - 1
                )
                c = ws.cell(row=1, column=col_idx, value=nombre_g)
                estilo_header(c, bg_h, txt_h, bold=True, size=9)
            col_idx += span

    ws.row_dimensions[1].height = 20

    # ---- Fila 2: nombres de columnas ----
    for ci, col in enumerate(cols_con_sep, start=1):
        if col == "__SEP__":
            c = ws.cell(row=2, column=ci, value="")
            estilo_sep(c)
        else:
            _, bg, txt, mono = col_grupo.get(col, ("", "FFFFFF", "000000", False))
            c = ws.cell(row=2, column=ci, value=col)
            estilo_header(c, bg, txt, bold=True, size=9)

    ws.row_dimensions[2].height = 20

    # ---- Filas de datos ----
    for ri, dato in enumerate(datos, start=3):
        valores = {
            "USER_ID":             dato["user_id"],
            "NOMBRE":              dato["nombre"],
            "MZ":                  dato["mz"],
            "LOTE":                dato["lote"],
            "TOTAL_APARICIONES":   dato["total_apariciones"],
            "MESES_EN_REGISTROS":  dato["meses_en_registros"],
            "FECHA_REGISTRO":      dato["fecha_registro"],
        }
        for i, ori in enumerate(dato["origenes"]):
            valores[f"ORIGEN_{i+1}"] = ori

        for ci, col in enumerate(cols_con_sep, start=1):
            if col == "__SEP__":
                c = ws.cell(row=ri, column=ci, value="")
                estilo_sep(c)
            else:
                _, bg, txt, mono = col_grupo.get(col, ("", "FFFFFF", "000000", False))
                val = valores.get(col, "")
                c   = ws.cell(row=ri, column=ci, value=val if val else "")
                align = "center" if col in cols_met else "left"
                estilo(c, bg=bg, txt=txt, align=align, mono=mono)

        ws.row_dimensions[ri].height = 18

    # ---- Anchos ----
    anchos = {
        "USER_ID": 10, "NOMBRE": 28, "MZ": 6, "LOTE": 7,
        "TOTAL_APARICIONES": 12, "MESES_EN_REGISTROS": 14,
        "FECHA_REGISTRO": 14,
    }
    for ci, col in enumerate(cols_con_sep, start=1):
        if col != "__SEP__":
            w = anchos.get(col, 24)
            ws.column_dimensions[get_column_letter(ci)].width = w

    wb.save(path)
    print(f"\n  ✔ origenes_yape.xlsx guardado: {path}")
    print(f"  ✔ {len(datos)} filas  |  {max_origenes} origen(es) máx")

# ========================EXPORTAR NO ENCONTRADOS============
def exportar_mz_no_encontradas(no_encontrados: dict):
    if not no_encontrados:
        print("  ✔ Sin MZ-LOTE no encontrados — mz_no_encontradas.xlsx no generado")
        return

    path = OUTPUT_DIR / "mz_no_encontradas.xlsx"

    cols = ["ORIGEN", "__SEP__", "MZ", "LOTE", "__SEP__",
            "TOTAL_APARICIONES", "MESES_EN_REGISTROS", "__SEP__", "FECHA_REGISTRO"]

    col_grupo = {
        "ORIGEN":              (COLOR_ID_C,  "991B1B", True),
        "MZ":                  (COLOR_UBI_C, COLOR_UBI_TXT, True),
        "LOTE":                (COLOR_UBI_C, COLOR_UBI_TXT, True),
        "TOTAL_APARICIONES":   (COLOR_MET_C, COLOR_MET_TXT, True),
        "MESES_EN_REGISTROS":  (COLOR_MET_C, COLOR_MET_TXT, True),
        "FECHA_REGISTRO":      (COLOR_CTL_C, COLOR_CTL_TXT, False),
    }

    grupos_h = [
        ("¿QUÉ VIO EL BANCO?",       1, "FEF2F2", "991B1B"),
        (None, 0, None, None),
        ("¿DÓNDE DIJO QUE VIVÍA?",   2, COLOR_UBI_H, COLOR_UBI_TXT),
        (None, 0, None, None),
        ("¿CUÁNTAS VECES LO VIMOS?", 2, COLOR_MET_H, COLOR_MET_TXT),
        (None, 0, None, None),
        ("¿CUÁNDO SE REGISTRÓ?",     1, COLOR_CTL_H, COLOR_CTL_TXT),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "mz_no_encontradas"
    ws.freeze_panes = "A3"

    # Fila 1: grupos
    col_idx = 1
    for nombre_g, span, bg_h, txt_h in grupos_h:
        if nombre_g is None:
            c = ws.cell(row=1, column=col_idx, value="")
            estilo_sep(c)
            ws.column_dimensions[get_column_letter(col_idx)].width = 1
            col_idx += 1
        else:
            if span == 1:
                c = ws.cell(row=1, column=col_idx, value=nombre_g)
                estilo_header(c, bg_h, txt_h, bold=True, size=9)
            else:
                ws.merge_cells(
                    start_row=1, start_column=col_idx,
                    end_row=1,   end_column=col_idx + span - 1
                )
                c = ws.cell(row=1, column=col_idx, value=nombre_g)
                estilo_header(c, bg_h, txt_h, bold=True, size=9)
                for i in range(1, span):
                    cc = ws.cell(row=1, column=col_idx + i)
                    estilo_header(cc, bg_h, txt_h)
            col_idx += span
    ws.row_dimensions[1].height = 20

    # Fila 2: nombres
    for ci, col in enumerate(cols, start=1):
        if col == "__SEP__":
            c = ws.cell(row=2, column=ci, value="")
            estilo_sep(c)
            ws.column_dimensions[get_column_letter(ci)].width = 1
        else:
            bg, txt, _ = col_grupo[col]
            c = ws.cell(row=2, column=ci, value=col)
            estilo_header(c, bg, txt, bold=True, size=9)
    ws.row_dimensions[2].height = 20

    # Filas de datos
    fecha = datetime.today().strftime("%d/%m/%Y")
    for ri, ((mz, lote), data) in enumerate(sorted(no_encontrados.items()), start=3):
        origenes_str = " · ".join(sorted(data["origenes"]))
        valores = {
            "ORIGEN":             origenes_str,
            "MZ":                 mz,
            "LOTE":               lote,
            "TOTAL_APARICIONES":  len(data["origenes"]),
            "MESES_EN_REGISTROS": len(data["archivos"]),
            "FECHA_REGISTRO":     fecha,
        }
        for ci, col in enumerate(cols, start=1):
            if col == "__SEP__":
                c = ws.cell(row=ri, column=ci, value="")
                estilo_sep(c)
            else:
                bg, txt, mono = col_grupo[col]
                val   = valores.get(col, "")
                align = "center" if col in ("TOTAL_APARICIONES", "MESES_EN_REGISTROS") else "left"
                c     = ws.cell(row=ri, column=ci, value=val)
                estilo(c, bg=bg, txt=txt, align=align, mono=mono)
        ws.row_dimensions[ri].height = 18

    # Anchos
    anchos = {
        "ORIGEN": 30, "MZ": 6, "LOTE": 7,
        "TOTAL_APARICIONES": 12, "MESES_EN_REGISTROS": 14, "FECHA_REGISTRO": 14,
    }
    for ci, col in enumerate(cols, start=1):
        if col != "__SEP__":
            ws.column_dimensions[get_column_letter(ci)].width = anchos.get(col, 20)

    wb.save(path)
    print(f"  ✔ mz_no_encontradas.xlsx guardado: {path}")
    print(f"  ✔ {len(no_encontrados)} MZ-LOTE no encontrados")

# ======================MAIN=================================
def main():
    print("=" * 55)
    print("  CREAR ORIGENES — acumulador de orígenes por MZ-LOTE")
    print("=" * 55)

    print("\n[1] Preparando carpeta de salida...")
    reset_output_folder(OUTPUT_DIR)

    print("\n[2] Cargando reportes y usuarios...")
    registros, no_encontrados, usuarios_id = cargar_reportes()

    print("\n[3] Construyendo origenes_yape...")
    datos = construir_origenes(registros, usuarios_id)

    print("\n[4] Exportando origenes_yape.xlsx...")
    exportar_origenes_yape(datos)

    print("\n[5] Exportando mz_no_encontradas.xlsx...")
    exportar_mz_no_encontradas(no_encontrados)

    print("\n" + "=" * 55)
    print("  Proceso completado exitosamente")
    print("=" * 55)

if __name__ == "__main__":
    main()
