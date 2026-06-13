# =========================IMPORTS===========================
import sys
import shutil
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# utils compartidos
sys.path.insert(0, str(Path(__file__).parent.parent))
from utils.consolidar import filtrar_origenes_validos

# ========================CONFIGURACION======================
BASE_DIR         = Path(__file__).parent
ORIGENES_PATH    = BASE_DIR.parent / "crear_origenes" / "outputs" / "origenes_yape.xlsx"
OUTPUT_DIR          = BASE_DIR / "outputs"
MAESTRO_PATH        = OUTPUT_DIR / "maestro_yape.xlsx"
SIN_CONFIRMAR_PATH  = OUTPUT_DIR / "sin_confirmar.xlsx"

# umbrales configurables en utils/consolidar.py

# ========================COLORES============================
COLOR_HEADER  = "4A235A"
COLOR_ID_H    = "F4ECF7"
COLOR_ID_C    = "FAF5FF"
COLOR_ID_TXT  = "5B21B6"
COLOR_UBI_H   = "E1F5EE"
COLOR_UBI_C   = "F0FFF8"
COLOR_UBI_TXT = "085041"
COLOR_ORI_H   = "FEF3E2"
COLOR_ORI_C   = "FFFBF0"
COLOR_ORI_TXT = "92400E"
COLOR_CON_H   = "FEF2F2"
COLOR_CON_C   = "FEF2F2"
COLOR_CON_TXT = "991B1B"
COLOR_CTL_H   = "F9FAFB"
COLOR_CTL_C   = "F9FAFB"
COLOR_CTL_TXT = "374151"
COLOR_SEP     = "F3F4F6"
COLOR_MET_H   = "EFF6FF"
COLOR_MET_C   = "EFF6FF"
COLOR_MET_TXT = "1D4ED8"
COLOR_DEC_H   = "FEF2F2"
COLOR_DEC_C   = "FEF2F2"
COLOR_DEC_TXT = "991B1B"
COLOR_SI      = "D1FAE5"
TXT_SI        = "065F46"



# ========================UTILIDADES=========================
def reset_output_folder(path: Path):
    if path.exists():
        shutil.rmtree(path)
    path.mkdir()

def lado(color="CCCCCC"):
    return Side(style="thin", color=color)

def estilo(cell, bold=False, bg=None, txt=None, align="left", mono=False, size=10):
    b = lado()
    cell.border    = Border(left=b, right=b, top=b, bottom=b)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    font_name      = "Courier New" if mono else "Arial"
    cell.font      = Font(name=font_name, bold=bold, size=size,
                          color=txt if txt else "000000")
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)

def estilo_header(cell, bg, txt, bold=True, size=9):
    b = lado("FFFFFF")
    cell.border    = Border(left=b, right=b, top=b, bottom=b)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font      = Font(name="Arial", bold=bold, size=size, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)

def estilo_sep(cell):
    cell.fill   = PatternFill("solid", start_color=COLOR_SEP)
    cell.border = Border(
        left=lado("CCCCCC"), right=lado("AAAAAA"),
        top=lado("CCCCCC"),  bottom=lado("CCCCCC")
    )

# ========================CARGA ORIGENES=====================
def cargar_origenes() -> list:
    if not ORIGENES_PATH.exists():
        raise FileNotFoundError(f"No se encontró {ORIGENES_PATH}")

    wb   = load_workbook(ORIGENES_PATH, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.values)
    wb.close()

    # Fila 1 es grupos, fila 2 es cabeceras reales
    if len(rows) < 3:
        raise ValueError("origenes_yape.xlsx vacío o sin datos")

    headers = [str(h).strip() if h else "" for h in rows[1]]

    datos = []
    for fila in rows[2:]:
        if not fila or not any(fila):
            continue
        row = dict(zip(headers, fila))

        uid    = str(row.get("USER_ID", "")).strip()
        nombre = str(row.get("NOMBRE",  "")).strip()
        mz     = str(row.get("MZ",      "")).strip().upper()
        lote   = str(row.get("LOTE",    "")).strip()

        if not uid or not mz or not lote:
            continue

        origenes = []
        i = 1
        while True:
            key = f"ORIGEN_{i}"
            if key not in row:
                break
            val = str(row[key]).strip() if row[key] else ""
            if val and val.upper() not in ("NONE", "NAN", ""):
                origenes.append(val)
            i += 1

        try:
            total_ap = int(float(str(row.get("TOTAL_APARICIONES",  0))))
            meses    = int(float(str(row.get("MESES_EN_REGISTROS", 0))))
        except:
            total_ap = 0
            meses    = 0

        datos.append({
            "user_id":            uid,
            "nombre":             nombre,
            "mz":                 mz,
            "lote":               lote,
            "origenes":           origenes,
            "total_apariciones":  total_ap,
            "meses_en_registros": meses,
        })

    print(f"  ✔ {len(datos)} filas cargadas desde origenes_yape.xlsx")
    return datos

# ========================VALIDACIONES MANUALES==============
def cargar_validados_manuales() -> dict:
    """
    Lee el maestro anterior y extrae los USER_ID+MZ+LOTE con VALIDADO_MANUAL = si.
    Retorna dict: {(user_id, mz, lote): True}
    """
    if not MAESTRO_PATH.exists():
        return {}

    wb   = load_workbook(MAESTRO_PATH, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.values)
    wb.close()

    if len(rows) < 3:
        return {}

    headers = [str(h).strip() if h else "" for h in rows[1]]
    manuales = {}

    for fila in rows[2:]:
        if not fila:
            continue
        row = dict(zip(headers, fila))
        vm  = str(row.get("VALIDADO_MANUAL", "")).strip().lower()
        if vm == "si":
            uid  = str(row.get("USER_ID", "")).strip()
            mz   = str(row.get("MZ",      "")).strip().upper()
            lote = str(row.get("LOTE",    "")).strip()
            if uid and mz and lote:
                manuales[(uid, mz, lote)] = True

    print(f"  ✔ {len(manuales)} validaciones manuales preservadas")
    return manuales

# ========================SIN CONFIRMAR ANTERIOR=============
def cargar_sin_confirmar_anterior() -> dict:
    """
    Lee sin_confirmar.xlsx anterior y extrae los autorizados.
    Retorna dict: {(user_id, mz, lote): origen_correcto}
    """
    if not SIN_CONFIRMAR_PATH.exists():
        return {}

    wb   = load_workbook(SIN_CONFIRMAR_PATH, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.values)
    wb.close()

    if len(rows) < 3:
        return {}

    headers   = [str(h).strip() if h else "" for h in rows[1]]
    autorizados = {}

    for fila in rows[2:]:
        if not fila:
            continue
        row  = dict(zip(headers, fila))
        auth = str(row.get("AUTORIZADO", "")).strip().lower()
        if auth == "si":
            uid    = str(row.get("USER_ID",          "")).strip()
            mz     = str(row.get("MZ",               "")).strip().upper()
            lote   = str(row.get("LOTE",             "")).strip()
            origen = str(row.get("ORIGEN_CORRECTO",  "")).strip()
            if uid and mz and lote and origen:
                autorizados[(uid, mz, lote)] = origen

    print(f"  ✔ {len(autorizados)} autorizados desde sin_confirmar anterior")
    return autorizados

# ========================CONSOLIDACION======================


def consolidar(datos: list, manuales: dict, autorizados: dict) -> tuple[list, list]:
    """
    Retorna (maestro, sin_confirmar)
    maestro       — tiene orígenes válidos · producción
    sin_confirmar — sin orígenes válidos · auditoría
    """
    maestro       = []
    sin_confirmar = []
    fecha         = datetime.today().strftime("%d/%m/%Y")

    for d in datos:
        clave = (d["user_id"], d["mz"], d["lote"])

        # Autorizado desde sin_confirmar anterior → sube al maestro
        if clave in autorizados:
            maestro.append({
                "user_id":          d["user_id"],
                "nombre":           d["nombre"],
                "mz":               d["mz"],
                "lote":             d["lote"],
                "origenes_validos": [autorizados[clave]],
                "validado_manual":  "si",
                "fecha_registro":   fecha,
            })
            continue

        # Filtrar orígenes válidos
        origenes_validos = filtrar_origenes_validos(d["origenes"], d["nombre"])

        # Preservar validación manual desde maestro anterior
        validado = "si" if clave in manuales else "no"

        if not origenes_validos:
            sin_confirmar.append({
                "user_id":            d["user_id"],
                "nombre":             d["nombre"],
                "mz":                 d["mz"],
                "lote":               d["lote"],
                "origenes":           d["origenes"],
                "total_apariciones":  d["total_apariciones"],
                "meses_en_registros": d["meses_en_registros"],
                "fecha_registro":     fecha,
            })
        else:
            maestro.append({
                "user_id":          d["user_id"],
                "nombre":           d["nombre"],
                "mz":               d["mz"],
                "lote":             d["lote"],
                "origenes_validos": origenes_validos,
                "validado_manual":  validado,
                "fecha_registro":   fecha,
            })

    print(f"  ✔ maestro_yape    : {len(maestro)} filas")
    print(f"  ✔ sin_confirmar   : {len(sin_confirmar)} filas pendientes")
    return maestro, sin_confirmar

# ========================EXPORTAR===========================
def exportar_maestro(maestro: list):
    max_origenes = max((len(m["origenes_validos"]) for m in maestro), default=1)

    cols_id  = ["USER_ID", "NOMBRE"]
    cols_ubi = ["MZ", "LOTE"]
    cols_ori = [f"ORIGEN_{i+1}" for i in range(max_origenes)]
    cols_con = ["VALIDADO_MANUAL"]
    cols_ctl = ["FECHA_REGISTRO"]

    # Columnas con separadores
    cols_con_sep = []
    for col in cols_id + cols_ubi + cols_ori + cols_con + cols_ctl:
        if col == "MZ":
            cols_con_sep.append("__SEP__")
        elif col == cols_ori[0]:
            cols_con_sep.append("__SEP__")
        elif col == "FECHA_REGISTRO":
            cols_con_sep.append("__SEP__")
        cols_con_sep.append(col)

    # Grupos fila 1
    grupos = [
        ("¿QUIÉN ES?",                    2,             COLOR_ID_H,  COLOR_ID_TXT),
        (None, 0, None, None),
        ("¿DÓNDE VIVE?",                  2,             COLOR_UBI_H, COLOR_UBI_TXT),
        (None, 0, None, None),
        ("¿CUÁLES SON SUS ORÍGENES VÁLIDOS?", max_origenes, COLOR_ORI_H, COLOR_ORI_TXT),
        (None, 0, None, None),
        ("¿VALIDADO MANUAL?",              1,             COLOR_CON_H, COLOR_CON_TXT),
        (None, 0, None, None),
        ("¿CUÁNDO SE CONSOLIDÓ?",         1,             COLOR_CTL_H, COLOR_CTL_TXT),
    ]

    # Mapeo col → estilo
    col_estilo = {}
    for c in cols_id:
        col_estilo[c] = (COLOR_ID_C,  COLOR_ID_TXT,  True)
    for c in cols_ubi:
        col_estilo[c] = (COLOR_UBI_C, COLOR_UBI_TXT, True)
    for c in cols_ori:
        col_estilo[c] = (COLOR_ORI_C, COLOR_ORI_TXT, True)
    for c in cols_con:
        col_estilo[c] = (COLOR_CON_C, COLOR_CON_TXT, True)
    for c in cols_ctl:
        col_estilo[c] = (COLOR_CTL_C, COLOR_CTL_TXT, False)

    wb = Workbook()
    ws = wb.active
    ws.title = "maestro_yape"
    ws.freeze_panes = "A3"

    # ---- Fila 1: grupos ----
    col_idx = 1
    for nombre_g, span, bg_h, txt_h in grupos:
        if nombre_g is None:
            c = ws.cell(row=1, column=col_idx, value="")
            estilo_sep(c)
            ws.column_dimensions[get_column_letter(col_idx)].width = 1
            col_idx += 1
        else:
            if span == 1:
                c = ws.cell(row=1, column=col_idx, value=nombre_g)
                estilo_header(c, bg_h, txt_h, bold=True, size=9)
            else:
                ws.merge_cells(
                    start_row=1, start_column=col_idx,
                    end_row=1,   end_column=col_idx + span - 1
                )
                c = ws.cell(row=1, column=col_idx, value=nombre_g)
                estilo_header(c, bg_h, txt_h, bold=True, size=9)
            col_idx += span
    ws.row_dimensions[1].height = 20

    # ---- Fila 2: nombres ----
    for ci, col in enumerate(cols_con_sep, start=1):
        if col == "__SEP__":
            c = ws.cell(row=2, column=ci, value="")
            estilo_sep(c)
            ws.column_dimensions[get_column_letter(ci)].width = 1
        else:
            bg, txt, mono = col_estilo.get(col, ("FFFFFF", "000000", False))
            c = ws.cell(row=2, column=ci, value=col)
            estilo_header(c, bg, txt, bold=True, size=9)
    ws.row_dimensions[2].height = 20

    # ---- Filas de datos ----
    for ri, m in enumerate(maestro, start=3):
        valores = {
            "USER_ID":         m["user_id"],
            "NOMBRE":          m["nombre"],
            "MZ":              m["mz"],
            "LOTE":            m["lote"],
            "VALIDADO_MANUAL": m["validado_manual"],
            "FECHA_REGISTRO":  m["fecha_registro"],
        }
        for i, ori in enumerate(m["origenes_validos"]):
            valores[f"ORIGEN_{i+1}"] = ori

        for ci, col in enumerate(cols_con_sep, start=1):
            if col == "__SEP__":
                c = ws.cell(row=ri, column=ci, value="")
                estilo_sep(c)
                continue

            val = valores.get(col, "")
            bg, txt, mono = col_estilo.get(col, ("FFFFFF", "000000", False))

            align = "center" if col in ("MZ", "LOTE", "VALIDADO_MANUAL") else "left"
            c = ws.cell(row=ri, column=ci, value=val if val else "")
            estilo(c, bg=bg, txt=txt, align=align, mono=mono)

        ws.row_dimensions[ri].height = 18

    # ---- Anchos ----
    anchos = {
        "USER_ID": 10, "NOMBRE": 28, "MZ": 6, "LOTE": 7,
        "VALIDADO_MANUAL": 14, "FECHA_REGISTRO": 14,
    }
    for ci, col in enumerate(cols_con_sep, start=1):
        if col != "__SEP__":
            ws.column_dimensions[get_column_letter(ci)].width = anchos.get(col, 24)

    wb.save(MAESTRO_PATH)
    print(f"\n  ✔ maestro_yape.xlsx guardado: {MAESTRO_PATH}")
    print(f"  ✔ {len(maestro)} filas  |  {max_origenes} origen(es) válido(s) máx")

# ========================EXPORTAR SIN CONFIRMAR=============
def exportar_sin_confirmar(sin_confirmar: list):
    if not sin_confirmar:
        print("  ✔ Sin pendientes — sin_confirmar.xlsx no generado")
        return

    path = SIN_CONFIRMAR_PATH
    max_origenes = max((len(d["origenes"]) for d in sin_confirmar), default=1)

    cols_id  = ["USER_ID", "NOMBRE"]
    cols_ubi = ["MZ", "LOTE"]
    cols_ori = [f"ORIGEN_{i+1}" for i in range(max_origenes)]
    cols_met = ["TOTAL_APARICIONES", "MESES_EN_REGISTROS"]
    cols_dec = ["ORIGEN_CORRECTO", "AUTORIZADO"]
    cols_ctl = ["FECHA_REGISTRO"]

    cols_con_sep = []
    for col in cols_id + cols_ubi + cols_ori + cols_met + cols_dec + cols_ctl:
        if col in ("MZ", cols_ori[0], "TOTAL_APARICIONES", "ORIGEN_CORRECTO", "FECHA_REGISTRO"):
            cols_con_sep.append("__SEP__")
        cols_con_sep.append(col)

    grupos = [
        ("¿QUIÉN ES?",              2,             COLOR_ID_H,  COLOR_ID_TXT),
        (None, 0, None, None),
        ("¿DÓNDE VIVE?",            2,             COLOR_UBI_H, COLOR_UBI_TXT),
        (None, 0, None, None),
        ("¿QUÉ ORÍGENES VIMOS?",    max_origenes,  COLOR_ORI_H, COLOR_ORI_TXT),
        (None, 0, None, None),
        ("¿CUÁNTO LO HEMOS VISTO?", 2,             COLOR_MET_H, COLOR_MET_TXT),
        (None, 0, None, None),
        ("¿QUÉ DECIDES?",           2,             COLOR_DEC_H, COLOR_DEC_TXT),
        (None, 0, None, None),
        ("¿CUÁNDO SE REGISTRÓ?",    1,             COLOR_CTL_H, COLOR_CTL_TXT),
    ]

    col_estilo = {}
    for c in cols_id:  col_estilo[c] = (COLOR_ID_C,  COLOR_ID_TXT,  True)
    for c in cols_ubi: col_estilo[c] = (COLOR_UBI_C, COLOR_UBI_TXT, True)
    for c in cols_ori: col_estilo[c] = (COLOR_ORI_C, COLOR_ORI_TXT, True)
    for c in cols_met: col_estilo[c] = (COLOR_MET_C, COLOR_MET_TXT, True)
    for c in cols_dec: col_estilo[c] = (COLOR_DEC_C, COLOR_DEC_TXT, True)
    for c in cols_ctl: col_estilo[c] = (COLOR_CTL_C, COLOR_CTL_TXT, False)

    wb = Workbook()
    ws = wb.active
    ws.title = "sin_confirmar"
    ws.freeze_panes = "A3"

    # Fila 1 grupos
    col_idx = 1
    for nombre_g, span, bg_h, txt_h in grupos:
        if nombre_g is None:
            c = ws.cell(row=1, column=col_idx, value="")
            estilo_sep(c)
            ws.column_dimensions[get_column_letter(col_idx)].width = 1
            col_idx += 1
        else:
            if span == 1:
                c = ws.cell(row=1, column=col_idx, value=nombre_g)
                estilo_header(c, bg_h, txt_h, bold=True, size=9)
            else:
                ws.merge_cells(start_row=1, start_column=col_idx,
                               end_row=1,   end_column=col_idx + span - 1)
                c = ws.cell(row=1, column=col_idx, value=nombre_g)
                estilo_header(c, bg_h, txt_h, bold=True, size=9)
            col_idx += span
    ws.row_dimensions[1].height = 20

    # Fila 2 nombres
    for ci, col in enumerate(cols_con_sep, start=1):
        if col == "__SEP__":
            c = ws.cell(row=2, column=ci, value="")
            estilo_sep(c)
            ws.column_dimensions[get_column_letter(ci)].width = 1
        else:
            bg, txt, mono = col_estilo.get(col, ("FFFFFF", "000000", False))
            c = ws.cell(row=2, column=ci, value=col)
            estilo_header(c, bg, txt, bold=True, size=9)
    ws.row_dimensions[2].height = 20

    # Filas datos
    for ri, d in enumerate(sin_confirmar, start=3):
        valores = {
            "USER_ID":            d["user_id"],
            "NOMBRE":             d["nombre"],
            "MZ":                 d["mz"],
            "LOTE":               d["lote"],
            "TOTAL_APARICIONES":  d["total_apariciones"],
            "MESES_EN_REGISTROS": d["meses_en_registros"],
            "ORIGEN_CORRECTO":    "",
            "AUTORIZADO":         "no",
            "FECHA_REGISTRO":     d["fecha_registro"],
        }
        for i, ori in enumerate(d["origenes"]):
            valores[f"ORIGEN_{i+1}"] = ori

        for ci, col in enumerate(cols_con_sep, start=1):
            if col == "__SEP__":
                c = ws.cell(row=ri, column=ci, value="")
                estilo_sep(c)
                continue
            val              = valores.get(col, "")
            bg, txt, mono    = col_estilo.get(col, ("FFFFFF", "000000", False))
            if col == "AUTORIZADO":
                bg  = COLOR_SI if str(val).lower() == "si" else COLOR_DEC_C
                txt = TXT_SI   if str(val).lower() == "si" else COLOR_DEC_TXT
            align = "center" if col in ("MZ", "LOTE", "AUTORIZADO",
                                        "TOTAL_APARICIONES", "MESES_EN_REGISTROS") else "left"
            c = ws.cell(row=ri, column=ci, value=val if val else "")
            estilo(c, bg=bg, txt=txt, align=align, mono=mono)
        ws.row_dimensions[ri].height = 18

    anchos = {
        "USER_ID": 10, "NOMBRE": 28, "MZ": 6, "LOTE": 7,
        "TOTAL_APARICIONES": 12, "MESES_EN_REGISTROS": 14,
        "ORIGEN_CORRECTO": 26, "AUTORIZADO": 12, "FECHA_REGISTRO": 14,
    }
    for ci, col in enumerate(cols_con_sep, start=1):
        if col != "__SEP__":
            ws.column_dimensions[get_column_letter(ci)].width = anchos.get(col, 24)

    wb.save(path)
    print(f"  ✔ sin_confirmar.xlsx guardado: {path}")
    print(f"  ✔ {len(sin_confirmar)} pendientes · completa ORIGEN_CORRECTO y pon AUTORIZADO = si")

# ======================MAIN=================================
def main():
    print("=" * 55)
    print("  CREAR MAESTRO — consolidador de orígenes Yape")
    print("=" * 55)

    print("\n[1] Leyendo validaciones manuales y autorizados anteriores...")
    manuales    = cargar_validados_manuales()
    autorizados = cargar_sin_confirmar_anterior()

    print("\n[2] Preparando carpeta de salida...")
    reset_output_folder(OUTPUT_DIR)

    print("\n[3] Cargando origenes_yape.xlsx...")
    datos = cargar_origenes()

    print("\n[4] Consolidando orígenes...")
    maestro, sin_confirmar = consolidar(datos, manuales, autorizados)

    print("\n[5] Exportando maestro_yape.xlsx...")
    exportar_maestro(maestro)

    print("\n[6] Exportando sin_confirmar.xlsx...")
    exportar_sin_confirmar(sin_confirmar)

    print("\n" + "=" * 55)
    print("  Proceso completado exitosamente")
    print("=" * 55)

if __name__ == "__main__":
    main()
