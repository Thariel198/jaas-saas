# =========================IMPORTS===========================
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========================CONFIGURACION======================
BASE_DIR      = Path(__file__).parent
SHARED_DIR    = BASE_DIR.parent.parent.parent / "shared"
MOTOR_DIR     = BASE_DIR.parent / "motor_matching" / "outputs"
REPORTE_DIR   = SHARED_DIR / "reporte_mes_crudo"
ACUMULADO_DIR = SHARED_DIR / "reporte_acumulado_procesado"
OUTPUT_DIR    = BASE_DIR / "outputs"

TEPAGO_FILE  = "pagos_yape_tepago.xlsx"
BLANCOS_FILE = "blancos_mes.xlsx"
TIPO_PAGO    = "TE PAGÓ"
TOLERANCIA   = 0.01

ALIAS_BANCO = {
    "tipo":    ["tipo de transacción", "tipo de transaccion"],
    "origen":  ["origen"],
    "monto":   ["monto"],
    "mensaje": ["mensaje"],
    "fecha":   ["fecha de operación", "fecha de operacion"],
}

# ========================UTILIDADES=========================
def limpiar_monto(val) -> float:
    try:
        return round(float(str(val).replace(",", ".").strip()), 2)
    except:
        return 0.0

def normalizar_columnas(df: pd.DataFrame, alias: dict) -> dict:
    cols_lower = {str(c).lower().strip(): c for c in df.columns
                  if c and str(c).strip() not in ("None", "nan")}
    mapa = {}
    for campo, variantes in alias.items():
        for v in variantes:
            if v in cols_lower:
                mapa[campo] = cols_lower[v]
                break
    return mapa

def parsear_fecha(val) -> str:
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y %H:%M:%S")
    s = str(val).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%d/%m/%Y %H:%M:%S")
        except:
            pass
    return s

# ========================CARGAR ANCLA======================
def cargar_ancla() -> datetime | None:
    archivos = sorted(ACUMULADO_DIR.glob("*.xlsx"))
    if not archivos:
        print("  ⚠ Sin ancla — se incluirán todos los registros del banco")
        return None
    archivo = archivos[-1]
    print(f"  Ancla desde: {archivo.name}")
    wb    = load_workbook(archivo, read_only=True, data_only=True)
    ws    = wb.active
    datos = list(ws.values)
    wb.close()
    if len(datos) < 2:
        return None
    headers = [str(h).strip().lower() if h else "" for h in datos[0]]
    col_f   = next((i for i, h in enumerate(headers) if "fecha" in h), None)
    if col_f is None:
        return None
    fechas = []
    for fila in datos[1:]:
        if not fila or col_f >= len(fila):
            continue
        val = fila[col_f]
        if isinstance(val, datetime):
            fechas.append(val)
            continue
        s = str(val).strip()
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
                    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                fechas.append(datetime.strptime(s, fmt))
                break
            except:
                pass
    if not fechas:
        return None
    ancla = max(fechas)
    print(f"  Ancla: {ancla.strftime('%d/%m/%Y %H:%M:%S')}")
    return ancla

# ========================CARGAR BANCO======================
def cargar_banco(ancla=None) -> tuple:
    archivos = sorted(REPORTE_DIR.glob("*.xlsx"))
    if not archivos:
        raise FileNotFoundError(f"Sin archivos en {REPORTE_DIR}")

    todos      = []
    mapa_final = {}
    for archivo in archivos:
        wb    = load_workbook(archivo, read_only=False, data_only=True)
        ws    = wb.active
        datos = []
        for row in ws.iter_rows(values_only=True):
            if any(v is not None for v in row):
                datos.append(row)
        wb.close()
        if len(datos) < 2:
            continue

        header_idx = 0
        for i, fila in enumerate(datos):
            if fila[0] and str(fila[0]).strip().lower() in (
                "tipo de transacción", "tipo de transaccion"
            ):
                header_idx = i
                break

        headers = datos[header_idx]
        filas   = datos[header_idx + 1:]
        n_cols  = len(headers)
        df      = pd.DataFrame([f[:n_cols] for f in filas], columns=headers)
        mapa    = normalizar_columnas(df, ALIAS_BANCO)
        if any(c not in mapa for c in ["tipo", "origen", "monto"]):
            continue
        df = df[df[mapa["tipo"]].astype(str).str.strip().str.upper() == TIPO_PAGO.upper()].copy()
        todos.append(df)
        mapa_final = mapa

    if not todos:
        raise ValueError("Sin registros TE PAGÓ en el reporte del banco")

    df_total   = pd.concat(todos, ignore_index=True)
    cols_dedup = [mapa_final.get(c) for c in ["origen", "monto", "fecha"] if mapa_final.get(c)]
    df_total   = df_total.drop_duplicates(subset=cols_dedup).reset_index(drop=True)

    if ancla is not None and mapa_final.get("fecha"):
        col_f = mapa_final["fecha"]
        def _parse(val):
            if isinstance(val, datetime): return val
            s = str(val).strip()
            for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
                        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try: return datetime.strptime(s, fmt)
                except: pass
            return None
        df_total["_DT"] = df_total[col_f].apply(_parse)
        df_total = df_total[df_total["_DT"].notna() & (df_total["_DT"] > ancla)]
        df_total = df_total.drop(columns=["_DT"]).reset_index(drop=True)

    print(f"  ✔ {len(df_total)} registros TE PAGÓ del banco")
    return df_total, mapa_final

# ========================CARGAR TEPAGO=====================
def cargar_tepago() -> pd.DataFrame:
    ruta = MOTOR_DIR / TEPAGO_FILE
    if not ruta.exists():
        raise FileNotFoundError(f"No existe {ruta} — corre motor_matching primero")

    wb    = load_workbook(ruta, read_only=True, data_only=True)
    ws    = wb.active
    datos = list(ws.values)
    wb.close()

    if len(datos) < 3:
        raise ValueError("pagos_yape_tepago.xlsx tiene menos de 3 filas")

    # Fila 0 = grupos, fila 1 = columnas, fila 2+ = datos
    headers = [
        str(h).strip().upper() if h and str(h).strip() else f"_SEP{i}"
        for i, h in enumerate(datos[1])
    ]
    n_cols = len(headers)
    df     = pd.DataFrame([list(f[:n_cols]) for f in datos[2:]], columns=headers)
    df     = df[[c for c in df.columns if not c.startswith("_SEP")]].copy()

    print(f"  ✔ {len(df)} registros en pagos_yape_tepago")
    return df

# ========================CARGAR BLANCOS MES================
def cargar_blancos_mes() -> pd.DataFrame:
    ruta = MOTOR_DIR / BLANCOS_FILE
    if not ruta.exists():
        print("  Sin blancos_mes.xlsx — 0 blancos este mes")
        return pd.DataFrame()

    wb    = load_workbook(ruta, read_only=True, data_only=True)
    ws    = wb.active
    datos = list(ws.values)
    wb.close()

    if len(datos) < 3:
        return pd.DataFrame()

    headers = [
        str(h).strip().upper() if h and str(h).strip() else f"_SEP{i}"
        for i, h in enumerate(datos[1])
    ]
    n_cols = len(headers)
    df     = pd.DataFrame([list(f[:n_cols]) for f in datos[2:]], columns=headers)
    df     = df[[c for c in df.columns if not c.startswith("_SEP")]].copy()

    print(f"  ✔ {len(df)} blancos este mes")
    return df

# ========================VALIDACIONES======================
def v1_conteo(df_banco: pd.DataFrame, df_tepago: pd.DataFrame,
              df_blancos: pd.DataFrame) -> dict:
    """
    len(tepago) + len(blancos) == len(banco) + particiones_extra
    Las particiones_extra nacen cuando 1 registro del banco se parte en N filas (pago múltiple).
    """
    n_banco   = len(df_banco)
    n_tepago  = len(df_tepago)
    n_blancos = len(df_blancos)

    extra = 0
    fuentes_mult = ("MULTIPLE_AUTO", "MULTIPLE_CORREGIDO")
    if all(c in df_tepago.columns for c in ["FUENTE", "ORIGEN", "FECHA"]):
        mask  = df_tepago["FUENTE"].astype(str).str.strip().str.upper().isin(fuentes_mult)
        df_m  = df_tepago[mask]
        if not df_m.empty:
            extra = int((df_m.groupby(["ORIGEN", "FECHA"]).size() - 1).sum())

    real     = n_tepago + n_blancos
    esperado = n_banco + extra
    return {
        "ok": real == esperado,
        "banco": n_banco, "tepago": n_tepago, "blancos": n_blancos,
        "extra_particiones": extra,
        "real": real, "esperado": esperado,
        "diferencia": real - esperado,
    }


def v2_montos(df_banco: pd.DataFrame, mapa_banco: dict,
              df_tepago: pd.DataFrame, df_blancos: pd.DataFrame) -> dict:
    """
    sum(tepago.MONTO_PAGO) + sum(blancos.MONTO) == sum(banco.MONTO)
    Las particiones reparten el monto original — la suma total no cambia.
    """
    total_banco   = round(df_banco[mapa_banco["monto"]].apply(limpiar_monto).sum(), 2)
    total_tepago  = round(df_tepago["MONTO_PAGO"].apply(limpiar_monto).sum(), 2) \
                    if "MONTO_PAGO" in df_tepago.columns else 0.0
    total_blancos = round(df_blancos["MONTO"].apply(limpiar_monto).sum(), 2) \
                    if not df_blancos.empty and "MONTO" in df_blancos.columns else 0.0
    total_real    = round(total_tepago + total_blancos, 2)
    diferencia    = round(total_real - total_banco, 2)

    return {
        "ok": abs(diferencia) <= TOLERANCIA,
        "total_banco":   total_banco,
        "total_tepago":  total_tepago,
        "total_blancos": total_blancos,
        "total_real":    total_real,
        "diferencia":    diferencia,
    }


def v3_inventados(df_banco: pd.DataFrame, mapa_banco: dict,
                  df_tepago: pd.DataFrame, df_blancos: pd.DataFrame) -> dict:
    """Cada ORIGEN+FECHA en tepago/blancos debe existir en banco (nada inventado)."""
    col_orig = mapa_banco["origen"]
    col_fec  = mapa_banco.get("fecha", "")

    claves_banco = {
        f"{str(r[col_orig]).strip().upper()}|{parsear_fecha(r[col_fec]) if col_fec else ''}"
        for _, r in df_banco.iterrows()
    }

    inventados = []
    for df, archivo in [(df_tepago, "tepago"), (df_blancos, "blancos")]:
        if df.empty:
            continue
        for _, row in df.iterrows():
            orig = str(row.get("ORIGEN", "")).strip().upper()
            fec  = parsear_fecha(row.get("FECHA", ""))
            if f"{orig}|{fec}" not in claves_banco:
                inventados.append({
                    "ORIGEN":  row.get("ORIGEN", ""),
                    "FECHA":   row.get("FECHA",  ""),
                    "MONTO":   row.get("MONTO_PAGO", row.get("MONTO", "")),
                    "MZ":      row.get("MZ",     ""),
                    "LOTE":    row.get("LOTE",   ""),
                    "FUENTE":  row.get("FUENTE", archivo),
                    "ARCHIVO": archivo,
                })

    return {"ok": len(inventados) == 0, "registros": inventados}


def v4_perdidos(df_banco: pd.DataFrame, mapa_banco: dict,
                df_tepago: pd.DataFrame, df_blancos: pd.DataFrame) -> dict:
    """Cada ORIGEN+FECHA del banco debe aparecer en tepago o blancos (nada perdido)."""
    col_orig = mapa_banco["origen"]
    col_fec  = mapa_banco.get("fecha", "")
    col_mnt  = mapa_banco["monto"]

    claves_out = set()
    for df in [df_tepago, df_blancos]:
        if df.empty:
            continue
        for _, row in df.iterrows():
            orig = str(row.get("ORIGEN", "")).strip().upper()
            fec  = parsear_fecha(row.get("FECHA", ""))
            claves_out.add(f"{orig}|{fec}")

    perdidos = []
    for _, row in df_banco.iterrows():
        orig = str(row[col_orig]).strip().upper()
        fec  = parsear_fecha(row[col_fec]) if col_fec else ""
        if f"{orig}|{fec}" not in claves_out:
            perdidos.append({
                "ORIGEN": row[col_orig],
                "FECHA":  row[col_fec] if col_fec else "",
                "MONTO":  row[col_mnt],
            })

    return {"ok": len(perdidos) == 0, "registros": perdidos}


def v5_particiones(df_banco: pd.DataFrame, mapa_banco: dict,
                   df_tepago: pd.DataFrame) -> dict:
    """Para pagos múltiples: sum(MONTO_PAGO del grupo) == MONTO del banco."""
    if "FUENTE" not in df_tepago.columns:
        return {"ok": True, "registros": [], "nota": "Sin columna FUENTE"}

    fuentes_mult = ("MULTIPLE_AUTO", "MULTIPLE_CORREGIDO")
    mask = df_tepago["FUENTE"].astype(str).str.strip().str.upper().isin(fuentes_mult)
    df_m = df_tepago[mask].copy()

    if df_m.empty:
        return {"ok": True, "registros": [], "nota": "Sin pagos múltiples"}

    col_orig = mapa_banco["origen"]
    col_fec  = mapa_banco.get("fecha", "")
    col_mnt  = mapa_banco["monto"]

    monto_banco = {
        f"{str(r[col_orig]).strip().upper()}|{parsear_fecha(r[col_fec]) if col_fec else ''}":
        limpiar_monto(r[col_mnt])
        for _, r in df_banco.iterrows()
    }

    df_m["_K"] = (df_m["ORIGEN"].astype(str).str.strip().str.upper() + "|" +
                  df_m["FECHA"].apply(parsear_fecha))
    df_m["_M"] = df_m["MONTO_PAGO"].apply(limpiar_monto)

    errores = []
    for k, grupo in df_m.groupby("_K"):
        suma   = round(grupo["_M"].sum(), 2)
        orig_b = monto_banco.get(k)
        if orig_b is None:
            continue
        if abs(suma - orig_b) > TOLERANCIA:
            errores.append({
                "ORIGEN":      grupo["ORIGEN"].iloc[0],
                "FECHA":       grupo["FECHA"].iloc[0],
                "MONTO_BANCO": orig_b,
                "SUMA_TEPAGO": suma,
                "DIFERENCIA":  round(suma - orig_b, 2),
                "LOTES":       ", ".join(
                    f"Mz{r.get('MZ','')} Lt{r.get('LOTE','')}"
                    for _, r in grupo.iterrows()
                ),
            })

    return {"ok": len(errores) == 0, "registros": errores}


def v6_duplicados(df_tepago: pd.DataFrame) -> dict:
    """Sin ORIGEN+FECHA+MZ+LOTE duplicados en pagos_yape_tepago."""
    cols_k    = [c for c in ["ORIGEN", "FECHA", "MZ", "LOTE"] if c in df_tepago.columns]
    dupes     = df_tepago[df_tepago.duplicated(subset=cols_k, keep=False)].copy()
    registros = []
    if not dupes.empty:
        extra = [c for c in ["MONTO_PAGO", "FUENTE"] if c in df_tepago.columns]
        for _, row in dupes.iterrows():
            registros.append({c: row.get(c, "") for c in cols_k + extra})
    return {"ok": len(registros) == 0, "registros": registros}

# ========================EXPORTAR DISCREPANCIAS============
C_HDR_REPORTE = "4A235A"
C_OK          = "D5F5E3"; C_OK_T  = "1E8449"
C_ERR         = "FADBD8"; C_ERR_T = "922B21"

def _lado_r(color="CCCCCC"):
    return Side(style="thin", color=color)

def _cel_r(ws, row, col, val, bg=None, bold=False, color="222222",
           align="left", mono=False):
    c = ws.cell(row=row, column=col, value=val)
    b = _lado_r()
    c.border    = Border(left=b, right=b, top=b, bottom=b)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.font      = Font(name="Courier New" if mono else "Arial",
                       bold=bold, size=10, color=color)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    return c

def _hdr_r(ws, row, headers, bg, color="FFFFFF"):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        b = _lado_r("FFFFFF")
        c.border    = Border(left=b, right=b, top=b, bottom=b)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font      = Font(name="Arial", bold=True, size=10, color=color)
        c.fill      = PatternFill("solid", start_color=bg)
    ws.row_dimensions[row].height = 20

def exportar_discrepancias(rv: dict):
    OUTPUT_DIR.mkdir(exist_ok=True)
    mes  = datetime.today().strftime("%Y_%m")
    ruta = OUTPUT_DIR / f"reporte_validacion_{mes}.xlsx"

    wb = Workbook()

    # ── Hoja Resumen ───────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    _hdr_r(ws, 1, ["VALIDACIÓN", "ESTADO", "DETALLE"], C_HDR_REPORTE)

    items = [
        ("V1 — Conteo de registros", rv["conteo"],
         f"banco={rv['conteo']['banco']} · tepago={rv['conteo']['tepago']} · "
         f"blancos={rv['conteo']['blancos']} · extra_part={rv['conteo']['extra_particiones']} · "
         f"real={rv['conteo']['real']} · esperado={rv['conteo']['esperado']}"),
        ("V2 — Monto total", rv["montos"],
         f"banco=S/{rv['montos']['total_banco']} · tepago=S/{rv['montos']['total_tepago']} · "
         f"blancos=S/{rv['montos']['total_blancos']} · diff=S/{rv['montos']['diferencia']}"),
        ("V3 — Registros inventados", rv["inventados"],
         f"{len(rv['inventados']['registros'])} registros en tepago/blancos sin par en banco"),
        ("V4 — Registros perdidos", rv["perdidos"],
         f"{len(rv['perdidos']['registros'])} registros del banco sin par en tepago/blancos"),
        ("V5 — Integridad particiones", rv["particiones"],
         rv["particiones"].get(
             "nota",
             f"{len(rv['particiones']['registros'])} pagos múltiples con suma incorrecta"
         )),
        ("V6 — Duplicados", rv["duplicados"],
         f"{len(rv['duplicados']['registros'])} filas duplicadas en tepago"),
    ]

    for ri, (nombre, r, detalle) in enumerate(items, 2):
        ok = r["ok"]
        bg = C_OK if ok else C_ERR
        _cel_r(ws, ri, 1, nombre,  bg=bg)
        _cel_r(ws, ri, 2, "✔ OK" if ok else "✗ ERROR",
               bg=bg, bold=True, align="center",
               color=C_OK_T if ok else C_ERR_T)
        _cel_r(ws, ri, 3, detalle, bg=bg)
        ws.row_dimensions[ri].height = 18

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 72

    # ── Hojas de detalle solo si hay errores ───────────────
    def _hoja(titulo, registros, headers):
        if not registros:
            return
        ws2 = wb.create_sheet(titulo)
        _hdr_r(ws2, 1, headers, C_ERR_T)
        for ri, reg in enumerate(registros, 2):
            for ci, h in enumerate(headers, 1):
                _cel_r(ws2, ri, ci, reg.get(h, ""), bg=C_ERR)
            ws2.row_dimensions[ri].height = 16
        for ci in range(1, len(headers) + 1):
            ws2.column_dimensions[get_column_letter(ci)].width = 22

    _hoja("Inventados",  rv["inventados"]["registros"],
          ["ORIGEN", "FECHA", "MONTO", "MZ", "LOTE", "FUENTE", "ARCHIVO"])
    _hoja("Perdidos",    rv["perdidos"]["registros"],
          ["ORIGEN", "FECHA", "MONTO"])
    _hoja("Particiones", rv["particiones"]["registros"],
          ["ORIGEN", "FECHA", "MONTO_BANCO", "SUMA_TEPAGO", "DIFERENCIA", "LOTES"])
    _hoja("Duplicados",  rv["duplicados"]["registros"],
          ["ORIGEN", "FECHA", "MZ", "LOTE", "MONTO_PAGO", "FUENTE"])

    wb.save(ruta)
    print(f"\n  ⚠ Discrepancias exportadas: {ruta.name}")

# ======================MAIN=================================
def main():
    print("=" * 55)
    print("  VALIDACIÓN YAPE — integridad del matching")
    print("=" * 55)

    print("\n[1] Cargando ancla de corte...")
    ancla = cargar_ancla()

    print("\n[2] Cargando reporte del banco...")
    df_banco, mapa_banco = cargar_banco(ancla)

    print("\n[3] Cargando pagos_yape_tepago...")
    df_tepago = cargar_tepago()

    print("\n[4] Cargando blancos_mes...")
    df_blancos = cargar_blancos_mes()

    print("\n[5] Ejecutando validaciones...")
    rv = {
        "conteo":      v1_conteo(df_banco, df_tepago, df_blancos),
        "montos":      v2_montos(df_banco, mapa_banco, df_tepago, df_blancos),
        "inventados":  v3_inventados(df_banco, mapa_banco, df_tepago, df_blancos),
        "perdidos":    v4_perdidos(df_banco, mapa_banco, df_tepago, df_blancos),
        "particiones": v5_particiones(df_banco, mapa_banco, df_tepago),
        "duplicados":  v6_duplicados(df_tepago),
    }

    print("\n" + "─" * 55)
    todo_ok = True
    lineas  = [
        ("V1 Conteo",      rv["conteo"],
         f"banco={rv['conteo']['banco']} tepago={rv['conteo']['tepago']} "
         f"blancos={rv['conteo']['blancos']} extra={rv['conteo']['extra_particiones']} "
         f"diff={rv['conteo']['diferencia']}"),
        ("V2 Montos",      rv["montos"],
         f"banco=S/{rv['montos']['total_banco']} "
         f"real=S/{rv['montos']['total_real']} "
         f"diff=S/{rv['montos']['diferencia']}"),
        ("V3 Inventados",  rv["inventados"],
         f"{len(rv['inventados']['registros'])} sin par en banco"),
        ("V4 Perdidos",    rv["perdidos"],
         f"{len(rv['perdidos']['registros'])} del banco sin par"),
        ("V5 Particiones", rv["particiones"],
         rv["particiones"].get(
             "nota",
             f"{len(rv['particiones']['registros'])} errores de suma"
         )),
        ("V6 Duplicados",  rv["duplicados"],
         f"{len(rv['duplicados']['registros'])} filas duplicadas"),
    ]

    for nombre, r, detalle in lineas:
        icono = "✔" if r["ok"] else "✗"
        print(f"  {icono}  {nombre:<22} {detalle}")
        if not r["ok"]:
            todo_ok = False

    print("─" * 55)

    if todo_ok:
        print("\n  ✔ Todo OK — el matching preservó la integridad del reporte")
    else:
        print("\n  ✗ Hay discrepancias — se exporta reporte de validación")
        exportar_discrepancias(rv)

    print("=" * 55)

if __name__ == "__main__":
    main()
