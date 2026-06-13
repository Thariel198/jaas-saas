# =========================IMPORTS===========================
from copy import copy as _copy
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========================COLORES DISEÑO=====================
# Siguiendo exactamente los HTMLs de diseño

# Grupos compartidos
C_ID_H    = "F4ECF7"; C_ID_C  = "FAF5FF"; C_ID_T  = "5B21B6"
C_BAN_H   = "E6F1FB"; C_BAN_C = "F0F8FF"; C_BAN_T = "185FA5"
C_UBI_H   = "E1F5EE"; C_UBI_C = "F0FFF8"; C_UBI_T = "065F46"
C_DED_H   = "FEF9E7"; C_DED_C = "FEFCE8"; C_DED_T = "7D6608"
C_CIC_H   = "FEF3E8"; C_CIC_C = "FEF3E8"; C_CIC_T = "7C3003"
C_MES_H   = "EFF6FF"; C_MES_C = "EFF6FF"; C_MES_T = "1D4ED8"
C_EST_H   = "FEF9E7"; C_EST_T = "7D6608"
C_LOT_H   = "E1F5EE"; C_LOT_C = "F0FFF8"; C_LOT_T = "065F46"
C_USR_H   = "E1F5EE"; C_USR_C = "F0FFF8"; C_USR_T = "065F46"
C_DEV_H   = "FEF3E8"; C_DEV_T = "7C3003"  # devoluciones tab
C_SEP     = "F3F4F6"

# Estados
EST_EXACTO  = "E1F5EE"; EST_EXACTO_T  = "085041"
EST_EXCESO  = "FAEEDA"; EST_EXCESO_T  = "854F0B"
EST_PARCIAL = "FCEBEB"; EST_PARCIAL_T = "A32D2D"
EST_PEND    = "FCEBEB"; EST_PEND_T    = "A32D2D"
EST_IDEN    = "FAEEDA"; EST_IDEN_T    = "854F0B"
EST_APLI    = "E1F5EE"; EST_APLI_T    = "085041"
EST_NOAP    = "FEFCE8"; EST_NOAP_T    = "7D6608"

# Fuentes
FUENTE_MSG = "E1F5EE"; FUENTE_MSG_T = "085041"
FUENTE_MAE = "E6F1FB"; FUENTE_MAE_T = "0C447C"
FUENTE_AMB = "FAEEDA"; FUENTE_AMB_T = "854F0B"
FUENTE_MUL = "F4ECF7"; FUENTE_MUL_T = "5B21B6"
FUENTE_MAN = "FEF3E8"; FUENTE_MAN_T = "7C3003"

MONTO_T = "065F46"

def _lado(c="CCCCCC"):
    return Side(style="thin", color=c)

def _cel(cell, bg, txt, mono=False, align="left", bold=False, size=10):
    b = _lado()
    cell.border    = Border(left=b, right=b, top=b, bottom=b)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.font      = Font(name="Courier New" if mono else "Arial",
                          bold=bold, size=size, color=txt)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)

def _hdr(cell, bg, txt, bold=True, size=9, align="center"):
    b = _lado("FFFFFF")
    cell.border    = Border(left=b, right=b, top=b, bottom=b)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.font      = Font(name="Arial", bold=bold, size=size, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)

def _sep(cell):
    cell.fill   = PatternFill("solid", start_color=C_SEP)
    cell.border = Border(left=_lado(), right=_lado("AAAAAA"),
                         top=_lado(),  bottom=_lado())

def _badge(cell, bg, txt, text, size=9):
    """Celda con estilo badge — texto centrado con color propio."""
    _cel(cell, bg, txt, mono=True, align="center", bold=True, size=size)
    cell.value = text

def _merge_hdr(ws, row, col_start, col_end, text, bg, txt, size=9):
    if col_start == col_end:
        c = ws.cell(row=row, column=col_start, value=text)
        _hdr(c, bg, txt, size=size)
    else:
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row,   end_column=col_end)
        c = ws.cell(row=row, column=col_start, value=text)
        _hdr(c, bg, txt, size=size)

def _sep_col(ws, row, ci, height=None):
    c = ws.cell(row=row, column=ci, value="")
    _sep(c)
    ws.column_dimensions[get_column_letter(ci)].width = 1

# ========================EXPORTAR pagos_yape_tepago=========
def exportar_pagos_tepago(todos: list, ciclo: int):
    """
    Exporta pagos_yape_tepago.xlsx siguiendo pagos_yape_tepago_diseno.html
    Narrativa: ¿Quién es? → ¿Qué hizo el banco? → ¿Dónde vive? → ¿Cuánto debe? → ¿Cuándo?
    """
    identificados = [r for r in todos if r.get("estado") == "identificado"]

    wb = Workbook()
    ws = wb.active
    ws.title = "pagos_yape_tepago"
    ws.freeze_panes = "A3"

    # Definicion de columnas con separadores
    # Grupos: quien(3) | sep | banco(7) | sep | donde(2) | sep | cuanto(3) | sep | ciclo(1)
    cols = [
        # quien
        ("USER_ID",        C_ID_C,  C_ID_T,  True,  "left"),
        ("NOMBRE",         C_ID_C,  C_ID_T,  False, "left"),
        ("FUENTE",         C_ID_C,  C_ID_T,  True,  "center"),
        "__SEP__",
        # banco
        ("TIPO",           C_BAN_C, C_BAN_T, True,  "left"),
        ("ORIGEN",         C_BAN_C, C_BAN_T, True,  "left"),
        ("DESTINO",        C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO_PAGO",     C_BAN_C, MONTO_T, True,  "right"),
        ("MONTO_ASIGNADO", C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE",        C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",          C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        # donde
        ("MZ",      C_UBI_C, C_UBI_T, True,  "center"),
        ("LOTE",    C_UBI_C, C_UBI_T, True,  "center"),
        ("CONCEPTO",C_UBI_C, C_UBI_T, False, "left"),
        "__SEP__",
        # cuanto
        ("DEUDA_TOTAL",  C_DED_C, C_DED_T, True, "right"),
        ("DIFERENCIA",   C_DED_C, C_DED_T, True, "right"),
        ("ESTADO_PAGO",  C_DED_C, C_DED_T, True, "center"),
        "__SEP__",
        # ciclo
        ("CICLO", C_CIC_C, C_CIC_T, True, "center"),
    ]

    grupos = [
        ("¿QUIÉN ES?",        3, C_ID_H,  C_ID_T),
        None,
        ("¿QUÉ HIZO EL BANCO?", 7, C_BAN_H, C_BAN_T),
        None,
        ("¿DÓNDE VIVE?",      3, C_UBI_H, C_UBI_T),
        None,
        ("¿CUÁNTO DEBE?",     3, C_DED_H, C_DED_T),
        None,
        ("¿CUÁNDO?",          1, C_CIC_H, C_CIC_T),
    ]

    _escribir_cabecera_doble(ws, cols, grupos)

    anchos = {
        "USER_ID":10, "NOMBRE":28, "FUENTE":10,
        "TIPO":10, "ORIGEN":22, "DESTINO":22,
        "MONTO_PAGO":12, "MONTO_ASIGNADO":14,
        "MENSAJE":35, "FECHA":20,
        "MZ":6, "LOTE":7, "CONCEPTO":25,
        "DEUDA_TOTAL":12, "DIFERENCIA":12, "ESTADO_PAGO":14,
        "CICLO":7,
    }

    fuente_colores = {
        "mensaje":             (FUENTE_MSG, FUENTE_MSG_T),
        "maestro":             (FUENTE_MAE, FUENTE_MAE_T),
        "manual":              (FUENTE_MAN, FUENTE_MAN_T),
        "maestro_inexacto":    (FUENTE_MAN, FUENTE_MAN_T),
        "ambiguo_auto":        (FUENTE_MAN, FUENTE_MAN_T),
        "correccion":          (FUENTE_MAN, FUENTE_MAN_T),
        "blanco":              (EST_EXCESO, EST_EXCESO_T),
    }

    estado_colores = {
        "exacto":   (EST_EXACTO,  EST_EXACTO_T),
        "exceso":   (EST_EXCESO,  EST_EXCESO_T),
        "parcial":  (EST_PARCIAL, EST_PARCIAL_T),
        "concepto": ("EFF6FF",    "1D4ED8"),
    }

    for ri, reg in enumerate(identificados, start=3):
        fuente = str(reg.get("fuente", "")).lower()
        f_bg, f_txt = fuente_colores.get(fuente, (C_ID_C, C_ID_T))
        ep = str(reg.get("estado_pago", "")).lower()
        ep_bg, ep_txt = estado_colores.get(ep, (C_DED_C, C_DED_T))

        valores = {
            "USER_ID":        reg.get("user_id", ""),
            "NOMBRE":         reg.get("nombre", ""),
            "FUENTE":         fuente,
            "TIPO":           "TE PAGÓ",
            "ORIGEN":         reg.get("origen", ""),
            "DESTINO":        "",
            "MONTO_PAGO":     reg.get("monto_pago", ""),
            "MONTO_ASIGNADO": reg.get("monto_asignado", ""),
            "MENSAJE":        reg.get("mensaje", ""),
            "FECHA":          reg.get("fecha", ""),
            "MZ":             reg.get("mz", ""),
            "LOTE":           reg.get("lote", ""),
            "CONCEPTO":       reg.get("concepto", ""),
            "DEUDA_TOTAL":    reg.get("deuda_total", ""),
            "DIFERENCIA":     reg.get("diferencia", ""),
            "ESTADO_PAGO":    ep,
            "CICLO":          ciclo,
        }

        ci = 1
        for col in cols:
            if col == "__SEP__":
                _sep_col(ws, ri, ci)
                ci += 1
                continue
            nombre, bg, txt, mono, align = col
            val = valores.get(nombre, "")

            # Colores especiales
            if nombre == "FUENTE":
                bg, txt = f_bg, f_txt
            elif nombre == "ESTADO_PAGO":
                bg, txt = ep_bg, ep_txt
            elif nombre in ("MONTO_PAGO", "MONTO_ASIGNADO", "DEUDA_TOTAL", "DIFERENCIA"):
                txt = MONTO_T

            c = ws.cell(row=ri, column=ci, value=val if val != "" else "")
            _cel(c, bg, txt, mono=mono, align=align)
            ci += 1

        ws.row_dimensions[ri].height = 18

    _aplicar_anchos(ws, cols, anchos)
    return wb

# ========================EXPORTAR pendientes================
def exportar_pendientes_diseño(sin_resolver: list, ambiguos: list = None,
                               maestro_inexacto: list = None,
                               pagaste: list = None,
                               preservados: dict = None) -> Workbook:
    """
    Exporta pendientes.xlsx con 3 hojas siguiendo pendientes_xlsx.html:
    - Sin_identificar: sin maestro ni mensaje válido
    - Ambiguos: maestro encontró 2+ candidatos
    - Maestro_inexacto: maestro encontró 1 candidato pero diff != 0

    preservados: dict {hoja: {clave_origen_fecha: {campo: valor_usuario}}}
                 Inyecta MZ/LOTE/CONCEPTO/MOTIVO/OK que el usuario llenó en el run
                 anterior pero no marcó como confirmados (OK=SI).
    """
    if ambiguos         is None: ambiguos         = []
    if maestro_inexacto is None: maestro_inexacto = []
    if pagaste          is None: pagaste          = []

    def _aplicar_preservado(valores: dict, reg: dict, hoja: str):
        if not preservados:
            return
        clave = f"{str(reg.get('origen','')).upper().strip()}|{str(reg.get('fecha','')).strip()}"
        p = preservados.get(hoja, {}).get(clave)
        if not p:
            return
        for campo in ("MZ", "LOTE", "CONCEPTO", "MOTIVO", "OK"):
            if campo in valores and p.get(campo) not in (None, ""):
                valores[campo] = p[campo]

    C_SUG_H = "FAEEDA"; C_SUG_C = "FFFBF0"; C_SUG_T = "854F0B"

    wb = Workbook()

    # ── Hoja 1: Sin_identificar ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Sin_identificar"
    ws1.freeze_panes = "A3"

    cols_si = [
        ("TIPO",    C_BAN_C, C_BAN_T, True,  "left"),
        ("ORIGEN",  C_BAN_C, C_BAN_T, True,  "left"),
        ("DESTINO", C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO",   C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE", C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",   C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MZ",      C_USR_C, C_USR_T, True,  "center"),
        ("LOTE",    C_USR_C, C_USR_T, True,  "center"),
        ("CONCEPTO",C_USR_C, C_USR_T, False, "left"),
        ("MOTIVO",  C_USR_C, C_USR_T, False, "left"),
        "__SEP__",
        ("OK",      C_ID_C,  C_ID_T,  True,  "center"),
    ]
    grupos_si = [
        ("← BANCO — NO EDITAR", 6, C_BAN_H, C_BAN_T),
        None,
        ("← TÚ COMPLETAS",      4, C_USR_H, C_USR_T),
        None,
        ("✓",                   1, C_ID_H,  C_ID_T),
    ]
    anchos_si = {
        "TIPO":10, "ORIGEN":28, "DESTINO":22,
        "MONTO":12, "MENSAJE":35, "FECHA":20,
        "MZ":10, "LOTE":10, "CONCEPTO":25, "MOTIVO":30, "OK":6,
    }
    _escribir_cabecera_doble(ws1, cols_si, grupos_si)
    tipo_col = {"TE PAGÓ": (EST_EXACTO, EST_EXACTO_T), "PAGASTE": (EST_EXCESO, EST_EXCESO_T)}
    for ri, reg in enumerate(sin_resolver, start=3):
        tipo_str = "PAGASTE" if "PAGASTE" in str(reg.get("tipo", reg.get("fuente",""))).upper() else "TE PAGÓ"
        t_bg, t_txt = tipo_col.get(tipo_str, (C_BAN_C, C_BAN_T))
        valores = {
            "TIPO": tipo_str, "ORIGEN": reg.get("origen",""),
            "DESTINO": reg.get("destino",""),
            "MONTO": reg.get("monto_pago", reg.get("monto","")),
            "MENSAJE": reg.get("mensaje",""), "FECHA": reg.get("fecha",""),
            "MZ":"", "LOTE":"", "CONCEPTO":"", "MOTIVO": reg.get("motivo",""), "OK":"",
        }
        _aplicar_preservado(valores, reg, "Sin_identificar")
        ci = 1
        for col in cols_si:
            if col == "__SEP__": _sep_col(ws1, ri, ci); ci += 1; continue
            nombre, bg, txt, mono, align = col
            val = valores.get(nombre, "")
            if nombre == "TIPO": bg, txt = t_bg, t_txt
            _cel(ws1.cell(row=ri, column=ci, value=val), bg, txt, mono=mono, align=align)
            ci += 1
        ws1.row_dimensions[ri].height = 18
    _aplicar_anchos(ws1, cols_si, anchos_si)

    # ── Hoja 2: Ambiguos ─────────────────────────────────────
    ws2 = wb.create_sheet("Ambiguos")
    ws2.freeze_panes = "A3"
    cols_amb = [
        ("USER_ID",  C_ID_C,  C_ID_T,  True,  "left"),
        ("NOMBRE",   C_ID_C,  C_ID_T,  False, "left"),
        "__SEP__",
        ("TIPO",     C_BAN_C, C_BAN_T, True,  "left"),
        ("ORIGEN",   C_BAN_C, C_BAN_T, True,  "left"),
        ("DESTINO",  C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO",    C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE",  C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",    C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MZ_SUG",   C_SUG_H, C_SUG_T, True,  "center"),
        ("LOTE_SUG", C_SUG_H, C_SUG_T, True,  "center"),
        ("CANDIDATOS",C_SUG_H,C_SUG_T, False, "left"),
        "__SEP__",
        ("MZ",       C_USR_C, C_USR_T, True,  "center"),
        ("LOTE",     C_USR_C, C_USR_T, True,  "center"),
        ("CONCEPTO", C_USR_C, C_USR_T, False, "left"),
        ("MOTIVO",   C_USR_C, C_USR_T, False, "left"),
        "__SEP__",
        ("OK",       C_ID_C,  C_ID_T,  True,  "center"),
    ]
    grupos_amb = [
        ("¿QUIÉN PAGÓ?",        2, C_ID_H,  C_ID_T),
        None,
        ("← BANCO — NO EDITAR", 6, C_BAN_H, C_BAN_T),
        None,
        ("← SISTEMA SUGIERE",   3, C_SUG_H, C_SUG_T),
        None,
        ("← TÚ COMPLETAS",      4, C_USR_H, C_USR_T),
        None,
        ("✓",                   1, C_ID_H,  C_ID_T),
    ]
    anchos_amb = {
        "USER_ID":10, "NOMBRE":24,
        "TIPO":10, "ORIGEN":28, "DESTINO":20,
        "MONTO":12, "MENSAJE":28, "FECHA":20,
        "MZ_SUG":8, "LOTE_SUG":9, "CANDIDATOS":32,
        "MZ":8, "LOTE":8, "CONCEPTO":20, "MOTIVO":25, "OK":6,
    }
    _escribir_cabecera_doble(ws2, cols_amb, grupos_amb)
    for ri, reg in enumerate(ambiguos, start=3):
        candidatos = reg.get("candidatos", [])
        partes = []
        for c in candidatos:
            deu = c.get("deuda", ""); dif = c.get("dif", "")
            deu_s = str(int(deu)) if isinstance(deu, (int,float)) and deu else str(deu)
            dif_s = (f"+{int(dif)}" if dif > 0 else str(int(dif))) if isinstance(dif, (int,float)) else str(dif)
            partes.append(f"{c.get('mz','')}-{c.get('lote','')}({deu_s},{dif_s})")
        valores = {
            "USER_ID": reg.get("user_id",""), "NOMBRE": reg.get("nombre",""),
            "TIPO": "TE PAGÓ", "ORIGEN": reg.get("origen",""),
            "DESTINO": reg.get("destino",""),
            "MONTO": reg.get("monto_pago", reg.get("monto","")),
            "MENSAJE": reg.get("mensaje",""), "FECHA": reg.get("fecha",""),
            "MZ_SUG": reg.get("mz_sug",""), "LOTE_SUG": reg.get("lote_sug",""),
            "CANDIDATOS": " · ".join(partes),
            "MZ":"", "LOTE":"", "CONCEPTO":"", "MOTIVO":"", "OK":"",
        }
        _aplicar_preservado(valores, reg, "Ambiguos")
        ci = 1
        for col in cols_amb:
            if col == "__SEP__": _sep_col(ws2, ri, ci); ci += 1; continue
            nombre, bg, txt, mono, align = col
            _cel(ws2.cell(row=ri, column=ci, value=valores.get(nombre,"")), bg, txt, mono=mono, align=align)
            ci += 1
        ws2.row_dimensions[ri].height = 18
    _aplicar_anchos(ws2, cols_amb, anchos_amb)

    # ── Hoja 3: Maestro_inexacto ─────────────────────────────
    ws3 = wb.create_sheet("Maestro_inexacto")
    ws3.freeze_panes = "A3"
    cols_mix = [
        ("USER_ID",  C_ID_C,  C_ID_T,  True,  "left"),
        ("NOMBRE",   C_ID_C,  C_ID_T,  False, "left"),
        "__SEP__",
        ("TIPO",     C_BAN_C, C_BAN_T, True,  "left"),
        ("ORIGEN",   C_BAN_C, C_BAN_T, True,  "left"),
        ("DESTINO",  C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO",    C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE",  C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",    C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MZ_SUG",   C_SUG_H, C_SUG_T, True,  "center"),
        ("LOTE_SUG", C_SUG_H, C_SUG_T, True,  "center"),
        ("DEUDA",    C_SUG_H, C_SUG_T, True,  "right"),
        ("DIFF",     C_SUG_H, C_SUG_T, True,  "right"),
        "__SEP__",
        ("MZ",       C_USR_C, C_USR_T, True,  "center"),
        ("LOTE",     C_USR_C, C_USR_T, True,  "center"),
        ("CONCEPTO", C_USR_C, C_USR_T, False, "left"),
        ("MOTIVO",   C_USR_C, C_USR_T, False, "left"),
        "__SEP__",
        ("OK",       C_ID_C,  C_ID_T,  True,  "center"),
    ]
    grupos_mix = [
        ("¿QUIÉN PAGÓ?",        2, C_ID_H,  C_ID_T),
        None,
        ("← BANCO — NO EDITAR", 6, C_BAN_H, C_BAN_T),
        None,
        ("← SISTEMA ENCONTRÓ",  4, C_SUG_H, C_SUG_T),
        None,
        ("← TÚ COMPLETAS",      4, C_USR_H, C_USR_T),
        None,
        ("✓",                   1, C_ID_H,  C_ID_T),
    ]
    anchos_mix = {
        "USER_ID":10, "NOMBRE":24,
        "TIPO":10, "ORIGEN":28, "DESTINO":20,
        "MONTO":12, "MENSAJE":28, "FECHA":20,
        "MZ_SUG":8, "LOTE_SUG":9, "DEUDA":12, "DIFF":10,
        "MZ":8, "LOTE":8, "CONCEPTO":20, "MOTIVO":25, "OK":6,
    }
    _escribir_cabecera_doble(ws3, cols_mix, grupos_mix)
    for ri, reg in enumerate(maestro_inexacto, start=3):
        diff_val = reg.get("diferencia", reg.get("diff", ""))
        valores = {
            "USER_ID": reg.get("user_id",""), "NOMBRE": reg.get("nombre",""),
            "TIPO": "TE PAGÓ", "ORIGEN": reg.get("origen",""),
            "DESTINO": reg.get("destino",""),
            "MONTO": reg.get("monto_pago", reg.get("monto","")),
            "MENSAJE": reg.get("mensaje",""), "FECHA": reg.get("fecha",""),
            "MZ_SUG": reg.get("mz",""), "LOTE_SUG": reg.get("lote",""),
            "DEUDA": reg.get("deuda_total",""), "DIFF": diff_val,
            "MZ":"", "LOTE":"", "CONCEPTO":"", "MOTIVO":"", "OK":"",
        }
        _aplicar_preservado(valores, reg, "Maestro_inexacto")
        ci = 1
        for col in cols_mix:
            if col == "__SEP__": _sep_col(ws3, ri, ci); ci += 1; continue
            nombre, bg, txt, mono, align = col
            val = valores.get(nombre, "")
            # DIFF en rojo si negativo, verde si positivo
            if nombre == "DIFF" and isinstance(val, (int, float)):
                txt = "065F46" if val >= 0 else "A32D2D"
            _cel(ws3.cell(row=ri, column=ci, value=val), bg, txt, mono=mono, align=align)
            ci += 1
        ws3.row_dimensions[ri].height = 18
    _aplicar_anchos(ws3, cols_mix, anchos_mix)

    # ── Hoja 4: Pagaste ──────────────────────────────────────
    if pagaste:
        ws4 = wb.create_sheet("Pagaste")
        ws4.freeze_panes = "A3"
        cols_pag = [
            ("TIPO",    C_BAN_C, C_BAN_T, True,  "left"),
            ("ORIGEN",  C_BAN_C, C_BAN_T, True,  "left"),
            ("DESTINO", C_BAN_C, C_BAN_T, True,  "left"),
            ("MONTO",   C_BAN_C, MONTO_T, True,  "right"),
            ("MENSAJE", C_BAN_C, C_BAN_T, False, "left"),
            ("FECHA",   C_BAN_C, C_BAN_T, True,  "left"),
            "__SEP__",
            ("MZ",       C_USR_C, C_USR_T, True,  "center"),
            ("LOTE",     C_USR_C, C_USR_T, True,  "center"),
            ("CONCEPTO", C_USR_C, C_USR_T, False, "left"),
            ("MOTIVO",   C_USR_C, C_USR_T, False, "left"),
            "__SEP__",
            ("OK", C_ID_C, C_ID_T, True, "center"),
        ]
        grupos_pag = [
            ("← BANCO — NO EDITAR", 6, C_BAN_H, C_BAN_T),
            None,
            ("← TÚ COMPLETAS",      4, C_USR_H, C_USR_T),
            None,
            ("✓",                   1, C_ID_H,  C_ID_T),
        ]
        anchos_pag = {
            "TIPO":10, "ORIGEN":28, "DESTINO":22,
            "MONTO":12, "MENSAJE":35, "FECHA":20,
            "MZ":6, "LOTE":6,
            "CONCEPTO":30, "MOTIVO":25, "OK":6,
        }
        _escribir_cabecera_doble(ws4, cols_pag, grupos_pag)
        for ri, reg in enumerate(pagaste, start=3):
            valores = {
                "TIPO":    "PAGASTE",
                "ORIGEN":  reg.get("origen",  ""),
                "DESTINO": reg.get("destino", ""),
                "MONTO":   reg.get("monto",   ""),
                "MENSAJE": reg.get("mensaje", ""),
                "FECHA":   reg.get("fecha",   ""),
                "MZ":       "",
                "LOTE":     "",
                "CONCEPTO": "",
                "MOTIVO":   "",
                "OK":       "",
            }
            _aplicar_preservado(valores, reg, "Pagaste")
            ci = 1
            for col in cols_pag:
                if col == "__SEP__":
                    _sep_col(ws4, ri, ci)
                    ci += 1
                    continue
                nombre, bg, txt, mono, align = col
                c = ws4.cell(row=ri, column=ci, value=valores.get(nombre, ""))
                _cel(c, bg, txt, mono=mono, align=align)
                ci += 1
            ws4.row_dimensions[ri].height = 18
        _aplicar_anchos(ws4, cols_pag, anchos_pag)

    return wb

# ========================EXPORTAR blancos_mes===============
def exportar_blancos_mes(blancos: list, ciclo: int) -> Workbook:
    """
    Exporta blancos_mes.xlsx siguiendo blancos_mes_diseno.html
    Narrativa: ¿Qué hizo el banco? → ¿Por qué BLANCO? → ¿Cuándo?
    """
    cols = [
        ("TIPO",    C_BAN_C, C_BAN_T, True,  "left"),
        ("ORIGEN",  C_BAN_C, C_BAN_T, True,  "left"),
        ("DESTINO", C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO",   C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE", C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",   C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MOTIVO",  C_DED_C, C_DED_T, False, "left"),
        "__SEP__",
        ("CICLO",   C_CIC_C, C_CIC_T, True,  "center"),
    ]

    grupos = [
        ("¿QUÉ HIZO EL BANCO?", 6, C_BAN_H, C_BAN_T),
        None,
        ("¿POR QUÉ BLANCO?",    1, C_DED_H, C_DED_T),
        None,
        ("¿CUÁNDO?",            1, C_CIC_H, C_CIC_T),
    ]

    anchos = {
        "TIPO":10, "ORIGEN":28, "DESTINO":22,
        "MONTO":12, "MENSAJE":35, "FECHA":20,
        "MOTIVO":35, "CICLO":7,
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "blancos_mes"
    ws.freeze_panes = "A3"

    _escribir_cabecera_doble(ws, cols, grupos)

    for ri, reg in enumerate(blancos, start=3):
        msg_raw = str(reg.get("mensaje", "")).strip()
        valores = {
            "TIPO":    "TE PAGÓ",
            "ORIGEN":  reg.get("origen",  ""),
            "DESTINO": reg.get("destino", ""),
            "MONTO":   reg.get("monto_pago", reg.get("monto", "")),
            "MENSAJE": "" if msg_raw.lower() in ("nan", "none", "") else msg_raw,
            "FECHA":   reg.get("fecha",   ""),
            "MOTIVO":  reg.get("motivo",  "marcado como blanco"),
            "CICLO":   ciclo,
        }
        ci = 1
        for col in cols:
            if col == "__SEP__":
                _sep_col(ws, ri, ci)
                ci += 1
                continue
            nombre, bg, txt, mono, align = col
            c = ws.cell(row=ri, column=ci, value=valores.get(nombre, ""))
            _cel(c, bg, txt, mono=mono, align=align)
            ci += 1
        ws.row_dimensions[ri].height = 18

    _aplicar_anchos(ws, cols, anchos)
    return wb

# ========================EXPORTAR blancos_acumulados========
def exportar_blancos_acumulados(wb_existente, nuevo_blanco: dict, mes: str) -> Workbook:
    """
    Agrega un registro a blancos_acumulados.xlsx siguiendo blancos_acumulados_diseno.html
    Narrativa: ¿De qué mes? → ¿Quién es? → ¿Qué hizo el banco? → ¿Cómo se identificó? → ¿Ya se aplicó?
    """
    cols = [
        ("MES",     C_MES_C, C_MES_T, True,  "center"),
        "__SEP__",
        ("USER_ID", C_ID_C,  C_ID_T,  True,  "left"),
        ("NOMBRE",  C_ID_C,  C_ID_T,  False, "left"),
        "__SEP__",
        ("TIPO",    C_BAN_C, C_BAN_T, True,  "left"),
        ("ORIGEN",  C_BAN_C, C_BAN_T, True,  "left"),
        ("DESTINO", C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO",   C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE", C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",   C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MZ",       C_UBI_C, C_UBI_T, True,  "center"),
        ("LOTE",     C_UBI_C, C_UBI_T, True,  "center"),
        ("CONCEPTO", C_UBI_C, C_UBI_T, False, "left"),
        ("MOTIVO",   C_UBI_C, C_UBI_T, False, "left"),
        "__SEP__",
        ("ESTADO",   C_EST_H, C_EST_T, True,  "center"),
    ]

    grupos = [
        ("¿DE QUÉ MES?",         1, C_MES_H, C_MES_T),
        None,
        ("¿QUIÉN ES?",           2, C_ID_H,  C_ID_T),
        None,
        ("¿QUÉ HIZO EL BANCO?",  6, C_BAN_H, C_BAN_T),
        None,
        ("¿CÓMO SE IDENTIFICÓ?", 4, C_UBI_H, C_UBI_T),
        None,
        ("¿YA SE APLICÓ?",       1, C_EST_H, C_EST_T),
    ]

    anchos = {
        "MES":10, "USER_ID":10, "NOMBRE":28,
        "TIPO":10, "ORIGEN":22, "DESTINO":22,
        "MONTO":12, "MENSAJE":30, "FECHA":20,
        "MZ":6, "LOTE":7, "CONCEPTO":25, "MOTIVO":30, "ESTADO":12,
    }

    # Si el workbook existe, leer filas existentes y extraer claves para dedup
    filas_existentes = []
    claves_existentes = set()
    if wb_existente is not None:
        ws_old = wb_existente.active
        rows = list(ws_old.values)
        # Fila 1 (index 0) = grupos, fila 2 (index 1) = cabeceras, fila 3+ = datos
        old_headers = [str(h).strip().upper() if h else "" for h in rows[1]] if len(rows) > 1 else []
        i_orig_old = old_headers.index("ORIGEN") if "ORIGEN" in old_headers else None
        i_fec_old  = old_headers.index("FECHA")  if "FECHA"  in old_headers else None
        for fila in rows[2:]:
            if fila and any(fila):
                filas_existentes.append(fila)
                orig = str(fila[i_orig_old]).strip().upper() if i_orig_old is not None and i_orig_old < len(fila) and fila[i_orig_old] else ""
                fec  = str(fila[i_fec_old]).strip()          if i_fec_old  is not None and i_fec_old  < len(fila) and fila[i_fec_old]  else ""
                if orig:
                    claves_existentes.add(f"{orig}|{fec}")

    # Verificar si el nuevo blanco ya existe (evita duplicados al re-correr el motor)
    nuevo_origen = str(nuevo_blanco.get("origen", "")).strip().upper()
    nuevo_fecha  = str(nuevo_blanco.get("fecha",  "")).strip()
    if f"{nuevo_origen}|{nuevo_fecha}" in claves_existentes:
        return wb_existente

    wb = Workbook()
    ws = wb.active
    ws.title = "blancos_acumulados"
    ws.freeze_panes = "A3"

    _escribir_cabecera_doble(ws, cols, grupos)

    # Reescribir filas existentes
    ri = 3
    for fila_vals in filas_existentes:
        ci = 1
        for col in cols:
            if col == "__SEP__":
                _sep_col(ws, ri, ci)
                ci += 1
                continue
            nombre, bg, txt, mono, align = col
            val = fila_vals[ci - 1] if ci - 1 < len(fila_vals) else ""
            # Color estado
            if nombre == "ESTADO":
                bg, txt = _color_estado_blanco(str(val))
            c = ws.cell(row=ri, column=ci, value=val)
            _cel(c, bg, txt, mono=mono, align=align)
            ci += 1
        ws.row_dimensions[ri].height = 18
        ri += 1

    # Agregar nuevo blanco
    msg_raw = str(nuevo_blanco.get("mensaje", "")).strip()
    valores = {
        "MES":     mes,
        "USER_ID": "",
        "NOMBRE":  "",
        "TIPO":    "TE PAGÓ",
        "ORIGEN":  nuevo_blanco.get("origen", ""),
        "DESTINO": nuevo_blanco.get("destino", ""),
        "MONTO":   nuevo_blanco.get("monto_pago", nuevo_blanco.get("monto", "")),
        "MENSAJE": "" if msg_raw.lower() in ("nan", "none", "") else msg_raw,
        "FECHA":   nuevo_blanco.get("fecha", ""),
        "MZ":       "",
        "LOTE":     "",
        "CONCEPTO": "",
        "MOTIVO":   "",
        "ESTADO":   "pendiente",
    }

    ci = 1
    for col in cols:
        if col == "__SEP__":
            _sep_col(ws, ri, ci)
            ci += 1
            continue
        nombre, bg, txt, mono, align = col
        val = valores.get(nombre, "")
        if nombre == "ESTADO":
            bg, txt = _color_estado_blanco(str(val))
        c = ws.cell(row=ri, column=ci, value=val)
        _cel(c, bg, txt, mono=mono, align=align)
        ci += 1
    ws.row_dimensions[ri].height = 18

    _aplicar_anchos(ws, cols, anchos)
    return wb

def _color_estado_blanco(estado: str):
    m = {"pendiente": (EST_PEND, EST_PEND_T),
         "identificado": (EST_IDEN, EST_IDEN_T),
         "aplicado": (EST_APLI, EST_APLI_T)}
    return m.get(estado.lower(), (C_EST_H, C_EST_T))

# ========================EXPORTAR devoluciones_acumulados===
def agregar_devoluciones_acumulados(wb_existente, nuevas: list, mes: str) -> Workbook:
    """
    Agrega registros PAGASTE a devoluciones_acumulados.xlsx
    siguiendo devoluciones_acumulados_diseno.html
    """
    cols = [
        ("MES",            C_MES_C, C_MES_T, True,  "center"),
        "__SEP__",
        ("USER_ID",        C_ID_C,  C_ID_T,  True,  "left"),
        ("NOMBRE",         C_ID_C,  C_ID_T,  False, "left"),
        "__SEP__",
        ("TIPO",           C_BAN_C, C_BAN_T, True,  "left"),
        ("ORIGEN",         C_BAN_C, C_BAN_T, True,  "left"),
        ("DESTINO",        C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO_PAGO",     C_BAN_C, MONTO_T, True,  "right"),
        ("MONTO_ASIGNADO", C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE",        C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",          C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MZ",      C_LOT_C, C_LOT_T, True,  "center"),
        ("LOTE",    C_LOT_C, C_LOT_T, True,  "center"),
        ("CONCEPTO",C_LOT_C, C_LOT_T, False, "left"),
        "__SEP__",
        ("ESTADO",  C_EST_H, C_EST_T, True,  "center"),
    ]

    grupos = [
        ("¿DE QUÉ MES?",         1, C_MES_H, C_MES_T),
        None,
        ("¿QUIÉN ES?",           2, C_ID_H,  C_ID_T),
        None,
        ("¿QUÉ HIZO EL BANCO?",  7, C_BAN_H, C_BAN_T),
        None,
        ("¿A QUÉ LOTE?",         3, C_LOT_H, C_LOT_T),
        None,
        ("¿YA SE APLICÓ?",       1, C_EST_H, C_EST_T),
    ]

    anchos = {
        "MES":10, "USER_ID":10, "NOMBRE":28,
        "TIPO":10, "ORIGEN":22, "DESTINO":22,
        "MONTO_PAGO":12, "MONTO_ASIGNADO":14,
        "MENSAJE":30, "FECHA":20,
        "MZ":6, "LOTE":7, "CONCEPTO":25, "ESTADO":12,
    }

    filas_existentes = []
    if wb_existente is not None:
        ws_old = wb_existente.active
        rows = list(ws_old.values)
        for fila in rows[2:]:
            if fila and any(fila):
                filas_existentes.append(list(fila))

    wb = Workbook()
    ws = wb.active
    ws.title = "devoluciones_acumulados"
    ws.freeze_panes = "A3"

    _escribir_cabecera_doble(ws, cols, grupos)

    ri = 3
    # Reescribir existentes
    for fila_vals in filas_existentes:
        ci = 1
        for col in cols:
            if col == "__SEP__":
                _sep_col(ws, ri, ci)
                ci += 1
                continue
            nombre, bg, txt, mono, align = col
            val = fila_vals[ci - 1] if ci - 1 < len(fila_vals) else ""
            if nombre == "ESTADO":
                bg, txt = _color_estado_dev(str(val))
            c = ws.cell(row=ri, column=ci, value=val)
            _cel(c, bg, txt, mono=mono, align=align)
            ci += 1
        ws.row_dimensions[ri].height = 18
        ri += 1

    # Agregar nuevas devoluciones
    for reg in nuevas:
        valores = {
            "MES":            mes,
            "USER_ID":        reg.get("user_id", ""),
            "NOMBRE":         reg.get("nombre", ""),
            "TIPO":           "PAGASTE",
            "ORIGEN":         reg.get("origen", ""),
            "DESTINO":        reg.get("destino", ""),
            "MONTO_PAGO":     reg.get("monto_pago", ""),
            "MONTO_ASIGNADO": reg.get("monto_asignado", ""),
            "MENSAJE":        reg.get("mensaje", ""),
            "FECHA":          reg.get("fecha", ""),
            "MZ":             reg.get("mz", ""),
            "LOTE":           reg.get("lote", ""),
            "CONCEPTO":       reg.get("concepto", ""),
            "ESTADO":         "pendiente",
        }
        ci = 1
        for col in cols:
            if col == "__SEP__":
                _sep_col(ws, ri, ci)
                ci += 1
                continue
            nombre, bg, txt, mono, align = col
            val = valores.get(nombre, "")
            if nombre == "ESTADO":
                bg, txt = _color_estado_dev(str(val))
            c = ws.cell(row=ri, column=ci, value=val)
            _cel(c, bg, txt, mono=mono, align=align)
            ci += 1
        ws.row_dimensions[ri].height = 18
        ri += 1

    _aplicar_anchos(ws, cols, anchos)
    return wb

def _color_estado_dev(estado: str):
    m = {"pendiente": (EST_PEND, EST_PEND_T),
         "aplicado":  (EST_APLI, EST_APLI_T),
         "no aplica": (EST_NOAP, EST_NOAP_T)}
    return m.get(estado.lower(), (C_EST_H, C_EST_T))

# ========================EXPORTAR pagos_yape_pagaste========
def exportar_pagos_pagaste(pagaste: list, ciclo: int) -> Workbook:
    """
    Exporta pagos_yape_pagaste.xlsx siguiendo pagos_yape_pagaste_diseno.html
    Narrativa: ¿Qué hizo el banco? → ¿A qué lote? → ¿Para qué? → ¿Cuándo?

    MZ+LOTE llenos → es devolución a un usuario (también irá a pagos_yape_devolucion).
    CONCEPTO lleno → es gasto JASS (queda solo aquí).
    """
    cols = [
        ("TIPO",    C_BAN_C, C_BAN_T, True,  "left"),
        ("ORIGEN",  C_BAN_C, C_BAN_T, True,  "left"),
        ("DESTINO", C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO",   C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE", C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",   C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MZ",       C_UBI_C, C_UBI_T, True,  "center"),
        ("LOTE",     C_UBI_C, C_UBI_T, True,  "center"),
        "__SEP__",
        ("CONCEPTO", C_DED_C, C_DED_T, False, "left"),
        ("MOTIVO",   C_DED_C, C_DED_T, False, "left"),
        "__SEP__",
        ("CICLO_CORRECCION", C_CIC_C, C_CIC_T, True, "center"),
    ]
    grupos = [
        ("¿QUÉ HIZO EL BANCO?", 6, C_BAN_H, C_BAN_T),
        None,
        ("¿A QUÉ LOTE?",        2, C_UBI_H, C_UBI_T),
        None,
        ("¿PARA QUÉ?",          2, C_DED_H, C_DED_T),
        None,
        ("¿CUÁNDO?",            1, C_CIC_H, C_CIC_T),
    ]
    anchos = {
        "TIPO":10, "ORIGEN":28, "DESTINO":22,
        "MONTO":12, "MENSAJE":35, "FECHA":20,
        "MZ":6, "LOTE":6,
        "CONCEPTO":30, "MOTIVO":25,
        "CICLO_CORRECCION":10,
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "pagos_yape_pagaste"
    ws.freeze_panes = "A3"

    _escribir_cabecera_doble(ws, cols, grupos)

    for ri, reg in enumerate(pagaste, start=3):
        valores = {
            "TIPO":             "PAGASTE",
            "ORIGEN":           reg.get("origen",   ""),
            "DESTINO":          reg.get("destino",  ""),
            "MONTO":            reg.get("monto",    ""),
            "MENSAJE":          reg.get("mensaje",  ""),
            "FECHA":            reg.get("fecha",    ""),
            "MZ":               reg.get("mz",       ""),
            "LOTE":             reg.get("lote",     ""),
            "CONCEPTO":         reg.get("concepto", ""),
            "MOTIVO":           reg.get("motivo",   ""),
            "CICLO_CORRECCION": reg.get("ciclo",    ciclo),
        }
        ci = 1
        for col in cols:
            if col == "__SEP__":
                _sep_col(ws, ri, ci)
                ci += 1
                continue
            nombre, bg, txt, mono, align = col
            c = ws.cell(row=ri, column=ci, value=valores.get(nombre, ""))
            _cel(c, bg, txt, mono=mono, align=align)
            ci += 1
        ws.row_dimensions[ri].height = 18

    _aplicar_anchos(ws, cols, anchos)
    return wb


# ========================EXPORTAR pagos_yape_devolucion=====
def exportar_pagos_devolucion(pagaste_confirmados: list, ciclo: int) -> Workbook:
    """
    Exporta pagos_yape_devolucion.xlsx siguiendo pagos_yape_devolucion_diseno.html
    Solo PAGASTE con MZ+LOTE identificados — los gastos JASS (CONCEPTO sin MZ+LOTE) se excluyen.
    Narrativa: ¿Qué hizo el banco? → ¿A quién se devolvió? → ¿Por qué? → ¿Cuándo?
    """
    devoluciones = [
        r for r in pagaste_confirmados
        if str(r.get("mz", "")).strip() and str(r.get("lote", "")).strip()
    ]

    cols = [
        ("TIPO",    C_BAN_C, C_BAN_T, True,  "left"),
        ("ORIGEN",  C_BAN_C, C_BAN_T, True,  "left"),
        ("DESTINO", C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO",   C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE", C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",   C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MZ",      C_UBI_C, C_UBI_T, True,  "center"),
        ("LOTE",    C_UBI_C, C_UBI_T, True,  "center"),
        "__SEP__",
        ("CONCEPTO", C_DED_C, C_DED_T, False, "left"),
        "__SEP__",
        ("CICLO_CORRECCION", C_CIC_C, C_CIC_T, True, "center"),
    ]
    grupos = [
        ("¿QUÉ HIZO EL BANCO?",   6, C_BAN_H, C_BAN_T),
        None,
        ("¿A QUIÉN SE DEVOLVIÓ?", 2, C_UBI_H, C_UBI_T),
        None,
        ("¿POR QUÉ?",             1, C_DED_H, C_DED_T),
        None,
        ("¿CUÁNDO?",              1, C_CIC_H, C_CIC_T),
    ]
    anchos = {
        "TIPO":10, "ORIGEN":28, "DESTINO":22,
        "MONTO":12, "MENSAJE":35, "FECHA":20,
        "MZ":6, "LOTE":6,
        "CONCEPTO":30,
        "CICLO_CORRECCION":10,
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "pagos_yape_devolucion"
    ws.freeze_panes = "A3"

    _escribir_cabecera_doble(ws, cols, grupos)

    for ri, reg in enumerate(devoluciones, start=3):
        valores = {
            "TIPO":             "PAGASTE",
            "ORIGEN":           reg.get("origen",   ""),
            "DESTINO":          reg.get("destino",  ""),
            "MONTO":            reg.get("monto",    ""),
            "MENSAJE":          reg.get("mensaje",  ""),
            "FECHA":            reg.get("fecha",    ""),
            "MZ":               reg.get("mz",       ""),
            "LOTE":             reg.get("lote",     ""),
            "CONCEPTO":         reg.get("concepto", ""),
            "CICLO_CORRECCION": reg.get("ciclo",    ciclo),
        }
        ci = 1
        for col in cols:
            if col == "__SEP__":
                _sep_col(ws, ri, ci)
                ci += 1
                continue
            nombre, bg, txt, mono, align = col
            c = ws.cell(row=ri, column=ci, value=valores.get(nombre, ""))
            _cel(c, bg, txt, mono=mono, align=align)
            ci += 1
        ws.row_dimensions[ri].height = 18

    _aplicar_anchos(ws, cols, anchos)
    return wb


# ========================EXPORTAR reporte_procesado=========
def exportar_reporte_procesado(ruta_tepago: Path, ruta_pagaste: Path) -> Workbook:
    """
    Combina TE_PAGÓ y PAGASTE en un workbook de 2 hojas para el archivo procesado mensual.
    Copia valores y estilos de los archivos de salida ya escritos.
    """
    wb_new = Workbook()

    def _copiar_hoja(ruta_src, wb_dst, titulo, es_primera):
        ws_dst = wb_dst.active if es_primera else wb_dst.create_sheet(titulo)
        ws_dst.title = titulo
        if not ruta_src or not Path(ruta_src).exists():
            return
        wb_src = load_workbook(ruta_src, data_only=True)
        ws_src = wb_src.active
        ws_dst.freeze_panes = ws_src.freeze_panes
        for k, v in ws_src.column_dimensions.items():
            ws_dst.column_dimensions[k].width = v.width
        for k, v in ws_src.row_dimensions.items():
            ws_dst.row_dimensions[k].height = v.height
        for rng in ws_src.merged_cells.ranges:
            ws_dst.merge_cells(str(rng))
        for row in ws_src.iter_rows():
            for cell in row:
                nc = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    nc.font          = _copy(cell.font)
                    nc.border        = _copy(cell.border)
                    nc.fill          = _copy(cell.fill)
                    nc.number_format = cell.number_format
                    nc.protection    = _copy(cell.protection)
                    nc.alignment     = _copy(cell.alignment)
        wb_src.close()

    _copiar_hoja(ruta_tepago,  wb_new, "TE_PAGÓ", True)
    _copiar_hoja(ruta_pagaste, wb_new, "PAGASTE",  False)
    return wb_new


# ========================HELPERS============================
def _escribir_cabecera_doble(ws, cols, grupos):
    """Escribe fila 1 (grupos) y fila 2 (nombres de columnas)."""
    # Fila 1: grupos
    ci = 1
    for g in grupos:
        if g is None:
            # separador
            c = ws.cell(row=1, column=ci, value="")
            _sep(c)
            ws.column_dimensions[get_column_letter(ci)].width = 1
            ci += 1
        else:
            texto, span, bg, txt = g
            if span == 1:
                c = ws.cell(row=1, column=ci, value=texto)
                _hdr(c, bg, txt, size=9)
            else:
                ws.merge_cells(start_row=1, start_column=ci,
                               end_row=1,   end_column=ci + span - 1)
                c = ws.cell(row=1, column=ci, value=texto)
                _hdr(c, bg, txt, size=9)
            ci += span
    ws.row_dimensions[1].height = 20

    # Fila 2: nombres
    ci = 1
    for col in cols:
        if col == "__SEP__":
            c = ws.cell(row=2, column=ci, value="")
            _sep(c)
            ws.column_dimensions[get_column_letter(ci)].width = 1
        else:
            nombre, bg, txt, mono, align = col
            c = ws.cell(row=2, column=ci, value=nombre)
            _hdr(c, bg, txt, size=9)
        ci += 1
    ws.row_dimensions[2].height = 20

def _aplicar_anchos(ws, cols, anchos):
    ci = 1
    for col in cols:
        if col != "__SEP__":
            nombre = col[0]
            ws.column_dimensions[get_column_letter(ci)].width = anchos.get(nombre, 16)
        ci += 1

# ========================EXPORTAR TRAZABILIDAD==============
# Siguiendo trazabilidad.html — 3 hojas: Sin_identificar · Ambiguos · Pagos_multiples

C_RES_H = "E1F5EE"; C_RES_C = "F0FFF8"; C_RES_T = "085041"
C_CAN_H = "FAEEDA"; C_CAN_C = "FFFBF0"; C_CAN_T = "854F0B"
C_CUA_H = "FEF3E8"; C_CUA_C = "FEF3E8"; C_CUA_T = "7C3003"

def exportar_trazabilidad(ruta, corr_simples: dict, validados_ambiguos: list,
                           corr_multiples: dict, ciclo: int, fecha_hoy: str,
                           validados_maestro_inexacto: list = None):
    """
    Genera trazabilidad_YYYY_MM.xlsx con 4 hojas siguiendo trazabilidad.html
    """
    if validados_maestro_inexacto is None:
        validados_maestro_inexacto = []
    from pathlib import Path
    wb = Workbook()

    # ── Hoja 1: Sin_identificar ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Sin_identificar"
    ws1.freeze_panes = "A3"

    cols_si = [
        ("USER_ID", C_ID_C,  C_ID_T,  True,  "left"),
        ("NOMBRE",  C_ID_C,  C_ID_T,  False, "left"),
        "__SEP__",
        ("TIPO",    C_BAN_C, C_BAN_T, True,  "left"),
        ("ORIGEN",  C_BAN_C, C_BAN_T, True,  "left"),
        ("DESTINO", C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO",   C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE", C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",   C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MZ",      C_RES_C, C_RES_T, True,  "center"),
        ("LOTE",    C_RES_C, C_RES_T, True,  "center"),
        ("CONCEPTO",C_RES_C, C_RES_T, False, "left"),
        ("MOTIVO",  C_RES_C, C_RES_T, False, "left"),
        "__SEP__",
        ("CICLO",          C_CUA_C, C_CUA_T, True, "center"),
        ("FECHA_CORRECCION",C_CUA_C, C_CUA_T, True, "left"),
    ]

    grupos_si = [
        ("¿QUIÉN ES?",        2, C_ID_H,  C_ID_T),
        None,
        ("¿QUÉ HIZO EL BANCO?", 6, C_BAN_H, C_BAN_T),
        None,
        ("¿CÓMO SE RESOLVIÓ?", 4, C_RES_H, C_RES_T),
        None,
        ("¿CUÁNDO?",           2, C_CUA_H, C_CUA_T),
    ]

    _escribir_cabecera_doble(ws1, cols_si, grupos_si)

    anchos_si = {
        "USER_ID":10, "NOMBRE":28,
        "TIPO":10, "ORIGEN":28, "DESTINO":22,
        "MONTO":12, "MENSAJE":35, "FECHA":20,
        "MZ":6, "LOTE":7, "CONCEPTO":20, "MOTIVO":30,
        "CICLO":7, "FECHA_CORRECCION":16,
    }

    ri = 3
    for origen, v in corr_simples.items():
        if v.get("_fuente") in ("ambiguo", "maestro_inexacto"):
            continue
        # La clave es "NOMBRE_PAGANTE|FECHA_PAGO" — separar para escribir columnas limpias.
        partes        = origen.rsplit("|", 1)
        origen_nombre = partes[0]
        fecha_pago    = partes[1] if len(partes) > 1 else ""
        valores = {
            "USER_ID":           v.get("user_id", ""),
            "NOMBRE":            v.get("nombre", ""),
            "TIPO":              v.get("tipo", "TE PAGÓ"),
            "ORIGEN":            origen_nombre,
            "DESTINO":           v.get("destino", ""),
            "MONTO":             v.get("monto", ""),
            "MENSAJE":           v.get("mensaje", ""),
            "FECHA":             fecha_pago,
            "MZ":                v.get("mz", ""),
            "LOTE":              v.get("lote", ""),
            "CONCEPTO":          v.get("concepto", ""),
            "MOTIVO":            v.get("motivo", ""),
            "CICLO":             ciclo,
            "FECHA_CORRECCION":  fecha_hoy,
        }
        ci = 1
        for col in cols_si:
            if col == "__SEP__":
                _sep_col(ws1, ri, ci)
                ci += 1
                continue
            nombre, bg, txt, mono, align = col
            c = ws1.cell(row=ri, column=ci, value=valores.get(nombre, ""))
            _cel(c, bg, txt, mono=mono, align=align)
            ci += 1
        ws1.row_dimensions[ri].height = 18
        ri += 1

    _aplicar_anchos(ws1, cols_si, anchos_si)

    # ── Hoja 2: Ambiguos ─────────────────────────────────────
    # Una fila por pago · CANDIDATOS = texto MZ-LOTE(deuda,diff) · siguiendo trazabilidad.html
    ws2 = wb.create_sheet("Ambiguos")
    ws2.freeze_panes = "A3"

    cols_amb = [
        ("USER_ID",          C_ID_C,  C_ID_T,  True,  "left"),
        ("NOMBRE",           C_ID_C,  C_ID_T,  False, "left"),
        "__SEP__",
        ("ORIGEN",           C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO",            C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE",          C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",            C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("CANDIDATOS",       C_CAN_C, C_CAN_T, False, "left"),
        "__SEP__",
        ("MZ_FINAL",         C_RES_C, C_RES_T, True,  "center"),
        ("LOTE_FINAL",       C_RES_C, C_RES_T, True,  "center"),
        "__SEP__",
        ("CICLO",            C_CUA_C, C_CUA_T, True,  "center"),
        ("FECHA_CORRECCION", C_CUA_C, C_CUA_T, True,  "left"),
    ]

    grupos_amb = [
        ("¿QUIÉN ES?",          2, C_ID_H,  C_ID_T),
        None,
        ("¿QUÉ HIZO EL BANCO?", 4, C_BAN_H, C_BAN_T),
        None,
        ("CANDIDATOS",          1, C_CAN_H, C_CAN_T),
        None,
        ("ELEGIDO",             2, C_RES_H, C_RES_T),
        None,
        ("¿CUÁNDO?",            2, C_CUA_H, C_CUA_T),
    ]

    _escribir_cabecera_doble(ws2, cols_amb, grupos_amb)

    anchos_amb = {
        "USER_ID":10, "NOMBRE":26,
        "ORIGEN":30, "MONTO":12, "MENSAJE":28, "FECHA":20,
        "CANDIDATOS":35,
        "MZ_FINAL":10, "LOTE_FINAL":10,
        "CICLO":7, "FECHA_CORRECCION":16,
    }

    ri = 3
    for reg in validados_ambiguos:
        valores = {
            "USER_ID":          reg.get("user_id", ""),
            "NOMBRE":           reg.get("nombre", ""),
            "ORIGEN":           reg.get("origen", ""),
            "MONTO":            reg.get("monto", ""),
            "MENSAJE":          reg.get("mensaje", ""),
            "FECHA":            reg.get("fecha", ""),
            "CANDIDATOS":       reg.get("candidatos", ""),
            "MZ_FINAL":         reg.get("mz_final", ""),
            "LOTE_FINAL":       reg.get("lote_final", ""),
            "CICLO":            ciclo,
            "FECHA_CORRECCION": fecha_hoy,
        }
        ci = 1
        for col in cols_amb:
            if col == "__SEP__":
                _sep_col(ws2, ri, ci)
                ci += 1
                continue
            nombre, bg, txt, mono, align = col
            c = ws2.cell(row=ri, column=ci, value=valores.get(nombre, ""))
            _cel(c, bg, txt, mono=mono, align=align)
            ci += 1
        ws2.row_dimensions[ri].height = 18
        ri += 1

    _aplicar_anchos(ws2, cols_amb, anchos_amb)

    # ── Hoja 3: Pagos_multiples ──────────────────────────────
    ws3 = wb.create_sheet("Pagos_multiples")
    ws3.freeze_panes = "A3"

    cols_mul = [
        ("USER_ID", C_ID_C,  C_ID_T,  True,  "left"),
        ("NOMBRE",  C_ID_C,  C_ID_T,  False, "left"),
        "__SEP__",
        ("ORIGEN",       C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO_TOTAL",  C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE",      C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",        C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MZ",             C_RES_C, C_RES_T, True, "center"),
        ("LOTE",           C_RES_C, C_RES_T, True, "center"),
        ("DEUDA",          C_RES_C, C_RES_T, True, "right"),
        ("MONTO_ASIGNADO", C_RES_C, MONTO_T, True, "right"),
        ("DIFF",           C_RES_C, C_RES_T, True, "right"),
        "__SEP__",
        ("CICLO",           C_CUA_C, C_CUA_T, True, "center"),
        ("FECHA_CORRECCION",C_CUA_C, C_CUA_T, True, "left"),
    ]

    grupos_mul = [
        ("¿QUIÉN ES?",          2, C_ID_H,  C_ID_T),
        None,
        ("¿QUÉ HIZO EL BANCO?", 4, C_BAN_H, C_BAN_T),
        None,
        ("ASIGNACIÓN POR LOTE", 5, C_RES_H, C_RES_T),
        None,
        ("¿CUÁNDO?",            2, C_CUA_H, C_CUA_T),
    ]

    _escribir_cabecera_doble(ws3, cols_mul, grupos_mul)

    anchos_mul = {
        "USER_ID":10, "NOMBRE":24,
        "ORIGEN":28, "MONTO_TOTAL":12, "MENSAJE":30, "FECHA":20,
        "MZ":6, "LOTE":7, "DEUDA":12,
        "MONTO_ASIGNADO":14, "DIFF":10,
        "CICLO":7, "FECHA_CORRECCION":16,
    }

    ri = 3
    for origen, items in corr_multiples.items():
        monto_total = sum(i.get("monto", 0) for i in items)
        for item in items:
            valores = {
                "USER_ID":          item.get("user_id", ""),
                "NOMBRE":           item.get("nombre", ""),
                "ORIGEN":           origen,
                "MONTO_TOTAL":      monto_total,
                "MENSAJE":          item.get("mensaje", ""),
                "FECHA":            item.get("fecha", ""),
                "MZ":               item.get("mz", ""),
                "LOTE":             item.get("lote", ""),
                "DEUDA":            item.get("deuda", ""),
                "MONTO_ASIGNADO":   item.get("monto", ""),
                "DIFF":             item.get("diff", ""),
                "CICLO":            ciclo,
                "FECHA_CORRECCION": fecha_hoy,
            }
            ci = 1
            for col in cols_mul:
                if col == "__SEP__":
                    _sep_col(ws3, ri, ci)
                    ci += 1
                    continue
                nombre, bg, txt, mono, align = col
                c = ws3.cell(row=ri, column=ci, value=valores.get(nombre, ""))
                _cel(c, bg, txt, mono=mono, align=align)
                ci += 1
            ws3.row_dimensions[ri].height = 18
            ri += 1

    _aplicar_anchos(ws3, cols_mul, anchos_mul)

    # ── Hoja 4: Maestro_inexacto ─────────────────────────────
    ws4 = wb.create_sheet("Maestro_inexacto")
    ws4.freeze_panes = "A3"

    cols_mix_traz = [
        ("USER_ID",          C_ID_C,  C_ID_T,  True,  "left"),
        ("NOMBRE",           C_ID_C,  C_ID_T,  False, "left"),
        "__SEP__",
        ("TIPO",             C_BAN_C, C_BAN_T, True,  "left"),
        ("ORIGEN",           C_BAN_C, C_BAN_T, True,  "left"),
        ("DESTINO",          C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO",            C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE",          C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",            C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MZ",               C_RES_C, C_RES_T, True,  "center"),
        ("LOTE",             C_RES_C, C_RES_T, True,  "center"),
        ("CONCEPTO",         C_RES_C, C_RES_T, False, "left"),
        ("DIFF",             C_RES_C, C_RES_T, True,  "right"),
        ("MOTIVO",           C_RES_C, C_RES_T, False, "left"),
        "__SEP__",
        ("CICLO",            C_CUA_C, C_CUA_T, True,  "center"),
        ("FECHA_CORRECCION", C_CUA_C, C_CUA_T, True,  "left"),
    ]

    grupos_mix_traz = [
        ("¿QUIÉN ES?",          2, C_ID_H,  C_ID_T),
        None,
        ("¿QUÉ HIZO EL BANCO?", 6, C_BAN_H, C_BAN_T),
        None,
        ("¿CÓMO SE RESOLVIÓ?",  5, C_RES_H, C_RES_T),
        None,
        ("¿CUÁNDO?",            2, C_CUA_H, C_CUA_T),
    ]

    _escribir_cabecera_doble(ws4, cols_mix_traz, grupos_mix_traz)

    anchos_mix_traz = {
        "USER_ID":10, "NOMBRE":28,
        "TIPO":10, "ORIGEN":28, "DESTINO":22,
        "MONTO":12, "MENSAJE":35, "FECHA":20,
        "MZ":6, "LOTE":7, "CONCEPTO":20, "DIFF":10, "MOTIVO":30,
        "CICLO":7, "FECHA_CORRECCION":16,
    }

    ri = 3
    for reg in validados_maestro_inexacto:
        diff_val = reg.get("diff", "")
        valores = {
            "USER_ID":           reg.get("user_id", ""),
            "NOMBRE":            reg.get("nombre", ""),
            "TIPO":              reg.get("tipo", "TE PAGÓ"),
            "ORIGEN":            reg.get("origen", ""),
            "DESTINO":           reg.get("destino", ""),
            "MONTO":             reg.get("monto", ""),
            "MENSAJE":           reg.get("mensaje", ""),
            "FECHA":             reg.get("fecha", ""),
            "MZ":                reg.get("mz", ""),
            "LOTE":              reg.get("lote", ""),
            "CONCEPTO":          reg.get("concepto", ""),
            "DIFF":              diff_val,
            "MOTIVO":            reg.get("motivo", ""),
            "CICLO":             ciclo,
            "FECHA_CORRECCION":  fecha_hoy,
        }
        ci = 1
        for col in cols_mix_traz:
            if col == "__SEP__":
                _sep_col(ws4, ri, ci)
                ci += 1
                continue
            nombre, bg, txt, mono, align = col
            val = valores.get(nombre, "")
            if nombre == "DIFF" and isinstance(val, (int, float)):
                txt = "065F46" if val >= 0 else "A32D2D"
            _cel(ws4.cell(row=ri, column=ci, value=val), bg, txt, mono=mono, align=align)
            ci += 1
        ws4.row_dimensions[ri].height = 18
        ri += 1

    _aplicar_anchos(ws4, cols_mix_traz, anchos_mix_traz)

    wb.save(ruta)
    print(f"  ✔ trazabilidad guardada: {Path(ruta).name}")
