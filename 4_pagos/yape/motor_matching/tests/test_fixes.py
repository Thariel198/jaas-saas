"""
Tests unitarios para los 3 fixes del motor matching.

Fix 1 — _leer_validados_ambiguos:
    · OK=SI  → pasa a validados (trazabilidad Ciclo 2+)
    · OK=NO  → excluido
    · sin hoja Ambiguos en workbook (simula Ciclo 1 sin pendientes.xlsx) → lista vacía

Fix 2 — _leer_validados_maestro_inexacto:
    · OK=SI  → pasa a validados
    · OK=NO  → excluido
    · sin hoja Maestro_inexacto (Ciclo 1) → lista vacía
    · DIFF = monto - deuda(MZ_final, LOTE_final), no de MZ_SUG/LOTE_SUG

Fix 3 — integración Ciclo 1:
    · si pendientes.xlsx no existe, leer_correcciones devuelve listas vacías

Fix 4 — resolución USER_ID en corr_multiples:
    · buscar_uid se invoca cuando user_id está vacío
    · NO se invoca si user_id ya está presente
    · múltiples lotes se resuelven de forma independiente
"""
import sys
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent))
import main


# ── helpers para construir workbooks de prueba ────────────────────────────────

def _wb_ambiguos(data_rows: list) -> Workbook:
    """Workbook con hoja Ambiguos en formato cabecera-doble (como pendientes.xlsx)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ambiguos"
    ws.append(["GRUPOS"])          # fila 0 — grupos, ignorada por la función
    ws.append([                    # fila 1 — cabecera de columnas
        "USER_ID", "NOMBRE", "",
        "TIPO", "ORIGEN", "DESTINO", "MONTO", "MENSAJE", "FECHA", "",
        "MZ_SUG", "LOTE_SUG", "CANDIDATOS", "",
        "MZ", "LOTE", "CONCEPTO", "MOTIVO", "",
        "OK",
    ])
    for row in data_rows:
        ws.append(row)
    return wb


def _wb_maestro_inexacto(data_rows: list) -> Workbook:
    """Workbook con hoja Maestro_inexacto en formato cabecera-doble."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Maestro_inexacto"
    ws.append(["GRUPOS"])
    ws.append([
        "USER_ID", "NOMBRE", "",
        "TIPO", "ORIGEN", "DESTINO", "MONTO", "MENSAJE", "FECHA", "",
        "MZ_SUG", "LOTE_SUG", "DEUDA", "DIFF", "",
        "MZ", "LOTE", "CONCEPTO", "MOTIVO", "",
        "OK",
    ])
    for row in data_rows:
        ws.append(row)
    return wb


def _fila_ambiguo(ok="SI", mz_final="D", lote_final="1") -> list:
    """Fila de datos para hoja Ambiguos. Posiciones alineadas con la cabecera."""
    return [
        "U001", "PEDRO PAZ", "",
        "TE PAGÓ", "Pedro P*", "Janet V*", 38.0, "—", "02/05/2026 23:02", "",
        "D", "1", "D-1(38,0) · D-19(47,-9)", "",
        mz_final, lote_final, "", "", "",
        ok,
    ]


def _fila_maestro_inexacto(ok="SI", mz_final="Q", lote_final="12",
                            monto=30.0, mz_sug="Q", lote_sug="12",
                            concepto="") -> list:
    """Fila de datos para hoja Maestro_inexacto. Posiciones alineadas con la cabecera."""
    return [
        "U002", "JUANA MORALES", "",
        "TE PAGÓ", "Jessica M*", "Janet V*", monto, "—", "04/05/2026 08:33", "",
        mz_sug, lote_sug, 43.0, -13.0, "",
        mz_final, lote_final, concepto, "confirmó pago parcial", "",
        ok,
    ]


# ── Fix 1: _leer_validados_ambiguos ──────────────────────────────────────────

class TestLeerValidadosAmbiguos(unittest.TestCase):

    def test_ok_si_pasa_a_trazabilidad(self):
        """Registro con OK=SI queda en la lista de validados."""
        wb = _wb_ambiguos([_fila_ambiguo(ok="SI")])
        result = main._leer_validados_ambiguos(wb)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["mz_final"], "D")
        self.assertEqual(result[0]["lote_final"], "1")
        self.assertEqual(result[0]["candidatos"], "D-1(38,0) · D-19(47,-9)")
        self.assertEqual(result[0]["origen"], "Pedro P*")

    def test_ok_no_excluido(self):
        """Registro con OK=NO queda excluido."""
        wb = _wb_ambiguos([_fila_ambiguo(ok="NO")])
        result = main._leer_validados_ambiguos(wb)
        self.assertEqual(result, [])

    def test_ok_vacio_excluido(self):
        """Registro con OK en blanco queda excluido."""
        wb = _wb_ambiguos([_fila_ambiguo(ok="")])
        result = main._leer_validados_ambiguos(wb)
        self.assertEqual(result, [])

    def test_sin_hoja_ambiguos_lista_vacia(self):
        """Sin hoja Ambiguos (Ciclo 1 sin pendientes.xlsx) → lista vacía."""
        wb = Workbook()
        wb.active.title = "Sin_identificar"
        result = main._leer_validados_ambiguos(wb)
        self.assertEqual(result, [])

    def test_mezcla_solo_ok_si_pasan(self):
        """De dos filas, solo la con OK=SI aparece en el resultado."""
        wb = _wb_ambiguos([
            _fila_ambiguo(ok="SI", mz_final="D", lote_final="1"),
            _fila_ambiguo(ok="NO", mz_final="D", lote_final="19"),
        ])
        result = main._leer_validados_ambiguos(wb)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["lote_final"], "1")


# ── Fix 2: _leer_validados_maestro_inexacto ──────────────────────────────────

class TestLeerValidadosMaestroInexacto(unittest.TestCase):

    PLANILLA = {
        ("Q", "12"): {"deuda_total": 43.0, "nombre": "JUANA MORALES", "mes_anterior": 0.0},
        ("B",  "2"): {"deuda_total": 30.0, "nombre": "JUAN PAZ",      "mes_anterior": 0.0},
        ("A",  "1"): {"deuda_total": 50.0, "nombre": "PEDRO SOL",     "mes_anterior": 0.0},
    }

    def test_ok_si_pasa_a_trazabilidad(self):
        """Registro con OK=SI queda en la lista de validados."""
        wb = _wb_maestro_inexacto([_fila_maestro_inexacto(ok="SI")])
        result = main._leer_validados_maestro_inexacto(wb, self.PLANILLA)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["mz"], "Q")
        self.assertEqual(result[0]["lote"], "12")
        self.assertEqual(result[0]["motivo"], "confirmó pago parcial")

    def test_ok_no_excluido(self):
        """Registro con OK=NO queda excluido."""
        wb = _wb_maestro_inexacto([_fila_maestro_inexacto(ok="NO")])
        result = main._leer_validados_maestro_inexacto(wb, self.PLANILLA)
        self.assertEqual(result, [])

    def test_ok_vacio_excluido(self):
        """Registro con OK en blanco queda excluido."""
        wb = _wb_maestro_inexacto([_fila_maestro_inexacto(ok="")])
        result = main._leer_validados_maestro_inexacto(wb, self.PLANILLA)
        self.assertEqual(result, [])

    def test_sin_hoja_maestro_inexacto_lista_vacia(self):
        """Sin hoja Maestro_inexacto (Ciclo 1 sin pendientes.xlsx) → lista vacía."""
        wb = Workbook()
        wb.active.title = "Sin_identificar"
        result = main._leer_validados_maestro_inexacto(wb, {})
        self.assertEqual(result, [])

    def test_diff_usa_mz_lote_final_no_sugerencia(self):
        """DIFF = monto - deuda(MZ_final, LOTE_final), no de MZ_SUG/LOTE_SUG.

        MZ_SUG=A, LOTE_SUG=1, deuda=50  →  diff_incorrecto = 35 - 50 = -15
        MZ=B,     LOTE=2,     deuda=30  →  diff_correcto   = 35 - 30 = +5
        """
        wb = _wb_maestro_inexacto([
            _fila_maestro_inexacto(
                ok="SI",
                mz_sug="A", lote_sug="1",    # sugerencia original (no debe usarse para DIFF)
                mz_final="B", lote_final="2", # elección final del usuario
                monto=35.0,
            )
        ])
        result = main._leer_validados_maestro_inexacto(wb, self.PLANILLA)
        self.assertEqual(len(result), 1)
        self.assertAlmostEqual(result[0]["diff"], 5.0)  # 35 - 30, no 35 - 50


# ── Fix 3: integración Ciclo 1 ───────────────────────────────────────────────

class TestLeerCorreccionesCiclo1(unittest.TestCase):

    def test_sin_pendientes_xlsx_validados_quedan_vacios(self):
        """Ciclo 1: si pendientes.xlsx no existe, validados_ambiguos y
        validados_maestro_inexacto son listas vacías."""
        import tempfile
        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("main.CORRECCIONES_DIR", Path(tmpdir)), \
                 patch("main.BASE_DIR", Path(tmpdir)):
                _, _, _, validados_amb, validados_mix, _ = main.leer_correcciones({})
        self.assertEqual(validados_amb, [])
        self.assertEqual(validados_mix, [])


# ── Fix 4: resolución USER_ID en corr_multiples ──────────────────────────────

class TestCorrMultiplesUserIdResolucion(unittest.TestCase):

    @staticmethod
    def _resolver(corr_multiples: dict) -> None:
        """Replica el loop de main() paso 11 que resuelve USER_ID."""
        for items in corr_multiples.values():
            for item in items:
                if not item.get("user_id") and item.get("mz") and item.get("lote"):
                    uid, nom = main.buscar_uid(item["mz"], item["lote"])
                    item["user_id"] = uid
                    if not item.get("nombre"):
                        item["nombre"] = nom

    def test_buscar_uid_se_llama_con_user_id_vacio(self):
        """buscar_uid se invoca y rellena user_id cuando está vacío."""
        corr_multiples = {
            "ORIGEN1": [{"mz": "E", "lote": "7", "monto": 17.0, "user_id": "", "nombre": ""}]
        }
        with patch("main.buscar_uid", return_value=("U370", "WILLIAM TRUJILLO")) as mock_uid:
            self._resolver(corr_multiples)
        mock_uid.assert_called_once_with("E", "7")
        self.assertEqual(corr_multiples["ORIGEN1"][0]["user_id"], "U370")
        self.assertEqual(corr_multiples["ORIGEN1"][0]["nombre"], "WILLIAM TRUJILLO")

    def test_buscar_uid_no_se_llama_si_user_id_presente(self):
        """buscar_uid NO se invoca si el ítem ya tiene user_id."""
        corr_multiples = {
            "ORIGEN1": [{"mz": "E", "lote": "7", "monto": 17.0, "user_id": "U370", "nombre": "YA TIENE"}]
        }
        with patch("main.buscar_uid", return_value=("OTRO", "OTRO")) as mock_uid:
            self._resolver(corr_multiples)
        mock_uid.assert_not_called()
        self.assertEqual(corr_multiples["ORIGEN1"][0]["user_id"], "U370")

    def test_multiples_lotes_resuelven_independientemente(self):
        """Cada lote de un pago múltiple se resuelve por separado."""
        corr_multiples = {
            "ORIGEN1": [
                {"mz": "E", "lote": "7",  "monto": 17.0, "user_id": "", "nombre": ""},
                {"mz": "P", "lote": "11", "monto": 13.0, "user_id": "", "nombre": ""},
            ]
        }

        def _side(mz, lote):
            return {"E": ("U370", "WILLIAM"), "P": ("U374", "PEDRO")}.get(mz, ("", ""))

        with patch("main.buscar_uid", side_effect=_side) as mock_uid:
            self._resolver(corr_multiples)

        self.assertEqual(mock_uid.call_count, 2)
        self.assertEqual(corr_multiples["ORIGEN1"][0]["user_id"], "U370")
        self.assertEqual(corr_multiples["ORIGEN1"][1]["user_id"], "U374")


# ── Fix 5: clave compuesta ORIGEN|FECHA en _leer_hoja_acumuladas ─────────────

def _wb_trazabilidad_si(filas: list) -> Workbook:
    """Workbook con hoja Sin_identificar en formato trazabilidad (doble cabecera)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sin_identificar"
    ws.append(["¿QUIÉN ES?", "", "", "¿QUÉ HIZO EL BANCO?", "", "", "", "", "", "¿CÓMO SE RESOLVIÓ?", "", "", "", "", "¿CUÁNDO?", ""])
    ws.append(["USER_ID", "NOMBRE", "", "TIPO", "ORIGEN", "DESTINO", "MONTO", "MENSAJE", "FECHA", "", "MZ", "LOTE", "CONCEPTO", "MOTIVO", "", "CICLO", "FECHA_CORRECCION"])
    for fila in filas:
        ws.append(fila)
    return wb


def _fila_traz(origen, fecha, mz="A", lote="1"):
    return ["U001", "NOMBRE TEST", "", "TE PAGÓ", origen, "", 30.0, "—", fecha, "", mz, lote, "", "", "", 1, "24/05/2026 10:00"]


class TestClaveCompuestaAcumuladas(unittest.TestCase):

    def test_mismo_origen_distinta_fecha_dos_entradas(self):
        """Mismo pagante, fechas distintas → 2 entradas separadas en corr_simples."""
        wb = _wb_trazabilidad_si([
            _fila_traz("Pedro P*", "02/05/2026 23:02:11", "A", "1"),
            _fila_traz("Pedro P*", "15/05/2026 08:30:45", "B", "3"),
        ])
        result = main._leer_hoja_acumuladas(wb, "Sin_identificar", tiene_instruccion=True)
        self.assertEqual(len(result), 2)
        self.assertIn("PEDRO P*|02/05/2026 23:02:11", result)
        self.assertIn("PEDRO P*|15/05/2026 08:30:45", result)
        self.assertEqual(result["PEDRO P*|02/05/2026 23:02:11"]["mz"], "A")
        self.assertEqual(result["PEDRO P*|15/05/2026 08:30:45"]["mz"], "B")

    def test_mismo_origen_misma_fecha_una_entrada(self):
        """Mismo pago duplicado en el archivo → 1 sola entrada, último gana."""
        wb = _wb_trazabilidad_si([
            _fila_traz("Pedro P*", "02/05/2026 23:02:11", "A", "1"),
            _fila_traz("Pedro P*", "02/05/2026 23:02:11", "B", "3"),
        ])
        result = main._leer_hoja_acumuladas(wb, "Sin_identificar", tiene_instruccion=True)
        self.assertEqual(len(result), 1)
        self.assertEqual(result["PEDRO P*|02/05/2026 23:02:11"]["mz"], "B")

    def test_tiene_instruccion_false_falla_con_doble_cabecera(self):
        """Con tiene_instruccion=False sobre doble cabecera → {} (headers son grupos, no columnas)."""
        wb = _wb_trazabilidad_si([_fila_traz("Pedro P*", "02/05/2026 23:02:11")])
        result = main._leer_hoja_acumuladas(wb, "Sin_identificar", tiene_instruccion=False)
        self.assertEqual(result, {})

    def test_pagantes_distintos_misma_fecha_dos_entradas(self):
        """Dos pagantes distintos con misma fecha → claves distintas, no se pisan."""
        wb = _wb_trazabilidad_si([
            _fila_traz("Ana R*",   "02/05/2026 10:00:00", "A", "1"),
            _fila_traz("Pedro P*", "02/05/2026 10:00:00", "B", "3"),
        ])
        result = main._leer_hoja_acumuladas(wb, "Sin_identificar", tiene_instruccion=True)
        self.assertEqual(len(result), 2)


# ── Fix 6: clave compuesta ORIGEN|FECHA en _leer_hoja_pendientes_ok ──────────

def _wb_pendientes_si_ok(filas: list) -> Workbook:
    """Workbook con hoja Sin_identificar en formato pendientes (doble cabecera con grupos)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sin_identificar"
    ws.append(["← BANCO — NO EDITAR", "", "", "", "", "", "← TÚ COMPLETAS", "", "", "", "✓"])
    ws.append(["TIPO", "ORIGEN", "DESTINO", "MONTO", "MENSAJE", "FECHA", "", "MZ", "LOTE", "CONCEPTO", "MOTIVO", "", "OK"])
    for fila in filas:
        ws.append(fila)
    return wb


def _fila_pend_ok(origen, fecha, mz="A", lote="1", ok="SI"):
    return ["TE PAGÓ", origen, "", 30.0, "—", fecha, "", mz, lote, "", "", "", ok]


class TestClaveCompuestaPendientesOk(unittest.TestCase):

    def test_mismo_origen_distinta_fecha_dos_entradas_ok_si(self):
        """Mismo pagante paga dos veces, ambos OK=SI → 2 correcciones separadas."""
        wb = _wb_pendientes_si_ok([
            _fila_pend_ok("Pedro P*", "02/05/2026 23:02:11", "A", "1"),
            _fila_pend_ok("Pedro P*", "15/05/2026 08:30:45", "B", "3"),
        ])
        result = main._leer_hoja_pendientes_ok(wb, "Sin_identificar")
        self.assertEqual(len(result), 2)
        self.assertIn("PEDRO P*|02/05/2026 23:02:11", result)
        self.assertIn("PEDRO P*|15/05/2026 08:30:45", result)

    def test_mismo_origen_misma_fecha_ok_no_excluido(self):
        """OK=NO queda excluido aunque mismo origen que un OK=SI."""
        wb = _wb_pendientes_si_ok([
            _fila_pend_ok("Pedro P*", "02/05/2026 23:02:11", "A", "1", ok="SI"),
            _fila_pend_ok("Pedro P*", "15/05/2026 08:30:45", "B", "3", ok="NO"),
        ])
        result = main._leer_hoja_pendientes_ok(wb, "Sin_identificar")
        self.assertEqual(len(result), 1)
        self.assertIn("PEDRO P*|02/05/2026 23:02:11", result)


# ── Fix 7: CONCEPTO leído del workbook en _leer_validados_maestro_inexacto ───

class TestMaestroInexactoConceptoLeido(unittest.TestCase):

    PLANILLA = {("Q", "12"): {"deuda_total": 43.0, "nombre": "JUANA MORALES", "mes_anterior": 0.0}}

    def test_concepto_lleno_se_lee(self):
        """CONCEPTO relleno en pendientes → aparece en el resultado."""
        wb = _wb_maestro_inexacto([
            _fila_maestro_inexacto(ok="SI", mz_final="Q", lote_final="12",
                                   concepto="devolución agua")
        ])
        result = main._leer_validados_maestro_inexacto(wb, self.PLANILLA)
        self.assertEqual(result[0]["concepto"], "devolución agua")

    def test_concepto_vacio_devuelve_cadena_vacia(self):
        """CONCEPTO en blanco → "" en el resultado (no None ni 'None')."""
        wb = _wb_maestro_inexacto([_fila_maestro_inexacto(ok="SI")])
        result = main._leer_validados_maestro_inexacto(wb, self.PLANILLA)
        self.assertEqual(result[0]["concepto"], "")


# ── Fix 8: round-trip trazabilidad — sin pérdida ni duplicación ──────────────

class TestTrazabilidadRoundTrip(unittest.TestCase):
    """
    Verifica que exportar_trazabilidad → leer_correcciones forme un ciclo sin pérdida.
    Clave: la clave compuesta ORIGEN|FECHA debe reconstruirse exactamente igual.
    """

    def _escribir_y_leer(self, corr_simples, tmpdir):
        from datetime import datetime
        traz_dir = Path(tmpdir) / "trazabilidad"
        traz_dir.mkdir(exist_ok=True)
        mes_str = datetime.today().strftime("%Y_%m")
        ruta    = traz_dir / f"trazabilidad_{mes_str}.xlsx"
        from exportar_motor import exportar_trazabilidad
        exportar_trazabilidad(str(ruta), corr_simples, [], {}, 1, "24/05/2026 10:00")
        corr_dir = Path(tmpdir) / "correcciones"
        corr_dir.mkdir(exist_ok=True)
        with patch("main.BASE_DIR",        Path(tmpdir)), \
             patch("main.CORRECCIONES_DIR", corr_dir):
            out, _, _, _, _, _ = main.leer_correcciones({})
        return out

    def test_dos_entradas_surviven_roundtrip(self):
        """Dos pagos distintos escritos → leídos → mismas 2 claves."""
        import tempfile
        corr = {
            "PEDRO P*|02/05/2026 23:02:11": {"mz": "A", "lote": "1"},
            "PEDRO P*|15/05/2026 08:30:45": {"mz": "B", "lote": "3"},
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            out = self._escribir_y_leer(corr, tmpdir)
        self.assertEqual(len(out), 2)
        self.assertIn("PEDRO P*|02/05/2026 23:02:11", out)
        self.assertIn("PEDRO P*|15/05/2026 08:30:45", out)

    def test_rerun_no_duplica_filas(self):
        """Escribir → leer → escribir → leer: mismo número de entradas, sin duplicados."""
        import tempfile
        corr = {
            "ANA R*|10/05/2026 09:15:00": {"mz": "C", "lote": "5"},
            "LUIS V*|12/05/2026 14:22:30": {"mz": "D", "lote": "8"},
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            primera = self._escribir_y_leer(corr, tmpdir)
            segunda = self._escribir_y_leer(primera, tmpdir)  # re-run con lo que leyó
        self.assertEqual(len(segunda), 2)

    def test_origen_limpio_en_trazabilidad(self):
        """ORIGEN en trazabilidad es el nombre del pagante, no la clave compuesta."""
        import tempfile
        from openpyxl import load_workbook
        from datetime import datetime
        corr = {"ROSA M*|10/05/2026 11:00:00": {"mz": "E", "lote": "2"}}
        with tempfile.TemporaryDirectory() as tmpdir:
            traz_dir = Path(tmpdir) / "trazabilidad"
            traz_dir.mkdir()
            ruta = traz_dir / f"trazabilidad_{datetime.today().strftime('%Y_%m')}.xlsx"
            from exportar_motor import exportar_trazabilidad
            exportar_trazabilidad(str(ruta), corr, [], {}, 1, "24/05/2026 10:00")
            wb = load_workbook(ruta, read_only=True, data_only=True)
            ws = wb["Sin_identificar"]
            filas = list(ws.values)
            wb.close()
        # fila 0 = grupos, fila 1 = cabeceras, fila 2 = primer dato
        headers = [str(h).strip().upper() if h else "" for h in filas[1]]
        datos   = dict(zip(headers, filas[2]))
        self.assertEqual(datos.get("ORIGEN"), "ROSA M*")          # nombre limpio
        self.assertEqual(datos.get("FECHA"),  "10/05/2026 11:00:00")  # fecha separada


# ── Fix 9: _fuente tagging — cada hoja etiqueta sus entradas ─────────────────

class TestFuenteTagging(unittest.TestCase):
    """
    leer_correcciones etiqueta cada entrada con _fuente según la hoja de origen:
    Sin_identificar → "sin_identificar", Ambiguos → "ambiguo", Maestro_inexacto → "maestro_inexacto"
    """

    def _leer_pend(self, wb):
        import tempfile
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            traz_dir = tmp / "trazabilidad"; traz_dir.mkdir()
            pend_dir = tmp / "correcciones"; pend_dir.mkdir()
            pend_path = pend_dir / "pendientes.xlsx"
            wb.save(str(pend_path))
            with patch("main.BASE_DIR", tmp), patch("main.CORRECCIONES_DIR", pend_dir):
                cs, _, _, _, _, _ = main.leer_correcciones({})
        return cs

    def test_sin_identificar_etiquetado(self):
        """Hoja Sin_identificar → _fuente = 'sin_identificar'."""
        wb = _wb_pendientes_si_ok([_fila_pend_ok("Ana R*", "10/05/2026 09:00:00", "C", "5")])
        cs = self._leer_pend(wb)
        clave = "ANA R*|10/05/2026 09:00:00"
        self.assertIn(clave, cs)
        self.assertEqual(cs[clave].get("_fuente"), "sin_identificar")

    def test_ambiguos_etiquetado(self):
        """Hoja Ambiguos → _fuente = 'ambiguo'."""
        wb = Workbook(); wb.active.title = "Sin_identificar"
        ws2 = wb.create_sheet("Ambiguos")
        ws2.append(["GRUPOS"])
        ws2.append(["USER_ID", "NOMBRE", "",
                    "TIPO", "ORIGEN", "DESTINO", "MONTO", "MENSAJE", "FECHA", "",
                    "MZ_SUG", "LOTE_SUG", "CANDIDATOS", "",
                    "MZ", "LOTE", "CONCEPTO", "MOTIVO", "",
                    "OK"])
        ws2.append(_fila_ambiguo(ok="SI", mz_final="D", lote_final="1"))
        cs = self._leer_pend(wb)
        # Key from _fila_ambiguo: ORIGEN="Pedro P*", FECHA="02/05/2026 23:02"
        for k, v in cs.items():
            if "PEDRO P" in k:
                self.assertEqual(v.get("_fuente"), "ambiguo")
                break
        else:
            self.fail("No se encontró entrada de Ambiguos en corr_simples")

    def test_maestro_inexacto_etiquetado(self):
        """Hoja Maestro_inexacto → _fuente = 'maestro_inexacto'."""
        wb = Workbook(); wb.active.title = "Sin_identificar"
        ws3 = wb.create_sheet("Maestro_inexacto")
        ws3.append(["GRUPOS"])
        ws3.append(["USER_ID", "NOMBRE", "",
                    "TIPO", "ORIGEN", "DESTINO", "MONTO", "MENSAJE", "FECHA", "",
                    "MZ_SUG", "LOTE_SUG", "DEUDA", "DIFF", "",
                    "MZ", "LOTE", "CONCEPTO", "MOTIVO", "",
                    "OK"])
        ws3.append(_fila_maestro_inexacto(ok="SI", mz_final="Q", lote_final="12"))
        cs = self._leer_pend(wb)
        for k, v in cs.items():
            if "JESSICA M" in k or "JUANA" in k or "Jessica" in k.title():
                self.assertEqual(v.get("_fuente"), "maestro_inexacto")
                break
        else:
            # fallback: check any entry has maestro_inexacto fuente
            fuentes = [v.get("_fuente") for v in cs.values()]
            self.assertIn("maestro_inexacto", fuentes, "No se encontró entrada de Maestro_inexacto")


# ── Fix 10: trazabilidad Sin_identificar no incluye ambiguo/maestro_inexacto ──

class TestTrazabilidadFuenteFiltro(unittest.TestCase):
    """exportar_trazabilidad sin_id sheet solo escribe entradas sin _fuente o con _fuente='sin_identificar'."""

    def _escribir_traz(self, corr_simples, tmpdir):
        from exportar_motor import exportar_trazabilidad
        from datetime import datetime
        traz_dir = Path(tmpdir) / "trazabilidad"; traz_dir.mkdir(exist_ok=True)
        ruta = traz_dir / f"trazabilidad_{datetime.today().strftime('%Y_%m')}.xlsx"
        exportar_trazabilidad(str(ruta), corr_simples, [], {}, 1, "24/05/2026 10:00")
        return ruta

    def test_ambiguo_no_aparece_en_sin_identificar(self):
        """Entrada con _fuente='ambiguo' no se escribe en hoja Sin_identificar."""
        import tempfile
        from openpyxl import load_workbook
        corr = {
            "PEDRO P*|02/05/2026 10:00:00": {"mz": "A", "lote": "1", "_fuente": "sin_identificar"},
            "JUANA M*|04/05/2026 08:33:00": {"mz": "D", "lote": "2", "_fuente": "ambiguo"},
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            ruta = self._escribir_traz(corr, tmpdir)
            wb = load_workbook(str(ruta), read_only=True, data_only=True)
            ws = wb["Sin_identificar"]
            filas = list(ws.values)
            wb.close()
        # fila 0=grupos, fila 1=cabeceras → datos desde fila 2
        origenes = [str(f[4]).strip() if f and len(f) > 4 else "" for f in filas[2:]]
        self.assertIn("PEDRO P*",  origenes)
        self.assertNotIn("JUANA M*", origenes)

    def test_maestro_inexacto_no_aparece_en_sin_identificar(self):
        """Entrada con _fuente='maestro_inexacto' no se escribe en Sin_identificar."""
        import tempfile
        from openpyxl import load_workbook
        corr = {
            "ANA R*|10/05/2026 09:00:00": {"mz": "C", "lote": "5", "_fuente": "sin_identificar"},
            "ROSA V*|15/05/2026 14:00:00": {"mz": "E", "lote": "2", "_fuente": "maestro_inexacto"},
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            ruta = self._escribir_traz(corr, tmpdir)
            wb = load_workbook(str(ruta), read_only=True, data_only=True)
            ws = wb["Sin_identificar"]
            filas = list(ws.values)
            wb.close()
        origenes = [str(f[4]).strip() if f and len(f) > 4 else "" for f in filas[2:]]
        self.assertIn("ANA R*",    origenes)
        self.assertNotIn("ROSA V*", origenes)

    def test_sin_fuente_aparece_en_sin_identificar(self):
        """Entrada sin _fuente se trata como sin_identificar y se escribe."""
        import tempfile
        from openpyxl import load_workbook
        corr = {"LUIS V*|12/05/2026 14:22:30": {"mz": "D", "lote": "8"}}
        with tempfile.TemporaryDirectory() as tmpdir:
            ruta = self._escribir_traz(corr, tmpdir)
            wb = load_workbook(str(ruta), read_only=True, data_only=True)
            ws = wb["Sin_identificar"]
            filas = list(ws.values)
            wb.close()
        origenes = [str(f[4]).strip() if f and len(f) > 4 else "" for f in filas[2:]]
        self.assertIn("LUIS V*", origenes)


# ── Fix 11: USER_ID/NOMBRE resueltos para corr_simples sin_identificar ────────

class TestCorrSimplesUserIdResolucion(unittest.TestCase):
    """Replica el loop de main() paso 11 que resuelve USER_ID para corr_simples."""

    @staticmethod
    def _resolver(corr_simples: dict) -> None:
        for v in corr_simples.values():
            if v.get("_fuente", "sin_identificar") == "sin_identificar":
                if not v.get("user_id") and v.get("mz") and v.get("lote") and v.get("mz") != "BLANCO":
                    uid, nom = main.buscar_uid(v["mz"], v["lote"])
                    v["user_id"] = uid
                    if not v.get("nombre"):
                        v["nombre"] = nom

    def test_buscar_uid_llamado_para_sin_identificar_sin_uid(self):
        """buscar_uid se llama para entrada sin_identificar sin user_id."""
        corr = {"ANA R*|10/05/2026 09:00:00": {"mz": "C", "lote": "5", "_fuente": "sin_identificar", "user_id": "", "nombre": ""}}
        with patch("main.buscar_uid", return_value=("U501", "ANA RAMOS")) as mock_uid:
            self._resolver(corr)
        mock_uid.assert_called_once_with("C", "5")
        self.assertEqual(corr["ANA R*|10/05/2026 09:00:00"]["user_id"], "U501")
        self.assertEqual(corr["ANA R*|10/05/2026 09:00:00"]["nombre"],  "ANA RAMOS")

    def test_buscar_uid_no_llamado_para_ambiguo(self):
        """buscar_uid NO se llama para entrada con _fuente='ambiguo'."""
        corr = {"JUANA M*|04/05/2026 08:33:00": {"mz": "D", "lote": "2", "_fuente": "ambiguo", "user_id": "", "nombre": ""}}
        with patch("main.buscar_uid", return_value=("U999", "X")) as mock_uid:
            self._resolver(corr)
        mock_uid.assert_not_called()

    def test_buscar_uid_no_llamado_para_blanco(self):
        """buscar_uid NO se llama cuando MZ='BLANCO'."""
        corr = {"PEDRO P*|02/05/2026 10:00:00": {"mz": "BLANCO", "lote": "", "_fuente": "sin_identificar", "user_id": "", "nombre": ""}}
        with patch("main.buscar_uid", return_value=("U999", "X")) as mock_uid:
            self._resolver(corr)
        mock_uid.assert_not_called()

    def test_buscar_uid_no_llamado_si_uid_ya_presente(self):
        """buscar_uid NO se llama si user_id ya está lleno."""
        corr = {"ROSA M*|10/05/2026 11:00:00": {"mz": "E", "lote": "2", "_fuente": "sin_identificar", "user_id": "U200", "nombre": "ROSA MILLA"}}
        with patch("main.buscar_uid", return_value=("OTRO", "OTRO")) as mock_uid:
            self._resolver(corr)
        mock_uid.assert_not_called()
        self.assertEqual(corr["ROSA M*|10/05/2026 11:00:00"]["user_id"], "U200")


# ── Fix 12: _leer_hoja_pagaste_ok ─────────────────────────────────────────────

def _wb_pagaste_pend(filas: list) -> Workbook:
    """Workbook con hoja Pagaste en formato cabecera-doble (schema v2 con MZ+LOTE)."""
    wb = Workbook()
    wb.active.title = "Sin_identificar"
    ws = wb.create_sheet("Pagaste")
    ws.append(["← BANCO — NO EDITAR", "", "", "", "", "", "", "← TÚ COMPLETAS", "", "", "", "", "✓"])
    ws.append(["TIPO", "ORIGEN", "DESTINO", "MONTO", "MENSAJE", "FECHA", "", "MZ", "LOTE", "CONCEPTO", "MOTIVO", "", "OK"])
    for fila in filas:
        ws.append(fila)
    return wb


class TestLeerHojaPagasteOk(unittest.TestCase):

    def test_sin_hoja_pagaste_retorna_vacio(self):
        wb = Workbook()
        wb.active.title = "Sin_identificar"
        self.assertEqual(main._leer_hoja_pagaste_ok(wb), {})

    def test_hoja_menos_de_3_filas_retorna_vacio(self):
        wb = Workbook()
        wb.active.title = "Sin_identificar"
        ws = wb.create_sheet("Pagaste")
        ws.append(["grupo"])
        ws.append(["TIPO", "ORIGEN", "DESTINO", "MONTO", "MENSAJE", "FECHA", "", "MZ", "LOTE", "CONCEPTO", "MOTIVO", "", "OK"])
        self.assertEqual(main._leer_hoja_pagaste_ok(wb), {})

    def test_ok_si_devuelve_entrada(self):
        filas = [["PAGASTE", "PROVEEDOR SA", "JASS", 150.0, "pago servicio", "10/05/2026", "", "", "", "MANT.", "reparacion", "", "SI"]]
        wb = _wb_pagaste_pend(filas)
        result = main._leer_hoja_pagaste_ok(wb)
        clave = "PROVEEDOR SA|150.0|10/05/2026"
        self.assertIn(clave, result)
        self.assertEqual(result[clave]["monto"], 150.0)
        self.assertEqual(result[clave]["concepto"], "MANT.")

    def test_ok_no_excluido(self):
        filas = [["PAGASTE", "PROVEEDOR SA", "JASS", 150.0, "pago servicio", "10/05/2026", "", "", "", "MANT.", "reparacion", "", "NO"]]
        wb = _wb_pagaste_pend(filas)
        self.assertEqual(main._leer_hoja_pagaste_ok(wb), {})

    def test_ok_vacio_excluido(self):
        filas = [["PAGASTE", "PROVEEDOR SA", "JASS", 150.0, "pago servicio", "10/05/2026", "", "", "", "MANT.", "reparacion", "", ""]]
        wb = _wb_pagaste_pend(filas)
        self.assertEqual(main._leer_hoja_pagaste_ok(wb), {})

    def test_clave_usa_origen_upper(self):
        filas = [["PAGASTE", "proveedor sa", "JASS", 50.0, "", "05/05/2026", "", "", "", "", "", "", "SI"]]
        wb = _wb_pagaste_pend(filas)
        result = main._leer_hoja_pagaste_ok(wb)
        self.assertIn("PROVEEDOR SA|50.0|05/05/2026", result)

    def test_mz_lote_se_leen_cuando_estan(self):
        filas = [["PAGASTE", "JASS Vil*", "Rosalina Cir*", 46.0, "devolucion adelantado", "02/06/2026", "", "B", "8", "", "pago adelantado", "", "SI"]]
        wb = _wb_pagaste_pend(filas)
        result = main._leer_hoja_pagaste_ok(wb)
        clave = "JASS VIL*|46.0|02/06/2026"
        self.assertEqual(result[clave]["mz"],   "B")
        self.assertEqual(result[clave]["lote"], "8")


# ── Fix 13: blanco → estado identificado ──────────────────────────────────────

class TestBlancoEstadoIdentificado(unittest.TestCase):

    def _ejecutar_con_blanco(self, origen, fecha, monto=38.0):
        import tempfile
        import pandas as pd
        base_row = {
            "tipo": "TE PAGÓ", "origen": origen, "destino": "JASS",
            "monto": monto, "mensaje": "", "fecha": fecha,
        }
        mapa = {"tipo": "tipo", "origen": "origen", "destino": "destino",
                "monto": "monto", "mensaje": "mensaje", "fecha": "fecha"}
        df = pd.DataFrame([base_row])
        clave = f"{origen.upper()}|{fecha}"
        corr_simples = {clave: {"mz": "BLANCO", "lote": "", "concepto": "", "motivo": "marcado como blanco"}}
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            with patch("main.OUTPUT_DIR", tmp), patch("main.CORRECCIONES_DIR", tmp), patch("main.BASE_DIR", tmp):
                # ciclo=2: las correcciones manuales sólo se aplican desde el ciclo 2
                todos = main.ejecutar_matching(df, mapa, {}, {}, corr_simples, {}, {}, set(), 2, {})
        return todos

    def test_blanco_estado_identificado(self):
        todos = self._ejecutar_con_blanco("JUAN P*", "10/05/2026 08:00:00")
        blancos = [r for r in todos if r.get("fuente") == "blanco"]
        self.assertEqual(len(blancos), 1)
        self.assertEqual(blancos[0]["estado"], "identificado")

    def test_blanco_concepto_BLANCO(self):
        todos = self._ejecutar_con_blanco("JUAN P*", "10/05/2026 08:00:00")
        blancos = [r for r in todos if r.get("fuente") == "blanco"]
        self.assertEqual(blancos[0]["concepto"], "BLANCO")

    def test_blanco_no_aparece_como_pendiente(self):
        todos = self._ejecutar_con_blanco("MARIA M*", "12/05/2026 09:00:00")
        sin_id = [
            r for r in todos
            if r.get("estado") == "pendiente"
            and r.get("fuente") not in ("ambiguo_auto", "maestro_inexacto")
        ]
        self.assertEqual(len(sin_id), 0)


# ── Fix 14: exportar_pagos_pagaste ────────────────────────────────────────────

class TestExportarPagosPagaste(unittest.TestCase):

    @staticmethod
    def _sample():
        return [{"origen": "PROVEEDOR SA", "destino": "JASS", "monto": 200.0,
                 "mensaje": "pago mes", "concepto": "MATERIAL", "motivo": "compra caños",
                 "fecha": "15/05/2026", "ciclo": 1}]

    def test_devuelve_workbook(self):
        from exportar_motor import exportar_pagos_pagaste
        wb = exportar_pagos_pagaste(self._sample(), 1)
        self.assertIsInstance(wb, Workbook)

    def test_titulo_hoja(self):
        from exportar_motor import exportar_pagos_pagaste
        wb = exportar_pagos_pagaste(self._sample(), 1)
        self.assertEqual(wb.active.title, "pagos_yape_pagaste")

    def test_cabecera_doble_tiene_origen_monto_ciclo(self):
        from exportar_motor import exportar_pagos_pagaste
        wb = exportar_pagos_pagaste(self._sample(), 1)
        ws = wb.active
        fila2 = [c.value for c in ws[2] if c.value]
        self.assertIn("ORIGEN",  fila2)
        self.assertIn("MONTO",   fila2)
        self.assertIn("CICLO_CORRECCION", fila2)
        self.assertIn("MZ",      fila2)
        self.assertIn("LOTE",    fila2)

    def test_dato_en_fila3(self):
        from exportar_motor import exportar_pagos_pagaste
        wb = exportar_pagos_pagaste(self._sample(), 1)
        ws = wb.active
        fila2 = [c.value for c in ws[2]]
        col_origen = fila2.index("ORIGEN") + 1
        self.assertEqual(ws.cell(row=3, column=col_origen).value, "PROVEEDOR SA")

    def test_lista_vacia_no_genera_fila3(self):
        from exportar_motor import exportar_pagos_pagaste
        wb = exportar_pagos_pagaste([], 1)
        ws = wb.active
        self.assertIsNone(ws.cell(row=3, column=1).value)


# ── Fix 15: exportar_reporte_procesado ────────────────────────────────────────

class TestExportarReporteProcesado(unittest.TestCase):

    def test_siempre_devuelve_2_hojas(self):
        from exportar_motor import exportar_reporte_procesado
        wb = exportar_reporte_procesado(None, None)
        self.assertEqual(len(wb.sheetnames), 2)
        self.assertIn("TE_PAGÓ",  wb.sheetnames)
        self.assertIn("PAGASTE",  wb.sheetnames)

    def test_archivo_inexistente_no_lanza_excepcion(self):
        from exportar_motor import exportar_reporte_procesado
        wb = exportar_reporte_procesado(Path("no_existe_tepago.xlsx"), Path("no_existe_pagaste.xlsx"))
        self.assertEqual(len(wb.sheetnames), 2)

    def test_copia_datos_de_pagaste(self):
        import tempfile
        from exportar_motor import exportar_pagos_pagaste, exportar_reporte_procesado
        pagaste = [{"origen": "PROV X", "destino": "JASS", "monto": 100.0,
                    "mensaje": "", "concepto": "SERV", "motivo": "", "fecha": "01/05/2026"}]
        with tempfile.TemporaryDirectory() as tmpdir:
            ruta_pagaste = Path(tmpdir) / "pagos_yape_pagaste.xlsx"
            exportar_pagos_pagaste(pagaste, 1).save(ruta_pagaste)
            wb = exportar_reporte_procesado(None, ruta_pagaste)
        fila2 = [c.value for c in wb["PAGASTE"][2]]
        self.assertIn("ORIGEN", fila2)


# ── Fix 16: hoja Pagaste en pendientes.xlsx ───────────────────────────────────

class TestPendientesPagasteHoja(unittest.TestCase):

    def test_sin_pagaste_no_crea_hoja(self):
        from exportar_motor import exportar_pendientes_diseño
        wb = exportar_pendientes_diseño([], [], [], None)
        self.assertNotIn("Pagaste", wb.sheetnames)

    def test_lista_vacia_no_crea_hoja(self):
        from exportar_motor import exportar_pendientes_diseño
        wb = exportar_pendientes_diseño([], [], [], [])
        self.assertNotIn("Pagaste", wb.sheetnames)

    def test_con_pagaste_crea_hoja(self):
        from exportar_motor import exportar_pendientes_diseño
        pagaste = [{"tipo": "PAGASTE", "origen": "PROV", "destino": "JASS",
                    "monto": 80.0, "mensaje": "", "fecha": "05/05/2026"}]
        wb = exportar_pendientes_diseño([], [], [], pagaste)
        self.assertIn("Pagaste", wb.sheetnames)

    def test_pagaste_hoja_cabecera_tiene_origen_y_ok(self):
        from exportar_motor import exportar_pendientes_diseño
        pagaste = [{"tipo": "PAGASTE", "origen": "PROV", "destino": "JASS",
                    "monto": 80.0, "mensaje": "", "fecha": "05/05/2026"}]
        wb = exportar_pendientes_diseño([], [], [], pagaste)
        ws = wb["Pagaste"]
        fila2 = [c.value for c in ws[2] if c.value]
        self.assertIn("ORIGEN", fila2)
        self.assertIn("OK",     fila2)


# ── Fix 17: leer_correcciones retorna 6-tupla con corr_pagaste ─────────────────

class TestLeerCorrecciones6Tupla(unittest.TestCase):

    def test_retorna_6_valores(self):
        import tempfile
        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("main.CORRECCIONES_DIR", Path(tmpdir)), \
                 patch("main.BASE_DIR", Path(tmpdir)):
                resultado = main.leer_correcciones({})
        self.assertEqual(len(resultado), 6)

    def test_sexto_es_dict(self):
        import tempfile
        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("main.CORRECCIONES_DIR", Path(tmpdir)), \
                 patch("main.BASE_DIR", Path(tmpdir)):
                *_, corr_pagaste = main.leer_correcciones({})
        self.assertIsInstance(corr_pagaste, dict)

    def test_sin_pendientes_corr_pagaste_vacio(self):
        import tempfile
        with tempfile.TemporaryDirectory() as tmpdir:
            with patch("main.CORRECCIONES_DIR", Path(tmpdir)), \
                 patch("main.BASE_DIR", Path(tmpdir)):
                *_, corr_pagaste = main.leer_correcciones({})
        self.assertEqual(corr_pagaste, {})

    def test_pagaste_ok_si_cargado_en_sexto_elemento(self):
        import tempfile
        from exportar_motor import exportar_pendientes_diseño
        pagaste_pend = [{"tipo": "PAGASTE", "origen": "PROVEEDOR A", "destino": "JASS",
                         "monto": 75.0, "mensaje": "", "fecha": "08/05/2026"}]
        wb = exportar_pendientes_diseño([], [], [], pagaste_pend)
        ws = wb["Pagaste"]
        fila2 = [c.value for c in ws[2]]
        col_ok = fila2.index("OK") + 1
        ws.cell(row=3, column=col_ok).value = "SI"
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            corr_dir = tmp / "correcciones"
            corr_dir.mkdir()
            wb.save(str(corr_dir / "pendientes.xlsx"))
            with patch("main.CORRECCIONES_DIR", corr_dir), \
                 patch("main.BASE_DIR", tmp):
                *_, corr_pagaste = main.leer_correcciones({})
        self.assertEqual(len(corr_pagaste), 1)
        self.assertIn("PROVEEDOR A|75.0|08/05/2026", corr_pagaste)


# ── Fix 18: cargar_reportes retorna 3-tupla ────────────────────────────────────

class TestCargaReportes3Tupla(unittest.TestCase):

    @staticmethod
    def _crear_reporte(tmpdir, filas):
        from openpyxl import Workbook as Wb
        wb = Wb()
        ws = wb.active
        ws.append(["Tipo de Transacción", "Origen", "Destino", "Monto", "Mensaje", "Fecha"])
        for fila in filas:
            ws.append(fila)
        ruta = Path(tmpdir) / "reporte_mayo.xlsx"
        wb.save(ruta)
        return ruta

    def test_retorna_3_tupla(self):
        import tempfile
        filas = [
            ["TE PAGÓ",  "ANA M*",  "JASS", 38.0,  "", "01/05/2026 08:00:00"],
            ["PAGASTE",  "PROV SA", "JASS", 100.0, "", "02/05/2026 09:00:00"],
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            self._crear_reporte(tmpdir, filas)
            with patch("main.REPORTE_DIR", Path(tmpdir)):
                resultado = main.cargar_reportes()
        self.assertEqual(len(resultado), 3)

    def test_tepago_y_pagaste_separados(self):
        import tempfile
        filas = [
            ["TE PAGÓ",  "ANA M*",  "JASS", 38.0,  "", "01/05/2026 08:00:00"],
            ["TE PAGÓ",  "ROSA R*", "JASS", 50.0,  "", "03/05/2026 10:00:00"],
            ["PAGASTE",  "PROV SA", "JASS", 100.0, "", "02/05/2026 09:00:00"],
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            self._crear_reporte(tmpdir, filas)
            with patch("main.REPORTE_DIR", Path(tmpdir)):
                df_tepago, df_pagaste, _ = main.cargar_reportes()
        self.assertEqual(len(df_tepago),  2)
        self.assertEqual(len(df_pagaste), 1)

    def test_sin_pagaste_df_pagaste_vacio(self):
        import tempfile
        filas = [
            ["TE PAGÓ", "ANA M*", "JASS", 38.0, "", "01/05/2026 08:00:00"],
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            self._crear_reporte(tmpdir, filas)
            with patch("main.REPORTE_DIR", Path(tmpdir)):
                df_tepago, df_pagaste, _ = main.cargar_reportes()
        self.assertEqual(len(df_tepago),  1)
        self.assertEqual(len(df_pagaste), 0)

    def test_mapa_tiene_tipo_y_origen(self):
        import tempfile
        filas = [["TE PAGÓ", "PEDRO P*", "JASS", 30.0, "", "05/05/2026 07:00:00"]]
        with tempfile.TemporaryDirectory() as tmpdir:
            self._crear_reporte(tmpdir, filas)
            with patch("main.REPORTE_DIR", Path(tmpdir)):
                _, _, mapa = main.cargar_reportes()
        self.assertIn("tipo",   mapa)
        self.assertIn("origen", mapa)


if __name__ == "__main__":
    unittest.main(verbosity=2)
