# Genera mesa_1.xlsx … mesa_7.xlsx en inputs/ con 3 hojas cada uno.
# También genera pagos_efectivo_devolucion.xlsx en outputs/ — template para retornos manuales.
# Correr una sola vez: python crear_templates.py
# Contrato visual: docs/formato_registro.html
# El dibujo de mesa_N vive en shared/utils_templates.py (compartido con 7_cierre,
# que resetea mesa_N al cerrar el período — ver docs/metodologia_desarrollo.md).

import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent.parent / "shared"))
from utils_templates import crear_mesa_vacio, filas_con_datos  # noqa: E402

INPUTS_DIR = Path(__file__).parent / "inputs"
INPUTS_DIR.mkdir(exist_ok=True)

N_MESAS = 7


# ── Helpers de estilo ───────────────────────────────────────────────────────

def _borde(color="FFFFFF"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _celda_grupo(cell, texto, bg, txt):
    cell.value     = texto
    cell.font      = Font(name="Arial", bold=True, size=9, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _borde()

def _celda_col(cell, texto, bg, txt):
    cell.value     = texto
    cell.font      = Font(name="Arial", bold=True, size=10, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _borde()

def _celda_ejemplo(cell, valor):
    cell.value     = valor
    cell.font      = Font(name="Arial", size=10, color="9CA3AF", italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")


# ── Crear un archivo de mesa ────────────────────────────────────────────────
# Dibujo delegado a shared/utils_templates.py (mismo primitivo que usa 7_cierre
# para resetear mesa_N al cerrar el período — cross-módulo va a shared, nunca
# import directo entre módulos hermanos, ver metodologia_desarrollo.md).

def crear_mesa(n: int):
    ruta = INPUTS_DIR / f"mesa_{n}.xlsx"
    crear_mesa_vacio(ruta)
    print(f"  OK mesa_{n}.xlsx  (3 hojas)")


# ── Template pagos_efectivo_devolucion.xlsx ─────────────────────────────────
# Schema: MZ | LOTE | NOMBRE | MONTO | FECHA | CONCEPTO
# La tesorera llena una fila por cada retorno en efectivo que haga.
# Este archivo va a 5_cobranza/inputs/pagos_efectivo/pagos_efectivo_devolucion.xlsx

_DEV_GRUPOS = [
    ("¿A quién se devolvió?", 3, "E1F5EE", "085041"),
    ("¿Cuánto y cuándo?",     2, "FEF9E7", "7D6608"),
    ("¿Por qué?",             1, "F4ECF7", "5B21B6"),
]

_DEV_COLUMNAS = [
    ("MZ",       "E1F5EE", "085041",  8),
    ("LOTE",     "E1F5EE", "085041",  8),
    ("NOMBRE",   "E1F5EE", "085041", 28),
    ("MONTO",    "FEF9E7", "7D6608", 12),
    ("FECHA",    "FEF9E7", "7D6608", 14),
    ("CONCEPTO", "F4ECF7", "5B21B6", 40),
]

_DEV_EJEMPLO = ["A", "7", "JUAN PEREZ GARCIA", "40.00", "12/06/2026", "Pago fuera de tiempo — retorno acordado con usuario"]


def crear_template_devolucion():
    OUTPUTS_DIR = Path(__file__).parent / "outputs"
    OUTPUTS_DIR.mkdir(exist_ok=True)
    ruta = OUTPUTS_DIR / "pagos_efectivo_devolucion.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "pagos_efectivo_devolucion"
    ws.freeze_panes = "A3"

    col = 1
    for texto, n_cols, bg, txt in _DEV_GRUPOS:
        if n_cols > 1:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + n_cols - 1)
        _celda_grupo(ws.cell(row=1, column=col), texto, bg, txt)
        col += n_cols
    ws.row_dimensions[1].height = 20

    for idx, (nombre, bg, txt, ancho) in enumerate(_DEV_COLUMNAS, start=1):
        _celda_col(ws.cell(row=2, column=idx), nombre, bg, txt)
        ws.column_dimensions[get_column_letter(idx)].width = ancho
    ws.row_dimensions[2].height = 22

    for idx, valor in enumerate(_DEV_EJEMPLO, start=1):
        _celda_ejemplo(ws.cell(row=3, column=idx), valor)
    ws.row_dimensions[3].height = 18

    wb.save(ruta)
    print(f"  OK pagos_efectivo_devolucion.xlsx  (outputs/)")


# ── Entry point ─────────────────────────────────────────────────────────────

def _mesas_con_datos() -> list[tuple[int, int]]:
    """[(n_mesa, filas)] de las mesas que hoy tienen cobros escritos."""
    return [(n, f) for n in range(1, N_MESAS + 1)
            if (f := filas_con_datos(INPUTS_DIR / f"mesa_{n}.xlsx")) > 0]


if __name__ == "__main__":
    # Guarda: este script pisa las 7 mesas. El 26/07/2026 se corrió para
    # preparar el ciclo de agosto y borró las 374 filas de julio que los
    # cobradores habían escrito a mano — con el split yape/efectivo y los
    # comentarios, que no se pueden reconstruir desde pagos_efectivo.xlsx
    # (recuperar_mesas.py lo dice: "MONTO_YAPE = 0, corregir a mano después").
    # Se recuperaron de una copia suelta del repo, por suerte.
    #
    # El reset legítimo de fin de ciclo lo hace 7_cierre, que archiva primero y
    # pide consentimiento. Este script es solo para el setup inicial.
    con_datos = _mesas_con_datos()
    if con_datos and "--force" not in sys.argv:
        print("\n  ALTO: las mesas tienen cobros escritos y este script las borra.\n")
        for n, f in con_datos:
            print(f"     mesa_{n}.xlsx  {f} filas de cobro")
        print(f"\n  Total: {sum(f for _, f in con_datos)} filas se perderian.\n"
              f"  El reset de fin de ciclo lo hace 7_cierre (archiva primero).\n"
              f"  Si igual queres regenerar los templates: python crear_templates.py --force\n"
              f"  (se respalda en backup/mesas_pre_reset_<fecha>/ de todas formas)\n")
        sys.exit(1)

    print("\nCreando templates de mesa...")
    for n in range(1, N_MESAS + 1):
        crear_mesa(n)
    print("\nCreando template de retornos en efectivo...")
    crear_template_devolucion()
    print(f"\nListo. Abre cada mesa_N.xlsx y llena tus cobros a partir de la fila 4 (la fila 3 es el ejemplo guía, déjala).")
    print("Para retornos: llena pagos_efectivo_devolucion.xlsx desde fila 4 y cópialo a 5_cobranza/inputs/pagos_efectivo/.\n")
