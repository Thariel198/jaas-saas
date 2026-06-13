"""DESECHABLE — inspecciona el estado actual de los mesa_N.xlsx para planear migración.
Se borra al terminar. NO toca ningún archivo, solo lee."""
from openpyxl import load_workbook
from pathlib import Path

ROOT = Path(__file__).parent

for n in range(1, 8):
    path = ROOT / "inputs" / f"mesa_{n}.xlsx"
    if not path.exists():
        continue
    wb = load_workbook(path, read_only=True, data_only=True)
    print(f"\n=== {path.name} ===")
    for hoja in wb.sheetnames:
        ws = wb[hoja]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue
        headers = [str(h or "").strip() for h in rows[1]]
        try:
            i_monto = headers.index("MONTO")
            i_efec  = headers.index("MONTO_EFECTIVO")
            i_yape  = headers.index("MONTO_YAPE")
            i_mz    = headers.index("MZ")
            i_lt    = headers.index("LT")
        except ValueError as e:
            print(f"  {hoja}: ERROR falta columna {e}")
            continue
        n_filas = n_monto = n_efec = n_yape = n_match = n_no_match = 0
        ejemplos = []
        for r in rows[3:]:
            if not r or all(v in (None, "") for v in r):
                continue
            mz_v = r[i_mz] if i_mz < len(r) else None
            if not mz_v:
                continue
            n_filas += 1
            mv = r[i_monto] if i_monto < len(r) else None
            ev = r[i_efec]  if i_efec  < len(r) else None
            yv = r[i_yape]  if i_yape  < len(r) else None
            if mv not in (None, "", 0): n_monto += 1
            if ev not in (None, "", 0): n_efec += 1
            if yv not in (None, "", 0): n_yape += 1
            try:
                fm = float(mv or 0); fe = float(ev or 0); fy = float(yv or 0)
                if abs(fm - (fe + fy)) < 0.01:
                    n_match += 1
                else:
                    n_no_match += 1
                    if len(ejemplos) < 3:
                        ejemplos.append(f"MZ={mz_v} LT={r[i_lt]} MONTO={fm} EFE={fe} YAPE={fy}")
            except (TypeError, ValueError):
                n_no_match += 1
        print(f"  {hoja}: filas={n_filas} | con_MONTO={n_monto} con_EFEC={n_efec} con_YAPE={n_yape} | match={n_match} no_match={n_no_match}")
        for ej in ejemplos:
            print(f"    ! {ej}")
    wb.close()
