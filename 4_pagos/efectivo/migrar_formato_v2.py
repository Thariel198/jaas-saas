# Migra mesa_1.xlsx … mesa_7.xlsx al formato v2.
# Cambio: agrega MONTO_EFECTIVO y MONTO_YAPE entre MONTO y FECHA.
# Resguarda originales en backup/migracion_v2_YYYY_MM/ antes de sobrescribir.
# Default: MONTO_EFECTIVO = MONTO, MONTO_YAPE = 0.
# Idempotente: si una hoja ya tiene cabecera MONTO_EFECTIVO, la omite.
#
# Uso: python migrar_formato_v2.py

from pathlib import Path
import shutil
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Reutiliza la misma definición de columnas que crear_templates.py
from crear_templates import GRUPOS, COLUMNAS, EJEMPLO, HOJAS, N_MESAS

INPUTS_DIR = Path(__file__).parent / "inputs"
BACKUP_DIR = (
    Path(__file__).parent / "backup" / f"migracion_v2_{datetime.now():%Y_%m}"
)

_MARKER = "MONTO_EFECTIVO"   # cabecera que distingue formato v2 del v1


# ── Helpers de estilo (espejo de crear_templates.py) ───────────────────────

def _estilo_grupo(cell, texto, bg, txt):
    cell.value     = texto
    cell.font      = Font(name="Arial", bold=True, size=9, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def _estilo_col(cell, texto, bg, txt, ancho, ws, col_idx):
    from openpyxl.utils import get_column_letter
    cell.value     = texto
    cell.font      = Font(name="Arial", bold=True, size=10, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(col_idx)].width = ancho

def _estilo_ejemplo(cell, valor):
    cell.value     = valor
    cell.font      = Font(name="Arial", size=10, color="9CA3AF", italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")


# ── Lectura de datos ────────────────────────────────────────────────────────

def _ya_migrada(ws) -> bool:
    filas = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    if not filas:
        return False
    headers = [str(h).strip().upper() if h else "" for h in filas[0]]
    return _MARKER in headers


def _leer_datos(ws) -> list[dict]:
    """Lee filas de datos (fila 4+) por nombre de columna. Omite filas vacías."""
    filas = list(ws.iter_rows(values_only=True))
    if len(filas) < 4:
        return []
    headers = [str(h).strip().upper() if h else "" for h in filas[1]]
    datos = []
    for fila in filas[3:]:
        if not fila or all(c is None for c in fila):
            continue
        row = {headers[i]: fila[i] for i in range(min(len(headers), len(fila)))}
        if not row.get("MZ") and not row.get("LT"):
            continue
        datos.append(row)
    return datos


# ── Escritura de hoja ───────────────────────────────────────────────────────

def _construir_hoja(ws, datos: list[dict]):
    # Fila 1 — grupos
    col = 1
    for texto, n_cols, bg, txt in GRUPOS:
        if n_cols > 1:
            ws.merge_cells(
                start_row=1, start_column=col,
                end_row=1,   end_column=col + n_cols - 1,
            )
        _estilo_grupo(ws.cell(row=1, column=col), texto, bg, txt)
        col += n_cols
    ws.row_dimensions[1].height = 20

    # Fila 2 — nombres de columna
    for idx, (nombre, bg, txt, ancho) in enumerate(COLUMNAS, start=1):
        _estilo_col(ws.cell(row=2, column=idx), nombre, bg, txt, ancho, ws, idx)
    ws.row_dimensions[2].height = 22

    # Fila 3 — ejemplo guía
    for idx, valor in enumerate(EJEMPLO, start=1):
        _estilo_ejemplo(ws.cell(row=3, column=idx), valor)
    ws.row_dimensions[3].height = 18

    ws.freeze_panes = "A3"

    # Filas de datos
    for fila_idx, d in enumerate(datos, start=4):
        monto      = d.get("MONTO")
        # Si ya existían las columnas (archivo v2 abierto dos veces), las conserva.
        # Si no, default: efectivo = monto total, yape = 0.
        monto_efec = d.get("MONTO_EFECTIVO")
        monto_yape = d.get("MONTO_YAPE")
        if monto_efec is None:
            monto_efec = monto
        if monto_yape is None:
            monto_yape = 0

        valores = [
            d.get("COBRADOR"),
            d.get("FECHA_REGISTRO"),
            d.get("MZ"),
            d.get("LT"),
            monto,
            monto_efec,
            monto_yape,
            d.get("FECHA"),
            d.get("COMENTARIO") or "",
        ]
        for col_idx, val in enumerate(valores, start=1):
            ws.cell(row=fila_idx, column=col_idx).value = val


# ── Migrar un archivo ───────────────────────────────────────────────────────

def _migrar_mesa(n: int) -> str:
    ruta = INPUTS_DIR / f"mesa_{n}.xlsx"
    if not ruta.exists():
        return f"  SKIP  mesa_{n}.xlsx — archivo no encontrado"

    wb_old = load_workbook(ruta)
    hojas_a_migrar   = []
    hojas_ya_v2      = []
    datos_por_hoja   = {}

    for hoja in HOJAS:
        if hoja not in wb_old.sheetnames:
            continue
        ws = wb_old[hoja]
        if _ya_migrada(ws):
            hojas_ya_v2.append(hoja)
            datos_por_hoja[hoja] = _leer_datos(ws)   # igual los leemos para reconstruir
        else:
            hojas_a_migrar.append(hoja)
            datos_por_hoja[hoja] = _leer_datos(ws)

    if not hojas_a_migrar:
        n_filas = sum(len(v) for v in datos_por_hoja.values())
        return f"  YA-V2 mesa_{n}.xlsx — {len(hojas_ya_v2)} hoja(s) · {n_filas} fila(s) sin cambios"

    # Backup del archivo original
    shutil.copy2(ruta, BACKUP_DIR / f"mesa_{n}.xlsx")

    # Reconstruir el workbook completo con el nuevo formato
    wb_new = Workbook()
    for i, hoja in enumerate(HOJAS):
        ws = wb_new.active if i == 0 else wb_new.create_sheet()
        ws.title = hoja
        _construir_hoja(ws, datos_por_hoja.get(hoja, []))

    wb_new.save(ruta)
    total_filas = sum(len(v) for v in datos_por_hoja.values())
    return (
        f"  OK    mesa_{n}.xlsx — "
        f"{len(hojas_a_migrar)} hoja(s) migradas · {total_filas} fila(s) preservadas"
        + (f" · {len(hojas_ya_v2)} ya era(n) v2" if hojas_ya_v2 else "")
    )


# ── Entry point ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"\nMigración formato v2 — {datetime.now():%Y-%m-%d %H:%M}")
    print(f"Backup en: {BACKUP_DIR}\n")

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    for n in range(1, N_MESAS + 1):
        print(_migrar_mesa(n))

    print("\nListo.")
    print("Revisión: abre cada mesa_N.xlsx y ajusta MONTO_YAPE donde corresponda")
    print("          (filas que venían de Yape quedaron con MONTO_EFECTIVO = MONTO).\n")
