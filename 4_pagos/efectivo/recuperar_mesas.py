# Reconstruye mesa_1.xlsx … mesa_7.xlsx desde outputs/pagos_efectivo.xlsx.
# Usar solo si los archivos en inputs/ fueron borrados accidentalmente.
# Default: MONTO_EFECTIVO = MONTO, MONTO_YAPE = 0  (corregir yape a mano después).
#
# Uso: python recuperar_mesas.py

from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from crear_templates import GRUPOS, COLUMNAS, EJEMPLO, HOJAS, N_MESAS

INPUTS_DIR  = Path(__file__).parent / "inputs"
PAGOS_PATH  = Path(__file__).parent / "outputs" / "pagos_efectivo.xlsx"


# ── Leer pagos_efectivo.xlsx ────────────────────────────────────────────────

def _leer_pagos() -> dict[str, list[dict]]:
    wb = load_workbook(PAGOS_PATH)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip().upper() if h else "" for h in filas[1]]

    por_mesa: dict[str, list[dict]] = {}
    for fila in filas[2:]:
        if not fila or all(c is None for c in fila):
            continue
        row = {headers[i]: fila[i] for i in range(min(len(headers), len(fila)))}
        mesa = str(row.get("MESA") or "").strip()
        if not mesa:
            continue
        por_mesa.setdefault(mesa, []).append(row)
    return por_mesa


# ── Helpers de estilo (espejo de crear_templates.py) ───────────────────────

def _construir_hoja(ws, datos: list[dict]):
    # Fila 1 — grupos
    col = 1
    for texto, n_cols, bg, txt in GRUPOS:
        if n_cols > 1:
            ws.merge_cells(
                start_row=1, start_column=col,
                end_row=1,   end_column=col + n_cols - 1,
            )
        cell = ws.cell(row=1, column=col)
        cell.value     = texto
        cell.font      = Font(name="Arial", bold=True, size=9, color=txt)
        cell.fill      = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col += n_cols
    ws.row_dimensions[1].height = 20

    # Fila 2 — nombres de columna
    for idx, (nombre, bg, txt, ancho) in enumerate(COLUMNAS, start=1):
        cell = ws.cell(row=2, column=idx)
        cell.value     = nombre
        cell.font      = Font(name="Arial", bold=True, size=10, color=txt)
        cell.fill      = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = ancho
    ws.row_dimensions[2].height = 22

    # Fila 3 — ejemplo guía
    for idx, valor in enumerate(EJEMPLO, start=1):
        cell = ws.cell(row=3, column=idx)
        cell.value     = valor
        cell.font      = Font(name="Arial", size=10, color="9CA3AF", italic=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 18

    ws.freeze_panes = "A3"

    # Filas de datos (fila 4+)
    for fila_idx, d in enumerate(datos, start=4):
        fecha = d.get("FECHA")
        if hasattr(fecha, "date"):
            fecha = fecha.date()
        valores = [
            d.get("COBRADOR"),
            None,               # FECHA_REGISTRO — no estaba en pagos_efectivo
            d.get("MZ"),
            d.get("LT"),
            d.get("MONTO"),
            None,               # MONTO_EFECTIVO — vacío; solo llenar en pagos mixtos
            None,               # MONTO_YAPE    — vacío; solo llenar en pagos mixtos
            fecha,
            d.get("COMENTARIO") or "",
        ]
        for col_idx, val in enumerate(valores, start=1):
            ws.cell(row=fila_idx, column=col_idx).value = val


# ── Reconstruir un mesa_N.xlsx ──────────────────────────────────────────────

def _reconstruir_mesa(n: int, por_mesa: dict):
    nombre_mesa = f"mesa_{n}"
    datos = por_mesa.get(nombre_mesa, [])

    wb = Workbook()
    for i, hoja in enumerate(HOJAS):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = hoja
        # Solo registro_1 lleva los datos recuperados; registro_2/3 quedan en blanco
        _construir_hoja(ws, datos if hoja == "registro_1" else [])

    ruta = INPUTS_DIR / f"{nombre_mesa}.xlsx"
    wb.save(ruta)
    return len(datos)


# ── Entry point ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"\nRecuperación de mesas — {datetime.now():%Y-%m-%d %H:%M}")
    print(f"Fuente: {PAGOS_PATH}\n")

    if not PAGOS_PATH.exists():
        print("ERROR: no se encontró pagos_efectivo.xlsx")
        raise SystemExit(1)

    por_mesa = _leer_pagos()
    total = 0
    for n in range(1, N_MESAS + 1):
        n_filas = _reconstruir_mesa(n, por_mesa)
        total += n_filas
        print(f"  OK  mesa_{n}.xlsx — {n_filas} filas recuperadas en registro_1")

    print(f"\nTotal: {total} filas recuperadas.")
    print("MONTO_EFECTIVO y MONTO_YAPE están vacíos — llena solo los pagos mixtos.\n")
