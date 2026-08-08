# 4_pagos/efectivo/verificar_lotes.py — ¿el MZ-LT que escribió el cobrador es el correcto?
#
# Los tres ejes que ya existen en el módulo no pueden atrapar un lote mal escrito:
# main.py cruza cobrador contra cobrador (inútil si la mesa tiene una sola hoja),
# arqueo.py cuadra la plata (cuadra perfecto: el monto está bien, el lote no) y
# reclamos.py detecta lo que el cobrador marcó (no marca nada, para él no pasó nada).
# Este cruza cada pago contra la boleta emitida y usa el monto como evidencia.
#
# Contrato visual:    docs/formato_verificacion_lotes.html
# Diagramas:          docs/diagrama_flujo_verificacion_lotes.html · docs/diagrama_verificacion_lotes.html
# Decisión de diseño: docs/decisiones/verificacion_lotes_efectivo.md
#
# Uso:
#   python verificar_lotes.py               # ciclo activo declarado por 1_lecturas
#   python verificar_lotes.py --mes 2026-08

import argparse
import logging
import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from itertools import combinations
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import entregas_repo as repo

log = logging.getLogger(__name__)

BASE_DIR = Path(__file__).parent
INPUTS_DIR = BASE_DIR / "inputs"
OUTPUTS_DIR = BASE_DIR / "outputs"
BACKUP_DIR = BASE_DIR / "backup"
SHARED_DIR = BASE_DIR.parent.parent / "shared"
USUARIOS_ID_FILE = SHARED_DIR / "usuarios_id.xlsx"
DATA_BOLETAS_FILE = BASE_DIR.parent.parent / "3_boletas" / "inputs" / "DATA_boletas.xlsx"

sys.path.insert(0, str(SHARED_DIR))
import ciclo as ciclo_activo  # noqa: E402

HOJAS = ["registro_1", "registro_2", "registro_3"]
TOL = 0.01

# Los 7 cargos de la boleta. (etiqueta corta para la columna COMO, columna en DATA_boletas).
# A escala multi-tenant esta lista es data del manifiesto de la JASS, no código
# — ver "Escala" en docs/decisiones/verificacion_lotes_efectivo.md.
CONCEPTOS = [
    ("consumo",   "Total mes actual"),
    ("mant",      "Mantenimiento"),
    ("anterior",  "MES ANTERIOR"),
    ("corte",     "Corte y reconexion"),
    ("convenio",  "Convenio"),
    ("multa",     "Multa (faena + reunión)"),
    ("cuota",     "Cuota directa"),
]

# Umbrales de la capa 2. Medidos contra 163 pagos reales del ciclo 2026-08:
# con estos valores la distribución es ALTA 25 · MEDIA 75 · BAJA 51 · NO CUADRA 6.
# Si BAJA supera el 50%, el monto dejó de discriminar y lo que falta es la
# columna NOMBRE en la hoja de papel — no afinar estos números.
RIVALES_ALTA = 5
RIVALES_MEDIA = 25

# Ventana de tolerancia después del vencimiento: un pago atrasado sigue siendo
# de esa boleta. Más allá, la boleta que tenemos es de otro ciclo.
DIAS_GRACIA = 45

# ── Tabla de confusión (capa 3) ──────────────────────────────────────────────
# Los grupos de letras son universales (cómo se confunde una U con una W no
# depende de la JASS). El sufijo "1" (A↔A1) SÍ depende de cómo esta JASS nombra
# sus manzanas: a escala eso va al manifiesto del tenant.
GRUPOS_MZ = [
    "UVW",    # verificado 2026-08: la Ц de U-2 leída como W
    "GOQC",   # trazo redondo
    "ILTJ",
    "PDOB",
    "AD",
    "BP",
    "EFG",    # verificado 2026-08: la G cuadrada de G-18/G-5/G-21 se lee como E
    "MNW",
    "SZ",
    "XYK",
    "RPB",
]
PARES_DIG = [("4", "9"), ("1", "7"), ("3", "8"), ("5", "6"), ("0", "6"),
             ("0", "8"), ("6", "8"), ("2", "7"), ("1", "4"), ("3", "5"),
             ("5", "8"), ("7", "9"), ("2", "3"), ("1", "2")]

_MZ_VEC = defaultdict(set)
for _g in GRUPOS_MZ:
    for _a in _g:
        _MZ_VEC[_a].update(c for c in _g if c != _a)

_DIG_VEC = defaultdict(set)
for _a, _b in PARES_DIG:
    _DIG_VEC[_a].add(_b)
    _DIG_VEC[_b].add(_a)

# ── Paleta (del contrato docs/formato_verificacion_lotes.html) ───────────────
_VI = ("F4ECF7", "5B21B6", "FAF5FF")   # ¿quién es?
_AM = ("FEF9E7", "7D6608", "FFFDF5")   # ¿quién cobró y cuándo?
_AZ = ("EBF5FB", "1A5276", "F4FAFF")   # ¿qué lote y cuánto?
_VE = ("E9F7EF", "1E5C3A", "F4FBF7")   # ¿cuadra con su boleta?
_NA = ("FFF7ED", "9A3412", "FFFBF5")   # ¿sospecha algún vecino?
_RE = ("ECFDF5", "065F46", "F0FFF8")   # resolución
_GR = ("F3F4F6", "374151", "F9FAFB")   # columna oculta

_EVID_COLOR = {
    "ALTA":       ("D5F5E3", "1E5C3A"),
    "MEDIA":      ("FEF3C7", "92400E"),
    "BAJA":       ("FEF9E7", "7D6608"),
    "NO CUADRA":  ("FEE2E2", "991B1B"),
    "OMITIDO":    ("F1F3F5", "868E96"),
    "SIN BOLETA": ("F1F3F5", "868E96"),
}

_COLS = [
    ("ID",               _VI, 10, "center"),
    ("NOMBRE",           _VI, 26, "left"),
    ("MESA",             _AM, 10, "center"),
    ("HOJA",             _AM, 12, "center"),
    ("COBRADOR",         _AM, 20, "left"),
    ("FECHA",            _AM, 12, "center"),
    ("MZ",               _AZ,  8, "center"),
    ("LT",               _AZ,  8, "center"),
    ("MONTO",            _AZ, 12, "right"),
    ("EVIDENCIA",        _VE, 14, "center"),
    ("COMO",             _VE, 28, "left"),
    ("IMPORTE_BOLETA",   _VE, 14, "right"),
    ("CANDIDATO",        _NA, 12, "center"),
    ("CANDIDATO_NOMBRE", _NA, 26, "left"),
    ("MOTIVO_CONFUSION", _NA, 18, "left"),
    ("RESOLUCION",       _RE, 16, "center"),
    ("MZ_CORRECTO",      _RE, 12, "center"),
    ("LT_CORRECTO",      _RE, 12, "center"),
    ("FILA_EXCEL",       _GR, 10, "center"),   # oculta — respaldo de la clave
]
_SECCIONES = [
    ("¿Quién es?",              "ID",         "NOMBRE",           _VI),
    ("¿Quién cobró y cuándo?",  "MESA",       "FECHA",            _AM),
    ("¿Qué lote y cuánto?",     "MZ",         "MONTO",            _AZ),
    ("¿Cuadra con su boleta?",  "EVIDENCIA",  "IMPORTE_BOLETA",   _VE),
    ("¿Sospecha algún vecino?", "CANDIDATO",  "MOTIVO_CONFUSION", _NA),
    ("Resolución — llenar",     "RESOLUCION", "LT_CORRECTO",      _RE),
]
COLS_HUMANAS = ("RESOLUCION", "MZ_CORRECTO", "LT_CORRECTO")

# ── Primitivos puros (sin I/O — el grueso de los tests vive acá) ─────────────

def _norm(v) -> str:
    """Normaliza MZ/LT. El '.0' aparece cuando openpyxl lee un lote numérico
    como float (DATA_boletas trae LT como número; las mesas, como texto)."""
    s = str(v).strip().upper() if v is not None else ""
    return s[:-2] if s.endswith(".0") else s


def clave(mz, lt) -> str:
    return f"{_norm(mz)}-{_norm(lt)}"


def mz_confundible(a: str, b: str) -> bool:
    """A↔A1 (sufijo que mete el autocompletar) · U↔W (trazo) · A1↔D1 (base confundible)."""
    if a == b:
        return False
    base_a, suf_a = a.rstrip("1"), a.endswith("1")
    base_b, suf_b = b.rstrip("1"), b.endswith("1")
    if base_a == base_b and suf_a != suf_b:
        return True
    return suf_a == suf_b and base_b in _MZ_VEC.get(base_a, ())


def lt_confundible(a: str, b: str) -> bool:
    """Un solo error: dígito parecido · dígito de más/de menos · transposición · sufijo."""
    if a == b:
        return False
    if len(a) == len(b):
        dif = [i for i in range(len(a)) if a[i] != b[i]]
        if len(dif) == 1 and b[dif[0]] in _DIG_VEC.get(a[dif[0]], ()):
            return True                                    # 19 ↔ 14
        if len(dif) == 2 and dif[1] == dif[0] + 1:         # 12 ↔ 21
            i = dif[0]
            return a[i] == b[i + 1] and a[i + 1] == b[i]
        return False
    if abs(len(a) - len(b)) == 1:                          # A-4 ↔ A-14 · 3 ↔ 3A
        largo, corto = (a, b) if len(a) > len(b) else (b, a)
        return any(largo[:i] + largo[i + 1:] == corto for i in range(len(largo)))
    return False


def confundible(k1: str, k2: str) -> tuple[str, int] | None:
    """(motivo, nivel) si k2 está a un error de tipeo de k1, o None.
    nivel 1 = error simple (una sola parte mal) · 2 = doble (manzana Y lote).
    El simple gana: en los 4 casos medidos el lote correcto siempre fue simple,
    y el nivel doble solo agregó ruido."""
    m1, l1 = k1.split("-", 1)
    m2, l2 = k2.split("-", 1)
    if m1 == m2 and lt_confundible(l1, l2):
        return f"lote {l1}→{l2}", 1
    if l1 == l2 and mz_confundible(m1, m2):
        return f"manzana {m1}→{m2}", 1
    if mz_confundible(m1, m2) and lt_confundible(l1, l2):
        return f"mz+lote {k1}→{k2}", 2
    return None


def subconjuntos(cargos: dict) -> dict:
    """{monto: etiqueta} de toda combinación de cargos > 0 que el predio podría pagar.
    7 conceptos → 127 combinaciones como máximo. Ante empate se queda la más corta
    (la primera que aparece): la ambigüedad no cambia el veredicto del lote, solo
    la etiqueta informativa."""
    activos = [(n, v) for n, v in cargos.items() if v > 0]
    out: dict = {}
    for n in range(1, len(activos) + 1):
        for combo in combinations(activos, n):
            monto = round(sum(v for _, v in combo), 2)
            out.setdefault(monto, "+".join(c for c, _ in combo))
    return out


def clasificar(salda_todo: bool, rivales: int) -> str:
    """Capa 2. El pago parcial baja un nivel: la ambigüedad de QUÉ pagó se suma
    a la de A QUIÉN pertenece."""
    if salda_todo:
        if rivales <= RIVALES_ALTA:
            return "ALTA"
        return "MEDIA" if rivales <= RIVALES_MEDIA else "BAJA"
    return "MEDIA" if rivales <= RIVALES_ALTA else "BAJA"

# ── Lectura ──────────────────────────────────────────────────────────────────

def leer_boletas() -> dict:
    """{clave: {nombre, total, cargos, emision, vencimiento}} desde DATA_boletas.

    Fuente deliberada: NO la planilla. Su TOTAL_A_PAGAR es una fórmula Excel y
    openpyxl/pandas la leen vacía; además la planilla sigue mutando después de
    emitida la boleta, y lo que el vecino pagó es lo que decía su recibo."""
    if not DATA_BOLETAS_FILE.exists():
        raise FileNotFoundError(
            f"Falta {DATA_BOLETAS_FILE}\n"
            f"  -> correr 3_boletas antes de verificar: sin la boleta emitida no hay "
            f"contra qué comparar el monto cobrado.")
    wb = load_workbook(DATA_BOLETAS_FILE, read_only=True, data_only=True)
    ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        raise ValueError(f"{DATA_BOLETAS_FILE.name} sin filas de datos")

    idx = {str(h).strip(): i for i, h in enumerate(rows[0]) if h}
    faltan = [c for _n, c in CONCEPTOS if c not in idx]
    if faltan:
        # No es fatal: un concepto ausente simplemente no participa de las
        # combinaciones. Sí hay que avisarlo — cambia qué montos pueden cuadrar.
        log.warning(f"DATA_boletas sin columna(s) {faltan} — se tratan como 0")
    for req in ("MZ", "LT", "Importe a pagar"):
        if req not in idx:
            raise ValueError(f"DATA_boletas sin columna requerida '{req}'")

    def g(fila, col):
        i = idx.get(col)
        return fila[i] if i is not None and i < len(fila) else None

    out = {}
    for fila in rows[1:]:
        if not fila or g(fila, "MZ") is None:
            continue
        k = clave(g(fila, "MZ"), g(fila, "LT"))
        out[k] = {
            "nombre": str(g(fila, "NOMBRES") or "").strip(),
            "total": repo._monto(g(fila, "Importe a pagar")),
            "cargos": {n: repo._monto(g(fila, c)) for n, c in CONCEPTOS},
            "emision": repo._fecha_key(g(fila, "FECHA DE EMISIÓN")),
            "vencimiento": repo._fecha_key(g(fila, "FECHA DE VENCIMIENTO")),
        }
    return out


def leer_mesas(mes: str) -> list:
    """Una entrada por fila de pago de todas las mesas del mes.

    FILA_EXCEL es la posición física en la hoja: entra en la clave de
    preservación como respaldo del caso que MZ+LT+MONTO+COBRADOR+FECHA no cubre
    — el mismo lote pagando el mismo importe dos veces el mismo día."""
    filas = []
    for n in range(1, 8):
        path = INPUTS_DIR / f"mesa_{n}.xlsx"
        if not path.exists():
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        for hoja in HOJAS:
            if hoja not in wb.sheetnames:
                continue
            rows = list(wb[hoja].values)
            if len(rows) < 4:
                continue
            headers = [str(h).strip().upper() if h else "" for h in rows[1]]
            idx = {h: i for i, h in enumerate(headers)}

            def g(fila, col):
                i = idx.get(col)
                return fila[i] if i is not None and i < len(fila) else None

            for pos, fila in enumerate(rows[3:], start=4):   # fila 3 = ejemplo guía
                if not fila or all(c is None for c in fila):
                    continue
                mz, lt = _norm(g(fila, "MZ")), _norm(g(fila, "LT"))
                if not mz or not lt:
                    continue
                fecha_key = repo._fecha_key(g(fila, "FECHA"))
                if mes and repo._mes_de(fecha_key) != mes:
                    continue
                filas.append({
                    "mesa": f"mesa_{n}",
                    "hoja": hoja,
                    "fila_excel": pos,
                    "cobrador": repo._norm_cobrador(g(fila, "COBRADOR")),
                    "fecha": fecha_key,
                    "mz": mz,
                    "lt": lt,
                    "monto": repo._monto(g(fila, "MONTO")),
                    "concepto": str(g(fila, "CONCEPTO") or "").strip(),
                })
        wb.close()
    return filas


_cache_usuarios = None

def _cargar_usuarios_id() -> dict:
    """(MZ,LT) -> (USER_ID, NOMBRE). Mismo mecanismo que buscar_quien() en main.py."""
    global _cache_usuarios
    if _cache_usuarios is not None:
        return _cache_usuarios
    mapa = {}
    if USUARIOS_ID_FILE.exists():
        wb = load_workbook(USUARIOS_ID_FILE, read_only=True, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        if len(rows) >= 2:
            hdr = {str(h).strip().upper(): i for i, h in enumerate(rows[0]) if h}

            def _get(fila, col):
                i = hdr.get(col)
                return str(fila[i]).strip() if i is not None and i < len(fila) and fila[i] else ""

            for fila in rows[1:]:
                if not fila:
                    continue
                uid, nom = _get(fila, "USER_ID"), _get(fila, "NOMBRE")
                for mz_c, lt_c in (("MZ", "LOTE"), ("MZ2", "LOTE2")):
                    mz, lt = _norm(_get(fila, mz_c)), _norm(_get(fila, lt_c))
                    if uid and mz and lt:
                        mapa.setdefault((mz, lt), (uid, nom))
    _cache_usuarios = mapa
    return mapa

# ── Guard de pipeline ────────────────────────────────────────────────────────

def guard_pipeline(filas: list, boletas: dict) -> None:
    """¿La boleta que tenemos es la del ciclo que se está cobrando?

    Se compara contra el dato real (fechas de la boleta vs fechas de los pagos)
    y no contra un archivo de estado: un JSON de "ya corrí" puede mentir si
    alguien lo edita o si un módulo corrió a medias; las fechas no."""
    emisiones = Counter(b["emision"] for b in boletas.values() if b["emision"])
    vencimientos = Counter(b["vencimiento"] for b in boletas.values() if b["vencimiento"])
    if not emisiones or not vencimientos:
        log.warning("DATA_boletas sin FECHA DE EMISIÓN/VENCIMIENTO — no se puede "
                    "validar que corresponda a este ciclo. Se continúa sin guard.")
        return

    emision = repo._parse_fecha(emisiones.most_common(1)[0][0])
    vence = repo._parse_fecha(vencimientos.most_common(1)[0][0])
    limite = vence + timedelta(days=DIAS_GRACIA)

    fechas = [repo._parse_fecha(f["fecha"]) for f in filas if f["fecha"]]
    if not fechas:
        return
    dentro = sum(1 for d in fechas if emision <= d <= limite)
    if dentro == 0:
        raise ValueError(
            f"DATA_boletas no corresponde a estos pagos.\n"
            f"  boleta   emitida {emision:%d/%m/%Y} · vence {vence:%d/%m/%Y}\n"
            f"  pagos    del {min(fechas):%d/%m/%Y} al {max(fechas):%d/%m/%Y}\n"
            f"  -> correr 3_boletas para el ciclo en curso antes de verificar.")
    if dentro < len(fechas):
        log.warning(f"{len(fechas) - dentro} de {len(fechas)} pagos caen fuera de la "
                    f"ventana de la boleta ({emision:%d/%m/%Y}–{limite:%d/%m/%Y}) — "
                    f"se verifican igual, pero revisá si son de otro ciclo.")

# ── Motor: las 4 capas ───────────────────────────────────────────────────────

def verificar(filas: list, boletas: dict) -> list:
    """Devuelve las filas con EVIDENCIA/COMO/CANDIDATO resueltos."""
    # Índices del pueblo, una sola vez.
    subconj = {k: subconjuntos(b["cargos"]) for k, b in boletas.items()}
    rivales_total = Counter(round(b["total"], 2) for b in boletas.values())
    pueden_pagar = defaultdict(list)          # monto -> [clave] (total o parcial)
    for k, mapa in subconj.items():
        for monto in mapa:
            pueden_pagar[monto].append(k)

    # ── Pase 1: capas 1 y 2 ──
    for f in filas:
        k = clave(f["mz"], f["lt"])
        f["clave"] = k
        f["candidato"] = f["candidato_nombre"] = f["motivo"] = ""

        if f["monto"] == 0 or f["concepto"]:
            # Plata que no es deuda de agua (aporte al tanque, honorario, gasto)
            # o visita sin cobro: nunca va a cuadrar contra la boleta. Marcarla
            # NO CUADRA sería ruido permanente — es el mismo campo CONCEPTO que
            # quedó sin marcar en el bug de A-4 (2026-08-06).
            f["evidencia"], f["como"], f["importe"] = "OMITIDO", "", None
            continue

        bo = boletas.get(k)
        if bo is None:
            f["evidencia"], f["como"], f["importe"] = "SIN BOLETA", "", None
            continue

        f["importe"] = bo["total"]
        como = subconj[k].get(round(f["monto"], 2))
        if como is None:
            f["evidencia"], f["como"] = "NO CUADRA", ""
        else:
            salda_todo = abs(f["monto"] - bo["total"]) < TOL
            f["como"] = "boleta completa" if salda_todo else como
            f["evidencia"] = clasificar(salda_todo, rivales_total.get(round(bo["total"], 2), 0))

    # ── Pase 2: capas 3 y 4 (necesitan el resultado de todas las filas) ──
    ya_pagados = {f["clave"] for f in filas if f["evidencia"] in ("ALTA", "MEDIA")}
    for f in filas:
        if f["evidencia"] != "NO CUADRA":
            continue
        k = f["clave"]
        cands = []
        for otro in pueden_pagar.get(round(f["monto"], 2), []):
            if otro == k or otro in ya_pagados:   # capa 4: el candidato debe estar impago
                continue
            hit = confundible(k, otro)
            if hit:
                cands.append((hit[1], otro, hit[0]))
        # Solo se propone un error SIMPLE. Un candidato con la manzana Y el lote
        # mal a la vez tiene tan poca evidencia que proponerlo es inventar con
        # cara de certeza: en los 4 casos medidos el correcto siempre fue simple
        # y el nivel doble solo aportó ruido (G-17→Q-12, Q-14→C-44 en el ciclo
        # 2026-08 — ninguno tenía nada que ver).
        simples = [c for c in cands if c[0] == 1]
        if len(simples) == 1:
            _n, otro, motivo = simples[0]
            f["candidato"] = otro
            f["candidato_nombre"] = boletas[otro]["nombre"]
            f["motivo"] = motivo
        elif simples:
            f["motivo"] = f"{len(simples)} candidatos"
        elif cands:
            f["motivo"] = f"{len(cands)} solo con doble error"
    return filas

# ── Preservación del trabajo manual (3 capas) ───────────────────────────────

def _clave_preserv(f: dict) -> tuple:
    """(MESA, COBRADOR, FECHA, MZ, LT, MONTO, FILA_EXCEL).

    Medida contra los 165 pagos reales del ciclo 2026-08: 0 colisiones con los
    primeros 6 campos. FILA_EXCEL entra como respaldo puro del caso que esos 6
    no cubren. La versión sin MZ/LT colisionaba en 62% de las filas."""
    return (f["mesa"], f["cobrador"], f["fecha"], f["mz"], f["lt"],
            round(f["monto"], 2), f["fila_excel"])


def leer_resoluciones(ruta: Path) -> dict:
    """Capa 2 de la preservación: lo que el supervisor escribió en la corrida
    anterior. Solo se preservan filas con AL MENOS un campo lleno — una fila que
    el humano miró y dejó vacía no debe ensuciar el mapa (mismo criterio que
    _leer_pendientes_preservados en motor_matching)."""
    if not ruta.exists():
        return {}
    try:
        wb = load_workbook(ruta, read_only=True, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        log.warning(f"No se pudo leer el reporte anterior para preservar: {e}")
        return {}
    if len(rows) < 3:
        return {}
    hdr = {str(h).strip().upper(): i for i, h in enumerate(rows[1]) if h}

    def g(fila, col):
        i = hdr.get(col)
        return fila[i] if i is not None and i < len(fila) else None

    out = {}
    for fila in rows[2:]:
        if not fila or all(c is None for c in fila):
            continue
        humanas = {c: str(g(fila, c) or "").strip() for c in COLS_HUMANAS}
        if not any(humanas.values()):
            continue
        k = (str(g(fila, "MESA") or ""), repo._norm_cobrador(g(fila, "COBRADOR")),
             repo._fecha_key(g(fila, "FECHA")), _norm(g(fila, "MZ")), _norm(g(fila, "LT")),
             repo._monto(g(fila, "MONTO")), int(g(fila, "FILA_EXCEL") or 0))
        out[k] = humanas
    return out


def _backup(ruta: Path) -> None:
    """Capa 1: copia antes de reescribir. Sin esto, un fallo a mitad de la
    escritura deja al supervisor sin punto de restauración."""
    if not ruta.exists():
        return
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = BACKUP_DIR / f"{ruta.stem}_{ts}{ruta.suffix}"
    shutil.copy2(ruta, destino)
    log.info(f"backup -> {destino.name}")

# ── Escritura ────────────────────────────────────────────────────────────────

def _argb(hex6):
    return "FF" + hex6.lstrip("#")


def _fill(hex6):
    return PatternFill("solid", fgColor=_argb(hex6))


def _hdr(cell, bg, fg, texto):
    cell.value = texto
    cell.fill = _fill(bg)
    cell.font = Font(color=_argb(fg), bold=True, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def escribir(filas: list, ruta: Path, previas: dict) -> int:
    """Escribe el reporte. Devuelve cuántas resoluciones humanas se preservaron."""
    quien = _cargar_usuarios_id()
    wb = Workbook()
    ws = wb.active
    ws.title = "verificacion"

    col_idx = {c[0]: i + 1 for i, c in enumerate(_COLS)}
    for label, ini, fin, sec in _SECCIONES:
        c1, c2 = col_idx[ini], col_idx[fin]
        if c1 != c2:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        _hdr(ws.cell(row=1, column=c1), sec[0], sec[1], label)
    ws.row_dimensions[1].height = 18
    for i, (nombre, sec, ancho, _a) in enumerate(_COLS, start=1):
        _hdr(ws.cell(row=2, column=i), sec[0], sec[1], nombre)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[2].height = 22
    # FILA_EXCEL existe para que la preservación nunca falle, pero el supervisor
    # busca por MZ — que no la vea ni la use.
    ws.column_dimensions[get_column_letter(col_idx["FILA_EXCEL"])].hidden = True
    ws.freeze_panes = "A3"

    n_preservadas = 0
    for r_off, f in enumerate(filas):
        row = r_off + 3
        uid, nom = quien.get((f["mz"], f["lt"]), ("", ""))
        humanas = previas.get(_clave_preserv(f), {})
        if humanas:
            n_preservadas += 1
        valores = {
            "ID": uid or "—", "NOMBRE": nom or "—",
            "MESA": f["mesa"], "HOJA": f["hoja"],
            "COBRADOR": f["cobrador"], "FECHA": f["fecha"],
            "MZ": f["mz"], "LT": f["lt"], "MONTO": f["monto"],
            "EVIDENCIA": f["evidencia"], "COMO": f["como"] or "—",
            "IMPORTE_BOLETA": f["importe"] if f["importe"] is not None else "—",
            "CANDIDATO": f["candidato"] or "—",
            "CANDIDATO_NOMBRE": f["candidato_nombre"] or "—",
            "MOTIVO_CONFUSION": f["motivo"] or "—",
            "RESOLUCION": humanas.get("RESOLUCION", ""),
            "MZ_CORRECTO": humanas.get("MZ_CORRECTO", ""),
            "LT_CORRECTO": humanas.get("LT_CORRECTO", ""),
            "FILA_EXCEL": f["fila_excel"],
        }
        for i, (nombre, sec, _ancho, align) in enumerate(_COLS, start=1):
            valor = valores[nombre]
            cell = ws.cell(row=row, column=i)
            if nombre == "FECHA" and valor:
                cell.value = repo._parse_fecha(valor)
                cell.number_format = "DD/MM/YYYY"
            else:
                cell.value = valor
                if nombre in ("MONTO", "IMPORTE_BOLETA") and valor != "—":
                    cell.number_format = '"S/ "#,##0.00'
            cell.fill = _fill(sec[2])
            cell.alignment = Alignment(horizontal=align, vertical="center")
            if nombre == "EVIDENCIA":
                bg, fg = _EVID_COLOR.get(valor, (sec[2], sec[1]))
                cell.fill = _fill(bg)
                cell.font = Font(color=_argb(fg), bold=True, size=10)
            elif nombre in COLS_HUMANAS and valor:
                cell.fill = _fill("D1FAE5")
                cell.font = Font(color=_argb("065F46"), bold=True, size=10)
            else:
                cell.font = Font(color=_argb(sec[1]), size=10)

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)
    return n_preservadas

# ── Orquestación ─────────────────────────────────────────────────────────────

def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    ap = argparse.ArgumentParser(
        description="Verifica que el MZ-LT escrito en la mesa sea el dueño del monto cobrado")
    ap.add_argument("--mes", default=None, help="AAAA-MM (default: ciclo activo)")
    args = ap.parse_args()

    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(OUTPUTS_DIR / "verificar_lotes_run.log", encoding="utf-8"),
        ],
        force=True,
    )

    # Guard, capa 1: sin ciclo declarado no se sabe qué mes verificar.
    mes = args.mes or ciclo_activo.activo(default=None,
                                          path=SHARED_DIR / "ciclo_activo.json")
    if not mes:
        log.error("No hay ciclo activo declarado (shared/ciclo_activo.json).\n"
                  "  -> correr 1_lecturas, o pasar --mes AAAA-MM")
        sys.exit(1)

    log.info(f"Verificación de lotes — ciclo {mes}")
    boletas = leer_boletas()
    filas = leer_mesas(mes)
    if not filas:
        log.warning(f"Sin pagos de efectivo para {mes}. No se genera reporte.")
        return
    log.info(f"{len(filas)} pagos · {len(boletas)} boletas emitidas")

    guard_pipeline(filas, boletas)      # guard, capa 2
    filas = verificar(filas, boletas)

    ruta = OUTPUTS_DIR / f"verificacion_lotes_{mes}.xlsx"
    previas = leer_resoluciones(ruta)
    _backup(ruta)
    n_preservadas = escribir(filas, ruta, previas)

    conteo = Counter(f["evidencia"] for f in filas)
    resumen = " · ".join(f"{k}={conteo[k]}" for k in
                         ("ALTA", "MEDIA", "BAJA", "NO CUADRA", "OMITIDO", "SIN BOLETA")
                         if conteo.get(k))
    log.info(f"verificacion_lotes_{mes}.xlsx — {len(filas)} filas · {resumen}")
    if n_preservadas:
        log.info(f"{n_preservadas} resolución(es) del supervisor preservadas")

    sin_resolver = [f for f in filas if f["evidencia"] == "NO CUADRA"]
    con_cand = [f for f in sin_resolver if f["candidato"]]
    if sin_resolver:
        log.warning(f"{len(sin_resolver)} fila(s) NO CUADRA — revisar antes de correr main.py"
                    + (f" ({len(con_cand)} con candidato único propuesto)" if con_cand else ""))
        for f in sin_resolver:
            extra = (f" -> ¿{f['candidato']} {f['candidato_nombre']}? [{f['motivo']}]"
                     if f["candidato"] else "")
            log.warning(f"   {f['mesa']} {f['clave']:>7} S/{f['monto']:>7.2f} "
                        f"(boleta S/{f['importe']:.2f}){extra}")

    baja = conteo.get("BAJA", 0)
    evaluables = sum(conteo[k] for k in ("ALTA", "MEDIA", "BAJA")) or 1
    if baja / evaluables > 0.5:
        log.warning(f"{baja}/{evaluables} filas con evidencia BAJA (>50%): el monto dejó "
                    f"de discriminar. Lo que falta es la columna NOMBRE en la hoja del "
                    f"cobrador, no afinar el algoritmo.")

    log.info(f"-> {ruta}")


if __name__ == "__main__":
    main()
