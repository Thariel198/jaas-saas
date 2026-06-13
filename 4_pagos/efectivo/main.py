# 4_pagos/efectivo — consolidación de cobros en efectivo
# Contratos visuales: docs/formato_registro.html
#                     docs/formato_pagos_efectivo.html
#                     docs/formato_discrepancias.html
#                     docs/formato_trazabilidad.html

import logging
import shutil
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Rutas ───────────────────────────────────────────────────────────────────

BASE_DIR    = Path(__file__).parent
INPUTS_DIR  = BASE_DIR / "inputs"
OUTPUTS_DIR = BASE_DIR / "outputs"
TRAZAB_DIR  = BASE_DIR / "trazabilidad"
DISC_FILE   = OUTPUTS_DIR / "discrepancias.xlsx"
OUTPUT_FILE = OUTPUTS_DIR / "pagos_efectivo.xlsx"

HOJAS_VALIDAS = ["registro_1", "registro_2", "registro_3"]

# ── Paleta (del contrato docs/formato_*.html) ────────────────────────────────

# pagos_efectivo — secciones (header bg, header txt, data bg)
PE = {
    "id":   ("EBF5FB", "1A5276", "F4FAFF"),   # MZ, LT
    "cob":  ("E9F7EF", "1E5C3A", "F4FBF7"),   # MONTO, FECHA
    "mesa": ("FEF9E7", "7D6608", "FFFDF5"),   # MESA
    "est":  ("F3E8FF", "5B21B6", "FAF5FF"),   # ESTADO
    "who":  ("EBF5FB", "1A5276", "F4FAFF"),   # COBRADOR
    "com":  ("F4ECF7", "5B21B6", "FAF5FF"),   # COMENTARIO
    "meta": ("F3E8FF", "5B21B6", "FAF5FF"),   # CICLO_CORRECCION (trazabilidad)
}

# discrepancias — secciones
DC = {
    "id":    ("EBF5FB", "1A5276", "F4FAFF"),  # MZ, LT
    "mesa":  ("FEF9E7", "7D6608", "FFFDF5"),  # MESA
    "disc":  ("FEF2F2", "991B1B", "FFF5F5"),  # COBRADOR_X, MONTO_X
    "cobro": ("E9F7EF", "1E5C3A", "F4FBF7"),  # MESA, COBRADOR, MONTO, FECHA (hoja pago_multi_mesa)
    "res":   ("ECFDF5", "065F46", "F0FFF8"),  # RESOLUCION, MONTO_CORRECTO
}
DC_RES_LLENO  = "D1FAE5"   # bg RESOLUCION cuando tiene valor
DC_MONTO_DISC = "B91C1C"   # txt montos discrepantes (más saturado)

# trazabilidad — secciones (header bg, header txt, data bg)
TR_DR = {  # discrepancias_resueltas
    "pred": ("EBF5FB", "1A5276", "F4FAFF"),
    "mesa": ("FEF9E7", "7D6608", "FFFDF5"),
    "disc": ("FEF2F2", "991B1B", "FFF5F5"),
    "res":  ("ECFDF5", "065F46", "F0FFF8"),
    "meta": ("F3E8FF", "5B21B6", "FAF5FF"),
}
TR_RES_LLENA = "D1FAE5"

TR_SC = {  # solo_un_cobrador y pago_multi_mesa_resueltas
    "orig":  ("FEF2F2", "991B1B", "FFF5F5"),  # predio original registrado (rojo)
    "pred":  ("EBF5FB", "1A5276", "F4FAFF"),  # predio imputado (azul)
    "mesa":  ("FEF9E7", "7D6608", "FFFDF5"),
    "cobro": ("E9F7EF", "1E5C3A", "F4FBF7"),
    "res":   ("ECFDF5", "065F46", "F0FFF8"),  # RESOLUCION (ok/rechaza)
    "meta":  ("F3E8FF", "5B21B6", "FAF5FF"),
}
TR_OK_BG      = "D1FAE5"   # bg RESOLUCION = ok
TR_RECHAZA_BG = "FEE2E2"   # bg RESOLUCION = rechaza


# ── Helpers de estilo ────────────────────────────────────────────────────────

def _borde(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, bg, txt, texto, bold=True, size=9):
    cell.value     = texto
    cell.font      = Font(name="Arial", bold=bold, size=size, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _borde("FFFFFF")

def _dat(cell, valor, bg, txt, align="left", bold=False, fmt=None, size=10):
    cell.value     = valor
    cell.font      = Font(name="Arial", bold=bold, size=size, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _borde()
    if fmt:
        cell.number_format = fmt

def _ancho(ws, col_idx, ancho):
    ws.column_dimensions[get_column_letter(col_idx)].width = ancho


# ── Utilidades de datos ──────────────────────────────────────────────────────

def _norm(val) -> str:
    return str(val).strip().upper() if val is not None else ""

def _monto(val) -> float:
    try:
        return round(float(str(val).replace(",", ".").strip()), 2)
    except Exception:
        return 0.0

def _to_datetime(val):
    """Convierte 'DD/MM/YYYY' o 'YYYY-MM-DD' a datetime. Si falla, devuelve el valor tal cual."""
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return s


def _leer_hoja_traz(path: Path, hoja: str) -> list:
    """Lee una hoja de trazabilidad existente. Devuelve lista de dicts."""
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    if hoja not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[hoja]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 3:
        return []
    headers = [str(h).strip().upper() if h else "" for h in rows[1]]
    out = []
    for r in rows[2:]:
        if not r or all(c is None for c in r):
            continue
        d = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
        # migración: normalizar columna legacy CICLO → CICLO_CORRECCION
        if "CICLO" in d and "CICLO_CORRECCION" not in d:
            d["CICLO_CORRECCION"] = d.pop("CICLO")
        out.append(d)
    return out


def detectar_ciclo(traz_file: Path) -> int:
    """
    Ciclo 1 si trazabilidad no existe. Si existe, max(CICLO_CORRECCION) + 1.
    Acepta también la columna legacy "CICLO" para migración de archivos existentes.
    """
    if not traz_file.exists():
        return 1
    wb = load_workbook(traz_file, read_only=True, data_only=True)
    max_ciclo = 0
    for hoja in wb.sheetnames:
        ws = wb[hoja]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        headers = [str(h).strip().upper() if h else "" for h in rows[1]]
        # acepta CICLO_CORRECCION (nuevo) o CICLO (legacy)
        if "CICLO_CORRECCION" in headers:
            idx = headers.index("CICLO_CORRECCION")
        elif "CICLO" in headers:
            idx = headers.index("CICLO")
        else:
            continue
        for r in rows[2:]:
            if r and len(r) > idx and r[idx] is not None:
                try:
                    max_ciclo = max(max_ciclo, int(r[idx]))
                except (ValueError, TypeError):
                    pass
    wb.close()
    return max_ciclo + 1 if max_ciclo > 0 else 1


def _leer_multi_mesa_resueltas_map(traz_file: Path) -> dict:
    """
    Lee trazabilidad pago_multi_mesa_resueltas y devuelve mapa de resoluciones previas.
    Key: (mz_orig, lt_orig, mesa, cobrador, round(monto, 2))
    Value: {"resolucion", "mz_imputado", "lt_imputado", "ciclo_original"}
    Permite que los cobros resueltos en ciclos previos no vuelvan a discrepancias.xlsx.
    """
    mapa = {}
    if not traz_file.exists():
        return mapa
    filas = _leer_hoja_traz(traz_file, "pago_multi_mesa_resueltas")
    for f in filas:
        mz_orig    = _norm(f.get("MZ_ORIGINAL"))
        lt_orig    = _norm(f.get("LT_ORIGINAL"))
        mesa       = str(f.get("MESA") or "").strip()
        cobrador   = str(f.get("COBRADOR") or "").strip()
        monto      = _monto(f.get("MONTO", 0))
        mz_imp     = _norm(f.get("MZ"))
        lt_imp     = _norm(f.get("LT"))
        resolucion = str(f.get("RESOLUCION") or "").strip().lower()
        ciclo      = f.get("CICLO_CORRECCION")
        if not (mz_orig and lt_orig and mesa and resolucion in ("si", "rechaza")):
            continue
        try:
            ciclo_int = int(ciclo) if ciclo is not None else 1
        except (ValueError, TypeError):
            ciclo_int = 1
        key = (mz_orig, lt_orig, mesa, cobrador, round(monto, 2))
        mapa[key] = {
            "resolucion":     resolucion,
            "mz_imputado":    mz_imp or mz_orig,
            "lt_imputado":    lt_imp or lt_orig,
            "ciclo_original": ciclo_int,
        }
    return mapa


def _construir_mapa_ciclo_original(traz_file: Path) -> dict:
    """
    Mapa (mz, lt, mesa, round(monto, 2)) → ciclo_correccion en que el cobro
    entró por última vez a pagos_efectivo (no el ciclo del run actual).

    Semántica: el "ciclo original" es el último ciclo en que el cobro fue VALIDADO
    para pagos. Si una fila apareció en Ciclo 1 como solo_un_cobrador pero quedó
    pendiente como multi_mesa, y se resolvió en Ciclo 2, su ciclo "oficial" es 2
    (cuando entró efectivamente a pagos_efectivo). Por eso se toma el MAX entre hojas.
    """
    mapa = {}
    if not traz_file.exists():
        return mapa
    for hoja in ("solo_un_cobrador", "discrepancias_resueltas", "pago_multi_mesa_resueltas"):
        filas = _leer_hoja_traz(traz_file, hoja)
        for f in filas:
            mz   = _norm(f.get("MZ"))
            lt   = _norm(f.get("LT"))
            mesa = str(f.get("MESA") or "").strip()
            # discrepancias_resueltas guarda MONTO_FINAL; las otras MONTO
            monto_raw = f.get("MONTO_FINAL") if "MONTO_FINAL" in f else f.get("MONTO")
            try:
                monto = round(float(monto_raw or 0), 2)
            except (ValueError, TypeError):
                continue
            ciclo_val = f.get("CICLO_CORRECCION")
            if ciclo_val is None or not (mz and lt and mesa):
                continue
            try:
                ciclo_int = int(ciclo_val)
            except (ValueError, TypeError):
                continue
            key = (mz, lt, mesa, monto)
            # Mantener el MAX: refleja el ciclo más reciente de validación oficial
            if key not in mapa or ciclo_int > mapa[key]:
                mapa[key] = ciclo_int
    return mapa


def _aplicar_multi_mesa_previas(confirmados: list, prev_map: dict) -> int:
    """
    Marca los confirmados que ya fueron resueltos en ciclos previos:
      - resolucion='si'      → mz/lt corregidos al imputado, estado='multi_mesa_resuelta',
                                ciclo_correccion=ciclo original, flag _ya_resuelto=True
      - resolucion='rechaza' → flag _ya_rechazado=True (se descartan en detectar_multi_mesa)
    Devuelve cantidad de filas marcadas.
    """
    marcadas = 0
    for r in confirmados:
        key = (
            _norm(r.get("mz", "")),
            _norm(r.get("lt", "")),
            str(r.get("mesa", "")).strip(),
            str(r.get("cobrador", "")).strip(),
            round(_monto(r.get("monto", 0)), 2),
        )
        prev = prev_map.get(key)
        if not prev:
            continue
        if prev["resolucion"] == "si":
            r["mz"]               = prev["mz_imputado"]
            r["lt"]               = prev["lt_imputado"]
            r["llave"]            = _llave(r["mz"], r["lt"])
            r["estado"]           = "multi_mesa_resuelta"
            r["ciclo_correccion"] = prev["ciclo_original"]
            r["_ya_resuelto"]     = True
            marcadas += 1
        elif prev["resolucion"] == "rechaza":
            r["_ya_rechazado"] = True
            marcadas += 1
    return marcadas


def _llave(mz, lt) -> str:
    return f"{_norm(mz)}-{_norm(lt)}"

def _mes_de_datos(rows: list) -> str:
    for r in rows:
        f = str(r.get("fecha", "")).strip()
        if len(f) >= 7:
            if "/" in f:
                partes = f.split("/")
                if len(partes) == 3:
                    return f"{partes[2]}-{partes[1].zfill(2)}"
            elif "-" in f:
                return f[:7]
    hoy = date.today()
    return f"{hoy.year}-{hoy.month:02d}"


# ── Leer inputs ──────────────────────────────────────────────────────────────

def leer_hoja(path: Path, hoja: str) -> list:
    """Lee una hoja de mesa_N.xlsx. Devuelve lista de registros válidos."""
    wb = load_workbook(path, read_only=True, data_only=True)
    if hoja not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[hoja]
    filas = list(ws.values)
    wb.close()

    if len(filas) < 4:   # secciones + columnas + ejemplo + al menos 1 dato
        return []

    # Fila 1 = secciones, Fila 2 = columnas, Fila 3 = ejemplo guía, Fila 4+ = datos
    headers = [str(h).strip().upper() if h else "" for h in filas[1]]
    cobrador_default = ""
    registros = []

    for fila in filas[3:]:
        if not fila or all(c is None for c in fila):
            continue
        row = {headers[i]: fila[i] for i in range(min(len(headers), len(fila)))}
        mz         = _norm(row.get("MZ"))
        lt         = _norm(row.get("LT"))
        monto_efec = _monto(row.get("MONTO_EFECTIVO"))
        if not mz or not lt or monto_efec <= 0:
            continue
        monto_total = _monto(row.get("MONTO"))
        monto_yape  = _monto(row.get("MONTO_YAPE"))
        if abs(monto_efec + monto_yape - monto_total) > 0.01:
            logging.warning(
                f"  [{path.name}/{hoja}] MZ={mz} LT={lt}: "
                f"MONTO_EFECTIVO({monto_efec})+MONTO_YAPE({monto_yape}) ≠ MONTO({monto_total})"
            )
        cobrador  = str(row.get("COBRADOR") or cobrador_default).strip()
        fecha_raw = row.get("FECHA")
        if isinstance(fecha_raw, datetime):
            fecha = fecha_raw.strftime("%d/%m/%Y")
        else:
            fecha = str(fecha_raw or "").strip()
        coment = str(row.get("COMENTARIO") or "").strip()
        registros.append({
            "llave":      _llave(mz, lt),
            "mz":         mz,
            "lt":         lt,
            "monto":      monto_efec,
            "fecha":      fecha,
            "cobrador":   cobrador,
            "comentario": coment,
        })

    return registros


def leer_mesas() -> list:
    """Devuelve lista de mesas con sus registros por hoja."""
    mesas = []
    for n in range(1, 8):
        path = INPUTS_DIR / f"mesa_{n}.xlsx"
        if not path.exists():
            continue
        hojas = {}
        for hoja in HOJAS_VALIDAS:
            regs = leer_hoja(path, hoja)
            if regs:
                hojas[hoja] = regs
        if hojas:
            mesas.append({"nombre": f"mesa_{n}", "hojas": hojas})

    if not mesas:
        raise FileNotFoundError(
            f"No se encontró ningún mesa_N.xlsx con datos en {INPUTS_DIR}"
        )
    return mesas


# ── Cross-check por mesa ─────────────────────────────────────────────────────

LETRAS = ["a", "b", "c"]


def cross_check_mesa(mesa: dict, ciclo: int, fecha_run: datetime) -> tuple:
    """
    Devuelve (confirmados, discrepancias, resueltas_auto, solo_cob).
    - confirmados: van a pagos_efectivo.xlsx
    - discrepancias: requieren resolución manual del supervisor
    - resueltas_auto: resoluciones automáticas por mayoría 2/3 (para trazabilidad)
    - solo_cob: cobros de mesas con 1 solo cobrador (para trazabilidad)
    """
    nombre  = mesa["nombre"]
    hojas   = mesa["hojas"]
    n_hojas = len(hojas)

    confirmados    = []
    discrepancias  = []
    resueltas_auto = []
    solo_cob       = []

    # 1 sola hoja → solo_un_cobrador
    if n_hojas == 1:
        regs = list(hojas.values())[0]
        for r in regs:
            confirmados.append({**r, "mesa": nombre, "estado": "solo_un_cobrador",
                                "ciclo_correccion": ciclo})
            solo_cob.append({
                "MZ":               r["mz"],
                "LT":               r["lt"],
                "MESA":             nombre,
                "COBRADOR":         r["cobrador"],
                "MONTO":            r["monto"],
                "FECHA_COBRO":      _to_datetime(r["fecha"]),
                "COMENTARIO":       r.get("comentario", ""),
                "CICLO_CORRECCION": ciclo,
            })
        return confirmados, discrepancias, resueltas_auto, solo_cob

    # 2-3 hojas → cross-check
    listas  = list(hojas.values())
    indices = [{r["llave"]: r for r in lst} for lst in listas]
    todas   = sorted(set().union(*[idx.keys() for idx in indices]))

    for llave in todas:
        presentes = [(i, idx[llave]) for i, idx in enumerate(indices) if llave in idx]
        n = len(presentes)

        if n == 1:
            _, r_pres = presentes[0]
            discrepancias.append(_hacer_disc(r_pres, [r for _, r in presentes], n_hojas, nombre))
            continue

        grupos = {}
        for i_pres, r_pres in presentes:
            grupos.setdefault(round(r_pres["monto"], 2), []).append((i_pres, r_pres))

        if len(grupos) == 1:
            r = presentes[0][1]
            confirmados.append({**r, "mesa": nombre, "estado": "confirmado",
                                "ciclo_correccion": ciclo})

        elif n_hojas == 3 and len(grupos) == 2:
            mayoria = [(monto, regs) for monto, regs in grupos.items() if len(regs) >= 2]
            if mayoria:
                monto_ganador, regs_ganadoras = mayoria[0]
                r_ganador = regs_ganadoras[0][1]
                confirmados.append({**r_ganador, "mesa": nombre, "estado": "mayoria_aplicada",
                                    "ciclo_correccion": ciclo})

                cobs  = ["", "", ""]
                monts = ["", "", ""]
                for i_pres, r_pres in presentes:
                    cobs[i_pres]  = r_pres["cobrador"]
                    monts[i_pres] = r_pres["monto"]
                letras_may = sorted(LETRAS[i] for i, _ in regs_ganadoras)
                resolucion = "mayoria_" + "_".join(letras_may)

                resueltas_auto.append({
                    "MZ":               r_ganador["mz"],
                    "LT":               r_ganador["lt"],
                    "MESA":             nombre,
                    "COBRADOR_A":       cobs[0],  "MONTO_A": monts[0],
                    "COBRADOR_B":       cobs[1],  "MONTO_B": monts[1],
                    "COBRADOR_C":       cobs[2],  "MONTO_C": monts[2],
                    "RESOLUCION":       resolucion,
                    "MONTO_FINAL":      monto_ganador,
                    "CICLO_CORRECCION": ciclo,
                    "FECHA_RESOLUCION": fecha_run,
                })
            else:
                discrepancias.append(_hacer_disc(presentes[0][1], [r for _, r in presentes], n_hojas, nombre))
        else:
            discrepancias.append(_hacer_disc(presentes[0][1], [r for _, r in presentes], n_hojas, nombre))

    return confirmados, discrepancias, resueltas_auto, solo_cob


def _hacer_disc(base: dict, presentes: list, n_hojas: int, mesa: str) -> dict:
    d = {
        "mz":    base["mz"],
        "lt":    base["lt"],
        "llave": base["llave"],
        "mesa":  mesa,
        "fecha": base["fecha"],
        "cobrador_a": "", "monto_a": "",
        "cobrador_b": "", "monto_b": "",
        "cobrador_c": "", "monto_c": "",
        "resolucion": "",
        "monto_correcto": "",
    }
    for i, r in enumerate(presentes[:3]):
        letra = LETRAS[i]
        d[f"cobrador_{letra}"] = r["cobrador"]
        d[f"monto_{letra}"]    = r["monto"]
    return d


# ── Detectar pago en múltiples mesas ────────────────────────────────────────

def detectar_multi_mesa(confirmados: list) -> tuple:
    """
    Separa filas donde el mismo MZ+LT aparece en mesas distintas.
    Las filas multi_mesa NO entran a pagos_efectivo — van a discrepancias.xlsx
    hoja pago_multi_mesa para que el supervisor marque cuál es el cobro válido.

    Respeta marcas de resolución previa:
      - r["_ya_resuelto"] → bypassa detección, va directo a limpios con sus datos imputados
      - r["_ya_rechazado"] → se descarta (no va a limpios ni a multi_mesa)
    """
    limpios_ya_resueltos = []
    procesables          = []
    for r in confirmados:
        if r.get("_ya_rechazado"):
            continue  # descartado por resolución previa
        if r.get("_ya_resuelto"):
            limpios_ya_resueltos.append(r)
            continue
        procesables.append(r)

    desde = {}
    for r in procesables:
        desde.setdefault(r["llave"], []).append(r)

    limpios    = []
    multi_mesa = []
    for llave, rows in desde.items():
        mesas_distintas = {r["mesa"] for r in rows}
        if len(mesas_distintas) > 1:
            for r in rows:
                multi_mesa.append({
                    "llave":    r["llave"],
                    "mz":       r["mz"],
                    "lt":       r["lt"],
                    "mesa":     r["mesa"],
                    "cobrador": r["cobrador"],
                    "monto":    r["monto"],
                    "fecha":    r["fecha"],
                })
        else:
            limpios.extend(rows)

    limpios.extend(limpios_ya_resueltos)
    return limpios, multi_mesa


# ── Leer y aplicar resoluciones del supervisor ───────────────────────────────

def leer_resoluciones() -> dict:
    """
    Lee RESOLUCION y MONTO_CORRECTO de discrepancias.xlsx hoja 'discrepancias'.
    Devuelve {llave: {resolucion, monto_correcto, cobrador_a/b/c, monto_a/b/c, fecha}}.
    """
    if not DISC_FILE.exists():
        return {}
    wb = load_workbook(DISC_FILE, read_only=True, data_only=True)
    if "discrepancias" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["discrepancias"]
    filas = list(ws.values)
    wb.close()

    if len(filas) < 3:
        return {}

    headers = [str(h).strip().upper() if h else "" for h in filas[1]]
    resoluciones = {}
    for fila in filas[2:]:
        if not fila:
            continue
        row = {headers[i]: fila[i] for i in range(min(len(headers), len(fila)))}
        mz  = _norm(row.get("MZ"))
        lt  = _norm(row.get("LT"))
        if not mz or not lt:
            continue
        llave = _llave(mz, lt)
        res   = str(row.get("RESOLUCION") or "").strip().lower()
        mc    = _monto(row.get("MONTO_CORRECTO", 0))
        resoluciones[llave] = {
            "resolucion":     res,
            "monto_correcto": mc,
            "monto_a": _monto(row.get("MONTO_A", 0)),
            "monto_b": _monto(row.get("MONTO_B", 0)),
            "monto_c": _monto(row.get("MONTO_C", 0)),
            "cobrador_a": str(row.get("COBRADOR_A") or "").strip(),
            "fecha":      str(row.get("FECHA") or "").strip(),
            "mesa":       str(row.get("MESA") or "").strip(),
        }
    return resoluciones


def aplicar_resoluciones(discrepancias: list, resoluciones: dict,
                         ciclo: int, fecha_run: datetime) -> tuple:
    """
    Aplica resoluciones del supervisor a las discrepancias pendientes.
    Devuelve (confirmados_resueltos, resueltas_traz, pendientes).
    """
    confirmados_resueltos = []
    resueltas_traz        = []
    pendientes            = []

    for d in discrepancias:
        res   = resoluciones.get(d["llave"], {})
        valor = res.get("resolucion", "").lower()

        if valor in ("acepta_a", "acepta_b", "acepta_c"):
            letra    = valor[-1]
            monto    = _monto(d.get(f"monto_{letra}", 0)) or _monto(res.get(f"monto_{letra}", 0))
            cobrador = d.get(f"cobrador_{letra}", "") or res.get("cobrador_a", "")
            confirmados_resueltos.append({
                "llave": d["llave"], "mz": d["mz"], "lt": d["lt"],
                "monto": monto, "fecha": d["fecha"], "mesa": d["mesa"],
                "cobrador": cobrador, "comentario": "",
                "estado": "discrepancia_resuelta",
                "ciclo_correccion": ciclo,
            })
            resueltas_traz.append(_traz_row(d, valor, monto, ciclo, fecha_run))

        elif valor == "corrige":
            mc = _monto(d.get("monto_correcto", 0)) or _monto(res.get("monto_correcto", 0))
            confirmados_resueltos.append({
                "llave": d["llave"], "mz": d["mz"], "lt": d["lt"],
                "monto": mc, "fecha": d["fecha"], "mesa": d["mesa"],
                "cobrador": d.get("cobrador_a", ""), "comentario": "",
                "estado": "discrepancia_resuelta",
                "ciclo_correccion": ciclo,
            })
            resueltas_traz.append(_traz_row(d, "corrige", mc, ciclo, fecha_run))

        else:
            pendientes.append(d)

    return confirmados_resueltos, resueltas_traz, pendientes


def _traz_row(d: dict, resolucion: str, monto_final, ciclo: int, fecha_run: datetime) -> dict:
    return {
        "MZ":               d["mz"],
        "LT":               d["lt"],
        "MESA":             d["mesa"],
        "COBRADOR_A":       d.get("cobrador_a", ""), "MONTO_A": d.get("monto_a", "") or "",
        "COBRADOR_B":       d.get("cobrador_b", ""), "MONTO_B": d.get("monto_b", "") or "",
        "COBRADOR_C":       d.get("cobrador_c", ""), "MONTO_C": d.get("monto_c", "") or "",
        "RESOLUCION":       resolucion,
        "MONTO_FINAL":      monto_final,
        "CICLO_CORRECCION": ciclo,
        "FECHA_RESOLUCION": fecha_run,
    }


def leer_resoluciones_multi_mesa() -> dict:
    """
    Lee hoja pago_multi_mesa de discrepancias.xlsx.
    Devuelve {(llave_original, mesa): {ok, mz_correcto, lt_correcto}}.
    Solo incluye filas con OK = 'ok' o 'rechaza'.
    Filas vacías quedan implícitamente pendientes.
    """
    if not DISC_FILE.exists():
        return {}
    wb = load_workbook(DISC_FILE, read_only=True, data_only=True)
    if "pago_multi_mesa" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["pago_multi_mesa"]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(filas) < 3:
        return {}
    headers = [str(h).strip().upper() if h else "" for h in filas[1]]
    resoluciones = {}
    for fila in filas[2:]:
        if not fila or all(c is None for c in fila):
            continue
        row  = {headers[i]: fila[i] for i in range(min(len(headers), len(fila)))}
        mz   = _norm(row.get("MZ"))
        lt   = _norm(row.get("LT"))
        mesa = str(row.get("MESA") or "").strip()
        ok   = str(row.get("OK") or "").strip().lower()
        if ok == "sí":          # normalizar tilde → todo downstream usa "si"
            ok = "si"
        if not (mz and lt and mesa):
            continue
        if ok not in ("si", "rechaza"):
            continue
        resoluciones[(_llave(mz, lt), mesa)] = {
            "ok":          ok,
            "mz_correcto": _norm(row.get("MZ_CORRECTO")),
            "lt_correcto": _norm(row.get("LT_CORRECTO")),
        }
    return resoluciones


def aplicar_resoluciones_multi_mesa(multi_mesa: list, resoluciones: dict,
                                     ciclo: int, fecha_run: datetime) -> tuple:
    """
    Aplica resoluciones del supervisor a los casos pago_multi_mesa — fila por fila.
    Cada fila se resuelve de forma independiente (no por grupo):
      - ok       → cobro entra a pagos_efectivo (con corrección de predio si aplica)
                   + se registra en trazabilidad con RESOLUCION=ok
      - rechaza  → cobro NO entra a pagos_efectivo; solo se registra en trazabilidad
                   con RESOLUCION=rechaza (auditoría, permite cerrar el módulo)
      - vacío    → fila sigue pendiente; vuelve a discrepancias.xlsx
    Devuelve (confirmados, traz_rows, pendientes).
    """
    confirmados = []
    traz_rows   = []
    pendientes  = []

    for r in multi_mesa:
        res = resoluciones.get((r["llave"], r["mesa"]), {})
        ok  = res.get("ok", "")

        if ok == "si":
            mz_corr = res.get("mz_correcto", "")
            lt_corr = res.get("lt_correcto", "")
            mz_final = mz_corr or r["mz"]
            lt_final = lt_corr or r["lt"]
            confirmados.append({
                "llave":            _llave(mz_final, lt_final),
                "mz":               mz_final,
                "lt":               lt_final,
                "mesa":             r["mesa"],
                "cobrador":         r["cobrador"],
                "monto":            r["monto"],
                "fecha":            r["fecha"],
                "comentario":       "",
                "estado":           "multi_mesa_resuelta",
                "ciclo_correccion": ciclo,
            })
            traz_rows.append({
                "MZ_ORIGINAL":      r["mz"],
                "LT_ORIGINAL":      r["lt"],
                "MZ":               mz_final,
                "LT":               lt_final,
                "MESA":             r["mesa"],
                "COBRADOR":         r["cobrador"],
                "MONTO":            r["monto"],
                "FECHA_COBRO":      _to_datetime(r["fecha"]),
                "RESOLUCION":       "si",
                "CICLO_CORRECCION": ciclo,
                "FECHA_RESOLUCION": fecha_run,
            })

        elif ok == "rechaza":
            traz_rows.append({
                "MZ_ORIGINAL":      r["mz"],
                "LT_ORIGINAL":      r["lt"],
                "MZ":               r["mz"],
                "LT":               r["lt"],
                "MESA":             r["mesa"],
                "COBRADOR":         r["cobrador"],
                "MONTO":            r["monto"],
                "FECHA_COBRO":      _to_datetime(r["fecha"]),
                "RESOLUCION":       "rechaza",
                "CICLO_CORRECCION": ciclo,
                "FECHA_RESOLUCION": fecha_run,
            })

        else:
            pendientes.append(r)

    return confirmados, traz_rows, pendientes


# ── Exportar pagos_efectivo.xlsx ─────────────────────────────────────────────
# Contrato: docs/formato_pagos_efectivo.html

def exportar_pagos_efectivo(rows: list, mapa_ciclo_orig: dict = None):
    """
    mapa_ciclo_orig: (mz, lt, mesa, round(monto, 2)) → ciclo_original.
    Si una fila está en el mapa, su CICLO_CORRECCION se sobreescribe al ciclo original
    (preserva el ciclo en que el cobro entró por primera vez, no el ciclo del run actual).
    """
    mapa_ciclo_orig = mapa_ciclo_orig or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "pagos_efectivo"
    ws.freeze_panes = "A3"

    secciones = [
        ("¿Dónde vive?",   2, "id"),
        ("¿Cuánto pagó?",  2, "cob"),
        ("¿De qué mesa?",  1, "mesa"),
        ("¿Es confiable?", 1, "est"),
        ("¿Quién cobró?",  1, "who"),
        ("¿Alguna nota?",  1, "com"),
        ("Trazabilidad",   1, "meta"),
    ]
    col = 1
    for texto, span, sec in secciones:
        bg, txt, _ = PE[sec]
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1,   end_column=col + span - 1)
        _hdr(ws.cell(row=1, column=col), bg, txt, texto)
        col += span
    ws.row_dimensions[1].height = 18

    cols = [
        ("MZ",               "id",   8,  "center"),
        ("LT",               "id",   8,  "center"),
        ("MONTO",            "cob",  12, "right"),
        ("FECHA",            "cob",  14, "center"),
        ("MESA",             "mesa", 10, "center"),
        ("ESTADO",           "est",  20, "center"),
        ("COBRADOR",         "who",  22, "left"),
        ("COMENTARIO",       "com",  30, "left"),
        ("CICLO_CORRECCION", "meta", 10, "center"),
    ]
    for ci, (nombre, sec, ancho, _) in enumerate(cols, start=1):
        bg, txt, _ = PE[sec]
        _hdr(ws.cell(row=2, column=ci), bg, txt, nombre)
        _ancho(ws, ci, ancho)
    ws.row_dimensions[2].height = 22

    for ri, r in enumerate(rows, start=3):
        # Preservar el ciclo original si la fila ya existía en trazabilidad
        key_orig = (
            _norm(r.get("mz", "")),
            _norm(r.get("lt", "")),
            str(r.get("mesa", "")).strip(),
            round(_monto(r.get("monto", 0)), 2),
        )
        ciclo_fila = mapa_ciclo_orig.get(key_orig, r.get("ciclo_correccion", ""))
        valores = [
            r.get("mz",         ""),
            r.get("lt",         ""),
            r.get("monto",      0),
            _to_datetime(r.get("fecha", "")),
            r.get("mesa",       ""),
            r.get("estado",     ""),
            r.get("cobrador",   ""),
            r.get("comentario", ""),
            ciclo_fila,
        ]
        formatos = [None, None, '"S/ "#,##0.00', "DD/MM/YYYY",
                    None, None, None, None, None]
        aligns   = ["center", "center", "right", "center",
                    "center", "center", "left",  "left", "center"]
        for ci, (nombre, sec, _, _align) in enumerate(cols, start=1):
            _, txt, bg_data = PE[sec]
            val = valores[ci - 1]
            _dat(ws.cell(row=ri, column=ci), val, bg_data, txt,
                 align=aligns[ci - 1], fmt=formatos[ci - 1])
        ws.row_dimensions[ri].height = 18

    OUTPUTS_DIR.mkdir(exist_ok=True)
    wb.save(OUTPUT_FILE)


# ── Exportar discrepancias.xlsx ──────────────────────────────────────────────
# Contrato: docs/formato_discrepancias.html
# 2 hojas:
#   discrepancias    — mismo predio + misma mesa + distinto monto
#   pago_multi_mesa  — mismo predio en mesas distintas (1 fila por aparición)

# Mismo patrón que motor_matching: protege el trabajo manual del supervisor
# cuando el módulo regenera discrepancias.xlsx. Sin esto, un valor no reconocido
# (typo, variante no aceptada) hace que la fila se reescriba en blanco y el
# supervisor pierde lo que escribió.

def _leer_disc_preservados(disc_file: Path) -> dict:
    """
    Lee discrepancias.xlsx existente y captura los campos que llenó el supervisor:
      - Hoja 'discrepancias':    (MZ,LT,MESA) → {resolucion, monto_correcto}
      - Hoja 'pago_multi_mesa':  (MZ,LT,MESA) → {ok, mz_correcto, lt_correcto}
    Solo preserva filas donde el supervisor escribió al menos un valor.
    """
    preservados = {"discrepancias": {}, "pago_multi_mesa": {}}
    if not disc_file.exists():
        return preservados
    try:
        wb = load_workbook(disc_file, read_only=True, data_only=True)
    except Exception:
        return preservados

    try:
        if "discrepancias" in wb.sheetnames:
            ws = wb["discrepancias"]
            filas = list(ws.iter_rows(values_only=True))
            if len(filas) >= 3:
                headers = [str(h).strip().upper() if h else "" for h in filas[1]]
                for fila in filas[2:]:
                    if not fila or all(c is None for c in fila):
                        continue
                    row = {headers[i]: fila[i] for i in range(min(len(headers), len(fila)))}
                    mz   = _norm(row.get("MZ"))
                    lt   = _norm(row.get("LT"))
                    mesa = str(row.get("MESA") or "").strip()
                    if not (mz and lt and mesa):
                        continue
                    res = str(row.get("RESOLUCION") or "").strip()
                    mc  = row.get("MONTO_CORRECTO")
                    mc_filled = mc not in (None, "", 0, 0.0)
                    if not res and not mc_filled:
                        continue
                    preservados["discrepancias"][(mz, lt, mesa)] = {
                        "resolucion":     res,
                        "monto_correcto": mc if mc_filled else "",
                    }

        if "pago_multi_mesa" in wb.sheetnames:
            ws = wb["pago_multi_mesa"]
            filas = list(ws.iter_rows(values_only=True))
            if len(filas) >= 3:
                headers = [str(h).strip().upper() if h else "" for h in filas[1]]
                for fila in filas[2:]:
                    if not fila or all(c is None for c in fila):
                        continue
                    row = {headers[i]: fila[i] for i in range(min(len(headers), len(fila)))}
                    mz   = _norm(row.get("MZ"))
                    lt   = _norm(row.get("LT"))
                    mesa = str(row.get("MESA") or "").strip()
                    if not (mz and lt and mesa):
                        continue
                    ok      = str(row.get("OK") or "").strip()
                    mz_corr = str(row.get("MZ_CORRECTO") or "").strip()
                    lt_corr = str(row.get("LT_CORRECTO") or "").strip()
                    if not (ok or mz_corr or lt_corr):
                        continue
                    preservados["pago_multi_mesa"][(mz, lt, mesa)] = {
                        "ok":          ok,
                        "mz_correcto": mz_corr,
                        "lt_correcto": lt_corr,
                    }
    finally:
        wb.close()

    return preservados


def _backup_discrepancias(disc_file: Path):
    """Copia discrepancias.xlsx a backup/ con timestamp antes de regenerarlo."""
    if not disc_file.exists():
        return
    backup_dir = BASE_DIR / "backup"
    backup_dir.mkdir(exist_ok=True)
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = backup_dir / f"discrepancias_{ts}.xlsx"
    try:
        shutil.copy2(disc_file, dst)
    except Exception:
        pass


def exportar_discrepancias(disc_rows: list, multi_mesa_rows: list) -> int:
    """
    Devuelve el número de filas preservadas (trabajo manual del ciclo anterior).
    """
    preservados = _leer_disc_preservados(DISC_FILE)
    _backup_discrepancias(DISC_FILE)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "discrepancias"
    ws1.freeze_panes = "A3"
    _hoja_discrepancias(ws1, disc_rows, preservados.get("discrepancias", {}))

    ws2 = wb.create_sheet("pago_multi_mesa")
    ws2.freeze_panes = "A3"
    _hoja_disc_pago_multi_mesa(ws2, multi_mesa_rows, preservados.get("pago_multi_mesa", {}))

    OUTPUTS_DIR.mkdir(exist_ok=True)
    wb.save(DISC_FILE)

    return sum(len(d) for d in preservados.values())


def _hoja_discrepancias(ws, rows: list, preservados: dict = None):
    preservados = preservados or {}
    secciones = [
        ("¿Cuál es el predio?",           2, "id"),
        ("¿De qué mesa?",                 1, "mesa"),
        ("¿Qué dijo cada cobrador?",      6, "disc"),
        ("Resolución — supervisor llena", 2, "res"),
    ]
    col = 1
    for texto, span, sec in secciones:
        bg, txt, _ = DC[sec]
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1,   end_column=col + span - 1)
        _hdr(ws.cell(row=1, column=col), bg, txt, texto)
        col += span
    ws.row_dimensions[1].height = 18

    cols = [
        ("MZ",             "id",   8,  "center"),
        ("LT",             "id",   8,  "center"),
        ("MESA",           "mesa", 10, "center"),
        ("COBRADOR_A",     "disc", 18, "left"),
        ("MONTO_A",        "disc", 12, "right"),
        ("COBRADOR_B",     "disc", 18, "left"),
        ("MONTO_B",        "disc", 12, "right"),
        ("COBRADOR_C",     "disc", 18, "left"),
        ("MONTO_C",        "disc", 12, "right"),
        ("RESOLUCION",     "res",  16, "center"),
        ("MONTO_CORRECTO", "res",  14, "right"),
    ]
    for ci, (nombre, sec, ancho, _) in enumerate(cols, start=1):
        bg, txt, _ = DC[sec]
        _hdr(ws.cell(row=2, column=ci), bg, txt, nombre)
        _ancho(ws, ci, ancho)
    ws.row_dimensions[2].height = 22

    for ri, d in enumerate(rows, start=3):
        key  = (_norm(d.get("mz","")), _norm(d.get("lt","")), str(d.get("mesa","")).strip())
        prev = preservados.get(key, {})
        valores = [
            d.get("mz",         ""),
            d.get("lt",         ""),
            d.get("mesa",       ""),
            d.get("cobrador_a", ""),
            d.get("monto_a",    ""),
            d.get("cobrador_b", ""),
            d.get("monto_b",    ""),
            d.get("cobrador_c", ""),
            d.get("monto_c",    ""),
            prev.get("resolucion", "") or d.get("resolucion", ""),
            prev.get("monto_correcto", "") or d.get("monto_correcto", ""),
        ]
        for ci, (nombre, sec, _, align) in enumerate(cols, start=1):
            bg_h, txt_h, bg_data = DC[sec]
            val = valores[ci - 1]

            if nombre in ("MONTO_A", "MONTO_B", "MONTO_C") and val != "":
                _dat(ws.cell(row=ri, column=ci), val, bg_data, DC_MONTO_DISC,
                     align=align, bold=True, fmt='"S/ "#,##0.00')
            elif nombre == "RESOLUCION":
                bg_res = DC_RES_LLENO if val else bg_data
                _dat(ws.cell(row=ri, column=ci), val, bg_res, txt_h,
                     align=align, bold=bool(val))
            elif nombre == "MONTO_CORRECTO" and val:
                _dat(ws.cell(row=ri, column=ci), val, "ECFDF5", "065F46",
                     align=align, fmt='"S/ "#,##0.00')
            else:
                _dat(ws.cell(row=ri, column=ci), val if val != "" else "",
                     bg_data, txt_h, align=align)
        ws.row_dimensions[ri].height = 18


def _hoja_disc_pago_multi_mesa(ws, rows: list, preservados: dict = None):
    """
    Hoja pago_multi_mesa dentro de discrepancias.xlsx.
    El supervisor llena fila por fila:
      - OK = si / rechaza / vacío
      - MZ_CORRECTO + LT_CORRECTO si el cobro fue a un predio mal registrado
    Si la fila ya tenía valores del ciclo anterior, se re-inyectan desde `preservados`.
    """
    preservados = preservados or {}
    secciones = [
        ("¿Cuál es el predio?",           2, "id"),
        ("¿Qué se registró?",             4, "cobro"),
        ("Resolución — supervisor llena", 3, "res"),
    ]
    col = 1
    for texto, span, sec in secciones:
        bg, txt, _ = DC[sec]
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1,   end_column=col + span - 1)
        _hdr(ws.cell(row=1, column=col), bg, txt, texto)
        col += span
    ws.row_dimensions[1].height = 18

    cols = [
        # (nombre, sección, ancho, align, fmt)
        ("MZ",          "id",    8,  "center", None),
        ("LT",          "id",    8,  "center", None),
        ("MESA",        "cobro", 12, "center", None),
        ("COBRADOR",    "cobro", 22, "left",   None),
        ("MONTO",       "cobro", 12, "right",  '"S/ "#,##0.00'),
        ("FECHA_COBRO", "cobro", 14, "center", "DD/MM/YYYY"),
        ("MZ_CORRECTO", "res",   12, "center", None),
        ("LT_CORRECTO", "res",   12, "center", None),
        ("OK",          "res",   10, "center", None),
    ]
    for ci, (nombre, sec, ancho, _, _) in enumerate(cols, start=1):
        bg, txt, _ = DC[sec]
        _hdr(ws.cell(row=2, column=ci), bg, txt, nombre)
        _ancho(ws, ci, ancho)
    ws.row_dimensions[2].height = 22

    for ri, r in enumerate(rows, start=3):
        key  = (_norm(r.get("mz","")), _norm(r.get("lt","")), str(r.get("mesa","")).strip())
        prev = preservados.get(key, {})
        for ci, (nombre, sec, _, align, fmt) in enumerate(cols, start=1):
            bg_h, txt_h, bg_data = DC[sec]

            if   nombre == "MZ":          val = r.get("mz", "")
            elif nombre == "LT":          val = r.get("lt", "")
            elif nombre == "MESA":        val = r.get("mesa", "")
            elif nombre == "COBRADOR":    val = r.get("cobrador", "")
            elif nombre == "MONTO":       val = r.get("monto", "")
            elif nombre == "FECHA_COBRO": val = _to_datetime(r.get("fecha", ""))
            elif nombre == "MZ_CORRECTO": val = prev.get("mz_correcto", "")
            elif nombre == "LT_CORRECTO": val = prev.get("lt_correcto", "")
            elif nombre == "OK":          val = prev.get("ok", "")
            else:                         val = ""

            if nombre == "OK":
                v_low = str(val).strip().lower()
                if v_low == "si":
                    bg_use = TR_OK_BG
                elif v_low == "rechaza":
                    bg_use = TR_RECHAZA_BG
                else:
                    bg_use = bg_data
                _dat(ws.cell(row=ri, column=ci), val, bg_use, txt_h,
                     align=align, bold=bool(val))
            else:
                _dat(ws.cell(row=ri, column=ci), val, bg_data, txt_h,
                     align=align, fmt=fmt)
        ws.row_dimensions[ri].height = 18


# ── Exportar trazabilidad ────────────────────────────────────────────────────
# Contrato: docs/formato_trazabilidad.html
# 3 hojas: discrepancias_resueltas · pago_multi_mesa_resueltas · solo_un_cobrador
# Las hojas *_resueltas acumulan entre ciclos.
# solo_un_cobrador se escribe en Ciclo 1 y se preserva en Ciclos 2+.

def exportar_trazabilidad(traz_file: Path, solo_cob: list, multi_resueltas: list,
                           resueltas_nuevas: list, ciclo: int):
    prev_resueltas  = _leer_hoja_traz(traz_file, "discrepancias_resueltas")
    todas_resueltas = prev_resueltas + resueltas_nuevas

    prev_multi_res  = _leer_hoja_traz(traz_file, "pago_multi_mesa_resueltas")
    todas_multi_res = prev_multi_res + multi_resueltas

    if ciclo == 1:
        final_solo = solo_cob
    else:
        final_solo = _leer_hoja_traz(traz_file, "solo_un_cobrador") or solo_cob

    TRAZAB_DIR.mkdir(exist_ok=True)
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "discrepancias_resueltas"
    _hoja_disc_resueltas(ws1, todas_resueltas)

    ws2 = wb.create_sheet("pago_multi_mesa_resueltas")
    _hoja_pago_multi_resueltas(ws2, todas_multi_res)

    ws3 = wb.create_sheet("solo_un_cobrador")
    _hoja_solo_cob(ws3, final_solo)

    wb.save(traz_file)


def _hoja_disc_resueltas(ws, rows: list):
    ws.freeze_panes = "A3"

    secs = [
        ("¿Cuál es el predio?",      2, "pred"),
        ("¿De qué mesa?",            1, "mesa"),
        ("¿Qué dijo cada cobrador?", 6, "disc"),
        ("¿Cómo se resolvió?",       2, "res"),
        ("Trazabilidad",             2, "meta"),
    ]
    col = 1
    for texto, span, sec in secs:
        bg, txt, _ = TR_DR[sec]
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1,   end_column=col + span - 1)
        _hdr(ws.cell(row=1, column=col), bg, txt, texto)
        col += span
    ws.row_dimensions[1].height = 18

    cols = [
        ("MZ",               "pred", 8,  "center", None),
        ("LT",               "pred", 8,  "center", None),
        ("MESA",             "mesa", 12, "center", None),
        ("COBRADOR_A",       "disc", 20, "left",   None),
        ("MONTO_A",          "disc", 12, "right",  '"S/ "#,##0.00'),
        ("COBRADOR_B",       "disc", 20, "left",   None),
        ("MONTO_B",          "disc", 12, "right",  '"S/ "#,##0.00'),
        ("COBRADOR_C",       "disc", 20, "left",   None),
        ("MONTO_C",          "disc", 12, "right",  '"S/ "#,##0.00'),
        ("RESOLUCION",       "res",  16, "center", None),
        ("MONTO_FINAL",      "res",  14, "right",  '"S/ "#,##0.00'),
        ("CICLO_CORRECCION", "meta", 10, "center", None),
        ("FECHA_RESOLUCION", "meta", 18, "center", "DD/MM/YYYY"),
    ]
    for ci, (nombre, sec, ancho, _, _) in enumerate(cols, start=1):
        bg, txt, _ = TR_DR[sec]
        _hdr(ws.cell(row=2, column=ci), bg, txt, nombre)
        _ancho(ws, ci, ancho)
    ws.row_dimensions[2].height = 22

    for ri, r in enumerate(rows, start=3):
        for ci, (nombre, sec, _, align, fmt) in enumerate(cols, start=1):
            bg_h, txt_h, bg_data = TR_DR[sec]
            val = r.get(nombre, "")
            if nombre == "FECHA_RESOLUCION" and val:
                val = _to_datetime(val)
            if nombre == "RESOLUCION" and val:
                _dat(ws.cell(row=ri, column=ci), val, TR_RES_LLENA, txt_h,
                     align=align, bold=True)
            else:
                _dat(ws.cell(row=ri, column=ci), val, bg_data, txt_h,
                     align=align, fmt=fmt)
        ws.row_dimensions[ri].height = 18


def _hoja_pago_multi_resueltas(ws, rows: list):
    """
    Hoja pago_multi_mesa_resueltas en trazabilidad. Acumula entre ciclos.
    Incluye filas 'ok' (cobro aceptado, va a pagos_efectivo) y 'rechaza'
    (cobro descartado, queda como auditoría para permitir cerrar el módulo).
    MZ_ORIGINAL/LT_ORIGINAL: predio tal como lo registró el cobrador.
    MZ/LT: predio imputado — distinto cuando el supervisor corrigió.
    """
    ws.freeze_panes = "A3"

    secs = [
        ("Predio original registrado",     2, "orig"),
        ("Predio imputado (puede diferir)", 2, "pred"),
        ("¿Qué se registró?",              4, "cobro"),
        ("¿Cómo se resolvió?",             1, "res"),
        ("Trazabilidad",                   2, "meta"),
    ]
    col = 1
    for texto, span, sec in secs:
        bg, txt, _ = TR_SC[sec]
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1,   end_column=col + span - 1)
        _hdr(ws.cell(row=1, column=col), bg, txt, texto)
        col += span
    ws.row_dimensions[1].height = 18

    cols = [
        ("MZ_ORIGINAL",      "orig",  12, "center", None),
        ("LT_ORIGINAL",      "orig",  12, "center", None),
        ("MZ",               "pred",   8, "center", None),
        ("LT",               "pred",   8, "center", None),
        ("MESA",             "cobro", 12, "center", None),
        ("COBRADOR",         "cobro", 22, "left",   None),
        ("MONTO",            "cobro", 12, "right",  '"S/ "#,##0.00'),
        ("FECHA_COBRO",      "cobro", 14, "center", "DD/MM/YYYY"),
        ("RESOLUCION",       "res",   10, "center", None),
        ("CICLO_CORRECCION", "meta",  10, "center", None),
        ("FECHA_RESOLUCION", "meta",  18, "center", "DD/MM/YYYY"),
    ]
    for ci, (nombre, sec, ancho, _, _) in enumerate(cols, start=1):
        bg, txt, _ = TR_SC[sec]
        _hdr(ws.cell(row=2, column=ci), bg, txt, nombre)
        _ancho(ws, ci, ancho)
    ws.row_dimensions[2].height = 22

    for ri, r in enumerate(rows, start=3):
        for ci, (nombre, sec, _, align, fmt) in enumerate(cols, start=1):
            bg_h, txt_h, bg_data = TR_SC[sec]
            val = r.get(nombre, "")
            if nombre in ("FECHA_COBRO", "FECHA_RESOLUCION") and val:
                val = _to_datetime(val)
            if nombre == "RESOLUCION":
                v_low = str(val).strip().lower()
                if v_low == "si":
                    bg_use = TR_OK_BG
                elif v_low == "rechaza":
                    bg_use = TR_RECHAZA_BG
                else:
                    bg_use = bg_data
                _dat(ws.cell(row=ri, column=ci), val, bg_use, txt_h,
                     align=align, bold=bool(val))
            else:
                _dat(ws.cell(row=ri, column=ci), val, bg_data, txt_h,
                     align=align, fmt=fmt)
        ws.row_dimensions[ri].height = 18


def _hoja_solo_cob(ws, rows: list):
    ws.freeze_panes = "A3"

    secs = [
        ("¿Cuál es el predio?", 2, "pred"),
        ("¿De qué mesa?",       1, "mesa"),
        ("¿Qué se cobró?",      4, "cobro"),
        ("Trazabilidad",        1, "meta"),
    ]
    col = 1
    for texto, span, sec in secs:
        bg, txt, _ = TR_SC[sec]
        if span > 1:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1,   end_column=col + span - 1)
        _hdr(ws.cell(row=1, column=col), bg, txt, texto)
        col += span
    ws.row_dimensions[1].height = 18

    cols = [
        ("MZ",          "pred",  8,  "center", None),
        ("LT",          "pred",  8,  "center", None),
        ("MESA",        "mesa",  12, "center", None),
        ("COBRADOR",    "cobro", 22, "left",   None),
        ("MONTO",       "cobro", 12, "right",  '"S/ "#,##0.00'),
        ("FECHA_COBRO", "cobro", 14, "center", "DD/MM/YYYY"),
        ("COMENTARIO",       "cobro", 30, "left",   None),
        ("CICLO_CORRECCION", "meta",  10, "center", None),
    ]
    for ci, (nombre, sec, ancho, _, _) in enumerate(cols, start=1):
        bg, txt, _ = TR_SC[sec]
        _hdr(ws.cell(row=2, column=ci), bg, txt, nombre)
        _ancho(ws, ci, ancho)
    ws.row_dimensions[2].height = 22

    for ri, r in enumerate(rows, start=3):
        for ci, (nombre, sec, _, align, fmt) in enumerate(cols, start=1):
            bg_h, txt_h, bg_data = TR_SC[sec]
            val = r.get(nombre, "")
            if nombre == "FECHA_COBRO" and val:
                val = _to_datetime(val)
            _dat(ws.cell(row=ri, column=ci), val, bg_data, txt_h, align=align, fmt=fmt)
        ws.row_dimensions[ri].height = 18


# ── cleanup_temporales — para que llame 7_cierre ─────────────────────────────

def cleanup_temporales():
    """
    Borra archivos temporales del mes tras cierre global (7_cierre post git-tag).
    Conserva: trazabilidad/, inputs/mesa_N.xlsx (sagrados).
    """
    for f in [DISC_FILE, OUTPUT_FILE]:
        if f.exists():
            f.unlink()
    print("  efectivo: temporales del mes limpiados")


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    OUTPUTS_DIR.mkdir(exist_ok=True)
    # Windows: forzar UTF-8 en stdout para que ═, acentos, etc. no rompan el log
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(OUTPUTS_DIR / "run.log", encoding="utf-8"),
        ],
        force=True,
    )
    log = logging.getLogger(__name__)

    log.info("══════════════════════════════════════")
    log.info("  4_pagos / efectivo")
    log.info("══════════════════════════════════════")

    # ── 1. Leer inputs ───────────────────────────────────────────────────────
    log.info("Cargando mesas...")
    mesas = leer_mesas()
    log.info(f"  {len(mesas)} mesa(s) con datos: {[m['nombre'] for m in mesas]}")
    for m in mesas:
        hojas_info = {h: len(r) for h, r in m["hojas"].items()}
        log.info(f"    {m['nombre']}: {hojas_info}")

    # ── 2. Determinar mes y ciclo ────────────────────────────────────────────
    todos_regs = [r for m in mesas for regs in m["hojas"].values() for r in regs]
    mes        = _mes_de_datos(todos_regs)
    traz_file  = TRAZAB_DIR / f"trazabilidad_{mes}.xlsx"
    ciclo      = detectar_ciclo(traz_file)
    fecha_run  = datetime.now()
    log.info(f"Mes: {mes}  |  Ciclo: {ciclo}")

    # ── 3. Cross-check por mesa ──────────────────────────────────────────────
    log.info("Cross-check por mesa...")
    todos_conf = []
    todas_disc = []
    auto_res   = []   # mayoría 2/3 automática → trazabilidad
    solo_cob   = []   # 1 cobrador → trazabilidad

    for mesa in mesas:
        conf, disc, res, sc = cross_check_mesa(mesa, ciclo, fecha_run)
        todos_conf.extend(conf)
        todas_disc.extend(disc)
        auto_res.extend(res)
        solo_cob.extend(sc)
        log.info(f"  {mesa['nombre']}: {len(conf)} confirmados, "
                 f"{len(disc)} discrepancias, {len(res)} mayoría_2/3, "
                 f"{len(sc)} solo_un_cobrador")

    # ── 3b. Re-incorporar resoluciones multi_mesa de ciclos previos ──────────
    # Bug: si en Ciclo N se resolvió un multi_mesa, en Ciclo N+1 vuelve a detectarse
    # desde los inputs y la resolución se "olvida". Fix: marcar las filas ya resueltas
    # leyendo trazabilidad/pago_multi_mesa_resueltas.
    prev_multi_map = _leer_multi_mesa_resueltas_map(traz_file)
    if prev_multi_map:
        n_marcadas = _aplicar_multi_mesa_previas(todos_conf, prev_multi_map)
        if n_marcadas:
            log.info(f"  Re-incorporadas {n_marcadas} resolución(es) multi_mesa de ciclos previos")

    # ── 4. Detectar pago_multi_mesa ──────────────────────────────────────────
    limpios, multi_mesa = detectar_multi_mesa(todos_conf)
    if multi_mesa:
        n_grupos = len({r["llave"] for r in multi_mesa})
        log.warning(f"  {len(multi_mesa)} fila(s) multi_mesa en {n_grupos} grupo(s) — excluidas de pagos_efectivo")

    # ── 5. Aplicar resoluciones del supervisor (Ciclo 2+) ────────────────────
    resueltas_disc  = []
    resueltas_multi = []

    # 5a. Discrepancias (mismo predio + misma mesa + distinto monto)
    resoluciones = leer_resoluciones()
    if resoluciones and todas_disc:
        log.info(f"Aplicando {len(resoluciones)} resolución(es) de discrepancias...")
        conf_res, res_traz, pendientes = aplicar_resoluciones(
            todas_disc, resoluciones, ciclo, fecha_run
        )
        limpios.extend(conf_res)
        resueltas_disc = res_traz
        todas_disc     = pendientes
        log.info(f"  resueltas: {len(conf_res)}  aún pendientes: {len(pendientes)}")

    # 5b. Multi-mesa (mismo predio en mesas distintas) — resolución fila por fila
    resoluciones_multi = leer_resoluciones_multi_mesa()
    if resoluciones_multi and multi_mesa:
        n_ok      = sum(1 for v in resoluciones_multi.values() if v["ok"] == "si")
        n_rechaza = sum(1 for v in resoluciones_multi.values() if v["ok"] == "rechaza")
        log.info(f"Aplicando resoluciones de pago_multi_mesa: {n_ok} ok, {n_rechaza} rechaza...")
        conf_multi, traz_multi, multi_pend = aplicar_resoluciones_multi_mesa(
            multi_mesa, resoluciones_multi, ciclo, fecha_run
        )
        limpios.extend(conf_multi)
        resueltas_multi = traz_multi
        multi_mesa      = multi_pend
        log.info(f"  multi_mesa: {len(conf_multi)} -> pagos_efectivo, "
                 f"{len(traz_multi) - len(conf_multi)} rechazadas (solo trazabilidad), "
                 f"{len(multi_pend)} filas aún pendientes")

    # ── 6. Exportar pagos_efectivo ───────────────────────────────────────────
    # Mapa de ciclos originales: preserva el CICLO_CORRECCION en que cada cobro fue
    # identificado por primera vez (en lugar de grabar todo con el ciclo del run actual).
    mapa_ciclo_orig = _construir_mapa_ciclo_original(traz_file)
    if limpios:
        exportar_pagos_efectivo(limpios, mapa_ciclo_orig)
        log.info(f"  -> pagos_efectivo.xlsx ({len(limpios)} cobros)")
    else:
        log.warning("  Sin cobros confirmados aún — pagos_efectivo.xlsx no generado")

    # ── 7. Exportar/eliminar discrepancias ───────────────────────────────────
    hay_disc       = bool(todas_disc)
    hay_multi      = bool(multi_mesa)
    hay_pendiente  = hay_disc or hay_multi

    if hay_pendiente:
        n_preservados = exportar_discrepancias(todas_disc, multi_mesa)
        if n_preservados:
            log.info(f"  ⤴ Preservadas {n_preservados} fila(s) con trabajo manual del ciclo anterior")
        if hay_disc:
            log.warning(f"  -> discrepancias.xlsx hoja 'discrepancias' ({len(todas_disc)} pendientes)")
        if hay_multi:
            n_grupos = len({r["llave"] for r in multi_mesa})
            log.warning(f"  -> discrepancias.xlsx hoja 'pago_multi_mesa' ({n_grupos} grupo(s), {len(multi_mesa)} filas)")
        log.warning("     Abre discrepancias.xlsx — discrepancias: llena RESOLUCION; pago_multi_mesa: llena OK (si/rechaza). Cada fila rechazada debe quedar marcada explícitamente para cerrar el módulo.")
    elif DISC_FILE.exists():
        DISC_FILE.unlink()
        log.info("  discrepancias.xlsx eliminado — todo resuelto")

    # ── 8. Exportar trazabilidad (acumula entre ciclos) ──────────────────────
    resueltas_nuevas = auto_res + resueltas_disc
    exportar_trazabilidad(traz_file, solo_cob, resueltas_multi, resueltas_nuevas, ciclo)
    log.info(f"  -> trazabilidad/trazabilidad_{mes}.xlsx")
    log.info(f"     discrepancias_resueltas:    +{len(resueltas_nuevas)} (ciclo={ciclo})")
    log.info(f"     pago_multi_mesa_resueltas:  +{len(resueltas_multi)} (ciclo={ciclo})")
    if ciclo == 1:
        log.info(f"     solo_un_cobrador:           {len(solo_cob)} filas")
    else:
        log.info(f"     solo_un_cobrador:           preservadas del Ciclo 1")

    # ── 9. Resumen final ─────────────────────────────────────────────────────
    log.info("══════════════════════════════════════")
    if not hay_pendiente:
        log.info("  pagos_efectivo.xlsx listo -> entregar a 5_cobranza")
    else:
        if hay_disc:
            log.warning(f"  Quedan {len(todas_disc)} discrepancia(s) sin resolver")
        if hay_multi:
            n_grupos = len({r["llave"] for r in multi_mesa})
            log.warning(f"  Quedan {n_grupos} grupo(s) pago_multi_mesa sin resolver")
    log.info("══════════════════════════════════════")


if __name__ == "__main__":
    main()
