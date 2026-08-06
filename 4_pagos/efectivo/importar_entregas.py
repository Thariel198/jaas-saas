# 4_pagos/efectivo/importar_entregas.py — puente Excel físico → ledger append-only
#
# La tesorera llena inputs/entregas_hoja.xlsx a mano (como una mesa). Este script
# lee cada fila y la vuelca al ledger append-only entregas.xlsx vía entregas_repo.
# El Excel es la superficie de captura; el ledger sigue siendo append-only y writer
# único. En SaaS, este import lo reemplaza un endpoint POST que llama la misma
# función registrar_entrega().
#
# Contrato visual (ledger):  docs/formato_entregas.html
# Decisión de diseño:        docs/decisiones/arqueo_efectivo.md
#
# Uso:
#   py importar_entregas.py          # 1a vez: crea el template. Luego: importa lo pendiente.
#
# Reglas de la hoja:
#   - Fila 3 = ejemplo guía (se ignora). Datos desde la fila 4.
#   - Fila normal (MOTIVO vacío): EFECTIVO/YAPE = lo que recibió. Se appendea tal cual.
#   - Fila de corrección (MOTIVO lleno): EFECTIVO/YAPE = el TOTAL CORRECTO de ese
#     (FECHA, COBRADOR). El script calcula el delta contra el ledger y appendea el delta.
#   - Columna IMPORTADO la llena el script (marca fecha + audit_ref). NO tocarla.
#   - Re-importar es seguro: idempotente por (source, audit_ref) — una fila no entra dos veces.

import argparse
import logging
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import entregas_repo as repo

log = logging.getLogger(__name__)

BASE_DIR = Path(__file__).parent
OUTPUTS_DIR = BASE_DIR / "outputs"
HOJA_PATH = BASE_DIR / "inputs" / "entregas_hoja.xlsx"
SHEET = "Entregas"

# (nombre, seccion_color, ancho, align) — paleta compartida con el ledger
_COLS = [
    ("FECHA",     repo._AZ, 14, "center"),
    ("COBRADOR",  repo._AZ, 22, "left"),
    ("EFECTIVO",  repo._VE, 12, "right"),
    ("YAPE",      repo._VE, 12, "right"),
    ("MOTIVO",    repo._RO, 30, "left"),
    ("IMPORTADO", repo._GR, 34, "left"),
]
_SECCIONES = [
    ("¿Quién y cuándo?",   "FECHA",     "COBRADOR",  repo._AZ),
    ("¿Cuánto recibió?",   "EFECTIVO",  "YAPE",      repo._VE),
    ("¿Corrección?",       "MOTIVO",    "MOTIVO",    repo._RO),
    ("Sistema — no tocar", "IMPORTADO", "IMPORTADO", repo._GR),
]
_EJEMPLO = ["04/07/2026", "Maximo Encarnacion", 2935, 200, "", ""]

# ── Crear template físico ─────────────────────────────────────────────────────

def _hdr(cell, bg, fg, texto):
    cell.value = texto
    cell.fill = repo._fill(bg)
    cell.font = Font(color=repo._argb(fg), bold=True, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def crear_template() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    col_idx = {c[0]: i + 1 for i, c in enumerate(_COLS)}

    for label, ini, fin, sec in _SECCIONES:
        c1, c2 = col_idx[ini], col_idx[fin]
        if c1 != c2:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        _hdr(ws.cell(row=1, column=c1), sec[0], sec[1], label)
    ws.row_dimensions[1].height = 18

    for i, (nombre, sec, ancho, _a) in enumerate(_COLS, start=1):
        _hdr(ws.cell(row=2, column=i), sec[0], sec[1], nombre)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[2].height = 22

    # Fila 3 = ejemplo guía (gris itálico, datos reales) — el operario la deja o la borra
    for i, valor in enumerate(_EJEMPLO, start=1):
        c = ws.cell(row=3, column=i, value=valor)
        c.font = Font(color="FF9CA3AF", italic=True, size=10)
        c.alignment = Alignment(horizontal=_COLS[i - 1][3], vertical="center")
        if _COLS[i - 1][0] in ("EFECTIVO", "YAPE"):
            c.number_format = "#,##0.00"

    ws.freeze_panes = "A4"
    HOJA_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(HOJA_PATH)

# ── Importar ──────────────────────────────────────────────────────────────────

def _slug(cobrador: str) -> str:
    s = "".join(ch for ch in cobrador.lower() if ch.isalnum())
    return s or "cob"


def importar() -> dict:
    wb = load_workbook(HOJA_PATH)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    headers = [str(c.value).strip().upper() if c.value else "" for c in ws[2]]
    idx = {h: i for i, h in enumerate(headers)}

    def celda(row, col):
        i = idx.get(col)
        return ws.cell(row=row, column=i + 1) if i is not None else None

    runts = datetime.now().strftime("%Y-%m-%d %H:%M")
    n_ok = n_skip = n_err = 0
    cambios = False

    for row in range(4, ws.max_row + 1):   # fila 3 = ejemplo guía, se ignora
        fecha_v = (celda(row, "FECHA").value if celda(row, "FECHA") else None)
        cob_v = (celda(row, "COBRADOR").value if celda(row, "COBRADOR") else None)
        efec = repo._monto(celda(row, "EFECTIVO").value if celda(row, "EFECTIVO") else None)
        yape = repo._monto(celda(row, "YAPE").value if celda(row, "YAPE") else None)
        motivo = str((celda(row, "MOTIVO").value if celda(row, "MOTIVO") else "") or "").strip()
        imp_c = celda(row, "IMPORTADO")

        if not cob_v and not fecha_v and efec == 0 and yape == 0:
            continue
        if imp_c and str(imp_c.value or "").strip():   # ya importada
            n_skip += 1
            continue

        try:
            fdt = repo._parse_fecha(fecha_v)
        except ValueError:
            log.warning(f"fila {row}: FECHA inválida ({fecha_v!r}) — no se importa")
            n_err += 1
            continue
        cob = repo._norm_cobrador(cob_v)
        if not cob:
            log.warning(f"fila {row}: COBRADOR vacío — no se importa")
            n_err += 1
            continue

        if motivo:   # corrección: EFECTIVO/YAPE son el total correcto → appendear el delta
            actual = repo.suma_entregas().get(
                (repo._fecha_key(fdt), cob.upper()), {"efectivo": 0.0, "yape": 0.0})
            efec_w = round(efec - actual["efectivo"], 2)
            yape_w = round(yape - actual["yape"], 2)
        else:
            efec_w, yape_w = efec, yape

        audit_ref = f"HOJA-{fdt:%Y%m%d}-{_slug(cob)}-r{row}"
        res = repo.registrar_entrega(fdt, cob, efec_w, yape_w,
                                     source="entregas_hoja", audit_ref=audit_ref, motivo=motivo)
        if imp_c:
            imp_c.value = f"{runts} [{audit_ref}]" + (" delta" if motivo else "")
            cambios = True
        n_skip += 1 if res["skipped"] else 0
        n_ok += 0 if res["skipped"] else 1

    if cambios:
        try:
            wb.save(HOJA_PATH)
        except PermissionError:
            log.error("entregas_hoja.xlsx abierto en Excel — cerralo. El ledger YA se escribió; "
                      "las marcas IMPORTADO no se guardaron, pero re-importar es seguro (idempotente).")
    return {"ok": n_ok, "skip": n_skip, "err": n_err}

# ── Orquestación ──────────────────────────────────────────────────────────────

def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    argparse.ArgumentParser(description="Importa entregas_hoja.xlsx al ledger append-only").parse_args()

    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(OUTPUTS_DIR / "importar_entregas_run.log", encoding="utf-8"),
        ],
        force=True,
    )

    if not HOJA_PATH.exists():
        crear_template()
        print(f"Template creado: {HOJA_PATH}")
        print("Llenalo (FECHA · COBRADOR · EFECTIVO · YAPE) desde la fila 4 y volvé a correr para importar.")
        return

    r = importar()
    log.info(f"import: {r['ok']} nuevas · {r['skip']} ya estaban · {r['err']} con error")
    print(f"Importadas {r['ok']} · saltadas {r['skip']} · errores {r['err']}")
    if r["ok"]:
        print("Ahora: py arqueo.py --mes AAAA-MM")


if __name__ == "__main__":
    main()
