# 4_pagos/efectivo/arqueo.py — cuadre de caja: papel (mesas) vs tesorera (entregas)
#
# Cruza, por (FECHA, COBRADOR), lo que el cobrador anotó en su mesa contra lo que
# la tesorera declaró haber recibido. Vista 100% regenerable — sin columnas humanas.
#
# Contrato visual:    docs/formato_arqueo.html
# Decisión de diseño: docs/decisiones/arqueo_efectivo.md
#
# Uso:
#   python arqueo.py               # mes actual
#   python arqueo.py --mes 2026-07

import argparse
import logging
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import entregas_repo as repo

log = logging.getLogger(__name__)

BASE_DIR = Path(__file__).parent
INPUTS_DIR = BASE_DIR / "inputs"
OUTPUTS_DIR = BASE_DIR / "outputs"
HOJAS = ["registro_1", "registro_2", "registro_3"]

TOL = 0.01  # tolerancia de centavos para declarar CUADRA

# (header_bg, header_fg, data_bg) — contrato: docs/formato_arqueo.html
_AZ = ("EBF5FB", "1A5276", "F4FAFF")   # ¿quién y cuándo?
_VE = ("E9F7EF", "1E5C3A", "F4FBF7")   # efectivo
_YA = ("E0F2FE", "075985", "F0F9FF")   # yape
_MO = ("F3E8FF", "5B21B6", "FAF5FF")   # ¿cuadra?

_DIF_CERO  = "1E5C3A"   # verde
_DIF_FALTA = "991B1B"   # rojo (negativo)
_DIF_SOBRA = "92400E"   # ámbar (positivo)

_COLS = [
    ("FECHA",             _AZ, 14, "center"),
    ("COBRADOR",          _AZ, 22, "left"),
    ("EFECTIVO_PAPEL",    _VE, 16, "right"),
    ("EFECTIVO_RECIBIDO", _VE, 18, "right"),
    ("DIF_EFECTIVO",      _VE, 14, "right"),
    ("YAPE_PAPEL",        _YA, 14, "right"),
    ("YAPE_RECIBIDO",     _YA, 16, "right"),
    ("DIF_YAPE",          _YA, 12, "right"),
    ("ESTADO",            _MO, 14, "center"),
]
_SECCIONES = [
    ("¿Quién y cuándo?", "FECHA",          "COBRADOR",          _AZ),
    ("Efectivo",         "EFECTIVO_PAPEL", "DIF_EFECTIVO",      _VE),
    ("Yape",             "YAPE_PAPEL",     "DIF_YAPE",          _YA),
    ("¿Cuadra?",         "ESTADO",         "ESTADO",            _MO),
]

# ── Lectura de mesas (crudo, sin dedupe — a diferencia de main.py) ────────────

def leer_mesas(mes: str | None = None) -> dict:
    """{(fecha_key, cobrador_upper): {'efectivo','yape','cobrador'}} sumando
    MONTO_EFECTIVO y MONTO_YAPE de todas las mesas. Es el 'PAPEL' del arqueo.

    A diferencia de main.py.leer_hoja NO descarta filas con MONTO_EFECTIVO<=0:
    un pago puro-yape (efectivo 0, yape > 0) también entra al cuadre."""
    out: dict = {}
    for n in range(1, 8):
        path = INPUTS_DIR / f"mesa_{n}.xlsx"
        if not path.exists():
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        sin_fecha = 0
        for hoja in HOJAS:
            if hoja not in wb.sheetnames:
                continue
            filas = list(wb[hoja].values)
            if len(filas) < 4:
                continue
            headers = [str(h).strip().upper() if h else "" for h in filas[1]]
            idx = {h: i for i, h in enumerate(headers)}

            def g(fila, col):
                i = idx.get(col)
                return fila[i] if i is not None and i < len(fila) else None

            for fila in filas[3:]:   # fila 3 = ejemplo guía
                if not fila or all(c is None for c in fila):
                    continue
                efectivo = repo._monto(g(fila, "MONTO_EFECTIVO"))
                yape = repo._monto(g(fila, "MONTO_YAPE"))
                if efectivo <= 0 and yape <= 0:
                    continue
                cobrador = repo._norm_cobrador(g(fila, "COBRADOR"))
                fecha_key = repo._fecha_key(g(fila, "FECHA"))
                if not fecha_key:
                    # Fila con plata pero sin FECHA: no entra al cuadre. Se cuenta
                    # y se avisa — una tesorería no puede cuadrar contra plata que
                    # el papel no fecha, y descartarla en silencio subcontaría el
                    # PAPEL y haría mentir al arqueo.
                    sin_fecha += 1
                    continue
                if not cobrador:
                    continue
                if mes and repo._mes_de(fecha_key) != mes:
                    continue
                key = (fecha_key, cobrador.upper())
                d = out.setdefault(key, {"efectivo": 0.0, "yape": 0.0, "cobrador": cobrador})
                d["efectivo"] = round(d["efectivo"] + efectivo, 2)
                d["yape"] = round(d["yape"] + yape, 2)
        wb.close()
        if sin_fecha:
            log.warning(f"mesa_{n}: {sin_fecha} fila(s) con plata pero sin FECHA "
                        f"— NO entran al cuadre. Completar FECHA en la mesa.")
    return out

# ── Cruce ─────────────────────────────────────────────────────────────────────

def cuadrar(papel: dict, recibido: dict) -> list:
    """Combina las claves de ambas fuentes y arma una fila por (FECHA, COBRADOR)."""
    filas = []
    for key in sorted(papel.keys() | recibido.keys(),
                      key=lambda k: (datetime.strptime(k[0], "%d/%m/%Y"), k[1])):
        fecha_key, _cob_up = key
        p = papel.get(key)
        r = recibido.get(key)
        cobrador = (p or r)["cobrador"]

        if p and r:
            dif_e = round(r["efectivo"] - p["efectivo"], 2)
            dif_y = round(r["yape"] - p["yape"], 2)
            estado = "CUADRA" if abs(dif_e) < TOL and abs(dif_y) < TOL else "DESCUADRE"
            fila = [fecha_key, cobrador, p["efectivo"], r["efectivo"], dif_e,
                    p["yape"], r["yape"], dif_y, estado]
        elif p and not r:
            fila = [fecha_key, cobrador, p["efectivo"], "—", "—",
                    p["yape"], "—", "—", "SIN_ENTREGA"]
        else:  # r and not p
            fila = [fecha_key, cobrador, "—", r["efectivo"], "—",
                    "—", r["yape"], "—", "SIN_MESA"]
        filas.append(fila)
    return filas

# ── Escritura del output ──────────────────────────────────────────────────────

def _argb(hex6):
    return "FF" + hex6.lstrip("#")


def _fill(hex6):
    return PatternFill("solid", fgColor=_argb(hex6))


def _hdr(cell, bg, fg, texto):
    cell.value = texto
    cell.fill = _fill(bg)
    cell.font = Font(color=_argb(fg), bold=True, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _color_dif(valor) -> str:
    if valor == "—":
        return _DIF_CERO
    if valor < 0:
        return _DIF_FALTA
    if valor > 0:
        return _DIF_SOBRA
    return _DIF_CERO


def escribir(filas: list, ruta: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Arqueo"

    col_idx = {c[0]: i + 1 for i, c in enumerate(_COLS)}
    for label, ini, fin, sec in _SECCIONES:
        c1, c2 = col_idx[ini], col_idx[fin]
        if c1 != c2:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        _hdr(ws.cell(row=1, column=c1), sec[0], sec[1], label)
    ws.row_dimensions[1].height = 18
    for i, (nombre, sec, ancho, _a) in enumerate(_COLS, start=1):
        _hdr(ws.cell(row=2, column=i), sec[0], sec[1], nombre)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    dif_cols = {"DIF_EFECTIVO", "DIF_YAPE"}
    num_cols = {"EFECTIVO_PAPEL", "EFECTIVO_RECIBIDO", "DIF_EFECTIVO",
                "YAPE_PAPEL", "YAPE_RECIBIDO", "DIF_YAPE"}
    for r_off, fila in enumerate(filas):
        row = r_off + 3
        for i, (nombre, sec, _ancho, align) in enumerate(_COLS, start=1):
            valor = fila[i - 1]
            cell = ws.cell(row=row, column=i)
            if nombre == "FECHA":
                cell.value = datetime.strptime(valor, "%d/%m/%Y")
                cell.number_format = "DD/MM/YYYY"
            else:
                cell.value = valor
                if nombre in num_cols and valor != "—":
                    cell.number_format = "#,##0.00"
            cell.fill = _fill(sec[2])
            cell.alignment = Alignment(horizontal=align, vertical="center")
            if nombre in dif_cols:
                cell.font = Font(color=_argb(_color_dif(valor)), bold=True, size=10)
            elif nombre == "ESTADO":
                cell.font = Font(color=_argb(sec[1]), bold=True, size=10)
            else:
                cell.font = Font(color=_argb(sec[1]), size=10)

    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ruta)

# ── Orquestación ──────────────────────────────────────────────────────────────

def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    ap = argparse.ArgumentParser(description="Cuadre de caja efectivo/yape por día y cobrador")
    hoy = date.today()
    ap.add_argument("--mes", default=f"{hoy.year}-{hoy.month:02d}",
                    help="AAAA-MM a cuadrar (default: mes actual)")
    args = ap.parse_args()
    mes = args.mes

    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(OUTPUTS_DIR / "arqueo_run.log", encoding="utf-8"),
        ],
        force=True,
    )

    log.info(f"Arqueo — mes {mes}")
    papel = leer_mesas(mes)
    recibido = repo.suma_entregas(mes)
    if not papel and not recibido:
        log.warning(f"Sin datos para {mes}: ni mesas ni entregas. No se genera arqueo.")
        return

    filas = cuadrar(papel, recibido)
    ruta = OUTPUTS_DIR / f"arqueo_{mes}.xlsx"
    escribir(filas, ruta)

    conteo: dict = {}
    for f in filas:
        conteo[f[-1]] = conteo.get(f[-1], 0) + 1
    resumen = " · ".join(f"{k}={v}" for k, v in sorted(conteo.items()))
    log.info(f"arqueo_{mes}.xlsx — {len(filas)} filas · {resumen}")
    log.info(f"-> {ruta}")


if __name__ == "__main__":
    main()
