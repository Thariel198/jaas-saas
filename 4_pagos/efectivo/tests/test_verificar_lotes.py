"""
test_verificar_lotes.py — Tests de 4_pagos/efectivo/verificar_lotes.py.

Cobertura (patrón 3.6c metodología — estado mínimo sintético + assert con mensaje claro):

  Primitivos puros (capa 3 — tabla de confusión):
    1. test_mz_confundible               — U/W sí · A/A1 sí (sufijo) · A/B no
    2. test_lt_confundible               — 19/14 · 4/14 (autocompletar) · 12/21 · 3/3A
    3. test_confundible_nivel            — simple = 1 · manzana Y lote = 2
    4. test_subconjuntos                 — toda combinación de cargos > 0
    5. test_clasificar                   — parcial baja un nivel respecto de saldar todo

  Capa 1 y 2 — cuadre y evidencia:
    6. test_cuadra_boleta_completa       — monto == total → COMO "boleta completa"
    7. test_cuadra_parcial               — monto == un cargo → COMO con el concepto
    8. test_no_cuadra                    — ningún subconjunto suma el monto
    9. test_omitido_monto_cero           — visita de reclamo sin cobro
   10. test_omitido_concepto             — CONCEPTO=tanque no es deuda de agua (bug A-4)
   11. test_sin_boleta                   — el MZ-LT escrito no existe en DATA_boletas

  CONTRAFACTUAL — los 2 errores reales del ciclo 2026-08:
   12. test_contrafactual_magda          — M-19 con S/9 → NO CUADRA + candidato M-14
   13. test_contrafactual_pedro          — G-13 con S/19 → NO CUADRA + candidato O-13

  Capa 3 y 4 — candidatos:
   14. test_doble_error_no_se_propone    — manzana Y lote mal → se cuenta, no se propone
   15. test_candidato_ya_pagado_descarta — un candidato ya confirmado no es la pista
   16. test_dos_candidatos_no_elige      — con 2 simples no se propone ninguno

  Preservación (3 capas):
   17. test_clave_preserv_no_colisiona   — dos lotes, mismo monto/día/cobrador
   18. test_preserva_resolucion          — roundtrip escribir → leer_resoluciones
   19. test_no_preserva_fila_vacia       — fila que el humano miró y dejó vacía no ensucia

  Guard de pipeline:
   20. test_guard_boleta_vieja_lanza     — pagos fuera de la ventana de la boleta
   21. test_guard_boleta_correcta_pasa   — pagos dentro de la ventana
"""

import shutil
import sys
from pathlib import Path

THIS = Path(__file__).resolve()
sys.path.insert(0, str(THIS.parent.parent))

import verificar_lotes as vl  # noqa: E402

TEST_ROOT = THIS.parent / "_tmp_verificar_lotes"


# ── Helpers ──────────────────────────────────────────────────────────────────

def _boleta(nombre="X", **cargos):
    """Boleta sintética: los cargos que se pasen, el total es su suma."""
    full = {n: 0.0 for n, _c in vl.CONCEPTOS}
    full.update(cargos)
    return {"nombre": nombre, "total": round(sum(full.values()), 2), "cargos": full,
            "emision": "27/07/2026", "vencimiento": "02/08/2026"}


def _pago(mz, lt, monto, mesa="mesa_1", hoja="registro_1", fila=4,
          cobrador="Tester", fecha="01/08/2026", concepto=""):
    return {"mesa": mesa, "hoja": hoja, "fila_excel": fila, "cobrador": cobrador,
            "fecha": fecha, "mz": mz, "lt": lt, "monto": monto, "concepto": concepto}


def _uno(filas, boletas):
    return vl.verificar(filas, boletas)[0]


def _setup():
    if TEST_ROOT.exists():
        shutil.rmtree(TEST_ROOT)
    TEST_ROOT.mkdir(parents=True)
    # Redirigir los directorios de ESCRITURA del módulo a la carpeta temporal.
    # La mayoría de estos tests son funciones puras que no tocan disco, pero
    # algunos escriben (escribir(), la preservación) y bastaba con que uno
    # apuntara al repo real para pisar las mesas del ciclo en curso — que es lo
    # que pasó el 12/08 con test_integracion (ver tests/conftest.py).
    (TEST_ROOT / "inputs").mkdir(exist_ok=True)
    (TEST_ROOT / "outputs").mkdir(exist_ok=True)
    (TEST_ROOT / "backup").mkdir(exist_ok=True)
    vl.BASE_DIR = TEST_ROOT
    vl.INPUTS_DIR = TEST_ROOT / "inputs"
    vl.OUTPUTS_DIR = TEST_ROOT / "outputs"
    vl.BACKUP_DIR = TEST_ROOT / "backup"


def _teardown():
    if TEST_ROOT.exists():
        shutil.rmtree(TEST_ROOT)


# ── 1-5 · Primitivos puros ───────────────────────────────────────────────────

def test_mz_confundible():
    assert vl.mz_confundible("W", "U"), "U y W están en el mismo grupo de trazo"
    assert vl.mz_confundible("G", "O"), "G y O se confunden"
    assert vl.mz_confundible("A", "A1"), "sufijo 1 del autocompletar"
    assert vl.mz_confundible("A1", "A"), "la relación es simétrica"
    assert not vl.mz_confundible("A", "Z"), "A y Z no se parecen"
    assert not vl.mz_confundible("A", "A"), "una manzana no se confunde consigo misma"
    print("  [OK] test_mz_confundible")


def test_lt_confundible():
    assert vl.lt_confundible("19", "14"), "9 y 4 se confunden a mano (caso Magda)"
    assert vl.lt_confundible("4", "14"), "dígito de más — el autocompletar"
    assert vl.lt_confundible("2", "12"), "idem B-2 / B-12"
    assert vl.lt_confundible("12", "21"), "transposición"
    assert vl.lt_confundible("3", "3A"), "sufijo de lote"
    assert not vl.lt_confundible("1", "8"), "1 y 8 no están emparejados"
    assert not vl.lt_confundible("5", "5"), "un lote no se confunde consigo mismo"
    print("  [OK] test_lt_confundible")


def test_confundible_nivel():
    assert vl.confundible("M-19", "M-14")[1] == 1, "solo el lote mal → simple"
    assert vl.confundible("G-13", "O-13")[1] == 1, "solo la manzana mal → simple"
    assert vl.confundible("G-17", "Q-12")[1] == 2, "manzana Y lote mal → doble"
    assert vl.confundible("A-1", "Z-99") is None, "nada en común"
    print("  [OK] test_confundible_nivel")


def test_subconjuntos():
    sub = vl.subconjuntos({"consumo": 5.0, "mant": 3.0, "anterior": 8.0, "convenio": 0.0})
    assert sub[5.0] == "consumo", "un solo cargo"
    assert sub[8.0] in ("mant+anterior", "anterior"), f"8 se arma de dos formas: {sub[8.0]}"
    assert 16.0 in sub, "la suma de los tres"
    assert 0.0 not in sub, "un cargo en 0 no participa"
    print("  [OK] test_subconjuntos")


def test_clasificar():
    assert vl.clasificar(True, 1) == "ALTA", "saldó todo y el monto es casi único"
    assert vl.clasificar(True, 20) == "MEDIA"
    assert vl.clasificar(True, 101) == "BAJA", "S/8 lo deben 101 lotes — no prueba nada"
    assert vl.clasificar(False, 1) == "MEDIA", "parcial baja un nivel aunque sea único"
    assert vl.clasificar(False, 20) == "BAJA"
    print("  [OK] test_clasificar")


# ── 6-11 · Capas 1 y 2 ───────────────────────────────────────────────────────

def test_cuadra_boleta_completa():
    boletas = {"A-1": _boleta(consumo=5.0, mant=3.0)}
    f = _uno([_pago("A", "1", 8.0)], boletas)
    assert f["evidencia"] in ("ALTA", "MEDIA", "BAJA"), f["evidencia"]
    assert f["como"] == "boleta completa", f["como"]
    print("  [OK] test_cuadra_boleta_completa")


def test_cuadra_parcial():
    boletas = {"A-1": _boleta(consumo=5.0, mant=3.0, convenio=50.0)}
    f = _uno([_pago("A", "1", 8.0)], boletas)
    assert f["como"] == "consumo+mant", f["como"]
    assert f["evidencia"] != "NO CUADRA"
    print("  [OK] test_cuadra_parcial")


def test_no_cuadra():
    boletas = {"A-1": _boleta(consumo=5.0, mant=3.0)}
    f = _uno([_pago("A", "1", 41.0)], boletas)
    assert f["evidencia"] == "NO CUADRA", f["evidencia"]
    assert f["como"] == ""
    print("  [OK] test_no_cuadra")


def test_omitido_monto_cero():
    boletas = {"A-1": _boleta(consumo=5.0)}
    f = _uno([_pago("A", "1", 0.0)], boletas)
    assert f["evidencia"] == "OMITIDO", "monto 0 = visita de reclamo, no se evalúa"
    print("  [OK] test_omitido_monto_cero")


def test_omitido_concepto():
    boletas = {"A-4": _boleta(consumo=5.0)}
    f = _uno([_pago("A", "4", 100.0, concepto="tanque")], boletas)
    assert f["evidencia"] == "OMITIDO", ("el aporte al tanque no es deuda de agua — "
                                         "marcarlo NO CUADRA sería ruido permanente")
    print("  [OK] test_omitido_concepto")


def test_sin_boleta():
    f = _uno([_pago("Z", "99", 10.0)], {"A-1": _boleta(consumo=5.0)})
    assert f["evidencia"] == "SIN BOLETA", f["evidencia"]
    print("  [OK] test_sin_boleta")


# ── 12-13 · CONTRAFACTUAL: los 2 errores reales del ciclo 2026-08 ───────────

def test_contrafactual_magda():
    """M-19 (Andres Quito, boleta 18) vs M-14 (Magda Blas, boleta 9). Se cobró
    S/9 y se anotó M-19. Datos reales de mesa_2 fila 76, 02/08/2026."""
    boletas = {
        "M-19": _boleta("ANDRES QUITO RODRIGUEZ", consumo=15.0, mant=3.0),
        "M-14": _boleta("MAGDA MARIA BLAS MORALES", consumo=6.0, mant=3.0),
    }
    f = _uno([_pago("M", "19", 9.0)], boletas)
    assert f["evidencia"] == "NO CUADRA", (
        f"el error de Magda tiene que salir marcado, salió {f['evidencia']}")
    assert f["candidato"] == "M-14", f"debía proponer M-14, propuso {f['candidato']!r}"
    assert f["motivo"] == "lote 19→14", f["motivo"]
    print("  [OK] test_contrafactual_magda")


def test_contrafactual_pedro():
    """G-13 (Reynaldo Melgarejo, boleta 37) vs O-13 (Pedro Mendoza, boleta 19).
    Se cobró S/19 y se anotó G-13. Datos reales de mesa_2 fila 107, 02/08/2026."""
    boletas = {
        "G-13": _boleta("REYNALDO MELGAREJO VILCARINO", consumo=34.0, mant=3.0),
        "O-13": _boleta("PEDRO MENDOZA MARQUINA", consumo=16.0, mant=3.0),
    }
    f = _uno([_pago("G", "13", 19.0)], boletas)
    assert f["evidencia"] == "NO CUADRA", (
        f"el error de Pedro tiene que salir marcado, salió {f['evidencia']}")
    assert f["candidato"] == "O-13", f"debía proponer O-13, propuso {f['candidato']!r}"
    assert f["motivo"] == "manzana G→O", f["motivo"]
    print("  [OK] test_contrafactual_pedro")


# ── 14-16 · Capas 3 y 4 ──────────────────────────────────────────────────────

def test_doble_error_no_se_propone():
    """G-17 → Q-12: G→Q y 17→12, los dos mal. Proponerlo sería inventar."""
    boletas = {
        "G-17": _boleta("CARLOS SIGUENAS", consumo=5.0, mant=3.0, convenio=132.0),
        "Q-12": _boleta("TEODORA MEZA", consumo=15.0, mant=3.0),
    }
    f = _uno([_pago("G", "17", 18.0)], boletas)
    assert f["evidencia"] == "NO CUADRA"
    assert f["candidato"] == "", f"no debía proponer nada, propuso {f['candidato']!r}"
    assert "doble" in f["motivo"], f"debía explicar por qué no propone: {f['motivo']!r}"
    print("  [OK] test_doble_error_no_se_propone")


def test_candidato_ya_pagado_descarta():
    """Si el vecino candidato ya tiene su pago confirmado en esta corrida, no es
    la pista — mandaría al supervisor a buscar donde no hay nada."""
    boletas = {
        "M-19": _boleta("ANDRES QUITO", consumo=15.0, mant=3.0),
        "M-14": _boleta("MAGDA BLAS", consumo=6.0, mant=3.0),
    }
    filas = vl.verificar([_pago("M", "19", 9.0, fila=4),
                          _pago("M", "14", 9.0, fila=5)], boletas)
    malo = [f for f in filas if f["clave"] == "M-19"][0]
    bueno = [f for f in filas if f["clave"] == "M-14"][0]
    assert bueno["evidencia"] != "NO CUADRA", "M-14 pagó su boleta exacta"
    assert malo["candidato"] == "", ("M-14 ya está pagado, no puede ser el candidato "
                                     f"de M-19; propuso {malo['candidato']!r}")
    print("  [OK] test_candidato_ya_pagado_descarta")


def test_dos_candidatos_no_elige():
    boletas = {
        "A-1": _boleta("TITULAR", consumo=99.0),
        "A-2": _boleta("VECINO A", consumo=10.0),
        "A-11": _boleta("VECINO B", consumo=10.0),
    }
    f = _uno([_pago("A", "1", 10.0)], boletas)
    assert f["evidencia"] == "NO CUADRA"
    assert f["candidato"] == "", "con 2 candidatos simples no se elige ninguno"
    assert "2 candidatos" in f["motivo"], f["motivo"]
    print("  [OK] test_dos_candidatos_no_elige")


# ── 17-19 · Preservación ─────────────────────────────────────────────────────

def test_clave_preserv_no_colisiona():
    """El bug que la clave sin MZ/LT tenía: 62% de las filas reales colisionaban."""
    a = _pago("I", "8", 8.0, fila=4)
    b = _pago("O", "5", 8.0, fila=5)
    assert vl._clave_preserv(a) != vl._clave_preserv(b), (
        "dos lotes distintos con el mismo monto, día y cobrador no pueden compartir clave")
    print("  [OK] test_clave_preserv_no_colisiona")


def test_preserva_resolucion():
    ruta = TEST_ROOT / "verificacion_lotes_2026-08.xlsx"
    boletas = {"M-19": _boleta("ANDRES QUITO", consumo=15.0, mant=3.0),
               "M-14": _boleta("MAGDA BLAS", consumo=6.0, mant=3.0)}
    filas = vl.verificar([_pago("M", "19", 9.0)], boletas)

    vl.escribir(filas, ruta, {})                       # corrida 1
    _escribir_resolucion(ruta, "corrige", "M", "14")   # el supervisor la llena

    previas = vl.leer_resoluciones(ruta)               # corrida 2
    assert previas, "la resolución del supervisor tiene que leerse"
    filas2 = vl.verificar([_pago("M", "19", 9.0)], boletas)
    n = vl.escribir(filas2, ruta, previas)
    assert n == 1, f"se debía preservar 1 resolución, se preservaron {n}"

    from openpyxl import load_workbook
    ws = load_workbook(ruta).active
    hdr = {str(c.value).strip(): i for i, c in enumerate(ws[2])}
    fila3 = [c.value for c in ws[3]]
    assert fila3[hdr["RESOLUCION"]] == "corrige", "la RESOLUCION sobrevivió la re-corrida"
    assert fila3[hdr["LT_CORRECTO"]] == "14", "el LT_CORRECTO también"
    print("  [OK] test_preserva_resolucion")


def test_no_preserva_fila_vacia():
    ruta = TEST_ROOT / "vacia.xlsx"
    boletas = {"A-1": _boleta("X", consumo=5.0, mant=3.0)}
    filas = vl.verificar([_pago("A", "1", 8.0)], boletas)
    vl.escribir(filas, ruta, {})
    previas = vl.leer_resoluciones(ruta)
    assert previas == {}, ("una fila que el humano miró y dejó vacía no debe entrar al "
                           f"mapa de preservación; entraron {len(previas)}")
    print("  [OK] test_no_preserva_fila_vacia")


def _escribir_resolucion(ruta: Path, resolucion, mz_c, lt_c) -> None:
    """Simula al supervisor llenando las 3 columnas humanas en Excel."""
    from openpyxl import load_workbook
    wb = load_workbook(ruta)
    ws = wb.active
    hdr = {str(c.value).strip(): i + 1 for i, c in enumerate(ws[2])}
    ws.cell(row=3, column=hdr["RESOLUCION"]).value = resolucion
    ws.cell(row=3, column=hdr["MZ_CORRECTO"]).value = mz_c
    ws.cell(row=3, column=hdr["LT_CORRECTO"]).value = lt_c
    wb.save(ruta)


# ── 20-21 · Guard de pipeline ────────────────────────────────────────────────

def test_guard_boleta_vieja_lanza():
    boletas = {"A-1": {"nombre": "X", "total": 8.0, "cargos": {},
                       "emision": "27/05/2026", "vencimiento": "02/06/2026"}}
    filas = [_pago("A", "1", 8.0, fecha="01/08/2026")]
    try:
        vl.guard_pipeline(filas, boletas)
    except ValueError as e:
        assert "no corresponde" in str(e), str(e)
        print("  [OK] test_guard_boleta_vieja_lanza")
        return
    raise AssertionError("una boleta de junio con pagos de agosto tenía que frenar la corrida")


def test_guard_boleta_correcta_pasa():
    boletas = {"A-1": _boleta("X", consumo=8.0)}       # emisión 27/07, vence 02/08
    vl.guard_pipeline([_pago("A", "1", 8.0, fecha="01/08/2026")], boletas)
    print("  [OK] test_guard_boleta_correcta_pasa")


# ── Runner ───────────────────────────────────────────────────────────────────

def main():
    tests = [
        test_mz_confundible,
        test_lt_confundible,
        test_confundible_nivel,
        test_subconjuntos,
        test_clasificar,
        test_cuadra_boleta_completa,
        test_cuadra_parcial,
        test_no_cuadra,
        test_omitido_monto_cero,
        test_omitido_concepto,
        test_sin_boleta,
        test_contrafactual_magda,
        test_contrafactual_pedro,
        test_doble_error_no_se_propone,
        test_candidato_ya_pagado_descarta,
        test_dos_candidatos_no_elige,
        test_clave_preserv_no_colisiona,
        test_preserva_resolucion,
        test_no_preserva_fila_vacia,
        test_guard_boleta_vieja_lanza,
        test_guard_boleta_correcta_pasa,
    ]
    for t in tests:
        _setup()
        try:
            t()
        except Exception:
            _teardown()
            raise
    _teardown()
    print(f"\n[OK] {len(tests)}/{len(tests)} tests de verificar_lotes pasaron")


if __name__ == "__main__":
    main()
