"""
test_unitarios.py — Tests unitarios de funciones individuales de 4_pagos/efectivo/main.py.

Cobertura (patrón 3.6c metodología — estado mínimo sintético + assert con mensaje claro):

  Utilidades:
    1. test_norm                          — _norm: strip + upper, None → ""
    2. test_llave                         — _llave: MZ-LT formato
    3. test_monto                         — _monto: coma/punto, inválido → 0, redondeo
    4. test_mes_de_datos                  — _mes_de_datos: extrae YYYY-MM de las filas

  Lectura inputs:
    5. test_leer_hoja_valida              — leer_hoja: parsea registros válidos
    6. test_leer_hoja_ignora_invalidos    — leer_hoja: descarta MZ/LT vacíos y monto<=0
    7. test_leer_mesas_sin_archivos       — leer_mesas: raise FileNotFoundError

  Cross-check:
    8. test_cross_check_1_hoja            — 1 hoja → solo_un_cobrador
    9. test_cross_check_2_hojas_coinciden — 2 hojas mismo monto → confirmado
   10. test_cross_check_2_hojas_difieren  — 2 hojas distinto monto → discrepancia
   11. test_cross_check_3_hojas_mayoria   — 2/3 coinciden → mayoria_aplicada
   12. test_cross_check_3_hojas_todas_distintas — 3 montos distintos → discrepancia
   13. test_cross_check_1_de_2_presente   — solo 1 hoja tiene el predio → discrepancia

  Multi-mesa:
   14. test_detectar_multi_mesa_separa    — mismo predio en 2 mesas → multi_mesa
   15. test_detectar_multi_mesa_misma     — mismo predio en 1 mesa → limpio
   16. test_detectar_multi_mesa_ya_resuelto — flag _ya_resuelto bypassa detección
   17. test_detectar_multi_mesa_ya_rechazado — flag _ya_rechazado descarta

  Resoluciones (ciclo 2+):
   18. test_aplicar_resoluciones_acepta_a — resolución acepta_a → confirmado_resuelto
   19. test_aplicar_resoluciones_corrige  — resolución corrige + MONTO_CORRECTO
   20. test_aplicar_resoluciones_vacia    — sin resolución → queda pendiente
   21. test_aplicar_multi_mesa_si         — ok=si → confirmado en pagos
   22. test_aplicar_multi_mesa_si_corrige_predio — ok=si + MZ_CORRECTO → imputa al corregido
   23. test_aplicar_multi_mesa_rechaza    — ok=rechaza → solo trazabilidad
   24. test_aplicar_multi_mesa_vacia      — sin ok → queda pendiente

  Ciclo:
   25. test_detectar_ciclo_sin_archivo    — sin trazabilidad → ciclo 1
   26. test_detectar_ciclo_acumula        — con trazabilidad ciclo 2 → 3
   27. test_detectar_ciclo_legacy         — acepta columna 'CICLO' antigua
"""

import shutil
import sys
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

THIS = Path(__file__).resolve()
sys.path.insert(0, str(THIS.parent.parent))

import main as efectivo  # noqa: E402  (módulo se llama main.py; alias por legibilidad)

TEST_ROOT = THIS.parent / "_tmp_unitarios"


# ── Helpers ──────────────────────────────────────────────────────────────────

def _setup():
    if TEST_ROOT.exists():
        shutil.rmtree(TEST_ROOT)
    TEST_ROOT.mkdir(parents=True)


def _teardown():
    if TEST_ROOT.exists():
        shutil.rmtree(TEST_ROOT)


_COLS_MESA = ["COBRADOR", "FECHA_REGISTRO", "MZ", "LT", "MONTO",
              "MONTO_EFECTIVO", "MONTO_YAPE", "FECHA", "COMENTARIO"]


def _crear_mesa(path: Path, hojas: dict):
    """
    Crea un mesa_N.xlsx con la estructura real: fila 1 = secciones, fila 2 = columnas,
    fila 3 = ejemplo (ignorada por leer_hoja), fila 4+ = datos reales.

    hojas: {"registro_1": [(cobrador, fecha_reg, mz, lt, monto, monto_efec, monto_yape, fecha, comentario), ...]}
    """
    wb = Workbook()
    wb.remove(wb.active)
    for nombre_hoja, filas in hojas.items():
        ws = wb.create_sheet(nombre_hoja)
        for ci, col in enumerate(_COLS_MESA, start=1):
            ws.cell(row=1, column=ci, value="")        # sección (placeholder)
            ws.cell(row=2, column=ci, value=col)        # nombre de columna
        # Fila 3 = ejemplo guía (leer_hoja la salta)
        ws.cell(row=3, column=1, value="(ejemplo)")
        for ri, fila in enumerate(filas, start=4):
            for ci, val in enumerate(fila, start=1):
                ws.cell(row=ri, column=ci, value=val)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _fila_mesa(mz, lt, monto, cobrador="Wilder", fecha="03/06/2026",
               monto_efec=None, monto_yape=0, comentario=""):
    """Fila de mesa con efectivo total por defecto."""
    if monto_efec is None:
        monto_efec = monto
    return (cobrador, "10/06/2026", mz, lt, monto, monto_efec, monto_yape, fecha, comentario)


def _registro(mz, lt, monto, mesa="mesa_1", cobrador="Wilder",
              fecha="03/06/2026", estado="confirmado", ciclo=1, comentario=""):
    """Dict de registro confirmado (formato interno de cross_check_mesa)."""
    return {
        "llave":            efectivo._llave(mz, lt),
        "mz":               mz,
        "lt":               lt,
        "monto":            monto,
        "fecha":            fecha,
        "cobrador":         cobrador,
        "comentario":       comentario,
        "mesa":             mesa,
        "estado":            estado,
        "ciclo_correccion": ciclo,
    }


def _crear_traz_disc_resueltas(path: Path, filas: list):
    """
    Crea trazabilidad/trazabilidad_YYYY-MM.xlsx con hoja discrepancias_resueltas.
    filas: lista de dicts con al menos MZ, LT, MESA, MONTO_FINAL, CICLO_CORRECCION.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "discrepancias_resueltas"
    cols = ["MZ", "LT", "MESA", "COBRADOR_A", "MONTO_A", "COBRADOR_B", "MONTO_B",
            "COBRADOR_C", "MONTO_C", "RESOLUCION", "MONTO_FINAL",
            "CICLO_CORRECCION", "FECHA_RESOLUCION"]
    for ci, c in enumerate(cols, start=1):
        ws.cell(row=1, column=ci, value="")
        ws.cell(row=2, column=ci, value=c)
    for ri, fila in enumerate(filas, start=3):
        for ci, c in enumerate(cols, start=1):
            ws.cell(row=ri, column=ci, value=fila.get(c, ""))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ── 1-4. Utilidades ──────────────────────────────────────────────────────────

def test_norm():
    assert efectivo._norm("a")     == "A"
    assert efectivo._norm("  a  ") == "A"
    assert efectivo._norm("8c")    == "8C"
    assert efectivo._norm(None)    == ""
    assert efectivo._norm(0)       == "0"
    print("  OK test_norm")


def test_llave():
    assert efectivo._llave("a", "8c")    == "A-8C"
    assert efectivo._llave(" A ", " 8C ") == "A-8C"
    assert efectivo._llave("P", "11A")    == "P-11A"
    # MZ distinto → llave distinta
    assert efectivo._llave("A", "1") != efectivo._llave("B", "1")
    print("  OK test_llave")


def test_monto():
    assert efectivo._monto(38.0)     == 38.0
    assert efectivo._monto("40.00")  == 40.0
    assert efectivo._monto("38,50")  == 38.5
    assert efectivo._monto("abc")    == 0.0
    assert efectivo._monto(None)     == 0.0
    assert efectivo._monto(20)       == 20.0
    assert efectivo._monto("38.256") == 38.26  # redondea a 2 decimales
    print("  OK test_monto")


def test_mes_de_datos():
    # Formato DD/MM/YYYY
    rows = [{"fecha": "03/06/2026"}, {"fecha": "05/06/2026"}]
    assert efectivo._mes_de_datos(rows) == "2026-06", \
        f"Esperaba '2026-06', llegó {efectivo._mes_de_datos(rows)!r}"
    # Formato YYYY-MM-DD
    rows = [{"fecha": "2026-07-15"}]
    assert efectivo._mes_de_datos(rows) == "2026-07"
    # Sin fechas válidas → fallback a fecha actual (al menos formato YYYY-MM)
    rows = [{"fecha": ""}]
    res = efectivo._mes_de_datos(rows)
    assert len(res) == 7 and res[4] == "-", f"Fallback inválido: {res!r}"
    print("  OK test_mes_de_datos")


# ── 5-7. Lectura inputs ──────────────────────────────────────────────────────

def test_leer_hoja_valida():
    p = TEST_ROOT / "mesa_1.xlsx"
    _crear_mesa(p, {"registro_1": [_fila_mesa("A", "8C", 38.0)]})
    regs = efectivo.leer_hoja(p, "registro_1")
    assert len(regs) == 1, f"Esperaba 1 registro, llegaron {len(regs)}"
    r = regs[0]
    assert r["mz"]       == "A"
    assert r["lt"]       == "8C"
    assert r["monto"]    == 38.0
    assert r["llave"]    == "A-8C"
    assert r["cobrador"] == "Wilder"
    print("  OK test_leer_hoja_valida")


def test_leer_hoja_ignora_invalidos():
    p = TEST_ROOT / "mesa_1.xlsx"
    _crear_mesa(p, {
        "registro_1": [
            _fila_mesa("A", "8C", 38.0),       # válido
            _fila_mesa("",  "1",  20.0),       # MZ vacío
            _fila_mesa("B", "",   20.0),       # LT vacío
            _fila_mesa("C", "5",  0.0),        # monto_efec = 0 (todo Yape)
        ]
    })
    regs = efectivo.leer_hoja(p, "registro_1")
    assert len(regs) == 1, f"Esperaba 1 válido, llegaron {len(regs)}: {regs}"
    assert regs[0]["mz"] == "A"
    print("  OK test_leer_hoja_ignora_invalidos")


def test_leer_mesas_sin_archivos():
    # Apuntar INPUTS_DIR a una carpeta vacía
    inputs_vacios = TEST_ROOT / "inputs_vacios"
    inputs_vacios.mkdir()
    orig = efectivo.INPUTS_DIR
    efectivo.INPUTS_DIR = inputs_vacios
    try:
        try:
            efectivo.leer_mesas()
            assert False, "Esperaba FileNotFoundError"
        except FileNotFoundError as e:
            assert "mesa_N.xlsx" in str(e), f"Mensaje inesperado: {e}"
    finally:
        efectivo.INPUTS_DIR = orig
    print("  OK test_leer_mesas_sin_archivos")


# ── 8-13. Cross-check por mesa ───────────────────────────────────────────────

def test_cross_check_1_hoja():
    mesa = {
        "nombre": "mesa_1",
        "hojas": {"registro_1": [{
            "llave": "A-8C", "mz": "A", "lt": "8C", "monto": 38.0,
            "fecha": "03/06/2026", "cobrador": "Wilder", "comentario": "",
        }]}
    }
    conf, disc, res, sc = efectivo.cross_check_mesa(mesa, ciclo=1, fecha_run=datetime.now())
    assert len(conf) == 1 and conf[0]["estado"] == "solo_un_cobrador"
    assert len(disc) == 0 and len(res) == 0
    assert len(sc) == 1 and sc[0]["MZ"] == "A"
    print("  OK test_cross_check_1_hoja")


def test_cross_check_2_hojas_coinciden():
    base = {"llave": "A-8C", "mz": "A", "lt": "8C", "monto": 38.0,
            "fecha": "03/06/2026", "comentario": ""}
    mesa = {
        "nombre": "mesa_1",
        "hojas": {
            "registro_1": [{**base, "cobrador": "Wilder"}],
            "registro_2": [{**base, "cobrador": "María"}],
        }
    }
    conf, disc, res, sc = efectivo.cross_check_mesa(mesa, ciclo=1, fecha_run=datetime.now())
    assert len(conf) == 1 and conf[0]["estado"] == "confirmado", \
        f"Esperaba 1 confirmado, llegó: {conf}"
    assert len(disc) == 0 and len(res) == 0 and len(sc) == 0
    print("  OK test_cross_check_2_hojas_coinciden")


def test_cross_check_2_hojas_difieren():
    base = {"llave": "A-8C", "mz": "A", "lt": "8C",
            "fecha": "03/06/2026", "comentario": ""}
    mesa = {
        "nombre": "mesa_1",
        "hojas": {
            "registro_1": [{**base, "monto": 38.0, "cobrador": "Wilder"}],
            "registro_2": [{**base, "monto": 40.0, "cobrador": "María"}],
        }
    }
    conf, disc, res, sc = efectivo.cross_check_mesa(mesa, ciclo=1, fecha_run=datetime.now())
    assert len(conf) == 0, f"No debería haber confirmados: {conf}"
    assert len(disc) == 1, f"Esperaba 1 discrepancia, llegaron: {disc}"
    d = disc[0]
    assert d["cobrador_a"] == "Wilder" and d["monto_a"] == 38.0
    assert d["cobrador_b"] == "María"  and d["monto_b"] == 40.0
    assert d["cobrador_c"] == ""       and d["monto_c"] == ""
    print("  OK test_cross_check_2_hojas_difieren")


def test_cross_check_3_hojas_mayoria():
    """2 hojas dicen 38, 1 dice 40 → mayoría_aplicada con 38."""
    base = {"llave": "A-8C", "mz": "A", "lt": "8C",
            "fecha": "03/06/2026", "comentario": ""}
    mesa = {
        "nombre": "mesa_1",
        "hojas": {
            "registro_1": [{**base, "monto": 38.0, "cobrador": "Wilder"}],
            "registro_2": [{**base, "monto": 38.0, "cobrador": "María"}],
            "registro_3": [{**base, "monto": 40.0, "cobrador": "Ana"}],
        }
    }
    conf, disc, res, sc = efectivo.cross_check_mesa(mesa, ciclo=1, fecha_run=datetime.now())
    assert len(conf) == 1 and conf[0]["estado"] == "mayoria_aplicada", \
        f"Esperaba mayoria_aplicada, llegó: {conf}"
    assert conf[0]["monto"] == 38.0, f"Esperaba 38.0 (mayoría), llegó {conf[0]['monto']}"
    assert len(res) == 1 and "mayoria_" in res[0]["RESOLUCION"]
    assert len(disc) == 0
    print("  OK test_cross_check_3_hojas_mayoria")


def test_cross_check_3_hojas_todas_distintas():
    """3 montos distintos → discrepancia (sin mayoría)."""
    base = {"llave": "A-8C", "mz": "A", "lt": "8C",
            "fecha": "03/06/2026", "comentario": ""}
    mesa = {
        "nombre": "mesa_1",
        "hojas": {
            "registro_1": [{**base, "monto": 38.0, "cobrador": "Wilder"}],
            "registro_2": [{**base, "monto": 40.0, "cobrador": "María"}],
            "registro_3": [{**base, "monto": 42.0, "cobrador": "Ana"}],
        }
    }
    conf, disc, res, sc = efectivo.cross_check_mesa(mesa, ciclo=1, fecha_run=datetime.now())
    assert len(conf) == 0 and len(res) == 0
    assert len(disc) == 1
    print("  OK test_cross_check_3_hojas_todas_distintas")


def test_cross_check_1_de_2_presente():
    """El predio solo aparece en 1 de las 2 hojas → discrepancia."""
    mesa = {
        "nombre": "mesa_1",
        "hojas": {
            "registro_1": [{"llave": "A-8C", "mz": "A", "lt": "8C", "monto": 38.0,
                            "fecha": "03/06/2026", "cobrador": "Wilder", "comentario": ""}],
            "registro_2": [{"llave": "B-1", "mz": "B", "lt": "1", "monto": 20.0,
                            "fecha": "03/06/2026", "cobrador": "María", "comentario": ""}],
        }
    }
    conf, disc, res, sc = efectivo.cross_check_mesa(mesa, ciclo=1, fecha_run=datetime.now())
    assert len(conf) == 0
    assert len(disc) == 2, f"Esperaba 2 discrepancias (1 por cada predio único), llegaron: {disc}"
    print("  OK test_cross_check_1_de_2_presente")


# ── 14-17. Multi-mesa ────────────────────────────────────────────────────────

def test_detectar_multi_mesa_separa():
    """Mismo predio (A-8C) cobrado en mesa_1 y mesa_2 → multi_mesa, nada limpio."""
    confirmados = [
        _registro("A", "8C", 38.0, mesa="mesa_1", cobrador="Wilder"),
        _registro("A", "8C", 40.0, mesa="mesa_2", cobrador="María"),
    ]
    limpios, multi = efectivo.detectar_multi_mesa(confirmados)
    assert len(limpios) == 0, f"No debería haber limpios: {limpios}"
    assert len(multi) == 2,   f"Esperaba 2 filas multi, llegaron: {multi}"
    mesas = {m["mesa"] for m in multi}
    assert mesas == {"mesa_1", "mesa_2"}
    print("  OK test_detectar_multi_mesa_separa")


def test_detectar_multi_mesa_misma():
    """Mismo predio en MISMA mesa (raro, pero válido) → limpio, no multi."""
    confirmados = [
        _registro("A", "8C", 38.0, mesa="mesa_1", cobrador="Wilder"),
        _registro("B", "1",  20.0, mesa="mesa_1", cobrador="María"),
    ]
    limpios, multi = efectivo.detectar_multi_mesa(confirmados)
    assert len(limpios) == 2
    assert len(multi)   == 0
    print("  OK test_detectar_multi_mesa_misma")


def test_detectar_multi_mesa_ya_resuelto():
    """Flag _ya_resuelto bypassa la detección y va directo a limpios."""
    r = _registro("A", "8C", 38.0, mesa="mesa_1")
    r["_ya_resuelto"] = True
    confirmados = [
        r,
        _registro("A", "8C", 38.0, mesa="mesa_2"),  # otro pago en otra mesa
    ]
    limpios, multi = efectivo.detectar_multi_mesa(confirmados)
    # El _ya_resuelto entra a limpios; el otro queda solo (sin par) → limpio también
    assert len(limpios) == 2, f"Esperaba 2 limpios, llegaron: {limpios}"
    assert len(multi) == 0
    print("  OK test_detectar_multi_mesa_ya_resuelto")


def test_detectar_multi_mesa_ya_rechazado():
    """Flag _ya_rechazado descarta la fila (no entra a limpios ni a multi)."""
    r = _registro("A", "8C", 38.0, mesa="mesa_1")
    r["_ya_rechazado"] = True
    limpios, multi = efectivo.detectar_multi_mesa([r])
    assert len(limpios) == 0
    assert len(multi)   == 0
    print("  OK test_detectar_multi_mesa_ya_rechazado")


# ── 18-24. Resoluciones ──────────────────────────────────────────────────────

def _disc_row(mz="A", lt="8C", mesa="mesa_1"):
    return {
        "mz": mz, "lt": lt, "llave": efectivo._llave(mz, lt),
        "mesa": mesa, "fecha": "03/06/2026",
        "cobrador_a": "Wilder", "monto_a": 38.0,
        "cobrador_b": "María",  "monto_b": 40.0,
        "cobrador_c": "",       "monto_c": "",
        "resolucion": "", "monto_correcto": "",
    }


def test_aplicar_resoluciones_acepta_a():
    discrepancias = [_disc_row()]
    resoluciones = {"A-8C": {
        "resolucion": "acepta_a", "monto_correcto": 0,
        "monto_a": 38.0, "monto_b": 40.0, "monto_c": 0,
        "cobrador_a": "Wilder", "fecha": "03/06/2026", "mesa": "mesa_1",
    }}
    conf, traz, pend = efectivo.aplicar_resoluciones(
        discrepancias, resoluciones, ciclo=2, fecha_run=datetime.now()
    )
    assert len(conf) == 1 and conf[0]["monto"] == 38.0
    assert conf[0]["estado"] == "discrepancia_resuelta"
    assert conf[0]["ciclo_correccion"] == 2
    assert len(traz) == 1 and traz[0]["RESOLUCION"] == "acepta_a"
    assert len(pend) == 0
    print("  OK test_aplicar_resoluciones_acepta_a")


def test_aplicar_resoluciones_corrige():
    discrepancias = [_disc_row()]
    resoluciones = {"A-8C": {
        "resolucion": "corrige", "monto_correcto": 39.5,
        "monto_a": 0, "monto_b": 0, "monto_c": 0,
        "cobrador_a": "", "fecha": "", "mesa": "",
    }}
    conf, traz, pend = efectivo.aplicar_resoluciones(
        discrepancias, resoluciones, ciclo=2, fecha_run=datetime.now()
    )
    assert len(conf) == 1 and conf[0]["monto"] == 39.5, \
        f"Esperaba monto corregido 39.5, llegó {conf[0]['monto']}"
    assert traz[0]["RESOLUCION"] == "corrige"
    print("  OK test_aplicar_resoluciones_corrige")


def test_aplicar_resoluciones_vacia():
    discrepancias = [_disc_row()]
    conf, traz, pend = efectivo.aplicar_resoluciones(
        discrepancias, resoluciones={}, ciclo=2, fecha_run=datetime.now()
    )
    assert len(conf) == 0
    assert len(traz) == 0
    assert len(pend) == 1, "Sin resolución → debería quedar pendiente"
    print("  OK test_aplicar_resoluciones_vacia")


def _multi_row(mz="A", lt="8C", mesa="mesa_1", monto=38.0, cobrador="Wilder"):
    return {
        "llave": efectivo._llave(mz, lt),
        "mz": mz, "lt": lt, "mesa": mesa,
        "cobrador": cobrador, "monto": monto, "fecha": "03/06/2026",
    }


def test_aplicar_multi_mesa_si():
    multi = [_multi_row()]
    resoluciones = {("A-8C", "mesa_1"): {"ok": "si", "mz_correcto": "", "lt_correcto": ""}}
    conf, traz, pend = efectivo.aplicar_resoluciones_multi_mesa(
        multi, resoluciones, ciclo=2, fecha_run=datetime.now()
    )
    assert len(conf) == 1 and conf[0]["estado"] == "multi_mesa_resuelta"
    assert conf[0]["mz"] == "A" and conf[0]["lt"] == "8C"  # sin corrección
    assert len(traz) == 1 and traz[0]["RESOLUCION"] == "si"
    assert len(pend) == 0
    print("  OK test_aplicar_multi_mesa_si")


def test_aplicar_multi_mesa_si_corrige_predio():
    multi = [_multi_row(mz="A", lt="8C")]
    resoluciones = {("A-8C", "mesa_1"): {"ok": "si", "mz_correcto": "B", "lt_correcto": "12"}}
    conf, traz, pend = efectivo.aplicar_resoluciones_multi_mesa(
        multi, resoluciones, ciclo=2, fecha_run=datetime.now()
    )
    assert len(conf) == 1
    assert conf[0]["mz"] == "B" and conf[0]["lt"] == "12", \
        f"Esperaba predio corregido B-12, llegó {conf[0]['mz']}-{conf[0]['lt']}"
    # Trazabilidad guarda el original
    assert traz[0]["MZ_ORIGINAL"] == "A" and traz[0]["LT_ORIGINAL"] == "8C"
    assert traz[0]["MZ"]          == "B" and traz[0]["LT"]          == "12"
    print("  OK test_aplicar_multi_mesa_si_corrige_predio")


def test_aplicar_multi_mesa_rechaza():
    multi = [_multi_row()]
    resoluciones = {("A-8C", "mesa_1"): {"ok": "rechaza", "mz_correcto": "", "lt_correcto": ""}}
    conf, traz, pend = efectivo.aplicar_resoluciones_multi_mesa(
        multi, resoluciones, ciclo=2, fecha_run=datetime.now()
    )
    assert len(conf) == 0, "Rechazado no debería ir a pagos"
    assert len(traz) == 1 and traz[0]["RESOLUCION"] == "rechaza", \
        "Rechazado SÍ debería ir a trazabilidad para auditoría"
    assert len(pend) == 0
    print("  OK test_aplicar_multi_mesa_rechaza")


def test_aplicar_multi_mesa_vacia():
    multi = [_multi_row()]
    conf, traz, pend = efectivo.aplicar_resoluciones_multi_mesa(
        multi, resoluciones={}, ciclo=2, fecha_run=datetime.now()
    )
    assert len(conf) == 0 and len(traz) == 0
    assert len(pend) == 1, "Sin ok → debería quedar pendiente"
    print("  OK test_aplicar_multi_mesa_vacia")


# ── 25-27. Detección de ciclo ────────────────────────────────────────────────

def test_detectar_ciclo_sin_archivo():
    p = TEST_ROOT / "no_existe.xlsx"
    assert efectivo.detectar_ciclo(p) == 1
    print("  OK test_detectar_ciclo_sin_archivo")


def test_detectar_ciclo_acumula():
    p = TEST_ROOT / "trazabilidad_2026-06.xlsx"
    _crear_traz_disc_resueltas(p, [
        {"MZ": "A", "LT": "8C", "MESA": "mesa_1", "MONTO_FINAL": 38.0, "CICLO_CORRECCION": 2},
    ])
    assert efectivo.detectar_ciclo(p) == 3, \
        f"Con max ciclo=2 debería retornar 3, llegó {efectivo.detectar_ciclo(p)}"
    print("  OK test_detectar_ciclo_acumula")


def test_detectar_ciclo_legacy():
    """detectar_ciclo acepta la columna legacy 'CICLO' para migración."""
    p = TEST_ROOT / "trazabilidad_legacy.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "discrepancias_resueltas"
    cols = ["MZ", "LT", "MESA", "RESOLUCION", "MONTO_FINAL", "CICLO"]  # CICLO antiguo
    for ci, c in enumerate(cols, start=1):
        ws.cell(row=1, column=ci, value="")
        ws.cell(row=2, column=ci, value=c)
    ws.cell(row=3, column=1, value="A")
    ws.cell(row=3, column=2, value="8C")
    ws.cell(row=3, column=3, value="mesa_1")
    ws.cell(row=3, column=4, value="acepta_a")
    ws.cell(row=3, column=5, value=38.0)
    ws.cell(row=3, column=6, value=1)
    wb.save(p)
    assert efectivo.detectar_ciclo(p) == 2, \
        f"Con CICLO legacy=1 debería retornar 2, llegó {efectivo.detectar_ciclo(p)}"
    print("  OK test_detectar_ciclo_legacy")


# ── Runner ───────────────────────────────────────────────────────────────────

def main():
    tests = [
        # Utilidades
        test_norm,
        test_llave,
        test_monto,
        test_mes_de_datos,
        # Lectura inputs
        test_leer_hoja_valida,
        test_leer_hoja_ignora_invalidos,
        test_leer_mesas_sin_archivos,
        # Cross-check
        test_cross_check_1_hoja,
        test_cross_check_2_hojas_coinciden,
        test_cross_check_2_hojas_difieren,
        test_cross_check_3_hojas_mayoria,
        test_cross_check_3_hojas_todas_distintas,
        test_cross_check_1_de_2_presente,
        # Multi-mesa
        test_detectar_multi_mesa_separa,
        test_detectar_multi_mesa_misma,
        test_detectar_multi_mesa_ya_resuelto,
        test_detectar_multi_mesa_ya_rechazado,
        # Resoluciones
        test_aplicar_resoluciones_acepta_a,
        test_aplicar_resoluciones_corrige,
        test_aplicar_resoluciones_vacia,
        test_aplicar_multi_mesa_si,
        test_aplicar_multi_mesa_si_corrige_predio,
        test_aplicar_multi_mesa_rechaza,
        test_aplicar_multi_mesa_vacia,
        # Ciclo
        test_detectar_ciclo_sin_archivo,
        test_detectar_ciclo_acumula,
        test_detectar_ciclo_legacy,
    ]
    for t in tests:
        _setup()
        try:
            t()
        except Exception:
            _teardown()
            raise
    _teardown()
    print(f"\n[OK] {len(tests)}/{len(tests)} tests unitarios pasaron")


if __name__ == "__main__":
    main()
