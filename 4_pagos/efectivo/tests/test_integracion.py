"""
test_integracion.py — Tests del flujo completo de 4_pagos/efectivo/main.main().

Cobertura (patrón 3.6c metodología — estado mínimo sintético + assert con mensaje claro):

  Ciclo 1 (primera corrida del mes):
    1. test_ciclo1_una_mesa_un_cobrador          — mesa con 1 hoja → solo_un_cobrador
    2. test_ciclo1_una_mesa_dos_hojas_coinciden  — 2 hojas mismo monto → confirmado
    3. test_ciclo1_una_mesa_dos_hojas_difieren   — 2 hojas distinto monto → discrepancia
    4. test_ciclo1_tres_hojas_mayoria            — 2/3 coinciden → mayoria_aplicada (auto)
    5. test_ciclo1_dos_mesas_mismo_predio        — predio en 2 mesas → pago_multi_mesa
    6. test_ciclo1_solo_cobradores_en_trazab     — solo_un_cobrador acumula en trazabilidad

  Ciclo 2 (supervisor resolvió):
    7. test_ciclo2_resolucion_discrepancia_acepta_a   — acepta_a → cobro entra a pagos
    8. test_ciclo2_resolucion_discrepancia_corrige    — corrige + MONTO_CORRECTO
    9. test_ciclo2_multi_mesa_si                       — ok=si → cobro entra
   10. test_ciclo2_multi_mesa_rechaza                  — ok=rechaza → solo trazabilidad
   11. test_ciclo2_multi_mesa_si_corrige_predio        — ok=si + MZ_CORRECTO → imputa al corregido
   12. test_ciclo2_pendiente_persiste                  — discrepancia sin RESOLUCION queda en .xlsx

  Preservación y acumulación:
   13. test_preserva_trabajo_manual_supervisor   — re-corrida sin tocar RESOLUCION conserva valor
   14. test_trazabilidad_acumula_entre_ciclos    — ciclo 2 no borra resoluciones del 1
   15. test_solo_un_cobrador_preserva_ciclo1     — solo_un_cobrador del ciclo 1 sobrevive en ciclo 2

  Re-incorporación de resoluciones previas:
   16. test_multi_mesa_resuelta_no_vuelve        — un multi_mesa resuelto en ciclo N no reaparece en N+1
"""

import logging
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

THIS = Path(__file__).resolve()
sys.path.insert(0, str(THIS.parent.parent))

import main as efectivo  # noqa: E402  (módulo se llama main.py; alias por legibilidad)

TEST_ROOT = THIS.parent / "_tmp_integracion"


# ── Helpers ──────────────────────────────────────────────────────────────────

def _cerrar_logging():
    """
    main.py registra un FileHandler sobre OUTPUTS_DIR/run.log; mientras esté abierto
    Windows bloquea el archivo y rmtree falla. Cerramos handlers entre tests.
    """
    for h in list(logging.root.handlers):
        try:
            h.close()
        except Exception:
            pass
        logging.root.removeHandler(h)


def _setup():
    _cerrar_logging()
    if TEST_ROOT.exists():
        shutil.rmtree(TEST_ROOT)
    (TEST_ROOT / "inputs").mkdir(parents=True)
    (TEST_ROOT / "outputs").mkdir(parents=True)
    (TEST_ROOT / "trazabilidad").mkdir(parents=True)
    # Redirigir rutas del módulo a las temporales
    efectivo.BASE_DIR    = TEST_ROOT
    efectivo.INPUTS_DIR  = TEST_ROOT / "inputs"
    efectivo.OUTPUTS_DIR = TEST_ROOT / "outputs"
    efectivo.TRAZAB_DIR  = TEST_ROOT / "trazabilidad"
    efectivo.DISC_FILE   = efectivo.OUTPUTS_DIR / "discrepancias.xlsx"
    efectivo.OUTPUT_FILE = efectivo.OUTPUTS_DIR / "pagos_efectivo.xlsx"


def _teardown():
    _cerrar_logging()
    if TEST_ROOT.exists():
        shutil.rmtree(TEST_ROOT)


_COLS_MESA = ["COBRADOR", "FECHA_REGISTRO", "MZ", "LT", "MONTO",
              "MONTO_EFECTIVO", "MONTO_YAPE", "FECHA", "COMENTARIO"]


def _crear_mesa(n: int, hojas: dict):
    """
    Crea inputs/mesa_N.xlsx con la estructura real:
    fila 1 = secciones, fila 2 = columnas, fila 3 = ejemplo (ignorada), fila 4+ = datos.

    hojas: {"registro_1": [(cobrador, fecha_reg, mz, lt, monto, m_efec, m_yape, fecha, com), ...]}
    """
    path = efectivo.INPUTS_DIR / f"mesa_{n}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    for nombre_hoja, filas in hojas.items():
        ws = wb.create_sheet(nombre_hoja)
        for ci, col in enumerate(_COLS_MESA, start=1):
            ws.cell(row=1, column=ci, value="")
            ws.cell(row=2, column=ci, value=col)
        ws.cell(row=3, column=1, value="(ejemplo guía)")
        for ri, fila in enumerate(filas, start=4):
            for ci, val in enumerate(fila, start=1):
                ws.cell(row=ri, column=ci, value=val)
    wb.save(path)


def _fila(mz, lt, monto, cobrador="Wilder", fecha="03/06/2026", comentario=""):
    """Fila de mesa con todo en efectivo (caso más común)."""
    return (cobrador, "10/06/2026", mz, lt, monto, monto, 0, fecha, comentario)


def _leer_pagos() -> list:
    """Lee filas de pagos_efectivo.xlsx (fila 1=secciones, fila 2=cols, fila 3+=data)."""
    p = efectivo.OUTPUT_FILE
    if not p.exists():
        return []
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 3:
        return []
    headers = [str(h).strip().upper() if h else "" for h in rows[1]]
    out = []
    for r in rows[2:]:
        if not r or all(c is None for c in r):
            continue
        out.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    return out


def _leer_disc(hoja: str = "discrepancias") -> list:
    p = efectivo.DISC_FILE
    if not p.exists():
        return []
    wb = load_workbook(p, read_only=True, data_only=True)
    if hoja not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[hoja]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 3:
        return []
    headers = [str(h).strip().upper() if h else "" for h in rows[1]]
    out = []
    for r in rows[2:]:
        if not r or all(c is None for c in r):
            continue
        out.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    return out


def _leer_traz(hoja: str, mes: str = "2026-06") -> list:
    p = efectivo.TRAZAB_DIR / f"trazabilidad_{mes}.xlsx"
    if not p.exists():
        return []
    return efectivo._leer_hoja_traz(p, hoja)


def _setear_resolucion(hoja: str, mz: str, lt: str, mesa: str, **campos):
    """
    Re-escribe discrepancias.xlsx llenando los campos de RESOLUCION/MONTO_CORRECTO/OK/etc.
    en la fila que coincide con (mz, lt, mesa). Simula al supervisor editando el archivo.
    """
    p = efectivo.DISC_FILE
    wb = load_workbook(p)
    ws = wb[hoja]
    headers = [c.value for c in ws[2]]
    for row in ws.iter_rows(min_row=3):
        d = {headers[i]: row[i].value for i in range(len(headers))}
        if (str(d.get("MZ", "")).strip().upper() == mz.upper()
            and str(d.get("LT", "")).strip().upper() == lt.upper()
            and str(d.get("MESA", "")).strip() == mesa):
            for campo, valor in campos.items():
                idx = headers.index(campo)
                row[idx].value = valor
            break
    wb.save(p)


# ── 1-6. Ciclo 1 ─────────────────────────────────────────────────────────────

def test_ciclo1_una_mesa_un_cobrador():
    _crear_mesa(1, {"registro_1": [_fila("A", "8C", 38.0)]})
    efectivo.main()

    pagos = _leer_pagos()
    assert len(pagos) == 1, f"Esperaba 1 pago, llegaron {len(pagos)}"
    assert pagos[0]["MZ"] == "A" and pagos[0]["LT"] == "8C"
    assert pagos[0]["ESTADO"] == "solo_un_cobrador"
    assert int(pagos[0]["CICLO_CORRECCION"]) == 1
    assert not efectivo.DISC_FILE.exists(), "Sin discrepancias no debería existir el archivo"
    print("  OK test_ciclo1_una_mesa_un_cobrador")


def test_ciclo1_una_mesa_dos_hojas_coinciden():
    _crear_mesa(1, {
        "registro_1": [_fila("A", "8C", 38.0, cobrador="Wilder")],
        "registro_2": [_fila("A", "8C", 38.0, cobrador="María")],
    })
    efectivo.main()

    pagos = _leer_pagos()
    assert len(pagos) == 1 and pagos[0]["ESTADO"] == "confirmado", \
        f"Esperaba 1 confirmado, llegó: {pagos}"
    assert float(pagos[0]["MONTO"]) == 38.0
    assert not efectivo.DISC_FILE.exists()
    print("  OK test_ciclo1_una_mesa_dos_hojas_coinciden")


def test_ciclo1_una_mesa_dos_hojas_difieren():
    _crear_mesa(1, {
        "registro_1": [_fila("A", "8C", 38.0, cobrador="Wilder")],
        "registro_2": [_fila("A", "8C", 40.0, cobrador="María")],
    })
    efectivo.main()

    pagos = _leer_pagos()
    assert len(pagos) == 0, f"No debería haber pagos confirmados: {pagos}"
    assert efectivo.DISC_FILE.exists(), "Debería existir discrepancias.xlsx"

    disc = _leer_disc("discrepancias")
    assert len(disc) == 1, f"Esperaba 1 discrepancia, llegaron: {disc}"
    assert disc[0]["MZ"] == "A" and disc[0]["LT"] == "8C"
    assert disc[0]["COBRADOR_A"] == "Wilder" and float(disc[0]["MONTO_A"]) == 38.0
    assert disc[0]["COBRADOR_B"] == "María"  and float(disc[0]["MONTO_B"]) == 40.0
    print("  OK test_ciclo1_una_mesa_dos_hojas_difieren")


def test_ciclo1_tres_hojas_mayoria():
    """2 hojas dicen 38, 1 dice 40 → mayoría automática → cobro entra como mayoria_aplicada."""
    _crear_mesa(1, {
        "registro_1": [_fila("A", "8C", 38.0, cobrador="Wilder")],
        "registro_2": [_fila("A", "8C", 38.0, cobrador="María")],
        "registro_3": [_fila("A", "8C", 40.0, cobrador="Ana")],
    })
    efectivo.main()

    pagos = _leer_pagos()
    assert len(pagos) == 1 and pagos[0]["ESTADO"] == "mayoria_aplicada"
    assert float(pagos[0]["MONTO"]) == 38.0, "Monto debería ser el de la mayoría (38)"
    assert not efectivo.DISC_FILE.exists(), "Mayoría auto resuelve, no debería haber discrepancias"

    # La resolución mayoría_x_y va a trazabilidad ciclo 1
    traz = _leer_traz("discrepancias_resueltas")
    assert len(traz) == 1 and "mayoria_" in traz[0]["RESOLUCION"], \
        f"Esperaba RESOLUCION mayoria_*, llegó: {traz[0]['RESOLUCION']!r}"
    print("  OK test_ciclo1_tres_hojas_mayoria")


def test_ciclo1_dos_mesas_mismo_predio():
    """Mismo predio en 2 mesas distintas → pago_multi_mesa (no entra a pagos)."""
    _crear_mesa(1, {"registro_1": [_fila("A", "8C", 38.0, cobrador="Wilder")]})
    _crear_mesa(2, {"registro_1": [_fila("A", "8C", 40.0, cobrador="María")]})
    efectivo.main()

    pagos = _leer_pagos()
    assert len(pagos) == 0, f"Multi-mesa pendiente no debería entrar a pagos: {pagos}"

    multi = _leer_disc("pago_multi_mesa")
    assert len(multi) == 2, f"Esperaba 2 filas (1 por aparición), llegaron: {multi}"
    mesas = {m["MESA"] for m in multi}
    assert mesas == {"mesa_1", "mesa_2"}
    print("  OK test_ciclo1_dos_mesas_mismo_predio")


def test_ciclo1_solo_cobradores_en_trazab():
    """solo_un_cobrador se acumula en hoja solo_un_cobrador de trazabilidad."""
    _crear_mesa(1, {"registro_1": [_fila("A", "8C", 38.0), _fila("B", "12", 16.0)]})
    efectivo.main()

    traz_sc = _leer_traz("solo_un_cobrador")
    assert len(traz_sc) == 2, f"Esperaba 2 filas en solo_un_cobrador, llegaron: {len(traz_sc)}"
    mzs = {r["MZ"] for r in traz_sc}
    assert mzs == {"A", "B"}
    print("  OK test_ciclo1_solo_cobradores_en_trazab")


# ── 7-12. Ciclo 2 ────────────────────────────────────────────────────────────

def test_ciclo2_resolucion_discrepancia_acepta_a():
    """
    Supervisor pone RESOLUCION=acepta_a → monto de A entra a pagos.
    Sembramos mesa_2 con solo_un_cobrador para que trazabilidad arranque en ciclo 1
    y detectar_ciclo retorne 2 en la segunda corrida (semántica realista).
    """
    _crear_mesa(1, {
        "registro_1": [_fila("A", "8C", 38.0, cobrador="Wilder")],
        "registro_2": [_fila("A", "8C", 40.0, cobrador="María")],
    })
    _crear_mesa(2, {"registro_1": [_fila("Z", "1", 22.0)]})  # seed para ciclo counter
    efectivo.main()
    assert efectivo.DISC_FILE.exists()

    # Supervisor edita discrepancias.xlsx
    _setear_resolucion("discrepancias", "A", "8C", "mesa_1", RESOLUCION="acepta_a")

    # Ciclo 2: aplica la resolución
    efectivo.main()
    pagos = _leer_pagos()
    # Esperamos 2: Z-1 (solo_un_cobrador, ciclo 1) + A-8C (discrepancia_resuelta, ciclo 2)
    assert len(pagos) == 2, f"Esperaba 2 pagos, llegaron: {pagos}"

    a8c = next(p for p in pagos if p["MZ"] == "A")
    assert float(a8c["MONTO"]) == 38.0, "MONTO debería ser el de Wilder (acepta_a)"
    assert a8c["ESTADO"] == "discrepancia_resuelta"
    assert int(a8c["CICLO_CORRECCION"]) == 2, \
        f"A-8C resuelto en ciclo 2, llegó CICLO_CORRECCION={a8c['CICLO_CORRECCION']}"

    z1 = next(p for p in pagos if p["MZ"] == "Z")
    assert int(z1["CICLO_CORRECCION"]) == 1, "Z-1 fue ciclo 1, debe preservarse"

    assert not efectivo.DISC_FILE.exists(), "Sin pendientes el archivo se borra"

    traz = _leer_traz("discrepancias_resueltas")
    assert len(traz) == 1 and traz[0]["RESOLUCION"] == "acepta_a"
    print("  OK test_ciclo2_resolucion_discrepancia_acepta_a")


def test_ciclo2_resolucion_discrepancia_corrige():
    """Supervisor pone RESOLUCION=corrige + MONTO_CORRECTO=39 → entra con 39."""
    _crear_mesa(1, {
        "registro_1": [_fila("A", "8C", 38.0, cobrador="Wilder")],
        "registro_2": [_fila("A", "8C", 40.0, cobrador="María")],
    })
    efectivo.main()

    _setear_resolucion("discrepancias", "A", "8C", "mesa_1",
                       RESOLUCION="corrige", MONTO_CORRECTO=39.0)

    efectivo.main()
    pagos = _leer_pagos()
    assert len(pagos) == 1 and float(pagos[0]["MONTO"]) == 39.0, \
        f"Esperaba MONTO=39 (corregido), llegó: {pagos}"
    print("  OK test_ciclo2_resolucion_discrepancia_corrige")


def test_ciclo2_multi_mesa_si():
    """OK=si en pago_multi_mesa → el cobro entra a pagos como multi_mesa_resuelta."""
    _crear_mesa(1, {"registro_1": [_fila("A", "8C", 38.0, cobrador="Wilder")]})
    _crear_mesa(2, {"registro_1": [_fila("A", "8C", 40.0, cobrador="María")]})
    efectivo.main()

    # Supervisor acepta solo el de mesa_1
    _setear_resolucion("pago_multi_mesa", "A", "8C", "mesa_1", OK="si")
    _setear_resolucion("pago_multi_mesa", "A", "8C", "mesa_2", OK="rechaza")

    efectivo.main()
    pagos = _leer_pagos()
    assert len(pagos) == 1, f"Esperaba 1 pago (solo mesa_1), llegaron: {pagos}"
    assert pagos[0]["MESA"] == "mesa_1"
    assert pagos[0]["ESTADO"] == "multi_mesa_resuelta"
    assert float(pagos[0]["MONTO"]) == 38.0

    # Trazabilidad guarda ambos (uno con si, otro con rechaza)
    traz = _leer_traz("pago_multi_mesa_resueltas")
    assert len(traz) == 2, f"Esperaba 2 resoluciones en trazabilidad, llegaron: {traz}"
    res = {r["MESA"]: r["RESOLUCION"] for r in traz}
    assert res == {"mesa_1": "si", "mesa_2": "rechaza"}, f"Resoluciones: {res}"
    print("  OK test_ciclo2_multi_mesa_si")


def test_ciclo2_multi_mesa_rechaza():
    """Todas las filas multi_mesa rechazadas → ningún cobro entra a pagos."""
    _crear_mesa(1, {"registro_1": [_fila("A", "8C", 38.0)]})
    _crear_mesa(2, {"registro_1": [_fila("A", "8C", 40.0)]})
    efectivo.main()

    _setear_resolucion("pago_multi_mesa", "A", "8C", "mesa_1", OK="rechaza")
    _setear_resolucion("pago_multi_mesa", "A", "8C", "mesa_2", OK="rechaza")

    efectivo.main()
    pagos = _leer_pagos()
    assert len(pagos) == 0, f"Todo rechazado, no debería haber pagos: {pagos}"

    traz = _leer_traz("pago_multi_mesa_resueltas")
    assert len(traz) == 2 and all(r["RESOLUCION"] == "rechaza" for r in traz), \
        "Las 2 rechazadas deben quedar en trazabilidad como auditoría"

    assert not efectivo.DISC_FILE.exists(), "Sin pendientes el archivo se borra"
    print("  OK test_ciclo2_multi_mesa_rechaza")


def test_ciclo2_multi_mesa_si_corrige_predio():
    """OK=si + MZ_CORRECTO/LT_CORRECTO → cobro se imputa al predio corregido."""
    _crear_mesa(1, {"registro_1": [_fila("A", "8C", 38.0)]})
    _crear_mesa(2, {"registro_1": [_fila("A", "8C", 38.0)]})
    efectivo.main()

    # Mesa_1: predio era B-12 (registró mal), aceptar con corrección
    _setear_resolucion("pago_multi_mesa", "A", "8C", "mesa_1",
                       OK="si", MZ_CORRECTO="B", LT_CORRECTO="12")
    _setear_resolucion("pago_multi_mesa", "A", "8C", "mesa_2", OK="rechaza")

    efectivo.main()
    pagos = _leer_pagos()
    assert len(pagos) == 1
    assert pagos[0]["MZ"] == "B" and pagos[0]["LT"] == "12", \
        f"Predio debería ser el corregido (B-12), llegó: {pagos[0]['MZ']}-{pagos[0]['LT']}"

    # Trazabilidad guarda original e imputado
    traz = _leer_traz("pago_multi_mesa_resueltas")
    fila_si = next(r for r in traz if r["RESOLUCION"] == "si")
    assert fila_si["MZ_ORIGINAL"] == "A" and fila_si["LT_ORIGINAL"] == "8C"
    assert fila_si["MZ"]          == "B" and fila_si["LT"]          == "12"
    print("  OK test_ciclo2_multi_mesa_si_corrige_predio")


def test_ciclo2_pendiente_persiste():
    """Una discrepancia sin RESOLUCION sigue en discrepancias.xlsx al re-correr."""
    _crear_mesa(1, {
        "registro_1": [_fila("A", "8C", 38.0)],
        "registro_2": [_fila("A", "8C", 40.0)],
    })
    efectivo.main()
    # Ciclo 2 sin tocar nada
    efectivo.main()
    assert efectivo.DISC_FILE.exists(), "Sin resolución, el archivo debe persistir"
    disc = _leer_disc("discrepancias")
    assert len(disc) == 1
    print("  OK test_ciclo2_pendiente_persiste")


# ── 13-15. Preservación y acumulación ────────────────────────────────────────

def test_preserva_trabajo_manual_supervisor():
    """Re-corrida regenera discrepancias.xlsx pero conserva el RESOLUCION del supervisor."""
    _crear_mesa(1, {
        "registro_1": [_fila("A", "8C", 38.0)],
        "registro_2": [_fila("A", "8C", 40.0)],
        # Discrepancia adicional para que la otra siga pendiente y el archivo no se borre
    })
    _crear_mesa(2, {
        "registro_1": [_fila("B", "12", 16.0)],
        "registro_2": [_fila("B", "12", 18.0)],
    })
    efectivo.main()
    assert len(_leer_disc("discrepancias")) == 2

    # Supervisor solo resuelve A-8C, deja B-12 pendiente
    _setear_resolucion("discrepancias", "A", "8C", "mesa_1", RESOLUCION="acepta_a")
    # Ciclo 2
    efectivo.main()

    pagos = _leer_pagos()
    assert len(pagos) == 1 and pagos[0]["MZ"] == "A"

    # B-12 sigue pendiente; A-8C ya no debería estar en discrepancias.xlsx
    disc = _leer_disc("discrepancias")
    assert len(disc) == 1 and disc[0]["MZ"] == "B", \
        f"Solo B-12 debería seguir pendiente: {disc}"
    print("  OK test_preserva_trabajo_manual_supervisor")


def test_trazabilidad_acumula_entre_ciclos():
    """
    Una resolución del ciclo 2 no borra mayorías automáticas del ciclo 1.

    Nota: cross_check_mesa es stateless, así que la mayoría auto se RE-detecta cada
    ciclo. Trazabilidad termina con la mayoría duplicada (1 del ciclo 1 preservada
    + 1 re-detectada en ciclo 2) más la nueva resolución del supervisor.
    """
    # Ciclo 1: una mayoría automática + una discrepancia pendiente
    _crear_mesa(1, {
        "registro_1": [_fila("A", "8C", 38.0)],
        "registro_2": [_fila("A", "8C", 38.0)],
        "registro_3": [_fila("A", "8C", 40.0)],  # → mayoria_aplicada
    })
    _crear_mesa(2, {
        "registro_1": [_fila("B", "12", 16.0)],
        "registro_2": [_fila("B", "12", 18.0)],  # → discrepancia pendiente
    })
    efectivo.main()
    traz_c1 = _leer_traz("discrepancias_resueltas")
    assert len(traz_c1) == 1 and "mayoria_" in traz_c1[0]["RESOLUCION"]

    # Ciclo 2: supervisor resuelve B-12
    _setear_resolucion("discrepancias", "B", "12", "mesa_2", RESOLUCION="acepta_a")
    efectivo.main()

    traz_c2 = _leer_traz("discrepancias_resueltas")
    # Contenido: la A de ciclo 1 sigue + la B resuelta en ciclo 2 (mínimo)
    mzs = [r["MZ"] for r in traz_c2]
    assert "A" in mzs, f"La mayoría auto de A del ciclo 1 debe persistir: {mzs}"
    assert "B" in mzs, f"La resolución acepta_a de B debe haberse agregado: {mzs}"

    b_row = next(r for r in traz_c2 if r["MZ"] == "B")
    assert b_row["RESOLUCION"] == "acepta_a"
    assert int(b_row["CICLO_CORRECCION"]) == 2
    print("  OK test_trazabilidad_acumula_entre_ciclos")


def test_solo_un_cobrador_preserva_ciclo1():
    """solo_un_cobrador se escribe en ciclo 1 y se preserva en ciclo 2+ aunque cambie input."""
    _crear_mesa(1, {"registro_1": [_fila("A", "8C", 38.0)]})
    _crear_mesa(2, {
        "registro_1": [_fila("B", "12", 16.0)],
        "registro_2": [_fila("B", "12", 18.0)],  # discrepancia para forzar ciclo 2
    })
    efectivo.main()
    sc1 = _leer_traz("solo_un_cobrador")
    assert len(sc1) == 1 and sc1[0]["MZ"] == "A"

    # Supervisor resuelve B-12
    _setear_resolucion("discrepancias", "B", "12", "mesa_2", RESOLUCION="acepta_a")
    efectivo.main()

    sc2 = _leer_traz("solo_un_cobrador")
    assert len(sc2) == 1 and sc2[0]["MZ"] == "A", \
        f"solo_un_cobrador del ciclo 1 debe persistir, llegó: {sc2}"
    print("  OK test_solo_un_cobrador_preserva_ciclo1")


# ── 16. Re-incorporación de resoluciones previas ─────────────────────────────

def test_multi_mesa_resuelta_no_vuelve():
    """
    Un multi_mesa resuelto en ciclo N no debe re-aparecer como pendiente en ciclo N+1
    aunque los inputs sigan teniendo el mismo pago.
    """
    _crear_mesa(1, {"registro_1": [_fila("A", "8C", 38.0)]})
    _crear_mesa(2, {"registro_1": [_fila("A", "8C", 38.0)]})
    efectivo.main()  # ciclo 1: ambos en pago_multi_mesa

    _setear_resolucion("pago_multi_mesa", "A", "8C", "mesa_1", OK="si")
    _setear_resolucion("pago_multi_mesa", "A", "8C", "mesa_2", OK="rechaza")
    efectivo.main()  # ciclo 2: resoluciones aplicadas

    pagos_c2 = _leer_pagos()
    assert len(pagos_c2) == 1 and pagos_c2[0]["MESA"] == "mesa_1"

    # Ciclo 3: SIN tocar nada — el sistema debe recordar que A-8C ya está resuelto
    efectivo.main()
    pagos_c3 = _leer_pagos()
    assert len(pagos_c3) == 1 and pagos_c3[0]["MESA"] == "mesa_1", \
        f"El pago resuelto en ciclo 2 debe persistir, llegaron: {pagos_c3}"
    assert not efectivo.DISC_FILE.exists(), \
        "No deberían reaparecer multi_mesa pendientes; el archivo debe seguir borrado"
    print("  OK test_multi_mesa_resuelta_no_vuelve")


# ── Runner ───────────────────────────────────────────────────────────────────

def main():
    tests = [
        # Ciclo 1
        test_ciclo1_una_mesa_un_cobrador,
        test_ciclo1_una_mesa_dos_hojas_coinciden,
        test_ciclo1_una_mesa_dos_hojas_difieren,
        test_ciclo1_tres_hojas_mayoria,
        test_ciclo1_dos_mesas_mismo_predio,
        test_ciclo1_solo_cobradores_en_trazab,
        # Ciclo 2
        test_ciclo2_resolucion_discrepancia_acepta_a,
        test_ciclo2_resolucion_discrepancia_corrige,
        test_ciclo2_multi_mesa_si,
        test_ciclo2_multi_mesa_rechaza,
        test_ciclo2_multi_mesa_si_corrige_predio,
        test_ciclo2_pendiente_persiste,
        # Preservación
        test_preserva_trabajo_manual_supervisor,
        test_trazabilidad_acumula_entre_ciclos,
        test_solo_un_cobrador_preserva_ciclo1,
        # Re-incorporación
        test_multi_mesa_resuelta_no_vuelve,
    ]
    for t in tests:
        _setup()
        try:
            t()
        except Exception:
            _teardown()
            raise
    _teardown()
    print(f"\n[OK] {len(tests)}/{len(tests)} tests de integración pasaron")


if __name__ == "__main__":
    main()
