# 4_pagos/efectivo/entregas_repo.py — escritor único de entregas.xlsx (append-only)
#
# Log de las entregas de la tesorera: efectivo + yape que cada cobrador entregó,
# por día. Mismo patrón que shared/seguimiento_repo.py.
#
# Contrato visual:    docs/formato_entregas.html
# Decisión de diseño: docs/decisiones/arqueo_efectivo.md
#
# API PÚBLICA:
#   registrar_entrega(fecha, cobrador, efectivo, yape, *, source, audit_ref, motivo="") -> dict
#   leer_entregas()          -> list[dict]  (una por fila, con fecha_key normalizada)
#   suma_entregas(mes=None)  -> dict {(fecha_key, cobrador_upper): {"efectivo","yape","cobrador"}}
#
# INVARIANTES:
#   - Único escritor de entregas.xlsx.
#   - Append-only: una fila nunca se modifica ni se borra.
#   - Idempotencia por (SOURCE, AUDIT_REF) — mismo evento no duplica.
#   - El arqueo SUMA todas las filas de un mismo (FECHA, COBRADOR): una segunda
#     entrega del día suma; una corrección es otra fila con el delta (MOTIVO).

import logging
import os
import time
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

BASE_DIR = Path(__file__).parent
ENTREGAS_PATH = BASE_DIR / "inputs" / "entregas.xlsx"
SHEET = "Entregas"

# (header_bg, header_fg, data_bg) — contrato: docs/formato_entregas.html
_AZ = ("EBF5FB", "1A5276", "F4FAFF")   # ¿quién y cuándo?
_VE = ("E9F7EF", "1E5C3A", "F4FBF7")   # ¿cuánto recibió?
_RO = ("FCE4EC", "880E4F", "FDF2F8")   # ¿corrección?
_GR = ("F3F4F6", "374151", "F9FAFB")   # auditoría

# (nombre, seccion_color, ancho, align)
_COLS = [
    ("FECHA",     _AZ, 14, "center"),
    ("COBRADOR",  _AZ, 22, "left"),
    ("EFECTIVO",  _VE, 12, "right"),
    ("YAPE",      _VE, 12, "right"),
    ("MOTIVO",    _RO, 30, "left"),
    ("SOURCE",    _GR, 24, "center"),
    ("AUDIT_REF", _GR, 28, "left"),
    ("TIMESTAMP", _GR, 20, "center"),
]
# (label, col_ini, col_fin, seccion_color)
_SECCIONES = [
    ("¿Quién y cuándo?", "FECHA",    "COBRADOR",  _AZ),
    ("¿Cuánto recibió?", "EFECTIVO", "YAPE",      _VE),
    ("¿Corrección?",     "MOTIVO",   "MOTIVO",    _RO),
    ("Auditoría",        "SOURCE",   "TIMESTAMP", _GR),
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def _argb(hex6: str) -> str:
    return "FF" + hex6.lstrip("#")


def _fill(hex6):
    return PatternFill("solid", fgColor=_argb(hex6))


def _save_atomic(wb, path: Path) -> None:
    """Escribe a un temporal en el mismo directorio y reemplaza con os.replace
    (atómico). Reintenta ante PermissionError transitorio de Windows (antivirus/
    indexador). Mismo patrón que seguimiento_repo._save_atomic."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.stem}.tmp{path.suffix}")
    wb.save(tmp)
    ultimo = None
    for intento in range(5):
        try:
            os.replace(tmp, path)
            return
        except PermissionError as e:
            ultimo = e
            time.sleep(0.3 * (intento + 1))
    raise ultimo


def _monto(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return round(float(str(v).replace("S/", "").replace(",", "").strip()), 2)
    except (ValueError, TypeError):
        return 0.0


def _norm_cobrador(v) -> str:
    return " ".join(str(v or "").split())


def _parse_fecha(v) -> datetime:
    """Acepta datetime, date o str (dd/mm/aaaa · aaaa-mm-dd · dd-mm-aaaa)."""
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    s = str(v or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt)
        except ValueError:
            continue
    raise ValueError(f"FECHA no parseable: {v!r} (usar dd/mm/aaaa)")


def _fecha_key(v) -> str:
    """Clave canónica de cruce 'dd/mm/aaaa'. '' si no se puede parsear."""
    try:
        return _parse_fecha(v).strftime("%d/%m/%Y")
    except ValueError:
        return ""


def _mes_de(fecha_key: str) -> str:
    try:
        d = datetime.strptime(fecha_key, "%d/%m/%Y")
        return f"{d.year}-{d.month:02d}"
    except ValueError:
        return ""

# ── Escritura de cabeceras ────────────────────────────────────────────────────

def _hdr(cell, bg, fg, texto):
    cell.value = texto
    cell.fill = _fill(bg)
    cell.font = Font(color=_argb(fg), bold=True, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_headers(ws) -> None:
    col_idx = {c[0]: i + 1 for i, c in enumerate(_COLS)}
    for label, ini, fin, sec in _SECCIONES:
        c1, c2 = col_idx[ini], col_idx[fin]
        if c1 != c2:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        _hdr(ws.cell(row=1, column=c1), sec[0], sec[1], label)
    ws.row_dimensions[1].height = 18
    for i, (nombre, sec, ancho, _a) in enumerate(_COLS, start=1):
        _hdr(ws.cell(row=2, column=i), sec[0], sec[1], nombre)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

# ── Lectura ───────────────────────────────────────────────────────────────────

def leer_entregas() -> list:
    """Todas las filas como lista de dicts. [] si el archivo no existe."""
    if not ENTREGAS_PATH.exists():
        return []
    wb = load_workbook(ENTREGAS_PATH, read_only=True, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    filas = list(ws.values)
    wb.close()
    if len(filas) < 3:
        return []
    headers = [str(h).strip().upper() if h else "" for h in filas[1]]
    idx = {h: i for i, h in enumerate(headers)}

    def g(fila, col):
        i = idx.get(col)
        return fila[i] if i is not None and i < len(fila) else None

    out = []
    for fila in filas[2:]:
        if not fila or all(c is None for c in fila):
            continue
        cobrador = _norm_cobrador(g(fila, "COBRADOR"))
        fecha_key = _fecha_key(g(fila, "FECHA"))
        if not cobrador or not fecha_key:
            continue
        out.append({
            "fecha_key": fecha_key,
            "cobrador":  cobrador,
            "efectivo":  _monto(g(fila, "EFECTIVO")),
            "yape":      _monto(g(fila, "YAPE")),
            "motivo":    str(g(fila, "MOTIVO") or "").strip(),
            "source":    str(g(fila, "SOURCE") or "").strip(),
            "audit_ref": str(g(fila, "AUDIT_REF") or "").strip(),
        })
    return out


def suma_entregas(mes: str | None = None) -> dict:
    """{(fecha_key, cobrador_upper): {'efectivo','yape','cobrador'}} sumando todas
    las filas de cada (FECHA, COBRADOR). Es el 'RECIBIDO' del arqueo."""
    out: dict = {}
    for r in leer_entregas():
        if mes and _mes_de(r["fecha_key"]) != mes:
            continue
        key = (r["fecha_key"], r["cobrador"].upper())
        d = out.setdefault(key, {"efectivo": 0.0, "yape": 0.0, "cobrador": r["cobrador"]})
        d["efectivo"] = round(d["efectivo"] + r["efectivo"], 2)
        d["yape"] = round(d["yape"] + r["yape"], 2)
    return out


def _ya_registrado(source: str, audit_ref: str) -> bool:
    for r in leer_entregas():
        if r["source"] == source and r["audit_ref"] == audit_ref:
            return True
    return False

# ── Escritura (única API que muta el archivo) ─────────────────────────────────

def registrar_entrega(fecha, cobrador, efectivo, yape, *, source: str,
                      audit_ref: str, motivo: str = "") -> dict:
    """Appendea una fila de entrega. NUNCA edita ni borra filas previas.
    Idempotente por (source, audit_ref). Una corrección se registra como otra
    llamada con el delta (efectivo/yape puede ser negativo) y un motivo."""
    if not source:
        raise ValueError("source no puede ser vacío")
    if not audit_ref:
        raise ValueError("audit_ref no puede ser vacío")
    cobrador = _norm_cobrador(cobrador)
    if not cobrador:
        raise ValueError("cobrador no puede ser vacío")
    fecha_dt = _parse_fecha(fecha)
    efectivo = _monto(efectivo)
    yape = _monto(yape)

    if _ya_registrado(source, audit_ref):
        log.info(f"entregas_repo: skip (idempotente) — {source}/{audit_ref}")
        return {"skipped": True}

    if ENTREGAS_PATH.exists():
        wb = load_workbook(ENTREGAS_PATH)
        ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
        next_row = max(ws.max_row + 1, 3)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET
        _write_headers(ws)
        next_row = 3

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    valores = {
        "FECHA": fecha_dt, "COBRADOR": cobrador, "EFECTIVO": efectivo, "YAPE": yape,
        "MOTIVO": motivo, "SOURCE": source, "AUDIT_REF": audit_ref, "TIMESTAMP": ts,
    }
    for ci, (nombre, sec, _ancho, align) in enumerate(_COLS, start=1):
        cell = ws.cell(row=next_row, column=ci, value=valores[nombre])
        cell.fill = _fill(sec[2])
        cell.font = Font(color=_argb(sec[1]), size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if nombre == "FECHA":
            cell.number_format = "DD/MM/YYYY"
        elif nombre in ("EFECTIVO", "YAPE"):
            cell.number_format = "#,##0.00"

    _save_atomic(wb, ENTREGAS_PATH)
    log.info(f"entregas_repo: entrega {fecha_dt:%d/%m/%Y} {cobrador} "
             f"efec={efectivo} yape={yape} [{audit_ref}]")
    return {"skipped": False}
