# migrar_formato_v3.py — agrega la columna CATEGORIA (col 11) a mesa_N.xlsx
#
# Schema v2 → v3: nueva sección "¿Qué pasó?" con columna CATEGORIA (dropdown
# reclamo/compromiso/otros, vacío = pago normal). Los datos existentes no se
# tocan; la columna nueva queda vacía (el cobrador decide — Protocolo M.6).
#
# Contrato visual: docs/formato_registro.html
# Protocolo de Migración (docs/metodologia_desarrollo.md):
#   M.1 backup previo en backup/migracion_2026-07/ (hecho antes de correr esto)
#   M.4 idempotente — si la hoja ya tiene CATEGORIA, se salta
#   M.7 probar en una: py migrar_formato_v3.py mesa_1  ·  todas: py migrar_formato_v3.py
#
# Uso:
#   py migrar_formato_v3.py           # migra mesa_1 … mesa_7
#   py migrar_formato_v3.py mesa_1    # migra solo mesa_1

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent.parent / "shared"))
from utils_templates import agregar_dropdown_categoria  # noqa: E402  (módulo puro, sin efectos al importar)

INPUTS_DIR = Path(__file__).parent / "inputs"

COL_CATEGORIA = 11          # después de CONCEPTO (col 10)
GRUPO_TXT  = "¿Qué pasó?"
BG, TXT    = "FCE4EC", "880E4F"
ANCHO      = 14
HOJAS      = ["registro_1", "registro_2", "registro_3"]


def _borde() -> Border:
    s = Side(style="thin", color="FFFFFF")
    return Border(left=s, right=s, top=s, bottom=s)


def _celda_header(cell, texto, size):
    cell.value     = texto
    cell.font      = Font(name="Arial", bold=True, size=size, color=TXT)
    cell.fill      = PatternFill("solid", start_color=BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _borde()


def migrar_hoja(ws) -> bool:
    """Agrega CATEGORIA a una hoja. Devuelve False si ya estaba migrada."""
    headers = [c.value for c in ws[2]]
    if "CATEGORIA" in headers:
        return False
    _celda_header(ws.cell(row=1, column=COL_CATEGORIA), GRUPO_TXT, size=9)
    _celda_header(ws.cell(row=2, column=COL_CATEGORIA), "CATEGORIA", size=10)
    ws.column_dimensions[get_column_letter(COL_CATEGORIA)].width = ANCHO
    # Datos existentes: la columna nueva queda vacía (M.6 — el cobrador llena)
    agregar_dropdown_categoria(ws)
    return True


def migrar_mesa(ruta: Path) -> None:
    wb = load_workbook(ruta)   # sin read_only: hay que escribir preservando estilos
    tocadas = []
    for hoja in HOJAS:
        if hoja not in wb.sheetnames:
            continue
        if migrar_hoja(wb[hoja]):
            tocadas.append(hoja)
    if tocadas:
        try:
            wb.save(ruta)
        except PermissionError:
            # Archivo abierto en Excel — no se migra; volver a correr tras cerrarlo
            print(f"  !! {ruta.name} BLOQUEADO (abierto en Excel) → cerrarlo y re-correr")
            return
        print(f"  OK {ruta.name}  ({', '.join(tocadas)})")
    else:
        print(f"  {ruta.name} ya v3 → skip")


def main():
    # Windows: forzar UTF-8 en stdout (mismo patrón que main.py)
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    solo = sys.argv[1] if len(sys.argv) > 1 else None
    nombres = [solo] if solo else [f"mesa_{n}" for n in range(1, 8)]
    print("Migrando mesas a schema v3 (+CATEGORIA)...")
    for nombre in nombres:
        ruta = INPUTS_DIR / f"{nombre}.xlsx"
        if not ruta.exists():
            print(f"  {nombre}.xlsx no existe → skip")
            continue
        migrar_mesa(ruta)
    print("Listo.")


if __name__ == "__main__":
    main()
