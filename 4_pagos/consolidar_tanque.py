# 4_pagos/consolidar_tanque.py — ledger canal-agnóstico de aportes al tanque comunitario
# Cosecha CONCEPTO=tanque desde las vistas operacionales de yape y efectivo.
# Writer único de outputs/aportes_tanque.xlsx (DE10).
# Contrato visual: docs/formato_aportes_tanque.html

import logging
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR          = Path(__file__).parent

# Los outputs de pagos llevan el ciclo en el nombre desde 2026-08; para los
# ciclos anteriores se acepta el nombre pelado (ver shared/ciclo.py).
sys.path.insert(0, str(BASE_DIR.parent / "shared"))
import ciclo as ciclo_activo  # noqa: E402

_MES_CICLO = ciclo_activo.activo(default=None,
                                 path=BASE_DIR.parent / "shared" / "ciclo_activo.json")


def _pago_path(carpeta: Path, base: str) -> Path:
    if _MES_CICLO is None:
        return carpeta / f"{base}.xlsx"
    return ciclo_activo.resolver(carpeta, base, _MES_CICLO,
                                 legacy_sin_periodo=ciclo_activo.acepta_legacy(_MES_CICLO))


YAPE_TEPAGO_PATH  = _pago_path(BASE_DIR / "yape" / "motor_matching" / "outputs", "pagos_yape_tepago")
EFECTIVO_PATH     = _pago_path(BASE_DIR / "efectivo" / "outputs", "pagos_efectivo")
OUTPUTS_DIR       = BASE_DIR / "outputs"
OUTPUT_FILE       = OUTPUTS_DIR / "aportes_tanque.xlsx"

# Aportes al tanque mal ubicados/desfasados, corregidos a mano — writer único
# humano (no lo escribe ningún main.py). Precursor de caja.MovimientoCaja
# BALDE=tanque (reasignado). Ver docs/decisiones y backfill_ledger/README.md.
APORTES_TANQUE_MANUALES_PATH = BASE_DIR.parent / "shared" / "aportes_tanque_manuales.xlsx"

log = logging.getLogger(__name__)

# ── Paleta — columnas cuentan una historia ───────────────────────────────────
GRUPOS = {
    "id":     ("EBF5FB", "1A5276", "F4FAFF"),  # ¿QUIÉN APORTÓ? — USER_ID, NOMBRE
    "origen": ("FEF9E7", "7D6608", "FFFDF5"),  # ¿DE DÓNDE VINO? — CANAL, REFERENCIA
    "ubic":   ("E9F7EF", "1E5C3A", "F4FBF7"),  # ¿DÓNDE VIVE? — MZ, LOTE
    "monto":  ("F3E8FF", "5B21B6", "FAF5FF"),  # ¿CUÁNTO Y CUÁNDO? — MONTO_PAGO, FECHA
    "tipo":   ("FFF7ED", "9A3412", "FFFBF7"),  # ¿PARA QUÉ? — CONCEPTO
}

COLUMNAS = [
    ("USER_ID",     "id"),
    ("NOMBRE",      "id"),
    ("CANAL",       "origen"),
    ("REFERENCIA",  "origen"),
    ("MZ",          "ubic"),
    ("LOTE",        "ubic"),
    ("MONTO_PAGO",  "monto"),
    ("FECHA",       "monto"),
    ("CONCEPTO",    "tipo"),
]

GRUPOS_TITULO = [
    ("¿QUIÉN APORTÓ?",     2),
    ("¿DE DÓNDE VINO?",    2),
    ("¿DÓNDE VIVE?",       2),
    ("¿CUÁNTO Y CUÁNDO?",  2),
    ("¿PARA QUÉ?",         1),
]


def _norm(v) -> str:
    return str(v).strip() if v is not None else ""


def _monto(v) -> float:
    try:
        return round(float(str(v).replace(",", ".").strip()), 2)
    except (TypeError, ValueError):
        return 0.0


def _leer_tanque_yape() -> list:
    """CONCEPTO=tanque desde pagos_yape_tepago.xlsx. FECHA se copia verbatim
    (string tal cual la fuente) — NO reparsear, evita repetir el bug de
    inversión día/mes de B2 (trazabilidad_cobranza)."""
    if not YAPE_TEPAGO_PATH.exists():
        return []
    wb = load_workbook(YAPE_TEPAGO_PATH, read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.values)
    wb.close()
    if len(filas) < 3:
        return []

    headers = [str(h).strip().upper() if h else "" for h in filas[1]]
    registros = []
    for fila in filas[2:]:
        if not fila or all(c is None for c in fila):
            continue
        row = {headers[i]: fila[i] for i in range(min(len(headers), len(fila)))}
        if _norm(row.get("CONCEPTO")).lower() != "tanque":
            continue
        origen = _norm(row.get("ORIGEN"))
        monto  = _monto(row.get("MONTO_PAGO"))
        registros.append({
            "user_id":    row.get("USER_ID"),
            "nombre":     row.get("NOMBRE"),
            "canal":      "yape",
            "referencia": f"{origen}-{monto:g}",
            "fecha":      _norm(row.get("FECHA")),
            "mz":         _norm(row.get("MZ")),
            "lote":       _norm(row.get("LOTE")),
            "monto":      monto,
        })
    return registros


def _leer_tanque_efectivo() -> list:
    """CONCEPTO=tanque desde pagos_efectivo.xlsx. USER_ID/NOMBRE quedan en
    blanco — efectivo no los rastrea (solo MZ/LT). Pendiente futuro:
    unificar USER_ID entre yape y efectivo (ver pendientes_plan.md)."""
    if not EFECTIVO_PATH.exists():
        return []
    wb = load_workbook(EFECTIVO_PATH, read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.values)
    wb.close()
    if len(filas) < 3:
        return []

    headers = [str(h).strip().upper() if h else "" for h in filas[1]]
    registros = []
    for fila in filas[2:]:
        if not fila or all(c is None for c in fila):
            continue
        row = {headers[i]: fila[i] for i in range(min(len(headers), len(fila)))}
        if _norm(row.get("CONCEPTO")).lower() != "tanque":
            continue
        mesa      = _norm(row.get("MESA"))
        cobrador  = _norm(row.get("COBRADOR"))
        monto     = _monto(row.get("MONTO"))
        fecha_raw = row.get("FECHA")
        fecha     = fecha_raw.strftime("%d/%m/%Y") if hasattr(fecha_raw, "strftime") else _norm(fecha_raw)
        registros.append({
            "user_id":    None,
            "nombre":     None,
            "canal":      "efectivo",
            "referencia": f"{mesa}-{cobrador}",
            "fecha":      fecha,
            "mz":         _norm(row.get("MZ")),
            "lote":       _norm(row.get("LT")),
            "monto":      monto,
        })
    return registros


def _mes_ciclo_actual() -> str:
    """Deriva el mes-ciclo vigente desde pagos_efectivo.xlsx (moda de FECHA,
    YYYY-MM) — este script no recibe --mes, así que no hay otra señal.
    Sirve solo para filtrar aportes_tanque_manuales.xlsx por MES_ANO_APLICA
    (se incorpora una sola vez, no en cada corrida futura)."""
    if not EFECTIVO_PATH.exists():
        return ""
    wb = load_workbook(EFECTIVO_PATH, read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.values)
    wb.close()
    if len(filas) < 3:
        return ""
    headers = [str(h).strip().upper() if h else "" for h in filas[1]]
    if "FECHA" not in headers:
        return ""
    idx = headers.index("FECHA")
    from collections import Counter
    meses = Counter()
    for fila in filas[2:]:
        if not fila or idx >= len(fila) or fila[idx] is None:
            continue
        v = fila[idx]
        s = v.strftime("%Y-%m") if hasattr(v, "strftime") else str(v)[:7]
        if len(s) == 7 and s[4] == "-":
            meses[s] += 1
    return meses.most_common(1)[0][0] if meses else ""


def _leer_tanque_manuales(mes_ano: str) -> list:
    """Aportes al tanque mal ubicados/desfasados corregidos a mano (ver
    shared/aportes_tanque_manuales.xlsx). Filtra por MES_ANO_APLICA: se
    incorpora solo en el ciclo al que se decidió aplicarlo, no en todos los
    ciclos futuros — evita doble-conteo en corridas siguientes."""
    if not APORTES_TANQUE_MANUALES_PATH.exists():
        return []
    wb = load_workbook(APORTES_TANQUE_MANUALES_PATH, read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.values)
    wb.close()
    if len(filas) < 3:
        return []

    headers = [str(h).strip().upper() if h else "" for h in filas[1]]
    registros = []
    for fila in filas[2:]:
        if not fila or all(c is None for c in fila):
            continue
        row = {headers[i]: fila[i] for i in range(min(len(headers), len(fila)))}
        if _norm(row.get("MES_ANO_APLICA")) != mes_ano:
            continue
        monto = _monto(row.get("MONTO"))
        fecha_raw = row.get("FECHA_REAL")
        fecha = fecha_raw.strftime("%d/%m/%Y") if hasattr(fecha_raw, "strftime") else _norm(fecha_raw)
        mz_o, lt_o = _norm(row.get("MZ_ORIGEN")), _norm(row.get("LT_ORIGEN"))
        registros.append({
            "user_id":    None,
            "nombre":     None,
            "canal":      _norm(row.get("CANAL")) or "efectivo",
            "referencia": f"reasignado de {mz_o}-{lt_o}" if mz_o else "manual",
            "fecha":      fecha,
            "mz":         _norm(row.get("MZ")),
            "lote":       _norm(row.get("LT")),
            "monto":      monto,
        })
    return registros


def _borde(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(cell, bg, txt, texto):
    cell.value     = texto
    cell.font      = Font(name="Arial", bold=True, size=9, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _borde("FFFFFF")


def _dat(cell, valor, bg, txt, align="left", fmt=None):
    cell.value     = valor
    cell.font      = Font(name="Arial", size=10, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _borde()
    if fmt:
        cell.number_format = fmt


def exportar(registros: list):
    OUTPUTS_DIR.mkdir(exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Aportes_tanque"

    # Fila 1 — grupos (la historia), Fila 2 — columnas reales
    col = 1
    for titulo, span in GRUPOS_TITULO:
        clave = COLUMNAS[col - 1][1]
        bg, txt, _ = GRUPOS[clave]
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
        _hdr(ws.cell(row=1, column=col), bg, txt, titulo)
        col += span

    for i, (nombre, clave) in enumerate(COLUMNAS, start=1):
        bg, txt, _ = GRUPOS[clave]
        _hdr(ws.cell(row=2, column=i), bg, txt, nombre)

    for r, reg in enumerate(registros, start=3):
        _, _, bg_id     = GRUPOS["id"]
        _, _, bg_origen = GRUPOS["origen"]
        _, _, bg_ubic   = GRUPOS["ubic"]
        _, _, bg_monto  = GRUPOS["monto"]
        _, _, bg_tipo   = GRUPOS["tipo"]
        _dat(ws.cell(row=r, column=1), reg["user_id"],    bg_id,     "1A5276")
        _dat(ws.cell(row=r, column=2), reg["nombre"],     bg_id,     "1A5276")
        _dat(ws.cell(row=r, column=3), reg["canal"],      bg_origen, "7D6608", align="center")
        _dat(ws.cell(row=r, column=4), reg["referencia"], bg_origen, "7D6608")
        _dat(ws.cell(row=r, column=5), reg["mz"],         bg_ubic,   "1E5C3A", align="center")
        _dat(ws.cell(row=r, column=6), reg["lote"],       bg_ubic,   "1E5C3A", align="center")
        _dat(ws.cell(row=r, column=7), reg["monto"],      bg_monto,  "5B21B6", align="right", fmt='"S/" #,##0.00')
        _dat(ws.cell(row=r, column=8), reg["fecha"],      bg_monto,  "5B21B6", align="center")
        _dat(ws.cell(row=r, column=9), "tanque",          bg_tipo,   "9A3412", align="center")

    anchos = [8, 26, 10, 20, 6, 6, 12, 18, 10]
    for i, a in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = a
    ws.freeze_panes = "A3"

    wb.save(OUTPUT_FILE)


def main():
    mes_ano = _mes_ciclo_actual()
    registros = _leer_tanque_yape() + _leer_tanque_efectivo() + _leer_tanque_manuales(mes_ano)
    exportar(registros)
    total = round(sum(r["monto"] for r in registros), 2)
    log.info(f"aportes_tanque.xlsx: {len(registros)} filas · S/ {total:.2f}")
    print(f"  aportes_tanque.xlsx generado: {len(registros)} filas · S/ {total:.2f}")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    main()
