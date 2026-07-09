# 4_pagos/consolidar_deuda_directiva.py — ledger append-only de deudas de la directiva
# Cosecha CONCEPTO=deuda_directiva desde las vistas de yape y efectivo del mes y
# hace APPEND (con dedup) a shared/deuda_directiva.xlsx. Writer único del ledger.
#
# A diferencia de consolidar_tanque.py (que REGENERA desde el mes actual y pierde
# los meses viejos al cerrar el ciclo — el "Gap" del tanque), este ledger es
# append-only: sobrevive el cierre. Por eso lleva columna CICLO y vive en shared/.
# Contrato visual: docs/formato_deuda_directiva.html

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR          = Path(__file__).parent
YAPE_TEPAGO_PATH  = BASE_DIR / "yape" / "motor_matching" / "outputs" / "pagos_yape_tepago.xlsx"
EFECTIVO_PATH     = BASE_DIR / "efectivo" / "outputs" / "pagos_efectivo.xlsx"
USUARIOS_ID_PATH  = BASE_DIR.parent / "shared" / "usuarios_id.xlsx"
OUTPUT_FILE       = BASE_DIR.parent / "shared" / "deuda_directiva.xlsx"

CONCEPTO_DEUDA    = "deuda_directiva"   # token canónico (no texto libre)

log = logging.getLogger(__name__)

GRUPOS = {
    "id":     ("EBF5FB", "1A5276", "F4FAFF"),  # ¿QUIÉN DEBÍA? — USER_ID, NOMBRE
    "origen": ("FEF9E7", "7D6608", "FFFDF5"),  # ¿DE DÓNDE VINO? — CANAL, REFERENCIA
    "monto":  ("F3E8FF", "5B21B6", "FAF5FF"),  # ¿CUÁNTO Y CUÁNDO? — MONTO_PAGO, FECHA
    "periodo":("E9F7EF", "1E5C3A", "F4FBF7"),  # ¿PERÍODO? — CICLO
    "tipo":   ("FFF7ED", "9A3412", "FFFBF7"),  # ¿PARA QUÉ? — CONCEPTO
}

COLUMNAS = [
    ("USER_ID",    "id"),    ("NOMBRE",     "id"),
    ("CANAL",      "origen"),("REFERENCIA", "origen"),
    ("MONTO_PAGO", "monto"), ("FECHA",      "monto"),
    ("CICLO",      "periodo"),
    ("CONCEPTO",   "tipo"),
]

GRUPOS_TITULO = [
    ("¿QUIÉN DEBÍA?",     2),
    ("¿DE DÓNDE VINO?",   2),
    ("¿CUÁNTO Y CUÁNDO?", 2),
    ("¿PERÍODO?",         1),
    ("¿PARA QUÉ?",        1),
]


def _norm(v) -> str:
    return str(v).strip() if v is not None else ""


def _monto(v) -> float:
    try:
        return round(float(str(v).replace(",", ".").strip()), 2)
    except (TypeError, ValueError):
        return 0.0


def _ciclo(fecha: str) -> str:
    """YYYY-MM del pago, derivado de la FECHA. Prueba los formatos que producen
    yape (ISO o dd/mm) y efectivo (dd/mm/yyyy). Si no parsea, retorna ''."""
    s = _norm(fecha)
    if not s:
        return ""
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m")
        except ValueError:
            continue
    return ""


def _cargar_usuarios_id() -> dict:
    """{NOMBRE_normalizado: USER_ID} para completar USER_ID cuando la fuente no lo trae."""
    if not USUARIOS_ID_PATH.exists():
        return {}
    wb = load_workbook(USUARIOS_ID_PATH, read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.values)
    wb.close()
    if not filas:
        return {}
    headers = [str(h).strip().upper() if h else "" for h in filas[0]]
    try:
        i_id  = headers.index("USER_ID")
        i_nom = headers.index("NOMBRE")
    except ValueError:
        return {}
    mapa = {}
    for fila in filas[1:]:
        if not fila or i_nom >= len(fila):
            continue
        nom = _norm(fila[i_nom]).upper()
        uid = _norm(fila[i_id]) if i_id < len(fila) else ""
        if nom and uid:
            mapa.setdefault(nom, uid)
    return mapa


def _leer_yape(usuarios: dict) -> list:
    if not YAPE_TEPAGO_PATH.exists():
        return []
    wb = load_workbook(YAPE_TEPAGO_PATH, read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.values)
    wb.close()
    if len(filas) < 3:
        return []
    headers = [str(h).strip().upper() if h else "" for h in filas[1]]
    registros = []
    for fila in filas[2:]:
        if not fila or all(c is None for c in fila):
            continue
        row = {headers[i]: fila[i] for i in range(min(len(headers), len(fila)))}
        if _norm(row.get("CONCEPTO")).lower() != CONCEPTO_DEUDA:
            continue
        origen = _norm(row.get("ORIGEN"))
        monto  = _monto(row.get("MONTO_PAGO"))
        nombre = _norm(row.get("NOMBRE"))
        uid    = _norm(row.get("USER_ID")) or usuarios.get(nombre.upper(), "")
        fecha  = _norm(row.get("FECHA"))
        registros.append({
            "user_id": uid, "nombre": nombre,
            "canal": "yape", "referencia": f"{origen}-{monto:g}",
            "monto": monto, "fecha": fecha, "ciclo": _ciclo(fecha),
        })
    return registros


def _leer_efectivo(usuarios: dict) -> list:
    if not EFECTIVO_PATH.exists():
        return []
    wb = load_workbook(EFECTIVO_PATH, read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.values)
    wb.close()
    if len(filas) < 3:
        return []
    headers = [str(h).strip().upper() if h else "" for h in filas[1]]
    registros = []
    for fila in filas[2:]:
        if not fila or all(c is None for c in fila):
            continue
        row = {headers[i]: fila[i] for i in range(min(len(headers), len(fila)))}
        if _norm(row.get("CONCEPTO")).lower() != CONCEPTO_DEUDA:
            continue
        mesa      = _norm(row.get("MESA"))
        cobrador  = _norm(row.get("COBRADOR"))
        monto     = _monto(row.get("MONTO"))
        nombre    = _norm(row.get("NOMBRE"))
        fecha_raw = row.get("FECHA")
        fecha     = fecha_raw.strftime("%d/%m/%Y") if hasattr(fecha_raw, "strftime") else _norm(fecha_raw)
        registros.append({
            "user_id": usuarios.get(nombre.upper(), ""), "nombre": nombre,
            "canal": "efectivo", "referencia": f"{mesa}-{cobrador}",
            "monto": monto, "fecha": fecha, "ciclo": _ciclo(fecha),
        })
    return registros


def _clave(reg: dict) -> tuple:
    return (reg["canal"], reg["referencia"], reg["monto"], reg["fecha"])


def _leer_ledger_existente() -> list:
    """Filas ya guardadas en el ledger (para el append idempotente)."""
    if not OUTPUT_FILE.exists():
        return []
    wb = load_workbook(OUTPUT_FILE, read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.values)
    wb.close()
    if len(filas) < 3:
        return []
    headers = [str(h).strip().upper() if h else "" for h in filas[1]]
    registros = []
    for fila in filas[2:]:
        if not fila or all(c is None for c in fila):
            continue
        row = {headers[i]: fila[i] for i in range(min(len(headers), len(fila)))}
        registros.append({
            "user_id": _norm(row.get("USER_ID")), "nombre": _norm(row.get("NOMBRE")),
            "canal": _norm(row.get("CANAL")), "referencia": _norm(row.get("REFERENCIA")),
            "monto": _monto(row.get("MONTO_PAGO")), "fecha": _norm(row.get("FECHA")),
            "ciclo": _norm(row.get("CICLO")),
        })
    return registros


def _borde(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(cell, bg, txt, texto):
    cell.value     = texto
    cell.font      = Font(name="Arial", bold=True, size=9, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _borde("FFFFFF")


def _dat(cell, valor, bg, txt, align="left", fmt=None):
    cell.value     = valor
    cell.font      = Font(name="Arial", size=10, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _borde()
    if fmt:
        cell.number_format = fmt


def exportar(registros: list):
    OUTPUT_FILE.parent.mkdir(exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Deuda_directiva"

    col = 1
    for titulo, span in GRUPOS_TITULO:
        clave = COLUMNAS[col - 1][1]
        bg, txt, _ = GRUPOS[clave]
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
        _hdr(ws.cell(row=1, column=col), bg, txt, titulo)
        col += span

    for i, (nombre, clave) in enumerate(COLUMNAS, start=1):
        bg, txt, _ = GRUPOS[clave]
        _hdr(ws.cell(row=2, column=i), bg, txt, nombre)

    for r, reg in enumerate(registros, start=3):
        _, _, bg_id      = GRUPOS["id"]
        _, _, bg_origen  = GRUPOS["origen"]
        _, _, bg_monto   = GRUPOS["monto"]
        _, _, bg_periodo = GRUPOS["periodo"]
        _, _, bg_tipo    = GRUPOS["tipo"]
        _dat(ws.cell(row=r, column=1), reg["user_id"],    bg_id,      "1A5276", align="center")
        _dat(ws.cell(row=r, column=2), reg["nombre"],     bg_id,      "1A5276")
        _dat(ws.cell(row=r, column=3), reg["canal"],      bg_origen,  "7D6608", align="center")
        _dat(ws.cell(row=r, column=4), reg["referencia"], bg_origen,  "7D6608")
        _dat(ws.cell(row=r, column=5), reg["monto"],      bg_monto,   "5B21B6", align="right", fmt='"S/" #,##0.00')
        _dat(ws.cell(row=r, column=6), reg["fecha"],      bg_monto,   "5B21B6", align="center")
        _dat(ws.cell(row=r, column=7), reg["ciclo"],      bg_periodo, "1E5C3A", align="center")
        _dat(ws.cell(row=r, column=8), CONCEPTO_DEUDA,    bg_tipo,    "9A3412", align="center")

    anchos = [8, 26, 10, 22, 12, 20, 10, 16]
    for i, a in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = a
    ws.freeze_panes = "A3"
    wb.save(OUTPUT_FILE)


def main():
    usuarios = _cargar_usuarios_id()
    nuevos   = _leer_yape(usuarios) + _leer_efectivo(usuarios)
    previos  = _leer_ledger_existente()

    vistos = {_clave(r) for r in previos}
    agregados = [r for r in nuevos if _clave(r) not in vistos]
    registros = previos + agregados

    exportar(registros)
    total = round(sum(r["monto"] for r in registros), 2)
    log.info(f"deuda_directiva.xlsx: {len(registros)} filas (+{len(agregados)} nuevas) · S/ {total:.2f}")
    print(f"  deuda_directiva.xlsx: {len(registros)} filas (+{len(agregados)} nuevas) · S/ {total:.2f}")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    main()
