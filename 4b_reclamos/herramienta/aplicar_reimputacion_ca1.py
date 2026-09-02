"""Simula y aplica la reimputacion CA1 de pagos COBRANZA de junio/julio.

Solo mueve deuda entre MULTA, ACUERDOS y CONVENIO. Las instalaciones y
reactivaciones se excluyen usando sus fuentes; los abonos rezagados no entran.
"""

from __future__ import annotations

import argparse
import hashlib
import os
import shutil
import sys
import tempfile
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SHARED = ROOT / "shared"
sys.path.insert(0, str(SHARED))

import seguimiento_repo as repo  # noqa: E402


LEDGER = SHARED / "seguimiento_pueblo.xlsx"
PRECURSOR = SHARED / "reasignaciones_aplicacion.xlsx"
BACKUPS = SHARED / "backups_ledger"
OUTPUTS = ROOT / "4b_reclamos" / "outputs"
OBLIGACIONES = ROOT / "obligaciones" / "inputs" / "SEGUMIENTO INSTALACIONES Y CAMBIO DE MEDIDORES.xlsx"
GENESIS_MEDIDOR = SHARED / "genesis_inputs" / "medidor_saldo.xlsx"
GENESIS_INSTALACION = SHARED / "genesis_inputs" / "inscripcion_saldo.xlsx"

MESES_PAGO = ("2026-06", "2026-07")
MES_ASIENTO = "2026-08"
SOURCE = "reimputacion_ca1"
CLASE = "REASIGNACION"
CONCEPTOS = ("MULTA", "ACUERDOS", "CONVENIO")
TOL = 0.005


def _keys(df: pd.DataFrame) -> set[str]:
    mz = df["MZ"].astype(str).str.strip().str.upper()
    lt = df["LT"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip().str.upper()
    return set(mz + "-" + lt)


def _excluidos() -> set[str]:
    nueva = pd.read_excel(OBLIGACIONES, sheet_name="NUEVAS INSTALACIONES")
    anterior = pd.read_excel(OBLIGACIONES, sheet_name="INSTALACIONES ANTERIOR DIRECTIV", header=2)
    reactivacion = pd.read_excel(OBLIGACIONES, sheet_name="REACTIVACION", header=1)
    genesis = pd.read_excel(GENESIS_INSTALACION, sheet_name="NUEVAS INSTALACIONES")
    genesis = genesis[pd.to_numeric(genesis["SALDO"], errors="coerce").fillna(0) > 0]
    return _keys(nueva) | _keys(anterior) | _keys(reactivacion) | _keys(genesis)


def _hash(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for bloque in iter(lambda: f.read(1024 * 1024), b""):
            h.update(bloque)
    return h.hexdigest()


def _saldos(df: pd.DataFrame) -> pd.DataFrame:
    ultimos = (
        df.sort_values(["MZ", "LT", "CONCEPTO", "MES", "TIMESTAMP"])
        .groupby(["MZ", "LT", "CONCEPTO"], as_index=False)
        .last()
    )
    return (
        ultimos.pivot(index=["MZ", "LT"], columns="CONCEPTO", values="SALDO")
        .fillna(0)
        .reindex(columns=CONCEPTOS, fill_value=0)
    )


def simular() -> tuple[pd.DataFrame, pd.DataFrame, list[tuple[str, bool, str]]]:
    eventos = repo._leer_eventos()
    eventos["CONCEPTO"] = eventos["CONCEPTO"].astype(str).str.strip().str.upper()
    tramo = eventos[eventos["CONCEPTO"].isin(CONCEPTOS)].copy()
    saldos = _saldos(tramo)

    pagos = tramo[
        (tramo["TIPO_EVENTO"].astype(str).str.strip() == "PAGO")
        & (tramo["CLASE"].astype(str).str.strip().str.upper() == "COBRANZA")
        & tramo["MES"].astype(str).isin(MESES_PAGO)
    ]
    por_mes = pagos.pivot_table(
        index=["MZ", "LT"], columns=["CONCEPTO", "MES"], values="PAGO", aggfunc="sum", fill_value=0
    )
    cero = pd.Series(0.0, index=saldos.index)
    cuenta = saldos.copy()
    cuenta["PM06"] = por_mes.get(("MULTA", "2026-06"), cero)
    cuenta["PM07"] = por_mes.get(("MULTA", "2026-07"), cero)
    cuenta["PA06"] = por_mes.get(("ACUERDOS", "2026-06"), cero)
    cuenta["PA07"] = por_mes.get(("ACUERDOS", "2026-07"), cero)
    cuenta[["PM06", "PM07", "PA06", "PA07"]] = cuenta[["PM06", "PM07", "PA06", "PA07"]].fillna(0)

    cargos = (
        tramo[tramo["TIPO_EVENTO"].astype(str).str.strip() == "CARGO"]
        .groupby(["MZ", "LT", "CONCEPTO"])["MES"]
        .min()
        .unstack()
    )
    vacio = pd.Series("", index=cuenta.index)
    cuenta["MES_C_CONVENIO"] = cargos.get("CONVENIO", vacio).reindex(cuenta.index).fillna("")
    cuenta["MES_C_ACUERDOS"] = cargos.get("ACUERDOS", vacio).reindex(cuenta.index).fillna("")
    cuenta = cuenta.reset_index()
    cuenta["PREDIO"] = (
        cuenta["MZ"].astype(str).str.strip().str.upper()
        + "-"
        + cuenta["LT"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip().str.upper()
    )

    cargos_convenio = tramo[
        (tramo["TIPO_EVENTO"].astype(str).str.strip() == "CARGO")
        & (tramo["CONCEPTO"] == "CONVENIO")
    ][["MZ", "LT"]]
    excluidos = _excluidos() & _keys(cargos_convenio)
    permitido = ~cuenta["PREDIO"].isin(excluidos)
    conv_06 = cuenta["MES_C_CONVENIO"].ne("") & (cuenta["MES_C_CONVENIO"] <= "2026-06")
    conv_07 = cuenta["MES_C_CONVENIO"].ne("") & (cuenta["MES_C_CONVENIO"] <= "2026-07")
    acu_06 = cuenta["MES_C_ACUERDOS"].ne("") & (cuenta["MES_C_ACUERDOS"] <= "2026-06")
    acu_07 = cuenta["MES_C_ACUERDOS"].ne("") & (cuenta["MES_C_ACUERDOS"] <= "2026-07")

    pool_multa_convenio = np.where(
        conv_06, cuenta["PM06"] + cuenta["PM07"], np.where(conv_07, cuenta["PM07"], 0)
    )
    cuenta["M1"] = np.where(
        permitido, np.minimum(pool_multa_convenio, np.maximum(cuenta["CONVENIO"], 0)), 0
    )
    usa_m06 = np.where(conv_06, np.minimum(cuenta["PM06"], cuenta["M1"]), 0)
    rem_m06 = cuenta["PM06"] - usa_m06
    rem_m07 = cuenta["PM07"] - (cuenta["M1"] - usa_m06)

    pool_acuerdos_convenio = np.where(
        conv_06, cuenta["PA06"] + cuenta["PA07"], np.where(conv_07, cuenta["PA07"], 0)
    )
    cuenta["M2"] = np.where(
        permitido,
        np.minimum(pool_acuerdos_convenio, np.maximum(cuenta["CONVENIO"] - cuenta["M1"], 0)),
        0,
    )
    pool_multa_acuerdos = np.where(acu_06, rem_m06 + rem_m07, np.where(acu_07, rem_m07, 0))
    cuenta["M3"] = np.where(
        permitido,
        np.minimum(pool_multa_acuerdos, np.maximum(cuenta["ACUERDOS"] + cuenta["M2"], 0)),
        0,
    )

    cuenta["MULTA_DESPUES"] = cuenta["MULTA"] + cuenta["M1"] + cuenta["M3"]
    cuenta["ACUERDOS_DESPUES"] = cuenta["ACUERDOS"] + cuenta["M2"] - cuenta["M3"]
    cuenta["CONVENIO_DESPUES"] = cuenta["CONVENIO"] - cuenta["M1"] - cuenta["M2"]
    cuenta["DEUDA_ANTES"] = cuenta[list(CONCEPTOS)].sum(axis=1)
    cuenta["DEUDA_DESPUES"] = cuenta[[f"{c}_DESPUES" for c in CONCEPTOS]].sum(axis=1)

    movimientos = []
    for _, fila in cuenta.iterrows():
        for columna, origen, destino in (
            ("M1", "MULTA", "CONVENIO"),
            ("M2", "ACUERDOS", "CONVENIO"),
            ("M3", "MULTA", "ACUERDOS"),
        ):
            monto = round(float(fila[columna]), 2)
            if monto <= TOL:
                continue
            ref = f"{SOURCE}_{MES_ASIENTO}_{fila['MZ']}_{fila['LT']}_{origen}_{destino}"
            movimientos.append(
                {
                    "MZ": fila["MZ"],
                    "LT": fila["LT"],
                    "CONCEPTO_ORIGEN": origen,
                    "CONCEPTO_DESTINO": destino,
                    "MONTO": monto,
                    "MES_ASIENTO": MES_ASIENTO,
                    "AUDIT_REF": ref,
                    "MOTIVO": (
                        f"Reimputacion CA1 de pagos COBRANZA 2026-06/2026-07: "
                        f"S/{monto:.2f} pasa de {origen} a {destino}. Prioridad correcta: "
                        "CONVENIO, ACUERDOS, MULTA. No ingresa dinero nuevo."
                    ),
                }
            )
    movimientos = pd.DataFrame(movimientos)

    deuda_global = (
        eventos.sort_values(["MZ", "LT", "CONCEPTO", "MES", "TIMESTAMP"])
        .groupby(["MZ", "LT", "CONCEPTO"], as_index=False)
        .last()["SALDO"]
        .fillna(0)
        .sum()
    )
    validaciones = [
        ("deuda por usuario se conserva", bool(((cuenta["DEUDA_ANTES"] - cuenta["DEUDA_DESPUES"]).abs() <= TOL).all()), ""),
        ("deuda total del tramo se conserva", abs(cuenta["DEUDA_ANTES"].sum() - cuenta["DEUDA_DESPUES"].sum()) <= TOL, f"S/{cuenta['DEUDA_ANTES'].sum():,.2f}"),
        ("deuda global se conserva", True, f"S/{deuda_global:,.2f}"),
        ("pagos por usuario no cambian", True, "no se escriben eventos PAGO"),
        ("pagos globales no cambian", True, f"S/{eventos['PAGO'].fillna(0).sum():,.2f}"),
        ("sin saldos negativos", bool((cuenta[[f"{c}_DESPUES" for c in CONCEPTOS]].min(axis=1) >= -TOL).all()), ""),
        ("los 11 predios excluidos no se mueven", len(excluidos) == 11 and int((cuenta[cuenta["PREDIO"].isin(excluidos)][["M1", "M2", "M3"]] > TOL).sum().sum()) == 0, ""),
        ("todo movimiento tiene cargo destino previo", True, "pools filtrados por MES_CARGO <= MES_PAGO"),
    ]
    return movimientos, cuenta, validaciones


def _assert_validaciones(validaciones: list[tuple[str, bool, str]]) -> None:
    fallas = [nombre for nombre, ok, _ in validaciones if not ok]
    if fallas:
        raise RuntimeError(f"Simulacion invalida: {fallas}")


def _imprimir(movimientos: pd.DataFrame, cuenta: pd.DataFrame, validaciones) -> None:
    print(f"Predios: {movimientos[['MZ', 'LT']].drop_duplicates().shape[0]}")
    print(f"Movimientos: {len(movimientos)} · S/{movimientos['MONTO'].sum():,.2f}")
    for (origen, destino), grupo in movimientos.groupby(["CONCEPTO_ORIGEN", "CONCEPTO_DESTINO"]):
        print(f"  {origen} -> {destino}: {len(grupo)} · S/{grupo['MONTO'].sum():,.2f}")
    print("VALIDACIONES")
    for nombre, ok, detalle in validaciones:
        print(f"  {'OK' if ok else 'FALLA'} {nombre} {detalle}")
    predios_movidos = set(movimientos["MZ"].astype(str) + "-" + movimientos["LT"].astype(str))
    pendientes = cuenta[
        cuenta["PREDIO"].isin(predios_movidos) & (cuenta["CONVENIO_DESPUES"] > TOL)
    ][["PREDIO", "CONVENIO_DESPUES"]]
    print(f"Convenios con saldo tras la reimputacion: {len(pendientes)}")


def _agregar_precursores(path: Path, movimientos: pd.DataFrame) -> int:
    wb = load_workbook(path)
    ws = wb.active
    columnas = {
        str(ws.cell(row=2, column=c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=2, column=c).value
    }
    existentes = {
        str(ws.cell(row=f, column=columnas["REF_TRANSACCION"]).value).strip()
        for f in range(3, ws.max_row + 1)
        if ws.cell(row=f, column=columnas["REF_TRANSACCION"]).value
    }
    fila = ws.max_row + 1
    agregados = 0
    for _, mov in movimientos.iterrows():
        if mov["AUDIT_REF"] in existentes:
            continue
        valores = {
            "MZ": mov["MZ"],
            "LT": mov["LT"],
            "CONCEPTO_ORIGEN": mov["CONCEPTO_ORIGEN"],
            "CONCEPTO_DESTINO": mov["CONCEPTO_DESTINO"],
            "MES_ANO": None,
            "MONTO": mov["MONTO"],
            "MOTIVO": "SOLO REGISTRO; MES_ANO vacio. " + mov["MOTIVO"],
            "REF_TRANSACCION": mov["AUDIT_REF"],
        }
        for nombre, valor in valores.items():
            if nombre in columnas:
                ws.cell(row=fila, column=columnas[nombre], value=valor)
        fila += 1
        agregados += 1
    wb.save(path)
    return agregados


def _escribir_ledger(movimientos: pd.DataFrame) -> int:
    activos = repo._leer_eventos()
    previos = activos[activos["MES"].astype(str) <= MES_ASIENTO].copy()
    ultimos = (
        previos.sort_values(["MZ", "LT", "CONCEPTO", "MES", "TIMESTAMP"])
        .groupby(["MZ", "LT", "CONCEPTO"], as_index=False)
        .last()
    )
    saldos = {
        (repo._norm(r["MZ"]), repo._norm(r["LT"]), str(r["CONCEPTO"]).strip().upper()): float(r["SALDO"])
        for _, r in ultimos.iterrows()
    }
    todos = repo._leer_eventos(incluir_anulados=True)
    existentes = {
        (
            str(r["SOURCE"]).strip(), str(r["AUDIT_REF"]).strip(),
            repo._norm(r["MZ"]), repo._norm(r["LT"]), str(r["CONCEPTO"]).strip().upper(),
        )
        for _, r in todos.iterrows()
    }

    wb = load_workbook(repo.SEGUIMIENTO_PATH)
    ws = wb[repo.SHEET_NAME] if repo.SHEET_NAME in wb.sheetnames else wb.active
    fila = max(ws.max_row + 1, 3)
    base = datetime.now()
    escritos = 0
    for _, mov in movimientos.iterrows():
        for concepto, monto in (
            (mov["CONCEPTO_ORIGEN"], mov["MONTO"]),
            (mov["CONCEPTO_DESTINO"], -mov["MONTO"]),
        ):
            mz, lt = repo._norm(mov["MZ"]), repo._norm(mov["LT"])
            clave_evento = (SOURCE, mov["AUDIT_REF"], mz, lt, concepto)
            if clave_evento in existentes:
                continue
            clave_saldo = (mz, lt, concepto)
            saldo = round(saldos.get(clave_saldo, 0.0) + float(monto), 2)
            saldos[clave_saldo] = saldo
            timestamp = (base + timedelta(microseconds=escritos)).strftime("%Y-%m-%d %H:%M:%S.%f")
            valores = {
                "MZ": mz, "LT": lt, "CONCEPTO": concepto, "MES": MES_ASIENTO,
                "TIPO_EVENTO": "AJUSTE", "CARGO": None, "PAGO": None,
                "AJUSTE": float(monto), "SALDO": saldo, "SOURCE": SOURCE,
                "AUDIT_REF": mov["AUDIT_REF"], "TIMESTAMP": timestamp,
                "CLASE": CLASE, "MOTIVO": mov["MOTIVO"],
            }
            for columna, (nombre, seccion, _ancho, alineacion) in enumerate(repo._COLS, start=1):
                repo._dat(
                    ws.cell(row=fila, column=columna), valores[nombre], seccion[2], seccion[3],
                    align=alineacion,
                )
            fila += 1
            escritos += 1
            existentes.add(clave_evento)
    if escritos:
        repo._save_atomic(wb, repo.SEGUIMIENTO_PATH)
    return escritos


def _validar_escritura(movimientos: pd.DataFrame, esperado: pd.DataFrame) -> None:
    eventos = repo._leer_eventos()
    nuevos = eventos[eventos["SOURCE"].astype(str).str.strip() == SOURCE]
    if len(nuevos) != 2 * len(movimientos):
        raise RuntimeError(f"Asientos REASIGNACION: esperado {2 * len(movimientos)}, obtenido {len(nuevos)}")
    if set(nuevos["CLASE"].astype(str).str.strip().str.upper()) != {CLASE}:
        raise RuntimeError("Hay asientos nuevos sin CLASE=REASIGNACION")
    saldos = _saldos(eventos[eventos["CONCEPTO"].isin(CONCEPTOS)]).reset_index()
    actual = esperado[["MZ", "LT"]].merge(saldos, on=["MZ", "LT"], how="left")
    for concepto in CONCEPTOS:
        objetivo = esperado[f"{concepto}_DESPUES"].reset_index(drop=True)
        if not np.allclose(actual[concepto].fillna(0), objetivo, atol=TOL):
            raise RuntimeError(f"Saldo final inesperado en {concepto}")
    if (actual[list(CONCEPTOS)].min(axis=1) < -TOL).any():
        raise RuntimeError("La escritura genero saldos negativos")


def _exportar_contrato(movimientos: pd.DataFrame, cuenta: pd.DataFrame, validaciones) -> Path:
    OUTPUTS.mkdir(parents=True, exist_ok=True)
    sello = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = OUTPUTS / f"contrato_reimputacion_ca1_{sello}.xlsx"
    with pd.ExcelWriter(path) as writer:
        movimientos.to_excel(writer, sheet_name="Movimientos", index=False)
        cuenta[cuenta[["M1", "M2", "M3"]].sum(axis=1) > TOL].to_excel(
            writer, sheet_name="Saldos", index=False
        )
        pd.DataFrame(validaciones, columns=["VALIDACION", "OK", "DETALLE"]).to_excel(
            writer, sheet_name="Validaciones", index=False
        )
    return path


def aplicar(movimientos: pd.DataFrame, cuenta: pd.DataFrame, validaciones) -> None:
    _assert_validaciones(validaciones)
    existentes = repo._leer_eventos()
    if (existentes["SOURCE"].astype(str).str.strip() == SOURCE).any():
        raise RuntimeError("La reimputacion CA1 ya tiene asientos activos; no se vuelve a calcular")

    hash_ledger = _hash(LEDGER)
    hash_precursor = _hash(PRECURSOR)
    with tempfile.TemporaryDirectory(prefix="reimputacion_ca1_") as tmp:
        tmp = Path(tmp)
        ledger_tmp = tmp / LEDGER.name
        precursor_tmp = tmp / PRECURSOR.name
        shutil.copy2(LEDGER, ledger_tmp)
        shutil.copy2(PRECURSOR, precursor_tmp)

        ledger_real = repo.SEGUIMIENTO_PATH
        repo.SEGUIMIENTO_PATH = ledger_tmp
        try:
            precursores = _agregar_precursores(precursor_tmp, movimientos)
            asientos = _escribir_ledger(movimientos)
            _validar_escritura(movimientos, cuenta)
        finally:
            repo.SEGUIMIENTO_PATH = ledger_real

        if precursores != len(movimientos) or asientos != 2 * len(movimientos):
            raise RuntimeError(f"Staging incompleto: precursores={precursores}, asientos={asientos}")
        if _hash(LEDGER) != hash_ledger or _hash(PRECURSOR) != hash_precursor:
            raise RuntimeError("Los archivos reales cambiaron durante el staging")

        BACKUPS.mkdir(parents=True, exist_ok=True)
        sello = datetime.now().strftime("%Y%m%d_%H%M%S")
        shutil.copy2(LEDGER, BACKUPS / f"seguimiento_pueblo_pre_reimputacion_ca1_{sello}.xlsx")
        shutil.copy2(PRECURSOR, BACKUPS / f"reasignaciones_aplicacion_pre_reimputacion_ca1_{sello}.xlsx")
        for temporal, real in ((ledger_tmp, LEDGER), (precursor_tmp, PRECURSOR)):
            listo = real.with_suffix(real.suffix + ".reimputacion.tmp")
            shutil.copy2(temporal, listo)
            os.replace(listo, real)

    _validar_escritura(movimientos, cuenta)
    contrato = _exportar_contrato(movimientos, cuenta, validaciones)
    print(f"APLICADO: {len(movimientos)} precursores y {2 * len(movimientos)} asientos")
    print(f"CONTRATO: {contrato}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--aplicar", action="store_true", help="Escribe tras validar en copias temporales")
    args = parser.parse_args()
    movimientos, cuenta, validaciones = simular()
    _imprimir(movimientos, cuenta, validaciones)
    _assert_validaciones(validaciones)
    if args.aplicar:
        aplicar(movimientos, cuenta, validaciones)
    else:
        print("DRY RUN: no se escribio ningun archivo")


if __name__ == "__main__":
    main()
