"""
4b_reclamos/herramienta/verificar_yape.py — ¿el yape que el cobrador anotó entró
de verdad a la cuenta de la JASS?

EL PROBLEMA QUE RESUELVE

    El cobrador tiene una hoja de EFECTIVO (mesa_N.xlsx) con 3 columnas de
    monto: MONTO, MONTO_EFECTIVO y MONTO_YAPE. Cuando un vecino le dice "ya te
    yapeé", el cobrador lo anota ahí — pero esa fila no la levanta nadie:

        4_pagos/efectivo  solo procesa MONTO_EFECTIVO (correcto: el yape no es suyo)
        motor_matching    lee el reporte del BANCO, nunca las mesas
                          ──────────────────────────────────────────
                          la fila cae entre los dos modulos

    Y hay dos desenlaces muy distintos que hay que separar:

      el yape SI entro a la cuenta  ->  motor_matching lo levanto del banco por
                                        su cuenta y el predio cobro igual. La
                                        anotacion del cobrador es un duplicado
                                        inofensivo.
      el yape NO entro              ->  el vecino cree que pago, el sistema no
                                        lo sabe, y el mes siguiente le llega la
                                        deuda igual. Puede haberle yapeado al
                                        telefono personal de alguien.

    Medido sobre junio/julio/agosto 2026: S/469 en 13 filas del segundo tipo, y
    en 7 de ellas el propio vecino reclama diciendo que yapeo. Tres de esos
    vecinos NUNCA reclamaron — ninguna herramienta que arranque desde el reclamo
    los va a encontrar. Por eso este barrido va sobre TODO el pueblo.

QUE NO HACE

    No dice quien se quedo con la plata. "No esta en la cuenta" admite varias
    explicaciones (el vecino yapeo a otro numero, el cobrador no rindio, error
    de anotacion) y el reporte no elige entre ellas: la captura del yape del
    vecino es la que decide, porque ahi sale el numero de destino.

Uso:
    py verificar_yape.py                # ciclo activo
    py verificar_yape.py --mes 2026-07
    py verificar_yape.py --todos         # los 3 ciclos con archivo disponible
"""

import argparse
import re
import shutil
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent          # 4b_reclamos/
REPO_DIR = BASE_DIR.parent                       # raíz del repo activo
SHARED_DIR = REPO_DIR / "shared"

sys.path.insert(0, str(BASE_DIR))
sys.path.insert(0, str(SHARED_DIR))
sys.path.insert(0, str(REPO_DIR / "4_pagos" / "efectivo"))

import ciclo                            # noqa: E402
import reporte_historico as rh          # noqa: E402
import reporte_referencias_pago as rrp   # noqa: E402
import verificar_lotes as vl            # noqa: E402  (leer_boletas)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

TOL = 0.005
HOJAS_MESA = ("registro_1", "registro_2", "registro_3")

_PATRON_YAPE = re.compile(r"\byape|\byapeo|\byapeó|\bplin\b", re.IGNORECASE)

# Un MZ-LT escrito dentro de un mensaje de yape. Dos formas, porque la gente
# escribe de las dos maneras y cada una destapó un falso positivo real:
#   con palabra   "Mz H lote 21" · "mz D1 lt 1" · "MZ.C1 LT.2" · "MZ,B1,Lt.3"
#   pelada        "H1-16" · "K-3"
# El separador incluye la coma: sin ella, "MZ,B1,Lt.3-Johan Rodriguez" no
# matcheaba y el pago de B1-3 se ofrecía como candidato de F1-13.
_PATRON_LOTE_MSG = re.compile(
    r"(?:MZ\.?[,\s]*)?\b([A-Z]{1,2}[0-9]?)\b[-.,\s]*(?:LT|LOTE|LTE)\.?[,\s]*(\d+[A-Z]?)\b",
    re.IGNORECASE)

# La forma pelada MZ-LT, sin la palabra "lt": "H1-16" en un mensaje suelto.
# Va aparte porque es mucho más laxa y hay que exigirle el guión.
_PATRON_LOTE_PELADO = re.compile(r"\b([A-Z]{1,2}[0-9]?)-(\d+[A-Z]?)\b")


# ── Normalización (las usa también buscar_pago.py, que importa de acá) ───────

def _norm(v) -> str:
    s = str(v).strip().upper() if v is not None else ""
    return s[:-2] if s.endswith(".0") else s


def _clave(mz, lt) -> str:
    return f"{_norm(mz)}-{_norm(lt)}"


def _numf(v) -> float:
    """float tolerante: acá entra 16.0, '16.0', '' y NaN."""
    try:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return 0.0
        s = str(v).strip().replace(",", "")
        return float(s) if s not in ("", "nan", "None", "NaT") else 0.0
    except (TypeError, ValueError):
        return 0.0


def _txt(v) -> str:
    s = str(v).strip() if v is not None else ""
    return "" if s in ("nan", "None", "NaT") else s


def _sin_tildes(s) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", str(s))
                   if unicodedata.category(c) != "Mn")


# El nombre del cobrador se tipea a mano en cada fila y sale mal escrito: en
# julio hay 3 filas que dicen "Yreald Romero" por "Yerald Romero". Sin unificar,
# el resumen parte a la misma persona en dos y ninguno de los dos totales sirve.
_ALIAS_COBRADOR = {
    "YREALD ROMERO": "Yerald Romero",
    "WILDER TRUJILLO ROSALES": "Wilder Trujillo",
}


def _cobrador_canon(nombre: str) -> str:
    n = _txt(nombre)
    return _ALIAS_COBRADOR.get(_sin_tildes(n).upper(), n)


def _fecha(v) -> pd.Timestamp | None:
    for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S"):
        f = pd.to_datetime(_txt(v), format=fmt, errors="coerce")
        if not pd.isna(f):
            return f
    f = pd.to_datetime(_txt(v), errors="coerce", dayfirst=True)
    return None if pd.isna(f) else f


def _menciona_lote(mensaje: str, mz: str, lt: str) -> bool:
    """¿El texto nombra este lote? También la usa buscar_pago para el
    multi-lote (un pago que cubre 2 predios, caso K-3/K-4)."""
    if not mensaje:
        return False
    t = _sin_tildes(mensaje).upper().replace(" ", "")
    m, l = _norm(mz), _norm(lt)
    return any(p in t for p in (f"{m}-{l}", f"MZ{m}LT{l}", f"{m}LT{l}", f"MZ{m}{l}"))


def _extractor_motor():
    """`extraer_mz_lote_mensaje` de motor_matching: 31 patrones afinados contra
    mensajes reales de yape durante meses. Saca cosas que un regex propio no
    ("Maria Rosa Jimenez Roca   M x L 11" -> X-11).

    Se carga por ruta explícita y NO con `import main`: motor_matching/main.py y
    4b_reclamos/main.py se llaman igual, y 4b_reclamos ya está en sys.path — un
    import por nombre resolvería al equivocado según el orden."""
    global _cache_extractor
    if _cache_extractor is None:
        import importlib.util
        ruta = REPO_DIR / "4_pagos" / "yape" / "motor_matching" / "main.py"
        try:
            spec = importlib.util.spec_from_file_location("_mm_extractor", ruta)
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            _cache_extractor = mod.extraer_mz_lote_mensaje
        except Exception:
            _cache_extractor = lambda _m: (None, None)   # noqa: E731
    return _cache_extractor


_cache_extractor = None


def _lotes_en_mensaje(mensaje: str) -> set[tuple[str, str]]:
    """Todos los (MZ, LT) que un mensaje nombra, juntando las dos fuentes.

    motor_matching devuelve UNO (el primero) y mis patrones barren todos, que es
    lo que hace falta para "¿nombra algún lote que no sea el mío?"."""
    t = _sin_tildes(mensaje)
    fuera = set()
    mz_m, lt_m = _extractor_motor()(mensaje)
    if mz_m:
        fuera.add((_norm(mz_m), _norm(lt_m)))
    for patron in (_PATRON_LOTE_MSG, _PATRON_LOTE_PELADO):
        for m in patron.finditer(t):
            fuera.add((m.group(1).upper(), m.group(2).upper()))
    return fuera


def _nombra_otro_lote(mensaje: str, mz: str, lt: str) -> bool:
    """¿El mensaje nombra un lote DISTINTO del buscado?

    Si lo hace, ese pago es de ese otro lote — por más que el monto coincida.
    Casos reales: S/36 "Roman Lozano Mz H lote 21" se reportaba como el yape de
    H1-15, y S/30 "mz D1 lt 1" como el de P-12. Los montos se repiten muchísimo
    (S/8 lo deben 101 predios), así que sin este chequeo la coincidencia por
    monto inventa dueños."""
    if not mensaje or _menciona_lote(mensaje, mz, lt):
        return False                      # nombra el mío: no es de otro
    return any(k != (_norm(mz), _norm(lt)) for k in _lotes_en_mensaje(mensaje))


def _nombra_a_otro(mensaje: str, nombre: str) -> bool:
    """¿El mensaje nombra a una PERSONA que no es el titular del predio?

    Caso real: S/36 con mensaje "pago de servicios de agua usuario Alejandro
    Melgarejo" matcheaba como el yape de H1-15, que es de Patricia Tarazona."""
    if not mensaje or not nombre:
        return False
    t = _sin_tildes(mensaje).upper()
    if any(p in t for p in {p for p in _sin_tildes(nombre).upper().split() if len(p) >= 4}):
        return False                      # nombra al titular: es suyo
    return bool(re.search(r"\bUSUARI[OA]\b", t))


# ── Fuentes ──────────────────────────────────────────────────────────────────

_cache_mesas: dict[str, pd.DataFrame] = {}
_cache_banco: pd.DataFrame | None = None


def pagos_de_mesas(mes: str) -> pd.DataFrame:
    """LA FUENTE: lo que el cobrador escribió a mano en mesa_N.xlsx.

    Un paso más atrás que pagos_efectivo.xlsx (que ya es salida de
    4_pagos/efectivo/main.py) y dos más atrás que planilla_cobrado:

        mesa_N.xlsx  ->  pagos_efectivo.xlsx  ->  planilla_cobrado.xlsx
         (fuente)         (consolidado)            (aplicado)

    Fila 3 de cada hoja es el ejemplo guía del template — se saltea igual que
    en 4_pagos/efectivo/main.py:leer_hoja, que lee `filas[3:]`."""
    if mes in _cache_mesas:
        return _cache_mesas[mes]
    carpeta = rh.REPOS_CICLO_CERRADO.get(mes, REPO_DIR) / "4_pagos" / "efectivo" / "inputs"
    filas = []
    for n in range(1, 8):
        ruta = carpeta / f"mesa_{n}.xlsx"
        if not ruta.exists():
            continue
        try:
            wb = load_workbook(ruta, read_only=True, data_only=True)
        except Exception:
            continue
        for hoja in wb.sheetnames:
            todas = list(wb[hoja].values)
            if len(todas) < 4:
                continue
            hdr = [str(c).strip().upper() if c else "" for c in todas[1]]
            for f in todas[3:]:
                if not f or all(c in (None, "") for c in f[:6]):
                    continue
                r = dict(zip(hdr, f))
                filas.append({
                    "MZ": _norm(r.get("MZ")), "LT": _norm(r.get("LT")),
                    "MONTO": _numf(r.get("MONTO")),
                    "EFECTIVO": _numf(r.get("MONTO_EFECTIVO")),
                    "YAPE": _numf(r.get("MONTO_YAPE")),
                    "COBRADOR": _cobrador_canon(r.get("COBRADOR")), "MESA": f"mesa_{n}",
                    "HOJA": hoja, "FECHA": _txt(r.get("FECHA")),
                    "COMENTARIO": _txt(r.get("COMENTARIO")),
                    "CATEGORIA": _txt(r.get("CATEGORIA")),
                })
        wb.close()
    _cache_mesas[mes] = pd.DataFrame(filas)
    return _cache_mesas[mes]


def _reporte_banco() -> pd.DataFrame:
    """Las transacciones crudas de la cuenta de la JASS (shared/reporte_mes_crudo).

    Es la ÚNICA prueba de que un yape entró: motor_matching sale de acá."""
    global _cache_banco
    if _cache_banco is not None:
        return _cache_banco
    filas = []
    for p in sorted((SHARED_DIR / "reporte_mes_crudo").glob("ReporteTransacciones*.xlsx")):
        try:
            df = pd.read_excel(p, sheet_name="Movimientos", header=4)
        except Exception:
            continue
        df.columns = [str(c).strip() for c in df.columns]
        if "Monto" not in df.columns:
            continue
        for _, r in df.iterrows():
            f = pd.to_datetime(_txt(r.get("Fecha de operación")),
                               format="%d/%m/%Y %H:%M:%S", errors="coerce")
            if pd.isna(f):
                continue
            filas.append({"TIPO": _txt(r.get("Tipo de Transacción")),
                          "ORIGEN": _txt(r.get("Origen")), "MONTO": _numf(r.get("Monto")),
                          "MENSAJE": _txt(r.get("Mensaje")), "FECHA": f})
    _cache_banco = pd.DataFrame(filas)
    return _cache_banco


_cache_ventanas: dict[str, tuple] | None = None


def ventana_del_ciclo(mes: str) -> tuple[pd.Timestamp | None, pd.Timestamp | None]:
    """(inicio, fin) de la ventana de cobro de un ciclo, según el ANCLA DE CORTE.

    El reporte del banco abarca ~3 meses, así que buscar un yape de julio sin
    acotar puede devolver uno de junio. La ventana sale de
    shared/reporte_acumulado_procesado/<mes>_procesado.xlsx, que es lo que
    motor_matching ya procesó para ese ciclo — su fecha máxima ES el ancla
    (`obtener_ancla()` en motor_matching/main.py).

        ciclo 2026-06   19/05 19:14 -> 15/06 21:13
        ciclo 2026-07   17/06 20:32 -> 20/07 22:48   <- ancla de agosto
        ciclo 2026-08   20/07 22:48 -> (abierto, hasta el fin del reporte)

    El ciclo abierto no tiene procesado propio: su ventana empieza en el ancla
    del último cerrado y termina donde termina el reporte."""
    global _cache_ventanas
    if _cache_ventanas is None:
        rangos = {}
        for p in sorted((SHARED_DIR / "reporte_acumulado_procesado").glob("????-??_procesado.xlsx")):
            try:
                df = pd.read_excel(p, header=1)
            except Exception:
                continue
            df.columns = [str(c).strip() for c in df.columns]
            if "FECHA" not in df.columns:
                continue
            d = pd.to_datetime(df["FECHA"], errors="coerce", dayfirst=True).dropna()
            if not d.empty:
                rangos[p.name[:7]] = (d.min(), d.max())
        _cache_ventanas = rangos

    if mes in _cache_ventanas:
        return _cache_ventanas[mes]
    # Ciclo abierto: desde el ancla del último cerrado hasta el fin del reporte.
    cerrados = sorted(m for m in _cache_ventanas if m < mes)
    inicio = _cache_ventanas[cerrados[-1]][1] if cerrados else None
    banco = _reporte_banco()
    fin = banco["FECHA"].max() if not banco.empty else None
    return inicio, fin


def _leer_precursor(nombre: str) -> pd.DataFrame:
    ruta = SHARED_DIR / f"{nombre}.xlsx"
    if not ruta.exists():
        return pd.DataFrame()
    try:
        df = pd.read_excel(ruta, header=1)
    except Exception:
        return pd.DataFrame()
    df.columns = [str(c).strip().upper() for c in df.columns]
    return df


# ── La verificación ──────────────────────────────────────────────────────────

def verificar_una(monto: float, fecha, mz: str, lt: str, dias: int = 3,
                  nombre: str = "", mes: str = "") -> dict:
    """¿Existe en el banco un yape que calce con lo que el cobrador anotó?

    Devuelve {"estado", "detalle"} con uno de:
        ENCONTRADO      un mensaje del banco nombra este lote
        POSIBLE         coincide monto y fecha, pero nada confirma que sea suyo
        NO_EXISTE       no hay transacción que calce: el yape no entró
        FUERA_DE_RANGO  el reporte del banco no cubre esa fecha
        SIN_FECHA       la hoja no trae fecha para poder buscar
        SIN_REPORTE     no hay reporte del banco

    La diferencia entre ENCONTRADO y POSIBLE importa: los montos se repiten
    tanto que una coincidencia de monto+fecha no prueba nada por sí sola. Y
    FUERA_DE_RANGO no es NO_EXISTE — no es lo mismo "no está" que "no lo puedo
    saber"."""
    banco = _reporte_banco()
    if banco.empty:
        return {"estado": "SIN_REPORTE", "detalle": "no hay reporte del banco para verificar"}

    # ACOTAR AL CICLO antes de cualquier búsqueda. El reporte abarca ~3 meses y
    # sin esto un yape de junio que nombre el lote se devolvía como prueba de un
    # pago de julio — la búsqueda por mensaje no filtraba fecha en absoluto.
    ini_c, fin_c = ventana_del_ciclo(mes) if mes else (None, None)
    ventana = ""
    if ini_c is not None and fin_c is not None:
        banco = banco[(banco["FECHA"] > ini_c) & (banco["FECHA"] <= fin_c)]
        ventana = f" [ventana del ciclo {mes}: {ini_c:%d/%m %H:%M}→{fin_c:%d/%m %H:%M}]"
        if banco.empty:
            return {"estado": "SIN_REPORTE",
                    "detalle": f"el reporte del banco no tiene transacciones en{ventana}"}

    por_lote = banco[banco["MENSAJE"].apply(lambda m: _menciona_lote(m, mz, lt))]
    if not por_lote.empty:
        r = por_lote.iloc[0]
        return {"estado": "ENCONTRADO",
                "detalle": f"S/{r['MONTO']:,.2f} de {r['ORIGEN']} el "
                           f"{r['FECHA']:%d/%m/%Y %H:%M} — su mensaje nombra el lote{ventana}"}

    f = _fecha(fecha)
    if f is None:
        return {"estado": "SIN_FECHA", "detalle": "la hoja no trae fecha para poder buscar"}

    # El rango se compara contra la VENTANA de búsqueda, no contra la fecha
    # exacta: un pago del 02/08 con el reporte terminando el 01/08 igual entra
    # en la ventana de ±3 días y hay que buscarlo.
    desde, hasta = f - pd.Timedelta(days=dias), f + pd.Timedelta(days=dias)
    ini, fin = banco["FECHA"].min(), banco["FECHA"].max()
    if fin < desde or ini > hasta:
        return {"estado": "FUERA_DE_RANGO",
                "detalle": f"el reporte del banco cubre {ini:%d/%m/%Y}→{fin:%d/%m/%Y} "
                           f"y el pago dice {f:%d/%m/%Y}: no se puede verificar"}

    cerca = banco[(banco["FECHA"] >= desde) & (banco["FECHA"] <= hasta) &
                  (banco["MONTO"].sub(monto).abs() < 0.01)]
    ajenos = cerca[cerca["MENSAJE"].apply(
        lambda m: _nombra_otro_lote(m, mz, lt) or _nombra_a_otro(m, nombre))]
    cerca = cerca.drop(ajenos.index)
    if not cerca.empty:
        r = cerca.iloc[0]
        return {"estado": "POSIBLE",
                "detalle": f"hay S/{r['MONTO']:,.2f} de {r['ORIGEN']} el "
                           f"{r['FECHA']:%d/%m/%Y %H:%M} (mensaje: "
                           f"{r['MENSAJE'][:35] or '(vacío)'}) — coincide monto y fecha "
                           f"pero nada confirma que sea de este predio{ventana}"}
    if not ajenos.empty:
        r = ajenos.iloc[0]
        return {"estado": "NO_EXISTE",
                "detalle": f"NO hay transaccion de S/{monto:,.2f} para este lote entre el "
                           f"{desde:%d/%m} y el {hasta:%d/%m}; la unica de ese monto "
                           f"({r['ORIGEN']}, {r['FECHA']:%d/%m}) nombra otro: "
                           f"\"{r['MENSAJE'][:45]}\"{ventana}"}
    return {"estado": "NO_EXISTE",
            "detalle": f"NO hay ninguna transaccion de S/{monto:,.2f} en la cuenta de la "
                       f"JASS entre el {desde:%d/%m} y el {hasta:%d/%m} — el yape no "
                       f"entro a la JASS{ventana}"}


def filas_yape(mz: str, lt: str, mes: str) -> list[dict]:
    """Las filas donde el cobrador anotó un MONTO_YAPE para este predio.

    Toma MONTO_YAPE y no MONTO a propósito: hay filas con MONTO=0 y
    MONTO_YAPE>0 (el cobrador la marcó como reclamo sin cobro y anotó el yape
    igual). Filtrar por MONTO las escondía — caso real W-5 de julio."""
    df = pagos_de_mesas(mes)
    if df.empty:
        return []
    sub = df[(df["MZ"] == _norm(mz)) & (df["LT"] == _norm(lt)) & (df["YAPE"] > TOL)]
    return [{"mes": mes, "monto": float(r["YAPE"]), "fecha": r["FECHA"],
             "cobrador": r["COBRADOR"], "mesa": r["MESA"],
             "comentario": r["COMENTARIO"]} for _, r in sub.iterrows()]


def barrer(meses: list[str]) -> pd.DataFrame:
    """Todo el pueblo: cada fila con MONTO_YAPE > 0 que NO tenga respaldo.

    Dos descuentos obligatorios antes de reportar algo, porque sin ellos el
    número se infla y acusa de más (medido: S/1,127 crudos vs S/469 reales):

      ya acreditado    el predio figura en pagos_yape_tepago de ese ciclo, o sea
                       que su yape SÍ entró y motor_matching lo levantó del
                       banco. La anotación del cobrador es un duplicado.
      ya regularizado  el monto está en abonos_rezagados (plata que el cobrador
                       retuvo y entregó después). En junio fueron 7 filas por
                       S/430 que se detectaron y se aplicaron en julio."""
    boletas = vl.leer_boletas()
    # Un ciclo esta CERRADO si tiene su <mes>_procesado.xlsx: ahi la
    # conciliacion del mes ya termino y el veredicto es definitivo.
    cerrados = set(rh.REPOS_CICLO_CERRADO)
    ar = _leer_precursor("abonos_rezagados")
    rez = {(_norm(r["MZ"]), _norm(r["LT"]), round(_numf(r["MONTO"]), 2))
           for _, r in ar.iterrows()} if not ar.empty else set()

    filas = []
    for mes in meses:
        mesas = pagos_de_mesas(mes)
        if mesas.empty:
            print(f"    {mes}: mesas vacias — no se puede verificar este ciclo")
            continue
        ya = rrp._cargar_pagos_yape_crudo(mes)
        acreditados = ({(_norm(r["MZ"]), _norm(r["LOTE"])) for _, r in ya.iterrows()}
                       if ya is not None and not ya.empty else set())
        con_yape = mesas[mesas["YAPE"] > TOL]
        print(f"    {mes}: {len(mesas)} filas en mesas · {len(con_yape)} con yape anotado")

        for _, r in con_yape.iterrows():
            mz, lt = r["MZ"], r["LT"]
            if not mz or mz == "BLANCO":
                continue
            k = _clave(mz, lt)
            nombre = boletas.get(k, {}).get("nombre", "")
            if (mz, lt) in acreditados:
                estado, detalle = "YA_ACREDITADO", "su yape entro y motor_matching lo levanto"
            elif (mz, lt, round(r["YAPE"], 2)) in rez:
                estado, detalle = "YA_REGULARIZADO", "registrado en abonos_rezagados"
            else:
                chk = verificar_una(r["YAPE"], r["FECHA"], mz, lt, dias=5,
                                    nombre=nombre, mes=mes)
                estado, detalle = chk["estado"], chk["detalle"]
                # En el ciclo ABIERTO la conciliacion todavia no termino: pueden
                # entrar transacciones despues del ultimo reporte del banco y
                # motor_matching puede volver a correr. Decir "el yape no entro"
                # ahi es prematuro -- se marca aparte y se concluye al cerrar.
                if estado == "NO_EXISTE" and mes not in cerrados:
                    estado = "NO_EXISTE_PROVISIONAL"
                    detalle = (f"{detalle} || OJO: el ciclo {mes} sigue ABIERTO, "
                               f"la conciliacion no termino. Confirmar al cerrar el mes")
            filas.append({
                "MES": mes, "MZ": mz, "LT": lt, "NOMBRE": nombre,
                "MONTO": r["MONTO"], "MONTO_YAPE": r["YAPE"], "COBRADOR": r["COBRADOR"],
                "MESA": r["MESA"], "FECHA": r["FECHA"], "COMENTARIO": r["COMENTARIO"],
                "ESTADO": estado, "EVIDENCIA": detalle,
            })
    return pd.DataFrame(filas)


# ── Salida ───────────────────────────────────────────────────────────────────

_SEC_QUIEN = ("EBF5FB", "1A5276", "F4FAFF")
_SEC_ANOTO = ("FEF3C7", "92400E", "FFFBEB")
_SEC_BANCO = ("E9F7EF", "1E5C3A", "F4FBF7")
_SEC_MANUAL = ("F3E8FF", "5B21B6", "FAF5FF")

_COLS = [
    ("MES",        _SEC_QUIEN,  10, "center", None),
    ("MZ",         _SEC_QUIEN,   6, "center", None),
    ("LT",         _SEC_QUIEN,   7, "center", None),
    ("NOMBRE",     _SEC_QUIEN,  28, "left",   None),
    ("MONTO",      _SEC_ANOTO,  11, "right",  '"S/ "#,##0.00'),
    ("MONTO_YAPE", _SEC_ANOTO,  13, "right",  '"S/ "#,##0.00'),
    ("COBRADOR",   _SEC_ANOTO,  20, "left",   None),
    ("MESA",       _SEC_ANOTO,  10, "center", None),
    ("FECHA",      _SEC_ANOTO,  13, "center", None),
    ("COMENTARIO", _SEC_ANOTO,  38, "left",   None),
    ("ESTADO",     _SEC_BANCO,  17, "center", None),
    ("EVIDENCIA",  _SEC_BANCO,  74, "left",   None),
    ("RESOLUCION", _SEC_MANUAL, 34, "left",   None),
]

_SECCIONES = [("¿De quién es?", "MES", "NOMBRE"),
              ("¿Qué anotó el cobrador?", "MONTO", "COMENTARIO"),
              ("¿Está en la cuenta de la JASS?", "ESTADO", "EVIDENCIA"),
              ("Resolución", "RESOLUCION", "RESOLUCION")]

_COLOR_ESTADO = {
    "NO_EXISTE":            ("FEE2E2", "991B1B"),   # ciclo cerrado: concluyente
    "NO_EXISTE_PROVISIONAL": ("FEF3C7", "92400E"),  # ciclo abierto: falta cerrar
    "POSIBLE":         ("FEF9E7", "7D6608"),
    "SIN_FECHA":       ("FEF9E7", "7D6608"),
    "FUERA_DE_RANGO":  ("F1F3F5", "868E96"),
    "ENCONTRADO":      ("D5F5E3", "1E5C3A"),
    "YA_ACREDITADO":   ("D5F5E3", "1E5C3A"),
    "YA_REGULARIZADO": ("D5F5E3", "1E5C3A"),
}


def _fill(hex6):
    return PatternFill("solid", fgColor=f"FF{hex6}")


def escribir(df: pd.DataFrame, sufijo: str) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Verificacion"
    idx = {c[0]: i + 1 for i, c in enumerate(_COLS)}

    for label, ini, fin in _SECCIONES:
        c1, c2 = idx[ini], idx[fin]
        sec = _COLS[c1 - 1][1]
        if c1 != c2:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        cell = ws.cell(row=1, column=c1, value=label)
        cell.fill, cell.font = _fill(sec[0]), Font(color=f"FF{sec[1]}", bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18

    for i, (nombre, sec, ancho, _a, _f) in enumerate(_COLS, start=1):
        cell = ws.cell(row=2, column=i, value=nombre)
        cell.fill, cell.font = _fill(sec[0]), Font(color=f"FF{sec[1]}", bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "D3"

    for ri, (_, row) in enumerate(df.iterrows(), start=3):
        for ci, (nombre, sec, _w, align, fmt) in enumerate(_COLS, start=1):
            val = row.get(nombre, "")
            if val is None or (isinstance(val, float) and pd.isna(val)):
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            bg, fg = sec[2], sec[1]
            if nombre == "ESTADO":
                bg, fg = _COLOR_ESTADO.get(str(val), (sec[2], sec[1]))
            cell.fill = _fill(bg)
            cell.font = Font(color=f"FF{fg}", size=10, bold=(nombre == "ESTADO"))
            cell.alignment = Alignment(horizontal=align, vertical="top",
                                       wrap_text=(nombre in ("COMENTARIO", "EVIDENCIA")))
            if fmt:
                cell.number_format = fmt

    out = BASE_DIR / "outputs" / f"verificacion_yape_{sufijo}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(out)
    except PermissionError:
        alt = out.with_name(f"{out.stem}_{datetime.now():%H%M%S}{out.suffix}")
        wb.save(alt)
        print(f"\n  AVISO: {out.name} esta abierto en Excel — se guardo como {alt.name}")
        return alt
    return out


def main(meses: list[str], sufijo: str) -> None:
    print(f"=== verificar_yape — ciclos: {' · '.join(meses)} ===")
    df = barrer(meses)
    if df.empty:
        print("  ninguna fila con MONTO_YAPE anotado — nada que verificar")
        return

    print("\n  estados:")
    for e, n in df["ESTADO"].value_counts().items():
        print(f"    {n:>3}  {e}")

    def _por_cobrador(g):
        for cob, sub in g.groupby("COBRADOR"):
            print(f"    {cob:<22} S/{sub['MONTO_YAPE'].sum():>8,.2f}  "
                  f"({len(sub)} filas: "
                  f"{' · '.join(f'{r.MZ}-{r.LT}' for r in sub.itertuples())})")

    perdidos = df[df["ESTADO"] == "NO_EXISTE"]
    if not perdidos.empty:
        print(f"\n  YAPE QUE NO ENTRO A LA JASS — ciclos CERRADOS, concluyente")
        print(f"  {len(perdidos)} filas · S/{perdidos['MONTO_YAPE'].sum():,.2f}")
        _por_cobrador(perdidos)
        print("\n  No dice quien se quedo con la plata: pedir la captura del yape al "
              "vecino,\n  ahi sale el numero de destino.")

    prov = df[df["ESTADO"] == "NO_EXISTE_PROVISIONAL"]
    if not prov.empty:
        print(f"\n  PROVISIONAL — ciclo ABIERTO, la conciliacion no termino")
        print(f"  {len(prov)} filas · S/{prov['MONTO_YAPE'].sum():,.2f}   "
              f"NO reclamar todavia: confirmar al cerrar el mes")
        _por_cobrador(prov)

    out = escribir(df, sufijo)
    print(f"\n  -> {out}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(
        description="Verifica contra el banco los yape anotados en las hojas de efectivo")
    ap.add_argument("--mes", default=None, help="Ciclo a verificar (YYYY-MM). Default: activo")
    ap.add_argument("--todos", action="store_true",
                    help="Todos los ciclos con archivo disponible (cerrados + activo)")
    args = ap.parse_args()

    activo = ciclo.activo(default=None)
    if args.todos:
        meses = sorted(set(rh.REPOS_CICLO_CERRADO) | ({activo} if activo else set()))
        sufijo = "todos"
    else:
        mes = args.mes or activo
        if not mes:
            ap.error("--mes requerido (no hay ciclo activo en shared/ciclo_activo.json)")
        meses, sufijo = [mes], mes
    main(meses, sufijo)
