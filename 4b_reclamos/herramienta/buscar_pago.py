"""
4b_reclamos/herramienta/buscar_pago.py — ¿existe el pago de "ya pagué mes anterior"?

Para cada reclamo de TIPO_RECLAMO=mes_anterior responde, con evidencia: ¿el pago
existe? ¿un bug lo ocultó? ¿está acreditado a otro predio? ¿nunca existió? Nunca
cierra el reclamo — eso lo hace el supervisor en resolucion_reclamos_YYYY-MM.xlsx.

SOLO mes_anterior, a propósito. "Ya pagué mi medidor" / "ya pagué techado y
campo" NO se auditan acá: su causa es anterior y distinta — el orden de la
cascada, que hoy cobra la multa antes del convenio y los acuerdos (ver el
bloque TIPO más abajo y reporte_reimputacion_cascada.py). Buscar el pago de un
convenio antes de ese reorden devuelve "pagó parte de sus cuotas", que es cierto
y no explica nada.

Diseño completo:  docs/decisiones/buscar_pago.md
Diagrama:         4b_reclamos/docs/diagrama_flujo_buscar_pago.html

EL EMBUDO (primer veredicto que matchea gana):

    GATE   el cargo disputado ya está en 0 en la boleta vigente
           -> RESUELTO_YA, no se busca nada más

    BLOQUE A — EXPLICAR (la plata ya está adentro del sistema)
      A1a  el pago de ese mes era aporte al tanque, no agua
           -> PAGO_PERO_NO_ERA_AGUA        (va PRIMERO: invalida la premisa
                                            "hubo pago de agua" de A0)
      A1b  abono rezagado: pagó en el mes disputado, se aplicó después
           -> PAGO_ANTES_APLICADO_DESPUES
      A0   el mes disputado, y A2 el ciclo del propio reclamo — misma regla
           (_clasificar_mes), en ese orden de fuerza:
             mes anterior recibió algo, quedó saldo  -> PAGO_PARCIAL
             quedó en 0 y la plata fue a multa /
             acuerdos / convenio                     -> CASCADA_FUERA_DE_ORDEN
                                                        (la anomalía real: esos
                                                         van DESPUÉS del arrastre)
             quedó en 0 y la plata fue a consumo +
             mantenimiento                           -> PAGO_SOLO_EL_MES
                                                        (cascada CORRECTA, no es
                                                         hallazgo: pagó el mes y
                                                         no pagó el arrastre
                                                         porque dice no deberlo.
                                                         14 de 29 en 2026-08)
      A1c  otro precursor de shared/ toca este predio en el mes que aplica
           -> EXPLICADO_POR_PRECURSOR

    BLOQUE B — BUSCAR (la plata no está: ¿dónde se fue?)
      B2b  el MENSAJE de un pago menciona este lote (1 transacción, 2 lotes)
           -> CANDIDATO_MULTILOTE          (caso real K-3/K-4, 2026-08-10)
      B2a  pago en un lote confundible cuyo monto cuadra con lo que este debía
           -> CANDIDATO_TIPEO
      B1   blanco sin reclamar, monto y ventana temporal plausibles
           -> CANDIDATO_BLANCO
      B3   exceso no resuelto de un predio confundible
           -> CANDIDATO_EXCESO

    SIN_EVIDENCIA -> pedir recibo o captura de yape

REGLA DE PROPUESTA (heredada de 4_pagos/efectivo/verificar_lotes.py): un
candidato del Bloque B se propone SOLO si la lista queda en exactamente 1. Con
2+ se reporta "N candidatos" sin elegir — proponer con esa evidencia es
inventar con cara de certeza.

VENTANA TEMPORAL (regla de negocio): el reclamo normal es de 1 mes de distancia
(en agosto reclaman julio, para no ser cortados sin pagar el mes). Un candidato
de 2+ meses atrás solo tiene sentido si el vecino ARRASTRÓ esa deuda todos los
meses intermedios (se niega a pagar y re-reclama). Si en algún mes intermedio no
debía nada, el candidato viejo se descarta.

Uso:
    py buscar_pago.py                # mes = ciclo activo (shared/ciclo_activo.json)
    py buscar_pago.py --mes 2026-08
"""

import argparse
import re
import sys
import unicodedata
from datetime import datetime
from itertools import combinations
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERRAMIENTA_DIR = Path(__file__).parent          # 4b_reclamos/herramienta/ (este archivo)
BASE_DIR = HERRAMIENTA_DIR.parent                # 4b_reclamos/
REPO_DIR = BASE_DIR.parent                       # raíz del repo activo
SHARED_DIR = REPO_DIR / "shared"

sys.path.insert(0, str(HERRAMIENTA_DIR))
sys.path.insert(0, str(BASE_DIR))
sys.path.insert(0, str(SHARED_DIR))
sys.path.insert(0, str(REPO_DIR / "4_pagos" / "efectivo"))

import ciclo                            # noqa: E402
import reporte_historico as rh          # noqa: E402
import verificar_lotes as vl            # noqa: E402  (confundible + subconjuntos + boletas)

# reporte_referencias_pago.py se absorbió acá el 2026-08-18 (ver mas abajo,
# seccion "REFERENCIA DE PAGO -- absorbido de reporte_referencias_pago.py"):
# menos archivos con importaciones cruzadas que se rompen si alguno se mueve
# de carpeta.

# La lectura de las mesas y toda la verificacion del yape contra el banco viven
# en verificar_yape.py, que ademas corre sola para barrer TODO el pueblo. Aca se
# importan sus primitivos para no tener dos copias que se desincronicen.
from verificar_yape import (  # noqa: E402
    _clave, _fecha, _menciona_lote, _norm, _numf, _sin_tildes, _txt,
    _nombra_a_otro, _nombra_otro_lote, _reporte_banco,
    _PATRON_YAPE, filas_yape, pagos_de_mesas, verificar_una,
)

# Los motivos de vl.confundible() traen "→"; la consola de Windows es cp1252.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

TOL = 0.005

# SOLO mes_anterior. Los demás TIPO_RECLAMO NO se auditan con esta herramienta:
# "ya pagué mi medidor" y "ya pagué techado y campo" tienen una causa distinta y
# anterior — el ORDEN de la cascada. Hoy reparte
#     consumo · mantenimiento · mes anterior · multa · acuerdos · convenio
# y va a repartir
#     consumo · mantenimiento · mes anterior · convenio · acuerdos · multa
# (la multa al final: es lo único que se puede cubrir con faena o exonerar).
# Mientras ese reorden no esté aplicado, buscar el pago de un convenio devuelve
# "pagó parte de sus cuotas", que es cierto y no explica nada. Es otro trabajo,
# ya simulado en 4b_reclamos/reporte_reimputacion_cascada.py.
TIPO = "mes_anterior"
CARGO_BOLETA = "anterior"        # clave en vl.leer_boletas()[k]["cargos"]
COL_TABLA = "MES_ANT"            # columna de rh.tabla_predio() — qué se aplicó
COL_PLANILLA = "MES_ANTERIOR"    # columna de shared/planilla_mes/ — qué se DEBÍA

# Conceptos que la cascada debe cobrar DESPUÉS del mes anterior. Que el pago del
# mes se vaya a consumo + mantenimiento es correcto y NO es un hallazgo — es la
# cascada funcionando (P1: agua del mes primero). La anomalía real sería que
# cobre multa, acuerdos o convenio dejando el mes anterior sin pagar: eso sí es
# fuera de orden. Medido sobre los 29 reclamos de 2026-08: 0 casos — los 14 que
# no pagaron mes anterior pagaron exactamente consumo+mantenimiento.
CONCEPTOS_POSTERIORES = ["MULTA", "ACUERDOS", "CONVENIO"]

# Precursores que se leen para el uso ① (explicar) y ⑤ (filtrar candidatos con
# dueño). deuda_directiva.xlsx queda fuera a propósito: se indexa por USER_ID,
# no por (MZ,LT), así que no se puede cruzar contra un reclamo de predio.
PRECURSORES_EXPLICAN = [
    ("ajustes_cargo",             "MES_ANO_APLICA"),
    ("deuda_correcciones",        "MES_ANO_APLICA"),
    ("genesis_tardia",            "MES_ANO_APLICA"),
    ("devoluciones_aplicadas",    "MES_ANO_APLICA"),
    ("reasignaciones_aplicacion", "MES_ANO"),
    ("exoneraciones_multa",       "MES_VIGENCIA"),
]


# ── Normalización ────────────────────────────────────────────────────────────

def _meses_esperados(hasta: str) -> list[str]:
    """Todos los meses en los que el predio DEBIÓ pagar, hasta el ciclo dado.

    Sale del primer archivo histórico (oct-2025) en adelante, que es donde
    arranca el historial que arma rh.tabla_predio()."""
    desde = rh._ARCHIVOS_HISTORICOS[0][1]
    out, m = [], desde
    while m <= hasta:
        out.append(m)
        y, mm = int(m[:4]), int(m[5:7])
        t = y * 12 + mm
        m = f"{t // 12:04d}-{t % 12 + 1:02d}"
    return out


def _mes_menos(mes_ano: str, n: int) -> str:
    y, m = int(mes_ano[:4]), int(mes_ano[5:7])
    t = y * 12 + (m - 1) - n
    return f"{t // 12:04d}-{t % 12 + 1:02d}"


def _dist_meses(a: str, b: str) -> int:
    """Meses de a − b. Positivo si a es posterior a b."""
    return (int(a[:4]) * 12 + int(a[5:7])) - (int(b[:4]) * 12 + int(b[5:7]))


# ── Lectura de inputs ────────────────────────────────────────────────────────

def _leer_reclamos(mes: str, tipo: str) -> pd.DataFrame:
    ruta = BASE_DIR / "outputs" / f"reclamos_{mes}.xlsx"
    if not ruta.exists():
        raise FileNotFoundError(
            f"Falta {ruta}\n"
            f"  -> correr 4b_reclamos/main.py --mes {mes} antes de buscar pagos.")
    df = pd.read_excel(ruta, sheet_name="Reclamos", header=1)
    df = df[df["TIPO_RECLAMO"].astype(str).str.strip() == tipo].reset_index(drop=True)
    return df


def _leer_precursor(nombre: str) -> pd.DataFrame:
    ruta = SHARED_DIR / f"{nombre}.xlsx"
    if not ruta.exists():
        return pd.DataFrame()
    try:
        df = pd.read_excel(ruta, header=1)
    except Exception:
        return pd.DataFrame()
    df.columns = [str(c).strip().upper() for c in df.columns]
    return df


def _pagos_del_ciclo(mes: str) -> pd.DataFrame:
    """Todos los pagos de UN ciclo (yape + efectivo) en un solo shape:
    MZ · LT · MONTO · ORIGEN · FECHA · MENSAJE · CANAL.

    El pool es TODO lo cobrado del mes (~300 filas), no una lista curada de
    "no identificados" — un pago que el vecino jura haber hecho casi nunca está
    marcado como perdido: está atribuido a un vecino (ver decisión de diseño)."""
    filas = []

    yape = _cargar_pagos_yape_crudo(mes)
    if yape is not None and not yape.empty:
        for _, r in yape.iterrows():
            filas.append({
                "MZ": _norm(r.get("MZ")), "LT": _norm(r.get("LOTE")),
                "MONTO": _numf(r.get("MONTO_ASIGNADO") if pd.notna(r.get("MONTO_ASIGNADO"))
                               else r.get("MONTO_PAGO")),
                "ORIGEN": _txt(r.get("ORIGEN")), "FECHA": _txt(r.get("FECHA")),
                "MENSAJE": _txt(r.get("MENSAJE")), "CANAL": "YAPE",
            })

    efec = _cargar_pagos_efectivo_crudo(mes)
    if efec is not None and not efec.empty:
        for _, r in efec.iterrows():
            filas.append({
                "MZ": _norm(r.get("MZ")), "LT": _norm(r.get("LT")),
                "MONTO": _numf(r.get("MONTO")),
                "ORIGEN": _txt(r.get("COBRADOR")), "FECHA": _txt(r.get("FECHA")),
                "MENSAJE": _txt(r.get("COMENTARIO")), "CANAL": "EFECTIVO",
            })

    return pd.DataFrame(filas)


def _excesos_no_resueltos() -> pd.DataFrame:
    """arrastre_devolucion de cada ciclo, solo las filas SIN resolver.

    Uso ⑤ del precursor: un exceso con ESTADO=resuelto ya tiene dueño y motivo
    escrito en REVISION — re-proponerlo como candidato sería repetir
    investigación ya cerrada."""
    filas = []
    dirs = {m: rh.REPOS_CICLO_CERRADO[m] / "5_cobranza" / "outputs" for m in rh.REPOS_CICLO_CERRADO}
    activo = ciclo.activo(default=None)
    if activo and activo not in dirs:
        dirs[activo] = REPO_DIR / "5_cobranza" / "outputs"

    for mes, carpeta in dirs.items():
        ruta = ciclo.resolver(carpeta, "arrastre_devolucion", mes)
        if not ruta.exists():
            continue
        df = pd.read_excel(ruta, header=1)
        df.columns = [str(c).strip().upper() for c in df.columns]
        for _, r in df.iterrows():
            if _txt(r.get("ESTADO")).lower() in ("resuelto", "resuelta"):
                continue
            filas.append({
                "MES": mes, "MZ": _norm(r.get("MZ")), "LT": _norm(r.get("LT")),
                "NOMBRE": _txt(r.get("NOMBRE")), "MONTO": _numf(r.get("MONTO")),
                "REFERENCIA": _txt(r.get("REFERENCIA")), "COMENTARIO": _txt(r.get("COMENTARIO")),
            })
    return pd.DataFrame(filas)


def _blancos_disponibles() -> pd.DataFrame:
    """Pool de pagos sin identificar que NADIE reclamó todavía:
    MES · MONTO · ORIGEN · MENSAJE · FUENTE.

    Uso ② del precursor: reidentificacion.xlsx dice qué pago ya se asignó a un
    predio concreto. Sin ese filtro el embudo puede proponer un blanco que ya
    tiene dueño y acreditar la plata dos veces (es la regla de R-7/M-7)."""
    reasignados = set()
    reid = _leer_precursor("reidentificacion")
    if not reid.empty and "MONTO" in reid.columns:
        for _, r in reid.iterrows():
            reasignados.add(round(_numf(r.get("MONTO")), 2))

    filas = []

    # yape sin identificar (motor_matching escribe acá cada corrida)
    ac = _leer_precursor("blancos_acumulados")
    if not ac.empty:
        for _, r in ac.iterrows():
            if _txt(r.get("ESTADO")).lower() == "aplicado":
                continue                       # ya tiene dueño
            if _norm(r.get("MZ")):
                continue                       # ya identificado a un lote
            filas.append({"MES": _txt(r.get("MES")), "MONTO": _numf(r.get("MONTO")),
                          "ORIGEN": _txt(r.get("ORIGEN")), "MENSAJE": _txt(r.get("MENSAJE")),
                          "FUENTE": "blancos_acumulados"})

    # efectivo que entró como blanco
    be = _leer_precursor("blancos_efectivo")
    if not be.empty:
        for _, r in be.iterrows():
            if _txt(r.get("MES_ANO_APLICA")):
                continue                       # apagado = ya se resolvió por otra vía
            filas.append({"MES": _txt(r.get("MES_CICLO")), "MONTO": _numf(r.get("MONTO")),
                          "ORIGEN": _txt(r.get("ORIGEN")), "MENSAJE": _txt(r.get("MOTIVO")),
                          "FUENTE": "blancos_efectivo"})

    # pre-mayo: la hoja "Reporte" de cada planilla histórica, mz="blanco"
    for archivo, mes in rh._ARCHIVOS_HISTORICOS:
        hojas = _cargar_hojas_historicas(archivo)
        dfr = hojas.get("Reporte")
        if dfr is None:
            continue
        cmz = _col(dfr, "mz")
        cmonto, cfecha = _col(dfr, "monto"), _col(dfr, "fecha de operacion")
        corigen, cmsg = _col(dfr, "origen"), _col(dfr, "mensaje")
        if cmz is None:
            continue
        sub = dfr[dfr[cmz].astype(str).str.strip().str.lower() == "blanco"]
        for _, r in sub.iterrows():
            filas.append({"MES": mes, "MONTO": _numf(r.get(cmonto)),
                          "ORIGEN": _txt(r.get(corigen)),
                          "MENSAJE": f"{_txt(r.get(cmsg))} {_txt(r.get(cfecha))}".strip(),
                          "FUENTE": f"hoja Reporte {mes}"})

    df = pd.DataFrame(filas)
    if not df.empty:
        df = df[~df["MONTO"].round(2).isin(reasignados)]
    return df.reset_index(drop=True)


# ── GATE ─────────────────────────────────────────────────────────────────────

def _cargo_vigente(mz: str, lt: str, boletas: dict) -> float | None:
    """Lo que la boleta VIGENTE todavía le cobra por el concepto disputado.
    None = el predio no tiene boleta este ciclo."""
    bo = boletas.get(_clave(mz, lt))
    if bo is None:
        return None
    return _numf(bo["cargos"].get(CARGO_BOLETA))


# ── BLOQUE A — explicar ──────────────────────────────────────────────────────

def _aporte_tanque(mz: str, lt: str, meses: set[str]) -> dict | None:
    """Uso ③ del precursor: la plata entró de verdad pero NO era agua. El
    vecino tiene razón en caja y no en deuda — y sin este chequeo A0 vería
    "hubo pago, el concepto quedó en 0" y diría MAL_IMPUTADO, que es falso."""
    df = _leer_precursor("aportes_tanque_manuales")
    if df.empty:
        return None
    for _, r in df.iterrows():
        if _norm(r.get("MZ")) != _norm(mz) or _norm(r.get("LT")) != _norm(lt):
            continue
        m = _txt(r.get("MES_ANO_APLICA")) or _txt(r.get("MES_CICLO"))
        if meses and m not in meses:
            continue
        return {"monto": _numf(r.get("MONTO")), "mes": m,
                "detalle": f"aporte al tanque S/{_numf(r.get('MONTO')):,.2f} "
                           f"({_txt(r.get('FECHA_REAL'))}) — {_txt(r.get('MOTIVO'))}"}
    return None


def _abono_rezagado(mz: str, lt: str, meses: set[str]) -> dict | None:
    """Uso ④ del precursor: el desfase de fechas ES la explicación. Un vecino
    que pagó el 28/06 y se le aplicó en julio dice "ya pagué mes anterior" y
    tiene razón — la plata existe, entró antes, se acreditó después."""
    df = _leer_precursor("abonos_rezagados")
    if df.empty:
        return None
    for _, r in df.iterrows():
        if _norm(r.get("MZ")) != _norm(mz) or _norm(r.get("LT")) != _norm(lt):
            continue
        aplica, real = _txt(r.get("MES_ANO_APLICA")), _txt(r.get("FECHA_REAL"))
        ciclo_pago = _txt(r.get("MES_CICLO"))
        if meses and not ({aplica, ciclo_pago} & meses):
            continue
        # Sin desfase real no hay nada que explicar: si la plata entró y se
        # aplicó en el mismo ciclo, el vecino no "pagó antes". Un predio puede
        # tener varias filas y solo alguna con desfase (caso I-9: una de junio
        # aplicada en julio, otra de julio aplicada en julio) — se sigue
        # buscando hasta encontrar la que sí lo tiene.
        if not (aplica and ciclo_pago) or _dist_meses(aplica, ciclo_pago) < 1:
            continue
        return {"monto": _numf(r.get("MONTO")), "mes": aplica,
                "detalle": f"abono rezagado S/{_numf(r.get('MONTO')):,.2f} — pagó {real} "
                           f"(ciclo {ciclo_pago}), se aplicó en {aplica}. "
                           f"Retenido por {_txt(r.get('RETENIDO_POR'))}"}
    return None


def _otro_precursor(mz: str, lt: str) -> list[dict]:
    """⑦ — TODOS los eventos de shared/ que tocan este predio, en cualquier mes.

    Antes devolvía el primero y filtraba por un mes: para un reclamo de mes
    anterior eso lo dejaba casi siempre en None, así que el reporte no decía
    nada de precursores aunque los hubiera. Un evento de cualquier mes puede
    explicar de dónde salió el arrastre."""
    out = []
    for nombre, col_mes in PRECURSORES_EXPLICAN:
        df = _leer_precursor(nombre)
        if df.empty or "MZ" not in df.columns:
            continue
        for _, r in df.iterrows():
            if _norm(r.get("MZ")) != _norm(mz) or _norm(r.get("LT")) != _norm(lt):
                continue
            m = _txt(r.get(col_mes))
            out.append({"fuente": nombre, "mes": m,
                        "detalle": f"{nombre} ({m}): {_txt(r.get('CONCEPTO'))} "
                                   f"S/{_numf(r.get('MONTO')):,.2f} — {_txt(r.get('MOTIVO'))[:90]}"})
    return out


def filas_yape_todos(mz: str, lt: str, mesas: dict) -> list[dict]:
    """filas_yape() de verificar_yape, pero barriendo todos los ciclos cargados."""
    return [f for mes in mesas for f in filas_yape(mz, lt, mes)]


def _cobrador_mencionado(texto: str, cobradores: set[str]) -> str:
    """El cobrador que el vecino nombra en su reclamo ("con Maximo", "le pagó a
    Wagner Trujillo", "yape a Janet").

    Sirve para EMPEZAR por ahí, no para limitar la búsqueda: el vecino a veces
    dice cualquier nombre para salir del paso cuando se le pregunta a quién le
    pagó. Si no aparece con ese cobrador, se busca igual en todas las mesas."""
    t = _sin_tildes(texto).upper()
    for c in cobradores:
        for tok in _sin_tildes(c).upper().split():
            if len(tok) >= 5 and tok in t:
                return c
    return ""


def _en_fuente_sin_consolidar(mz: str, lt: str, tabla: pd.DataFrame,
                               mesas: dict[str, pd.DataFrame],
                               cobrador_dicho: str = "",
                               ya_explicados: set[float] = frozenset()) -> list[dict]:
    """② — ¿aparece en la FUENTE (mesa_N.xlsx) en un mes en el que el historial
    dice que no pagó nada? Eso es plata cobrada que nunca se aplicó.

    El cobrador que el vecino nombra ordena los resultados primero (el gol
    rápido), pero se barren TODAS las mesas igual.

    `ya_explicados` son los montos que un precursor ya se llevó (uso ⑤): sin
    ese filtro, un abono rezagado se reporta como pago perdido. Caso real I-9:
    S/86 anotados por Wagner en mesa_6 de junio, que efectivamente NO figuran en
    junio... porque abonos_rezagados los aplicó en julio. Ya tiene dueño."""
    pagados = set(tabla[tabla["TOTAL"] > TOL]["MES"].astype(str))
    out = []
    for mes, df in mesas.items():
        if df.empty or mes in pagados:
            continue
        sub = df[(df["MZ"] == _norm(mz)) & (df["LT"] == _norm(lt))]
        for _, p in sub.iterrows():
            monto = float(p["MONTO"])
            if round(monto, 2) in ya_explicados:
                continue
            coincide = bool(cobrador_dicho) and p["COBRADOR"] == cobrador_dicho
            # Un yape anotado en la hoja de efectivo cae entre dos módulos:
            # 4_pagos/efectivo solo procesa MONTO_EFECTIVO, y motor_matching lee
            # el reporte del banco, no las mesas. Nadie lo levanta (caso F1-8).
            solo_yape = p["YAPE"] > TOL and p["EFECTIVO"] <= TOL
            out.append({
                "mes": mes, "monto": monto, "coincide_cobrador": coincide,
                "detalle": f"{mes}: S/{monto:,.2f} escrito por {p['COBRADOR']} "
                           f"en {p['MESA']}/{p['HOJA']} ({p['FECHA']})"
                           + (" — Y ES EL COBRADOR QUE NOMBRA EL VECINO" if coincide else "")
                           + (f" — anotado como YAPE en la hoja de efectivo: 4_pagos/efectivo "
                              f"solo procesa efectivo y motor_matching no lee las mesas, "
                              f"así que no lo levanta ningún módulo" if solo_yape else "")
                           + f" · el historial de {mes} dice que no pagó"})
    out.sort(key=lambda d: not d["coincide_cobrador"])
    return out


def _historico_mes_anterior(tabla: pd.DataFrame) -> tuple[str, str]:
    """(etiqueta corta para la columna, texto largo para la evidencia).

    ¿Alguna vez, en todo el historial, se le aplicó plata al mes anterior? Es el
    discriminador del reclamo "ya pagué meses anteriores", y parte los casos en
    dos grupos con acción distinta (medido sobre los 14 de 2026-08):

      nunca (8)      el arrastre nunca se pagó: el reclamo no tiene respaldo en
                     el sistema. Contra el grupo de control (60 predios que
                     deben arrastre y NO reclamaron) esto pasa en 18%, así que
                     entre los reclamantes está sobre-representado — es el perfil
                     del vecino que paga el mes y se niega al arrastre.
      viene pagando  (6) Q-9 pagó S/76 de arrastre en 4 meses distintos, Z-17
                     S/101 en 4, H-13 S/96 en 3. Estos SÍ vienen pagando
                     arrastres y aun así les sigue apareciendo — acá es donde
                     buscar un problema real de reconciliación, no en los otros 8."""
    con = tabla[tabla[COL_TABLA] > TOL]
    if con.empty:
        return "nunca", "en todo el historial nunca se le aplicó nada a mes anterior"
    total, meses = float(con[COL_TABLA].sum()), con["MES"].astype(str).tolist()
    return (f"S/{total:,.2f} en {len(meses)} meses",
            f"histórico: mes anterior recibió S/{total:,.2f} en {' · '.join(meses)}")


def _clasificar_mes(mz: str, lt: str, mes: str, tabla: pd.DataFrame,
                    cargos: dict, nota_tanque: str = "") -> dict | None:
    """Qué pasó con el pago de ESE mes respecto del mes anterior.

    Tres desenlaces, y la diferencia entre el 2do y el 3ro es la que importa:

      PAGO_PARCIAL            mes anterior recibió algo y quedó saldo.

      CASCADA_FUERA_DE_ORDEN  mes anterior quedó en 0 pero la plata SÍ fue a
                              multa / acuerdos / convenio. Esa es la anomalía
                              real: esos tres se cobran DESPUÉS del arrastre.
                              (0 casos en 2026-08 — la cascada está bien.)

      PAGO_SOLO_EL_MES        mes anterior quedó en 0 y la plata fue a consumo +
                              mantenimiento. Eso es la cascada CORRECTA (P1:
                              agua del mes primero) — no es un hallazgo. Lo que
                              pasó es que el vecino pagó el mes y NO pagó el
                              arrastre, precisamente porque su reclamo dice que
                              no lo debe (ya lo pagó antes). Lo que hay que
                              verificar entonces es el ORIGEN del arrastre, no
                              re-imputar este pago.

    Nota sobre el monto: cuando consumo+mantenimiento suma lo mismo que el cargo
    de mes anterior (6 de 14 casos en 2026-08: O-28, U-2, Y-3, T-15, O-21, F1-1
    tienen 5+3=8 y anterior=8), el monto pagado NO permite decidir qué quiso
    pagar. Se dice explícitamente en vez de afirmar una de las dos lecturas."""
    if COL_TABLA not in tabla.columns:
        return None
    filas = tabla[(tabla["MES"] == mes) & (tabla["TOTAL"] > TOL)]
    if filas.empty:
        return None                        # no hubo pago real ese mes
    f = filas.iloc[0]
    pagado, aplicado = float(f["TOTAL"]), float(f[COL_TABLA])

    if aplicado > TOL:
        return {"veredicto": "PAGO_PARCIAL", "monto": pagado,
                "detalle": f"en {mes} pagó S/{pagado:,.2f}, de los cuales "
                           f"S/{aplicado:,.2f} fueron a mes anterior — quedó saldo"
                           + nota_tanque}

    # Que mes anterior reciba 0 solo se evalúa si ese mes SE DEBÍA.
    debia = _debia(mz, lt, mes, COL_PLANILLA)
    if debia is not None and debia <= TOL:
        return None

    posteriores = {c: float(f[c]) for c in CONCEPTOS_POSTERIORES
                   if c in tabla.columns and float(f[c]) > TOL}
    if posteriores:
        detalle = " + ".join(f"{c} S/{v:,.2f}" for c, v in posteriores.items())
        return {"veredicto": "CASCADA_FUERA_DE_ORDEN", "monto": pagado,
                "detalle": f"en {mes} pagó S/{pagado:,.2f}, mes anterior recibió S/0.00 "
                           f"y en cambio se cobró {detalle} — esos van DESPUÉS del "
                           f"arrastre. Revisar el orden de aplicación" + nota_tanque}

    cm = _numf(cargos.get("consumo")) + _numf(cargos.get("mant"))
    ambiguo = abs(cm - _numf(cargos.get(CARGO_BOLETA))) < 0.01
    return {"veredicto": "PAGO_SOLO_EL_MES", "monto": pagado,
            "detalle": f"en {mes} pagó S/{pagado:,.2f} y fue a consumo+mantenimiento "
                       f"(cascada correcta); mes anterior quedó sin pagar. "
                       + (f"Ojo: consumo+mant suma S/{cm:,.2f}, lo mismo que el cargo de "
                          f"mes anterior — el monto no dice cuál de los dos quiso pagar. "
                          if ambiguo else "")
                       + _historico_mes_anterior(tabla)[1]
                       + ". Verificar el origen del arrastre, no re-imputar este pago"
                       + nota_tanque}


# ── Ventana temporal ─────────────────────────────────────────────────────────

def _debia(mz: str, lt: str, mes: str, col: str) -> float | None:
    """Lo que el predio DEBÍA ese mes según shared/planilla_mes (el CARGO real
    de 2_planilla). None = no hay planilla de ese mes para probar nada."""
    df = rh._cargar_planilla_correcta(mes)
    if df is None or col not in df.columns:
        return None
    fila = df[(df["MZ"] == _norm(mz)) & (df["LT"] == _norm(lt))]
    return _numf(fila.iloc[0].get(col)) if not fila.empty else None


def _plausible(mz: str, lt: str, mes_cand: str, mes_reclamo: str) -> tuple[bool, str]:
    """La regla del negocio: el reclamo normal mira 1 mes atrás (en agosto
    reclaman julio, para no ser cortados sin pagar el mes). Un candidato de 2+
    meses solo tiene sentido si arrastró esa deuda TODOS los meses intermedios
    — si en alguno no debía nada, la deuda se cerró y el candidato viejo no
    puede ser lo que reclama."""
    if not mes_cand or len(mes_cand) < 7:
        return False, "candidato sin mes"
    d = _dist_meses(mes_reclamo, mes_cand)
    if d < 0:
        return False, f"candidato de {mes_cand} es posterior al reclamo"
    if d <= 1:
        return True, ""

    col = COL_PLANILLA
    for i in range(1, d):
        m = _mes_menos(mes_reclamo, i)
        debia = _debia(mz, lt, m, col)
        if debia is None:
            return False, f"{d} meses atrás y sin planilla de {m} para probar el arrastre"
        if debia <= TOL:
            return False, f"{d} meses atrás y en {m} no debía nada — no arrastró"
    return True, f"{d} meses atrás, pero arrastró la deuda en todos los meses intermedios"


# ── BLOQUE B — buscar ────────────────────────────────────────────────────────

def _montos_que_cubren(cargos: dict, concepto: str) -> dict:
    """{monto: etiqueta} de las combinaciones de cargos que INCLUYEN el concepto
    disputado.

    vl.subconjuntos() devuelve TODAS las combinaciones, y eso acá genera falsos:
    un pago que cuadra con "mant" (S/3) no puede ser el pago de mes anterior
    (S/20) que el vecino reclama. Medido en la 1a corrida sobre los 29 reclamos
    reales: sin este filtro, Q-9 proponía un pago de S/3 y Z-17 uno de S/34
    (consumo+mant) para disputas de S/20 y S/36 de mes anterior."""
    if cargos.get(concepto, 0.0) <= TOL:
        return {}
    base = cargos[concepto]
    otros = [(n, v) for n, v in cargos.items() if n != concepto and v > TOL]
    out = {round(base, 2): concepto}
    for n in range(1, len(otros) + 1):
        for combo in combinations(otros, n):
            monto = round(base + sum(v for _, v in combo), 2)
            out.setdefault(monto, "+".join([concepto] + [c for c, _ in combo]))
    return out


# cargo interno -> columna de shared/planilla_mes/planilla_YYYY-MM.xlsx
_MAPA_PLANILLA_CARGO = {
    "consumo": "MES_ACTUAL", "mant": "MANTENIMIENTO", "anterior": "MES_ANTERIOR",
    "corte": "CORTE_RECONEXION", "convenio": "CONVENIO", "multa": "MULTA",
    "cuota": "ACUERDOS_ASAMBLEA",
}


def _cargos_planilla(mz: str, lt: str, mes: str) -> dict | None:
    """Lo que un predio DEBÍA en un mes concreto (2_planilla).

    Hace falta porque los pagos que se buscan son de junio/julio y la boleta
    de vl.leer_boletas() es la del ciclo VIGENTE: comparar un pago de julio
    contra los cargos de agosto no dice nada del julio de ese predio."""
    df = rh._cargar_planilla_correcta(mes)
    if df is None:
        return None
    fila = df[(df["MZ"] == _norm(mz)) & (df["LT"] == _norm(lt))]
    if fila.empty:
        return None
    r = fila.iloc[0]
    return {k: _numf(r.get(col)) for k, col in _MAPA_PLANILLA_CARGO.items()}


def _explica_su_propio_pago(otro: str, monto: float, mes: str, boletas: dict) -> bool:
    """Capa 4 de verificar_lotes.py, corregida: el candidato tiene que estar
    IMPAGO — y eso se mide contra lo que ESE lote debía EN EL MES DEL PAGO.

    Si el monto cuadra con alguna combinación de sus propios cargos de ese mes,
    el pago es plausiblemente suyo y proponerlo como error de tipeo es ruido.
    La versión anterior solo comparaba contra el TOTAL de su boleta vigente, que
    es el mes equivocado y además ignora los pagos parciales (alguien que paga
    solo su consumo+mantenimiento)."""
    mz_o, lt_o = otro.split("-", 1)
    cargos = _cargos_planilla(mz_o, lt_o, mes)
    if cargos is None:                       # sin planilla de ese mes: se cae al
        bo = boletas.get(otro)               # total de la boleta vigente
        return bo is not None and abs(monto - bo["total"]) < 0.01
    return vl.subconjuntos(cargos).get(round(monto, 2)) is not None


def _b2_otro_predio(mz: str, lt: str, mes_pago: str, cargo_reclamado: float,
                    boletas: dict, pagos: pd.DataFrame) -> tuple[list[dict], list[dict]]:
    """(multilote, tipeo) — pagos de OTRO predio que podrían ser de este.

    Dos ramas que se distinguen por la evidencia, no por el monto:
      multilote  el mensaje del pago nombra este lote  -> es suyo, sin duda
      tipeo      el lote escrito es confundible con este, el monto cuadra con lo
                 que YO debía ese mes, NO cuadra con lo que EL OTRO debía ese
                 mes, y alcanza para cubrir el arrastre que se reclama.

    Las tres condiciones del tipeo se miden contra la planilla DEL MES DEL PAGO,
    no contra la boleta vigente: los pagos que se buscan son de junio/julio y
    comparar contra los cargos de agosto no dice nada de ese mes."""
    if pagos.empty:
        return [], []
    yo = _clave(mz, lt)
    multilote, tipeo = [], []

    # Lo que YO debía ese mes: si pagué ese mes, el pago tuvo que cuadrar con
    # alguna combinación de esos cargos. Sin planilla del mes se cae a la boleta
    # vigente (mejor que no buscar, pero se dice en el detalle).
    mis_cargos = _cargos_planilla(mz, lt, mes_pago)
    aprox = mis_cargos is None
    if aprox:
        bo_yo = boletas.get(yo)
        mis_cargos = bo_yo["cargos"] if bo_yo else {}
    montos_posibles = vl.subconjuntos(mis_cargos) if mis_cargos else {}

    for _, p in pagos.iterrows():
        otro = _clave(p["MZ"], p["LT"])
        if otro == yo or not p["MZ"]:
            continue

        if _menciona_lote(p["MENSAJE"], mz, lt):
            multilote.append({
                "candidato": otro, "monto": p["MONTO"], "mes": mes_pago,
                "detalle": f"pago de S/{p['MONTO']:,.2f} acreditado a {otro} "
                           f"({p['CANAL']}, {p['ORIGEN']}, {p['FECHA']}) — "
                           f"su mensaje nombra {yo}: \"{p['MENSAJE']}\""})
            continue

        hit = vl.confundible(yo, otro)
        if not hit or hit[1] != 1:        # solo error simple: el doble es ruido
            continue
        como = montos_posibles.get(round(p["MONTO"], 2))
        if como is None:                  # no cuadra con lo que yo debía ese mes
            continue
        if p["MONTO"] < cargo_reclamado - 0.01:
            continue                      # no alcanza para cubrir el arrastre que reclama
        if _explica_su_propio_pago(otro, p["MONTO"], mes_pago, boletas):
            continue                      # capa 4: es plausiblemente el pago de ese lote
        tipeo.append({
            "candidato": otro, "monto": p["MONTO"], "mes": mes_pago,
            "detalle": f"S/{p['MONTO']:,.2f} acreditado a {otro} ({p['CANAL']}, "
                       f"{p['ORIGEN']}, {p['FECHA']}) — {hit[0]}; en {mes_pago} yo debía "
                       f"{como}{' (aprox: sin planilla de ese mes)' if aprox else ''} y "
                       f"{otro} no debía nada que cuadre con ese monto"})
    return multilote, tipeo


def _b1_blancos(mz: str, lt: str, mes_reclamo: str,
                blancos: pd.DataFrame, montos_posibles: dict) -> list[dict]:
    if blancos.empty:
        return []
    out = []
    for _, b in blancos.iterrows():
        como = montos_posibles.get(round(b["MONTO"], 2))
        if como is None:
            continue
        ok, nota = _plausible(mz, lt, b["MES"], mes_reclamo)
        if not ok:
            continue
        out.append({"candidato": f"blanco {b['FUENTE']}", "monto": b["MONTO"], "mes": b["MES"],
                    "detalle": f"blanco de S/{b['MONTO']:,.2f} en {b['MES']} "
                               f"({b['ORIGEN']}) cuadra con mi {como}"
                               + (f" — {nota}" if nota else "")})
    return out


def _b3_excesos(mz: str, lt: str, mes_reclamo: str,
                excesos: pd.DataFrame, montos_posibles: dict) -> list[dict]:
    if excesos.empty:
        return []
    yo = _clave(mz, lt)
    out = []
    for _, e in excesos.iterrows():
        otro = _clave(e["MZ"], e["LT"])
        if otro == yo:
            continue
        hit = vl.confundible(yo, otro)
        if not hit or hit[1] != 1:
            continue
        if montos_posibles.get(round(e["MONTO"], 2)) is None:
            continue
        ok, nota = _plausible(mz, lt, e["MES"], mes_reclamo)
        if not ok:
            continue
        out.append({"candidato": otro, "monto": e["MONTO"], "mes": e["MES"],
                    "detalle": f"exceso sin resolver de S/{e['MONTO']:,.2f} en {otro} "
                               f"({e['NOMBRE']}, {e['MES']}) — {hit[0]}"
                               + (f" — {nota}" if nota else "")})
    return out


def _resolver_candidatos(nombre: str, cands: list[dict]) -> dict | None:
    """Regla de propuesta de verificar_lotes.py: se propone SOLO si queda 1.
    Con 2+ se dice cuántos hay y no se elige — con esa evidencia, elegir es
    inventar con cara de certeza."""
    if not cands:
        return None
    if len(cands) == 1:
        c = cands[0]
        return {"veredicto": nombre, "candidato": c["candidato"],
                "monto": c["monto"], "detalle": c["detalle"]}
    return {"veredicto": nombre, "candidato": f"{len(cands)} candidatos", "monto": None,
            "detalle": " || ".join(c["detalle"] for c in cands[:4])}


# ── Motor ────────────────────────────────────────────────────────────────────

def investigar(r: pd.Series, ctx: dict) -> dict:
    mz, lt = _norm(r.get("MZ")), _norm(r.get("LT"))
    mes_reclamo = _txt(r.get("MES_ANO_ORIGEN")) or ctx["mes"]
    if len(mes_reclamo) < 7:
        mes_reclamo = ctx["mes"]
    # "Ya pagué mes anterior" apunta siempre a un mes concreto: el previo al
    # ciclo en que se registró el reclamo. Los reclamos arrastrados traen su
    # MES_ANO_ORIGEN viejo, así que un reclamo de julio disputa junio, no julio.
    disputado = _mes_menos(mes_reclamo, 1)
    meses_foco = {disputado}

    fila = {"MZ": mz, "LT": lt, "NOMBRE": "", "MES_RECLAMO": mes_reclamo,
            "MES_DISPUTADO": disputado, "RECLAMO": _txt(r.get("RECLAMO")),
            "CARGO_VIGENTE": None, "MESES_SIN_PAGO": "", "PAGO_ARRASTRES_ANTES": "",
            "COBRADOR_QUE_NOMBRA": "—", "EN_FUENTE_MESA": "—", "YAPE_EN_BANCO": "—",
            "BLANCOS": "—", "EXCESOS": "—",
            "OTROS_LOTES": "—", "PAGO_MULTIPLE": "—", "PRECURSORES": "—",
            "CONCLUSION": ""}

    bo = ctx["boletas"].get(_clave(mz, lt))
    fila["NOMBRE"] = bo["nombre"] if bo else ""

    cargo = _cargo_vigente(mz, lt, ctx["boletas"])
    fila["CARGO_VIGENTE"] = cargo
    if cargo is None:
        fila["CONCLUSION"] = "SIN BOLETA — el predio no existe en DATA_boletas, revisar MZ/LT"
        return fila
    if cargo <= TOL:
        fila["CONCLUSION"] = "RESUELTO — la boleta vigente ya no le cobra mes anterior"
        return fila

    tabla = ctx["tabla_de"](mz, lt)

    # ① historial completo: en qué meses pagó y en cuáles no.
    # Se compara contra TODOS los meses esperados, no contra los que la tabla
    # trae: un mes sin ningún movimiento no genera fila, así que recorrer
    # tabla["MES"] lo vuelve invisible. Caso real J-8 — pagó de oct-2025 a
    # jun-2026 y en agosto, julio no tiene fila, y el reporte decía "pagó todos
    # los meses" mientras le cobraban S/16 de arrastre nacido justo en julio.
    pagados = set(tabla[tabla["TOTAL"] > TOL]["MES"].astype(str))
    sin_pago = [m for m in _meses_esperados(mes_reclamo) if m not in pagados]
    fila["MESES_SIN_PAGO"] = " · ".join(sin_pago) or "(pagó todos los meses)"
    fila["PAGO_ARRASTRES_ANTES"] = _historico_mes_anterior(tabla)[0]

    # ⑦ precursores primero: TODOS los eventos que tocan el predio, en cualquier
    # mes. Se calculan acá arriba porque ② los necesita para no reportar como
    # pago perdido algo que un precursor ya se llevó (uso ⑤).
    precursores = _otro_precursor(mz, lt)
    t = _aporte_tanque(mz, lt, set())
    a = _abono_rezagado(mz, lt, set())
    if t:
        precursores.append({"fuente": "aportes_tanque_manuales", "monto": t["monto"],
                            "detalle": t["detalle"][:140]})
    if a:
        precursores.append({"fuente": "abonos_rezagados", "monto": a["monto"],
                            "detalle": a["detalle"][:140]})
    if precursores:
        fila["PRECURSORES"] = " || ".join(p["detalle"] for p in precursores[:4])
    ya_explicados = {round(float(p["monto"]), 2) for p in precursores if p.get("monto")}

    # ② la FUENTE (mesa_N.xlsx), en meses donde el historial dice que no pagó.
    # Se empieza por el cobrador que el vecino nombra, pero se barren todas.
    dicho = _cobrador_mencionado(fila["RECLAMO"], ctx["cobradores"])
    fila["COBRADOR_QUE_NOMBRA"] = dicho or "—"
    no_consolidado = _en_fuente_sin_consolidar(mz, lt, tabla, ctx["mesas"], dicho,
                                                ya_explicados)
    if no_consolidado:
        fila["EN_FUENTE_MESA"] = " || ".join(c["detalle"] for c in no_consolidado[:4])

    # ⑧ si alguien dice "yape" —el vecino en su reclamo o el cobrador en su
    # hoja— hay que ir al reporte del banco, que es la única prueba de que el
    # yape entró a la cuenta de la JASS. Si no está, el veredicto es ese.
    yapes = filas_yape_todos(mz, lt, ctx["mesas"])
    dice_yape = bool(_PATRON_YAPE.search(fila["RECLAMO"])) or bool(yapes)
    if dice_yape:
        if yapes:
            y = yapes[0]
            chk = verificar_una(y["monto"], y["fecha"], mz, lt,
                                 nombre=fila["NOMBRE"])
            # Uso ⑤ otra vez: si un precursor ya se llevó ese monto, que no esté
            # en el banco deja de ser un hallazgo — pasa a ser la explicación.
            # Caso real I-9: S/86 anotados como yape que abonos_rezagados
            # registra como RETENIDOS por Wagner y aplicados en julio; no están
            # en el banco justamente porque fue efectivo mal anotado como yape.
            if round(y["monto"], 2) in ya_explicados and chk["estado"] == "NO_EXISTE":
                chk = {"estado": "EXPLICADO",
                       "detalle": f"no está en el banco, pero un precursor ya registra "
                                  f"esos S/{y['monto']:,.2f} — probablemente fue efectivo "
                                  f"anotado como yape en la hoja"}
            fila["YAPE_EN_BANCO"] = (f"[{chk['estado']}] anotado S/{y['monto']:,.2f} por "
                                      f"{y['cobrador']} ({y['fecha']}): {chk['detalle']}")
        else:
            chk = verificar_una(cargo, "", mz, lt, nombre=fila["NOMBRE"])
            fila["YAPE_EN_BANCO"] = f"[{chk['estado']}] el vecino dice yape: {chk['detalle']}"
        fila["_yape_estado"] = chk["estado"]

    # ③④⑤⑥ la búsqueda de la plata, en TODOS los meses con pool disponible.
    # Llave dura: las combinaciones de cargos de ESTE predio que incluyen el mes
    # anterior (un pago que solo cuadra con "mant" no puede ser el pago que
    # reclama). El monto que el vecino DICE en el texto queda como pista blanda
    # a propósito — muchos reclaman desde la emoción, no desde el monto exacto.
    montos = _montos_que_cubren(bo["cargos"], CARGO_BOLETA)
    multilote, tipeo = [], []
    for mes_pool, pool in ctx["pools"].items():
        ml, tp = _b2_otro_predio(mz, lt, mes_pool, cargo, ctx["boletas"], pool)
        multilote += ml
        tipeo += tp
    blancos = _b1_blancos(mz, lt, mes_reclamo, ctx["blancos"], montos)
    excesos = _b3_excesos(mz, lt, mes_reclamo, ctx["excesos"], montos)

    for col, cands in (("PAGO_MULTIPLE", multilote), ("OTROS_LOTES", tipeo),
                       ("BLANCOS", blancos), ("EXCESOS", excesos)):
        if cands:
            fila[col] = " || ".join(c["detalle"] for c in cands[:4])

    fila["CONCLUSION"] = _concluir(cargo, disputado, tabla, bo["cargos"],
                                    no_consolidado, multilote, tipeo, blancos, excesos,
                                    precursores, fila.pop("_yape_estado", ""))
    return fila


def _concluir(cargo, disputado, tabla, cargos, no_consolidado,
              multilote, tipeo, blancos, excesos, precursores,
              yape_estado: str = "") -> str:
    """Qué hacer con este reclamo, juntando TODO lo que encontraron las 8
    verificaciones. Es el único campo que interpreta; los demás son hallazgos
    crudos para que el supervisor los pueda contradecir."""
    # El yape que no está en el banco manda sobre todo lo demás: si alguien
    # anotó un yape que nunca entró a la cuenta de la JASS, no hay nada que
    # buscar adentro del sistema — el problema está afuera.
    if yape_estado == "NO_EXISTE":
        return ("YAPE QUE NO ENTRO A LA JASS — el cobrador lo anoto pero no hay ninguna "
                "transaccion que calce en la cuenta. Pedir la captura del yape: pudo haber "
                "ido a un telefono personal")
    if no_consolidado:
        return ("PAGO SIN CONSOLIDAR — está en la mesa del cobrador y no en el historial: "
                "el vecino tiene razón y el sistema no lo registró. Revisar por qué")

    candidatos = len(multilote) + len(tipeo) + len(blancos) + len(excesos)
    if candidatos == 1:
        c = (multilote + tipeo + blancos + excesos)[0]
        return f"REVISAR CANDIDATO ÚNICO: {c.get('candidato', '')} — {c['detalle'][:110]}"
    if candidatos > 1:
        return (f"{candidatos} candidatos posibles (ver columnas) — no se elige uno: "
                f"con esta evidencia elegir sería inventar")

    # Sin candidatos: ¿el propio historial explica el arrastre?
    m = _clasificar_mes("", "", disputado, tabla, cargos) if disputado else None
    if m and m["veredicto"] == "CASCADA_FUERA_DE_ORDEN":
        return f"BUG DE APLICACIÓN — {m['detalle'][:130]}"
    if m and m["veredicto"] == "PAGO_PARCIAL":
        return f"PAGÓ PARTE — {m['detalle'][:130]}"

    sin_pago = [x for x in tabla["MES"].astype(str)
                if x not in set(tabla[tabla["TOTAL"] > TOL]["MES"].astype(str))]
    origen = f"el arrastre nace en {' · '.join(sin_pago)}" if sin_pago else ""
    if precursores:
        return (f"PEDIR RECIBO — no hay rastro del pago en ningún lado; {origen}. "
                f"Hay precursores que tocan el predio (ver columna), revisar si explican algo")
    return (f"PEDIR RECIBO O CAPTURA DE YAPE — no aparece en planillas, ni en crudos, "
            f"ni en blancos, ni en excesos, ni en otros lotes, ni hay precursor. {origen}")


def buscar(mes: str) -> pd.DataFrame:
    reclamos = _leer_reclamos(mes, TIPO)
    print(f"  reclamos {TIPO}: {len(reclamos)}")
    if reclamos.empty:
        return pd.DataFrame()

    boletas = vl.leer_boletas()
    print(f"  boletas del ciclo: {len(boletas)}")
    blancos = _blancos_disponibles()
    excesos = _excesos_no_resueltos()
    print(f"  pool: {len(blancos)} blancos sin reclamar - {len(excesos)} excesos sin resolver")

    # Caches: una corrida toca los mismos ciclos y predios muchas veces.
    historicos, eventos, mapa_raw = rh._cargar_historicos(), None, rh._cargar_mapa_raw()
    dfp = pd.read_excel(REPO_DIR / "5_cobranza" / "outputs" / "planilla_cobrado.xlsx",
                        sheet_name="planilla_cobrado", header=1)
    _tablas, _pagos = {}, {}

    def tabla_de(mz, lt):
        nonlocal eventos
        if (mz, lt) not in _tablas:
            if eventos is None:
                import seguimiento_repo as srepo
                eventos = srepo._leer_eventos()
            _tablas[(mz, lt)] = rh.tabla_predio(mz, lt, historicos, eventos, dfp, mapa_raw)
        return _tablas[(mz, lt)]

    # Pools crudos de TODOS los ciclos con archivo disponible (los cerrados de
    # rh.REPOS_CICLO_CERRADO + el activo). Antes se cargaba solo el ciclo del
    # reclamo, así que un pago reclamado de julio se buscaba en agosto: el mes
    # equivocado. Se cargan una vez y se reusan para los 29 reclamos.
    ciclos = sorted(set(rh.REPOS_CICLO_CERRADO) | {mes})
    pools = {c: _pagos_del_ciclo(c) for c in ciclos}
    mesas = {c: pagos_de_mesas(c) for c in ciclos}
    for c in ciclos:
        print(f"    {c}: {len(pools[c])} pagos consolidados · {len(mesas[c])} filas en las mesas (fuente)")

    # Cobradores reales, sacados de las propias mesas: sirven para detectar a
    # quién nombra el vecino en el texto del reclamo. "María García" queda fuera
    # porque es la fila de ejemplo del template, no una persona.
    cobradores = {c for df in mesas.values() if not df.empty
                  for c in df["COBRADOR"].dropna().unique()
                  if c and "MARIA GARCIA" not in _sin_tildes(c).upper()}
    print(f"    cobradores detectados: {' · '.join(sorted(cobradores))}")

    ctx = {"mes": mes, "boletas": boletas, "blancos": blancos, "excesos": excesos,
           "tabla_de": tabla_de, "pools": pools, "mesas": mesas, "cobradores": cobradores}

    filas = []
    for n, (_, r) in enumerate(reclamos.iterrows(), 1):
        filas.append(investigar(r, ctx))
        if n % 10 == 0:
            print(f"    {n}/{len(reclamos)}...")
    return pd.DataFrame(filas)


# ── Salida ───────────────────────────────────────────────────────────────────

_SEC_PREDIO = ("EBF5FB", "1A5276", "F4FAFF")
_SEC_QUE    = ("FEF3C7", "92400E", "FFFBEB")
_SEC_HALLA  = ("E9F7EF", "1E5C3A", "F4FBF7")
_SEC_MANUAL = ("F3E8FF", "5B21B6", "FAF5FF")

_SEC_BUSCA  = ("EBF5FB", "1A5276", "F4FAFF")

# Una columna por verificación: se ve QUÉ se buscó y qué salió en cada lado, no
# un veredicto único que esconde las búsquedas que no se hicieron.
_COLS = [
    ("MZ",                      _SEC_PREDIO,  6, "center", None),
    ("LT",                      _SEC_PREDIO,  7, "center", None),
    ("NOMBRE",                  _SEC_PREDIO, 26, "left",   None),
    ("MES_RECLAMO",             _SEC_QUE,    12, "center", None),
    ("MES_DISPUTADO",           _SEC_QUE,    13, "center", None),
    ("CARGO_VIGENTE",           _SEC_QUE,    13, "right",  '"S/ "#,##0.00'),
    ("RECLAMO",                 _SEC_QUE,    46, "left",   None),
    ("MESES_SIN_PAGO",          _SEC_HALLA,  30, "left",   None),   # ①
    ("PAGO_ARRASTRES_ANTES",    _SEC_HALLA,  17, "center", None),   # ①
    ("COBRADOR_QUE_NOMBRA",     _SEC_BUSCA,  16, "center", None),   # ②
    ("EN_FUENTE_MESA",          _SEC_BUSCA,  46, "left",   None),   # ②
    ("YAPE_EN_BANCO",           _SEC_BUSCA,  52, "left",   None),   # ⑧
    ("BLANCOS",                 _SEC_BUSCA,  36, "left",   None),   # ③
    ("EXCESOS",                 _SEC_BUSCA,  36, "left",   None),   # ④
    ("OTROS_LOTES",             _SEC_BUSCA,  40, "left",   None),   # ⑤
    ("PAGO_MULTIPLE",           _SEC_BUSCA,  36, "left",   None),   # ⑥
    ("PRECURSORES",             _SEC_BUSCA,  44, "left",   None),   # ⑦
    ("CONCLUSION",              _SEC_MANUAL, 60, "left",   None),
    ("RESOLUCION",              _SEC_MANUAL, 34, "left",   None),
]

_SECCIONES = [("¿Cuál es el predio?", "MZ", "NOMBRE"),
              ("¿Qué reclama?", "MES_RECLAMO", "RECLAMO"),
              ("① su historial", "MESES_SIN_PAGO", "PAGO_ARRASTRES_ANTES"),
              ("②-⑦ ¿dónde está la plata?", "COBRADOR_QUE_NOMBRA", "PRECURSORES"),
              ("Qué hacer", "CONCLUSION", "RESOLUCION")]

# El color de CONCLUSION se elige por su primera palabra: dice de un vistazo si
# hay algo que revisar, si hay un pago perdido, o si toca pedir comprobante.
_COLOR_CONCLUSION = {
    "YAPE":     ("FEE2E2", "991B1B"),   # yape que no entro a la cuenta
    "PAGO":     ("FEE2E2", "991B1B"),   # pago sin consolidar — el mas grave
    "BUG":      ("FEE2E2", "991B1B"),
    "REVISAR":  ("D5F5E3", "1E5C3A"),   # hay un candidato concreto
    "PAGÓ":     ("FEF9E7", "7D6608"),
    "PEDIR":    ("F1F3F5", "868E96"),   # no hay rastro: trabajo de campo
    "RESUELTO": ("F1F3F5", "495057"),
    "SIN":      ("F1F3F5", "868E96"),
}


def _fill(hex6):
    return PatternFill("solid", fgColor=f"FF{hex6}")


def escribir(df: pd.DataFrame, mes: str) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Busqueda"
    idx = {c[0]: i + 1 for i, c in enumerate(_COLS)}

    for label, ini, fin in _SECCIONES:
        c1, c2 = idx[ini], idx[fin]
        sec = _COLS[c1 - 1][1]
        if c1 != c2:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        cell = ws.cell(row=1, column=c1, value=label)
        cell.fill, cell.font = _fill(sec[0]), Font(color=f"FF{sec[1]}", bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18

    for i, (nombre, sec, ancho, _a, _f) in enumerate(_COLS, start=1):
        cell = ws.cell(row=2, column=i, value=nombre)
        cell.fill, cell.font = _fill(sec[0]), Font(color=f"FF{sec[1]}", bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "C3"

    for ri, (_, row) in enumerate(df.iterrows(), start=3):
        for ci, (nombre, sec, _w, align, fmt) in enumerate(_COLS, start=1):
            val = row.get(nombre, "")
            if val is None or (isinstance(val, float) and pd.isna(val)):
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            bg, fg = sec[2], sec[1]
            if nombre == "CONCLUSION":
                bg, fg = _COLOR_CONCLUSION.get(str(val).split(" ")[0], (sec[2], sec[1]))
            elif str(val) not in ("", "—") and nombre in (
                    "EN_FUENTE_MESA", "YAPE_EN_BANCO", "BLANCOS", "EXCESOS",
                    "OTROS_LOTES", "PAGO_MULTIPLE"):
                bg, fg = "D5F5E3", "1E5C3A"      # se encontró algo: resaltar
            cell.fill = _fill(bg)
            cell.font = Font(color=f"FF{fg}", size=10,
                             bold=(nombre == "CONCLUSION"))
            cell.alignment = Alignment(horizontal=align, vertical="top",
                                       wrap_text=(nombre not in ("MZ", "LT", "CARGO_VIGENTE")))
            if fmt:
                cell.number_format = fmt

    out = BASE_DIR / "outputs" / f"busqueda_pago_{TIPO}_{mes}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(out)
    except PermissionError:
        # El archivo está abierto en Excel. Perder la corrida entera (~1 min de
        # lectura de 3 repos + 3 pools) por eso es absurdo: se guarda al lado y
        # se avisa, en vez de tirar el trabajo.
        alt = out.with_name(f"{out.stem}_{datetime.now():%H%M%S}{out.suffix}")
        wb.save(alt)
        print(f"\n  AVISO: {out.name} esta abierto en Excel — se guardo como {alt.name}")
        return alt
    return out


# ============================================================================
# REFERENCIA DE PAGO -- absorbido de reporte_referencias_pago.py (2026-08-18)
#
# "De donde vino cada pago": yape con fecha/hora exacta, o efectivo, y su
# monto -- lo que reporte_historico.py imprime debajo del historial mensual.
# Vivia en un archivo aparte que reporte_historico.py y verificar_yape.py
# importaban -- se junta acá porque buscar_pago.py ya reusaba 2 de sus
# funciones (_cargar_pagos_yape_crudo / _cargar_pagos_efectivo_crudo) y
# quedar repartido en 3 archivos que se importan entre si es justo el patron
# fragil que se rompio el 2026-08-17 cuando uno de los 3 se movio de carpeta.
#
# Fuentes por rango de fecha:
#   - Oct 2025 -> Ene 2026: hoja "Cobranza" (Estado=c + Medio) decide yape/efectivo;
#     si es yape, la hoja "Reporte" del mismo archivo trae fecha/hora exacta.
#   - Feb 2026 -> May 2026: mismo criterio, pero el monto de efectivo sale de la
#     hoja "Efectivo" dedicada (mas preciso que el total de Cobranza).
#   - Jun 2026 -> Jul 2026: no hay archivo que preserve fecha/hora por transaccion
#     para la mayoria de pagos yape (motor_matching no lo persiste) -- se muestra
#     el total del ciclo desde planilla_cobrado.xlsx, sin desglose de hora.
# ============================================================================

HIST_DIR_REFS = BASE_DIR.parent / "obligaciones" / "inputs" / "planillas anteriores"
_ARCHIVOS_HISTORICOS_REFS = rh._ARCHIVOS_HISTORICOS

_AZUL = (26/255, 82/255, 118/255)
_AZUL_BG = (235/255, 245/255, 251/255)
_GRIS = (0.42, 0.45, 0.5)
_NEGRO = (0.12, 0.16, 0.22)
_VERDE = (0.02, 0.37, 0.27)
_ZEBRA = (243/255, 244/255, 246/255)


def _norm_col(s: str) -> str:
    """Normaliza NOMBRE DE COLUMNA (sin tildes/mayus) para _col() -- distinto
    de _norm() (arriba, importado de verificar_yape.py), que normaliza
    VALORES de MZ/LT/nombre. Mismo nombre en los dos archivos originales,
    renombrado acá para que puedan convivir en uno solo."""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return s.lower().replace("\n", " ").strip()


def _col(df: pd.DataFrame, *candidatos: str):
    normed = {}
    for c in df.columns:
        k = _norm_col(c)
        if k not in normed:
            normed[k] = c
    for cand in candidatos:
        cn = _norm_col(cand)
        if cn in normed:
            return normed[cn]
    for cand in candidatos:
        cn = _norm_col(cand)
        for k, real in normed.items():
            if cn in k:
                return real
    return None


_cache_hojas_refs: dict[str, dict] = {}
_avisados_refs: set[str] = set()


def _avisar_falta_refs(p: Path, que: str) -> None:
    """Un archivo de ciclo cerrado que no esta donde dice el path se avisa UNA
    vez por corrida (esto se llama por predio). Fallaba en silencio hasta el
    04/08/2026 y las filas sin origen se leian como pagos fantasma."""
    if str(p) in _avisados_refs:
        return
    _avisados_refs.add(str(p))
    print(f"  AVISO: falta {que} -> {p}")


def _cargar_hojas_historicas(archivo: str) -> dict:
    """Cachea Cobranza/Reporte/Efectivo de UN archivo historico a nivel modulo
    -- sin esto, una corrida por lote reabre los mismos 8 excels una vez POR
    PREDIO (~500 predios x 8 archivos = ~4000 lecturas en vez de 8)."""
    if archivo in _cache_hojas_refs:
        return _cache_hojas_refs[archivo]
    p = HIST_DIR_REFS / archivo
    hojas = {}
    if p.exists():
        hojas["Cobranza"] = pd.read_excel(p, sheet_name="Cobranza", header=0)
        try:
            hojas["Reporte"] = pd.read_excel(p, sheet_name="Reporte", header=0)
        except Exception:
            hojas["Reporte"] = None
    _cache_hojas_refs[archivo] = hojas
    return hojas


def referencias_pago(mz: str, lt: str, tabla: pd.DataFrame | None = None,
                     incluir_overlays: bool = True) -> list[dict]:
    """Lista de {MES, MEDIO, FECHA_HORA, MONTO} -- una entrada por pago detectado.

    El MONTO de cada mes SIEMPRE es el TOTAL que ya calculo _fila_historica()
    (mismo numero que aparece en la tabla de arriba) -- no una relectura aparte
    de una columna suelta. Hasta mayo 2026 ese total ya incluye mantenimiento
    arrastrado y exceso devuelto via "mes anterior" (asi se manejaba antes);
    la hoja Reporte/Efectivo solo aporta DONDE vino la plata (yape con hora,
    o efectivo), no el monto -- eso evita que un monto parcial de una sola
    transaccion descuadre contra el total real del mes."""
    out = []

    mapa_raw = rh._cargar_mapa_raw()
    mz_hist, lt_hist = mapa_raw.get((mz, lt), (mz, lt))

    for archivo, mes in _ARCHIVOS_HISTORICOS_REFS:
        hojas = _cargar_hojas_historicas(archivo)
        dfc = hojas.get("Cobranza")
        if dfc is None:
            continue
        cmz, clt, cest, cmed = _col(dfc, "mz"), _col(dfc, "lt", "lote"), _col(dfc, "estado"), _col(dfc, "medio")
        sub = dfc[(dfc[cmz].astype(str).str.strip() == mz_hist) & (dfc[clt].astype(str).str.strip() == lt_hist)]
        if sub.empty or str(sub.iloc[0][cest]).strip().lower() != "c":
            continue

        fila_mes = None
        if tabla is not None:
            m = tabla[tabla["MES"] == mes]
            if not m.empty:
                fila_mes = m.iloc[0]
        monto_mes = float(fila_mes["TOTAL"]) if fila_mes is not None else None
        if monto_mes is None or monto_mes <= 0.005:
            continue

        medio = str(sub.iloc[0][cmed]).strip().lower()
        if medio == "y":
            dfr = hojas.get("Reporte")
            if dfr is None:
                out.append({"MES": mes, "MEDIO": "YAPE", "FECHA_HORA": "sin hoja Reporte", "MONTO": monto_mes})
                continue
            cmzr, cltr = _col(dfr, "mz"), _col(dfr, "lote", "lt")
            ctr, cfr = _col(dfr, "tipo de transaccion"), _col(dfr, "fecha de operacion")
            subr = dfr[(dfr[cmzr].astype(str).str.strip() == mz_hist) & (dfr[cltr].astype(str).str.strip() == lt_hist)]
            subr = subr[subr[ctr].astype(str).str.upper().str.startswith("TE PAG", na=False)]
            fecha = str(subr.iloc[0][cfr]) if not subr.empty else "sin match en Reporte"
            out.append({"MES": mes, "MEDIO": "YAPE", "FECHA_HORA": fecha, "MONTO": monto_mes})
        else:
            out.append({"MES": mes, "MEDIO": "EFECTIVO", "FECHA_HORA": "--", "MONTO": monto_mes})

    # Cada ciclo posterior a mayo: total real pagado (yape+efectivo) desde el
    # planilla_cobrado DE SU PROPIO CICLO (los cerrados en su repo congelado,
    # el vigente en el repo activo -- ver _ciclos_recientes()) vs. lo que la
    # tabla de arriba muestra aplicado a conceptos -- desde junio 2026 el
    # exceso se RETIENE (no se aplica ni se muestra solo, ver nota del
    # reporte), asi que puede haber una diferencia real y legitima. Se
    # muestra como linea aparte, nunca mezclada.
    _PLANILLAS_RECIENTES = _ciclos_recientes()
    for mes_ciclo, f in _PLANILLAS_RECIENTES:
        # Los overlays entran a la cascada del ciclo pero NO viven en
        # planilla_cobrado -- van primero para que aparezcan aunque el ciclo
        # no tenga fila de yape/efectivo.
        if incluir_overlays:
            out.extend(_overlays_de_plata(mz, lt, mes_ciclo))
        if not f.exists():
            _avisar_falta_refs(f, f"referencias de pago de {mes_ciclo}")
            continue
        dfp = pd.read_excel(f, sheet_name="planilla_cobrado", header=1)
        subp = dfp[(dfp["MZ"].astype(str).str.strip() == mz) & (dfp["LT"].astype(str).str.strip() == lt)]
        if subp.empty:
            continue
        r = subp.iloc[0]
        yape = rh._numf(r.get("MONTO_YAPE"))
        efectivo = rh._numf(r.get("MONTO_EFECTIVO"))
        total_real = yape + efectivo
        total_aplicado = 0.0
        if tabla is not None:
            m = tabla[tabla["MES"] == mes_ciclo]
            if not m.empty:
                total_aplicado = float(m.iloc[0]["TOTAL"])
        if yape > 0.005:
            # pagos_yape_tepago.xlsx directo -- ya viene resuelto 1 a 1 por
            # MZ/LOTE, no hace falta adivinar nada. _yape_fecha_hora (cruce
            # de nombre por prefijo contra el banco crudo) se sacó del
            # camino: cuando el origen registrado en maestro_yape.xlsx es
            # generico (ej. J-1: origen = "Janet Villanueva", la tesorera --
            # su nombre aparece en decenas de transacciones ajenas) pegaba
            # fechas de pagos de otros predios (bug encontrado 01/08/2026).
            fecha_yape = _fecha_yape_cruda(mz, lt, mes_ciclo)
            out.append({"MES": mes_ciclo, "MEDIO": "YAPE",
                        "FECHA_HORA": fecha_yape or "sin match en pagos_yape_tepago.xlsx (total del ciclo)",
                        "MONTO": yape})
        if efectivo > 0.005:
            dia, cobrador = _cobrador_efectivo(mz, lt, mes_ciclo)
            detalle = f"{dia} · {cobrador}" if cobrador else "sin detalle de transaccion (total del ciclo)"
            out.append({"MES": mes_ciclo, "MEDIO": "EFECTIVO", "FECHA_HORA": detalle, "MONTO": efectivo})
        # Si total_real y total_aplicado no coinciden (en cualquier sentido),
        # no se afirma de donde viene ni a donde fue la diferencia -- esto
        # solia mostrar una nota especulando el origen (blancos/abonos/
        # condonacion/convenio instalacion) y generaba confusion, porque esa
        # explicacion no esta confirmada (a veces es solo lo que la secretaria
        # declaro despues). Se deja en blanco a proposito: se muestra
        # unicamente el pago real encontrado, sin inventar el resto.

    return out


_ABONOS_REZAGADOS_REFS = BASE_DIR.parent / "shared" / "abonos_rezagados.xlsx"
_BLANCOS_EFECTIVO_REFS = BASE_DIR.parent / "shared" / "blancos_efectivo.xlsx"
_cache_overlays_refs: dict[str, pd.DataFrame] = {}


def _cargar_overlay(p: Path) -> pd.DataFrame:
    if str(p) not in _cache_overlays_refs:
        if not p.exists():
            _avisar_falta_refs(p, "overlay de pagos")
            _cache_overlays_refs[str(p)] = pd.DataFrame()
        else:
            df = pd.read_excel(p, header=1)
            df.columns = [str(c).strip().upper() for c in df.columns]
            _cache_overlays_refs[str(p)] = df
    return _cache_overlays_refs[str(p)]


def _overlays_de_plata(mz: str, lt: str, mes_ciclo: str) -> list[dict]:
    """Plata real que entra a la cascada del ciclo pero NO esta en
    MONTO_YAPE/MONTO_EFECTIVO de planilla_cobrado: abonos que el cobrador
    retuvo y se regularizaron despues, y efectivo que habia entrado como
    blanco. Sin estas lineas el pago que salda la deuda queda sin origen
    visible y se lee como pago fantasma (caso encontrado 05/08/2026: los 4
    abonos retenidos por Wagner en julio -- S-5, D-16, D1-6, L-4)."""
    out = []
    for p, medio, quien_cols in ((_ABONOS_REZAGADOS_REFS, "ABONO REZ.", ("RETENIDO_POR", "CANAL_ORIGEN")),
                                  (_BLANCOS_EFECTIVO_REFS, "BLANCO EF.", ("ORIGEN", "CANAL"))):
        df = _cargar_overlay(p)
        if df.empty or "MES_ANO_APLICA" not in df.columns:
            continue
        sub = df[(df["MZ"].astype(str).str.strip() == mz) &
                 (df["LT"].apply(_norm_lote) == _norm_lote(lt)) &
                 (df["MES_ANO_APLICA"].astype(str).str.strip() == mes_ciclo)]
        for _, r in sub.iterrows():
            quien = " · ".join(str(r[c]).strip() for c in quien_cols
                               if c in df.columns and pd.notna(r.get(c)))
            fecha_raw = r.get("FECHA_REAL", "")
            fecha = str(fecha_raw).strip() if pd.notna(fecha_raw) else "sin fecha registrada"
            out.append({"MES": mes_ciclo, "MEDIO": medio,
                        "FECHA_HORA": f"{fecha} · del ciclo {r.get('MES_CICLO')} · {quien}",
                        "MONTO": float(r["MONTO"])})
    return out


def _repo_de_ciclo(mes_ano: str) -> Path:
    """Repo donde vive el output crudo de ese ciclo: el suyo propio congelado
    si ya cerró (rh.REPOS_CICLO_CERRADO), o el repo activo si es el ciclo
    vigente -- así ningún ciclo depende de hardcodear "cuál mes es hoy"."""
    return rh.REPOS_CICLO_CERRADO.get(mes_ano, BASE_DIR.parent)


def _ciclos_recientes() -> list[tuple[str, Path]]:
    """(mes_ano, ruta a planilla_cobrado) de cada ciclo posterior a mayo: los
    cerrados de rh.REPOS_CICLO_CERRADO + el vigente de shared/ciclo_activo.json
    (si no es ya uno de los cerrados). Se recalcula en cada llamada -- barato,
    y así recoge el ciclo activo sin reiniciar el proceso cuando el mes rueda."""
    ciclos = {m: rh._planilla_ciclo_cerrado(m) for m in rh.REPOS_CICLO_CERRADO}
    activo = ciclo.activo(default=None)
    if activo and activo not in ciclos:
        ciclos[activo] = BASE_DIR.parent / "5_cobranza" / "outputs" / "planilla_cobrado.xlsx"
    return sorted(ciclos.items())


def _pagos_efectivo_crudo_path(mes_ciclo: str) -> Path:
    """Repo propio si el ciclo cerró, repo activo si es el vigente (ver
    _repo_de_ciclo). legacy_sin_periodo=True porque el ciclo activo todavía
    escribe pagos_efectivo.xlsx sin periodo además del nombre canónico
    (migra en la tanda B) y los repos cerrados de junio/julio quedaron con
    esa convención -- ciclo.resolver prueba el nombre con periodo primero,
    así que no cambia nada para los ciclos que ya migraron."""
    return ciclo.resolver(_repo_de_ciclo(mes_ciclo) / "4_pagos" / "efectivo" / "outputs",
                          "pagos_efectivo", mes_ciclo, legacy_sin_periodo=True)


_cache_pagos_efectivo_refs: dict[str, pd.DataFrame] = {}


def _cargar_pagos_efectivo_crudo(mes_ciclo: str) -> pd.DataFrame | None:
    """FECHA real por transaccion -- trazabilidad_{mes}.xlsx (hoja
    solo_un_cobrador) tiene FECHA_COBRO vacio para TODO julio (bug de
    construccion, 221/221 filas verificado 01/08/2026) aunque el archivo
    crudo que la alimenta si trae fecha. Se usa como fuente de la fecha en
    vez de trazabilidad; trazabilidad sigue siendo la fuente del COBRADOR
    resuelto (para pagos divididos entre mesas)."""
    if mes_ciclo not in _cache_pagos_efectivo_refs:
        p = _pagos_efectivo_crudo_path(mes_ciclo)
        if not p.exists():
            _avisar_falta_refs(p, f"pagos_efectivo de {mes_ciclo}")
            _cache_pagos_efectivo_refs[mes_ciclo] = pd.DataFrame()
        else:
            df = pd.read_excel(p, header=1)
            df.columns = [str(c).strip() for c in df.columns]
            _cache_pagos_efectivo_refs[mes_ciclo] = df
    return _cache_pagos_efectivo_refs[mes_ciclo]


def _fecha_efectivo_cruda(mz: str, lt: str, mes_ciclo: str, cobrador: str) -> str | None:
    df = _cargar_pagos_efectivo_crudo(mes_ciclo)
    if df is None or df.empty or "FECHA" not in df.columns:
        return None
    sub = df[(df["MZ"].astype(str).str.strip() == mz) & (df["LT"].astype(str).str.strip() == str(lt))]
    if sub.empty:
        return None
    if cobrador and "COBRADOR" in sub.columns:
        con_cobrador = sub[sub["COBRADOR"].astype(str).str.strip() == cobrador]
        if not con_cobrador.empty:
            sub = con_cobrador
    fechas = sub["FECHA"].dropna()
    if fechas.empty:
        return None
    return " · ".join(sorted({str(f) for f in fechas}))


def _cobrador_fecha_efectivo_crudo(mz: str, lt: str, mes_ciclo: str) -> tuple[str, str] | None:
    """Respaldo cuando el predio NO aparece en trazabilidad_{mes}.xlsx en
    absoluto (135 de 357 casos de julio, verificado 01/08/2026 -- gap
    distinto al de FECHA_COBRO vacio: acá falta la fila entera). Se toma
    directo de pagos_efectivo.xlsx filtrando ESTADO=solo_un_cobrador (mismo
    criterio que usaría trazabilidad si el predio estuviera ahí)."""
    df = _cargar_pagos_efectivo_crudo(mes_ciclo)
    if df is None or df.empty or "ESTADO" not in df.columns:
        return None
    sub = df[(df["MZ"].astype(str).str.strip() == mz) & (df["LT"].astype(str).str.strip() == str(lt)) &
              (df["ESTADO"] == "solo_un_cobrador")]
    if sub.empty:
        return None
    r = sub.iloc[0]
    fecha = r.get("FECHA")
    dia = str(fecha) if pd.notna(fecha) else "(fecha no registrada)"
    cobrador = str(r.get("COBRADOR", "")).strip()
    return dia, cobrador


def _pagos_yape_crudo_path(mes_ciclo: str) -> Path:
    """Mismo criterio que _pagos_efectivo_crudo_path: repo propio si el ciclo
    cerró, repo activo si es el vigente."""
    return ciclo.resolver(_repo_de_ciclo(mes_ciclo) / "4_pagos" / "yape" / "motor_matching" / "outputs",
                          "pagos_yape_tepago", mes_ciclo, legacy_sin_periodo=True)


_cache_pagos_yape_crudo_refs: dict[str, pd.DataFrame] = {}


def _cargar_pagos_yape_crudo(mes_ciclo: str) -> pd.DataFrame | None:
    """FECHA real por transaccion, directo de pagos_yape_tepago.xlsx (lo que
    alimenta maestro_yape.xlsx) -- respaldo cuando _yape_fecha_hora() no
    encuentra match cruzando maestro_yape contra el banco crudo (cruce por
    nombre truncado + fecha, se cae seguido). Un paso mas atras que ese
    cruce: esta fecha ya viene resuelta contra MZ/LOTE, sin necesidad de
    adivinar el origen."""
    if mes_ciclo not in _cache_pagos_yape_crudo_refs:
        p = _pagos_yape_crudo_path(mes_ciclo)
        if not p.exists():
            _avisar_falta_refs(p, f"pagos_yape_tepago de {mes_ciclo}")
            _cache_pagos_yape_crudo_refs[mes_ciclo] = pd.DataFrame()
        else:
            df = pd.read_excel(p, header=1)
            df.columns = [str(c).strip() for c in df.columns]
            _cache_pagos_yape_crudo_refs[mes_ciclo] = df
    return _cache_pagos_yape_crudo_refs[mes_ciclo]


def _norm_lote(v) -> str:
    """pagos_yape_tepago.xlsx guarda LOTE como numero (4.0) para lotes sin
    letra -- comparar como texto plano ("4") sin esto nunca matchea, aunque
    la fila SI exista (bug encontrado 01/08/2026, caso I-4: el dato estaba,
    la comparacion de texto lo escondia)."""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _fecha_yape_cruda(mz: str, lt: str, mes_ciclo: str) -> str | None:
    df = _cargar_pagos_yape_crudo(mes_ciclo)
    if df is None or df.empty or "FECHA" not in df.columns or "LOTE" not in df.columns:
        return None
    sub = df[(df["MZ"].astype(str).str.strip() == mz) & (df["LOTE"].apply(_norm_lote) == _norm_lote(lt))]
    fechas = sub["FECHA"].dropna()
    if fechas.empty:
        return None
    return " · ".join(sorted({str(f) for f in fechas}))


def _cobrador_efectivo(mz: str, lt: str, mes_ciclo: str) -> tuple[str, str]:
    """Dia de cobro y cobrador para un pago en efectivo de junio/julio 2026.
    Fuente primaria: 4_pagos/efectivo/trazabilidad/trazabilidad_{mes}.xlsx
    (hoja solo_un_cobrador). Dos gaps verificados en julio (01/08/2026), los
    dos con respaldo en pagos_efectivo.xlsx (el crudo que alimenta la
    trazabilidad):
      1. El predio SI esta en trazabilidad pero FECHA_COBRO vacio (221/221
         filas de julio) -> _fecha_efectivo_cruda().
      2. El predio NO esta en trazabilidad en absoluto (135/357 de julio,
         gap distinto) -> _cobrador_fecha_efectivo_crudo().
    Un solo fix acá cubre cualquier reporte que llame referencias_pago()."""
    f = BASE_DIR.parent / "4_pagos" / "efectivo" / "trazabilidad" / f"trazabilidad_{mes_ciclo}.xlsx"
    df = None
    if f.exists():
        try:
            df = pd.read_excel(f, sheet_name="solo_un_cobrador", header=1)
        except Exception:
            df = None

    sub = None
    if df is not None:
        sub = df[(df["MZ"].astype(str).str.strip() == mz) & (df["LT"].astype(str).str.strip() == str(lt))]

    if sub is None or sub.empty:
        crudo = _cobrador_fecha_efectivo_crudo(mz, lt, mes_ciclo)
        return crudo if crudo else ("", "")

    r = sub.iloc[0]
    cobrador = str(r.get("COBRADOR", "")).strip()
    fecha = r.get("FECHA_COBRO")
    if pd.notna(fecha):
        dia = str(fecha)
    else:
        dia = _fecha_efectivo_cruda(mz, lt, mes_ciclo, cobrador) or "(fecha no registrada)"
    return dia, cobrador


def _dibujar_tabla_referencias(page, x: float, y: float, w: float, refs: list[dict]) -> float:
    headers = ["Mes", "Medio", "Fecha / hora", "Monto"]
    anchos = [70, 80, w - 70 - 80 - 110, 110]
    rh_row = 16

    page.insert_text((x, y), "Referencia de pago (de donde vino cada pago)", fontsize=9, fontname="hebo", color=_AZUL)
    y += 16

    page.draw_rect(fitz.Rect(x, y, x + w, y + rh_row), fill=_AZUL_BG, color=None)
    cx = x
    for h, cw in zip(headers, anchos):
        page.insert_text((cx + 4, y + rh_row - 5), h, fontsize=8, fontname="hebo", color=_AZUL)
        cx += cw
    y += rh_row

    for n, r in enumerate(refs):
        if n % 2 == 1:
            page.draw_rect(fitz.Rect(x, y, x + w, y + rh_row), fill=_ZEBRA, color=None)
        cx = x
        page.insert_text((cx + 4, y + rh_row - 5), str(r["MES"]), fontsize=8, fontname="hebo", color=_NEGRO)
        cx += anchos[0]
        es_nota = r["MEDIO"] == "(nota)"
        color_medio = _GRIS if es_nota else (_VERDE if r["MEDIO"] == "YAPE" else _GRIS)
        page.insert_text((cx + 4, y + rh_row - 5), r["MEDIO"], fontsize=8, fontname="hebo", color=color_medio)
        cx += anchos[1]
        page.insert_text((cx + 4, y + rh_row - 5), str(r["FECHA_HORA"]), fontsize=7.5 if es_nota else 8,
                          fontname="helv", color=_GRIS if es_nota else _NEGRO)
        cx += anchos[2]
        if not es_nota:
            texto_monto = f"S/ {r['MONTO']:,.2f}"
            tw = fitz.get_text_length(texto_monto, fontname="hebo", fontsize=8)
            page.insert_text((cx + anchos[-1] - 6 - tw, y + rh_row - 5), texto_monto, fontsize=8, fontname="hebo", color=_NEGRO)
        y += rh_row

    return y


def verificar_predio(mz: str, lt: str, tabla: pd.DataFrame | None = None, refs: list[dict] | None = None) -> list[dict]:
    """Compara, mes a mes, el TOTAL de la tabla de historial contra la suma de
    referencias de pago de ese mes. Devuelve solo los meses que NO cuadran
    (lista vacia = todo cuadra)."""
    if tabla is None:
        tabla = rh.tabla_predio(mz, lt)
        tabla = rh.corregir_tabla_por_redirects(tabla, mz, lt, rh._cargar_redirects())
    if refs is None:
        refs = referencias_pago(mz, lt)

    ref_por_mes: dict[str, float] = {}
    for r in refs:
        ref_por_mes[r["MES"]] = ref_por_mes.get(r["MES"], 0.0) + r["MONTO"]

    _MESES_RECIENTES = {"2026-06", "2026-07"}

    problemas = []
    for _, row in tabla.iterrows():
        mes = row["MES"]
        total_tabla = float(row["TOTAL"])
        total_ref = ref_por_mes.get(mes, 0.0)
        # jun-jul: no se afirma el origen de una diferencia (ver
        # referencias_pago) -- un descuadre en cualquier sentido ahi es
        # esperado, no se marca como "problema" del reporte.
        if mes in _MESES_RECIENTES:
            continue
        if abs(total_tabla - total_ref) > 0.5:
            problemas.append({"MES": mes, "TOTAL_TABLA": total_tabla, "TOTAL_REFERENCIA": total_ref,
                               "DIFERENCIA": round(total_tabla - total_ref, 2)})
    return problemas


def main(mes: str) -> None:
    print(f"=== buscar_pago — {TIPO} · ciclo {mes} ===")
    df = buscar(mes)
    if df.empty:
        print("  sin reclamos de ese tipo — nada que buscar")
        return

    # Cuántas veces encontró algo cada verificación. Un 0 acá no es un bug: es
    # el dato de que en esa fuente no está la plata, y es lo que justifica
    # terminar pidiendo el comprobante.
    print("\n  qué encontró cada búsqueda:")
    for col, etiqueta in (("EN_FUENTE_MESA", "② pago en la FUENTE (mesa_N) sin consolidar"),
                          ("YAPE_EN_BANCO", "⑧ se verifico el yape contra el banco"),
                          ("BLANCOS", "③ blancos que calzan"),
                          ("EXCESOS", "④ excesos que calzan"),
                          ("OTROS_LOTES", "⑤ pagos en lotes confundibles"),
                          ("PAGO_MULTIPLE", "⑥ pagos que nombran el lote"),
                          ("PRECURSORES", "⑦ precursores que tocan el predio")):
        n = int((df[col].astype(str) != "—").sum())
        print(f"    {n:>3}  {etiqueta}")

    print("\n  conclusiones:")
    for c, n in df["CONCLUSION"].str.split(" —").str[0].value_counts().items():
        print(f"    {n:>3}  {str(c)[:78]}")

    out = escribir(df, mes)
    print(f"\n  -> {out}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(
        description=f"Busca evidencia del pago que afirma un reclamo de {TIPO}")
    ap.add_argument("--mes", default=None, help="Ciclo a auditar (YYYY-MM). Default: ciclo activo")
    args = ap.parse_args()
    mes = args.mes or ciclo.activo(default=None)
    if not mes:
        ap.error("--mes requerido (no hay ciclo activo en shared/ciclo_activo.json)")
    main(mes)
