"""
test_reclamos.py — Tests sintéticos para main.py (ciclo de reclamos).

Cobertura (patrón 3.6c metodología — estado mínimo sintético + assert con mensaje claro):

  1. test_detecta_solo_reclamos      — filtra COMENTARIO ≈ "reclamo"
  2. test_caso_vacio                 — sin reclamos → vista vacía sin error
  3. test_preserva_trabajo_manual    — re-corrida del mes conserva RECLAMO/ESTADO/FECHA_RESOLUCION
  4. test_cierra_a_trazabilidad      — RESUELTO/RECHAZADO salen de vista, entran a trazabilidad
  5. test_arrastra_de_mes_anterior   — PENDIENTE sin match en mes actual se arrastra (con MES_ANO_ORIGEN preservado)
  6. test_clave_distingue_mesas      — mismo MZ+LT en 2 mesas distintas se trata como filas separadas
  7. test_trazabilidad_acumula       — corridas sucesivas no borran filas viejas de trazabilidad
  8. test_mergea_correccion          — RESUELTO toma CAMPO/VALOR_*; RECHAZADO los deja vacíos
  9. test_sync_estado_desde_resolucion — ESTADO solo en resolucion → main.py sincroniza y cierra
"""

import shutil
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

THIS = Path(__file__).resolve()
sys.path.insert(0, str(THIS.parent.parent))

import main as reclamos  # noqa: E402  (módulo se llama main.py; alias por legibilidad)

TEST_ROOT = THIS.parent / "_tmp_reclamos"


# ── Helpers ──────────────────────────────────────────────────────────────────

def _setup():
    if TEST_ROOT.exists():
        shutil.rmtree(TEST_ROOT)
    (TEST_ROOT / "outputs").mkdir(parents=True)
    (TEST_ROOT / "trazabilidad").mkdir(parents=True)
    reclamos.OUTPUTS_DIR = TEST_ROOT / "outputs"
    reclamos.TRAZAB_DIR  = TEST_ROOT / "trazabilidad"
    # En producción PAGOS_EFECTIVO_PATH apunta a 4_pagos/efectivo/outputs/;
    # en tests redirigimos a nuestra OUTPUTS_DIR sintética.
    reclamos.PAGOS_EFECTIVO_PATH = TEST_ROOT / "outputs" / "pagos_efectivo.xlsx"


def _teardown():
    if TEST_ROOT.exists():
        shutil.rmtree(TEST_ROOT)


_COLS_PAGOS = ["MZ", "LT", "MONTO", "FECHA", "MESA",
               "ESTADO", "COBRADOR", "COMENTARIO", "CICLO_CORRECCION"]


def _crear_pagos(filas: list):
    """
    Crea outputs/pagos_efectivo.xlsx con la estructura real:
    fila 1 = secciones (placeholder vacío), fila 2 = nombres de columna, fila 3+ = data.
    Esto coincide con header=1 de main.py._cargar_detectados.
    """
    wb = Workbook()
    ws = wb.active
    for ci, nombre in enumerate(_COLS_PAGOS, start=1):
        ws.cell(row=1, column=ci, value="")        # sección (placeholder)
        ws.cell(row=2, column=ci, value=nombre)    # nombre de columna
    for ri, fila in enumerate(filas, start=3):
        for ci, val in enumerate(fila, start=1):
            ws.cell(row=ri, column=ci, value=val)
    p = reclamos.PAGOS_EFECTIVO_PATH
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)


def _crear_vista_previa(mes: str, filas: list):
    """
    Crea outputs/reclamos_{mes}.xlsx simulando un estado previo (trabajo del supervisor).
    Lo escribe usando el writer del módulo, para que el formato sea exactamente el que el
    código produce y luego volverá a leer.
    """
    cols = ["MZ", "LT", "MESA", "COBRADOR", "FECHA_COBRO", "MONTO",
            "MES_ANO_DETECTADO", "MES_ANO_ORIGEN",
            "RECLAMO", "ESTADO", "FECHA_RESOLUCION"]
    df = pd.DataFrame(filas, columns=cols)
    reclamos._write_vista(df, mes)


def _leer_vista(mes: str) -> pd.DataFrame:
    p = reclamos.OUTPUTS_DIR / f"reclamos_{mes}.xlsx"
    if not p.exists():
        return pd.DataFrame()
    # header=1 → la fila 2 (1-indexed) contiene los nombres de columna
    return pd.read_excel(p, sheet_name="Reclamos", header=1, dtype=str).fillna("")


def _leer_trazab() -> pd.DataFrame:
    p = reclamos.TRAZAB_DIR / "trazabilidad_reclamos.xlsx"
    if not p.exists():
        return pd.DataFrame()
    return pd.read_excel(p, sheet_name="Trazabilidad", header=1, dtype=str).fillna("")


_COLS_RESOL = ["MZ", "LT", "NOMBRE", "MESA", "FECHA_COBRO", "MES_ANO_ORIGEN",
               "RECLAMO", "TIPO_RECLAMO", "CAMPO", "VALOR_ACTUAL",
               "VALOR_A_CORREGIR", "RESOLUCION", "ESTADO", "FECHA_RESOLUCION"]


def _crear_resolucion(mes: str, filas: list):
    """
    Crea outputs/resolucion_reclamos_{mes}.xlsx con la estructura que
    main._build_resolucion_lookup espera (header=1, sheet "Correcciones").
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Correcciones"
    for ci, nombre in enumerate(_COLS_RESOL, start=1):
        ws.cell(row=1, column=ci, value="")        # sección (placeholder)
        ws.cell(row=2, column=ci, value=nombre)
    for ri, fila in enumerate(filas, start=3):
        for ci, val in enumerate(fila, start=1):
            ws.cell(row=ri, column=ci, value=val)
    p = reclamos.OUTPUTS_DIR / f"resolucion_reclamos_{mes}.xlsx"
    wb.save(p)


# ── Tests ────────────────────────────────────────────────────────────────────

def test_detecta_solo_reclamos():
    mes = "2026-06"
    _crear_pagos([
        ["A", "5",  "38", "2026-06-03", "mesa_1", "solo_un_cobrador", "Wilder", "reclamo monto", "1"],
        ["B", "3",  "42", "2026-06-04", "mesa_1", "solo_un_cobrador", "Wilder", "",              "1"],
        ["C", "7",  "35", "2026-06-05", "mesa_2", "solo_un_cobrador", "Yerald", "pago mixto",    "1"],
        ["D", "11", "50", "2026-06-06", "mesa_2", "solo_un_cobrador", "Yerald", "RECLAMO mayús", "1"],
    ])
    reclamos.main(mes)
    v = _leer_vista(mes)
    assert len(v) == 2, f"Esperaba 2 reclamos, llegaron {len(v)}: {v[['MZ','LT']].to_dict('records')}"
    pares = set(zip(v["MZ"], v["LT"]))
    assert pares == {("A", "5"), ("D", "11")}, f"Pares incorrectos: {pares}"
    assert all(v["ESTADO"] == "PENDIENTE"), f"Estados iniciales no son PENDIENTE: {v['ESTADO'].tolist()}"
    print("  OK test_detecta_solo_reclamos")


def test_caso_vacio():
    mes = "2026-06"
    _crear_pagos([
        ["A", "5", "38", "2026-06-03", "mesa_1", "solo_un_cobrador", "Wilder", "",       "1"],
        ["B", "3", "42", "2026-06-04", "mesa_1", "solo_un_cobrador", "Wilder", "normal", "1"],
    ])
    reclamos.main(mes)
    v = _leer_vista(mes)
    assert v.empty, f"Esperaba vista vacía, llegó: {v.to_dict('records')}"
    p = reclamos.TRAZAB_DIR / "trazabilidad_reclamos.xlsx"
    assert not p.exists(), "Trazabilidad no debería crearse cuando no hay reclamos"
    print("  OK test_caso_vacio")


def test_preserva_trabajo_manual():
    mes = "2026-06"
    # 1ª corrida: detecta el reclamo en estado PENDIENTE
    _crear_pagos([
        ["A", "5", "38", "2026-06-03", "mesa_1", "solo_un_cobrador", "Wilder", "reclamo", "1"],
    ])
    reclamos.main(mes)

    # Supervisor llena el archivo (RECLAMO + cambio a EN_REVISION)
    p = reclamos.OUTPUTS_DIR / f"reclamos_{mes}.xlsx"
    df = pd.read_excel(p, sheet_name="Reclamos", header=1, dtype=str)
    df.loc[0, "RECLAMO"] = "dice que ya pagó en mayo"
    df.loc[0, "ESTADO"]  = "EN_REVISION"
    # Re-escribimos preservando el formato
    reclamos._write_vista(df.fillna(""), mes)

    # 2ª corrida: el mismo pago_efectivo (no cambia), debe preservar el trabajo manual
    reclamos.main(mes)

    v = _leer_vista(mes)
    assert len(v) == 1, f"Esperaba 1 fila, llegaron {len(v)}"
    assert v.loc[0, "RECLAMO"]  == "dice que ya pagó en mayo", f"RECLAMO perdido: {v.loc[0, 'RECLAMO']!r}"
    assert v.loc[0, "ESTADO"]   == "EN_REVISION",              f"ESTADO perdido: {v.loc[0, 'ESTADO']!r}"
    print("  OK test_preserva_trabajo_manual")


def test_cierra_a_trazabilidad():
    mes = "2026-06"
    _crear_pagos([
        ["A", "5", "38", "2026-06-03", "mesa_1", "solo_un_cobrador", "Wilder", "reclamo A", "1"],
        ["B", "7", "42", "2026-06-04", "mesa_2", "solo_un_cobrador", "Yerald", "reclamo B", "1"],
    ])
    reclamos.main(mes)

    # Supervisor cierra: A=RESUELTO, B sigue PENDIENTE
    p = reclamos.OUTPUTS_DIR / f"reclamos_{mes}.xlsx"
    df = pd.read_excel(p, sheet_name="Reclamos", header=1, dtype=str).fillna("")
    df.loc[df["MZ"] == "A", "ESTADO"]           = "RESUELTO"
    df.loc[df["MZ"] == "A", "RECLAMO"]          = "ya pagó, comprobante visto"
    df.loc[df["MZ"] == "A", "FECHA_RESOLUCION"] = "2026-06-10"
    reclamos._write_vista(df, mes)

    # 2ª corrida: cierre debe moverse a trazabilidad
    reclamos.main(mes)

    v = _leer_vista(mes)
    assert len(v) == 1 and v.loc[0, "MZ"] == "B", \
        f"Vista debería tener solo B PENDIENTE, llegó: {v[['MZ','ESTADO']].to_dict('records')}"

    t = _leer_trazab()
    assert len(t) == 1 and t.loc[0, "MZ"] == "A", \
        f"Trazabilidad debería tener solo A RESUELTO, llegó: {t[['MZ','ESTADO_FINAL']].to_dict('records')}"
    assert t.loc[0, "ESTADO_FINAL"] == "RESUELTO", f"ESTADO_FINAL: {t.loc[0, 'ESTADO_FINAL']!r}"
    assert t.loc[0, "MES_CIERRE"]   == mes,        f"MES_CIERRE: {t.loc[0, 'MES_CIERRE']!r}"
    print("  OK test_cierra_a_trazabilidad")


def test_arrastra_de_mes_anterior():
    mes_prev = "2026-05"
    mes      = "2026-06"

    # Vista de mayo con 1 reclamo PENDIENTE (que nadie cerró)
    _crear_vista_previa(mes_prev, [
        ["A", "5", "mesa_1", "Wilder", "2026-05-03", "38",
         "2026-05", "2026-05",
         "no le tomaron la lectura", "PENDIENTE", ""],
    ])

    # En junio: pagos_efectivo no incluye ese predio
    _crear_pagos([
        ["B", "3", "42", "2026-06-04", "mesa_1", "solo_un_cobrador", "Wilder", "", "1"],
    ])
    reclamos.main(mes)

    v = _leer_vista(mes)
    assert len(v) == 1, f"Esperaba 1 arrastre, llegaron {len(v)}: {v.to_dict('records')}"
    assert v.loc[0, "MZ"] == "A" and v.loc[0, "LT"] == "5", f"Predio arrastrado incorrecto: {v.loc[0].to_dict()}"
    assert v.loc[0, "MES_ANO_ORIGEN"]    == "2026-05", f"MES_ANO_ORIGEN: {v.loc[0, 'MES_ANO_ORIGEN']!r}"
    assert v.loc[0, "MES_ANO_DETECTADO"] == "2026-06", f"MES_ANO_DETECTADO: {v.loc[0, 'MES_ANO_DETECTADO']!r}"
    assert v.loc[0, "ESTADO"]            == "PENDIENTE"
    print("  OK test_arrastra_de_mes_anterior")


def test_clave_distingue_mesas():
    """Mismo MZ+LT en 2 mesas distintas → 2 reclamos separados con trabajo manual independiente."""
    mes = "2026-06"
    _crear_pagos([
        ["A", "5", "38", "2026-06-03", "mesa_1", "pago_multi_mesa", "Wilder", "reclamo m1", "1"],
        ["A", "5", "40", "2026-06-04", "mesa_2", "pago_multi_mesa", "Yerald", "reclamo m2", "1"],
    ])
    reclamos.main(mes)

    # Supervisor cierra solo el de mesa_1
    p = reclamos.OUTPUTS_DIR / f"reclamos_{mes}.xlsx"
    df = pd.read_excel(p, sheet_name="Reclamos", header=1, dtype=str).fillna("")
    df.loc[df["MESA"] == "mesa_1", "ESTADO"]   = "RECHAZADO"
    df.loc[df["MESA"] == "mesa_1", "RECLAMO"]  = "el usuario reconoció el pago"
    reclamos._write_vista(df, mes)

    reclamos.main(mes)

    v = _leer_vista(mes)
    assert len(v) == 1, f"Esperaba 1 activo (mesa_2), llegaron {len(v)}"
    assert v.loc[0, "MESA"] == "mesa_2", f"Activo incorrecto: {v.loc[0, 'MESA']!r}"

    t = _leer_trazab()
    assert len(t) == 1 and t.loc[0, "MESA"] == "mesa_1", \
        f"Trazabilidad debería tener mesa_1, llegó: {t['MESA'].tolist()}"
    print("  OK test_clave_distingue_mesas")


def test_trazabilidad_acumula():
    """Segunda corrida con un nuevo cierre no borra los cierres previos."""
    mes = "2026-06"

    # Corrida 1: A se cierra como RESUELTO
    _crear_pagos([
        ["A", "5", "38", "2026-06-03", "mesa_1", "solo_un_cobrador", "Wilder", "reclamo A", "1"],
        ["B", "7", "42", "2026-06-04", "mesa_2", "solo_un_cobrador", "Yerald", "reclamo B", "1"],
    ])
    reclamos.main(mes)

    p = reclamos.OUTPUTS_DIR / f"reclamos_{mes}.xlsx"
    df = pd.read_excel(p, sheet_name="Reclamos", header=1, dtype=str).fillna("")
    df.loc[df["MZ"] == "A", "ESTADO"] = "RESUELTO"
    reclamos._write_vista(df, mes)

    reclamos.main(mes)
    t1 = _leer_trazab()
    assert len(t1) == 1, f"Trazabilidad después del 1er cierre: esperaba 1, llegó {len(t1)}"

    # Ahora cerramos B como RECHAZADO
    df = pd.read_excel(p, sheet_name="Reclamos", header=1, dtype=str).fillna("")
    df.loc[df["MZ"] == "B", "ESTADO"] = "RECHAZADO"
    reclamos._write_vista(df, mes)

    reclamos.main(mes)
    t2 = _leer_trazab()

    assert len(t2) == 2, f"Trazabilidad después del 2do cierre: esperaba 2, llegó {len(t2)}"
    mzs = set(t2["MZ"].tolist())
    assert mzs == {"A", "B"}, f"Trazabilidad perdió alguno: {mzs}"
    print("  OK test_trazabilidad_acumula")


def test_mergea_correccion_a_trazabilidad():
    """
    RESUELTO toma CAMPO/VALOR_ANTERIOR/VALOR_APLICADO de resolucion_reclamos.
    RECHAZADO los deja vacíos aunque resolucion_reclamos tenga datos para esa fila.
    """
    mes = "2026-06"

    _crear_pagos([
        ["A", "5",  "38", "2026-06-03", "mesa_1", "solo_un_cobrador", "Wilder", "reclamo A", "1"],
        ["B", "12", "44", "2026-06-04", "mesa_2", "solo_un_cobrador", "Yerald", "reclamo B", "1"],
    ])
    reclamos.main(mes)

    # Supervisor clasifica TIPO_RECLAMO + marca ESTADO en reclamos.xlsx
    p_rec = reclamos.OUTPUTS_DIR / f"reclamos_{mes}.xlsx"
    df = pd.read_excel(p_rec, sheet_name="Reclamos", header=1, dtype=str).fillna("")
    df.loc[df["MZ"] == "A", ["TIPO_RECLAMO", "ESTADO", "RESOLUCION"]] = [
        "mes_anterior", "RESUELTO", "comprobante visto"
    ]
    df.loc[df["MZ"] == "B", ["TIPO_RECLAMO", "ESTADO", "RESOLUCION"]] = [
        "convenio", "RECHAZADO", "deuda activa"
    ]
    reclamos._write_vista(df, mes)

    # Supervisor genera resolucion_reclamos.xlsx con la corrección.
    # B (RECHAZADO) también tiene valores — main.py debe ignorarlos por ESTADO.
    _crear_resolucion(mes, [
        ["A", "5",  "Juan A", "mesa_1", "2026-06-03", "2026-06",
         "reclamo A", "mes_anterior", "MES ANTERIOR", "35.00", "0.00",
         "comprobante visto", "RESUELTO",  "2026-06-10"],
        ["B", "12", "Rosa B", "mesa_2", "2026-06-04", "2026-06",
         "reclamo B", "convenio",     "Convenio",     "44.00", "44.00",
         "deuda activa",      "RECHAZADO", "2026-06-11"],
    ])

    # 2ª corrida: cierres pasan a trazabilidad
    reclamos.main(mes)

    t = _leer_trazab()
    assert len(t) == 2, f"Esperaba 2 filas en trazabilidad, llegaron {len(t)}"

    fila_a = t[t["MZ"] == "A"].iloc[0]
    fila_b = t[t["MZ"] == "B"].iloc[0]

    # A (RESUELTO): corrección mergeada desde resolucion_reclamos
    assert fila_a["ESTADO_FINAL"] == "RESUELTO"
    assert fila_a["CAMPO"]        == "MES ANTERIOR", f"CAMPO A: {fila_a['CAMPO']!r}"
    assert float(fila_a["VALOR_ANTERIOR"]) == 35.0, f"VALOR_ANTERIOR A: {fila_a['VALOR_ANTERIOR']!r}"
    assert float(fila_a["VALOR_APLICADO"]) == 0.0,  f"VALOR_APLICADO A: {fila_a['VALOR_APLICADO']!r}"

    # B (RECHAZADO): las 3 columnas vacías aunque resolucion_reclamos tuviera datos
    assert fila_b["ESTADO_FINAL"]   == "RECHAZADO"
    assert fila_b["CAMPO"]          == "", f"CAMPO B debería estar vacío: {fila_b['CAMPO']!r}"
    assert fila_b["VALOR_ANTERIOR"] == "", f"VALOR_ANTERIOR B: {fila_b['VALOR_ANTERIOR']!r}"
    assert fila_b["VALOR_APLICADO"] == "", f"VALOR_APLICADO B: {fila_b['VALOR_APLICADO']!r}"

    print("  OK test_mergea_correccion_a_trazabilidad")


def test_sync_estado_desde_resolucion():
    """
    Workflow nuevo: el supervisor solo edita resolucion_reclamos.xlsx.
    reclamos.xlsx queda con ESTADO=PENDIENTE; main.py debe leer ESTADO desde
    resolucion_reclamos.xlsx y propagar el cierre.
    """
    mes = "2026-06"

    _crear_pagos([
        ["A", "5", "38", "2026-06-03", "mesa_1", "solo_un_cobrador", "Wilder", "reclamo A", "1"],
    ])
    reclamos.main(mes)

    # NOTA: deliberadamente NO tocamos reclamos.xlsx. Sigue como PENDIENTE.
    # El supervisor solo edita resolucion_reclamos.xlsx.
    _crear_resolucion(mes, [
        ["A", "5", "Juan A", "mesa_1", "2026-06-03", "2026-06",
         "reclamo A", "mes_anterior", "MES ANTERIOR", "35.00", "0.00",
         "comprobante visto", "RESUELTO", "2026-06-10"],
    ])

    # main.py debe sincronizar: ver ESTADO=RESUELTO en resolucion → cerrar
    reclamos.main(mes)

    v = _leer_vista(mes)
    assert v.empty, f"Vista debería estar vacía tras cierre, llegó: {v.to_dict('records')}"

    t = _leer_trazab()
    assert len(t) == 1, f"Trazabilidad debería tener 1 fila, llegaron {len(t)}"
    fila = t.iloc[0]
    assert fila["ESTADO_FINAL"]   == "RESUELTO",         f"ESTADO_FINAL: {fila['ESTADO_FINAL']!r}"
    assert fila["RESOLUCION"]     == "comprobante visto", f"RESOLUCION: {fila['RESOLUCION']!r}"
    assert fila["CAMPO"]          == "MES ANTERIOR",     f"CAMPO: {fila['CAMPO']!r}"
    assert float(fila["VALOR_ANTERIOR"]) == 35.0
    assert float(fila["VALOR_APLICADO"]) == 0.0
    print("  OK test_sync_estado_desde_resolucion")


# ── Runner ───────────────────────────────────────────────────────────────────

def main():
    tests = [
        test_detecta_solo_reclamos,
        test_caso_vacio,
        test_preserva_trabajo_manual,
        test_cierra_a_trazabilidad,
        test_arrastra_de_mes_anterior,
        test_clave_distingue_mesas,
        test_trazabilidad_acumula,
        test_mergea_correccion_a_trazabilidad,
        test_sync_estado_desde_resolucion,
    ]
    for t in tests:
        _setup()
        try:
            t()
        except Exception:
            _teardown()
            raise
    _teardown()
    print(f"\n[OK] {len(tests)}/{len(tests)} tests pasaron")


if __name__ == "__main__":
    main()
