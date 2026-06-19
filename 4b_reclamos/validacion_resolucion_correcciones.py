"""
validacion_resolucion_correcciones.py — Verifica RESUELTO → DATA_boletas (4b_reclamos)

USO:
    python validacion_resolucion_correcciones.py --mes 2026-06

LOGICA:
    1. Lee trazabilidad/trazabilidad_reclamos.xlsx → filtra ESTADO_FINAL=RESUELTO y MES_CIERRE=mes.
    2. Lee 3_boletas/inputs/DATA_boletas.xlsx (lectura directa — valor vigente).
    3. Para cada fila, cruza (MZ, LT) y compara VALOR_APLICADO vs DATA_boletas[CAMPO].
    4. Clasifica: OK / ERR / AUSENTE / SIN_MAP.
    5. Escribe outputs/validacion_correcciones_{mes}.xlsx con 2 hojas (Resumen + Detalle).

NO MODIFICA NINGÚN ARCHIVO — solo lectura sobre trazabilidad y DATA_boletas.

Contrato visual: docs/formato_validacion_correcciones.html
"""

import argparse
import logging
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent.parent / "shared"))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

from main import (
    OUTPUTS_DIR, TRAZAB_DIR,
    SEC_PREDIO, SEC_TIPO, SEC_MANUAL,
    _norm, _clean, _argb,
    _hdr, _dat, _write_headers, _write_fila,
)
from resolucion import SEC_AUTO
from data_boletas_repo import get_predio_lookup

log = logging.getLogger(__name__)

# ── Paleta de verificación ────────────────────────────────────────────────────
# Sección "Verificación" — cabecera neutra; las celdas se tintan por estado.
SEC_VERIF = ("4B5563", "FFFFFF", "F9FAFB", "374151")

TINT_OK      = ("D1FAE5", "065F46")
TINT_ERR     = ("FEE2E2", "991B1B")
TINT_AUSENTE = ("FEF3C7", "92400E")
TINT_SIN_MAP = ("F3F4F6", "6B7280")

TINT_POR_ESTADO = {
    "OK":      TINT_OK,
    "ERR":     TINT_ERR,
    "AUSENTE": TINT_AUSENTE,
    "SIN_MAP": TINT_SIN_MAP,
}

DISPLAY_POR_ESTADO = {
    "OK":      "✔ OK",
    "ERR":     "✗ ERR",
    "AUSENTE": "✗ AUSENTE",
    "SIN_MAP": "⊘ SIN_MAP",
}

TOLERANCIA_NUMERICA = 0.01

# ── Schema (contrato: docs/formato_validacion_correcciones.html) ─────────────
_COLS_DETALLE = [
    ("MZ",                  SEC_PREDIO, 6,  "center", None),
    ("LT",                  SEC_PREDIO, 7,  "center", None),
    ("NOMBRE",              SEC_PREDIO, 28, "left",   None),
    ("TIPO_RECLAMO",        SEC_TIPO,   20, "center", None),
    ("CAMPO",               SEC_TIPO,   24, "center", None),
    ("FECHA_RESOLUCION",    SEC_TIPO,   16, "center", "DD/MM/YYYY"),
    ("VALOR_APLICADO",      SEC_MANUAL, 16, "right",  '#,##0.00'),
    ("VALOR_EN_BOLETAS",    SEC_AUTO,   18, "right",  '#,##0.00'),
    ("ESTADO_VERIFICACION", SEC_VERIF,  20, "center", None),
    ("DIFERENCIA",          SEC_VERIF,  12, "right",  '#,##0.00'),
]

_SECCIONES_DETALLE = [
    ("¿Cuál es el predio?",      "MZ",                  "NOMBRE"),
    ("¿Qué se cerró?",           "TIPO_RECLAMO",        "FECHA_RESOLUCION"),
    ("¿Qué se decidió?",         "VALOR_APLICADO",      "VALOR_APLICADO"),
    ("¿Qué hay en boletas hoy?", "VALOR_EN_BOLETAS",    "VALOR_EN_BOLETAS"),
    ("Verificación",             "ESTADO_VERIFICACION", "DIFERENCIA"),
]

COL_ESTADO = next(i for i, c in enumerate(_COLS_DETALLE, 1) if c[0] == "ESTADO_VERIFICACION")
COL_DIFF   = next(i for i, c in enumerate(_COLS_DETALLE, 1) if c[0] == "DIFERENCIA")

# ── Rutas ─────────────────────────────────────────────────────────────────────

def _ruta_trazab() -> Path:
    return TRAZAB_DIR / "trazabilidad_reclamos.xlsx"

def _ruta_salida(mes: str) -> Path:
    return OUTPUTS_DIR / f"validacion_correcciones_{mes}.xlsx"

# ── Carga ─────────────────────────────────────────────────────────────────────

def _leer_trazabilidad(mes: str) -> pd.DataFrame:
    """Lee trazabilidad y filtra ESTADO_FINAL=RESUELTO y MES_CIERRE=mes."""
    p = _ruta_trazab()
    if not p.exists():
        log.warning(f"trazabilidad_reclamos.xlsx no encontrado: {p}")
        return pd.DataFrame()
    try:
        df = pd.read_excel(p, sheet_name="Trazabilidad", header=1, dtype=str)
    except Exception as e:
        log.error(f"Error leyendo trazabilidad: {e}")
        return pd.DataFrame()
    df = df.fillna("")
    mask = (df["ESTADO_FINAL"].str.strip().str.upper() == "RESUELTO") & \
           (df["MES_CIERRE"].str.strip() == mes)
    return df[mask].copy()


# ── Comparación ──────────────────────────────────────────────────────────────

def _try_float(s) -> float | None:
    """Convierte string a float. None si no es numérico o vacío."""
    if s is None:
        return None
    txt = str(s).strip().replace(",", "")
    if txt in ("", "nan", "None", "NaT"):
        return None
    try:
        return float(txt)
    except (ValueError, TypeError):
        return None


def _comparar(aplicado: str, en_boletas: str) -> tuple[str, float | None]:
    """
    Compara dos valores.
        - Si ambos numéricos → tolerancia ±0.01, retorna diff (b - a).
        - Si no → strings normalizados (strip + upper + sin espacios).
    Retorna (estado, diferencia_o_None).
    """
    a = _try_float(aplicado)
    b = _try_float(en_boletas)
    if a is not None and b is not None:
        diff = b - a
        return ("OK" if abs(diff) <= TOLERANCIA_NUMERICA else "ERR"), diff

    if _norm(aplicado) == _norm(en_boletas):
        return "OK", None
    return "ERR", None

# ── Verificación ─────────────────────────────────────────────────────────────

def _verificar(traz: pd.DataFrame, boletas: dict) -> list[dict]:
    """Genera una fila del Detalle por cada fila RESUELTO de trazabilidad."""
    filas = []
    for _, r in traz.iterrows():
        mz = _norm(r.get("MZ", ""))
        lt = _norm(r.get("LT", ""))
        campo    = _clean(r.get("CAMPO", ""))
        aplicado = _clean(r.get("VALOR_APLICADO", ""))

        boleta = boletas.get((mz, lt))

        if boleta is None:
            estado, nombre, en_boletas, diff = "AUSENTE", "", "", None
        elif not campo:
            estado = "SIN_MAP"
            nombre = _clean(boleta.get("NOMBRES", ""))
            en_boletas, diff = "", None
        else:
            nombre = _clean(boleta.get("NOMBRES", ""))
            en_boletas = _clean(boleta.get(campo, ""))
            estado, diff = _comparar(aplicado, en_boletas)

        filas.append({
            "MZ":                  _clean(r.get("MZ", "")),
            "LT":                  _clean(r.get("LT", "")),
            "NOMBRE":              nombre,
            "TIPO_RECLAMO":        _clean(r.get("TIPO_RECLAMO", "")),
            "CAMPO":               campo,
            "FECHA_RESOLUCION":    _clean(r.get("FECHA_RESOLUCION", "")),
            "VALOR_APLICADO":      _try_float(aplicado) if _try_float(aplicado) is not None else aplicado,
            "VALOR_EN_BOLETAS":    _try_float(en_boletas) if _try_float(en_boletas) is not None else en_boletas,
            "ESTADO_VERIFICACION": estado,
            "DIFERENCIA":          diff,
        })
    return filas

# ── Escritura ────────────────────────────────────────────────────────────────

def _write_resumen(ws, conteos: dict, total: int, mes: str) -> None:
    """Hoja 1 — Resumen: contadores por categoría."""
    ws.merge_cells("A1:C1")
    _hdr(ws.cell(row=1, column=1), "1E3A5F", "FFFFFF",
         f"Verificación de correcciones — mes {mes}")
    ws.row_dimensions[1].height = 22

    _hdr(ws.cell(row=2, column=1), "F3F4F6", "374151", "Categoría")
    _hdr(ws.cell(row=2, column=2), "F3F4F6", "374151", "Cantidad")
    _hdr(ws.cell(row=2, column=3), "F3F4F6", "374151", "Descripción")
    ws.row_dimensions[2].height = 20

    filas = [
        ("✔ Aplicadas correctamente", conteos["OK"],
         "VALOR_EN_BOLETAS coincide con VALOR_APLICADO (tolerancia ±0.01)", TINT_OK),
        ("✗ Discrepancias", conteos["ERR"],
         "VALOR_EN_BOLETAS difiere de VALOR_APLICADO — revisar si la corrección se aplicó", TINT_ERR),
        ("✗ Predios ausentes", conteos["AUSENTE"],
         "(MZ, LT) no encontrado en DATA_boletas — revisar si el predio cambió de clave", TINT_AUSENTE),
        ("⊘ Sin mapeo", conteos["SIN_MAP"],
         "TIPO_RECLAMO sin columna en DATA_boletas (no verificable)", TINT_SIN_MAP),
    ]
    for i, (cat, n, desc, tint) in enumerate(filas, start=3):
        _dat(ws.cell(row=i, column=1), cat, "F4FAFF", "1F2937", align="left")
        _dat(ws.cell(row=i, column=2), n, tint[0], tint[1], align="right")
        ws.cell(row=i, column=2).font = Font(color=_argb(tint[1]), size=10, bold=True)
        _dat(ws.cell(row=i, column=3), desc, "F4FAFF", "1F2937", align="left")

    r = 3 + len(filas)
    _dat(ws.cell(row=r, column=1), "TOTAL RESUELTOS verificados", "E5E7EB", "111111", align="left")
    ws.cell(row=r, column=1).font = Font(color=_argb("111111"), size=10, bold=True)
    _dat(ws.cell(row=r, column=2), total, "E5E7EB", "111111", align="right")
    ws.cell(row=r, column=2).font = Font(color=_argb("111111"), size=10, bold=True)
    _dat(ws.cell(row=r, column=3),
         "Filas con ESTADO_FINAL=RESUELTO y MES_CIERRE=mes en trazabilidad",
         "E5E7EB", "555555", align="left")

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 70


def _write_detalle(ws, filas: list[dict]) -> None:
    """Hoja 2 — Detalle: cabecera + filas tintadas por estado de verificación."""
    _write_headers(ws, _COLS_DETALLE, _SECCIONES_DETALLE)

    for ri, fila in enumerate(filas, start=3):
        _write_fila(ws, ri, fila, _COLS_DETALLE)

        # Re-tinte de ESTADO_VERIFICACION y DIFERENCIA según estado
        estado = fila["ESTADO_VERIFICACION"]
        tint = TINT_POR_ESTADO[estado]
        display = DISPLAY_POR_ESTADO[estado]

        _dat(ws.cell(row=ri, column=COL_ESTADO), display, tint[0], tint[1], align="center")
        ws.cell(row=ri, column=COL_ESTADO).font = Font(color=_argb(tint[1]), size=10, bold=True)

        diff = fila["DIFERENCIA"]
        if diff is None:
            _dat(ws.cell(row=ri, column=COL_DIFF), None, tint[0], tint[1], align="right")
        else:
            _dat(ws.cell(row=ri, column=COL_DIFF), diff, tint[0], tint[1],
                 align="right", fmt='#,##0.00')
            ws.cell(row=ri, column=COL_DIFF).font = Font(color=_argb(tint[1]), size=10, bold=True)


def _write_salida(filas: list[dict], conteos: dict, mes: str) -> None:
    wb = Workbook()
    ws_res = wb.active
    ws_res.title = "Resumen"
    _write_resumen(ws_res, conteos, len(filas), mes)

    ws_det = wb.create_sheet("Detalle")
    _write_detalle(ws_det, filas)

    p = _ruta_salida(mes)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    log.info(f"Validación guardada: {p.name} ({len(filas)} filas)")

# ── Main ─────────────────────────────────────────────────────────────────────

def main(mes: str) -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
        force=True,
    )
    log.info(f"=== 4b_reclamos/validacion_resolucion_correcciones.py — mes {mes} ===")

    traz = _leer_trazabilidad(mes)
    log.info(f"RESUELTO en {mes}: {len(traz)}")

    conteos = {"OK": 0, "ERR": 0, "AUSENTE": 0, "SIN_MAP": 0}

    if traz.empty:
        log.info("Sin filas RESUELTO para verificar — se escribe Resumen vacío")
        _write_salida([], conteos, mes)
        return

    boletas = get_predio_lookup()
    log.info(f"DATA_boletas (vía repo): {len(boletas)} predios disponibles")

    filas = _verificar(traz, boletas)
    for f in filas:
        conteos[f["ESTADO_VERIFICACION"]] += 1

    log.info(f"Resumen: OK={conteos['OK']}  ERR={conteos['ERR']}  "
             f"AUSENTE={conteos['AUSENTE']}  SIN_MAP={conteos['SIN_MAP']}")

    _write_salida(filas, conteos, mes)
    log.info("=== completado ===")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Verifica que correcciones RESUELTO se aplicaron en DATA_boletas")
    parser.add_argument("--mes", required=True, help="Mes a procesar (YYYY-MM)")
    args = parser.parse_args()
    main(args.mes)
