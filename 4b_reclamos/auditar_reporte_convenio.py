"""
4b_reclamos/auditar_reporte_convenio.py — Auditoria INTERNA de
`reporte_convenio_multa_referencias_2026-07.pdf`.

Regla: solo se mira lo que el PDF imprime -- portada + las dos tablas de cada
pagina (historial mensual y "Referencia de pago"). NO se cruza contra el
ledger, la planilla ni los precursores: la pregunta es si el reporte se
contradice A SI MISMO.

Chequeos:
  1_ARITMETICA_FILA    -- TOTAL PAGADO de una fila != suma de sus 7 conceptos
  2_DESCUADRE_TABLAS   -- lo que dice la tabla de arriba para un mes no
                          coincide con la suma de "Referencia de pago" del
                          mismo mes (las dos tablas de la misma hoja)
  3_PAGO_SIN_REFERENCIA-- fila con TOTAL > 0 y marca "OK" pero abajo no hay
                          ninguna referencia de pago de ese mes
  4_REFERENCIA_HUERFANA-- referencia de pago de un mes que arriba figura sin pago
  5_PORTADA_VS_PAGINA  -- "Multa (ya pago)" de la portada != suma de la
                          columna Multa del historial de esa misma pagina
  6_PORTADA_ARITMETICA -- DIFERENCIA != Multa - Convenio, o "Cubriria" no
                          coincide con los numeros que la propia portada muestra
  7_FECHA_VS_MES       -- la fecha impresa en la referencia cae en un mes
                          distinto al mes de la fila
  8_MONTO_NEGATIVO     -- monto negativo impreso (deuda o pago en negativo)

Uso: py 4b_reclamos/auditar_reporte_convenio.py
"""

import sys
from pathlib import Path

import pandas as pd

BASE_DIR = Path(__file__).parent
ROOT = BASE_DIR.parent
sys.path.insert(0, str(ROOT / "shared"))
sys.path.insert(0, str(BASE_DIR))
import seguimiento_repo as repo  # noqa: E402
import reporte_historico as rh  # noqa: E402
import reporte_convenio_multa as rcm  # noqa: E402
import reporte_referencias_pago as rrp  # noqa: E402
import agregar_lote_al_reporte_convenio as alrc  # noqa: E402

MES_ANO = "2026-07"
TOL = 0.005
# Desde junio 2026 el exceso se retiene y no se muestra (lo dice el pie del
# propio reporte), asi que un descuadre tabla-vs-referencia en jun/jul es
# esperado y no se marca. Antes de junio no hay retencion: ahi si debe cuadrar.
MESES_CON_RETENCION = {"2026-06", "2026-07"}


def _f(v) -> float:
    n = pd.to_numeric(v, errors="coerce")
    return 0.0 if pd.isna(n) else float(n)


def _mes_de_fecha(txt: str) -> str | None:
    """El reporte imprime dos formatos: ISO (2026-07-04, viene de efectivo) y
    dd/mm/yyyy (04/07/2026, viene de yape). Parsear todo con dayfirst=True
    rompe el ISO -- lee 2026-07-04 como 7 de ABRIL y marcaba 44 predios como
    anomalos por un bug del auditor, no del reporte (visto 01/08/2026)."""
    s = str(txt).split("·")[0].strip()
    iso = bool(pd.Series([s]).str.match(r"^\d{4}-\d{2}-\d{2}").iloc[0])
    ts = pd.to_datetime(s, dayfirst=not iso, errors="coerce")
    return None if pd.isna(ts) else f"{ts.year:04d}-{ts.month:02d}"


def _meses_validos(mes: str) -> set[str]:
    """El ciclo N se cobra desde mediados del mes anterior hasta fin del mes
    del ciclo (verificado: ciclo 2026-07 va del 17/06 al 20/07). Que la fecha
    caiga en el mes anterior es lo NORMAL, no una anomalia."""
    p = pd.Period(mes, freq="M")
    return {str(p - 1), str(p)}


def auditar() -> pd.DataFrame:
    hallazgos: list[dict] = []

    def add(tipo, sev, predio, nombre, mes, detalle, esperado="", encontrado=""):
        hallazgos.append({
            "TIPO": tipo, "SEVERIDAD": sev, "PREDIO": predio, "NOMBRE": nombre,
            "MES": mes, "ESPERADO": esperado, "ENCONTRADO": encontrado,
            "DETALLE": detalle,
        })

    # ---- exactamente los datos que el PDF imprime -------------------------
    df_rep = alrc.calcular_tabla_con_lote(MES_ANO)          # portada
    historicos = rh._cargar_historicos()
    eventos = repo._leer_eventos()
    mapa_raw = rh._cargar_mapa_raw()
    dfp = pd.read_excel(ROOT / "5_cobranza" / "outputs" / "planilla_cobrado.xlsx",
                        sheet_name="planilla_cobrado", header=1)
    nombres = repo._lookup_nombres()
    redirects = rcm._cargar_redirects()

    for _, rp in df_rep.iterrows():
        mz, lt = rp["MZ"], rp["LT"]
        predio = f"{mz}-{lt}"
        nombre = str(rp.get("NOMBRE") or "")

        # las 2 tablas de la pagina, tal cual se dibujan
        tabla = rh.tabla_predio(mz, lt, historicos, eventos, dfp, mapa_raw, nombre)
        tabla = rcm.corregir_tabla_por_redirects(tabla, mz, lt, redirects)
        tabla = alrc.quitar_pago_fantasma_julio(tabla, mz, lt)
        refs = rrp.referencias_pago(mz, lt, tabla=tabla)

        ref_por_mes: dict[str, float] = {}
        for r in refs:
            ref_por_mes[r["MES"]] = ref_por_mes.get(r["MES"], 0.0) + _f(r["MONTO"])

        # --- 1 / 8 / 3 : por fila del historial ---------------------------
        for _, fila in tabla.iterrows():
            mes = str(fila["MES"])
            suma = round(sum(_f(fila[c]) for c in rh.CONCEPTOS_TABLA), 2)
            total = round(_f(fila["TOTAL"]), 2)
            if abs(suma - total) > TOL:
                add("1_ARITMETICA_FILA", "ALTA", predio, nombre, mes,
                    "En la tabla de arriba, TOTAL PAGADO no es la suma de los "
                    "conceptos de esa misma fila.", f"{suma:g}", f"{total:g}")

            # MES_ANT negativo NO es anomalia: hasta mayo 2026 el exceso pagado
            # se devolvia al vecino descontandolo del "mes anterior" del ciclo
            # siguiente, y quedaba grabado como negativo (confirmado por el
            # usuario 01/08/2026). Solo se marca informativo porque en una tabla
            # titulada "Pago por mes y concepto" un numero negativo se lee raro.
            for c in rh.CONCEPTOS_TABLA + ["TOTAL"]:
                v = _f(fila[c])
                if v < -TOL:
                    es_devolucion = (c == "MES_ANT" and mes < "2026-06")
                    add("8_MONTO_NEGATIVO", "INFO" if es_devolucion else "ALTA",
                        predio, nombre, mes,
                        ("Exceso devuelto al vecino via 'mes anterior' (mecanismo "
                         "viejo, legitimo) -- se imprime en negativo dentro de una "
                         "tabla de pagos, conviene explicarlo al entregar."
                         if es_devolucion else
                         f"La columna {c} del historial imprime un monto negativo."),
                        ">= 0", f"{c}={v:g}")

            if total > TOL and bool(fila["PAGO_COMPLETO"]) and mes not in ref_por_mes:
                add("3_PAGO_SIN_REFERENCIA", "MEDIA", predio, nombre, mes,
                    "Arriba figura pago de ese mes, pero abajo no hay ninguna "
                    "linea que diga de donde vino.", "1 referencia", "ninguna")

        # --- 2 / 4 : cruce entre las dos tablas de la misma hoja ----------
        meses_tabla = {str(f["MES"]): round(_f(f["TOTAL"]), 2) for _, f in tabla.iterrows()}
        for mes, monto_ref in ref_por_mes.items():
            total_tabla = meses_tabla.get(mes)
            if total_tabla is None or total_tabla <= TOL:
                add("4_REFERENCIA_HUERFANA", "MEDIA", predio, nombre, mes,
                    "Abajo hay una referencia de pago de ese mes, pero arriba el "
                    "mes figura sin pago.", "TOTAL > 0", f"S/{monto_ref:g} sin fila")
                continue
            if mes in MESES_CON_RETENCION:
                continue
            if abs(total_tabla - monto_ref) > 0.5:
                add("2_DESCUADRE_TABLAS", "ALTA", predio, nombre, mes,
                    "Las dos tablas de la misma pagina no coinciden para ese mes.",
                    f"{total_tabla:g}", f"{monto_ref:g}")

        # --- 7 : la fecha impresa no pertenece al mes de la fila ----------
        for r in refs:
            m_fecha = _mes_de_fecha(r["FECHA_HORA"])
            if m_fecha and m_fecha not in _meses_validos(r["MES"]):
                add("7_FECHA_VS_MES", "MEDIA", predio, nombre, r["MES"],
                    f"La referencia dice mes {r['MES']} pero la fecha impresa es de "
                    f"{m_fecha} -- fuera de la ventana del ciclo (S/{_f(r['MONTO']):g}, "
                    f"{r['MEDIO']}).",
                    "/".join(sorted(_meses_validos(r["MES"]))),
                    str(r["FECHA_HORA"]).split("·")[0].strip())

        # --- 5 / 6 : portada contra la pagina del predio ------------------
        multa_portada = round(_f(rp["MULTA_PAGO"]), 2)
        multa_pagina = round(_f(tabla["MULTA"].sum()), 2)
        if abs(multa_portada - multa_pagina) > TOL:
            add("5_PORTADA_VS_PAGINA", "ALTA", predio, nombre, "-",
                "La portada dice una multa pagada distinta a la que suma el "
                "historial de su propia pagina.", f"{multa_pagina:g}", f"{multa_portada:g}")

        conv = round(_f(rp["CONVENIO_SALDO"]), 2)
        dif_calc = round(multa_portada - conv, 2)
        if abs(dif_calc - round(_f(rp["DIFERENCIA"]), 2)) > TOL:
            add("6_PORTADA_ARITMETICA", "ALTA", predio, nombre, "-",
                "En la portada, DIFERENCIA no es Multa menos Convenio.",
                f"{dif_calc:g}", f"{_f(rp['DIFERENCIA']):g}")
        if bool(rp["CUBRIRIA"]) != (multa_portada >= conv - TOL):
            add("6_PORTADA_ARITMETICA", "ALTA", predio, nombre, "-",
                "La columna '¿Cubriria convenio?' no coincide con los propios "
                "montos de la portada.",
                "SI" if multa_portada >= conv - TOL else "no", str(rp["CUBRIRIA"]))
        if conv <= TOL:
            add("6_PORTADA_ARITMETICA", "MEDIA", predio, nombre, "-",
                "Sale en un reporte de 'convenio pendiente' con convenio en cero.",
                "> 0", f"{conv:g}")

    df = pd.DataFrame(hallazgos)
    if df.empty:
        return df
    df["_s"] = df["SEVERIDAD"].map({"ALTA": 0, "MEDIA": 1, "BAJA": 2, "INFO": 3})
    df = df.sort_values(["_s", "TIPO", "PREDIO", "MES"]).drop(columns="_s")
    return df.reset_index(drop=True)


def exportar(df: pd.DataFrame, salida: Path | None = None) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    salida = salida or (BASE_DIR / "outputs" / f"auditoria_reporte_convenio_{MES_ANO}.xlsx")
    salida.parent.mkdir(exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "hallazgos"
    cols = ["TIPO", "SEVERIDAD", "PREDIO", "NOMBRE", "MES", "ESPERADO", "ENCONTRADO", "DETALLE"]
    anchos = [24, 11, 9, 32, 10, 14, 20, 78]

    ws.append([f"Auditoria INTERNA de reporte_convenio_multa_referencias_{MES_ANO}.pdf — "
               f"{len(df)} hallazgo(s). Solo se comparo el reporte consigo mismo "
               f"(portada vs paginas, tabla de arriba vs referencias de abajo). "
               f"No se cruzo contra ledger, planilla ni precursores."])
    ws.append(cols)
    for _, r in df.iterrows():
        ws.append([r[c] for c in cols])

    ws.cell(row=1, column=1).font = Font(italic=True, size=9)
    for i, w in enumerate(anchos, start=1):
        c = ws.cell(row=2, column=i)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1A5276")
        ws.column_dimensions[c.column_letter].width = w

    colores = {"ALTA": "F8D7DA", "MEDIA": "FFF3CD", "BAJA": "E2E3E5"}
    for row in range(3, ws.max_row + 1):
        fill = PatternFill("solid", fgColor=colores.get(ws.cell(row=row, column=2).value, "FFFFFF"))
        for col in range(1, len(cols) + 1):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=(col == 8))
    ws.freeze_panes = "A3"
    wb.save(salida)
    return salida


if __name__ == "__main__":
    df = auditar()
    if df.empty:
        print("Sin hallazgos: el reporte es internamente coherente.")
    else:
        print(df.groupby(["TIPO", "SEVERIDAD"]).size().to_string())
        print(f"\nTOTAL: {len(df)} hallazgos en {df['PREDIO'].nunique()} predios")
        print(f"-> {exportar(df)}")
