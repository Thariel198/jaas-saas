"""
main.py — Ciclo de vida de reclamos (4b_reclamos)

USO:
    python main.py --mes 2026-06

LOGICA:
    1. Lee 4_pagos/efectivo/outputs/pagos_efectivo.xlsx → filtra CATEGORIA="reclamo"
       (fallback jul-2026: CATEGORIA vacía + COMENTARIO ≈ "reclamo").
    2. Lee outputs/reclamos_YYYY-MM.xlsx existente → preserva RECLAMO/TIPO_RECLAMO/
       RESOLUCION/ESTADO/FECHA_RESOLUCION por clave (MESA, MZ, LT, FECHA_COBRO).
    3. Arrastra PENDIENTE/EN_REVISION sin match del mes actual.
    4. Si existe outputs/resolucion_reclamos_YYYY-MM.xlsx, sincroniza ESTADO/
       FECHA_RESOLUCION/RESOLUCION desde ese archivo hacia reclamos.xlsx — el
       supervisor solo toca resolucion_reclamos.xlsx, este script lo propaga.
       Además mergea CAMPO/VALOR_ACTUAL/VALOR_A_CORREGIR para las filas RESUELTO.
    5. RESUELTO/RECHAZADO → trazabilidad/trazabilidad_reclamos.xlsx (acumula, nunca borra).
       Para RESUELTO: registra CAMPO/VALOR_ANTERIOR/VALOR_APLICADO de la corrección.
       Para RECHAZADO: esas tres columnas quedan vacías.
    6. Escribe vista operacional: outputs/reclamos_YYYY-MM.xlsx (solo PENDIENTE+EN_REVISION)
       con los ESTADO ya sincronizados.

Contratos visuales: docs/formato_reclamos.html
                    docs/formato_trazabilidad_reclamos.html
                    docs/formato_resolucion_reclamos.html
                    docs/diagrama_4b_reclamos.html
"""

import argparse
import logging
import shutil
import sys
from datetime import date, datetime
from pathlib import Path

import sys
sys.path.insert(0, str(Path(__file__).parent.parent / "shared"))
from utils_lote import leer_correcciones_lote

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

log = logging.getLogger(__name__)

BASE_DIR    = Path(__file__).parent
OUTPUTS_DIR = BASE_DIR / "outputs"
TRAZAB_DIR  = BASE_DIR / "trazabilidad"
LOGS_DIR    = BASE_DIR / "logs"

# Cruza límites de módulo: pagos_efectivo vive en 4_pagos/efectivo/outputs/ y
# desde 2026-08 lleva el ciclo en el nombre (ver shared/ciclo.py).
sys.path.insert(0, str(BASE_DIR.parent / "shared"))
import ciclo as ciclo_activo  # noqa: E402

_MES_CICLO = ciclo_activo.activo(default=None,
                                 path=BASE_DIR.parent / "shared" / "ciclo_activo.json")
_EFEC_DIR = BASE_DIR.parent / "4_pagos" / "efectivo" / "outputs"
PAGOS_EFECTIVO_PATH = (
    _EFEC_DIR / "pagos_efectivo.xlsx" if _MES_CICLO is None else
    ciclo_activo.resolver(_EFEC_DIR, "pagos_efectivo", _MES_CICLO,
                          legacy_sin_periodo=ciclo_activo.acepta_legacy(_MES_CICLO))
)

# Reclamos que llegan por fuera del sistema (secretaria, verbal, WhatsApp) — no
# nacen de un cobro. Writer único humano (nadie lo regenera); se sintetiza
# MESA/FECHA_COBRO para reusar la misma clave de identidad de _cargar_detectados.
RECLAMOS_MANUALES_PATH = BASE_DIR / "inputs" / "reclamos_manuales.xlsx"

# ── Paleta (del contrato docs/formato_reclamos.html y formato_trazabilidad_reclamos.html) ────

# (header_bg, header_fg, data_bg, data_fg)
SEC_PREDIO = ("EBF5FB", "1A5276", "F4FAFF", "1A5276")
SEC_COBRO  = ("E9F7EF", "1E5C3A", "F4FBF7", "1E5C3A")
SEC_MES    = ("F3F4F6", "374151", "F9FAFB", "374151")
SEC_TIPO   = ("FEF3C7", "92400E", "FFFBEB", "92400E")
SEC_MANUAL = ("F3E8FF", "5B21B6", "FAF5FF", "5B21B6")
SEC_CORR   = ("FEF0E0", "7C2D12", "FFFAF5", "7C2D12")  # Corrección aplicada (naranja)
SEC_CIERRE = ("D6EAF8", "1A5276", "EBF5FB", "1A5276")

ESTADOS_VALIDOS = ["PENDIENTE", "EN_REVISION", "RESUELTO", "RECHAZADO", "INFORMADO"]
ESTADOS_ACTIVOS = {"PENDIENTE", "EN_REVISION"}
ESTADOS_CERRADOS = {"RESUELTO", "RECHAZADO", "INFORMADO"}

# Estados cuya corrección persiste en trazabilidad con CAMPO + VALOR_ANTERIOR.
# RESUELTO también lleva VALOR_APLICADO; INFORMADO no (no hubo corrección al repo).
ESTADOS_CON_CAMPO = {"RESUELTO", "INFORMADO"}

# Clasificación del supervisor — qué componente de DATA_boletas está disputado
TIPOS_RECLAMO_VALIDOS = [
    "mes_anterior", "mes_actual", "convenio", "multa",
    "corte_reconexion", "cuota", "mantenimiento",
]

# ── Rutas ─────────────────────────────────────────────────────────────────────

def _ruta_vista(mes: str) -> Path:
    return OUTPUTS_DIR / f"reclamos_{mes}.xlsx"

def _ruta_trazab() -> Path:
    return TRAZAB_DIR / "trazabilidad_reclamos.xlsx"

def _ruta_pagos() -> Path:
    return PAGOS_EFECTIVO_PATH

def _ruta_resolucion(mes: str) -> Path:
    return OUTPUTS_DIR / f"resolucion_reclamos_{mes}.xlsx"

def _ruta_manuales() -> Path:
    return RECLAMOS_MANUALES_PATH

def _backup_con_timestamp(mes: str) -> Path | None:
    """
    Fase 1 del patrón de preservación: copia reclamos_{mes}.xlsx a
    backup/reclamos/reclamos_{mes}_YYYY-MM-DD_HHMMSS.xlsx antes de sobreescribir.
    Retorna el Path del backup, o None si el archivo no existía.
    """
    src = _ruta_vista(mes)
    if not src.exists():
        return None
    backup_dir = BASE_DIR / "backup" / "reclamos"
    backup_dir.mkdir(parents=True, exist_ok=True)
    ts   = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    dest = backup_dir / f"{src.stem}_{ts}{src.suffix}"
    shutil.copy2(src, dest)
    return dest

# ── Normalización ─────────────────────────────────────────────────────────────

def _norm(v) -> str:
    return str(v).strip().upper().replace(" ", "")

def _norm_fecha(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s in ("", "nan", "None", "NaT"):
        return ""
    try:
        if hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        dt = pd.to_datetime(v)
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return s

def _pres_key(mesa, mz, lt, fecha) -> tuple:
    """Clave 'evento reclamo': identifica un cobro único (mesa+predio+fecha).
    Un evento puede tener N correcciones (N TIPO_RECLAMO distintos)."""
    return (_norm(mesa), _norm(mz), _norm(lt), _norm_fecha(fecha))

def _full_key(mesa, mz, lt, fecha, tipo_reclamo) -> tuple:
    """Clave 'corrección única': evento + TIPO_RECLAMO.
    Permite que X-10/mesa_3 tenga multa + mes_actual como dos correcciones separadas."""
    return (_norm(mesa), _norm(mz), _norm(lt), _norm_fecha(fecha), _norm(tipo_reclamo))

def _to_datetime(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "nan", "None", "NaT"):
        return None
    try:
        if isinstance(v, datetime):
            return v
        if isinstance(v, date):
            return datetime(v.year, v.month, v.day)
        return pd.to_datetime(v).to_pydatetime()
    except Exception:
        return None

def _clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("nan", "None", "NaT") else s

def _to_numeric(v):
    """Convierte a float. Retorna None si no es numérico → celda vacía en Excel."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "nan", "None", "NaT"):
        return None
    try:
        return float(s.replace(",", ""))
    except (ValueError, TypeError):
        return None

# ── Carga de datos ────────────────────────────────────────────────────────────

def _cargar_detectados(mes: str) -> pd.DataFrame:
    p = _ruta_pagos()
    if not p.exists():
        log.warning(f"pagos_efectivo.xlsx no encontrado: {p}")
        return pd.DataFrame()

    # header=1: fila 0 son secciones merged, fila 1 son los nombres de columna
    df = pd.read_excel(p, header=1, dtype=str)
    if "COMENTARIO" not in df.columns:
        log.warning("pagos_efectivo.xlsx no tiene columna COMENTARIO -> 0 reclamos detectados")
        return pd.DataFrame()

    # Detección por CATEGORIA=reclamo (schema v3 de mesa_N). Fallback de
    # transición jul-2026: CATEGORIA vacía + COMENTARIO contiene "reclamo"
    # (mesas llenadas antes de la migración) — retirar el fallback en agosto.
    mask_coment = df["COMENTARIO"].str.contains("reclamo", case=False, na=False)
    if "CATEGORIA" in df.columns:
        cat = df["CATEGORIA"].fillna("").astype(str).str.strip().str.lower()
        mask_fallback = (cat == "") & mask_coment
        if mask_fallback.any():
            log.warning(f"{int(mask_fallback.sum())} reclamo(s) detectados por COMENTARIO "
                        f"(fallback) — marcar CATEGORIA=reclamo en la mesa")
        mask = (cat == "reclamo") | mask_fallback
    else:
        log.warning("pagos_efectivo.xlsx sin columna CATEGORIA (schema viejo) "
                    "-> deteccion solo por COMENTARIO")
        mask = mask_coment
    det = df[mask].copy()
    if det.empty:
        log.info("Ningún reclamo detectado en pagos_efectivo.xlsx")
        return pd.DataFrame()

    # pagos_efectivo usa "FECHA"; reclamos usa "FECHA_COBRO"
    if "FECHA" in det.columns and "FECHA_COBRO" not in det.columns:
        det = det.rename(columns={"FECHA": "FECHA_COBRO"})

    det["MES_ANO_DETECTADO"] = mes
    det["MES_ANO_ORIGEN"]    = mes
    det["TIPO_RECLAMO"]      = ""  # supervisor clasifica
    # RECLAMO se auto-pobla con el COMENTARIO del cobrador — el supervisor puede editar
    det["RECLAMO"]           = det["COMENTARIO"].fillna("").astype(str)
    det["RESOLUCION"]        = ""
    det["ESTADO"]            = "PENDIENTE"
    det["FECHA_RESOLUCION"]  = ""

    for c in ["MZ", "LT", "MESA", "COBRADOR", "FECHA_COBRO", "MONTO",
              "MES_ANO_DETECTADO", "MES_ANO_ORIGEN",
              "TIPO_RECLAMO", "RECLAMO", "RESOLUCION", "ESTADO", "FECHA_RESOLUCION"]:
        if c not in det.columns:
            det[c] = ""

    # Aplicar remapeos de correcciones_lote.xlsx (ej. B-21 → B-14)
    correcciones = leer_correcciones_lote()
    if correcciones:
        for i, row in det.iterrows():
            key = (_norm(str(row.get("MZ", ""))), _norm(str(row.get("LT", ""))))
            if key in correcciones:
                det.at[i, "MZ"] = correcciones[key][0]
                det.at[i, "LT"] = correcciones[key][1]

    return det[["MZ", "LT", "MESA", "COBRADOR", "FECHA_COBRO", "MONTO",
                "MES_ANO_DETECTADO", "MES_ANO_ORIGEN",
                "TIPO_RECLAMO", "RECLAMO", "RESOLUCION", "ESTADO", "FECHA_RESOLUCION"]].reset_index(drop=True)


def _cargar_manuales(mes: str) -> pd.DataFrame:
    """Reclamos tipeados a mano en inputs/reclamos_manuales.xlsx (llegan por fuera
    del sistema: secretaria, verbal, WhatsApp) — no nacen de un cobro en
    pagos_efectivo.xlsx, así que no tienen MESA/COBRADOR/MONTO reales. Se
    sintetizan (MESA="MANUAL", FECHA_COBRO=FECHA del reporte) para producir el
    mismo shape que _cargar_detectados y reusar sin cambios toda la maquinaria
    de preservación/arrastre/dedup/trazabilidad (clave = MESA+MZ+LT+FECHA_COBRO).
    """
    p = _ruta_manuales()
    if not p.exists():
        return pd.DataFrame()

    df = pd.read_excel(p, header=1, dtype=str)
    df = df.fillna("")
    df = df[df.get("MZ", "").astype(str).str.strip() != ""]
    if df.empty:
        return pd.DataFrame()

    det = pd.DataFrame()
    det["MZ"]                = df["MZ"]
    det["LT"]                = df["LT"]
    det["MESA"]               = "MANUAL"
    det["COBRADOR"]           = df["QUIEN_REPORTA"] if "QUIEN_REPORTA" in df.columns else ""
    det["FECHA_COBRO"]        = df["FECHA"] if "FECHA" in df.columns else ""
    det["MONTO"]              = ""
    det["MES_ANO_DETECTADO"]  = mes
    det["MES_ANO_ORIGEN"]     = mes
    det["TIPO_RECLAMO"]       = ""
    det["RECLAMO"]            = df["RECLAMO"] if "RECLAMO" in df.columns else ""
    det["RESOLUCION"]         = ""
    det["ESTADO"]             = "PENDIENTE"
    det["FECHA_RESOLUCION"]   = ""

    # Aplicar remapeos de correcciones_lote.xlsx (igual que _cargar_detectados)
    correcciones = leer_correcciones_lote()
    if correcciones:
        for i, row in det.iterrows():
            key = (_norm(str(row.get("MZ", ""))), _norm(str(row.get("LT", ""))))
            if key in correcciones:
                det.at[i, "MZ"] = correcciones[key][0]
                det.at[i, "LT"] = correcciones[key][1]

    return det[["MZ", "LT", "MESA", "COBRADOR", "FECHA_COBRO", "MONTO",
                "MES_ANO_DETECTADO", "MES_ANO_ORIGEN",
                "TIPO_RECLAMO", "RECLAMO", "RESOLUCION", "ESTADO", "FECHA_RESOLUCION"]].reset_index(drop=True)


def _mes_anterior(mes: str) -> str:
    y, m = mes.split("-")
    y, m = int(y), int(m)
    if m == 1:
        return f"{y - 1}-12"
    return f"{y}-{m - 1:02d}"


def _leer_vista(path: Path) -> pd.DataFrame:
    # header=1 → la fila 1 (secciones merged) se ignora; la fila 2 son los nombres de columna
    try:
        df = pd.read_excel(path, sheet_name="Reclamos", header=1, dtype=str)
        return df.fillna("")
    except Exception as e:
        log.warning(f"No se pudo leer {path.name}: {e}")
        return pd.DataFrame()


def _cargar_existente(mes: str) -> pd.DataFrame:
    """
    Si existe reclamos_{mes}.xlsx (re-corrida del mismo mes), lo usa — ya contiene
    los arrastres consolidados de meses previos.

    Si NO existe (primera corrida del mes), busca reclamos_{mes-1}.xlsx para que
    los PENDIENTE/EN_REVISION del mes anterior se arrastren automáticamente.
    """
    p_actual = _ruta_vista(mes)
    if p_actual.exists():
        df = _leer_vista(p_actual)
        log.info(f"Estado del mes {mes}: {len(df)} filas (re-corrida)")
        return df

    mes_prev = _mes_anterior(mes)
    p_prev = _ruta_vista(mes_prev)
    if p_prev.exists():
        df = _leer_vista(p_prev)
        log.info(f"Estado heredado de {mes_prev}: {len(df)} filas (primera corrida del mes)")
        return df

    return pd.DataFrame()

# ── Merge: aplicar columnas manuales del supervisor ──────────────────────────

def _aplicar_manual(detectados: pd.DataFrame, existente: pd.DataFrame) -> pd.DataFrame:
    """
    Merge trabajo manual del supervisor desde existente hacia detectados.

    Si existente tiene N filas con misma clave_base pero distinto TIPO_RECLAMO
    (caso: supervisor copió la fila para registrar 2 correcciones en un solo predio),
    expande detectados emitiendo N filas, una por cada TIPO_RECLAMO existente.
    """
    if existente.empty:
        return detectados

    # Indexa existente por clave_base — colecta TODAS las filas por evento
    by_base = {}
    for _, row in existente.iterrows():
        bk = _pres_key(row.get("MESA", ""), row.get("MZ", ""),
                       row.get("LT", ""), row.get("FECHA_COBRO", ""))
        estado = _clean(row.get("ESTADO", "PENDIENTE"))
        if estado not in ESTADOS_VALIDOS:
            estado = "PENDIENTE"
        by_base.setdefault(bk, []).append({
            "TIPO_RECLAMO":     _clean(row.get("TIPO_RECLAMO", "")),
            "RECLAMO":          _clean(row.get("RECLAMO", "")),
            "RESOLUCION":       _clean(row.get("RESOLUCION", "")),
            "ESTADO":           estado,
            "FECHA_RESOLUCION": _clean(row.get("FECHA_RESOLUCION", "")),
            "MES_ANO_ORIGEN":   _clean(row.get("MES_ANO_ORIGEN", "")),
        })

    cols = list(detectados.columns)
    salida = []
    for _, det in detectados.iterrows():
        bk = _pres_key(det.get("MESA", ""), det.get("MZ", ""),
                       det.get("LT", ""), det.get("FECHA_COBRO", ""))
        manuales = by_base.get(bk, [])
        if not manuales:
            salida.append(det.to_dict())
            continue
        # Expande: una fila por cada (TIPO_RECLAMO) en existente
        for m in manuales:
            fila = det.to_dict()
            if m["TIPO_RECLAMO"]:
                fila["TIPO_RECLAMO"] = m["TIPO_RECLAMO"]
            # RECLAMO ya no se preserva: desde que mesa_N.xlsx tiene CATEGORIA, el
            # detalle nace en el input y debe reflejar siempre el COMENTARIO fresco
            # (ver docs/aprendizaje/.../placeholder_disfrazado_de_manual_20260710.html)
            fila["RESOLUCION"]       = m["RESOLUCION"]
            fila["ESTADO"]           = m["ESTADO"]
            fila["FECHA_RESOLUCION"] = m["FECHA_RESOLUCION"]
            if m["MES_ANO_ORIGEN"]:
                fila["MES_ANO_ORIGEN"] = m["MES_ANO_ORIGEN"]
            salida.append(fila)

    df = pd.DataFrame(salida)
    # Preserva orden de columnas original
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    return df[cols].reset_index(drop=True)


def _cargar_arrastres(existente: pd.DataFrame, detectados: pd.DataFrame, mes: str) -> pd.DataFrame:
    if existente.empty:
        return pd.DataFrame()

    claves_det = set()
    for _, row in detectados.iterrows():
        claves_det.add(_pres_key(row["MESA"], row["MZ"], row["LT"], row["FECHA_COBRO"]))

    arrastres = []
    for _, row in existente.iterrows():
        k = _pres_key(row.get("MESA", ""), row.get("MZ", ""),
                      row.get("LT", ""), row.get("FECHA_COBRO", ""))
        estado = _clean(row.get("ESTADO", "PENDIENTE"))
        if k not in claves_det and estado in ESTADOS_ACTIVOS:
            d = {c: _clean(row.get(c, "")) for c in row.index}
            d["MES_ANO_DETECTADO"] = mes  # se re-detecta este mes
            arrastres.append(d)

    if not arrastres:
        return pd.DataFrame()

    df = pd.DataFrame(arrastres)
    for c in ["MZ", "LT", "MESA", "COBRADOR", "FECHA_COBRO", "MONTO",
              "MES_ANO_DETECTADO", "MES_ANO_ORIGEN",
              "TIPO_RECLAMO", "RECLAMO", "RESOLUCION", "ESTADO", "FECHA_RESOLUCION"]:
        if c not in df.columns:
            df[c] = ""
    return df[["MZ", "LT", "MESA", "COBRADOR", "FECHA_COBRO", "MONTO",
               "MES_ANO_DETECTADO", "MES_ANO_ORIGEN",
               "TIPO_RECLAMO", "RECLAMO", "RESOLUCION", "ESTADO", "FECHA_RESOLUCION"]].reset_index(drop=True)

# ── Depuración de duplicados por FECHA_COBRO corregida a mano ──────────────────

def _reconciliar_duplicados(todas: pd.DataFrame, detectados: pd.DataFrame, mes: str):
    """
    Reconcilia filas duplicadas por predio nacidas de una FECHA_COBRO corregida a mano.

    El reclamo se identifica por (MESA,MZ,LT,FECHA_COBRO). Si el cobrador corrige o
    completa la FECHA de un pago en mesa_N DESPUÉS de la primera detección, la fila
    vieja queda huérfana (arrastre) y la detección fresca entra como nueva → dos filas
    del mismo predio que nunca se vuelven a unir. Este paso las reconcilia ANTES de
    escribir reclamos.xlsx, así el próximo run no las regenera.

    Criterio por grupo (MESA,MZ,LT) — todo automático salvo el caso ambiguo:
      grupo de 1 fila                         → intacto (arrastre legítimo)
      grupo >1:
        confirmada = su FECHA_COBRO está en el input fresco (detectados) del predio
        stale      = NO confirmada + MES_ANO_ORIGEN == mes (mismo cobro, fecha corregida)
        · hay ≥1 confirmada Y ≥1 stale → conserva la confirmada, migra trabajo manual
          de stale→confirmada, elimina la stale (motivo FECHA_CORREGIDA)
        · arrastre legítimo de mes anterior (MES_ANO_ORIGEN ≠ mes) → nunca se toca
        · 2+ confirmadas con fechas distintas (dos pagos reales) → intacto, legítimo
        · grupo >1 sin confirmada que lo ancle y con algo originado este mes → REVISAR
          (no se elimina; queda en el log para mirada humana)

    No escribe archivos — función pura. Retorna (df_limpio, log_entries).
    """
    if todas.empty:
        return todas, []

    # Fechas confirmadas por el input fresco, por predio
    fresh = {}
    for _, r in detectados.iterrows():
        pk = (_norm(r.get("MESA", "")), _norm(r.get("MZ", "")), _norm(r.get("LT", "")))
        fresh.setdefault(pk, set()).add(_norm_fecha(r.get("FECHA_COBRO", "")))

    MANUALES = ["TIPO_RECLAMO", "RESOLUCION", "ESTADO", "FECHA_RESOLUCION"]
    mes_n = _norm(mes)

    grupos = {}
    for idx, r in todas.iterrows():
        pk = (_norm(r.get("MESA", "")), _norm(r.get("MZ", "")), _norm(r.get("LT", "")))
        grupos.setdefault(pk, []).append(idx)

    drop_idx = []
    log_entries = []

    def _reg(i, keep, motivo, migro):
        return {
            "mesa": _clean(todas.at[i, "MESA"]), "mz": _clean(todas.at[i, "MZ"]),
            "lt": _clean(todas.at[i, "LT"]),
            "fecha_del": _norm_fecha(todas.at[i, "FECHA_COBRO"]) or "(vacía)",
            "fecha_keep": (_norm_fecha(todas.at[keep, "FECHA_COBRO"]) or "(vacía)") if keep is not None else "",
            "motivo": motivo, "migro": migro,
        }

    for pk, idxs in grupos.items():
        if len(idxs) < 2:
            continue
        fresh_fechas = fresh.get(pk, set())
        confirmadas = [i for i in idxs
                       if fresh_fechas and _norm_fecha(todas.at[i, "FECHA_COBRO"]) in fresh_fechas]
        stale = [i for i in idxs
                 if i not in confirmadas and _norm(todas.at[i, "MES_ANO_ORIGEN"]) == mes_n]

        if confirmadas and stale:
            keep = confirmadas[0]  # ancla = detección fresca
            for i in stale:
                migro = False
                for c in MANUALES:
                    keep_v, stale_v = _clean(todas.at[keep, c]), _clean(todas.at[i, c])
                    if not stale_v:
                        continue
                    # ESTADO nunca está vacío: migrar si la fresca sigue en PENDIENTE
                    # y la stale trae un estado más avanzado (el supervisor lo trabajó).
                    needs = (not keep_v) or (c == "ESTADO" and keep_v == "PENDIENTE" and stale_v != "PENDIENTE")
                    if needs:
                        todas.at[keep, c] = stale_v
                        migro = True
                drop_idx.append(i)
                log_entries.append(_reg(i, keep, "FECHA_CORREGIDA", migro))
        elif not confirmadas and any(_norm(todas.at[i, "MES_ANO_ORIGEN"]) == mes_n for i in idxs):
            # grupo >1 sin ancla fresca pero con algo de este mes → no se puede decidir
            for i in idxs:
                log_entries.append(_reg(i, None, "REVISAR", False))

    if drop_idx:
        todas = todas.drop(index=drop_idx).reset_index(drop=True)
    return todas, log_entries


def _append_log_duplicados(entries: list, mes: str) -> None:
    """Log de diagnóstico append-only. NO es auditoría de negocio: registra qué borró
    el sistema al reconciliar duplicados, para poder rastrear si un reclamo desaparece.
    (ver docs/aprendizaje/.../auditoria_negocio_vs_log_diagnostico_20260710.html)"""
    if not entries:
        return
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    with open(LOGS_DIR / "duplicados.log", "a", encoding="utf-8") as f:
        for e in entries:
            predio = f"{e['mesa']} {e['mz']}-{e['lt']}"
            if e["motivo"] == "REVISAR":
                f.write(f"{ts} · {mes} · {predio} · REVISAR · "
                        f"fecha {e['fecha_del']} sin confirmar en input · no se tocó\n")
            else:
                f.write(f"{ts} · {mes} · {predio} · eliminada {e['fecha_del']} · "
                        f"conservada {e['fecha_keep'] or '—'} · {e['motivo']} · "
                        f"migro_manual={'SI' if e['migro'] else 'NO'}\n")


# ── Escritura Excel ───────────────────────────────────────────────────────────

def _argb(hex6: str) -> str:
    return "FF" + hex6.lstrip("#")

def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=_argb(hex6))

def _hdr(cell, bg, fg, texto):
    cell.value = texto
    cell.fill  = _fill(bg)
    cell.font  = Font(color=_argb(fg), bold=True, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def _dat(cell, valor, bg, fg, align="center", fmt=None):
    cell.value = valor
    cell.fill  = _fill(bg)
    cell.font  = Font(color=_argb(fg), size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt


# Definición de columnas: (nombre, sección, ancho, alineación, formato_número)
_COLS_VISTA = [
    ("MZ",                SEC_PREDIO, 6,  "center", None),
    ("LT",                SEC_PREDIO, 7,  "center", None),
    ("MESA",              SEC_COBRO,  12, "center", None),
    ("COBRADOR",          SEC_COBRO,  22, "left",   None),
    ("FECHA_COBRO",       SEC_COBRO,  14, "center", "DD/MM/YYYY"),
    ("MONTO",             SEC_COBRO,  12, "right",  '"S/ "#,##0.00'),
    ("MES_ANO_DETECTADO", SEC_MES,    16, "center", None),
    ("MES_ANO_ORIGEN",    SEC_MES,    14, "center", None),
    ("TIPO_RECLAMO",      SEC_TIPO,   20, "center", None),
    ("RECLAMO",           SEC_MANUAL, 40, "left",   None),
    ("RESOLUCION",        SEC_MANUAL, 40, "left",   None),
    ("ESTADO",            SEC_MANUAL, 14, "center", None),
    ("FECHA_RESOLUCION",  SEC_MANUAL, 16, "center", "DD/MM/YYYY"),
]

_SECCIONES_VISTA = [
    ("¿Cuál es el predio?",    "MZ",               "LT"),
    ("¿Quién cobró?",          "MESA",              "MONTO"),
    ("Período",                "MES_ANO_DETECTADO", "MES_ANO_ORIGEN"),
    ("¿Qué tipo?",             "TIPO_RECLAMO",      "TIPO_RECLAMO"),
    ("Reclamo y resolución",   "RECLAMO",           "FECHA_RESOLUCION"),
]

_COLS_TRAZAB = [
    ("MZ",                    SEC_PREDIO, 6,  "center", None),
    ("LT",                    SEC_PREDIO, 7,  "center", None),
    ("MESA",                  SEC_COBRO,  12, "center", None),
    ("COBRADOR",              SEC_COBRO,  22, "left",   None),
    ("FECHA_COBRO",           SEC_COBRO,  14, "center", "DD/MM/YYYY"),
    ("MONTO",                 SEC_COBRO,  12, "right",  '"S/ "#,##0.00'),
    ("MES_ANO_DETECTADO",     SEC_MES,    16, "center", None),
    ("MES_ANO_ORIGEN",        SEC_MES,    14, "center", None),
    ("TIPO_RECLAMO",          SEC_TIPO,   20, "center", None),
    ("RECLAMO",               SEC_MANUAL, 40, "left",   None),
    ("RESOLUCION",            SEC_MANUAL, 40, "left",   None),
    ("ESTADO_FINAL",          SEC_MANUAL, 14, "center", None),
    ("CAMPO",                 SEC_CORR,   24, "center", None),
    ("VALOR_ANTERIOR",        SEC_CORR,   16, "right",  '#,##0.00'),
    ("VALOR_APLICADO",        SEC_CORR,   16, "right",  '#,##0.00'),
    ("FECHA_RESOLUCION",      SEC_CIERRE, 16, "center", "DD/MM/YYYY"),
    ("MES_CIERRE",            SEC_CIERRE, 12, "center", None),
    ("FECHA_REGISTRO_CIERRE", SEC_CIERRE, 20, "center", "DD/MM/YYYY"),
]

_SECCIONES_TRAZAB = [
    ("¿Cuál es el predio?", "MZ",               "LT"),
    ("¿Quién cobró?",       "MESA",             "MONTO"),
    ("Período",             "MES_ANO_DETECTADO","MES_ANO_ORIGEN"),
    ("¿Qué tipo?",          "TIPO_RECLAMO",     "TIPO_RECLAMO"),
    ("Reclamo registrado",  "RECLAMO",          "ESTADO_FINAL"),
    ("Corrección aplicada", "CAMPO",            "VALOR_APLICADO"),
    ("Cierre",              "FECHA_RESOLUCION", "FECHA_REGISTRO_CIERRE"),
]


def _write_headers(ws, cols, secciones):
    col_idx = {c[0]: i + 1 for i, c in enumerate(cols)}
    for label, start, end in secciones:
        c1 = col_idx[start]
        c2 = col_idx[end]
        sec = cols[c1 - 1][1]
        if c1 != c2:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        _hdr(ws.cell(row=1, column=c1), sec[0], sec[1], label)
    ws.row_dimensions[1].height = 18

    for i, (nombre, sec, ancho, _, _f) in enumerate(cols, start=1):
        _hdr(ws.cell(row=2, column=i), sec[0], sec[1], nombre)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"


def _write_fila(ws, ri, row, cols):
    FECHA_COLS   = {"FECHA_COBRO", "FECHA_RESOLUCION", "FECHA_REGISTRO_CIERRE"}
    NUMERIC_COLS = {"MONTO", "VALOR_ANTERIOR", "VALOR_APLICADO",
                    "VALOR_ACTUAL", "VALOR_A_CORREGIR"}
    for ci, (nombre, sec, _, align, fmt) in enumerate(cols, start=1):
        val = row.get(nombre, "")
        if nombre in FECHA_COLS:
            val = _to_datetime(val)
        elif nombre in NUMERIC_COLS:
            val = _to_numeric(val)
        else:
            if not isinstance(val, str):
                try:
                    val = "" if pd.isna(val) else val
                except TypeError:
                    pass
            if isinstance(val, str) and val in ("nan", "None", "NaT"):
                val = ""
        _dat(ws.cell(row=ri, column=ci), val, sec[2], sec[3], align=align, fmt=fmt)


def _write_vista(df: pd.DataFrame, mes: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reclamos"

    _write_headers(ws, _COLS_VISTA, _SECCIONES_VISTA)

    for ri, (_, row) in enumerate(df.iterrows(), start=3):
        _write_fila(ws, ri, row.to_dict(), _COLS_VISTA)

    # Dropdowns: ESTADO (obligatorio) + TIPO_RECLAMO (opcional, lo clasifica el supervisor)
    if len(df) > 0:
        col_idx = {c[0]: i + 1 for i, c in enumerate(_COLS_VISTA)}

        est_col = get_column_letter(col_idx["ESTADO"])
        estados_csv = ",".join(ESTADOS_VALIDOS)
        dv_estado = DataValidation(
            type="list",
            formula1=f'"{estados_csv}"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Estado inválido",
            error=f"Use uno de: {', '.join(ESTADOS_VALIDOS)}",
        )
        ws.add_data_validation(dv_estado)
        dv_estado.sqref = f"{est_col}3:{est_col}{len(df) + 2}"

        tipo_col = get_column_letter(col_idx["TIPO_RECLAMO"])
        tipos_csv = ",".join(TIPOS_RECLAMO_VALIDOS)
        dv_tipo = DataValidation(
            type="list",
            formula1=f'"{tipos_csv}"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Tipo de reclamo inválido",
            error=f"Use uno de: {tipos_csv}",
        )
        ws.add_data_validation(dv_tipo)
        dv_tipo.sqref = f"{tipo_col}3:{tipo_col}{len(df) + 2}"

    out = _ruta_vista(mes)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    log.info(f"Vista operacional guardada: {out.name} ({len(df)} filas)")


def _build_resolucion_lookup(mes: str) -> dict:
    """
    Lee resolucion_reclamos_{mes}.xlsx (si existe) y construye un lookup:
        (MESA, MZ, LT, FECHA_COBRO) -> {
            CAMPO, VALOR_ANTERIOR, VALOR_APLICADO,   # para trazabilidad
            ESTADO, FECHA_RESOLUCION, RESOLUCION,    # para sincronizar a reclamos.xlsx
        }

    VALOR_ANTERIOR = VALOR_ACTUAL del archivo (lo que registraba DATA_boletas al cruzar).
    VALOR_APLICADO = VALOR_A_CORREGIR del archivo (lo que decidió el supervisor).

    ESTADO/FECHA_RESOLUCION/RESOLUCION son las columnas que el supervisor escribe en
    resolucion_reclamos.xlsx; main.py las propaga a reclamos.xlsx via _aplicar_resolucion.

    Si el archivo no existe (primera corrida antes de resolucion.py), retorna {}.
    """
    p = _ruta_resolucion(mes)
    if not p.exists():
        return {}
    try:
        df = pd.read_excel(p, sheet_name="Correcciones", header=1, dtype=str)
    except Exception as e:
        log.warning(f"No se pudo leer {p.name}: {e}")
        return {}

    lookup = {}
    for _, row in df.iterrows():
        k = _full_key(row.get("MESA", ""), row.get("MZ", ""),
                      row.get("LT", ""), row.get("FECHA_COBRO", ""),
                      row.get("TIPO_RECLAMO", ""))
        lookup[k] = {
            "CAMPO":            _clean(row.get("CAMPO", "")),
            "VALOR_ANTERIOR":   _clean(row.get("VALOR_ACTUAL", "")),
            "VALOR_APLICADO":   _clean(row.get("VALOR_A_CORREGIR", "")),
            "ESTADO":           _clean(row.get("ESTADO", "")),
            "FECHA_RESOLUCION": _clean(row.get("FECHA_RESOLUCION", "")),
            "RESOLUCION":       _clean(row.get("RESOLUCION", "")),
        }
    return lookup


def _aplicar_resolucion(df: pd.DataFrame, resolucion_lookup: dict) -> pd.DataFrame:
    """
    Sincroniza desde resolucion_reclamos.xlsx hacia reclamos.xlsx:
    sobrescribe ESTADO / FECHA_RESOLUCION / RESOLUCION cuando el supervisor las llenó allí.

    Regla: solo sobrescribe si en resolucion el valor es no-vacío. Vacío = "aún no decidido"
    y conserva lo que ya tiene reclamos.xlsx (ej. PENDIENTE inicial).

    Esto permite que el supervisor toque un solo archivo: resolucion_reclamos.xlsx.
    """
    if not resolucion_lookup or df.empty:
        return df

    def _override(row):
        k = _full_key(row.get("MESA", ""), row.get("MZ", ""),
                      row.get("LT", ""), row.get("FECHA_COBRO", ""),
                      row.get("TIPO_RECLAMO", ""))
        r = resolucion_lookup.get(k)
        if r is None:
            return row
        row = row.copy()
        # ESTADO: si reclamos.xlsx ya tiene un estado CERRADO, gana
        # (el supervisor lo decidió explícitamente ahí; no debe revertirse desde resolucion stale).
        estado_reclamos = str(row.get("ESTADO", "")).strip().upper()
        if estado_reclamos not in ESTADOS_CERRADOS and r["ESTADO"]:
            row["ESTADO"] = r["ESTADO"]
        if r["FECHA_RESOLUCION"]:
            row["FECHA_RESOLUCION"] = r["FECHA_RESOLUCION"]
        if r["RESOLUCION"]:
            row["RESOLUCION"] = r["RESOLUCION"]
        return row

    return df.apply(_override, axis=1)


def _trazabilidad_keys_existentes() -> set:
    """
    Retorna set de (MESA, MZ, LT, FECHA_COBRO, TIPO_RECLAMO) ya presentes en trazabilidad.
    Usa clave_completa para que X-10 con multa + mes_actual se traten como dos
    correcciones distintas (ambas se cierran, ninguna se duplica).
    """
    p = _ruta_trazab()
    if not p.exists():
        return set()
    try:
        df = pd.read_excel(p, sheet_name="Trazabilidad", header=1, dtype=str).fillna("")
    except Exception as e:
        log.warning(f"No se pudo leer trazabilidad para chequear duplicados: {e}")
        return set()
    return {
        _full_key(r.get("MESA", ""), r.get("MZ", ""),
                  r.get("LT", ""), r.get("FECHA_COBRO", ""),
                  r.get("TIPO_RECLAMO", ""))
        for _, r in df.iterrows()
    }


def _trazabilidad_base_keys() -> set:
    """
    Retorna set de (MESA, MZ, LT, FECHA_COBRO) ya presentes en trazabilidad.

    Sirve para filtrar detecciones de pagos_efectivo cuyo evento de cobro ya fue
    procesado en ciclos anteriores. Si una sub-corrección sigue pendiente, vendrá
    por _cargar_arrastres desde reclamos.xlsx existente, no por re-detección.
    """
    p = _ruta_trazab()
    if not p.exists():
        return set()
    try:
        df = pd.read_excel(p, sheet_name="Trazabilidad", header=1, dtype=str).fillna("")
    except Exception as e:
        log.warning(f"No se pudo leer trazabilidad para chequear base_keys: {e}")
        return set()
    return {
        _pres_key(r.get("MESA", ""), r.get("MZ", ""),
                  r.get("LT", ""), r.get("FECHA_COBRO", ""))
        for _, r in df.iterrows()
    }


def _append_trazabilidad(df_cerrados: pd.DataFrame, mes: str, resolucion_lookup: dict) -> None:
    if df_cerrados.empty:
        return

    # Idempotencia por clave_completa: re-correr main.py no añade duplicados
    # de RESUELTO/RECHAZADO/INFORMADO ya cerrados. Misma clave_base con distinto
    # TIPO_RECLAMO = correcciones distintas = entradas distintas.
    existentes = _trazabilidad_keys_existentes()
    if existentes:
        antes = len(df_cerrados)
        mask = df_cerrados.apply(
            lambda r: _full_key(r.get("MESA", ""), r.get("MZ", ""),
                                r.get("LT", ""), r.get("FECHA_COBRO", ""),
                                r.get("TIPO_RECLAMO", "")) not in existentes,
            axis=1,
        )
        df_cerrados = df_cerrados[mask]
        saltadas = antes - len(df_cerrados)
        if saltadas:
            log.info(f"Saltadas {saltadas} filas ya en trazabilidad (idempotencia)")
        if df_cerrados.empty:
            log.info("Todas las cerradas ya estaban en trazabilidad — nada que hacer")
            return

    hoy = date.today()
    p = _ruta_trazab()
    p.parent.mkdir(parents=True, exist_ok=True)

    if p.exists():
        wb = load_workbook(p)
        ws = wb["Trazabilidad"] if "Trazabilidad" in wb.sheetnames else wb.active
        next_row = ws.max_row + 1
        es_nuevo = False
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Trazabilidad"
        next_row = 3
        es_nuevo = True

    if es_nuevo or next_row <= 2:
        _write_headers(ws, _COLS_TRAZAB, _SECCIONES_TRAZAB)
        next_row = 3

    missing_corr = 0
    for ri, (_, row) in enumerate(df_cerrados.iterrows(), start=next_row):
        estado_final = _clean(row.get("ESTADO", ""))

        # CAMPO + VALOR_ANTERIOR: para RESUELTO e INFORMADO (informe operario los necesita).
        # VALOR_APLICADO: solo RESUELTO (INFORMADO no muta DATA_boletas).
        # RECHAZADO: las 3 columnas quedan vacías.
        if estado_final in ESTADOS_CON_CAMPO:
            k = _full_key(row.get("MESA", ""), row.get("MZ", ""),
                          row.get("LT", ""), row.get("FECHA_COBRO", ""),
                          row.get("TIPO_RECLAMO", ""))
            corr = resolucion_lookup.get(k)
            if corr is None:
                missing_corr += 1
                campo = valor_anterior = valor_aplicado = ""
            else:
                campo          = corr["CAMPO"]
                valor_anterior = corr["VALOR_ANTERIOR"]
                valor_aplicado = corr["VALOR_APLICADO"] if estado_final == "RESUELTO" else ""
        else:
            campo = valor_anterior = valor_aplicado = ""

        fila = {
            "MZ":                    _clean(row.get("MZ", "")),
            "LT":                    _clean(row.get("LT", "")),
            "MESA":                  _clean(row.get("MESA", "")),
            "COBRADOR":              _clean(row.get("COBRADOR", "")),
            "FECHA_COBRO":           _clean(row.get("FECHA_COBRO", "")),
            "MONTO":                 _clean(row.get("MONTO", "")),
            "MES_ANO_DETECTADO":     _clean(row.get("MES_ANO_DETECTADO", mes)),
            "MES_ANO_ORIGEN":        _clean(row.get("MES_ANO_ORIGEN", mes)),
            "TIPO_RECLAMO":          _clean(row.get("TIPO_RECLAMO", "")),
            "RECLAMO":               _clean(row.get("RECLAMO", "")),
            "RESOLUCION":            _clean(row.get("RESOLUCION", "")),
            "ESTADO_FINAL":          estado_final,
            "CAMPO":                 campo,
            "VALOR_ANTERIOR":        valor_anterior,
            "VALOR_APLICADO":        valor_aplicado,
            "FECHA_RESOLUCION":      _clean(row.get("FECHA_RESOLUCION", "")),
            "MES_CIERRE":            mes,
            "FECHA_REGISTRO_CIERRE": str(hoy),
        }
        _write_fila(ws, ri, fila, _COLS_TRAZAB)

    wb.save(p)
    log.info(f"Trazabilidad: +{len(df_cerrados)} filas -> {p.name}")
    if missing_corr:
        log.warning(
            f"{missing_corr} reclamos RESUELTO sin corrección en resolucion_reclamos_{mes}.xlsx "
            "— CAMPO/VALOR_ANTERIOR/VALOR_APLICADO quedaron vacíos. "
            "¿Corriste resolucion.py antes de cerrar?"
        )


# ── Main ──────────────────────────────────────────────────────────────────────

def main(mes: str) -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
        force=True,
    )
    log.info(f"=== 4b_reclamos/main.py — mes {mes} ===")

    detectados = _cargar_detectados(mes)
    manuales   = _cargar_manuales(mes)
    log.info(f"Detectados en pagos_efectivo: {len(detectados)} · manuales (fuera del sistema): {len(manuales)}")
    if not manuales.empty:
        detectados = pd.concat([detectados, manuales], ignore_index=True) if not detectados.empty else manuales

    # Eventos ya procesados en ciclos anteriores: skip re-detección.
    # Si alguna sub-corrección sigue activa, vendrá por _cargar_arrastres desde
    # reclamos.xlsx existente (no por pagos_efectivo).
    trazab_base = _trazabilidad_base_keys()
    if trazab_base and not detectados.empty:
        antes = len(detectados)
        mask = detectados.apply(
            lambda r: _pres_key(r.get("MESA", ""), r.get("MZ", ""),
                                r.get("LT", ""), r.get("FECHA_COBRO", "")) not in trazab_base,
            axis=1,
        )
        detectados = detectados[mask].reset_index(drop=True)
        saltados = antes - len(detectados)
        if saltados:
            log.info(f"Saltados {saltados} eventos ya procesados en trazabilidad")

    # Fase 2: leer trabajo manual existente (RECLAMO, RESOLUCION, ESTADO, FECHA_RESOLUCION)
    existente = _cargar_existente(mes)

    # Limpiar existente: filas ya procesadas en trazabilidad o stale
    # (caso típico: reclamos.xlsx quedó con filas de detección pre-filtro).
    # Regla: SKIP si (full_key en trazab) O (TIPO_RECLAMO vacío Y base_key en trazab).
    # KEEP si TIPO_RECLAMO clasificado y full_key NO en trazab → sub-corrección pendiente legítima.
    if trazab_base and not existente.empty:
        trazab_full = _trazabilidad_keys_existentes()
        antes = len(existente)
        def _conservar(r):
            tipo = _clean(r.get("TIPO_RECLAMO", ""))
            bk = _pres_key(r.get("MESA", ""), r.get("MZ", ""),
                           r.get("LT", ""), r.get("FECHA_COBRO", ""))
            fk = _full_key(r.get("MESA", ""), r.get("MZ", ""),
                           r.get("LT", ""), r.get("FECHA_COBRO", ""), tipo)
            if fk in trazab_full:
                return False  # esta corrección ya está cerrada en trazab
            if not tipo and bk in trazab_base:
                return False  # detección stale pre-filtro de un evento ya procesado
            return True
        existente = existente[existente.apply(_conservar, axis=1)].reset_index(drop=True)
        depurados = antes - len(existente)
        if depurados:
            log.info(f"Limpiadas {depurados} filas stale del estado del mes (ya en trazabilidad)")

    # Fase 1: backup antes de sobreescribir (el archivo ya fue leído en Fase 2)
    backup = _backup_con_timestamp(mes)
    if backup:
        log.info(f"Backup: {backup.name}")

    # Fase 3: preservar columnas manuales al regenerar
    if not detectados.empty and not existente.empty:
        detectados = _aplicar_manual(detectados, existente)

    arrastres = _cargar_arrastres(existente, detectados, mes)
    if not arrastres.empty:
        log.info(f"Arrastres de mes anterior: {len(arrastres)}")

    fuentes = [df for df in [detectados, arrastres] if not df.empty]
    if not fuentes:
        log.info("Sin reclamos — se escribe vista vacía")
        _write_vista(pd.DataFrame(columns=[c[0] for c in _COLS_VISTA]), mes)
        return

    todas = pd.concat(fuentes, ignore_index=True)

    # Aplicar correcciones de lote también a arrastres heredados de ciclos anteriores
    correcciones = leer_correcciones_lote()
    if correcciones:
        for i, row in todas.iterrows():
            key = (_norm(str(row.get("MZ", ""))), _norm(str(row.get("LT", ""))))
            if key in correcciones:
                todas.at[i, "MZ"] = correcciones[key][0]
                todas.at[i, "LT"] = correcciones[key][1]
    if "RESOLUCION" not in todas.columns:
        todas["RESOLUCION"] = ""

    # Sync ESTADO/FECHA_RESOLUCION/RESOLUCION desde resolucion_reclamos.xlsx
    # (el supervisor solo toca ese archivo; main.py propaga aquí)
    resolucion_lookup = _build_resolucion_lookup(mes)
    if resolucion_lookup:
        log.info(f"Correcciones disponibles en resolucion_reclamos_{mes}.xlsx: {len(resolucion_lookup)}")
        todas = _aplicar_resolucion(todas, resolucion_lookup)

    # Reconciliar duplicados por FECHA_COBRO corregida a mano (canal efectivo).
    # Corrige en memoria antes de escribir → el próximo run no los regenera.
    todas, dup_log = _reconciliar_duplicados(todas, detectados, mes)
    if dup_log:
        _append_log_duplicados(dup_log, mes)
        elim  = [e for e in dup_log if e["motivo"] == "FECHA_CORREGIDA"]
        revis = [e for e in dup_log if e["motivo"] == "REVISAR"]
        if elim:
            predios = ", ".join(f"{e['mesa']} {e['mz']}-{e['lt']}" for e in elim[:8])
            extra = f" (+{len(elim) - 8} más)" if len(elim) > 8 else ""
            log.info(f"Depurados {len(elim)} duplicados por fecha corregida: {predios}{extra}")
            log.info("  detalle -> logs/duplicados.log")
        if revis:
            pr = sorted({(e["mesa"], e["mz"], e["lt"]) for e in revis})
            log.warning(f"[REVISAR] {len(pr)} predio(s) no se tocaron: "
                        + ", ".join(f"{m} {mz}-{lt}" for m, mz, lt in pr))

    mask_activos = todas["ESTADO"].isin(ESTADOS_ACTIVOS)
    df_activos   = todas[mask_activos].copy().reset_index(drop=True)
    df_cerrados  = todas[~mask_activos].copy().reset_index(drop=True)

    log.info(f"Activos  (PENDIENTE/EN_REVISION):           {len(df_activos)}")
    log.info(f"Cerrados (RESUELTO/RECHAZADO/INFORMADO):    {len(df_cerrados)}")

    if not resolucion_lookup and df_cerrados["ESTADO"].isin(ESTADOS_CON_CAMPO).any():
        log.warning(
            f"Hay reclamos RESUELTO/INFORMADO pero resolucion_reclamos_{mes}.xlsx no existe — "
            "corré resolucion.py antes para registrar CAMPO/VALOR_ACTUAL."
        )

    _append_trazabilidad(df_cerrados, mes, resolucion_lookup)
    _write_vista(df_activos, mes)

    log.info("=== completado ===")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Genera vista de reclamos del mes")
    parser.add_argument("--mes", required=True, help="Mes a procesar (YYYY-MM)")
    args = parser.parse_args()
    main(args.mes)
