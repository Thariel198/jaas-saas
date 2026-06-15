"""
resolucion.py — Genera hoja de corrección para reclamos clasificados (4b_reclamos)

USO:
    python resolucion.py --mes 2026-06

LOGICA:
    1. Lee outputs/reclamos_{mes}.xlsx.
    2. Lee 3_boletas/inputs/DATA_boletas.xlsx (lectura directa — valor vigente).
    3. Para cada reclamo, cruza por (MZ, LT) contra DATA_boletas:
        - NOMBRE: columna NOMBRES en DATA_boletas.
        - CAMPO + VALOR_ACTUAL: solo si TIPO_RECLAMO está clasificado.
          TIPO_RECLAMO → mapeo a columna en DATA_boletas → valor de esa columna.
    4. Preserva VALOR_A_CORREGIR + RESOLUCION del archivo previo (re-corridas).
    5. Backup + escribe outputs/resolucion_reclamos_{mes}.xlsx.

Contratos visuales: docs/formato_resolucion_reclamos.html
                    docs/diagrama_4b_reclamos.html
"""

import argparse
import logging
import shutil
import sys
from datetime import datetime
from pathlib import Path

# Permite importar main.py cuando se ejecuta desde cualquier CWD
sys.path.insert(0, str(Path(__file__).parent))

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from main import (
    BASE_DIR, OUTPUTS_DIR,
    SEC_PREDIO, SEC_COBRO, SEC_MES, SEC_TIPO, SEC_MANUAL, SEC_CIERRE,
    TIPOS_RECLAMO_VALIDOS, ESTADOS_VALIDOS,
    _norm, _clean, _pres_key,
    _write_headers, _write_fila,
)

log = logging.getLogger(__name__)

# ── Paleta extendida ──────────────────────────────────────────────────────────
# Sección "Dato en sistema" — cyan, auto-llenado por resolucion.py
SEC_AUTO = ("ECFEFF", "0E7490", "F0FDFF", "0E7490")

# ── Rutas externas (lectura directa, sin copia a inputs/) ─────────────────────
DATA_BOLETAS_PATH = BASE_DIR.parent / "3_boletas" / "inputs" / "DATA_boletas.xlsx"

# ── Mapeo TIPO_RECLAMO → columna exacta en DATA_boletas ──────────────────────
# El nombre de la columna se usa también como CAMPO (display al supervisor).
TIPO_TO_COL_BOLETAS = {
    "mes_anterior":     "MES ANTERIOR",
    "mes_actual":       "Total mes actual",
    "convenio":         "Convenio",
    "multa":            "Multa (faena + reunión)",
    "corte_reconexion": "Corte y reconexion",
    "cuota":            "Cuota directa",
    "mantenimiento":    "Mantenimiento",
}

# ── Schema (contrato: docs/formato_resolucion_reclamos.html) ─────────────────
_COLS_RESOL = [
    ("MZ",               SEC_PREDIO,  6,  "center", None),
    ("LT",               SEC_PREDIO,  7,  "center", None),
    ("NOMBRE",           SEC_PREDIO,  28, "left",   None),
    ("MESA",             SEC_COBRO,   12, "center", None),
    ("FECHA_COBRO",      SEC_COBRO,   14, "center", "DD/MM/YYYY"),
    ("MES_ANO_ORIGEN",   SEC_MES,     14, "center", None),
    ("RECLAMO",          SEC_TIPO,    40, "left",   None),
    ("TIPO_RECLAMO",     SEC_TIPO,    20, "center", None),
    ("CAMPO",            SEC_AUTO,    24, "center", None),
    ("VALOR_ACTUAL",     SEC_AUTO,    16, "right",  '#,##0.00'),
    ("VALOR_A_CORREGIR", SEC_MANUAL,  18, "right",  '#,##0.00'),
    ("RESOLUCION",       SEC_MANUAL,  44, "left",   None),
    ("ESTADO",           SEC_CIERRE,  14, "center", None),
    ("FECHA_RESOLUCION", SEC_CIERRE,  16, "center", "DD/MM/YYYY"),
]

_SECCIONES_RESOL = [
    ("¿Cuál es el predio?", "MZ",               "NOMBRE"),
    ("¿Quién cobró?",       "MESA",             "FECHA_COBRO"),
    ("Período",             "MES_ANO_ORIGEN",   "MES_ANO_ORIGEN"),
    ("¿Qué reclamaron?",    "RECLAMO",          "TIPO_RECLAMO"),
    ("Dato en sistema",     "CAMPO",            "VALOR_ACTUAL"),
    ("Corrección",          "VALOR_A_CORREGIR", "RESOLUCION"),
    ("Cierre",              "ESTADO",           "FECHA_RESOLUCION"),
]

# Columnas numéricas que merecen conversión a número (para que VALOR_ACTUAL
# del MES ANTERIOR — guardado como número en DATA_boletas — vuelva a ser número).
NUMERIC_VALOR_COLS = {"VALOR_ACTUAL", "VALOR_A_CORREGIR"}

# ── Rutas ─────────────────────────────────────────────────────────────────────

def _ruta_reclamos(mes: str) -> Path:
    return OUTPUTS_DIR / f"reclamos_{mes}.xlsx"

def _ruta_resol(mes: str) -> Path:
    return OUTPUTS_DIR / f"resolucion_reclamos_{mes}.xlsx"

def _backup_con_timestamp(mes: str) -> Path | None:
    src = _ruta_resol(mes)
    if not src.exists():
        return None
    backup_dir = BASE_DIR / "backup" / "resolucion"
    backup_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    dest = backup_dir / f"{src.stem}_{ts}{src.suffix}"
    shutil.copy2(src, dest)
    return dest

# ── Carga ─────────────────────────────────────────────────────────────────────

def _leer_reclamos(mes: str) -> pd.DataFrame:
    p = _ruta_reclamos(mes)
    if not p.exists():
        log.warning(f"reclamos_{mes}.xlsx no encontrado: {p}")
        return pd.DataFrame()
    try:
        df = pd.read_excel(p, sheet_name="Reclamos", header=1, dtype=str)
        return df.fillna("")
    except Exception as e:
        log.error(f"Error leyendo {p.name}: {e}")
        return pd.DataFrame()


def _build_boletas_lookup() -> dict:
    """
    Lookup por (MZ_norm, LT_norm) → dict completo de la fila en DATA_boletas.
    (MZ, LT) es único en el padrón — 575 predios, 0 duplicados.
    """
    p = DATA_BOLETAS_PATH
    if not p.exists():
        log.warning(f"DATA_boletas.xlsx no encontrado: {p}")
        return {}
    try:
        df = pd.read_excel(p, sheet_name="Data", dtype=str)
    except Exception as e:
        log.error(f"Error leyendo DATA_boletas.xlsx: {e}")
        return {}

    lookup = {}
    for _, row in df.iterrows():
        mz = _norm(row.get("MZ", ""))
        lt = _norm(row.get("LT", ""))
        if not mz or not lt:
            continue
        lookup[(mz, lt)] = row.to_dict()
    return lookup


def _leer_existente(mes: str) -> pd.DataFrame:
    p = _ruta_resol(mes)
    if not p.exists():
        return pd.DataFrame()
    try:
        df = pd.read_excel(p, sheet_name="Correcciones", header=1, dtype=str)
        return df.fillna("")
    except Exception as e:
        log.warning(f"No se pudo leer {p.name}: {e}")
        return pd.DataFrame()

# ── Auto-fill ────────────────────────────────────────────────────────────────

def _autofill(row: dict, boletas_lookup: dict) -> tuple[str, str, str]:
    """
    Retorna (NOMBRE, CAMPO, VALOR_ACTUAL) para la fila del reclamo.

    - Si (MZ, LT) no está en DATA_boletas → ("", "", "").
    - Si está pero TIPO_RECLAMO vacío → (NOMBRE, "", "").
    - Si está y TIPO_RECLAMO clasificado → (NOMBRE, CAMPO, VALOR_ACTUAL).
    - Si TIPO_RECLAMO no tiene mapeo conocido → warning + (NOMBRE, "", "").
    """
    mz = _norm(row.get("MZ", ""))
    lt = _norm(row.get("LT", ""))
    boleta = boletas_lookup.get((mz, lt))
    if boleta is None:
        return "", "", ""

    nombre = _clean(boleta.get("NOMBRES", ""))

    tipo = _clean(row.get("TIPO_RECLAMO", ""))
    if not tipo:
        return nombre, "", ""

    col = TIPO_TO_COL_BOLETAS.get(tipo)
    if col is None:
        log.warning(f"TIPO_RECLAMO '{tipo}' sin mapeo a DATA_boletas (MZ={mz}, LT={lt})")
        return nombre, "", ""

    valor = _clean(boleta.get(col, ""))
    return nombre, col, valor

# ── Preservación de trabajo manual ───────────────────────────────────────────

def _preservar_manual(df: pd.DataFrame, existente: pd.DataFrame) -> pd.DataFrame:
    """
    Re-corridas: preserva las columnas que el supervisor escribió en resolucion.xlsx:
        VALOR_A_CORREGIR, RESOLUCION, ESTADO, FECHA_RESOLUCION
    Clave de preservación: (MESA, MZ, LT, FECHA_COBRO).

    ESTADO/FECHA_RESOLUCION solo se sobrescriben si la versión existente tiene valor —
    así una re-corrida no degrada un RESUELTO a PENDIENTE cuando el supervisor todavía
    no ha corrido main.py para sincronizar.
    """
    if existente.empty:
        return df

    lookup = {}
    for _, row in existente.iterrows():
        k = _pres_key(row.get("MESA", ""), row.get("MZ", ""),
                      row.get("LT", ""), row.get("FECHA_COBRO", ""))
        lookup[k] = {
            "VALOR_A_CORREGIR": _clean(row.get("VALOR_A_CORREGIR", "")),
            "RESOLUCION":       _clean(row.get("RESOLUCION", "")),
            "ESTADO":           _clean(row.get("ESTADO", "")),
            "FECHA_RESOLUCION": _clean(row.get("FECHA_RESOLUCION", "")),
        }

    def _merge(row):
        k = _pres_key(row["MESA"], row["MZ"], row["LT"], row["FECHA_COBRO"])
        if k in lookup:
            row = row.copy()
            row["VALOR_A_CORREGIR"] = lookup[k]["VALOR_A_CORREGIR"]
            row["RESOLUCION"]       = lookup[k]["RESOLUCION"]
            if lookup[k]["ESTADO"]:
                row["ESTADO"] = lookup[k]["ESTADO"]
            if lookup[k]["FECHA_RESOLUCION"]:
                row["FECHA_RESOLUCION"] = lookup[k]["FECHA_RESOLUCION"]
        return row

    return df.apply(_merge, axis=1)

# ── Escritura ────────────────────────────────────────────────────────────────

def _write_resolucion(df: pd.DataFrame, mes: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Correcciones"

    _write_headers(ws, _COLS_RESOL, _SECCIONES_RESOL)

    for ri, (_, row) in enumerate(df.iterrows(), start=3):
        _write_fila(ws, ri, row.to_dict(), _COLS_RESOL)

    if len(df) > 0:
        col_idx = {c[0]: i + 1 for i, c in enumerate(_COLS_RESOL)}

        # Dropdown ESTADO (mismo conjunto que reclamos.xlsx — el master state
        # vive en reclamos.xlsx, esto es informativo para el supervisor)
        est_col = get_column_letter(col_idx["ESTADO"])
        dv_estado = DataValidation(
            type="list",
            formula1='"' + ",".join(ESTADOS_VALIDOS) + '"',
            allow_blank=True,
        )
        ws.add_data_validation(dv_estado)
        dv_estado.sqref = f"{est_col}3:{est_col}{len(df) + 2}"

        # Dropdown TIPO_RECLAMO (informativo, copia de reclamos.xlsx)
        tipo_col = get_column_letter(col_idx["TIPO_RECLAMO"])
        dv_tipo = DataValidation(
            type="list",
            formula1='"' + ",".join(TIPOS_RECLAMO_VALIDOS) + '"',
            allow_blank=True,
        )
        ws.add_data_validation(dv_tipo)
        dv_tipo.sqref = f"{tipo_col}3:{tipo_col}{len(df) + 2}"

    out = _ruta_resol(mes)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    log.info(f"Resolución guardada: {out.name} ({len(df)} filas)")

# ── Main ─────────────────────────────────────────────────────────────────────

def main(mes: str) -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
        force=True,
    )
    log.info(f"=== 4b_reclamos/resolucion.py — mes {mes} ===")

    reclamos = _leer_reclamos(mes)
    log.info(f"Reclamos en {mes}: {len(reclamos)}")

    cols_orden = [c[0] for c in _COLS_RESOL]

    if reclamos.empty:
        log.info("Sin reclamos — se escribe hoja vacía")
        _write_resolucion(pd.DataFrame(columns=cols_orden), mes)
        return

    boletas_lookup = _build_boletas_lookup()
    log.info(f"DATA_boletas: {len(boletas_lookup)} predios disponibles")

    # Auto-fill por fila
    rows = []
    sin_match = 0
    sin_clasificar = 0
    for _, r in reclamos.iterrows():
        # Punto de partida: limpieza de campos heredados de reclamos.xlsx
        row = {col: _clean(r.get(col, "")) for col in cols_orden}

        nombre, campo, valor = _autofill(r.to_dict(), boletas_lookup)
        row["NOMBRE"]       = nombre
        row["CAMPO"]        = campo
        row["VALOR_ACTUAL"] = valor

        if not nombre and _norm(r.get("MZ", "")) and _norm(r.get("LT", "")):
            sin_match += 1
        if not _clean(r.get("TIPO_RECLAMO", "")):
            sin_clasificar += 1

        rows.append(row)

    df = pd.DataFrame(rows, columns=cols_orden)

    if sin_match:
        log.warning(f"{sin_match} predios no encontrados en DATA_boletas — NOMBRE/CAMPO/VALOR_ACTUAL vacíos")
    if sin_clasificar:
        log.info(f"{sin_clasificar} reclamos sin TIPO_RECLAMO clasificado — CAMPO/VALOR_ACTUAL vacíos")

    # Preservar trabajo manual del supervisor antes de sobreescribir
    existente = _leer_existente(mes)
    if not existente.empty:
        df = _preservar_manual(df, existente)
        log.info(f"Preservado trabajo manual de {len(existente)} filas previas")

    # Backup
    backup = _backup_con_timestamp(mes)
    if backup:
        log.info(f"Backup: {backup.name}")

    _write_resolucion(df, mes)
    log.info("=== completado ===")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Genera hoja de corrección para reclamos clasificados")
    parser.add_argument("--mes", required=True, help="Mes a procesar (YYYY-MM)")
    args = parser.parse_args()
    main(args.mes)
