import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
ROOT          = Path(__file__).parent
COBRANZA_DIR  = ROOT.parent / "04_cobranza"
BOLETAS_DIR   = ROOT.parent / "06_boletas"

CONFIG_PATH   = ROOT / "inputs" / "config_mes.xlsx"
PLANILLA_PATH = COBRANZA_DIR / "inputs" / "planilla_base" / "planilla_base.xlsx"
OUTPUT_PATH   = BOLETAS_DIR / "inputs" / "data_boletas.xlsx"
LOG_PATH      = ROOT / "inputs" / "run.log"

COSTO_M3 = 1.0   # S/ por m3 — debe coincidir con 04_cobranza

CONFIG_REQUERIDOS = [
    "PERIODO", "FECHA_VENCIMIENTO", "FECHA_EMISION",
    "LECTURA_ANT_FECHA", "LECTURA_ACT_FECHA",
    "FECHA_PAGO", "HORA_PAGO", "LUGAR_PAGO", "TELEFONO",
    "NUMERO_RECIBO_INICIO",
]

# ── LOGGING ───────────────────────────────────────────────────────────────────
def _init_logging():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(LOG_PATH, encoding="utf-8"),
        ],
    )

log = logging.getLogger(__name__)

# ── UTILIDADES ────────────────────────────────────────────────────────────────
def _float(val) -> float:
    try:
        return float(str(val).replace(",", ".").strip())
    except (ValueError, TypeError):
        return 0.0

def _norm_lt(val) -> str:
    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN"):
        return ""
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s.upper()

def _str(val, fallback="") -> str:
    s = str(val).strip() if val is not None else ""
    return fallback if s in ("", "nan", "None", "NaT") else s

def _estado_boleta(arrastre: float, corte_reconexion: float,
                   convenio: float, reunion_faena: float, techado: float) -> str:
    deuda_acumulada = arrastre + corte_reconexion + convenio + reunion_faena + techado
    return "USTED ESTÁ AL DÍA" if deuda_acumulada <= 0.005 else "NO ESTÁ AL DÍA"

def _fecha_pago_str(config: dict) -> str:
    return (f"PAGO PRESENCIAL:  {config['FECHA_PAGO']} de {config['HORA_PAGO']} "
            f"({config['LUGAR_PAGO']}), PAGO YAPE: {config['TELEFONO']}")

# ── VALIDACIÓN ────────────────────────────────────────────────────────────────
def _validar_inputs():
    for ruta, desc in [
        (CONFIG_PATH,   "crear con crear_config.py y completar"),
        (PLANILLA_PATH, "asegurar que planilla_base.xlsx está lista para este mes"),
    ]:
        if not ruta.exists():
            raise FileNotFoundError(f"Falta: {ruta}\n  → {desc}")

    df = pd.read_excel(CONFIG_PATH, dtype=str)
    df.columns = [str(c).strip().upper() for c in df.columns]
    faltantes = set(CONFIG_REQUERIDOS) - set(df.columns)
    if faltantes:
        raise ValueError(f"config_mes.xlsx — columnas faltantes: {faltantes}")
    if df.shape[0] == 0:
        raise ValueError("config_mes.xlsx está vacío — completar con los datos del mes")
    log.info("Inputs validados correctamente")

# ── CARGA CONFIG ──────────────────────────────────────────────────────────────
def _cargar_config() -> dict:
    df = pd.read_excel(CONFIG_PATH, dtype=str)
    df.columns = [str(c).strip().upper() for c in df.columns]
    fila = df.iloc[0]
    config = {col: _str(fila.get(col, "")) for col in CONFIG_REQUERIDOS}
    config["NUMERO_RECIBO_INICIO"] = int(_float(fila.get("NUMERO_RECIBO_INICIO", 1)))
    log.info(f"Config cargada: {config['PERIODO']} — recibo inicio {config['NUMERO_RECIBO_INICIO']}")
    return config

# ── CARGA PLANILLA ────────────────────────────────────────────────────────────
def _cargar_planilla() -> list[dict]:
    df = pd.read_excel(PLANILLA_PATH, dtype=str)
    df.columns = [str(c).strip().upper() for c in df.columns]
    usuarios = []
    for _, f in df.iterrows():
        mz = str(f.get("MZ", "")).strip().upper()
        lt = _norm_lt(f.get("LT", ""))
        if not mz or not lt:
            continue
        marc_ant = _float(f.get("MARC_ANT", 0))
        marc_act = _float(f.get("MARC_ACT", 0))
        m3       = round(marc_act - marc_ant, 3)
        arrastre       = _float(f.get("ARRASTRE",       0))
        convenio       = _float(f.get("CONVENIO",       0))
        mant           = _float(f.get("MANT",           3))
        corte_reconex  = _float(f.get("CORTE_RECONEXION", 0))
        reunion_faena  = _float(f.get("REUNION_FAENA",  0))
        techado        = _float(f.get("TECHADO",        0))
        devolucion     = _float(f.get("DEVOLUCION",     0))
        ajuste         = _float(f.get("AJUSTE",         0))
        usuarios.append({
            "mz":             mz,
            "lt":             lt,
            "nombre":         str(f.get("NOMBRE", "")).strip(),
            "marc_ant":       marc_ant,
            "marc_act":       marc_act,
            "m3":             m3,
            "total_mes":      round(m3 * COSTO_M3, 2),
            "arrastre":       arrastre,
            "convenio":       convenio,
            "mant":           mant,
            "corte_reconex":  corte_reconex,
            "reunion_faena":  reunion_faena,
            "techado":        techado,
            "devolucion":     devolucion,
            "ajuste":         ajuste,
        })
    log.info(f"Planilla base: {len(usuarios)} usuarios cargados")
    return usuarios

# ── ENRIQUECER ────────────────────────────────────────────────────────────────
def _enriquecer(usuarios: list[dict], config: dict) -> list[dict]:
    fecha_pago_txt = _fecha_pago_str(config)
    recibo_inicio  = config["NUMERO_RECIBO_INICIO"]
    registros = []
    for idx, u in enumerate(usuarios):
        total = round(
            u["total_mes"] + u["arrastre"] + u["convenio"] + u["mant"]
            + u["corte_reconex"] + u["reunion_faena"] + u["techado"]
            - u["devolucion"] + u["ajuste"],
            2,
        )
        total = max(0.0, total)  # nunca negativo en el recibo
        estado = _estado_boleta(
            u["arrastre"], u["corte_reconex"],
            u["convenio"], u["reunion_faena"], u["techado"],
        )
        registros.append({
            "NUMERO DE RECIBO":    recibo_inicio + idx,
            "PERIODO":             config["PERIODO"],
            "FECHA DE VENCIMIENTO": config["FECHA_VENCIMIENTO"],
            "FECHA DE EMISIÓN":    config["FECHA_EMISION"],
            "LECTURA ANTERIOR":    config["LECTURA_ANT_FECHA"],
            "LECTURA ACTUAL":      config["LECTURA_ACT_FECHA"],
            "NOMBRES":             u["nombre"],
            "MZ":                  u["mz"],
            "LT":                  u["lt"],
            "Marcación anterior":  u["marc_ant"],
            "Marcacion altual":    u["marc_act"],
            "M3":                  u["m3"],
            "Total mes actual":    u["total_mes"],
            "MES ANTERIOR":        u["arrastre"],
            "Corte y reconexion":  u["corte_reconex"],
            "Convenio":            u["convenio"],
            "Mantenimiento":       u["mant"],
            "Total":               total,
            "Estado":              estado,
            "fecha pago":          fecha_pago_txt,
            "Multa (faena + reunión)": u["reunion_faena"],
            "Cuota directa":       u["techado"],
            "Importe a pagar":     total,
            # Variables para 06_boletas (leídas por fila, no hardcoded)
            "FECHA_PAGO":          config["FECHA_PAGO"],
            "HORA_PAGO":           config["HORA_PAGO"],
        })
    return registros

# ── EXPORTAR ──────────────────────────────────────────────────────────────────
def _exportar(registros: list[dict]):
    df = pd.DataFrame(registros)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Data", index=False)
        ws = writer.sheets["Data"]
        _estilizar(ws, len(df))
    log.info(f"data_boletas.xlsx → {len(registros)} registros → {OUTPUT_PATH}")

def _estilizar(ws, n_filas: int):
    b = Side(style="thin", color="CCCCCC")
    borde = Border(left=b, right=b, top=b, bottom=b)
    HEAD_BG = "1E3A5F"

    for cell in ws[1]:
        cell.font      = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", start_color=HEAD_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = borde
        ws.column_dimensions[cell.column_letter].width = max(
            12, len(str(cell.value or "")) + 2
        )
    ws.row_dimensions[1].height = 20

    for row in ws.iter_rows(min_row=2, max_row=n_filas + 1):
        for cell in row:
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = borde
        ws.row_dimensions[row[0].row].height = 16

    ws.freeze_panes = "A2"

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("\n" + "═" * 57)
    print("  06a_enriquecimiento — Preparación DATA_boletas")
    print("═" * 57)
    _init_logging()

    print("\n[1/4] Validando inputs...")
    _validar_inputs()

    print("\n[2/4] Cargando config y planilla base...")
    config   = _cargar_config()
    usuarios = _cargar_planilla()

    print(f"\n[3/4] Enriqueciendo {len(usuarios)} registros...")
    registros = _enriquecer(usuarios, config)

    print("\n[4/4] Exportando data_boletas.xlsx...")
    _exportar(registros)

    print("\n" + "═" * 57)
    print(f"  {len(registros)} recibos preparados — recibos {config['NUMERO_RECIBO_INICIO']}"
          f" al {config['NUMERO_RECIBO_INICIO'] + len(registros) - 1}")
    print(f"  → {OUTPUT_PATH.name}")
    print("  Siguiente paso: correr 06_boletas/main.py")
    print("═" * 57 + "\n")


if __name__ == "__main__":
    main()
