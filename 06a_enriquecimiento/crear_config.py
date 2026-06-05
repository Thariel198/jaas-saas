# Crea config_mes.xlsx vacío con los campos correctos.
# Correr una sola vez: python crear_config.py
# Luego completar los datos del mes en inputs/config_mes.xlsx.
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DESTINO = Path(__file__).parent / "inputs" / "config_mes.xlsx"

COLS = [
    # (nombre, descripción, ejemplo, ancho)
    ("PERIODO",              "Período del recibo",                     "11/03/2026 al 10/04/2026", 26),
    ("FECHA_VENCIMIENTO",    "Fecha límite de pago",                   "2026-05-02",               16),
    ("FECHA_EMISION",        "Fecha de emisión del recibo",            "2026-04-26",               16),
    ("LECTURA_ANT_FECHA",    "Fecha de lectura anterior (día/mes/año)","2026-03-10",               18),
    ("LECTURA_ACT_FECHA",    "Fecha de lectura actual (día/mes/año)",  "2026-04-10",               18),
    ("FECHA_PAGO",           "Día y mes de cobro (ej: 02/05)",         "02/05",                    12),
    ("HORA_PAGO",            "Hora de cobro",                          "4-6 pm",                   12),
    ("LUGAR_PAGO",           "Lugar de cobro presencial",              "LOCAL DEL PUEBLO",         20),
    ("TELEFONO",             "Número Yape para pagos",                 "948 227 636",              14),
    ("NUMERO_RECIBO_INICIO", "Número del primer recibo de este mes",   16900,                      20),
]

C_HEAD_BG = "1E3A5F"
C_DESC_BG = "EFF6FF"
C_EJMP_BG = "FAFAFA"


def main():
    DESTINO.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "config"
    ws.freeze_panes = "A3"

    b = Side(style="thin", color="CCCCCC")
    borde = Border(left=b, right=b, top=b, bottom=b)

    for ci, (nombre, _, _, ancho) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=nombre)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", start_color=C_HEAD_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = borde
        ws.column_dimensions[c.column_letter].width = ancho

    for ci, (_, desc, _, _) in enumerate(COLS, 1):
        c = ws.cell(row=2, column=ci, value=desc)
        c.font      = Font(name="Arial", italic=True, color="666666", size=9)
        c.fill      = PatternFill("solid", start_color=C_DESC_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = borde

    for ci, (_, _, ejemplo, _) in enumerate(COLS, 1):
        c = ws.cell(row=3, column=ci, value=ejemplo)
        c.font      = Font(name="Arial", color="999999", size=9, italic=True)
        c.fill      = PatternFill("solid", start_color=C_EJMP_BG)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = borde

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    nota = ws.cell(row=4, column=1,
                   value="← COMPLETAR ESTA FILA con los datos del mes. Eliminar filas 2 y 3.")
    nota.font = Font(name="Arial", color="AAAAAA", italic=True, size=8)

    wb.save(DESTINO)
    print(f"Template creado: {DESTINO}")
    print("Completar fila 4. Eliminar filas 2 y 3 antes de correr main.py.")


if __name__ == "__main__":
    main()
