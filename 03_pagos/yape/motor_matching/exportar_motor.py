# =========================IMPORTS===========================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========================COLORES DISEÑO=====================
# Siguiendo exactamente los HTMLs de diseño

# Grupos compartidos
C_ID_H    = "F4ECF7"; C_ID_C  = "FAF5FF"; C_ID_T  = "5B21B6"
C_BAN_H   = "E6F1FB"; C_BAN_C = "F0F8FF"; C_BAN_T = "185FA5"
C_UBI_H   = "E1F5EE"; C_UBI_C = "F0FFF8"; C_UBI_T = "065F46"
C_DED_H   = "FEF9E7"; C_DED_C = "FEFCE8"; C_DED_T = "7D6608"
C_CIC_H   = "FEF3E8"; C_CIC_C = "FEF3E8"; C_CIC_T = "7C3003"
C_MES_H   = "EFF6FF"; C_MES_C = "EFF6FF"; C_MES_T = "1D4ED8"
C_EST_H   = "FEF9E7"; C_EST_T = "7D6608"
C_LOT_H   = "E1F5EE"; C_LOT_C = "F0FFF8"; C_LOT_T = "065F46"
C_USR_H   = "E1F5EE"; C_USR_C = "F0FFF8"; C_USR_T = "065F46"
C_DEV_H   = "FEF3E8"; C_DEV_T = "7C3003"  # devoluciones tab
C_SEP     = "F3F4F6"

# Estados
EST_EXACTO  = "E1F5EE"; EST_EXACTO_T  = "085041"
EST_EXCESO  = "FAEEDA"; EST_EXCESO_T  = "854F0B"
EST_PARCIAL = "FCEBEB"; EST_PARCIAL_T = "A32D2D"
EST_PEND    = "FCEBEB"; EST_PEND_T    = "A32D2D"
EST_IDEN    = "FAEEDA"; EST_IDEN_T    = "854F0B"
EST_APLI    = "E1F5EE"; EST_APLI_T    = "085041"
EST_NOAP    = "FEFCE8"; EST_NOAP_T    = "7D6608"

# Fuentes
FUENTE_MSG = "E1F5EE"; FUENTE_MSG_T = "085041"
FUENTE_MAE = "E6F1FB"; FUENTE_MAE_T = "0C447C"
FUENTE_AMB = "FAEEDA"; FUENTE_AMB_T = "854F0B"
FUENTE_MUL = "F4ECF7"; FUENTE_MUL_T = "5B21B6"
FUENTE_MAN = "FEF3E8"; FUENTE_MAN_T = "7C3003"

MONTO_T = "065F46"

def _lado(c="CCCCCC"):
    return Side(style="thin", color=c)

def _cel(cell, bg, txt, mono=False, align="left", bold=False, size=10):
    b = _lado()
    cell.border    = Border(left=b, right=b, top=b, bottom=b)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.font      = Font(name="Courier New" if mono else "Arial",
                          bold=bold, size=size, color=txt)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)

def _hdr(cell, bg, txt, bold=True, size=9, align="center"):
    b = _lado("FFFFFF")
    cell.border    = Border(left=b, right=b, top=b, bottom=b)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.font      = Font(name="Arial", bold=bold, size=size, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)

def _sep(cell):
    cell.fill   = PatternFill("solid", start_color=C_SEP)
    cell.border = Border(left=_lado(), right=_lado("AAAAAA"),
                         top=_lado(),  bottom=_lado())

def _badge(cell, bg, txt, text, size=9):
    """Celda con estilo badge — texto centrado con color propio."""
    _cel(cell, bg, txt, mono=True, align="center", bold=True, size=size)
    cell.value = text

def _merge_hdr(ws, row, col_start, col_end, text, bg, txt, size=9):
    if col_start == col_end:
        c = ws.cell(row=row, column=col_start, value=text)
        _hdr(c, bg, txt, size=size)
    else:
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row,   end_column=col_end)
        c = ws.cell(row=row, column=col_start, value=text)
        _hdr(c, bg, txt, size=size)

def _sep_col(ws, row, ci, height=None):
    c = ws.cell(row=row, column=ci, value="")
    _sep(c)
    ws.column_dimensions[get_column_letter(ci)].width = 1

# ========================EXPORTAR pagos_yape_tepago=========
def exportar_pagos_tepago(todos: list, ciclo: int):
    """
    Exporta pagos_yape_tepago.xlsx siguiendo pagos_yape_tepago_diseno.html
    Narrativa: ¿Quién es? → ¿Qué hizo el banco? → ¿Dónde vive? → ¿Cuánto debe? → ¿Cuándo?
    """
    identificados = [r for r in todos if r.get("estado") == "identificado"]

    wb = Workbook()
    ws = wb.active
    ws.title = "pagos_yape_tepago"
    ws.freeze_panes = "A3"

    # Definicion de columnas con separadores
    # Grupos: quien(3) | sep | banco(7) | sep | donde(2) | sep | cuanto(3) | sep | ciclo(1)
    cols = [
        # quien
        ("USER_ID",        C_ID_C,  C_ID_T,  True,  "left"),
        ("NOMBRE",         C_ID_C,  C_ID_T,  False, "left"),
        ("FUENTE",         C_ID_C,  C_ID_T,  True,  "center"),
        "__SEP__",
        # banco
        ("TIPO",           C_BAN_C, C_BAN_T, True,  "left"),
        ("ORIGEN",         C_BAN_C, C_BAN_T, True,  "left"),
        ("DESTINO",        C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO_PAGO",     C_BAN_C, MONTO_T, True,  "right"),
        ("MONTO_ASIGNADO", C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE",        C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",          C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        # donde
        ("MZ",   C_UBI_C, C_UBI_T, True, "center"),
        ("LOTE", C_UBI_C, C_UBI_T, True, "center"),
        "__SEP__",
        # cuanto
        ("DEUDA_TOTAL",  C_DED_C, C_DED_T, True, "right"),
        ("DIFERENCIA",   C_DED_C, C_DED_T, True, "right"),
        ("ESTADO_PAGO",  C_DED_C, C_DED_T, True, "center"),
        "__SEP__",
        # ciclo
        ("CICLO", C_CIC_C, C_CIC_T, True, "center"),
    ]

    grupos = [
        ("¿QUIÉN ES?",        3, C_ID_H,  C_ID_T),
        None,
        ("¿QUÉ HIZO EL BANCO?", 7, C_BAN_H, C_BAN_T),
        None,
        ("¿DÓNDE VIVE?",      2, C_UBI_H, C_UBI_T),
        None,
        ("¿CUÁNTO DEBE?",     3, C_DED_H, C_DED_T),
        None,
        ("¿CUÁNDO?",          1, C_CIC_H, C_CIC_T),
    ]

    _escribir_cabecera_doble(ws, cols, grupos)

    anchos = {
        "USER_ID":10, "NOMBRE":28, "FUENTE":10,
        "TIPO":10, "ORIGEN":22, "DESTINO":22,
        "MONTO_PAGO":12, "MONTO_ASIGNADO":14,
        "MENSAJE":35, "FECHA":20,
        "MZ":6, "LOTE":7,
        "DEUDA_TOTAL":12, "DIFERENCIA":12, "ESTADO_PAGO":14,
        "CICLO":7,
    }

    fuente_colores = {
        "mensaje":   (FUENTE_MSG, FUENTE_MSG_T),
        "maestro":   (FUENTE_MAE, FUENTE_MAE_T),
        "ambiguo_auto": (FUENTE_AMB, FUENTE_AMB_T),
        "multiple_auto": (FUENTE_MUL, FUENTE_MUL_T),
        "multiple_corregido": (FUENTE_MUL, FUENTE_MUL_T),
        "correccion": (FUENTE_MAN, FUENTE_MAN_T),
        "blanco":    (EST_EXCESO, EST_EXCESO_T),
    }

    estado_colores = {
        "exacto":  (EST_EXACTO,  EST_EXACTO_T),
        "exceso":  (EST_EXCESO,  EST_EXCESO_T),
        "parcial": (EST_PARCIAL, EST_PARCIAL_T),
    }

    for ri, reg in enumerate(identificados, start=3):
        fuente = str(reg.get("fuente", "")).lower()
        f_bg, f_txt = fuente_colores.get(fuente, (C_ID_C, C_ID_T))
        ep = str(reg.get("estado_pago", "")).lower()
        ep_bg, ep_txt = estado_colores.get(ep, (C_DED_C, C_DED_T))

        valores = {
            "USER_ID":        reg.get("user_id", ""),
            "NOMBRE":         reg.get("nombre", ""),
            "FUENTE":         fuente,
            "TIPO":           "TE PAGÓ",
            "ORIGEN":         reg.get("origen", ""),
            "DESTINO":        "",
            "MONTO_PAGO":     reg.get("monto_pago", ""),
            "MONTO_ASIGNADO": reg.get("monto_asignado", ""),
            "MENSAJE":        reg.get("mensaje", ""),
            "FECHA":          reg.get("fecha", ""),
            "MZ":             reg.get("mz", ""),
            "LOTE":           reg.get("lote", ""),
            "DEUDA_TOTAL":    reg.get("deuda_total", ""),
            "DIFERENCIA":     reg.get("diferencia", ""),
            "ESTADO_PAGO":    ep,
            "CICLO":          ciclo,
        }

        ci = 1
        for col in cols:
            if col == "__SEP__":
                _sep_col(ws, ri, ci)
                ci += 1
                continue
            nombre, bg, txt, mono, align = col
            val = valores.get(nombre, "")

            # Colores especiales
            if nombre == "FUENTE":
                bg, txt = f_bg, f_txt
            elif nombre == "ESTADO_PAGO":
                bg, txt = ep_bg, ep_txt
            elif nombre in ("MONTO_PAGO", "MONTO_ASIGNADO", "DEUDA_TOTAL", "DIFERENCIA"):
                txt = MONTO_T

            c = ws.cell(row=ri, column=ci, value=val if val != "" else "")
            _cel(c, bg, txt, mono=mono, align=align)
            ci += 1

        ws.row_dimensions[ri].height = 18

    _aplicar_anchos(ws, cols, anchos)
    return wb

# ========================EXPORTAR pendientes================
def exportar_pendientes_diseño(sin_resolver: list) -> Workbook:
    """
    Exporta pendientes.xlsx siguiendo pendientes_xlsx.html
    Narrativa: banco (no editar) | tú completas
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sin_identificar"
    ws.freeze_panes = "A3"

    cols = [
        ("TIPO",    C_BAN_C, C_BAN_T, True,  "left"),
        ("ORIGEN",  C_BAN_C, C_BAN_T, True,  "left"),
        ("DESTINO", C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO",   C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE", C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",   C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MZ",      C_USR_C, C_USR_T, True,  "center"),
        ("LOTE",    C_USR_C, C_USR_T, True,  "center"),
        ("CONCEPTO",C_USR_C, C_USR_T, False, "left"),
        ("MOTIVO",  C_USR_C, C_USR_T, False, "left"),
    ]

    grupos = [
        ("← BANCO — NO EDITAR", 6, C_BAN_H, C_BAN_T),
        None,
        ("← TÚ COMPLETAS",      4, C_USR_H, C_USR_T),
    ]

    _escribir_cabecera_doble(ws, cols, grupos)

    anchos = {
        "TIPO":10, "ORIGEN":28, "DESTINO":22,
        "MONTO":12, "MENSAJE":35, "FECHA":20,
        "MZ":10, "LOTE":10, "CONCEPTO":25, "MOTIVO":30,
    }

    tipo_col = {"TE PAGÓ": (EST_EXACTO, EST_EXACTO_T),
                "PAGASTE": (EST_EXCESO, EST_EXCESO_T)}

    for ri, reg in enumerate(sin_resolver, start=3):
        tipo = str(reg.get("tipo", reg.get("fuente", "TE PAGÓ"))).upper()
        if "PAGASTE" in tipo:
            tipo_str = "PAGASTE"
        else:
            tipo_str = "TE PAGÓ"
        t_bg, t_txt = tipo_col.get(tipo_str, (C_BAN_C, C_BAN_T))

        valores = {
            "TIPO":    tipo_str,
            "ORIGEN":  reg.get("origen", ""),
            "DESTINO": reg.get("destino", ""),
            "MONTO":   reg.get("monto_pago", reg.get("monto", "")),
            "MENSAJE": reg.get("mensaje", ""),
            "FECHA":   reg.get("fecha", ""),
            "MZ":      "",
            "LOTE":    "",
            "CONCEPTO":"",
            "MOTIVO":  reg.get("motivo", ""),
        }

        ci = 1
        for col in cols:
            if col == "__SEP__":
                _sep_col(ws, ri, ci)
                ci += 1
                continue
            nombre, bg, txt, mono, align = col
            val = valores.get(nombre, "")
            if nombre == "TIPO":
                bg, txt = t_bg, t_txt
            c = ws.cell(row=ri, column=ci, value=val)
            _cel(c, bg, txt, mono=mono, align=align)
            ci += 1
        ws.row_dimensions[ri].height = 18

    _aplicar_anchos(ws, cols, anchos)
    return wb

# ========================EXPORTAR blancos_acumulados========
def exportar_blancos_acumulados(wb_existente, nuevo_blanco: dict, mes: str) -> Workbook:
    """
    Agrega un registro a blancos_acumulados.xlsx siguiendo blancos_acumulados_diseno.html
    Narrativa: ¿De qué mes? → ¿Quién es? → ¿Qué hizo el banco? → ¿Cómo se identificó? → ¿Ya se aplicó?
    """
    cols = [
        ("MES",     C_MES_C, C_MES_T, True,  "center"),
        "__SEP__",
        ("USER_ID", C_ID_C,  C_ID_T,  True,  "left"),
        ("NOMBRE",  C_ID_C,  C_ID_T,  False, "left"),
        "__SEP__",
        ("TIPO",    C_BAN_C, C_BAN_T, True,  "left"),
        ("ORIGEN",  C_BAN_C, C_BAN_T, True,  "left"),
        ("DESTINO", C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO",   C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE", C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",   C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MZ",      C_UBI_C, C_UBI_T, True,  "center"),
        ("LOTE",    C_UBI_C, C_UBI_T, True,  "center"),
        ("MOTIVO",  C_UBI_C, C_UBI_T, False, "left"),
        "__SEP__",
        ("ESTADO",  C_EST_H, C_EST_T, True,  "center"),
    ]

    grupos = [
        ("¿DE QUÉ MES?",         1, C_MES_H, C_MES_T),
        None,
        ("¿QUIÉN ES?",           2, C_ID_H,  C_ID_T),
        None,
        ("¿QUÉ HIZO EL BANCO?",  6, C_BAN_H, C_BAN_T),
        None,
        ("¿CÓMO SE IDENTIFICÓ?", 3, C_UBI_H, C_UBI_T),
        None,
        ("¿YA SE APLICÓ?",       1, C_EST_H, C_EST_T),
    ]

    anchos = {
        "MES":10, "USER_ID":10, "NOMBRE":28,
        "TIPO":10, "ORIGEN":22, "DESTINO":22,
        "MONTO":12, "MENSAJE":30, "FECHA":20,
        "MZ":6, "LOTE":7, "MOTIVO":30, "ESTADO":12,
    }

    # Si el workbook existe, leer filas existentes
    filas_existentes = []
    if wb_existente is not None:
        ws_old = wb_existente.active
        rows = list(ws_old.values)
        # Saltar cabeceras (fila 1 y 2)
        for fila in rows[2:]:
            if fila and any(fila):
                filas_existentes.append(fila)

    wb = Workbook()
    ws = wb.active
    ws.title = "blancos_acumulados"
    ws.freeze_panes = "A3"

    _escribir_cabecera_doble(ws, cols, grupos)

    # Reescribir filas existentes
    ri = 3
    for fila_vals in filas_existentes:
        ci = 1
        for col in cols:
            if col == "__SEP__":
                _sep_col(ws, ri, ci)
                ci += 1
                continue
            nombre, bg, txt, mono, align = col
            val = fila_vals[ci - 1 - cols[:ci-1].count("__SEP__")] if ci <= len(fila_vals) else ""
            # Color estado
            if nombre == "ESTADO":
                bg, txt = _color_estado_blanco(str(val))
            c = ws.cell(row=ri, column=ci, value=val)
            _cel(c, bg, txt, mono=mono, align=align)
            ci += 1
        ws.row_dimensions[ri].height = 18
        ri += 1

    # Agregar nuevo blanco
    valores = {
        "MES":     mes,
        "USER_ID": "",
        "NOMBRE":  "",
        "TIPO":    "TE PAGÓ",
        "ORIGEN":  nuevo_blanco.get("origen", ""),
        "DESTINO": nuevo_blanco.get("destino", ""),
        "MONTO":   nuevo_blanco.get("monto_pago", ""),
        "MENSAJE": nuevo_blanco.get("mensaje", ""),
        "FECHA":   nuevo_blanco.get("fecha", ""),
        "MZ":      "",
        "LOTE":    "",
        "MOTIVO":  "",
        "ESTADO":  "pendiente",
    }

    ci = 1
    for col in cols:
        if col == "__SEP__":
            _sep_col(ws, ri, ci)
            ci += 1
            continue
        nombre, bg, txt, mono, align = col
        val = valores.get(nombre, "")
        if nombre == "ESTADO":
            bg, txt = _color_estado_blanco(str(val))
        c = ws.cell(row=ri, column=ci, value=val)
        _cel(c, bg, txt, mono=mono, align=align)
        ci += 1
    ws.row_dimensions[ri].height = 18

    _aplicar_anchos(ws, cols, anchos)
    return wb

def _color_estado_blanco(estado: str):
    m = {"pendiente": (EST_PEND, EST_PEND_T),
         "identificado": (EST_IDEN, EST_IDEN_T),
         "aplicado": (EST_APLI, EST_APLI_T)}
    return m.get(estado.lower(), (C_EST_H, C_EST_T))

# ========================EXPORTAR devoluciones_acumulados===
def agregar_devoluciones_acumulados(wb_existente, nuevas: list, mes: str) -> Workbook:
    """
    Agrega registros PAGASTE a devoluciones_acumulados.xlsx
    siguiendo devoluciones_acumulados_diseno.html
    """
    cols = [
        ("MES",            C_MES_C, C_MES_T, True,  "center"),
        "__SEP__",
        ("USER_ID",        C_ID_C,  C_ID_T,  True,  "left"),
        ("NOMBRE",         C_ID_C,  C_ID_T,  False, "left"),
        "__SEP__",
        ("TIPO",           C_BAN_C, C_BAN_T, True,  "left"),
        ("ORIGEN",         C_BAN_C, C_BAN_T, True,  "left"),
        ("DESTINO",        C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO_PAGO",     C_BAN_C, MONTO_T, True,  "right"),
        ("MONTO_ASIGNADO", C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE",        C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",          C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MZ",      C_LOT_C, C_LOT_T, True,  "center"),
        ("LOTE",    C_LOT_C, C_LOT_T, True,  "center"),
        ("CONCEPTO",C_LOT_C, C_LOT_T, False, "left"),
        "__SEP__",
        ("ESTADO",  C_EST_H, C_EST_T, True,  "center"),
    ]

    grupos = [
        ("¿DE QUÉ MES?",         1, C_MES_H, C_MES_T),
        None,
        ("¿QUIÉN ES?",           2, C_ID_H,  C_ID_T),
        None,
        ("¿QUÉ HIZO EL BANCO?",  7, C_BAN_H, C_BAN_T),
        None,
        ("¿A QUÉ LOTE?",         3, C_LOT_H, C_LOT_T),
        None,
        ("¿YA SE APLICÓ?",       1, C_EST_H, C_EST_T),
    ]

    anchos = {
        "MES":10, "USER_ID":10, "NOMBRE":28,
        "TIPO":10, "ORIGEN":22, "DESTINO":22,
        "MONTO_PAGO":12, "MONTO_ASIGNADO":14,
        "MENSAJE":30, "FECHA":20,
        "MZ":6, "LOTE":7, "CONCEPTO":25, "ESTADO":12,
    }

    filas_existentes = []
    if wb_existente is not None:
        ws_old = wb_existente.active
        rows = list(ws_old.values)
        for fila in rows[2:]:
            if fila and any(fila):
                filas_existentes.append(list(fila))

    wb = Workbook()
    ws = wb.active
    ws.title = "devoluciones_acumulados"
    ws.freeze_panes = "A3"

    _escribir_cabecera_doble(ws, cols, grupos)

    ri = 3
    # Reescribir existentes
    col_names = [c[0] if c != "__SEP__" else "__SEP__" for c in cols]
    for fila_vals in filas_existentes:
        ci = 1
        idx = 0
        for col in cols:
            if col == "__SEP__":
                _sep_col(ws, ri, ci)
                ci += 1
                continue
            nombre, bg, txt, mono, align = col
            val = fila_vals[idx] if idx < len(fila_vals) else ""
            idx += 1
            if nombre == "ESTADO":
                bg, txt = _color_estado_dev(str(val))
            c = ws.cell(row=ri, column=ci, value=val)
            _cel(c, bg, txt, mono=mono, align=align)
            ci += 1
        ws.row_dimensions[ri].height = 18
        ri += 1

    # Agregar nuevas devoluciones
    for reg in nuevas:
        valores = {
            "MES":            mes,
            "USER_ID":        reg.get("user_id", ""),
            "NOMBRE":         reg.get("nombre", ""),
            "TIPO":           "PAGASTE",
            "ORIGEN":         reg.get("origen", ""),
            "DESTINO":        reg.get("destino", ""),
            "MONTO_PAGO":     reg.get("monto_pago", ""),
            "MONTO_ASIGNADO": reg.get("monto_asignado", ""),
            "MENSAJE":        reg.get("mensaje", ""),
            "FECHA":          reg.get("fecha", ""),
            "MZ":             reg.get("mz", ""),
            "LOTE":           reg.get("lote", ""),
            "CONCEPTO":       reg.get("concepto", ""),
            "ESTADO":         "pendiente",
        }
        ci = 1
        for col in cols:
            if col == "__SEP__":
                _sep_col(ws, ri, ci)
                ci += 1
                continue
            nombre, bg, txt, mono, align = col
            val = valores.get(nombre, "")
            if nombre == "ESTADO":
                bg, txt = _color_estado_dev(str(val))
            c = ws.cell(row=ri, column=ci, value=val)
            _cel(c, bg, txt, mono=mono, align=align)
            ci += 1
        ws.row_dimensions[ri].height = 18
        ri += 1

    _aplicar_anchos(ws, cols, anchos)
    return wb

def _color_estado_dev(estado: str):
    m = {"pendiente": (EST_PEND, EST_PEND_T),
         "aplicado":  (EST_APLI, EST_APLI_T),
         "no aplica": (EST_NOAP, EST_NOAP_T)}
    return m.get(estado.lower(), (C_EST_H, C_EST_T))

# ========================HELPERS============================
def _escribir_cabecera_doble(ws, cols, grupos):
    """Escribe fila 1 (grupos) y fila 2 (nombres de columnas)."""
    # Fila 1: grupos
    ci = 1
    for g in grupos:
        if g is None:
            # separador
            c = ws.cell(row=1, column=ci, value="")
            _sep(c)
            ws.column_dimensions[get_column_letter(ci)].width = 1
            ci += 1
        else:
            texto, span, bg, txt = g
            if span == 1:
                c = ws.cell(row=1, column=ci, value=texto)
                _hdr(c, bg, txt, size=9)
            else:
                ws.merge_cells(start_row=1, start_column=ci,
                               end_row=1,   end_column=ci + span - 1)
                c = ws.cell(row=1, column=ci, value=texto)
                _hdr(c, bg, txt, size=9)
            ci += span
    ws.row_dimensions[1].height = 20

    # Fila 2: nombres
    ci = 1
    for col in cols:
        if col == "__SEP__":
            c = ws.cell(row=2, column=ci, value="")
            _sep(c)
            ws.column_dimensions[get_column_letter(ci)].width = 1
        else:
            nombre, bg, txt, mono, align = col
            c = ws.cell(row=2, column=ci, value=nombre)
            _hdr(c, bg, txt, size=9)
        ci += 1
    ws.row_dimensions[2].height = 20

def _aplicar_anchos(ws, cols, anchos):
    ci = 1
    for col in cols:
        if col != "__SEP__":
            nombre = col[0]
            ws.column_dimensions[get_column_letter(ci)].width = anchos.get(nombre, 16)
        ci += 1

# ========================EXPORTAR TRAZABILIDAD==============
# Siguiendo trazabilidad.html — 3 hojas: Sin_identificar · Ambiguos · Pagos_multiples

C_RES_H = "E1F5EE"; C_RES_C = "F0FFF8"; C_RES_T = "085041"
C_CAN_H = "FAEEDA"; C_CAN_C = "FFFBF0"; C_CAN_T = "854F0B"
C_CUA_H = "FEF3E8"; C_CUA_C = "FEF3E8"; C_CUA_T = "7C3003"

def exportar_trazabilidad(ruta, corr_simples: dict, corr_ambiguos: dict,
                           corr_multiples: dict, ciclo: int, fecha_hoy: str):
    """
    Genera trazabilidad_YYYY_MM.xlsx con 3 hojas siguiendo trazabilidad.html
    """
    from pathlib import Path
    wb = Workbook()

    # ── Hoja 1: Sin_identificar ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Sin_identificar"
    ws1.freeze_panes = "A3"

    cols_si = [
        ("USER_ID", C_ID_C,  C_ID_T,  True,  "left"),
        ("NOMBRE",  C_ID_C,  C_ID_T,  False, "left"),
        "__SEP__",
        ("TIPO",    C_BAN_C, C_BAN_T, True,  "left"),
        ("ORIGEN",  C_BAN_C, C_BAN_T, True,  "left"),
        ("DESTINO", C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO",   C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE", C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",   C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MZ",      C_RES_C, C_RES_T, True,  "center"),
        ("LOTE",    C_RES_C, C_RES_T, True,  "center"),
        ("CONCEPTO",C_RES_C, C_RES_T, False, "left"),
        ("MOTIVO",  C_RES_C, C_RES_T, False, "left"),
        "__SEP__",
        ("CICLO",          C_CUA_C, C_CUA_T, True, "center"),
        ("FECHA_CORRECCION",C_CUA_C, C_CUA_T, True, "left"),
    ]

    grupos_si = [
        ("¿QUIÉN ES?",        2, C_ID_H,  C_ID_T),
        None,
        ("¿QUÉ HIZO EL BANCO?", 6, C_BAN_H, C_BAN_T),
        None,
        ("¿CÓMO SE RESOLVIÓ?", 4, C_RES_H, C_RES_T),
        None,
        ("¿CUÁNDO?",           2, C_CUA_H, C_CUA_T),
    ]

    _escribir_cabecera_doble(ws1, cols_si, grupos_si)

    anchos_si = {
        "USER_ID":10, "NOMBRE":28,
        "TIPO":10, "ORIGEN":28, "DESTINO":22,
        "MONTO":12, "MENSAJE":35, "FECHA":20,
        "MZ":6, "LOTE":7, "CONCEPTO":20, "MOTIVO":30,
        "CICLO":7, "FECHA_CORRECCION":16,
    }

    ri = 3
    for origen, v in corr_simples.items():
        valores = {
            "USER_ID":           v.get("user_id", ""),
            "NOMBRE":            v.get("nombre", ""),
            "TIPO":              v.get("tipo", "TE PAGÓ"),
            "ORIGEN":            origen,
            "DESTINO":           v.get("destino", ""),
            "MONTO":             v.get("monto", ""),
            "MENSAJE":           v.get("mensaje", ""),
            "FECHA":             v.get("fecha", ""),
            "MZ":                v.get("mz", ""),
            "LOTE":              v.get("lote", ""),
            "CONCEPTO":          v.get("concepto", ""),
            "MOTIVO":            v.get("motivo", ""),
            "CICLO":             ciclo,
            "FECHA_CORRECCION":  fecha_hoy,
        }
        ci = 1
        for col in cols_si:
            if col == "__SEP__":
                _sep_col(ws1, ri, ci)
                ci += 1
                continue
            nombre, bg, txt, mono, align = col
            c = ws1.cell(row=ri, column=ci, value=valores.get(nombre, ""))
            _cel(c, bg, txt, mono=mono, align=align)
            ci += 1
        ws1.row_dimensions[ri].height = 18
        ri += 1

    _aplicar_anchos(ws1, cols_si, anchos_si)

    # ── Hoja 2: Ambiguos ─────────────────────────────────────
    ws2 = wb.create_sheet("Ambiguos")
    ws2.freeze_panes = "A3"

    cols_amb = [
        ("USER_ID", C_ID_C,  C_ID_T,  True,  "left"),
        ("NOMBRE",  C_ID_C,  C_ID_T,  False, "left"),
        "__SEP__",
        ("ORIGEN",  C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO",   C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE", C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",   C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MZ_CAND",   C_CAN_C, C_CAN_T, True, "center"),
        ("LOTE_CAND", C_CAN_C, C_CAN_T, True, "center"),
        ("NOMBRE_CAND",C_CAN_C,C_CAN_T, False,"left"),
        ("DEUDA_CAND", C_CAN_C, C_CAN_T, True, "right"),
        ("DIFF_CAND",  C_CAN_C, C_CAN_T, True, "right"),
        "__SEP__",
        ("ELEGIDO",  C_RES_C, C_RES_T, True, "center"),
        ("MZ_FINAL", C_RES_C, C_RES_T, True, "center"),
        ("LOTE_FINAL",C_RES_C,C_RES_T, True, "center"),
        "__SEP__",
        ("CICLO",          C_CUA_C, C_CUA_T, True, "center"),
        ("FECHA_CORRECCION",C_CUA_C,C_CUA_T, True, "left"),
    ]

    grupos_amb = [
        ("¿QUIÉN ES?",          2, C_ID_H,  C_ID_T),
        None,
        ("¿QUÉ HIZO EL BANCO?", 4, C_BAN_H, C_BAN_T),
        None,
        ("CANDIDATOS EVALUADOS",5, C_CAN_H, C_CAN_T),
        None,
        ("RESULTADO FINAL",     3, C_RES_H, C_RES_T),
        None,
        ("¿CUÁNDO?",            2, C_CUA_H, C_CUA_T),
    ]

    _escribir_cabecera_doble(ws2, cols_amb, grupos_amb)

    anchos_amb = {
        "USER_ID":10, "NOMBRE":24,
        "ORIGEN":28, "MONTO":12, "MENSAJE":30, "FECHA":20,
        "MZ_CAND":8, "LOTE_CAND":8, "NOMBRE_CAND":20,
        "DEUDA_CAND":12, "DIFF_CAND":10,
        "ELEGIDO":8, "MZ_FINAL":8, "LOTE_FINAL":8,
        "CICLO":7, "FECHA_CORRECCION":16,
    }

    ri = 3
    for origen, v in corr_ambiguos.items():
        candidatos = v.get("candidatos", [{"mz": v.get("mz",""), "lote": v.get("lote",""),
                                            "nombre":"", "deuda":0, "diff":0}])
        mz_final   = v.get("mz", "")
        lote_final = v.get("lote", "")
        for cand in candidatos:
            elegido = "✅" if (cand.get("mz","") == mz_final and
                              cand.get("lote","") == lote_final) else ""
            valores = {
                "USER_ID":          v.get("user_id", ""),
                "NOMBRE":           v.get("nombre", ""),
                "ORIGEN":           origen,
                "MONTO":            v.get("monto", ""),
                "MENSAJE":          v.get("mensaje", ""),
                "FECHA":            v.get("fecha", ""),
                "MZ_CAND":          cand.get("mz", ""),
                "LOTE_CAND":        cand.get("lote", ""),
                "NOMBRE_CAND":      cand.get("nombre", ""),
                "DEUDA_CAND":       cand.get("deuda", ""),
                "DIFF_CAND":        cand.get("diff", ""),
                "ELEGIDO":          elegido,
                "MZ_FINAL":         mz_final,
                "LOTE_FINAL":       lote_final,
                "CICLO":            ciclo,
                "FECHA_CORRECCION": fecha_hoy,
            }
            ci = 1
            for col in cols_amb:
                if col == "__SEP__":
                    _sep_col(ws2, ri, ci)
                    ci += 1
                    continue
                nombre, bg, txt, mono, align = col
                c = ws2.cell(row=ri, column=ci, value=valores.get(nombre, ""))
                _cel(c, bg, txt, mono=mono, align=align)
                ci += 1
            ws2.row_dimensions[ri].height = 18
            ri += 1

    _aplicar_anchos(ws2, cols_amb, anchos_amb)

    # ── Hoja 3: Pagos_multiples ──────────────────────────────
    ws3 = wb.create_sheet("Pagos_multiples")
    ws3.freeze_panes = "A3"

    cols_mul = [
        ("USER_ID", C_ID_C,  C_ID_T,  True,  "left"),
        ("NOMBRE",  C_ID_C,  C_ID_T,  False, "left"),
        "__SEP__",
        ("ORIGEN",       C_BAN_C, C_BAN_T, True,  "left"),
        ("MONTO_TOTAL",  C_BAN_C, MONTO_T, True,  "right"),
        ("MENSAJE",      C_BAN_C, C_BAN_T, False, "left"),
        ("FECHA",        C_BAN_C, C_BAN_T, True,  "left"),
        "__SEP__",
        ("MZ",             C_RES_C, C_RES_T, True, "center"),
        ("LOTE",           C_RES_C, C_RES_T, True, "center"),
        ("DEUDA",          C_RES_C, C_RES_T, True, "right"),
        ("MONTO_ASIGNADO", C_RES_C, MONTO_T, True, "right"),
        ("DIFF",           C_RES_C, C_RES_T, True, "right"),
        "__SEP__",
        ("CICLO",           C_CUA_C, C_CUA_T, True, "center"),
        ("FECHA_CORRECCION",C_CUA_C, C_CUA_T, True, "left"),
    ]

    grupos_mul = [
        ("¿QUIÉN ES?",          2, C_ID_H,  C_ID_T),
        None,
        ("¿QUÉ HIZO EL BANCO?", 4, C_BAN_H, C_BAN_T),
        None,
        ("ASIGNACIÓN POR LOTE", 5, C_RES_H, C_RES_T),
        None,
        ("¿CUÁNDO?",            2, C_CUA_H, C_CUA_T),
    ]

    _escribir_cabecera_doble(ws3, cols_mul, grupos_mul)

    anchos_mul = {
        "USER_ID":10, "NOMBRE":24,
        "ORIGEN":28, "MONTO_TOTAL":12, "MENSAJE":30, "FECHA":20,
        "MZ":6, "LOTE":7, "DEUDA":12,
        "MONTO_ASIGNADO":14, "DIFF":10,
        "CICLO":7, "FECHA_CORRECCION":16,
    }

    ri = 3
    for origen, items in corr_multiples.items():
        monto_total = sum(i.get("monto", 0) for i in items)
        for item in items:
            valores = {
                "USER_ID":          item.get("user_id", ""),
                "NOMBRE":           item.get("nombre", ""),
                "ORIGEN":           origen,
                "MONTO_TOTAL":      monto_total,
                "MENSAJE":          item.get("mensaje", ""),
                "FECHA":            item.get("fecha", ""),
                "MZ":               item.get("mz", ""),
                "LOTE":             item.get("lote", ""),
                "DEUDA":            item.get("deuda", ""),
                "MONTO_ASIGNADO":   item.get("monto", ""),
                "DIFF":             item.get("diff", ""),
                "CICLO":            ciclo,
                "FECHA_CORRECCION": fecha_hoy,
            }
            ci = 1
            for col in cols_mul:
                if col == "__SEP__":
                    _sep_col(ws3, ri, ci)
                    ci += 1
                    continue
                nombre, bg, txt, mono, align = col
                c = ws3.cell(row=ri, column=ci, value=valores.get(nombre, ""))
                _cel(c, bg, txt, mono=mono, align=align)
                ci += 1
            ws3.row_dimensions[ri].height = 18
            ri += 1

    _aplicar_anchos(ws3, cols_mul, anchos_mul)

    wb.save(ruta)
    print(f"  ✔ trazabilidad guardada: {Path(ruta).name}")
