# =========================IMPORTS===========================
import re
import shutil
from datetime import datetime

import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========================CONFIGURACION======================
BASE_DIR         = Path(".")
INPUT_DIR        = BASE_DIR / "Inputs"
MAESTRO_DIR      = INPUT_DIR / "Maestro"
REPORTE_DIR      = INPUT_DIR / "Reporte_mes"
PLANILLA_DIR     = INPUT_DIR / "Planilla_mes"
PLANILLA_ANT_DIR = INPUT_DIR / "Planilla_anterior"
CORRECCIONES_DIR = BASE_DIR / "Correcciones"
OUTPUT_DIR       = BASE_DIR / "Outputs"

MAESTRO_FILE    = "maestro_yape.xlsx"
OUTPUT_FILE     = "pagos_yape.xlsx"
PENDIENTES_FILE = "pendientes.xlsx"
TIPO_PAGO       = "TE PAGÓ"

ALIAS_BANCO = {
    "tipo":    ["tipo de transacción", "tipo de transaccion"],
    "origen":  ["origen"],
    "monto":   ["monto"],
    "mensaje": ["mensaje"],
    "fecha":   ["fecha de operación", "fecha de operacion"],
}

ALIAS_PLANILLA = {
    "mz":           ["mz"],
    "lote":         ["lt", "lote"],
    "nombre":       ["nombres", "nombre"],
    "total":        ["total"],
    "mes_anterior": ["mes anterior"],
}

# Patrones para extraer mz y lote del mensaje
PATRONES_MENSAJE = [
    re.compile(r'mz\.?\s*["\']([ A-Z0-9]+)["\']?\s*(?:lt\.?|lote\.?|lte\.?)\s*["\']?(\d+[A-Z]?)["\']?', re.IGNORECASE),
    re.compile(r'mz\.?\s*([A-Z0-9\-]+)[_]+(?:lote|lt|lte)\.?\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'maz\.?\s*([A-Z0-9]+)\.?\s*(?:lot\.?|lote\.?|lt\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz[na]+\.?\s*([A-Z0-9\-]+)\.?\s*(?:lt\.?|lot\.?|lote\.?|lte\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'man[sz]?[aio]?[aon]?[aáñ]?\s+([A-Z0-9\-]+)\s+(?:lt\.?|lot\.?|lote\.?|lte\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.([A-Z0-9\-]+)\.?\s*(?:lt\.?|lte\.?|lote\.?)\s*\.?\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz([A-Z][A-Z0-9]*)\.?\s*(?:lt\.?|lote\.?|lte\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?\s*\.?\s*([A-Z][A-Z0-9]*)\.?\s*(?:tl\.?|lt\.?|lte\.?)\s*\.?\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?\s*([A-Z0-9\-]+),\s*(?:lt\.?|lte\.?|lote\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?-?([A-Z][A-Z0-9]*)\s*[-/]\s*(?:lt\.?|lte\.?|l\.?)-?\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?\s*:?\s*([A-Z0-9\-]+)\s*:?\s*(?:lt\.?|lote\.?|lte\.?)\s*:?\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?\s*([A-Z][A-Z0-9]*)\.?\s+(?:lt\.?|lte\.?)\s*\.?\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?\s*([A-Z][A-Z0-9]*)\s+l\.?\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?\s*([A-Z0-9\-]+)\s*(?:late\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?\s*([A-Z0-9]+)\s+la\s+(\d+)', re.IGNORECASE),
    re.compile(r'lote\s+([A-Z])\s+(\d+)', re.IGNORECASE),
    re.compile(r'\bm\.([A-Z0-9]+)\s*(?:lote\.?|lt\.?|lte\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'\bm\s+([A-Z][A-Z0-9]*)\.?\s+(?:lt\.?|lte\.?|l\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'^mz?([A-Z][A-Z0-9]*)\s+(?:lt\.?|lte\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'^([A-Z][A-Z0-9]+)\s+(?:lt\.?|lote\.?|lte\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'^([A-Z][A-Z0-9]*)-(\d+[A-Z]?)(?:\s|$|,)', re.IGNORECASE),
    re.compile(r'^([A-Z])0*(\d+)(?:\s|$)', re.IGNORECASE),
    re.compile(r'^([A-Z][A-Z0-9]*)\s+(\d+)', re.IGNORECASE),
]

COLOR_CABECERA     = "4A235A"
COLOR_IDENTIFICADO = "D5F5E3"
COLOR_PENDIENTE    = "FADBD8"
COLOR_ORIGEN       = "D6EAF8"
COLOR_UBICACION    = "FDEBD0"
COLOR_CONFIANZA    = "D7BDE2"
COLOR_DEUDA        = "FDEDEC"

# ========================UTILIDADES=========================
def reset_output_folder(path: Path):
    if path.exists():
        shutil.rmtree(path)
    path.mkdir()

def estilo_celda(cell, bg_color=None, align="left"):
    lado = Side(style="thin", color="CCCCCC")
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = Border(left=lado, right=lado, top=lado, bottom=lado)
    if bg_color:
        cell.fill = PatternFill("solid", start_color=bg_color)

def cabecera_celda(cell, texto):
    cell.value     = texto
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", start_color=COLOR_CABECERA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    lado = Side(style="thin", color="FFFFFF")
    cell.border    = Border(left=lado, right=lado, top=lado, bottom=lado)

def limpiar_lote(val) -> str:
    s = str(val).strip()
    if not s or s in ("None", "NAN", ""):
        return ""
    try:
        return str(int(float(s)))
    except:
        return s

def limpiar_monto(val) -> float:
    try:
        return float(str(val).replace(",", ".").strip())
    except:
        return 0.0

# ====================ANCLA DE CORTE=========================
def obtener_ancla() -> datetime | None:
    archivos = list(PLANILLA_ANT_DIR.glob("*.xlsx"))
    if not archivos:
        print(f"  ⚠ No hay archivo en {PLANILLA_ANT_DIR} — sin ancla de corte")
        return None

    archivo = archivos[0]
    print(f"  Planilla anterior: {archivo.name}")

    wb    = load_workbook(archivo, read_only=True, data_only=True)
    hojas = [h for h in wb.sheetnames if "reporte" in h.lower()]
    if not hojas:
        wb.close()
        print("  ⚠ No se encontró hoja 'Reporte' en planilla anterior — sin ancla")
        return None

    ws    = wb[hojas[0]]
    datos = list(ws.values)
    wb.close()

    if len(datos) < 2:
        return None

    headers   = [str(h).strip().lower() if h else "" for h in datos[0]]
    col_fecha = next((i for i, h in enumerate(headers) if "fecha" in h), None)
    if col_fecha is None:
        print("  ⚠ No se encontró columna fecha en Reporte anterior — sin ancla")
        return None

    fechas = []
    for fila in datos[1:]:
        if not fila or col_fecha >= len(fila):
            continue
        val = fila[col_fecha]
        if isinstance(val, datetime):
            fechas.append(val)
            continue
        s = str(val).strip()
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
                    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                fechas.append(datetime.strptime(s, fmt))
                break
            except:
                pass

    if not fechas:
        print("  ⚠ Sin fechas válidas en Reporte anterior — sin ancla")
        return None

    ancla = max(fechas)
    print(f"  Ancla de corte: {ancla.strftime('%d/%m/%Y %H:%M:%S')}")
    return ancla

# ====================EXTRACCION MENSAJE=====================
def extraer_mz_lote_mensaje(mensaje: str) -> tuple:
    """
    Intenta extraer mz y lote del mensaje. Siempre tiene prioridad sobre el origen.
    Retorna (mz, lote) si encuentra, (None, None) si no.
    """
    if not mensaje or str(mensaje).strip() in ("", "nan", "None"):
        return None, None

    msg = str(mensaje).strip()

    for patron in PATRONES_MENSAJE:
        m = patron.search(msg)
        if m:
            grupos = [g for g in m.groups() if g is not None]
            if len(grupos) >= 2:
                mz  = grupos[0].strip().upper()
                try:
                    lote = str(int(grupos[1].strip()))
                except:
                    lote = grupos[1].strip()
                if mz and lote:
                    return mz, lote

    return None, None

# ====================NORMALIZACION MZ=======================
def normalizar_mz(mz: str, mzs_validas: set) -> str:
    """
    Limpia la MZ extraída del mensaje:
    1. Quita cualquier carácter que no sea letra o número (guiones, puntos, etc.)
    2. Si el resultado no existe en planilla pero reemplazando 0→O sí existe → corrige
    3. Si sigue sin existir → retorna la MZ limpia igual (será pendiente)
    """
    if not mz:
        return mz

    # Paso 1: quitar todo lo que no sea letra o número
    limpia = re.sub(r'[^A-Z0-9]', '', mz.upper())
    if not limpia:
        return mz

    # Paso 2: si ya existe → devolver
    if limpia in mzs_validas:
        return limpia

    # Paso 3: intentar 0 → O
    con_o = limpia.replace('0', 'O')
    if con_o in mzs_validas:
        return con_o

    return limpia

# ====================NORMALIZACION==========================
def normalizar_columnas(df: pd.DataFrame, alias: dict) -> dict:
    cols_lower = {str(c).lower().strip(): c for c in df.columns
                  if c and str(c).strip() not in ("None", "nan")}
    mapa = {}
    for campo, variantes in alias.items():
        for v in variantes:
            if v in cols_lower:
                mapa[campo] = cols_lower[v]
                break
    return mapa

# ====================CARGA DE DATOS=========================
def cargar_maestro() -> dict:
    path = MAESTRO_DIR / MAESTRO_FILE
    if not path.exists():
        raise FileNotFoundError(f"No se encontró {path}")

    wb    = load_workbook(path, read_only=True, data_only=True)
    ws    = wb.active
    datos = list(ws.values)
    wb.close()

    headers = [str(h).strip().upper() if h else "" for h in datos[0]]

    try:
        col_uid   = headers.index("USER_ID")
        col_mz    = headers.index("MZ")
        col_lote  = headers.index("LOTE")
        col_nivel = headers.index("NIVEL_CONFIANZA")
    except ValueError as e:
        raise ValueError(f"Columna no encontrada en maestro: {e}")

    cols_origen = [i for i, h in enumerate(headers) if h.startswith("ORIGEN_")]

    indice = {}
    for fila in datos[1:]:
        if not fila:
            continue
        uid   = str(fila[col_uid]).strip()        if fila[col_uid]   else ""
        mz    = str(fila[col_mz]).strip().upper() if fila[col_mz]    else ""
        lote  = limpiar_lote(fila[col_lote])
        nivel = str(fila[col_nivel]).strip()      if fila[col_nivel] else ""

        if not mz or not lote:
            continue

        for ci in cols_origen:
            if ci < len(fila) and fila[ci]:
                origen_key = str(fila[ci]).strip().upper()
                if origen_key:
                    indice[origen_key] = {
                        "user_id":         uid,
                        "mz":              mz,
                        "lote":            lote,
                        "nivel_confianza": nivel,
                    }

    print(f"  ✔ {len(indice)} orígenes indexados desde el maestro")
    return indice

def cargar_planilla() -> tuple[dict, set]:
    archivos = list(PLANILLA_DIR.glob("*.xlsx"))
    if not archivos:
        print(f"  ⚠ No hay planilla en {PLANILLA_DIR} — se omite cruce de deuda")
        return {}, set()

    archivo = archivos[0]
    print(f"  Leyendo planilla: {archivo.name}")

    wb    = load_workbook(archivo, read_only=True, data_only=True)
    ws    = wb.active
    datos = []
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row):
            datos.append(row)
    wb.close()

    if len(datos) < 2:
        return {}, set()

    n_cols = len(datos[0])
    df     = pd.DataFrame([f[:n_cols] for f in datos[1:]], columns=datos[0])
    mapa   = normalizar_columnas(df, ALIAS_PLANILLA)

    faltantes = [c for c in ["mz", "lote", "total"] if c not in mapa]
    if faltantes:
        print(f"  ⚠ Columnas faltantes en planilla: {faltantes}")
        return {}, set()

    planilla    = {}
    mzs_validas = set()

    for _, fila in df.iterrows():
        mz   = str(fila[mapa["mz"]]).strip().upper() if mapa.get("mz")   else ""
        lote = limpiar_lote(fila[mapa["lote"]])       if mapa.get("lote") else ""
        if not mz or not lote or mz in ("NAN", "NONE", ""):
            continue
        mzs_validas.add(mz)
        nombre       = str(fila[mapa["nombre"]]).strip() if mapa.get("nombre") else ""
        total        = limpiar_monto(fila[mapa["total"]])
        mes_anterior = limpiar_monto(fila[mapa["mes_anterior"]]) if mapa.get("mes_anterior") else 0.0
        planilla[(mz, lote)] = {
            "nombre":       nombre,
            "deuda_total":  total,
            "mes_anterior": mes_anterior,
        }

    print(f"  ✔ {len(planilla)} usuarios | {len(mzs_validas)} manzanas válidas")
    return planilla, mzs_validas

def cargar_reportes(ancla=None) -> tuple:
    archivos = sorted(REPORTE_DIR.glob("*.xlsx"))
    if not archivos:
        raise FileNotFoundError(f"No hay .xlsx en {REPORTE_DIR}")

    print(f"  Archivos encontrados: {len(archivos)}")
    todos      = []
    mapa_final = {}

    for archivo in archivos:
        print(f"  Leyendo: {archivo.name}")

        wb    = load_workbook(archivo, read_only=False, data_only=True)
        ws    = wb.active
        datos = []
        for row in ws.iter_rows(values_only=True):
            if any(v is not None for v in row):
                datos.append(row)
        wb.close()

        if len(datos) < 2:
            continue

        header_idx = 0
        for i, fila in enumerate(datos):
            if fila[0] and str(fila[0]).strip().lower() in (
                "tipo de transacción", "tipo de transaccion"
            ):
                header_idx = i
                break

        headers = datos[header_idx]
        filas   = datos[header_idx + 1:]
        n_cols  = len(headers)
        df      = pd.DataFrame([f[:n_cols] for f in filas], columns=headers)

        mapa = normalizar_columnas(df, ALIAS_BANCO)
        faltantes = [c for c in ["tipo", "origen", "monto"] if c not in mapa]
        if faltantes:
            print(f"  ⚠ Saltado — columnas faltantes: {faltantes}")
            continue

        df = df[df[mapa["tipo"]].astype(str).str.strip().str.upper() == TIPO_PAGO.upper()].copy()
        df = df.reset_index(drop=True)
        print(f"  ✔ {len(df)} filas TE PAGÓ")
        todos.append(df)
        mapa_final = mapa

    if not todos:
        raise ValueError("Ningún archivo válido en Reporte_mes")

    df_total = pd.concat(todos, ignore_index=True)

    cols_dedup = [mapa_final.get(c) for c in ["origen", "monto", "fecha"] if mapa_final.get(c)]
    antes = len(df_total)
    df_total = df_total.drop_duplicates(subset=cols_dedup).reset_index(drop=True)
    if antes > len(df_total):
        print(f"  ⚠ {antes - len(df_total)} duplicados eliminados")

    # Filtrar por ancla de corte
    if ancla is not None and mapa_final.get("fecha"):
        col_f = mapa_final["fecha"]

        def _parsear(val):
            if isinstance(val, datetime):
                return val
            s = str(val).strip()
            for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
                        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    return datetime.strptime(s, fmt)
                except:
                    pass
            return None

        df_total["_DT"] = df_total[col_f].apply(_parsear)
        antes_ancla = len(df_total)
        df_total = df_total[df_total["_DT"].notna() & (df_total["_DT"] > ancla)]
        df_total = df_total.drop(columns=["_DT"]).reset_index(drop=True)
        excluidos = antes_ancla - len(df_total)
        if excluidos:
            print(f"  ⚠ {excluidos} filas anteriores al ancla excluidas")

    print(f"  Total final: {len(df_total)} filas")
    return df_total, mapa_final

def leer_correcciones() -> dict:
    """Retorna dict: origen_upper -> {"mz": str, "lote": str}"""
    ruta = CORRECCIONES_DIR / PENDIENTES_FILE
    if not ruta.exists():
        return {}

    try:
        # Fila 1 = instrucción, fila 2 = cabecera real
        df = pd.read_excel(ruta, header=1, dtype=str)
        df.columns = [c.strip().upper() for c in df.columns]
    except Exception as e:
        print(f"  ⚠ No se pudo leer correcciones: {e}")
        return {}

    if "MZ" not in df.columns or "ORIGEN" not in df.columns:
        print("  ⚠ pendientes.xlsx no tiene columnas MZ/ORIGEN")
        return {}

    correcciones = {}
    for _, fila in df.iterrows():
        mz = str(fila.get("MZ", "")).strip().upper()
        if not mz or mz in ("NAN", "NONE", ""):
            continue
        origen = str(fila.get("ORIGEN", "")).strip().upper()
        if not origen or origen in ("NAN", "NONE", ""):
            continue
        lote = limpiar_lote(fila.get("LOTE", "")) if mz != "BLANCO" else ""
        correcciones[origen] = {"mz": mz, "lote": lote}

    if correcciones:
        n_blanco = sum(1 for v in correcciones.values() if v["mz"] == "BLANCO")
        n_real   = len(correcciones) - n_blanco
        print(f"  Correcciones: {n_real} identificados + {n_blanco} blanco")
    else:
        print("  Sin correcciones previas")
    return correcciones

# ====================VALIDACION=============================
def validar_reporte(df: pd.DataFrame, mapa: dict) -> pd.DataFrame:
    antes = len(df)
    col   = mapa["origen"]
    df    = df[df[col].astype(str).str.strip() != ""].copy()
    df    = df[df[col].astype(str).str.strip().str.upper() != "NAN"].copy()
    df    = df.reset_index(drop=True)
    if antes - len(df):
        print(f"  ⚠ {antes - len(df)} filas descartadas")
    print(f"  Filas válidas: {len(df)}")
    return df

# ====================PROCESAMIENTO==========================
def ejecutar_matching(df: pd.DataFrame, mapa: dict,
                      indice: dict, planilla: dict,
                      correcciones: dict = None,
                      mzs_validas: set = None) -> list:
    if correcciones is None:
        correcciones = {}
    if mzs_validas is None:
        mzs_validas = set(k[0] for k in planilla.keys()) if planilla else set()
    todos = []

    for _, fila in df.iterrows():
        origen     = str(fila[mapa["origen"]]).strip()
        origen_key = origen.upper()
        monto_pago = limpiar_monto(fila[mapa["monto"]]) if mapa.get("monto") else 0.0
        mensaje    = str(fila[mapa["mensaje"]]) if mapa.get("mensaje") and mapa["mensaje"] in fila.index else ""
        fecha      = str(fila[mapa["fecha"]])   if mapa.get("fecha")   and mapa["fecha"]   in fila.index else ""

        mensaje = "" if mensaje == "None" else mensaje
        fecha   = "" if fecha   == "None" else fecha

        base = {
            "origen":     origen,
            "monto_pago": monto_pago,
            "mensaje":    mensaje,
            "fecha":      fecha,
        }

        # ── Capa 1: mensaje siempre primero ──────────────────
        mz_msg, lote_msg = extraer_mz_lote_mensaje(mensaje)

        # Normalizar MZ extraída del mensaje
        if mz_msg:
            mz_msg = normalizar_mz(mz_msg, mzs_validas)

        # ── Capa 2: origen en maestro como respaldo ──────────
        match_origen = indice.get(origen_key)

        # ── Capa 3: correcciones manuales ─────────────────────
        corr = correcciones.get(origen_key)

        if mz_msg and lote_msg:
            mz     = mz_msg
            lote   = lote_msg
            uid    = match_origen["user_id"] if match_origen and match_origen["mz"] == mz and match_origen["lote"] == lote else ""
            nivel  = "por mensaje"
            fuente = "mensaje"
        elif match_origen:
            mz     = match_origen["mz"]
            lote   = match_origen["lote"]
            uid    = match_origen["user_id"]
            nivel  = match_origen["nivel_confianza"]
            fuente = "maestro"
        elif corr and corr["mz"] == "BLANCO":
            todos.append({
                **base,
                "user_id": "", "nombre": "", "mz": "BLANCO", "lote": "",
                "nivel_confianza": "", "deuda_total": "", "mes_anterior": "",
                "diferencia": "", "estado_pago": "", "fuente": "blanco",
                "estado": "pendiente",
            })
            continue
        elif corr and corr["mz"] and corr["lote"]:
            mz     = corr["mz"]
            lote   = corr["lote"]
            uid    = ""
            nivel  = "correccion"
            fuente = "correccion"
        else:
            todos.append({
                **base,
                "user_id": "", "nombre": "", "mz": "", "lote": "",
                "nivel_confianza": "", "deuda_total": "", "mes_anterior": "",
                "diferencia": "", "estado_pago": "", "fuente": "pendiente",
                "estado": "pendiente",
            })
            continue

        # ── Validar que MZ/LOTE existan en planilla ───────────
        if planilla and (mz, lote) not in planilla:
            todos.append({
                **base,
                "user_id": uid, "nombre": "", "mz": mz, "lote": lote,
                "nivel_confianza": nivel, "deuda_total": "", "mes_anterior": "",
                "diferencia": "", "estado_pago": "no en planilla", "fuente": fuente,
                "estado": "pendiente",
            })
            continue

        # Cruce con planilla
        datos_p      = planilla.get((mz, lote), {})
        deuda_total  = datos_p.get("deuda_total",  0.0)
        mes_anterior = datos_p.get("mes_anterior", 0.0)
        nombre       = datos_p.get("nombre", "")

        if deuda_total > 0:
            diferencia = round(monto_pago - deuda_total, 2)
            if diferencia == 0:   estado_pago = "exacto"
            elif diferencia > 0:  estado_pago = "exceso"
            else:                 estado_pago = "parcial"
        else:
            diferencia  = 0.0
            estado_pago = "sin deuda en planilla"

        todos.append({
            **base,
            "user_id":         uid,
            "nombre":          nombre,
            "mz":              mz,
            "lote":            lote,
            "nivel_confianza": nivel,
            "deuda_total":     deuda_total,
            "mes_anterior":    mes_anterior,
            "diferencia":      diferencia,
            "estado_pago":     estado_pago,
            "fuente":          fuente,
            "estado":          "identificado",
        })

    identificados = [r for r in todos if r["estado"] == "identificado"]
    pendientes    = [r for r in todos if r["estado"] == "pendiente"]
    print(f"  Identificados : {len(identificados)}")
    print(f"  Pendientes    : {len(pendientes)}")
    return todos

# ====================EXPORTAR EXCEL========================
def exportar_excel(todos: list):
    # Identificados primero, luego pendientes al final
    ordenados = [r for r in todos if r["estado"] == "identificado"] + \
                [r for r in todos if r["estado"] == "pendiente"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos"
    ws.freeze_panes = "A2"

    cols = [
        "user_id", "nombre", "nivel_confianza", "fuente",
        "origen", "monto_pago", "deuda_total", "mes_anterior",
        "diferencia", "estado_pago", "mz", "lote", "mensaje", "fecha", "estado",
    ]

    for ci, cab in enumerate(cols, start=1):
        cabecera_celda(ws.cell(row=1, column=ci), cab.upper())
    ws.row_dimensions[1].height = 22

    color_col = {
        "user_id":         COLOR_ORIGEN,
        "nombre":          COLOR_ORIGEN,
        "nivel_confianza": COLOR_CONFIANZA,
        "fuente":          COLOR_CONFIANZA,
        "mz":              COLOR_UBICACION,
        "lote":            COLOR_UBICACION,
        "deuda_total":     COLOR_DEUDA,
        "mes_anterior":    COLOR_DEUDA,
        "diferencia":      COLOR_DEUDA,
        "estado_pago":     COLOR_CONFIANZA,
        "estado":          COLOR_CONFIANZA,
    }

    for ri, reg in enumerate(ordenados, start=2):
        es_pendiente = reg.get("estado") == "pendiente"
        for ci, col in enumerate(cols, start=1):
            c = ws.cell(row=ri, column=ci, value=reg.get(col, ""))
            if es_pendiente:
                estilo_celda(c, bg_color=COLOR_PENDIENTE)
            else:
                estilo_celda(c, bg_color=color_col.get(col, COLOR_IDENTIFICADO))
        ws.row_dimensions[ri].height = 17

    anchos = {
        "user_id":10, "nombre":30, "nivel_confianza":16, "fuente":12,
        "origen":32, "monto_pago":12, "deuda_total":12, "mes_anterior":14,
        "diferencia":12, "estado_pago":18, "mz":8, "lote":8,
        "mensaje":40, "fecha":22, "estado":14,
    }
    for ci, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = anchos.get(col, 14)

    wb.save(OUTPUT_DIR / OUTPUT_FILE)

    n_id  = len([r for r in ordenados if r["estado"] == "identificado"])
    n_pen = len([r for r in ordenados if r["estado"] == "pendiente"])
    print(f"\n  ✔ {OUTPUT_FILE}")
    print(f"     Identificados : {n_id}")
    print(f"     Pendientes    : {n_pen}  ← filtrar por estado=pendiente")
    print(f"     Total         : {len(ordenados)}")

def exportar_pendientes(todos: list):
    # Excluye BLANCO (ya decididos); solo los que aún faltan identificar
    sin_resolver = [
        r for r in todos
        if r["estado"] == "pendiente" and str(r.get("mz", "")).upper() != "BLANCO"
    ]

    CORRECCIONES_DIR.mkdir(exist_ok=True)
    ruta = CORRECCIONES_DIR / PENDIENTES_FILE

    if not sin_resolver:
        print("  ✔ Sin pendientes sin resolver")
        if ruta.exists():
            ruta.unlink()
        return

    cols = ["ORIGEN", "MONTO_PAGO", "MENSAJE", "FECHA", "MZ", "LOTE"]
    key  = {c: c.lower() for c in cols}

    wb = Workbook()
    ws = wb.active
    ws.title        = "Pendientes"
    ws.freeze_panes = "A2"

    instruccion = "👉 Completa MZ y LOTE. Escribe BLANCO en MZ si no se identifica al dueño. Guarda y vuelve a correr."
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    ws["A1"]           = instruccion
    ws["A1"].font      = Font(name="Arial", bold=True, size=10, color="92400E")
    ws["A1"].fill      = PatternFill("solid", start_color="FEF3C7")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for ci, col in enumerate(cols, 1):
        cabecera_celda(ws.cell(2, ci), col)
    ws.row_dimensions[2].height = 20

    anchos = {"ORIGEN": 35, "MONTO_PAGO": 12, "MENSAJE": 40,
              "FECHA": 22, "MZ": 10, "LOTE": 10}

    for ri, reg in enumerate(sin_resolver, 3):
        for ci, col in enumerate(cols, 1):
            c = ws.cell(ri, ci, value=reg.get(key[col], ""))
            bg = "FFF9C4" if col in ("MZ", "LOTE") else COLOR_PENDIENTE
            estilo_celda(c, bg_color=bg)
        ws.row_dimensions[ri].height = 17

    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = anchos.get(col, 14)

    wb.save(ruta)
    print(f"  ⚠ {len(sin_resolver)} pendientes → {PENDIENTES_FILE}")

# ======================MAIN=================================
def main():
    print("=" * 55)
    print("  MOTOR MATCHING — identificación de pagos Yape")
    print("=" * 55)

    # Ciclo 1 si aún no hay output; ciclos siguientes sobreescriben
    ciclo = 1 if not (OUTPUT_DIR / OUTPUT_FILE).exists() else 2
    print(f"\n  CICLO {ciclo}")

    print("\n[1] Preparando carpeta de salida...")
    if ciclo == 1:
        reset_output_folder(OUTPUT_DIR)
    else:
        OUTPUT_DIR.mkdir(exist_ok=True)

    print("\n[2] Leyendo ancla de corte (Planilla_anterior)...")
    ancla = obtener_ancla()

    print("\n[3] Cargando maestro...")
    indice = cargar_maestro()

    print("\n[4] Cargando planilla del mes...")
    planilla, mzs_validas = cargar_planilla()

    print("\n[5] Cargando reportes del banco...")
    df, mapa = cargar_reportes(ancla)

    print("\n[6] Validando reporte...")
    df = validar_reporte(df, mapa)

    print("\n[7] Leyendo correcciones manuales...")
    correcciones = leer_correcciones()

    print("\n[8] Ejecutando matching...")
    todos = ejecutar_matching(df, mapa, indice, planilla, correcciones, mzs_validas)

    print("\n[9] Exportando resultados...")
    exportar_excel(todos)

    print("\n[10] Exportando pendientes a Correcciones/...")
    exportar_pendientes(todos)

    total = len(todos)
    n_id  = len([r for r in todos if r["estado"] == "identificado"])
    pct   = round(n_id / total * 100, 1) if total else 0

    print("\n" + "=" * 55)
    print(f"  Proceso completado — {pct}% identificado automáticamente")
    print("=" * 55)

if __name__ == "__main__":
    main()
