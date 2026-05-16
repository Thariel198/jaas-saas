# =========================IMPORTS===========================
import re
import shutil
from datetime import datetime

import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Funciones de exportación con diseño HTML
import sys
sys.path.insert(0, str(Path(__file__).parent))
from exportar_motor import (
    exportar_trazabilidad,
    exportar_pagos_tepago, exportar_pendientes_diseño,
    exportar_blancos_acumulados, agregar_devoluciones_acumulados,
    _cel, _hdr, _sep, _sep_col, _lado, _escribir_cabecera_doble, _aplicar_anchos,
    C_ID_H, C_ID_C, C_ID_T, C_BAN_H, C_BAN_C, C_BAN_T,
    C_UBI_H, C_UBI_C, C_UBI_T, C_DED_H, C_DED_C, C_DED_T,
    C_CIC_H, C_CIC_C, C_CIC_T, C_MES_H, C_MES_C, C_MES_T,
    C_LOT_H, C_LOT_C, C_LOT_T, C_SEP, MONTO_T,
    EST_EXACTO, EST_EXACTO_T, EST_EXCESO, EST_EXCESO_T,
    EST_PARCIAL, EST_PARCIAL_T, EST_PEND, EST_PEND_T,
    EST_IDEN, EST_IDEN_T, EST_APLI, EST_APLI_T
)

# ========================CONFIGURACION======================
BASE_DIR         = Path(__file__).parent
SHARED_DIR       = BASE_DIR.parent.parent.parent / "shared"
MAESTRO_DIR      = BASE_DIR.parent / "construir_maestro" / "crear_maestro" / "outputs"
REPORTE_DIR      = SHARED_DIR / "reporte_mes_crudo"
PLANILLA_DIR     = SHARED_DIR / "planilla_mes"
PLANILLA_ANT_DIR = SHARED_DIR / "planilla_acumulado"
CORRECCIONES_DIR = BASE_DIR / "correcciones"
OUTPUT_DIR       = BASE_DIR / "outputs"

MAESTRO_FILE    = "maestro_yape.xlsx"
OUTPUT_FILE     = "pagos_yape_tepago.xlsx"
PENDIENTES_FILE = "pendientes.xlsx"
TIPO_PAGO       = "TE PAGÓ"

ALIAS_BANCO = {
    "tipo":    ["tipo de transacción", "tipo de transaccion"],
    "origen":  ["origen"],
    "monto":   ["monto"],
    "mensaje": ["mensaje"],
    "fecha":   ["fecha de operación", "fecha de operacion"],
}

ALIAS_PLANILLA = {
    "mz":           ["mz"],
    "lote":         ["lt", "lote"],
    "nombre":       ["nombres", "nombre"],
    "total":        ["total"],
    "mes_anterior": ["mes anterior"],
}

# Patrones para extraer mz y lote del mensaje
PATRONES_MENSAJE = [
    # Mz:Y. Lote  9 — punto después de MZ, múltiples espacios antes del número
    re.compile(r'mz:?\s*([A-Z][A-Z0-9]*)\.?\s+lote\s+(\d+[A-Z]?)', re.IGNORECASE),
    # Mz:Y. Lote 9 con dos puntos y punto
    re.compile(r'mz:([A-Z][A-Z0-9]*)\.?\s*(?:lt\.?|lote\.?|lte\.?)\s+(\d+[A-Z]?)', re.IGNORECASE),
    # M.Z V.LT.16 — mz con puntos intercalados
    re.compile(r'm\.z\.?\s*([A-Z][A-Z0-9]*)\.?\s*(?:lt\.?|lte\.?)\s*\.?\s*(\d+[A-Z]?)', re.IGNORECASE),
    # Mz F . Lt 10 — punto separador entre MZ y Lt
    re.compile(r'mz\.?\s*([A-Z][A-Z0-9]*)\s*\.\s*(?:lt\.?|lte\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    # mz:w. lt:17 — dos puntos en lt también
    re.compile(r'mz:?\s*([A-Z][A-Z0-9]*)\.?\s*lt:(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?\s*["\']([ A-Z0-9]+)["\']?\s*(?:lt\.?|lote\.?|lte\.?)\s*["\']?(\d+[A-Z]?)["\']?', re.IGNORECASE),
    re.compile(r'mz\.?\s*([A-Z0-9\-]+)[_]+(?:lote|lt|lte)\.?\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'maz\.?\s*([A-Z0-9]+)\.?\s*(?:lot\.?|lote\.?|lt\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz[na]+\.?\s*([A-Z0-9\-]+)\.?\s*(?:lt\.?|lot\.?|lote\.?|lte\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'man[sz]?[aio]?[aon]?[aáñ]?\s+([A-Z0-9\-]+)\s+(?:lt\.?|lot\.?|lote\.?|lte\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.([A-Z0-9\-]+)\.?\s*(?:lt\.?|lte\.?|lote\.?)\s*\.?\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz([A-Z][A-Z0-9]*)\.?\s*(?:lt\.?|lote\.?|lte\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?\s*\.?\s*([A-Z][A-Z0-9]*)\.?\s*(?:tl\.?|lt\.?|lte\.?)\s*\.?\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?\s*([A-Z0-9\-]+),\s*(?:lt\.?|lte\.?|lote\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?-?([A-Z][A-Z0-9]*)\s*[-/]\s*(?:lt\.?|lte\.?|l\.?)-?\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?\s*:?\s*([A-Z0-9\-]+)\s*:?\s*(?:lt\.?|lote\.?|lte\.?)\s*:?\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?\s*([A-Z][A-Z0-9]*)\.?\s+(?:lt\.?|lte\.?)\s*\.?\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?\s*([A-Z][A-Z0-9]*)\s+l\.?\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?\s*([A-Z0-9\-]+)\s*(?:late\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'mz\.?\s*([A-Z0-9]+)\s+la\s+(\d+)', re.IGNORECASE),
    re.compile(r'lote\s+([A-Z])\s+(\d+)', re.IGNORECASE),
    re.compile(r'\bm\.([A-Z0-9]+)\s*(?:lote\.?|lt\.?|lte\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'\bm\s+([A-Z][A-Z0-9]*)\.?\s+(?:lt\.?|lte\.?|l\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'^mz?([A-Z][A-Z0-9]*)\s+(?:lt\.?|lte\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'^([A-Z][A-Z0-9]+)\s+(?:lt\.?|lote\.?|lte\.?)\s*(\d+[A-Z]?)', re.IGNORECASE),
    re.compile(r'^([A-Z][A-Z0-9]*)-(\d+[A-Z]?)(?:\s|$|,)', re.IGNORECASE),
    re.compile(r'^([A-Z])0*(\d+)(?:\s|$)', re.IGNORECASE),
    re.compile(r'^([A-Z][A-Z0-9]*)\s+(\d+)', re.IGNORECASE),
    # Formato: Mz: D1 - Lt: 3 · Mz: i - Lote: 4
    re.compile(r'mz:?\s*([A-Z][A-Z0-9]*)\s*-\s*(?:lt|lote):?\s*(\d+[A-Z]?)', re.IGNORECASE),
]

COLOR_CABECERA     = "4A235A"
COLOR_IDENTIFICADO = "D5F5E3"
COLOR_PENDIENTE    = "FADBD8"
COLOR_ORIGEN       = "D6EAF8"
COLOR_UBICACION    = "FDEBD0"
COLOR_CONFIANZA    = "D7BDE2"
COLOR_DEUDA        = "FDEDEC"

# ========================UTILIDADES=========================
def reset_output_folder(path: Path):
    if path.exists():
        shutil.rmtree(path)
    path.mkdir()

def estilo_celda(cell, bg_color=None, align="left"):
    lado = Side(style="thin", color="CCCCCC")
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = Border(left=lado, right=lado, top=lado, bottom=lado)
    if bg_color:
        cell.fill = PatternFill("solid", start_color=bg_color)

def cabecera_celda(cell, texto):
    cell.value     = texto
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", start_color=COLOR_CABECERA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    lado = Side(style="thin", color="FFFFFF")
    cell.border    = Border(left=lado, right=lado, top=lado, bottom=lado)

def limpiar_lote(val) -> str:
    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN", "", "LT", "LOTE"):
        return ""
    # Intentar número puro primero
    try:
        return str(int(float(s)))
    except:
        # Lote con letra al final: 8a → 8A, 12b → 12B
        return s.strip().upper()

def limpiar_monto(val) -> float:
    try:
        return float(str(val).replace(",", ".").strip())
    except:
        return 0.0

# ====================ANCLA DE CORTE=========================
def obtener_ancla() -> datetime | None:
    """
    Lee reporte_acumulado_crudo/ — toma el archivo más reciente
    y extrae la fecha máxima de sus registros como ancla de corte.
    """
    archivos = sorted(SHARED_DIR.joinpath("reporte_acumulado_procesado").glob("*.xlsx"))
    if not archivos:
        print(f"  ⚠ No hay archivos en reporte_acumulado_procesado/ — sin ancla de corte")
        return None

    # Tomar el más reciente por nombre
    archivo = archivos[-1]
    print(f"  Reporte procesado más reciente: {archivo.name}")

    wb    = load_workbook(archivo, read_only=True, data_only=True)
    ws    = wb.active
    datos = list(ws.values)
    wb.close()

    if len(datos) < 2:
        return None

    headers   = [str(h).strip().lower() if h else "" for h in datos[0]]
    col_fecha = next((i for i, h in enumerate(headers) if "fecha" in h), None)
    if col_fecha is None:
        print("  ⚠ No se encontró columna fecha en reporte acumulado — sin ancla")
        return None

    fechas = []
    for fila in datos[1:]:
        if not fila or col_fecha >= len(fila):
            continue
        val = fila[col_fecha]
        if isinstance(val, datetime):
            fechas.append(val)
            continue
        s = str(val).strip()
        for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
                    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                fechas.append(datetime.strptime(s, fmt))
                break
            except:
                pass

    if not fechas:
        print("  ⚠ Sin fechas válidas en reporte acumulado — sin ancla")
        return None

    ancla = max(fechas)
    print(f"  Ancla de corte: {ancla.strftime('%d/%m/%Y %H:%M:%S')}")
    return ancla

# ====================EXTRACCION MENSAJE=====================
def extraer_mz_lote_mensaje(mensaje: str) -> tuple:
    """
    Intenta extraer mz y lote del mensaje. Siempre tiene prioridad sobre el origen.
    Retorna (mz, lote) si encuentra, (None, None) si no.
    """
    if not mensaje or str(mensaje).strip() in ("", "nan", "None"):
        return None, None

    msg = str(mensaje).strip()

    for patron in PATRONES_MENSAJE:
        m = patron.search(msg)
        if m:
            grupos = [g for g in m.groups() if g is not None]
            if len(grupos) >= 2:
                mz  = grupos[0].strip().upper()
                try:
                    lote = str(int(grupos[1].strip()))
                except:
                    lote = grupos[1].strip()
                if mz and lote:
                    return mz, lote

    return None, None

# ====================NORMALIZACION MZ=======================
def normalizar_mz(mz: str, mzs_validas: set) -> str:
    """
    Limpia la MZ extraída del mensaje:
    1. Quita cualquier carácter que no sea letra o número (guiones, puntos, etc.)
    2. Si el resultado no existe en planilla pero reemplazando 0→O sí existe → corrige
    3. Si sigue sin existir → retorna la MZ limpia igual (será pendiente)
    """
    if not mz:
        return mz

    # Paso 1: quitar todo lo que no sea letra o número
    limpia = re.sub(r'[^A-Z0-9]', '', mz.upper())
    if not limpia:
        return mz

    # Paso 2: si ya existe → devolver
    if limpia in mzs_validas:
        return limpia

    # Paso 3: intentar 0 → O
    con_o = limpia.replace('0', 'O')
    if con_o in mzs_validas:
        return con_o

    return limpia

# ====================NORMALIZACION==========================
def normalizar_columnas(df: pd.DataFrame, alias: dict) -> dict:
    cols_lower = {str(c).lower().strip(): c for c in df.columns
                  if c and str(c).strip() not in ("None", "nan")}
    mapa = {}
    for campo, variantes in alias.items():
        for v in variantes:
            if v in cols_lower:
                mapa[campo] = cols_lower[v]
                break
    return mapa

# ====================BUSCAR EN USUARIOS_ID==================
_cache_uid = None

def cargar_uid_cache() -> dict:
    global _cache_uid
    if _cache_uid is not None:
        return _cache_uid
    path = SHARED_DIR / "usuarios_id.xlsx"
    if not path.exists():
        _cache_uid = {}
        return _cache_uid
    wb   = load_workbook(path, read_only=True, data_only=True)
    ws   = wb.active
    rows = list(ws.values)
    wb.close()
    if len(rows) < 2:
        _cache_uid = {}
        return _cache_uid
    headers = {str(h).strip().upper(): i for i, h in enumerate(rows[0]) if h}
    mapa = {}
    for fila in rows[1:]:
        if not fila:
            continue
        def get(key):
            i = headers.get(key)
            return str(fila[i]).strip() if i is not None and i < len(fila) and fila[i] else ""
        uid = get("USER_ID")
        nom = get("NOMBRE")
        mz  = get("MZ").upper()
        lt  = limpiar_lote(get("LOTE") or get("LT"))
        mz2 = get("MZ2").upper()
        lt2 = limpiar_lote(get("LOTE2") or get("LT2"))
        if uid and mz and lt:
            mapa[(mz, lt)] = (uid, nom)
        if uid and mz2 and lt2:
            mapa[(mz2, lt2)] = (uid, nom)
    _cache_uid = mapa
    return mapa

def buscar_uid(mz: str, lote: str) -> tuple[str, str]:
    cache = cargar_uid_cache()
    return cache.get((mz.upper().strip(), limpiar_lote(lote)), ("", ""))

# ====================CARGA DE DATOS=========================
def cargar_maestro() -> tuple[dict, dict]:
    """
    Retorna:
      indice         — origen_key → {user_id, mz, lote, nivel_confianza}  (1 solo lote)
      indice_ambiguo — origen_key → lista de candidatos [{mz, lote, nombre, nivel}]
    """
    path = MAESTRO_DIR / MAESTRO_FILE
    if not path.exists():
        raise FileNotFoundError(f"No se encontró {path}")

    wb    = load_workbook(path, read_only=True, data_only=True)
    ws    = wb.active
    datos = list(ws.values)
    wb.close()

    # maestro_yape tiene cabecera doble: fila 0 = grupos · fila 1 = columnas reales
    headers = [str(h).strip().upper() if h else "" for h in datos[1]]

    try:
        col_uid  = headers.index("USER_ID")
        col_mz   = headers.index("MZ")
        col_lote = headers.index("LOTE")
    except ValueError as e:
        raise ValueError(f"Columna no encontrada en maestro: {e}")
    col_nivel = None

    col_nombre  = headers.index("NOMBRE") if "NOMBRE" in headers else None
    cols_origen = [i for i, h in enumerate(headers) if h.startswith("ORIGEN_")]

    # Acumular todos los lotes por origen — saltar filas de cabecera (0 y 1)
    origen_a_lotes = {}
    for fila in datos[2:]:
        if not fila:
            continue
        uid    = str(fila[col_uid]).strip()        if fila[col_uid]   else ""
        mz     = str(fila[col_mz]).strip().upper() if fila[col_mz]    else ""
        lote   = limpiar_lote(fila[col_lote])
        nivel  = ""
        nombre = str(fila[col_nombre]).strip()     if col_nombre and fila[col_nombre] else ""

        if not mz or not lote:
            continue

        for ci in cols_origen:
            if ci < len(fila) and fila[ci]:
                origen_key = str(fila[ci]).strip().upper()
                if origen_key:
                    if origen_key not in origen_a_lotes:
                        origen_a_lotes[origen_key] = []
                    if not any(c["mz"]==mz and c["lote"]==lote for c in origen_a_lotes[origen_key]):
                        origen_a_lotes[origen_key].append({
                            "user_id":         uid,
                            "mz":              mz,
                            "lote":            lote,
                            "nivel_confianza": nivel,
                            "nombre":          nombre,
                        })

    # Separar únicos de ambiguos
    indice         = {}
    indice_ambiguo = {}

    for origen_key, candidatos in origen_a_lotes.items():
        if len(candidatos) == 1:
            indice[origen_key] = candidatos[0]
        else:
            indice_ambiguo[origen_key] = candidatos

    print(f"  ✔ {len(indice)} orígenes únicos indexados")
    if indice_ambiguo:
        print(f"  ⚠ {len(indice_ambiguo)} orígenes ambiguos → irán a pendientes")
    return indice, indice_ambiguo

def cargar_planilla() -> tuple[dict, set]:
    archivos = list(PLANILLA_DIR.glob("*.xlsx"))
    if not archivos:
        print(f"  ⚠ No hay planilla en {PLANILLA_DIR} — se omite cruce de deuda")
        return {}, set()

    archivo = archivos[0]
    print(f"  Leyendo planilla: {archivo.name}")

    wb    = load_workbook(archivo, read_only=True, data_only=True)
    ws    = wb.active
    datos = []
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row):
            datos.append(row)
    wb.close()

    if len(datos) < 2:
        return {}, set()

    n_cols = len(datos[0])
    df     = pd.DataFrame([f[:n_cols] for f in datos[1:]], columns=datos[0])
    mapa   = normalizar_columnas(df, ALIAS_PLANILLA)

    faltantes = [c for c in ["mz", "lote", "total"] if c not in mapa]
    if faltantes:
        print(f"  ⚠ Columnas faltantes en planilla: {faltantes}")
        return {}, set()

    planilla    = {}
    mzs_validas = set()

    for _, fila in df.iterrows():
        mz   = str(fila[mapa["mz"]]).strip().upper() if mapa.get("mz")   else ""
        lote = limpiar_lote(fila[mapa["lote"]])       if mapa.get("lote") else ""
        if not mz or not lote or mz in ("NAN", "NONE", ""):
            continue
        mzs_validas.add(mz)
        nombre       = str(fila[mapa["nombre"]]).strip() if mapa.get("nombre") else ""
        total        = limpiar_monto(fila[mapa["total"]])
        mes_anterior = limpiar_monto(fila[mapa["mes_anterior"]]) if mapa.get("mes_anterior") else 0.0
        planilla[(mz, lote)] = {
            "nombre":       nombre,
            "deuda_total":  total,
            "mes_anterior": mes_anterior,
        }

    print(f"  ✔ {len(planilla)} usuarios | {len(mzs_validas)} manzanas válidas")
    return planilla, mzs_validas

def cargar_reportes(ancla=None) -> tuple:
    archivos = sorted(REPORTE_DIR.glob("*.xlsx"))
    if not archivos:
        raise FileNotFoundError(f"No hay .xlsx en {REPORTE_DIR}")

    print(f"  Archivos encontrados: {len(archivos)}")
    todos      = []
    mapa_final = {}

    for archivo in archivos:
        print(f"  Leyendo: {archivo.name}")

        wb    = load_workbook(archivo, read_only=False, data_only=True)
        ws    = wb.active
        datos = []
        for row in ws.iter_rows(values_only=True):
            if any(v is not None for v in row):
                datos.append(row)
        wb.close()

        if len(datos) < 2:
            continue

        header_idx = 0
        for i, fila in enumerate(datos):
            if fila[0] and str(fila[0]).strip().lower() in (
                "tipo de transacción", "tipo de transaccion"
            ):
                header_idx = i
                break

        headers = datos[header_idx]
        filas   = datos[header_idx + 1:]
        n_cols  = len(headers)
        df      = pd.DataFrame([f[:n_cols] for f in filas], columns=headers)

        mapa = normalizar_columnas(df, ALIAS_BANCO)
        faltantes = [c for c in ["tipo", "origen", "monto"] if c not in mapa]
        if faltantes:
            print(f"  ⚠ Saltado — columnas faltantes: {faltantes}")
            continue

        df = df[df[mapa["tipo"]].astype(str).str.strip().str.upper() == TIPO_PAGO.upper()].copy()
        df = df.reset_index(drop=True)
        print(f"  ✔ {len(df)} filas TE PAGÓ")
        todos.append(df)
        mapa_final = mapa

    if not todos:
        raise ValueError("Ningún archivo válido en Reporte_mes")

    df_total = pd.concat(todos, ignore_index=True)

    cols_dedup = [mapa_final.get(c) for c in ["origen", "monto", "fecha"] if mapa_final.get(c)]
    antes = len(df_total)
    df_total = df_total.drop_duplicates(subset=cols_dedup).reset_index(drop=True)
    if antes > len(df_total):
        print(f"  ⚠ {antes - len(df_total)} duplicados eliminados")

    # Filtrar por ancla de corte
    if ancla is not None and mapa_final.get("fecha"):
        col_f = mapa_final["fecha"]

        def _parsear(val):
            if isinstance(val, datetime):
                return val
            s = str(val).strip()
            for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
                        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    return datetime.strptime(s, fmt)
                except:
                    pass
            return None

        df_total["_DT"] = df_total[col_f].apply(_parsear)
        antes_ancla = len(df_total)
        df_total = df_total[df_total["_DT"].notna() & (df_total["_DT"] > ancla)]
        df_total = df_total.drop(columns=["_DT"]).reset_index(drop=True)
        excluidos = antes_ancla - len(df_total)
        if excluidos:
            print(f"  ⚠ {excluidos} filas anteriores al ancla excluidas")

    print(f"  Total final: {len(df_total)} filas")
    return df_total, mapa_final

ACUMULADAS_FILE = "trazabilidad_{}.xlsx".format(datetime.today().strftime("%Y_%m"))

# ── Detector de pagos múltiples ──────────────────────────
PATRON_MULTIPLE = re.compile(
    r'mz\.?\s*([A-Z][A-Z0-9]*)\s*\.?\s*(?:lt\.?|lte\.?|lote\.?)\s*\.?\s*(\d+[A-Z]?)',
    re.IGNORECASE
)

# Patrón para misma MZ con múltiples lotes: "Mz K 3 y 4" → K-3 y K-4
PATRON_MISMA_MZ = re.compile(
    r'mz\.?\s*([A-Z][A-Z0-9]*)\s+(\d+[A-Z]?)\s*(?:y|,)\s*(\d+[A-Z]?)',
    re.IGNORECASE
)

def extraer_multiples(mensaje: str) -> list:
    """
    Extrae todos los pares MZ-LOTE de un mensaje con múltiples lotes.
    Casos:
      'MzE Lt7, MzP Lt11A y MzM Lt18' → [(E,7),(P,11A),(M,18)]
      'Mz K 3 y 4'                     → [(K,3),(K,4)]
    Retorna lista si hay 2+ pares, lista vacía si hay 0 o 1.
    """
    if not mensaje or str(mensaje).strip() in ("", "nan", "None"):
        return []

    # Caso 1: múltiples MZ-LOTE distintos
    pares = PATRON_MULTIPLE.findall(mensaje)
    if len(pares) >= 2:
        return [(mz.strip().upper(), lote.strip().upper()) for mz, lote in pares]

    # Caso 2: misma MZ con múltiples lotes "Mz K 3 y 4"
    m = PATRON_MISMA_MZ.search(mensaje)
    if m:
        mz    = m.group(1).strip().upper()
        lote1 = m.group(2).strip().upper()
        lote2 = m.group(3).strip().upper()
        return [(mz, lote1), (mz, lote2)]

    return []


# ── Leer acumuladas (3 hojas) ────────────────────────────
def _leer_hoja_acumuladas(wb, nombre_hoja: str, tiene_instruccion: bool = False) -> dict:
    """Lee una hoja → dict: origen_upper → {mz, lote, fecha}"""
    if nombre_hoja not in wb.sheetnames:
        return {}
    ws   = wb[nombre_hoja]
    data = list(ws.values)
    # Si tiene instruccion: fila 0=instruccion, fila 1=cabecera, fila 2+=datos
    # Si no tiene instruccion: fila 0=cabecera, fila 1+=datos
    if tiene_instruccion:
        if len(data) < 3:
            return {}
        headers = [str(h).strip().upper() if h else "" for h in data[1]]
        filas   = data[2:]
    else:
        if len(data) < 2:
            return {}
        headers = [str(h).strip().upper() if h else "" for h in data[0]]
        filas   = data[1:]

    resultado = {}
    for fila in filas:
        if not fila:
            continue
        row    = dict(zip(headers, fila))
        origen = str(row.get("ORIGEN", "")).strip().upper()
        mz     = str(row.get("MZ", "")).strip().upper()
        if not origen or not mz or origen in ("NAN","") or mz in ("NAN",""):
            continue
        lote = limpiar_lote(row.get("LOTE", "")) if mz != "BLANCO" else ""
        resultado[origen] = {"mz": mz, "lote": lote}
    return resultado or {}


def _leer_hoja_multiples(wb) -> dict:
    """
    Lee hoja Pagos_multiples — una fila por lote.
    Solo lee filas con AUTORIZADO=1.
    Retorna dict: origen_upper → lista de {mz, lote, monto}
    donde monto = MONTO_MANUAL si existe, sino MONTO_SISTEMA
    """
    nombre = "Pagos_multiples"
    if nombre not in wb.sheetnames:
        return {}
    ws   = wb[nombre]
    data = list(ws.values)
    if len(data) < 2:
        return {}
    # Fila 1 = instrucción, fila 2 = cabecera
    headers = [str(h).strip().upper() if h else "" for h in data[1]]
    resultado = {}
    for fila in data[2:]:
        if not fila:
            continue
        row = dict(zip(headers, fila))
        # Solo filas autorizadas
        aut = str(row.get("AUTORIZADO","")).strip()
        if aut != "1":
            continue
        origen = str(row.get("ORIGEN","")).strip().upper()
        mz     = str(row.get("MZ","")).strip().upper()
        if not origen or not mz or origen in ("NAN","") or mz in ("NAN",""):
            continue
        lote = limpiar_lote(row.get("LOTE",""))
        # Prioridad: MONTO_MANUAL sobre MONTO_SISTEMA
        monto_manual  = row.get("MONTO_MANUAL","")
        monto_sistema = row.get("MONTO_SISTEMA","")
        try:
            monto = float(str(monto_manual).replace(",",".")) if monto_manual and str(monto_manual).strip() not in ("","nan","None") else float(str(monto_sistema).replace(",","."))
        except:
            monto = 0.0
        if origen not in resultado:
            resultado[origen] = []
        resultado[origen].append({"mz": mz, "lote": lote, "monto": monto})
    return resultado if resultado else {}


def leer_correcciones(planilla: dict = None) -> tuple[dict, dict, dict]:
    """
    Lee correcciones_acumuladas.xlsx + pendientes.xlsx.
    Valida MZ-LOTE contra planilla antes de aceptar correcciones manuales.
    """
    ruta_acum = BASE_DIR / "trazabilidad" / ACUMULADAS_FILE
    ruta_pend = CORRECCIONES_DIR / PENDIENTES_FILE
    if planilla is None:
        planilla = {}

    corr_simples   = {}
    corr_ambiguos  = {}
    corr_multiples = {}

    if ruta_acum.exists():
        try:
            wb_acum = load_workbook(ruta_acum, read_only=True, data_only=True)
            corr_simples   = _leer_hoja_acumuladas(wb_acum, "Sin_identificar", tiene_instruccion=False)
            corr_ambiguos  = _leer_hoja_acumuladas(wb_acum, "Ambiguos",        tiene_instruccion=False)
            corr_multiples = _leer_hoja_multiples(wb_acum)
            wb_acum.close()
        except Exception as e:
            print(f"  ⚠ No se pudo leer acumuladas: {e}")

    # Leer pendientes completados por el usuario
    n_nuevas    = 0
    n_rechazadas = 0

    if ruta_pend.exists():
        try:
            wb_pend = load_workbook(ruta_pend, read_only=True, data_only=True)

            # Hoja Sin_identificar
            nuevas_si = _leer_hoja_acumuladas(wb_pend, "Sin_identificar", tiene_instruccion=True)
            for k, v in nuevas_si.items():
                mz   = v.get("mz","")
                lote = v.get("lote","")
                if not mz or mz in ("NAN","NONE",""):
                    continue
                # BLANCO siempre se acepta
                if mz == "BLANCO":
                    if k not in corr_simples:
                        n_nuevas += 1
                    corr_simples[k] = v
                    continue
                # Validar contra planilla
                if planilla and (mz, lote) not in planilla:
                    n_rechazadas += 1
                    print(f"  ✗ Rechazada: {k[:30]} → {mz}-{lote} no existe en planilla")
                    continue
                if k not in corr_simples:
                    n_nuevas += 1
                corr_simples[k] = v

            # Hoja Ambiguos (por compatibilidad)
            nuevas_amb = _leer_hoja_acumuladas(wb_pend, "Ambiguos", tiene_instruccion=True)
            for k, v in nuevas_amb.items():
                mz   = v.get("mz","")
                lote = v.get("lote","")
                if not mz or mz in ("NAN","NONE",""):
                    continue
                if planilla and mz != "BLANCO" and (mz, lote) not in planilla:
                    n_rechazadas += 1
                    print(f"  ✗ Rechazada: {k[:30]} → {mz}-{lote} no existe en planilla")
                    continue
                if k not in corr_ambiguos:
                    n_nuevas += 1
                corr_ambiguos[k] = v

            wb_pend.close()
        except Exception as e:
            print(f"  ⚠ No se pudo leer pendientes: {e}")

    if n_rechazadas:
        print(f"  ⚠ {n_rechazadas} correcciones rechazadas — MZ-LOTE no existe en planilla")
    if n_nuevas > 0:
        # Guardar trazabilidad con diseño HTML en carpeta trazabilidad/
        mes_str      = datetime.today().strftime("%Y_%m")
        ruta_traz    = BASE_DIR / "trazabilidad" / f"trazabilidad_{mes_str}.xlsx"
        ruta_traz.parent.mkdir(exist_ok=True)
        fecha_hoy    = datetime.today().strftime("%d/%m/%Y %H:%M")
        exportar_trazabilidad(ruta_traz, corr_simples, corr_ambiguos,
                              corr_multiples, ciclo, fecha_hoy)

    n_si   = len(corr_simples)
    n_amb  = len(corr_ambiguos)
    n_mult = len(corr_multiples)
    print(f"  Acumuladas: {n_si} sin_id + {n_amb} ambiguos + {n_mult} multiples ({n_nuevas} nuevas)")
    return corr_simples, corr_ambiguos, corr_multiples


def _guardar_acumuladas(corr_simples: dict, corr_ambiguos: dict, corr_multiples: dict):
    if corr_simples   is None: corr_simples   = {}
    if corr_ambiguos  is None: corr_ambiguos  = {}
    if corr_multiples is None: corr_multiples = {}
    """Guarda correcciones_acumuladas.xlsx con 3 hojas"""
    CORRECCIONES_DIR.mkdir(exist_ok=True)
    ruta = BASE_DIR / "trazabilidad" / ACUMULADAS_FILE
    ruta.parent.mkdir(exist_ok=True)
    fecha_hoy = datetime.now().strftime("%d/%m/%Y %H:%M")

    wb = Workbook()

    def _escribir_hoja_simple(ws, datos: dict):
        for ci, h in enumerate(["ORIGEN","MZ","LOTE","FECHA"], 1):
            cabecera_celda(ws.cell(1, ci), h)
        ws.row_dimensions[1].height = 18
        for ri, (origen, v) in enumerate(datos.items(), 2):
            ws.cell(ri, 1, value=origen)
            ws.cell(ri, 2, value=v.get("mz",""))
            ws.cell(ri, 3, value=v.get("lote",""))
            ws.cell(ri, 4, value=v.get("fecha", fecha_hoy))
            ws.row_dimensions[ri].height = 16
        for ci, w in enumerate([35,8,8,18], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    # Hoja 1: Sin_identificar
    ws1 = wb.active
    ws1.title = "Sin_identificar"
    _escribir_hoja_simple(ws1, corr_simples)

    # Hoja 2: Ambiguos
    ws2 = wb.create_sheet("Ambiguos")
    _escribir_hoja_simple(ws2, corr_ambiguos)

    # Hoja 3: Pagos_multiples
    ws3 = wb.create_sheet("Pagos_multiples")
    for ci, h in enumerate(["ORIGEN","MZ","LOTE","MONTO_ASIGNADO","FECHA"], 1):
        cabecera_celda(ws3.cell(1, ci), h)
    ws3.row_dimensions[1].height = 18
    ri = 2
    for origen, lista in corr_multiples.items():
        for item in lista:
            ws3.cell(ri, 1, value=origen)
            ws3.cell(ri, 2, value=item.get("mz",""))
            ws3.cell(ri, 3, value=item.get("lote",""))
            ws3.cell(ri, 4, value=item.get("monto",""))
            ws3.cell(ri, 5, value=item.get("fecha", fecha_hoy))
            ws3.row_dimensions[ri].height = 16
            ri += 1
    for ci, w in enumerate([35,8,8,14,18], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    wb.save(ruta)



# ====================VALIDACION=============================
def validar_reporte(df: pd.DataFrame, mapa: dict) -> pd.DataFrame:
    antes = len(df)
    col   = mapa["origen"]
    df    = df[df[col].astype(str).str.strip() != ""].copy()
    df    = df[df[col].astype(str).str.strip().str.upper() != "NAN"].copy()
    df    = df.reset_index(drop=True)
    if antes - len(df):
        print(f"  ⚠ {antes - len(df)} filas descartadas")
    print(f"  Filas válidas: {len(df)}")
    return df

# ====================PROCESAMIENTO==========================
def ejecutar_matching(df: pd.DataFrame, mapa: dict,
                      indice: dict, planilla: dict,
                      corr_simples: dict = None,
                      corr_ambiguos: dict = None,
                      corr_multiples: dict = None,
                      mzs_validas: set = None,
                      ciclo: int = 1,
                      indice_ambiguo: dict = None) -> list:
    if corr_simples   is None: corr_simples   = {}
    if corr_ambiguos  is None: corr_ambiguos  = {}
    if corr_multiples is None: corr_multiples = {}
    if indice_ambiguo is None: indice_ambiguo = {}
    if mzs_validas    is None: mzs_validas    = set(k[0] for k in planilla.keys()) if planilla else set()
    todos = []

    for _, fila in df.iterrows():
        origen     = str(fila[mapa["origen"]]).strip()
        origen_key = origen.upper()
        monto_pago = limpiar_monto(fila[mapa["monto"]]) if mapa.get("monto") else 0.0
        mensaje    = str(fila[mapa["mensaje"]]) if mapa.get("mensaje") and mapa["mensaje"] in fila.index else ""
        fecha      = str(fila[mapa["fecha"]])   if mapa.get("fecha")   and mapa["fecha"]   in fila.index else ""
        mensaje = "" if mensaje == "None" else mensaje
        fecha   = "" if fecha   == "None" else fecha

        base = {"origen": origen, "monto_pago": monto_pago, "mensaje": mensaje, "fecha": fecha}

        # Unir todas las correcciones para este origen
        corr_simple  = corr_simples.get(origen_key) or corr_ambiguos.get(origen_key)
        corr_mult    = corr_multiples.get(origen_key)
        match_origen = indice.get(origen_key)

        # ── Ciclo 2+: corrección manual tiene prioridad máxima ──
        if ciclo > 1 and corr_mult:
            # Pago múltiple ya corregido → generar una fila por lote
            for item in corr_mult:
                mz   = item["mz"]
                lote = item["lote"]
                monto_item = float(item.get("monto") or 0)
                datos_p    = planilla.get((mz, lote), {})
                deuda      = datos_p.get("deuda_total", 0.0)
                nombre     = datos_p.get("nombre", "")
                diferencia = round(monto_item - deuda, 2) if deuda else 0.0
                estado_p   = "exacto" if diferencia==0 else ("exceso" if diferencia>0 else "parcial")
                todos.append({
                    **base,
                    "monto_pago": monto_item,
                    "user_id": "", "nombre": nombre, "mz": mz, "lote": lote,
                    "nivel_confianza": "multiple_corregido",
                    "deuda_total": deuda, "mes_anterior": datos_p.get("mes_anterior",0),
                    "diferencia": diferencia, "estado_pago": estado_p,
                    "fuente": "multiple_corregido", "motivo": "", "estado": "identificado",
                })
            continue

        if ciclo > 1 and corr_simple:
            if corr_simple["mz"] == "BLANCO":
                todos.append({**base, "user_id":"","nombre":"","mz":"BLANCO","lote":"",
                    "nivel_confianza":"","deuda_total":"","mes_anterior":"",
                    "diferencia":"","estado_pago":"","fuente":"blanco",
                    "motivo":"marcado como blanco","estado":"pendiente"})
                continue
            elif corr_simple["mz"] and corr_simple["lote"]:
                mz = corr_simple["mz"]; lote = corr_simple["lote"]
                datos_p = planilla.get((mz,lote),{})
                deuda   = datos_p.get("deuda_total",0.0)
                nombre  = datos_p.get("nombre","")
                mes_ant = datos_p.get("mes_anterior",0.0)
                if deuda > 0:
                    dif = round(monto_pago - deuda, 2)
                    ep  = "exacto" if dif==0 else ("exceso" if dif>0 else "parcial")
                else:
                    dif = 0.0; ep = "sin deuda en planilla"
                todos.append({**base, "user_id":"","nombre":nombre,"mz":mz,"lote":lote,
                    "nivel_confianza":"correccion","deuda_total":deuda,"mes_anterior":mes_ant,
                    "diferencia":dif,"estado_pago":ep,"fuente":"correccion",
                    "motivo":"","estado":"identificado"})
                continue

        # ── Ciclo 1 o sin corrección: flujo normal ───────────────
        # Capa 1: detectar pagos múltiples en mensaje — siempre automático
        multiples = extraer_multiples(mensaje)
        if len(multiples) >= 2:
            deudas = []
            for mz_m, lote_m in multiples:
                lote_m  = limpiar_lote(lote_m)
                datos_c = planilla.get((mz_m, lote_m), {})
                deudas.append({
                    "mz":     mz_m,
                    "lote":   lote_m,
                    "nombre": datos_c.get("nombre", ""),
                    "deuda":  datos_c.get("deuda_total", 0.0),
                    "mes_ant":datos_c.get("mes_anterior", 0.0),
                })

            total_deudas = sum(d["deuda"] for d in deudas)
            resto        = round(monto_pago - total_deudas, 2)

            for i, d in enumerate(deudas):
                es_ultimo  = (i == len(deudas) - 1)
                # Último lote absorbe la diferencia si no cuadra
                monto_item = round(d["deuda"] + (resto if es_ultimo else 0), 2)
                dif_item   = round(monto_item - d["deuda"], 2)
                ep         = "exacto" if dif_item==0 else ("exceso" if dif_item>0 else "parcial")

                todos.append({
                    **base,
                    "monto_pago":      monto_item,
                    "user_id":         d.get("user_id", ""),
                    "nombre":          d.get("nombre", d["nombre"]),
                    "mz":              d["mz"],
                    "lote":            d["lote"],
                    "nivel_confianza": "multiple_auto",
                    "deuda_total":     d["deuda"],
                    "mes_anterior":    d["mes_ant"],
                    "diferencia":      dif_item,
                    "estado_pago":     ep,
                    "fuente":          "multiple_auto",
                    "motivo":          f"pago multiple — total:{monto_pago} deudas:{total_deudas}",
                    "estado":          "identificado",
                })

            # Registrar en acumulado
            if origen_key not in corr_multiples:
                corr_multiples[origen_key] = [
                    {"mz": d["mz"], "lote": d["lote"],
                     "monto": round(d["deuda"] + (resto if i==len(deudas)-1 else 0), 2)}
                    for i, d in enumerate(deudas)
                ]
            continue

        # Capa 2: mensaje simple → regex extrae 1 MZ-LOTE
        mz_msg, lote_msg = extraer_mz_lote_mensaje(mensaje)
        if mz_msg:
            mz_msg = normalizar_mz(mz_msg, mzs_validas)

        if mz_msg and lote_msg:
            # Mensaje dio resultado → usar directo, no es ambiguo
            mz     = mz_msg
            lote   = lote_msg
            uid    = match_origen["user_id"] if match_origen and match_origen["mz"] == mz and match_origen["lote"] == lote else ""
            nivel  = "por mensaje"
            fuente = "mensaje"

        # Capa 3: sin mensaje → ambiguo → elegir mejor candidato automático
        elif origen_key in indice_ambiguo:
            candidatos = indice_ambiguo[origen_key]

            # Calcular diferencia de cada candidato
            scored = []
            for c in candidatos:
                datos_c = planilla.get((c["mz"], c["lote"]), {})
                deuda_c = datos_c.get("deuda_total", 0.0)
                dif_c   = round(monto_pago - deuda_c, 2) if deuda_c else 9999.0
                nivel_c = c.get("nivel_confianza", "nuevo")
                # Prioridad: 1) dif=0, 2) menor dif absoluta, 3) mayor nivel confianza
                scored.append({
                    "mz":     c["mz"],
                    "lote":   c["lote"],
                    "nombre": c.get("nombre",""),
                    "uid":    c.get("user_id",""),
                    "nivel":  nivel_c,
                    "deuda":  deuda_c,
                    "dif":    dif_c,
                })

            # Ordenar: dif=0 primero, luego menor dif absoluta, luego nivel confianza
            nivel_orden = {"confirmado": 0, "en proceso": 1, "nuevo": 2, "correccion": 3}
            scored.sort(key=lambda x: (
                0 if x["dif"] == 0 else 1,
                abs(x["dif"]),
                nivel_orden.get(x["nivel"], 9)
            ))

            mejor = scored[0]
            mz    = mejor["mz"]
            lote  = mejor["lote"]
            uid   = mejor["uid"]
            nivel = mejor["nivel"]
            fuente= "ambiguo_auto"

            # Registrar en acumulado automáticamente
            # (se acumula como corrección para que ciclos futuros lo usen directo)
            if origen_key not in corr_simples:
                corr_simples[origen_key] = {"mz": mz, "lote": lote}

        elif match_origen:
            mz     = match_origen["mz"]
            lote   = match_origen["lote"]
            uid    = match_origen["user_id"]
            nivel  = match_origen["nivel_confianza"]
            fuente = "maestro"
        else:
            todos.append({
                **base,
                "user_id": "", "nombre": "", "mz": "", "lote": "",
                "nivel_confianza": "", "deuda_total": "", "mes_anterior": "",
                "diferencia": "", "estado_pago": "", "fuente": "pendiente",
                "motivo": "sin mensaje ni maestro",
                "estado": "pendiente",
            })
            continue

        # ── Validar que MZ/LOTE existan en planilla ───────────
        if planilla and (mz, lote) not in planilla:
            todos.append({
                **base,
                "user_id": uid, "nombre": "", "mz": mz, "lote": lote,
                "nivel_confianza": nivel, "deuda_total": "", "mes_anterior": "",
                "diferencia": "", "estado_pago": "no en planilla", "fuente": fuente,
                "motivo": f"{mz}-{lote} no existe en planilla",
                "estado": "pendiente",
            })
            continue

        # Cruce con planilla
        datos_p      = planilla.get((mz, lote), {})
        deuda_total  = datos_p.get("deuda_total",  0.0)
        mes_anterior = datos_p.get("mes_anterior", 0.0)
        nombre       = datos_p.get("nombre", "")

        if deuda_total > 0:
            diferencia = round(monto_pago - deuda_total, 2)
            if diferencia == 0:   estado_pago = "exacto"
            elif diferencia > 0:  estado_pago = "exceso"
            else:                 estado_pago = "parcial"
        else:
            diferencia  = 0.0
            estado_pago = "sin deuda en planilla"

        todos.append({
            **base,
            "user_id":         uid,
            "nombre":          nombre,
            "mz":              mz,
            "lote":            lote,
            "nivel_confianza": nivel,
            "deuda_total":     deuda_total,
            "mes_anterior":    mes_anterior,
            "diferencia":      diferencia,
            "estado_pago":     estado_pago,
            "fuente":          fuente,
            "motivo":          "",
            "estado":          "identificado",
        })

    identificados = [r for r in todos if r["estado"] == "identificado"]
    pendientes    = [r for r in todos if r["estado"] == "pendiente"]
    print(f"  Identificados : {len(identificados)}")
    print(f"  Pendientes    : {len(pendientes)}")
    return todos

# ====================EXPORTAR EXCEL========================
def exportar_excel(todos: list, ciclo: int = 1):
    """Exporta pagos_yape_tepago.xlsx con diseño HTML."""
    wb = exportar_pagos_tepago(todos, ciclo)
    wb.save(OUTPUT_DIR / OUTPUT_FILE)
    n_id  = len([r for r in todos if r.get("estado") == "identificado"])
    n_pen = len([r for r in todos if r.get("estado") == "pendiente"])
    print(f"\n  ✔ {OUTPUT_FILE}")
    print(f"     Identificados : {n_id}")
    print(f"     Pendientes    : {n_pen}")
    print(f"     Total         : {n_id + n_pen}")

def exportar_pendientes(todos: list):
    """Exporta pendientes.xlsx con diseño HTML."""
    sin_resolver = [
        r for r in todos
        if r.get("estado") == "pendiente" and str(r.get("mz", "")).upper() != "BLANCO"
    ]

    CORRECCIONES_DIR.mkdir(exist_ok=True)
    ruta = CORRECCIONES_DIR / PENDIENTES_FILE

    if not sin_resolver:
        print("  ✔ Sin pendientes sin resolver")
        if ruta.exists():
            ruta.unlink()
        return

    wb = exportar_pendientes_diseño(sin_resolver)
    wb.save(ruta)
    print(f"  ⚠ Pendientes: {len(sin_resolver)} sin identificar → {PENDIENTES_FILE}")

# ======================MAIN=================================
def main():
    print("=" * 55)
    print("  MOTOR MATCHING — identificación de pagos Yape")
    print("=" * 55)

    # Ciclo 1 si aún no hay output; ciclos siguientes sobreescriben
    ciclo = 1 if not (OUTPUT_DIR / OUTPUT_FILE).exists() else 2
    print(f"\n  CICLO {ciclo}")

    print("\n[1] Preparando carpeta de salida...")
    if ciclo == 1:
        reset_output_folder(OUTPUT_DIR)
        if CORRECCIONES_DIR.exists():
            shutil.rmtree(CORRECCIONES_DIR)
        CORRECCIONES_DIR.mkdir()
        print("  ✔ outputs y correcciones limpiados para Ciclo 1")
    else:
        OUTPUT_DIR.mkdir(exist_ok=True)
        CORRECCIONES_DIR.mkdir(exist_ok=True)

    print("\n[2] Leyendo ancla de corte (Planilla_anterior)...")
    ancla = obtener_ancla()

    print("\n[3] Cargando maestro...")
    indice, indice_ambiguo = cargar_maestro()

    print("\n[4] Cargando planilla del mes...")
    planilla, mzs_validas = cargar_planilla()

    print("\n[5] Cargando reportes del banco...")
    df, mapa = cargar_reportes(ancla)

    print("\n[6] Validando reporte...")
    df = validar_reporte(df, mapa)

    print("\n[7] Leyendo correcciones manuales...")
    corr_simples, corr_ambiguos, corr_multiples = leer_correcciones(planilla)

    print("\n[8] Ejecutando matching...")
    n_simples_antes  = len(corr_simples)
    n_multiples_antes = len(corr_multiples)
    todos = ejecutar_matching(df, mapa, indice, planilla,
                              corr_simples, corr_ambiguos, corr_multiples,
                              mzs_validas, ciclo, indice_ambiguo)

    # Solo guardar si el matching agregó nuevas entradas al acumulado
    if len(corr_simples) > n_simples_antes or len(corr_multiples) > n_multiples_antes:
        _guardar_acumuladas(corr_simples, corr_ambiguos, corr_multiples)  # guarda en trazabilidad/
        print(f"  Acumuladas actualizadas: {len(corr_simples)} sin_id + {len(corr_ambiguos)} ambiguos + {len(corr_multiples)} multiples")

    # Resolver USER_ID faltantes en un solo lugar
    for reg in todos:
        if not reg.get("user_id") and reg.get("mz") and reg.get("lote"):
            uid, nom = buscar_uid(reg["mz"], reg["lote"])
            reg["user_id"] = uid
            if not reg.get("nombre"):
                reg["nombre"] = nom

    # Resolver USER_ID faltantes — un solo lugar independiente del tipo de matching
    for reg in todos:
        if not reg.get("user_id") and reg.get("mz") and reg.get("lote"):
            uid, nom = buscar_uid(reg["mz"], reg["lote"])
            reg["user_id"] = uid
            if not reg.get("nombre"):
                reg["nombre"] = nom

    print("\n[9] Exportando resultados...")
    exportar_excel(todos, ciclo)

    print("\n[10] Exportando pendientes a Correcciones/...")
    exportar_pendientes(todos)

    total = len(todos)
    n_id  = len([r for r in todos if r["estado"] == "identificado"])
    pct   = round(n_id / total * 100, 1) if total else 0

    print("\n" + "=" * 55)
    print(f"  Proceso completado — {pct}% identificado automáticamente")
    print("=" * 55)

if __name__ == "__main__":
    main()
