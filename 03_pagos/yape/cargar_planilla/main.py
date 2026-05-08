# ======================== IMPORTS ========================
import shutil
import pandas as pd
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ======================== CONFIGURACION ========================
BASE_DIR         = Path(__file__).parent
INPUT_DIR        = BASE_DIR / "Inputs"
PLANILLA_MES_DIR = INPUT_DIR / "planilla_mes"
MOTOR_OUTPUT_DIR = BASE_DIR.parent / "motor_matching" / "Outputs"
CORRECCIONES_DIR = BASE_DIR / "Correcciones"
OUTPUT_DIR       = BASE_DIR / "Outputs"

PAGOS_YAPE_FILE = "pagos_yape.xlsx"
RESUMEN_FILE    = "resumen_validacion.txt"
SALDOS_FILE     = "saldos_pendientes.xlsx"

COLOR_CABECERA = "4A235A"
COL_YAPE_REP   = "Yape rep"
COL_AJUSTE     = "Ajuste"
COL_MEDIO      = "Medio"
COL_ESTADO     = "Estado"
COL_LLAVE      = "LLAVE (Mz, Lt)"

# ======================== UTILIDADES ========================
def lado():
    return Side(style="thin", color="CCCCCC")

def estilo(cell, bg=None, align="left"):
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    b = lado()
    cell.border    = Border(left=b, right=b, top=b, bottom=b)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)

def cabecera(cell, texto):
    cell.value     = texto
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", start_color=COLOR_CABECERA)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    b = Side(style="thin", color="FFFFFF")
    cell.border    = Border(left=b, right=b, top=b, bottom=b)

def limpiar_lote(val) -> str:
    s = str(val).strip()
    if not s or s.upper() in ("NONE", "NAN", ""):
        return ""
    try:
        return str(int(float(s)))
    except:
        return s.strip().upper()

def normalizar_llave(mz, lote) -> str:
    mz   = str(mz).strip().upper()
    lote = limpiar_lote(lote)
    return f"{mz}-{lote}" if mz and lote else ""

# ======================== CICLO ========================
def leer_ciclo() -> int:
    ruta = OUTPUT_DIR / RESUMEN_FILE
    if not ruta.exists():
        return 1
    try:
        for linea in ruta.read_text(encoding="utf-8").splitlines():
            if "Ciclo" in linea:
                return int(linea.split(":")[1].strip()) + 1
    except:
        pass
    return 1

# ======================== PASO 1: CARGAR PAGOS_YAPE ========================
def cargar_pagos() -> pd.DataFrame:
    ruta = MOTOR_OUTPUT_DIR / PAGOS_YAPE_FILE
    if not ruta.exists():
        raise FileNotFoundError(f"No se encontro {ruta}\n  Corre motor_matching primero.")
    df = pd.read_excel(ruta, sheet_name="Pagos", dtype=str)
    df.columns = [c.strip().upper() for c in df.columns]
    print(f"  Total pagos_yape  : {len(df)}")
    identificados = df[df["ESTADO"].str.strip().str.lower() == "identificado"].copy()
    print(f"  Identificados     : {len(identificados)}")
    print(f"  Blanco/pendiente  : {len(df) - len(identificados)} ignorados")
    return identificados

# ======================== PASO 2: CALCULAR PAGOS POR LLAVE ========================
def calcular_pagos(identificados: pd.DataFrame) -> dict:
    pagos = {}
    for _, fila in identificados.iterrows():
        mz   = str(fila.get("MZ", "")).strip().upper()
        lote = limpiar_lote(fila.get("LOTE", ""))
        if not mz or not lote or mz in ("NAN", "BLANCO", ""):
            continue
        llave = f"{mz}-{lote}"
        try:
            monto = float(str(fila.get("MONTO_PAGO", 0)).replace(",", "."))
        except:
            monto = 0.0
        pagos[llave] = pagos.get(llave, 0.0) + monto
    print(f"  Llaves unicas con pago: {len(pagos)}")
    return pagos

# ======================== PASO 3: LEER LLAVES DE COBRANZA ========================
def leer_llaves_cobranza(ruta_planilla: Path):
    wb_data = load_workbook(ruta_planilla, data_only=True)
    hojas   = [h for h in wb_data.sheetnames if "cobranza" in h.lower()]
    if not hojas:
        raise ValueError("No se encontro hoja Cobranza")
    ws = wb_data[hojas[0]]

    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    col_llave = headers.get(COL_LLAVE)
    col_mz    = headers.get("MZ")
    col_lt    = headers.get("LT")
    col_total = headers.get("Total")

    llaves  = {}
    totales = {}

    for row in ws.iter_rows(min_row=2):
        llave_val = row[col_llave - 1].value if col_llave else None
        if llave_val and str(llave_val).strip().upper() not in ("", "NAN", "NONE"):
            llave = str(llave_val).strip().upper()
        else:
            mz_v  = row[col_mz - 1].value if col_mz else None
            lt_v  = row[col_lt - 1].value if col_lt else None
            llave = normalizar_llave(mz_v, lt_v) if mz_v and lt_v else ""

        if llave:
            llaves[llave] = row[0].row
            if col_total:
                try:
                    totales[llave] = float(row[col_total - 1].value or 0)
                except:
                    totales[llave] = 0.0

    wb_data.close()
    return llaves, headers, totales

def limpiar_columnas_cobranza(ws, headers: dict):
    """Limpia Yape rep, Ajuste, Medio, Estado en todas las filas"""
    col_yape_rep = headers.get(COL_YAPE_REP)
    col_ajuste   = headers.get(COL_AJUSTE)
    col_medio    = headers.get(COL_MEDIO)
    col_estado   = headers.get(COL_ESTADO)

    for row in ws.iter_rows(min_row=2):
        if col_yape_rep:
            ws.cell(row=row[0].row, column=col_yape_rep).value = None
        if col_ajuste:
            ws.cell(row=row[0].row, column=col_ajuste).value   = None
        if col_medio:
            ws.cell(row=row[0].row, column=col_medio).value    = None
        if col_estado:
            ws.cell(row=row[0].row, column=col_estado).value   = None
    print("  Columnas Yape rep/Ajuste/Medio/Estado limpiadas")

# ======================== PASO 4: ESCRIBIR COBRANZA ========================
def escribir_cobranza(ruta_output: Path, pagos: dict, ciclo: int):
    print(f"  Planilla mes: {ruta_output.name}")
    llaves, headers, totales = leer_llaves_cobranza(ruta_output)
    print(f"  Llaves leidas en Cobranza: {len(llaves)}")

    wb = load_workbook(ruta_output)
    hojas = [h for h in wb.sheetnames if "cobranza" in h.lower()]
    ws = wb[hojas[0]]

    col_yape_rep = headers.get(COL_YAPE_REP)
    col_ajuste   = headers.get(COL_AJUSTE)
    col_medio    = headers.get(COL_MEDIO)
    col_estado   = headers.get(COL_ESTADO)

    # Ciclo 1 → limpiar columnas antes de escribir
    if ciclo == 1:
        limpiar_columnas_cobranza(ws, headers)

    cargados = 0
    for llave, num_fila in llaves.items():
        if llave not in pagos:
            continue
        monto  = pagos[llave]
        total  = totales.get(llave, 0.0)
        ajuste = round(monto - total, 2)

        if col_yape_rep:
            ws.cell(row=num_fila, column=col_yape_rep).value = monto
        if col_ajuste:
            ws.cell(row=num_fila, column=col_ajuste).value   = ajuste
        if col_medio:
            ws.cell(row=num_fila, column=col_medio).value    = "Y"
        if col_estado:
            ws.cell(row=num_fila, column=col_estado).value   = "C"
        cargados += 1

    print(f"  Filas cargadas en Cobranza: {cargados}")
    return wb, cargados

# ======================== PASO 5: VALIDAR SALDOS POR MZ ========================
def leer_autorizaciones() -> dict:
    ruta = CORRECCIONES_DIR / SALDOS_FILE
    if not ruta.exists():
        return {}
    try:
        df = pd.read_excel(ruta, header=1, dtype=str)
        df.columns = [c.strip().upper() for c in df.columns]
        if "MZ" not in df.columns:
            return {}
        result = {}
        for _, r in df.iterrows():
            mz = str(r.get("MZ", "")).strip().upper()
            if mz:
                result[mz] = {
                    "autorizado": str(r.get("AUTORIZADO", "")).strip(),
                    "comentario": str(r.get("COMENTARIO", "")).strip(),
                }
        return result
    except:
        return {}


def validar_saldos(ruta_output: Path):
    """
    Por cada MZ compara:
      Suma Yape rep (Cobranza) vs Suma MONTO_PAGO (pagos_yape del mes)
    Si difieren → va a saldos_pendientes.xlsx
    """
    autorizaciones = leer_autorizaciones()

    # Leer Yape rep por MZ desde Cobranza
    wb_data = load_workbook(ruta_output, data_only=True)
    hojas   = [h for h in wb_data.sheetnames if "cobranza" in h.lower()]
    if not hojas:
        wb_data.close()
        return True, []

    ws = wb_data[hojas[0]]
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    col_mz       = headers.get("MZ")
    col_yape_rep = headers.get(COL_YAPE_REP)

    suma_cobranza = {}
    for row in ws.iter_rows(min_row=2):
        mz_val = row[col_mz - 1].value if col_mz else None
        yr_val = row[col_yape_rep - 1].value if col_yape_rep else None
        if not mz_val:
            continue
        mz = str(mz_val).strip().upper()
        if mz in ("NAN", "NONE", ""):
            continue
        try:
            suma_cobranza[mz] = suma_cobranza.get(mz, 0.0) + float(yr_val or 0)
        except:
            pass
    wb_data.close()

    # Leer MONTO_PAGO por MZ desde pagos_yape
    ruta_yape = MOTOR_OUTPUT_DIR / PAGOS_YAPE_FILE
    df_yape   = pd.read_excel(ruta_yape, sheet_name="Pagos", dtype=str)
    df_yape.columns = [c.strip().upper() for c in df_yape.columns]
    df_yape   = df_yape[df_yape["ESTADO"].str.strip().str.lower() == "identificado"].copy()
    df_yape["MZ"]         = df_yape["MZ"].astype(str).str.strip().str.upper()
    df_yape["MONTO_PAGO"] = pd.to_numeric(df_yape["MONTO_PAGO"], errors="coerce").fillna(0)
    suma_banco = df_yape.groupby("MZ")["MONTO_PAGO"].sum().to_dict()

    # Comparar por MZ
    todas_mz = set(list(suma_cobranza.keys()) + list(suma_banco.keys()))
    errores  = []
    for mz in sorted(todas_mz):
        cob   = round(suma_cobranza.get(mz, 0.0), 2)
        banco = round(suma_banco.get(mz, 0.0), 2)
        dif   = round(cob - banco, 2)
        if dif != 0:
            aut = autorizaciones.get(mz, {}).get("autorizado", "")
            if aut != "1":
                errores.append({
                    "MZ":         mz,
                    "Cobranza":   cob,
                    "Banco":      banco,
                    "Diferencia": dif,
                })

    if errores:
        _exportar_saldos(errores, autorizaciones)
        return False, [
            f"SALDOS != 0 en {len(errores)} manzanas "
            f"-> abre Correcciones/saldos_pendientes.xlsx · investiga · pon AUTORIZADO=1 · vuelve a correr"
        ]

    ruta_saldos = CORRECCIONES_DIR / SALDOS_FILE
    if ruta_saldos.exists():
        ruta_saldos.unlink()
    return True, []
    return True, []

def _exportar_saldos(errores: list, autorizaciones: dict):
    CORRECCIONES_DIR.mkdir(exist_ok=True)
    ruta = CORRECCIONES_DIR / SALDOS_FILE

    wb = Workbook(); ws = wb.active
    ws.title = "Saldos MZ"; ws.freeze_panes = "A2"

    instruccion = "Suma Yape rep (Cobranza) vs Suma pagos_yape (banco) por MZ. Diferencia != 0 = revisar. Pon 1 en AUTORIZADO para cerrar."
    ws.merge_cells(f"A1:{get_column_letter(6)}1")
    ws["A1"]           = instruccion
    ws["A1"].font      = Font(name="Arial", bold=True, size=10, color="1E40AF")
    ws["A1"].fill      = PatternFill("solid", start_color="DBEAFE")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    cols = ["MZ", "Cobranza", "Banco", "Diferencia", "AUTORIZADO", "COMENTARIO"]
    for ci, col in enumerate(cols, 1):
        cabecera(ws.cell(2, ci), col)
    ws.row_dimensions[2].height = 18

    anchos = {"MZ": 8, "Cobranza": 14, "Banco": 14,
              "Diferencia": 14, "AUTORIZADO": 13, "COMENTARIO": 38}

    for ri, fila in enumerate(errores, 3):
        mz   = fila.get("MZ", "")
        prev = autorizaciones.get(mz, {})
        dif  = fila.get("Diferencia", 0)
        for ci, col in enumerate(cols, 1):
            if col == "AUTORIZADO":
                val = prev.get("autorizado", "")
                bg  = "FFF9C4"
            elif col == "COMENTARIO":
                val = prev.get("comentario", "")
                bg  = "FFF9C4"
            elif col == "Diferencia":
                val = dif
                bg  = "FECACA" if dif != 0 else "D1FAE5"
            else:
                val = fila.get(col, "")
                bg  = "FEF2F2"
            c = ws.cell(ri, ci, value=val)
            c.font      = Font(name="Arial", size=10)
            c.fill      = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(
                horizontal="right" if col in ("Cobranza","Banco","Diferencia") else "center",
                vertical="center"
            )
            if isinstance(val, float):
                c.number_format = '#,##0.00'
        ws.row_dimensions[ri].height = 16

    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = anchos.get(col, 14)

    wb.save(ruta)
    print(f"  Saldos pendientes: {len(errores)} manzanas -> Correcciones/saldos_pendientes.xlsx")

# ======================== PASO 6: RESUMEN ========================
def escribir_resumen(ciclo: int, ok: bool, errores: list, n_cargados: int):
    OUTPUT_DIR.mkdir(exist_ok=True)
    ruta = OUTPUT_DIR / RESUMEN_FILE
    lineas = [
        "=" * 55,
        "  RESUMEN — cargar_planilla",
        f"  {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        "=" * 55,
        f"  Ciclo          : {ciclo}",
        f"  Filas cargadas : {n_cargados}",
        "",
        "  ESTADO: " + ("CERRADO — todo validado" if ok else "ABIERTO — revisa errores"),
    ]
    if errores:
        lineas += ["", "  ERRORES:"] + [f"    -> {e}" for e in errores]
    lineas.append("=" * 55)
    ruta.write_text("\n".join(lineas), encoding="utf-8")
    print(f"  Resumen guardado: {RESUMEN_FILE}")

# ======================== MAIN ========================
def main():
    print("=" * 55)
    print("  CARGAR PLANILLA — pagos Yape")
    print("=" * 55)

    ciclo = leer_ciclo()
    print(f"\n  CICLO {ciclo}")

    if ciclo == 1:
        if OUTPUT_DIR.exists():
            shutil.rmtree(OUTPUT_DIR)
        OUTPUT_DIR.mkdir()
        print("\n  Outputs limpiado para Ciclo 1")
    else:
        OUTPUT_DIR.mkdir(exist_ok=True)
        CORRECCIONES_DIR.mkdir(exist_ok=True)

    print("\n[1] Cargando pagos_yape...")
    identificados = cargar_pagos()

    print("\n[2] Calculando pagos por llave...")
    pagos = calcular_pagos(identificados)

    print("\n[3] Preparando planilla cargada...")
    archivos = list(PLANILLA_MES_DIR.glob("*.xlsx"))
    if not archivos:
        raise FileNotFoundError(f"No hay planilla en {PLANILLA_MES_DIR}")
    ruta_planilla = archivos[0]
    ruta_output   = OUTPUT_DIR / ("planilla_cargada_" + ruta_planilla.name)
    shutil.copy2(ruta_planilla, ruta_output)

    print("\n[4] Escribiendo Cobranza...")
    wb, n_cargados = escribir_cobranza(ruta_output, pagos, ciclo)

    print("\n[5] Guardando planilla cargada...")
    wb.save(ruta_output)
    print(f"  {ruta_output.name}")

    print("\n[6] Validando saldos Yape por MZ...")
    ok, errores = validar_saldos(ruta_output)

    print("\n[7] Guardando resumen...")
    escribir_resumen(ciclo, ok, errores, n_cargados)

    print("\n" + "=" * 55)
    if ok:
        print("  VALIDACION COMPLETA — planilla lista")
    else:
        print("  CICLO INCOMPLETO:")
        for e in errores:
            print(f"    -> {e}")
        print("\n  Corrige y vuelve a correr main.py")
    print("=" * 55)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        print("\n" + "=" * 55)
        print("  ERROR FATAL:")
        print(f"  {e}")
        print("\n  Detalle:")
        traceback.print_exc()
        print("=" * 55)
