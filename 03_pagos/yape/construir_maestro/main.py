# =========================IMPORTS===========================
import shutil
import pandas as pd
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ========================CONFIGURACION======================
BASE_DIR        = Path(".")
INPUT_DIR       = BASE_DIR / "Inputs"
PLANILLA_DIR    = INPUT_DIR / "Planilla"
USUARIO_ID_DIR  = INPUT_DIR / "Usuario id"
OUTPUT_DIR      = BASE_DIR / "Outputs"

MAESTRO_PATH    = OUTPUT_DIR / "maestro_yape.xlsx"
USUARIO_ID_FILE = "usuarios_id.xlsx"

HOJA_REPORTE    = "Reporte"
TIPO_PAGO       = "TE PAGÓ"

ALIAS_COLUMNAS = {
    "tipo":    ["tipo de transacción", "tipo de transaccion"],
    "origen":  ["origen"],
    "monto":   ["monto"],
    "mz":      ["mz", "manzana"],
    "lote":    ["lt", "lote"],
    "mensaje": ["mensaje"],
    "fecha":   ["fecha de operación", "fecha de operacion"],
}

COLUMNAS_OBLIGATORIAS = ["tipo", "origen", "mz", "lote"]

COLOR_CABECERA   = "4A235A"
COLOR_IDENTIDAD  = "D6EAF8"
COLOR_UBICACION  = "D5F5E3"
COLOR_ORIGENES   = "FDEBD0"
COLOR_CONFIANZA  = "FDEDEC"
COLOR_NUEVO      = "FADBD8"
COLOR_EN_PROCESO = "FDEBD0"
COLOR_CONFIRMADO = "D5F5E3"
COLOR_VALIDADO   = "D7BDE2"
COLOR_SIN_ID     = "F2F3F4"

# ========================UTILIDADES=========================
def reset_output_folder(path: Path):
    if path.exists():
        shutil.rmtree(path)
    path.mkdir()

def nivel_confianza(meses: int) -> str:
    if meses >= 3:   return "confirmado"
    elif meses == 2: return "en proceso"
    else:            return "nuevo"

def color_nivel(nivel: str) -> str:
    return {
        "nuevo":       COLOR_NUEVO,
        "en proceso":  COLOR_EN_PROCESO,
        "confirmado":  COLOR_CONFIRMADO,
        "validado":    COLOR_VALIDADO,
    }.get(nivel, COLOR_NUEVO)

def estilo_celda(cell, bold=False, bg_color=None, align="left"):
    lado = Side(style="thin", color="CCCCCC")
    cell.font      = Font(name="Arial", bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = Border(left=lado, right=lado, top=lado, bottom=lado)
    if bg_color:
        cell.fill = PatternFill("solid", start_color=bg_color)

def limpiar_lote(val) -> str:
    s = str(val).strip()
    if not s or s in ("None", "NAN", "LT", "LOTE"):
        return ""
    try:
        return str(int(float(s)))
    except:
        return s

# ====================NORMALIZACION==========================
def normalizar_columnas(df: pd.DataFrame) -> dict:
    cols_lower = {c.lower().strip(): c for c in df.columns if c}
    mapa = {}
    for campo, variantes in ALIAS_COLUMNAS.items():
        for v in variantes:
            if v in cols_lower:
                mapa[campo] = cols_lower[v]
                break
    return mapa

def validar_columnas_obligatorias(mapa: dict, archivo: str) -> bool:
    faltantes = [c for c in COLUMNAS_OBLIGATORIAS if c not in mapa]
    if faltantes:
        print(f"  ⚠ Saltado — columnas faltantes: {faltantes} ({archivo})")
        return False
    return True

# ====================CARGA DE DATOS=========================
def cargar_usuario_id() -> dict:
    path = USUARIO_ID_DIR / USUARIO_ID_FILE
    if not path.exists():
        print(f"  ⚠ No se encontró {path} — maestro sin user_id")
        return {}

    wb   = load_workbook(path, read_only=True, data_only=True)
    ws   = wb.active
    datos = list(ws.values)
    wb.close()

    if len(datos) < 2:
        return {}

    headers = [str(h).strip().upper() if h else "" for h in datos[0]]
    try:
        col_id   = headers.index("USER_ID")
        col_mz   = headers.index("MZ")
        col_lote = headers.index("LOTE")
    except ValueError as e:
        print(f"  ⚠ usuarios_id.xlsx columna no encontrada: {e}")
        return {}

    mapa = {}
    for fila in datos[1:]:
        if not fila or len(fila) <= max(col_id, col_mz, col_lote):
            continue
        uid  = str(fila[col_id]).strip()         if fila[col_id]   else ""
        mz   = str(fila[col_mz]).strip().upper() if fila[col_mz]   else ""
        lote = limpiar_lote(fila[col_lote])
        if uid and mz and lote:
            mapa[(mz, lote)] = uid

    print(f"  ✔ {len(mapa)} usuarios cargados desde usuarios_id.xlsx")
    return mapa

def cargar_reportes() -> list:
    archivos = sorted(PLANILLA_DIR.glob("*.xlsx"))
    if not archivos:
        raise FileNotFoundError(f"No hay .xlsx en {PLANILLA_DIR}")

    registros = []

    for archivo in archivos:
        nombre = archivo.name
        print(f"\n  Leyendo: {nombre}")

        try:
            wb = load_workbook(archivo, read_only=True, data_only=True)
        except Exception as e:
            print(f"  ⚠ No se pudo abrir: {e}")
            continue

        hojas = [s for s in wb.sheetnames if s.lower() == HOJA_REPORTE.lower()]
        if not hojas:
            print(f"  ⚠ Saltado — sin hoja '{HOJA_REPORTE}'")
            wb.close()
            continue

        ws    = wb[hojas[0]]
        datos = list(ws.values)
        wb.close()

        if len(datos) < 2:
            print(f"  ⚠ Saltado — hoja vacía")
            continue

        n_cols = len(datos[0])
        df     = pd.DataFrame([f[:n_cols] for f in datos[1:]], columns=datos[0])

        mapa = normalizar_columnas(df)
        if not validar_columnas_obligatorias(mapa, nombre):
            continue

        df = df[df[mapa["tipo"]].astype(str).str.strip().str.upper() == TIPO_PAGO.upper()]
        if df.empty:
            print(f"  ⚠ Saltado — sin filas '{TIPO_PAGO}'")
            continue

        print(f"  ✔ {len(df)} filas válidas")

        for _, fila in df.iterrows():
            mz   = str(fila[mapa["mz"]]).strip().upper() if mapa.get("mz")   else ""
            lote = limpiar_lote(fila[mapa["lote"]])       if mapa.get("lote") else ""

            if not mz or mz in ("NONE", "NAN", "") or mz.startswith("="):
                continue
            if not lote:
                continue

            origen = str(fila[mapa["origen"]]).strip() if mapa.get("origen") else ""
            if not origen or origen in ("NONE", "NAN", ""):
                continue

            monto = None
            if mapa.get("monto"):
                try:
                    monto = float(str(fila[mapa["monto"]]).replace(",", "."))
                except:
                    pass

            registros.append({
                "origen":  origen,
                "mz":      mz,
                "lote":    lote,
                "monto":   monto,
                "archivo": nombre,
            })

    print(f"\n  Total registros cargados: {len(registros)}")
    return registros

# ====================VALIDACION=============================
def validar_registros(registros: list) -> list:
    validos = [r for r in registros if r["origen"] and r["mz"] and r["lote"]]
    print(f"  Registros tras validación: {len(validos)}")
    return validos

# ====================PROCESAMIENTO==========================
def construir_maestro(registros: list, usuarios_id: dict) -> list:
    grupos = defaultdict(lambda: {
        "origenes":  set(),
        "archivos":  set(),
        "monto_ref": None,
    })

    for r in registros:
        clave = (r["mz"], r["lote"])
        grupos[clave]["origenes"].add(r["origen"])
        grupos[clave]["archivos"].add(r["archivo"])
        if grupos[clave]["monto_ref"] is None and r["monto"]:
            grupos[clave]["monto_ref"] = r["monto"]

    maestro = []
    sin_id  = 0

    for (mz, lote), data in sorted(grupos.items()):
        meses    = len(data["archivos"])
        nivel    = nivel_confianza(meses)
        origenes = sorted(data["origenes"])
        uid      = usuarios_id.get((mz, lote), "")
        if not uid:
            sin_id += 1

        maestro.append({
            "user_id":            uid,
            "mz":                 mz,
            "lote":               lote,
            "monto_ref":          data["monto_ref"],
            "meses_en_registros": meses,
            "nivel_confianza":    nivel,
            "validado_manual":    "no",
            "fecha_registro":     datetime.today().strftime("%d/%m/%Y"),
            "origenes":           origenes,
        })

    print(f"  Usuarios en maestro : {len(maestro)}")
    print(f"  Con user_id         : {len(maestro) - sin_id}")
    print(f"  Sin user_id (nuevos): {sin_id}")
    return maestro

# ====================EXPORTAR EXCEL========================
def exportar_maestro(maestro: list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Maestro"
    ws.freeze_panes = "A2"

    max_origenes    = max((len(u["origenes"]) for u in maestro), default=1)
    cabeceras_fijas = [
        "user_id", "mz", "lote", "monto_ref",
        "meses_en_registros", "nivel_confianza",
        "validado_manual", "fecha_registro",
    ]
    cabeceras = cabeceras_fijas + [f"origen_{i+1}" for i in range(max_origenes)]

    color_col = {
        "user_id":            COLOR_IDENTIDAD,
        "mz":                 COLOR_UBICACION,
        "lote":               COLOR_UBICACION,
        "monto_ref":          COLOR_UBICACION,
        "meses_en_registros": COLOR_CONFIANZA,
        "nivel_confianza":    COLOR_CONFIANZA,
        "validado_manual":    COLOR_CONFIANZA,
        "fecha_registro":     COLOR_CONFIANZA,
    }

    # Cabecera
    for col_idx, cab in enumerate(cabeceras, start=1):
        c = ws.cell(row=1, column=col_idx, value=cab.upper())
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", start_color=COLOR_CABECERA)
        c.alignment = Alignment(horizontal="center", vertical="center")
        lado = Side(style="thin", color="FFFFFF")
        c.border = Border(left=lado, right=lado, top=lado, bottom=lado)
    ws.row_dimensions[1].height = 22

    # Filas
    for row_idx, usuario in enumerate(maestro, start=2):
        nivel    = usuario["nivel_confianza"]
        bg_nivel = color_nivel(nivel)
        sin_id   = not usuario["user_id"]

        valores = [
            usuario["user_id"],
            usuario["mz"],
            usuario["lote"],
            usuario["monto_ref"],
            usuario["meses_en_registros"],
            nivel,
            usuario["validado_manual"],
            usuario["fecha_registro"],
        ]

        for col_idx, val in enumerate(valores, start=1):
            cab = cabeceras_fijas[col_idx - 1]
            if sin_id:
                bg = COLOR_SIN_ID
            elif cab in ("nivel_confianza", "meses_en_registros", "validado_manual"):
                bg = bg_nivel
            else:
                bg = color_col.get(cab, "FFFFFF")
            estilo_celda(ws.cell(row=row_idx, column=col_idx, value=val), bg_color=bg)

        for o_idx, origen in enumerate(usuario["origenes"]):
            bg = COLOR_SIN_ID if sin_id else COLOR_ORIGENES
            estilo_celda(
                ws.cell(row=row_idx, column=len(cabeceras_fijas) + o_idx + 1, value=origen),
                bg_color=bg
            )

        ws.row_dimensions[row_idx].height = 18

    # Anchos
    anchos = {
        "user_id": 10, "mz": 8, "lote": 8, "monto_ref": 12,
        "meses_en_registros": 10, "nivel_confianza": 16,
        "validado_manual": 14, "fecha_registro": 14,
    }
    for col_idx, cab in enumerate(cabeceras, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = anchos.get(cab, 30)

    wb.save(MAESTRO_PATH)
    print(f"\n  ✔ Maestro guardado: {MAESTRO_PATH}")
    print(f"  ✔ {len(maestro)} filas  |  {max_origenes} origen(es) máx")

# ======================MAIN=================================
def main():
    print("=" * 55)
    print("  MAESTRO YAPE — construcción de maestro de usuarios")
    print("=" * 55)

    print("\n[1] Preparando carpeta de salida...")
    reset_output_folder(OUTPUT_DIR)

    print("\n[2] Cargando usuarios_id...")
    usuarios_id = cargar_usuario_id()

    print("\n[3] Cargando reportes históricos...")
    registros = cargar_reportes()

    print("\n[4] Validando registros...")
    registros = validar_registros(registros)

    print("\n[5] Construyendo maestro...")
    maestro = construir_maestro(registros, usuarios_id)

    print("\n[6] Exportando Excel...")
    exportar_maestro(maestro)

    print("\n" + "=" * 55)
    print("  Proceso completado exitosamente")
    print("=" * 55)

if __name__ == "__main__":
    main()
