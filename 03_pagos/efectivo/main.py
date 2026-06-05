# =========================IMPORTS===========================
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ========================CONFIGURACION======================
BASE_DIR       = Path(__file__).parent
INPUTS_DIR     = BASE_DIR / "inputs"
OUTPUTS_DIR    = BASE_DIR / "outputs"
REGISTRO_MIA   = INPUTS_DIR / "registro_mia.xlsx"
REGISTRO_AMIGA = INPUTS_DIR / "registro_amiga.xlsx"
DISC_FILE      = OUTPUTS_DIR / "discrepancias.xlsx"
OUTPUT_FILE    = OUTPUTS_DIR / "pagos_efectivo.xlsx"

# ========================COLORES============================
C_LOT_H = "E1F5EE"; C_LOT_C = "F0FFF8"; C_LOT_T = "065F46"
C_CAD_H = "FEF9E7"; C_CAD_C = "FEFCE8"; C_CAD_T = "7D6608"
C_COR_H = "FEF3E8"; C_COR_C = "FEF3E8"; C_COR_T = "7C3003"
C_CIC_H = "F4ECF7"; C_CIC_C = "FAF5FF"; C_CIC_T = "5B21B6"
C_SEP   = "F3F4F6"

TIPO_COLORES = {
    "diff_monto":  ("FEF3C7", "92400E"),
    "solo_mia":    ("DBEAFE", "1E40AF"),
    "solo_amiga":  ("FCE7F3", "9D174D"),
}

# ========================ESTILOS============================
def _lado(c="CCCCCC"):
    return Side(style="thin", color=c)

def _cel(cell, bg, txt, mono=False, align="left", bold=False, size=10):
    b = _lado()
    cell.border    = Border(left=b, right=b, top=b, bottom=b)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.font      = Font(name="Courier New" if mono else "Arial",
                          bold=bold, size=size, color=txt)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)

def _hdr(cell, bg, txt, bold=True, size=9, align="center"):
    b = _lado("FFFFFF")
    cell.border    = Border(left=b, right=b, top=b, bottom=b)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.font      = Font(name="Arial", bold=bold, size=size, color=txt)
    cell.fill      = PatternFill("solid", start_color=bg)

def _sep(cell):
    cell.fill   = PatternFill("solid", start_color=C_SEP)
    cell.border = Border(left=_lado(), right=_lado("AAAAAA"),
                         top=_lado(),  bottom=_lado())

def _sep_col(ws, row, ci):
    c = ws.cell(row=row, column=ci, value="")
    _sep(c)
    ws.column_dimensions[get_column_letter(ci)].width = 1

def _escribir_cabecera_doble(ws, cols, grupos):
    ci = 1
    for g in grupos:
        if g is None:
            c = ws.cell(row=1, column=ci, value="")
            _sep(c)
            ws.column_dimensions[get_column_letter(ci)].width = 1
            ci += 1
        else:
            texto, span, bg, txt = g
            if span > 1:
                ws.merge_cells(start_row=1, start_column=ci,
                               end_row=1,   end_column=ci + span - 1)
            c = ws.cell(row=1, column=ci, value=texto)
            _hdr(c, bg, txt, size=9)
            ci += span
    ws.row_dimensions[1].height = 20
    ci = 1
    for col in cols:
        if col == "__SEP__":
            c = ws.cell(row=2, column=ci, value="")
            _sep(c)
            ws.column_dimensions[get_column_letter(ci)].width = 1
        else:
            nombre, bg, txt, mono, align = col
            c = ws.cell(row=2, column=ci, value=nombre)
            _hdr(c, bg, txt, size=9)
        ci += 1
    ws.row_dimensions[2].height = 20

def _aplicar_anchos(ws, cols, anchos):
    ci = 1
    for col in cols:
        if col != "__SEP__":
            ws.column_dimensions[get_column_letter(ci)].width = anchos.get(col[0], 16)
        ci += 1

# ========================UTILIDADES=========================
def normalizar_llave(mz, lt) -> str:
    return str(mz).strip().upper() + "-" + str(lt).strip().upper()

def limpiar_monto(val) -> float:
    try:
        return round(float(str(val).replace(",", ".").strip()), 2)
    except Exception:
        return 0.0

def es_ok(val) -> bool:
    return str(val).strip().lower() in ("sí", "si", "s", "yes", "y", "1")

# ========================CARGAR REGISTROS===================
def cargar_registro(path: Path, etiqueta: str) -> list:
    if not path.exists():
        raise FileNotFoundError(f"No se encontró {etiqueta}: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    datos = list(ws.values)
    wb.close()
    if len(datos) < 2:
        return []
    headers = [str(h).strip().upper() if h else "" for h in datos[0]]
    registros = []
    for fila in datos[1:]:
        if not fila or all(c is None for c in fila):
            continue
        row = dict(zip(headers, fila))
        mz    = str(row.get("MZ")    or "").strip().upper()
        lt    = str(row.get("LT")    or "").strip().upper()
        monto = limpiar_monto(row.get("MONTO", 0))
        fecha = str(row.get("FECHA") or "").strip()
        if not mz or not lt or monto <= 0:
            continue
        registros.append({
            "llave": normalizar_llave(mz, lt),
            "mz":    mz,
            "lt":    lt,
            "monto": monto,
            "fecha": fecha,
        })
    return registros

def cargar_discrepancias() -> list:
    """Lee discrepancias.xlsx (doble cabecera). Fila 1=grupos, fila 2=nombres, datos desde fila 3."""
    if not DISC_FILE.exists():
        return []
    wb = load_workbook(DISC_FILE, read_only=True, data_only=True)
    ws = wb.active
    datos = list(ws.values)
    wb.close()
    if len(datos) < 3:
        return []
    headers = [str(h).strip().upper() if h else "" for h in datos[1]]
    resultado = []
    for fila in datos[2:]:
        if not fila:
            continue
        row = {headers[i]: fila[i] for i in range(min(len(headers), len(fila)))}
        mz = str(row.get("MZ", "")).strip().upper()
        lt = str(row.get("LT", "")).strip().upper()
        if not mz or not lt:
            continue
        resultado.append({
            "tipo":        str(row.get("TIPO", "")).strip(),
            "mz":          mz,
            "lt":          lt,
            "llave":       str(row.get("LLAVE", "")).strip().upper(),
            "monto_mia":   row.get("MONTO_MIA", ""),
            "monto_amiga": row.get("MONTO_AMIGA", ""),
            "monto_final": limpiar_monto(row.get("MONTO_FINAL", 0)),
            "ok":          str(row.get("OK", "")).strip(),
            "fecha":       str(row.get("FECHA", "")).strip(),
            "ciclo":       int(row.get("CICLO") or 1),
        })
    return resultado

def cargar_pagos_existentes() -> list:
    """Lee pagos_efectivo.xlsx para acumular registros de ciclos anteriores."""
    if not OUTPUT_FILE.exists():
        return []
    wb = load_workbook(OUTPUT_FILE, read_only=True, data_only=True)
    ws = wb.active
    datos = list(ws.values)
    wb.close()
    if len(datos) < 3:
        return []
    headers = [str(h).strip().upper() if h else "" for h in datos[1]]
    registros = []
    for fila in datos[2:]:
        if not fila:
            continue
        row = {headers[i]: fila[i] for i in range(min(len(headers), len(fila)))}
        mz = str(row.get("MZ", "")).strip().upper()
        lt = str(row.get("LT", "")).strip().upper()
        if not mz or not lt:
            continue
        registros.append({
            "mz":    mz,
            "lt":    lt,
            "llave": str(row.get("LLAVE", normalizar_llave(mz, lt))).strip().upper(),
            "monto": limpiar_monto(row.get("MONTO", 0)),
            "fecha": str(row.get("FECHA", "")).strip(),
            "ciclo": int(row.get("CICLO") or 1),
        })
    return registros

# ========================COMPARAR (CICLO 1)=================
def comparar_fresco(mia: list, amiga: list) -> tuple:
    """
    Ciclo 1: compara los dos registros desde cero.
    Retorna (confirmados, pendientes).
    """
    mia_idx   = {r["llave"]: r for r in mia}
    amiga_idx = {r["llave"]: r for r in amiga}
    todas     = sorted(set(mia_idx) | set(amiga_idx))

    confirmados = []
    pendientes  = []

    for llave in todas:
        en_mia   = llave in mia_idx
        en_amiga = llave in amiga_idx

        if en_mia and en_amiga:
            rm = mia_idx[llave]
            ra = amiga_idx[llave]
            if abs(rm["monto"] - ra["monto"]) < 0.01:
                confirmados.append({
                    "mz": rm["mz"], "lt": rm["lt"], "llave": llave,
                    "monto": rm["monto"], "fecha": rm["fecha"], "ciclo": 1,
                })
            else:
                pendientes.append({
                    "tipo": "diff_monto", "mz": rm["mz"], "lt": rm["lt"], "llave": llave,
                    "monto_mia": rm["monto"], "monto_amiga": ra["monto"],
                    "monto_final": ra["monto"],  # amiga es referencia
                    "ok": "", "fecha": rm["fecha"], "ciclo": 1,
                })
        elif en_mia:
            rm = mia_idx[llave]
            pendientes.append({
                "tipo": "solo_mia", "mz": rm["mz"], "lt": rm["lt"], "llave": llave,
                "monto_mia": rm["monto"], "monto_amiga": "",
                "monto_final": rm["monto"],
                "ok": "", "fecha": rm["fecha"], "ciclo": 1,
            })
        else:
            ra = amiga_idx[llave]
            pendientes.append({
                "tipo": "solo_amiga", "mz": ra["mz"], "lt": ra["lt"], "llave": llave,
                "monto_mia": "", "monto_amiga": ra["monto"],
                "monto_final": ra["monto"],
                "ok": "", "fecha": ra["fecha"], "ciclo": 1,
            })

    return confirmados, pendientes

# ========================EXPORTAR===========================
def exportar_discrepancias_xlsx(pendientes: list):
    cols = [
        ("TIPO",        "FEF9E7", "7D6608", False, "center"),
        ("MZ",          C_LOT_C,  C_LOT_T,  True,  "center"),
        ("LT",          C_LOT_C,  C_LOT_T,  True,  "center"),
        ("LLAVE",       C_LOT_C,  C_LOT_T,  True,  "center"),
        "__SEP__",
        ("MONTO_MIA",   C_CAD_C,  C_CAD_T,  True,  "right"),
        ("MONTO_AMIGA", C_CAD_C,  C_CAD_T,  True,  "right"),
        "__SEP__",
        ("MONTO_FINAL", C_COR_C,  C_COR_T,  True,  "right"),
        ("OK",          C_COR_C,  C_COR_T,  False, "center"),
        "__SEP__",
        ("FECHA",       C_CIC_C,  C_CIC_T,  False, "left"),
        ("CICLO",       C_CIC_C,  C_CIC_T,  True,  "center"),
    ]
    grupos = [
        ("¿CUÁL ES EL LOTE?",        4, C_LOT_H, C_LOT_T),
        None,
        ("¿QUÉ DICE CADA CUADERNO?", 2, C_CAD_H, C_CAD_T),
        None,
        ("¿CUÁL ES EL CORRECTO?",    2, C_COR_H, C_COR_T),
        None,
        ("¿CUÁNDO?",                 2, C_CIC_H, C_CIC_T),
    ]
    anchos = {
        "TIPO": 12, "MZ": 6, "LT": 7, "LLAVE": 12,
        "MONTO_MIA": 13, "MONTO_AMIGA": 14,
        "MONTO_FINAL": 14, "OK": 6,
        "FECHA": 16, "CICLO": 7,
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "discrepancias"
    ws.freeze_panes = "A3"
    _escribir_cabecera_doble(ws, cols, grupos)
    for ri, reg in enumerate(pendientes, start=3):
        t_bg, t_txt = TIPO_COLORES.get(reg["tipo"], ("F3F4F6", "374151"))
        valores = {
            "TIPO":        reg["tipo"],
            "MZ":          reg["mz"],
            "LT":          reg["lt"],
            "LLAVE":       reg["llave"],
            "MONTO_MIA":   reg["monto_mia"],
            "MONTO_AMIGA": reg["monto_amiga"],
            "MONTO_FINAL": reg["monto_final"],
            "OK":          reg["ok"],
            "FECHA":       reg["fecha"],
            "CICLO":       reg["ciclo"],
        }
        ci = 1
        for col in cols:
            if col == "__SEP__":
                _sep_col(ws, ri, ci); ci += 1; continue
            nombre, bg, txt, mono, align = col
            if nombre == "TIPO":
                bg, txt = t_bg, t_txt
            val = valores.get(nombre, "")
            _cel(ws.cell(row=ri, column=ci, value=val if val != "" else ""),
                 bg, txt, mono=mono, align=align)
            ci += 1
        ws.row_dimensions[ri].height = 18
    _aplicar_anchos(ws, cols, anchos)
    OUTPUTS_DIR.mkdir(exist_ok=True)
    wb.save(DISC_FILE)

def exportar_pagos_efectivo_xlsx(confirmados: list):
    cols = [
        ("MZ",    C_LOT_C, C_LOT_T, True,  "center"),
        ("LT",    C_LOT_C, C_LOT_T, True,  "center"),
        ("LLAVE", C_LOT_C, C_LOT_T, True,  "center"),
        "__SEP__",
        ("MONTO", C_CAD_C, C_CAD_T, True,  "right"),
        "__SEP__",
        ("FECHA", C_COR_C, C_COR_T, False, "left"),
        ("CICLO", C_COR_C, C_COR_T, True,  "center"),
    ]
    grupos = [
        ("¿DÓNDE VIVE?",  3, C_LOT_H, C_LOT_T),
        None,
        ("¿CUÁNTO PAGÓ?", 1, C_CAD_H, C_CAD_T),
        None,
        ("¿CUÁNDO?",      2, C_COR_H, C_COR_T),
    ]
    anchos = {
        "MZ": 6, "LT": 7, "LLAVE": 12,
        "MONTO": 12,
        "FECHA": 16, "CICLO": 7,
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "pagos_efectivo"
    ws.freeze_panes = "A3"
    _escribir_cabecera_doble(ws, cols, grupos)
    for ri, reg in enumerate(confirmados, start=3):
        valores = {
            "MZ":    reg["mz"],
            "LT":    reg["lt"],
            "LLAVE": reg["llave"],
            "MONTO": reg["monto"],
            "FECHA": reg["fecha"],
            "CICLO": reg["ciclo"],
        }
        ci = 1
        for col in cols:
            if col == "__SEP__":
                _sep_col(ws, ri, ci); ci += 1; continue
            nombre, bg, txt, mono, align = col
            val = valores.get(nombre, "")
            _cel(ws.cell(row=ri, column=ci, value=val if val != "" else ""),
                 bg, txt, mono=mono, align=align)
            ci += 1
        ws.row_dimensions[ri].height = 18
    _aplicar_anchos(ws, cols, anchos)
    OUTPUTS_DIR.mkdir(exist_ok=True)
    wb.save(OUTPUT_FILE)

# ========================MAIN===============================
def main():
    print("\n══════════════════════════════════════")
    print("  03_pagos / efectivo")
    print("══════════════════════════════════════")

    # ── Ciclo 1: comparación fresca ─────────────────────────
    if not DISC_FILE.exists():
        print("\n  Ciclo 1 — comparación inicial")
        print("\n── Cargando registros ──")
        mia   = cargar_registro(REGISTRO_MIA,   "registro_mia")
        amiga = cargar_registro(REGISTRO_AMIGA, "registro_amiga")
        print(f"  registro_mia:   {len(mia)} cobros")
        print(f"  registro_amiga: {len(amiga)} cobros")

        print("\n── Comparando ──")
        confirmados, pendientes = comparar_fresco(mia, amiga)
        print(f"  ✔ coinciden:      {len(confirmados)}")
        if pendientes:
            por_tipo = {}
            for p in pendientes:
                por_tipo[p["tipo"]] = por_tipo.get(p["tipo"], 0) + 1
            for t, n in por_tipo.items():
                print(f"  ⚠ {t}: {n}")

        print("\n── Exportando ──")
        if confirmados:
            exportar_pagos_efectivo_xlsx(confirmados)
            print(f"  → pagos_efectivo.xlsx ({len(confirmados)} cobros confirmados en ciclo 1)")
        if pendientes:
            exportar_discrepancias_xlsx(pendientes)
            print(f"  → discrepancias.xlsx  ({len(pendientes)} pendientes)")
            print(f"\n  ⚠ Abre discrepancias.xlsx · revisa MONTO_FINAL · escribe Sí en OK · vuelve a ejecutar.")
        elif not confirmados:
            print("  Sin cobros registrados.")
        else:
            print(f"\n  ✔ Sin discrepancias. pagos_efectivo.xlsx listo → entrega a 04_cobranza.")

    # ── Ciclo 2+: procesar discrepancias resueltas ──────────
    else:
        disc = cargar_discrepancias()
        ciclo_actual = (max(r["ciclo"] for r in disc) + 1) if disc else 2
        print(f"\n  Ciclo {ciclo_actual} — procesando discrepancias")

        resueltas  = [r for r in disc if es_ok(r["ok"])]
        aun_pend   = [r for r in disc if not es_ok(r["ok"])]
        print(f"\n── Discrepancias ──")
        print(f"  resueltas (OK=Sí): {len(resueltas)}")
        print(f"  pendientes:        {len(aun_pend)}")

        print("\n── Exportando ──")
        # Construir lista completa: registros anteriores + nuevos resueltos
        existentes = cargar_pagos_existentes()
        llaves_existentes = {r["llave"] for r in existentes}

        nuevos_confirmados = []
        for r in resueltas:
            if r["llave"] in llaves_existentes:
                continue
            if r["monto_final"] <= 0:
                continue
            nuevos_confirmados.append({
                "mz":    r["mz"],
                "lt":    r["lt"],
                "llave": r["llave"],
                "monto": r["monto_final"],
                "fecha": r["fecha"],
                "ciclo": ciclo_actual,
            })

        todos_confirmados = existentes + nuevos_confirmados
        if todos_confirmados:
            exportar_pagos_efectivo_xlsx(todos_confirmados)
            print(f"  → pagos_efectivo.xlsx ({len(todos_confirmados)} cobros · +{len(nuevos_confirmados)} nuevos)")

        if aun_pend:
            exportar_discrepancias_xlsx(aun_pend)
            print(f"  → discrepancias.xlsx  ({len(aun_pend)} pendientes)")
            print(f"\n  ⚠ Aún quedan discrepancias. Completa las filas restantes y vuelve a ejecutar.")
        else:
            DISC_FILE.unlink()
            print(f"  ✔ discrepancias.xlsx eliminado — todo resuelto")
            print(f"\n  ✔ pagos_efectivo.xlsx listo → entrega a 04_cobranza.")

    print("\n══════════════════════════════════════\n")

if __name__ == "__main__":
    main()
