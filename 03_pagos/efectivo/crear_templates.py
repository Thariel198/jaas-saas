# Script de uso único — genera los templates vacíos de registro_mia y registro_amiga.
# Ejecutar una sola vez: python crear_templates.py
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INPUTS_DIR = Path(__file__).parent / "inputs"
INPUTS_DIR.mkdir(exist_ok=True)

C_H = "E1F5EE"; C_T = "085041"

def lado():
    return Side(style="thin", color="FFFFFF")

def hdr(cell, texto):
    cell.value     = texto
    cell.font      = Font(name="Arial", bold=True, size=10, color=C_T)
    cell.fill      = PatternFill("solid", start_color=C_H)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    b = lado()
    cell.border    = Border(left=b, right=b, top=b, bottom=b)

def ejemplo(cell, texto):
    cell.value     = texto
    cell.font      = Font(name="Arial", size=10, color="9CA3AF", italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def crear(nombre):
    wb = Workbook()
    ws = wb.active
    ws.title = nombre.replace(".xlsx", "")
    ws.row_dimensions[1].height = 22
    for i, col in enumerate(["MZ", "LT", "MONTO", "FECHA"], start=1):
        hdr(ws.cell(row=1, column=i), col)
    anchos = {"A": 8, "B": 8, "C": 12, "D": 16}
    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = ancho
    # fila de ejemplo en gris claro para guiar
    for texto, col in [("A", 1), ("8C", 2), ("38.00", 3), ("03/05/2026", 4)]:
        ejemplo(ws.cell(row=2, column=col), texto)
    wb.save(INPUTS_DIR / nombre)
    print(f"  ✔ {nombre}")

print("\nCreando templates de entrada...")
crear("registro_mia.xlsx")
crear("registro_amiga.xlsx")
print("Listo. Abre los archivos, borra la fila de ejemplo y llena tus cobros.\n")
