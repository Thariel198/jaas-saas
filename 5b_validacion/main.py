import json
import logging
import sys
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
ROOT      = Path(__file__).parent
OUT_DIR   = ROOT / "outputs"

sys.path.insert(0, str(ROOT.parent / "shared"))
import utils_estado_ciclo as repo_estado  # noqa: E402
MOD04      = ROOT.parent / "5_cobranza"
MOD04_EFEC = ROOT.parent / "4_pagos" / "efectivo" / "outputs"
MOD04_YAPE = ROOT.parent / "4_pagos" / "yape" / "motor_matching" / "outputs"
MOD04_PAGOS = ROOT.parent / "4_pagos" / "outputs"
SHARED     = ROOT.parent / "shared"

CRUDO_DIR         = SHARED / "reporte_mes_crudo"
ESTADO_CICLO_PATH = SHARED / "reporte_acumulado_procesado" / "estado_ciclo.json"
# Los outputs de pagos llevan el ciclo en el nombre desde 2026-08; para los
# ciclos anteriores se acepta el nombre pelado (ver shared/ciclo.py).
import ciclo as ciclo_activo  # noqa: E402

_MES_CICLO = ciclo_activo.activo(default=None, path=SHARED / "ciclo_activo.json")


def _pago_path(carpeta, base: str):
    if _MES_CICLO is None:
        return carpeta / f"{base}.xlsx"
    return ciclo_activo.resolver(carpeta, base, _MES_CICLO,
                                 legacy_sin_periodo=ciclo_activo.acepta_legacy(_MES_CICLO))


YAPE_PATH        = _pago_path(MOD04_YAPE, "pagos_yape_tepago")
YAPE_RET_PATH    = _pago_path(MOD04_YAPE, "pagos_yape_retorno")
YAPE_DEV_PATH    = _pago_path(MOD04_YAPE, "pagos_yape_devolucion")
TANQUE_PATH      = MOD04_PAGOS / "aportes_tanque.xlsx"  # canal-agnóstico (DE10) — reemplaza pagos_yape_tanque.xlsx
BLANCOS_MES_PATH = MOD04_YAPE / "blancos_mes.xlsx"
PAGASTE_PATH     = _pago_path(MOD04_YAPE, "pagos_yape_pagaste")
EFEC_PATH        = _pago_path(MOD04_EFEC, "pagos_efectivo")
COB_PATH         = MOD04 / "outputs" / "planilla_cobrado.xlsx"

TOLERANCIA = 0.005  # diferencia máxima considerada OK

# ── PALETA ────────────────────────────────────────────────────────────────────
GH_PERIODO  = ("F3F4F6", "374151")
GH_QUIEN    = ("F4ECF7", "5B21B6")
GH_REPORTE  = ("E6F1FB", "0C447C")
GH_PLANILLA = ("FEF9E7", "7D6608")
GH_CUADRA   = ("E1F5EE", "085041")
GH_PORQUE   = ("FEF2F2", "991B1B")

TD_PERIODO  = "F9FAFB"
TD_QUIEN    = "FAF5FF"
TD_REPORTE  = "F0F8FF"
TD_PLANILLA = "FEFCE8"
TD_OK       = "E1F5EE"
TD_ERR      = "FEF2F2"

# ── LOGGING ───────────────────────────────────────────────────────────────────
def _init_logging():
    OUT_DIR.mkdir(exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)s  %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(OUT_DIR / "run.log", encoding="utf-8"),
        ],
    )

log = logging.getLogger(__name__)

# ── ESTILO ────────────────────────────────────────────────────────────────────
def _borde():
    b = Side(style="thin", color="CCCCCC")
    return Border(left=b, right=b, top=b, bottom=b)

def _c(ws, row, col, value=None, bg=None, txt="333333",
       bold=False, align="left", mono=False, size=9):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Consolas" if mono else "Arial",
                       size=size, bold=bold, color=txt)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _borde()
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    return c

def _gh(ws, row, cs, ce, texto, bg, txt):
    ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
    c = ws.cell(row=row, column=cs, value=texto)
    c.font      = Font(name="Arial", size=8, bold=True, color=txt)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", start_color=bg)
    c.border    = _borde()

def _ch(ws, row, col, texto, bg, txt):
    _c(ws, row, col, texto, bg=bg, txt=txt, bold=True, align="center")

def _sep(ws, col, row_end):
    ws.column_dimensions[get_column_letter(col)].width = 0.8
    b_sep = Border(left=Side(style="thin", color="D1D5DB"),
                   right=Side(style="thin", color="D1D5DB"))
    for r in range(1, row_end + 1):
        c = ws.cell(row=r, column=col)
        c.fill   = PatternFill("solid", start_color="F3F4F6")
        c.border = b_sep

def _w(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

# ── UTILIDADES ────────────────────────────────────────────────────────────────
def _float(val) -> float:
    try:
        s = str(val).replace(",", ".").strip()
        if s.lower() in ("nan", "none", ""):
            return 0.0
        return float(s)
    except (ValueError, TypeError):
        return 0.0

def _norm_mz(val) -> str:
    return str(val).strip().upper()

def _norm_cols(df) -> list:
    return [str(c).strip().upper() if str(c) not in ("NAN", "NONE") else f"_C{i}"
            for i, c in enumerate(df.columns)]

def _norm_tipo(val) -> str:
    """Normaliza tipo de transacción quitando tildes para comparar."""
    return (str(val).strip().upper()
            .replace("Ó", "O").replace("É", "E")
            .replace("Á", "A").replace("Í", "I").replace("Ú", "U"))

# ── ANCLA DE CORTE ────────────────────────────────────────────────────────────
def _obtener_ancla() -> pd.Timestamp | None:
    """Max fecha del reporte_acumulado_procesado — mismo criterio que motor_matching."""
    proc_dir = SHARED / "reporte_acumulado_procesado"
    archivos = sorted(p for p in proc_dir.glob("*_procesado.xlsx")
                      if not p.name.startswith("~$"))
    if not archivos:
        return None
    df = pd.read_excel(archivos[-1], header=0, dtype=str)
    col_f = next((c for c in df.columns if "fecha" in str(c).lower()), None)
    if col_f is None:
        log.warning("No se pudo determinar ancla de corte")
        return None
    fechas = pd.to_datetime(df[col_f], errors="coerce", dayfirst=True).dropna()
    if fechas.empty:
        log.warning("No se pudo determinar ancla de corte")
        return None
    ancla = fechas.max()
    log.info(f"Ancla de corte: {ancla}")
    return ancla

# ── VALIDAR INPUTS ────────────────────────────────────────────────────────────
def _validar_inputs():
    errores = []
    crudos = list(CRUDO_DIR.glob("*.xlsx")) if CRUDO_DIR.exists() else []
    if not crudos:
        errores.append(f"Falta reporte crudo en: {CRUDO_DIR}")
    for ruta in [YAPE_PATH, EFEC_PATH, COB_PATH]:
        if not ruta.exists():
            errores.append(f"Falta: {ruta}")
    if errores:
        for e in errores:
            log.error(e)
        raise FileNotFoundError("Inputs faltantes — ver errores arriba")
    log.info("Inputs validados")

# ── CARGA: PAGOS YAPE TEPAGO ──────────────────────────────────────────────────
def _cargar_yape_tepago() -> tuple[dict, dict, dict, float, tuple]:
    """Retorna ({mz: monto_neto}, {(mz,lote): monto}, {(mz,lote): [ORIGEN_MONTO_FECHA]},
    total, (fecha_min, fecha_max)).
    Suma TE PAGÓ y resta retornos por MZ para que el reporte refleje
    lo mismo que planilla_cobrado (que también descuenta retornos)."""
    df = pd.read_excel(YAPE_PATH, header=1, dtype=str)
    df.columns = _norm_cols(df)

    totales  = {}
    por_lote = {}
    detalle  = {}
    fechas   = []
    for _, f in df.iterrows():
        tipo = _norm_tipo(f.get("TIPO", ""))
        if tipo != "TE PAGO":
            continue
        mz   = _norm_mz(f.get("MZ", ""))
        lote = str(f.get("LOTE", "")).strip().upper()
        conc = str(f.get("CONCEPTO", "")).strip().upper()
        if not mz or mz == "NAN" or mz == "BLANCO" or not lote or lote == "NAN":
            continue
        if conc and conc not in ("NAN", "NONE", ""):
            continue
        monto = _float(f.get("MONTO_PAGO", 0))
        totales[mz] = totales.get(mz, 0.0) + monto
        key = (mz, lote)
        por_lote[key] = por_lote.get(key, 0.0) + monto
        origen    = str(f.get("ORIGEN", "")).strip()
        fecha_raw = str(f.get("FECHA", "")).strip()
        detalle.setdefault(key, []).append(f"{origen}_{monto:.2f}_{fecha_raw}")
        if fecha_raw and fecha_raw not in ("NAN", ""):
            try:
                fechas.append(pd.to_datetime(fecha_raw, dayfirst=True))
            except Exception:
                pass

    # Restar retornos Yape por MZ
    if YAPE_RET_PATH.exists():
        ret = pd.read_excel(YAPE_RET_PATH, header=1, dtype=str)
        ret.columns = _norm_cols(ret)
        total_ret = 0.0
        for _, f in ret.iterrows():
            mz = _norm_mz(f.get("MZ", ""))
            if not mz or mz == "NAN":
                continue
            monto = _float(f.get("MONTO", 0))
            totales[mz] = totales.get(mz, 0.0) - monto
            total_ret += monto
        if total_ret:
            log.info(f"Yape retornos descontados: S/ {total_ret:.2f}")

    periodo = (min(fechas), max(fechas)) if fechas else (None, None)
    total   = round(sum(totales.values()), 2)
    log.info(f"Yape tepago: {len(totales)} MZ · S/ {total:.2f} "
             f"· periodo {periodo[0]} → {periodo[1]}")
    return totales, por_lote, detalle, total, periodo


# ── CARGA: CRUDO BANCO — ambos TIPO ──────────────────────────────────────────
def _cargar_crudo_por_tipo(periodo: tuple, periodo_pagaste: tuple = None) -> tuple[float, float]:
    """Retorna (total_te_pago, total_pagaste) del reporte crudo.
    TE PAGÓ y PAGASTE se mueven en ventanas de fecha independientes — cada
    tipo se filtra con su propio periodo, no con el del otro."""
    fecha_min, fecha_max = periodo
    fecha_min_pg, fecha_max_pg = periodo_pagaste if periodo_pagaste else (None, None)
    crudo_file = sorted(p for p in CRUDO_DIR.glob("*.xlsx") if not p.name.startswith("~$"))[-1]
    df = pd.read_excel(crudo_file, header=4, dtype=str)
    df.columns = [
        str(c).strip().upper()
        .replace("Ó","O").replace("É","E").replace("Á","A")
        .replace("Ú","U").replace("Ñ","N").replace(" ","_")
        for c in df.columns
    ]
    col_tipo  = next((c for c in df.columns if "TIPO" in c), None)
    col_monto = next((c for c in df.columns if "MONTO" in c), None)
    col_fecha = next((c for c in df.columns if "FECHA" in c), None)
    if not col_tipo or not col_monto:
        raise ValueError(f"Columnas TIPO/MONTO no encontradas en crudo: {df.columns.tolist()}")

    te_pago = 0.0; pagaste = 0.0
    for _, f in df.iterrows():
        tipo = _norm_tipo(f.get(col_tipo, ""))
        if tipo not in ("TE PAGO", "PAGASTE"):
            continue
        f_min, f_max = (fecha_min, fecha_max) if tipo == "TE PAGO" else (fecha_min_pg, fecha_max_pg)
        if col_fecha and f_min and f_max:
            try:
                fecha = pd.to_datetime(str(f[col_fecha]), dayfirst=True)
                if not (f_min <= fecha <= f_max):
                    continue
            except Exception:
                pass
        monto = _float(f.get(col_monto, 0))
        if tipo == "TE PAGO":
            te_pago += monto
        else:
            pagaste += monto

    te_pago = round(te_pago, 2); pagaste = round(pagaste, 2)
    log.info(f"Crudo banco · TE PAGÓ: S/ {te_pago:.2f} · PAGASTE: S/ {pagaste:.2f}")
    return te_pago, pagaste


# ── CARGA: BLANCOS TOTAL (ambas fuentes) ─────────────────────────────────────
def _cargar_blancos_total() -> float:
    """Lee blancos_mes.MONTO — fuente canónica generada por motor_matching."""
    total = 0.0
    if BLANCOS_MES_PATH.exists():
        df = pd.read_excel(BLANCOS_MES_PATH, header=1, dtype=str)
        df.columns = _norm_cols(df)
        total += df["MONTO"].apply(_float).sum() if "MONTO" in df.columns else 0.0
    total = round(total, 2)
    log.info(f"Blancos total: S/ {total:.2f}")
    return total


# ── CARGA: DEVUELTOS TOTAL ────────────────────────────────────────────────────
def _cargar_devueltos_total() -> float:
    if not YAPE_DEV_PATH.exists():
        return 0.0
    df = pd.read_excel(YAPE_DEV_PATH, header=1, dtype=str)
    df.columns = _norm_cols(df)
    total = round(df["MONTO"].apply(_float).sum() if "MONTO" in df.columns else 0.0, 2)
    log.info(f"Devueltos total: S/ {total:.2f}")
    return total


def _netear_devoluciones(yape_mz: dict) -> dict:
    """Resta devoluciones Yape por MZ — SOLO para comparar contra planilla_cobrado
    (Nivel 2), que ya las netea. NO usar para Nivel 1a: ahí agua_total debe
    quedar crudo, igual que crudo_tepago (el pago original es legítimo en TE PAGÓ
    del banco, la devolución es un evento PAGASTE aparte con su propia
    reconciliación en Nivel 1b)."""
    if not YAPE_DEV_PATH.exists():
        return dict(yape_mz)
    dev = pd.read_excel(YAPE_DEV_PATH, header=1, dtype=str)
    dev.columns = _norm_cols(dev)
    neto = dict(yape_mz)
    total_dev = 0.0
    for _, f in dev.iterrows():
        mz = _norm_mz(f.get("MZ", ""))
        if not mz or mz == "NAN":
            continue
        monto = _float(f.get("MONTO", 0))
        neto[mz] = neto.get(mz, 0.0) - monto
        total_dev += monto
    if total_dev:
        log.info(f"Yape devoluciones descontadas (solo Nivel 2): S/ {total_dev:.2f}")
    return neto


# ── CARGA: TANQUE TOTAL ───────────────────────────────────────────────────────
def _cargar_tanque_total() -> float:
    """Tanque de Yape únicamente — alimenta Nivel 1a, que reconcilia contra el
    crudo bancario de TE PAGÓ (solo Yape). aportes_tanque.xlsx es canal-agnóstico
    (yape + efectivo, ver DE10) — filtrar CANAL='yape' aquí para no mezclar
    dinero de efectivo en una reconciliación que es específica del banco."""
    if not TANQUE_PATH.exists():
        return 0.0
    df = pd.read_excel(TANQUE_PATH, header=1, dtype=str)
    df.columns = _norm_cols(df)
    if "CANAL" in df.columns:
        df = df[df["CANAL"].str.strip().str.lower() == "yape"]
    # MZ=="BLANCO" ya se cuenta vía blancos_tot (blancos_mes.xlsx) — excluir aquí
    # para no duplicarlo.
    if "MZ" in df.columns:
        df = df[df["MZ"].apply(_norm_mz) != "BLANCO"]
    col = next((c for c in df.columns if "MONTO" in c), None)
    total = round(df[col].apply(_float).sum() if col else 0.0, 2)
    log.info(f"Tanque total (yape): S/ {total:.2f}")
    return total


def _cargar_otros_conceptos() -> float:
    """Balde genérico para Nivel 1a: Σ TE PAGÓ Yape con CONCEPTO no vacío que NO
    sea tanque (ese ya lo cuenta _cargar_tanque_total). Captura deuda_directiva,
    y cualquier concepto libre futuro, sin parchar la fórmula cada vez. Sin este
    balde, esa plata está en el crudo del banco pero en ningún componente → falso
    descuadre 1a (caso "saldo Ronel", -62)."""
    df = pd.read_excel(YAPE_PATH, header=1, dtype=str)
    df.columns = _norm_cols(df)
    total = 0.0
    for _, f in df.iterrows():
        if _norm_tipo(f.get("TIPO", "")) != "TE PAGO":
            continue
        conc = str(f.get("CONCEPTO", "")).strip().lower()
        if conc in ("", "nan", "none", "tanque"):
            continue
        total += _float(f.get("MONTO_PAGO", 0))
    total = round(total, 2)
    log.info(f"Otros conceptos (yape, no-tanque): S/ {total:.2f}")
    return total


# ── CARGA: PAGASTE PROCESADO ──────────────────────────────────────────────────
def _cargar_pagaste_total() -> tuple[float, tuple]:
    """Retorna (total, (fecha_min, fecha_max)) — PAGASTE tiene su propia ventana
    de fechas, independiente de TE PAGÓ (no se mueven en el mismo rango)."""
    if not PAGASTE_PATH.exists():
        return 0.0, (None, None)
    df = pd.read_excel(PAGASTE_PATH, header=1, dtype=str)
    df.columns = _norm_cols(df)
    total = round(df["MONTO"].apply(_float).sum() if "MONTO" in df.columns else 0.0, 2)
    log.info(f"Pagaste procesado total: S/ {total:.2f}")
    fechas = pd.to_datetime(df["FECHA"], dayfirst=True, errors="coerce").dropna() if "FECHA" in df.columns else pd.Series([], dtype="datetime64[ns]")
    periodo = (fechas.min(), fechas.max()) if len(fechas) else (None, None)
    return total, periodo


# ── CARGA: PAGOS EFECTIVO ─────────────────────────────────────────────────────
CORR_PATH = MOD04 / "inputs" / "correcciones_lote.xlsx"

def _cargar_efectivo() -> tuple[dict, dict, dict, float]:
    """Retorna ({mz: monto}, {(mz,lt): monto}, {(mz,lt): [MESA_COBRADOR_LT]}, total)
    desde pagos_efectivo.xlsx.
    Excluye MZ=BLANCO, excluye CONCEPTO no vacío (tanque/honorario/gasto/etc. —
    igual que _cargar_yape_tepago, ver E1) y aplica correcciones de lote
    cross-MZ para que el reporte refleje el mismo lote destino que usa
    planilla_cobrado."""
    df = pd.read_excel(EFEC_PATH, header=1, dtype=str)
    df.columns = _norm_cols(df)

    # totales por (mz, lt) para poder aplicar correcciones por lote
    por_lote = {}
    detalle  = {}
    blancos  = 0.0
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ", ""))
        lt = str(f.get("LT", "")).strip().upper()
        conc = str(f.get("CONCEPTO", "")).strip().upper()
        if not mz or mz == "NAN":
            continue
        if conc and conc not in ("NAN", "NONE", ""):
            continue
        if mz == "BLANCO":
            blancos += _float(f.get("MONTO", 0))
            continue
        key = (mz, lt)
        por_lote[key] = por_lote.get(key, 0.0) + _float(f.get("MONTO", 0))
        mesa      = str(f.get("MESA", "")).strip()
        cobrador  = str(f.get("COBRADOR", "")).strip()
        detalle.setdefault(key, []).append(f"{mesa}_{cobrador}_LT{lt}")

    # Aplicar correcciones cross-MZ desde correcciones_lote.xlsx
    if CORR_PATH.exists():
        corr = pd.read_excel(CORR_PATH, dtype=str)
        for _, c in corr.iterrows():
            mz_o = _norm_mz(c.get("MZ_ORIGEN",  ""))
            lt_o = str(c.get("LT_ORIGEN",  "")).strip().upper()
            mz_d = _norm_mz(c.get("MZ_DESTINO", ""))
            lt_d = str(c.get("LT_DESTINO", "")).strip().upper()
            if mz_o == mz_d:
                continue  # misma MZ — no afecta totales por MZ
            key_o = (mz_o, lt_o)
            if key_o not in por_lote:
                continue
            monto = por_lote.pop(key_o)
            key_d = (mz_d, lt_d)
            por_lote[key_d] = por_lote.get(key_d, 0.0) + monto
            refs = detalle.pop(key_o, [])
            detalle.setdefault(key_d, []).extend(refs)
            log.info(f"Corrección aplicada en reporte: {mz_o}-{lt_o} → {mz_d}-{lt_d} S/ {monto:.2f}")

    # Agregar por MZ
    totales = {}
    for (mz, _), monto in por_lote.items():
        totales[mz] = totales.get(mz, 0.0) + monto

    total = round(sum(totales.values()), 2)
    log.info(f"Efectivo: {len(totales)} MZ · S/ {total:.2f}")
    if blancos:
        log.info(f"Efectivo BLANCO excluido (sin dueño): S/ {blancos:.2f}")
    return totales, por_lote, detalle, total

# ── CARGA: COBRANZA FINAL ─────────────────────────────────────────────────────
def _cargar_cobranza() -> tuple[dict, dict, dict, dict, float, float]:
    """Retorna ({mz: yape}, {mz: efectivo}, {(mz,lt): yape}, {(mz,lt): efectivo},
    total_yape, total_efectivo)."""
    df = pd.read_excel(COB_PATH, header=1, dtype=str)
    df.columns = _norm_cols(df)
    yape_mz, efec_mz = {}, {}
    yape_lote, efec_lote = {}, {}
    for _, f in df.iterrows():
        mz = _norm_mz(f.get("MZ", ""))
        if not mz or mz == "NAN":
            continue
        lt = str(f.get("LT", "")).strip().upper()
        y = _float(f.get("MONTO_YAPE",     0))
        e = _float(f.get("MONTO_EFECTIVO", 0))
        yape_mz[mz] = yape_mz.get(mz, 0.0) + y
        efec_mz[mz] = efec_mz.get(mz, 0.0) + e
        key = (mz, lt)
        yape_lote[key] = yape_lote.get(key, 0.0) + y
        efec_lote[key] = efec_lote.get(key, 0.0) + e
    t_yape = round(sum(yape_mz.values()), 2)
    t_efec = round(sum(efec_mz.values()), 2)
    log.info(f"Cobranza → YAPE: S/ {t_yape:.2f} · EFECTIVO: S/ {t_efec:.2f}")
    return yape_mz, efec_mz, yape_lote, efec_lote, t_yape, t_efec

# ── CALCULAR DIFERENCIAS ─────────────────────────────────────────────────────
def _referencia_por_mz(reporte_lote: dict, planilla_lote: dict, detalle: dict) -> dict:
    """Para cada MZ, cuenta la historia de POR QUÉ no cuadra: lista las
    referencias (detalle crudo) de los lotes cuyo monto difiere entre
    reporte y planilla. MZs sin diferencia por lote no aparecen."""
    keys = set(reporte_lote) | set(planilla_lote)
    por_mz: dict = {}
    for (mz, lt) in sorted(keys):
        r = round(reporte_lote.get((mz, lt), 0.0), 2)
        p = round(planilla_lote.get((mz, lt), 0.0), 2)
        if abs(r - p) < TOLERANCIA:
            continue
        refs = detalle.get((mz, lt), [f"LT{lt}"])
        por_mz.setdefault(mz, []).extend(refs)
    return {mz: "; ".join(refs) for mz, refs in por_mz.items()}


def _diferencias_por_mz(reporte: dict, planilla: dict, referencia: dict = None) -> list[dict]:
    """DIFERENCIA = PLANILLA − REPORTE. Incluye MZs de ambas fuentes."""
    referencia = referencia or {}
    mzs = sorted(set(reporte) | set(planilla))
    return [
        {
            "mz":        mz,
            "reporte":   round(reporte.get(mz, 0.0), 2),
            "planilla":  round(planilla.get(mz, 0.0), 2),
            "diferencia": round(planilla.get(mz, 0.0) - reporte.get(mz, 0.0), 2),
            "ok":        abs(planilla.get(mz, 0.0) - reporte.get(mz, 0.0)) < TOLERANCIA,
            "referencia": referencia.get(mz, ""),
        }
        for mz in mzs
    ]

# ── EXPORT: HOJA RESUMEN ──────────────────────────────────────────────────────
def _gh_resumen(ws, row, texto, bg, txt):
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value=texto)
    c.font      = Font(name="Arial", size=8, bold=True, color=txt)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = _borde()
    ws.row_dimensions[row].height = 14

def _fila_resumen(ws, row, seccion, componente, monto, estado=None):
    bg_sec  = "F9FAFB"; txt_sec  = "6B7280"
    bg_comp = "FFFFFF"; txt_comp = "111111"
    _c(ws, row, 1, seccion,    bg=bg_sec,  txt=txt_sec,  size=9)
    _c(ws, row, 2, componente, bg=bg_comp, txt=txt_comp, size=9)
    _c(ws, row, 3, monto,      bg="F0F8FF", txt="185FA5", mono=True, align="right")
    if estado is None:
        _c(ws, row, 4, "",     bg="F9FAFB", txt="9CA3AF", align="center", size=8)
        _c(ws, row, 5, "",     bg="F9FAFB")
    else:
        ok = abs(monto) < TOLERANCIA if isinstance(monto, (int, float)) else False
        dif_bg  = TD_OK  if ok else TD_ERR
        dif_txt = "085041" if ok else "991B1B"
        _c(ws, row, 4, monto,           bg=dif_bg, txt=dif_txt, mono=True, align="right", bold=True)
        est = ws.cell(row=row, column=5, value="OK" if ok else "ALERTA")
        est.font      = Font(name="Arial", size=9, bold=True, color=dif_txt)
        est.fill      = PatternFill("solid", start_color=dif_bg)
        est.alignment = Alignment(horizontal="center", vertical="center")
        est.border    = _borde()
    ws.row_dimensions[row].height = 16


def _hoja_resumen(wb, crudo_tepago, crudo_pagaste,
                  agua, agua_neto, blancos_tot, devueltos_tot, tanque_tot, otros_tot, pagaste_proc,
                  cob_yape_total, efec_total, cob_efec_total, periodo):
    ws = wb.create_sheet("resumen", 0)
    ws.freeze_panes = "A2"
    fecha_min, fecha_max = periodo
    periodo_str = (f"{fecha_min.strftime('%Y-%m-%d')} / {fecha_max.strftime('%Y-%m-%d')}"
                   if fecha_min else "—")

    ws.merge_cells("A1:E1")
    t = ws.cell(row=1, column=1, value=f"Resumen validación — {periodo_str}")
    t.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    t.fill      = PatternFill("solid", start_color="1E3A5F")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    cabeceras = ["SECCIÓN", "COMPONENTE", "MONTO", "DIFERENCIA", "ESTADO"]
    for ci, h in enumerate(cabeceras, 1):
        _c(ws, 2, ci, h, bg="F3F4F6", txt="374151", bold=True, align="center")
    ws.row_dimensions[2].height = 18

    ri = 3

    # ── Nivel 1a: TE PAGÓ ──────────────────────────────────────
    _gh_resumen(ws, ri, "Nivel 1a — Banco crudo (TE PAGÓ)  vs  agua + blancos + tanque + otros", "E6F1FB", "0C447C")
    ri += 1
    _fila_resumen(ws, ri, "banco crudo", "Total TE PAGÓ en crudo del banco", crudo_tepago)
    ri += 1
    _fila_resumen(ws, ri, "  agua",      "Pagos de agua identificados a un lote", agua)
    ri += 1
    _fila_resumen(ws, ri, "  blancos",   "Sin dueño identificado (blancos_mes + tepago MZ=BLANCO)", blancos_tot)
    ri += 1
    _fila_resumen(ws, ri, "  tanque",   "Aportes al tanque comunitario (CONCEPTO=tanque)", tanque_tot)
    ri += 1
    _fila_resumen(ws, ri, "  otros conceptos", "CONCEPTO libre ≠ tanque (deuda_directiva, etc.)", otros_tot)
    ri += 1
    suma_partes = round(agua + blancos_tot + tanque_tot + otros_tot, 2)
    dif_1a = round(suma_partes - crudo_tepago, 2)
    _fila_resumen(ws, ri, "∑ partes", "agua + blancos + tanque + otros", suma_partes)
    ri += 1
    _fila_resumen(ws, ri, "diferencia 1a", "∑ partes − banco crudo", dif_1a, estado=True)
    ri += 1
    _fila_resumen(ws, ri, "  devueltos (informativo)", "Devueltos al remitente — reconciliado en Nivel 1b, no en 1a", devueltos_tot)
    ri += 1

    # ── Nivel 1b: PAGASTE ──────────────────────────────────────
    _gh_resumen(ws, ri, "Nivel 1b — Banco crudo (PAGASTE)  vs  pagaste procesado", "F5F3FF", "4C1D95")
    ri += 1
    _fila_resumen(ws, ri, "banco PAGASTE",    "Total PAGASTE en crudo del banco", crudo_pagaste)
    ri += 1
    _fila_resumen(ws, ri, "pagaste procesado","Gastos clasificados por motor_matching", pagaste_proc)
    ri += 1
    dif_1b = round(pagaste_proc - crudo_pagaste, 2)
    _fila_resumen(ws, ri, "diferencia 1b", "pagaste procesado − banco PAGASTE", dif_1b, estado=True)
    ri += 1

    # ── Nivel 2: agua vs planilla yape ────────────────────────
    _gh_resumen(ws, ri, "Nivel 2 — Agua procesada  vs  planilla cobrada Yape", "FEF9E7", "7D6608")
    ri += 1
    _fila_resumen(ws, ri, "agua procesada (neto)", "pagos_yape_tepago CONCEPTO=vacío − devoluciones", agua_neto)
    ri += 1
    _fila_resumen(ws, ri, "planilla yape",  "planilla_cobrado.MONTO_YAPE", cob_yape_total)
    ri += 1
    dif_2 = round(cob_yape_total - agua_neto, 2)
    _fila_resumen(ws, ri, "diferencia 2", "planilla − agua neto", dif_2, estado=True)
    ri += 1

    # ── Efectivo ───────────────────────────────────────────────
    _gh_resumen(ws, ri, "Efectivo — procesado  vs  planilla cobrada", "ECFDF5", "085041")
    ri += 1
    _fila_resumen(ws, ri, "efectivo procesado", "pagos_efectivo (excl. BLANCO)", efec_total)
    ri += 1
    _fila_resumen(ws, ri, "planilla efectivo",  "planilla_cobrado.MONTO_EFECTIVO", cob_efec_total)
    ri += 1
    dif_efec = round(cob_efec_total - efec_total, 2)
    _fila_resumen(ws, ri, "diferencia efectivo", "planilla − procesado", dif_efec, estado=True)

    for ci, w in enumerate([20, 46, 14, 14, 10], 1):
        _w(ws, ci, w)

# ── EXPORT: HOJA POR MZ ───────────────────────────────────────────────────────
# Col layout:
#  1    ¿Qué período?   PERIODO
#  2    sep
#  3    ¿Quién?         MZ
#  4    sep
#  5    ¿Qué entró?     REPORTE
#  6    sep
#  7    ¿Qué quedó?     PLANILLA
#  8    sep
#  9-10 ¿Cuadra?        DIFERENCIA  ESTADO
#  11   sep
#  12   ¿Por qué?       REFERENCIA

_GRUPOS = [
    (1, 1, "¿Qué período?",  GH_PERIODO[0],  GH_PERIODO[1]),
    (3, 3, "¿Quién?",        GH_QUIEN[0],    GH_QUIEN[1]),
    (5, 5, "¿Qué entró?",   GH_REPORTE[0],  GH_REPORTE[1]),
    (7, 7, "¿Qué quedó?",   GH_PLANILLA[0], GH_PLANILLA[1]),
    (9,10, "¿Cuadra?",       GH_CUADRA[0],   GH_CUADRA[1]),
    (12,12,"¿Por qué?",      GH_PORQUE[0],   GH_PORQUE[1]),
]
_SEP_COLS = [2, 4, 6, 8, 11]

def _hoja_por_mz(wb, sheet_name, lbl_reporte, lbl_planilla, filas, periodo):
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A3"
    fecha_min, fecha_max = periodo
    periodo_str = (f"{fecha_min.strftime('%Y-%m-%d')} / {fecha_max.strftime('%Y-%m-%d')}"
                   if fecha_min else "—")
    last_row = len(filas) + 3  # 2 cabeceras + datos + fila total

    # Fila 1 — grupos
    for cs, ce, texto, bg, txt in _GRUPOS:
        _gh(ws, 1, cs, ce, texto, bg, txt)

    # Fila 2 — columnas
    col_defs = [
        (1,  "PERIODO",       GH_PERIODO[0],  GH_PERIODO[1],  24),
        (3,  "MZ",            GH_QUIEN[0],    GH_QUIEN[1],     8),
        (5,  lbl_reporte,     GH_REPORTE[0],  GH_REPORTE[1],  16),
        (7,  lbl_planilla,    GH_PLANILLA[0], GH_PLANILLA[1], 16),
        (9,  "DIFERENCIA",    GH_CUADRA[0],   GH_CUADRA[1],   13),
        (10, "ESTADO",        GH_CUADRA[0],   GH_CUADRA[1],   10),
        (12, "REFERENCIA",    GH_PORQUE[0],   GH_PORQUE[1],   55),
    ]
    for col, nombre, bg, txt, ancho in col_defs:
        _ch(ws, 2, col, nombre, bg, txt)
        _w(ws, col, ancho)

    # Separadores (todos los rows incluyendo total)
    for sc in _SEP_COLS:
        _sep(ws, sc, last_row)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 20

    # Datos
    for ri, r in enumerate(filas, 3):
        ok      = r["ok"]
        dif_bg  = TD_OK  if ok else TD_ERR
        dif_txt = "085041" if ok else "991B1B"
        _c(ws, ri, 1,  periodo_str,    TD_PERIODO,  "374151", mono=True, size=9)
        _c(ws, ri, 3,  r["mz"],        TD_QUIEN,    "5B21B6", mono=True, align="center")
        _c(ws, ri, 5,  r["reporte"],   TD_REPORTE,  "185FA5", mono=True, align="right")
        _c(ws, ri, 7,  r["planilla"],  TD_PLANILLA, "7D6608", mono=True, align="right")
        _c(ws, ri, 9,  r["diferencia"],dif_bg,      dif_txt,  bold=True, mono=True, align="right")
        c_e = ws.cell(row=ri, column=10, value="OK" if ok else "ALERTA")
        c_e.font      = Font(name="Arial", size=9, bold=True, color=dif_txt)
        c_e.fill      = PatternFill("solid", start_color=dif_bg)
        c_e.alignment = Alignment(horizontal="center", vertical="center")
        c_e.border    = _borde()
        _c(ws, ri, 12, r.get("referencia", ""), dif_bg, dif_txt, align="left", size=8)
        ws.row_dimensions[ri].height = 17

    # Fila de totales
    tr      = last_row
    t_rep   = round(sum(r["reporte"]   for r in filas), 2)
    t_pla   = round(sum(r["planilla"]  for r in filas), 2)
    t_dif   = round(t_pla - t_rep, 2)
    ok_tot  = abs(t_dif) < TOLERANCIA
    dif_bg  = TD_OK  if ok_tot else TD_ERR
    dif_txt = "085041" if ok_tot else "991B1B"

    for col in range(1, 13):
        _c(ws, tr, col, None, bg="F3F4F6")
    _c(ws, tr, 1, "TOTAL",  "F3F4F6", "374151", bold=True)
    _c(ws, tr, 5, t_rep,    TD_REPORTE,  "185FA5", bold=True, mono=True, align="right")
    _c(ws, tr, 7, t_pla,    TD_PLANILLA, "7D6608", bold=True, mono=True, align="right")
    _c(ws, tr, 9, t_dif,    dif_bg,      dif_txt,  bold=True, mono=True, align="right")
    c_e = ws.cell(row=tr, column=10, value="OK" if ok_tot else "ALERTA")
    c_e.font      = Font(name="Arial", size=9, bold=True, color=dif_txt)
    c_e.fill      = PatternFill("solid", start_color=dif_bg)
    c_e.alignment = Alignment(horizontal="center", vertical="center")
    c_e.border    = _borde()
    ws.row_dimensions[tr].height = 18

    for sc in _SEP_COLS:
        _sep(ws, sc, last_row)

# ── MAIN ──────────────────────────────────────────────────────────────────────
def _sellar_estado_ciclo():
    """Marca arrastre.validado=true en estado_ciclo.json para el ciclo que tiene generado=true."""
    sellado = repo_estado.sellar_validado(ruta=ESTADO_CICLO_PATH)
    if not sellado:
        log.info("estado_ciclo.json — ningún arrastre pendiente de sello")
        return
    log.info(f"estado_ciclo.json → sellado validado=true · ciclos: {', '.join(sellado)}")


def main():
    print("\n" + "═" * 55)
    print("  5b_validacion — Validación de cobranza")
    print("═" * 55)
    _init_logging()

    print("\n[1/4] Validando inputs...")
    _validar_inputs()

    print("\n[2/4] Cargando datos...")
    yape_mz, yape_lote, yape_detalle, agua_total, periodo  = _cargar_yape_tepago()
    pagaste_proc, periodo_pagaste = _cargar_pagaste_total()
    crudo_tepago, crudo_pagaste   = _cargar_crudo_por_tipo(periodo, periodo_pagaste)
    blancos_tot                   = _cargar_blancos_total()
    devueltos_tot                 = _cargar_devueltos_total()
    tanque_tot                    = _cargar_tanque_total()
    otros_tot                     = _cargar_otros_conceptos()
    efec_mz, efec_lote, efec_detalle, efec_total = _cargar_efectivo()
    cob_yape_mz, cob_efec_mz, cob_yape_lote, cob_efec_lote, cob_yape_total, cob_efec_total = _cargar_cobranza()

    print("\n[3/4] Calculando diferencias...")
    # Nivel 2 compara contra planilla_cobrado, que ya netea devoluciones
    # (MONTO_YAPE queda vacío para el lote devuelto) — agua_total de Nivel 1a
    # NO se toca (debe quedar crudo para cuadrar contra crudo_tepago).
    yape_mz_neto   = _netear_devoluciones(yape_mz)
    agua_total_neto = round(sum(yape_mz_neto.values()), 2)
    ref_yape = _referencia_por_mz(yape_lote, cob_yape_lote, yape_detalle)
    ref_efec = _referencia_por_mz(efec_lote, cob_efec_lote, efec_detalle)
    dif_yape = _diferencias_por_mz(yape_mz_neto, cob_yape_mz, ref_yape)
    dif_efec = _diferencias_por_mz(efec_mz, cob_efec_mz, ref_efec)

    # devueltos_tot NO entra: es plata PAGASTE (salida), tiene su propia
    # reconciliación en Nivel 1b. El pago original de Rosalina ya está
    # contado una vez en agua_total y otra en crudo_tepago — sumar
    # devueltos_tot aquí agregaba una tercera cantidad ajena a TE PAGÓ.
    dif_1a       = round((agua_total + blancos_tot + tanque_tot + otros_tot) - crudo_tepago, 2)
    dif_1b       = round(pagaste_proc - crudo_pagaste, 2)
    dif_2        = round(cob_yape_total - agua_total_neto, 2)
    dif_efec_tot = round(cob_efec_total - efec_total, 2)

    ok_1a        = abs(dif_1a) < TOLERANCIA
    ok_1b        = abs(dif_1b) < TOLERANCIA
    ok_2         = abs(dif_2) < TOLERANCIA
    ok_efec_tot  = abs(dif_efec_tot) < TOLERANCIA
    alertas_efec = sum(1 for r in dif_efec if not r["ok"])

    print("\n[4/4] Exportando reporte...")
    wb = Workbook()
    wb.remove(wb.active)
    _hoja_resumen(wb, crudo_tepago, crudo_pagaste,
                  agua_total, agua_total_neto, blancos_tot, devueltos_tot, tanque_tot, otros_tot, pagaste_proc,
                  cob_yape_total, efec_total, cob_efec_total, periodo)
    _hoja_por_mz(wb, "yape_por_mz",     "AGUA_PROCESADA",   "PLANILLA_YAPE",     dif_yape, periodo)
    _hoja_por_mz(wb, "efectivo_por_mz", "REPORTE_EFECTIVO", "PLANILLA_EFECTIVO", dif_efec, periodo)
    wb.save(OUT_DIR / "validacion_diferencias.xlsx")

    todo_ok = ok_1a and ok_1b and ok_2 and ok_efec_tot
    print("\n" + "=" * 55)
    if todo_ok:
        print("  VALIDACION OK — todos los niveles cuadran")
        log.info("VALIDACION OK")
        _sellar_estado_ciclo()
    else:
        if not ok_1a:
            print(f"  ALERTA  Nivel 1a (TE PAGO): dif={dif_1a:+.2f}")
            log.warning(f"Nivel 1a no cuadra: dif={dif_1a:.2f}")
        if not ok_1b:
            print(f"  ALERTA  Nivel 1b (PAGASTE): dif={dif_1b:+.2f}")
            log.warning(f"Nivel 1b no cuadra: dif={dif_1b:.2f}")
        if not ok_2:
            print(f"  ALERTA  Nivel 2 (agua vs planilla): dif={dif_2:+.2f}")
            log.warning(f"Nivel 2 no cuadra: dif={dif_2:.2f}")
        if not ok_efec_tot:
            print(f"  ALERTA  Efectivo total: dif={dif_efec_tot:+.2f}")
            log.warning(f"Efectivo total no cuadra: dif={dif_efec_tot:.2f}")
        if alertas_efec:
            print(f"  ALERTA  Efectivo por MZ: {alertas_efec} MZ(s) con diferencia")
        print("  Revisar outputs/validacion_diferencias.xlsx")
    print("=" * 55 + "\n")


if __name__ == "__main__":
    main()
